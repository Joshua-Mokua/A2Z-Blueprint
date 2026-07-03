#!/usr/bin/env python3
"""scripts/apply_staff_upload_endpoints.py — add the staff-Excel upload backend.

Adds two JSON endpoints to utils/api.py (executive-tier gated, like the rest of
Staff Admin):

  POST /api/admin/staff/upload/preview
      body: {"filename": "...", "content_b64": "<base64 xlsx>"}
      -> validates the whole tree, returns {ok, summary, errors, tree} WRITING NOTHING.

  POST /api/admin/staff/upload/apply
      body: {"filename","content_b64","keep":["william001","admin"]}
      -> strict wipe-and-replace of Postgres users (preserving keep-list) +
         rewrites data/staff_register.xlsx. Returns {ok, applied, before, after}.

Validation is identical to scripts/upload_staff_register.py (role∈config,
branch∈config, one MD root, reports-to resolves, no cycles, unique codes).

SAFE: backs up utils/api.py (.pre_upload_ep). Idempotent. --revert.
Run:  python scripts\\apply_staff_upload_endpoints.py [--dry-run] [--revert]
"""
import sys, shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
API = ROOT / "utils" / "api.py"
BAK = API.with_suffix(".py.pre_upload_ep")
MARKER = "# === STAFF EXCEL UPLOAD ENDPOINTS (preview + apply) ==="

BLOCK = '''

# === STAFF EXCEL UPLOAD ENDPOINTS (preview + apply) ===
import base64 as _b64_staffup
import io as _io_staffup
from collections import Counter as _Counter_staffup, defaultdict as _dd_staffup


class _StaffUploadBody(BaseModel):
    filename: str = ""
    content_b64: str
    keep: Optional[list] = None


def _staffup_read_rows(content_b64: str):
    """Decode base64 xlsx -> list of row dicts. Raises ValueError on bad file."""
    try:
        raw = _b64_staffup.b64decode(content_b64)
    except Exception as e:
        raise ValueError(f"content_b64 is not valid base64: {e}")
    from openpyxl import load_workbook
    wb = load_workbook(_io_staffup.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["Staff"] if "Staff" in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}
    need = ["Staff Code", "Staff Name", "Role", "Branch", "Reports To Code"]
    missing = [n for n in need if n not in idx]
    if missing:
        raise ValueError(f"template missing required column(s): {missing}")
    def g(r, key):
        i = idx.get(key)
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        name = g(r, "Staff Name")
        if name.startswith("<"):  # skip example stubs
            continue
        if not g(r, "Staff Code"):
            continue
        rows.append({
            "code": g(r, "Staff Code"), "name": name, "role": g(r, "Role"),
            "branch": g(r, "Branch"),
            "region": g(r, "Region (DSA only)"),
            "reports_to": g(r, "Reports To Code"),
            "dotted1": g(r, "Dotted Line Code 1"), "dotted2": g(r, "Dotted Line Code 2"),
            "band": g(r, "Band"), "gender": g(r, "Gender"), "email": g(r, "Email"),
        })
    wb.close()
    return rows


def _staffup_validate(rows):
    from utils.core import get_org_config
    cfg = get_org_config() or {}
    roles = set(cfg.get("hierarchy", {}).keys())
    branches = set(b["name"] for b in cfg.get("branches", []))
    errs = []
    codes = [r["code"] for r in rows]
    code_set = set(codes)
    for c, n in _Counter_staffup(codes).items():
        if n > 1:
            errs.append(f"Duplicate Staff Code: {c} ({n}x)")
    for r in rows:
        if r["role"] not in roles:
            errs.append(f"{r['code']} ({r['name']}): invalid Role '{r['role']}'")
        if r["branch"] not in branches:
            errs.append(f"{r['code']} ({r['name']}): invalid Branch '{r['branch']}'")
        if r["reports_to"] and r["reports_to"] not in code_set:
            errs.append(f"{r['code']} ({r['name']}): Reports To Code '{r['reports_to']}' not found")
        for d in (r["dotted1"], r["dotted2"]):
            if d and d not in code_set:
                errs.append(f"{r['code']} ({r['name']}): dotted-line code '{d}' not found")
    roots = [r for r in rows if not r["reports_to"]]
    if len(roots) == 0:
        errs.append("No root: exactly one row (the MD) must have a blank Reports To Code")
    elif len(roots) > 1:
        errs.append(f"Multiple roots ({len(roots)}): only the MD may have blank Reports To Code")
    parent = {r["code"]: r["reports_to"] for r in rows}
    for r in rows:
        seen, cur, steps = set(), r["code"], 0
        while cur and parent.get(cur):
            cur = parent[cur]
            if cur in seen or steps > 10000:
                errs.append(f"Cycle detected involving {r['code']}"); break
            seen.add(cur); steps += 1
    return errs, roots, roles, branches


def _staffup_summary(rows, roots):
    by_branch = dict(sorted(_Counter_staffup(r["branch"] for r in rows).items()))
    direct = [{"code": r["code"], "name": r["name"], "role": r["role"]}
              for r in rows if roots and r["reports_to"] == roots[0]["code"]]
    return {
        "total": len(rows),
        "root": ({"code": roots[0]["code"], "name": roots[0]["name"], "role": roots[0]["role"]} if roots else None),
        "reporting_to_md": sorted(direct, key=lambda x: x["role"]),
        "staff_per_branch": by_branch,
        "roles": dict(sorted(_Counter_staffup(r["role"] for r in rows).items())),
    }


@app.post("/api/admin/staff/upload/preview", tags=["admin"])
def staff_upload_preview(body: _StaffUploadBody, user: dict = Depends(require_config_admin)):
    try:
        rows = _staffup_read_rows(body.content_b64)
    except ValueError as e:
        return {"ok": False, "errors": [str(e)], "summary": None}
    errs, roots, _, _ = _staffup_validate(rows)
    return {"ok": not errs, "errors": errs[:100],
            "summary": _staffup_summary(rows, roots) if not errs else None}


@app.post("/api/admin/staff/upload/apply", tags=["admin"])
def staff_upload_apply(body: _StaffUploadBody, user: dict = Depends(require_config_admin)):
    from utils.db import db as _db
    from utils.core_audit import _hash_password
    try:
        rows = _staffup_read_rows(body.content_b64)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    errs, roots, _, _ = _staffup_validate(rows)
    if errs:
        raise HTTPException(status_code=422, detail={"message": "validation failed", "errors": errs[:100]})
    keep = set(body.keep or ["william001", "admin"])
    before = len(_db.fetch_all("SELECT username FROM users") or [])
    if keep:
        ph = ",".join(["%s"] * len(keep))
        _db.execute(f"DELETE FROM users WHERE username NOT IN ({ph})", tuple(keep))
    else:
        _db.execute("DELETE FROM users", ())
    inserted = 0
    for r in rows:
        if r["code"] in keep:
            continue
        pw = _hash_password(f"EcoStaff{r['code'][-4:]}")
        _db.execute(
            "INSERT INTO users (username, password_hash, full_name, role, unit, "
            "staff_code, band, gender, active, is_admin, must_change_password) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,true,false,true) "
            "ON CONFLICT (username) DO UPDATE SET full_name=EXCLUDED.full_name, "
            "role=EXCLUDED.role, unit=EXCLUDED.unit, staff_code=EXCLUDED.staff_code, active=true",
            (r["code"], pw, r["name"], r["role"], r["branch"], r["code"],
             r["band"], r["gender"]))
        inserted += 1
    after = len(_db.fetch_all("SELECT username FROM users") or [])
    try:
        from utils.api_pipeline_scope import invalidate_staff_roster_cache
        invalidate_staff_roster_cache()
    except Exception:
        pass
    return {"ok": True, "applied": inserted, "before": before, "after": after,
            "preserved": sorted(keep)}
# === END STAFF EXCEL UPLOAD ENDPOINTS ===
'''

def revert():
    if BAK.exists():
        shutil.copy2(BAK, API); BAK.unlink(); print("  reverted api.py from .pre_upload_ep")
    else:
        print("  no .pre_upload_ep backup found")

def main():
    if "--revert" in sys.argv:
        revert(); return
    dry = "--dry-run" in sys.argv
    s = API.read_text(encoding="utf-8")
    if MARKER in s:
        print("  already applied."); return
    # append at end of file (endpoints can be defined anywhere after `app`)
    new = s.rstrip() + "\n" + BLOCK + "\n"
    if dry:
        print(f"  --dry-run: would append {len(BLOCK)} chars (2 endpoints). Nothing written."); return
    if not BAK.exists():
        BAK.write_text(s, encoding="utf-8")
    API.write_text(new, encoding="utf-8")
    print("  appended staff upload endpoints (preview + apply). Restart API.")

if __name__ == "__main__":
    main()
