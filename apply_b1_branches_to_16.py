#!/usr/bin/env python3
"""scripts/apply_b1_branches_to_16.py — Phase B1 of branch reconciliation.

Make org_config.json hold the REAL 16 branches (the canonical set the 956 staff
already use), derived FROM the staff register — the working source of truth.
Branches have NO region tier anymore (they roll up directly to Head of Branches),
so every branch carries a uniform region value to satisfy code that reads
b["region"], without re-introducing the retired geographic regions.

This is ADDITIVE + SURGICAL: only the `branches` key of org_config.json is
replaced. All other keys (hierarchy, roles, role_categories, bank_name, DSA
regions, etc.) are preserved byte-for-byte.

SAFE: backs up org_config.json (.pre_b1). Idempotent. --revert restores backup.
Does NOT touch the hardcoded DEFAULT_ORG_CONFIG (that is Phase B2). Does NOT
touch staff reporting lines (that is Phase B-DSA).

Run:  python scripts\\apply_b1_branches_to_16.py [--dry-run] [--revert]
"""
import json, sys, shutil
from pathlib import Path
from collections import OrderedDict

ROOT = Path(__file__).resolve().parent.parent
ORG = ROOT / "data" / "org_config.json"
BAK = ORG.with_suffix(".json.pre_b1")
REGISTER = ROOT / "data" / "staff_register.xlsx"

# Branches have no region tier now; this uniform marker keeps b["region"] valid.
BRANCH_ROLLUP = "Head of Branches"
HEAD_OFFICE = "Head Office"


def derive_16_from_register():
    """Read distinct Units (=branches) from the staff register. Returns ordered
    list of branch names (Head Office first, then the 16 alphabetically)."""
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    unit_idx = hdr.index("Unit")
    units = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        u = row[unit_idx]
        if u and str(u).strip() and str(u) not in seen:
            seen.add(str(u)); units.append(str(u).strip())
    wb.close()
    # Head Office first, rest sorted
    branches = [HEAD_OFFICE] if HEAD_OFFICE in seen else []
    branches += sorted(u for u in units if u != HEAD_OFFICE)
    return branches


def build_branch_dicts(names):
    out = []
    n = 1
    for nm in names:
        is_ho = (nm == HEAD_OFFICE)
        out.append({
            "code": f"BRN{n:03d}",
            "name": nm,
            "region": "Head Office" if is_ho else BRANCH_ROLLUP,
            "county": "",
            "type": "HO" if is_ho else "Branch",
            "tier": 1 if is_ho else 2,
        })
        n += 1
    return out


def revert():
    if BAK.exists():
        shutil.copy2(BAK, ORG); BAK.unlink()
        print("  reverted org_config.json from .pre_b1")
    else:
        print("  no .pre_b1 backup found")


def main():
    if "--revert" in sys.argv:
        revert(); return
    dry = "--dry-run" in sys.argv

    names = derive_16_from_register()
    new_branches = build_branch_dicts(names)
    print(f"  derived {len(names)} branches from staff register:")
    for b in new_branches:
        print(f"      {b['code']}  {b['name']:<24} ({b['type']})")

    cfg = json.loads(ORG.read_text(encoding="utf-8"), object_pairs_hook=OrderedDict)
    old_count = len(cfg.get("branches", []))
    print(f"\n  org_config currently has {old_count} branches -> will become {len(new_branches)}")

    if dry:
        print("\n  --dry-run: no file written. Other keys preserved:",
              [k for k in cfg.keys() if k != "branches"])
        return

    if not BAK.exists():
        shutil.copy2(ORG, BAK)
    cfg["branches"] = new_branches  # ONLY this key changes
    ORG.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n  wrote {len(new_branches)} branches to org_config.json (other keys untouched)")
    print("  backup: data/org_config.json.pre_b1")
    print("\n  RESTART the API so rebuild_branch_maps() picks up the new list,")
    print("  or call utils.core.rebuild_branch_maps() in a fresh process.")


if __name__ == "__main__":
    main()
