"""
compute_actuals.py — Compute KPI actuals from CBS data.
Reads cbs_data/ and produces cbs_data/actuals_YYYY_MM.xlsx
that the A2Z app can load directly — no manual Excel upload needed.

Run AFTER generate_cbs.py and generate_staff.py.
"""
import csv, json
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

CBS  = Path(__file__).parent / "cbs_data"
DATA = Path(__file__).parent / "a2z" / "data"

TODAY      = date(2025, 12, 31)
YEAR       = 2025
CURR_MONTH = "Dec-25"

print("="*60)
print("A2Z Blueprint — CBS Actuals Computation")
print(f"Period: {CURR_MONTH} | Year: {YEAR}")
print("="*60)

# ── Load staff register ───────────────────────────────────────────────
print("\nLoading staff register...")
staff_list = []
sr_path = DATA/"staff_register.xlsx" if (DATA/"staff_register.xlsx").exists() else CBS/"staff_register.xlsx"

if sr_path.exists():
    import openpyxl
    wb = openpyxl.load_workbook(sr_path)
    ws = wb.active
    headers = [ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    for row in ws.iter_rows(min_row=2,values_only=True):
        r = dict(zip(headers,row))
        if r.get("Staff Code"): staff_list.append(r)
else:
    sr_path = CBS/"staff_register.csv"
    with open(sr_path,encoding="utf-8") as f:
        staff_list = list(csv.DictReader(f))

staff_by_code  = {str(s["Staff Code"]): s for s in staff_list}
staff_by_branch= defaultdict(list)
for s in staff_list:
    staff_by_branch[str(s.get("Branch Code",""))].append(str(s["Staff Code"]))

print(f"  {len(staff_list)} staff loaded")

# ── Load KPI map ──────────────────────────────────────────────────────
kpi_map = json.loads((CBS/"staff_kpi_map.json").read_text())
DEFAULT_KPIS = ["Compliance Score","Diligence Score"]


# KPI_WEIGHTS replaced by ROLE_KPI_WEIGHTS below

ROLE_KPI_WEIGHTS = {
    # ── 2026 BSC — real bank KPIs and weights ──────────────────────────
    "Branch Manager": {"PBT":0.12, "Total NFI":0.05, "Retail & MSME Deposit Growth":0.07, "Commercial Deposit Growth":0.04, "CASA Ratio":0.03, "Top 100 Customers Deposit":0.04, "Collection Throughput":0.04, "Number of Business Borrowers":0.03, "Disbursements Retail Loans":0.05, "Disbursements MSME Loans":0.05, "Disbursements Corporate Loans":0.03, "Loan Book Growth":0.05, "NPL Ratio":0.03, "PAR":0.05, "Account Dormancy":0.05, "Channel Dormancy":0.03, "New Accounts":0.03, "CX Score":0.03, "Compliance Score":0.03, "Audit Score":0.03, "Staff Productivity":0.12},
    "Senior Branch Manager": {"PBT":0.12, "Total NFI":0.05, "Retail & MSME Deposit Growth":0.07, "Commercial Deposit Growth":0.04, "CASA Ratio":0.03, "Top 100 Customers Deposit":0.04, "Collection Throughput":0.04, "Number of Business Borrowers":0.03, "Disbursements Retail Loans":0.05, "Disbursements MSME Loans":0.05, "Disbursements Corporate Loans":0.03, "Loan Book Growth":0.05, "NPL Ratio":0.03, "PAR":0.05, "Account Dormancy":0.05, "Channel Dormancy":0.03, "New Accounts":0.03, "CX Score":0.03, "Compliance Score":0.03, "Audit Score":0.03, "Staff Productivity":0.12},
    "Branch Operations Manager": {"Retail & MSME Deposit Growth":0.08, "CASA Ratio":0.03, "Total NFI":0.08, "Number of Business Borrowers":0.03, "Disbursements Retail Loans":0.05, "Disbursements MSME Loans":0.05, "Loan Book Growth":0.05, "NPL Ratio":0.05, "PAR":0.05, "Account Dormancy":0.05, "Channel Dormancy":0.03, "New Accounts":0.03, "CX Score":0.03, "Compliance Score":0.03, "Audit Score":0.03, "Collection Throughput":0.06, "Staff Productivity":0.22, "Top 100 Customers Deposit":0.03, "Commercial Deposit Growth":0.02},
    "Branch Operations Supervisor": {"PBT":0.0851, "Total NFI":0.0426, "Retail & MSME Deposit Growth":0.1064, "Commercial Deposit Growth":0.0213, "CASA Ratio":0.0319, "Top 100 Customers Deposit":0.0213, "Collection Throughput":0.0426, "Number of Business Borrowers":0.0213, "Disbursements Retail Loans":0.0532, "Disbursements MSME Loans":0.0532, "Disbursements Corporate Loans":0.0213, "Loan Book Growth":0.0532, "NPL Ratio":0.0319, "PAR":0.0532, "Account Dormancy":0.0532, "Channel Dormancy":0.0319, "New Accounts":0.0319, "CX Score":0.0319, "Compliance Score":0.0426, "Audit Score":0.0426, "Staff Productivity":0.1277},
    "Teller": {"PBT":0.0748, "Total NFI":0.0467, "Retail & MSME Deposit Growth":0.1402, "Commercial Deposit Growth":0.0093, "CASA Ratio":0.028, "Top 100 Customers Deposit":0.0374, "Collection Throughput":0.0374, "Number of Business Borrowers":0.028, "Disbursements Retail Loans":0.0467, "Disbursements MSME Loans":0.0467, "Disbursements Corporate Loans":0.0093, "Loan Book Growth":0.0467, "NPL Ratio":0.028, "PAR":0.0467, "Account Dormancy":0.0748, "Channel Dormancy":0.0374, "New Accounts":0.0561, "CX Score":0.028, "Compliance Score":0.0374, "Audit Score":0.028, "Staff Productivity":0.1121},
    "Customer Service Officer": {"PBT":0.0748, "Total NFI":0.0467, "Retail & MSME Deposit Growth":0.1402, "Commercial Deposit Growth":0.0093, "CASA Ratio":0.028, "Top 100 Customers Deposit":0.0374, "Collection Throughput":0.0374, "Number of Business Borrowers":0.028, "Disbursements Retail Loans":0.0467, "Disbursements MSME Loans":0.0467, "Disbursements Corporate Loans":0.0093, "Loan Book Growth":0.0467, "NPL Ratio":0.028, "PAR":0.0467, "Account Dormancy":0.0748, "Channel Dormancy":0.0374, "New Accounts":0.0561, "CX Score":0.028, "Compliance Score":0.0374, "Audit Score":0.028, "Staff Productivity":0.1121},
    "Branch Relationship Manager": {"PBT":0.0714, "Total NFI":0.0446, "Retail & MSME Deposit Growth":0.1071, "Commercial Deposit Growth":0.0357, "CASA Ratio":0.0268, "Top 100 Customers Deposit":0.0357, "Collection Throughput":0.0357, "Number of Business Borrowers":0.0536, "Disbursements Retail Loans":0.0714, "Disbursements MSME Loans":0.0714, "Disbursements Corporate Loans":0.0268, "Loan Book Growth":0.0446, "NPL Ratio":0.0268, "PAR":0.0446, "Account Dormancy":0.0446, "Channel Dormancy":0.0268, "New Accounts":0.0446, "CX Score":0.0268, "Compliance Score":0.0268, "Audit Score":0.0268, "Staff Productivity":0.1071},
    "Branch Senior Relationship Officer": {"PBT":0.0714, "Total NFI":0.0446, "Retail & MSME Deposit Growth":0.1071, "Commercial Deposit Growth":0.0357, "CASA Ratio":0.0268, "Top 100 Customers Deposit":0.0357, "Collection Throughput":0.0357, "Number of Business Borrowers":0.0536, "Disbursements Retail Loans":0.0714, "Disbursements MSME Loans":0.0714, "Disbursements Corporate Loans":0.0268, "Loan Book Growth":0.0446, "NPL Ratio":0.0268, "PAR":0.0446, "Account Dormancy":0.0446, "Channel Dormancy":0.0268, "New Accounts":0.0446, "CX Score":0.0268, "Compliance Score":0.0268, "Audit Score":0.0268, "Staff Productivity":0.1071},
    "Relationship Officer-Personal Banker": {"PBT":0.0714, "Total NFI":0.0446, "Retail & MSME Deposit Growth":0.1071, "Commercial Deposit Growth":0.0357, "CASA Ratio":0.0268, "Top 100 Customers Deposit":0.0357, "Collection Throughput":0.0357, "Number of Business Borrowers":0.0536, "Disbursements Retail Loans":0.0714, "Disbursements MSME Loans":0.0714, "Disbursements Corporate Loans":0.0268, "Loan Book Growth":0.0446, "NPL Ratio":0.0268, "PAR":0.0446, "Account Dormancy":0.0446, "Channel Dormancy":0.0268, "New Accounts":0.0446, "CX Score":0.0268, "Compliance Score":0.0268, "Audit Score":0.0268, "Staff Productivity":0.1071},
    "Relationship Officer-Business Banker": {"PBT":0.0714, "Total NFI":0.0446, "Retail & MSME Deposit Growth":0.1071, "Commercial Deposit Growth":0.0357, "CASA Ratio":0.0268, "Top 100 Customers Deposit":0.0357, "Collection Throughput":0.0357, "Number of Business Borrowers":0.0536, "Disbursements Retail Loans":0.0714, "Disbursements MSME Loans":0.0714, "Disbursements Corporate Loans":0.0268, "Loan Book Growth":0.0446, "NPL Ratio":0.0268, "PAR":0.0446, "Account Dormancy":0.0446, "Channel Dormancy":0.0268, "New Accounts":0.0446, "CX Score":0.0268, "Compliance Score":0.0268, "Audit Score":0.0268, "Staff Productivity":0.1071},
    "Direct Sales Representative - Assets & Liabilities": {"PBT":0.12, "Total NFI":0.05, "Retail & MSME Deposit Growth":0.07, "Commercial Deposit Growth":0.04, "CASA Ratio":0.03, "Top 100 Customers Deposit":0.04, "Collection Throughput":0.04, "Number of Business Borrowers":0.03, "Disbursements Retail Loans":0.05, "Disbursements MSME Loans":0.05, "Disbursements Corporate Loans":0.03, "Loan Book Growth":0.05, "NPL Ratio":0.03, "PAR":0.05, "Account Dormancy":0.05, "Channel Dormancy":0.03, "New Accounts":0.03, "CX Score":0.03, "Compliance Score":0.03, "Audit Score":0.03, "Staff Productivity":0.12},
    "Senior Digital Channels Officer": {"Retail & MSME Deposit Growth":0.1, "Channel Dormancy":0.15, "New Accounts":0.1, "Account Dormancy":0.08, "Collection Throughput":0.12, "CX Score":0.05, "Total NFI":0.08, "CASA Ratio":0.05, "Compliance Score":0.05, "Audit Score":0.05, "Staff Productivity":0.12, "Commercial Deposit Growth":0.05},
    "Area Manager": {"PBT":0.1, "Retail & MSME Deposit Growth":0.15, "Commercial Deposit Growth":0.05, "Loan Book Growth":0.1, "Total NFI":0.08, "NPL Ratio":0.08, "PAR":0.07, "New Accounts":0.05, "Account Dormancy":0.05, "Channel Dormancy":0.03, "CX Score":0.03, "Compliance Score":0.03, "Audit Score":0.04, "Staff Productivity":0.12, "CASA Ratio":0.02},
    "Chief Executive & Managing Director": {"PBT":0.15, "Retail & MSME Deposit Growth":0.1, "Commercial Deposit Growth":0.05, "Total NFI":0.1, "Loan Book Growth":0.1, "NPL Ratio":0.08, "PAR":0.05, "Collection Throughput":0.04, "New Accounts":0.04, "Account Dormancy":0.04, "CX Score":0.04, "Compliance Score":0.03, "Audit Score":0.03, "Staff Productivity":0.12, "CASA Ratio":0.03},
    "Relationship Manager - Corporate Banking": {"Commercial Deposit Growth":0.2, "Disbursements Corporate Loans":0.15, "Total NFI":0.15, "Loan Book Growth":0.1, "NPL Ratio":0.1, "PAR":0.05, "Number of Business Borrowers":0.08, "New Accounts":0.05, "CX Score":0.04, "Compliance Score":0.04, "Audit Score":0.04},
    "Relationship Manager - SME": {"Retail & MSME Deposit Growth":0.15, "Disbursements MSME Loans":0.15, "Total NFI":0.1, "Loan Book Growth":0.1, "NPL Ratio":0.1, "PAR":0.05, "Number of Business Borrowers":0.1, "New Accounts":0.08, "CX Score":0.05, "Compliance Score":0.05, "Audit Score":0.07},
    "Relationship Officer Bancassurance": {"Total NFI":0.3, "New Accounts":0.2, "Retail & MSME Deposit Growth":0.15, "CX Score":0.1, "Account Dormancy":0.1, "Compliance Score":0.08, "Audit Score":0.07},
}
DEFAULT_KPI_WEIGHTS = {"Compliance Score":0.50,"Diligence Score":0.50}

# ── Load baseline (31 Dec 2025 snapshot) ─────────────────────────────
print("\nLoading baseline...")
_baseline = {"rm":{}, "branch":{}}
for _bl_path in [CBS/"baseline_2025_Dec.json", DATA/"baseline_2025_Dec.json"]:
    if Path(_bl_path).exists():
        _baseline = json.loads(Path(_bl_path).read_text())
        print(f"  ✅ Baseline loaded from {_bl_path}")
        break
else:
    print("  ⚠️  Baseline not found — run compute_baseline.py first. Using zero baseline.")

_rm_base    = _baseline.get("rm",{})
_br_base    = _baseline.get("branch",{})

# ── Accumulators ──────────────────────────────────────────────────────
# Per RM: financial metrics
rm_metrics = defaultdict(lambda: {
    "loan_outstanding":0.0, "loan_new":0.0, "deposit_bal":0.0,
    "fee_income":0.0, "npl_loan":0.0, "total_loan":0.0,
    "acct_count":0, "loan_count":0, "cifs":set(),
})
# Per branch: operational metrics
br_metrics = defaultdict(lambda: {
    "new_customers":0, "dormant_reactivated":0,
    "kyc_verified":0, "kyc_total":0,
    "txn_count":0, "dfs_volume":0.0, "digital_cifs":set(),
})

print("\nProcessing accounts (streaming 1.2M records)...")
period_start = date(YEAR, 1, 1)

with open(CBS/"accounts.csv", encoding="utf-8") as f:
    for i,row in enumerate(csv.DictReader(f)):
        rm  = row.get("relationship_manager_code","")
        bc  = row.get("branch_code","")
        cat = row.get("category","")
        bal = float(row.get("current_balance") or 0)
        loan_out = float(row.get("loan_outstanding") or 0)
        loan_amt = float(row.get("loan_amount") or 0)
        fee      = float(row.get("fee_income_ytd") or 0)
        npl      = row.get("npl_status","")
        opened   = row.get("date_opened","")
        dormancy = row.get("dormancy_status","")
        last_txn = row.get("last_transaction_date","")
        cif      = row.get("cif","")

        r = rm_metrics[rm]
        r["cifs"].add(cif)
        r["acct_count"] += 1

        if cat in ("CASA","Term Deposit"):
            r["deposit_bal"] += bal

        if cat == "Loan":
            r["total_loan"]    += loan_out
            r["loan_outstanding"] += loan_out
            r["fee_income"]    += fee
            r["loan_count"]    += 1
            if npl == "NPL":
                r["npl_loan"]  += loan_out
            # New loans this year
            if opened and opened[:4] == str(YEAR):
                r["loan_new"]  += loan_amt
        else:
            r["fee_income"] += fee * 0.3  # partial fee for non-loan products

        # Branch operational metrics
        b = br_metrics[bc]
        if dormancy == "Dormant" and last_txn:
            try:
                lt = date.fromisoformat(last_txn)
                if lt >= date(YEAR, 10, 1):  # reactivated in last quarter
                    b["dormant_reactivated"] += 1
            except: pass

        if (i+1) % 200000 == 0:
            print(f"  {i+1:,} accounts processed...")

print("  ✅ Accounts processed")

print("\nProcessing customers...")
with open(CBS/"customers.csv", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        bc = row.get("branch_code","")
        kyc = row.get("kyc_status","")
        onboarded = row.get("date_onboarded","")
        b = br_metrics[bc]
        b["kyc_total"] += 1
        if kyc == "Verified": b["kyc_verified"] += 1
        if onboarded and onboarded[:4] == str(YEAR):
            b["new_customers"] += 1

print("  ✅ Customers processed")

print("\nProcessing transactions...")
with open(CBS/"transactions_sample.csv", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        bc  = row.get("branch_code","")
        ch  = row.get("txn_channel","")
        amt = float(row.get("amount") or 0)
        cif = row.get("cif","")
        b   = br_metrics[bc]
        b["txn_count"] += 1
        if ch == "Mobile Banking":
            b["dfs_volume"] += amt
            b["digital_cifs"].add(cif)

print("  ✅ Transactions processed")

# ── Build KPI actuals per staff member ───────────────────────────────
print("\nBuilding KPI actuals per staff member...")

# Annualise transaction count (sample is ~90 days → ×4)
TXN_SCALE = 4.0

rows = []
for s in staff_list:
    sc   = str(s["Staff Code"])
    role = str(s.get("Role",""))
    unit = str(s.get("Unit",""))
    bc   = str(s.get("Branch Code",""))

    rm  = rm_metrics.get(sc, rm_metrics["__default__"] if "__default__" in rm_metrics else defaultdict(float))
    br  = br_metrics.get(bc, {})

    # Compute actuals
    n_cifs    = len(rm.get("cifs",set())) or 1
    n_cifs_br = br.get("kyc_total",1) or 1

    # Baseline for this RM
    _rm_b  = _rm_base.get(sc, {})
    _br_b  = _br_base.get(bc, {})

    # Current portfolio values
    _dep_now  = rm.get("deposit_bal",0)
    _loan_now = rm.get("loan_outstanding",0)
    _fee_now  = rm.get("fee_income",0)

    # Baseline values (31 Dec 2025)
    _dep_base  = _rm_b.get("deposit_bal",0)
    _loan_base = _rm_b.get("loan_bal",0)
    _fee_base  = _rm_b.get("fee_income",0)
    _acct_base = _rm_b.get("active_accounts",0)

    # Growth = current - baseline (can be negative = contraction)
    _dep_growth  = _dep_now  - _dep_base
    _loan_growth = _loan_now - _loan_base
    _fee_growth  = _fee_now  - _fee_base
    _acct_growth = rm.get("active_accounts",0) - _acct_base

    _share = len(rm.get("cifs",set())) / max(n_cifs_br,1)

    actuals_by_kpi = {
        "Loan Book Growth":          _loan_growth,   # KES growth vs baseline
        "Loans Disbursement":        rm.get("loan_new",0),
        "Deposit Growth":            _dep_growth,    # KES growth vs baseline
        "Fees and Commission":       _fee_now,       # absolute YTD (not growth)
        "DFS Revenue":               br.get("dfs_volume",0) * _share,
        "NPL Ratio":                 (rm.get("npl_loan",0)/max(rm.get("total_loan",1),1))*100,
        "Active Account Growth":     _acct_growth,   # net active account change
        "Transactions":              br.get("txn_count",0) * TXN_SCALE * (1/max(len(staff_by_branch.get(bc,[1])),1)),
        "New Customer Acquisition":  br.get("new_customers",0) * _share,
        "Dormancy Reactivation":     br.get("dormant_reactivated",0) * _share,
        "Compliance Score":          (br.get("kyc_verified",0)/max(br.get("kyc_total",1),1))*100,
        "Digital Active Customers":  len(br.get("digital_cifs",set())) * _share,
        "Diligence Score":           85.0,
        "Staff Satisfaction Index":  78.0,
        "Trade Finance":             0.0,  # manual — treasury data
        "PBT":                       0.0,  # manual — finance data
        "NPS Score":                 0.0,  # manual — survey data
        "Audit Score":               0.0,  # manual — audit data
        "SLA Adherence Score":       0.0,  # manual — ops data
    }

    # Use role-specific KPI weights — keys are KPI names, values sum to 1.0
    role_wt_map = ROLE_KPI_WEIGHTS.get(role, DEFAULT_KPI_WEIGHTS)

    PILLAR_MAP = {
                "Total NFI":"Financial","Retail & MSME Deposit Growth":"Financial",
        "Commercial Deposit Growth":"Financial","Top 100 Customers Deposit":"Financial",
        "Collection Throughput":"Financial","Number of Business Borrowers":"Financial",
        "Disbursements Retail Loans":"Financial","Disbursements MSME Loans":"Financial",
        "Disbursements Corporate Loans":"Financial","Loan Book Growth":"Financial",
        "PAR":"Financial",
        "Account Dormancy":"Customer Focus","Channel Dormancy":"Customer Focus",
        "New Accounts":"Customer Focus","CX Score":"Customer Focus",
        "Staff Productivity":"People & Learning",
        "Audit Score":"Operational Excellence","Compliance Score":"Operational Excellence",
        "CASA Ratio":"Financial","PBT":"Financial",
        "NPL Ratio":"Financial","Fees and Commission":"Financial",
        "DFS Revenue":"Financial","Trade Finance":"Financial","Treasury Revenue":"Financial",
        "Bancassurance":"Financial","Ecosystem Banking":"Customer Focus",
        "Diligence Score":"People & Learning","Digital Active Customers":"Customer Focus",
        "SLA Adherence Score":"Operational Excellence","Audit Coverage Rate":"Operational Excellence",
    }

    for kpi, weight in role_wt_map.items():
        actual = actuals_by_kpi.get(kpi, 0.0)
        pillar = PILLAR_MAP.get(kpi, "Operational Excellence")
        rows.append({
            "Staff Code":   sc,
            "Staff Name":   s.get("Staff Name",""),
            "Role":         role,
            "Unit":         unit,
            "Category":     s.get("Category",""),
            "Staff Status": "Active",
            "KPI":          kpi,
            "Pillar":       pillar,
            "Weight":       round(weight, 4),
            "Annual Target":0,        # filled by cascade
            "YTD_Actual":   round(actual,2),
            CURR_MONTH:     round(actual/12,2),
            "Annual Actual":round(actual,2),
        })

print(f"  {len(rows)} KPI rows generated for {len(staff_list)} staff")

# ── Inject initiative KPIs for HO staff ──────────────────────────────
try:
    import sys as _sys_ik
    _sys_ik.path.insert(0, str(DATA.parent))
    from utils.core import compute_initiative_kpis as _cik
    from utils.core import get_kpi_library as _gkl_ik
    _lib_ik  = _gkl_ik()
    _rk_ik   = _lib_ik.get('role_kpis', {})
    _pw_ik   = _lib_ik.get('kpi_weights', {})
    _seen    = set()   # (staff_name, kpi) already added
    _init_rows = []
    for _r in rows:
        _sn   = _r.get('Staff Name', '')
        _role = _r.get('Role', '')
        _key  = (_sn, 'INIT_STATUS')
        if _key in _seen: continue
        _role_kpis = _rk_ik.get(_role, [])
        if not any(k in _role_kpis for k in ['INIT_STATUS','INIT_COUNT']):
            continue
        _seen.add(_key)
        _ia = _cik(_sn)
        _base = {k:v for k,v in _r.items()
                 if k not in ('KPI','Pillar','Weight','Annual Target','YTD_Actual',
                              CURR_MONTH,'Annual Actual')}
        for _kid, _kname, _pillar, _tgt in [
            ('INIT_STATUS','Initiative Implementation Score','Operational Excellence',100),
            ('INIT_COUNT', 'Active Initiatives Count',       'Operational Excellence',5),
        ]:
            if _kid in _role_kpis:
                _actual = _ia.get(_kname, 0)
                _wt     = _pw_ik.get(_kid, 0.03)
                _init_rows.append({
                    **_base,
                    'KPI':          _kname,
                    'Pillar':       _pillar,
                    'Weight':       _wt,
                    'Annual Target':_tgt,
                    'YTD_Actual':   _actual,
                    CURR_MONTH:     _actual,
                    'Annual Actual':_actual,
                })
    rows.extend(_init_rows)
    if _init_rows:
        print(f"  + {len(_init_rows)} initiative KPI rows added for HO staff")
except Exception as _ik_e:
    print(f"  ⚠️  Initiative KPI injection skipped: {_ik_e}")

# ── Save as Excel ─────────────────────────────────────────────────────
print("\nSaving actuals Excel file...")
out_name = f"actuals_{YEAR}_{CURR_MONTH.replace('-','_')}.xlsx"
out_path = CBS / out_name

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Data"

    if rows:
        headers = list(rows[0].keys())
        # Header
        for ci,h in enumerate(headers,1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="center")
        # Data
        for ri,row in enumerate(rows,3):
            for ci,h in enumerate(headers,1):
                ws.cell(row=ri, column=ci, value=row.get(h,""))
        # Widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 25

    wb.save(out_path)
    print(f"  ✅ {out_name} — {len(rows):,} rows")
except Exception as e:
    # CSV fallback
    import csv as _csv
    out_path = CBS / out_name.replace(".xlsx",".csv")
    with open(out_path,"w",newline="",encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"  ✅ {out_name.replace('.xlsx','.csv')} — {len(rows):,} rows")

# Also copy to data/ folder so app picks it up automatically
import shutil
dest = DATA / out_name
shutil.copy2(out_path, dest)
print(f"  ✅ Copied to a2z/data/{out_name}")

print("\n" + "="*60)
print("ACTUALS SUMMARY")
print("="*60)
total_loans    = sum(rm_metrics[k].get("loan_outstanding",0) for k in rm_metrics)
total_deposits = sum(rm_metrics[k].get("deposit_bal",0)      for k in rm_metrics)
total_fees     = sum(rm_metrics[k].get("fee_income",0)       for k in rm_metrics)
total_npl      = sum(rm_metrics[k].get("npl_loan",0)         for k in rm_metrics)
print(f"  Loan book:        KES {total_loans/1e9:.1f}B")
print(f"  Deposit book:     KES {total_deposits/1e9:.1f}B")
print(f"  Fees & Commission:KES {total_fees/1e6:.1f}M")
print(f"  NPL:              KES {total_npl/1e9:.1f}B ({total_npl/max(total_loans,1)*100:.1f}%)")
print(f"  KPI rows:         {len(rows):,}")
print("="*60)
print(f"\nNext step: open A2Z app → Admin → Upload → select {out_name}")
print("Or: app will auto-load from a2z/data/ on next restart")
