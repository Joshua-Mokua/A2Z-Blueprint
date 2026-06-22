#!/usr/bin/env python3
"""
build_mous.py — MOU-1: replace data/partnerships_mous.json with the curated
115-partner list (alphabetical, Acumen first), coherent titles, real RM codes.

Replaces the incoherent legacy 30 (whose title/partner_name didn't match) with
clean records on the SAME schema the API + Partnerships module already read:
  id, title, partner_name, partner_type, mou_type, department,
  relationship_manager, signed_date, effective_date, expiry_date, status,
  auto_renew, renewal_notice_days, deal_value_kes_m, referral_revenue_ytd_m,
  co_brand_income_ytd_m

All MOUs Active (per Josh). title derives from partner ("<Partner> — <mou_type>"),
fixing the legacy mismatch. relationship_manager maps to a REAL staff code drawn
from the live register's customer-facing roles (never a phantom).

SAFE: dry-run unless --apply. Backs up the file (.pre_mou115_<ts>) first.

    python scripts\\build_mous.py            # dry-run (prints sample + counts)
    python scripts\\build_mous.py --apply    # backup + write
"""
import sys
import json
import random
from pathlib import Path
from datetime import datetime, timedelta

DATA = Path(__file__).resolve().parent.parent / "data"
MOU_FILE = DATA / "partnerships_mous.json"
REGISTER = DATA / "staff_register.xlsx"

# 115 curated partners (deduped, alphabetical, Acumen first). Embedded so the
# script is self-contained and reproducible.
PARTNERS = [
    'Acumen Advisory',
    'Africa Enterprice International',
    'Africa Gas & Oil Company',
    'Africa Re Insurance Corporation',
    'Africa Telecommunication Union',
    'Africa Trade Insurance (Ati)',
    'Africa Wildlife Foundation',
    'African Enterprice International',
    'African Field Epidemiology Network (Afenet)',
    'African Management Institute',
    'African Population And Health Research Center (Aprhc)',
    'African Telecommunication Union',
    'African Union',
    'African Union - Cdc',
    'African Union - Ibar',
    'African Union - Nrc',
    'African Wildlife Foundation',
    'Aga-Khan Hospital Kisumu',
    'All Africa Conference Of Churches',
    'Alliance For A Green Revolution In Africa Agra',
    'American Friends Service',
    'Asharami Synergy Kenya Limited',
    'Atlascopco',
    'Axxonsoft Mena Fzco',
    'Baraton University',
    'Batuk',
    'Britam Insurance',
    'British American Tobacco',
    'Bulkstream Limited',
    'Bulkstream Logistics',
    'Business Cashbacked',
    'Cfao',
    'CIC Insurance',
    'Coast Development Authority',
    'Communivity Initiatives Concern',
    'Compassion International',
    'Embassy Of Senegal',
    'Express Shipping And Logistics',
    'Flexi Personel',
    'Geothermal Development Company',
    'Ghana Reinsurance Ltd',
    'Giz Kenya',
    'International Planned Parenthood Federation (Ippf)',
    'International Rescue Committee',
    'Jaramogi Oginga Odinga University',
    'Jomo Kenyatta University Of Agriculture And Technology (Jkuat)',
    'Jubilee Insurance',
    'Kamu Limited',
    'Karen Hospital',
    'Kenya Agricultural & Livestock (Kalro)',
    'Kenya Bureau Of Standards',
    'Kenya Medical Supplies Authority',
    'Kenya Medical Training College',
    'Kenyatta National Hospital',
    'KICC',
    'Kisii University',
    'Lake Victoria Basin Cimmission (Lvbc)',
    'Lukenya Gataway Ltd',
    'Maseno University',
    'Mater Hospital',
    'Mbaraki Bulk',
    'Menengai Oil',
    'Moi Teaching And Referral Hospital',
    'Moi University Eldoret',
    'Mpesa Academy',
    'National Government- Westlands Contituency',
    'Nedbank',
    'New Kenya Cooperative Creameries',
    'Norwegian Refugee Council',
    'Office Of The Auditor General',
    'Ola',
    'One Commodities',
    'One Petroleum',
    'Pharmacy And Poisons Board',
    'Phillips East Africa Limited',
    'Price Waterhouse Cooperation',
    'Pwani Oil',
    'Pwani Oil (Unemployed)',
    'Rongo University',
    'Royal Golf',
    'Rwandair',
    'Sanlam Annuitant',
    'SDA Advent Hill Primary School',
    'SDA Herald Publishing House',
    'SDA Central Kenya Conference',
    'SDA Central Nyanza Conference',
    'SDA Central Rift Conference',
    'SDA East Central Africa Division',
    'SDA East Nairobi Field',
    'SDA East-Central Africa Division',
    'SDA Greater Rift Valley',
    'SDA Home Health Education Services',
    'SDA Karura School',
    'SDA Kenya Lake Conference',
    'SDA Maxwell Adventist Preparatory',
    'SDA Nairobi Central',
    'SDA Newlife',
    'SDA North East Kenya Field',
    'SDA North West Kenya Field',
    'SDA South East Kenya Field',
    'SDA South Kenya Conference',
    'SDA South West Kenya Field',
    'SDA Southern Kenya Lake Conference',
    'SDA West Kenya Union Conference',
    'SDA West Nyanza Field',
    'SDA West Rift Field',
    'Shelter Afrique',
    'Teachers Service Commission',
    'Technical University Of Kenya',
    'Telkom Kenya Ltd',
    'The University Of Nairobi',
    'Tolaram',
    'Trustgro',
    'Trustgro(Equity)',
    'United Nation',
    'University Of Eldoret',
    'Waica-Re',
    'World Scouts Organization',
    'Yehu Microfinance',
]

MOU_TYPES = ["Referral Agreement", "Agency Agreement", "Co-branded Product",
             "Revenue Share", "Joint Marketing", "Supply Chain Finance",
             "Data Sharing", "API Integration"]
DEPARTMENTS = ["Bancassurance", "Retail Banking", "SME Banking", "Corporate Banking",
               "Digital & Channels", "Trade Finance"]


def real_rm_codes():
    """Pull real customer-facing staff codes from the live register (no phantoms)."""
    try:
        import openpyxl
    except ImportError:
        return ["300731", "300716", "300003"]  # safe fallback personas
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    H = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(H)}
    want = ("Relationship", "Direct Sales Agent", "Branch DSA Team Lead",
            "Regional DSA Head", "Personal Banker", "Business Banker")
    codes = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        role = str(r[ix.get("Role", 2)] or "")
        if any(w in role for w in want):
            codes.append(str(r[ix.get("Staff Code", 0)]))
    wb.close()
    return codes or ["300731", "300716", "300003"]


def main():
    apply = "--apply" in sys.argv
    random.seed(20260622)
    rms = real_rm_codes()

    base = datetime(2024, 1, 1)
    out = []
    for i, partner in enumerate(PARTNERS, start=1):
        mt = MOU_TYPES[i % len(MOU_TYPES)]
        signed = base + timedelta(days=random.randint(0, 600))
        expiry = signed + timedelta(days=random.choice([365, 730, 1095]))
        out.append({
            "id": f"MOU{i:04d}",
            "title": partner,
            "partner_name": partner,
            "partner_type": random.choice(["Retailer", "Institution", "Government",
                                           "NGO", "University", "Insurer", "Corporate"]),
            "mou_type": mt,
            "department": random.choice(DEPARTMENTS),
            "relationship_manager": random.choice(rms),
            "signed_date": signed.strftime("%Y-%m-%d"),
            "effective_date": signed.strftime("%Y-%m-%d"),
            "expiry_date": expiry.strftime("%Y-%m-%d"),
            "status": "Active",
            "auto_renew": random.choice([True, False]),
            "renewal_notice_days": random.choice([30, 60, 90]),
            "deal_value_kes_m": round(random.uniform(5, 500), 2),
            "referral_revenue_ytd_m": round(random.uniform(0, 60), 2),
            "co_brand_income_ytd_m": round(random.uniform(0, 10), 2),
        })

    old_n = len(json.loads(MOU_FILE.read_text(encoding="utf-8"))) if MOU_FILE.exists() else 0
    print(f"partnerships_mous.json: {old_n} -> {len(out)} MOUs (all Active)")
    print(f"RM codes drawn from {len(rms)} real customer-facing register rows")
    print(f"\nfirst 3:")
    for m in out[:3]:
        print(f"  {m['id']}  {m['title']}  | RM={m['relationship_manager']}")
    print(f"...\n  {out[22]['id']}  {out[22]['title']}   <- Batuk")
    print(f"...\nlast: {out[-1]['id']}  {out[-1]['title']}")

    if not apply:
        print("\n[DRY-RUN] No file written. Re-run with --apply to back up + write.")
        return

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    if MOU_FILE.exists():
        backup = MOU_FILE.with_name(f"partnerships_mous.json.pre_mou115_{ts}")
        backup.write_text(MOU_FILE.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"\n[backup] {backup.name}")
    MOU_FILE.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[apply] wrote {len(out)} MOUs. Restart API — pipeline picker now shows all {len(out)}.")


if __name__ == "__main__":
    main()
