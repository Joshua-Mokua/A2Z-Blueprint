#!/usr/bin/env python3
"""
wl_owner_split_diag.py — READ-ONLY. Classify watchlist rows by the REGISTER ROLE
of their rm_code, to decide delete-vs-resync. Writes nothing.

Buckets every row:
  EXEC/NON-RM owner  -> deletion candidate (credit accounts shouldn't sit under
                       MD/Directors/Heads; this is the stale 'Veronica Mutai' class)
  REAL RM owner      -> keep; resync rm_name from register if stale
  WILLIAM (300001)   -> ALWAYS protected, reported separately, never touched

    python scripts\\wl_owner_split_diag.py
"""
import sys
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA_DIR / "staff_register.xlsx"

PROTECT = {"300001"}  # William Mwanake, MD — never touch

# Register roles that are NOT account-owning RMs (exec / management / non-credit)
EXEC_HINTS = ("managing director", "chief", "director", "head of", "head,",
              "general manager", "company secretary", "board")
RM_HINTS = ("relationship", "officer", "credit", "sales", "personal banker",
            "rm ", "ro ", "dso", "branch manager")


def register_lookup():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = headers.index("Staff Code") if "Staff Code" in headers else 0
    ni = headers.index("Staff Name") if "Staff Name" in headers else 1
    ri = headers.index("Role") if "Role" in headers else 2
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[ci]:
            out[str(row[ci]).strip()] = (str(row[ni] or "").strip(), str(row[ri] or "").strip())
    wb.close()
    return out


def classify(role: str) -> str:
    rl = (role or "").lower()
    if any(h in rl for h in EXEC_HINTS):
        return "EXEC"
    if any(h in rl for h in RM_HINTS):
        return "RM"
    return "OTHER"


def main():
    from utils.db import db
    reg = register_lookup()
    rows = db.fetch_all("SELECT id, rm_code, rm_name FROM watchlist", ())
    print(f"register codes: {len(reg)} | watchlist rows: {len(rows)}\n")

    buckets = Counter()
    exec_rows, rm_resync, other_rows, protected = [], [], [], 0
    for r in rows:
        code = str(r.get("rm_code") or "").strip()
        cur = str(r.get("rm_name") or "").strip()
        if code in PROTECT:
            protected += 1
            continue
        reg_name, reg_role = reg.get(code, (None, None))
        if reg_name is None:
            buckets["UNRESOLVED"] += 1
            other_rows.append((code, cur, "(not in register)"))
            continue
        kind = classify(reg_role)
        buckets[kind] += 1
        if kind == "EXEC":
            exec_rows.append((code, cur, reg_role))
        elif kind == "RM":
            if reg_name != cur:
                rm_resync.append((code, cur, reg_name))
        else:
            other_rows.append((code, cur, reg_role))

    print(f"PROTECTED (William 300001):     {protected}  (never touched)")
    print(f"EXEC-owned rows (DELETE cand):  {buckets['EXEC']}")
    print(f"RM-owned rows:                  {buckets['RM']}  "
          f"(of which {len(rm_resync)} need name resync)")
    print(f"OTHER role rows:                {buckets['OTHER']}")
    print(f"UNRESOLVED rm_code:             {buckets['UNRESOLVED']}\n")

    if exec_rows:
        print("--- EXEC-owned (delete candidates): distinct code -> role / stale name ---")
        seen = set()
        for code, cur, role in exec_rows:
            if code not in seen:
                print(f"  {code}: role='{role}'  stale rm_name='{cur}'")
                seen.add(code)
    if other_rows:
        print("\n--- OTHER/unresolved (review): ---")
        seen = set()
        for code, cur, role in other_rows:
            if code not in seen:
                print(f"  {code}: {role}  rm_name='{cur}'")
                seen.add(code)


if __name__ == "__main__":
    main()
