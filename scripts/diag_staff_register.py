#!/usr/bin/env python3
"""
diag_staff_register.py — READ-ONLY. Reads the staff register the SAME way the app
auto-loader does (openpyxl, wb.active, headers on row 1, data from row 2) so the
B1 remap is built against the real columns. Writes nothing.

Run (in .venv):  python scripts\diag_staff_register.py
"""
import json
from pathlib import Path
from collections import Counter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"


def main():
    sr = DATA_DIR / "staff_register.xlsx"
    print(f"DATA_DIR = {DATA_DIR}")
    print(f"staff_register.xlsx exists: {sr.exists()}")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(sr), read_only=True)
        ws = wb.active
        print(f"active sheet: {ws.title!r} | sheets: {wb.sheetnames}")
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        print(f"\n--- headers (row 1, {len(headers)} cols) ---\n{headers}")

        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        print(f"\n--- data rows: {len(rows)} ---")

        def idx(name, default=None):
            return headers.index(name) if name in headers else default

        i_code = idx("Staff Code", 0)
        i_name = idx("Staff Name", idx("Name", 1))
        i_role = idx("Role", 2)
        i_bname = idx("Branch Name")
        i_bcode = idx("Branch Code")

        if i_bname is not None:
            counts = Counter(str(r[i_bname]) for r in rows if r[i_bname])
            print(f"\n--- staff per Branch Name ({len(counts)} branches) ---")
            for name, n in sorted(counts.items()):
                code = ""
                if i_bcode is not None:
                    cs = {str(r[i_bcode]) for r in rows
                          if str(r[i_bname]) == name and r[i_bcode]}
                    code = f"[{'/'.join(sorted(cs))}]" if cs else ""
                print(f"   {n:>4}  {code:<12} {name}")

        nums = [int(str(r[i_code])) for r in rows if str(r[i_code]).isdigit()]
        if nums:
            print(f"\n--- staff code range: {min(nums)} .. {max(nums)} ---")

        print("\n--- persona rows ---")
        for target in ("300731", "300716", "300001"):
            hit = next((r for r in rows if str(r[i_code]) == target), None)
            if hit:
                print(f"   {target}: " + " | ".join(
                    f"{h}={hit[k]}" for k, h in enumerate(headers) if hit[k] is not None))
            else:
                print(f"   {target}: NOT FOUND")
        wb.close()
    except Exception as e:
        print(f"[register read error] {type(e).__name__}: {e}")

    uj = DATA_DIR / "users.json"
    print(f"\n\nusers.json exists: {uj.exists()}")
    try:
        users = json.loads(uj.read_text(encoding="utf-8"))
        recs = users.get("users", users) if isinstance(users, dict) else users
        keys = list(recs.keys()) if isinstance(recs, dict) else [u.get("username") for u in recs]
        print(f"users count: {len(keys)}")
        sample = recs.get("frank0731") if isinstance(recs, dict) else None
        if sample is None and isinstance(recs, dict):
            sample = next(iter(recs.values()))
        print(f"\n--- sample user record (keys) ---\n{list(sample.keys()) if sample else '?'}")
        if sample:
            safe = {k: ("<hash>" if "pass" in k.lower() else v) for k, v in sample.items()}
            print(f"frank0731 (redacted): {safe}")
        for u in ("frank0731", "immaculate0716", "william001"):
            print(f"   {u}: {'present' if u in keys else 'MISSING'}")
    except Exception as e:
        print(f"[users read error] {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
