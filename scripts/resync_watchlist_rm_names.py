#!/usr/bin/env python3
"""
resync_watchlist_rm_names.py — fix surviving watchlist rows whose rm_name is
stale/invented even though rm_code is a real register code.

The branch remap (remap_watchlist_db.py) rewrote branch fields but left rm_name
as-is, so rows like rm_code=300xxx / rm_name="Elijah Nasambu" (a name not in the
register for that code) still display the wrong owner. This authoritatively
re-derives rm_name from staff_register.xlsx, keyed on rm_code.

REPORT (default): lists how many rows have a name mismatch, with examples.
--apply: backs up the table, then UPDATEs rm_name to the register name for every
row whose rm_code resolves. Rows whose rm_code is NOT in the register are left
(should be none post-remap) and reported.

    python scripts\\resync_watchlist_rm_names.py            # dry-run
    python scripts\\resync_watchlist_rm_names.py --apply    # backup + fix
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA_DIR / "staff_register.xlsx"


def register_names():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = headers.index("Staff Code") if "Staff Code" in headers else 0
    ni = headers.index("Staff Name") if "Staff Name" in headers else 1
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[ci]:
            out[str(row[ci]).strip()] = str(row[ni] or "").strip()
    wb.close()
    return out


def main():
    apply = "--apply" in sys.argv
    from utils.db import db

    reg = register_names()
    print(f"register codes: {len(reg)}")

    rows = db.fetch_all("SELECT id, rm_code, rm_name FROM watchlist", ())
    print(f"watchlist rows: {len(rows)}\n")

    mismatched, unresolved, ok = [], [], 0
    for r in rows:
        code = str(r.get("rm_code") or "").strip()
        cur = str(r.get("rm_name") or "").strip()
        reg_name = reg.get(code)
        if reg_name is None:
            unresolved.append((r["id"], code, cur))
        elif reg_name != cur:
            mismatched.append((r["id"], code, cur, reg_name))
        else:
            ok += 1

    print(f"rm_name already correct:        {ok}")
    print(f"rm_name MISMATCH (will fix):    {len(mismatched)}")
    print(f"rm_code not in register:        {len(unresolved)} (expected 0 post-remap)\n")

    if mismatched:
        print("--- sample mismatches (current -> register) ---")
        seen = Counter()
        for _id, code, cur, regn in mismatched:
            key = (cur, regn)
            if seen[key] == 0:
                print(f"  {code}: '{cur}'  ->  '{regn}'")
            seen[key] += 1
    if unresolved:
        print("\n--- UNRESOLVED rm_codes (not in register) ---")
        for _id, code, cur in unresolved[:10]:
            print(f"  {code}: '{cur}'  (row {_id})")

    if not apply:
        print("\n[DRY-RUN] No DB change. Re-run with --apply to back up + fix names.")
        return

    if not mismatched:
        print("Nothing to fix.")
        return

    full = db.fetch_all("SELECT * FROM watchlist", ())
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = DATA_DIR / f"watchlist_db_table.pre_namesync_{ts}.json"
    backup.write_text(json.dumps(full, default=str, indent=2), encoding="utf-8")
    print(f"\n[backup] {len(full)} rows -> {backup.name}")

    # group ids by target name for batched updates
    by_name = {}
    for _id, code, cur, regn in mismatched:
        by_name.setdefault(regn, []).append(_id)
    with db.transaction() as conn:
        for regn, ids in by_name.items():
            for i in range(0, len(ids), 500):
                chunk = ids[i:i + 500]
                ph = ",".join(["%s"] * len(chunk))
                db.execute(f"UPDATE watchlist SET rm_name=%s WHERE id IN ({ph})",
                           (regn, *chunk), conn=conn)
    print(f"[apply] resynced rm_name on {len(mismatched)} rows.")
    print("Restart API + run harness — credit drill RM names now match the register.")


if __name__ == "__main__":
    main()
