#!/usr/bin/env python3
"""
cbs_diag.py — READ-ONLY. Inventory the credit accounts store before the stable
branch/owner remap. Writes nothing.

The credit analytics read path (api.py _acquire_credit_accounts) resolves, in
order: credit_watchlist (PG) -> watchlist (PG) -> credit_monitoring.json (disk).
This probe reports which store is live, its row count, and the current
distribution of branch_name / branch_code / region / rm_code / rm_name — so the
remap targets real values, not assumed ones.

    python scripts\\cbs_diag.py
"""
import sys
import json
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"

FIELDS = ("branch_name", "branch_code", "region", "rm_code", "rm_name")


def _rows_from_db():
    try:
        from utils.db import db
    except Exception as e:
        print(f"[db import error] {e}")
        return None, None
    for table, sql in (
        ("credit_watchlist",
         "SELECT account_number, cif, branch_code, branch_name, region, "
         "rm_code, rm_name, risk_data FROM credit_watchlist"),
        ("watchlist", "SELECT * FROM watchlist"),
    ):
        try:
            rows = db.fetch_all(sql, ())
        except Exception as e:
            print(f"[{table}] not readable: {type(e).__name__}: {e}")
            continue
        if rows:
            out = []
            for r in rows:
                a = dict(r)
                rd = a.get("risk_data")
                if isinstance(rd, str):
                    try:
                        rd = json.loads(rd)
                    except Exception:
                        rd = {}
                if isinstance(rd, dict):
                    for k in FIELDS + ("account_number", "cif"):
                        if a.get(k) is None and rd.get(k) is not None:
                            a[k] = rd.get(k)
                out.append(a)
            return table, out
    return None, None


def _rows_from_disk():
    p = DATA_DIR / "credit_monitoring.json"
    if not p.exists():
        return None, None
    raw = json.loads(p.read_text(encoding="utf-8"))
    rows = raw if isinstance(raw, list) else raw.get("watchlist", [])
    return "credit_monitoring.json", rows


def dist(rows, field, top=25):
    c = Counter(str(r.get(field) or "∅") for r in rows)
    return c.most_common(top), len(c)


def main():
    src, rows = _rows_from_db()
    if not rows:
        src, rows = _rows_from_disk()
    if not rows:
        print("No credit accounts found in any store (DB or disk).")
        return

    print(f"LIVE SOURCE: {src}")
    print(f"ROWS: {len(rows)}\n")

    for f in ("branch_name", "branch_code", "region"):
        items, n = dist(rows, f)
        print(f"--- {f}: {n} distinct ---")
        for v, c in items:
            print(f"  {c:>6}  {v}")
        print()

    # RM coverage — how many rm_codes, and do they look like real staff codes?
    rm_items, rm_n = dist(rows, "rm_code", top=10)
    print(f"--- rm_code: {rm_n} distinct (top 10 shown) ---")
    for v, c in rm_items:
        print(f"  {c:>6}  {v}")
    name_items, name_n = dist(rows, "rm_name", top=10)
    print(f"\n--- rm_name: {name_n} distinct (top 10 shown) ---")
    for v, c in name_items:
        print(f"  {c:>6}  {v}")

    # how many rm_codes resolve to the LIVE register (300xxx) vs invented?
    live_codes = set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_DIR / "staff_register.xlsx", read_only=True)
        ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                live_codes.add(str(row[0]).strip())
        wb.close()
    except Exception as e:
        print(f"\n[register read error] {e}")
    if live_codes:
        rm_codes_in_data = {str(r.get("rm_code") or "") for r in rows if r.get("rm_code")}
        resolves = rm_codes_in_data & live_codes
        print(f"\n--- RM ROSTER RESOLUTION ---")
        print(f"  distinct rm_codes in credit data: {len(rm_codes_in_data)}")
        print(f"  live register codes:              {len(live_codes)}")
        print(f"  resolve to live register:         {len(resolves)} "
              f"({100*len(resolves)//max(1,len(rm_codes_in_data))}%)")
        print(f"  NOT in register (invented/stale): {len(rm_codes_in_data - live_codes)}")


if __name__ == "__main__":
    main()
