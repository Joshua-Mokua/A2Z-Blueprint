"""scripts/seed_branch_test_logins.py — real-staff test logins for ONE branch
chain (frontline -> CEO), drawn from data/staff_register.xlsx.

Why: william001 (staff_code 0001) is a SYNTHETIC test account not present in the
register hierarchy, which is why scope needed band-aids. These logins use the
REAL register staff (300xxx codes), so cascade scope resolves correctly:
- The register root (Chief Executive & Managing Director) is all-view (B1).
- Mid-level roles scope by the reporting tree (full per-level cascade lands with
  the mid-level scope rebuild; until then mid-level roles see self-only).

Password convention: EcoStaff + last-4 of staff_code (matches utils/test_logins).
Username: lowercase first name + last-4 of staff_code (e.g. william0001).

Safe: backs up users.json (timestamped) before writing; idempotent (re-runnable);
aborts if users.json is missing/empty so it can't trigger a defaults overwrite.

USAGE (project root, venv active):
    python scripts/seed_branch_test_logins.py --branch Thika
    python scripts/seed_branch_test_logins.py --branch Thika --list
"""
from __future__ import annotations
import argparse, json, shutil
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
REGISTER = DATA / "staff_register.xlsx"
USERS = DATA / "users.json"


def _load_register():
    from openpyxl import load_workbook
    wb = load_workbook(REGISTER, read_only=True); ws = wb.active
    it = ws.iter_rows(values_only=True); hdr = list(next(it))
    rows = [dict(zip(hdr, r)) for r in it]
    for r in rows:
        for k in ("Staff Code", "Staff Name", "Role", "Unit", "Region", "Reports To"):
            r[k] = "" if r.get(k) is None else str(r[k]).strip()
    return rows


def _chain_for_branch(rows, branch):
    branch_staff = [r for r in rows if r["Unit"] == branch]
    if not branch_staff:
        raise SystemExit(f"No staff found for branch Unit '{branch}'. "
                         f"Sample units: {sorted({r['Unit'] for r in rows})[:10]}")

    def find_mgr(child):
        cand = [r for r in rows if r["Role"] == child["Reports To"]]
        if not cand:
            return None
        same = [c for c in cand if c["Region"] == child["Region"]]
        return (same or cand)[0]

    # representative frontline + every branch level, deduped by staff code
    picked, seen = [], set()
    def add(r):
        if r and r["Staff Code"] not in seen:
            seen.add(r["Staff Code"]); picked.append(r)

    # one of each branch role (frontline first)
    want_roles = ["Teller", "Customer Service Officer",
                  "Relationship Officer-Personal Banker",
                  "Branch Operations Supervisor", "Branch Relationship Manager",
                  "Branch Operations Manager", "Senior Branch Manager", "Branch Manager"]
    for role in want_roles:
        for r in branch_staff:
            if r["Role"] == role:
                add(r); break

    # walk upward from the branch's top manager to the CEO (root)
    top = next((r for r in branch_staff
                if r["Role"] in ("Senior Branch Manager", "Branch Manager")), None)
    node = top or (branch_staff[0] if branch_staff else None)
    guard = 0
    while node and guard < 20:
        guard += 1
        add(node)
        if node["Reports To"] in ("", "None", "nan"):
            break
        node = find_mgr(node)
    return picked


def _username(r):
    first = r["Staff Name"].split()[0].lower() if r["Staff Name"] else "staff"
    return f"{first}{r['Staff Code'][-4:]}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--branch", default="Thika")
    ap.add_argument("--list", action="store_true", help="just print, don't write")
    args = ap.parse_args()

    rows = _load_register()
    chain = _chain_for_branch(rows, args.branch)

    creds = []
    new_users = {}
    for r in chain:
        code = r["Staff Code"]; uname = _username(r)
        pw = "EcoStaff" + code[-4:]
        is_root = r["Reports To"] in ("", "None", "nan")
        new_users[uname] = {
            "username": uname, "password": pw,
            "full_name": r["Staff Name"], "role": r["Role"],
            "staff_code": code, "unit": r["Unit"], "region": r["Region"],
            "active": True, "must_change_password": False,
            "_protected": True,
            # scope is role-based; root (CEO) is all-view via B1. Others scope
            # by the reporting tree. can_view_all left False so the test
            # reflects real cascade scope, not a blanket override.
            "can_view_all": bool(is_root),
        }
        creds.append((uname, pw, r["Role"], code, r["Staff Name"]))

    if args.list:
        print(f"Branch chain '{args.branch}' ({len(creds)} logins):")
        for u, p, role, code, name in creds:
            print(f"  {u:18s} {p:14s} {code}  {role:34s} {name}")
        return

    if not USERS.exists() or USERS.stat().st_size == 0:
        raise SystemExit("users.json missing/empty — run scripts/seed_test_logins.py "
                         "first (won't overwrite a missing file).")

    backup = USERS.with_suffix(f".json.bak_{datetime.now():%Y%m%d_%H%M%S}")
    shutil.copy2(USERS, backup)
    data = json.loads(USERS.read_text(encoding="utf-8"))
    users = data if isinstance(data, dict) else {}
    users.update(new_users)        # idempotent: re-running just refreshes
    USERS.write_text(json.dumps(users, indent=2), encoding="utf-8")

    doc = ROOT / "BRANCH_TEST_LOGINS.md"
    lines = [f"# Branch test logins — {args.branch} chain (frontline → CEO)\n",
             f"_Generated {datetime.now():%Y-%m-%d %H:%M}. Password = EcoStaff + last-4 of staff code._\n",
             "| Username | Password | Staff code | Role | Name |", "|---|---|---|---|---|"]
    for u, p, role, code, name in creds:
        lines.append(f"| `{u}` | `{p}` | {code} | {role} | {name} |")
    doc.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Backed up users.json -> {backup.name}")
    print(f"Seeded {len(new_users)} branch logins into users.json")
    print(f"Wrote {doc.name}")
    for u, p, role, *_ in creds:
        print(f"  {u:18s} {p:14s} {role}")


if __name__ == "__main__":
    main()
