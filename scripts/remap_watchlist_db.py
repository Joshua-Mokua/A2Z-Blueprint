#!/usr/bin/env python3
"""
remap_watchlist_db.py — stable CBS credit-book remap, operating DIRECTLY on the
Postgres `watchlist` table (the live store the credit analytics reads).

The JSON (credit_monitoring.json) is empty on this install (a reset script wiped
it), so the 5,001 rows live only in the DB. This script therefore mutates the
table in place:

  DELETE rows at the 18 old branches with no 16-branch survivor.
  DELETE rows owned by phantom rm_codes absent from staff_register.xlsx.
  UPDATE survivors: branch_name -> canonical 16 name, branch_code -> P-code,
                    region -> 11-region taxonomy.

Financials + identity (account_number, cif, metadata, outstanding, collateral,
dpd, dates) are never touched.

SAFE: dry-run unless --apply. Before any mutation it dumps the whole table to a
timestamped JSON backup. All deletes/updates run inside one transaction.

    python scripts\\remap_watchlist_db.py            # dry-run report
    python scripts\\remap_watchlist_db.py --apply    # backup + mutate
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA_DIR / "staff_register.xlsx"

SURVIVE = {
    "Towers": ("P01", "Nairobi CBD"), "Plaza": ("P03", "Nairobi CBD"),
    "Industrial Area": ("P11", "Nairobi CBD"), "Westlands": ("P13", "Nairobi Metro"),
    "Upper Hill": ("P22", "Nairobi CBD"), "Valley Arcade": ("P23", "Nairobi Metro"),
    "Karen": ("P24", "Nairobi Metro"), "Fortis Office Park": ("P30", "Nairobi Metro"),
    "Mombasa Moi Avenue": ("P02", "Coast"), "Thika": ("P06", "Mt Kenya West"),
    "Eldoret": ("P07", "North Rift"), "Kisumu": ("P08", "West Kenya"),
    "Kisii": ("P09", "South Rift"), "Karatina": ("P12", "Mt Kenya East"),
    "Nakuru": ("P15", "North Rift"), "Nyeri": ("P17", "Mt Kenya East"),
}
ALIASES = {"Sarit Centre": "Westlands", "Gigiri": "Westlands",
           "Thika Road Mall": "Thika", "Mombasa Road": None}


def target_branch(old_name: str):
    s = (old_name or "").replace(" Branch", "").replace(" Main", "") \
        .replace(" Mega", "").replace(" West", "").strip()
    if s in SURVIVE:
        return s
    if s in ALIASES:
        return ALIASES[s]
    return None


def load_register_codes():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            codes.add(str(row[0]).strip())
    wb.close()
    return codes


def main():
    apply = "--apply" in sys.argv
    from utils.db import db

    rows = db.fetch_all(
        "SELECT id, branch_name, rm_code, rm_name FROM watchlist", ())
    print(f"watchlist rows: {len(rows)}")
    live_codes = load_register_codes()
    print(f"live register codes: {len(live_codes)}\n")

    del_branch_ids, del_phantom_ids = [], []
    update_plan = {}   # id -> (new_name, pcode, region)
    branch_after = Counter()
    for r in rows:
        tgt = target_branch(r.get("branch_name"))
        if tgt is None:
            del_branch_ids.append(r["id"])
            continue
        if str(r.get("rm_code") or "") not in live_codes:
            del_phantom_ids.append(r["id"])
            continue
        p, reg = SURVIVE[tgt]
        update_plan[r["id"]] = (tgt, p, reg)
        branch_after[tgt] += 1

    print(f"DELETE (dead branch):   {len(del_branch_ids)}")
    print(f"DELETE (phantom RM):    {len(del_phantom_ids)}")
    print(f"UPDATE (remap survivor):{len(update_plan)}")
    print(f"final row count:        {len(update_plan)}\n")
    print("--- post-remap branch distribution ---")
    for name, cnt in branch_after.most_common():
        p, reg = SURVIVE[name]
        print(f"  {cnt:>5}  {name} ({p}, {reg})")

    if not apply:
        print("\n[DRY-RUN] No DB change. Re-run with --apply to back up + mutate.")
        return

    # ---- backup whole table first ----
    full = db.fetch_all("SELECT * FROM watchlist", ())
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = DATA_DIR / f"watchlist_db_table.pre_remap_{ts}.json"
    backup.write_text(json.dumps(full, default=str, indent=2), encoding="utf-8")
    print(f"\n[backup] {len(full)} rows -> {backup.name}")

    # ---- mutate in one transaction ----
    with db.transaction() as conn:
        # deletes
        for ids in (del_branch_ids, del_phantom_ids):
            for i in range(0, len(ids), 500):
                chunk = ids[i:i + 500]
                ph = ",".join(["%s"] * len(chunk))
                db.execute(f"DELETE FROM watchlist WHERE id IN ({ph})",
                           tuple(chunk), conn=conn)
        # updates grouped by target for fewer statements
        by_target = {}
        for rid, (name, p, reg) in update_plan.items():
            by_target.setdefault((name, p, reg), []).append(rid)
        for (name, p, reg), ids in by_target.items():
            for i in range(0, len(ids), 500):
                chunk = ids[i:i + 500]
                ph = ",".join(["%s"] * len(chunk))
                db.execute(
                    f"UPDATE watchlist SET branch_name=%s, branch_code=%s, "
                    f"region=%s WHERE id IN ({ph})",
                    (name, p, reg, *chunk), conn=conn)

    after = db.fetch_one("SELECT COUNT(*) AS n FROM watchlist", ())
    print(f"[apply] watchlist now: {after['n']} rows "
          f"(was {len(rows)}; deleted {len(del_branch_ids)+len(del_phantom_ids)}).")
    print("\nRestart API, then run the harness — credit drill should show 16 branches + real RMs.")


if __name__ == "__main__":
    main()
