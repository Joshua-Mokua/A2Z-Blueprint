#!/usr/bin/env python3
"""
director_rename_diag.py — READ-ONLY. Complete inventory of the two director role
strings to be renamed, across every store. Writes nothing.

  'Director Retail Banking'     -> 'Director Consumer & Commercial Banking (CCB)'
  'Director Commercial Banking' -> 'Director Corporate & Investment Banking (CIB)'

Also flags the bare 'Director Retail' (no 'Banking') orphan so it isn't missed.

    python scripts\\director_rename_diag.py
"""
import sys
import json
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

OLD_RETAIL = "Director Retail Banking"
OLD_COMM = "Director Commercial Banking"
BARE_RETAIL = re.compile(r"Director Retail(?! Banking)")  # bare, not followed by Banking


def scan_text(path: Path):
    try:
        txt = path.read_text(encoding="utf-8")
    except Exception:
        return None
    return (txt.count(OLD_RETAIL), txt.count(OLD_COMM),
            len(BARE_RETAIL.findall(txt)))


def main():
    print("=== DATA FILES (tracked config + runtime) ===")
    data_hits = {}
    for p in sorted(DATA.glob("*.json")):
        r = scan_text(p)
        if r and (r[0] or r[1] or r[2]):
            data_hits[p.name] = r
            print(f"  {p.name:<34} retail={r[0]:<3} commercial={r[1]:<3} bare_retail={r[2]}")
    if not data_hits:
        print("  (none)")

    # users.json — count by record role
    print("\n=== users.json (role field per record) ===")
    uj = DATA / "users.json"
    if uj.exists():
        try:
            users = json.loads(uj.read_text(encoding="utf-8"))
            rcnt = ccnt = 0
            for u, rec in users.items():
                role = str(rec.get("role", ""))
                if role == OLD_RETAIL:
                    rcnt += 1
                    print(f"  {u}: role='{role}'")
                elif role == OLD_COMM:
                    ccnt += 1
                    print(f"  {u}: role='{role}'")
            print(f"  -> {rcnt} retail-director logins, {ccnt} commercial-director logins")
        except Exception as e:
            print(f"  users.json parse error: {e}")

    # staff_register.xlsx — count by Role
    print("\n=== staff_register.xlsx (Role column) ===")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA / "staff_register.xlsx", read_only=True)
        ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ri = headers.index("Role") if "Role" in headers else 2
        ci = headers.index("Staff Code") if "Staff Code" in headers else 0
        rcnt = ccnt = bare = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            role = str(row[ri] or "")
            if role == OLD_RETAIL:
                rcnt += 1
                print(f"  {row[ci]}: '{role}'")
            elif role == OLD_COMM:
                ccnt += 1
                print(f"  {row[ci]}: '{role}'")
            elif BARE_RETAIL.search(role):
                bare += 1
                print(f"  {row[ci]}: BARE '{role}'")
        wb.close()
        print(f"  -> retail={rcnt} commercial={ccnt} bare_retail={bare}")
    except Exception as e:
        print(f"  register read error: {e}")

    # code files (report only — Josh hand-edits these)
    print("\n=== CODE FILES (hand-edit separately) ===")
    code_total_r = code_total_c = code_total_bare = 0
    for sub in ("utils", "pages"):
        d = ROOT / sub
        if not d.exists():
            continue
        for p in d.rglob("*.py"):
            if "pycache" in str(p):
                continue
            r = scan_text(p)
            if r and (r[0] or r[1] or r[2]):
                code_total_r += r[0]
                code_total_c += r[1]
                code_total_bare += r[2]
                print(f"  {p.relative_to(ROOT)}: retail={r[0]} commercial={r[1]} bare={r[2]}")
    print(f"  -> CODE TOTALS: retail={code_total_r} commercial={code_total_c} "
          f"bare_retail={code_total_bare}")

    print("\nData/register/users renames -> scripted. Code renames -> hand-edited.")


if __name__ == "__main__":
    main()
