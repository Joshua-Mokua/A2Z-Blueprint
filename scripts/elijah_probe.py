#!/usr/bin/env python3
"""elijah_probe.py — READ-ONLY. Find every watchlist row with rm_name like
'Elijah', show its rm_code, and what the register says that code's name is.
Resolves why the credit drill still surfaces 'Elijah Nasambu'. Writes nothing.

    python scripts\\elijah_probe.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA_DIR / "staff_register.xlsx"


def register_lookup():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = headers.index("Staff Code") if "Staff Code" in headers else 0
    ni = headers.index("Staff Name") if "Staff Name" in headers else 1
    ri = headers.index("Role") if "Role" in headers else 2
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[ci]:
            out[str(row[ci]).strip()] = (str(row[ni] or "").strip(), str(row[ri] or "").strip())
    wb.close()
    return out


def main():
    from utils.db import db
    reg = register_lookup()

    rows = db.fetch_all(
        "SELECT rm_code, rm_name, COUNT(*) AS n, SUM(outstanding) AS tot "
        "FROM watchlist WHERE rm_name ILIKE %s GROUP BY rm_code, rm_name",
        ("%Elijah%",))
    if not rows:
        print("No watchlist rows with rm_name like 'Elijah'. "
              "It may be coming from a non-watchlist source.")
    else:
        print("watchlist rows with 'Elijah':")
        for r in rows:
            code = str(r.get("rm_code") or "")
            regname, regrole = reg.get(code, ("(not in register)", ""))
            print(f"  rm_code={code} rm_name='{r.get('rm_name')}' n={r.get('n')} "
                  f"out={r.get('tot')}  | register: '{regname}' ({regrole})")

    # also: is 'Elijah Nasambu' a register name at all?
    hits = [(c, n) for c, (n, _) in reg.items() if "elijah" in n.lower()]
    print(f"\nregister entries containing 'Elijah': {hits or 'NONE'}")

    # top RM by outstanding in the whole watchlist (what the drill would surface)
    top = db.fetch_all(
        "SELECT rm_code, rm_name, SUM(outstanding) AS tot, COUNT(*) AS n "
        "FROM watchlist GROUP BY rm_code, rm_name ORDER BY tot DESC LIMIT 8", ())
    print("\ntop RMs by outstanding (drill picks brm[0]):")
    for r in top:
        code = str(r.get("rm_code") or "")
        regname = reg.get(code, ("(not in register)", ""))[0]
        flag = "  <-- MISMATCH" if regname != str(r.get("rm_name") or "") else ""
        print(f"  {code}  '{r.get('rm_name')}'  out={r.get('tot')}  reg='{regname}'{flag}")


if __name__ == "__main__":
    main()
