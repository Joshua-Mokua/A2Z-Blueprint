#!/usr/bin/env python3
"""
remap_deals.py — B-deals: re-home pipeline deals onto the surviving 16-branch RMs.

After B1 dropped ~1,033 staff, deals owned by them orphaned ("Unassigned"). This
reassigns each orphaned deal to a surviving branch RM (round-robin across the 16,
balanced), refreshes the deal's unit/staff_name from the live roster, and carries
the portfolio owner along. Deals owned by surviving staff keep their owner (unit
refreshed). Admin-owned + the harness personas (300731 Frank, 300716 Immaculate)
are NEVER re-home recipients, so their deal counts stay stable for the harness.

Backup-first, dry-run aware.
    python scripts\\remap_deals.py --dry-run
    python scripts\\remap_deals.py
"""
import argparse, datetime, json, itertools
from pathlib import Path
from collections import Counter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
PERSONAS = {"300731", "300716", "300001"}   # never re-home recipients
RM_ROLES = {
    "Relationship Officer-Personal Banker",
    "Relationship Officer-Business Banker",
    "Branch Relationship Manager",
    "Branch Senior Relationship Officer",
    "Relationship Officer Bancassurance",
}


def load_roster():
    import openpyxl
    wb = openpyxl.load_workbook(str(DATA_DIR / "staff_register.xlsx"), read_only=True)
    ws = wb.active
    H = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = [dict(zip(H, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    wb.close()
    by_code = {str(r["Staff Code"]): r for r in rows}
    # recipients: branch RMs, excluding personas, sorted to round-robin by branch
    recips = [r for r in rows
              if str(r.get("Category")) == "Branch"
              and r.get("Role") in RM_ROLES
              and str(r["Staff Code"]) not in PERSONAS]
    recips.sort(key=lambda r: (str(r.get("Unit")), str(r["Staff Code"])))
    # interleave across branches so re-homed deals spread evenly
    by_branch = {}
    for r in recips:
        by_branch.setdefault(str(r.get("Unit")), []).append(r)
    ring = []
    for tup in itertools.zip_longest(*by_branch.values()):
        ring.extend(x for x in tup if x is not None)
    return by_code, ring


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    by_code, ring = load_roster()
    survivors = set(by_code)
    print(f"surviving roster: {len(survivors)} | re-home recipients (branch RMs): {len(ring)}")
    if not ring:
        print("[abort] no recipient RMs found — run remap_staff first."); return

    pf = DATA_DIR / "pipeline_deals.json"
    raw = json.loads(pf.read_text(encoding="utf-8"))
    deals = raw.get("deals", raw) if isinstance(raw, dict) else raw

    rr = itertools.cycle(ring)
    kept, refreshed, rehomed = 0, 0, 0
    rehome_to = Counter()
    for d in deals:
        sc = str(d.get("staff_code") or "")
        if sc.upper().startswith("ADMIN"):
            kept += 1
            continue
        if sc in survivors:
            # owner survives — refresh unit + name to current branch naming
            r = by_code[sc]
            d["unit"] = r.get("Unit")
            d["staff_name"] = r.get("Staff Name")
            # carry portfolio owner if it also survives, else point at the RM
            poc = str(d.get("portfolio_owner_code") or "")
            if poc and poc not in survivors and not poc.upper().startswith("ADMIN"):
                d["portfolio_owner_code"] = sc
                d["portfolio_owner_name"] = r.get("Staff Name")
            refreshed += 1
        else:
            # orphaned — reassign to next recipient RM
            r = next(rr)
            nc = str(r["Staff Code"])
            d["staff_code"] = nc
            d["staff_name"] = r.get("Staff Name")
            d["unit"] = r.get("Unit")
            d["portfolio_owner_code"] = nc
            d["portfolio_owner_name"] = r.get("Staff Name")
            rehomed += 1
            rehome_to[r.get("Unit")] += 1

    print(f"deals: {len(deals)}  (admin-kept {kept} | owner-refreshed {refreshed} | "
          f"re-homed {rehomed})")
    if rehome_to:
        print("  re-homed spread across branches:")
        for unit, n in sorted(rehome_to.items()):
            print(f"     {n:>4}  {unit}")
    # post-condition: zero deals point outside the roster (excl admin)
    orphan_left = [str(d.get("staff_code")) for d in deals
                   if not str(d.get("staff_code") or "").upper().startswith("ADMIN")
                   and str(d.get("staff_code")) not in survivors]
    print(f"  orphans remaining (should be 0): {len(orphan_left)}")
    # persona deal counts (should be unchanged — never recipients)
    for p in ("300731", "300716"):
        print(f"  deals owned by {p}: {sum(1 for d in deals if str(d.get('staff_code'))==p)}")

    if args.dry_run:
        print("\n[dry-run] nothing written.")
        return

    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    bak = pf.with_suffix(f".json.pre_branch16_{stamp}")
    bak.write_text(pf.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"[backup] {bak.name}")
    if isinstance(raw, dict):
        raw["deals"] = deals
        pf.write_text(json.dumps(raw, indent=2), encoding="utf-8")
    else:
        pf.write_text(json.dumps(deals, indent=2), encoding="utf-8")
    print(f"[ok] pipeline_deals.json re-homed ({len(deals)} deals).")


if __name__ == "__main__":
    main()
