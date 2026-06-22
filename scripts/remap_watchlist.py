#!/usr/bin/env python3
"""
remap_watchlist.py — stable CBS credit-book remap onto the 16 branches + live roster.

Operates on data/credit_monitoring.json (the canonical watchlist the migrator
loads into Postgres). Does NOT mint new customers/accounts — preserves every
account_number, cif, and all financial/risk fields. Only branch identity and
ownership are rewritten; non-conforming rows are deleted (simulated data, per
Joshua's call).

Two deletes:
  1. Rows at the 18 old branches with no 16-branch survivor.
  2. Rows owned by phantom rm_codes absent from the live staff_register.xlsx
     (incl. those at surviving branches — "fully real book").

Survivors get: branch_name -> canonical 16 name, branch_code -> P-code,
region -> 11-region taxonomy (derived from the new branch, auto-consistent).

SAFE: dry-run unless --apply. Backs up the JSON (.pre_remap_<ts>) before writing.
After --apply, reload Postgres from the JSON via the migrator (separate step).

    python scripts\\remap_watchlist.py            # dry-run report
    python scripts\\remap_watchlist.py --apply    # backup + rewrite JSON
"""
import sys
import json
import random
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
WL_FILE = DATA_DIR / "credit_monitoring.json"
REGISTER = DATA_DIR / "staff_register.xlsx"

random.seed(1719)  # deterministic owner reassignment

# Canonical 16 survivors: name -> (P-code, 11-region)
SURVIVE = {
    "Towers": ("P01", "Nairobi CBD"), "Plaza": ("P03", "Nairobi CBD"),
    "Industrial Area": ("P11", "Nairobi CBD"), "Westlands": ("P13", "Nairobi Metro"),
    "Upper Hill": ("P22", "Nairobi CBD"), "Valley Arcade": ("P23", "Nairobi Metro"),
    "Karen": ("P24", "Nairobi Metro"), "Fortis Office Park": ("P30", "Nairobi Metro"),
    "Mombasa Moi Avenue": ("P02", "Coast"), "Thika": ("P06", "Mt Kenya West"),
    "Eldoret": ("P07", "North Rift"), "Kisumu": ("P08", "West Kenya"),
    "Kisii": ("P09", "South Rift"), "Karatina": ("P12", "Mt Kenya East"),
    "Nakuru": ("P15", "North Rift"), "Nyeri": ("P17", "Mt Kenya East"),
}

# Explicit aliases (confirmed by Joshua). None => DELETE.
ALIASES = {
    "Sarit Centre": "Westlands", "Gigiri": "Westlands",
    "Thika Road Mall": "Thika", "Mombasa Road": None,
}


def target_branch(old_name: str):
    """Return canonical survivor name, or None if the row should be deleted."""
    s = (old_name or "").replace(" Branch", "").replace(" Main", "") \
        .replace(" Mega", "").replace(" West", "").strip()
    if s in SURVIVE:
        return s
    if s in ALIASES:
        return ALIASES[s]
    return None


def load_register_codes_and_units():
    """staff_code -> Unit (branch) for live RMs, plus the full code set."""
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    code_i = idx.get("Staff Code", 0)
    unit_i = idx.get("Unit", 3)
    role_i = idx.get("Role", 2)
    name_i = idx.get("Staff Name", 1)
    codes, code_unit, code_name, unit_rms = set(), {}, {}, {}
    RM_HINT = ("relationship", "officer", "rm ", "sales", "personal banker", "dso", "ro ")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[code_i]:
            continue
        code = str(row[code_i]).strip()
        unit = str(row[unit_i] or "").strip()
        role = str(row[role_i] or "").strip()
        name = str(row[name_i] or "").strip()
        codes.add(code)
        code_unit[code] = unit
        code_name[code] = name
        if any(h in role.lower() for h in RM_HINT):
            unit_rms.setdefault(unit, []).append((code, name))
    wb.close()
    return codes, code_unit, code_name, unit_rms


def main():
    apply = "--apply" in sys.argv
    raw = json.loads(WL_FILE.read_text(encoding="utf-8"))
    is_wrapped = isinstance(raw, dict) and "watchlist" in raw
    wl = raw["watchlist"] if is_wrapped else raw
    print(f"watchlist rows: {len(wl)}")

    live_codes, code_unit, code_name, unit_rms = load_register_codes_and_units()
    print(f"live register codes: {len(live_codes)} | branches w/ RMs: {len(unit_rms)}\n")

    # Pass 1: classify each row
    kept, del_branch, del_phantom = [], 0, 0
    branch_after = Counter()
    for r in wl:
        tgt = target_branch(r.get("branch_name"))
        if tgt is None:
            del_branch += 1
            continue
        if str(r.get("rm_code") or "") not in live_codes:
            del_phantom += 1
            continue
        # survivor — rewrite branch identity
        pcode, region = SURVIVE[tgt]
        r2 = dict(r)
        r2["branch_name"] = tgt
        r2["branch_code"] = pcode
        r2["region"] = region
        kept.append(r2)
        branch_after[tgt] += 1

    print(f"DELETE (dead branch):   {del_branch}")
    print(f"DELETE (phantom RM):    {del_phantom}")
    print(f"KEEP (remapped):        {len(kept)}\n")
    print("--- post-remap branch distribution (16) ---")
    for name, cnt in branch_after.most_common():
        p, reg = SURVIVE[name]
        print(f"  {cnt:>5}  {name} ({p}, {reg})")

    # sanity: every kept row's RM is real, branch is one of 16
    bad = [r for r in kept if r["branch_name"] not in SURVIVE
           or str(r.get("rm_code")) not in live_codes]
    print(f"\nintegrity check — non-conforming survivors: {len(bad)} (must be 0)")

    if not apply:
        print("\n[DRY-RUN] No file written. Re-run with --apply to back up + rewrite.")
        return

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = WL_FILE.with_name(f"credit_monitoring.json.pre_remap_{ts}")
    backup.write_text(WL_FILE.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"\n[backup] {backup.name}")

    if is_wrapped:
        raw["watchlist"] = kept
        raw["last_updated"] = ts
        out = raw
    else:
        out = kept
    WL_FILE.write_text(json.dumps(out, indent=2), encoding="utf-8")
    # validate round-trip
    chk = json.loads(WL_FILE.read_text(encoding="utf-8"))
    chk_n = len(chk["watchlist"] if is_wrapped else chk)
    print(f"[apply] wrote {chk_n} rows to {WL_FILE.name} (was {len(wl)}).")
    print("\nNEXT: reload Postgres from the JSON, then restart API + run harness.")


if __name__ == "__main__":
    main()
