#!/usr/bin/env python3
"""
build_dsa_tree.py — B2a: build the Direct Sales Agent (DSA) matrix tree in the
staff register. ROSTER ONLY — mints NO logins (those fold into the auth finale).

Structure built (per Josh sign-off):
  DSA Head (1)                         reports to: Director Consumer & Commercial Banking (CCB)
    Regional DSA Head (4)              reports to: DSA Head
      Branch DSA Team Lead (16, 1/br)  reports to: Regional DSA Head
        Direct Sales Agent (~30/br)    solid-line: Branch Manager (branch P&L scope)
                                       dotted-line: Branch DSA Team Lead (sales rollup)

Treatment of existing Direct-Sales staff in the LIVE register:
  * 13 'Direct Sales Representative - Assets & Liabilities' -> renamed 'Direct Sales Agent'
  * 1  'Senior Manager Direct Sales Force'                  -> promoted 'DSA Head'
  * any DSR on 'Head Office' (no sales structure)           -> moved to Upper Hill
  * top up each branch to TARGET_PER_BRANCH agents; add the 16 leads + 4 regionals

Canonical 16 branches + 4 regional clusters are defined below (from set_branches.py).

SAFE: dry-run unless --apply. Backs up register (.pre_dsa_<ts>) before writing.
Roster-only: users.json is NOT touched.

    python scripts\\build_dsa_tree.py            # dry-run (prints full plan)
    python scripts\\build_dsa_tree.py --apply    # backup + write register
"""
import sys
import random
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

DATA = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA / "staff_register.xlsx"

TARGET_PER_BRANCH = 30
DSR_OLD = "Direct Sales Representative - Assets & Liabilities"
SR_MGR_OLD = "Senior Manager Direct Sales Force"
ROLE_AGENT = "Direct Sales Agent"
ROLE_LEAD = "Branch DSA Team Lead"
ROLE_REGIONAL = "Regional DSA Head"
ROLE_HEAD = "DSA Head"
DIRECTOR_CCB = "Director Consumer & Commercial Banking (CCB)"

# Canonical 16 branches -> region (from scripts/set_branches.py SPEC).
BRANCH_REGION = {
    "Towers": "Nairobi CBD", "Plaza": "Nairobi CBD", "Industrial Area": "Nairobi CBD",
    "Westlands": "Nairobi Metro", "Upper Hill": "Nairobi CBD", "Valley Arcade": "Nairobi Metro",
    "Karen": "Nairobi Metro", "Fortis Office Park": "Nairobi Metro",
    "Mombasa Moi Avenue": "Coast", "Thika": "Mt Kenya West", "Eldoret": "North Rift",
    "Kisumu": "West Kenya", "Kisii": "South Rift", "Karatina": "Mt Kenya East",
    "Nakuru": "North Rift", "Nyeri": "Mt Kenya East",
}
CANON_BRANCHES = list(BRANCH_REGION.keys())

# 4 Regional DSA clusters (split each area in two).
REGIONAL_CLUSTERS = {
    "Regional DSA Head - Nairobi CBD":   ["Towers", "Plaza", "Industrial Area", "Upper Hill"],
    "Regional DSA Head - Nairobi Metro": ["Westlands", "Valley Arcade", "Karen", "Fortis Office Park"],
    "Regional DSA Head - Mt Kenya & Coast": ["Thika", "Karatina", "Nyeri", "Mombasa Moi Avenue"],
    "Regional DSA Head - Rift & West":   ["Eldoret", "Kisumu", "Kisii", "Nakuru"],
}
BRANCH_TO_CLUSTER = {b: c for c, bs in REGIONAL_CLUSTERS.items() for b in bs}
HO_REASSIGN_TO = "Upper Hill"   # the HO-based DSR moves here

FIRST = ["Aisha","Brian","Cynthia","David","Esther","Felix","Grace","Henry","Irene","James",
         "Joy","Kevin","Lucy","Mark","Nancy","Oscar","Pauline","Quincy","Rose","Samuel",
         "Tabitha","Victor","Winnie","Xavier","Yvonne","Zachary","Beatrice","Collins","Diana","Edwin"]
LAST = ["Achieng","Barasa","Chebet","Diing","Etyang","Furaha","Gathoni","Hassan","Imali","Juma",
        "Kiprono","Lagat","Mwangi","Njoroge","Otieno","Pkemoi","Quaint","Ruto","Simiyu","Tirop",
        "Ueru","Vihiga","Wekesa","Xolani","Yego","Zawadi","Anyango","Bett","Cheruiyot","Dada"]


def load_register():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and any(x is not None for x in r):
            rows.append({headers[i]: r[i] for i in range(len(headers))})
    return wb, ws, headers, rows


def main():
    apply = "--apply" in sys.argv
    random.seed(20260622)
    wb, ws, headers, rows = load_register()

    codes = [int(str(r["Staff Code"])) for r in rows if str(r["Staff Code"]).isdigit()]
    next_code = max(codes) + 1
    print(f"LIVE register: {len(rows)} staff, codes {min(codes)}..{max(codes)}")

    # ---- 1) rename existing DSRs, promote Sr Manager, fix off-canon placement ----
    renamed = promoted = ho_moved = 0
    per_branch = Counter()
    _rr = []  # round-robin pointer across canonical branches for off-canon DSRs
    _rr_i = 0
    for r in rows:
        role = str(r.get("Role") or "")
        if role == DSR_OLD:
            r["Role"] = ROLE_AGENT
            renamed += 1
            unit = str(r.get("Unit") or "")
            if unit == "Head Office" or unit not in BRANCH_REGION:
                # distribute off-canon DSRs round-robin (don't dump all on one branch)
                target = CANON_BRANCHES[_rr_i % len(CANON_BRANCHES)]
                _rr_i += 1
                r["Unit"] = target
                r["Region"] = BRANCH_REGION[target]
                ho_moved += 1
            r["Reports To"] = ROLE_LEAD  # dotted-line to branch team lead
            per_branch[str(r["Unit"])] += 1
        elif role == SR_MGR_OLD:
            r["Role"] = ROLE_HEAD
            r["Reports To"] = DIRECTOR_CCB
            r["Unit"] = "Head Office"
            promoted += 1

    print(f"\nrenamed DSR->Agent: {renamed}   promoted Sr Mgr->DSA Head: {promoted}   "
          f"off-canon DSRs redistributed: {ho_moved}")
    over = [(b, per_branch[b]) for b in per_branch if per_branch[b] > TARGET_PER_BRANCH]
    if over:
        print(f"  NOTE over-target branches (kept as-is, no top-up): {over}")

    # ---- 2) generate Regional Heads (4) + Team Leads (16) ----
    new_rows = []

    def mint(role, unit, region, reports_to, band, dept="Sales"):
        nonlocal next_code
        code = str(next_code); next_code += 1
        nm = f"{random.choice(FIRST)} {random.choice(LAST)}"
        return {"Staff Code": code, "Staff Name": nm, "Role": role, "Unit": unit,
                "Region": region, "Category": "Branch" if role != ROLE_HEAD else "Head Office",
                "Department": dept, "Band": band, "Gender": random.choice(["M", "F"]),
                "Reports To": reports_to,
                "Date of Employment": datetime(2024, 1, 1).strftime("%Y-%m-%d")}

    for cluster in REGIONAL_CLUSTERS:
        lead_branch = REGIONAL_CLUSTERS[cluster][0]
        new_rows.append(mint(ROLE_REGIONAL, lead_branch, BRANCH_REGION[lead_branch],
                             ROLE_HEAD, "M4"))
    for branch in CANON_BRANCHES:
        new_rows.append(mint(ROLE_LEAD, branch, BRANCH_REGION[branch], ROLE_REGIONAL, "M3"))

    # ---- 3) top up Agents to TARGET_PER_BRANCH per branch ----
    gen_per_branch = Counter()
    for branch in CANON_BRANCHES:
        have = per_branch.get(branch, 0)
        need = max(0, TARGET_PER_BRANCH - have)
        for _ in range(need):
            new_rows.append(mint(ROLE_AGENT, branch, BRANCH_REGION[branch], ROLE_LEAD, "M1"))
            gen_per_branch[branch] += 1

    print(f"\nGENERATED: {len(REGIONAL_CLUSTERS)} Regional Heads + {len(CANON_BRANCHES)} Team Leads "
          f"+ {sum(gen_per_branch.values())} Agents")
    print("\nper-branch agent counts (existing + generated = target):")
    for b in CANON_BRANCHES:
        print(f"  {b:<22} existing={per_branch.get(b,0):>2}  +gen={gen_per_branch.get(b,0):>2}  "
              f"= {per_branch.get(b,0)+gen_per_branch.get(b,0):>2}")

    total_dsa = renamed + sum(gen_per_branch.values())
    print(f"\nTOTAL DSA tree: {total_dsa} Agents + 16 Leads + 4 Regionals + 1 Head "
          f"= {total_dsa + 21} sales staff")
    print(f"register: {len(rows)} -> {len(rows) + len(new_rows)} staff "
          f"(roster-only, NO logins minted)")

    if not apply:
        print("\n[DRY-RUN] No file written. Re-run with --apply to back up + write register.")
        return

    # ---- apply: backup + write all rows back ----
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = REGISTER.with_name(f"staff_register.xlsx.pre_dsa_{ts}")
    import shutil
    shutil.copy2(REGISTER, backup)
    print(f"\n[backup] {backup.name}")

    # rewrite sheet: clear data rows, write existing (mutated) + new
    import openpyxl
    wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = "Staff Register"
    ws2.append(headers)
    for r in rows + new_rows:
        ws2.append([r.get(h, "") for h in headers])
    wb2.save(REGISTER)
    print(f"[apply] register written: {len(rows)+len(new_rows)} staff. "
          f"Next: wire role strings into core.py hierarchy.")


if __name__ == "__main__":
    main()
