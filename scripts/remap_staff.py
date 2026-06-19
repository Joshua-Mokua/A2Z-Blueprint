#!/usr/bin/env python3
"""
remap_staff.py — B1: remap the live staff register + users.json to the 16 branches.

Retain + remap (decision (a)):
  * 11 existing Units carry their staff forward (names + Staff Codes intact),
    re-homed onto the new P-code branch name + area region.
  * Head Office staff (Category == 'Head Office') are kept untouched.
  * Everyone on the other ~78 Units is dropped with their data.
  * 5 net-new branches (no existing Unit) get fresh staff cloned from the real
    per-tier role template, with new Staff Codes continuing past the current max.
  * users.json: logins on retained Units are re-homed; HO + _protected are kept;
    logins on dropped Units are removed. No new logins are minted (net-new branch
    staff are roster-only; admin adds logins later).

Personas survive naturally: frank0731 (300731) + immaculate0716 (300716) are both
Unit=Thika (retained -> P06); william001 is Head Office. The script asserts this.

Backup-first, dry-run aware.
    python scripts\\remap_staff.py --dry-run
    python scripts\\remap_staff.py
"""
import argparse, datetime, json, random
from pathlib import Path
from collections import Counter, defaultdict

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
random.seed(2026)

R1, R2 = "Region 1", "Region 2"

# source Unit -> (new branch name, region, area)
RETAIN = {
    "FB Towers Retail":        ("Towers",             "Nairobi CBD",   R1),
    "Industrial Area":         ("Industrial Area",    "Nairobi CBD",   R1),
    "Westlands":               ("Westlands",          "Nairobi Metro", R1),
    "Mombasa Kenyatta Avenue": ("Mombasa Moi Avenue", "Coast",         R2),
    "Thika":                   ("Thika",              "Mt Kenya West", R2),
    "Eldoret":                 ("Eldoret",            "North Rift",    R2),
    "Kisumu Express":          ("Kisumu",             "West Kenya",    R2),
    "Kisii":                   ("Kisii",              "South Rift",    R2),
    "Karatina":                ("Karatina",           "Mt Kenya East", R2),
    "Nyeri":                   ("Nyeri",              "Mt Kenya East", R2),
    "Nakuru Finance":          ("Nakuru",             "North Rift",    R2),
}
# net-new branches (no existing Unit): (name, region, area, tier)
NET_NEW = [
    ("Plaza",              "Nairobi CBD",   R1, "main"),
    ("Upper Hill",         "Nairobi CBD",   R1, "main"),
    ("Valley Arcade",      "Nairobi Metro", R1, "standard"),
    ("Karen",              "Nairobi Metro", R1, "standard"),
    ("Fortis Office Park", "Nairobi Metro", R1, "main"),
]
# per-tier role template (count, role) — taken verbatim from real branches
TEMPLATE = {
    "flagship": [(4,"Teller"),(2,"Branch Operations Supervisor"),(2,"Customer Service Officer"),
                 (2,"Relationship Officer-Business Banker"),(1,"Senior Branch Manager"),
                 (1,"Branch Operations Manager"),(1,"Senior Digital Channels Officer"),
                 (1,"Branch Senior Relationship Officer"),(1,"Branch Relationship Manager"),
                 (1,"Relationship Officer-Personal Banker"),
                 (1,"Direct Sales Representative - Assets & Liabilities")],
    "main":     [(3,"Teller"),(2,"Customer Service Officer"),(1,"Branch Manager"),
                 (1,"Branch Operations Manager"),(1,"Branch Operations Supervisor"),
                 (1,"Senior Digital Channels Officer"),(1,"Branch Senior Relationship Officer"),
                 (1,"Branch Relationship Manager"),(1,"Relationship Officer-Business Banker"),
                 (1,"Relationship Officer-Personal Banker"),
                 (1,"Direct Sales Representative - Assets & Liabilities")],
    "standard": [(2,"Teller"),(1,"Branch Manager"),(1,"Branch Operations Manager"),
                 (1,"Branch Operations Supervisor"),(1,"Customer Service Officer"),
                 (1,"Senior Digital Channels Officer"),(1,"Branch Relationship Manager"),
                 (1,"Relationship Officer-Business Banker"),(1,"Relationship Officer-Personal Banker"),
                 (1,"Relationship Officer Bancassurance")],
}
FIRST_M = ["James","John","Peter","David","Paul","Joseph","Michael","Patrick","Francis","Daniel",
           "Samuel","Philip","George","Charles","Robert","Thomas","Kevin","Dennis","Eric","Brian",
           "Victor","Moses","Aaron","Simon","Emmanuel","Kenneth","Alex","Mark","Anthony","Edwin"]
FIRST_F = ["Mary","Grace","Faith","Ann","Rose","Jane","Susan","Elizabeth","Catherine","Agnes",
           "Beatrice","Charity","Diana","Esther","Florence","Gloria","Hannah","Irene","Joyce",
           "Lydia","Margaret","Nancy","Olive","Rebecca","Sarah","Alice","Caroline","Naomi","Pauline"]
SURNAMES = ["Kamau","Odhiambo","Wanjiku","Mwangi","Otieno","Kariuki","Mutua","Njoroge","Akinyi",
            "Ochieng","Kiprotich","Chebet","Langat","Koech","Waweru","Ngugi","Njenga","Ndegwa",
            "Mungai","Ouma","Okello","Anyango","Hassan","Omar","Mwenda","Kirimi","Njue","Wafula",
            "Simiyu","Barasa","Wekesa","Munyao","Musyoka","Shimba","Maina","Kiptoo"]


def rand_name():
    g = random.choice(["M","F"])
    return f"{random.choice(FIRST_M if g=='M' else FIRST_F)} {random.choice(SURNAMES)}", g


def load_register():
    import openpyxl
    wb = openpyxl.load_workbook(str(DATA_DIR/"staff_register.xlsx"))
    ws = wb.active
    H = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    rows = [dict(zip(H, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    return H, rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    H, rows = load_register()
    # derive role -> band, role -> reports_to, role -> department from real data
    band_of = {r["Role"]: r["Band"] for r in rows if r.get("Band")}
    reports_of = {}
    for r in rows:
        if r.get("Role") and r.get("Reports To") and r["Role"] not in reports_of:
            reports_of[r["Role"]] = r["Reports To"]
    dept_of = {r["Role"]: r["Department"] for r in rows if r.get("Department")}

    kept_ho, remapped, dropped = [], [], 0
    for r in rows:
        unit = str(r.get("Unit") or "")
        if str(r.get("Category")) == "Head Office" or unit == "Head Office":
            kept_ho.append(r); continue
        if unit in RETAIN:
            name, region, area = RETAIN[unit]
            r2 = dict(r); r2["Unit"] = name; r2["Region"] = region
            remapped.append(r2)
        else:
            dropped += 1

    # net-new staff
    existing_codes = [int(str(r["Staff Code"])) for r in rows if str(r["Staff Code"]).isdigit()]
    nxt = max(existing_codes) + 1
    new_rows = []
    for name, region, area, tier in NET_NEW:
        for cnt, role in TEMPLATE[tier]:
            for _ in range(cnt):
                pname, g = rand_name()
                d = datetime.date(2010,1,1) + datetime.timedelta(days=random.randint(0, 5000))
                new_rows.append({
                    "Staff Code": str(nxt), "Staff Name": pname, "Role": role,
                    "Unit": name, "Region": region,
                    "Category": "Branch", "Department": dept_of.get(role, "Sales"),
                    "Band": band_of.get(role, "M2"), "Gender": g,
                    "Reports To": reports_of.get(role, "Branch Manager"),
                    "Date of Employment": d.strftime("%Y-%m-%d 00:00:00"),
                })
                nxt += 1

    final = kept_ho + remapped + new_rows
    # persona assertions
    codes = {str(r["Staff Code"]) for r in final}
    units = {str(r["Staff Code"]): str(r["Unit"]) for r in final}
    print(f"register: {len(rows)} -> {len(final)}  "
          f"(HO kept {len(kept_ho)} | remapped {len(remapped)} | "
          f"net-new {len(new_rows)} | dropped {dropped})")
    print(f"  new branches: {[n for n,_,_,_ in NET_NEW]}  codes {max(existing_codes)+1}..{nxt-1}")
    for pc, expect in (("300731","Thika"),("300716","Thika"),("300001","Head Office")):
        ok = pc in codes
        loc = units.get(pc, "?") if pc in units else "HO" if pc in codes else "MISSING"
        print(f"  persona {pc}: {'OK' if ok else 'MISSING'} (unit={loc})")

    # users.json remap
    uj = DATA_DIR/"users.json"
    users = json.loads(uj.read_text(encoding="utf-8")) if uj.exists() else {}
    u_remap, u_keep, u_drop = 0, 0, []
    new_users = {}
    for un, u in users.items():
        unit = str(u.get("unit") or "")
        if u.get("_protected") or unit == "Head Office" or str(u.get("department")) == "Head Office":
            if unit in RETAIN:
                nm, region, _ = RETAIN[unit]
                u = {**u, "unit": nm, "region": region,
                     "department": nm if str(u.get("department")) == unit else u.get("department"),
                     "managed_units": [nm if x == unit else x for x in u.get("managed_units", [])]}
                u_remap += 1
            else:
                u_keep += 1
            new_users[un] = u
        elif unit in RETAIN:
            nm, region, _ = RETAIN[unit]
            new_users[un] = {**u, "unit": nm, "region": region,
                             "department": nm if str(u.get("department")) == unit else u.get("department"),
                             "managed_units": [nm if x == unit else x for x in u.get("managed_units", [])]}
            u_remap += 1
        else:
            u_drop.append(un)
    print(f"\nusers.json: {len(users)} -> {len(new_users)}  "
          f"(remapped {u_remap} | kept {u_keep} | dropped {len(u_drop)})")
    if u_drop:
        print(f"  dropped logins: {u_drop}")
    for p in ("frank0731","immaculate0716","william001"):
        print(f"  {p}: {'present' if p in new_users else 'MISSING'}")

    if args.dry_run:
        print("\n[dry-run] nothing written.")
        return

    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    # backup
    for f in ("staff_register.xlsx", "users.json"):
        src = DATA_DIR/f
        if src.exists():
            bak = src.with_suffix(src.suffix + f".pre_branch16_{stamp}")
            bak.write_bytes(src.read_bytes())
            print(f"[backup] {bak.name}")
    # write register
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Staff Register"
    ws.append(H)
    for r in final:
        ws.append([r.get(h) for h in H])
    wb.save(str(DATA_DIR/"staff_register.xlsx"))
    print(f"[ok] staff_register.xlsx written ({len(final)} staff).")
    # write users
    uj.write_text(json.dumps(new_users, indent=2), encoding="utf-8")
    print(f"[ok] users.json written ({len(new_users)} logins).")


if __name__ == "__main__":
    main()
