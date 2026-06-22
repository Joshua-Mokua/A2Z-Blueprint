#!/usr/bin/env python3
"""
clean_watchlist_owners.py — final ownership cleanup of the watchlist credit book.

After the branch remap, surviving rows are owned by a mix of real RMs and
non-account-owning staff (C-suite, Heads, Ops, Tellers, IT). A credit-monitoring
account should only sit under a relationship/credit RM. This deletes every row
whose rm_code resolves to a NON-RM register role, keeping only genuine
account-owning RMs. William Mwanake (300001, MD) is hard-protected and never
touched.

The keep-set is ROLE-BASED and printed in the dry-run so you can see exactly
which register roles survive before anything is deleted — no silent rules.

    python scripts\\clean_watchlist_owners.py            # dry-run + role breakdown
    python scripts\\clean_watchlist_owners.py --apply    # backup + delete non-RM rows
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
REGISTER = DATA_DIR / "staff_register.xlsx"

PROTECT = {"300001"}  # William Mwanake, MD — never delete

# Account-owning roles (KEEP). Substring match, case-insensitive.
KEEP_ROLE_HINTS = (
    "relationship", "personal banker", "relationship officer",
    "relationship manager", "credit analyst", "credit officer",
    "ro pb", "ro bb", "dso", "direct sales",
)


def register_lookup():
    import openpyxl
    wb = openpyxl.load_workbook(REGISTER, read_only=True)
    ws = wb["Staff Register"] if "Staff Register" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = headers.index("Staff Code") if "Staff Code" in headers else 0
    ni = headers.index("Staff Name") if "Staff Name" in headers else 1
    ri = headers.index("Role") if "Role" in headers else 2
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[ci]:
            out[str(row[ci]).strip()] = (str(row[ni] or "").strip(), str(row[ri] or "").strip())
    wb.close()
    return out


def is_keep_role(role: str) -> bool:
    rl = (role or "").lower()
    return any(h in rl for h in KEEP_ROLE_HINTS)


def main():
    apply = "--apply" in sys.argv
    from utils.db import db
    reg = register_lookup()
    rows = db.fetch_all("SELECT id, rm_code, rm_name FROM watchlist", ())
    print(f"register codes: {len(reg)} | watchlist rows: {len(rows)}\n")

    keep_ids, del_ids = [], []
    keep_roles, del_roles = Counter(), Counter()
    protected = 0
    for r in rows:
        code = str(r.get("rm_code") or "").strip()
        if code in PROTECT:
            protected += 1
            keep_ids.append(r["id"])
            continue
        name, role = reg.get(code, (None, None))
        if name is not None and is_keep_role(role):
            keep_ids.append(r["id"])
            keep_roles[role] += 1
        else:
            del_ids.append(r["id"])
            del_roles[role if name is not None else "(rm_code not in register)"] += 1

    print(f"PROTECTED (William):  {protected}")
    print(f"KEEP (real RM):       {len(keep_ids) - protected}")
    print(f"DELETE (non-RM):      {len(del_ids)}")
    print(f"final row count:      {len(keep_ids)}\n")

    print("--- KEEP roles (these survive) ---")
    for role, n in keep_roles.most_common():
        print(f"  {n:>5}  {role}")
    print("\n--- DELETE roles (removed) ---")
    for role, n in del_roles.most_common():
        print(f"  {n:>5}  {role}")

    if not apply:
        print("\n[DRY-RUN] No DB change. Re-run with --apply to back up + delete.")
        return
    if not del_ids:
        print("\nNothing to delete.")
        return

    full = db.fetch_all("SELECT * FROM watchlist", ())
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = DATA_DIR / f"watchlist_db_table.pre_ownerclean_{ts}.json"
    backup.write_text(json.dumps(full, default=str, indent=2), encoding="utf-8")
    print(f"\n[backup] {len(full)} rows -> {backup.name}")

    with db.transaction() as conn:
        for i in range(0, len(del_ids), 500):
            chunk = del_ids[i:i + 500]
            ph = ",".join(["%s"] * len(chunk))
            db.execute(f"DELETE FROM watchlist WHERE id IN ({ph})", tuple(chunk), conn=conn)

    after = db.fetch_one("SELECT COUNT(*) AS n FROM watchlist", ())
    print(f"[apply] deleted {len(del_ids)} non-RM rows. watchlist now: {after['n']} rows.")
    print("Restart API + run harness — every credit account now owned by a real RM.")


if __name__ == "__main__":
    main()
