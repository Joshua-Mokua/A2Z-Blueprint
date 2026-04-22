#!/usr/bin/env python3
"""
generate_propositions.py — Simulate proposition overlay data.
Run from a2z/ directory: python generate_propositions.py
"""
import json, random, csv
from pathlib import Path
from datetime import date, timedelta
from collections import defaultdict

random.seed(2026)
DATA = Path('data')
CBS  = Path('../cbs_data')
today = date.today()

def rdate(start=180, end=0):
    return (today - timedelta(days=random.randint(end, start))).isoformat()

# ── Load base data ─────────────────────────────────────────────────
users = json.loads((DATA/'users.json').read_text())
import openpyxl
wb = openpyxl.load_workbook(str(DATA/'staff_register.xlsx'))
ws = wb.active; hdr = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
sc_i = hdr.index('Staff Code'); role_i = hdr.index('Role'); unit_i = hdr.index('Unit')
all_staff = [{'code':str(r[sc_i]),'role':str(r[role_i] or ''),'unit':str(r[unit_i] or '')}
             for r in ws.iter_rows(min_row=2, values_only=True) if r[sc_i]]

# ── 1. Proposition definitions ─────────────────────────────────────
PROPOSITIONS = {
    'WB':  {'name':'Women Banking',       'icon':'👩', 'color':'#EC4899',
             'head_role':'Head Of Women Banking',
             'rm_roles':['Relationship','Business Banker','Personal Banker'],
             'description':'Banking proposition for women-owned businesses and women customers'},
    'DIA': {'name':'Diaspora Banking',    'icon':'✈️', 'color':'#0EA5E9',
             'head_role':'Senior Manager Diaspora Banking',
             'rm_roles':['Diaspora','Relationship'],
             'description':'Financial services for Kenyan diaspora and returnees'},
    'SME': {'name':'SME Banking',         'icon':'🏪', 'color':'#F59E0B',
             'head_role':'Head of MSME',
             'rm_roles':['SME','Business Banker','MSME'],
             'description':'Dedicated SME and MSME banking solutions'},
    'AGR': {'name':'Agribusiness',        'icon':'🌾', 'color':'#10B981',
             'head_role':'Head of MSME',
             'rm_roles':['Agribusiness','Relationship'],
             'description':'Agricultural finance and agri-business banking'},
    'TF':  {'name':'Trade Finance',       'icon':'🚢', 'color':'#6366F1',
             'head_role':'Head Of Corporates & Trade Finance',
             'rm_roles':['Trade Finance','Corporate'],
             'description':'Import/export finance, letters of credit, guarantees'},
    'GOV': {'name':'Govt & Institutional','icon':'🏛️', 'color':'#64748B',
             'head_role':'Head of Government & Institutional Banking',
             'rm_roles':['Public Sector','Institutional','Government'],
             'description':'Government parastatals, NGOs, SACCOs, faith-based'},
    'BNC': {'name':'Bancassurance',       'icon':'🛡️', 'color':'#8B5CF6',
             'head_role':'General Manager - Bancassurance',
             'rm_roles':['Bancassurance'],
             'description':'Embedded insurance products through banking channels'},
    'DFS': {'name':'Digital Finance',     'icon':'📱', 'color':'#06B6D4',
             'head_role':'Head of Digital Financial Services',
             'rm_roles':['Digital','Agency'],
             'description':'Mobile banking, agency network, digital lending'},
}

# ── 2. KPI definitions per proposition (OVERLAY — no double-counting) ──
PROPOSITION_KPIS = {
    'WB': [
        {'id':'WB_CUST_ACQ',  'name':'WB Customers Acquired',    'unit':'count','direction':'higher','weight':0.20,'target':1500,  'actual_range':(800,1600)},
        {'id':'WB_ACTIVE',    'name':'WB Active Customers',       'unit':'count','direction':'higher','weight':0.15,'target':8000,  'actual_range':(5000,9000)},
        {'id':'WB_WALLET_SH', 'name':'WB Deposit Wallet Share',   'unit':'%',   'direction':'higher','weight':0.20,'target':65.0,  'actual_range':(50,75)},
        {'id':'WB_LOAN_PEN',  'name':'WB Loan Penetration',       'unit':'%',   'direction':'higher','weight':0.20,'target':35.0,  'actual_range':(20,45)},
        {'id':'WB_XSELL',     'name':'WB Cross-sell Score',       'unit':'count','direction':'higher','weight':0.10,'target':2.5,  'actual_range':(1.8,3.2)},
        {'id':'WB_EVENTS',    'name':'WB Events Held',            'unit':'count','direction':'higher','weight':0.05,'target':24,   'actual_range':(10,30)},
        {'id':'WB_NPS',       'name':'WB NPS Score',              'unit':'score','direction':'higher','weight':0.10,'target':55.0, 'actual_range':(40,70)},
    ],
    'DIA': [
        {'id':'DIA_ACQ',      'name':'Diaspora Customers Acquired','unit':'count','direction':'higher','weight':0.20,'target':500,  'actual_range':(200,600)},
        {'id':'DIA_REMIT',    'name':'Diaspora Remittances Volume','unit':'KES', 'direction':'higher','weight':0.30,'target':2e9,  'actual_range':(1e9,3e9)},
        {'id':'DIA_LOANS',    'name':'Diaspora Loan Book',         'unit':'KES', 'direction':'higher','weight':0.20,'target':1.5e9,'actual_range':(8e8,2e9)},
        {'id':'DIA_ACTIV',    'name':'Diaspora Account Activation','unit':'%',   'direction':'higher','weight':0.15,'target':70.0, 'actual_range':(50,85)},
        {'id':'DIA_DIGITAL',  'name':'Diaspora Digital Adoption',  'unit':'%',   'direction':'higher','weight':0.15,'target':80.0, 'actual_range':(60,90)},
    ],
    'SME': [
        {'id':'SME_ACQ',      'name':'SME Customers Acquired',    'unit':'count','direction':'higher','weight':0.20,'target':800,  'actual_range':(400,1000)},
        {'id':'SME_BOOK',     'name':'SME Loan Book',             'unit':'KES', 'direction':'higher','weight':0.25,'target':5e9,  'actual_range':(3e9,7e9)},
        {'id':'SME_NPL',      'name':'SME NPL Ratio',             'unit':'%',   'direction':'lower', 'weight':0.20,'target':4.0,  'actual_range':(3.0,8.0)},
        {'id':'SME_TRAINING', 'name':'SME Training Events',       'unit':'count','direction':'higher','weight':0.10,'target':36,   'actual_range':(20,50)},
        {'id':'SME_XSELL',    'name':'SME Cross-sell',            'unit':'count','direction':'higher','weight':0.15,'target':3.0,  'actual_range':(2.0,4.0)},
        {'id':'SME_REPAY',    'name':'SME Repayment Rate',        'unit':'%',   'direction':'higher','weight':0.10,'target':92.0, 'actual_range':(80,98)},
    ],
    'AGR': [
        {'id':'AGR_ACQ',      'name':'Agri Customers Acquired',   'unit':'count','direction':'higher','weight':0.20,'target':600,  'actual_range':(300,800)},
        {'id':'AGR_BOOK',     'name':'Agri Loan Book',            'unit':'KES', 'direction':'higher','weight':0.30,'target':3e9,  'actual_range':(1.5e9,4e9)},
        {'id':'AGR_NPL',      'name':'Agri NPL Ratio',            'unit':'%',   'direction':'lower', 'weight':0.20,'target':5.0,  'actual_range':(3.0,10.0)},
        {'id':'AGR_SEASONAL', 'name':'Seasonal Loan Recovery',    'unit':'%',   'direction':'higher','weight':0.30,'target':88.0, 'actual_range':(70,95)},
    ],
    'TF': [
        {'id':'TF_VOLUME',    'name':'Trade Finance Volume',      'unit':'KES', 'direction':'higher','weight':0.35,'target':8e9,  'actual_range':(4e9,12e9)},
        {'id':'TF_CLIENTS',   'name':'Active TF Clients',         'unit':'count','direction':'higher','weight':0.20,'target':120,  'actual_range':(60,160)},
        {'id':'TF_TAT',       'name':'LC/LG Processing TAT',      'unit':'days','direction':'lower', 'weight':0.25,'target':3.0,  'actual_range':(2,7)},
        {'id':'TF_INCOME',    'name':'TF Fee Income',             'unit':'KES', 'direction':'higher','weight':0.20,'target':500e6,'actual_range':(200e6,700e6)},
    ],
    'GOV': [
        {'id':'GOV_CLIENTS',  'name':'Govt/PSI Clients',          'unit':'count','direction':'higher','weight':0.25,'target':80,   'actual_range':(40,120)},
        {'id':'GOV_DEPOSITS', 'name':'Govt Deposits',             'unit':'KES', 'direction':'higher','weight':0.35,'target':15e9, 'actual_range':(8e9,25e9)},
        {'id':'GOV_LOANS',    'name':'Govt Loan Book',            'unit':'KES', 'direction':'higher','weight':0.25,'target':5e9,  'actual_range':(2e9,8e9)},
        {'id':'GOV_TENDERS',  'name':'Tender Bonds Issued',       'unit':'count','direction':'higher','weight':0.15,'target':200,  'actual_range':(100,300)},
    ],
    'BNC': [
        {'id':'BNC_POLICIES', 'name':'Policies Sold',             'unit':'count','direction':'higher','weight':0.30,'target':5000, 'actual_range':(2000,7000)},
        {'id':'BNC_PREMIUM',  'name':'Bancassurance Premium',     'unit':'KES', 'direction':'higher','weight':0.30,'target':800e6,'actual_range':(400e6,1.2e9)},
        {'id':'BNC_PENETR',   'name':'BNC Penetration Rate',      'unit':'%',   'direction':'higher','weight':0.20,'target':18.0, 'actual_range':(10,25)},
        {'id':'BNC_CLAIMS',   'name':'Claims Ratio',              'unit':'%',   'direction':'lower', 'weight':0.20,'target':55.0, 'actual_range':(40,70)},
    ],
    'DFS': [
        {'id':'DFS_MOBILE',   'name':'Mobile Banking Active Users','unit':'count','direction':'higher','weight':0.25,'target':150000,'actual_range':(80000,200000)},
        {'id':'DFS_MIGRATION','name':'Digital Txn Migration',     'unit':'%',   'direction':'higher','weight':0.25,'target':60.0, 'actual_range':(40,80)},
        {'id':'DFS_AGENTS',   'name':'Active Agency Agents',      'unit':'count','direction':'higher','weight':0.20,'target':800,  'actual_range':(400,1200)},
        {'id':'DFS_MOBILE_L', 'name':'Mobile Loan Uptake',        'unit':'%',   'direction':'higher','weight':0.15,'target':25.0, 'actual_range':(15,40)},
        {'id':'DFS_UPTIME',   'name':'Platform Uptime',           'unit':'%',   'direction':'higher','weight':0.15,'target':99.5, 'actual_range':(97,100)},
    ],
}

# ── 3. Generate proposition performance data ─────────────────────
def gen_actual(kpi):
    lo, hi = kpi['actual_range']
    raw = random.uniform(lo, hi)
    if kpi['unit'] in ('count','KES'): return round(raw, 0)
    return round(raw, 2)

proposition_data = {}
for tag, kpis in PROPOSITION_KPIS.items():
    prop = PROPOSITIONS[tag]
    head_user = next(((u,d) for u,d in users.items()
                      if d.get('role','') == prop['head_role']), (None,{}))
    head_sc = str(head_user[1].get('staff_code','')) if head_user[0] else ''
    head_name = head_user[1].get('full_name','') if head_user[0] else ''

    # Generate RM-level staff who contribute to this proposition
    prop_rms = [s for s in all_staff
                if any(kw.lower() in s['role'].lower() for kw in prop['rm_roles'])]
    # Sample a representative subset
    sample_rms = random.sample(prop_rms, min(len(prop_rms), 15))

    # Monthly trend (last 12 months)
    months = [(today - timedelta(days=30*i)).strftime('%b-%y') for i in range(11, -1, -1)]

    # KPI actuals with monthly breakdown
    kpi_records = []
    for kpi in kpis:
        actual = gen_actual(kpi)
        # Monthly: build trend toward actual
        monthly = []
        base = actual * 0.7
        for i, m in enumerate(months):
            progress = base + (actual - base) * (i / 11) * random.uniform(0.8, 1.2)
            monthly.append({'month': m, 'value': round(progress, 2)})

        # Achievement
        if kpi['direction'] == 'lower':
            ach = round(kpi['target'] / max(actual, 0.001) * 100, 1)
        else:
            ach = round(actual / max(kpi['target'], 0.001) * 100, 1)

        score = (5.0 if ach >= 130 else 4.5 if ach >= 120 else 4.0 if ach >= 110
                 else 3.5 if ach >= 100 else 3.0 if ach >= 91 else 2.5 if ach >= 61
                 else 2.0 if ach >= 51 else 1.5)

        kpi_records.append({**kpi, 'actual': actual, 'achievement': ach,
                             'score': score, 'monthly_trend': monthly})

    # Overall proposition score
    total_w = sum(k['weight'] for k in kpi_records)
    prop_score = sum(k['score'] * k['weight'] for k in kpi_records) / max(total_w, 0.01)

    # Branch contribution breakdown (which branches have tagged customers)
    branches = list(set(s['unit'] for s in all_staff if s['unit'] not in ('Head Office','HO','')))
    branch_contribs = []
    for branch in random.sample(branches, min(15, len(branches))):
        branch_contribs.append({
            'branch': branch,
            'tagged_customers': random.randint(50, 500),
            'proposition_score': round(random.uniform(2.5, 4.5), 2),
            'champion_rm': random.choice(sample_rms)['code'] if sample_rms else '',
        })

    proposition_data[tag] = {
        'tag': tag,
        'name': prop['name'],
        'icon': prop['icon'],
        'color': prop['color'],
        'description': prop['description'],
        'head_staff_code': head_sc,
        'head_name': head_name,
        'proposition_score': round(prop_score, 2),
        'kpis': kpi_records,
        'branch_contributions': branch_contribs,
        'contributing_rms': [s['code'] for s in sample_rms],
        'total_tagged_customers': sum(b['tagged_customers'] for b in branch_contribs),
        'last_updated': today.isoformat(),
        'period': '2026',
    }

(DATA/'proposition_performance.json').write_text(json.dumps(proposition_data, indent=2))
print(f"✅ proposition_performance.json: {len(proposition_data)} propositions")
for tag, d in proposition_data.items():
    print(f"  {d['icon']} {d['name']:<25} score={d['proposition_score']:.2f} "
          f"customers={d['total_tagged_customers']:,}")

# ── 4. Segment tags on customers (CIF tagging) ─────────────────────
# Tag a sample of CBS CIFs with proposition codes
segment_tags = {}
if (CBS/'accounts.csv').exists():
    cifs = set()
    with open(CBS/'accounts.csv') as f:
        for row in csv.DictReader(f):
            cif = row.get('cif','').strip()
            if cif: cifs.add(cif)
    cif_list = sorted(cifs)
    print(f"\nTagging {len(cif_list):,} CIFs with proposition tags...")
    # Each CIF gets 0-2 tags (most get 0, some get 1, few get 2)
    for cif in random.sample(cif_list, min(50000, len(cif_list))):
        r = random.random()
        if r < 0.05:      segment_tags[cif] = ['WB']
        elif r < 0.08:    segment_tags[cif] = ['SME']
        elif r < 0.10:    segment_tags[cif] = ['DFS']
        elif r < 0.115:   segment_tags[cif] = ['AGR']
        elif r < 0.13:    segment_tags[cif] = ['DIA']
        elif r < 0.135:   segment_tags[cif] = ['GOV']
        elif r < 0.14:    segment_tags[cif] = ['TF']
        elif r < 0.145:   segment_tags[cif] = ['BNC']
        elif r < 0.15:    segment_tags[cif] = ['WB','SME']  # multi-tag
    (DATA/'segment_tags.json').write_text(json.dumps(segment_tags, indent=2))
    tag_counts = defaultdict(int)
    for tags in segment_tags.values():
        for t in tags: tag_counts[t] += 1
    print(f"  Tagged: {len(segment_tags):,} CIFs")
    for tag, cnt in sorted(tag_counts.items(), key=lambda x:-x[1]):
        print(f"  {tag}: {cnt:,}")
else:
    # Without CBS, create a representative sample
    sample_cifs = [f'CIF{100000+i}' for i in range(50000)]
    tags_list = ['WB','SME','DFS','AGR','DIA','GOV','TF','BNC']
    for cif in random.sample(sample_cifs, 7500):
        segment_tags[cif] = [random.choice(tags_list)]
    (DATA/'segment_tags.json').write_text(json.dumps(segment_tags, indent=2))
    print(f"  Sample tags: {len(segment_tags):,} CIFs")

print(f"\n✅ Done. Files created:")
for fname in ['proposition_performance.json', 'segment_tags.json']:
    p = DATA / fname
    print(f"  {fname}: {p.stat().st_size/1024:.0f} KB")
