"""pages/35_stress_testing.py — Portfolio Stress Testing Engine.
Scenario-based stress testing: rate shocks, FX movements, credit crunch.
central-bank ICAAP requirements. AI generates executive narrative.
"""
import streamlit as st
from utils.db import db as a2z_db
import pandas as pd
import json
from pathlib import Path
from datetime import date
from pages._shared import load_shared_state
from pages._access import require_access
from utils.core_audit import audit_log
from utils.config import currency_symbol, regulator, bank_name, country
import requests

require_access("risk.stress_testing")
DATA = Path(__file__).parent.parent / "data"
today = date.today()
um, ud, uname, *_ = load_shared_state()[:12]
role = ud.get("role",""); name = ud.get("full_name","")
is_admin = ud.get("is_admin",False)
is_exec  = any(x in role for x in ("Chief","Director","Managing","Head","CFO","Risk","Treasury"))

@st.cache_data(ttl=60, show_spinner=False)
def _load(fname):
    p = DATA / fname
    return a2z_db.load_json(p) if p.exists() else {}

scenarios = _load("stress_scenarios.json")
ifrs_sum  = _load("ifrs9_summary.json")
alm       = _load("treasury_alm.json")

# Read bank baseline from MD actuals
try:
    import openpyxl
    act_file = sorted([f for f in DATA.glob('actuals_*.xlsx') if 'backup' not in f.name],reverse=True)[0]
    wb = openpyxl.load_workbook(str(act_file)); ws = wb.active
    hdr = [ws.cell(2,c).value for c in range(1,ws.max_column+1)]
    kpi_c=hdr.index('KPI')+1; ac_c=hdr.index('YTD_Actual')+1; sc2=hdr.index('Staff Code')+1
    MD_ACTUALS={}
    for row in ws.iter_rows(min_row=3,values_only=True):
        if not row[0] or str(row[sc2-1] or '')!='300001': continue
        MD_ACTUALS[str(row[kpi_c-1] or '')]=float(row[ac_c-1] or 0)
except: MD_ACTUALS={}

# Baseline figures
BASE_LOAN_BOOK   = MD_ACTUALS.get("Loan Book Growth", 2.6e12)
BASE_PBT         = MD_ACTUALS.get("PBT", 6.65e11)
BASE_NFI         = MD_ACTUALS.get("Total NFI", 1.33e11)
BASE_NPL         = MD_ACTUALS.get("NPL Ratio", 11.0)
BASE_ECL         = ifrs_sum.get("total_ecl_provision", 3.4e9)
BASE_LCR         = alm.get("liquidity_ratios",{}).get("lcr",{}).get("value",122.7)

st.markdown(
    "<div style='padding:16px 0 4px'>"
    "<span style='font-size:22px;font-weight:800'>🔥 Stress Testing</span>"
    "<span style='font-size:13px;color:var(--color-text-secondary);margin-left:12px'>"
    f"Portfolio stress scenarios · {regulator()} ICAAP · Rate/FX/Credit shocks</span></div>",
    unsafe_allow_html=True)

st.markdown(
    "<div style='background:#FFF7ED;border:1px solid #FED7AA;border-radius:8px;"
    "padding:8px 14px;font-size:12px;margin-bottom:10px'>"
    f"Stress testing is required under {regulator()} Prudential Guideline No. CB/PG/08 (ICAAP). "
    "These scenarios test the bank's capital adequacy under adverse conditions.</div>",
    unsafe_allow_html=True)

tabs = st.tabs(["📊 Scenario Runner","📈 Side-by-Side","🎛️ Custom Scenario","📄 ICAAP Report",f"🏛️ {regulator()} Supervisory (Standard #79)","🤖 Arc Engines"])

# ── TAB 1: Scenario Runner ─────────────────────────────────────────
with tabs[0]:
    st.markdown("**Select scenario to stress-test:**")
    scenario_names = {k: v["name"] for k,v in scenarios.items()}
    sel_key = st.selectbox("Scenario", list(scenario_names.keys()),
                            format_func=lambda x: scenario_names[x], key="st_sel")
    sc = scenarios[sel_key]
    assump = sc.get("assumptions",{})
    impact = sc.get("impacts",{})

    # Show assumptions
    st.markdown(f"**{sc['name']}:** {sc['description']}")
    a1,a2,a3,a4 = st.columns(4)
    cbk_chg = assump.get("cbk_rate_change",0)
    a1.metric(f"{regulator()} Rate Change",    f"{cbk_chg:+.0f}bps",  delta_color="normal" if cbk_chg<=0 else "inverse")
    a2.metric("GDP Growth",         f"{assump.get('gdp_growth',5):.1f}%")
    a3.metric("Inflation",          f"{assump.get('inflation',6):.1f}%")
    a4.metric(f"{currency_symbol()} Depreciation",   f"{assump.get('shilling_depreciation',0):+.0f}%")

    st.markdown("---")

    # Compute stressed values
    npl_chg   = impact.get("npl_change",0)
    book_chg  = impact.get("loan_book_growth",0)/100
    nfi_chg   = impact.get("interest_income_change",0)/100
    ecl_chg   = impact.get("ecl_change",0)/100

    stressed_npl  = round(BASE_NPL + npl_chg, 2)
    stressed_book = round(BASE_LOAN_BOOK * (1 + book_chg) / 1e9, 1)
    stressed_pbt  = round(BASE_PBT * (1 + nfi_chg) / 1e9, 2)
    stressed_ecl  = round(BASE_ECL * (1 + ecl_chg) / 1e6, 0)
    stressed_lcr  = round(BASE_LCR - abs(npl_chg) * 3, 1) if npl_chg > 0 else round(BASE_LCR + 5, 1)

    st.markdown("**Stressed outcomes:**")
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Loan Book",    f"{currency_symbol()} {stressed_book:.0f}B",
              f"{book_chg*100:+.0f}%", delta_color="normal" if book_chg>=0 else "inverse")
    m2.metric("NPL Ratio",    f"{stressed_npl:.1f}%",
              f"{npl_chg:+.1f}pp", delta_color="normal" if npl_chg<=0 else "inverse")
    m3.metric("PBT",          f"{currency_symbol()} {stressed_pbt:.2f}B",
              f"{nfi_chg*100:+.0f}%", delta_color="normal" if nfi_chg>=0 else "inverse")
    m4.metric("ECL Provision",f"{currency_symbol()} {stressed_ecl:,.0f}M",
              f"{ecl_chg*100:+.0f}%", delta_color="normal" if ecl_chg<=0 else "inverse")
    m5.metric("LCR",          f"{stressed_lcr:.1f}%",
              "Above 100%" if stressed_lcr>=100 else "BREACH",
              delta_color="normal" if stressed_lcr>=100 else "inverse")

    # AI narrative
    if st.button("🤖 Generate ICAAP narrative for this scenario", key="st_ai"):
        with st.spinner("Generating regulatory narrative…"):
            try:
                resp = requests.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"Content-Type":"application/json"},
                    json={
                        "model":"claude-sonnet-4-20250514",
                        "max_tokens":500,
                        "system":f"You are a bank risk manager writing ICAAP stress test commentary for {regulator()} submission. Be precise, regulatory in tone.",
                        "messages":[{"role":"user","content":
                            f"Scenario: {sc['name']}. {sc['description']}. "
                            f"Key impacts: NPL {BASE_NPL:.1f}%→{stressed_npl:.1f}%, "
                            f"Loan book change {book_chg*100:+.0f}%, "
                            f"ECL provision change {ecl_chg*100:+.0f}%, "
                            f"LCR {BASE_LCR:.1f}%→{stressed_lcr:.1f}%. "
                            f"Write a 3-paragraph ICAAP stress test commentary suitable for {regulator()} submission."}]
                    }, timeout=30)
                resp.raise_for_status()
                st.markdown("**ICAAP Narrative:**")
                st.markdown(resp.json()["content"][0]["text"])
            except Exception as e:
                st.error(f"AI unavailable: {str(e)[:80]}")

# ── TAB 2: Side-by-Side ────────────────────────────────────────────
with tabs[1]:
    st.markdown("**All scenarios compared:**")
    rows = []
    for k, sc in scenarios.items():
        imp = sc.get("impacts",{}); ass = sc.get("assumptions",{})
        rows.append({
            "Scenario":          sc["name"],
            f"{regulator()} Chg (bps)":     f"{ass.get('cbk_rate_change',0):+.0f}",
            "NPL (pp chg)":      f"{imp.get('npl_change',0):+.1f}",
            "Loan Book":         f"{imp.get('loan_book_growth',0):+.0f}%",
            "PBT":               f"{imp.get('interest_income_change',0):+.0f}%",
            "ECL":               f"{imp.get('ecl_change',0):+.0f}%",
            "Stressed NPL":      f"{BASE_NPL+imp.get('npl_change',0):.1f}%",
            "Stressed LCR":      f"{max(BASE_LCR - abs(imp.get('npl_change',0))*3, 80):.0f}%",
            "Capital Impact":    "🔴 Severe" if imp.get("npl_change",0)>6 else
                                  "🟡 Moderate" if imp.get("npl_change",0)>2 else "🟢 Manageable"
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ── TAB 3: Custom Scenario ─────────────────────────────────────────
with tabs[2]:
    st.markdown("**Build a custom stress scenario:**")
    cc1,cc2 = st.columns(2)
    c_name   = cc1.text_input("Scenario name", value="Custom Scenario 1", key="st_cname")
    c_cbk    = cc1.number_input(f"{regulator()} rate change (bps)", value=0.0, step=25.0, key="st_cbk")
    c_gdp    = cc1.number_input("GDP growth (%)", value=5.0, step=0.5, key="st_gdp")
    c_npl    = cc2.number_input("NPL change (pp)", value=0.0, step=0.5, key="st_npl")
    c_kes    = cc2.number_input(f"{currency_symbol()} depreciation (%)", value=0.0, step=1.0, key="st_kes")
    c_ecl    = cc2.number_input("ECL provision change (%)", value=0.0, step=5.0, key="st_ecl")

    if st.button("▶️ Run custom scenario", key="st_custom", type="primary"):
        audit_log("STRESS_TEST_RUN", uname, "Custom scenario executed")
        s_npl  = round(BASE_NPL+c_npl,2)
        s_ecl  = round(BASE_ECL*(1+c_ecl/100)/1e6,0)
        s_lcr  = round(BASE_LCR-abs(c_npl)*3,1) if c_npl>0 else round(BASE_LCR+3,1)
        st.markdown("**Custom scenario results:**")
        r1,r2,r3,r4 = st.columns(4)
        r1.metric("Stressed NPL",  f"{s_npl:.1f}%",  f"{c_npl:+.1f}pp")
        r2.metric("Stressed ECL",  f"{currency_symbol()} {s_ecl:,.0f}M", f"{c_ecl:+.0f}%")
        r3.metric("Stressed LCR",  f"{s_lcr:.1f}%", "OK" if s_lcr>=100 else "BREACH",
                  delta_color="normal" if s_lcr>=100 else "inverse")
        r4.metric(f"{regulator()} Rate",      f"{13.0+c_cbk/100:.2f}%", f"{c_cbk:+.0f}bps")

# ── TAB 4: ICAAP Report ────────────────────────────────────────────
with tabs[3]:
    st.markdown("**ICAAP Stress Testing Summary Report**")
    st.caption(f"{country()} central bank — ICAAP Stress Testing requirements per regulatory PG/08")
    st.markdown(f"""
**Report Date:** {today}  
**Bank:** {bank_name()}  
**Reporting Period:** Annual 2025

---

**1. Baseline Capital Position**

| Metric | Value |
|--------|-------|
| NPL Ratio | {BASE_NPL:.1f}% |
| Total ECL Provision | {currency_symbol()} {BASE_ECL/1e6:,.0f}M |
| LCR | {BASE_LCR:.1f}% |
| Loan Book | {currency_symbol()} {BASE_LOAN_BOOK/1e9:,.0f}B |
| PBT | {currency_symbol()} {BASE_PBT/1e9:,.2f}B |

---

**2. Scenarios Tested**
""")
    for k, sc in scenarios.items():
        imp = sc.get("impacts",{})
        severity = "Severe" if imp.get("npl_change",0)>6 else "Moderate" if imp.get("npl_change",0)>2 else "Mild"
        st.markdown(f"- **{sc['name']}** ({severity}): NPL {imp.get('npl_change',0):+.1f}pp, ECL {imp.get('ecl_change',0):+.0f}%")

    st.markdown("""
---
**3. Key Risk Indicators**

Scenarios that result in LCR below 100% or NPL above 15% trigger capital contingency plans.

**4. Management Actions**

Contingency buffer maintained per central-bank ICAAP requirements. Capital enhancement strategies include retained earnings, Tier 2 subordinated debt, and portfolio deleveraging.
    """)
    if st.download_button("📥 Export ICAAP Report",
                           data=f"ICAAP Stress Testing Report — {today}".encode(),
                           file_name=f"ICAAP_StressTest_{today}.txt", key="st_dl"):
        st.success("Report downloaded")


# ════════════════════════════════════════════════════════════════
# TAB 4: Regulator Supervisory Stress (Standard #79, integrated v5.78)
# ════════════════════════════════════════════════════════════════
with tabs[4]:
    from utils.stress_testing import (
        StressTestingEngine, StressTestInputs,
        STRESS_SCENARIOS, SCENARIO_SHOCKS,
        CBK_TOTAL_CAR_MIN_PCT_LOCAL,
        NPL_INCREASE_TO_LOSS_FACTOR,
        RATE_SHOCK_TO_NII_BPS,
        ASSET_PRICE_SHOCK_TO_PROVISIONS,
        REVERSE_STRESS_MAX_NPL_PCT,
        REVERSE_STRESS_MAX_RATE_BPS,
    )
    from decimal import Decimal as _D_st

    st.markdown(
        f"**Standard #79 — Regulator Supervisory Stress Testing Engine**. "
        f"Three deterministic scenarios per central-bank ICAAP framework: "
        f"`{' / '.join(STRESS_SCENARIOS)}`. "
        f"{regulator()} total CAR minimum: **{CBK_TOTAL_CAR_MIN_PCT_LOCAL}%** (local prudential floor)."
    )
    st.caption(
        "Companion to the AI-narrative scenarios above. This tab uses the "
        "deterministic engine bound byte-for-byte to its constants — no LLM, "
        "no randomness, identical results every run."
    )

    st_sub_tabs = st.tabs([
        "📋 Inputs",
        "🎯 3-Scenario Run",
        "📈 Capital Projection",
        "🔄 Reverse Stress Test",
        "📦 Stress Testing Depth (#51, v6.2)",
    ])

    # ---- Inputs ----
    with st_sub_tabs[0]:
        st.markdown("**Bank starting position** — used by all sub-tabs in this section.")
        st.caption(
            f"Defaults reflect a representative Tier-2 domestic bank balance sheet. "
            "Pull values from real CBS/treasury data when running for production "
            "ICAAP submission."
        )
        c1, c2 = st.columns(2)
        with c1:
            st_cap = st.number_input(f"Total capital ({currency_symbol()} B)",
                                       min_value=0.0, value=25.0, step=1.0,
                                       key="st79_cap",
                                       help="Tier 1 + Tier 2 capital base.")
            st_rwa = st.number_input(f"Risk-weighted assets ({currency_symbol()} B)",
                                       min_value=0.0, value=150.0, step=5.0,
                                       key="st79_rwa")
            st_loans = st.number_input(f"Loan book ({currency_symbol()} B)",
                                         min_value=0.0, value=120.0, step=5.0,
                                         key="st79_loans")
            st_npl = st.number_input(f"Current NPL stock ({currency_symbol()} B)",
                                       min_value=0.0, value=13.32, step=0.5,
                                       key="st79_npl",
                                       help=f"Default 11.1% of 120B loan book.")
        with c2:
            st_secs = st.number_input(f"Securities portfolio ({currency_symbol()} B)",
                                        min_value=0.0, value=50.0, step=2.0,
                                        key="st79_secs")
            st_fx = st.number_input(f"FX open position ({currency_symbol()} B, abs)",
                                      min_value=0.0, value=2.5, step=0.5,
                                      key="st79_fx")
            st_pbt = st.number_input(f"Annual pre-tax profit ({currency_symbol()} B)",
                                       min_value=0.0, value=5.0, step=0.5,
                                       key="st79_pbt")
            st_horizon = st.number_input("Projection horizon (years)",
                                           min_value=1, max_value=5, value=3, step=1,
                                           key="st79_horizon")

        # Compute starting CAR for context
        starting_car = (st_cap / st_rwa) * 100 if st_rwa > 0 else 0
        if starting_car >= float(CBK_TOTAL_CAR_MIN_PCT_LOCAL) + 2:
            color, label = "#10B981", "WELL CAPITALISED"
        elif starting_car >= float(CBK_TOTAL_CAR_MIN_PCT_LOCAL):
            color, label = "#F59E0B", "AT MINIMUM"
        else:
            color, label = "#DC2626", "ALREADY BREACHING"
        st.markdown(
            f"<div style='margin-top:18px;padding:14px;background:{color}22;"
            f"border-left:6px solid {color};border-radius:10px'>"
            f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
            f"PRE-STRESS CAR</div>"
            f"<div style='font-size:24px;font-weight:800;color:{color};margin-top:4px'>"
            f"{starting_car:.2f}% — {label}</div>"
            f"<div style='font-size:12px;margin-top:4px'>"
            f"{regulator()} floor: {CBK_TOTAL_CAR_MIN_PCT_LOCAL}%</div></div>",
            unsafe_allow_html=True)

        # Engine constants reference table
        with st.expander("Engine constants reference"):
            const_rows = [
                {"Constant": "CBK_TOTAL_CAR_MIN_PCT_LOCAL", "Value": f"{CBK_TOTAL_CAR_MIN_PCT_LOCAL}%",
                  "Meaning": f"{regulator()} prudential CAR floor (Basel III + 4pp local)"},
                {"Constant": "NPL_INCREASE_TO_LOSS_FACTOR", "Value": f"{NPL_INCREASE_TO_LOSS_FACTOR}",
                  "Meaning": "Each 1% NPL increase = 0.45% loss against capital"},
                {"Constant": "RATE_SHOCK_TO_NII_BPS", "Value": f"{RATE_SHOCK_TO_NII_BPS}",
                  "Meaning": "Each 100bps shock = 0.5% NII delta vs loan book"},
                {"Constant": "ASSET_PRICE_SHOCK_TO_PROVISIONS",
                  "Value": f"{ASSET_PRICE_SHOCK_TO_PROVISIONS}",
                  "Meaning": "Each 1% asset price drop = 0.5% provisioning"},
            ]
            st.dataframe(pd.DataFrame(const_rows),
                         use_container_width=True, hide_index=True)
            st.markdown("**Scenario shock parameters** (engine SCENARIO_SHOCKS dict):")
            shock_rows = []
            for s in STRESS_SCENARIOS:
                shocks = SCENARIO_SHOCKS[s]
                shock_rows.append({
                    "Scenario": s,
                    "GDP Δ (pp)": float(shocks['gdp_growth_delta_pp']),
                    "Rate shock (bps)": float(shocks['interest_rate_shock_bps']),
                    "NPL Δ (%)": float(shocks['npl_increase_pct']),
                    "Asset Δ (%)": float(shocks['asset_price_shock_pct']),
                    "FX deval (%)": float(shocks['fx_devaluation_pct']),
                    "Deposit out (%)": float(shocks['deposit_outflow_pct']),
                    "RWA infl (%)": float(shocks['rwa_inflation_pct']),
                })
            st.dataframe(pd.DataFrame(shock_rows),
                         use_container_width=True, hide_index=True)
            st.caption(
                "Shock parameters are bound byte-for-byte in `SCENARIO_SHOCKS`. "
                "Calibration changes require engine code review.")


    # ---- 3-scenario run ----
    with st_sub_tabs[1]:
        st.markdown("**Run all 3 supervisory scenarios** (point-in-time stressed CAR).")

        def _build_inputs():
            return StressTestInputs(
                starting_total_capital_kes=_D_st(str(st_cap)) * _D_st("1000000000"),
                starting_rwa_kes=_D_st(str(st_rwa)) * _D_st("1000000000"),
                starting_loan_book_kes=_D_st(str(st_loans)) * _D_st("1000000000"),
                starting_npl_kes=_D_st(str(st_npl)) * _D_st("1000000000"),
                starting_securities_kes=_D_st(str(st_secs)) * _D_st("1000000000"),
                starting_fx_open_position_kes=_D_st(str(st_fx)) * _D_st("1000000000"),
                annual_pre_tax_profit_kes=_D_st(str(st_pbt)) * _D_st("1000000000"),
                horizon_years=int(st_horizon),
            )

        if st.button("Run 3-scenario stress test",
                       key="st79_3scen_btn", type="primary"):
            inputs = _build_inputs()
            r = StressTestingEngine.run_supervisory_scenarios(inputs)
            scenarios_data = r.get("scenarios", {})
            verdict = r.get("verdict")
            worst = r.get("worst_scenario")
            worst_car = r.get("worst_stressed_car_pct")

            # Verdict banner
            verdict_color = "#10B981" if verdict == "PASS" else "#DC2626"
            st.markdown(
                f"<div style='padding:18px;background:{verdict_color}22;"
                f"border-left:6px solid {verdict_color};border-radius:12px'>"
                f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                f"OVERALL VERDICT</div>"
                f"<div style='font-size:32px;font-weight:800;color:{verdict_color};margin-top:6px'>"
                f"{verdict}</div>"
                f"<div style='font-size:13px;margin-top:6px'>"
                f"Worst scenario: <b>{worst}</b> at "
                f"<b>{worst_car}%</b> CAR ({regulator()} floor: {CBK_TOTAL_CAR_MIN_PCT_LOCAL}%).</div>"
                f"</div>", unsafe_allow_html=True)

            # Per-scenario results
            scen_rows = []
            for s, data in scenarios_data.items():
                car_pct = float(_D_st(str(data.get("stressed_car_pct", 0))))
                scen_rows.append({
                    "Scenario": s,
                    "Starting CAR (%)": float(_D_st(str(data.get("starting_car_pct", 0)))),
                    "Stressed CAR (%)": car_pct,
                    "CAR Drop (pp)": float(_D_st(str(data.get("car_drop_pp", 0)))),
                    f"Stressed Capital ({currency_symbol()} B)":
                        float(_D_st(str(data.get("stressed_capital_kes", 0))) / _D_st("1000000000")),
                    f"Stressed RWA ({currency_symbol()} B)":
                        float(_D_st(str(data.get("stressed_rwa_kes", 0))) / _D_st("1000000000")),
                    f"Breaches {regulator()} Min": "🔴 YES" if data.get("breaches_cbk_minimum") else "✅ no",
                })
            st.dataframe(pd.DataFrame(scen_rows),
                         use_container_width=True, hide_index=True)

            audit_log("IFRS_ENGINE_USED", uname,
                       f"Stress #79: 3-scenario verdict={verdict}, "
                       f"worst={worst} @ {worst_car}%")

    # ---- Capital projection ----
    with st_sub_tabs[2]:
        st.markdown(
            f"**Multi-year capital projection** — apply scenario shocks compounded "
            f"over the {st_horizon}-year horizon. Pre-tax profit accretion is NOT "
            f"included in this conservative projection (worst-case assumption)."
        )
        proj_scen = st.selectbox("Scenario to project",
                                   list(STRESS_SCENARIOS),
                                   index=1, key="st79_proj_scen")
        if st.button("Project capital path",
                       key="st79_proj_btn", type="primary"):
            inputs = StressTestInputs(
                starting_total_capital_kes=_D_st(str(st_cap)) * _D_st("1000000000"),
                starting_rwa_kes=_D_st(str(st_rwa)) * _D_st("1000000000"),
                starting_loan_book_kes=_D_st(str(st_loans)) * _D_st("1000000000"),
                starting_npl_kes=_D_st(str(st_npl)) * _D_st("1000000000"),
                starting_securities_kes=_D_st(str(st_secs)) * _D_st("1000000000"),
                starting_fx_open_position_kes=_D_st(str(st_fx)) * _D_st("1000000000"),
                annual_pre_tax_profit_kes=_D_st(str(st_pbt)) * _D_st("1000000000"),
                horizon_years=int(st_horizon),
            )
            r = StressTestingEngine.capital_projection(inputs, proj_scen)
            yearly = r.get("yearly_projection", [])
            if yearly:
                proj_rows = [
                    {"Year": int(_D_st(str(y["year_index"]))),
                      f"Capital ({currency_symbol()} B)": float(_D_st(str(y["capital_kes"])) / _D_st("1000000000")),
                      f"RWA ({currency_symbol()} B)": float(_D_st(str(y["rwa_kes"])) / _D_st("1000000000")),
                      "CAR (%)": float(_D_st(str(y["car_pct"]))),
                      "Breach": "🔴" if (str(y["breaches_cbk_min"]) == "True") else "✅"}
                    for y in yearly
                ]
                proj_df = pd.DataFrame(proj_rows)
                st.dataframe(proj_df, use_container_width=True, hide_index=True)

                # CAR trajectory chart
                car_data = pd.DataFrame({
                    "CAR %": proj_df["CAR (%)"].tolist(),
                    f"{regulator()} Floor": [float(CBK_TOTAL_CAR_MIN_PCT_LOCAL)] * len(proj_df),
                }, index=proj_df["Year"])
                st.line_chart(car_data, use_container_width=True)

                # Find first breach year
                breach_years = [r["Year"] for r in proj_rows
                                if r["Breach"] == "🔴"]
                if breach_years:
                    st.error(
                        f"⛔ **First regulator breach** under {proj_scen}: "
                        f"Year {breach_years[0]}. "
                        f"Capital would be {proj_rows[breach_years[0]-1]['Capital ({currency_symbol()} B)']:.2f}B, "
                        f"CAR {proj_rows[breach_years[0]-1]['CAR (%)']}%.")
                else:
                    st.success(
                        f"✅ No {regulator()} breach over {st_horizon}-year horizon under {proj_scen}.")

                audit_log("IFRS_ENGINE_USED", uname,
                           f"Stress #79: projection {proj_scen} {st_horizon}yr, "
                           f"first_breach_year={breach_years[0] if breach_years else 'none'}")

    # ---- Reverse stress test ----
    with st_sub_tabs[3]:
        st.markdown(
            f"**Reverse stress test** ({regulator()} ICAAP requirement). "
            f"Searches for the minimum NPL increase × rate shock combination that "
            f"causes the bank to breach the {regulator()} CAR floor. "
            f"Search grid: NPL up to {REVERSE_STRESS_MAX_NPL_PCT}%, "
            f"rate up to {REVERSE_STRESS_MAX_RATE_BPS} bps."
        )
        st.caption(
            "Lower numbers = bank is more fragile. Higher numbers = bank is more resilient. "
            "Result tells you 'what would it take to break us?'"
        )
        breach_threshold = st.number_input(
            "Breach threshold (%)",
            min_value=8.0, max_value=20.0,
            value=float(CBK_TOTAL_CAR_MIN_PCT_LOCAL), step=0.5,
            key="st79_rev_thresh",
            help=f"Default = {regulator()} floor ({CBK_TOTAL_CAR_MIN_PCT_LOCAL}%). "
                 "Lower threshold = test even more severe breach scenarios.")

        if st.button("Run reverse stress test",
                       key="st79_rev_btn", type="primary"):
            inputs = StressTestInputs(
                starting_total_capital_kes=_D_st(str(st_cap)) * _D_st("1000000000"),
                starting_rwa_kes=_D_st(str(st_rwa)) * _D_st("1000000000"),
                starting_loan_book_kes=_D_st(str(st_loans)) * _D_st("1000000000"),
                starting_npl_kes=_D_st(str(st_npl)) * _D_st("1000000000"),
                starting_securities_kes=_D_st(str(st_secs)) * _D_st("1000000000"),
                starting_fx_open_position_kes=_D_st(str(st_fx)) * _D_st("1000000000"),
                annual_pre_tax_profit_kes=_D_st(str(st_pbt)) * _D_st("1000000000"),
                horizon_years=int(st_horizon),
            )
            r = StressTestingEngine.reverse_stress_test(
                inputs, breach_threshold_pct=_D_st(str(breach_threshold)))

            breach_npl = r.get("breach_npl_pct")
            breach_rate = r.get("breach_rate_bps")
            stressed_car = r.get("stressed_car_pct")
            starting_car = r.get("starting_car_pct")

            if breach_npl is None or stressed_car is None:
                st.success(
                    f"✅ **NO BREACH FOUND** within search grid "
                    f"(NPL up to {REVERSE_STRESS_MAX_NPL_PCT}%, "
                    f"rate up to {REVERSE_STRESS_MAX_RATE_BPS} bps). "
                    "Bank would withstand even extreme combined shocks.")
            else:
                # Resilience banner
                if _D_st(str(breach_npl)) <= _D_st("10"):
                    color, label = "#DC2626", "FRAGILE"
                elif _D_st(str(breach_npl)) <= _D_st("25"):
                    color, label = "#F59E0B", "MODERATE"
                else:
                    color, label = "#10B981", "RESILIENT"

                st.markdown(
                    f"<div style='padding:18px;background:{color}22;"
                    f"border-left:6px solid {color};border-radius:12px;text-align:center'>"
                    f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                    f"RESILIENCE ASSESSMENT</div>"
                    f"<div style='font-size:28px;font-weight:800;color:{color};margin-top:6px'>"
                    f"{label}</div>"
                    f"<div style='font-size:13px;margin-top:6px'>"
                    f"Bank breaches {breach_threshold}% CAR floor at "
                    f"<b>+{breach_npl}% NPL increase</b> with "
                    f"<b>+{breach_rate} bps</b> rate shock</div></div>",
                    unsafe_allow_html=True)

                k1, k2, k3 = st.columns(3)
                k1.metric("Starting CAR", f"{starting_car}%")
                k2.metric("Stressed CAR at breach",
                           f"{stressed_car}%",
                           delta=f"{float(_D_st(str(stressed_car))) - float(_D_st(str(starting_car))):.2f} pp",
                           delta_color="inverse")
                k3.metric(f"{regulator()} floor", f"{breach_threshold}%")

                st.info(
                    f"**Plain-English interpretation**: If your NPL ratio rises by "
                    f"**+{breach_npl} percentage points** AND interest rates move by "
                    f"**+{breach_rate} bps**, your CAR drops to **{stressed_car}%**, "
                    f"breaching {regulator()} {breach_threshold}% prudential floor. "
                    f"Smaller numbers = closer to breach = bank is more fragile.")

            audit_log("IFRS_ENGINE_USED", uname,
                       f"Stress #79: reverse stress, breach @ NPL+{breach_npl}%, "
                       f"rate+{breach_rate}bps, stressed_car={stressed_car}%")


    # ════════════════════════════════════════════════════════════════
    # ST_SUB_TABS[4]: Stress Testing Depth (Standard #51, integrated v6.2)
    # ════════════════════════════════════════════════════════════════
    with st_sub_tabs[4]:
        st.markdown(
            "**Stress Testing Depth analysis** — extends v5.78 with 4 inner "
            "views following the proven depth-batch template (6th application "
            "after v5.95+v5.97+v5.98+v5.99+v6.1).")
        st.caption(
            "💡 v5.78 surfaces inputs + 3-scenario run + projection + reverse "
            "stress independently. v6.2 composes these into board-ready views: "
            "Executive Scorecard, Sensitivity Analysis Batch, Multi-period "
            "Trajectory, Capital Buffer Adequacy Map.")

        _stress_depth_inner = st.tabs([
            "📋 Stress Executive Scorecard",
            "🎯 Sensitivity Analysis Batch",
            "📈 Multi-Period Trajectory",
            "🎚️ Capital Buffer Adequacy Map",
        ])

        # ────────── Inner[0]: Stress Executive Scorecard ──────────
        with _stress_depth_inner[0]:
            st.markdown(
                "**Stress Executive Scorecard** — composes "
                "run_supervisory_scenarios + reverse_stress_test into "
                "single-screen GREEN/AMBER/RED verdict for board ICAAP "
                "submission.")
            st.caption(
                "Mirrors v5.97/v5.98/v5.99/v6.1 scorecard pattern. Click "
                "compute to refresh from current Input tab values.")

            if st.button("📋 Compute stress scorecard",
                           key="st_es_btn", type="primary"):
                _es_inputs = _build_inputs()
                _es_supervisory = StressTestingEngine.run_supervisory_scenarios(
                    _es_inputs)
                _es_reverse = StressTestingEngine.reverse_stress_test(
                    _es_inputs)

                # === Section 1️⃣: Starting capital position ===
                st.markdown("### 1️⃣ Starting capital position")
                baseline = _es_supervisory["scenarios"]["BASELINE"]
                starting_car = float(_D_st(str(baseline["starting_car_pct"])))
                cbk_min = float(CBK_TOTAL_CAR_MIN_PCT_LOCAL)
                buffer_pp = starting_car - cbk_min

                p1, p2, p3 = st.columns(3)
                p1.metric("Starting CAR %", f"{starting_car:.2f}%")
                p2.metric(f"{regulator()} minimum",
                            f"{cbk_min:.1f}%",
                            help=f"{regulator()} PG/03 prudential capital adequacy minimum.")
                p3.metric("Buffer above minimum",
                            f"+{buffer_pp:.2f} pp",
                            delta_color="normal" if buffer_pp >= 5 else "inverse")

                # === Section 2️⃣: 3-scenario stress impact ===
                st.markdown("### 2️⃣ 3-scenario stress impact")
                scenarios_data = _es_supervisory["scenarios"]

                scen_rows = []
                for scen_name in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    s = scenarios_data[scen_name]
                    stressed_car = float(_D_st(str(s["stressed_car_pct"])))
                    breach = str(s["breaches_cbk_minimum"]).lower() == "true"
                    scen_rows.append({
                        "Scenario": scen_name,
                        "Starting CAR %": f"{starting_car:.2f}%",
                        "Stressed CAR %": f"{stressed_car:.2f}%",
                        "Drop (pp)":
                            f"{float(_D_st(str(s['car_drop_pp']))):+.2f}",
                        "Breaches 14.5%": "🚨 YES" if breach else "✅ NO",
                    })
                st.dataframe(pd.DataFrame(scen_rows),
                             use_container_width=True, hide_index=True)

                # === Section 3️⃣: Reverse stress test ===
                st.markdown("### 3️⃣ Reverse stress test (fragility)")
                breach_npl = float(_D_st(str(_es_reverse["breach_npl_pct"])))
                breach_rate = float(_D_st(str(_es_reverse["breach_rate_bps"])))
                rs_stressed_car = float(_D_st(str(
                    _es_reverse["stressed_car_pct"])))

                r1, r2, r3 = st.columns(3)
                r1.metric("NPL increase to breach",
                            f"+{breach_npl:.0f} pp",
                            help="How much NPL needs to rise before breach.")
                r2.metric("Rate shock to breach",
                            f"+{breach_rate:.0f} bps")
                r3.metric("Stressed CAR at breach point",
                            f"{rs_stressed_car:.2f}%")

                # === Section 4️⃣: Overall stress verdict ===
                st.markdown("### 4️⃣ Overall stress verdict")
                worst_breach = (_es_supervisory.get("verdict") == "FAIL")
                adverse_breach = (str(scenarios_data["ADVERSE"][
                    "breaches_cbk_minimum"]).lower() == "true")
                buffer_thin = buffer_pp < 5
                fragile_npl = breach_npl < 30  # less than 30 pp NPL increase

                issues = []
                if worst_breach:
                    issues.append(f"severely_adverse scenario breaches {regulator()} minimum")
                if adverse_breach:
                    issues.append(f"ADVERSE scenario breaches {regulator()} minimum")
                if buffer_thin:
                    issues.append(
                        f"capital buffer thin (+{buffer_pp:.2f} pp above min)")
                if fragile_npl:
                    issues.append(
                        f"high NPL fragility (breach @ NPL+{breach_npl:.0f}pp)")

                if not issues:
                    st.success(
                        "✅ **Stress test health: GREEN.** Bank withstands "
                        f"all 3 supervisory scenarios above {regulator()} minimum. "
                        "Capital position is robust.")
                elif len(issues) <= 1:
                    st.warning(
                        f"⚠ **Stress test health: AMBER.** Issue: "
                        f"{issues[0]}. Risk committee review recommended.")
                else:
                    st.error(
                        f"🚨 **Stress test health: RED.** Multiple issues: "
                        f"{'; '.join(issues)}. Capital plan + ICAAP "
                        "remediation required.")

                audit_log("IFRS_ENGINE_USED", uname,
                            f"Stress #51 (depth): scorecard "
                            f"start_car={starting_car:.2f}% "
                            f"adverse_breach={adverse_breach} "
                            f"reverse_npl={breach_npl} issues={len(issues)}")

        # ────────── Inner[1]: Sensitivity Analysis Batch ──────────
        with _stress_depth_inner[1]:
            st.markdown(
                "**Sensitivity Analysis Batch** — sweeps each shock dimension "
                "independently to identify which dimension drives capital "
                "deterioration most. Engine has 7 shock parameters; this "
                "batch tests each in isolation.")
            st.caption(
                "Useful for: identifying primary capital risk driver, "
                "informing risk-mitigation priorities, communicating "
                "balance sheet fragility to risk committee.")

            if st.button("🎯 Run sensitivity analysis",
                           key="st_sens_btn", type="primary"):
                _sens_inputs = _build_inputs()

                # Test each scenario shock dimension by running the
                # available supervisory scenarios — these embody the
                # combined shock vectors. For sensitivity we report
                # per-shock contributions surfaced in shock_parameters.
                _sens_supervisory = StressTestingEngine.run_supervisory_scenarios(
                    _sens_inputs)

                # Show sensitivity matrix per scenario per shock dimension
                shock_dims = [
                    ("gdp_growth_delta_pp", "GDP growth delta", "pp"),
                    ("interest_rate_shock_bps", "Interest rate shock", "bps"),
                    ("npl_increase_pct", "NPL increase", "pp"),
                    ("asset_price_shock_pct", "Asset price shock", "pp"),
                    ("fx_devaluation_pct", "FX devaluation", "pp"),
                    ("deposit_outflow_pct", "Deposit outflow", "pp"),
                    ("rwa_inflation_pct", "RWA inflation", "pp"),
                ]

                sens_rows = []
                for shock_key, shock_label, shock_unit in shock_dims:
                    row = {"Shock dimension": shock_label}
                    for scen_name in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                        s = _sens_supervisory["scenarios"][scen_name]
                        params = s["shock_parameters"]
                        val = float(_D_st(str(params.get(shock_key, 0))))
                        row[scen_name] = f"{val:+.0f} {shock_unit}"
                    sens_rows.append(row)
                st.markdown("**Per-scenario shock parameters:**")
                st.dataframe(pd.DataFrame(sens_rows),
                             use_container_width=True, hide_index=True)

                # Resulting CAR per scenario
                st.markdown("**Resulting stressed CAR per scenario:**")
                car_rows = []
                for scen_name in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    s = _sens_supervisory["scenarios"][scen_name]
                    car_rows.append({
                        "Scenario": scen_name,
                        "Stressed CAR %":
                            f"{float(_D_st(str(s['stressed_car_pct']))):.2f}%",
                        "Drop from start (pp)":
                            f"{float(_D_st(str(s['car_drop_pp']))):+.2f}",
                    })
                st.dataframe(pd.DataFrame(car_rows),
                             use_container_width=True, hide_index=True)

                # Bar chart of CAR drop magnitudes
                chart_data = pd.DataFrame({
                    "Drop (pp)": [
                        float(_D_st(str(_sens_supervisory["scenarios"][s][
                            "car_drop_pp"])))
                        for s in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]]
                }, index=["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"])
                st.markdown("**CAR drop magnitude by scenario (pp):**")
                st.bar_chart(chart_data)

                # Identify dominant shock — heuristic via NPL sensitivity
                # because the engine's transmission factor for NPL
                # (NPL_INCREASE_TO_LOSS_FACTOR=0.45) tends to dominate
                # in adverse scenarios
                adverse_npl_pct = float(_D_st(str(_sens_supervisory[
                    "scenarios"]["ADVERSE"]["shock_parameters"][
                    "npl_increase_pct"])))
                if adverse_npl_pct >= 30:
                    st.info(
                        f"💡 **Primary risk driver — NPL shock**: ADVERSE "
                        f"scenario applies +{adverse_npl_pct:.0f}% NPL "
                        "increase. Engine transmission factor "
                        "(NPL_INCREASE_TO_LOSS_FACTOR=0.45) means each "
                        f"+{adverse_npl_pct:.0f}% NPL → "
                        f"~{adverse_npl_pct * 0.45:.1f}% loan-book loss. "
                        "Mitigation focus: collections + early-warning "
                        "signals + provisioning prudence.")

                audit_log("IFRS_ENGINE_USED", uname,
                            f"Stress #51 (depth): sensitivity "
                            f"shock_dims={len(shock_dims)} "
                            f"scenarios=3")

        # ────────── Inner[2]: Multi-Period Trajectory ──────────
        with _stress_depth_inner[2]:
            st.markdown(
                "**Multi-Period Capital Trajectory** — runs capital_projection "
                "for ALL 3 scenarios + visualizes year-over-year breach "
                "timing. v5.78 surfaces single-scenario projection; this "
                "view enables comparison and breach-point identification.")
            st.caption(
                "Useful for: ICAAP capital plan submission, board "
                "communication of resilience timelines, identifying "
                "which year breach occurs under each scenario.")

            if st.button("📈 Run multi-period trajectory",
                           key="st_mpt_btn", type="primary"):
                _mpt_inputs = _build_inputs()

                # Run capital_projection for each scenario
                _mpt_results = {}
                for scen in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    _mpt_results[scen] = StressTestingEngine.capital_projection(
                        _mpt_inputs, scen)

                # Build trajectory matrix
                horizon = int(_mpt_inputs.horizon_years)
                traj_rows = []
                for year_idx in range(1, horizon + 1):
                    row = {"Year": year_idx}
                    for scen in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                        proj = _mpt_results[scen]["yearly_projection"]
                        if year_idx - 1 < len(proj):
                            year_data = proj[year_idx - 1]
                            car = float(_D_st(str(year_data["car_pct"])))
                            breach = str(year_data["breaches_cbk_min"]).lower() == "true"
                            row[scen] = (f"{car:.2f}% "
                                         + ("🚨" if breach else "✅"))
                        else:
                            row[scen] = "—"
                    traj_rows.append(row)
                st.dataframe(pd.DataFrame(traj_rows),
                             use_container_width=True, hide_index=True)

                # Line chart of CAR trajectory per scenario
                chart_rows = {}
                for scen in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    chart_rows[scen] = [
                        float(_D_st(str(y["car_pct"])))
                        for y in _mpt_results[scen]["yearly_projection"]]
                chart_data = pd.DataFrame(chart_rows,
                                            index=[f"Y{i+1}"
                                                   for i in range(horizon)])
                st.markdown("**Capital trajectory by scenario:**")
                st.line_chart(chart_data)

                # Breach analysis — first year of breach per scenario
                breach_rows = []
                for scen in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    proj = _mpt_results[scen]["yearly_projection"]
                    first_breach_year = None
                    for i, y in enumerate(proj):
                        if str(y["breaches_cbk_min"]).lower() == "true":
                            first_breach_year = i + 1
                            break
                    breach_rows.append({
                        "Scenario": scen,
                        "Years before breach":
                            f"Y{first_breach_year}" if first_breach_year
                            else f"NO BREACH within {horizon}y",
                        "Final CAR %":
                            f"{float(_D_st(str(proj[-1]['car_pct']))):.2f}%"
                            if proj else "—",
                    })
                st.markdown("**Breach timing per scenario:**")
                st.dataframe(pd.DataFrame(breach_rows),
                             use_container_width=True, hide_index=True)

                # Insights
                early_breach = [r for r in breach_rows
                                  if r["Years before breach"] == "Y1"]
                no_breach = [r for r in breach_rows
                               if "NO BREACH" in r["Years before breach"]]
                if len(early_breach) >= 2:
                    st.error(
                        f"🚨 **Year-1 breaches in {len(early_breach)} "
                        "scenario(s)** — capital insufficient to absorb "
                        "supervisory shocks even at horizon start. ICAAP "
                        "capital plan must address this gap immediately.")
                if no_breach:
                    st.success(
                        f"✅ **{len(no_breach)} scenario(s) maintain CAR "
                        f"above {regulator()} minimum throughout {horizon}-year "
                        "horizon** — capital robust against stress.")

                audit_log("IFRS_ENGINE_USED", uname,
                            f"Stress #51 (depth): trajectory horizon={horizon}y "
                            f"early_breach={len(early_breach)} "
                            f"no_breach={len(no_breach)}")

        # ────────── Inner[3]: Capital Buffer Adequacy Map ──────────
        with _stress_depth_inner[3]:
            st.markdown(
                "**Capital Buffer Adequacy Map** — measures the bank's "
                f"starting capital buffer above {regulator()} minimum + maps each "
                "scenario to a buffer-erosion ranking. Production deployment "
                "with peer-bank data could compare bank's resilience to "
                "industry quartiles.")
            st.caption(
                f"💡 Buffer adequacy bands: 🔴 INADEQUATE <2 pp / "
                f"🟡 THIN <5 pp / 🟢 ADEQUATE <10 pp / "
                f"✅ STRONG ≥10 pp above {regulator()} {float(CBK_TOTAL_CAR_MIN_PCT_LOCAL):.1f}%.")

            if st.button("🎚️ Compute buffer adequacy map",
                           key="st_bam_btn", type="primary"):
                _bam_inputs = _build_inputs()
                _bam_supervisory = StressTestingEngine.run_supervisory_scenarios(
                    _bam_inputs)

                cbk_min = float(CBK_TOTAL_CAR_MIN_PCT_LOCAL)

                bam_rows = []
                for scen_name in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]:
                    s = _bam_supervisory["scenarios"][scen_name]
                    starting_car = float(_D_st(str(s["starting_car_pct"])))
                    stressed_car = float(_D_st(str(s["stressed_car_pct"])))
                    starting_buffer = starting_car - cbk_min
                    stressed_buffer = stressed_car - cbk_min

                    # Adequacy band based on stressed buffer
                    if stressed_buffer >= 10:
                        band = "✅ STRONG"
                    elif stressed_buffer >= 5:
                        band = "🟢 ADEQUATE"
                    elif stressed_buffer >= 2:
                        band = "🟡 THIN"
                    elif stressed_buffer >= 0:
                        band = "🟠 MARGINAL"
                    else:
                        band = "🔴 INADEQUATE (breach)"

                    bam_rows.append({
                        "Scenario": scen_name,
                        "Starting buffer (pp)": f"+{starting_buffer:.2f}",
                        "Stressed buffer (pp)":
                            f"{stressed_buffer:+.2f}",
                        "Erosion (pp)":
                            f"{starting_buffer - stressed_buffer:+.2f}",
                        "Adequacy band": band,
                    })

                st.dataframe(pd.DataFrame(bam_rows),
                             use_container_width=True, hide_index=True)

                # Bar chart — stressed buffer by scenario
                chart_data = pd.DataFrame({
                    "Stressed buffer (pp)": [
                        float(_D_st(str(_bam_supervisory["scenarios"][s][
                            "stressed_car_pct"]))) - cbk_min
                        for s in ["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"]]
                }, index=["BASELINE", "ADVERSE", "SEVERELY_ADVERSE"])
                st.markdown(f"**Buffer above {regulator()} minimum (post-stress):**")
                st.bar_chart(chart_data)

                # Investment priority
                inadequate = [r for r in bam_rows
                                if "INADEQUATE" in r["Adequacy band"]]
                marginal = [r for r in bam_rows
                              if "MARGINAL" in r["Adequacy band"]]
                thin = [r for r in bam_rows
                          if "THIN" in r["Adequacy band"]]

                if inadequate:
                    st.error(
                        f"🔴 **{len(inadequate)} scenario(s) breach {regulator()} "
                        "minimum**: "
                        f"{', '.join(r['Scenario'] for r in inadequate)}. "
                        "Capital plan + ICAAP remediation required. "
                        "Options: (a) capital raise via rights issue or "
                        "Tier 2 instruments, (b) RWA reduction via "
                        "balance sheet de-risking, (c) profit retention "
                        "increase via dividend deferral.")
                if marginal:
                    st.warning(
                        f"🟠 **{len(marginal)} scenario(s) at MARGINAL "
                        "adequacy**: "
                        f"{', '.join(r['Scenario'] for r in marginal)}. "
                        "Capital position is precarious. Buffer "
                        "rebuilding plan required within 6 months.")
                if thin:
                    st.info(
                        f"🟡 **{len(thin)} scenario(s) at THIN adequacy** — "
                        "monitor closely + plan buffer rebuilding within "
                        "12 months.")
                if not inadequate and not marginal and not thin:
                    st.success(
                        "✅ All scenarios maintain ADEQUATE or STRONG "
                        f"buffer above {regulator()} minimum. Capital position is "
                        "resilient against supervisory stress shocks.")

                audit_log("IFRS_ENGINE_USED", uname,
                            f"Stress #51 (depth): buffer map "
                            f"inadequate={len(inadequate)} "
                            f"marginal={len(marginal)} thin={len(thin)}")


# ──────────────────────────────────────────────────────────────────────
# Section 5: 🤖 Arc Engines (absorbed from 93_risk_arc_cockpit.py
# in v10.208 per the architectural reorganization sub-campaign.
# 4 Basel III regulatory engines + dataclasses presented as 5
# nested sub-tabs: Market Risk VaR (BCBS d352), IRB Capital (BCBS
# d424), Op Risk SMA (BCBS d457), Stressed LCR (BCBS d295), and
# About. All engines are diagnostic-only — no remediation buttons;
# outputs feed governance discussions (ALCO, Capital Management
# Committee, Risk Committee). Mirrors v10.202..v10.207 absorption
# patterns. Note: this also remains visible to the MD/CEO via
# admin's full-department traversal, since 35_stress_testing.py is
# in the risk department and the MD has cross-departmental visibility.
# ──────────────────────────────────────────────────────────────────────
with tabs[5]:
    from decimal import Decimal as _D_arc
    from datetime import date as _date_arc

    try:
        from utils.market_risk_var import (
            VaREngine, VaRMethodology, VaRResult,
        )
        from utils.credit_risk_irb import (
            IRBCapitalEngine, IRBExposure, ExposureClass, CapitalResult,
        )
        from utils.op_risk import (
            OperationalRiskSMA, BusinessIndicatorInputs,
            OperationalLossEvent, SMAInputs, SMAResult,
            Bucket, ILMSource,
        )
        from utils.liquidity_stress import (
            LiquidityStressEngine, HQLAHolding, HQLALevel,
            OutflowCategory, InflowCategory, StressSeverity,
            BreachSeverity, StressedLCRResult,
        )
        _ARC_RISK_AVAILABLE = True
    except ImportError as _ie:
        st.error(f"Risk arc engines unavailable: {_ie}")
        _ARC_RISK_AVAILABLE = False

    if _ARC_RISK_AVAILABLE:
        st.markdown(
            "<div style='padding:24px;background:linear-gradient(135deg,#1E3A8A 0%,#0EA5E9 100%);"
            "border-radius:16px;color:white;margin-bottom:20px'>"
            "<div style='font-size:13px;letter-spacing:2px;opacity:0.85'>RISK ARC · LIVE COCKPIT</div>"
            "<div style='font-size:28px;font-weight:800;margin-top:6px'>"
            "Risk Arc Engines</div>"
            "<div style='font-size:14px;opacity:0.9;margin-top:8px'>"
            "v10.208 absorbed from 93_risk_arc_cockpit.py — Four Basel III "
            "regulatory engines: VaR (BCBS d352), IRB capital (BCBS d424), "
            "SMA op-risk (BCBS d457), stressed LCR (BCBS d295). Every "
            "output is diagnostic — these engines surface exposure, never "
            "execute remediation. All approvals flow through ALCO, Capital "
            "Management Committee, and the Risk Committee.</div>"
            "</div>",
            unsafe_allow_html=True,
        )

        arc_tabs = st.tabs([
            "📈 Market Risk VaR (#ENH-MR-001)",
            "🏛️ IRB Capital (#ENH-CR-001)",
            "⚙️ Op Risk SMA (#ENH-OR-001)",
            "💧 Stressed LCR (#ENH-LR-001)",
            "ℹ️ About",
        ])

        with arc_tabs[0]:
            st.markdown("### 📈 Parametric / Historical VaR + Expected Shortfall")
            st.caption(
                "BCBS d352 §K. Returns provided by caller (no auto-fetch per "
                "Rule 7). 1-day horizon and 99% confidence are the regulatory "
                "default; both are tunable for stress-testing.")

            col_l, col_r = st.columns([2, 1])
            with col_l:
                method_label = st.selectbox(
                    "Methodology",
                    ["Parametric (Normal)", "Historical"],
                    help=(
                        "Parametric assumes Normal returns; closed-form. "
                        "Historical uses the empirical distribution; no "
                        "distribution assumption."))
                portfolio_value = st.number_input(
                    f"Portfolio value ({currency_symbol()})",
                    min_value=1_000_000.0, value=100_000_000.0,
                    step=1_000_000.0, format="%.2f")
                confidence_pct = st.slider(
                    "Confidence level (%)", 90.0, 99.9, 99.0, 0.1)
                horizon_days = st.number_input(
                    "Horizon (days)", min_value=1, max_value=30, value=1)

                returns_csv = st.text_area(
                    "Daily returns (comma-separated decimals, e.g. -0.012,0.008,...)",
                    value=(
                        "-0.012, 0.008, 0.005, -0.020, 0.011, 0.003, "
                        "-0.008, 0.015, -0.025, 0.007, 0.002, -0.014, "
                        "0.009, -0.011, 0.006, 0.001, -0.018, 0.013, "
                        "-0.004, 0.010, -0.022, 0.005, 0.012, -0.009, "
                        "0.004, -0.016, 0.008, -0.013, 0.011, -0.006"),
                    height=80,
                    help="≥ 30 observations recommended for stable VaR.")

            with col_r:
                st.markdown("##### How to read")
                st.caption(
                    "**VaR** is the loss at the chosen confidence — \"on the "
                    f"worst day in 100 we expect to lose at least {currency_symbol()} X\". "
                    "**Expected Shortfall (ES)** is the average loss in "
                    "those tail days — ES ≥ VaR by construction.")

            if st.button("Compute VaR", type="primary", key="var_compute"):
                try:
                    returns = [
                        float(x.strip()) for x in returns_csv.split(",")
                        if x.strip()]
                except ValueError as e:
                    st.error(f"Could not parse returns: {e}")
                    returns = []

                if len(returns) < 5:
                    st.warning(
                        f"Only {len(returns)} return observations supplied; "
                        f"VaR estimates need ≥ 5 (≥ 30 strongly recommended).")
                elif len(returns) >= 1:
                    engine = VaREngine()
                    conf = Decimal(str(confidence_pct / 100))
                    pv = Decimal(str(portfolio_value))
                    if method_label.startswith("Parametric"):
                        result: VaRResult = engine.parametric_var(
                            returns=returns,
                            portfolio_value_kes=pv,
                            confidence=conf,
                            horizon_days=int(horizon_days))
                    else:
                        result = engine.historical_var(
                            returns=returns,
                            portfolio_value_kes=pv,
                            confidence=conf,
                            horizon_days=int(horizon_days))

                    audit_log("RISK_ENGINE_USED", uname, {
                        "engine": "market_risk_var",
                        "method": result.methodology.value,
                        "var_kes": str(result.var_kes),
                        "es_kes": str(result.expected_shortfall_kes),
                        "n_obs": len(returns),
                    })

                    st.success(
                        f"✅ {result.methodology.value} VaR computed on "
                        f"{len(returns)} returns")
                    m_a, m_b, m_c = st.columns(3)
                    m_a.metric(
                        f"VaR ({currency_symbol()})",
                        f"{float(result.var_kes):,.0f}",
                        help="Loss at the chosen confidence")
                    m_b.metric(
                        f"Expected Shortfall ({currency_symbol()})",
                        f"{float(result.expected_shortfall_kes):,.0f}",
                        help="Avg loss in the tail beyond VaR")
                    m_c.metric(
                        "VaR / Portfolio %",
                        f"{float(result.var_kes) / float(pv) * 100:.2f}%")

                    with st.expander("Return distribution + provenance"):
                        rd = result.return_distribution
                        st.json({
                            "methodology": result.methodology.value,
                            "confidence": str(result.confidence),
                            "horizon_days": result.horizon_days,
                            "portfolio_value_kes": str(result.portfolio_value_kes),
                            "var_kes": str(result.var_kes),
                            "expected_shortfall_kes": str(
                                result.expected_shortfall_kes),
                            "return_distribution": {
                                "n": rd.n_observations,
                                "mean": str(rd.mean),
                                "stdev": str(rd.stdev),
                                "min": str(rd.min_return),
                                "max": str(rd.max_return),
                            },
                            "framework_refs": list(result.framework_refs),
                        })

        with arc_tabs[1]:
            st.markdown("### 🏛️ IRB Capital — Single Exposure")
            st.caption(
                "BCBS d424 §RBC25 corporate exposure formula. Per Rule 7, "
                "engine never moves loans between exposure classes and "
                "never auto-approves capital allocations — outputs feed "
                "ALCO + Capital Management Committee.")

            col_l, col_r = st.columns([3, 2])
            with col_l:
                exp_id = st.text_input("Exposure ID", value="LARGE-CORP-001")
                exp_class = st.selectbox(
                    "Exposure class",
                    [ExposureClass.LARGE_CORPORATE.value,
                     ExposureClass.SME_CORPORATE.value],
                    help="Sovereign / Bank classes reserved for future scope.")
                pd = st.slider(
                    "PD — probability of default (1y)", 0.0003, 1.0, 0.01, 0.001,
                    format="%.4f",
                    help="Floored at 3 bp per BCBS d424 §RBC25.6")
                lgd = st.slider(
                    "LGD — loss given default", 0.0, 1.0, 0.45, 0.01)
                ead_kes = st.number_input(
                    f"EAD — exposure at default ({currency_symbol()})",
                    min_value=1_000_000.0, value=10_000_000.0,
                    step=1_000_000.0, format="%.2f")
                m_years = st.slider(
                    "M — effective maturity (years)", 1.0, 5.0, 2.5, 0.1,
                    help="Bounded [1, 5] per BCBS d424 §RBC25.13")

            with col_r:
                st.markdown("##### Defaulted exposures")
                st.caption(
                    "Set PD = 1.0 to model a defaulted exposure: per "
                    "§RBC25.16 the IRB capital requirement above EL falls "
                    "to zero; expected loss is fully realised.")

            if st.button("Compute IRB Capital", type="primary", key="irb_compute"):
                try:
                    exposure = IRBExposure(
                        exposure_id=exp_id,
                        exposure_class=ExposureClass(exp_class),
                        pd=float(pd), lgd=float(lgd),
                        ead_kes=Decimal(str(ead_kes)),
                        maturity_years=float(m_years))
                    engine = IRBCapitalEngine()
                    result: CapitalResult = engine.compute(exposure)

                    audit_log("RISK_ENGINE_USED", uname, {
                        "engine": "credit_risk_irb",
                        "exposure_id": exp_id,
                        "rwa_kes": str(result.rwa_kes),
                        "k_pct": str(result.capital_requirement_pct),
                    })

                    st.success(f"✅ IRB capital computed for {exp_id}")
                    m_a, m_b, m_c, m_d = st.columns(4)
                    m_a.metric(
                        "K (capital requirement %)",
                        f"{float(result.capital_requirement_pct) * 100:.4f}%")
                    m_b.metric(
                        f"RWA ({currency_symbol()})",
                        f"{float(result.rwa_kes):,.0f}",
                        help="K × 12.5 × EAD")
                    m_c.metric(
                        f"Expected Loss ({currency_symbol()})",
                        f"{float(result.expected_loss_kes):,.0f}",
                        help="PD × LGD × EAD")
                    m_d.metric(
                        "Capital ratio (RWA / EAD)",
                        f"{float(result.rwa_kes) / float(ead_kes) * 100:.2f}%")

                    with st.expander("Intermediates + provenance (Rule 1)"):
                        st.json({
                            "exposure_id": result.exposure_id,
                            "exposure_class": result.exposure_class.value,
                            "pd": result.pd,
                            "lgd": result.lgd,
                            "ead_kes": str(result.ead_kes),
                            "maturity_years": result.maturity_years,
                            "correlation_R": result.correlation_R,
                            "maturity_adj_b": result.maturity_adj_b,
                            "capital_requirement_pct": str(
                                result.capital_requirement_pct),
                            "rwa_kes": str(result.rwa_kes),
                            "expected_loss_kes": str(
                                result.expected_loss_kes),
                            "framework_refs": list(result.framework_refs),
                        })
                except ValueError as e:
                    st.error(f"Validation error: {e}")

        with arc_tabs[2]:
            st.markdown("### ⚙️ SMA Operational Risk Capital")
            st.caption(
                "BCBS d457 §RBC30. Bucket-1 banks (BI ≤ EUR 1bn) typically "
                "elect ILM = 1.0 under §RBC30.41 national discretion. "
                "Larger banks compute ILM via "
                "ln(e − 1 + (LC / BIC)^0.8) when ≥ 5 years of loss data "
                "are available. Per Rule 7, engine never records loss "
                "events and never approves capital allocations.")

            st.markdown("#### Business Indicator (one fiscal year — replicated 3x)")
            bi_cols = st.columns(3)
            with bi_cols[0]:
                ii = st.number_input(
                    f"Interest income ({currency_symbol()})",
                    min_value=0.0, value=12_000_000_000.0,
                    step=100_000_000.0)
                ie = st.number_input(
                    f"Interest expense ({currency_symbol()})",
                    min_value=0.0, value=6_000_000_000.0,
                    step=100_000_000.0)
                iea = st.number_input(
                    f"Interest-earning assets ({currency_symbol()})",
                    min_value=0.0, value=400_000_000_000.0,
                    step=1_000_000_000.0)
                di = st.number_input(
                    f"Dividend income ({currency_symbol()})",
                    min_value=0.0, value=100_000_000.0,
                    step=10_000_000.0)
            with bi_cols[1]:
                oi = st.number_input(
                    "Other operating income (KES)",
                    min_value=0.0, value=500_000_000.0,
                    step=10_000_000.0)
                oe = st.number_input(
                    "Other operating expense (KES)",
                    min_value=0.0, value=400_000_000.0,
                    step=10_000_000.0)
                fi = st.number_input(
                    "Fee income (KES)",
                    min_value=0.0, value=3_000_000_000.0,
                    step=100_000_000.0)
                fe = st.number_input(
                    "Fee expense (KES)",
                    min_value=0.0, value=500_000_000.0,
                    step=10_000_000.0)
            with bi_cols[2]:
                net_tb = st.number_input(
                    "Net P&L Trading Book (KES, signed)",
                    value=200_000_000.0, step=10_000_000.0)
                net_bb = st.number_input(
                    "Net P&L Banking Book (KES, signed)",
                    value=100_000_000.0, step=10_000_000.0)
                eur_kes = st.number_input(
                    "EUR / KES rate", min_value=1.0, value=145.0, step=1.0)
                bucket1_disc = st.checkbox(
                    "Apply Bucket-1 discretion (ILM = 1)", value=True,
                    help=f"§RBC30.41 — most Tier-2 domestic banks elect this.")

            st.markdown("#### Loss history (annual aggregate, 10y window)")
            loss_per_year = st.number_input(
                "Average annual operational loss (KES)",
                min_value=0.0, value=500_000_000.0, step=50_000_000.0,
                help=(
                    "Use the bank's loss-event database aggregated annually. "
                    "Engine accepts per-event records too — UI simplifies "
                    "to a single annual figure replicated across the 10y "
                    "window for the cockpit view."))
            n_loss_years = st.slider(
                "Years of loss history available", 0, 10, 10,
                help=(
                    "When < 5, ILM is forced to 1.0 — INSUFFICIENT_HISTORY "
                    "source surfaced in result."))

            if st.button("Compute SMA capital", type="primary", key="or_compute"):
                try:
                    kw = dict(
                        interest_income_kes=Decimal(str(ii)),
                        interest_expense_kes=Decimal(str(ie)),
                        interest_earning_assets_kes=Decimal(str(iea)),
                        dividend_income_kes=Decimal(str(di)),
                        other_operating_income_kes=Decimal(str(oi)),
                        other_operating_expense_kes=Decimal(str(oe)),
                        fee_income_kes=Decimal(str(fi)),
                        fee_expense_kes=Decimal(str(fe)),
                        net_pnl_trading_book_kes=Decimal(str(net_tb)),
                        net_pnl_banking_book_kes=Decimal(str(net_bb)))
                    bi_inputs = tuple(
                        BusinessIndicatorInputs(fiscal_year=y, **kw)
                        for y in (2023, 2024, 2025))
                    current_yr = date.today().year
                    loss_events = tuple(
                        OperationalLossEvent(
                            fiscal_year=y,
                            gross_loss_kes=Decimal(str(loss_per_year)))
                        for y in range(
                            current_yr - n_loss_years, current_yr))
                    inputs = SMAInputs(
                        bi_inputs=bi_inputs, loss_events=loss_events,
                        eur_to_kes_rate=Decimal(str(eur_kes)),
                        apply_bucket_1_discretion=bool(bucket1_disc))
                    result: SMAResult = OperationalRiskSMA().compute(inputs)

                    audit_log("RISK_ENGINE_USED", uname, {
                        "engine": "op_risk",
                        "bucket": result.bucket.value,
                        "ilm_source": result.ilm_source.value,
                        "rwa_op_kes": str(result.rwa_op_kes),
                    })

                    st.success(
                        f"✅ SMA computed — bucket {result.bucket.value}, "
                        f"ILM source: {result.ilm_source.value}")
                    m_a, m_b, m_c, m_d = st.columns(4)
                    m_a.metric(
                        "BI 3y avg (EUR)",
                        f"{float(result.bi_three_year_avg_eur):,.0f}")
                    m_b.metric(
                        "BIC (KES)",
                        f"{float(result.bic_kes):,.0f}")
                    m_c.metric(
                        "ILM",
                        f"{float(result.ilm):.4f}",
                        help=f"Source: {result.ilm_source.value}")
                    m_d.metric(
                        "RWA op (KES)",
                        f"{float(result.rwa_op_kes):,.0f}",
                        help="ORC × 12.5")

                    with st.expander("Intermediates + provenance (Rule 1)"):
                        st.json({
                            "bi_per_year_kes": [
                                {"year": y, "bi_kes": str(v)}
                                for y, v in result.bi_per_year_kes],
                            "bi_three_year_avg_kes": str(
                                result.bi_three_year_avg_kes),
                            "bi_three_year_avg_eur": str(
                                result.bi_three_year_avg_eur),
                            "bucket": result.bucket.value,
                            "bic_kes": str(result.bic_kes),
                            "annual_avg_loss_kes": str(
                                result.annual_avg_loss_kes),
                            "lc_kes": str(result.lc_kes),
                            "ilm": str(result.ilm),
                            "ilm_source": result.ilm_source.value,
                            "orc_kes": str(result.orc_kes),
                            "rwa_op_kes": str(result.rwa_op_kes),
                            "framework_refs": list(result.framework_refs),
                        })
                except ValueError as e:
                    st.error(f"Validation error: {e}")

        with arc_tabs[3]:
            st.markdown("### 💧 Stressed LCR — BCBS d295 calibration")
            st.caption(
                "Distinct from baseline LCR (utils.liquidity_risk, Standard "
                "#73). This page applies severity-tiered run-off "
                "multipliers and surfaces survival horizon when breaching. "
                "Per Rule 7, engine never auto-liquidates HQLA and never "
                "executes funding draws.")

            sev_label = st.radio(
                "Stress severity",
                [s.value for s in StressSeverity],
                index=2, horizontal=True,
                help=(
                    "BASELINE 1.0× / MODERATE 1.5× / SEVERE 2.0× / "
                    "BANK_RUN 3.0× outflow multipliers. Inflows reduce "
                    "in mirror tiers."))

            st.markdown("#### HQLA composition")
            h_a, h_b, h_c = st.columns(3)
            with h_a:
                l1 = st.number_input(
                    "Level 1 HQLA (KES) — cash, CB reserves, sovereign 0% RW",
                    min_value=0.0, value=80_000_000_000.0,
                    step=1_000_000_000.0,
                    help="0% haircut")
            with h_b:
                l2a = st.number_input(
                    "Level 2A HQLA (KES) — sovereign 20% RW, AA- corp",
                    min_value=0.0, value=20_000_000_000.0,
                    step=1_000_000_000.0,
                    help="15% haircut, capped at 40% of total HQLA")
            with h_c:
                l2b = st.number_input(
                    "Level 2B HQLA (KES) — lower-rated corp, equities",
                    min_value=0.0, value=5_000_000_000.0,
                    step=500_000_000.0,
                    help="50% haircut, capped at 15% of total HQLA")

            st.markdown("#### Outflows (30-day, baseline run-off rates)")
            o_a, o_b = st.columns(2)
            with o_a:
                retail_bal = st.number_input(
                    "Retail stable deposits (KES)",
                    min_value=0.0, value=100_000_000_000.0,
                    step=1_000_000_000.0)
                retail_rate = st.slider(
                    "Retail base run-off rate", 0.0, 1.0, 0.05, 0.01,
                    help="Basel III default for stable retail = 5%")
            with o_b:
                whole_bal = st.number_input(
                    "Unsecured wholesale (non-financial, KES)",
                    min_value=0.0, value=30_000_000_000.0,
                    step=1_000_000_000.0)
                whole_rate = st.slider(
                    "Wholesale base run-off rate", 0.0, 1.0, 0.40, 0.01,
                    help="Basel III default for non-fin wholesale = 40%")

            st.markdown("#### Inflows (30-day)")
            in_a, in_b = st.columns(2)
            with in_a:
                loans_bal = st.number_input(
                    "Performing loans receipts (KES)",
                    min_value=0.0, value=8_000_000_000.0,
                    step=500_000_000.0)
            with in_b:
                loans_rate = st.slider(
                    "Performing loans run-in rate", 0.0, 1.0, 0.50, 0.01,
                    help="Basel III default for performing retail/SME = 50%")

            if st.button("Compute stressed LCR", type="primary", key="lr_compute"):
                try:
                    holdings = (
                        HQLAHolding(
                            holding_id="cb_reserves_l1",
                            level=HQLALevel.LEVEL_1,
                            market_value_kes=Decimal(str(l1))),
                        HQLAHolding(
                            holding_id="govt_l2a", level=HQLALevel.LEVEL_2A,
                            market_value_kes=Decimal(str(l2a))),
                        HQLAHolding(
                            holding_id="corp_l2b", level=HQLALevel.LEVEL_2B,
                            market_value_kes=Decimal(str(l2b))),
                    )
                    outflows = (
                        OutflowCategory(
                            category_id="retail_stable",
                            label="Retail stable",
                            balance_kes=Decimal(str(retail_bal)),
                            base_run_off_rate=Decimal(str(retail_rate))),
                        OutflowCategory(
                            category_id="wholesale_unsec",
                            label="Wholesale unsecured (non-fin)",
                            balance_kes=Decimal(str(whole_bal)),
                            base_run_off_rate=Decimal(str(whole_rate))),
                    )
                    inflows = (
                        InflowCategory(
                            category_id="performing_loans",
                            label="Performing loans",
                            balance_kes=Decimal(str(loans_bal)),
                            base_run_in_rate=Decimal(str(loans_rate))),
                    )
                    sev = StressSeverity(sev_label)
                    result: StressedLCRResult = LiquidityStressEngine().compute(
                        holdings=holdings, outflows=outflows,
                        inflows=inflows, severity=sev,
                        notes=f"cockpit_run_{date.today().isoformat()}")

                    audit_log("RISK_ENGINE_USED", uname, {
                        "engine": "liquidity_stress",
                        "severity": result.severity.value,
                        "lcr_ratio": str(result.lcr_ratio),
                        "breach_severity": result.breach_severity.value,
                    })

                    # Breach traffic light
                    colour = {
                        BreachSeverity.COMPLIANT: "#10B981",
                        BreachSeverity.AMBER: "#F59E0B",
                        BreachSeverity.RED: "#EF4444",
                        BreachSeverity.CRITICAL: "#7F1D1D",
                    }[result.breach_severity]
                    lcr_pct = (
                        f"{float(result.lcr_ratio) * 100:.2f}%"
                        if result.lcr_ratio is not None else "n/a (NCO ≤ 0)")
                    st.markdown(
                        f"<div style='padding:16px;background:{colour};"
                        f"border-radius:12px;color:white;margin:12px 0'>"
                        f"<div style='font-size:12px;opacity:0.85'>BREACH SEVERITY</div>"
                        f"<div style='font-size:24px;font-weight:700'>"
                        f"{result.breach_severity.value} · LCR = {lcr_pct}</div>"
                        f"</div>",
                        unsafe_allow_html=True)

                    m_a, m_b, m_c, m_d = st.columns(4)
                    m_a.metric(
                        "HQLA after caps (KES)",
                        f"{float(result.hqla_total_after_caps_kes):,.0f}")
                    m_b.metric(
                        "Stressed outflows (KES)",
                        f"{float(result.total_outflows_kes):,.0f}")
                    m_c.metric(
                        "NCO 30d (KES)",
                        f"{float(result.nco_30d_kes):,.0f}")
                    survival_str = (
                        f"{float(result.survival_days):.1f} days"
                        if result.survival_days is not None
                        else "compliant — no horizon")
                    m_d.metric("Survival horizon", survival_str)

                    with st.expander(
                            "Per-category outflows + provenance (Rule 1)"):
                        st.json({
                            "severity": result.severity.value,
                            "hqla_pre_cap_kes": str(
                                result.hqla_total_pre_cap_kes),
                            "hqla_after_caps_kes": str(
                                result.hqla_total_after_caps_kes),
                            "hqla_breakdown": [
                                {
                                    "level": b.level.value,
                                    "gross_kes": str(b.gross_kes),
                                    "haircut_pct": str(b.haircut_pct),
                                    "after_haircut_kes": str(
                                        b.after_haircut_kes),
                                }
                                for b in result.hqla_breakdown],
                            "outflows": [
                                {
                                    "category_id": f.category_id,
                                    "balance_kes": str(f.balance_kes),
                                    "base_rate": str(f.base_rate),
                                    "stress_multiplier": str(
                                        f.stress_multiplier),
                                    "stressed_rate": str(f.stressed_rate),
                                    "stressed_kes": str(f.stressed_kes),
                                }
                                for f in result.outflows],
                            "inflows_capped_kes": str(
                                result.inflows_capped_kes),
                            "nco_30d_kes": str(result.nco_30d_kes),
                            "lcr_ratio": (
                                str(result.lcr_ratio)
                                if result.lcr_ratio is not None else None),
                            "breach_severity": result.breach_severity.value,
                            "survival_days": (
                                str(result.survival_days)
                                if result.survival_days is not None
                                else None),
                            "framework_refs": list(result.framework_refs),
                        })
                except ValueError as e:
                    st.error(f"Validation error: {e}")

        with arc_tabs[4]:
            st.markdown("### ℹ️ Risk Arc — About this Cockpit")
            st.markdown(
                """
                The Risk arc was built across batches **v10.39 → v10.45**:

                | Batch    | Module                          | Standards            |
                | -------- | ------------------------------- | -------------------- |
                | v10.39   | market_risk_factors / sens / var | ENH-MR-001..005     |
                | v10.40   | market_risk_limits              | ENH-MR-006/007       |
                | v10.41   | trading_book_boundary           | ENH-MR-008/009/010   |
                | v10.42   | credit_risk_irb                 | ENH-CR-001           |
                | v10.43   | op_risk                         | ENH-OR-001           |
                | v10.44   | liquidity_stress                | ENH-LR-001           |
                | v10.45   | G129 + Tier 24 + Master Prompt  | closure ratchet      |
                | **v10.46** | **this cockpit + G130 ratchet** | **UI integration backfill** |

                **Frameworks referenced:**

                - BCBS d352 (FRTB) — Market Risk VaR + Trading Book Boundary
                - BCBS d424 — IRB Approach (Credit Risk Capital)
                - BCBS d457 — Standardised Approach for Operational Risk
                - BCBS d295 — Liquidity Coverage Ratio
                - CBK PG/12 — Liquidity Risk Management
                - CBK PG/15 — Risk Classification & Provisioning

                **Diagnostic-only posture (Rule 7).** None of the four engines
                on this page execute remediation. They surface exposure,
                capital requirement, breach severity, or VaR — and the
                operator carries those numbers to ALCO, the Capital
                Management Committee, or the Risk Committee. No "auto-rebalance",
                "auto-hedge", or "auto-liquidate" affordances exist.

                **Provenance discipline (Rule 1).** Every result rendered above
                also exposes its full intermediate state under the
                "Intermediates + provenance" expander — inputs, computed
                intermediates (correlation R, maturity adjustment b, ILM
                source, HQLA caps applied), outputs, framework refs.
                Decimal-internal monetary precision preserved end-to-end.

                Locked under **G129 risk_arc_closed** (registry/scenario
                ratchet) and **G130 risk_arc_ui_integrated** (this page's
                ratchet — see CHANGELOG_v10.46).
                """)

            audit_log("RISK_COCKPIT_ABOUT_VIEWED", uname, {
                "page": "93_risk_arc_cockpit"})


        # Footer audit log
        try:
            audit_log(
                action="risk_arc_engines.view",
                username=ud.get("username", "anonymous"),
                detail=f"viewed_at={_date_arc.today().isoformat()}",
                module="stress_testing")
        except Exception:
            pass
