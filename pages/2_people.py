"""pages/2_people.py — People & HR Intelligence Module."""
import streamlit as st
from pathlib import Path
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
from utils.core import *
from pages._shared import load_shared_state
from pages._access import require_access, get_my_scope

def _safe_date(s, fallback=None):
    """Safe date parsing — returns fallback on invalid/None input."""
    try:
        from datetime import date as _d
        return _d.fromisoformat(str(s)) if s else (fallback or _d.today())
    except Exception:
        from datetime import date as _d
        return fallback or _d.today()


require_access("people_hr.dashboard")
DATA = Path(__file__).parent.parent / "data"


# Fallbacks in case core.py hasn't been updated yet on this machine
try:
    _ = LEAVE_TYPES
except NameError:
    LEAVE_TYPES = {"Annual Leave":{"days_entitled":21,"description":"21 days","paid":True,"affects_performance":False,"compensation":"pro_rata","color":"var(--brand-primary,#006B3F)"},"Sick Leave":{"days_entitled":14,"description":"14 days","paid":True,"affects_performance":True,"compensation":"exclude_month","color":"#E24B4A"}}
    EXIT_REASONS = ["Resignation — Better opportunity","Resignation — Salary/compensation","Dismissal — Gross misconduct","Dismissal — Performance","Retirement — Mandatory","Contract end — Not renewed","Mutual separation"]
    TRANSFER_REASONS = ["Performance improvement","Branch need","Staff request","Rotational development","Promotion transfer"]
    DISCIPLINARY_CATEGORIES = ["Gross misconduct","Insubordination","Absenteeism","Fraud","Policy violation","Performance negligence"]
    DISCIPLINARY_STAGES = ["Verbal warning","Written warning (1st)","Written warning (final)","Show cause notice","Suspension pending investigation","Disciplinary hearing scheduled","Disciplinary hearing held","Decision — Dismissed","Decision — Exonerated"]
    PIP_DURATIONS = [30, 60, 90]
    PIP_OUTCOMES = ["In progress","Successfully completed","Extended","Terminated — dismissal","Converted to final warning"]
    PIP_REVIEW_FREQUENCIES = ["Weekly","Fortnightly","Monthly"]


st.markdown(
    "<div style='padding:16px 0 4px'>"
    "<span style='font-size:22px;font-weight:800'>👥 People</span>"
    "<span style='font-size:13px;color:var(--color-text-secondary);margin-left:12px'>"
    "HR · Leave management · Performance · Succession planning</span></div>",
    unsafe_allow_html=True)

um, ud, uname, em, ri_pm, prod_m, pm, lm, hr_m, casc, vm, rlm = load_shared_state()
staff_scores  = st.session_state.get("staff_scores",  pd.DataFrame())
df_proc       = st.session_state.get("df_processed",  pd.DataFrame())
filtered      = st.session_state.get("filtered_staff", pd.DataFrame())
action_plans  = st.session_state.get("sbu_action_plans", {})

role_l  = str(ud.get("role","")).lower()
can_all = ud.get("can_view_all", False) or any(k in role_l for k in ("admin","director","md","hr"))

if len(staff_scores) == 0:
    st.markdown(
        "<div style='padding:40px;text-align:center;background:var(--brand-light,#E8F5EE);"
        "border-radius:12px;border:1px solid var(--brand-primary,#006B3F)33'>"
        "<div style='font-size:32px;margin-bottom:12px'>👥</div>"
        "<div style='font-size:18px;font-weight:500;color:var(--brand-primary,#006B3F)'>Upload BSC data to activate People module</div>"
        "</div>", unsafe_allow_html=True)
    st.stop()

# Header
st.markdown(
    "<div style=\'padding:16px 22px;background:#2C3E50;border-radius:12px;margin-bottom:20px;box-shadow:0 2px 12px rgba(0,0,0,0.15)\'><div style=\'display:flex;align-items:center;justify-content:space-between\'><div><div style=\'color:var(--color-background-primary);font-size:16px;font-weight:700;letter-spacing:-0.2px\'>People & HR Intelligence</div><div style=\'color:rgba(255,255,255,0.65);font-size:11px;margin-top:3px;font-weight:400\'>Leave · Exits · Disciplinary · PIP · Diligence scores</div></div><div style=\'opacity:0.12;font-size:36px;line-height:1;color:white\'>◆</div></div></div>",
    unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# v7.8: Workforce Health Composite (composite_scores.py surfacing)
# ─────────────────────────────────────────────────────────────────
with st.expander("📊 Workforce Health Composite (v6.0 / v7.8 surfaced)", expanded=False):
    from utils.composite_scores import workforce_health_composite

    st.caption(
        "v7.8 surfacing of `composite_scores.workforce_health_composite()` on this "
        "domain page (per Charter §13). Composes engagement + eNPS + weakest "
        "driver + flight risk into a single 0-100 score with HEALTHY / MODERATE / "
        "LOW severity bands. Inputs below are illustrative healthy-bank profile; "
        "production deployment will read live values from engagement surveys + "
        "succession_engine flight risk outputs."
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Inputs (illustrative healthy bank):**")
        wf_engagement = st.slider("Engagement score (0-100)", 0, 100, 78,
                                   key="wf_health_engagement")
        wf_enps = st.slider("eNPS (-100 to +100)", -100, 100, 35,
                            key="wf_health_enps")
        wf_weakest = st.slider("Weakest driver score (0-100)", 0, 100, 65,
                                key="wf_health_weakest")
        wf_flight = st.slider("Flight risk HIGH %", 0, 100, 8,
                               key="wf_health_flight")

    with c2:
        wf_result = workforce_health_composite(
            engagement_score=float(wf_engagement),
            enps=float(wf_enps),
            weakest_driver_score=float(wf_weakest),
            flight_risk_high_pct=float(wf_flight),
        )
        wf_score = wf_result.get("score")
        wf_severity = wf_result.get("severity")
        sev_color = {"HEALTHY": "✅", "MODERATE": "🟡",
                     "LOW": "🚨", "UNKNOWN": "⚠"}.get(wf_severity, "")
        st.metric("Workforce Health score",
                  f"{wf_score:.1f}/100" if wf_score is not None else "—",
                  wf_severity)
        st.markdown(f"**{sev_color} {wf_severity}**")

        if wf_result.get("components"):
            st.markdown("**Component scores:**")
            for k, v in wf_result["components"].items():
                st.markdown(f"- `{k}`: {v:.1f}")

# ─────────────────────────────────────────────────────────────────
# Restructured: 2-level navigation for clarity
# ─────────────────────────────────────────────────────────────────
sections = st.tabs([
    "📊 Insights",
    "👥 Records",
    "🏖️ Leave",
    "📋 Discipline & Dev",
    "🏆 Recognition",
    "🌿 Wellness",
])

# ── Section 0: 📊 Insights ─────────────────────────────
with sections[0]:
    sub = st.tabs([
        "📊 HR overview",
        "📈 Team insights",
        "⚖️ Compensation Equity (Standard #63)",
        "🎯 Engagement & Performance (Standard #64)",
        "🔮 Predictive Performance (Standards #20 + #21)",
    ])
    with sub[0]:
        total_staff  = len(staff_scores)
        on_leave     = len(lm.get_active_leave()) if lm else 0
        exits_12m    = len(hr_m.get_exits(12))    if hr_m else 0
        active_pips  = len(hr_m.get_active_pips()) if hr_m else 0
        disc_cases   = len(hr_m.get_active_cases()) if hr_m else 0
        at_risk      = int((staff_scores['Final_BSC_Score'] < 2.5).sum()) if len(staff_scores) else 0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total staff",     total_staff)
        c2.metric("Currently on leave", on_leave,
                  delta=f"{on_leave/total_staff*100:.1f}%" if total_staff else "0%")
        c3.metric("Exits (12 months)", exits_12m,
                  delta=f"-{exits_12m}" if exits_12m else "0", delta_color="inverse")
        c4.metric("Staff on PIP",    active_pips,
                  delta=f"-{active_pips}" if active_pips else "0", delta_color="inverse")
        c5.metric("Open disc. cases",disc_cases,
                  delta=f"-{disc_cases}" if disc_cases else "0", delta_color="inverse")
        c6.metric("At-risk BSC <2.5",at_risk,
                  delta=f"-{at_risk}" if at_risk else "0", delta_color="inverse")

        st.markdown("---")
        ov1, ov2 = st.columns(2)

        with ov1:
            # BSC distribution
            if len(staff_scores):
                bands = pd.DataFrame([
                    {'Band':'Exceeded (≥3.1)','Count':int((staff_scores['Final_BSC_Score']>=3.1).sum()),'Color':'var(--brand-primary,#006B3F)'},
                    {'Band':'Met (3.0–3.1)',  'Count':int(staff_scores['Final_BSC_Score'].between(3.0,3.1).sum()),'Color':'var(--brand-mid,#1D9E75)'},
                    {'Band':'Below (2.5–3.0)','Count':int(staff_scores['Final_BSC_Score'].between(2.5,3.0).sum()),'Color':'#F5A623'},
                    {'Band':'At risk (<2.5)', 'Count':int((staff_scores['Final_BSC_Score']<2.5).sum()),'Color':'#E24B4A'},
                ])
                fig = px.bar(bands, x='Band', y='Count', color='Band',
                             color_discrete_map={r['Band']:r['Color'] for _,r in bands.iterrows()},
                             title='Performance distribution')
                fig.update_layout(height=260, showlegend=False,
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    margin=dict(l=0,r=0,t=40,b=0))
                st.plotly_chart(fig, use_container_width=True)

        with ov2:
            # Leave types in use
            if lm and lm.records:
                lt_counts = {}
                for r in lm.records:
                    lt = r.get('leave_type','Unknown')
                    lt_counts[lt] = lt_counts.get(lt,0) + 1
                lt_df = pd.DataFrame(list(lt_counts.items()), columns=['Leave type','Count'])
                lt_df['Color'] = lt_df['Leave type'].map(lambda x: LEAVE_TYPES.get(x,{}).get('color','#888'))
                fig2 = px.pie(lt_df, names='Leave type', values='Count',
                              title='Leave types (all records)',
                              color='Leave type',
                              color_discrete_map={r['Leave type']:r['Color'] for _,r in lt_df.iterrows()})
                fig2.update_layout(height=260, margin=dict(l=0,r=0,t=40,b=0))
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("No leave records yet.")

        # Trend alerts
        st.markdown("#### HR Alerts")
        alerts = []

        # Declining performance — within band but trending down
        if len(df_proc) and 'Staff Name' in df_proc.columns:
            month_cols = [c for c in df_proc.columns
                          if any(m in str(c) for m in ['Jan','Feb','Mar','Apr','May','Jun',
                                                         'Jul','Aug','Sep','Oct','Nov','Dec'])
                          and 'Target' not in str(c)]
            if len(month_cols) >= 3:
                for name in staff_scores['Staff Name'].tolist():
                    rows = df_proc[df_proc['Staff Name']==name]
                    if len(rows)==0: continue
                    monthly_avgs = []
                    for mc in month_cols[-3:]:
                        if mc in rows.columns:
                            tgt_col = mc.replace(' Actual','') + ' Target' if ' Actual' in mc else None
                            act_vals = pd.to_numeric(rows[mc], errors='coerce').dropna()
                            if len(act_vals): monthly_avgs.append(act_vals.mean())
                    if len(monthly_avgs) == 3 and monthly_avgs[2] < monthly_avgs[0]:
                        drop = monthly_avgs[0] - monthly_avgs[2]
                        if drop / max(monthly_avgs[0], 1) > 0.1:
                            bsc_row = staff_scores[staff_scores['Staff Name']==name]
                            bsc_score = bsc_row['Final_BSC_Score'].values[0] if len(bsc_row) else 0
                            if 2.5 <= bsc_score < 3.2:  # within band but declining
                                alerts.append({
                                    'type':'⚠️ Declining trend',
                                    'staff': name,
                                    'detail': f"BSC {bsc_score:.2f} — performance declining over last 3 months",
                                    'color':'#FAEEDA','border':'#F5A623'
                                })

        # PIP nearing deadline
        if hr_m:
            for pip in hr_m.get_active_pips():
                days_left = hr_m.pip_days_remaining(pip)
                if days_left <= 14:
                    alerts.append({
                        'type':'🚨 PIP deadline',
                        'staff': pip['staff_name'],
                        'detail': f"PIP ends in {days_left} day(s) — review required",
                        'color':'#FDEDEC','border':'#E24B4A'
                    })

        # Disc cases open > 30 days
        if hr_m:
            for case in hr_m.get_active_cases():
                try:
                    opened = datetime.fromisoformat(case['recorded_at'])
                    days_open = (datetime.now() - opened).days
                    if days_open > 30:
                        alerts.append({
                            'type':'🔴 Stalled case',
                            'staff': case['staff_name'],
                            'detail': f"Disciplinary case {case['id']} open for {days_open} days — no resolution",
                            'color':'#FDEDEC','border':'#E24B4A'
                        })
                except: pass

        if not alerts:
            st.success("No HR alerts at this time.")
        else:
            for a in alerts[:10]:
                st.markdown(
                    f"<div style='padding:8px 12px;background:{a['color']};"
                    f"border-left:3px solid {a['border']};border-radius:0 4px 4px 0;"
                    f"font-size:12px;margin:3px 0'>"
                    f"<b>{a['type']}</b> — {a['staff']}: {a['detail']}</div>",
                    unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # ════════════════════════════════════════════════════════════════
        # TAB 2 — STAFF DIRECTORY
        # ════════════════════════════════════════════════════════════════
    with sub[1]:
        st.subheader("Team performance insights")
        team_df = filtered.copy()
        all_names = sorted(team_df['Staff Name'].tolist()) if len(team_df) else []

        if not all_names:
            st.info("No team data available.")
        else:
            tc1,tc2,tc3,tc4,tc5 = st.columns(5)
            tc1.metric("Team size",       len(team_df))
            tc2.metric("Avg BSC",         fmt_score(team_df['Final_BSC_Score'].mean()))
            tc3.metric("Exceeded",        int((team_df['Final_BSC_Score']>=3.1).sum()))
            tc4.metric("Below target",    int((team_df['Final_BSC_Score']<3.0).sum()))
            tc5.metric("Critical (<2.5)", int((team_df['Final_BSC_Score']<2.5).sum()))

            team_kpis = df_proc[df_proc['Staff Name'].isin(all_names)].copy()

            if not team_kpis.empty and 'KPI' in team_kpis.columns:
                kpi_summary = team_kpis.groupby(['KPI','Pillar']).agg(
                    Avg_Score=('Score','mean'), Avg_Achievement=('Percent_Achieved','mean'),
                    Staff_Count=('Staff Name','nunique'),
                    Below_Target=('Score', lambda x: (x < 3.0).sum()),
                    Critical=('Score', lambda x: (x < 2.5).sum()),
                ).reset_index()
                kpi_summary['Below_Target_Pct'] = (kpi_summary['Below_Target']/kpi_summary['Staff_Count']*100).round(0)
                kpi_summary = kpi_summary.sort_values('Avg_Score')

                st.markdown("---")
                st.markdown("### Team-wide weaknesses")
                team_weak = kpi_summary[kpi_summary['Avg_Score'] < 3.0].head(8)
                for _, r in team_weak.iterrows():
                    colour = "#E74C3C" if r['Avg_Score'] < 2.5 else "#F39C12"
                    st.markdown(
                        f"<div style='padding:9px 14px;background:{colour}15;"
                        f"border-left:4px solid {colour};border-radius:5px;margin:4px 0'>"
                        f"<strong>{r['KPI']}</strong> <span style='color:#888;font-size:12px'>({r['Pillar']})</span><br>"
                        f"Avg: <strong>{r['Avg_Score']:.2f}</strong> | {r['Avg_Achievement']:.1f}% achieved | "
                        f"{int(r['Below_Target_Pct'])}% of team below target</div>",
                        unsafe_allow_html=True)

                st.markdown("### Team-wide strengths")
                team_strong = kpi_summary[kpi_summary['Avg_Score']>=3.5].sort_values('Avg_Score', ascending=False).head(5)
                for _, r in team_strong.iterrows():
                    st.markdown(
                        f"<div style='padding:9px 14px;background:#2ECC7115;"
                        f"border-left:4px solid #2ECC71;border-radius:5px;margin:4px 0'>"
                        f"<strong>{r['KPI']}</strong> ({r['Pillar']}) — "
                        f"Avg: <strong>{r['Avg_Score']:.2f}</strong> | {r['Avg_Achievement']:.1f}%"
                        f"</div>", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("### Individual focus areas")
            for _, staff_row in team_df.sort_values('Final_BSC_Score').iterrows():
                sname = staff_row['Staff Name']
                score = staff_row['Final_BSC_Score']
                remark = staff_row['Performance_Remark']
                rank = staff_row['Overall_Rank']
                with st.expander(
                    f"{'🔴' if score<2.5 else '🟡' if score<3.0 else '🟢'} "
                    f"{sname} — {fmt_score(score)} | {remark} | Rank #{rank}",
                    expanded=(score<2.5)):
                    s_kpis = df_proc[df_proc['Staff Name']==sname]
                    s_insights = get_kpi_insights(s_kpis)
                    render_insight_card(s_insights, sname)

            if not team_kpis.empty and 'Pillar' in team_kpis.columns:
                st.markdown("---")
                st.markdown("### Pillar heatmap")
                pillar_staff = team_kpis.groupby(['Staff Name','Pillar'])['Score'].mean().reset_index()
                pivot = pillar_staff.pivot(index='Staff Name', columns='Pillar', values='Score').fillna(0)
                fig_heat = px.imshow(pivot,
                    color_continuous_scale=[[0,"#E74C3C"],[0.5,"#F39C12"],[0.6,"#FFE4B5"],[1,"#2ECC71"]],
                    zmin=1, zmax=5, title="Score per pillar per staff", aspect="auto", text_auto=".2f")
                fig_heat.update_layout(height=max(300, len(pivot)*28))
                st.plotly_chart(fig_heat, use_container_width=True)

        # ── Payroll Export (end-of-month) ─────────────────────────────────
        # This is in a new sub-tab of the BSC section — add to end of page
        _payroll_expander = st.expander("💰 Payroll Export — BSC scores + commission for payroll", expanded=False)

    # ════════════════════════════════════════════════════════════════
    # SUB-TAB 2: Compensation Equity (Standard #63, integrated v5.79)
    # ════════════════════════════════════════════════════════════════
    with sub[2]:
        from utils.compensation_equity import (
            CompensationEquityEngine, CompensationRecord,
            PAY_GAP_FAIR_MAX_PCT, PAY_GAP_MODERATE_MAX_PCT,
            CEO_RATIO_HEALTHY_MAX, CEO_RATIO_HIGH_THRESHOLD,
            COMPA_RATIO_HEALTHY_MIN, COMPA_RATIO_HEALTHY_MAX,
        )
        from decimal import Decimal as _D_ce

        st.markdown(
            f"**Standard #63 — Compensation Equity**. "
            f"Pay-gap thresholds: FAIR ≤{PAY_GAP_FAIR_MAX_PCT}% / MODERATE ≤{PAY_GAP_MODERATE_MAX_PCT}% / HIGH > {PAY_GAP_MODERATE_MAX_PCT}%. "
            f"CEO-to-median ratio: HEALTHY ≤{CEO_RATIO_HEALTHY_MAX}× / HIGH > {CEO_RATIO_HIGH_THRESHOLD}×. "
            f"Compa-ratio band: {COMPA_RATIO_HEALTHY_MIN}-{COMPA_RATIO_HEALTHY_MAX}."
        )

        ce_sub_tabs = st.tabs([
            "👫 Gender Pay Gap",
            "👑 CEO-to-Median Ratio",
            "📐 Internal Equity (Compa-ratio)",
            "📊 Pay Distribution by Grade",
            "📦 Compensation Depth (#63, v5.97)",
        ])

        # Try to load real comp records from staff_register; fall back to demo
        @st.cache_data(ttl=120, show_spinner=False)
        def _load_comp_records():
            """Build CompensationRecord list from staff_register if available."""
            try:
                import openpyxl
                sf = DATA / "staff_register.xlsx"
                if not sf.exists():
                    return None
                wb = openpyxl.load_workbook(str(sf), data_only=True)
                ws = wb.active
                hdr = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
                # Required columns
                col_map = {}
                for c, h in enumerate(hdr):
                    h_low = h.lower().strip()
                    if "staff_code" in h_low or h_low == "staff code":
                        col_map["staff_id"] = c
                    elif h_low in ("salary", "base_salary", "base_salary_kes", "monthly_salary"):
                        col_map["salary"] = c
                    elif h_low == "grade":
                        col_map["grade"] = c
                    elif h_low == "role":
                        col_map["role"] = c
                    elif h_low == "branch_code" or h_low == "branch":
                        col_map["branch"] = c
                    elif h_low == "gender":
                        col_map["gender"] = c
                if not all(k in col_map for k in ("staff_id", "salary", "grade", "role")):
                    return None
                rows = []
                for r in range(2, min(ws.max_row + 1, 600)):
                    sid = ws.cell(r, col_map["staff_id"] + 1).value
                    sal = ws.cell(r, col_map["salary"] + 1).value
                    grd = ws.cell(r, col_map["grade"] + 1).value
                    role = ws.cell(r, col_map["role"] + 1).value
                    if not (sid and sal and grd):
                        continue
                    try:
                        sal_d = _D_ce(str(sal))
                    except Exception:
                        continue
                    rec = CompensationRecord(
                        staff_id=str(sid),
                        base_salary_kes=sal_d,
                        grade=str(grd),
                        role=str(role or ""),
                        branch_code=str(ws.cell(r, col_map.get("branch", 0) + 1).value or "")
                                     if "branch" in col_map else "",
                        gender=str(ws.cell(r, col_map.get("gender", 0) + 1).value or "")
                                if "gender" in col_map else None,
                    )
                    rows.append(rec)
                return rows if rows else None
            except Exception:
                return None

        comp_records = _load_comp_records()
        # Demo fallback if no register or columns missing
        if not comp_records:
            comp_records = [
                CompensationRecord("S001", _D_ce("8000000"), "EXEC", "MD", "100", "M",
                                     grade_midpoint_kes=_D_ce("7500000")),
                CompensationRecord("S002", _D_ce("4500000"), "G7", "Director", "100", "F",
                                     grade_midpoint_kes=_D_ce("4500000")),
                CompensationRecord("S003", _D_ce("3000000"), "G6", "Head", "100", "M",
                                     grade_midpoint_kes=_D_ce("3000000")),
                CompensationRecord("S004", _D_ce("2000000"), "G5", "Manager", "200", "F",
                                     grade_midpoint_kes=_D_ce("2000000")),
                CompensationRecord("S005", _D_ce("800000"), "G4", "RM", "200", "M",
                                     grade_midpoint_kes=_D_ce("800000")),
                CompensationRecord("S006", _D_ce("750000"), "G4", "RM", "200", "F",
                                     grade_midpoint_kes=_D_ce("800000")),
                CompensationRecord("S007", _D_ce("400000"), "G3", "CSO", "300", "F"),
                CompensationRecord("S008", _D_ce("400000"), "G3", "Teller", "300", "M"),
            ]
            st.caption(
                f"⚠ Using **demo dataset** ({len(comp_records)} records) — "
                "`staff_register.xlsx` did not have salary/grade/gender columns. "
                "Add those columns to switch to live data.")
        else:
            st.caption(
                f"📂 Loaded **{len(comp_records)} records** from "
                f"`staff_register.xlsx`.")

        # ──────── Gender pay gap ────────
        with ce_sub_tabs[0]:
            st.markdown("**Gender Pay Gap** — overall + grade-adjusted")
            st.caption(
                "Raw gap reflects role-mix differences (more women in junior grades = higher raw gap "
                "even when pay is fair within grades). Grade-adjusted gap is the within-grade comparison.")

            if st.button("Compute gender pay gap",
                           key="ce_gpg_btn", type="primary"):
                r = CompensationEquityEngine.gender_pay_gap(
                    comp_records, by_grade=True)
                k1, k2, k3 = st.columns(3)
                k1.metric("Male count", r.get("male_count"),
                           help="Records with gender=M")
                k2.metric("Female count", r.get("female_count"))
                k3.metric("Unknown gender", r.get("unknown_gender_count"))

                raw_gap = r.get("raw_gap_pct")
                adj_gap = r.get("adjusted_gap_pct")
                raw_sev = r.get("raw_gap_severity")
                adj_sev = r.get("adjusted_gap_severity")
                colors = {"FAIR": "#10B981", "MODERATE": "#F59E0B",
                          "HIGH": "#DC2626", None: "#6B7280"}

                k1, k2 = st.columns(2)
                with k1:
                    color = colors.get(raw_sev, "#6B7280")
                    st.markdown(
                        f"<div style='padding:14px;background:{color}22;"
                        f"border-left:6px solid {color};border-radius:10px'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                        f"RAW GAP (overall)</div>"
                        f"<div style='font-size:24px;font-weight:800;color:{color};margin-top:4px'>"
                        f"{raw_gap}% — {raw_sev or '—'}</div>"
                        f"<div style='font-size:11px;margin-top:4px;opacity:0.85'>"
                        f"Reflects role mix; can be high even with fair within-grade pay.</div></div>",
                        unsafe_allow_html=True)
                with k2:
                    color = colors.get(adj_sev, "#6B7280")
                    st.markdown(
                        f"<div style='padding:14px;background:{color}22;"
                        f"border-left:6px solid {color};border-radius:10px'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                        f"ADJUSTED GAP (within grade)</div>"
                        f"<div style='font-size:24px;font-weight:800;color:{color};margin-top:4px'>"
                        f"{adj_gap if adj_gap is not None else '—'}% — {adj_sev or '—'}</div>"
                        f"<div style='font-size:11px;margin-top:4px;opacity:0.85'>"
                        f"True equity test — should be ≤{PAY_GAP_FAIR_MAX_PCT}% for FAIR.</div></div>",
                        unsafe_allow_html=True)

                # Per-grade table
                per_grade = r.get("per_grade", [])
                if per_grade:
                    st.markdown("**Per-grade breakdown:**")
                    pg_rows = [
                        {"Grade": g["grade"],
                          "Male count": g["male_count"],
                          "Female count": g["female_count"],
                          "Male median (KES)": g["male_median"] or "—",
                          "Female median (KES)": g["female_median"] or "—",
                          "Within-grade gap %":
                              f"{g['gap_pct']}%" if g["gap_pct"] is not None else "—"}
                        for g in per_grade
                    ]
                    st.dataframe(pd.DataFrame(pg_rows),
                                 use_container_width=True, hide_index=True)

                audit_log("IFRS_ENGINE_USED", uname,
                           f"Comp #63: gender pay gap raw={raw_gap}% adj={adj_gap}% "
                           f"raw_sev={raw_sev} adj_sev={adj_sev}")

        # ──────── CEO-to-median ratio ────────
        with ce_sub_tabs[1]:
            st.markdown("**CEO-to-Median Pay Ratio** (transparency disclosure)")
            st.caption(
                f"Healthy: ≤{CEO_RATIO_HEALTHY_MAX}× · High: > {CEO_RATIO_HIGH_THRESHOLD}×. "
                "Reflects pay equity between top of organisation and median worker.")

            ceo_id = st.text_input("CEO staff ID",
                                     value="S001",
                                     key="ce_ceo_id",
                                     help="Use exact staff_code from your staff register.")
            if st.button("Compute CEO ratio",
                           key="ce_ceo_btn", type="primary"):
                r = CompensationEquityEngine.ceo_to_median_ratio(
                    comp_records, ceo_staff_id=ceo_id)
                ratio = r.get("ratio")
                severity = r.get("severity")
                if ratio is None:
                    st.error(
                        f"⛔ Could not compute — CEO staff ID `{ceo_id}` not found "
                        f"in records.")
                else:
                    colors = {"HEALTHY": "#10B981", "ELEVATED": "#F59E0B",
                              "HIGH": "#DC2626", "EXTREME": "#7C2D12"}
                    color = colors.get(severity, "#6B7280")
                    k1, k2, k3 = st.columns(3)
                    k1.metric("CEO salary",
                               f"KES {_D_ce(str(r['ceo_salary_kes'])):,.0f}")
                    k2.metric("Median employee",
                               f"KES {_D_ce(str(r['median_employee_salary_kes'])):,.0f}")
                    with k3:
                        st.markdown(
                            f"<div style='padding:8px 12px;background:{color}22;"
                            f"border-left:4px solid {color};border-radius:8px;text-align:center'>"
                            f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                            f"RATIO</div>"
                            f"<div style='font-size:24px;font-weight:800;color:{color}'>"
                            f"{ratio:.1f}× — {severity}</div></div>",
                            unsafe_allow_html=True)
                    audit_log("IFRS_ENGINE_USED", uname,
                               f"Comp #63: CEO ratio {ratio} ({severity})")

        # ──────── Internal equity / compa-ratio ────────
        with ce_sub_tabs[2]:
            st.markdown(
                f"**Compa-Ratio Analysis** = salary / grade midpoint. "
                f"Healthy band: {COMPA_RATIO_HEALTHY_MIN}-{COMPA_RATIO_HEALTHY_MAX}.")
            st.caption(
                "Below 0.8 → underpaid for grade. Above 1.2 → overpaid for grade. "
                "Records without a `grade_midpoint_kes` are excluded with count surfaced.")

            if st.button("Compute internal equity",
                           key="ce_eq_btn", type="primary"):
                r = CompensationEquityEngine.internal_equity_ratios(comp_records)
                k1, k2, k3, k4 = st.columns(4)
                k1.metric("Below band", r.get("below_band_count"),
                           delta_color="inverse" if r.get("below_band_count", 0) > 0 else "normal")
                k2.metric("In band", r.get("in_band_count"))
                k3.metric("Above band", r.get("above_band_count"),
                           delta_color="inverse" if r.get("above_band_count", 0) > 0 else "normal")
                k4.metric("No midpoint", r.get("no_midpoint_count"),
                           help="Excluded — records missing grade_midpoint_kes")

                rec_list = r.get("records", [])
                if rec_list:
                    rec_rows = [
                        {"Staff ID": rr["staff_id"],
                          "Grade": rr["grade"],
                          "Compa-ratio": rr["compa_ratio"],
                          "Band": rr["band"],
                          "Status": "🔴" if rr["band"] != "IN_BAND" else "✅"}
                        for rr in rec_list
                    ]
                    st.dataframe(pd.DataFrame(rec_rows),
                                 use_container_width=True, hide_index=True)
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Comp #63: internal equity below={r.get('below_band_count')} "
                           f"above={r.get('above_band_count')}")

        # ──────── Pay distribution by grade ────────
        with ce_sub_tabs[3]:
            st.markdown("**Pay Distribution by Grade** (P25 / median / P75 / IQR)")
            grades = sorted({r.grade for r in comp_records if r.grade})
            if not grades:
                st.info("No grades in dataset.")
            else:
                grade = st.selectbox("Grade", grades, key="ce_dist_grade")
                if st.button("Compute distribution",
                               key="ce_dist_btn", type="primary"):
                    r = CompensationEquityEngine.pay_distribution_by_grade(
                        comp_records, grade=grade)
                    if r.get("headcount", 0) == 0:
                        st.warning(f"No records with valid salary in grade {grade}.")
                    else:
                        k1, k2, k3 = st.columns(3)
                        k1.metric("Headcount", r.get("headcount"),
                                   help=f"Excluded (no salary): {r.get('headcount_excluded_no_salary', 0)}")
                        k2.metric("Median (KES)",
                                   f"{_D_ce(str(r['median'])):,.0f}")
                        k3.metric("IQR (P75-P25)",
                                   f"KES {_D_ce(str(r['iqr'])):,.0f}")
                        st.markdown(
                            f"**Range**: {_D_ce(str(r['min'])):,.0f} (min) → "
                            f"{_D_ce(str(r['p25'])):,.0f} (P25) → "
                            f"{_D_ce(str(r['median'])):,.0f} (median) → "
                            f"{_D_ce(str(r['p75'])):,.0f} (P75) → "
                            f"{_D_ce(str(r['max'])):,.0f} (max)")
                        audit_log("IFRS_ENGINE_USED", uname,
                                   f"Comp #63: distribution {grade} hc={r['headcount']}")

        # ════════════════════════════════════════════════════════════════
        # CE_SUB_TABS[4]: Compensation Depth (Standard #63, integrated v5.97)
        # ════════════════════════════════════════════════════════════════
        with ce_sub_tabs[4]:
            st.markdown(
                "**Compensation Depth analysis** — extends v5.79 with 4 inner views: "
                "executive scorecard combining all 4 engine paths, branch-level pay "
                "analytics, position-in-band concentration, and underpaid-uplift "
                "scenario simulator.")
            st.caption(
                "💡 v5.79 surfaces each engine path independently. v5.97 combines "
                "them and adds branch-level + band-position cuts that the engine "
                "supports via existing CompensationRecord fields (`branch_code`, "
                "`position_in_band`) but v5.79 doesn't expose.")

            _ce_depth_inner = st.tabs([
                "📋 Executive Scorecard",
                "🏢 Branch-Level Analytics",
                "📊 Position-in-Band",
                "🎯 Underpaid-Uplift Simulator",
            ])

            # ────────── Inner[0]: Executive Scorecard ──────────
            with _ce_depth_inner[0]:
                st.markdown(
                    "**Single-screen executive summary** — combines all 4 engine "
                    "paths into one pay-equity scorecard for board reporting.")
                st.caption(
                    "Each tile has a traffic-light severity. Click compute to refresh "
                    "all metrics. Useful for monthly board pack or annual sustainability "
                    "report.")

                ces_ceo_id = st.text_input(
                    "CEO staff ID (for ratio calc)",
                    value="S001", key="ces_ceo_id")

                if st.button("📋 Compute scorecard",
                               key="ces_compute_btn", type="primary"):
                    # Run all 4 engine paths in one shot
                    gpg = CompensationEquityEngine.gender_pay_gap(
                        comp_records, by_grade=True)
                    ceo = CompensationEquityEngine.ceo_to_median_ratio(
                        comp_records, ces_ceo_id)
                    eq = CompensationEquityEngine.internal_equity_ratios(comp_records)

                    # Severity → color
                    sev_colors = {"FAIR": "#10B981", "MODERATE": "#F59E0B",
                                   "HIGH": "#DC2626",
                                   "HEALTHY": "#10B981", "ELEVATED": "#F59E0B",
                                   "EXTREME": "#7C2D12"}

                    st.markdown("### 1️⃣ Gender pay gap")
                    raw_gap = gpg.get("raw_gap_pct")
                    adj_gap = gpg.get("adjusted_gap_pct")
                    raw_sev = gpg.get("raw_gap_severity") or "—"
                    adj_sev = gpg.get("adjusted_gap_severity") or "—"

                    g1, g2, g3 = st.columns(3)
                    g1.metric("Headcount (M / F / Unknown)",
                                f"{gpg.get('male_count')} / "
                                f"{gpg.get('female_count')} / "
                                f"{gpg.get('unknown_gender_count')}")
                    raw_color = sev_colors.get(raw_sev, "#6B7280")
                    g2.markdown(
                        f"<div style='padding:8px;background:{raw_color}22;"
                        f"border-left:4px solid {raw_color};border-radius:6px'>"
                        f"<div style='font-size:10px;opacity:0.7'>RAW GAP</div>"
                        f"<div style='font-size:18px;font-weight:700;color:{raw_color}'>"
                        f"{raw_gap}% — {raw_sev}</div></div>",
                        unsafe_allow_html=True)
                    adj_color = sev_colors.get(adj_sev, "#6B7280")
                    g3.markdown(
                        f"<div style='padding:8px;background:{adj_color}22;"
                        f"border-left:4px solid {adj_color};border-radius:6px'>"
                        f"<div style='font-size:10px;opacity:0.7'>ADJUSTED GAP</div>"
                        f"<div style='font-size:18px;font-weight:700;color:{adj_color}'>"
                        f"{adj_gap if adj_gap is not None else '—'}% — {adj_sev}</div></div>",
                        unsafe_allow_html=True)

                    st.markdown("### 2️⃣ CEO-to-median ratio")
                    if ceo.get("ratio") is None:
                        st.error(f"⛔ CEO `{ces_ceo_id}` not found in records.")
                    else:
                        ceo_sev = ceo.get("severity", "—")
                        ceo_color = sev_colors.get(ceo_sev, "#6B7280")
                        c1, c2, c3 = st.columns(3)
                        c1.metric("CEO salary",
                                    f"KES {_D_ce(str(ceo['ceo_salary_kes'])):,.0f}")
                        c2.metric("Median employee",
                                    f"KES {_D_ce(str(ceo['median_employee_salary_kes'])):,.0f}")
                        c3.markdown(
                            f"<div style='padding:8px;background:{ceo_color}22;"
                            f"border-left:4px solid {ceo_color};border-radius:6px'>"
                            f"<div style='font-size:10px;opacity:0.7'>RATIO</div>"
                            f"<div style='font-size:18px;font-weight:700;color:{ceo_color}'>"
                            f"{float(ceo['ratio']):.1f}× — {ceo_sev}</div></div>",
                            unsafe_allow_html=True)

                    st.markdown("### 3️⃣ Internal equity (compa-ratio)")
                    in_band = int(eq.get("in_band_count", 0))
                    below = int(eq.get("below_band_count", 0))
                    above = int(eq.get("above_band_count", 0))
                    no_mid = int(eq.get("no_midpoint_count", 0))
                    total_with_midpoint = in_band + below + above
                    in_pct = (in_band/total_with_midpoint*100
                                if total_with_midpoint else 0)

                    e1, e2, e3, e4 = st.columns(4)
                    e1.metric("In band (0.8-1.2)", f"{in_band} ({in_pct:.0f}%)")
                    e2.metric("Below 0.8 (underpaid)", below,
                                delta_color="inverse" if below > 0 else "normal")
                    e3.metric("Above 1.2 (overpaid)", above,
                                delta_color="inverse" if above > 0 else "normal")
                    e4.metric("No midpoint", no_mid,
                                help="Excluded from analysis — engine surfaces (Rule 6)")

                    st.markdown("### 4️⃣ Overall scorecard verdict")
                    issues = []
                    if adj_sev in ("MODERATE", "HIGH"):
                        issues.append(f"adjusted gender pay gap is {adj_sev}")
                    if ceo.get("severity") in ("HIGH", "EXTREME"):
                        issues.append(
                            f"CEO ratio is {ceo.get('severity')} "
                            f"({float(ceo.get('ratio', 0)):.1f}×)")
                    if below > 0:
                        issues.append(f"{below} staff below compa-ratio band")
                    if above > 0:
                        issues.append(f"{above} staff above compa-ratio band")

                    if not issues:
                        st.success(
                            "✅ **Pay-equity health: GREEN.** All metrics in healthy "
                            "ranges. Maintain via annual review cycle.")
                    elif len(issues) <= 1:
                        st.warning(
                            f"⚠ **Pay-equity health: AMBER.** Issue: "
                            f"{issues[0]}. Targeted remediation recommended.")
                    else:
                        st.error(
                            f"🚨 **Pay-equity health: RED.** Multiple issues: "
                            f"{'; '.join(issues)}. Comprehensive review required.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Comp #63 (depth): scorecard issues={len(issues)} "
                                f"adj_gap={adj_gap} ceo_ratio={ceo.get('ratio')} "
                                f"below={below} above={above}")

            # ────────── Inner[1]: Branch-Level Analytics ──────────
            with _ce_depth_inner[1]:
                st.markdown(
                    "**Branch-level pay analytics** — `CompensationRecord` has a "
                    "`branch_code` field that v5.79 doesn't surface. Run the "
                    "engine's distribution method per branch to identify branches "
                    "with pay compression or pay anomalies.")
                st.caption(
                    "💡 Useful for: branch RM compensation reviews, identifying "
                    "regional pay patterns (HQ vs branches), branch managers "
                    "checking their team's pay distribution.")

                branches = sorted({r.branch_code for r in comp_records
                                    if r.branch_code})
                if not branches:
                    st.info("No branch_code values in dataset.")
                else:
                    # Per-branch distribution table
                    if st.button("🏢 Compute per-branch pay analytics",
                                   key="ce_branch_btn", type="primary"):
                        branch_rows = []
                        for br in branches:
                            br_records = [r for r in comp_records
                                            if r.branch_code == br]
                            br_count = len(br_records)
                            br_salaries = [float(r.base_salary_kes)
                                            for r in br_records
                                            if r.base_salary_kes]

                            if br_salaries:
                                br_min = min(br_salaries)
                                br_max = max(br_salaries)
                                br_median = sorted(br_salaries)[len(br_salaries)//2]
                                br_total = sum(br_salaries)
                                br_avg = br_total / len(br_salaries)
                                # Compa ratios for this branch
                                eq_br = CompensationEquityEngine.internal_equity_ratios(
                                    br_records)
                                below_br = int(eq_br.get("below_band_count", 0))
                                above_br = int(eq_br.get("above_band_count", 0))
                            else:
                                br_min = br_max = br_median = br_total = br_avg = 0
                                below_br = above_br = 0

                            branch_rows.append({
                                "Branch": br,
                                "Headcount": br_count,
                                "Total salary (KES)": f"{br_total:,.0f}",
                                "Avg salary (KES)": f"{br_avg:,.0f}",
                                "Median (KES)": f"{br_median:,.0f}",
                                "Range (KES)":
                                    f"{br_min:,.0f} – {br_max:,.0f}"
                                    if br_salaries else "—",
                                "Below band": below_br,
                                "Above band": above_br,
                                "Status": ("🔴" if (below_br + above_br) > 0
                                            else "✅"),
                            })

                        st.dataframe(pd.DataFrame(branch_rows),
                                     use_container_width=True, hide_index=True)

                        # Bar chart of branch-level total payroll
                        chart_data = pd.DataFrame({
                            "Total salary (KES)":
                                [sum(float(r.base_salary_kes)
                                      for r in comp_records
                                      if r.branch_code == br
                                      and r.base_salary_kes)
                                  for br in branches]
                        }, index=branches)
                        st.markdown("**Total payroll by branch:**")
                        st.bar_chart(chart_data)

                        # Branches with band issues
                        issue_branches = [br_row for br_row in branch_rows
                                            if br_row["Status"] == "🔴"]
                        if issue_branches:
                            st.warning(
                                f"⚠ **{len(issue_branches)} branch(es) have "
                                "compa-ratio outliers** — review pay-band "
                                "alignment with HQ standards.")

                        audit_log("IFRS_ENGINE_USED", uname,
                                    f"Comp #63 (depth): branch analytics "
                                    f"branches={len(branches)} "
                                    f"with_issues={len(issue_branches)}")

            # ────────── Inner[2]: Position-in-Band ──────────
            with _ce_depth_inner[2]:
                st.markdown(
                    "**Position-in-band concentration analysis**. "
                    "`CompensationRecord` has a `position_in_band` field "
                    "(typically P25 / P50 / P75) indicating where staff sit "
                    "relative to grade-band quartiles. v5.79 doesn't analyze this.")
                st.caption(
                    "💡 **Concentration patterns matter**: if 80% of senior "
                    "staff sit at P75 (top of band), the bank is approaching "
                    "**pay compression** — high performers can't be rewarded "
                    "without crossing into a higher grade. If 80% sit at P25 "
                    "(bottom), there's headroom for merit increases without "
                    "structural changes.")

                # Position-in-band distribution
                positions = {}
                for r in comp_records:
                    pos = r.position_in_band
                    if pos:
                        positions[pos] = positions.get(pos, 0) + 1

                if not positions:
                    st.info(
                        "No `position_in_band` values in dataset. Production "
                        "deployment needs HR data with band-position assignments.")
                else:
                    total_with_pos = sum(positions.values())
                    pos_rows = [
                        {"Position": pos,
                          "Count": cnt,
                          "% of staff": f"{cnt/total_with_pos*100:.1f}%"}
                        for pos, cnt in sorted(positions.items())
                    ]
                    st.dataframe(pd.DataFrame(pos_rows),
                                 use_container_width=True, hide_index=True)

                    # Bar chart
                    chart_pos = pd.DataFrame({
                        "Count": [positions[p]
                                    for p in sorted(positions.keys())]
                    }, index=sorted(positions.keys()))
                    st.markdown("**Position distribution:**")
                    st.bar_chart(chart_pos)

                    # Per-grade position breakdown
                    st.markdown("**Position breakdown by grade:**")
                    grade_pos = {}
                    for r in comp_records:
                        if r.grade and r.position_in_band:
                            grade_pos.setdefault(r.grade, {})
                            grade_pos[r.grade][r.position_in_band] = (
                                grade_pos[r.grade].get(r.position_in_band, 0) + 1)

                    grade_pos_rows = []
                    all_positions = sorted({pos for d in grade_pos.values()
                                              for pos in d.keys()})
                    for grade in sorted(grade_pos.keys()):
                        row = {"Grade": grade}
                        total_grade = sum(grade_pos[grade].values())
                        for pos in all_positions:
                            count = grade_pos[grade].get(pos, 0)
                            row[pos] = (f"{count} ({count/total_grade*100:.0f}%)"
                                          if count else "—")
                        grade_pos_rows.append(row)
                    st.dataframe(pd.DataFrame(grade_pos_rows),
                                 use_container_width=True, hide_index=True)

                    # Compression detection
                    p75_pct = positions.get("P75", 0) / total_with_pos * 100
                    p25_pct = positions.get("P25", 0) / total_with_pos * 100
                    if p75_pct >= 60:
                        st.warning(
                            f"⚠ **Pay compression risk** — {p75_pct:.0f}% of "
                            "staff sit at P75 (top of band). High performers "
                            "have limited upside without grade-shift; consider "
                            "band-range adjustments.")
                    elif p25_pct >= 60:
                        st.info(
                            f"💡 **Pay headroom** — {p25_pct:.0f}% of staff "
                            "sit at P25 (bottom of band). Merit increases can "
                            "proceed without structural pay-band changes.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Comp #63 (depth): position analytics "
                                f"P25={p25_pct:.0f}% P75={p75_pct:.0f}%")

            # ────────── Inner[3]: Underpaid-Uplift Simulator ──────────
            with _ce_depth_inner[3]:
                st.markdown(
                    "**Underpaid-uplift cost scenario**. "
                    "Find all staff with compa-ratio below the healthy band "
                    "(<0.8) and simulate the cost of uplifting them to a target "
                    "compa-ratio. Useful for board approval of remediation budget.")
                st.caption(
                    "💡 Uses internal_equity_ratios output to identify underpaid "
                    "staff, then computes target_salary = grade_midpoint × "
                    "target_compa_ratio. Cost = sum of (target_salary - "
                    "current_salary) across affected staff.")

                target_compa = st.slider(
                    "Target compa-ratio for uplift",
                    min_value=0.80, max_value=1.10,
                    value=0.95, step=0.01,
                    key="ce_uplift_target",
                    help="Default 0.95 brings underpaid staff to mid-band. "
                         "0.80 = minimum healthy. 1.0 = grade midpoint.")

                if st.button("🎯 Run uplift simulation",
                               key="ce_uplift_btn", type="primary"):
                    eq_r = CompensationEquityEngine.internal_equity_ratios(
                        comp_records)
                    underpaid_records = []
                    target_d = _D_ce(str(target_compa))

                    for record in comp_records:
                        # Find this record in eq_r results
                        match = next((rr for rr in eq_r.get("records", [])
                                       if rr["staff_id"] == record.staff_id),
                                      None)
                        if (match and match["band"] == "BELOW_BAND"
                                and record.grade_midpoint_kes):
                            current = float(record.base_salary_kes)
                            target_salary = float(record.grade_midpoint_kes) * target_compa
                            uplift = target_salary - current
                            underpaid_records.append({
                                "Staff ID": record.staff_id,
                                "Grade": record.grade,
                                "Branch": record.branch_code,
                                "Current salary (KES)": f"{current:,.0f}",
                                "Current compa": match["compa_ratio"],
                                "Target salary (KES)": f"{target_salary:,.0f}",
                                "Uplift (KES)": uplift,
                            })

                    if not underpaid_records:
                        st.success(
                            "✅ **No underpaid staff** — all in-band records "
                            "have compa-ratio ≥0.8. No uplift required.")
                    else:
                        total_uplift = sum(r["Uplift (KES)"]
                                            for r in underpaid_records)
                        n_affected = len(underpaid_records)
                        avg_uplift = total_uplift / n_affected

                        k1, k2, k3 = st.columns(3)
                        k1.metric("Underpaid staff", n_affected)
                        k2.metric("Total uplift cost (KES)",
                                    f"{total_uplift:,.0f}")
                        k3.metric("Avg uplift per staff (KES)",
                                    f"{avg_uplift:,.0f}")

                        # Display table with formatted uplift column
                        display_rows = []
                        for r in underpaid_records:
                            display_rows.append({
                                **{k: v for k, v in r.items()
                                    if k != "Uplift (KES)"},
                                "Uplift (KES)": f"{r['Uplift (KES)']:,.0f}",
                            })
                        st.dataframe(pd.DataFrame(display_rows),
                                     use_container_width=True, hide_index=True)

                        # Annual cost vs payroll context
                        total_payroll = sum(float(r.base_salary_kes)
                                              for r in comp_records
                                              if r.base_salary_kes)
                        uplift_pct = (total_uplift / total_payroll * 100
                                       if total_payroll else 0)
                        st.caption(
                            f"💡 Uplift represents **{uplift_pct:.2f}%** of "
                            f"total annual payroll (KES {total_payroll:,.0f}). "
                            "For board context: typical merit-increase budgets "
                            "are 3-5% — this remediation fits within a single "
                            "merit cycle if needed.")

                        audit_log("IFRS_ENGINE_USED", uname,
                                    f"Comp #63 (depth): uplift target={target_compa} "
                                    f"affected={n_affected} cost={total_uplift:.0f} "
                                    f"pct={uplift_pct:.2f}")


    # ════════════════════════════════════════════════════════════════
    # SUB-TAB 3: Engagement & Performance (Standard #64, integrated v5.79)
    # ════════════════════════════════════════════════════════════════
    with sub[3]:
        from utils.employee_engagement import (
            EmployeeEngagementEngine, SurveyResponse, StaffSignals,
            ENGAGEMENT_DRIVERS, ENGAGEMENT_HIGH_THRESHOLD,
            ENGAGEMENT_MODERATE_THRESHOLD,
            FLIGHT_RISK_HIGH_THRESHOLD, FLIGHT_RISK_MEDIUM_THRESHOLD,
            LIKERT_MIN, LIKERT_MAX,
        )
        from utils.performance_talent import (
            PerformanceTalentEngine, PerformanceReview, SuccessionPlan,
            RATING_LEVELS, READINESS_LEVELS,
            BENCH_AT_RISK_PCT, BENCH_HEALTHY_PCT, CALIBRATION_TARGETS,
        )

        st.markdown(
            f"**Standard #64 — Performance & Engagement**. "
            f"Engagement scoring per {LIKERT_MIN}-{LIKERT_MAX} Likert across "
            f"{len(ENGAGEMENT_DRIVERS)} drivers. "
            f"Performance ratings: {' / '.join(RATING_LEVELS)}. "
            f"Bench strength: AT_RISK <{BENCH_AT_RISK_PCT}% / HEALTHY ≥{BENCH_HEALTHY_PCT}%."
        )

        ep_sub_tabs = st.tabs([
            "💚 Engagement Score",
            "👍 eNPS",
            "🎚️ Driver Breakdown",
            "💬 Sentiment",
            "🚨 Flight Risk + Depth (#64, v5.98)",
            "📊 Rating Distribution",
            "🔄 Succession Bench",
        ])

        # Demo survey responses (single-quarter snapshot)
        @st.cache_data(ttl=300, show_spinner=False)
        def _demo_survey():
            return [
                SurveyResponse("R1", "S001", "2026-Q1", overall_likert=4, enps_score=8,
                                driver_scores={"LEADERSHIP": 4, "COMPENSATION": 5,
                                                "GROWTH_DEVELOPMENT": 4, "WORK_LIFE_BALANCE": 3,
                                                "RECOGNITION": 4, "PURPOSE_MEANING": 5}),
                SurveyResponse("R2", "S002", "2026-Q1", overall_likert=5, enps_score=10,
                                driver_scores={"LEADERSHIP": 5, "COMPENSATION": 5,
                                                "GROWTH_DEVELOPMENT": 4, "WORK_LIFE_BALANCE": 4,
                                                "RECOGNITION": 5, "PURPOSE_MEANING": 5}),
                SurveyResponse("R3", "S003", "2026-Q1", overall_likert=3, enps_score=6,
                                driver_scores={"LEADERSHIP": 3, "COMPENSATION": 2,
                                                "GROWTH_DEVELOPMENT": 3, "WORK_LIFE_BALANCE": 4,
                                                "RECOGNITION": 2, "PURPOSE_MEANING": 4}),
                SurveyResponse("R4", "S004", "2026-Q1", overall_likert=2, enps_score=4,
                                driver_scores={"LEADERSHIP": 2, "COMPENSATION": 1,
                                                "GROWTH_DEVELOPMENT": 2, "WORK_LIFE_BALANCE": 3,
                                                "RECOGNITION": 1, "PURPOSE_MEANING": 3},
                                text_response="I feel stressed and underpaid; thinking of leaving"),
                SurveyResponse("R5", "S005", "2026-Q1", overall_likert=5, enps_score=10,
                                driver_scores={"LEADERSHIP": 5, "COMPENSATION": 4,
                                                "GROWTH_DEVELOPMENT": 5, "WORK_LIFE_BALANCE": 5,
                                                "RECOGNITION": 5, "PURPOSE_MEANING": 5},
                                text_response="Great team and amazing growth opportunities"),
            ]

        responses = _demo_survey()
        st.caption(
            f"📊 Using **{len(responses)}-respondent demo dataset** — "
            "production deployment would feed survey data via "
            "`employee_survey_responses.json`.")

        # ──── Engagement score ────
        with ep_sub_tabs[0]:
            st.markdown("**Engagement Score** = average of overall Likert × 20 (out of 100)")
            if st.button("Compute engagement",
                           key="ep_eng_btn", type="primary"):
                r = EmployeeEngagementEngine.engagement_score(responses)
                score = r.get("score")
                severity = r.get("severity")
                colors = {"HIGH": "#10B981", "MODERATE": "#F59E0B",
                          "LOW": "#DC2626", None: "#6B7280"}
                color = colors.get(severity, "#6B7280")

                k1, k2, k3 = st.columns(3)
                k1.metric("Respondents", r.get("respondents"))
                k2.metric("Abstained", r.get("abstained"))
                with k3:
                    st.markdown(
                        f"<div style='padding:14px;background:{color}22;"
                        f"border-left:6px solid {color};border-radius:10px;text-align:center'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>SCORE</div>"
                        f"<div style='font-size:32px;font-weight:800;color:{color};margin-top:4px'>"
                        f"{score} — {severity}</div>"
                        f"<div style='font-size:11px;margin-top:4px;opacity:0.85'>"
                        f"HIGH ≥{ENGAGEMENT_HIGH_THRESHOLD} · MODERATE ≥{ENGAGEMENT_MODERATE_THRESHOLD}</div></div>",
                        unsafe_allow_html=True)
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Engagement #64: score={score} severity={severity}")

        # ──── eNPS ────
        with ep_sub_tabs[1]:
            st.markdown(
                "**eNPS** = % Promoters − % Detractors. "
                "Promoters score 9-10, Detractors 0-6, Passives 7-8.")
            if st.button("Compute eNPS",
                           key="ep_enps_btn", type="primary"):
                r = EmployeeEngagementEngine.enps(responses)
                enps_v = r.get("enps")
                colors = {None: "#6B7280"}
                if enps_v is not None:
                    color = "#10B981" if enps_v > 30 else "#F59E0B" if enps_v > 0 else "#DC2626"
                else:
                    color = "#6B7280"

                k1, k2, k3, k4 = st.columns(4)
                k1.metric("Respondents", r.get("respondents"))
                k2.metric("Promoters (9-10)", r.get("promoter_count"))
                k3.metric("Detractors (0-6)", r.get("detractor_count"))
                with k4:
                    st.markdown(
                        f"<div style='padding:8px 12px;background:{color}22;"
                        f"border-left:4px solid {color};border-radius:8px;text-align:center'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>eNPS</div>"
                        f"<div style='font-size:24px;font-weight:800;color:{color}'>"
                        f"{enps_v if enps_v is not None else '—'}</div></div>",
                        unsafe_allow_html=True)
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Engagement #64: eNPS={enps_v}")

        # ──── Driver breakdown ────
        with ep_sub_tabs[2]:
            st.markdown(
                f"**Driver Breakdown** — score per "
                f"{len(ENGAGEMENT_DRIVERS)} engagement drivers")
            if st.button("Compute drivers",
                           key="ep_drv_btn", type="primary"):
                r = EmployeeEngagementEngine.drivers_breakdown(responses)
                rows = []
                for drv in ENGAGEMENT_DRIVERS:
                    info = r.get(drv, {})
                    score = info.get("score")
                    rows.append({
                        "Driver": drv.replace("_", " ").title(),
                        "Respondents": info.get("respondents", 0),
                        "Score (0-100)": score,
                        "Verdict": "🟢 Strong" if (score or 0) >= ENGAGEMENT_HIGH_THRESHOLD else
                                    "🟡 Moderate" if (score or 0) >= ENGAGEMENT_MODERATE_THRESHOLD else
                                    "🔴 Weak",
                    })
                df_drv = pd.DataFrame(rows)
                st.dataframe(df_drv, use_container_width=True, hide_index=True)

                # Bar chart
                chart_data = pd.DataFrame({
                    "Score": [r["Score (0-100)"] or 0 for r in rows]
                }, index=[r["Driver"] for r in rows])
                st.bar_chart(chart_data)
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Engagement #64: drivers across {len(rows)}")

        # ──── Sentiment ────
        with ep_sub_tabs[3]:
            st.markdown(
                "**Sentiment Score** (rule-based keyword scoring)")
            st.caption(
                "Engine returns rule-based sentiment ∈ {-1.0, 0.0, 1.0}. "
                "ML-based sentiment classification is deferred to v7+ "
                "(spec deviation #7 documented).")
            text = st.text_area("Survey text response",
                                  value="I love the team and feel supported by my manager",
                                  height=80, key="ep_sent_text")
            if st.button("Compute sentiment", key="ep_sent_btn", type="primary"):
                r = EmployeeEngagementEngine.sentiment_score(text)
                sentiment = r.get("rule_based_sentiment")
                meta = r.get("rule_based_meta", {})
                color = "#10B981" if sentiment == 1.0 else "#DC2626" if sentiment == -1.0 else "#6B7280"
                label = "POSITIVE" if sentiment == 1.0 else "NEGATIVE" if sentiment == -1.0 else "NEUTRAL"

                st.markdown(
                    f"<div style='padding:14px;background:{color}22;"
                    f"border-left:6px solid {color};border-radius:10px;text-align:center'>"
                    f"<div style='font-size:24px;font-weight:800;color:{color}'>"
                    f"{label}</div>"
                    f"<div style='font-size:14px;margin-top:6px'>Score: {sentiment}</div></div>",
                    unsafe_allow_html=True)

                pos_hits = meta.get("positive_hits", [])
                neg_hits = meta.get("negative_hits", [])
                if pos_hits:
                    st.success("**Positive keywords:** " + ", ".join(pos_hits))
                if neg_hits:
                    st.error("**Negative keywords:** " + ", ".join(neg_hits))
                if not pos_hits and not neg_hits:
                    st.info("ℹ No positive or negative keywords detected — neutral.")
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Engagement #64: sentiment={sentiment}")

        # ──── Flight risk + v5.98 depth ────
        with ep_sub_tabs[4]:
            _fr_inner = st.tabs([
                "🚨 Single-Staff Flight Risk (existing)",
                "📋 Engagement Executive Scorecard (v5.98)",
                "🎯 Flight Risk Batch (v5.98)",
                "💬 Aggregate Sentiment (v5.98)",
                "🎚️ Driver Investment Map (v5.98)",
            ])

            with _fr_inner[0]:
                st.markdown(
                    "**Flight Risk Indicators** — composite signal across 5 factors")
                st.caption(
                    f"HIGH ≥{FLIGHT_RISK_HIGH_THRESHOLD} · MEDIUM ≥{FLIGHT_RISK_MEDIUM_THRESHOLD}. "
                    "Used for proactive retention conversations.")
                c1, c2 = st.columns(2)
                with c1:
                    fr_id = st.text_input("Staff ID", value="S004", key="ep_fr_id")
                    fr_eng = st.number_input("Engagement score (0-100)",
                                                min_value=0.0, max_value=100.0,
                                                value=40.0, step=5.0, key="ep_fr_eng")
                    fr_promo = st.number_input("Years since last promotion",
                                                  min_value=0.0, value=4.5, step=0.5,
                                                  key="ep_fr_promo")
                with c2:
                    fr_comp = st.number_input("Compensation percentile (0-100)",
                                                 min_value=0.0, max_value=100.0,
                                                 value=20.0, step=5.0, key="ep_fr_comp",
                                                 help="20 = bottom quintile of pay band.")
                    fr_tenure = st.number_input("Tenure (years)",
                                                  min_value=0.0, value=6.0, step=0.5,
                                                  key="ep_fr_tenure")
                    fr_ratings = st.multiselect("Last two ratings",
                                                  list(RATING_LEVELS),
                                                  default=["DEVELOPING", "DEVELOPING"],
                                                  key="ep_fr_ratings",
                                                  max_selections=2)

                if st.button("Compute flight risk",
                               key="ep_fr_btn", type="primary"):
                    signals = StaffSignals(
                        staff_id=fr_id,
                        engagement_score=fr_eng,
                        last_promotion_years_ago=fr_promo,
                        compensation_percentile=fr_comp,
                        last_two_ratings=list(fr_ratings),
                        tenure_years=fr_tenure,
                    )
                    r = EmployeeEngagementEngine.flight_risk_indicators(signals)
                    score = r.get("score")
                    severity = r.get("severity")
                    triggered = r.get("triggered_factors", [])
                    colors = {"HIGH": "#DC2626", "MEDIUM": "#F59E0B",
                              "LOW": "#10B981", None: "#6B7280"}
                    color = colors.get(severity, "#6B7280")

                    k1, k2 = st.columns(2)
                    k1.metric("Score", score)
                    with k2:
                        st.markdown(
                            f"<div style='padding:8px 12px;background:{color}22;"
                            f"border-left:4px solid {color};border-radius:8px;text-align:center'>"
                            f"<div style='font-size:18px;font-weight:800;color:{color}'>"
                            f"{severity}</div></div>", unsafe_allow_html=True)

                    if triggered:
                        st.markdown("**Triggered factors:**")
                        for f in triggered:
                            st.markdown(f"- {f.replace('_', ' ')}")
                    missing = r.get("missing_signals", [])
                    if missing:
                        st.warning(
                            "⚠ Missing signals (Rule 6 transparency): "
                            + ", ".join(missing))
                    audit_log("IFRS_ENGINE_USED", uname,
                               f"Engagement #64: flight risk {fr_id} score={score} {severity}")

            # ────────── _fr_inner[1]: Engagement Executive Scorecard ──────────
            with _fr_inner[1]:
                st.markdown(
                    "**Engagement Executive Scorecard** — single-screen summary "
                    "combining engagement_score + eNPS + drivers_breakdown into "
                    "GREEN/AMBER/RED verdict for board reporting (mirrors v5.97 "
                    "Compensation Executive Scorecard pattern).")
                st.caption(
                    "Uses the same demo survey responses from sub-tabs above. "
                    "Click compute to refresh all 3 engine paths in one shot.")

                if st.button("📋 Compute engagement scorecard",
                               key="ep_es_btn", type="primary"):
                    es_responses = _demo_survey()
                    eng_r = EmployeeEngagementEngine.engagement_score(es_responses)
                    enps_r = EmployeeEngagementEngine.enps(es_responses)
                    drv_r = EmployeeEngagementEngine.drivers_breakdown(es_responses)

                    eng_score = float(eng_r.get("score", 0))
                    eng_sev = eng_r.get("severity") or "—"
                    enps_v = enps_r.get("enps")
                    promoters = int(enps_r.get("promoter_count", 0))
                    detractors = int(enps_r.get("detractor_count", 0))

                    # Severity → color
                    sev_colors = {"HIGH": "#10B981", "MODERATE": "#F59E0B",
                                   "LOW": "#DC2626"}

                    st.markdown("### 1️⃣ Engagement score")
                    eng_color = sev_colors.get(eng_sev, "#6B7280")
                    e1, e2, e3 = st.columns(3)
                    e1.metric("Respondents", eng_r.get("respondents"))
                    e2.metric("Abstained", eng_r.get("abstained"))
                    e3.markdown(
                        f"<div style='padding:8px;background:{eng_color}22;"
                        f"border-left:4px solid {eng_color};border-radius:6px'>"
                        f"<div style='font-size:10px;opacity:0.7'>SCORE / SEVERITY</div>"
                        f"<div style='font-size:18px;font-weight:700;color:{eng_color}'>"
                        f"{eng_score:.0f} — {eng_sev}</div></div>",
                        unsafe_allow_html=True)

                    st.markdown("### 2️⃣ eNPS")
                    enps_color = ("#10B981" if (enps_v or 0) > 30
                                    else "#F59E0B" if (enps_v or 0) > 0
                                    else "#DC2626")
                    n1, n2, n3, n4 = st.columns(4)
                    n1.metric("Promoters", promoters)
                    n2.metric("Passives", enps_r.get("passive_count"))
                    n3.metric("Detractors", detractors,
                                delta_color="inverse" if detractors > 0 else "normal")
                    with n4:
                        st.markdown(
                            f"<div style='padding:8px;background:{enps_color}22;"
                            f"border-left:4px solid {enps_color};border-radius:6px'>"
                            f"<div style='font-size:10px;opacity:0.7'>eNPS</div>"
                            f"<div style='font-size:18px;font-weight:700;color:{enps_color}'>"
                            f"{enps_v if enps_v is not None else '—'}</div></div>",
                            unsafe_allow_html=True)

                    st.markdown("### 3️⃣ Drivers — strongest & weakest")
                    drv_pairs = [(d, float(drv_r.get(d, {}).get("score", 0)))
                                  for d in ENGAGEMENT_DRIVERS]
                    drv_pairs.sort(key=lambda p: -p[1])
                    strongest = drv_pairs[0]
                    weakest = drv_pairs[-1]

                    d1, d2 = st.columns(2)
                    d1.markdown(
                        f"<div style='padding:8px;background:#10B98122;"
                        f"border-left:4px solid #10B981;border-radius:6px'>"
                        f"<div style='font-size:10px;opacity:0.7'>STRONGEST</div>"
                        f"<div style='font-size:14px;font-weight:700;color:#10B981'>"
                        f"{strongest[0]} ({strongest[1]:.0f})</div></div>",
                        unsafe_allow_html=True)
                    d2.markdown(
                        f"<div style='padding:8px;background:#DC262622;"
                        f"border-left:4px solid #DC2626;border-radius:6px'>"
                        f"<div style='font-size:10px;opacity:0.7'>WEAKEST (invest here)</div>"
                        f"<div style='font-size:14px;font-weight:700;color:#DC2626'>"
                        f"{weakest[0]} ({weakest[1]:.0f})</div></div>",
                        unsafe_allow_html=True)

                    # Overall verdict
                    st.markdown("### 4️⃣ Overall scorecard verdict")
                    issues = []
                    if eng_sev == "LOW":
                        issues.append(f"engagement score is LOW ({eng_score:.0f})")
                    if (enps_v or 0) <= 0:
                        issues.append(f"eNPS is ≤0 ({enps_v})")
                    if weakest[1] < 50:
                        issues.append(
                            f"{weakest[0]} driver is critically low ({weakest[1]:.0f})")
                    if detractors > promoters:
                        issues.append(
                            f"detractors ({detractors}) outnumber promoters ({promoters})")

                    if not issues:
                        st.success(
                            "✅ **Engagement health: GREEN.** All metrics in healthy "
                            "ranges. Maintain via quarterly pulse + driver focus.")
                    elif len(issues) <= 1:
                        st.warning(
                            f"⚠ **Engagement health: AMBER.** Issue: "
                            f"{issues[0]}. Targeted intervention recommended.")
                    else:
                        st.error(
                            f"🚨 **Engagement health: RED.** Multiple issues: "
                            f"{'; '.join(issues)}. Comprehensive engagement "
                            "review required.")

                    # ── v6.0 Composite Scoring Layer ──
                    st.markdown("### 5️⃣ Workforce Health Composite (v6.0)")
                    st.caption(
                        "💡 **v6.0 composite scoring layer** — pure-Python "
                        "composition over existing engine outputs, no engine "
                        "modifications. Combines engagement_score (0.40) + "
                        "eNPS normalised (0.25) + weakest driver (0.20) + "
                        "inverse flight risk (0.15) into single 0-100 board-"
                        "ready number with HEALTHY/MODERATE/LOW severity bands.")

                    from utils.composite_scores import (workforce_health_composite,
                                                          WORKFORCE_HEALTH_WEIGHTS)
                    # Approximate flight_risk_high_pct from sample portfolio
                    # (not ideal but sufficient for demo composite)
                    fr_high_estimate = 25.0  # placeholder — production would
                    # compute from actual flight risk batch
                    composite = workforce_health_composite(
                        engagement_score=eng_score,
                        enps=float(enps_v) if enps_v is not None else None,
                        weakest_driver_score=weakest[1],
                        flight_risk_high_pct=fr_high_estimate,
                    )

                    comp_score = composite.get("score")
                    comp_severity = composite.get("severity", "UNKNOWN")
                    comp_color = {"HEALTHY": "#10B981", "MODERATE": "#F59E0B",
                                    "LOW": "#DC2626", "UNKNOWN": "#6B7280"}[comp_severity]

                    cs1, cs2 = st.columns([1, 2])
                    with cs1:
                        comp_display = (f"{comp_score:.1f}"
                                         if comp_score is not None else "—")
                        st.markdown(
                            f"<div style='padding:14px;background:{comp_color}22;"
                            f"border-left:6px solid {comp_color};border-radius:10px;text-align:center'>"
                            f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                            f"COMPOSITE SCORE</div>"
                            f"<div style='font-size:32px;font-weight:800;color:{comp_color};margin-top:4px'>"
                            f"{comp_display}</div>"
                            f"<div style='font-size:13px;margin-top:6px'>"
                            f"<b>{comp_severity}</b></div></div>",
                            unsafe_allow_html=True)
                    with cs2:
                        st.markdown("**Component contributions** (after weight):")
                        comp_rows = []
                        for comp_name, comp_value in composite.get(
                                "components", {}).items():
                            comp_rows.append({
                                "Component": comp_name.replace("_", " ").title(),
                                "Contribution": f"{comp_value:.1f}",
                            })
                        if comp_rows:
                            st.dataframe(pd.DataFrame(comp_rows),
                                         use_container_width=True,
                                         hide_index=True)

                    if composite.get("missing_inputs"):
                        st.caption(
                            f"⚠ Rule 6 transparency — missing inputs: "
                            f"{', '.join(composite['missing_inputs'])}. "
                            "Composite renormalised over available components.")

                    st.caption(
                        f"💡 Flight risk estimate ({fr_high_estimate}%) is a "
                        "placeholder. Production deployment would compute "
                        "from actual flight risk batch results across staff.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Engagement #64 (depth): scorecard issues={len(issues)} "
                                f"score={eng_score:.0f} enps={enps_v} "
                                f"weakest_driver={weakest[0]}={weakest[1]:.0f} "
                                f"composite={comp_score} sev={comp_severity}")

            # ────────── _fr_inner[2]: Flight Risk Batch ──────────
            with _fr_inner[2]:
                st.markdown(
                    "**Flight Risk Batch Analysis** — runs flight_risk_indicators "
                    "across a portfolio of staff signals. v5.79 surfaces single-staff "
                    "flight risk; this batch view enables proactive retention list "
                    "generation across an entire team or branch.")
                st.caption(
                    "Synthetic 8-staff portfolio with varied signal profiles. "
                    "Production deployment would feed via staff_register joined to "
                    "BSC engagement scores, compensation percentiles, and rating "
                    "history.")

                if st.button("🎯 Run flight risk batch",
                               key="ep_frb_btn", type="primary"):
                    # Synthetic 8-staff signals with variety
                    batch_signals = [
                        StaffSignals("S001", engagement_score=82, last_promotion_years_ago=1.5,
                                       compensation_percentile=60, last_two_ratings=["EXCEEDS", "EXCEEDS"],
                                       tenure_years=4.0),  # LOW
                        StaffSignals("S002", engagement_score=70, last_promotion_years_ago=2.0,
                                       compensation_percentile=50, last_two_ratings=["MEETS_PLUS", "MEETS"],
                                       tenure_years=3.0),  # LOW/MEDIUM
                        StaffSignals("S003", engagement_score=45, last_promotion_years_ago=3.5,
                                       compensation_percentile=30, last_two_ratings=["MEETS", "MEETS"],
                                       tenure_years=3.5),  # MEDIUM
                        StaffSignals("S004", engagement_score=35, last_promotion_years_ago=4.5,
                                       compensation_percentile=20, last_two_ratings=["DEVELOPING", "BELOW"],
                                       tenure_years=3.5),  # HIGH
                        StaffSignals("S005", engagement_score=55, last_promotion_years_ago=2.5,
                                       compensation_percentile=40, last_two_ratings=["MEETS", "MEETS"],
                                       tenure_years=2.5),  # MEDIUM
                        StaffSignals("S006", engagement_score=38, last_promotion_years_ago=4.0,
                                       compensation_percentile=22, last_two_ratings=["DEVELOPING", "DEVELOPING"],
                                       tenure_years=4.5),  # HIGH
                        StaffSignals("S007", engagement_score=78, last_promotion_years_ago=1.0,
                                       compensation_percentile=70, last_two_ratings=["EXCEEDS", "MEETS_PLUS"],
                                       tenure_years=2.0),  # LOW
                        StaffSignals("S008", engagement_score=30, last_promotion_years_ago=5.0,
                                       compensation_percentile=18, last_two_ratings=["DEVELOPING", "DEVELOPING"],
                                       tenure_years=3.0),  # HIGH
                    ]

                    batch_results = []
                    for sig in batch_signals:
                        r = EmployeeEngagementEngine.flight_risk_indicators(sig)
                        batch_results.append({
                            "Staff ID": sig.staff_id,
                            "Engagement": sig.engagement_score,
                            "Comp percentile": sig.compensation_percentile,
                            "Score": int(r.get("score", 0)),
                            "Severity": r.get("severity") or "—",
                            "Triggered factors": len(r.get("triggered_factors", [])),
                        })

                    # Sort by score desc (highest risk first)
                    batch_results.sort(key=lambda x: -x["Score"])

                    # Severity counts
                    high_count = sum(1 for r in batch_results
                                       if r["Severity"] == "HIGH")
                    medium_count = sum(1 for r in batch_results
                                         if r["Severity"] == "MEDIUM")
                    low_count = sum(1 for r in batch_results
                                      if r["Severity"] == "LOW")

                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Total assessed", len(batch_results))
                    k2.metric("HIGH risk", high_count,
                                delta_color="inverse" if high_count > 0 else "normal")
                    k3.metric("MEDIUM risk", medium_count)
                    k4.metric("LOW risk", low_count)

                    # Table with severity-tinted rows
                    st.markdown("**Sorted by flight risk (highest first):**")
                    display_rows = []
                    sev_emoji = {"HIGH": "🚨", "MEDIUM": "⚠", "LOW": "✅"}
                    for r in batch_results:
                        display_rows.append({
                            "Status": sev_emoji.get(r["Severity"], "⚪"),
                            "Staff ID": r["Staff ID"],
                            "Engagement": r["Engagement"],
                            "Comp percentile": r["Comp percentile"],
                            "Score": r["Score"],
                            "Severity": r["Severity"],
                            "Factors triggered": r["Triggered factors"],
                        })
                    st.dataframe(pd.DataFrame(display_rows),
                                 use_container_width=True, hide_index=True)

                    # Recommendation
                    if high_count > 0:
                        st.error(
                            f"🚨 **{high_count} HIGH-risk staff** require "
                            "immediate retention conversations. Prioritize 1:1s "
                            "with manager + comp/promotion review.")
                    if medium_count >= 2:
                        st.warning(
                            f"⚠ **{medium_count} MEDIUM-risk staff** — "
                            "schedule check-ins within 30 days. Pulse-check "
                            "engagement drivers.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Engagement #64 (depth): flight risk batch "
                                f"high={high_count} medium={medium_count} "
                                f"low={low_count}")

            # ────────── _fr_inner[3]: Aggregate Sentiment ──────────
            with _fr_inner[3]:
                st.markdown(
                    "**Aggregate Sentiment Analysis** — runs sentiment_score across "
                    "all survey text responses, surfaces overall sentiment "
                    "distribution + most-cited keywords.")
                st.caption(
                    "v5.79 surfaces single-text sentiment. This batch view enables "
                    "macro analysis: \"how does the workforce feel overall?\" + "
                    "\"which themes (keywords) come up most often?\"")

                if st.button("💬 Compute aggregate sentiment",
                               key="ep_as_btn", type="primary"):
                    as_responses = _demo_survey()
                    sentiments = []
                    all_pos_hits = []
                    all_neg_hits = []
                    for resp in as_responses:
                        if resp.text_response:
                            r = EmployeeEngagementEngine.sentiment_score(
                                resp.text_response)
                            sentiments.append({
                                "Staff": resp.staff_id,
                                "Text": resp.text_response,
                                "Sentiment": r.get("rule_based_sentiment"),
                                "Pos hits": r.get("rule_based_meta", {}).get("positive_hits", []),
                                "Neg hits": r.get("rule_based_meta", {}).get("negative_hits", []),
                            })
                            all_pos_hits.extend(
                                r.get("rule_based_meta", {}).get("positive_hits", []))
                            all_neg_hits.extend(
                                r.get("rule_based_meta", {}).get("negative_hits", []))

                    # Distribution
                    pos = sum(1 for s in sentiments if s["Sentiment"] == 1.0)
                    neg = sum(1 for s in sentiments if s["Sentiment"] == -1.0)
                    neu = sum(1 for s in sentiments if s["Sentiment"] == 0.0)
                    total = len(sentiments)

                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Total responses", total)
                    k2.metric("Positive 😊", pos,
                                f"{pos/total*100:.0f}%" if total else "—")
                    k3.metric("Neutral 😐", neu,
                                f"{neu/total*100:.0f}%" if total else "—")
                    k4.metric("Negative 😟", neg,
                                f"{neg/total*100:.0f}%" if total else "—",
                                delta_color="inverse" if neg > 0 else "normal")

                    # Net sentiment
                    net_sentiment = (pos - neg) / total * 100 if total else 0
                    net_color = ("#10B981" if net_sentiment > 30
                                  else "#F59E0B" if net_sentiment > 0
                                  else "#DC2626")
                    st.markdown(
                        f"<div style='padding:14px;background:{net_color}22;"
                        f"border-left:6px solid {net_color};border-radius:10px'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                        f"NET SENTIMENT</div>"
                        f"<div style='font-size:24px;font-weight:800;color:{net_color}'>"
                        f"{net_sentiment:+.0f}</div>"
                        f"<div style='font-size:11px;margin-top:4px;opacity:0.85'>"
                        f"(% positive – % negative)</div></div>",
                        unsafe_allow_html=True)

                    # Keyword frequencies
                    from collections import Counter as _C
                    pos_freq = _C(all_pos_hits).most_common(10)
                    neg_freq = _C(all_neg_hits).most_common(10)

                    kc1, kc2 = st.columns(2)
                    with kc1:
                        st.markdown("**Top positive keywords:**")
                        if pos_freq:
                            pos_rows = [{"Keyword": k, "Mentions": v}
                                         for k, v in pos_freq]
                            st.dataframe(pd.DataFrame(pos_rows),
                                         use_container_width=True,
                                         hide_index=True)
                        else:
                            st.caption("None.")
                    with kc2:
                        st.markdown("**Top negative keywords:**")
                        if neg_freq:
                            neg_rows = [{"Keyword": k, "Mentions": v}
                                         for k, v in neg_freq]
                            st.dataframe(pd.DataFrame(neg_rows),
                                         use_container_width=True,
                                         hide_index=True)
                        else:
                            st.caption("None.")

                    # Per-staff breakdown
                    with st.expander("Per-staff sentiment breakdown"):
                        per_rows = [{
                            "Staff": s["Staff"],
                            "Sentiment": ("POS" if s["Sentiment"] == 1.0
                                            else "NEG" if s["Sentiment"] == -1.0
                                            else "NEU"),
                            "Text": (s["Text"][:60] + "...")
                                    if len(s["Text"]) > 60 else s["Text"],
                        } for s in sentiments]
                        st.dataframe(pd.DataFrame(per_rows),
                                     use_container_width=True, hide_index=True)

                    # Spec deviation reminder
                    st.caption(
                        "💡 Per Rule 7 + spec deviation #7 — engine uses "
                        "rule-based keyword sentiment. ML sentiment is "
                        "downstream work. Production deployment can plug in "
                        "ML model via `ml_sentiment_fn` callback per call.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Engagement #64 (depth): aggregate sentiment "
                                f"pos={pos} neu={neu} neg={neg} net={net_sentiment:.0f}")

            # ────────── _fr_inner[4]: Driver Investment Map ──────────
            with _fr_inner[4]:
                st.markdown(
                    "**Driver Investment Map** — drivers_breakdown ranked by score "
                    "with investment priority guidance. v5.79 surfaces driver "
                    "scores; this view ranks them and surfaces actionable "
                    "investment priorities.")
                st.caption(
                    f"6 engagement drivers: {', '.join(ENGAGEMENT_DRIVERS)}. "
                    "Investment priority = inverse rank (weakest driver gets "
                    "highest priority).")

                if st.button("🎚️ Compute driver investment map",
                               key="ep_dim_btn", type="primary"):
                    dim_responses = _demo_survey()
                    drv = EmployeeEngagementEngine.drivers_breakdown(dim_responses)

                    # Build ranked list
                    driver_pairs = [(d, float(drv.get(d, {}).get("score", 0)),
                                      int(drv.get(d, {}).get("respondents", 0)),
                                      int(drv.get(d, {}).get("missing_count", 0)))
                                     for d in ENGAGEMENT_DRIVERS]
                    driver_pairs.sort(key=lambda p: p[1])  # ascending by score

                    # Investment priority bands
                    priority_rows = []
                    for rank, (driver, score, respondents, missing) in enumerate(driver_pairs, 1):
                        if score < 50:
                            priority = "🔴 CRITICAL — invest immediately"
                            color = "#DC2626"
                        elif score < 65:
                            priority = "🟡 IMPORTANT — invest within 6 months"
                            color = "#F59E0B"
                        elif score < 75:
                            priority = "🟢 MONITOR — annual review sufficient"
                            color = "#10B981"
                        else:
                            priority = "✅ STRONG — maintain current programs"
                            color = "#10B981"
                        priority_rows.append({
                            "Rank": rank,
                            "Driver": driver,
                            "Score": score,
                            "Respondents": respondents,
                            "Missing": missing,
                            "Investment priority": priority,
                            "_color": color,
                        })

                    # Display table
                    display_rows = [{k: v for k, v in r.items() if k != "_color"}
                                     for r in priority_rows]
                    st.dataframe(pd.DataFrame(display_rows),
                                 use_container_width=True, hide_index=True)

                    # Bar chart
                    chart_data = pd.DataFrame({
                        "Score": [p[1] for p in driver_pairs]
                    }, index=[p[0] for p in driver_pairs])
                    st.markdown("**Driver scores (ascending — weakest first):**")
                    st.bar_chart(chart_data)

                    # Critical drivers callout
                    critical = [r for r in priority_rows if "CRITICAL" in r["Investment priority"]]
                    important = [r for r in priority_rows if "IMPORTANT" in r["Investment priority"]]
                    if critical:
                        st.error(
                            f"🔴 **{len(critical)} driver(s) at CRITICAL level**: "
                            f"{', '.join(r['Driver'] for r in critical)}. "
                            "Immediate action required — these are the largest "
                            "engagement levers.")
                    if important:
                        st.warning(
                            f"🟡 **{len(important)} driver(s) at IMPORTANT level**: "
                            f"{', '.join(r['Driver'] for r in important)}. "
                            "Plan investment within 6-month cycle.")
                    if not critical and not important:
                        st.success(
                            "✅ All drivers are at MONITOR or STRONG levels. "
                            "Maintain current engagement programs.")

                    # Concentration insight
                    weakest_score = driver_pairs[0][1]
                    strongest_score = driver_pairs[-1][1]
                    spread = strongest_score - weakest_score
                    if spread > 25:
                        st.info(
                            f"💡 **Wide driver spread ({spread:.0f} points)** — "
                            "engagement is uneven across drivers. Targeted "
                            "intervention on weakest drivers will lift overall "
                            "score most efficiently.")

                    audit_log("IFRS_ENGINE_USED", uname,
                                f"Engagement #64 (depth): driver map "
                                f"weakest={driver_pairs[0][0]}={driver_pairs[0][1]:.0f} "
                                f"critical={len(critical)} important={len(important)}")

        # ──── Rating distribution ────
        with ep_sub_tabs[5]:
            st.markdown("**Performance Rating Distribution** vs CALIBRATION_TARGETS")
            st.caption(
                f"5-level rating scale: {' / '.join(RATING_LEVELS)}. "
                "Targets enforce calibration discipline — bell-curve distribution rather than rating drift.")

            # Use demo reviews for current period
            demo_reviews = [
                PerformanceReview("V1", "S001", "2026-Q1", "EXCEEDS", "M001"),
                PerformanceReview("V2", "S002", "2026-Q1", "MEETS_PLUS", "M001"),
                PerformanceReview("V3", "S003", "2026-Q1", "MEETS", "M001"),
                PerformanceReview("V4", "S004", "2026-Q1", "DEVELOPING", "M002"),
                PerformanceReview("V5", "S005", "2026-Q1", "MEETS", "M002"),
                PerformanceReview("V6", "S006", "2026-Q1", "MEETS", "M002"),
                PerformanceReview("V7", "S007", "2026-Q1", "EXCEEDS", "M003"),
                PerformanceReview("V8", "S008", "2026-Q1", "MEETS_PLUS", "M003"),
            ]
            if st.button("Compute distribution",
                           key="ep_rd_btn", type="primary"):
                r = PerformanceTalentEngine.rating_distribution(
                    demo_reviews, "2026-Q1")
                k1, k2 = st.columns(2)
                k1.metric("Total rated", r.get("total_rated"))
                k2.metric("Calibration compliant",
                           "✅ YES" if r.get("calibration_compliant") else "🔴 NO")

                rows = []
                for level in RATING_LEVELS:
                    info = r.get("distribution", {}).get(level, {})
                    pct = info.get("pct", 0)
                    target = info.get("target_range_pct", [0, 0])
                    rows.append({
                        "Rating": level,
                        "Count": info.get("count", 0),
                        "Actual %": pct,
                        "Target %": f"{target[0]}-{target[1]}",
                        "In target": "✅" if info.get("in_target") else "🔴",
                    })
                st.dataframe(pd.DataFrame(rows),
                             use_container_width=True, hide_index=True)

                # High-potential pipeline
                hp = PerformanceTalentEngine.high_potential_pipeline(
                    demo_reviews, periods_required=2)
                st.caption(
                    f"**High-potential pipeline** (≥2 EXCEEDS periods): "
                    f"{hp.get('hipo_count', 0)} hipos identified — "
                    + ", ".join(hp.get("hipo_staff_ids", [])))
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Performance #64: rating dist 2026-Q1 compliant={r.get('calibration_compliant')}")

        # ──── Succession bench ────
        with ep_sub_tabs[6]:
            st.markdown(
                "**Succession Bench Strength** — % of critical roles with READY_NOW successor")
            st.caption(
                f"AT_RISK <{BENCH_AT_RISK_PCT}% · CRITICAL <{(BENCH_AT_RISK_PCT/2):.0f}% · "
                f"HEALTHY ≥{BENCH_HEALTHY_PCT}%.")

            # Demo succession plans
            demo_plans = [
                SuccessionPlan("P1", "MD", "S001", "S002", "READY_NOW", is_critical_role=True),
                SuccessionPlan("P2", "Head IT", "S010", "S011", "READY_1_YEAR", is_critical_role=True),
                SuccessionPlan("P3", "CRO", "S020", "S021", "READY_2_YEAR", is_critical_role=True),
                SuccessionPlan("P4", "Branch Mgr", "S030", "S031", "NOT_READY", is_critical_role=False),
            ]
            if st.button("Compute bench strength",
                           key="ep_bench_btn", type="primary"):
                r = PerformanceTalentEngine.succession_bench_strength(demo_plans)
                bench_pct = r.get("bench_strength_pct")
                severity = r.get("severity")
                colors = {"HEALTHY": "#10B981", "AT_RISK": "#F59E0B",
                          "CRITICAL": "#DC2626"}
                color = colors.get(severity, "#6B7280")

                k1, k2, k3 = st.columns(3)
                k1.metric("Critical roles", r.get("critical_role_count"))
                k2.metric("READY_NOW successors",
                           r.get("roles_with_ready_now_successor"))
                with k3:
                    st.markdown(
                        f"<div style='padding:8px 12px;background:{color}22;"
                        f"border-left:4px solid {color};border-radius:8px;text-align:center'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>BENCH</div>"
                        f"<div style='font-size:20px;font-weight:800;color:{color}'>"
                        f"{bench_pct}% — {severity}</div></div>",
                        unsafe_allow_html=True)

                roles_at_risk = r.get("roles_at_risk", [])
                if roles_at_risk:
                    st.error(
                        "🔴 **Roles at risk** (no READY_NOW successor): "
                        + ", ".join(roles_at_risk))
                else:
                    st.success("✅ All critical roles have READY_NOW successors.")
                audit_log("IFRS_ENGINE_USED", uname,
                           f"Performance #64: bench strength {bench_pct}% {severity}")

    # ════════════════════════════════════════════════════════════════
    # SUB-TAB 4: Predictive Performance (Standards #20 + #21, integrated v5.84)
    # ════════════════════════════════════════════════════════════════
    with sub[4]:
        from utils.predictive_performance import (
            PredictivePerformance, KPIPrediction,
            ACCURACY_TOLERANCE_PCT, SPEC_ACCURACY_TARGET,
            DEFAULT_BASE_SPREAD,
        )
        from utils.performance_insights import (
            get_performance_insights,
            STRENGTH_THRESHOLD_PCT, DEFAULT_MAX_STRENGTHS,
        )
        from datetime import date as _date_pp, timedelta as _td_pp
        from decimal import Decimal as _D_pp

        st.markdown(
            f"**Standards #20 + #21 — Forward-Looking Performance Analytics**. "
            f"#20 Performance Insights (strengths, promotion readiness, overall score) + "
            f"#21 KPI Prediction (linear extrapolation of pace-to-target with probability)."
        )
        st.caption(
            f"Predictive engine: linear_extrapolation model · "
            f"base_spread={DEFAULT_BASE_SPREAD} · "
            f"accuracy tolerance ±{int(ACCURACY_TOLERANCE_PCT*100)}% · "
            f"spec accuracy target {int(SPEC_ACCURACY_TARGET*100)}%. "
            f"Strengths threshold ≥{STRENGTH_THRESHOLD_PCT}% achievement, "
            f"top {DEFAULT_MAX_STRENGTHS} returned."
        )

        pp_sub_tabs = st.tabs([
            "🔮 KPI Achievement Forecast (#21)",
            "✨ Performance Insights (#20)",
            "🌳 Engine Reference",
        ])

        # ──────── KPI Achievement Forecast ────────
        with pp_sub_tabs[0]:
            st.markdown(
                "**Predict end-of-period KPI achievement** based on current pace.")
            st.caption(
                "Linear extrapolation: pace = current_value / days_elapsed × total_days. "
                "Probability factors in days-remaining (more spread early in period). "
                "Engine surfaces `kpis_skipped` when target/actual missing (Rule 6).")

            c1, c2 = st.columns(2)
            with c1:
                pp_staff = st.text_input("Staff code",
                                            value="300001",
                                            key="pp_staff",
                                            help="Use staff_code from users register.")
                pp_today_str = st.date_input("Snapshot date",
                                                value=_date_pp.today(),
                                                key="pp_today",
                                                help="Date for the prediction snapshot.")
            with c2:
                # Demo KPI inputs — user editable
                st.caption("**Demo KPI inputs** (edit as needed):")
                kpi_n = st.number_input("Number of KPIs",
                                          min_value=1, max_value=8, value=3,
                                          key="pp_kpi_n")

            # Build KPI table editor
            st.markdown("**Demo KPI definitions:**")
            kpi_default = [
                ("K001", "Revenue Growth", 100.0, 60.0),
                ("K002", "NPS Score", 70.0, 65.0),
                ("K003", "Cost Income Ratio", 40.0, 38.0),
                ("K004", "Customer Retention", 90.0, 85.0),
                ("K005", "Loan Quality", 5.0, 8.0),
                ("K006", "Cross-Sell Ratio", 1.5, 1.8),
                ("K007", "Digital Adoption", 75.0, 60.0),
                ("K008", "Branch Compliance", 95.0, 92.0),
            ]
            user_kpis = []
            for i in range(int(kpi_n)):
                kc1, kc2, kc3, kc4 = st.columns([1, 2, 1, 1])
                with kc1:
                    kid = st.text_input(f"ID {i+1}",
                                         value=kpi_default[i][0],
                                         key=f"pp_kid_{i}")
                with kc2:
                    knm = st.text_input(f"Name {i+1}",
                                         value=kpi_default[i][1],
                                         key=f"pp_knm_{i}")
                with kc3:
                    ktg = st.number_input(f"Target {i+1}",
                                            value=kpi_default[i][2], step=1.0,
                                            key=f"pp_ktg_{i}")
                with kc4:
                    kac = st.number_input(f"Actual {i+1}",
                                            value=kpi_default[i][3], step=1.0,
                                            key=f"pp_kac_{i}")
                if kid.strip():
                    user_kpis.append((kid.strip(), knm, ktg, kac))

            # Period configuration
            with st.expander("Period configuration"):
                col1, col2 = st.columns(2)
                period_start = col1.date_input("Period start",
                    value=_date_pp(2026, 1, 1), key="pp_pstart")
                period_end = col2.date_input("Period end",
                    value=_date_pp(2026, 3, 31), key="pp_pend")

            if st.button("🔮 Predict achievement",
                           key="pp_pred_btn", type="primary"):
                # Wire the engine with closures
                kpis_data = {kid: (knm, _D_pp(str(ktg)), _D_pp(str(kac)))
                              for (kid, knm, ktg, kac) in user_kpis}

                def _kpis_fn(sc):
                    return [{"id": kid} for kid in kpis_data]

                def _target_fn(sc, kid, period):
                    info = kpis_data.get(kid)
                    return info[1] if info else None

                def _actual_fn(sc, kid, period):
                    info = kpis_data.get(kid)
                    return info[2] if info else None

                def _period_fn(today): return "2026-Q1"

                def _bounds_fn(period):
                    return (period_start, period_end)

                def _elapsed_fn(period, today):
                    delta = (today - period_start).days
                    return max(0, delta)

                pp = PredictivePerformance(
                    active_kpis_fn=_kpis_fn,
                    target_lookup_fn=_target_fn,
                    actual_lookup_fn=_actual_fn,
                    period_fn=_period_fn,
                    period_bounds_fn=_bounds_fn,
                    days_elapsed_fn=_elapsed_fn,
                )

                r = pp.predict_achievement(pp_staff, today=pp_today_str)
                meta = r.get("meta", {})
                preds = r.get("predictions", {})
                overall = float(r.get("overall_prediction", 0))

                # Period progress banner
                total_days = int(meta.get("total_days", 0))
                elapsed = int(meta.get("days_elapsed", 0))
                remaining = int(meta.get("days_remaining", 0))
                progress_pct = (elapsed / total_days * 100) if total_days else 0

                st.markdown(
                    f"<div style='padding:12px;background:rgba(59,130,246,0.1);"
                    f"border-left:4px solid #3B82F6;border-radius:8px'>"
                    f"<b>Period progress</b>: day {elapsed} of {total_days} ({progress_pct:.0f}%) — "
                    f"<b>{remaining} days remaining</b> until period end "
                    f"({meta.get('period_end', '')})"
                    f"</div>", unsafe_allow_html=True)

                # Overall prediction
                if overall >= 0.85:
                    color, label = "#10B981", "ON TRACK"
                elif overall >= 0.50:
                    color, label = "#F59E0B", "AT RISK"
                else:
                    color, label = "#DC2626", "OFF TRACK"

                k1, k2, k3 = st.columns(3)
                k1.metric("Overall prediction probability",
                           f"{overall*100:.1f}%")
                k2.metric("KPIs predicted", meta.get("kpis_predicted"))
                k3.metric("KPIs skipped (Rule 6)",
                           meta.get("kpis_skipped"),
                           help="Skipped due to missing target or actual.")
                st.markdown(
                    f"<div style='padding:14px;background:{color}22;"
                    f"border-left:6px solid {color};border-radius:10px;text-align:center'>"
                    f"<div style='font-size:24px;font-weight:800;color:{color}'>"
                    f"{label}</div></div>", unsafe_allow_html=True)

                # Per-KPI predictions table
                if preds:
                    rows = []
                    for kid, p in preds.items():
                        kid_name = kpis_data.get(kid, (kid, None, None))[0]
                        prob = float(p.get("probability", 0)) * 100
                        prob_emoji = ("🟢" if prob >= 85
                                       else "🟡" if prob >= 50
                                       else "🔴")
                        rows.append({
                            "KPI ID": kid,
                            "Name": kpis_data.get(kid, (kid, "", 0, 0))[0]
                                if kid in kpis_data else kid,
                            "Current": float(p.get("current_value", 0)),
                            "Target": float(p.get("target", 0)),
                            "Predicted": round(float(p.get("predicted_value", 0)), 2),
                            "Pace/day": round(float(p.get("pace_per_day", 0)), 3),
                            "Probability": f"{prob_emoji} {prob:.1f}%",
                        })
                    st.dataframe(pd.DataFrame(rows),
                                 use_container_width=True, hide_index=True)
                    st.caption(
                        "**Reading the table**: Current = today's actual. "
                        "Target = end-of-period goal. Predicted = where pace will land. "
                        "Probability = chance of hitting target by period end based on current pace.")

                audit_log("IFRS_ENGINE_USED", uname,
                           f"Predictive #21: staff={pp_staff} predicted={meta.get('kpis_predicted')} "
                           f"skipped={meta.get('kpis_skipped')} overall={overall:.3f}")

        # ──────── Performance Insights ────────
        with pp_sub_tabs[1]:
            st.markdown(
                "**Performance Insights** — strengths, promotion readiness, overall score")
            st.caption(
                f"Strengths: KPIs ≥{STRENGTH_THRESHOLD_PCT}% achievement, "
                f"top {DEFAULT_MAX_STRENGTHS} returned sorted desc. "
                f"Promotion readiness from growth plan, clamped [0, 1]. "
                f"Engine validates staff_code via users register.")

            pi_staff = st.text_input("Staff code",
                                       value="300001",
                                       key="pi_staff")

            with st.expander("Demo KPI achievement data (editable)"):
                pi_kpi_data = []
                pi_defaults = [
                    ("K001", "Revenue Growth", 125),
                    ("K002", "NPS Score", 111.4),
                    ("K003", "Cost Income Ratio", 95),
                    ("K004", "Customer Retention", 83),
                    ("K005", "Loan Quality", 62.5),
                    ("K006", "Cross-Sell", 120),
                ]
                for i, (kid, knm, kpct) in enumerate(pi_defaults):
                    pc1, pc2, pc3 = st.columns([1, 2, 1])
                    with pc1:
                        kid_v = st.text_input(f"ID", value=kid, key=f"pi_kid_{i}")
                    with pc2:
                        knm_v = st.text_input(f"Name", value=knm, key=f"pi_knm_{i}")
                    with pc3:
                        kpct_v = st.number_input(f"Achievement %",
                                                  value=float(kpct), step=1.0,
                                                  key=f"pi_kpct_{i}")
                    if kid_v.strip():
                        pi_kpi_data.append({
                            "kpi_id": kid_v.strip(),
                            "kpi": knm_v,
                            "achievement_pct": kpct_v,
                        })

            promo_readiness_v = st.slider(
                "Promotion readiness (from growth plan)",
                0.0, 1.0, 0.75, 0.05,
                key="pi_promo")
            overall_v = st.slider("Overall BSC score (0-5)",
                                    0.0, 5.0, 3.6, 0.1, key="pi_overall")

            if st.button("✨ Compute insights",
                           key="pi_btn", type="primary"):
                def _kpi_status_fn(sc):
                    return pi_kpi_data

                def _growth_plan_fn(sc):
                    return {"promotion_readiness": promo_readiness_v}

                def _overall_score_fn(sc, period):
                    return overall_v

                r = get_performance_insights(
                    pi_staff,
                    today=_date_pp.today(),
                    kpi_status_fn=_kpi_status_fn,
                    growth_plan_fn=_growth_plan_fn,
                    overall_score_fn=_overall_score_fn)

                if not r:
                    st.error(
                        f"⛔ Unknown staff_code `{pi_staff}` — engine validates "
                        "against users.json. Use a real staff_code "
                        "(try `300001` for William Mwanake).")
                else:
                    overall = r.get("overall_score", 0)
                    strengths = r.get("strengths", [])
                    promo = r.get("promotion_readiness", 0)
                    meta = r.get("meta", {})

                    k1, k2, k3 = st.columns(3)
                    k1.metric("Overall BSC score",
                               f"{overall:.2f} / 5",
                               help="From BSC engine.")
                    k2.metric("Promotion readiness",
                               f"{promo:.0%}",
                               help="From growth plan, clamped [0, 1].")
                    k3.metric("Strengths identified",
                               len(strengths))

                    # Staff name
                    st.caption(
                        f"📋 Staff: **{meta.get('staff_name', '—')}** "
                        f"({meta.get('staff_code', '—')}) — "
                        f"period {meta.get('period', '—')} · "
                        f"{meta.get('kpi_count', 0)} KPIs evaluated.")

                    # Strengths section
                    if strengths:
                        st.markdown("### 💪 Strengths")
                        st.caption(
                            f"KPIs at ≥{STRENGTH_THRESHOLD_PCT}% achievement, "
                            "sorted desc:")
                        # Match strengths to original KPI names
                        kpi_lookup = {k["kpi_id"]: k for k in pi_kpi_data}
                        strength_rows = []
                        for s_id in strengths:
                            kpi = kpi_lookup.get(s_id, {})
                            strength_rows.append({
                                "KPI ID": s_id,
                                "Name": kpi.get("kpi", "—"),
                                "Achievement %": kpi.get("achievement_pct", 0),
                            })
                        st.dataframe(pd.DataFrame(strength_rows),
                                     use_container_width=True, hide_index=True)
                    else:
                        st.info(
                            f"ℹ No KPIs at ≥{STRENGTH_THRESHOLD_PCT}% achievement — "
                            "no strengths identified this period.")

                    # Signal transparency
                    sig = meta.get("signals_present", {})
                    if not all(sig.values()):
                        missing = [k for k, v in sig.items() if not v]
                        st.warning(
                            f"⚠ Some signals missing: {', '.join(missing)} "
                            "(Rule 6 transparency).")

                    audit_log("IFRS_ENGINE_USED", uname,
                               f"Insights #20: staff={pi_staff} "
                               f"score={overall} strengths={len(strengths)} "
                               f"promo={promo}")

        # ──────── Engine Reference ────────
        with pp_sub_tabs[2]:
            st.markdown("**Engine Constants Reference** (single source of truth)")

            st.markdown("**Predictive Performance (#21):**")
            pp_ref_rows = [
                {"Constant": "DEFAULT_BASE_SPREAD",
                  "Value": DEFAULT_BASE_SPREAD,
                  "Meaning": "Base uncertainty in probability calculation"},
                {"Constant": "ACCURACY_TOLERANCE_PCT",
                  "Value": f"{ACCURACY_TOLERANCE_PCT*100:.0f}%",
                  "Meaning": "Acceptable prediction-to-actual variance"},
                {"Constant": "SPEC_ACCURACY_TARGET",
                  "Value": f"{SPEC_ACCURACY_TARGET*100:.0f}%",
                  "Meaning": "Spec target for prediction accuracy"},
                {"Constant": "Model",
                  "Value": "linear_extrapolation",
                  "Meaning": "Pace = actual / days_elapsed × total_days"},
            ]
            st.dataframe(pd.DataFrame(pp_ref_rows),
                         use_container_width=True, hide_index=True)

            st.markdown("**Performance Insights (#20):**")
            pi_ref_rows = [
                {"Constant": "STRENGTH_THRESHOLD_PCT",
                  "Value": f"{STRENGTH_THRESHOLD_PCT}%",
                  "Meaning": "Minimum achievement % to qualify as strength"},
                {"Constant": "DEFAULT_MAX_STRENGTHS",
                  "Value": DEFAULT_MAX_STRENGTHS,
                  "Meaning": "Maximum strengths returned per staff"},
                {"Constant": "Promotion readiness clamp",
                  "Value": "[0, 1]",
                  "Meaning": "Growth plan value bounded to [0, 1]"},
                {"Constant": "Overall score scale",
                  "Value": "0-5",
                  "Meaning": "BSC engine returns this scale"},
            ]
            st.dataframe(pd.DataFrame(pi_ref_rows),
                         use_container_width=True, hide_index=True)

            st.caption(
                "**Forward-looking analytics** complement the **retrospective analytics** "
                "in Engagement & Performance (#64) sub-tab. Predictive helps *plan* "
                "and *prioritize*; retrospective helps *grade* and *reward*. "
                "Both rely on validated source data — running predictions on stale or "
                "incomplete actuals will produce misleading projections.")

# ── Section 1: 👥 Records ─────────────────────────────
with sections[1]:
    sub = st.tabs([
        "👤 Staff directory",
        "🚪 Exits & attrition",
        "🔄 Transfers",
    ])
    with sub[0]:
        st.subheader("Staff directory")

        if len(staff_scores) == 0:
            st.info("Upload BSC data to view staff directory.")
        else:
            _vis_dir = staff_scores.copy()
            if not can_all:
                from utils.core_audit import get_visible_staff as _gvs2
                _vis_dir = _gvs2(ud, staff_scores)

            # Search and filter
            d1, d2, d3 = st.columns([2, 1, 1])
            _dsearch = d1.text_input("🔍 Search name or unit", key="dir_search")
            _drole   = d2.selectbox("Role", ["All roles"] + sorted(_vis_dir["Role"].unique().tolist()), key="dir_role")
            _dunit   = d3.selectbox("Unit", ["All units"] + sorted(_vis_dir["Unit"].unique().tolist()), key="dir_unit")

            _dir_df = _vis_dir.copy()
            if _dsearch:
                _dir_df = _dir_df[_dir_df["Staff Name"].str.contains(_dsearch, case=False, na=False) |
                                   _dir_df["Unit"].str.contains(_dsearch, case=False, na=False)]
            if _drole != "All roles":
                _dir_df = _dir_df[_dir_df["Role"] == _drole]
            if _dunit != "All units":
                _dir_df = _dir_df[_dir_df["Unit"] == _dunit]

            st.caption(f"Showing {len(_dir_df)} of {len(_vis_dir)} staff")

            # Card grid — 3 per row
            _dir_rows = [_dir_df.iloc[i:i+3] for i in range(0, len(_dir_df), 3)]
            for _drow in _dir_rows:
                _dcols = st.columns(3)
                for _dci, (_, _dr) in enumerate(zip(_dcols, _drow.iterrows())):
                    _dc  = _dcols[_dci]
                    _, _dr = _dr
                    _name = _dr.get("Staff Name","")
                    _role = _dr.get("Role","")
                    _unit = _dr.get("Unit","")
                    _code = str(_dr.get("Staff Code",""))
                    _bsc  = float(_dr.get("Final_BSC_Score",0) or 0)
                    _rem  = _dr.get("Performance_Remark","—")
                    _clr_map = {"Exceeded By Far":"var(--brand-primary,#006B3F)","Exceeded":"var(--brand-mid,#1D9E75)",
                                "Met":"#F5A623","Partially Met":"#E67E22","Unmet":"#E24B4A"}
                    _bclr = _clr_map.get(_rem,"#9CA3AF")
                    _init = "".join(p[0].upper() for p in _name.split()[:2]) if _name else "?"
                    # Check if on leave
                    _on_lv = ""
                    if lm:
                        _lv_active = [l for l in lm.get_active_leave()
                                      if l.get("staff_name","").lower() == _name.lower()]
                        if _lv_active:
                            _on_lv = f"<span style='background:#FEF3C7;color:#92400E;font-size:9px;padding:1px 5px;border-radius:8px;margin-left:4px'>🏖️ On leave</span>"

                    _dc.markdown(
                        f"<div style='padding:14px;background:var(--color-background-primary);border:0.5px solid var(--color-border-tertiary);"
                        f"border-radius:10px;margin-bottom:8px;'>"
                        f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:8px'>"
                        f"<div style='width:38px;height:38px;border-radius:50%;background:{_bclr}20;"
                        f"border:2px solid {_bclr};display:flex;align-items:center;justify-content:center;"
                        f"font-size:13px;font-weight:700;color:{_bclr};flex-shrink:0'>{_init}</div>"
                        f"<div><div style='font-weight:700;font-size:12px;color:var(--color-text-primary)'>{_name}{_on_lv}</div>"
                        f"<div style='font-size:10px;color:var(--color-text-secondary)'>{_role}</div></div></div>"
                        f"<div style='font-size:10px;color:var(--color-text-tertiary);margin-bottom:4px'>📍 {_unit}</div>"
                        f"<div style='display:flex;justify-content:space-between;align-items:center'>"
                        f"<span style='font-size:10px;font-weight:600;color:{_bclr}'>BSC {_bsc:.2f}</span>"
                        f"<span style='font-size:9px;background:{_bclr}15;color:{_bclr};"
                        f"padding:2px 7px;border-radius:8px;font-weight:600'>{_rem}</span>"
                        f"</div></div>",
                        unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 3 — LEAVE MANAGEMENT (was TAB 2)
        # ════════════════════════════════════════════════════════════════
        # TAB 2 — LEAVE MANAGEMENT
        # ════════════════════════════════════════════════════════════════
    with sub[1]:
        st.subheader("Staff exits & attrition analysis")

        ex1, ex2 = st.columns([2,3])

        with ex1:
            st.markdown("#### Record exit")
            with st.form("exit_form"):
                all_staff_names_ex = sorted(staff_scores['Staff Name'].tolist()) if len(staff_scores) else []
                exit_staff  = st.selectbox("Staff member", [""] + all_staff_names_ex, key="ex_staff")
                exit_date   = st.date_input("Exit date", value=date.today(), key="ex_date")
                exit_reason = st.selectbox("Primary reason", EXIT_REASONS, key="ex_reason")
                exit_detail = st.text_area("Detail / exit interview notes", height=80, key="ex_detail")
                ec1, ec2   = st.columns(2)
                tenure     = ec1.number_input("Tenure (years)", min_value=0.0, step=0.5, key="ex_tenure")
                rehire     = ec2.checkbox("Eligible for rehire", value=True, key="ex_rehire")

                if st.form_submit_button("Record exit", type="primary"):
                    if exit_staff:
                        sc_row = staff_scores[staff_scores['Staff Name']==exit_staff]
                        sc  = str(sc_row['Staff Code'].values[0]) if len(sc_row) and 'Staff Code' in sc_row.columns else ''
                        unit = str(sc_row['Unit'].values[0]) if len(sc_row) and 'Unit' in sc_row.columns else ''
                        bsc  = float(sc_row['Final_BSC_Score'].values[0]) if len(sc_row) else None
                        hr_m.record_exit({
                            'staff_code': sc, 'staff_name': exit_staff, 'unit': unit,
                            'exit_date': exit_date, 'reason': exit_reason,
                            'reason_detail': exit_detail, 'final_bsc': bsc,
                            'tenure_years': tenure, 'rehire_eligible': rehire,
                            'recorded_by': uname,
                        })
                        audit_log("EXIT_RECORDED", uname, f"{exit_staff}:{exit_reason}")
                        st.success(f"Exit recorded for {exit_staff}")
                        st.rerun()

        with ex2:
            st.markdown("#### Attrition analytics")
            analytics = hr_m.exit_analytics() if hr_m else {}

            if not analytics or analytics.get('total',0) == 0:
                st.info("No exit records yet. Record exits to see analytics.")
            else:
                ec1,ec2,ec3 = st.columns(3)
                ec1.metric("Total exits (all time)", analytics['total'])
                ec2.metric("Avg tenure at exit", f"{analytics['avg_tenure']:.1f} yrs")
                if analytics.get('avg_bsc_at_exit'):
                    ec3.metric("Avg BSC at exit", f"{analytics['avg_bsc_at_exit']:.2f}")

                if analytics.get('by_reason'):
                    by_r = pd.DataFrame(list(analytics['by_reason'].items()),
                                        columns=['Reason','Count']).sort_values('Count', ascending=False)
                    fig = px.bar(by_r, x='Count', y='Reason', orientation='h',
                                 title='Exits by reason', color='Count',
                                 color_continuous_scale=['#E8F5EE','var(--brand-primary,#006B3F)'])
                    fig.update_layout(height=300, showlegend=False,
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        margin=dict(l=0,r=0,t=40,b=0))
                    st.plotly_chart(fig, use_container_width=True)

                if analytics.get('by_unit'):
                    by_u = pd.DataFrame(list(analytics['by_unit'].items()),
                                        columns=['Unit','Exits']).sort_values('Exits', ascending=False)
                    st.markdown("**Exits by unit (top 10)**")
                    st.dataframe(by_u.head(10), hide_index=True, use_container_width=True)

        # Exit records table
        st.markdown("---")
        st.markdown("#### Exit records")
        exits_list = hr_m.exits if hr_m else []
        if exits_list:
            ex_df = pd.DataFrame([{
                'ID': e['id'], 'Name': e['staff_name'], 'Unit': e['unit'],
                'Exit date': e['exit_date'], 'Reason': e['reason'],
                'Tenure': f"{e.get('tenure_years',0):.1f}y",
                'BSC': f"{e['final_bsc']:.2f}" if e.get('final_bsc') else '—',
                'Rehire': '✅' if e.get('rehire_eligible') else '❌',
            } for e in exits_list])
            st.dataframe(ex_df, use_container_width=True, hide_index=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 4 — TRANSFERS
        # ════════════════════════════════════════════════════════════════
    with sub[2]:
        st.subheader("Staff transfers")
        st.caption("Track transfers and analyse their impact on performance.")

        tr1, tr2 = st.columns([2,3])

        with tr1:
            st.markdown("#### Record transfer")
            all_units = sorted(set(
                list(BRANCH_REGION.keys()) +
                ['ICT','HR','Finance','Operations','Products','DFS','SME Banking',
                 'Corporate Banking','Commercial Banking','Procurement']
            ))
            with st.form("transfer_form"):
                tr_staff   = st.selectbox("Staff member", [""] + sorted(staff_scores['Staff Name'].tolist() if len(staff_scores) else []))
                tc1, tc2   = st.columns(2)
                tr_from    = tc1.selectbox("From unit", [""] + all_units, key="tr_from")
                tr_to      = tc2.selectbox("To unit",   [""] + all_units, key="tr_to")
                tr_date    = st.date_input("Transfer date", value=date.today())
                tr_reason  = st.selectbox("Reason", TRANSFER_REASONS)
                tc3, tc4   = st.columns(2)
                bsc_before = tc3.number_input("BSC score before", min_value=0.0, max_value=5.0, step=0.01, value=0.0)
                bsc_after  = tc4.number_input("BSC score after (if known)", min_value=0.0, max_value=5.0, step=0.01, value=0.0)

                if st.form_submit_button("Record transfer", type="primary"):
                    if tr_staff and tr_from and tr_to:
                        sc_row = staff_scores[staff_scores['Staff Name']==tr_staff]
                        sc = str(sc_row['Staff Code'].values[0]) if len(sc_row) and 'Staff Code' in sc_row.columns else ''
                        hr_m.record_transfer({
                            'staff_code': sc, 'staff_name': tr_staff,
                            'from_unit': tr_from, 'to_unit': tr_to,
                            'from_region': BRANCH_REGION.get(tr_from,'Head Office'),
                            'to_region': BRANCH_REGION.get(tr_to,'Head Office'),
                            'transfer_date': tr_date, 'reason': tr_reason,
                            'bsc_before': bsc_before or None,
                            'bsc_after': bsc_after or None,
                            'initiated_by': uname,
                        })
                        audit_log("TRANSFER_RECORDED", uname, f"{tr_staff}:{tr_from}→{tr_to}")
                        st.success(f"Transfer recorded for {tr_staff}")
                        st.rerun()
                    else:
                        st.error("Staff, from unit and to unit are required.")

        with tr2:
            st.markdown("#### Transfer impact analysis")
            transfers = hr_m.get_transfers(24) if hr_m else []

            if not transfers:
                st.info("No transfer records yet.")
            else:
                # Impact: BSC before vs after
                impact_data = [t for t in transfers
                               if t.get('bsc_before') and t.get('bsc_after')]
                if impact_data:
                    impact_df = pd.DataFrame([{
                        'Staff': t['staff_name'],
                        'From': t['from_unit'],'To': t['to_unit'],
                        'Reason': t['reason'],
                        'BSC before': t['bsc_before'],
                        'BSC after': t['bsc_after'],
                        'Change': round(float(t['bsc_after'])-float(t['bsc_before']),2),
                    } for t in impact_data])
                    impact_df['Outcome'] = impact_df['Change'].apply(
                        lambda x: '✅ Improved' if x > 0.1 else ('❌ Declined' if x < -0.1 else '➡️ Neutral'))
                    st.dataframe(impact_df, use_container_width=True, hide_index=True)

                    avg_change = impact_df['Change'].mean()
                    pos = (impact_df['Change'] > 0.1).sum()
                    neg = (impact_df['Change'] < -0.1).sum()
                    tc1,tc2,tc3 = st.columns(3)
                    tc1.metric("Avg BSC change", f"{avg_change:+.2f}")
                    tc2.metric("Improved after transfer", pos)
                    tc3.metric("Declined after transfer", neg)
                else:
                    tr_df = pd.DataFrame([{
                        'Staff': t['staff_name'], 'From': t['from_unit'],
                        'To': t['to_unit'], 'Date': t['transfer_date'],
                        'Reason': t['reason'],
                    } for t in transfers])
                    st.dataframe(tr_df, use_container_width=True, hide_index=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 5 — DISCIPLINARY
        # ════════════════════════════════════════════════════════════════

# ── Section 2: 🏖️ Leave ─────────────────────────────
with sections[2]:
    sub = st.tabs([
        "🏖️ Leave management",
        "📅 Leave calendar",
    ])
    with sub[0]:
        st.subheader("Leave management")

        # ── Who can do what ───────────────────────────────────────────────
        _lv_role     = str(ud.get("role","")).lower()
        _lv_is_mgr   = (ud.get("is_admin") or
                        any(k in _lv_role for k in ("manager","director","head of","regional","hr","admin")))
        _lv_my_name  = ud.get("full_name","")
        _lv_my_sc    = str(ud.get("staff_code","") or uname)
        _lv_my_unit  = ud.get("unit","")
        _lv_my_role  = ud.get("role","")

        lv_tabs = st.tabs(["📝 Apply for leave",
                           "✅ Manager approvals" if _lv_is_mgr else "📋 My leave",
                           "📊 Team overview",
                           "📜 All records (HR)"])

        # ── TAB A: Staff applies for leave ────────────────────────────────
        with lv_tabs[0]:
            st.markdown("#### Apply for leave")
            st.caption("Submit a leave application. Your line manager will receive it for approval.")

            with st.form("leave_apply_form"):
                la1, la2 = st.columns(2)
                la_type   = la1.selectbox("Leave type", list(LEAVE_TYPES.keys()))
                la_relief = la2.text_input("Relief staff (optional)",
                    placeholder="Who covers your duties?")
                la_start  = la1.date_input("From", value=date.today())
                la_end    = la2.date_input("To", value=date.today())
                la_reason = st.text_area("Reason *", height=70,
                    placeholder="Please provide a brief reason for your leave request")

                lt_info = LEAVE_TYPES.get(la_type, {})
                _entitled = lt_info.get('days_entitled', 0) or 0
                # Calculate days already taken this year
                _yr_taken = sum(r.get("days",0) for r in lm.records
                               if r.get("staff_code")==_lv_my_sc
                               and r.get("leave_type")==la_type
                               and str(r.get("start_date",""))[:4] == str(date.today().year)
                               and r.get("approved") is not False)
                _balance = max(0, _entitled - _yr_taken)
                _req_days = (la_end - la_start).days + 1 if la_end >= la_start else 0

                st.markdown(
                    f"<div style='padding:8px 12px;background:var(--brand-light,#E8F5EE);"
                    f"border-left:3px solid var(--brand-primary,#006B3F);border-radius:0 6px 6px 0;"
                    f"font-size:11px;margin:4px 0'>"
                    f"<b>{la_type}</b> · {lt_info.get('description','')} · "
                    f"Paid: {'✅' if lt_info.get('paid') else '❌'} · "
                    f"Balance: <b>{_balance}/{_entitled} days</b> remaining · "
                    f"Requesting: <b>{_req_days} day(s)</b>"
                    f"{'<span style="color:#E24B4A"> ⚠️ Exceeds balance</span>' if _req_days > _balance and _entitled > 0 else ''}"
                    f"</div>", unsafe_allow_html=True)

                if st.form_submit_button("📤 Submit leave application", type="primary",
                                          use_container_width=True):
                    if not la_reason.strip():
                        st.error("Please provide a reason.")
                    elif la_end < la_start:
                        st.error("End date must be on or after start date.")
                    else:
                        lid = lm.apply_leave(
                            _lv_my_sc, _lv_my_name, _lv_my_role, _lv_my_unit,
                            la_type, la_start, la_end, la_reason, la_relief)
                        audit_log("LEAVE_APPLIED", uname, f"{la_type}:{la_start}:{la_end}")
                        st.success(f"✅ Leave application {lid} submitted. "
                                   "Your manager will review and approve it.")
                        st.rerun()

            # Show my own pending/approved leaves
            st.markdown("#### My recent leave requests")
            _my_leaves = lm.get_staff_leave(_lv_my_sc)
            if _my_leaves:
                for _lr in sorted(_my_leaves, key=lambda x: x.get("applied_at",""), reverse=True)[:8]:
                    _st  = _lr.get("status","")
                    _clr = {"Approved":"var(--brand-primary,#006B3F)","Rejected":"#E24B4A","Pending":"#F5A623"}.get(_st,"#9CA3AF")
                    _bg  = {"Approved":"#F0FDF4","Rejected":"#FEF2F2","Pending":"#FFFBEB"}.get(_st,"#F9FAFB")
                    st.markdown(
                        f"<div style='padding:8px 12px;background:{_bg};"
                        f"border-left:3px solid {_clr};border-radius:0 6px 6px 0;margin:4px 0;font-size:11px'>"
                        f"<b>{_lr.get('leave_type','')}</b> · "
                        f"{_lr.get('start_date','')} → {_lr.get('end_date','')} ({_lr.get('days',0)}d) · "
                        f"<span style='color:{_clr};font-weight:700'>{_st}</span>"
                        f"{'  ·  Reason: ' + _lr.get('rejection_reason','') if _st=='Rejected' else ''}"
                        f"</div>", unsafe_allow_html=True)
            else:
                st.info("No leave records yet.")

            st.markdown("#### Leave entitlements (Kenya Employment Act 2007)")
            ent_data = [[lt, info.get('days_entitled','Discretionary') or 'Discretionary',
                         '✅' if info.get('paid') else '❌',
                         info.get('description','')]
                        for lt, info in LEAVE_TYPES.items() if lt != 'Public Holiday']
            st.dataframe(pd.DataFrame(ent_data, columns=['Leave type','Days','Paid','Notes']),
                         hide_index=True, use_container_width=True)

        # ── TAB B: Manager approval queue ────────────────────────────────
        with lv_tabs[1]:
            if _lv_is_mgr:
                st.markdown("#### Pending leave approvals")
                _pending = lm.get_pending_approvals(manager_unit=_lv_my_unit if not can_all else None)
                if not _pending:
                    st.success("✅ No pending leave requests.")
                else:
                    st.info(f"**{len(_pending)} request(s) awaiting your approval.**")
                    for _pr in _pending:
                        _pr_id = str(_pr.get("id",""))
                        with st.expander(
                                f"{'🕐'} {_pr.get('staff_name','')} — {_pr.get('leave_type','')} · "
                                f"{_pr.get('start_date','')} to {_pr.get('end_date','')} ({_pr.get('days',0)}d)",
                                expanded=True):
                            pa1, pa2 = st.columns(2)
                            pa1.markdown(
                                "**Staff:** " + _pr.get("staff_name","") + "  \n"
                                "**Role:** " + _pr.get("staff_role","") + "  \n"
                                "**Unit:** " + _pr.get("staff_unit","") + "  \n"
                                "**Applied:** " + str(_pr.get("applied_at",""))[:10])
                            pa2.markdown(
                                "**Leave type:** " + _pr.get("leave_type","") + "  \n"
                                "**Period:** " + _pr.get("start_date","") + " to " + _pr.get("end_date","") + "  \n"
                                "**Days:** " + str(_pr.get("days",0)) + "  \n"
                                "**Relief:** " + (_pr.get("relief_staff","—") or "—"))
                            st.markdown(f"**Reason:** {_pr.get('reason','')}")

                            _rej_key = f"rej_reason_{_pr_id}"
                            _rej_txt = st.text_input("Rejection reason (if rejecting)",
                                                       key=_rej_key, placeholder="Optional")
                            _ab1, _ab2 = st.columns(2)
                            if _ab1.button("✅ Approve", key=f"appr_{_pr_id}",
                                           type="primary", use_container_width=True):
                                lm.approve_leave(_pr_id, uname, approve=True)
                                audit_log("LEAVE_APPROVED", uname,
                                          f"{_pr.get('staff_name','')}:{_pr.get('leave_type','')}")
                                st.toast("✅ Leave approved", icon="✅")
                                st.rerun()
                            if _ab2.button("❌ Reject", key=f"rejt_{_pr_id}",
                                           use_container_width=True):
                                lm.approve_leave(_pr_id, uname, approve=False,
                                                reason=st.session_state.get(_rej_key,""))
                                audit_log("LEAVE_REJECTED", uname,
                                          f"{_pr.get('staff_name','')}")
                                st.toast("❌ Leave rejected", icon="❌")
                                st.rerun()

                # HR record: recently approved (auto-populates)
                st.markdown("---")
                st.markdown("#### Recently approved (HR auto-record)")
                _recent_appr = [r for r in lm.records
                                if r.get("approved") is True
                                and r.get("hr_notified")]
                if _recent_appr:
                    _hr_df = pd.DataFrame([{
                        "ID": r.get("id",""), "Staff": r.get("staff_name",""),
                        "Type": r.get("leave_type",""),
                        "From": r.get("start_date",""), "To": r.get("end_date",""),
                        "Days": r.get("days",0), "Approved by": r.get("approved_by",""),
                    } for r in sorted(_recent_appr,
                                      key=lambda x: x.get("approved_at",""), reverse=True)[:20]])
                    st.dataframe(_hr_df, hide_index=True, use_container_width=True)
                else:
                    st.info("No approved leaves yet.")
            else:
                # Non-manager: show their own leave history
                st.markdown("#### My leave history")
                _my_all = lm.get_staff_leave(_lv_my_sc)
                if _my_all:
                    _mh_df = pd.DataFrame([{
                        "Type": r.get("leave_type",""),
                        "From": r.get("start_date",""), "To": r.get("end_date",""),
                        "Days": r.get("days",0),
                        "Status": r.get("status","Pending"),
                        "Approved by": r.get("approved_by","—"),
                    } for r in _my_all])
                    def _hl_st(v):
                        if v=="Approved": return "color:var(--brand-primary,#006B3F);font-weight:600"
                        if v=="Rejected": return "color:#E24B4A"
                        return "color:#F5A623"
                    st.dataframe(_mh_df.style.map(_hl_st, subset=["Status"]),
                                 hide_index=True, use_container_width=True)
                else:
                    st.info("No leave history.")

        with lv_tabs[2]:
            st.markdown("#### Leave register")
            all_leave = lm.records if lm else []

            # Filter options
            fc1, fc2 = st.columns(2)
            filter_type   = fc1.selectbox("Filter by type", ["All"] + list(LEAVE_TYPES.keys()), key="lv_type")
            filter_status = fc2.selectbox("Status", ["All","Active","Upcoming","Completed"], key="lv_status")

            disp_leave = all_leave.copy()
            if filter_type   != "All": disp_leave = [r for r in disp_leave if r.get('leave_type')==filter_type]
            if filter_status != "All": disp_leave = [r for r in disp_leave if r.get('status')==filter_status]

            if not disp_leave:
                st.info("No leave records matching filter.")
            else:
                leave_df = pd.DataFrame([{
                    'Staff':      r.get('staff_name',''),
                    'Type':       r.get('leave_type',''),
                    'Start':      r.get('start_date',''),
                    'End':        r.get('end_date',''),
                    'Days':       r.get('days',0),
                    'Status':     r.get('status',''),
                    'Approved by':r.get('approved_by',''),
                } for r in disp_leave])
                st.dataframe(leave_df, use_container_width=True, hide_index=True)

            # Currently on leave highlight
            active_now = lm.get_active_leave() if lm else []
            if active_now:
                st.markdown(f"**{len(active_now)} staff currently on leave:**")
                for r in active_now:
                    lt_color = LEAVE_TYPES.get(r.get('leave_type',''),{}).get('color','#888')
                    st.markdown(
                        f"<div style='padding:5px 10px;background:#F8F8F8;"
                        f"border-left:4px solid {lt_color};font-size:12px;margin:2px 0'>"
                        f"<b>{r['staff_name']}</b> — {r['leave_type']} "
                        f"({r['start_date']} to {r['end_date']}, {r['days']} days)"
                        f"</div>", unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # ════════════════════════════════════════════════════════════════
        # TAB 4 — LEAVE CALENDAR
        # ════════════════════════════════════════════════════════════════
    with sub[1]:
        st.subheader("Leave calendar")
        if not lm or not lm.records:
            st.info("No leave records yet.")
        else:
            import calendar as _cal
            _today_lc = date.today()
            _lc1, _lc2 = st.columns([1, 3])
            _sel_month = _lc1.selectbox("Month", list(range(1,13)),
                index=_today_lc.month-1, format_func=lambda m: _cal.month_name[m], key="lc_month")
            _sel_year  = _lc1.selectbox("Year", [2025, 2026, 2027],
                index=1, key="lc_year")

            # Get leave in this month
            _mo_start = date(_sel_year, _sel_month, 1)
            _mo_end   = date(_sel_year, _sel_month, _cal.monthrange(_sel_year, _sel_month)[1])

            _mo_leaves = []
            for _r in lm.records:
                try:
                    _s = _safe_date(str(_r.get("start_date",""))[:10])
                    _e = _safe_date(str(_r.get("end_date",""))[:10])
                    if _s <= _mo_end and _e >= _mo_start:
                        _mo_leaves.append(_r)
                except: pass

            # Build a simple calendar grid
            _days_in_month = _cal.monthrange(_sel_year, _sel_month)[1]
            # Show coloured leave bars per staff
            if _mo_leaves:
                _lc_staff = sorted(set(l.get("staff_name","") for l in _mo_leaves))
                _lc2.caption(f"{len(_mo_leaves)} leave record(s) · {len(_lc_staff)} staff on leave this month")

                # Simple table: rows = staff, columns = day numbers
                _lcols_hdr = ["Staff","Role"] + [str(d) for d in range(1, _days_in_month+1)]
                _lc_rows = []
                for _snm in _lc_staff:
                    _snm_leaves = [l for l in _mo_leaves if l.get("staff_name","")==_snm]
                    _row = {"Staff":_snm,
                            "Role":_snm_leaves[0].get("role","") if _snm_leaves else ""}
                    for _d in range(1, _days_in_month+1):
                        _dt = date(_sel_year, _sel_month, _d)
                        _on = any(_safe_date(str(l.get("start_date",""))[:10]) <= _dt <=
                                  _safe_date(str(l.get("end_date",""))[:10])
                                  for l in _snm_leaves)
                        _row[str(_d)] = "●" if _on else ""
                    _lc_rows.append(_row)

                if _lc_rows:
                    _lc_df = pd.DataFrame(_lc_rows)
                    def _lc_hl(v):
                        return "color:var(--brand-primary,#006B3F);font-weight:700" if v=="●" else ""
                    day_cols = [str(d) for d in range(1,_days_in_month+1)]
                    _lc2.dataframe(
                        _lc_df.style.map(_lc_hl, subset=day_cols),
                        use_container_width=True, hide_index=True, height=350)
            else:
                _lc2.info(f"No leave records for {_cal.month_name[_sel_month]} {_sel_year}.")

            # Summary stats
            _lc_st1, _lc_st2, _lc_st3 = st.columns(3)
            _lc_st1.metric("On leave this month", len(_mo_leaves))
            _lc_st2.metric("Total leave days", sum(
                (min(_mo_end, _safe_date(str(l.get("end_date",""))[:10])) -
                 max(_mo_start, _safe_date(str(l.get("start_date",""))[:10]))).days + 1
                for l in _mo_leaves
                if l.get("start_date") and l.get("end_date")))
            _approved = sum(1 for l in _mo_leaves if l.get("approved"))
            _lc_st3.metric("Approved", _approved)

        # TAB 3 — EXITS & ATTRITION (now TAB 5)
        # TAB 3 — EXITS & ATTRITION
        # ════════════════════════════════════════════════════════════════

# ── Section 3: 📋 Discipline & Dev ─────────────────────────────
with sections[3]:
    sub = st.tabs([
        "⚖️ Disciplinary",
        "📋 PIP management",
        "🎯 Diligence scores",
        "🤝 Coaching Intelligence (Standard #11, v5.93)",
    ])
    with sub[0]:
        st.subheader("Disciplinary case management")
        st.caption("All cases are confidential. Access controlled by HR and administration roles.")

        if not can_all:
            st.warning("Disciplinary records are restricted to HR and administration.")
            st.stop()

        dc1, dc2 = st.columns([2,3])

        with dc1:
            st.markdown("#### Open new case")
            with st.form("disc_form"):
                dc_staff  = st.selectbox("Staff member", [""] + sorted(staff_scores['Staff Name'].tolist() if len(staff_scores) else []))
                dc_cat    = st.selectbox("Category", DISCIPLINARY_CATEGORIES)
                dc_date   = st.date_input("Incident date", value=date.today())
                dc_desc   = st.text_area("Description of incident", height=100)

                if st.form_submit_button("Open case", type="primary"):
                    if dc_staff and dc_desc:
                        sc_row = staff_scores[staff_scores['Staff Name']==dc_staff]
                        sc   = str(sc_row['Staff Code'].values[0]) if len(sc_row) and 'Staff Code' in sc_row.columns else ''
                        unit = str(sc_row['Unit'].values[0]) if len(sc_row) and 'Unit' in sc_row.columns else ''
                        hr_m.open_case({
                            'staff_code': sc, 'staff_name': dc_staff,
                            'unit': unit, 'category': dc_cat,
                            'incident_date': dc_date,
                            'description': dc_desc, 'opened_by': uname,
                        })
                        audit_log("DISC_CASE_OPENED", uname, f"{dc_staff}:{dc_cat}")
                        st.success(f"Case opened for {dc_staff}")
                        st.rerun()
                    else:
                        st.error("Staff and description are required.")

        with dc2:
            st.markdown("#### Active cases")
            active_cases = hr_m.get_active_cases() if hr_m else []

            if not active_cases:
                st.success("No open disciplinary cases.")
            else:
                for case in active_cases:
                    opened  = datetime.fromisoformat(case['recorded_at'])
                    days_open = (datetime.now() - opened).days
                    urgency = '🔴' if days_open > 30 else ('🟡' if days_open > 14 else '🟢')

                    with st.expander(
                        f"{urgency} {case['id']} — {case['staff_name']} | {case['category']} | "
                        f"Stage: {case['stage']} | {days_open}d open", expanded=(days_open>30)):

                        st.caption(f"Unit: {case['unit']} | Incident: {case['incident_date']}")
                        st.markdown(f"**Incident:** {case['description']}")

                        # Stage progression
                        st.markdown("**Advance stage:**")
                        with st.form(f"disc_advance_{case['id']}"):
                            cur_idx = DISCIPLINARY_STAGES.index(case['stage']) if case['stage'] in DISCIPLINARY_STAGES else 0
                            next_stages = DISCIPLINARY_STAGES[cur_idx+1:] if cur_idx+1 < len(DISCIPLINARY_STAGES) else []
                            if next_stages:
                                new_stage = st.selectbox("Next stage", next_stages, key=f"ns_{case['id']}")
                                stage_note= st.text_input("Note", key=f"sn_{case['id']}")
                                if st.form_submit_button("Advance", type="primary"):
                                    hr_m.advance_stage(case['id'], new_stage, stage_note, uname)
                                    audit_log("DISC_STAGE", uname, f"{case['id']}:{new_stage}")
                                    st.success("Stage updated")
                                    st.rerun()
                            else:
                                st.info("Case is at final stage.")

                        # Stage history
                        if case.get('stage_history'):
                            st.markdown("**Stage history:**")
                            for s in case['stage_history']:
                                st.caption(f"  {s['date']} — {s['stage']} (by {s['by']})")

        # All cases table
        st.markdown("---")
        all_cases = hr_m.disciplinary if hr_m else []
        if all_cases:
            cases_df = pd.DataFrame([{
                'ID': c['id'], 'Staff': c['staff_name'], 'Unit': c['unit'],
                'Category': c['category'], 'Stage': c['stage'],
                'Status': c['status'], 'Incident date': c['incident_date'],
                'Outcome': c.get('outcome','—'),
            } for c in all_cases])
            st.dataframe(cases_df, use_container_width=True, hide_index=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 6 — PIP MANAGEMENT
        # ════════════════════════════════════════════════════════════════
    with sub[1]:
        st.subheader("Performance Improvement Plan (PIP) management")
        st.caption("Structured, time-bound improvement plans with milestone tracking and HR discussion support.")

        pip_tabs = st.tabs(["📋 Active PIPs", "➕ Open new PIP", "📊 PIP analytics"])

        with pip_tabs[0]:
            active_pips = hr_m.get_active_pips() if hr_m else []
            if not active_pips:
                st.success("No active PIPs. All staff are performing within expectations.")
            else:
                for pip in active_pips:
                    days_left = hr_m.pip_days_remaining(pip)
                    urgency   = '🔴' if days_left <= 7 else ('🟡' if days_left <= 14 else '🟢')
                    prog_pct  = max(0, round((pip['duration_days']-days_left)/pip['duration_days']*100,0))

                    with st.expander(
                        f"{urgency} {pip['id']} — {pip['staff_name']} | {pip['unit']} | "
                        f"{days_left}d remaining | {prog_pct:.0f}% elapsed",
                        expanded=(days_left <= 14)):

                        pc1, pc2, pc3 = st.columns(3)
                        pc1.metric("Days remaining", days_left)
                        pc2.metric("Current BSC",
                                   f"{pip['current_bsc']:.2f}" if pip.get('current_bsc') else '—')
                        pc3.metric("Target BSC",
                                   f"{pip['target_bsc']:.2f}" if pip.get('target_bsc') else '3.00')

                        st.progress(prog_pct/100, text=f"PIP progress: {prog_pct:.0f}%")

                        # Performance gaps
                        if pip.get('performance_gaps'):
                            st.markdown("**Performance gaps addressed:**")
                            for g in pip['performance_gaps']:
                                st.markdown(f"  • {g}")

                        # Objectives
                        if pip.get('objectives'):
                            st.markdown("**SMART objectives:**")
                            for o in pip['objectives']:
                                st.markdown(f"  ✓ {o}")

                        # Support offered
                        if pip.get('support_offered'):
                            st.markdown(f"**Support offered:** {pip['support_offered']}")

                        # Review history
                        if pip.get('reviews'):
                            st.markdown(f"**Review history ({len(pip['reviews'])} reviews):**")
                            for rv in pip['reviews'][-3:]:
                                status_ico = '✅' if rv.get('score',0) >= float(pip.get('target_bsc',3)) else '⚠️'
                                st.markdown(
                                    f"<div style='padding:6px 10px;background:#F8F8F8;"
                                    f"border-left:3px solid var(--brand-primary,#006B3F);font-size:12px;margin:2px 0'>"
                                    f"{status_ico} <b>{rv['date']}</b> — Score: {rv.get('score','—')} | "
                                    f"{rv.get('progress','')} (by {rv.get('by','')})</div>",
                                    unsafe_allow_html=True)

                        # Discussion helper
                        with st.expander("💬 HR discussion guide for this PIP review"):
                            st.markdown("""
        **Opening the review:**
        - "This is a safe space to discuss your progress and challenges."
        - "I want to start by acknowledging what you've achieved since our last meeting."

        **Progress assessment:**
        - "Walk me through your performance against each objective."
        - "What has worked well? What has been most challenging?"
        - "Are there any barriers I haven't addressed that are preventing your progress?"

        **If progressing well:**
        - "Your improvement shows real commitment. Let's talk about sustaining this."
        - "What support do you need to maintain this trajectory?"

        **If not progressing:**
        - "I'm concerned about [specific gap]. What's your perspective on why this is happening?"
        - "What would need to change for you to meet this objective?"
        - "I want to be honest — if we don't see improvement in [specific area] by [date], we'll need to discuss next steps."

        **Closing:**
        - "Let's agree on the specific actions between now and our next review."
        - "My door is always open — please reach out if you hit a wall before our next meeting."
                            """)

                        # Add review
                        st.markdown("---")
                        st.markdown("**Add review entry:**")
                        with st.form(f"pip_review_{pip['id']}"):
                            rv1, rv2 = st.columns(2)
                            rv_score    = rv1.number_input("BSC score at review", 0.0, 5.0, step=0.01, key=f"rvs_{pip['id']}")
                            rv_outcome  = rv2.selectbox("Outcome", PIP_OUTCOMES, key=f"rvo_{pip['id']}")
                            rv_progress = st.text_area("Progress notes", height=60, key=f"rvp_{pip['id']}")
                            rv_concerns = st.text_area("Concerns / support needed", height=60, key=f"rvc_{pip['id']}")

                            if st.form_submit_button("Save review", type="primary"):
                                hr_m.add_pip_review(pip['id'], {
                                    'score': rv_score, 'outcome': rv_outcome,
                                    'progress': rv_progress, 'concerns': rv_concerns,
                                    'by': uname,
                                })
                                audit_log("PIP_REVIEW", uname, f"{pip['id']}:{rv_outcome}")
                                st.success("Review recorded")
                                st.rerun()

        with pip_tabs[1]:
            st.markdown("#### Open a new Performance Improvement Plan")
            all_staff_pip = sorted(staff_scores['Staff Name'].tolist()) if len(staff_scores) else []

            # Pre-fill with at-risk staff
            at_risk_staff = staff_scores[staff_scores['Final_BSC_Score'] < 2.8]['Staff Name'].tolist() if len(staff_scores) else []
            if at_risk_staff:
                st.markdown(
                    f"<div style='padding:8px 12px;background:#FFFBF0;"
                    f"border-left:3px solid #F5A623;font-size:12px;margin:4px 0'>"
                    f"⚠️ <b>{len(at_risk_staff)} staff</b> scoring below 2.8 — consider PIP: "
                    f"{', '.join(at_risk_staff[:5])}"
                    f"{'...' if len(at_risk_staff)>5 else ''}</div>",
                    unsafe_allow_html=True)

            with st.form("pip_open_form"):
                pip_staff   = st.selectbox("Staff member *", [""] + all_staff_pip)
                pp1, pp2, pp3 = st.columns(3)
                pip_start   = pp1.date_input("PIP start date", value=date.today())
                pip_dur     = pp2.selectbox("Duration (days)", PIP_DURATIONS)
                pip_freq    = pp3.selectbox("Review frequency", PIP_REVIEW_FREQUENCIES)

                pip_reason  = st.text_area("Reason for PIP *",
                    placeholder="Describe the performance gaps that necessitate this PIP", height=80)

                st.markdown("**Performance gaps (one per line)**")
                pip_gaps_text = st.text_area("Gaps", height=80, placeholder="e.g. Deposit growth at 45% of target for 3 consecutive months")

                st.markdown("**SMART objectives (one per line)**")
                pip_obj_text = st.text_area("Objectives", height=80,
                    placeholder="e.g. Achieve minimum 70% deposit target by end of PIP period")

                pip_support = st.text_area("Support and resources offered",
                    placeholder="e.g. Weekly coaching with Branch Manager, sales training, paired with top performer", height=60)

                pip_manager = st.text_input("PIP manager (reviewing manager)", value=uname)
                pip_hr      = st.text_input("HR officer", value="")

                sc_row = staff_scores[staff_scores['Staff Name']==pip_staff] if pip_staff else pd.DataFrame()
                current_bsc = float(sc_row['Final_BSC_Score'].values[0]) if len(sc_row) else None

                if st.form_submit_button("Open PIP", type="primary"):
                    if pip_staff and pip_reason:
                        sc = str(sc_row['Staff Code'].values[0]) if len(sc_row) and 'Staff Code' in sc_row.columns else ''
                        unit = str(sc_row['Unit'].values[0]) if len(sc_row) and 'Unit' in sc_row.columns else ''
                        role = str(sc_row['Role'].values[0]) if len(sc_row) and 'Role' in sc_row.columns else ''
                        hr_m.open_pip({
                            'staff_code': sc, 'staff_name': pip_staff, 'unit': unit, 'role': role,
                            'manager': pip_manager, 'hr_officer': pip_hr,
                            'start_date': pip_start, 'duration_days': pip_dur,
                            'reason': pip_reason,
                            'performance_gaps': [g.strip() for g in pip_gaps_text.split('\n') if g.strip()],
                            'objectives':       [o.strip() for o in pip_obj_text.split('\n') if o.strip()],
                            'support_offered':  pip_support,
                            'review_frequency': pip_freq,
                            'current_bsc': current_bsc,
                            'target_bsc': 3.0,
                            'opened_by': uname,
                        })
                        audit_log("PIP_OPENED", uname, f"{pip_staff}")
                        st.success(f"PIP opened for {pip_staff}. Duration: {pip_dur} days.")
                        st.rerun()
                    else:
                        st.error("Staff member and reason are required.")

        with pip_tabs[2]:
            all_pips = hr_m.pips if hr_m else []
            if not all_pips:
                st.info("No PIP records yet.")
            else:
                p1,p2,p3,p4 = st.columns(4)
                completed = [p for p in all_pips if p['outcome']=='Successfully completed']
                dismissed = [p for p in all_pips if 'dismissal' in p['outcome'].lower()]
                extended  = [p for p in all_pips if p['outcome']=='Extended']
                p1.metric("Total PIPs", len(all_pips))
                p2.metric("Completed successfully", len(completed))
                p3.metric("Terminated/dismissed", len(dismissed))
                p4.metric("Extended", len(extended))

                outcomes = {}
                for p in all_pips:
                    o = p['outcome']; outcomes[o] = outcomes.get(o,0)+1
                out_df = pd.DataFrame(list(outcomes.items()), columns=['Outcome','Count'])
                fig = px.pie(out_df, names='Outcome', values='Count', title='PIP outcomes')
                fig.update_layout(height=280, margin=dict(l=0,r=0,t=40,b=0))
                st.plotly_chart(fig, use_container_width=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 7 — DILIGENCE SCORES
        # ════════════════════════════════════════════════════════════════
    with sub[2]:
        st.subheader("Staff diligence scores")
        st.caption("Composite score (0–100) measuring execution, accountability, conduct, and attendance.")

        st.markdown(
            "<div style='padding:10px 14px;background:var(--brand-light,#E8F5EE);"
            "border-left:3px solid var(--brand-primary,#006B3F);font-size:12px;margin-bottom:12px'>"
            "<b>Diligence score components:</b> "
            "Milestone on-time rate (30%) · Action plan acceptance (20%) · "
            "Leave compliance (15%) · Disciplinary clean (20%) · Not on PIP (15%)"
            "</div>", unsafe_allow_html=True)

        # Compute for all visible staff
        diligence_rows = []
        for _, staff_row in filtered.iterrows():
            sc = str(staff_row.get('Staff Code',''))
            name = staff_row.get('Staff Name','')
            score = hr_m.compute_diligence(sc, em, action_plans) if hr_m else 100.0
            bsc   = staff_row.get('Final_BSC_Score', 0)
            on_pip = hr_m.staff_on_pip(sc) if hr_m else False
            has_disc = hr_m.staff_has_active_case(sc) if hr_m else False
            on_leave = lm.is_on_leave(sc) if lm else False

            diligence_rows.append({
                'Staff Name':  name,
                'Unit':        staff_row.get('Unit',''),
                'BSC Score':   round(bsc, 2),
                'Diligence':   score,
                'On PIP':      '⚠️ Yes' if on_pip else '✅ No',
                'Disc. case':  '🔴 Active' if has_disc else '✅ Clear',
                'On leave':    '🏖️ Yes' if on_leave else '—',
            })

        if diligence_rows:
            dil_df = pd.DataFrame(diligence_rows).sort_values('Diligence')

            # Chart
            dil_df['Color'] = dil_df['Diligence'].apply(
                lambda x: '#E24B4A' if x < 50 else ('#F5A623' if x < 75 else 'var(--brand-primary,#006B3F)'))
            fig = go.Figure()
            fig.add_bar(x=dil_df['Staff Name'], y=dil_df['Diligence'],
                        marker_color=dil_df['Color'],
                        text=[f"{v:.0f}" for v in dil_df['Diligence']],
                        textposition='outside')
            fig.add_hline(y=75, line_dash='dash', line_color='#F5A623',
                           annotation_text='75 — target')
            fig.update_layout(height=350, xaxis_tickangle=-35, yaxis_range=[0,110],
                yaxis_title='Diligence score', title='Staff diligence scores',
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig, use_container_width=True)

            # Table
            disp_cols = ['Staff Name','Unit','BSC Score','Diligence','On PIP','Disc. case','On leave']
            st.dataframe(dil_df[disp_cols].sort_values('Diligence'),
                         use_container_width=True, hide_index=True)

        # ════════════════════════════════════════════════════════════════
        # TAB 8 — TEAM INSIGHTS (existing, preserved)
        # ════════════════════════════════════════════════════════════════

with _payroll_expander:
    st.markdown("**Month-end payroll export — download BSC score + commission tier for payroll processing:**")
    try:
        import io as _io_pr, pandas as _pd_pr
        _scores_pr = json.loads((DATA/"feb_2026_staff_scores.json").read_text()) if (DATA/"feb_2026_staff_scores.json").exists() else {}
        _comm_pr   = json.loads((DATA/"commission_records.json").read_text()) if (DATA/"commission_records.json").exists() else []
        _comm_map  = {str(c.get("staff_code","")): c for c in _comm_pr}
        _payroll_rows = []
        for sc_v, s in _scores_pr.items():
            c = _comm_map.get(str(sc_v), {})
            _payroll_rows.append({
                "Staff Code":    sc_v,
                "Full Name":     s.get("name","")[:30],
                "Role":          s.get("role","")[:30],
                "Unit":          s.get("unit","")[:25],
                "BSC Score":     s.get("final_score",0),
                "Tier":          c.get("tier","—"),
                "Commission (KES)": c.get("total_commission",0),
                "Payment Status":   c.get("status","Pending"),
            })
        _df_pr = _pd_pr.DataFrame(_payroll_rows)
        st.dataframe(_df_pr.head(20), use_container_width=True, hide_index=True)
        st.caption(f"{len(_payroll_rows):,} staff in payroll export")
        _buf_pr = _io_pr.BytesIO()
        _df_pr.to_excel(_buf_pr, index=False, sheet_name="Payroll", engine="openpyxl")
        _buf_pr.seek(0)
        st.download_button("📥 Download Payroll File",
                            data=_buf_pr.getvalue(),
                            file_name=f"Payroll_BSC_{date.today().strftime('%Y%m')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="payroll_dl")
    except Exception as _e:
        st.error(f"Payroll export error: {str(_e)[:80]}")


    # ════════════════════════════════════════════════════════════════
    # SECTION 3 SUB[3] — COACHING INTELLIGENCE (Standard #11, v5.93)
    # ════════════════════════════════════════════════════════════════
    with sub[3]:
        from utils.coaching_intelligence import (
            CoachingIntelligence, CoachingScript,
            DEFAULT_AGENDA_MIN, DEFAULT_AGENDA_MAX,
            DEFAULT_TALKING_POINTS_MIN, DEFAULT_TALKING_POINTS_MAX,
            DEFAULT_ACTIONS_MIN, DEFAULT_ACTIONS_MAX,
        )
        from datetime import date as _date_ci
        from utils.core_audit import audit_log as _audit_log_ci

        st.markdown(
            f"**Standard #11 — Coaching Intelligence Engine**. "
            f"Generates structured 1:1 coaching scripts for managers "
            f"based on staff signals (KPI status, nudges, growth plan, "
            f"micro-tasks, learning cards).")
        st.caption(
            f"Script structure: {DEFAULT_AGENDA_MIN}-{DEFAULT_AGENDA_MAX} agenda items, "
            f"{DEFAULT_TALKING_POINTS_MIN}-{DEFAULT_TALKING_POINTS_MAX} talking points, "
            f"{DEFAULT_ACTIONS_MIN}-{DEFAULT_ACTIONS_MAX} recommended actions. "
            f"Engine enforces authorization (manager must be staff's direct report) "
            f"and Rule 6 transparency (returns {{}} for unknown staff).")

        ci_sub_tabs = st.tabs([
            "🤝 Generate Coaching Script",
            "🎯 Demo Scenario Builder",
            "🌳 Engine Reference",
        ])

        # ──────── Generate Coaching Script ────────
        with ci_sub_tabs[0]:
            st.markdown(
                "**Generate a coaching script for a 1:1 meeting** — provide the "
                "manager and staff codes; engine assembles agenda + talking points "
                "+ recommended actions from the staff's current signals.")
            st.caption(
                "💡 Engine queries 7 signal sources (staff lookup, direct-report check, "
                "KPI status, active nudges, growth plan, microtasks, learning cards). "
                "Production deployment would auto-pull from CBS/HR systems via DI callbacks; "
                "this UI uses demo data with realistic signal mix.")

            cc1, cc2 = st.columns(2)
            with cc1:
                ci_mgr = st.text_input("Manager code",
                                         value="robert002",
                                         key="ci_mgr",
                                         help="Manager initiating the coaching session.")
                ci_staff = st.text_input("Staff code",
                                            value="STAFF_DEMO_001",
                                            key="ci_staff",
                                            help="Staff member being coached.")
            with cc2:
                ci_for_date = st.date_input("Meeting date",
                                              value=_date_ci(2026, 5, 1),
                                              key="ci_for_date")

            st.markdown("**Demo signal data:**")
            with st.expander("Signal source values (used by engine)"):
                ci_kpi_behind = st.number_input(
                    "KPIs 'behind' (count)", min_value=0, max_value=5,
                    value=2, key="ci_kpi_behind",
                    help="Number of KPIs at < 80% of target.")
                ci_kpi_exceeded = st.number_input(
                    "KPIs 'exceeded' (count)", min_value=0, max_value=5,
                    value=1, key="ci_kpi_exceeded",
                    help="Number of KPIs at > 100% of target.")
                ci_nudges = st.number_input(
                    "Active nudges (count)", min_value=0, max_value=10,
                    value=2, key="ci_nudges")
                ci_skill_gaps = st.number_input(
                    "Skill gaps (count)", min_value=0, max_value=5,
                    value=2, key="ci_skill_gaps")
                ci_microtasks = st.number_input(
                    "Outstanding microtasks (count)", min_value=0, max_value=10,
                    value=2, key="ci_microtasks")

            if st.button("🤝 Generate coaching script",
                           key="ci_gen_btn", type="primary"):
                # Build DI callbacks from inputs
                def _is_direct_report(mgr, staff):
                    return True  # demo allows all

                def _staff_lookup(staff_code):
                    return {
                        "staff_code": staff_code,
                        "full_name": f"Staff {staff_code}",
                        "role": "Branch Manager",
                        "unit": "Branch 015",
                    }

                def _kpi_status(staff_code):
                    rows = []
                    behind_kpis = ["deposits_growth_pct", "nps_score",
                                    "loan_book_growth_pct", "cir_pct", "active_customers"]
                    exceeded_kpis = ["fee_income_growth", "digital_adoption",
                                      "complaint_resolution_pct", "training_hours"]
                    for i in range(int(ci_kpi_behind)):
                        rows.append({"kpi_id": behind_kpis[i],
                                      "current": 60 + i * 2,
                                      "target": 100,
                                      "achievement_pct": 60 + i * 2,
                                      "status": "behind"})
                    for i in range(int(ci_kpi_exceeded)):
                        rows.append({"kpi_id": exceeded_kpis[i],
                                      "current": 110 + i * 5,
                                      "target": 100,
                                      "achievement_pct": 110 + i * 5,
                                      "status": "exceeded"})
                    return rows

                def _nudges(staff_code):
                    nudge_pool = [
                        {"id": f"N{i:03d}",
                          "text": ["Review pipeline staleness — 3 deals over 60d",
                                   "Schedule complaint review — 2 open >14d",
                                   "Confirm Q1 budget reconciliation",
                                   "Update risk register",
                                   "Customer sat survey response overdue"][i % 5],
                          "severity": "high" if i % 2 == 0 else "medium",
                          "kpi": ["loan_book_growth_pct", "nps_score",
                                   "cir_pct", "ops_risk", "csat_score"][i % 5]}
                        for i in range(int(ci_nudges))
                    ]
                    return nudge_pool

                def _growth_plan(staff_code):
                    skill_pool = [
                        {"skill": "credit_analysis", "current": 3, "required": 4},
                        {"skill": "team_leadership", "current": 2, "required": 4},
                        {"skill": "stakeholder_management", "current": 3, "required": 5},
                        {"skill": "data_analytics", "current": 2, "required": 3},
                    ]
                    return {
                        "skill_gaps": skill_pool[:int(ci_skill_gaps)],
                        "promotion_readiness_pct": 65,
                        "next_role": "Regional Head",
                    }

                def _microtasks(staff_code):
                    return [
                        {"id": f"MT{i:03d}",
                          "title": ["Complete IFRS 9 refresher quiz",
                                    "Submit Q1 customer review",
                                    "Update branch risk assessment",
                                    "Sign off Q1 timesheet"][i % 4],
                          "due_date": "2026-05-15",
                          "category": "compliance" if i % 2 == 0 else "performance"}
                        for i in range(int(ci_microtasks))
                    ]

                def _learning_cards(staff_code):
                    return [
                        {"id": "LC001",
                          "title": "Difficult conversations workshop",
                          "topic": "team_leadership",
                          "duration_min": 30}
                    ]

                engine = CoachingIntelligence(
                    is_direct_report_fn=_is_direct_report,
                    staff_lookup_fn=_staff_lookup,
                    kpi_status_fn=_kpi_status,
                    nudges_fn=_nudges,
                    growth_plan_fn=_growth_plan,
                    microtasks_fn=_microtasks,
                    learning_cards_fn=_learning_cards,
                )
                r = engine.generate_coaching_script(
                    ci_mgr, ci_staff, today=ci_for_date)

                if not r:
                    st.error(
                        "⛔ Engine returned empty — check authorization "
                        "(manager must be direct supervisor) and staff existence.")
                else:
                    # Header banner
                    meta = r.get("meta", {})
                    staff_name = meta.get("staff_name", "—")
                    staff_role = meta.get("staff_role", "—")
                    staff_unit = meta.get("staff_unit", "—")

                    st.markdown(
                        f"<div style='padding:16px;background:#3B82F622;"
                        f"border-left:6px solid #3B82F6;border-radius:10px'>"
                        f"<div style='font-size:11px;letter-spacing:1.5px;opacity:0.7'>"
                        f"COACHING SCRIPT</div>"
                        f"<div style='font-size:20px;font-weight:700;color:#1E40AF;margin-top:4px'>"
                        f"🤝 {staff_name}</div>"
                        f"<div style='font-size:13px;margin-top:4px'>"
                        f"{staff_role} · {staff_unit} · "
                        f"Meeting date: <b>{meta.get('for_date', '—')}</b></div></div>",
                        unsafe_allow_html=True)

                    # Meeting Agenda
                    agenda = r.get("meeting_agenda", [])
                    st.markdown("### 📋 Meeting Agenda")
                    if agenda:
                        for i, item in enumerate(agenda, 1):
                            st.markdown(f"**{i}.** {item}")
                    else:
                        st.caption("(no agenda items)")

                    # Talking Points
                    talking = r.get("talking_points", [])
                    st.markdown("### 💬 Talking Points")
                    if talking:
                        for tp in talking:
                            st.markdown(f"- {tp}")
                    else:
                        st.caption("(no talking points)")

                    # Recommended Actions
                    actions = r.get("recommended_actions", [])
                    st.markdown("### ✅ Recommended Actions")
                    if actions:
                        for i, a in enumerate(actions, 1):
                            st.markdown(f"**{i}.** {a}")
                    else:
                        st.caption("(no recommended actions)")

                    # Signals used (transparency)
                    signals = meta.get("signals_used", {})
                    if signals:
                        st.markdown("---")
                        st.caption("**Signal sources used by engine:**")
                        sig_cols = st.columns(len(signals))
                        for i, (k, v) in enumerate(signals.items()):
                            sig_cols[i].metric(
                                k.replace("_", " ").title(),
                                int(v) if str(v).isdigit() else str(v))

                    with st.expander("Engine metadata"):
                        st.json({
                            "manager_code": meta.get("manager_code"),
                            "staff_code": meta.get("staff_code"),
                            "for_date": meta.get("for_date"),
                            "generated_at": meta.get("generated_at"),
                            "agenda_count": len(agenda),
                            "talking_points_count": len(talking),
                            "actions_count": len(actions),
                        })

                    _audit_log_ci("IFRS_ENGINE_USED", uname,
                                   f"Coaching #11: script {ci_mgr}→{ci_staff} "
                                   f"agenda={len(agenda)} talking={len(talking)} "
                                   f"actions={len(actions)}")

        # ──────── Demo Scenario Builder ────────
        with ci_sub_tabs[1]:
            st.markdown(
                "**Pre-configured scenarios** — see how coaching scripts vary "
                "based on different signal mixes.")

            scenario = st.selectbox(
                "Scenario",
                [
                    "High performer (mostly exceeded KPIs)",
                    "Underperformer (multiple behind KPIs + nudges)",
                    "Development focus (skill gaps + learning cards)",
                    "Compliance pressure (microtasks + nudges)",
                    "Authorization refused (not direct report)",
                    "Unknown staff (Rule 6)",
                ],
                key="ci_demo_scenario")

            scenarios = {
                "High performer (mostly exceeded KPIs)": {
                    "is_direct_report": True,
                    "kpi_status": [
                        {"kpi_id": "deposits_growth_pct", "current": 15, "target": 12,
                          "achievement_pct": 125, "status": "exceeded"},
                        {"kpi_id": "loan_book_growth_pct", "current": 14, "target": 12,
                          "achievement_pct": 117, "status": "exceeded"},
                        {"kpi_id": "fee_income_growth", "current": 18, "target": 15,
                          "achievement_pct": 120, "status": "exceeded"},
                        {"kpi_id": "cir_pct", "current": 53, "target": 55,
                          "achievement_pct": 96, "status": "on_track"},
                    ],
                    "nudges": [],
                    "skill_gaps": [
                        {"skill": "strategic_thinking", "current": 3, "required": 4}],
                    "microtasks": [],
                    "learning_cards": [
                        {"id": "LC001", "title": "Strategic finance certification",
                          "topic": "strategic_thinking", "duration_min": 60}],
                },
                "Underperformer (multiple behind KPIs + nudges)": {
                    "is_direct_report": True,
                    "kpi_status": [
                        {"kpi_id": "deposits_growth_pct", "current": 6, "target": 12,
                          "achievement_pct": 50, "status": "behind"},
                        {"kpi_id": "loan_book_growth_pct", "current": 5, "target": 12,
                          "achievement_pct": 42, "status": "behind"},
                        {"kpi_id": "nps_score", "current": 35, "target": 60,
                          "achievement_pct": 58, "status": "behind"},
                        {"kpi_id": "cir_pct", "current": 65, "target": 55,
                          "achievement_pct": 85, "status": "behind"},
                    ],
                    "nudges": [
                        {"id": "N001", "text": "Critical: 3 deals stale >90d",
                          "severity": "high", "kpi": "loan_book_growth_pct"},
                        {"id": "N002", "text": "5 unresolved complaints",
                          "severity": "high", "kpi": "nps_score"},
                        {"id": "N003", "text": "Cost overrun on overtime budget",
                          "severity": "medium", "kpi": "cir_pct"},
                    ],
                    "skill_gaps": [],
                    "microtasks": [],
                    "learning_cards": [],
                },
                "Development focus (skill gaps + learning cards)": {
                    "is_direct_report": True,
                    "kpi_status": [
                        {"kpi_id": "deposits_growth_pct", "current": 11, "target": 12,
                          "achievement_pct": 92, "status": "on_track"},
                    ],
                    "nudges": [],
                    "skill_gaps": [
                        {"skill": "credit_analysis", "current": 2, "required": 4},
                        {"skill": "team_leadership", "current": 2, "required": 4},
                    ],
                    "microtasks": [],
                    "learning_cards": [
                        {"id": "LC001", "title": "IFRS 9 deep-dive",
                          "topic": "credit_analysis", "duration_min": 90}],
                },
                "Compliance pressure (microtasks + nudges)": {
                    "is_direct_report": True,
                    "kpi_status": [
                        {"kpi_id": "compliance_score", "current": 75, "target": 95,
                          "achievement_pct": 79, "status": "behind"},
                    ],
                    "nudges": [
                        {"id": "N001", "text": "5 KYC reviews overdue",
                          "severity": "high", "kpi": "compliance_score"},
                        {"id": "N002", "text": "Annual mandatory training pending",
                          "severity": "high", "kpi": "training_completion"},
                    ],
                    "skill_gaps": [],
                    "microtasks": [
                        {"id": "MT001", "title": "Complete AML refresher",
                          "due_date": "2026-05-10", "category": "compliance"},
                        {"id": "MT002", "title": "Sign code of conduct",
                          "due_date": "2026-05-15", "category": "compliance"},
                        {"id": "MT003", "title": "Submit quarterly compliance attestation",
                          "due_date": "2026-05-20", "category": "compliance"},
                    ],
                    "learning_cards": [],
                },
                "Authorization refused (not direct report)": {
                    "is_direct_report": False,
                    "kpi_status": [],
                    "nudges": [],
                    "skill_gaps": [],
                    "microtasks": [],
                    "learning_cards": [],
                },
                "Unknown staff (Rule 6)": {
                    "is_direct_report": True,
                    "staff_lookup_returns_none": True,
                    "kpi_status": [],
                    "nudges": [],
                    "skill_gaps": [],
                    "microtasks": [],
                    "learning_cards": [],
                },
            }
            cfg = scenarios[scenario]

            with st.expander("Scenario configuration"):
                st.json({k: (str(v) if not isinstance(v, list) else f"{len(v)} items")
                          for k, v in cfg.items()})

            if st.button("🎯 Run scenario",
                           key="ci_demo_btn", type="primary"):
                def _is_dr(mgr, staff): return cfg["is_direct_report"]

                def _staff(staff_code):
                    if cfg.get("staff_lookup_returns_none"):
                        return None
                    return {"staff_code": staff_code,
                              "full_name": "Demo Staff",
                              "role": "Branch Manager",
                              "unit": "Branch 015"}

                engine = CoachingIntelligence(
                    is_direct_report_fn=_is_dr,
                    staff_lookup_fn=_staff,
                    kpi_status_fn=lambda s: cfg["kpi_status"],
                    nudges_fn=lambda s: cfg["nudges"],
                    growth_plan_fn=lambda s: {
                        "skill_gaps": cfg["skill_gaps"],
                        "promotion_readiness_pct": 50,
                        "next_role": "Senior role"},
                    microtasks_fn=lambda s: cfg["microtasks"],
                    learning_cards_fn=lambda s: cfg["learning_cards"],
                )
                r = engine.generate_coaching_script(
                    "MGR_DEMO", "STAFF_DEMO", today=_date_ci(2026, 5, 1))

                if not r:
                    if not cfg["is_direct_report"]:
                        st.error(
                            "⛔ **Authorization refused** — engine returned `{}` "
                            "because manager is NOT this staff's direct supervisor. "
                            "**Engine enforces hierarchical access control** to prevent "
                            "managers coaching outside their reporting chain.")
                    elif cfg.get("staff_lookup_returns_none"):
                        st.error(
                            "⛔ **Staff not found** — engine returned `{}` per Rule 6. "
                            "Production deployment must ensure staff register is current.")
                    else:
                        st.error("⛔ Engine returned empty (unknown reason).")
                else:
                    agenda = r.get("meeting_agenda", [])
                    talking = r.get("talking_points", [])
                    actions = r.get("recommended_actions", [])
                    signals = r.get("meta", {}).get("signals_used", {})

                    k1, k2, k3 = st.columns(3)
                    k1.metric("Agenda items", len(agenda))
                    k2.metric("Talking points", len(talking))
                    k3.metric("Actions", len(actions))

                    st.markdown("**Meeting agenda:**")
                    for i, item in enumerate(agenda, 1):
                        st.markdown(f"{i}. {item}")

                    st.markdown("**Talking points:**")
                    for tp in talking:
                        st.markdown(f"- {tp}")

                    st.markdown("**Recommended actions:**")
                    for a in actions:
                        st.markdown(f"- {a}")

                    st.caption(f"Signals used: {dict(signals)}")

                _audit_log_ci("IFRS_ENGINE_USED", uname,
                                f"Coaching #11: scenario '{scenario[:30]}'")

        # ──────── Engine Reference ────────
        with ci_sub_tabs[2]:
            st.markdown("**Engine Constants Reference** (single source of truth)")

            st.markdown("**Output structure constraints:**")
            const_rows = [
                {"Section": "📋 Meeting Agenda",
                  "Min items": DEFAULT_AGENDA_MIN,
                  "Max items": DEFAULT_AGENDA_MAX,
                  "Purpose": "Structured talking framework for the 1:1"},
                {"Section": "💬 Talking Points",
                  "Min items": DEFAULT_TALKING_POINTS_MIN,
                  "Max items": DEFAULT_TALKING_POINTS_MAX,
                  "Purpose": "Specific Q-style prompts surfacing signals"},
                {"Section": "✅ Recommended Actions",
                  "Min items": DEFAULT_ACTIONS_MIN,
                  "Max items": DEFAULT_ACTIONS_MAX,
                  "Purpose": "Concrete next steps with ownership"},
            ]
            st.dataframe(pd.DataFrame(const_rows),
                         use_container_width=True, hide_index=True)

            st.markdown("**7 DI callbacks** (engine integrates with HR ecosystem):")
            di_rows = [
                {"Callback": "is_direct_report_fn(manager, staff)",
                  "Returns": "bool — whether manager directly supervises staff",
                  "Default source": "target_cascade.json"},
                {"Callback": "staff_lookup_fn(staff_code)",
                  "Returns": "dict | None — staff record",
                  "Default source": "staff_register.xlsx"},
                {"Callback": "kpi_status_fn(staff_code)",
                  "Returns": "list[dict] — current KPI status",
                  "Default source": "target_cascade + bsc_engine"},
                {"Callback": "nudges_fn(staff_code)",
                  "Returns": "list[dict] — active nudges",
                  "Default source": "data/nudges.json"},
                {"Callback": "growth_plan_fn(staff_code)",
                  "Returns": "dict — skill_gaps, promotion_readiness",
                  "Default source": "data/growth_plans.json"},
                {"Callback": "microtasks_fn(staff_code)",
                  "Returns": "list[dict] — outstanding tasks",
                  "Default source": "data/microtasks.json"},
                {"Callback": "learning_cards_fn(staff_code)",
                  "Returns": "list[dict] — relevant learning",
                  "Default source": "data/learning_cards.json"},
            ]
            st.dataframe(pd.DataFrame(di_rows),
                         use_container_width=True, hide_index=True)

            st.markdown("**Authorization model:**")
            st.caption(
                "Engine enforces that the manager must be the staff's direct "
                "supervisor (per cascade-derived hierarchy). Manager A cannot "
                "generate a coaching script for Manager B's report. **Returns `{}` "
                "if authorization fails** — caller can detect this empty response "
                "and surface a permission error to the user.")

            st.markdown("**Rule 6 transparency:**")
            st.caption(
                "Engine returns `{}` for: unknown staff (staff_lookup returned None), "
                "unauthorized manager (not direct report), or any other inability to "
                "build a meaningful script. **The `signals_used` dict in meta** "
                "shows which signal sources contributed — caller can detect data "
                "gaps (e.g. growth_plan_present=0 means HR hasn't built a growth "
                "plan for this staff yet) and prompt remediation.")

            st.caption(
                "💡 **Strategic context — completes HR axis**: v5.79 People Section 0 "
                "added retrospective HR analytics (compensation equity, engagement, "
                "predictive performance). v5.84 added forward-looking HR planning. "
                "v5.93 Coaching Intelligence now adds **action-oriented coaching support** "
                "— turning insights into structured 1:1 conversations. The HR axis is "
                "now: retrospective (v5.79) + forward-looking (v5.84) + action-oriented "
                "(v5.93).")


# ════════════════════════════════════════════════════════════════════
# v10.438 — Section 4: 🏆 Recognition (Std #17 GamificationEngine)
# ════════════════════════════════════════════════════════════════════

with sections[4]:
    st.markdown("### 🏆 Recognition & Gamification")
    st.caption(
        "Badges awarded based on verifiable performance triggers. "
        "Std #17 (GamificationEngine) — badge accuracy ≥90% per audit gate G28."
    )

    try:
        from utils.gamification import (
            list_badges_for_staff, GamificationEngine, Badge,
        )
    except Exception as exc:  # noqa: BLE001
        st.error(f"Gamification engine unavailable: {exc}")
    else:
        recog_tabs = st.tabs([
            "🎖️ My Badges",
            "🏅 Team Leaderboard",
            "⚙️ Admin",
        ])

        # ── Tab 0: My Badges ──
        with recog_tabs[0]:
            staff_code_g = str(ud.get("staff_code", ""))
            if staff_code_g:
                my_badges = list_badges_for_staff(staff_code_g)
                if my_badges:
                    st.markdown(f"**{len(my_badges)} badge(s) earned:**")
                    badge_rows = [{
                        "Badge": b.get("badge_type", "")[:30],
                        "Title": b.get("title", "")[:40],
                        "Awarded": (b.get("awarded_at", "") or "")[:10],
                        "Reason": (b.get("reason", "") or "")[:60],
                        "Period": b.get("period", ""),
                    } for b in my_badges]
                    st.dataframe(pd.DataFrame(badge_rows),
                                use_container_width=True, hide_index=True)

                    # Badge type counts
                    from collections import Counter as _C
                    type_counts = _C(b.get("badge_type", "Unknown") for b in my_badges)
                    st.markdown("**Badge collection:**")
                    cols = st.columns(min(6, max(1, len(type_counts))))
                    for i, (btype, cnt) in enumerate(type_counts.most_common()):
                        cols[i % len(cols)].metric(btype[:25], cnt)
                else:
                    st.info(
                        "No badges yet. Badges award automatically when "
                        "you meet thresholds (100% achiever, most improved, "
                        "consistent high, comeback kid, team player, "
                        "perfect quarter)."
                    )
            else:
                st.warning("Cannot identify your staff code.")

        # ── Tab 1: Leaderboard ──
        with recog_tabs[1]:
            st.markdown("**Top performers — badge count this quarter**")
            try:
                from utils.db import db as _db
                all_badges = _db.load_json("badges.json", default=[])
                if isinstance(all_badges, list) and all_badges:
                    from collections import Counter as _C
                    badge_counts = _C(
                        b.get("staff_code", "")
                        for b in all_badges
                        if isinstance(b, dict) and b.get("staff_code")
                    )
                    top = badge_counts.most_common(15)
                    if top:
                        # Lookup names from registry if present
                        try:
                            from utils.db import db as _db2
                            register = _db2.load_dataframe("staff_register.xlsx")
                            name_map = dict(zip(
                                register["Staff Code"].astype(str),
                                register["Staff Name"].astype(str),
                            ))
                        except Exception:  # noqa: BLE001
                            name_map = {}
                        lb_rows = [{
                            "Rank": i + 1,
                            "Staff": name_map.get(str(code), str(code))[:30],
                            "Code": code,
                            "Badges": cnt,
                        } for i, (code, cnt) in enumerate(top)]
                        st.dataframe(pd.DataFrame(lb_rows),
                                    use_container_width=True, hide_index=True)
                    else:
                        st.info("No badges awarded yet.")
                else:
                    st.info("No badges in registry. Admin can trigger evaluation.")
            except Exception as exc:  # noqa: BLE001
                st.error(f"Leaderboard unavailable: {exc}")

        # ── Tab 2: Admin ──
        with recog_tabs[2]:
            if is_admin or is_hr:
                st.markdown("**Trigger badge evaluation**")
                st.write("Evaluates all badge types for all staff. "
                        "Idempotent — won't double-award.")
                if st.button("Evaluate badges (all staff)", key="ppl_eval_badges"):
                    try:
                        engine = GamificationEngine()
                        # Run on a few sample staff for demo
                        sample = str(ud.get("staff_code", "300001"))
                        awarded = engine.evaluate_all_badges(sample)
                        st.success(f"✓ Evaluated for {sample}: "
                                  f"{len(awarded)} badge(s) qualifying.")
                    except Exception as exc:  # noqa: BLE001
                        st.error(f"Evaluation failed: {exc}")
            else:
                st.info("Admin and HR only.")


# ════════════════════════════════════════════════════════════════════
# v10.440 — Section 5: 🌿 Wellness (Std #19 WellnessEngine)
# ════════════════════════════════════════════════════════════════════

with sections[5]:
    st.markdown("### 🌿 Wellness & Burnout Risk Monitoring")
    st.caption(
        "Defensive monitoring across 4 signals: escalation frequency, "
        "stale micro-tasks, declining trajectory, alert clustering. "
        "Std #19 (WellnessEngine) — ethical guardrails: never produces "
        "medical/emotional speculation; opt-out respected."
    )

    try:
        from utils.wellness import WellnessEngine, list_alerts_for_manager
    except Exception as exc:  # noqa: BLE001
        st.error(f"Wellness engine unavailable: {exc}")
    else:
        wellness_tabs = st.tabs([
            "🙋 My wellness check",
            "👀 Team alerts (manager)",
            "ℹ️ How this works",
        ])

        # ── Tab 0: My wellness check ──
        with wellness_tabs[0]:
            staff_code_w = str(ud.get("staff_code", ""))
            if not staff_code_w:
                st.warning("Cannot identify your staff code.")
            else:
                st.write(
                    "Your wellness signals are computed from your work "
                    "data — not from self-reports. You can opt out by "
                    "setting `wellness_monitoring_disabled: true` on "
                    "your user record."
                )
                if st.button("Check my wellness signals", key="ppl_wellness_self"):
                    try:
                        engine = WellnessEngine()
                        result = engine.assess_burnout_risk(staff_code_w)
                        if not result:
                            st.info(
                                "No wellness data available — you may "
                                "have opted out, or your record isn't "
                                "in the lookup. This is a feature, "
                                "not a bug."
                            )
                        else:
                            risk_score = result.get("risk_score", 0)
                            risk_level = result.get("risk_level", "unknown")
                            signals = result.get("signals", {}) or {}
                            recs = result.get("recommendations", []) or []

                            level_color = {
                                "low": "🟢", "moderate": "🟡", "high": "🔴",
                            }.get(risk_level, "⚪")

                            c1, c2 = st.columns([1, 3])
                            c1.metric(
                                "Risk level",
                                f"{level_color} {risk_level.capitalize()}",
                                delta=f"score {risk_score:.2f}/1.0",
                            )
                            c2.markdown("**Signals contributing:**")
                            for sig_name, sig_value in signals.items():
                                c2.write(f"  • {sig_name}: {sig_value}")

                            if recs:
                                st.markdown("**Recommendations:**")
                                for r in recs:
                                    st.write(f"  • {r}")
                    except Exception as exc:  # noqa: BLE001
                        st.error(f"Wellness check failed: {exc}")

        # ── Tab 1: Team alerts (manager) ──
        with wellness_tabs[1]:
            mgr_code = str(ud.get("staff_code", ""))
            if mgr_code:
                try:
                    alerts = list_alerts_for_manager(mgr_code)
                    if alerts:
                        st.markdown(
                            f"**{len(alerts)} active wellness alert(s) "
                            f"for your team:**"
                        )
                        alert_rows = [{
                            "Staff": a.get("staff_name", "")[:25],
                            "Code": a.get("staff_code", ""),
                            "Risk": a.get("risk_level", ""),
                            "Score": round(a.get("risk_score", 0), 2),
                            "Triggered": (a.get("triggered_at", "") or "")[:10],
                            "Primary signal": a.get("primary_signal", "")[:30],
                        } for a in alerts]
                        st.dataframe(pd.DataFrame(alert_rows),
                                    use_container_width=True, hide_index=True)
                        st.warning(
                            "These alerts are not diagnoses. Have a "
                            "supportive conversation; consider workload "
                            "rebalancing; route to HR if patterns persist."
                        )
                    else:
                        st.success("✅ No active wellness alerts for your team.")
                except Exception as exc:  # noqa: BLE001
                    st.error(f"Alert listing failed: {exc}")
            else:
                st.info("Manager view requires staff code.")

        # ── Tab 2: How this works ──
        with wellness_tabs[2]:
            st.markdown("""
**Four signals computed from your work data:**

1. **Escalation frequency** — 8+ Std #11 alerts in 30 days
2. **Stale micro-tasks** — 5+ incomplete tasks older than 14 days
3. **Declining trajectory** — 3+ consecutive BSC achievement decreases
4. **Pace deficit** — recent achievement well below role baseline

**Risk levels:**
- 🟢 **Low**: score < 0.4
- 🟡 **Moderate**: 0.4 ≤ score < 0.7
- 🔴 **High**: score ≥ 0.7

**Ethical safeguards (G30 audit-gated):**
- Never produces medical or emotional speculation
- Forbidden words verified absent: `depressed`, `burnt out`, `stress disorder`, `mental health`, `anxiety`
- Opt-out respected: `wellness_monitoring_disabled` on user record
- High-risk alerts route to manager (Std #19 escalation)
- Recommendations focus on workload/process, not personal traits
            """)
