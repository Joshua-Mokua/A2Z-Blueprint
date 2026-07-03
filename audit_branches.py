#!/usr/bin/env python3
"""scripts/audit_branches.py — read-only audit of branch + reporting-line state.

Gathers the facts for the branch/reporting-line governance audit. WRITES NOTHING.
Run from the repo root:  python scripts\\audit_branches.py
"""
import json, os, collections
from utils.core import get_org_config

print("=" * 68)
print("BRANCH + REPORTING-LINE AUDIT (read-only)")
print("=" * 68)

# 1. org_config branches (the configured source of truth)
cfg = get_org_config() or {}
org_branches = [str(b.get("name", "")) for b in cfg.get("branches", [])]
org_regions = cfg.get("regions", [])
print(f"\n[1] org_config.json branches: {len(org_branches)}")
for b in sorted(org_branches):
    print(f"      - {b}")
print(f"    regions: {org_regions}")

# 2. staff register: Unit (=branch) and Region distribution
try:
    import openpyxl
    wb = openpyxl.load_workbook("data/staff_register.xlsx")
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    units = collections.Counter()
    regions = collections.Counter()
    reports = collections.Counter()
    dsa_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        unit = str(r[idx["Unit"]]) if r[idx["Unit"]] is not None else "(blank)"
        reg = str(r[idx["Region"]]) if r[idx["Region"]] is not None else "(blank)"
        rep = r[idx["Reports To"]]
        role = str(r[idx["Role"]]) if r[idx["Role"]] is not None else ""
        units[unit] += 1
        regions[reg] += 1
        reports[("set" if rep not in (None, "") else "BLANK")] += 1
        if "dsa" in role.lower() or "direct sales" in role.lower():
            dsa_rows.append((r[idx["Staff Code"]], role, unit, reg, rep))
    print(f"\n[2] staff register: {sum(units.values())} staff")
    print(f"    distinct Units (=branches): {len(units)}")
    for u, n in sorted(units.items()):
        print(f"      {n:>4}  {u}")
    print(f"\n    distinct Regions: {len(regions)}")
    for rg, n in sorted(regions.items()):
        print(f"      {n:>4}  {rg}")
    print(f"\n    Reports To populated: {dict(reports)}")

    # 3. ORPHAN CHECK — staff Units not in org_config branches
    orphan_units = sorted(set(units) - set(org_branches) - {"(blank)"})
    print(f"\n[3] ORPHAN staff Units (not in org_config branches): {len(orphan_units)}")
    for u in orphan_units:
        print(f"      - {u}  ({units[u]} staff)")
    empty_branches = sorted(set(org_branches) - set(units))
    print(f"    org_config branches with NO staff: {len(empty_branches)}")
    for b in empty_branches:
        print(f"      - {b}")

    # 4. DSA reporting lines (the dual-line concern)
    print(f"\n[4] DSA-type staff: {len(dsa_rows)}")
    for code, role, unit, reg, rep in dsa_rows[:30]:
        print(f"      {code} | {role} | unit={unit} | region={reg} | reports_to={rep}")
except Exception as e:
    print(f"\n[2-4] staff register read failed: {type(e).__name__}: {e}")

# 5. users table (Postgres) branch alignment — does staff Unit match there too?
try:
    from utils.db import db
    rows = db.fetch_all("SELECT unit, COUNT(*) AS n FROM users GROUP BY unit ORDER BY n DESC")
    print(f"\n[5] Postgres users table — distinct units: {len(rows)}")
    for r in rows[:50]:
        print(f"      {r['n']:>4}  {r['unit']}")
except Exception as e:
    print(f"\n[5] users-table unit read failed: {type(e).__name__}: {e}")

print("\n" + "=" * 68)
print("END AUDIT — no files were modified.")
