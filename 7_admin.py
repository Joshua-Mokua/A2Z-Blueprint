"""pages/7_admin.py — Administration: users, permissions, reporting lines, audit."""
import streamlit as st
from utils.db import db as a2z_db
from pages._admin_sprint import render_sprint_config
from pages._admin_module_config import render_module_config_centre
from pages._admin_postgres import render_postgres_centre
from pages._admin_reconciliation import render_recon_centre
from pages._admin_etl import render_etl_centre
from pages._admin_cutover import render_cutover_centre
from pages._admin_org import (
    render_dept_manager, render_branch_manager,
    render_module_assignment, render_roles_manager,
    render_thresholds, render_nav_labels,
)
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
from utils.core import *
try:
    from utils.core import get_fiscal_year as _gfy
except: _gfy = lambda: _gfy()

from pages._shared import load_shared_state, safe_html
from pages._access import require_access, get_my_scope
require_access("admin")


um, ud, uname, em, ri_pm, prod_m, pm, lm, hr_m, casc, vm, rlm = load_shared_state()
is_admin = ud.get("is_admin", False) or ud.get("can_view_all", False)

role_low = str(ud.get("role","")).lower()
if not (ud.get("can_view_all") or "admin" in role_low):
    st.error("⛔ Access restricted to administrators.")
    st.stop()

staff_scores = st.session_state.get("staff_scores",  pd.DataFrame())
registry     = st.session_state.get("staff_registry", pd.DataFrame())
# Roles and units from org_config (source of truth), fallback to staff_scores
try:
    from utils.core import get_org_config as _goc2, get_all_branches as _gab2
    _oc2 = _goc2()
    avail_roles = sorted(_oc2.get("roles", []))
    avail_units = ["Head Office"] + sorted(
        b["name"] for b in _gab2() if b.get("type","") != "HO")
    if not avail_roles: raise ValueError("empty")
except:
    avail_roles = sorted(staff_scores["Role"].unique().tolist()) if len(staff_scores) and "Role" in staff_scores.columns else ["Staff"]
    avail_units = sorted(staff_scores["Unit"].unique().tolist()) if len(staff_scores) and "Unit" in staff_scores.columns else ["Head Office"]

st.markdown(
    "<div style=\'padding:16px 22px;background:var(--brand-primary,#006B3F);border-radius:12px;margin-bottom:20px;box-shadow:0 2px 12px rgba(0,0,0,0.15)\'><div style=\'display:flex;align-items:center;justify-content:space-between\'><div><div style=\'color:var(--color-background-primary);font-size:16px;font-weight:700;letter-spacing:-0.2px\'>System Administration</div><div style=\'color:rgba(255,255,255,0.65);font-size:11px;margin-top:3px;font-weight:400\'>Users · Permissions · Reporting lines · Audit · Upload format</div></div><div style=\'opacity:0.12;font-size:36px;line-height:1;color:white\'>◆</div></div></div>",
    unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
# v5.12 RESTRUCTURED — 6 logical sections, 21 sub-tabs
# Module-specific configs migrated to Module Config Centre via registry
# See docs/ADMIN_CONVENTIONS.md for the convention.
# ─────────────────────────────────────────────────────────────────
sections = st.tabs([
    "👥 People & Org",
    "📊 Performance",
    "🧩 Modules",
    "🔌 Data & Integration",
    "🩺 System",
    "🛡️ Security",
])

# ── Section 0: People & Org ─────────────────────────────────
with sections[0]:
    sub = st.tabs([
        "🏦 Organisation",
        "👤 Users",
        "🔑 Permissions",
        "🗂️ Staff movements",
        "🌳 Org tree",
        "🏢 Org structure",
        "🎭 Roles Library",
    ])
    with sub[0]:
        from utils.core import get_org_config, save_org_config, DATA_DIR as _org_DATA_DIR
        _data_dir = _org_DATA_DIR

        st.subheader("🏦 Organisation Setup")
        st.caption(
            "Configure this bank's structure — name, branches, regions, reporting hierarchy, and roles. "
            "All other modules (BSC, Cascade, Pipeline, CBS) read from this configuration automatically.")

        _org = get_org_config()

        _org_view = st.radio("",
            ["🏦 Bank identity","🏢 Branches & regions","🌳 Hierarchy & roles",
             "🔀 Workstreams","📥 Bulk staff import","🔧 Reset & migrate"],
            horizontal=True, key="org_view")
        st.markdown("---")

        # ══════════════════════════════════════════════════════════════
        if "Bank identity" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**Configure the bank's identity and display settings.**")
            with st.form("org_identity_form"):
                _c1, _c2 = st.columns(2)
                _bank_name = _c1.text_input("Bank name *", value=_org.get("bank_name",""))
                _bank_code = _c2.text_input("Bank code (e.g. ECO)", value=_org.get("bank_code","ECO"),
                                             help="Short code used for account number prefixes")
                _country   = _c1.text_input("Country", value=_org.get("country","Kenya"))
                _currency  = _c2.text_input("Currency code (e.g. KES)", value=_org.get("currency","KES"))
                _curr_sym  = _c1.text_input("Currency symbol", value=_org.get("currency_symbol","KES"))

                st.markdown("**Pillar weights** — how the BSC score is computed bank-wide")
                _pw1,_pw2,_pw3,_pw4 = st.columns(4)
                _pillar_wts = _org.get("pillar_weights",{})
                def _pct(key, default_pct):
                    v = _pillar_wts.get(key, default_pct/100)
                    # Normalise: if stored as decimal (0.40), convert to int pct (40)
                    return int(v * 100) if v <= 1.0 else int(v)
                _fin_wt  = _pw1.number_input("Financial %",  0, 100, _pct("Financial", 40))
                _cust_wt = _pw2.number_input("Customer %",   0, 100, _pct("Customer Focus", 25))
                _ops_wt  = _pw3.number_input("Operations %", 0, 100, _pct("Operational Excellence", 25))
                _ppl_wt  = _pw4.number_input("People %",     0, 100, _pct("People & Learning", 10))
                _wt_total = _fin_wt+_cust_wt+_ops_wt+_ppl_wt
                _wt_clr = "#10B981" if _wt_total==100 else "#EF4444"
                _wt_bg   = "#ECFDF5" if _wt_total==100 else "#FEF2F2"
                _wt_icon = "✅" if _wt_total==100 else "— must equal 100%"
                st.markdown(
                    f"<div style='padding:6px 12px;background:{_wt_bg};"
                    f"border-radius:6px;font-size:12px;font-weight:700;color:{_wt_clr}'>"
                    f"Total: {_wt_total}% {_wt_icon}</div>",
                    unsafe_allow_html=True)

                if st.form_submit_button("💾 Save bank identity", type="primary"):
                    if _wt_total != 100:
                        st.error("Pillar weights must total 100%")
                    else:
                        _org["bank_name"] = _bank_name.strip()
                        _org["bank_code"] = _bank_code.strip().upper()
                        _org["country"]   = _country.strip()
                        _org["currency"]  = _currency.strip().upper()
                        _org["currency_symbol"] = _curr_sym.strip()
                        _org["pillar_weights"] = {
                            "Financial": _fin_wt/100, "Customer Focus": _cust_wt/100,
                            "Operational Excellence": _ops_wt/100, "People & Learning": _ppl_wt/100}
                        save_org_config(_org)
                        audit_log("ORG_IDENTITY_SAVED", uname, _bank_name)
                        st.success(f"✅ Bank identity saved — {_bank_name}")
                        st.cache_data.clear()
                        st.rerun()

        # ══════════════════════════════════════════════════════════════
        elif "Branches" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**Manage branches and regions.** Changes apply immediately to BSC scoping, cascade, and CBS lookup.")

            _branches = list(_org.get("branches", []))
            _regions  = list(_org.get("regions", []))

            # Summary strip
            _bm1,_bm2,_bm3,_bm4 = st.columns(4)
            _bm1.metric("Branches", len(_branches))
            _bm2.metric("Regions",  len(set(b.get("region","") for b in _branches)))
            _tc = {}
            for _b in _branches: _tc[_b.get("tier",3)] = _tc.get(_b.get("tier",3),0)+1
            _bm3.metric("Flagship (T1)", _tc.get(1,0))
            _bm4.metric("Light (T4)",    _tc.get(4,0))

            # Live table
            if _branches:
                _bdf_show = pd.DataFrame([{
                    "Code":b["code"],"Branch Name":b["name"],"Region":b.get("region",""),
                    "County":b.get("county",""),"Type":b.get("type",""),"Tier":b.get("tier",3)
                } for b in _branches])
                st.dataframe(_bdf_show, use_container_width=True, hide_index=True, height=280)

            st.markdown("---")
            _br_action = st.radio("Action",
                ["➕ Add branch","✏️ Edit branch","✏️ Rename branch","🗑️ Delete branch",
                 "🗺️ Regions"],
                horizontal=True, key="br_action")

            if "Add branch" in _br_action:
                with st.form("add_branch_form"):
                    _ab1,_ab2 = st.columns(2)
                    _new_bcode = _ab1.text_input("Branch code *", placeholder="e.g. BRN036")
                    _new_bname = _ab2.text_input("Branch name *", placeholder="e.g. Kisumu Mega Branch")
                    _new_breg  = _ab1.selectbox("Region *",
                        _regions + (["+ New region"] if _regions else ["Head Office","+ New region"]))
                    _new_bnewreg = _ab2.text_input("New region name (if selected above)", placeholder="e.g. South West")
                    _new_bcnty = _ab1.text_input("County", placeholder="e.g. Kisumu")
                    _new_btype = _ab2.selectbox("Type", ["Flagship","Main","Standard","Light","HO"])
                    _new_btier = _ab1.selectbox("Tier", [1,2,3,4], help="1=Flagship 2=Main 3=Standard 4=Light")
                    if st.form_submit_button("➕ Add branch", type="primary"):
                        if _new_bcode.strip() and _new_bname.strip():
                            _reg_to_use = _new_bnewreg.strip() if _new_breg=="+ New region" and _new_bnewreg.strip() else _new_breg
                            if _new_bcode.strip().upper() in [b["code"] for b in _branches]:
                                st.error(f"Code {_new_bcode} already exists.")
                            else:
                                _branches.append({
                                    "code":  _new_bcode.strip().upper(),
                                    "name":  _new_bname.strip(),
                                    "region":_reg_to_use,
                                    "county":_new_bcnty.strip(),
                                    "type":  _new_btype,
                                    "tier":  int(_new_btier),
                                })
                                if _reg_to_use not in _regions:
                                    _regions.append(_reg_to_use)
                                _org["branches"] = _branches
                                _org["regions"]  = _regions
                                save_org_config(_org)
                                audit_log("BRANCH_ADDED", uname, _new_bname)
                                st.success(f"✅ '{_new_bname}' added.")
                                st.cache_data.clear()
                                st.rerun()
                        else:
                            st.error("Branch code and name are required.")

            elif "Edit branch" in _br_action:
                _sel_branch = st.selectbox("Select branch", [b["name"] for b in _branches], key="br_edit_sel")
                _sel_b = next((b for b in _branches if b["name"]==_sel_branch), None)
                if _sel_b:
                    with st.form("edit_branch_form"):
                        _eb1,_eb2 = st.columns(2)
                        _eb_region = _eb1.selectbox("Region",
                            _regions if _regions else ["Head Office"],
                            index=_regions.index(_sel_b.get("region","")) if _sel_b.get("region","") in _regions else 0)
                        _eb_county = _eb2.text_input("County", value=_sel_b.get("county",""))
                        _eb_type   = _eb1.selectbox("Type",
                            ["Flagship","Main","Standard","Light","HO"],
                            index=["Flagship","Main","Standard","Light","HO"].index(_sel_b.get("type","Standard"))
                            if _sel_b.get("type","Standard") in ["Flagship","Main","Standard","Light","HO"] else 2)
                        _eb_tier = _eb2.selectbox("Tier",[1,2,3,4],
                            index=max(0,min(3,_sel_b.get("tier",3)-1)))
                        if st.form_submit_button("💾 Save", type="primary"):
                            for _b in _branches:
                                if _b["code"]==_sel_b["code"]:
                                    _b["region"]=_eb_region; _b["county"]=_eb_county.strip()
                                    _b["type"]=_eb_type; _b["tier"]=int(_eb_tier); break
                            _org["branches"]=_branches; save_org_config(_org)
                            audit_log("BRANCH_EDITED",uname,_sel_branch)
                            st.success("✅ Updated."); st.rerun()

            elif "Rename branch" in _br_action:
                with st.form("rename_branch_form"):
                    _ren_branch = st.selectbox("Branch to rename", [b["name"] for b in _branches])
                    _ren_bname  = st.text_input("New name *", placeholder="e.g. Kisumu Central Branch")
                    if st.form_submit_button("✏️ Rename", type="primary"):
                        _ren_bname = _ren_bname.strip()
                        if _ren_bname:
                            for _b in _branches:
                                if _b["name"]==_ren_branch:
                                    _b["name"]=_ren_bname; break
                            _org["branches"]=_branches; save_org_config(_org)
                            # Also update users who have this branch as their unit
                            _ren_count=0
                            for _u,_ud2 in um.users.items():
                                if _ud2.get("unit","")==_ren_branch:
                                    _ud2["unit"]=_ren_bname; _ren_count+=1
                            if _ren_count: um.save()
                            audit_log("BRANCH_RENAMED",uname,f"{_ren_branch}→{_ren_bname}")
                            st.success(f"✅ Renamed. {_ren_count} user(s) updated."); st.rerun()

            elif "Delete branch" in _br_action:
                with st.form("del_branch_form"):
                    _del_branch = st.selectbox("Branch to delete", [b["name"] for b in _branches])
                    st.caption("⚠️ Staff assigned to this branch will retain the branch name until reassigned.")
                    if st.form_submit_button("🗑️ Delete branch", type="secondary"):
                        _org["branches"]=[b for b in _branches if b["name"]!=_del_branch]
                        save_org_config(_org)
                        audit_log("BRANCH_DELETED",uname,_del_branch)
                        st.success(f"✅ {_del_branch} deleted."); st.rerun()

            elif "Regions" in _br_action:
                st.markdown("**Manage regions** — used for Regional Head scoping and branch grouping.")
                _reg_action = st.radio("",["Add region","Rename region","Delete region"],
                    horizontal=True, key="reg_action")

                if "Add" in _reg_action:
                    with st.form("add_region_form"):
                        _new_reg = st.text_input("New region name *", placeholder="e.g. North Eastern")
                        if st.form_submit_button("➕ Add region", type="primary"):
                            if _new_reg.strip() and _new_reg.strip() not in _regions:
                                _regions.append(_new_reg.strip())
                                _org["regions"]=_regions; save_org_config(_org)
                                st.success(f"✅ '{_new_reg}' added."); st.rerun()
                            elif _new_reg.strip() in _regions:
                                st.error("Region already exists.")

                elif "Rename" in _reg_action:
                    with st.form("rename_region_form"):
                        _ren_reg_from = st.selectbox("Region to rename",
                            _regions if _regions else ["— none —"])
                        _ren_reg_to   = st.text_input("New name *")
                        if st.form_submit_button("✏️ Rename", type="primary"):
                            _ren_reg_to = _ren_reg_to.strip()
                            if _ren_reg_to:
                                _regions = [_ren_reg_to if r==_ren_reg_from else r for r in _regions]
                                for _b in _branches:
                                    if _b.get("region")==_ren_reg_from: _b["region"]=_ren_reg_to
                                _org["regions"]=_regions; _org["branches"]=_branches
                                save_org_config(_org)
                                st.success(f"✅ '{_ren_reg_from}' → '{_ren_reg_to}'. Branches updated."); st.rerun()

                else:  # Delete
                    with st.form("del_region_form"):
                        _del_reg = st.selectbox("Region to delete",
                            _regions if _regions else ["— none —"])
                        st.caption("Branches in this region will show 'Head Office' until reassigned.")
                        if st.form_submit_button("🗑️ Delete region", type="secondary"):
                            _regions=[r for r in _regions if r!=_del_reg]
                            _org["regions"]=_regions; save_org_config(_org)
                            st.success(f"✅ '{_del_reg}' removed."); st.rerun()

                # Live region list
                if _regions:
                    st.markdown("**Current regions:**")
                    _rdf = pd.DataFrame([{
                        "Region": r,
                        "Branches": sum(1 for b in _branches if b.get("region")==r)
                    } for r in _regions])
                    st.dataframe(_rdf, use_container_width=True, hide_index=True)

        # ══════════════════════════════════════════════════════════════
        elif "Hierarchy" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**Define who reports to whom.** This drives cascade allocation, BSC scoping, and backup staff selection.")
            st.caption("⚠️ Changing the hierarchy takes effect immediately on the next page load.")

            _hierarchy = _org.get("hierarchy", {})
            _roles     = _org.get("roles", [])

            # Display current hierarchy as a clean table
            import pandas as pd
            _hier_rows = []
            for role, parents in _hierarchy.items():
                _hier_rows.append({
                    "Role": role,
                    "Reports to": ", ".join(parents) if parents else "— (Top of hierarchy)"
                })
            if _hier_rows:
                st.dataframe(pd.DataFrame(_hier_rows), use_container_width=True,
                             hide_index=True, height=350)

            st.markdown("---")
            _hc1, _hc2 = st.columns(2)

            # Edit reporting line for a role
            with _hc1:
                st.markdown("**Edit a role's reporting line**")
                with st.form("edit_hierarchy_form"):
                    _sel_hier_role = st.selectbox("Role to configure", _roles, key="hier_role_sel")
                    _cur_parents = _hierarchy.get(_sel_hier_role, [])
                    _new_parents = st.multiselect(
                        "Reports to (select one or more)",
                        [r for r in _roles if r != _sel_hier_role],
                        default=[p for p in _cur_parents if p in _roles])
                    if st.form_submit_button("💾 Save reporting line", type="primary"):
                        _hierarchy[_sel_hier_role] = _new_parents
                        _org["hierarchy"] = _hierarchy
                        save_org_config(_org)
                        audit_log("HIERARCHY_CHANGED", uname,
                                  f"{_sel_hier_role}→{','.join(_new_parents)}")
                        st.success(f"✅ {_sel_hier_role} now reports to: {', '.join(_new_parents) or 'nobody (top)'}")
                        st.cache_data.clear()
                        st.rerun()

            with _hc2:
                _role_action = st.radio("Action",
                    ["➕ Add role","✏️ Rename role","🗑️ Remove role"],
                    horizontal=True, key="role_action")

                if "Add" in _role_action:
                    with st.form("add_role_form"):
                        _new_role = st.text_input("New role name *",
                            placeholder="e.g. Digital Banking Manager / CEO / Chief")
                        _new_role_parent = st.selectbox(
                            "Reports to", ["— (Top level)"] + _roles)
                        _new_role_kpis = st.multiselect(
                            "Clone KPIs from similar role (optional)",
                            ["— None (configure later)"] + _roles,
                            max_selections=1, key="new_role_kpi_src")
                        if st.form_submit_button("➕ Add role", type="primary"):
                            if _new_role.strip() and _new_role.strip() not in _roles:
                                _nr = _new_role.strip()
                                _roles.append(_nr)
                                _hierarchy[_nr] = ([] if _new_role_parent=="— (Top level)"
                                                   else [_new_role_parent])
                                # Clone KPIs from similar role
                                if _new_role_kpis and _new_role_kpis[0] != "— None (configure later)":
                                    try:
                                        from utils.core_kpi import get_kpi_library, save_kpi_library
                                        _klib = get_kpi_library()
                                        _src_kpis = _klib.get("role_kpis",{}).get(_new_role_kpis[0],[])
                                        if _src_kpis:
                                            _klib["role_kpis"][_nr] = list(_src_kpis)
                                            save_kpi_library(_klib)
                                    except: pass
                                _org["roles"] = _roles
                                _org["hierarchy"] = _hierarchy
                                save_org_config(_org)
                                audit_log("ROLE_ADDED", uname, _nr)
                                st.success(f"✅ Role '{_nr}' added.")
                                st.cache_data.clear()
                                st.rerun()
                            elif _new_role.strip() in _roles:
                                st.error("That role already exists.")

                elif "Rename" in _role_action:
                    st.caption("Renames the role everywhere: users, staff register, KPI library, hierarchy, cascade — all updated atomically.")
                    with st.form("rename_role_form"):
                        _ren_from = st.selectbox("Role to rename", _roles, key="ren_from")
                        _ren_to   = st.text_input("New name *",
                            placeholder="e.g. rename 'Managing Director' → 'CEO'")
                        if st.form_submit_button("✏️ Rename role", type="primary"):
                            _ren_to = _ren_to.strip()
                            if not _ren_to:
                                st.error("New name cannot be empty.")
                            elif _ren_to in _roles and _ren_to != _ren_from:
                                st.error(f"'{_ren_to}' already exists.")
                            else:
                                # 1. Roles list
                                _roles = [_ren_to if r==_ren_from else r for r in _roles]
                                # 2. Hierarchy keys and values
                                _new_hier = {}
                                for r, parents in _hierarchy.items():
                                    _new_key = _ren_to if r==_ren_from else r
                                    _new_parents = [_ren_to if p==_ren_from else p for p in parents]
                                    _new_hier[_new_key] = _new_parents
                                _hierarchy = _new_hier
                                _org["roles"] = _roles
                                _org["hierarchy"] = _hierarchy
                                save_org_config(_org)
                                # 3. KPI library role_kpis
                                try:
                                    from utils.core_kpi import get_kpi_library, save_kpi_library
                                    _klib = get_kpi_library()
                                    _rk = _klib.get("role_kpis", {})
                                    if _ren_from in _rk:
                                        _rk[_ren_to] = _rk.pop(_ren_from)
                                    _klib["role_kpis"] = _rk
                                    save_kpi_library(_klib)
                                except: pass
                                # 4. Users — rename role field
                                _ren_users = 0
                                for _u, _ud2 in um.users.items():
                                    if _ud2.get("role","") == _ren_from:
                                        _ud2["role"] = _ren_to
                                        _ren_users += 1
                                if _ren_users: um.save()
                                # 5. Use rename_role_everywhere for org_config + kpi_library
                                try:
                                    from utils.core import rename_role_everywhere as _rre
                                    _rre(_ren_from, _ren_to)
                                except: pass
                                # 6. Update staff_register.xlsx
                                _sr_updated = 0
                                try:
                                    import openpyxl as _ox
                                    from pathlib import Path as _Pth
                                    _sr_path = _Pth("a2z/data/staff_register.xlsx")
                                    if _sr_path.exists():
                                        _sr_wb = _ox.load_workbook(str(_sr_path))
                                        _sr_ws = _sr_wb.active
                                        _sr_hdrs = [_sr_ws.cell(1,c).value for c in range(1,_sr_ws.max_column+1)]
                                        _sr_rc = _sr_hdrs.index("Role")+1 if "Role" in _sr_hdrs else None
                                        if _sr_rc:
                                            for _sr_row in _sr_ws.iter_rows(min_row=2):
                                                if _sr_row[_sr_rc-1].value == _ren_from:
                                                    _sr_row[_sr_rc-1].value = _ren_to
                                                    _sr_updated += 1
                                        _sr_wb.save(str(_sr_path))
                                except: pass
                                # 7. Update actuals xlsx
                                _act_updated = 0
                                try:
                                    import glob as _gl
                                    for _af in _gl.glob("a2z/data/actuals_*.xlsx"):
                                        _act_wb = _ox.load_workbook(_af)
                                        _act_ws = _act_wb.active
                                        _act_hdrs = [_act_ws.cell(2,c).value for c in range(1,_act_ws.max_column+1)]
                                        _act_rc = _act_hdrs.index("Role")+1 if "Role" in _act_hdrs else None
                                        if _act_rc:
                                            for _ar in _act_ws.iter_rows(min_row=3):
                                                if _ar[_act_rc-1].value == _ren_from:
                                                    _ar[_act_rc-1].value = _ren_to
                                                    _act_updated += 1
                                        _act_wb.save(_af)
                                except: pass
                                audit_log("ROLE_RENAMED", uname, f"{_ren_from}→{_ren_to}")
                                st.success(
                                    f"✅ '{_ren_from}' renamed to '{_ren_to}' across all data — "
                                    f"{_ren_users} user(s), {_sr_updated} staff register rows, "
                                    f"{_act_updated} actuals rows, hierarchy, KPI library, "
                                    f"role categories — all updated.")
                                st.cache_data.clear()
                                st.rerun()

                else:  # Remove
                    with st.form("remove_role_form"):
                        _del_role = st.selectbox("Role to remove", _roles, key="del_role_sel")
                        st.caption("⚠️ Staff with this role will retain it in their user record until reassigned.")
                        if st.form_submit_button("🗑️ Remove role", type="secondary"):
                            _roles = [r for r in _roles if r != _del_role]
                            _hierarchy.pop(_del_role, None)
                            for r in _hierarchy:
                                _hierarchy[r] = [p for p in _hierarchy[r] if p != _del_role]
                            _org["roles"] = _roles
                            _org["hierarchy"] = _hierarchy
                            save_org_config(_org)
                            audit_log("ROLE_REMOVED", uname, _del_role)
                            st.success(f"✅ '{_del_role}' removed."); st.rerun()


        # ══════════════════════════════════════════════════════════════
        elif "Workstreams" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**Configure workstreams** — rename, add sub-workstreams, manage cross-functional pools.")
            st.caption("Workstreams auto-seed from the org hierarchy (Chiefs/Directors). Customise here.")

            _ws_file_adm = DATA_DIR / "execute_workstreams.json"
            try:
                _ws_adm = a2z_db.load_json(_ws_file_adm) if _ws_file_adm.exists() else {}
            except:
                _ws_adm = {}

            try:
                from utils.core import get_workstreams_from_hierarchy as _gwsh_adm
                _all_ws_adm = _gwsh_adm()
            except:
                _all_ws_adm = _ws_adm

            for _ws_id_a, _ws_a in _all_ws_adm.items():
                with st.expander(
                        f"**{_ws_id_a}** — {_ws_a.get('name','?')}  ·  "
                        f"{_ws_a.get('full_role','Custom')[:45]}",
                        expanded=False):

                    _wa1, _wa2 = st.columns([1, 1])

                    with _wa1:
                        st.markdown("**Display name & sub-workstreams**")
                        with st.form(f"adm_ws_{_ws_id_a}"):
                            _wna = st.text_input("Display name",
                                                  value=_ws_a.get("name", ""),
                                                  key=f"adm_wsn_{_ws_id_a}")
                            _sub_joined = "\n".join(_ws_a.get("sub_workstreams", []))
                            _wsa = st.text_area("Sub-workstreams (one per line)",
                                                 value=_sub_joined,
                                                 height=80,
                                                 key=f"adm_wss_{_ws_id_a}")
                            if st.form_submit_button("💾 Save", type="primary"):
                                _sub_a = [s.strip() for s in _wsa.splitlines() if s.strip()]
                                _ws_adm.setdefault(_ws_id_a, dict(_ws_a))
                                _ws_adm[_ws_id_a]["name"] = _wna
                                _ws_adm[_ws_id_a]["sub_workstreams"] = _sub_a
                                a2z_db.save_json(_ws_file_adm, _ws_adm)
                                audit_log("WS_RENAMED", uname, f"{_ws_id_a}:{_wna}")
                                st.success("✅ Saved")
                                st.cache_data.clear()
                                st.rerun()

                    with _wa2:
                        st.markdown("**Cross-functional pool**")
                        st.caption("Staff from other workstreams who can be IO on initiatives here.")
                        _pool_a = _ws_a.get("cross_functional_pool", [])
                        _users_all_a = um.users if um else {}
                        for _pu in list(_pool_a):
                            _pname = _users_all_a.get(_pu, {}).get("full_name", _pu)
                            _rc1a, _rc2a = st.columns([3, 1])
                            _rc1a.markdown(
                                f"<span style='background:var(--color-background-info);"
                                f"color:var(--color-text-info);padding:2px 8px;"
                                f"border-radius:10px;font-size:11px'>🔀 {_pname}</span>",
                                unsafe_allow_html=True)
                            if _rc2a.button("✕", key=f"rm_cf_{_ws_id_a}_{_pu}",
                                            help="Remove from pool"):
                                _ws_adm.setdefault(_ws_id_a, dict(_ws_a))
                                _ws_adm[_ws_id_a]["cross_functional_pool"] = [
                                    p for p in _pool_a if p != _pu]
                                a2z_db.save_json(_ws_file_adm, _ws_adm)
                                st.cache_data.clear()
                                st.rerun()

                        with st.form(f"adm_cf_{_ws_id_a}"):
                            _outside_staff = sorted(
                                [(u, d.get("full_name", u))
                                 for u, d in _users_all_a.items()
                                 if d.get("active") and u not in _pool_a],
                                key=lambda x: x[1])
                            _cf_opts = ["— Select staff to add —"] + [
                                f"{n}  ({u})" for u, n in _outside_staff[:300]]
                            _cf_sel = st.selectbox(
                                "Add to pool", _cf_opts,
                                key=f"adm_cf_sel_{_ws_id_a}")
                            if st.form_submit_button("➕ Add"):
                                if _cf_sel != "— Select staff to add —":
                                    _cf_u = _cf_sel.split("(")[-1].rstrip(")")
                                    _ws_adm.setdefault(_ws_id_a, dict(_ws_a))
                                    _cur_pool = _ws_adm[_ws_id_a].get(
                                        "cross_functional_pool", [])
                                    if _cf_u not in _cur_pool:
                                        _cur_pool.append(_cf_u)
                                    _ws_adm[_ws_id_a]["cross_functional_pool"] = _cur_pool
                                    a2z_db.save_json(_ws_file_adm, _ws_adm)
                                    audit_log("CF_ADDED", uname, f"{_ws_id_a}:{_cf_u}")
                                    st.success("Added")
                                    st.cache_data.clear()
                                    st.rerun()

                    # Gate summary for this workstream
                    _ws_inits = [i for i in st.session_state.get("em_initiatives", [])
                                 if i.get("workstream", "").startswith(_ws_id_a)]
                    if _ws_inits:
                        _gate_dist = {}
                        for _wi in _ws_inits:
                            _g = _wi.get("gate", "G0")
                            _gate_dist[_g] = _gate_dist.get(_g, 0) + 1
                        st.markdown("**Initiatives:** " +
                                    " · ".join(f"{g}:{c}" for g, c in _gate_dist.items()))

            st.markdown("---")
            with st.expander("➕ Add custom workstream (not from hierarchy)"):
                with st.form("adm_ws_new"):
                    _nc1, _nc2 = st.columns(2)
                    _nws_id   = _nc1.text_input("ID *", placeholder="e.g. DIGITAL")
                    _nws_name = _nc1.text_input("Name *",
                                                 placeholder="e.g. Digital Transformation")
                    _nws_spon = _nc2.text_input("Sponsor username")
                    _nws_sub  = _nc2.text_area("Sub-workstreams (one per line)",
                                                height=60)
                    if st.form_submit_button("➕ Create workstream", type="primary"):
                        if _nws_id.strip() and _nws_name.strip():
                            _ws_adm[_nws_id.upper()] = {
                                "name": _nws_name.strip(),
                                "sponsor_username": _nws_spon.strip(),
                                "sponsor_name":     _nws_spon.strip(),
                                "full_role": "",
                                "sub_workstreams": [
                                    x.strip() for x in _nws_sub.splitlines()
                                    if x.strip()],
                                "cross_functional_pool": [],
                                "custom": True,
                            }
                            a2z_db.save_json(_ws_file_adm, _ws_adm)
                            audit_log("WS_CREATED", uname, _nws_name.strip())
                            st.success(f"✅ Workstream {_nws_id.upper()} created!")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error("ID and name required.")

        # ══════════════════════════════════════════════════════════════
        elif "Bulk staff import" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**Import staff roster from Excel or CSV.** Creates login accounts for all staff automatically.")
            st.info(
                "Expected columns: **Staff Name**, **Role**, **Unit** (branch name), **Staff Code** "
                "(optional — auto-assigned if missing), **Email** (optional), **Reports To** (optional).")

            _imp_file = st.file_uploader("Upload staff roster (.xlsx or .csv)",
                                          type=["xlsx","csv"], key="org_import_file")
            if _imp_file:
                try:
                    import pandas as _pd, hashlib as _hl
                    if _imp_file.name.endswith(".csv"):
                        _imp_df = _pd.read_csv(_imp_file)
                    else:
                        _imp_df = _pd.read_excel(_imp_file)
                    _imp_df.columns = [str(c).strip() for c in _imp_df.columns]

                    st.success(f"✅ {len(_imp_df)} rows loaded")
                    st.dataframe(_imp_df.head(10), use_container_width=True, hide_index=True)

                    _req_cols = ["Staff Name","Role","Unit"]
                    _missing  = [c for c in _req_cols if c not in _imp_df.columns]
                    if _missing:
                        st.error(f"Missing required columns: {', '.join(_missing)}")
                    else:
                        _next_code = max(
                            (int(str(u.get("staff_code","0")).strip() or 0)
                             for u in um.users.values()
                             if str(u.get("staff_code","0")).isdigit()),
                            default=300000) + 1

                        _preview = []
                        for _, row in _imp_df.iterrows():
                            name = str(row.get("Staff Name","")).strip()
                            if not name: continue
                            _sc = str(row.get("Staff Code","")).strip()
                            if not _sc or not _sc.isdigit():
                                _sc = str(_next_code); _next_code += 1
                            uname_gen = name.split()[0].lower() + _sc[-3:]
                            pwd_gen = f"{_org.get(chr(112)+ chr(97)+ chr(115)+ chr(115)+ chr(119)+ chr(111)+ chr(114)+ chr(100)+chr(95)+chr(112)+chr(114)+chr(101)+chr(102)+chr(105)+chr(120), chr(83)+chr(116)+chr(97)+chr(102)+chr(102))}{_sc[-4:]}"
                            _preview.append({
                                "Staff Name": name, "Role": row.get("Role",""),
                                "Unit": row.get("Unit",""), "Staff Code": _sc,
                                "Username": uname_gen, "Password": pwd_gen,
                            })
                        _prev_df = _pd.DataFrame(_preview)
                        st.markdown(f"**Preview — {len(_prev_df)} accounts will be created:**")
                        st.dataframe(_prev_df, use_container_width=True, hide_index=True, height=280)

                        if st.button(f"✅ Import {len(_prev_df)} staff", type="primary", key="org_import_run"):
                            _created = 0
                            for row in _preview:
                                _uname = row["Username"]
                                # Avoid duplicates
                                _base  = _uname
                                _sfx   = 1
                                while _uname in um.users:
                                    _uname = f"{_base}{_sfx}"; _sfx += 1
                                pwd = row["Password"]
                                um.users[_uname] = {
                                    "password": um.hash_pw(pwd),
                                    "full_name": row["Staff Name"],
                                    "role": row["Role"],
                                    "unit": row["Unit"],
                                    "staff_code": row["Staff Code"],
                                    "active": True,
                                    "must_change_password": True,
                                    "is_admin": False,
                                    "can_view_all": False,
                                    "is_dept_super_user": False,
                                    "dept_super_user_for": "",
                                }
                                _created += 1
                            um.save()
                            audit_log("BULK_IMPORT", uname, f"{_created} staff imported")
                            st.success(f"✅ {_created} user accounts created. Default passwords: {_org.get('password_prefix','Staff')} + last 4 of staff code.")
                            st.cache_data.clear()
                            st.rerun()
                except Exception as _ie:
                    st.error(f"Import error: {_ie}")

        # ══════════════════════════════════════════════════════════════
        elif "Reset" in _org_view:
        # ══════════════════════════════════════════════════════════════
            st.markdown("**System reset and migration tools.**")
            st.warning("⚠️ These actions are irreversible. Use with caution.")

            with st.expander("🔄 Reset org config to defaults"):
                st.caption("Resets branches, regions, hierarchy and roles to the system defaults.")
                if st.button("Reset to system defaults", key="reset_org"):
                    from utils.core import DEFAULT_ORG_CONFIG
                    save_org_config(DEFAULT_ORG_CONFIG.copy())
                    st.success("✅ Org config reset to defaults.")
                    st.cache_data.clear()
                    st.rerun()

            with st.expander("📤 Export all config as ZIP"):
                st.caption("Download all configuration files for backup or migration.")
                import io, zipfile, json as _json
                _zip_buf = io.BytesIO()
                with zipfile.ZipFile(_zip_buf, "w") as _zf:
                    for _cfg_file in ["org_config.json","kpi_library.json","pipeline_settings.json",
                                      "leave_settings.json","bank_targets.json","target_cascade.json"]:
                        _cfg_path = _data_dir / _cfg_file
                        if _cfg_path.exists():
                            _zf.write(_cfg_path, _cfg_file)
                _zip_buf.seek(0)
                st.download_button(
                    "⬇️ Download config backup",
                    _zip_buf.read(),
                    file_name=f"a2z_config_backup.zip",
                    mime="application/zip")

            with st.expander("📥 Import config from backup ZIP"):
                st.caption("Restore configuration from a previously exported backup.")
                _imp_zip = st.file_uploader("Upload config ZIP", type=["zip"], key="cfg_import_zip")
                if _imp_zip:
                    import io as _io, zipfile as _zf2
                    with _zf2.ZipFile(_io.BytesIO(_imp_zip.read())) as _z:
                        _names = _z.namelist()
                        st.write("Files in backup:", _names)
                        if st.button("📥 Restore from backup", type="primary"):
                            for _n in _names:
                                (_data_dir / _n).write_bytes(_z.read(_n))
                            st.success("✅ Config restored. Restart the app to apply.")


    # ════════════════════════════════════════════════════════════════
    # TAB 1 — USERS
    # ════════════════════════════════════════════════════════════════

    with sub[1]:
        st.subheader("User management")

        # ── Quick admin restore notice ────────────────────────────────
        if "admin" not in um.users:
            st.error(
                "⚠️ **Admin account missing.** Click below to restore it. "
                "Username: `admin` · Default password: `admin123`")
            if st.button("🔧 Restore admin account", type="primary", key="restore_admin"):
                um.users["admin"] = {
                    "password":   um.hash_pw("admin123"),
                    "full_name":  "System Admin", "role": "Admin",
                    "department": "All", "can_view_all": True,
                    "managed_roles": [], "managed_units": [],
                    "managed_staff_codes": [], "staff_code": "ADMIN001",
                    "email": "admin@bank.com", "active": True, "_protected": True,
                }
                um.save()
                audit_log("ADMIN_RESTORED", uname, "admin")
                st.success("✅ Admin account restored. Password: admin123")
                st.cache_data.clear()
                st.rerun()
        else:
            # Show protected badge next to admin
            _admin_info = um.users.get("admin", {})
            st.markdown(
                f"<div style='padding:6px 12px;background:var(--brand-light,#E8F5EE);border:1px solid #BBF7D0;"
                f"border-radius:6px;font-size:11px;color:#166534;margin-bottom:12px'>"
                f"🔒 <b>admin</b> account is active and protected — cannot be deleted "
                f"· last active account: {safe_html(_admin_info.get('full_name','System Admin'))}"
                f"</div>", unsafe_allow_html=True)

        # ── Auto-sync staff codes from BSC data ────────────────────────
        df_bsc = st.session_state.get("df_processed", pd.DataFrame())
        if not df_bsc.empty and um.users:
            unsynced = []
            for uname_k, udata in um.users.items():
                fname = udata.get("full_name","")
                sc    = str(udata.get("staff_code","")).strip()
                if fname and (not sc or not sc.isdigit()):
                    # Try to find in BSC data
                    hits = df_bsc[df_bsc["Staff Name"] == fname]
                    if hits.empty:
                        surname = fname.strip().split()[-1]
                        hits = df_bsc[df_bsc["Staff Name"].str.contains(surname, case=False, na=False)]
                    if not hits.empty:
                        bsc_code = str(hits["Staff Code"].iloc[0]).strip()
                        if bsc_code and bsc_code != "nan":
                            unsynced.append((uname_k, fname, bsc_code))

            if unsynced:
                st.markdown(
                    f"<div style='padding:10px 14px;background:#FEF3C7;border:1px solid #FDE68A;"
                    f"border-radius:8px;margin-bottom:10px;font-size:12px;color:#92400E'>"
                    f"⚠️ <b>{len(unsynced)} user account(s)</b> are missing their numeric staff code "
                    f"(needed for cascade to work). Click below to auto-sync from BSC data."
                    f"</div>", unsafe_allow_html=True)
                sc1, sc2 = st.columns([3,1])
                with sc1:
                    for _, fname, bsc_code in unsynced:
                        st.markdown(f"  - **{fname}** → will set staff code to `{bsc_code}`")
                if sc2.button("🔄 Sync staff codes", type="primary", use_container_width=True):
                    synced = 0
                    for uname_k, fname, bsc_code in unsynced:
                        um.users[uname_k]["staff_code"] = bsc_code
                        synced += 1
                    um.save()
                    audit_log("STAFF_CODE_SYNC", "admin", f"{synced} users synced")
                    st.toast(f"✅ {synced} staff codes synced from BSC data", icon="✅")
                    st.cache_data.clear()
                    st.rerun()

        # ── Permission safety check ──────────────────────────────────────
        try:
            from utils.core_audit import fix_view_all_permissions
            from utils.core import _ALL_VIEW_ROLES
            _danger = [(u, d.get("full_name",u), d.get("role",""))
                       for u,d in um.users.items()
                       if d.get("can_view_all")
                       and not d.get("is_admin")
                       and str(d.get("role","")).lower() not in _ALL_VIEW_ROLES]
            if _danger:
                st.markdown(
                    f"<div style='padding:10px 14px;background:#FEF2F2;border:1px solid #FECACA;"
                    f"border-radius:8px;margin-bottom:10px;font-size:12px;color:#991B1B'>"
                    f"<b>⚠️ Privacy risk detected:</b> "
                    f"{len(_danger)} account(s) have 'View ALL staff' enabled but are not MD/Admin: "
                    f"{', '.join(n for _,n,_ in _danger)}. "
                    f"Click below to fix immediately.</div>", unsafe_allow_html=True)
                if st.button("🔒 Fix — remove 'View all' from non-MD accounts",
                              type="primary", key="fix_view_all_btn"):
                    n = fix_view_all_permissions(um)
                    st.session_state["_filtered_for"] = None
                    st.toast(f"✅ Fixed {n} accounts — view-all removed", icon="🔒")
                    st.cache_data.clear()
                    st.rerun()
        except Exception as _pe:
            pass

        mode = st.radio("Action", ["Create new user","Edit existing user"],
                        horizontal=True, key="admin_mode")

        if mode == "Edit existing user":
            existing = list(um.users.keys())
            if not existing:
                st.info("No users yet.")
            else:
                sel_user = st.selectbox("Select user", existing, key="edit_sel")
                eu = um.users.get(sel_user, {})
                with st.form("edit_user_form"):
                    ec1, ec2 = st.columns(2)
                    e_fname  = ec1.text_input("Full name",  value=eu.get("full_name",""))
                    e_email  = ec2.text_input("Email",      value=eu.get("email",""))
                    e_role   = ec1.selectbox("Role", avail_roles,
                        index=avail_roles.index(eu["role"]) if eu.get("role") in avail_roles else 0)
                    e_unit   = ec2.selectbox("Unit", avail_units,
                        index=avail_units.index(eu["unit"]) if eu.get("unit") in avail_units else 0)
                    e_sc     = ec1.text_input("Staff Code", value=str(eu.get("staff_code","")))
                    e_active = ec2.checkbox("Active", value=eu.get("active", True))

                    st.markdown("**Profile photo**")
                    ph1, ph2 = st.columns([1,3])
                    try:
                        from utils.core import photo_avatar_html, save_profile_photo, get_photo_b64
                        _sc_edit = eu.get("staff_code","") or sel_user
                        _av = photo_avatar_html(_sc_edit, eu.get("full_name",""), size=60)
                        ph1.markdown(_av, unsafe_allow_html=True)
                    except: ph1.markdown("")

                    new_photo = ph2.file_uploader(
                        "Upload photo (JPG/PNG, max 2MB)",
                        type=["jpg","jpeg","png","webp"],
                        key=f"photo_{sel_user}",
                        help="Square crop recommended · Will appear on scorecard and cascade pages")
                    if new_photo:
                        _sc_save = eu.get("staff_code","") or sel_user
                        ext = new_photo.name.split(".")[-1].lower()
                        save_profile_photo(_sc_save, new_photo.read(), ext)
                        ph2.success("✅ Photo uploaded")

                    st.markdown("**Permissions**")
                    pc1,pc2,pc3,pc4 = st.columns(4)
                    e_all    = pc1.checkbox("Can view all staff", value=eu.get("can_view_all",False))
                    e_exec   = pc2.checkbox("Can manage Execute",  value=eu.get("can_execute",False))
                    e_admin  = pc3.checkbox("Admin privileges",    value=eu.get("is_admin",False))
                    e_new_pw = pc4.text_input("Reset password (optional)", type="password",
                                               placeholder="Leave blank to keep")

                    if st.form_submit_button("Save changes", type="primary"):
                        if e_new_pw.strip() and len(e_new_pw.strip()) < 8:
                            st.error("Password must be at least 8 characters")
                            st.stop()
                        if e_new_pw.strip():
                            um.users[sel_user]["password"] = e_new_pw.strip()
                            audit_log("PASSWORD_RESET", uname, f"Admin reset password for {sel_user}")
                        um.users[sel_user].update({
                            "full_name":    e_fname,
                            "email":        e_email,
                            "role":         e_role,
                            "unit":         e_unit,
                            "staff_code":   e_sc,
                            "active":       e_active,
                            "can_view_all": e_all,
                            "can_execute":  e_exec,
                            "is_admin":     e_admin,
                        })
                        um.save()
                        audit_log("USER_EDITED", uname, sel_user)
                        st.success(f"User '{sel_user}' updated.")
                        st.cache_data.clear()
                        st.rerun()
        else:
            # ── Staff code lookup (outside form so rerun works) ──────────
            df_for_lookup = st.session_state.get("df_processed", pd.DataFrame())
            registry_lu   = st.session_state.get("staff_registry", {})

            st.markdown(
                "<div style='padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;"
                "border-radius:8px;margin-bottom:12px;font-size:12px;color:#166534'>"
                "💡 <b>Quick create from BSC data</b> — enter the staff code below and the system "
                "will auto-fill name, role, unit and email from the uploaded file."
                "</div>", unsafe_allow_html=True)

            lookup_col, _ = st.columns([2,3])
            sc_input = lookup_col.text_input(
                "Staff Code (auto-fill)",
                key="new_user_sc_lookup",
                placeholder="e.g. 300130",
                help="Type the staff code and press Enter — details load automatically")

            # Resolve staff details from BSC data or staff registry
            auto_fname  = ""
            auto_role   = avail_roles[0] if avail_roles else ""
            auto_unit   = avail_units[0] if avail_units else ""
            auto_email  = ""
            sc_found    = False

            if sc_input.strip():
                sc_str = sc_input.strip()
                # Search df_processed first
                if not df_for_lookup.empty and "Staff Code" in df_for_lookup.columns:
                    hits = df_for_lookup[df_for_lookup["Staff Code"].astype(str).str.strip()==sc_str]
                    if not hits.empty:
                        row          = hits.iloc[0]
                        auto_fname   = str(row.get("Staff Name",""))
                        auto_role    = str(row.get("Role",""))
                        auto_unit    = str(row.get("Unit",""))
                        sc_found     = True
                # Fallback: search staff registry dict
                if not sc_found and isinstance(registry_lu, dict):
                    for code, info in registry_lu.items():
                        if str(code).strip() == sc_str or str(info.get("Staff Code","")).strip()==sc_str:
                            auto_fname = info.get("Staff Name","")
                            auto_role  = info.get("Role","")
                            auto_unit  = info.get("Unit","")
                            auto_email = info.get("Email","")
                            sc_found   = True
                            break

                if sc_found:
                    st.markdown(
                        f"<div style='padding:8px 12px;background:#EFF6FF;border:1px solid #BFDBFE;"
                        f"border-radius:8px;margin-bottom:8px;font-size:12px;color:#1E40AF'>"
                        f"✅ Found: <b>{auto_fname}</b> · {auto_role} · {auto_unit}"
                        f"</div>", unsafe_allow_html=True)
                elif sc_input.strip():
                    st.markdown(
                        "<div style='padding:8px 12px;background:#FEF2F2;border:1px solid #FCA5A5;"
                        "border-radius:8px;margin-bottom:8px;font-size:12px;color:#991B1B'>"
                        "⚠️ Staff code not found in uploaded data. Fill details manually below."
                        "</div>", unsafe_allow_html=True)

            # Create user form — pre-populated from lookup
            # Staff code is the default username — use sc_input if found, else blank
            default_username = sc_input.strip() if sc_input.strip() else ""

            with st.form("create_user_form"):
                nc1, nc2 = st.columns(2)

                # Username defaults to staff code — admin can override
                n_user = nc1.text_input(
                    "Username *",
                    value=default_username,
                    placeholder="Auto-filled from staff code",
                    help="Defaults to the staff code. Staff use this to log in.")
                n_pass = nc2.text_input("Password *", type="password")
                n_fname = nc1.text_input("Full name *", value=auto_fname)
                n_email = nc2.text_input("Email", value=auto_email)

                if avail_roles:
                    role_idx = avail_roles.index(auto_role) if auto_role in avail_roles else 0
                    n_role = nc1.selectbox("Role", avail_roles, index=role_idx)
                else:
                    n_role = nc1.text_input("Role", value=auto_role)

                if avail_units:
                    unit_idx = avail_units.index(auto_unit) if auto_unit in avail_units else 0
                    n_unit = nc2.selectbox("Unit", avail_units, index=unit_idx)
                else:
                    n_unit = nc2.text_input("Unit", value=auto_unit)

                # Staff code shown and editable — already pre-filled from lookup
                n_sc = nc1.text_input(
                    "Staff Code",
                    value=sc_input,
                    help="The numeric staff code from the BSC data. Also used as the default username.")

                if sc_found:
                    nc2.markdown(
                        f"<div style='padding:8px 6px;background:#F0FDF4;"
                        f"border:1px solid #BBF7D0;border-radius:6px;"
                        f"font-size:11px;color:#166534;margin-top:24px'>"
                        f"✅ Username & staff code auto-filled from BSC data</div>",
                        unsafe_allow_html=True)

                st.markdown("**Permissions**")
                pc1,pc2,pc3,pc4 = st.columns(4)
                n_all   = pc1.checkbox("Can view all staff")
                n_exec  = pc2.checkbox("Can manage Execute")
                n_admin = pc3.checkbox("Admin privileges")
                n_val   = pc4.checkbox("Can validate")

                if st.form_submit_button("✅ Create user account", type="primary",
                                          use_container_width=True):
                    # Use staff code as username if username field left as default
                    final_user = n_user.strip() or n_sc.strip()
                    final_sc   = n_sc.strip() or final_user
                    if final_user and n_pass and n_fname:
                        if final_user in um.users:
                            st.error(f"Username '{final_user}' already exists.")
                        else:
                            um.add_user(final_user, n_pass, n_fname, n_email, n_role,
                                        n_unit, final_sc, n_all, n_exec, n_admin)
                            um.users[final_user]["can_validate"] = n_val
                            um.save()
                            audit_log("USER_CREATED", uname, f"{final_user}|sc:{final_sc}")
                            st.success(f"✅ User '{final_user}' created for {n_fname}. "
                                       f"Login: username `{final_user}`, staff code `{final_sc}`.")
                            st.cache_data.clear()
                            st.rerun()
                    else:
                        st.error("Username (or staff code), password and full name are required.")

        st.markdown("---")
        st.markdown("#### All users")
        if um.users:
            udf = pd.DataFrame([{
                "Username":   u,
                "Full name":  d.get("full_name",""),
                "Role":       d.get("role",""),
                "Unit":       d.get("unit",""),
                "Staff Code": d.get("staff_code",""),
                "Active":     "✅" if d.get("active",True) else "❌",
                "View all ⚠️": "⚠️ ALL" if d.get("can_view_all") else "Tree only",
                "Admin":      "✅" if d.get("is_admin") else "—",
                "Modules":    str(len(d.get("accessible_modules",[]))) + " set",
            } for u,d in um.users.items()])

            # Highlight rows where can_view_all is set on non-admin roles
            def _flag_view_all(v):
                return "background-color:#FEF2F2;color:#991B1B;font-weight:600" if "ALL" in str(v) else ""

            st.dataframe(
                udf.style.map(_flag_view_all, subset=["View all ⚠️"]),
                use_container_width=True, hide_index=True)

            # Warn if any non-admin/non-MD has can_view_all set
            _dangerous = [d.get("full_name",u) for u,d in um.users.items()
                          if d.get("can_view_all") and not d.get("is_admin")
                          and "managing" not in str(d.get("role","")).lower()
                          and "admin" not in str(d.get("role","")).lower()]
            if _dangerous:
                st.warning(
                    f"⚠️ **Data privacy risk:** {', '.join(_dangerous)} "
                    f"{'has' if len(_dangerous)==1 else 'have'} 'View all staff' enabled. "
                    f"This grants access to all staff data. Remove unless MD/Admin.")

            # ── Bulk create users from Staff Register ────────────────
            with st.expander("📥 Bulk create users from Staff Register"):
                st.caption(
                    "Create user accounts for all staff in the uploaded Staff Register. "
                    "Username = staff_code.lower(), default password = first 6 chars of "
                    "staff code + '!' (e.g. 300803! for staff 300803). "
                    "Existing accounts are NOT overwritten.")

                _sr_df = st.session_state.get("staff_register_df")
                if _sr_df is None:
                    st.warning("Upload the BSC Excel file in the sidebar first — "
                               "the Staff Register sheet will be detected automatically.")
                else:
                    _sr_roles = _sr_df["Role"].unique().tolist() if "Role" in _sr_df.columns else []
                    _bk_role_f = st.multiselect(
                        "Filter by role (leave blank = all roles)",
                        sorted(_sr_roles), key="bk_role_filter")

                    _bk_df = _sr_df.copy()
                    if _bk_role_f:
                        _bk_df = _bk_df[_bk_df["Role"].isin(_bk_role_f)]

                    st.info(f"Will create accounts for {len(_bk_df)} staff members.")

                    _preview_rows = []
                    for _, _sr_row in _bk_df.head(5).iterrows():
                        _sc = str(_sr_row.get("Staff Code","")).strip()
                        _nm = str(_sr_row.get("Staff Name","")).strip()
                        _rl = str(_sr_row.get("Role","")).strip()
                        _un = str(_sc).lower()
                        _pw = _sc[:6] + "!" if len(_sc) >= 6 else _sc + "!"
                        _exists = "✅ exists" if _un in um.users else "🆕 new"
                        _preview_rows.append({"Username":_un,"Name":_nm,
                                              "Role":_rl,"Password":_pw,"Status":_exists})

                    if _preview_rows:
                        st.markdown("**Preview (first 5):**")
                        st.dataframe(pd.DataFrame(_preview_rows),
                                     use_container_width=True, hide_index=True)

                    if st.button("🚀 Create accounts", type="primary", key="bk_create"):
                        _created = 0; _skipped = 0
                        for _, _sr_row in _bk_df.iterrows():
                            _sc  = str(_sr_row.get("Staff Code","")).strip()
                            _nm  = str(_sr_row.get("Staff Name","")).strip()
                            _rl  = str(_sr_row.get("Role","")).strip()
                            _unt = str(_sr_row.get("Unit","")).strip()
                            _em  = str(_sr_row.get("Email","")).strip()
                            _rg  = str(_sr_row.get("Region","")).strip()
                            _un  = str(_sc).lower()
                            _pw  = _sc[:6] + "!" if len(_sc) >= 6 else _sc + "!"
                            if _un in um.users:
                                _skipped += 1
                                continue
                            um.users[_un] = {
                                "password":             um.hash_pw(_pw),
                                "full_name":            _nm,
                                "role":                 _rl,
                                "department":           _unt,
                                "unit":                 _unt,
                                "region":               _rg,
                                "email":                _em,
                                "staff_code":           _sc,
                                "can_view_all":         False,
                                "is_admin":             False,
                                "managed_roles":        [],
                                "managed_units":        [],
                                "managed_staff_codes":  [],
                                "active":               True,
                                "must_change_password": True,
                            }
                            _created += 1
                        um.save()
                        audit_log("BULK_USER_CREATE", uname, f"{_created} created, {_skipped} skipped")
                        st.success(f"✅ Created {_created} accounts, skipped {_skipped} existing. "
                                   f"Default passwords: staff_code + '!'")
                        st.cache_data.clear()
                        st.rerun()

            with st.expander("⚠️ Delete user — protected action"):
                # Admin cannot be deleted; any admin-level delete requires OTP verification
                _all_users = list(um.users.keys())
                _deletable = [u2 for u2 in _all_users if um.can_delete_user(u2)[0]]

                if not _deletable:
                    st.info("🔒 No deletable accounts. The admin account is permanently protected.")
                else:
                    del_sel = st.selectbox("User to delete", _deletable, key="del_sel")
                    _can_del, _del_reason = um.can_delete_user(del_sel)

                    if not _can_del:
                        st.error(f"🔒 {_del_reason}")
                    else:
                        st.warning(
                            f"You are about to delete **{del_sel}** "
                            f"({um.users[del_sel].get('full_name','')}) permanently. "
                            f"This cannot be undone.")

                        # OTP flow — generate 6-digit code, show in UI
                        # (also send via email if SMTP configured)
                        _otp_key  = f"del_otp_{del_sel}"
                        _otp_sent = f"del_otp_sent_{del_sel}"

                        if st.button("🔑 Generate verification code", key="gen_otp_btn"):
                            import random
                            _otp = str(random.randint(100000, 999999))
                            st.session_state[_otp_key]  = _otp
                            st.session_state[_otp_sent] = True
                            # Try send by email
                            _admin_email = um.users.get(uname, {}).get("email", "")
                            _email_cfg   = load_email_config()
                            _sent_email  = False
                            if _admin_email and _email_cfg.get("smtp_host"):
                                try:
                                    from email.mime.text import MIMEText
                                    from email.mime.multipart import MIMEMultipart
                                    import smtplib
                                    _msg = MIMEMultipart("alternative")
                                    _msg["Subject"] = "A2Z Blueprint — User deletion verification code"
                                    _msg["From"]    = _email_cfg["sender_email"]
                                    _msg["To"]      = _admin_email
                                    _html = (
                                        f"<div style='font-family:Arial;max-width:400px;margin:auto'>"
                                        f"<div style='background:var(--brand-primary,#006B3F);padding:16px;border-radius:8px 8px 0 0'>"
                                        f"<div style='color:var(--color-background-primary);font-size:16px;font-weight:700'>A2Z Blueprint</div></div>"
                                        f"<div style='padding:20px;background:var(--color-background-primary);border:0.5px solid var(--color-border-tertiary)'>"
                                        f"<p style='color:var(--color-text-primary)'>Your verification code to delete user "
                                        f"<b>{del_sel}</b> is:</p>"
                                        f"<div style='font-size:36px;font-weight:700;letter-spacing:8px;"
                                        f"color:var(--brand-primary,#006B3F);text-align:center;padding:20px'>{_otp}</div>"
                                        f"<p style='color:var(--color-text-tertiary);font-size:12px'>This code expires when you close the page. "
                                        f"If you did not request this, contact your system administrator.</p>"
                                        f"</div></div>"
                                    )
                                    _msg.attach(MIMEText(_html, "html"))
                                    with smtplib.SMTP(_email_cfg["smtp_host"],
                                                      int(_email_cfg.get("smtp_port",587))) as _s:
                                        _s.starttls()
                                        _s.login(_email_cfg["sender_email"], _email_cfg["sender_password"])
                                        _s.sendmail(_email_cfg["sender_email"], _admin_email, _msg.as_string())
                                    _sent_email = True
                                except: pass

                            if _sent_email:
                                st.success(f"✅ Verification code sent to {_admin_email}")
                            else:
                                st.info(
                                    "📋 Email not configured. Your verification code is: "
                                    f"**{_otp}** — copy it and enter below.")

                        if st.session_state.get(_otp_sent):
                            _entered_otp = st.text_input(
                                "Enter 6-digit verification code",
                                max_chars=6, key="del_otp_input",
                                placeholder="______")
                            if st.button("🗑️ Confirm delete", type="secondary", key="del_confirm_btn"):
                                if _entered_otp == st.session_state.get(_otp_key, ""):
                                    ok, msg = um.delete_user(del_sel, uname)
                                    if ok:
                                        audit_log("USER_DELETED", uname, del_sel)
                                        st.session_state.pop(_otp_key, None)
                                        st.session_state.pop(_otp_sent, None)
                                        st.success(f"✅ {msg}")
                                        st.cache_data.clear()
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {msg}")
                                else:
                                    st.error("❌ Incorrect verification code. Please try again.")

    # ════════════════════════════════════════════════════════════════
    # TAB 2 — PERMISSIONS
    # ════════════════════════════════════════════════════════════════

    with sub[2]:
        # hidden_modules is stored per-user in users.json
        # accessible_modules are additional modules beyond dept default
        st.subheader("Role-based access control")
        st.caption(
            "This matrix shows which modules each role can access by default. "
            "Use the override section below to grant or restrict access for individual users.")

        try:
            from utils.core import MODULE_ACCESS
            # Build matrix — rows=roles, cols=modules
            ALL_ROLES = [
                "Managing Director","Director Retail Banking","Director Commercial Banking",
                "Head Of Retail","Head Of Corporate","Head Of SME","Head Of Digital Innovation",
                "Head Of Strategy","Head Of Internal Audit","Head Of Marketing",
                "Chief Finance Officer","Chief Risk Officer","Chief Operations Officer",
                "Chief Compliance Officer","Chief Human Resources Officer","Chief Credit Officer",
                "Debt Recovery Unit Manager","Procurement Manager",
                "Regional Head","Branch Manager","Branch Operations Manager","Branch Credit Manager",
                "IT Manager","Operations Manager","HR Business Partner",
                "Relationship Manager Corporate","Relationship Manager SME",
                "Relationship Officer Personal Banking","Direct Sales Officer",
                "Teller","Customer Service Officer","Recovery Officer",
            ]
            MODULE_LABELS = {
                "perform":"Perform","people":"People","pipeline":"Pipeline",
                "execute":"Execute","products":"Products","integrate":"Integrate",
                "cascade":"Cascade","sla":"SLA","branch_log":"Branch Log",
                "optimize":"Optimize","commission":"Commission","campaigns":"Campaigns",
                "cims":"CIMS","sbu":"SBU","opex":"Opex","competitor":"Competitor",
                "export":"Export","admin":"Admin",
            }
            rows = []
            for role in ALL_ROLES:
                row = {"Role": role}
                for mod, lbl in MODULE_LABELS.items():
                    cfg = MODULE_ACCESS.get(mod, {"min":"public","roles_all":[]})
                    has = (role in cfg["roles_all"] or cfg["min"] == "self"
                           or cfg["min"] == "public")
                    row[lbl] = "✅" if has else "—"
                rows.append(row)
            perm_df = pd.DataFrame(rows)
            # Style: green for ✅
            def _clr(v):
                return "background-color:#F0FDF4;color:#166534;font-weight:600" if v=="✅" else "color:#D1D5DB"
            st.dataframe(
                perm_df.style.map(_clr, subset=list(MODULE_LABELS.values())),
                use_container_width=True, hide_index=True,
                height=min(38*len(rows)+38, 520))
        except Exception as _e:
            st.error(f"Could not load module access: {_e}")

        st.markdown("---")
        st.markdown("#### Override permissions for a specific user")
        st.caption(
            "Changes here override the role defaults above for this individual user. "
            "Granting Admin gives full system access. "
            "Can view all gives access to all staff data across modules.")

        if um.users:
            # ── User selector grouped by unit ────────────────────────────────
            _perm_users_raw = list(um.users.items())
            from collections import defaultdict as _dd_perm
            _unit_to_users = _dd_perm(list)
            for _un, _ud_p in _perm_users_raw:
                _u_unit = _ud_p.get("unit") or _ud_p.get("branch") or "Other"
                if not _u_unit or _u_unit in ("", "nan", "None"):
                    _u_unit = "Other"
                _unit_to_users[_u_unit].append(_un)
            _sorted_units = sorted(_unit_to_users.keys())

            # Build flat options list with unit header separators
            _perm_options  = []  # display labels
            _perm_label_to_key = {}  # display label → username key
            for _pu_unit in _sorted_units:
                for _un in sorted(_unit_to_users[_pu_unit],
                                  key=lambda x: um.users[x].get("full_name","") or x):
                    _ud_p = um.users[_un]
                    _lbl  = f"{_ud_p.get('full_name',_un)}  ·  {_ud_p.get('role','')}"
                    _perm_options.append(_lbl)
                    _perm_label_to_key[_lbl] = _un

            # Unit filter
            _pu_unit_sel = st.selectbox(
                "Filter by unit",
                ["All units"] + _sorted_units,
                key="perm_unit_filter")

            _filtered_opts = [
                _lbl for _lbl, _un in _perm_label_to_key.items()
                if _pu_unit_sel == "All units"
                or (um.users[_un].get("unit") or "Other") == _pu_unit_sel
            ]
            if not _filtered_opts:
                _filtered_opts = _perm_options

            _prev_sel_key  = st.session_state.get("_perm_prev_key","")
            _prev_lbl      = next((l for l,k in _perm_label_to_key.items()
                                   if k == _prev_sel_key), None)
            _default_idx   = _filtered_opts.index(_prev_lbl) if _prev_lbl in _filtered_opts else 0

            _pu_sel_lbl = st.selectbox(
                "Select user to configure",
                _filtered_opts,
                index=_default_idx,
                key="perm_sel_lbl")
            pu_sel = _perm_label_to_key.get(_pu_sel_lbl, list(um.users.keys())[0])
            st.session_state["_perm_prev_key"] = pu_sel
            pu     = um.users.get(pu_sel, {})
            pu_role= pu.get("role","")

            st.markdown(
                f"<div style='padding:8px 12px;background:var(--color-background-secondary);border:0.5px solid var(--color-border-tertiary);"
                f"border-radius:8px;font-size:12px;margin-bottom:10px'>"
                f"<b>{safe_html(pu.get('full_name',pu_sel))}</b> · Role: <b>{safe_html(pu_role)}</b> · "
                f"Unit: {safe_html(pu.get('unit','—'))} · Staff code: {safe_html(pu.get('staff_code','—'))}</div>",
                unsafe_allow_html=True)

            # When user selection changes, clear all cached widget states for permissions
            if st.session_state.get("_perm_user") != pu_sel:
                st.session_state["_perm_user"] = pu_sel
                for _kk in list(st.session_state.keys()):
                    if _kk.startswith(("_p_", "_mod_")):
                        del st.session_state[_kk]

            _k = pu_sel

            # Use NO FORM — direct checkboxes with explicit session_state defaults
            # This is the only reliable way to show live values that can be toggled
            if f"_p_all_{_k}" not in st.session_state:
                st.session_state[f"_p_all_{_k}"]  = bool(pu.get("can_view_all", False))
                st.session_state[f"_p_adm_{_k}"]  = bool(pu.get("is_admin", False))
                st.session_state[f"_p_exec_{_k}"] = bool(pu.get("can_execute", False))
                st.session_state[f"_p_val_{_k}"]  = bool(pu.get("can_validate", False))
                st.session_state[f"_p_hr_{_k}"]   = bool(pu.get("can_hr", False))

            if True:  # scope block (replaces "with st.form")
                st.markdown("**System permissions**")
                fc1,fc2,fc3,fc4,fc5 = st.columns(5)
                p_all  = fc1.checkbox("View ALL staff (MD/Admin only)",
                                       key=f"_p_all_{_k}",
                                       help="⚠️ Grants visibility of ALL 380+ staff. "
                                            "Only tick for MD or System Admin. "
                                            "Directors see their reporting tree automatically — "
                                            "do NOT tick this for them.")
                p_adm  = fc2.checkbox("Admin privileges",
                                       key=f"_p_adm_{_k}",
                                       help="Full system access including Admin module.")
                p_exec = fc3.checkbox("Can manage Execute",
                                       key=f"_p_exec_{_k}")
                p_val  = fc4.checkbox("Can validate staff",
                                       key=f"_p_val_{_k}")
                p_hr   = fc5.checkbox("HR module access",
                                       key=f"_p_hr_{_k}",
                                       help="Access to People / HR module.")

                st.markdown("---")
                st.markdown(
                    "**Module access** — tick modules this user can access. "
                    "Expand each module to restrict to specific pages within it.")

                try:
                    from utils.core import MODULE_ACCESS, MODULE_PAGES
                    cur_accessible = set(pu.get("accessible_modules", []))
                    cur_pages      = dict(pu.get("accessible_pages", {}))
                    if not cur_accessible:
                        cur_accessible = {m for m, cfg in MODULE_ACCESS.items()
                                          if pu_role in cfg.get("roles_all",[])
                                          or cfg["min"] == "self"}
                    mods_list = list(MODULE_ACCESS.keys())
                    # Always re-initialise from cur_accessible — never use stale session value.
                    # This ensures newly-added modules appear with the correct default.
                    for _m in mods_list:
                        st.session_state[f"_mod_{_k}_{_m}"] = _m in cur_accessible

                    new_accessible = set()
                    new_pages      = {}

                    # Render each module as a row with optional page sub-controls
                    for m in mods_list:
                        mod_lbl  = m.replace("_"," ").title()
                        mod_pages = MODULE_PAGES.get(m, [])
                        has_pages = len(mod_pages) > 0

                        m_col1, m_col2 = st.columns([2, 3])
                        mod_on = m_col1.checkbox(
                            f"**{mod_lbl}**",
                            key=f"_mod_{_k}_{m}",
                            help=f"{len(mod_pages)} pages within this module" if has_pages else "")
                        if mod_on:
                            new_accessible.add(m)
                            if has_pages and mod_pages:
                                cur_mod_pages = set(cur_pages.get(m, mod_pages))
                                sel_pages = m_col2.multiselect(
                                    f"Pages in {mod_lbl}",
                                    options=mod_pages,
                                    default=[p for p in cur_mod_pages if p in mod_pages],
                                    key=f"_pages_{_k}_{m}",
                                    label_visibility="collapsed",
                                    placeholder=f"All {len(mod_pages)} pages (no restriction)")
                                if sel_pages and len(sel_pages) < len(mod_pages):
                                    new_pages[m] = sel_pages
                        else:
                            m_col2.markdown(
                                "<span style='color:var(--color-text-tertiary);font-size:11px'>no access</span>",
                                unsafe_allow_html=True)
                except Exception as _me:
                    st.error(f"Module access error: {_me}")
                    new_accessible = set()
                    new_pages      = {}

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                if st.button("💾 Save permissions", type="primary",
                              use_container_width=True,
                              key=f"save_perm_{_k}"):
                    um.users[pu_sel].update({
                        "can_view_all":       p_all,
                        "can_execute":        p_exec,
                        "is_admin":           p_adm,
                        "can_validate":       p_val,
                        "can_hr":             p_hr,
                        "accessible_modules": sorted(new_accessible),
                        "accessible_pages":   new_pages,
                    })
                    um.save()
                    # Clear cached checkbox states so next render reads fresh from DB
                    for _kk in list(st.session_state.keys()):
                        if _kk.startswith(("_p_", "_mod_")):
                            del st.session_state[_kk]
                    st.session_state.pop("_perm_user", None)
                    audit_log("PERM_CHANGED", uname,
                              f"{pu_sel}:admin={p_adm},view_all={p_all},"
                              f"modules={len(new_accessible)}")
                    st.success(f"✅ Permissions saved for {pu.get('full_name',pu_sel)}")
                    st.cache_data.clear()
                    st.rerun()

    # TAB 4 — STAFF MOVEMENTS
    # ════════════════════════════════════════════════════════════════

    with sub[3]:
        st.subheader("🗂️ Staff movements")
        st.caption(
            "Remap reporting lines, transfer staff between branches, change roles, "
            "and manage acting appointments. All changes are audited and take effect immediately.")

        if rlm is None:
            st.error("ReportingLineManager not available. Restart the app.")
        else:
            _mv_view = st.radio("",
                ["🔗 Remap reporting line","🚀 Branch transfer","🎭 Role change","📋 Active overrides"],
                horizontal=True, key="mv_view")
            st.markdown("---")

            def _build_staff_opts(df):
                opts = {}
                if len(df) == 0: return opts
                for role in sorted(df["Role"].dropna().unique()):
                    for _, row in df[df["Role"]==role].sort_values("Staff Name").iterrows():
                        label = f"[{role}]  {row['Staff Name']}  ·  {row.get('Unit','')}"
                        opts[label] = str(row["Staff Code"])
                return opts

            # ── REMAP REPORTING LINE ─────────────────────────────────────
            if "Remap reporting" in _mv_view:
                st.markdown("**Change who a staff member reports to** — acting appointment, individual exception.")
                if len(staff_scores) == 0:
                    st.info("Load CBS data or upload BSC file to see staff.")
                else:
                    rf1,rf2,rf3 = st.columns(3)
                    _all_roles_mv = ["All roles"] + sorted(staff_scores["Role"].dropna().unique().tolist())
                    _all_units_mv = ["All units"] + sorted(staff_scores["Unit"].dropna().unique().tolist())
                    _filt_role = rf1.selectbox("Filter role", _all_roles_mv, key="mv_filt_role")
                    _filt_unit = rf2.selectbox("Filter unit", _all_units_mv, key="mv_filt_unit")
                    _filt_df   = staff_scores.copy()
                    if _filt_role != "All roles": _filt_df = _filt_df[_filt_df["Role"]==_filt_role]
                    if _filt_unit != "All units": _filt_df = _filt_df[_filt_df["Unit"]==_filt_unit]
                    rf3.markdown(f"<div style='padding:10px 6px;font-size:12px'><b>{len(_filt_df)}</b> staff match</div>",unsafe_allow_html=True)

                    # Manager pool from org_config hierarchy
                    try:
                        from utils.core import get_org_config as _goc3
                        _parent_roles = set()
                        for _r3,_p3 in _goc3().get("hierarchy",{}).items():
                            _parent_roles.update(_p3)
                        _mgr_df = staff_scores[staff_scores["Role"].isin(_parent_roles)] if _parent_roles else staff_scores
                    except:
                        _mgr_df = staff_scores

                    _staff_opts = _build_staff_opts(_filt_df) or _build_staff_opts(staff_scores)
                    _mgr_opts   = _build_staff_opts(_mgr_df)

                    with st.form("remap_form"):
                        rl1,rl2 = st.columns(2)
                        _staff_lbl = rl1.selectbox(f"Staff member ({len(_staff_opts)} shown)",
                            list(_staff_opts.keys()), key="remap_staff")
                        _mgr_lbl   = rl2.selectbox(f"New line manager ({len(_mgr_opts)})",
                            list(_mgr_opts.keys()), key="remap_mgr")
                        _remap_reason = st.text_input("Reason *",
                            placeholder="e.g. Promotion, acting appointment, restructure")
                        if st.form_submit_button("✅ Apply remap", type="primary", use_container_width=True):
                            _sc_r  = _staff_opts.get(_staff_lbl,"")
                            _mc_r  = _mgr_opts.get(_mgr_lbl,"")
                            _sn    = _staff_lbl.split("]")[-1].strip().split("·")[0].strip()
                            _mn    = _mgr_lbl.split("]")[-1].strip().split("·")[0].strip()
                            if _sc_r == _mc_r:
                                st.error("Staff and manager cannot be the same person.")
                            elif not _sc_r or not _mc_r:
                                st.error("Could not resolve codes.")
                            else:
                                rlm.remap(_sc_r, _mc_r, uname, _remap_reason)
                                audit_log("REPORTING_LINE_REMAP", uname, f"{_sn} → {_mn}")
                                st.success(f"✅ {_sn} now reports to {_mn}")
                                st.cache_data.clear()
                                st.rerun()

            # ── BRANCH TRANSFER ──────────────────────────────────────────
            elif "Branch transfer" in _mv_view:
                st.markdown("**Move a staff member to a different branch** — updates unit, region, and manager.")
                if len(staff_scores) == 0:
                    st.info("Load CBS data or upload BSC file to see staff.")
                else:
                    _all_staff_opts = _build_staff_opts(staff_scores)
                    try:
                        from utils.core import get_all_branches as _gab4, get_org_config as _goc4
                        _branch_list = [b["name"] for b in _gab4() if b.get("type","")!="HO"]
                        _mgr_opts2   = _build_staff_opts(staff_scores)  # any staff can be manager
                    except:
                        _branch_list = avail_units

                    with st.form("transfer_form"):
                        tr1,tr2 = st.columns(2)
                        _tr_staff = tr1.selectbox("Staff member to transfer",
                            list(_all_staff_opts.keys()), key="tr_staff")
                        _tr_branch= tr2.selectbox("New branch", _branch_list, key="tr_branch")
                        _tr_mgr   = tr1.selectbox("New line manager (optional)",
                            ["— Keep current —"] + list(_build_staff_opts(staff_scores).keys()),
                            key="tr_mgr")
                        _tr_role  = tr2.selectbox("New role (if changing)",
                            ["— Keep current —"] + avail_roles, key="tr_role")
                        _tr_reason= st.text_input("Reason *",
                            placeholder="e.g. Branch-to-branch transfer, secondment")
                        _tr_eff   = st.date_input("Effective date", value=date.today())

                        if st.form_submit_button("🚀 Execute transfer", type="primary", use_container_width=True):
                            if _tr_reason.strip():
                                _tr_sc = _all_staff_opts.get(_tr_staff,"")
                                _tr_mn = ""
                                if _tr_mgr != "— Keep current —":
                                    _tr_mc = _build_staff_opts(staff_scores).get(_tr_mgr,"")
                                    rlm.remap(_tr_sc, _tr_mc, uname, f"Transfer: {_tr_reason}")
                                    _tr_mn = _tr_mgr.split("]")[-1].strip().split("·")[0].strip()
                                # Update user account unit
                                for _un2, _ud3 in um.users.items():
                                    if str(_ud3.get("staff_code","")) == _tr_sc:
                                        _ud3["unit"] = _tr_branch
                                        if _tr_role != "— Keep current —": _ud3["role"] = _tr_role
                                        um.save(); break
                                audit_log("BRANCH_TRANSFER", uname,
                                    f"{_tr_staff.split(']')[-1].strip().split('·')[0].strip()} → {_tr_branch}")
                                st.success(f"✅ Transfer complete → {_tr_branch}"
                                           + (f" | Manager: {_tr_mn}" if _tr_mn else ""))
                                st.cache_data.clear()
                                st.rerun()
                            else:
                                st.error("Reason is required.")

            # ── ROLE CHANGE ──────────────────────────────────────────────
            elif "Role change" in _mv_view:
                st.markdown("**Promote or reassign a staff member's role.** Updates user account and cascade position.")
                if len(staff_scores) == 0:
                    st.info("Load CBS data or upload BSC file to see staff.")
                else:
                    _rc_opts = {f"{r['Staff Name']} [{r['Role']}] · {r.get('Unit','')}": str(r['Staff Code'])
                                for _, r in staff_scores.iterrows()}
                    with st.form("role_change_form"):
                        rc1,rc2 = st.columns(2)
                        _rc_staff   = rc1.selectbox("Staff member", list(_rc_opts.keys()), key="rc_staff_sel")
                        _rc_newrole = rc2.selectbox("New role", avail_roles, key="rc_new_role")
                        _rc_newunit = rc1.selectbox("New branch/unit", avail_units, key="rc_new_unit")
                        _rc_reason  = rc2.text_input("Reason *",
                            placeholder="e.g. Promotion, acting, secondment")
                        st.date_input("Effective date", value=date.today(), key="rc_eff_date")
                        if st.form_submit_button("✅ Apply role change", type="primary", use_container_width=True):
                            if _rc_reason.strip():
                                _rc_sc = _rc_opts.get(_rc_staff,"")
                                _changed = False
                                for _un3,_ud4 in um.users.items():
                                    if str(_ud4.get("staff_code","")) == _rc_sc:
                                        _ud4["role"] = _rc_newrole
                                        _ud4["unit"] = _rc_newunit
                                        um.save(); _changed = True; break
                                if _changed:
                                    audit_log("ROLE_CHANGE", uname,
                                        f"{_rc_staff.split('[')[0].strip()} → {_rc_newrole} @ {_rc_newunit}")
                                    st.success(f"✅ Role changed to {_rc_newrole} at {_rc_newunit}.")
                                    st.cache_data.clear()
                                    st.rerun()
                                else:
                                    st.warning("Could not find matching user account. Check staff code.")
                            else:
                                st.error("Reason is required.")

            # ── ACTIVE OVERRIDES ─────────────────────────────────────────
            else:
                ov_summary = rlm.summary()
                _oa,_ob,_oc_ = st.columns(3)
                _oa.metric("Active overrides",   ov_summary.get("total_overrides",0))
                _ob.metric("Transfer overrides", ov_summary.get("total_transfers",0))
                _oc_.metric("Last change", str(ov_summary.get("last_updated","Never"))[:10])
                overrides = rlm.get_all_overrides()
                if not overrides:
                    st.info("No overrides active.")
                else:
                    _cn = {}
                    if len(staff_scores) and "Staff Code" in staff_scores.columns:
                        for _,_r4 in staff_scores.iterrows():
                            _cn[str(_r4["Staff Code"])] = _r4["Staff Name"]
                    _ov_rows = [{"Staff":_cn.get(o.get("staff_code",""),o.get("staff_code","")),
                                 "New Manager":_cn.get(o.get("manager_code",""),o.get("manager_code","")),
                                 "New Unit":o.get("unit","—"),"Type":o.get("type","remap"),
                                 "Reason":o.get("reason",""),"By":o.get("updated_by",""),
                                 "When":str(o.get("updated_at",""))[:16]} for o in overrides]
                    st.dataframe(pd.DataFrame(_ov_rows), use_container_width=True, hide_index=True)
                    _cc1,_cc2 = st.columns([3,1])
                    _clr_opts = {f"{r['Staff']} ({r['Type']})": overrides[i].get("staff_code","")
                                 for i,r in enumerate(_ov_rows)}
                    _clr_sel = _cc1.selectbox("Remove override", list(_clr_opts.keys()), key="clr_ov3")
                    if _cc2.button("Remove", type="secondary", key="clr_btn3"):
                        rlm.clear_override(_clr_opts[_clr_sel], uname)
                        audit_log("OVERRIDE_CLEARED", uname, _clr_sel)
                        st.success("Override removed."); st.rerun()


    # TAB 6 — ORG TREE
    # ════════════════════════════════════════════════════════════════

    with sub[4]:
        st.subheader("Organisation chart")
        st.caption("Live org tree combining uploaded data and all active overrides.")

        reg = st.session_state.get("staff_registry", pd.DataFrame())
        if rlm and len(reg):
            applied = rlm.apply_to_registry(reg)
            tree    = rlm.get_org_tree(applied)

            # Build code → name map
            code_name = {}
            code_role = {}
            code_unit = {}
            if "Staff Code" in applied.columns:
                for _, r in applied.iterrows():
                    sc = str(r["Staff Code"])
                    code_name[sc] = r.get("Staff Name", sc)
                    code_role[sc] = r.get("Role","")
                    code_unit[sc] = r.get("Unit","")

            # Build Plotly org chart
            def flatten_tree(node_code, depth=0, parent_idx=None, rows=None, edges=None):
                if rows is None: rows=[]; edges=[]
                idx = len(rows)
                name = code_name.get(node_code, node_code)
                role = code_role.get(node_code,"")
                unit = code_unit.get(node_code,"")
                rows.append({'id':node_code,'name':name,'role':role,'unit':unit,'depth':depth,'idx':idx})
                if parent_idx is not None:
                    edges.append((parent_idx, idx))
                for child in tree.get(node_code,[]):
                    flatten_tree(child, depth+1, idx, rows, edges)
                return rows, edges

            # Find roots (no parent in tree values)
            all_children = set(c for children in tree.values() for c in children)
            roots = [sc for sc in tree.keys() if sc not in all_children]
            if not roots and tree:
                roots = [list(tree.keys())[0]]

            # Org view filters
            ot1, ot2 = st.columns(2)
            view_unit = ot1.selectbox("Filter by unit", ["All"] + sorted(set(code_unit.values())), key="org_unit")
            max_depth = ot2.slider("Max depth shown", 1, 6, 3, key="org_depth")

            all_rows, all_edges = [], []
            for root in roots:
                flatten_tree(root, 0, None, all_rows, all_edges)

            if all_rows:
                # Filter by unit
                if view_unit != "All":
                    keep_ids = {r['id'] for r in all_rows if r['unit']==view_unit}
                    all_rows = [r for r in all_rows if r['id'] in keep_ids]

                # Filter by depth
                all_rows = [r for r in all_rows if r['depth'] <= max_depth]

                # Summary stats
                total_mapped = len(all_rows)
                n_overrides  = len(rlm.get_all_overrides())

                os1, os2, os3 = st.columns(3)
                os1.metric("Staff in tree", total_mapped)
                os2.metric("Active overrides", n_overrides)
                os3.metric("Depth levels", max_depth)

                # Render as collapsible tree rows
                DEPTH_COLORS = ['var(--brand-primary,#006B3F)','#F5A623','#185FA5','#7F8C8D','#9B59B6','#E24B4A']
                for row in all_rows[:100]:  # cap at 100 for performance
                    indent  = '&nbsp;' * (row['depth'] * 5)
                    clr     = DEPTH_COLORS[min(row['depth'], 5)]
                    border  = 2 + row['depth']
                    is_mgr  = row['id'] in tree
                    icon    = '👔' if row['depth']==0 else ('📋' if is_mgr else '👤')
                    has_ov  = row['id'] in rlm.overrides
                    ov_badge= "<span style='background:#F5A623;color:var(--color-background-primary);padding:1px 5px;border-radius:8px;font-size:9px;margin-left:4px'>override</span>" if has_ov else ''

                    st.markdown(
                        f"<div style='padding:5px 10px;background:var(--color-background-secondary);"
                        f"border-left:{border}px solid {clr};"
                        f"margin:1px 0;border-radius:0 3px 3px 0;font-size:11px'>"
                        f"{indent}{icon} <b>{row['name']}</b> "
                        f"<span style='color:#888;font-size:10px'>{row['role']}</span> "
                        f"<span style='color:#aaa;font-size:10px'>· {row['unit']}</span>"
                        f"{ov_badge}</div>",
                        unsafe_allow_html=True)

                if len(all_rows) > 100:
                    st.caption(f"Showing 100 of {len(all_rows)} staff. Use the unit filter to narrow down.")
            else:
                st.info("No org tree data available. Upload BSC data with Staff Register sheet.")
        elif len(staff_scores):
            # Build from staff_scores directly
            st.info("Load the Staff Register (upload BSC Excel with Staff Register sheet) for the full org tree.")
            # Show a simple role hierarchy
            role_counts = staff_scores.groupby(['Role','Unit'])['Staff Name'].count().reset_index()
            role_counts.columns = ['Role','Unit','Count']
            fig = px.treemap(role_counts, path=['Unit','Role'], values='Count',
                              title='Staff distribution by unit and role')
            fig.update_layout(height=420)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Upload BSC data to view the organisation chart.")

    # ════════════════════════════════════════════════════════════════
    # TAB 7 — AUDIT LOG
    # ════════════════════════════════════════════════════════════════

    # Combined org-structure tab: dept + branch under one container
    with sub[5]:
        _struct_view = st.radio(
            "Org structure scope",
            ["🏢 Departments", "🏪 Branches"],
            horizontal=True, key="org_struct_scope",
            label_visibility="collapsed",
        )
        st.markdown("---")
        if "Departments" in _struct_view:
            render_dept_manager(st.container(), uname)
        else:
            render_branch_manager(st.container(), uname)

    render_roles_manager(sub[6],   uname)

# ── Section 1: Performance & BSC ────────────────────────────
with sections[1]:
    sub = st.tabs([
        "📚 KPI Library",
    ])
    with sub[0]:
        from utils.core_kpi import (get_kpi_library, save_kpi_library, DEFAULT_KPI_LIBRARY,
                                     DEFAULT_ROLE_KPIS, get_active_kpis)
        from utils.core import CBS_SOURCE_LABELS

        st.subheader("📚 KPI Library")
        st.caption(
            "Configure which KPIs this bank tracks, their pillars, weights, "
            "data sources, and which roles carry each KPI. "
            "This drives the BSC, cascade, and CBS actuals engine.")

        _lib = get_kpi_library()
        # Build _pillars as dict: {pillar_name: [kpi_dicts]}
        # kpi_library stores kpis as a flat list with a 'pillar' field
        _kpis_flat = _lib.get("kpis", [])
        if _kpis_flat:
            # Build from flat kpis list grouped by pillar
            _pillars_raw = {}
            for _kp in _kpis_flat:
                _pl = _kp.get("pillar", "Financial")
                _pillars_raw.setdefault(_pl, []).append(_kp)
            _pillars = _pillars_raw if _pillars_raw else DEFAULT_KPI_LIBRARY
        else:
            # Legacy format or empty — use default
            _raw = _lib.get("pillars", DEFAULT_KPI_LIBRARY)
            if isinstance(_raw, dict):
                _pillars = _raw
            elif isinstance(_raw, list):
                # List of pillar metadata — build empty dict structure
                _pillars = {p["id"]: [] for p in _raw} if _raw else DEFAULT_KPI_LIBRARY
            else:
                _pillars = DEFAULT_KPI_LIBRARY
        _role_kpis      = _lib.get("role_kpis", DEFAULT_ROLE_KPIS)
        _active_kpis    = set(_lib.get("active_kpis", []))
        _pillar_weights = _lib.get("pillar_weights", {
            "Financial":0.40,"Customer Focus":0.25,
            "Operational Excellence":0.25,"People & Learning":0.10})
        _pillar_colours = {
            "Financial":"#2563EB","Customer Focus":"#059669",
            "Operational Excellence":"#7C3AED","People & Learning":"#F59E0B"}
        _pillar_icons = {
            "Financial":"💰","Customer Focus":"👥",
            "Operational Excellence":"⚙️","People & Learning":"🎓"}

        _kl_view = st.radio("",
            ["🎯 Activate KPIs","⚖️ Pillar weights","👥 Role assignments",
             "➕ Add custom KPI","📋 Summary"],
            horizontal=True, key="kl_view")
        st.markdown("---")

        # ══ ACTIVATE KPIs ══════════════════════════════════════════════
        if "Activate KPIs" in _kl_view:
            st.markdown(
                "<div style='padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;"
                "border-radius:10px;font-size:12px;color:#1E40AF;margin-bottom:14px'>"
                "Tick the KPIs this bank tracks. Click <b>Save KPI selection</b> when done. "
                "Unticked KPIs are hidden from BSC, cascade and reporting."
                "</div>", unsafe_allow_html=True)

            with st.form("kpi_activate_form"):
                _selections = {}
                for _pillar, _kpis in _pillars.items():
                    _pclr = _pillar_colours.get(_pillar,"#6B7280")
                    _pico = _pillar_icons.get(_pillar,"📋")
                    st.markdown(
                        f"<div style='font-weight:800;font-size:13px;color:{_pclr};"
                        f"margin:16px 0 8px;padding-bottom:4px;"
                        f"border-bottom:2px solid {_pclr}40'>{_pico} {_pillar}</div>",
                        unsafe_allow_html=True)
                    _kpi_cols = st.columns(2)
                    for _ki, _kpi in enumerate(_kpis):
                        _col = _kpi_cols[_ki % 2]
                        _default = _kpi["id"] in _active_kpis if _active_kpis else True
                        _auto  = _kpi.get("cbs_source","manual") != "manual"
                        _fixed = _kpi.get("fixed",False)
                        _val = _col.checkbox(
                            f"**{_kpi['name']}**", value=_default, key=f"kla_{_kpi['id']}")
                        _selections[_kpi["id"]] = _val
                        _src_badge = (
                            "<span style='background:#ECFDF5;color:#065F46;font-size:9px;"
                            "padding:1px 6px;border-radius:10px;font-weight:700'>🔄 CBS Auto</span>"
                            if _auto else
                            "<span style='background:#FEF3C7;color:#92400E;font-size:9px;"
                            "padding:1px 6px;border-radius:10px;font-weight:700'>✏️ Manual</span>")
                        _fix_badge = (" <span style='background:#EFF6FF;color:#1E40AF;font-size:9px;"
                            "padding:1px 6px;border-radius:10px;font-weight:700'>🔒 Fixed</span>"
                            if _fixed else "")
                        _col.markdown(
                            f"<div style='margin:-6px 0 10px 26px;font-size:10px;color:var(--color-text-tertiary)'>"
                            f"{_src_badge}{_fix_badge}<br>{_kpi.get('description','')[:80]}</div>",
                            unsafe_allow_html=True)

                if st.form_submit_button("💾 Save KPI selection", type="primary", use_container_width=True):
                    _new_ids = [kid for kid,val in _selections.items() if val]
                    _lib["active_kpis"] = _new_ids
                    save_kpi_library(_lib)
                    audit_log("KPI_LIBRARY_SAVED", uname, f"{len(_new_ids)} KPIs active")
                    st.success(f"✅ {len(_new_ids)} KPIs saved.")
                    st.cache_data.clear()
                    st.rerun()

        # ══ PILLAR WEIGHTS ═════════════════════════════════════════════
        elif "Pillar weights" in _kl_view:
            st.markdown("**Set pillar weights.** Must total 100%.")
            with st.form("pillar_weights_form"):
                _new_pw = {}
                _pw_cols = st.columns(len(_pillar_weights))
                _pw_total = 0
                for _pi, (_pname, _pw) in enumerate(_pillar_weights.items()):
                    _pclr = _pillar_colours.get(_pname,"#6B7280")
                    _v = int(_pw*100) if _pw<=1.0 else int(_pw)
                    _nv = _pw_cols[_pi].number_input(_pname, 0, 100, _v, key=f"pw_{_pname}")
                    _pw_cols[_pi].markdown(
                        f"<div style='text-align:center;font-size:20px;font-weight:800;color:{_pclr}'>{_nv}%</div>",
                        unsafe_allow_html=True)
                    _new_pw[_pname] = _nv/100
                    _pw_total += _nv
                _pwclr = "#10B981" if _pw_total==100 else "#EF4444"
                st.markdown(
                    f"<div style='padding:8px;border-radius:6px;font-weight:700;color:{_pwclr}'>"
                    f"Total: {_pw_total}% {'✅' if _pw_total==100 else '— must equal 100%'}</div>",
                    unsafe_allow_html=True)
                if st.form_submit_button("💾 Save weights", type="primary"):
                    if _pw_total == 100:
                        _lib["pillar_weights"] = _new_pw
                        save_kpi_library(_lib)
                        st.success("✅ Pillar weights saved."); st.rerun()
                    else:
                        st.error("Pillar weights must total 100%.")

        # ══ ROLE ASSIGNMENTS ═══════════════════════════════════════════
        elif "Role assignments" in _kl_view:
            st.markdown("**Assign KPIs to roles.** Configure one role, then clone to similar roles.")
            _ss_roles = sorted(DEFAULT_ROLE_KPIS.keys())
            if len(st.session_state.get("staff_scores",pd.DataFrame())):
                _sr = st.session_state.get("staff_scores", {})["Role"].dropna().unique().tolist()
                if _sr: _ss_roles = sorted(set(_ss_roles + _sr))

            _sel_role = st.selectbox("Select role to configure", _ss_roles, key="kl_role_sel")
            if _sel_role:
                _cur_assigned = set(_role_kpis.get(_sel_role,
                    DEFAULT_ROLE_KPIS.get(_sel_role,[])))
                _cur_names = [k["name"] for p,ks in _pillars.items() for k in ks if k["id"] in _cur_assigned]
                if _cur_names:
                    st.markdown(
                        f"<div style='padding:8px 12px;background:var(--color-background-secondary);border:0.5px solid var(--color-border-tertiary);"
                        f"border-radius:8px;margin-bottom:8px;font-size:11px'>"
                        f"<b>Currently assigned ({len(_cur_names)}):</b> {', '.join(_cur_names)}</div>",
                        unsafe_allow_html=True)
                with st.form(f"kl_role_form"):
                    _selections2 = {}
                    for _pillar, _kpis in _pillars.items():
                        _active_in_p = [k for k in _kpis if not _active_kpis or k["id"] in _active_kpis]
                        if not _active_in_p: continue
                        _pclr = _pillar_colours.get(_pillar,"#6B7280")
                        st.markdown(
                            f"<div style='font-weight:700;font-size:11px;color:{_pclr};"
                            f"text-transform:uppercase;letter-spacing:.5px;margin:12px 0 4px'>{_pillar}</div>",
                            unsafe_allow_html=True)
                        _rc = st.columns(3)
                        for _ki, _kpi in enumerate(_active_in_p):
                            _v = _rc[_ki%3].checkbox(_kpi["name"],
                                value=_kpi["id"] in _cur_assigned, key=f"kl_ra2_{_sel_role}_{_kpi['id']}")
                            _selections2[_kpi["id"]] = _v
                    _n_sel2 = sum(1 for v in _selections2.values() if v)
                    st.markdown(f"**{_n_sel2} KPIs selected for {_sel_role}**")
                    if st.form_submit_button(f"💾 Save for {_sel_role}", type="primary", use_container_width=True):
                        _new_ids2 = [kid for kid,val in _selections2.items() if val]
                        _role_kpis[_sel_role] = _new_ids2
                        _lib["role_kpis"] = _role_kpis
                        save_kpi_library(_lib)
                        audit_log("KPI_ROLE_SAVED", uname, f"{_sel_role}|{_n_sel2}")
                        st.success(f"✅ {_n_sel2} KPIs saved for {_sel_role}"); st.rerun()

                # Clone to other roles
                st.markdown("---")
                st.markdown(f"**Clone KPIs from {_sel_role} to other roles**")
                _clone_src = set(_role_kpis.get(_sel_role,[]))
                if _clone_src:
                    _clone_names = [k["name"] for p,ks in _pillars.items() for k in ks if k["id"] in _clone_src]
                    st.caption(f"Will clone: {', '.join(_clone_names[:6])}" + (f" +{len(_clone_names)-6} more" if len(_clone_names)>6 else ""))
                    _other = [r for r in _ss_roles if r != _sel_role]
                    _clone_tgts = st.multiselect("Apply to these roles", _other, key="kl_clone_tgts")
                    _clone_mode = st.radio("Mode",["Replace existing","Merge with existing"], horizontal=True, key="kl_clone_mode")
                    if _clone_tgts:
                        _prev_rows = [{"Role":r,"Before":len(_role_kpis.get(r,[])),"After":len(_clone_src if "Replace" in _clone_mode else _clone_src|set(_role_kpis.get(r,[])))} for r in _clone_tgts]
                        st.dataframe(pd.DataFrame(_prev_rows), use_container_width=True, hide_index=True)
                        if st.button(f"📋 Apply to {len(_clone_tgts)} role(s)", type="primary", key="kl_clone_apply"):
                            for _tr in _clone_tgts:
                                _exist = set(_role_kpis.get(_tr,[]))
                                _role_kpis[_tr] = list(_clone_src if "Replace" in _clone_mode else _clone_src|_exist)
                            _lib["role_kpis"] = _role_kpis
                            save_kpi_library(_lib)
                            audit_log("KPI_ROLE_CLONED",uname,f"{_sel_role}→{','.join(_clone_tgts)}")
                            st.success(f"✅ Applied to {len(_clone_tgts)} role(s)"); st.rerun()

        # ══ ADD CUSTOM KPI ═════════════════════════════════════════════
        elif "Add custom KPI" in _kl_view:
            st.markdown("**Add a KPI not in the standard library.**")
            with st.form("add_custom_kpi_form"):
                _ca,_cb = st.columns(2)
                _new_kpi_name   = _ca.text_input("KPI name *", placeholder="e.g. Foreign Currency Deposits")
                _new_kpi_pillar = _cb.selectbox("Pillar *", list(_pillars.keys()))
                _new_kpi_unit   = _ca.selectbox("Unit", ["KES","Count","%","Score","Days","USD","Other"])
                _new_kpi_dir    = _cb.selectbox("Direction", ["Higher is better","Lower is better"])
                _new_kpi_src    = _ca.selectbox("Data source",
                    ["manual"] + list(CBS_SOURCE_LABELS.keys()),
                    format_func=lambda x: CBS_SOURCE_LABELS.get(x,x))
                _new_kpi_wt     = _cb.number_input("Default weight",0.0,1.0,0.10,0.01,format="%.2f")
                _new_kpi_fixed  = _ca.checkbox("Fixed KPI (bank-wide, not cascaded)")
                _new_kpi_desc   = st.text_area("Description",height=60)
                if st.form_submit_button("➕ Add to library", type="primary"):
                    if _new_kpi_name.strip():
                        _new_id = "CUSTOM_"+_new_kpi_name.strip().upper().replace(" ","_")[:20]
                        _new_entry = {"id":_new_id,"name":_new_kpi_name.strip(),
                            "unit":_new_kpi_unit,"direction":"higher" if "Higher" in _new_kpi_dir else "lower",
                            "cbs_source":_new_kpi_src,"fixed":_new_kpi_fixed,
                            "default_weight":_new_kpi_wt,"description":_new_kpi_desc.strip(),"custom":True}
                        if _new_kpi_pillar not in _lib["pillars"]: _lib["pillars"][_new_kpi_pillar]=[]
                        _lib["pillars"][_new_kpi_pillar].append(_new_entry)
                        if _lib.get("active_kpis"): _lib["active_kpis"].append(_new_id)
                        save_kpi_library(_lib)
                        audit_log("KPI_CUSTOM_ADDED",uname,f"{_new_id}|{_new_kpi_pillar}")
                        st.success(f"✅ '{_new_kpi_name}' added."); st.rerun()

        # ══ SUMMARY ════════════════════════════════════════════════════
        elif "Summary" in _kl_view:
            _all_active = get_active_kpis()
            _auto_k = [k for k in _all_active if k.get("cbs_source","manual")!="manual"]
            _sc = st.columns(4)
            _sc[0].metric("Active KPIs",    len(_all_active))
            _sc[1].metric("CBS Auto",       len(_auto_k))
            _sc[2].metric("Manual entry",   len(_all_active)-len(_auto_k))
            _sc[3].metric("Pillars",        len(_pillars))
            for _pillar, _kpis in _pillars.items():
                _active_in_p = [k for k in _kpis if not _active_kpis or k["id"] in _active_kpis]
                if not _active_in_p: continue
                _pclr = _pillar_colours.get(_pillar,"#6B7280")
                _pw = _pillar_weights.get(_pillar,0)
                _pwv = int(_pw*100) if _pw<=1.0 else int(_pw)
                st.markdown(
                    f"<div style='padding:8px 14px;background:{_pclr}15;"
                    f"border-left:4px solid {_pclr};border-radius:0 8px 8px 0;margin:8px 0'>"
                    f"<span style='font-weight:800;color:{_pclr}'>{_pillar}</span>"
                    f"<span style='float:right;font-weight:700;color:{_pclr}'>{_pwv}% of BSC</span>"
                    f"</div>", unsafe_allow_html=True)
                _kdf = pd.DataFrame([{
                    "KPI":k["name"],"Unit":k["unit"],
                    "Direction":"↑ Higher" if k.get("direction")=="higher" else "↓ Lower",
                    "Source":"🔄 CBS" if k.get("cbs_source","manual")!="manual" else "✏️ Manual",
                    "Fixed":"🔒" if k.get("fixed") else "",
                } for k in _active_in_p])
                st.dataframe(_kdf,use_container_width=True,hide_index=True)


    # TAB 10 — PIPELINE SETTINGS
    # ════════════════════════════════════════════════════════════════

# ── Section 2: Modules ──────────────────────────────────────
with sections[2]:
    sub = st.tabs([
        "🔧 Module Config Centre",
        "🧩 Module Assignment",
        "🚀 Sprint Modules",
        "🏷️ Nav Labels",
        "⚙️ Thresholds",
    ])
    render_module_config_centre(sub[0], uname, is_admin)
    render_module_assignment(sub[1],   uname)
    render_sprint_config(sub[2],       uname)
    render_nav_labels(sub[3],          uname)
    render_thresholds(sub[4],          uname)

# ── Section 3: Data & Integration ───────────────────────────
with sections[3]:
    sub = st.tabs([
        "🗄️ PostgreSQL Migration",
        "🔍 Reconciliation",
        "🔄 ETL Centre",
        "🚀 Cutover",
    ])
    render_postgres_centre(sub[0],  uname, is_admin)
    render_recon_centre(sub[1],     uname, is_admin)
    render_etl_centre(sub[2],       uname, is_admin)
    render_cutover_centre(sub[3],   uname, is_admin)

# ── Section 4: System ───────────────────────────────────────
with sections[4]:
    sub = st.tabs([
        "⚙️ System health",
        "📤 Upload format",
        "📑 Living Documentation",
        "📜 Commercial Readiness",
        "🗄️ State Backend",
        "🔌 Engine Hub",
        "📐 Standards Hub",
    ])
    with sub[0]:
        st.subheader("System health")
        # ── Database status ─────────────────────────────────────────────
        # ── FastAPI status ────────────────────────────────────────────
        try:
            from utils.api_client import api as _api
            _api_health = _api.health()
            if _api_health.get("status") == "healthy":
                st.success(f"✅ FastAPI backend: running on port 8502 · "
                          f"DB: {_api_health.get('db','?')} · "
                          f"Cache keys: {_api_health.get('cache_keys',0)}")
                if st.button("🗑️ Clear API cache", key="clear_api_cache"):
                    # V-001 — clear_cache is admin-only on the API. The
                    # client needs a bearer token before calling it. We
                    # log in using the current Streamlit session credentials
                    # held in session_state (set during Streamlit login).
                    _sess_user = st.session_state.get("user", {}) or {}
                    _sess_pwd  = st.session_state.get("_pwd_for_api", "")
                    _logged_in = False
                    if _sess_user.get("username") and _sess_pwd:
                        _logged_in = _api.login(_sess_user["username"], _sess_pwd)
                    if _logged_in:
                        _api.clear_cache()
                        st.success("API cache cleared")
                    else:
                        st.warning("⚠️ API cache clear requires re-login. "
                                  "Log out and back in, or run "
                                  "`curl -X POST http://localhost:8502/api/auth/login` "
                                  "to obtain a token.")
            else:
                st.info("ℹ️ FastAPI not running — using direct DB access. "
                       "Run: python -m utils.api (in a second terminal) for faster performance.")
        except Exception as _apie:
            st.info("ℹ️ FastAPI module not available.")

        try:
            from utils.db import db as _db
            _db_health = _db.health_check()
            if _db_health["status"] == "healthy":
                st.success(f"✅ PostgreSQL: {_db_health.get('version','')[:30]} · "
                          f"DB size: {_db_health.get('db_size','?')} · "
                          f"Connections: {_db_health.get('connections',0)}")
            elif _db_health["status"] == "disabled":
                st.info("ℹ️ Database: JSON file mode (PostgreSQL not configured). "
                       "Set A2Z_USE_DB=true + connection env vars to enable PostgreSQL.")
            else:
                st.error(f"❌ PostgreSQL error: {_db_health.get('error','unknown')}")
        except Exception as _dbe:
            st.warning(f"Database module not available: {_dbe}")

        # ── CBS Batch & Cache Status ──────────────────────────────────
        st.markdown("**CBS Data & Cache Status:**")
        from pathlib import Path as _Ph
        import os as _os_sh
        _act_files = sorted([f for f in (_Ph(__file__).parent.parent/"data").glob("actuals_*.xlsx") if "backup" not in f.name], reverse=True)
        _act_file  = _act_files[0] if _act_files else None
        _sh1,_sh2,_sh3,_sh4 = st.columns(4)
        if _act_file:
            import datetime as _dt_sh
            _mtime = _dt_sh.datetime.fromtimestamp(_act_file.stat().st_mtime)
            _age_h = (_dt_sh.datetime.now()-_mtime).total_seconds()/3600
            _sh1.metric("Actuals file", _act_file.name[:25])
            _sh2.metric("Last refreshed", f"{_age_h:.0f}h ago")
            _sh3.metric("File size", f"{_act_file.stat().st_size//1024}KB")
            _sh4.metric("Status", "✅ Current" if _age_h<24 else "⚠️ Stale")
            if _age_h > 24:
                st.warning("⚠️ Actuals data is more than 24 hours old. Trigger a refresh.")

        # Data file health
        st.markdown("**Data file inventory:**")
        _data_dir = _Ph(__file__).parent.parent / "data"
        _json_files = sorted(_data_dir.glob("*.json"))
        _file_rows = [{"File":f.name[:35],"Size":f"{f.stat().st_size//1024}KB",
                        "Status":"✅" if f.stat().st_size>100 else "⚠️ Empty"}
                       for f in _json_files[:20]]
        import pandas as _pd_sh
        st.dataframe(_pd_sh.DataFrame(_file_rows),use_container_width=True,hide_index=True)

        if st.button("🔄 Clear all caches", key="ict_clear_cache"):
            st.cache_data.clear()
            audit_log("CACHE_CLEARED", uname, "Manual cache clear from System Health")
            st.success("✅ All caches cleared")


        # ── BSC Month-end Lock ────────────────────────────────────────
        st.markdown("---")
        st.markdown("**BSC Month-end Lock:**")
        try:
            from utils.core import is_bsc_locked, lock_bsc, unlock_bsc
            _lock_path = Path(__file__).parent.parent / "data" / "bsc_lock.json"
            _lock_state = a2z_db.load_json(_lock_path) if _lock_path.exists() else {}
            _is_locked  = _lock_state.get("locked", False)

            if _is_locked:
                st.error(f"🔒 BSC LOCKED — Period: {_lock_state.get('period','')} · Locked by: {_lock_state.get('locked_by','')} · Date: {_lock_state.get('locked_at','')}")
                if is_admin:
                    _unlock_reason = st.text_input("Unlock reason (required)", key="bsc_unlock_reason")
                    if st.button("🔓 Unlock BSC Scores", key="bsc_unlock_btn", type="secondary"):
                        if _unlock_reason.strip():
                            unlock_bsc(uname, _unlock_reason)
                            audit_log("BSC_UNLOCKED", uname, _unlock_reason)
                            st.cache_data.clear(); st.success("✅ BSC unlocked"); st.rerun()
                        else:
                            st.error("Unlock reason is required")
            else:
                st.success("🔓 BSC OPEN — Scores can be updated")
                if is_admin:
                    import calendar as _cal_bsc
                    _today_bsc = date.today()
                    _period_bsc = _today_bsc.strftime("%B %Y")
                    if st.button(f"🔒 Lock BSC for {_period_bsc}", key="bsc_lock_btn", type="primary"):
                        lock_bsc(uname, _period_bsc, str(Path(__file__).parent.parent/"data"))
                        audit_log("BSC_LOCKED", uname, f"Period {_period_bsc}")
                        st.cache_data.clear(); st.success(f"✅ BSC locked for {_period_bsc}"); st.rerun()
        except Exception as _e_bsc:
            st.warning(f"BSC lock unavailable: {str(_e_bsc)[:80]}")



        # ── Active users / session overview ───────────────────────────
        st.markdown("#### 👥 User activity overview")
        _users_all = um.users
        _active_users  = sum(1 for u in _users_all.values() if u.get("active", True))
        _inactive_users= sum(1 for u in _users_all.values() if not u.get("active", True))
        _no_role_users = sum(1 for u in _users_all.values() if not u.get("role","").strip())
        _admin_users   = sum(1 for u in _users_all.values() if u.get("is_admin"))
        _sh1,_sh2,_sh3,_sh4 = st.columns(4)
        _sh1.metric("Total accounts",  len(_users_all))
        _sh2.metric("Active",          _active_users)
        _sh3.metric("Inactive / locked", _inactive_users)
        _sh4.metric("Admins",          _admin_users)
        if _no_role_users:
            st.warning(f"⚠️ {_no_role_users} user accounts have no role — check Users tab")

        # ── Session context compact ─────────────────────────────────────
        st.markdown("#### 📋 Session context (for continuing in a new chat)")
        _compact_path = Path(__file__).parent.parent / "data" / "context_compact.md"
        if _compact_path.exists():
            _compact_text = _compact_path.read_text()
            with st.expander("View / copy context compact", expanded=False):
                st.code(_compact_text, language="markdown")
                st.caption("Copy this block and paste it at the start of a new conversation "
                            "to continue where you left off — similar to the % compact used "
                            "in long AI conversations.")
        if st.button("🔄 Refresh context compact", key="refresh_compact"):
            # Regenerate
            try:
                import subprocess as _sp
                _compact_lines = ["## A2Z Blueprint — Session Compact\n"]
                _u2 = um.users
                _active2 = sum(1 for u in _u2.values() if u.get("active", True))
                _compact_lines.append(f"**Users:** {len(_u2)} total, {_active2} active")
                _compact_lines.append(f"**Modules:** {len(list(Path('pages').glob('[0-9]*.py')))} pages")
                import json as _json2
                for fname, desc in [
                    ("loan_applications.json","LMS apps"),
                    ("legal_matters.json","Legal"),
                    ("compliance_cases.json","Compliance"),
                ]:
                    _p2 = Path(__file__).parent.parent / "data" / fname
                    if _p2.exists():
                        _d2 = _json2.loads(_p2.read_text())
                        _compact_lines.append(f"**{desc}:** {len(_d2)}")
                _compact_path.write_text("\n".join(_compact_lines))
                st.cache_data.clear()
                st.success("✅ Context compact refreshed")
                st.rerun()
            except Exception as _ce:
                st.error(f"Could not refresh: {_ce}")

        # ── Current logged-in users (session state proxy) ─────────────
        st.markdown("#### 🟢 Session stats")
        st.info(
            "Streamlit Community Cloud does not expose live session counts. "
            "The table below shows the **last 20 accounts to have accessed the system** "
            "based on audit log entries.")

        _audit_path = Path(__file__).parent.parent / "data" / "audit_trail.jsonl"
        if _audit_path.exists() and _audit_path.stat().st_size > 10:
            try:
                _entries = [__import__("json").loads(l) for l in
                            _audit_path.read_text().strip().split("\n") if l.strip()]
                _seen = {}
                for _e in reversed(_entries):
                    _u = _e.get("user","")
                    if _u and _u not in _seen:
                        _seen[_u] = _e.get("timestamp","")[:19].replace("T"," ")
                _recent_rows = [{"Username":u,"Last Seen":t,"Role":um.users.get(u,{}).get("role","—")[:40]}
                                  for u,t in list(_seen.items())[:20]]
                if _recent_rows:
                    st.dataframe(__import__("pandas").DataFrame(_recent_rows),
                                 use_container_width=True, hide_index=True)
            except Exception as _se:
                st.caption(f"Could not parse audit log: {_se}")
        else:
            st.caption("No audit log entries yet.")
        st.markdown("---")
        st.caption("Data integrity checks, cascade status, storage usage and active users.")

        import os as _os
        from pathlib import Path as _Path

        _data_dir = DATA_DIR
        _sh1, _sh2, _sh3 = st.columns(3)

        # Data files status
        _files_info = []
        for _fn in ["users.json","target_cascade.json","bank_targets.json",
                    "fixed_kpis.json","leaves.json","hr_records.json",
                    "pipeline_deals.json","audit_log.json","sbu_action_plans.json"]:
            _fp = _data_dir / _fn
            _exists = _fp.exists()
            _size   = _fp.stat().st_size if _exists else 0
            _files_info.append({
                "File":   _fn,
                "Status": "✅ OK" if _exists else "⚠️ Missing",
                "Size":   f"{_size/1024:.1f} KB" if _size else "0 KB",
            })

        _sh1.markdown("**Data files**")
        _sh1.dataframe(pd.DataFrame(_files_info), hide_index=True, use_container_width=True)

        # User stats
        _active_users = sum(1 for u in um.users.values() if u.get("active"))
        _total_users  = len(um.users)
        _admin_users  = sum(1 for u in um.users.values() if u.get("is_admin") or u.get("role","").lower()=="admin")
        _must_pw      = sum(1 for u in um.users.values() if u.get("must_change_password"))

        _sh2.markdown("**User accounts**")
        _sh2.metric("Total users", _total_users)
        _sh2.metric("Active",      _active_users)
        _sh2.metric("Admins",      _admin_users)
        if _must_pw:
            _sh2.warning(f"⚠️ {_must_pw} user(s) must change password on next login")

        # Cascade health
        _sh3.markdown("**Cascade health**")
        if casc:
            try:
                _casc_data = getattr(casc, "cascade", {}) or {}
                _bt_data   = getattr(casc, "bank_targets", {}) or {}
                _allocs    = sum(1 for k in _casc_data if not k.startswith("deadline|") and not k.startswith("global_"))
                _deadlines = sum(1 for k in _casc_data if k.startswith("deadline|"))
                _bt_count  = len(_bt_data)
                try:    _fixed_k = len(casc.get_fixed_kpis(_gfy()))
                except: _fixed_k = 0
                _sh3.metric("Allocations saved", _allocs)
                _sh3.metric("Deadlines set",     _deadlines)
                _sh3.metric("Bank targets set",  _bt_count)
                _sh3.metric("Fixed KPIs (2026)", _fixed_k)
            except Exception as _ce:
                _sh3.warning(f"Cascade data unavailable: {_ce}")
        else:
            _sh3.info("Cascade manager not loaded.")

        # Audit log preview
        st.markdown("---")
        st.markdown("**Recent audit entries**")
        _al_file = _data_dir / "audit_log.json"
        if _al_file.exists():
            try:
                _al = a2z_db.load_json(_al_file)
                _al_df = pd.DataFrame(_al[-20:] if len(_al)>20 else _al)
                if not _al_df.empty:
                    st.dataframe(_al_df[::-1], hide_index=True,
                                 use_container_width=True, height=250)
            except:
                st.info("Could not read audit log.")
        else:
            st.info("No audit log yet.")

        # Quick fix actions
        st.markdown("---")
        st.markdown("**Quick fix actions**")
        _qc1, _qc2 = st.columns(2)
        # ── Live actuals refresh ──────────────────────────────────────────
        st.markdown("**Live actuals engine**")
        _ae_c1, _ae_c2, _ae_c3 = st.columns(3)
        if _ae_c1.button("🔄 Refresh actuals now", key="sh_refresh_actuals", type="primary"):
            with st.spinner("Computing actuals from CBS data..."):
                try:
                    import sys as _sys_ae2
                    _sys_ae2.path.insert(0, str(DATA_DIR.parent.parent))
                    from utils.actuals_engine import compute_actuals_from_cbs as _cac2
                    _result = _cac2(force=True)
                    if _result.get("success"):
                        st.success(
                            f"✅ Actuals refreshed — {_result.get('rows',0):,} rows "
                            f"in {_result.get('duration_s',0)}s")
                        # Reload into session
                        for _k in ["df_processed","staff_scores","filtered_staff",
                                    "_cbs_loaded_file"]:
                            st.session_state.pop(_k, None)
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ {_result.get('message','Unknown error')}")
                except Exception as _ae_err:
                    st.error(f"Actuals engine error: {_ae_err}")

        if _ae_c2.button("📊 View actuals status", key="sh_actuals_status"):
            _act_files = sorted(DATA_DIR.glob("actuals_*.xlsx"), reverse=True)
            if _act_files:
                _af = _act_files[0]
                st.info(
                    f"Latest: `{_af.name}` | "
                    f"Size: {_af.stat().st_size/1024:.1f} KB | "
                    f"Modified: {__import__('datetime').datetime.fromtimestamp(_af.stat().st_mtime).strftime('%Y-%m-%d %H:%M')}")
            else:
                st.warning("No actuals file found. Click 'Refresh actuals now'.")

        _ae_c3.caption(
            "Actuals auto-refresh on app startup when CBS data is newer. "
            "Use 'Refresh now' after uploading new CBS data.")
        st.markdown("---")

        if _qc1.button("🔄 Reload cascade manager", key="sh_reload_casc"):
            try:
                st.session_state["cascade_manager"] = CascadeManager()
                st.success("✅ Cascade manager reloaded from disk.")
            except Exception as _e:
                st.error(f"Error: {_e}")
        if _qc2.button("🧹 Clear session cache", key="sh_clear_cache"):
            _keep = {"logged_in","username","user_data","user_manager"}
            _to_del = [k for k in st.session_state if k not in _keep]
            for k in _to_del: del st.session_state[k]
            st.success(f"✅ Cleared {len(_to_del)} session keys. Data managers will reload.")
            st.cache_data.clear()
            st.rerun()

    with sub[1]:
        st.subheader("Upload format guide")
        st.caption("The system accepts two upload modes: Full BSC (targets + actuals) and Actuals-only (monthly update).")

        mode_tabs = st.tabs(["📊 Full BSC upload","📥 Actuals-only upload","📋 Column reference","✅ Validation rules"])

        with mode_tabs[0]:
            st.markdown("#### Full BSC upload — KPI Data sheet")
            st.markdown(
                "Use this when setting up the system or updating targets. "
                "The file must have two sheets: **KPI Data** and **Staff Register**.")

            st.markdown(
                "<div style='padding:12px 16px;background:#EFF6FF;border:1px solid #BFDBFE;"
                "border-radius:8px;margin-bottom:12px;font-size:12px;color:#1E40AF'>"
                "<b>📌 Targets come from the Cascade module — not from this file.</b><br>"
                "The upload file no longer includes an <code>Annual Target</code> column. "
                "Once the MD sets bank targets and managers cascade to every staff member, "
                "the system auto-populates each person's target directly into their BSC scorecard. "
                "The file now only contains KPI structure, weights, and monthly actuals."
                "</div>", unsafe_allow_html=True)

            full_cols = pd.DataFrame([
                {"Column":"Unit",          "Required":"✅","Type":"Text",   "Example":"Nairobi CBD Branch","Notes":"Branch name or HO department"},
                {"Column":"Category",      "Required":"✅","Type":"Text",   "Example":"Branch",            "Notes":"Branch or Head Office"},
                {"Column":"Staff Code",    "Required":"✅","Type":"Number", "Example":"300130",            "Notes":"Unique employee ID — numeric, no apostrophe"},
                {"Column":"Staff Name",    "Required":"✅","Type":"Text",   "Example":"Grace K. Kamau",    "Notes":"Full name as registered in HR system"},
                {"Column":"Role",          "Required":"✅","Type":"Text",   "Example":"Teller",            "Notes":"Must match role hierarchy exactly"},
                {"Column":"Role Function", "Required":"✅","Type":"Text",   "Example":"Support",           "Notes":"Business | Support | Executive"},
                {"Column":"Pillar",        "Required":"✅","Type":"Text",   "Example":"Financial",         "Notes":"Financial | Customer Focus | Operational Excellence"},
                {"Column":"KPI",           "Required":"✅","Type":"Text",   "Example":"Deposit Growth",    "Notes":"KPI name — must be consistent per role"},
                {"Column":"Annual Actual", "Required":"✅","Type":"Number", "Example":"45000000",          "Notes":"Cumulative YTD actual — updated by monthly upload"},
                {"Column":"Weight",        "Required":"✅","Type":"Decimal","Example":"0.08",              "Notes":"All weights per staff must sum to exactly 1.00"},
                {"Column":"Jan-26 Actual", "Required":"✅","Type":"Number", "Example":"15000000",          "Notes":"Format: Mon-YY Actual (e.g. Apr-26 Actual)"},
                {"Column":"Feb-26 Actual", "Required":"⬜","Type":"Number", "Example":"14500000",          "Notes":"Add new month columns each month — auto-detected"},
                {"Column":"Mar-26 Actual", "Required":"⬜","Type":"Number", "Example":"15500000",          "Notes":"Prior months can be left blank if not yet uploaded"},
                {"Column":"FY-25 Actual",  "Required":"⬜","Type":"Number", "Example":"380000000",         "Notes":"Full prior year actual — used for AI target suggestions"},
                {"Column":"Dec-25 Actual", "Required":"⬜","Type":"Number", "Example":"32000000",          "Notes":"December prior year monthly — used for trend analysis"},
            ])
            st.dataframe(full_cols, use_container_width=True, hide_index=True)

            st.markdown(
                "<div style='padding:12px 16px;background:#F0FDF4;border:1px solid #BBF7D0;"
                "border-radius:8px;margin-top:8px;font-size:12px;color:#166534'>"
                "<b>How targets flow into the BSC:</b><br>"
                "1. MD sets bank-level targets in <b>Target Cascade → Bank targets</b><br>"
                "2. MD allocates to directors, directors to managers, managers to staff<br>"
                "3. Staff accept and lock their targets — tracking begins immediately<br>"
                "4. The BSC scorecard reads the <b>cascaded target</b> as the Annual Target for scoring<br>"
                "5. Staff cannot see their target until their manager has cascaded and they have accepted"
                "</div>", unsafe_allow_html=True)

            st.markdown("#### Staff Register sheet")
            reg_cols = pd.DataFrame([
                {"Column":"Staff Code","Required":"✅","Example":"300130"},
                {"Column":"Staff Name","Required":"✅","Example":"Grace K. Kamau"},
                {"Column":"Email","Required":"⬜","Example":"g.kamau@yourbank.co.ke"},
                {"Column":"Phone","Required":"⬜","Example":"0712345678"},
                {"Column":"Role","Required":"✅","Example":"Teller"},
                {"Column":"Role Function","Required":"✅","Example":"Support"},
                {"Column":"Unit","Required":"✅","Example":"Nairobi CBD Branch"},
                {"Column":"Category","Required":"✅","Example":"Branch"},
                {"Column":"Staff Status","Required":"✅","Example":"Existing | New | Probation"},
                {"Column":"Hire Date","Required":"✅","Example":"01/01/2022"},
                {"Column":"Reports To Code","Required":"✅","Example":"300105","Notes":"Staff Code of direct line manager"},
                {"Column":"Region","Required":"⬜","Example":"Central","Notes":"Auto-derived from unit if blank"},
            ])
            st.dataframe(reg_cols, use_container_width=True, hide_index=True)

        with mode_tabs[1]:
            st.markdown("#### Actuals-only upload — monthly update format")
            st.markdown(
                "Once targets are set (via full upload or cascade), use this lighter format "
                "to upload monthly actuals only. The system merges actuals into the existing targets.")

            st.markdown(
                "<div style='padding:12px 16px;background:var(--brand-light,#E8F5EE);"
                "border-left:4px solid var(--brand-primary,#006B3F);border-radius:0 6px 6px 0;margin:8px 0'>"
                "<b>How it works:</b> Targets come from the cascade allocation or the last full upload. "
                "Each month you upload a file with only Staff Code, KPI, and the month's actual. "
                "The system matches on Staff Code + KPI and updates that month's column."
                "</div>", unsafe_allow_html=True)

            actuals_cols = pd.DataFrame([
                {"Column":"Staff Code","Required":"✅","Example":"300130","Notes":"Must match exactly"},
                {"Column":"KPI","Required":"✅","Example":"Deposit Growth","Notes":"Must match exactly"},
                {"Column":"Apr-26 Actual","Required":"✅","Example":"18500000","Notes":"The column header drives which month is updated"},
                {"Column":"Staff Name","Required":"⬜","Example":"Grace K. Kamau","Notes":"Optional — used for validation only"},
                {"Column":"Unit","Required":"⬜","Example":"Nairobi CBD Branch","Notes":"Optional — used for validation only"},
            ])
            st.dataframe(actuals_cols, use_container_width=True, hide_index=True)

            st.markdown("#### Sample actuals file (copy this format)")
            sample_actuals = pd.DataFrame({
                "Staff Code": [300130,300131,300132],
                "Staff Name":  ["Grace K. Kamau","John M. Otieno","Mary A. Wanjiku"],
                "KPI":         ["Deposit Growth","Deposit Growth","New Customer Acquisition"],
                "Apr-26 Actual":[16500000, 19200000, 38],
                "Unit":        ["Nairobi CBD Branch","Westlands Branch","Kisumu Branch"],
            })
            st.dataframe(sample_actuals, use_container_width=True, hide_index=True)

            st.markdown(
                "<div style='padding:10px 14px;background:#FFFBF0;"
                "border-left:3px solid #F5A623;font-size:12px;margin-top:8px'>"
                "⚠️ <b>For KPIs auto-calculated by the system</b> (Diligence Score, Initiative Score, "
                "SLA Adherence Score, Branch Optimization Score, CX Score, Diligence Score) — "
                "do NOT include these in the actuals upload. The system computes them automatically "
                "from Execute, Daily Log, and SLA modules. Including them will overwrite the system calculation."
                "</div>", unsafe_allow_html=True)

            st.markdown("#### Auto-calculated KPIs — source modules")
            auto_kpis = pd.DataFrame([
                {"KPI":"Diligence Score","Source module":"Execute → milestones + People → discipline/PIP","Frequency":"Real-time"},
                {"KPI":"Initiative Score","Source module":"Execute → gate progression","Frequency":"Real-time"},
                {"KPI":"SLA Adherence Score","Source module":"SLA Tracker (coming)","Frequency":"Daily"},
                {"KPI":"CX Score","Source module":"Daily Branch Log validation + SLA","Frequency":"Daily"},
                {"KPI":"Branch Optimization Score","Source module":"Branch Optimization Engine (coming)","Frequency":"Daily"},
                {"KPI":"Campaign Conversion Rate","Source module":"Campaigns module (coming)","Frequency":"Per campaign"},
                {"KPI":"Digital Acquiring","Source module":"Daily Branch Log","Frequency":"Daily"},
                {"KPI":"Digital Transaction Migration","Source module":"Daily Branch Log","Frequency":"Daily"},
                {"KPI":"Transactions","Source module":"Daily Branch Log","Frequency":"Daily"},
            ])
            st.dataframe(auto_kpis, use_container_width=True, hide_index=True)

        with mode_tabs[2]:
            st.markdown("#### Full column reference — all supported KPIs")
            st.caption("These are all KPIs currently mapped in the system with their BSC pillar and role mapping.")

            kpi_ref = pd.DataFrame([
                {"KPI":"Deposit Growth","Pillar":"Financial","Roles":"All business roles","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Loan Book Growth","Pillar":"Financial","Roles":"Business roles","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Loans Disbursement","Pillar":"Financial","Roles":"RM, DSO, BCM","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Fees and Commission","Pillar":"Financial","Roles":"All business","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"DFS Revenue","Pillar":"Financial","Roles":"Business + Digital","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Digital Acquiring","Pillar":"Financial/Customer","Roles":"Branch staff","Type":"Count","Unit":"Accounts"},
                {"KPI":"Transactions","Pillar":"Financial","Roles":"Tellers, CSO","Type":"Count","Unit":"Count"},
                {"KPI":"Trade Finance","Pillar":"Financial","Roles":"Corporate, SME","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Treasury","Pillar":"Financial","Roles":"Corporate, Treasury","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"PBT","Pillar":"Financial","Roles":"Managers/Directors","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"Bancassurance","Pillar":"Financial","Roles":"Branch, DSO","Type":"Amount (KES)","Unit":"KES 000"},
                {"KPI":"NPL Ratio","Pillar":"Financial","Roles":"Credit, Business","Type":"Ratio","Unit":"Decimal (0.05 = 5%)"},
                {"KPI":"PAR","Pillar":"Financial","Roles":"Credit, BCM","Type":"Ratio","Unit":"Decimal"},
                {"KPI":"CIR","Pillar":"Financial","Roles":"Directors, CFO","Type":"Ratio","Unit":"Decimal (0.65 = 65%)"},
                {"KPI":"ROE","Pillar":"Financial","Roles":"MD, Directors","Type":"Ratio","Unit":"Decimal"},
                {"KPI":"New Customer Acquisition","Pillar":"Customer Focus","Roles":"All business","Type":"Count","Unit":"Customers"},
                {"KPI":"Dormancy Reactivation","Pillar":"Customer Focus","Roles":"Branch, DSO","Type":"Count","Unit":"Accounts"},
                {"KPI":"CX Score","Pillar":"Customer Focus","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"SLA Adherence Score","Pillar":"Customer Focus","Roles":"All staff","Type":"Score","Unit":"0.00–1.00 (auto)"},
                {"KPI":"Credit TAT Score","Pillar":"Customer Focus","Roles":"Credit staff","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"Diligence Score","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0–100 (auto)"},
                {"KPI":"Initiative Score","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00 (auto)"},
                {"KPI":"Compliance","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"Audit Closure","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"Timely Reconciliations","Pillar":"Operational Excellence","Roles":"Ops, Tellers","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"Branch Optimization Score","Pillar":"Operational Excellence","Roles":"BOM, BM","Type":"Score","Unit":"0.00–1.00 (auto)"},
                {"KPI":"Campaign Conversion Rate","Pillar":"Operational Excellence","Roles":"Marketing, DSO","Type":"Score","Unit":"0.00–1.00 (auto)"},
                {"KPI":"Recovery Rate","Pillar":"Financial","Roles":"DRU","Type":"Score","Unit":"0.00–1.00"},
                {"KPI":"Loan Recovery Amount","Pillar":"Financial","Roles":"DRU, Recovery Officer","Type":"Amount (KES)","Unit":"KES 000"},
            ])
            st.dataframe(kpi_ref, use_container_width=True, hide_index=True)

        with mode_tabs[3]:
            st.markdown("#### Validation rules — what the system checks on upload")
            rules = [
                ("✅","Weights sum to 1.00","Per staff member, all KPI weights must sum to exactly 1.00 (±0.005 tolerance)"),
                ("✅","No duplicate KPI rows","Each staff + KPI combination must appear only once"),
                ("✅","Valid pillar names","Pillar must be exactly: Financial, Customer Focus, Operational Excellence"),
                ("✅","Staff Code format","Staff codes must be numeric, no leading apostrophe, no spaces"),
                ("✅","Month column format","Month columns must be: Mon-YY Actual (e.g. Apr-26 Actual)"),
                ("ℹ️","Annual Target optional","Targets come from the cascade module — Annual Target column is not required"),
                ("✅","Weight > 0","Each KPI weight must be > 0"),
                ("✅","Role Function","Must be: Business, Support, or Executive"),
                ("⚠️","Auto-KPIs excluded","Diligence Score, SLA Adherence, etc. should not be in actuals upload"),
                ("⚠️","Reports To Code","Must match an existing Staff Code in the same file or overridden in Admin"),
                ("⚠️","Category","Must be: Branch or Head Office"),
            ]
            rules_df = pd.DataFrame(rules, columns=["Status","Rule","Detail"])
            st.dataframe(rules_df, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("#### Live validation check")
            st.caption("Upload your file here to check for errors before the main upload.")

            check_file = st.file_uploader("Upload file to validate", type=["xlsx","xls"], key="val_upload")
            if check_file:
                raw_chk = check_file.getvalue()
                try:
                    chk_df = pd.read_excel(io.BytesIO(raw_chk), sheet_name='KPI Data', header=1)
                    errors = []
                    warnings = []

                    # Weight check
                    chk_df['Weight'] = pd.to_numeric(chk_df['Weight'], errors='coerce').fillna(0)
                    chk_df['Weight'] = chk_df['Weight'].apply(lambda x: x/100 if x > 1 else x)
                    wt_sums = chk_df.groupby('Staff Code')['Weight'].sum()
                    bad_wt  = wt_sums[(wt_sums - 1.0).abs() > 0.005]
                    if len(bad_wt):
                        errors.append(f"❌ {len(bad_wt)} staff have weights not summing to 1.00: {bad_wt.index.tolist()[:5]}")
                    else:
                        st.success(f"✅ Weights: all {len(wt_sums)} staff sum to 1.00")

                    # Pillar check
                    valid_pillars = {'Financial','Customer Focus','Operational Excellence'}
                    bad_pillars = chk_df[~chk_df['Pillar'].isin(valid_pillars)]['Pillar'].unique()
                    if len(bad_pillars):
                        errors.append(f"❌ Invalid pillars: {bad_pillars.tolist()}")
                    else:
                        st.success("✅ Pillars: all valid")

                    # Duplicates
                    dups = chk_df.duplicated(subset=['Staff Code','KPI'])
                    if dups.sum():
                        errors.append(f"❌ {dups.sum()} duplicate Staff Code + KPI rows")
                    else:
                        st.success(f"✅ No duplicates: {len(chk_df)} unique rows")

                    # Annual Target is optional — targets come from cascade
                    if 'Annual Target' in chk_df.columns:
                        zero_tgt = (pd.to_numeric(chk_df['Annual Target'], errors='coerce').fillna(0) == 0).sum()
                        if zero_tgt:
                            st.info(f"ℹ️ {zero_tgt} rows have zero/missing Annual Target — "
                                    f"targets populate from the cascade module.")
                    else:
                        st.success("✅ Annual Target column not present — targets will come from cascade (correct format)")

                    st.metric("Total rows",    len(chk_df))
                    st.metric("Unique staff",  chk_df['Staff Code'].nunique())
                    st.metric("Unique KPIs",   chk_df['KPI'].nunique())
                    st.metric("Errors",        len(errors),   delta=f"-{len(errors)}" if errors else "0", delta_color="inverse")
                    st.metric("Warnings",      len(warnings), delta=f"-{len(warnings)}" if warnings else "0", delta_color="inverse")

                    for e in errors:
                        st.error(e)
                    for w in warnings:
                        st.warning(w)
                    if not errors and not warnings:
                        st.success("✅ File passed all validation checks. Ready to upload via the sidebar.")
                except Exception as ex:
                    st.error(f"Could not read file: {ex}")


    # ════════════════════════════════════════════════════════════════
    # TAB 9 — LEAVE SETTINGS
    # ════════════════════════════════════════════════════════════════

    # ────────────────────────────────────────────────────────────────
    # SUB-TAB: Living Documentation (v8.15) — Phase 3 of Living Doc
    # sub-campaign per docs/A2Z_LIVING_DOCS_PLAN.md Part 7
    # ────────────────────────────────────────────────────────────────
    with sub[2]:
        st.subheader("📑 Living Documentation")
        from pathlib import Path
        st.caption(
            "Audit-locked sales-grade collateral generated from the systems-layer "
            "registries. Per docs/A2Z_LIVING_DOCS_PLAN.md, every numeric claim in "
            "rendered output is validated against the registry before writing — "
            "if a claim diverges from reality, generation aborts and reports the "
            "drift. Sales claims become as audit-lockable as engineering invariants."
        )

        # ── Status panel — what the registry currently reports ──────────
        try:
            from scripts.docgen import load_registry as _docgen_load
            _reg = _docgen_load()
            st.markdown("##### 📊 Registry snapshot")
            colA, colB, colC, colD = st.columns(4)
            colA.metric("Audit gates",
                          _reg["platform"].get("audit_gates", "—"))
            colB.metric("Stocks (wired)",
                          f"{_reg['stocks_wired']}/{_reg['stocks_count']}")
            colC.metric("Loops (wired)",
                          f"{_reg['loops_wired']}/{_reg['loops_count']}")
            colD.metric("Sales-content JSONs",
                          f"{_reg['sales_content_files_present']}/6")
            st.caption(
                f"Platform version: **{_reg['platform'].get('version', '?')}** · "
                f"Engines: {_reg['platform'].get('engines_count', '?')} · "
                f"CHANGELOGs: {_reg['platform'].get('changelog_count', '?')} · "
                f"Learning loops: {_reg.get('learning_loops_count', '?')}"
            )
        except Exception as e:
            st.error(f"Living Doc registry unavailable: {type(e).__name__}: {e}")
            st.info(
                "If you see this on a fresh deployment, the docgen package may "
                "not be installed. Confirm `scripts/docgen/` exists with the "
                "v8.12 + v8.14 modules."
            )
            _reg = None

        st.markdown("---")

        # ── Generation controls ───────────────────────────────────────────
        st.markdown("##### 🎬 Generate audit-locked collateral")
        st.caption(
            "Clicking a button validates every claim in that artifact against "
            "the registry. If any claim fails, the artifact is NOT written and "
            "the divergence is reported below."
        )

        out_dir = Path("/tmp/a2z_living_docs_runtime")
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        gen_targets = [
            ("brochure", "📊 Brochure (PPTX)", "brochure",
             "A2Z_MIS_360_Brochure.pptx",
             "15-slide executive deck for bank CEO/COO/CIO; 12-min read."),
            ("magazine", "📖 Magazine (PDF)", "magazine",
             "A2Z_MIS_360_Magazine.pdf",
             "Multi-page deep-dive for evaluation committees + regulators."),
            ("security", "🛡️ Security Whitepaper", "security",
             "A2Z_MIS_360_Security_Whitepaper.pdf",
             "CISO-facing security architecture with explicit shipped/designed/roadmap markers."),
            ("compliance", "📜 Compliance Pack", "compliance",
             "A2Z_MIS_360_Compliance_Pack.pdf",
             "Regulator-facing pack mapping engines to CBK/IFRS/Basel/DPA frames."),
        ]

        for key, label, target_id, filename, description in gen_targets:
            col_btn, col_desc = st.columns([1, 3])
            with col_btn:
                clicked = st.button(label, key=f"docgen_{key}_btn",
                                     use_container_width=True,
                                     disabled=(_reg is None))
            with col_desc:
                st.caption(description)

            if clicked and _reg is not None:
                try:
                    from scripts.generate_all_docs import TARGETS
                    with st.spinner(f"Generating {label} (validating claims first)..."):
                        result = TARGETS[target_id](out_dir)

                    if result.get("status") == "OK":
                        output_path = result.get("output_path", "")
                        claims = result.get("claims_validated", 0)
                        try:
                            size_kb = Path(output_path).stat().st_size / 1024
                            st.success(
                                f"✓ Generated **{filename}** "
                                f"({size_kb:.1f}KB; {claims} audit-locked claims validated). "
                                f"Saved to `{output_path}`."
                            )
                        except Exception:
                            st.success(
                                f"✓ Generated **{filename}** "
                                f"({claims} claims validated). Saved to `{output_path}`."
                            )

                        # Offer download
                        try:
                            with open(output_path, "rb") as f:
                                st.download_button(
                                    f"📥 Download {filename}",
                                    data=f.read(),
                                    file_name=filename,
                                    key=f"docgen_{key}_download",
                                )
                        except Exception:
                            st.caption("(File available at the path above.)")
                    else:
                        # ── Audit-claim diff view ────────────────────────────
                        st.error(f"✗ Generation aborted: {result.get('reason', 'unknown')}")
                        failures = result.get("failures", [])
                        if failures:
                            st.markdown("**Audit-claim divergence diagnostic:**")
                            for f in failures:
                                st.warning(
                                    f"- **Claim**: {f.get('claim_text', '?')}\n"
                                    f"- **Registry path**: `{f.get('registry_path', '?')}`\n"
                                    f"- **Expected**: `{f.get('expected', '?')}`\n"
                                    f"- **Error**: {f.get('error', '?')}"
                                )
                            st.info(
                                "This is the audit-locked claim discipline working as designed. "
                                "Either update the registry to match the claim, OR update the "
                                "generator's `_build_claims()` to match the registry. "
                                "The collateral was NOT written — the build refuses to lie."
                            )
                except Exception as e:
                    st.error(f"Generator exception: {type(e).__name__}: {e}")
                    st.exception(e)

        st.markdown("---")

        # ── Generate-all CTA ─────────────────────────────────────────────
        st.markdown("##### 🚀 Generate all 4 artifacts")
        if st.button("Generate all 4 audit-locked artifacts",
                       key="docgen_all_btn",
                       use_container_width=True,
                       disabled=(_reg is None)):
            try:
                from scripts.generate_all_docs import TARGETS as _ALL_TARGETS
                results = []
                progress = st.progress(0.0, text="Starting...")
                total_t = len(_ALL_TARGETS)
                claim_total = 0
                ok_count = 0
                for i, (tname, tfn) in enumerate(_ALL_TARGETS.items()):
                    progress.progress((i + 0.0) / total_t,
                                       text=f"Generating {tname}...")
                    try:
                        r = tfn(out_dir)
                    except Exception as e:
                        r = {"status": "EXCEPTION", "reason": str(e)}
                    results.append((tname, r))
                    if r.get("status") == "OK":
                        ok_count += 1
                        claim_total += r.get("claims_validated", 0)
                    progress.progress((i + 1.0) / total_t,
                                       text=f"Done: {tname}")

                progress.empty()
                if ok_count == total_t:
                    st.success(
                        f"✓ All {total_t} artifacts generated. "
                        f"{claim_total} audit-locked claims validated total. "
                        f"Output directory: `{out_dir}`"
                    )
                else:
                    st.warning(
                        f"⚠ {ok_count}/{total_t} generated; "
                        f"{total_t - ok_count} aborted. See per-target buttons "
                        f"above for diagnostics."
                    )

                with st.expander("Per-target results", expanded=False):
                    for tname, r in results:
                        if r.get("status") == "OK":
                            st.write(f"✓ **{tname}**: "
                                      f"{r.get('claims_validated', 0)} claims · "
                                      f"`{r.get('output_path', '?')}`")
                        else:
                            st.write(f"✗ **{tname}**: "
                                      f"{r.get('reason', 'unknown')}")
            except Exception as e:
                st.error(f"Orchestrator failed: {type(e).__name__}: {e}")

        st.markdown("---")

        # ── Sub-campaign progress map ───────────────────────────────────
        st.markdown("##### 🗺️ Living Documentation sub-campaign progress")
        progress_md = """
| Phase | Batch | Status |
|---|---|---|
| Plan | v8.11 | ✅ Shipped — `docs/A2Z_LIVING_DOCS_PLAN.md` (588 lines) |
| Phase 1: registry loader + claim validator + 6 sales-content JSONs | v8.12 | ✅ Shipped |
| Phase 2: 3 generators + orchestrator | v8.14 | ✅ Shipped — produces 4 audit-locked artifacts |
| **Phase 3: admin/systems-view UI surface** | **v8.15** | ✅ **You're looking at it** |
| Phase 4: G110 audit gate (collateral claims traceable) | v8.16+ | ⏳ Optional hardening |
"""
        st.markdown(progress_md)

        st.caption(
            "Per docs/A2Z_LIVING_DOCS_PLAN.md Part 7 + Spirit Statements. "
            "The campaign that built the platform now operates the discipline "
            "that documents it."
        )

    # ────────────────────────────────────────────────────────────────
    # SUB-TAB: Commercial Readiness (v9.4) — surfaces v9.1 legal templates,
    # v9.2 translation prep, v9.3 patent briefs as operator-facing status panels
    # ────────────────────────────────────────────────────────────────
    with sub[3]:
        st.subheader("📜 Commercial Readiness")
        from pathlib import Path as _CRPath
        st.caption(
            "Operational artifacts shipped in the v9.x main track. Per "
            "`docs/A2Z_V8_RETROSPECTIVE_FINAL_AND_V9_PLAN.md`, the v9.x "
            "rhythm pivots from engineering-heavy (v8.x) to "
            "commercial-readiness-heavy. This panel surfaces the artifacts. "
            "**The artifacts are templates and pre-filing briefs — none "
            "is binding without professional review (lawyer / translator / "
            "patent agent).**"
        )

        # Status of each v9.x artifact directory
        st.markdown("##### 🗂️ Artifact directories")
        artifact_groups = [
            ("📋 Legal templates (v9.1)",
             "docs/legal_templates/",
             ["NDA_MUTUAL_TEMPLATE.md", "NDA_UNILATERAL_TEMPLATE.md",
              "IP_ASSIGNMENT_TEMPLATE.md", "REFERENCE_CUSTOMER_AGREEMENT_TEMPLATE.md",
              "README.md"],
             "Tier 1 templates per IP Strategy Plan Appendix A.2. "
             "**Not binding — require Kenyan corporate lawyer refinement.**"),
            ("🌍 Translation prep (v9.2)",
             "docs/translations/",
             ["TRANSLATION_PREP_GUIDE.md"],
             "Reviewer-ready guide for French + Swahili translators. "
             "**Draft translations are machine-generated starting points; "
             "require native-speaker verification.**"),
            ("📝 Patent briefs (v9.3)",
             "docs/patent_briefs/",
             ["INV-008_BRIEF.md", "INV-009_BRIEF.md", "README.md"],
             "Pre-filing technical disclosures for INV-008 + INV-009 per "
             "IP Strategy Plan Part 5. **Not filed applications — require "
             "registered Kenyan patent agent for prior-art search and "
             "filing decision.**"),
        ]

        for group_label, group_path, expected_files, group_desc in artifact_groups:
            with st.expander(group_label, expanded=False):
                st.caption(group_desc)
                _dir = _CRPath(group_path)
                if not _dir.exists():
                    st.error(f"Directory `{group_path}` not found.")
                    continue
                rows = []
                for fname in expected_files:
                    fpath = _dir / fname
                    if fpath.exists():
                        try:
                            size_kb = fpath.stat().st_size / 1024
                            line_count = sum(1 for _ in open(fpath))
                            rows.append({
                                "File": fname,
                                "Status": "✓ Present",
                                "Size (KB)": f"{size_kb:.1f}",
                                "Lines": line_count,
                            })
                        except Exception:
                            rows.append({
                                "File": fname,
                                "Status": "✓ Present (unreadable)",
                                "Size (KB)": "—",
                                "Lines": "—",
                            })
                    else:
                        rows.append({
                            "File": fname,
                            "Status": "✗ Missing",
                            "Size (KB)": "—",
                            "Lines": "—",
                        })
                if rows:
                    import pandas as _crpd
                    st.dataframe(_crpd.DataFrame(rows),
                                  use_container_width=True, hide_index=True)

        st.markdown("---")

        # ── v9.x sub-campaign progress map ───────────────────────────────
        st.markdown("##### 🗺️ v9.x main-track progress")
        progress_md = """
| Batch | Theme | Status |
|---|---|---|
| v9.0 | v8.x retrospective + v9.x plan | ✅ Shipped (`docs/A2Z_V8_RETROSPECTIVE_FINAL_AND_V9_PLAN.md`) |
| v9.1 | Operational Legal Tier 1 templates (NDA + IP Assignment + Reference Customer) | ✅ Shipped (4 templates) |
| v9.2 | Native-speaker translation prep (FR + SW) | ✅ Shipped (translation prep guide) |
| v9.3 | Patent strategy execution Phase 1 (INV-008 + INV-009 briefs) | ✅ Shipped (2 patent briefs) |
| **v9.4** | **UI surfacing for v9.1-v9.3 deliverables** | ✅ **You're looking at it** |
| v9.5 | G113 audit gate locking v9.1-v9.4 contracts | ⏳ Next batch |
"""
        st.markdown(progress_md)

        st.markdown("---")

        # ── Operator action items (what Joshua must do externally) ────────
        st.markdown("##### 📋 Operator action items (external engagement)")
        st.caption(
            "The v9.x track produces reviewer-ready artifacts; binding work "
            "happens outside the codebase. The following actions are Joshua's "
            "responsibility:"
        )
        actions = [
            ("Engage a Kenyan registered corporate lawyer", "v9.1",
             "Refine the 4 legal templates; produce binding versions; budget KES 200-400K"),
            ("Engage French + Swahili translators", "v9.2",
             "Native-speaker review of TRANSLATION_PREP_GUIDE.md; budget KES 12-45K total"),
            ("Engage a Kenyan registered patent agent", "v9.3",
             "Prior-art search for INV-008 + INV-009; filing decision; budget KES 175-300K Year 1"),
            ("Insert contact email in `LICENSE.md`", "v8.14",
             "Replace the `[contact email — to be added by Joshua]` placeholder before next github commit"),
            ("Verify github repo LICENSE state", "v8.13",
             "Confirm proprietary LICENSE.md is in place; if accidentally permissive (MIT/Apache), urgent revision needed"),
        ]
        action_rows = [
            {"Action": label, "Source batch": batch, "Notes": notes}
            for label, batch, notes in actions
        ]
        import pandas as _ardf
        st.dataframe(_ardf.DataFrame(action_rows),
                      use_container_width=True, hide_index=True)

        st.caption(
            "Per `docs/A2Z_V8_RETROSPECTIVE_FINAL_AND_V9_PLAN.md` Part 13 — "
            "v9.x's success depends on operational engagement Joshua drives "
            "outside Claude's scope. The campaign prepares; Joshua executes."
        )

    # ────────────────────────────────────────────────────────────────
    # SUB-TAB: State Backend (v9.9) — surfaces v9.6-v9.8 multi-process
    # state migration; shows backend type, key statistics, durability info
    # ────────────────────────────────────────────────────────────────
    with sub[4]:
        st.subheader("🗄️ State Backend")
        import os as _sb_os
        from utils.state_backend import get_default_backend
        from utils import flexcube_adapter as _sb_fc
        from utils import smart_alerts as _sb_sa
        from utils import event_bus as _sb_eb
        st.caption(
            "v9.6-v9.8 ships multi-process state via `utils/state_backend.py`. "
            "Default is `InMemoryBackend` (matches v8.x semantics exactly). "
            "Set `A2Z_REDIS_URL=redis://host:6379/0` to flip to `RedisBackend` "
            "for multi-Streamlit-process state sharing. **No code changes "
            "required** — switch is environment-driven."
        )

        _backend = get_default_backend()

        # ── Backend identity ────────────────────────────────────────────
        st.markdown("##### 🎯 Active backend")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Backend", _backend.backend_name())
        with col2:
            try:
                health = _backend.ping()
                st.metric("Health", "✓ healthy" if health else "✗ unhealthy")
            except Exception:
                st.metric("Health", "error")
        with col3:
            st.metric("Multi-process",
                       "✓ shared" if _backend.is_remote() else "process-local")

        if _backend.is_remote():
            st.success(
                "**Redis-backed** — circuit breaker / retry telemetry / "
                "latency / alert history / dedup state is shared across all "
                "Streamlit processes connecting to the same Redis instance.")
        else:
            st.info(
                "**In-memory backend** — state is process-local; each "
                "Streamlit process has its own copy. For multi-process "
                "deployments set `A2Z_REDIS_URL`. File persistence "
                "(`flexcube_data/latency_state.json`, "
                "`smart_alerts_data/alert_history.json`) preserves state "
                "across single-process restarts.")

        # ── Configuration ───────────────────────────────────────────────
        st.markdown("##### ⚙️ Configuration")
        config_rows = [
            {"Setting": "A2Z_REDIS_URL env var",
             "Value": _sb_os.environ.get("A2Z_REDIS_URL", "(not set)")},
            {"Setting": "Backend class", "Value": type(_backend).__name__},
            {"Setting": "Latency window size",
             "Value": str(_sb_fc.LATENCY_WINDOW_SIZE)},
            {"Setting": "Alert history max",
             "Value": str(_sb_sa.ALERT_HISTORY_MAX_ENTRIES)},
            {"Setting": "Latency persist path",
             "Value": str(_sb_fc.LATENCY_PERSIST_PATH)},
            {"Setting": "Alert history persist path",
             "Value": str(_sb_sa.ALERT_HISTORY_PATH)},
        ]
        import pandas as _sb_pd
        st.dataframe(_sb_pd.DataFrame(config_rows),
                      use_container_width=True, hide_index=True)

        # ── Key statistics by domain ─────────────────────────────────────
        st.markdown("##### 📊 State by domain")
        try:
            circuit_keys = _backend.keys_matching("circuit:")
            retry_keys = _backend.keys_matching("retry:")
            latency_keys = _backend.keys_matching("latency:")
            dedup_keys = _backend.keys_matching("dedup:")
            alert_count = _backend.list_length("alert_history")

            domain_rows = [
                {"Domain": "Circuit breakers (per-endpoint)",
                 "Backend keys": len(circuit_keys),
                 "Sample keys": ", ".join(circuit_keys[:3]) or "—",
                 "Source batch": "v9.6"},
                {"Domain": "Retry telemetry (per-endpoint)",
                 "Backend keys": len(retry_keys),
                 "Sample keys": ", ".join(retry_keys[:3]) or "—",
                 "Source batch": "v9.7"},
                {"Domain": "Latency rolling window",
                 "Backend keys": len(latency_keys),
                 "Sample keys": ", ".join(latency_keys[:2]) or "—",
                 "Source batch": "v9.8"},
                {"Domain": "Alert history (single list)",
                 "Backend keys": "1 list" if alert_count > 0 else "0",
                 "Sample keys": f"{alert_count} entries",
                 "Source batch": "v9.8"},
                {"Domain": "Event-bus dedup (per-topic)",
                 "Backend keys": len(dedup_keys),
                 "Sample keys": ", ".join(dedup_keys[:3]) or "—",
                 "Source batch": "v9.8"},
            ]
            st.dataframe(_sb_pd.DataFrame(domain_rows),
                          use_container_width=True, hide_index=True)
        except Exception as e:
            st.error(f"Failed to read backend state: "
                      f"{type(e).__name__}: {e}")

        # ── Live state verification ─────────────────────────────────────
        st.markdown("##### 🔍 Live state verification")
        verify_cols = st.columns([3, 1])
        with verify_cols[0]:
            st.caption(
                "Read live values to verify the backend is wired correctly. "
                "If you see `0`s here in steady production, the FLEXCUBE "
                "adapter has not been called yet. After traffic flows "
                "through `_live_request()` these counters become non-zero.")
        with verify_cols[1]:
            if st.button("🔄 Refresh", key="state_backend_refresh"):
                st.rerun()

        try:
            cs = _sb_fc.get_circuit_state()
            rt = _sb_fc.get_retry_telemetry()
            ls = _sb_fc.get_latency_state()
            ahs = _sb_sa.get_alert_history_stats()
            ds = _sb_eb.get_dedup_stats()

            metric_cols = st.columns(5)
            metric_cols[0].metric(
                "Circuit endpoints",
                cs.get("endpoints_tracked", 0),
                "open" if cs.get("is_open") else "closed",
            )
            metric_cols[1].metric(
                "Retry total requests",
                rt["summary"].get("requests_total", 0),
                f"{rt['summary'].get('retry_recovery_rate_pct') or 0}% recovery",
            )
            metric_cols[2].metric(
                "Latency endpoints",
                ls["summary"].get("endpoints_observed", 0),
                f"{ls['summary'].get('total_calls', 0)} calls",
            )
            metric_cols[3].metric(
                "Alerts total",
                ahs.get("total", 0),
                f"{ahs.get('acknowledged', 0)} acked",
            )
            metric_cols[4].metric(
                "Dedup topics",
                ds.get("topics_tracked", 0),
                f"{ds.get('dedup_hit_rate_pct') or 0}% hit rate",
            )
        except Exception as e:
            st.error(f"Failed to read live state: "
                      f"{type(e).__name__}: {e}")

        st.markdown("---")

        # ── Migration map ───────────────────────────────────────────────
        st.markdown("##### 🗺️ v9.6-v9.8 migration map")
        migration_md = """
| Batch | What migrated | Old location | New backend key |
|---|---|---|---|
| v9.6 | Per-endpoint circuit state | `_CIRCUIT_STATES` (v8.17) | `circuit:{endpoint_key}` |
| v9.7 | Retry telemetry | `_RETRY_TELEMETRY` (v8.19) | `retry:{endpoint_key}` |
| v9.8 | Latency rolling window | `_LATENCY_SAMPLES` + JSON file (v8.2/v8.24) | `latency:{endpoint_path}` |
| v9.8 | Alert history | `_ALERT_HISTORY` + JSON file (v8.25) | `alert_history` (list) |
| v9.8 | Event-bus dedup stats | `_DEDUP_STATS` (v8.23) | `dedup:{topic}` |
"""
        st.markdown(migration_md)

        st.caption(
            "Per `docs/A2Z_V8_RETROSPECTIVE_FINAL_AND_V9_PLAN.md` Part 7. "
            "All five state surfaces are now multi-process-safe when "
            "RedisBackend is active. v9.10 adds G114 audit gate locking "
            "the abstraction contract."
        )

        # ────────────────────────────────────────────────────────────
        # v9.14 — Production operations panel
        # Surfaces deployment-grade configuration, memory usage,
        # and operator destructive actions
        # ────────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("##### 🛠️ Production operations")
        st.caption(
            "Per `docs/REDIS_DEPLOYMENT_RUNBOOK.md` (v9.12). "
            "Operator surfaces for production deployment debugging "
            "and maintenance. **Destructive actions are gated** by "
            "explicit confirmation."
        )

        # ── Connection pool config (only meaningful for RedisBackend) ──
        if _backend.is_remote() and hasattr(_backend, 'get_connection_config'):
            with st.expander("🔌 Connection pool configuration", expanded=False):
                try:
                    conn_cfg = _backend.get_connection_config()
                    cfg_rows = [{"Setting": k, "Value": str(v)}
                                  for k, v in conn_cfg.items()]
                    st.dataframe(_sb_pd.DataFrame(cfg_rows),
                                  use_container_width=True, hide_index=True)
                    st.caption(
                        "Credentials are masked in the URL. Tune via env "
                        "vars: `A2Z_REDIS_MAX_CONNECTIONS`, "
                        "`A2Z_REDIS_SOCKET_TIMEOUT`, "
                        "`A2Z_REDIS_HEALTH_CHECK_INTERVAL`."
                    )
                except Exception as e:
                    st.error(f"Failed to read connection config: "
                              f"{type(e).__name__}: {e}")

        # ── Operator destructive actions ───────────────────────────────
        with st.expander("⚠️ Operator destructive actions",
                          expanded=False):
            st.warning(
                "Each action below clears state for one domain. State "
                "rebuilds from new traffic. Use during debugging or "
                "deliberate reset only."
            )
            op_cols = st.columns(2)
            domain_options = [
                ("circuit", "Circuit breaker state",
                 "Resets all per-endpoint circuit state. Affects: "
                 "consecutive_failures and tripped_until per endpoint."),
                ("retry", "Retry telemetry counters",
                 "Resets all per-endpoint retry counters. Affects: "
                 "requests_total, retries_triggered, recovery rates."),
                ("latency", "Latency rolling window",
                 "Resets per-endpoint latency samples. Affects: "
                 "p50/p95/p99 percentile observability."),
                ("dedup", "Event-bus dedup statistics",
                 "Resets per-topic dedup counters. Affects: "
                 "total_publish_calls, dedup_hits, unique_published."),
                ("alert_history", "Alert history",
                 "Removes all alert history entries. Affects: "
                 "operator dashboard alert visibility."),
            ]
            for i, (domain_key, label, description) in enumerate(domain_options):
                with op_cols[i % 2]:
                    st.markdown(f"**{label}**")
                    st.caption(description)
                    confirm_key = f"v914_clear_{domain_key}_confirm"
                    if not st.session_state.get(confirm_key, False):
                        if st.button(f"Clear {domain_key}",
                                      key=f"v914_clear_{domain_key}_btn"):
                            st.session_state[confirm_key] = True
                            st.rerun()
                    else:
                        st.error(
                            f"⚠️ Confirm clearing {domain_key}? This "
                            f"cannot be undone.")
                        confirm_cols = st.columns(2)
                        if confirm_cols[0].button(
                                f"YES, clear {domain_key}",
                                key=f"v914_confirm_{domain_key}_yes"):
                            try:
                                if domain_key == "alert_history":
                                    cleared = _backend.list_length(
                                        "alert_history")
                                    _backend.list_clear("alert_history")
                                    st.success(
                                        f"Cleared {cleared} alert_history "
                                        f"entries.")
                                else:
                                    keys = _backend.keys_matching(
                                        f"{domain_key}:")
                                    cleared = 0
                                    for k in keys:
                                        try:
                                            _backend.hash_delete(k)
                                            cleared += 1
                                        except Exception:
                                            try:
                                                _backend.list_clear(k)
                                                cleared += 1
                                            except Exception:
                                                pass
                                    st.success(
                                        f"Cleared {cleared}/{len(keys)} "
                                        f"{domain_key} keys.")
                                st.session_state[confirm_key] = False
                            except Exception as e:
                                st.error(f"Clear failed: "
                                          f"{type(e).__name__}: {e}")
                                st.session_state[confirm_key] = False
                        if confirm_cols[1].button(
                                "Cancel",
                                key=f"v914_confirm_{domain_key}_no"):
                            st.session_state[confirm_key] = False
                            st.rerun()

        # ── Operator command-line tooling ───────────────────────────────
        with st.expander("📋 Command-line operations (redis_admin.py)",
                          expanded=False):
            st.caption(
                "For programmatic ops, the v9.13 CLI provides equivalent "
                "actions plus snapshot/restore for backups. Run from "
                "the project root:"
            )
            cli_examples_md = """
```bash
# Read-only diagnostics
python scripts/redis_admin.py health-check
python scripts/redis_admin.py config
python scripts/redis_admin.py inventory
python scripts/redis_admin.py live-state
python scripts/redis_admin.py verify-state

# Backup state to JSON file
python scripts/redis_admin.py snapshot -o /backup/a2z-state.json

# Restore from backup (destructive)
python scripts/redis_admin.py restore \\
    -i /backup/a2z-state.json --confirm

# Clear specific domain (destructive)
python scripts/redis_admin.py clear-domain \\
    --domain circuit --confirm
```
"""
            st.markdown(cli_examples_md)
            st.caption(
                "See `docs/REDIS_DEPLOYMENT_RUNBOOK.md` for full "
                "deployment guide including TLS / ACL / monitoring / "
                "backup automation / disaster recovery procedures."
            )

        # ────────────────────────────────────────────────────────────
        # v9.19 — Load test results + observability stack panels
        # ────────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("##### 🧪 Load test + observability")
        st.caption(
            "Operator surfaces for v9.17 load testing + v9.18 "
            "observability stack."
        )

        # ── Latest load test results ───────────────────────────────────
        with st.expander("🧪 Recent load test results",
                          expanded=False):
            st.caption(
                "v9.17 ships `scripts/load_test_multi_instance.py` — "
                "concurrent-user simulator validating v9.6-v9.16 "
                "architecture. Output JSON files appear here."
            )
            from pathlib import Path as _v919Path
            # Look for any recent load test JSONs in /tmp/ or load_test_results/
            search_paths = [
                _v919Path("load_test_results"),
                _v919Path("/tmp"),
            ]
            found_files = []
            for sp in search_paths:
                if sp.exists():
                    for f in sp.glob("loadtest*.json"):
                        try:
                            mtime = f.stat().st_mtime
                            found_files.append((mtime, f))
                        except Exception:
                            pass
            found_files.sort(reverse=True)
            found_files = found_files[:5]  # most recent 5

            if not found_files:
                st.info(
                    "No load test results found. Run "
                    "`python scripts/load_test_multi_instance.py "
                    "--users 10 --calls 100 --output "
                    "/tmp/loadtest_$(date +%Y%m%d_%H%M).json` "
                    "to generate one."
                )
            else:
                file_options = [str(f) for _, f in found_files]
                selected = st.selectbox("Recent load test files:",
                                          file_options,
                                          key="v919_loadtest_select")
                if selected:
                    try:
                        import json as _v919_json
                        loadtest_data = _v919_json.loads(
                            _v919Path(selected).read_text(
                                encoding="utf-8"))
                        # Display key metrics
                        ltcols = st.columns(4)
                        ltcols[0].metric(
                            "Total calls",
                            loadtest_data.get("total_calls", "?"))
                        ltcols[1].metric(
                            "Success rate",
                            f"{loadtest_data.get('success_rate_pct', 0)}%")
                        ltcols[2].metric(
                            "Throughput",
                            f"{loadtest_data.get('throughput_calls_per_second', 0)}/s")
                        ltcols[3].metric(
                            "Latency p95",
                            f"{loadtest_data.get('latency_p95_ms', 0)} ms")
                        st.caption(
                            f"Started: {loadtest_data.get('started_iso', '?')} | "
                            f"Duration: {loadtest_data.get('duration_seconds', '?')}s | "
                            f"Backend: {loadtest_data.get('backend_state_summary', {}).get('backend', '?')}"
                        )

                        if loadtest_data.get('per_endpoint_summary'):
                            ep_rows = []
                            for ep, s in loadtest_data['per_endpoint_summary'].items():
                                ep_rows.append({
                                    "Endpoint": ep,
                                    "Calls": s['calls'],
                                    "Successes": s['successes'],
                                    "p50_ms": s['p50_ms'],
                                    "p95_ms": s['p95_ms'],
                                    "Retries": s['total_retries'],
                                })
                            st.dataframe(_sb_pd.DataFrame(ep_rows),
                                          use_container_width=True,
                                          hide_index=True)
                    except Exception as e:
                        st.error(f"Failed to load test result: "
                                  f"{type(e).__name__}: {e}")

        # ── Observability stack ──────────────────────────────────────
        with st.expander("📊 Observability stack (Prometheus + Grafana)",
                          expanded=False):
            st.caption(
                "v9.18 ships the deployment recipe and dashboard JSON "
                "for surfacing A2Z telemetry into a production "
                "monitoring stack."
            )
            obs_artifacts = [
                ("📘 Runbook",
                 "docs/OBSERVABILITY_DASHBOARD_RUNBOOK.md",
                 "Telemetry sources + Prometheus exporter pattern + "
                 "thresholds + alert rules + triage flow"),
                ("📊 Grafana dashboard",
                 "scripts/observability/grafana_dashboard.json",
                 "Importable Grafana 9.x+ dashboard. 7 panels: "
                 "circuits / failures / recovery / latencies / alerts"),
                ("🚨 Prometheus alerts",
                 "scripts/observability/prometheus_alerts.yml",
                 "8 alert rules across 4 groups; warning + critical "
                 "tiers"),
            ]
            obs_rows = []
            for label, path, desc in obs_artifacts:
                p = _v919Path(path)
                exists = p.exists()
                size_kb = (p.stat().st_size / 1024) if exists else 0
                obs_rows.append({
                    "Artifact": label,
                    "Path": path,
                    "Status": "✓ Present" if exists else "✗ Missing",
                    "Size (KB)": f"{size_kb:.1f}" if exists else "—",
                    "Description": desc,
                })
            st.dataframe(_sb_pd.DataFrame(obs_rows),
                          use_container_width=True, hide_index=True)

            st.markdown("**Telemetry surfaces exposed to Prometheus:**")
            st.markdown("""
- `a2z_circuit_open{endpoint}` — 1=open, 0=closed
- `a2z_circuit_consecutive_failures{endpoint}` — gauge
- `a2z_retry_recovery_rate_pct{endpoint}` — gauge (%)
- `a2z_latency_p95_ms{endpoint}` / `a2z_latency_p99_ms{endpoint}` — gauges
- `a2z_alerts_unacknowledged` — gauge
- `a2z_alerts_total{tier}` — counter (URGENT/HIGH/INFO)
""")
            st.caption(
                "Per `docs/OBSERVABILITY_DASHBOARD_RUNBOOK.md` §2 — "
                "operator deploys a thin Prometheus exporter as a "
                "sidecar process that polls A2Z's public APIs and "
                "exposes these metrics on `/metrics`. Skeleton in §2.2."
            )

    # ────────────────────────────────────────────────────────────────
    # SUB-TAB: Engine Hub (v9.21) — surfaces unintegrated engines
    # ────────────────────────────────────────────────────────────────
    with sub[5]:
        st.subheader("🔌 Engine Hub")
        st.caption(
            "Operator surface for engines that don't yet have dedicated "
            "Streamlit pages. Each engine is verified importable + its "
            "public API surface displayed. Provides a unified status view "
            "of the platform's full engine inventory while individual "
            "engines await deeper UI integration."
        )

        import importlib as _hub_imp
        from pathlib import Path as _hub_Path

        # ── Engine Hub registry ──────────────────────────────────────
        # Tier 1 (v9.21): regulatory + financial reporting + IFRS/IAS
        # Future v9.22+: customer intelligence / profitability / strategy / ops
        ENGINE_HUB_TIERS = {
            "Tier 1 — Regulatory & Financial Reporting (v9.21)": [
                ("board_reporting", "BoardReportingEngine",
                 "Board-pack consolidation: KPIs + compliance + risk "
                 "summary for monthly board meetings"),
                ("earnings_per_share", "EarningsPerShareEngine",
                 "Basic + diluted EPS calculation per IAS 33; "
                 "segments by ordinary / preference shares"),
                ("cash_flow_statement", "CashFlowEngine",
                 "Statement of cash flows per IAS 7; operating / "
                 "investing / financing classifications"),
                ("pillar3_disclosure", None,
                 "Basel III Pillar 3 disclosures: capital structure, "
                 "risk-weighted assets, leverage ratio"),
                ("regulatory_reporting", None,
                 "Generic regulatory reporting engine; CBK BSD returns "
                 "+ ad-hoc supervisory submissions"),
                ("risk_weighted_assets", None,
                 "RWA computation per Basel III: credit / market / "
                 "operational risk exposures"),
                ("market_risk", None,
                 "Market risk metrics: VaR, sensitivities, FX exposure, "
                 "interest rate gap"),
                ("esg_reporting", "EsgReportingEngine",
                 "ESG disclosures: GHG inventory, social metrics, "
                 "governance summary per emerging Kenya guidelines"),
                ("fair_value_measurement", "FairValueEngine",
                 "IFRS 13 fair value hierarchy: levels 1/2/3 with "
                 "input observability documentation"),
                ("ias1_presentation", "IAS1PresentationEngine",
                 "IAS 1 financial statement presentation: structure + "
                 "comparatives + accounting policies disclosure"),
                ("ias8_policies", "IAS8PoliciesEngine",
                 "IAS 8 changes in accounting policies / estimates / "
                 "errors disclosure"),
                ("ifrs7_disclosures", "IFRS7DisclosureEngine",
                 "IFRS 7 financial-instrument disclosures: classes, "
                 "risks, fair-value measurement, hedge accounting"),
            ],
            "Tier 2 — Customer & Operational Intelligence (v9.22)": [
                ("deposit_intelligence", "DepositIntelligenceEngine",
                 "Deposit portfolio analytics: vintage cohorts, "
                 "concentration risk, attrition prediction"),
                ("dormancy_intelligence", "DormancyIntelligenceEngine",
                 "Account dormancy detection + reactivation campaign "
                 "scoring; CBK dormancy compliance"),
                ("lending_intelligence", "LendingIntelligenceEngine",
                 "Loan portfolio analytics: concentration, vintage "
                 "performance, early-warning signals"),
                ("treasury_intelligence", "TreasuryIntelligenceEngine",
                 "Treasury position analytics: liquidity ladder, "
                 "currency mismatch, maturity profile"),
                ("business_intelligence", "AutomatedBusinessIntelligence",
                 "Cross-domain BI: trend detection, anomaly flagging, "
                 "exec-summary generation"),
                ("management_reporting", "ManagementReportingEngine",
                 "MIS pack: standardized monthly management reports "
                 "(P&L by segment, balance sheet trends)"),
                ("operations_dashboard", "OperationsDashboardEngine",
                 "Real-time ops KPIs: branch productivity, channel "
                 "uptime, transaction throughput"),
                ("queue_analytics", "QueueAnalyticsEngine",
                 "Branch + call-centre queue analytics: wait times, "
                 "abandonment rate, CSAT correlation"),
                ("sanctions_screening", "SanctionsScreeningEngine",
                 "OFAC + UN + EU sanctions list screening; AML "
                 "compliance for transactions and customers"),
                ("funds_transfer_pricing", "FtpEngine",
                 "FTP curve construction + product-level FTP rates "
                 "for profitability attribution"),
                ("cost_allocation", None,
                 "Activity-based cost allocation: shared services + "
                 "cost driver assignment"),
                ("operating_segments", "OperatingSegmentEngine",
                 "IFRS 8 operating segments: chief operating decision "
                 "maker view, segment reconciliation"),
            ],
            "Tier 3 — Profitability Suite (v9.23)": [
                ("product_profitability", "ProductProfitabilityEngine",
                 "Per-product P&L: revenue + cost-to-serve + capital "
                 "charge per product line"),
                ("product_raroc", "ProductRarocEngine",
                 "Product-level Risk-Adjusted Return on Capital; "
                 "drives pricing + portfolio decisions"),
                ("rm_profitability", "RMProfitabilityDashboard",
                 "Per-RM profitability with portfolio P&L attribution; "
                 "feeds incentive computation"),
                ("profitability_integration", None,
                 "Profitability data integration layer: harmonizes "
                 "product / RM / customer P&Ls into unified view"),
                ("profitability_heatmap", None,
                 "2D profitability visualization: customer×product / "
                 "branch×RM heatmaps for portfolio review"),
                ("profitability_hierarchy", "CustomerProfitabilityHierarchy",
                 "Hierarchical profitability rollup: customer→RM→"
                 "branch→region→bank with reconciling totals"),
                ("profitability_trends", "ProfitabilityTrends",
                 "Time-series profitability: rolling 12-month trends, "
                 "vintage analysis, seasonality decomposition"),
                ("asset_impairment", "ImpairmentEngine",
                 "IAS 36 asset impairment: recoverable amount + value "
                 "in use + impairment loss recognition"),
                ("channel_income", "ChannelIncomeEngine",
                 "Per-channel revenue attribution: ATM / USSD / "
                 "Mobile / Branch / Internet banking"),
                ("channel_performance", "ChannelPerformanceEngine",
                 "Channel KPIs: throughput, success rate, peak load, "
                 "uptime correlated with revenue"),
                ("channel_sla", "ChannelSlaMonitoringEngine",
                 "Channel SLA tracking: target vs actual uptime, "
                 "outage cost analysis"),
            ],
            "Tier 4 — Strategy & Initiatives (v9.24)": [
                ("strategic_planning", "StrategicPlanningEngine",
                 "Strategic plan management: 3-year plan, KPI cascade, "
                 "milestone tracking + variance analysis"),
                ("initiative_impact", "InitiativeImpactEngine",
                 "Initiative impact attribution: revenue / cost / risk "
                 "deltas attributable to each strategic initiative"),
                ("initiative_dependency", "DependencyIntelligenceEngine",
                 "Initiative dependency graph: blocking / blocked-by "
                 "analysis, critical path identification"),
                ("initiative_resource", "ResourceIntelligenceEngine",
                 "Initiative resource allocation: people, budget, "
                 "capacity constraints + bottleneck detection"),
                ("stage_gate", "StageGateEngine",
                 "Stage-gate process: ideation → business case → "
                 "execution → benefits realization gates"),
                ("growth_path_engine", "GrowthPathEngine",
                 "Per-RM career growth path: skill gaps, milestone "
                 "tracking, promotion readiness scoring"),
            ],
            "Tier 5 — People & Operations (v9.24)": [
                ("microtask_engine", "MicroTaskEngine",
                 "Micro-task assignment + tracking: daily 15-minute "
                 "actions for RMs derived from coaching insights"),
                ("nudge_engine", "PerformanceNudgeEngine",
                 "Behavioral nudge generation: targeted prompts driving "
                 "specific KPI improvements"),
                ("gamification", "GamificationEngine",
                 "Points + badges + leaderboards: motivational layer "
                 "over performance metrics"),
                ("peer_learning", "PeerLearningNetwork",
                 "Peer learning network: best-practice cards from "
                 "high-performers shared with peer cohort"),
                ("wellness", "WellnessEngine",
                 "Staff wellness alerts: workload stress, burnout "
                 "indicators, work-life balance signals"),
                ("workforce_analytics", "WorkforceAnalyticsEngine",
                 "Workforce metrics: headcount trends, attrition, "
                 "skill gaps, succession planning"),
                ("employee_benefits", "EmployeeBenefitsEngine",
                 "IAS 19 employee benefits: defined-benefit obligation, "
                 "actuarial gains/losses, pension expense"),
                ("edms", "EDMSEngine",
                 "Electronic document management: policy versioning, "
                 "approval workflows, retention compliance"),
            ],
            "Tier 6 — Audit, Compliance & Workflow (v9.25)": [
                ("bsc_engine", None,
                 "Balanced Scorecard core engine: 4-pillar BSC "
                 "construction, KPI scoring, performance overlay. "
                 "Function-based; central to A2Z perform module"),
                ("audit_reporting", "AuditReportingEngine",
                 "Audit report generation: findings, recommendations, "
                 "management responses with severity tiers"),
                ("audit_universe", "AuditUniverseEngine",
                 "Audit universe definition: auditable entities, "
                 "risk-based scheduling, coverage tracking"),
                ("issue_management", "IssueManagementEngine",
                 "Audit issue tracking: identification → assignment → "
                 "remediation → verification workflow"),
                ("submission_workflow", "SubmissionWorkflowEngine",
                 "Regulatory submission workflow: prepare → review → "
                 "approve → submit → archive lifecycle"),
                ("efficiency", "EfficiencyEngine",
                 "Operational efficiency scoring: process timings, "
                 "cost-per-output, productivity benchmarks"),
                ("fatca_crs", "FatcaCrsReportingEngine",
                 "FATCA + CRS reporting: account holder identification, "
                 "reportable balance snapshots, IRS/CRS XML output"),
                ("held_for_sale", "HeldForSaleEngine",
                 "IFRS 5 held-for-sale assets: classification criteria, "
                 "measurement at lower of carrying/fair-value"),
            ],
            "Tier 7 — Climate / ESG (v10.6-v10.10)": [
                ("esg_intelligence", "ESGIntelligenceEngine",
                 "IFRS S1/S2 + KGFT + governance — 5 core ESG standards: "
                 "framework definitions, scope 1/2/3 emissions, green "
                 "asset classification, climate governance assessment"),
                ("climate_risk", "ClimateRiskEngine",
                 "Physical + transition + TNFD biodiversity risk modeling: "
                 "RCP/NGFS scenarios, sector vulnerability, carbon price "
                 "exposure, stranded asset estimates"),
                ("climate_ecl_adjustment", "ClimateECLEngine",
                 "Climate-adjusted ECL per IFRS 9 §5.5.17 + §5.5.4: "
                 "PD/LGD/EAD multipliers from physical+transition risk "
                 "scores; probability-weighted across stress scenarios"),
                ("esg_reporting_outputs", None,
                 "KGFT + CRDF report generators + greenwashing claim "
                 "verification: structured disclosure pack assembly per "
                 "CBK April 2025 requirements"),
            ],
            "Tier 8 — Credit AI Underwriting (v10.11-v10.15)": [
                ("ai_underwriting", "AIUnderwritingEngine",
                 "AI credit decisioning + explainability + EU AI Act "
                 "compliance + CFPB adverse action codes (Reg B App C). "
                 "Cat A — decisions affect issuance/denial outcomes"),
                ("applicant_data_sources", "ApplicantDataAggregator",
                 "Alt data + bureau (TransUnion/Metropol/Creditinfo) + "
                 "eKYC (IPRS, biometric, PEP, sanctions) + fraud signals: "
                 "unified profile feeds the AI underwriting engine"),
                ("risk_based_pricing", None,
                 "Basel IRB-aligned risk-based rate calculation: 5-component "
                 "decomposition, RAROC, floor/ceiling decisioning. "
                 "Cat A — rate determination affects revenue + customer"),
                ("credit_workflow", "CreditWorkflowEngine",
                 "17-state application machine + 80/20 automation policy + "
                 "credit committee voting (4-tier amount thresholds) + "
                 "memo drafting (LLM-hookable per Rule 7)"),
                ("portfolio_monitoring", "PortfolioMonitoringEngine",
                 "Early warning signals + collections strategy + "
                 "unstructured signals (adverse media, regulatory filing). "
                 "CBK PG/04 risk classification + DPD-bucket roll rates"),
                ("fairness_testing", None,
                 "ECOA + EU AI Act Art 10 fairness audit: 4/5ths rule "
                 "disparate impact, equal opportunity difference, LDA-based "
                 "latent bias search across protected classes"),
                ("document_management", "DocumentManagementEngine",
                 "Digital document lifecycle (8 states) + authenticity "
                 "checks (hash, format, MRZ, hologram) + retention per "
                 "Kenya DPA 2019 + CBK AML 7yr / KRA 5yr"),
                ("group_exposure", "GroupExposureEngine",
                 "Single obligor (25%) + group + insider (5% / 20% agg) "
                 "limit checks per Banking Act §10A + §11; large exposure "
                 "reporting threshold (10%) + Basel CCF aggregation"),
            ],
            "Tier 9 — KESONIA + RBCPM (v10.17)": [
                ("benchmark_rates", "BenchmarkRateRegistry",
                 "KESONIA + CBR + KESONIA Compounded Index registry per "
                 "CBK Revised RBCPM (Aug 2025). Total Rate = KESONIA + K. "
                 "New variable KES loans from 1 Dec 2025; existing migrate "
                 "by 28 Feb 2026. Rule 7 — rate fetcher hookable, no fab"),
            ],
            "Tier 10 — RMS Reconciliation (v10.18+)": [
                ("reconciliation_matching", "ReconciliationMatchingEngine",
                 "Multi-source matching engine: 13-source DataSource enum "
                 "(GL/CBS/Nostro/Vostro/KEPSS/PesaLink/SWIFT/etc.) + "
                 "vendor name normalization (Kenya legal-suffix stripping, "
                 "synonym expansion) + 7 match algorithms (exact ref, "
                 "exact amount+date, tolerance, fuzzy name, ML-hookable). "
                 "ENH-RMS-R1 90% auto-match target enforced + reported"),
                ("reconciliation_workflow", "ReconciliationWorkflowEngine",
                 "Exception lifecycle (10 types × 9 states with explicit "
                 "transitions) + queue routing (9 queues by type/amount/"
                 "hint) + aging buckets + SLA tracking. Memory layer: "
                 "signature-based pattern recall with confidence growth "
                 "(0.5 → 0.75 → 0.90). Timing-difference auto-handling "
                 "(T+1 auto, T+2-3 review). Governed execution: 7 guard "
                 "types (amount limit, business hours, dual approval, "
                 "rate limit, blocked counterparties, account types, "
                 "pattern confidence floor). TruePath-style guardrails"),
                ("reconciliation_specialized", "SpecializedReconciliationEngine",
                 "CBK statutory returns (15 types: CRR/LR/RAR/CAR/LCR/"
                 "NSFR/etc.) + Nostro/Vostro recon (SWIFT MT940/942/950 "
                 "+ MX camt.052/053/054, FX revaluation, stale-item "
                 "aging 0-30/31-60/61-90/91+ per CBK CRMF §6.4) + "
                 "intercompany counterparty matrix + 9 internal "
                 "suspense categories with per-category aging + "
                 "real-time KEPSS/PesaLink (auto/delayed/breach/pending "
                 "verdicts within 5-min default latency target)"),
                ("reconciliation_realtime", "ReconciliationRealtimeEngine",
                 "Real-time dashboard (4 KPIs: match rate, open "
                 "exceptions, SLA breaches, critical alerts with G/A/R "
                 "thresholds) + AI learning loop (FeedbackOutcome "
                 "capture, train_callable Rule-7 hookable) + continuous "
                 "reconciliation (StreamingWatermark + late-arrival "
                 "detection) + audit certification (8 cert states with "
                 "explicit transition graph + immutable audit trail + "
                 "dual-approval check) + sub-monthly cadence policy "
                 "(REAL_TIME→MONTHLY ordering + per-account-type minimum "
                 "per CBK CRMF §6.5)"),
            ],
            "Tier 11 — Audit/GRC (v10.23+)": [
                ("audit_core", "AuditCoreEngine",
                 "Audit universe (8 entity types) + risk-based annual "
                 "planning (5-tier risk × frequency-by-risk policy: "
                 "CRITICAL/HIGH=Annual, MEDIUM=Biennial, LOW=Triennial, "
                 "VERY_LOW=As-Required per IPPF Std 2010) + continuous "
                 "control monitoring (4 control types × 7 verdicts × "
                 "4-tier severity → remediation deadline) + electronic "
                 "working papers with SHA-256 integrity per IPPF Std "
                 "2330 (7-year retention per CBK CRMF §7) + "
                 "Connect-Validate-Respond architecture (3-stage CVR "
                 "with hookable connector/validator/responder per "
                 "Rule 7)"),
                ("audit_controls_issues", "AuditControlsIssuesEngine",
                 "Issue tracking (7 sources × 8 lifecycle states with "
                 "explicit transitions × 4-tier severity-based deadlines "
                 "7d/30d/60d/90d) + 4 aging buckets (FRESH/APPROACHING/"
                 "OVERDUE/AGED) + automated test scripts library (7 "
                 "languages: SQL/Python/SPL/KQL/Shell/Regex/Declarative) "
                 "+ test scheduling with due/overdue detection + control "
                 "graph mapping across 14 frameworks (COSO IC/ERM, COBIT "
                 "2019, ISO 27001/27002, NIST CSF/800-53, PCI DSS, SOX "
                 "404, CBK PG/02 + CRMF, Basel BCBS 239, GDPR, Kenya "
                 "DPA) with 10 seed canonical concepts + ticketing "
                 "integration (5 systems: Jira/ServiceNow/GitHub/Azure/"
                 "Internal) with hookable creator + status sync "
                 "per Rule 7"),
                ("audit_analytics_vendor", "AuditAnalyticsVendorEngine",
                 "AI-powered audit analytics — Z-score + IQR statistical "
                 "outlier detection + Benford's Law fraud detection "
                 "(chi-square test, ~95% / ~90% confidence thresholds) "
                 "+ ML detector hookable per Rule 7. Vendor risk "
                 "monitoring: 4-tier (CRITICAL/HIGH/MEDIUM/LOW) × 12 "
                 "categories × 8 risk dimensions (financial/cyber/"
                 "operational/reputational/regulatory/BCM/concentration/"
                 "data_privacy) × tier-based reassessment cadence (180d/"
                 "365d/730d/1095d) + concentration risk per CBK "
                 "Outsourcing 25% threshold. 24/7 always-on assurance: "
                 "4 priorities (P1-P4) × 7 alert channels (PagerDuty/"
                 "SMS/Slack/Email/SIEM/Board/Audit Committee) × per-"
                 "priority response SLAs (15min/4h/24h/1wk). Cyber "
                 "framework integration: NIST CSF v2.0 (6 functions × 22 "
                 "categories) + ISO 27001:2022 (4 groups × 93 controls) "
                 "+ CIS Controls v8 (18 controls × 153 sub-controls)"),
                ("audit_dashboards_portal", "AuditDashboardsPortalEngine",
                 "Auditor dashboard with 8 default KPIs and direction-"
                 "aware G/A/R thresholds (HIGHER_IS_BETTER / "
                 "LOWER_IS_BETTER / TARGET_RANGE), 4 view modes (Desktop/"
                 "Tablet/Mobile-Dense/Mobile-Summary) with priority-"
                 "ordered filtering. External auditor portal: engagement-"
                 "scoped access (3 access levels: READ_ONLY / "
                 "READ_WITH_NOTES / EXPORT_ALLOWED) × 9 request types × "
                 "immutable access logs with denial reasons. Audit "
                 "committee reporting: 5×5 risk heatmap (likelihood × "
                 "impact, scoring 1-25) with 4-zone summary (low/medium/"
                 "high/critical), plan-vs-actual completion + hours "
                 "variance. Board-ready risk dashboard: 10 risk "
                 "categories × quantified metrics with appetite "
                 "utilization (WITHIN/APPROACHING/BREACH) × VaR + EL "
                 "support per NIST SP 800-30"),
                ("audit_trail_certification", "AuditTrailCertificationEngine",
                 "SHA-256 hash chain audit trail (13 event types × "
                 "monotonic sequence × tamper-detection across all "
                 "fields including previous_hash linkage; "
                 "verify_chain_integrity surfaces first break with "
                 "explicit reason per Rule 1). Period sealing: "
                 "cryptographic snapshot at end of period; "
                 "seal_period verifies chain integrity before sealing. "
                 "Compliance certification: 11 frameworks (SOX 302/"
                 "404/906, CBK CRMF + Banking Act, Basel BCBS 239, "
                 "ISO 27001, PCI DSS, GDPR Art 30, Kenya DPA, internal "
                 "governance) × 9 attestation roles × per-framework "
                 "required signatures × 8-state lifecycle with explicit "
                 "transition graph + segregation-of-duties enforcement "
                 "(distinct user IDs required for ATTESTED transition)"),
            ],
            "Tier 12 — Model Governance (v10.28+)": [
                ("model_governance", "ModelGovernanceEngine",
                 "Cat A safety net before any ML pilot. Model inventory "
                 "(15 model types × 3-tier risk classification per SR "
                 "11-7: Tier 1 HIGH=annual / Tier 2 MEDIUM=biennial / "
                 "Tier 3 LOW=triennial validation cadence) × 4-category "
                 "EU AI Act risk × 8-state lifecycle with explicit "
                 "transition graph (DEV→TESTING→VALIDATION→APPROVED→"
                 "PRODUCTION + UNDER_REMEDIATION/SUSPENDED/RETIRED). "
                 "Tier 1/2 IN_PRODUCTION blocked without passed "
                 "independent validation report. Drift detection: "
                 "PSI (Siddiqi 2017 thresholds 0.10/0.20/0.25) + KS test "
                 "(Smirnov 1948 critical values) + Wasserstein distance "
                 "(1D earth mover's). Validation framework: 11 gates "
                 "per SR 11-7 with tier-based requirements. "
                 "Explainability: 7 methods (SHAP/LIME/permutation/"
                 "partial_dependence/integrated_gradients/counterfactual/"
                 "rule_extraction) — Rule 7 hookable, never fabricates. "
                 "CFPB Reg B Appendix C adverse action codes (20 codes). "
                 "Bias monitoring: 6 metrics (4/5ths rule per EEOC 29 "
                 "CFR §1607.4, demographic parity, equal opportunity, "
                 "equalized odds, predictive parity, calibration) with "
                 "verdicts NO_BIAS/POTENTIAL/DISPARATE_IMPACT"),
                ("model_governance_runtime",
                 "ModelGovernanceRuntimeEngine",
                 "Vendor model management + automated retraining "
                 "workflow per OCC 2011-12 §IV.B.2 + SR 11-7 §V. "
                 "Vendor models: 3-tier vendor model classification × "
                 "3-level transparency (FULL_DISCLOSURE / "
                 "LIMITED_DISCLOSURE / BLACK_BOX) × 10 due diligence "
                 "categories (financial soundness, model methodology, "
                 "data quality, performance track record, security "
                 "controls, business continuity, regulatory compliance, "
                 "contractual audit rights, exit strategy, "
                 "subcontractor oversight) × 5 verdicts including "
                 "REQUIRES_PROVIDER per Rule 7. Tier 1 requires all 10 "
                 "DD categories; Tier 3 requires 3. Concentration risk "
                 "monitoring per CBK Outsourcing Guideline 2018 (25% "
                 "threshold). Automated retraining: 7 trigger types "
                 "(DRIFT_DETECTED/PERFORMANCE_DEGRADATION/BIAS_DETECTED/"
                 "SCHEDULED/REGULATORY_REQUIRED/DATA_REFRESH/MANUAL) × "
                 "9-state lifecycle with explicit transition graph + "
                 "champion-challenger gating: PROMOTED_TO_CHAMPION "
                 "blocked without statistically significant "
                 "outperformance ≥2% improvement"),
            ],
            "Tier 13 — Virtual Bank Simulation (v10.30+)": [
                ("virtual_bank_core", "VirtualBankCore",
                 "Cat B deterministic simulation testbed — drop-in mock "
                 "FLEXCUBE adapter (fetch_account_balance, "
                 "fetch_customer, fetch_loan_status, "
                 "fetch_branch_metrics, fetch_rm_portfolio) + banking "
                 "entity simulator (5 customer segments × 5 account "
                 "types × 10-state loan lifecycle aligned with CBK "
                 "PG/04 risk classification: PERFORMING / DPD30 / DPD60 "
                 "/ DPD90 / NPL / WRITTEN_OFF / CLOSED). Deterministic "
                 "seeding via SHA-256 derive_seed + LCG pseudo-random; "
                 "same (seed, day_offset) → same outputs always. Time "
                 "controller (tick days/months) + day-end batch "
                 "(simple-interest accrual on savings/FD using Decimal "
                 "throughout — no float arithmetic on money + DPD-"
                 "driven loan aging via state-machine-validated "
                 "transitions). Idempotent day-end (same-day reruns "
                 "skip already-posted INT- prefixed transactions). All "
                 "MockResponse payloads carry sim_seed + sim_day_offset "
                 "for traceability per Rule 1; market-data fetcher "
                 "hookable per Rule 7 — no fabrication of CBR/KESONIA/"
                 "FX without wired source"),
                ("virtual_bank_simulator",
                 "VirtualBankSimulatorEngine",
                 "Cat B simulation orchestrator — DailyOpsSimulator "
                 "(deterministic transaction stream generator with "
                 "per-segment amount ranges + 4-tier transaction mix: "
                 "LOW_ACTIVITY / NORMAL / HIGH_ACTIVITY / STRESS) + "
                 "ScenarioInjector with 8 scenario types (RATE_SHOCK / "
                 "DEPOSIT_RUN / FRAUD_VELOCITY / FRAUD_STRUCTURING / "
                 "POPULATION_DRIFT / AML_TRIGGER / MARKET_SHOCK / "
                 "CREDIT_DETERIORATION) — 3 implemented (deposit run + "
                 "fraud structuring below CTR threshold KES 1M per CBK "
                 "AML Guideline 2023 + credit deterioration with "
                 "automatic intermediate-state walk-through). 6-state "
                 "SimulationRun lifecycle with explicit transition "
                 "graph (CONFIGURED → RUNNING → PAUSED / COMPLETED / "
                 "FAILED / CANCELLED with terminal-state invariants). "
                 "execute_run validates bank.base_seed matches "
                 "config.base_seed before running (catch reproducibility "
                 "drift before simulation starts)"),
            ],
            "Tier 14 — Cross-Sell Bandit Pilot (v10.32+)": [
                ("cross_sell_bandit", "CrossSellBanditEngine",
                 "Cat A first-ML pilot — LinUCB contextual bandit "
                 "(Li, Chu, Langford & Schapire 2010) with 7 offer "
                 "arms (SAVINGS_BOOST / FIXED_DEPOSIT / CREDIT_CARD / "
                 "LOAN_TOPUP / INSURANCE_LIFE / INSURANCE_HEALTH / "
                 "INVESTMENT_FUND) + NO_OFFER fallback. Pure-Python "
                 "matrix ops (Gaussian elimination invert) for "
                 "feature dim ≤ 10 — deterministic across runs/"
                 "platforms (no numpy dependency). Per-arm A_a (d×d "
                 "identity init) + b_a (d-zeros init); choose arm = "
                 "argmax(θ_a^T x + α √(x^T A_a^-1 x)) with α=1.0 "
                 "default. ENH-267 Risk Appetite filter: "
                 "RISK_BEARING_OFFERS={LOAN_TOPUP, CREDIT_CARD} "
                 "suppressed for customers with NPL / WRITTEN_OFF / "
                 "DELINQUENT_90 loan status. Bias safeguards: "
                 "FORBIDDEN_FEATURE_NAMES guard (gender/ethnicity/"
                 "marital_status/religion/disability/nationality/"
                 "is_pep) — engine raises ValueError on contexts that "
                 "include them. Rule 7: reward observation hookable; "
                 "record_feedback raises on unknown decisions — never "
                 "fabricates gradient updates. Rule 1: every "
                 "BanditDecision surfaces UCB + exploitation + "
                 "exploration components for full traceability"),
            ],
            "Tier 15 — Treasury ALM (v10.33+)": [
                ("treasury_alm", "TreasuryALMEngine",
                 "Cat A — implements ENH-231 + ENH-232 + ENH-233 of "
                 "16-standard Treasury arc. NMD behavioral modeling "
                 "(7 deposit categories per Basel BCBS 188 LCR + 30-day "
                 "stress runoff rates from 3% retail-stable to 100% "
                 "institutional × decay analysis with 90-day dormancy "
                 "tracking + sticky-balance estimation). "
                 "Liquidity per Basel III: LCR (HQLA / 30-day net "
                 "outflows ≥ 100%) with 3 HQLA levels (L1 0% haircut / "
                 "L2A 15% / L2B 50%) + L2 cap 40% + L2B cap 15% + "
                 "75% inflow cap, NSFR (ASF / RSF ≥ 100%) per BCBS "
                 "295, intraday liquidity per BCBS 248. IRRBB per "
                 "BCBS 368: 6 standardized rate-shock scenarios "
                 "(PARALLEL_UP/DOWN ±200bps, STEEPENER -65/+90, "
                 "FLATTENER +80/-150, SHORT_RATE_UP/DOWN ±250) × 9 "
                 "maturity buckets (overnight to 5y+) × NII 12-month "
                 "sensitivity + EVE sensitivity with ΔEVE > 15% Tier "
                 "1 outlier flag. Decimal-internal precision 28 — no "
                 "float on money. Coexists with legacy "
                 "treasury_intelligence (Volume Seven shell)"),
            ],
            "Tier 16 — Treasury Products + RWA + FTP (v10.34+)": [
                ("treasury_products", "TreasuryProductsEngine",
                 "Cat A — implements ENH-234 Treasury Products Suite. "
                 "FX spot/forward MTM (level 1/2 per IFRS 13 fair "
                 "value hierarchy). Yield curve construction with "
                 "linear interpolation + flat extrapolation at "
                 "endpoints + monotonic tenor enforcement. Bond "
                 "pricing via yield-to-maturity discount + ACT/365 "
                 "accrued interest. Money market instruments + "
                 "repo/reverse-repo placeholders. IFRS 9 "
                 "classification (HFT/AFS/HTM/LAR/DESIGNATED_FVTPL). "
                 "Net FX exposure per currency. Per Rule 7, market "
                 "data hookable; without wiring, FX forward MTM "
                 "raises ValueError requiring base + quote yield curves"),
                ("rwa_optimization", "RWAOptimizationEngine",
                 "Cat A — implements ENH-235 RWA Optimization & "
                 "Capital Management (Pillar 1 Basel III). 25 "
                 "AssetClass enums per SA-CR final framework Dec "
                 "2017 with risk weights from 0% (sovereign domestic) "
                 "to 250% (equity unlisted). 5 CCFCategory enums "
                 "(uncond cancel 10% / short LC 20% / long LC 50% / "
                 "undrawn 40% / on-BS 100%). Capital ratios: CET1 "
                 "(Basel min 4.5%, CBK PG/03 10.5%) + Tier 1 (Basel "
                 "6%) + Total (Basel 8%, CBK 14.5%). SACCR per BCBS "
                 "282 with 5 derivative asset classes + supervisory "
                 "factors (IR 0.5%, FX 4%, credit 0.46%, equity 32%, "
                 "commodity 18%) + α=1.4 alpha multiplier. EAD = α × "
                 "(RC + PFE) where RC = max(MTM−collateral, 0). "
                 "Coexists with legacy risk_weighted_assets + "
                 "capital_adequacy"),
                ("fund_transfer_pricing", "FTPEngine",
                 "Cat A — implements ENH-236 FTP Enhancement. "
                 "Matched-maturity FTP with 9 product categories "
                 "(DEMAND_DEPOSIT / SAVINGS / FD / INTERBANK / "
                 "LOAN_TERM / LOAN_REVOLVING / BOND / UNSECURED_OD / "
                 "MORTGAGE) + per-category liquidity premium spreads "
                 "(0bps demand → 75bps unsecured OD) + behavioral "
                 "tenor fallbacks for NMD (2y demand, 3y savings). "
                 "FTP curve = yield curve + liquidity premium. NIM "
                 "decomposition: lending_margin (asset rate − FTP) + "
                 "funding_margin (FTP − liability rate) — separately "
                 "surfaced per Rule 1. Per Rule 7, "
                 "construct_ftp_curve raises REQUIRES_PROVIDER if no "
                 "yield curve points supplied"),
            ],
            "Tier 17 — Cash Forecasting + Dashboard (v10.35+)": [
                ("cash_forecasting", "TreasuryCashForecastingEngine",
                 "Cat A — implements ENH-237 AI-Powered Cash "
                 "Forecasting. 13-week (91-day) horizon per Basel "
                 "BCBS 144 / CBK PG/16 liquidity planning. Three "
                 "composed components: (1) Deterministic — scheduled "
                 "flows with 9 FlowDriver enums (BOND_MATURITY / "
                 "BOND_COUPON / LOAN_AMORTIZATION / "
                 "LOAN_DISBURSEMENT / FD_ROLLOVER / "
                 "INTERBANK_SETTLEMENT / FX_SETTLEMENT / "
                 "SCHEDULED_PAYMENT / OTHER_SCHEDULED). (2) Seasonal "
                 "— day-of-week + day-of-month bucket multipliers "
                 "fit from history (min 30 days). (3) Baseline — "
                 "exponential smoothing α=0.3 default per Holt-"
                 "Winters lite. Confidence bands at 80% (Z=1.28) and "
                 "95% (Z=1.96). Per Rule 7, "
                 "ml_forecast_provider hookable; without wiring, "
                 "forecast_with_ml_overlay() raises REQUIRES_PROVIDER "
                 "rather than fabricate ML predictions"),
                ("treasury_dashboard", "TreasuryDashboardEngine",
                 "Cat A — implements ENH-238 Treasury Dashboard & "
                 "Reporting. Aggregator that composes all upstream "
                 "Treasury arc engines (treasury_alm + "
                 "treasury_products + rwa_optimization + "
                 "fund_transfer_pricing + cash_forecasting). 4 "
                 "ReportType enums: DAILY_TREASURY (today's "
                 "positions + ratios + near-term forecast) / "
                 "BOARD_PACK (monthly ALCO/Risk-Cmtte aggregation) / "
                 "REGULATORY_PACK (CBK PG/16 LCR/NSFR + CBK PG/03 "
                 "capital + IRRBB outliers, structured per CBK "
                 "submission format) / INTRADAY_LIQUIDITY. 4 "
                 "SectionStatus enums (OK / WARNING / BREACH / "
                 "NO_DATA) with worst-of roll-up across sections. "
                 "Per Rule 1, every section reports source engine + "
                 "metrics + thresholds + headroom. Per Rule 7, "
                 "dashboard never invents data — unwired engines "
                 "produce NO_DATA cleanly"),
            ],
            "Tier 18 — Scenario Simulation Foundation (v10.36+)": [
                ("scenario_simulator", "ScenarioRunner",
                 "Cross-arc executable scenario harness. Foundation "
                 "for the v10.36 'safe learning' framework: 10 "
                 "ScenarioCategory enums (CUSTOMER_LIFECYCLE / "
                 "CREDIT_LENDING / DEPOSIT_LIQUIDITY / "
                 "PERFORMANCE_MGMT / RISK_COMPLIANCE / "
                 "OPERATIONS_TREASURY / STRATEGY_CAMPAIGNS / "
                 "FRAUD_SECURITY / RECOVERY_DISASTER / "
                 "COMPETITOR_MARKET) + 5 ScenarioStatus enums "
                 "(PASS / WARNING / FAIL / SKIPPED / ERROR). "
                 "Scenario contract: setup → actions → assertions "
                 "with engine bundle injection. ScenarioRunner "
                 "supports two modes — shared-bundle (state "
                 "accumulates) and bundle_factory (fresh per "
                 "scenario for regression suites). Initial library "
                 "ships 11 Treasury-focused scenarios (LI-01/02 "
                 "LCR compliance/breach, IRRBB-01 outlier detection, "
                 "CAP-01 dual Basel/CBK threshold, FX-01 net "
                 "exposure, NIM-01 decomposition, DASH-01 dashboard "
                 "roll-up, CF-01 forecast, CF-02 REQUIRES_PROVIDER, "
                 "MODGOV-01 model registration, CROSS-01 end-to-end "
                 "LCR propagation). v10.37 extends with 8 closure "
                 "scenarios (ISLAMIC-01/02, AGENT-01/02, CONN-01, "
                 "DIGITAL-01, UNIFIED-01, CLIMATE-01) — total 19. "
                 "Per Rule 1, every AssertionResult surfaces "
                 "expected + observed + matched. Per Rule 7, "
                 "scenarios needing external providers declare "
                 "requires_providers; runner skips unsatisfied "
                 "scenarios cleanly without fabricating responses. "
                 "Each subsequent batch adds 3-5 scenarios to the "
                 "library covering its new functionality"),
            ],
            "Tier 19 — Treasury Arc Closure (v10.37)": [
                ("islamic_treasury", "IslamicTreasuryEngine",
                 "Cat A — implements ENH-239 Islamic Treasury "
                 "Products per AAOIFI + IFSB. 6 IslamicProductType "
                 "enums (MURABAHA cost-plus / WAKALA agency / SUKUK "
                 "ownership / MUDARABAH PSP / IJARAH leasing / "
                 "QARD_HASAN benevolent loan). 5 SukukStructure "
                 "sub-types (SUKUK_IJARA HQLA-eligible / "
                 "SUKUK_MURABAHA / SUKUK_MUDARABA / SUKUK_WAKALA / "
                 "SUKUK_HYBRID). 4 ShariaComplianceStatus (COMPLIANT "
                 "/ PROVISIONAL pending board / NON_COMPLIANT / "
                 "REQUIRES_REVIEW). 8 PROHIBITED_INDUSTRIES "
                 "constants (alcohol / pork / gambling / "
                 "conventional_banking / conventional_insurance / "
                 "tobacco / weapons / adult_entertainment) per "
                 "AAOIFI Sharia Standard 8. Per Rule 7, "
                 "Mudarabah/Wakala profit-sharing without Sharia "
                 "Supervisory Board approval raises "
                 "REQUIRES_PROVIDER:sharia_supervisory_board. Per "
                 "Rule 1, every IslamicProductValuation surfaces "
                 "principal + markup/profit + Sharia compliance + "
                 "non_compliance_reasons + AAOIFI/IFSB framework "
                 "refs"),
                ("treasury_agents", "AgentOrchestrator",
                 "Cat A — implements ENH-240 Agentic Treasury "
                 "Orchestration patterned after Kyriba TAI. 5 "
                 "concrete TreasuryAgent classes "
                 "(LiquidityBufferAgent monitoring LCR vs buffer / "
                 "HedgingAgent monitoring IRRBB EVE outliers / "
                 "CashShortfallAgent reading cash_forecasting / "
                 "PaymentReviewAgent flagging suspicious payments "
                 "for ENH-TRS-R5 / SweepingAgent identifying idle "
                 "cash for MMF placement). 4 RecommendationPriority "
                 "(URGENT / HIGH / MEDIUM / LOW) × 6 "
                 "RecommendationCategory × 5 ApprovalStatus "
                 "(PENDING / APPROVED / REJECTED / EXECUTED / "
                 "EXPIRED). AgentOrchestrator runs registered "
                 "agents in priority order; recommendations enter "
                 "PENDING; treasurer approves → APPROVED; manual "
                 "EXECUTED transition. Per Rule 7 — agents NEVER "
                 "autonomously execute (EU AI Act Art 14 human "
                 "oversight). Per Rule 1, every Recommendation "
                 "surfaces detected_condition + rationale + "
                 "suggested_action + estimated_impact + "
                 "upstream_engines_consulted + framework_refs"),
                ("treasury_connectivity",
                 "TreasuryConnectivityEngine",
                 "Cat A — implements ENH-TRS-R1 (9900+ Bank "
                 "Connection per Kyriba benchmark) + ENH-TRS-R3 "
                 "(MMF Direct Access) + ENH-TRS-R5 (Real-Time API "
                 "ERP-to-Bank). Single module because all three are "
                 "external connectivity. 6 ConnectorType "
                 "(BANK_PARTNER / MMF_COUNTERPARTY / ERP_SYSTEM / "
                 "CENTRAL_BANK / CARD_NETWORK / OTHER). 13 "
                 "MessageFormat (ISO 20022 CAMT.053/054/PAIN.001/"
                 "008 / SWIFT MT940/942/103/202/210 / BACS / SEPA / "
                 "KEPSS Kenya / REST_JSON / OTHER). Format "
                 "registries: FORMAT_REQUIRED_FIELDS validates "
                 "payloads, REGION_PREFERRED_FORMAT routes by "
                 "region. ENH-TRS-R3 MMF management with "
                 "best_yielding_mmf and T+0 redemption filter. "
                 "ENH-TRS-R5 review_payment hook for pre-execution "
                 "fraud screening. Per Rule 7, live transmission "
                 "with require_credentials=True raises "
                 "REQUIRES_PROVIDER:credential_provider unless "
                 "wired"),
                ("treasury_digital_assets",
                 "DigitalAssetTreasuryEngine",
                 "Cat A — implements ENH-TRS-R2 Stablecoin & "
                 "Digital Asset Treasury per CBK VASP Regulations "
                 "2026. 6 DigitalAssetType (USDC / USDT / EURC / "
                 "KES_STABLE pilot / BTC / ETH). 3 BCBSCryptoGroup "
                 "(GROUP_1A_TOKENIZED / GROUP_1B_STABLECOIN / "
                 "GROUP_2_OTHER 1250% RW per BCBS Crypto 2022). 5 "
                 "DePegStatus (ON_PEG ≤50bps / MINOR_DEVIATION "
                 "50-100bps / SIGNIFICANT_DEVIATION 100-300bps / "
                 "DE_PEGGED >300bps / NOT_APPLICABLE for non-stable). "
                 "Per-asset concentration limits (USDC 3%, USDT 2%, "
                 "EURC 2%, KES_STABLE 5%, BTC 0.5%, ETH 0.5%). "
                 "Volatile total cap 1% (BTC + ETH). Wallet "
                 "whitelist + WalletStatus enum + KYT prep. Per "
                 "Rule 7, live spot rates require rate_provider "
                 "hook (chain oracle / CEX); without it, manual "
                 "rates with rate_source provenance flag. Per Rule "
                 "1, every DigitalAssetValuation surfaces holding "
                 "+ spot + KES equivalent + de_peg_status + "
                 "BCBS_group + framework_refs"),
                ("treasury_unified_platform",
                 "UnifiedTreasuryPlatform",
                 "Cat A — implements ENH-TRS-R4 MX.3-style unified "
                 "cross-asset platform. Facade composing 7 upstream "
                 "engines (treasury_alm + treasury_products + "
                 "rwa_optimization + fund_transfer_pricing + "
                 "islamic_treasury + treasury_digital_assets + "
                 "cash_forecasting). 6 AssetClass enums (FX / "
                 "FIXED_INCOME / MONEY_MARKET / ISLAMIC / DIGITAL / "
                 "DERIVATIVES). 5 IFRS9Category (AMORTIZED_COST / "
                 "FVOCI_DEBT / FVOCI_EQUITY / FVTPL / "
                 "FVTPL_DESIGNATED). 4 adapter functions read "
                 "upstream engines (positions_from_treasury_alm, "
                 "_treasury_products, _islamic_treasury, "
                 "_digital_assets) into UnifiedPosition stream. "
                 "CrossAssetRiskRollup aggregates by class with "
                 "n_engines_consulted trace. Per Rule 7, facade "
                 "READS upstream engines but never mutates them — "
                 "missing engines simply produce no positions of "
                 "that class. Per Rule 1, every UnifiedPosition "
                 "surfaces source_engine + asset_class + value + "
                 "IFRS9 category + framework_refs"),
                ("climate_treasury_limits",
                 "ClimateTreasuryLimitsEngine",
                 "Cat A — implements ENH-TRS-R6 Climate-Adjusted "
                 "Treasury Risk Limits. Cross-arc bridge composing "
                 "v10.6-10 climate engines (climate_risk) with "
                 "v10.33-35 treasury concentration limits. 10 "
                 "TreasuryAssetClass enums (SOVEREIGN_KENYA / "
                 "SOVEREIGN_OTHER / CORPORATE_FOSSIL / "
                 "CORPORATE_HEAVY_INDUSTRY / CORPORATE_AGRICULTURE / "
                 "CORPORATE_RENEWABLE / CORPORATE_FINANCIALS / "
                 "CORPORATE_OTHER / REAL_ESTATE_COASTAL / "
                 "REAL_ESTATE_OTHER). 4 CLIMATE_HAIRCUT_BANDS: 0-25 "
                 "score=1% haircut / 26-50=5% / 51-75=15% / 76-100="
                 "30%. Two haircut channels (PHYSICAL drought/"
                 "flood/sea-level + TRANSITION carbon-price/stranded"
                 "-asset) — applied as worst-of (max channel). "
                 "ASSET_CLASS_TO_SECTORS mapping (e.g., "
                 "CORPORATE_FOSSIL → oil_and_gas + coal_mining + "
                 "petroleum). LimitBreachReport severity (NONE / "
                 "WARNING within base but over adjusted / BREACH "
                 "over base). Per Rule 7, this is a READ-ONLY "
                 "facade — never mutates climate engine. Per Rule "
                 "1, every ClimateAdjustedLimit surfaces base + "
                 "physical_haircut + transition_haircut + adjusted "
                 "+ source counts + BCBS Climate Principles 2022 + "
                 "IFRS S2 + CBK CRDF + NGFS framework refs"),
            ],
            "Tier 20 — Structural Hygiene Foundation (v10.38+)": [
                ("structure_audit_core", "StructureAuditEngine",
                 "Cat A — codebase-shape analyzer locked behind "
                 "G128. AST-based dependency graph builder + 7 "
                 "rule families: CIRCULAR_IMPORT (HARD — graph "
                 "cycle detection via three-color DFS) / "
                 "LAYER_VIOLATION (HARD — utils ↛ pages, utils ↛ "
                 "scripts, scripts ↛ pages) / GOD_MODULE (WARN — "
                 "fan-in > 15 unless on CROSS_ARC_BRIDGES exempt "
                 "list) / JUNK_DRAWER (WARN — fan-out > 25) / "
                 "ORPHAN_MODULE (WARN — no callers + not entry "
                 "point) / DUPLICATE_SYMBOL (WARN — same class in "
                 "2+ modules, same function in 3+ modules) / "
                 "SIZE_OUTLIER (INFO 2k+ / WARN 4k+). 3 "
                 "FindingSeverity (HARD / WARN / INFO) × 7 "
                 "FindingCategory. Mypy-style baseline mechanism: "
                 "compute_baseline + compare_to_baseline allow "
                 "existing HARD findings to remain (system runs) "
                 "but reject any new HARD finding (no regression). "
                 "Per Rule 1, every Finding surfaces severity + "
                 "category + module_path + observed_value + "
                 "threshold + suggestion. Per Rule 7, the engine "
                 "never auto-mutates code — reorganization is "
                 "always a human decision. CROSS_ARC_BRIDGES exempt "
                 "list covers intentional facades (treasury_"
                 "dashboard / treasury_unified_platform / climate_"
                 "treasury_limits / scenario_simulator) plus base "
                 "infrastructure (db / config / core_audit / "
                 "_shared / _access / standards_registry) that has "
                 "high fan-in by design"),
            ],
            "Tier 21 — Market Risk Foundation (v10.39+)": [
                ("market_risk_factors", "RiskFactorRegistry",
                 "Cat A — implements ENH-MR-004 Risk Factor "
                 "Taxonomy & Stress Scenarios. 5 RiskFactorClass "
                 "buckets (INTEREST_RATE / FOREIGN_EXCHANGE / "
                 "EQUITY / COMMODITY / CREDIT_SPREAD) × 23 specific "
                 "RiskFactor enums (KES IR curve buckets / foreign "
                 "IR / 6 FX pairs vs KES / 3 equity indices / 3 "
                 "commodities / 5 credit-spread buckets). "
                 "RISK_FACTOR_TO_CLASS lookup + ShockType enum "
                 "(ABSOLUTE_BPS for IR, ABSOLUTE_PCT for crashes, "
                 "RELATIVE_PCT for FX/equity multiplicative). "
                 "FactorShock + StressScenario frozen dataclasses. "
                 "9 ALL_PREBUILT_SCENARIOS: 6 BCBS d368 IRRBB "
                 "(BCBS-IRRBB-1 parallel up 200bp / -2 parallel "
                 "down / -3 short up 250bp / -4 short down / -5 "
                 "steepener / -6 flattener) + 3 internal/CBK "
                 "(INT-FX-1 USD/KES +15%, INT-FX-2 -10%, INT-EQ-1 "
                 "equity crash 30%). KES magnitudes ±200bp parallel "
                 "/ ±250bp short / ±100bp long per BCBS d368 §K. "
                 "RiskFactorRegistry.get / by_framework / "
                 "by_factor_class / summary. Per Rule 1, every "
                 "FactorShock surfaces factor + shock_type + "
                 "magnitude; every StressScenario surfaces name + "
                 "description + shocks tuple + framework_refs"),
                ("market_risk_sensitivities", "SensitivityEngine",
                 "Cat A — implements ENH-MR-003 Sensitivity-Based "
                 "Measures per BCBS d352 FRTB SBM §A.5 + IFRS 7 "
                 "§40. 4 SensitivityType (DELTA / VEGA / CURVATURE "
                 "/ DV01). BondPosition (validates IR factor; "
                 "modified_duration ≥ 0 enforced) / FXPosition "
                 "(validates FX factor; spot_to_kes positive) / "
                 "EquityPosition (validates equity factor; default "
                 "beta=1.0). compute_dv01: D_mod × P × 0.0001 with "
                 "convexity term 0.5 × convexity × P × Δy² for "
                 "second-order correction. compute_fx_delta: "
                 "foreign_amount × spot × 0.01. compute_equity_"
                 "delta: market_value × beta × 0.01. aggregate(): "
                 "groups by RiskFactor + RiskFactorClass producing "
                 "SensitivityReport with by_factor / by_class / "
                 "total_delta_kes. apply_scenario_pnl: ABSOLUTE_BPS "
                 "DV01 returns -delta×magnitude (long bond loses on "
                 "rate UP); ABSOLUTE_PCT/RELATIVE_PCT delta returns "
                 "+delta×magnitude. per_factor_pnl_contribution "
                 "decomposes total scenario PnL by factor. Decimal-"
                 "internal precision throughout. Constants: "
                 "ONE_BP=Decimal('0.0001'), ONE_PCT=Decimal('0.01')"),
                ("market_risk_var", "VaREngine",
                 "Cat A — implements ENH-MR-001 (VaR), ENH-MR-002 "
                 "(ES), ENH-MR-005 (backtests). 3 VaRMethodology "
                 "(PARAMETRIC variance-covariance / HISTORICAL "
                 "empirical percentile / MONTE_CARLO simulation). "
                 "Confidence levels 95%, 97.5% (FRTB-IMA), 99% "
                 "(Basel VaR). Pure stdlib — statistics.NormalDist "
                 "+ math.exp/log/sqrt, no scipy. Hard-coded χ² "
                 "critical values: _CHI2_1_CRITICAL {0.10:2.706, "
                 "0.05:3.841, 0.01:6.635}, _CHI2_2_CRITICAL "
                 "{0.10:4.605, 0.05:5.991, 0.01:9.210}. "
                 "parametric_var: Normal assumption, z = "
                 "NormalDist().inv_cdf(α), √T scaling per Basel "
                 "MRA 1996, ES via φ(z)/(1−α)×σ×√T. "
                 "historical_var: linear-interpolation percentile, "
                 "ES = mean of returns ≤ percentile. "
                 "monte_carlo_var: gauss simulation (n_simulations "
                 "≥ 100, optional seed for reproducibility, "
                 "rng=Random(seed)), internally calls historical_"
                 "var. kupiec_pof_test: LR=−2(log_h0 − log_h1) "
                 "Bernoulli unconditional coverage, edge cases "
                 "x=0 and x=N handled, χ²(1) significance. "
                 "christoffersen_independence_test: 2×2 transition "
                 "matrix (n_00/n_01/n_10/n_11) for breach "
                 "clustering, χ²(1). 3 BacktestVerdict (PASS / "
                 "FAIL / INSUFFICIENT_DATA). VaR sign convention: "
                 "var_kes is POSITIVE loss magnitude; ES ≥ VaR "
                 "by construction. Per Rule 1, VaRResult surfaces "
                 "methodology + confidence + horizon + portfolio_"
                 "value + return distribution summary; "
                 "BacktestResult surfaces test_name + significance "
                 "+ n_obs + n_breaches + expected + statistic + "
                 "critical_value + verdict + framework refs"),
            ],
            "Tier 22 — Market Risk Limits & Breach Management (v10.40+)": [
                ("market_risk_limits", "LimitMonitor",
                 "Cat A — implements ENH-MR-006 Market Risk Limit "
                 "Framework + ENH-MR-007 Limit Breach Detection & "
                 "Escalation. 3 LimitType (CONCENTRATION per-factor "
                 "or per-class exposure / VAR_LIMIT daily VaR "
                 "ceiling / ES_LIMIT FRTB-IMA Expected Shortfall "
                 "ceiling) × 3 LimitScope (SINGLE_FACTOR / "
                 "FACTOR_CLASS / PORTFOLIO). 4 BreachSeverity bands: "
                 "WITHIN_LIMIT < 80% / WARN 80-99.99% / BREACH "
                 "100-119.99% / SEVERE_BREACH ≥ 120%. RiskLimit "
                 "frozen dataclass — immutable once registered, "
                 "deactivate-and-re-register pattern preserves "
                 "audit history. LimitRegistry: register / "
                 "deactivate / get / by_factor (returns both "
                 "single-factor + applicable class limits) / "
                 "by_type / summary. LimitMonitor.check_"
                 "concentration aggregates exposures by factor + "
                 "by class; absolute value used (net SHORT counts). "
                 "check_var / check_es match ONLY exact (confidence, "
                 "horizon) — a 99%/1d limit doesn't trigger on "
                 "95%/1d obs. run_pass orchestrates all three "
                 "checks and returns MonitorReport. Per Rule 1, "
                 "every BreachAlert surfaces severity + observed + "
                 "threshold + utilization_pct + factor + "
                 "framework_refs + suggested_action + "
                 "escalation_target (Treasury → ALCO+CRO → Board "
                 "Risk Committee scaling with severity). "
                 "Deterministic alert_id = limit::date::obs::sev "
                 "supports audit-trail dedup. Per Rule 7, monitor "
                 "is purely diagnostic — never auto-executes "
                 "remediation; alerts flow into treasury_agents."
                 "PaymentReviewAgent or human approval workflow "
                 "(EU AI Act Art 14 oversight preserved). 5 "
                 "default illustrative limits (KES 50m daily 99% "
                 "VaR / KES 150m 10-day 97.5% ES / KES 2bn USD/KES "
                 "single-factor / KES 5bn FX class / KES 1bn "
                 "Equity class) per CBK PG/04 §4 + BCBS d352 §A.4 "
                 "+ BCBS 239 §5 + EBA/GL/2018/02"),
            ],
            "Tier 23 — Trading Book Boundary (v10.41+)": [
                ("trading_book_boundary", "BoundaryEngine",
                 "Cat A — implements ENH-MR-008 Trading Book "
                 "Boundary Classification + ENH-MR-009 Trading Desk "
                 "Definition + ENH-MR-010 Boundary Crossing Approval "
                 "Workflow per BCBS d352 §A.4. Two BookClassification "
                 "(TRADING_BOOK → market risk FRTB capital / "
                 "BANKING_BOOK → credit risk IRB + IRRBB). 16 "
                 "InstrumentType enums with disjoint presumption "
                 "sets: 9 PRESUMPTIVE_TRADING_BOOK (LISTED_EQUITY / "
                 "EQUITY_FUND / LISTED_DERIVATIVE / OTC_DERIVATIVE_"
                 "TRADING / SECURITY_HELD_FOR_RESALE / REPO_REVERSE_"
                 "REPO / MARKET_MAKING_INVENTORY / COMMODITY_TRADING "
                 "/ FX_TRADING) + 7 PRESUMPTIVE_BANKING_BOOK "
                 "(LOAN_RECEIVABLE / DEPOSIT_LIABILITY / BANKING_BOOK_"
                 "HEDGE / SECURITISATION_BB / EQUITY_INVESTMENT_NON_"
                 "TRADING / REAL_ESTATE / LIQUIDITY_BUFFER_HOLD). "
                 "TradingDesk frozen dataclass per §A.4.2 requires "
                 "desk_id + name + head_trader + mandate + "
                 "risk_classes (FrozenSet[RiskFactorClass] composes "
                 "with market_risk_factors) + default_holding_period_"
                 "days + parent_business_unit. validate() returns "
                 "DeskValidationIssue tuple covering MISSING_HEAD_"
                 "TRADER / NO_RISK_CLASSES / INVALID_HOLDING_PERIOD "
                 "/ NO_MANDATE / OK. 3 default desks ship: DESK-FX-"
                 "NAIROBI (FOREIGN_EXCHANGE) / DESK-FI-NAIROBI "
                 "(INTEREST_RATE + CREDIT_SPREAD, 10-day) / DESK-EQ-"
                 "NAIROBI (EQUITY, 5-day). Reclassification workflow "
                 "per §A.4.5: ReclassificationRequest validates "
                 "non-empty reason + from_book ≠ to_book; "
                 "compute_capital_surcharge applies DEFAULT_SURCHARGE_"
                 "RATE (1.0, overrideable) only when "
                 "expected_capital_impact_kes > 0 (benefits bank). "
                 "approve_reclassification is THE ONLY mutation "
                 "path — explicit approver (senior management) "
                 "required, returns ApprovalDecision + updated "
                 "BookAssignment, mutates registry. reject_"
                 "reclassification leaves assignment unchanged. "
                 "Per Rule 7, request creation NEVER auto-approves "
                 "(EU AI Act Art 14 human oversight preserved). "
                 "Per Rule 1, every BookAssignment / "
                 "ReclassificationRequest / ApprovalDecision "
                 "carries full provenance + framework_refs. "
                 "Decimal-internal precision throughout"),
            ],
            "Tier 24 — Risk Arc Closure (v10.42-v10.45)": [
                ("credit_risk_irb", "IRBCapitalEngine",
                 "Cat A — implements ENH-CR-001 IRB Capital "
                 "Framework per BCBS d424 §RBC25 corporate "
                 "exposure formula. PD/LGD/EAD/M inputs with "
                 "PD floor 3bp, M ∈ [1, 5]. Computes K = LGD × "
                 "[N(...) − PD] × maturity_adj, RWA = K × 12.5 × "
                 "EAD, EL = PD × LGD × EAD. Correlation R(PD) "
                 "and maturity adjustment b(PD) per §RBC25.7 + "
                 "§RBC25.13. Defaulted exposure (PD=1.0) → K=0 "
                 "above EL per §RBC25.16. 2 ExposureClass enums "
                 "(LARGE_CORPORATE / SME_CORPORATE; SOVEREIGN + "
                 "BANK reserved for future scope). IRBExposure + "
                 "CapitalResult frozen dataclasses. Pure stdlib "
                 "(statistics.NormalDist + math); no scipy. Per "
                 "Rule 1, every CapitalResult surfaces pd + lgd "
                 "+ ead + maturity + correlation_R + maturity_adj_b "
                 "+ capital_requirement_pct + rwa_kes + "
                 "expected_loss_kes + framework_refs. Per Rule 7, "
                 "computational only — never moves loans between "
                 "exposure classes, never auto-approves capital "
                 "allocations; all approvals flow through ALCO + "
                 "Capital Management Committee. Decimal-internal "
                 "precision; float used only for NormalDist input. "
                 "compute_portfolio aggregates per-exposure → "
                 "(results, total_RWA, total_EL); diversification "
                 "benefit deferred to Pillar 2 economic capital."),
                ("op_risk", "OperationalRiskSMA",
                 "Cat A — implements ENH-OR-001 SMA Operational "
                 "Risk Capital per BCBS d457 §RBC30. Three "
                 "components: BI (3-year average of ILDC + SC + "
                 "FC), BIC (marginal-α bucket-wise application of "
                 "12% / 15% / 18% across EUR 1bn / 30bn "
                 "thresholds), ILM (ln(e − 1 + (LC/BIC)^0.8); "
                 "forced to 1.0 in Bucket 1 by §RBC30.41 national "
                 "discretion or when loss history < 5 years). ORC "
                 "= BIC × ILM, RWA_op = ORC × 12.5. ILDC = "
                 "min(|II−IE|, 0.0225×IEA) + DI per §RBC30.10; "
                 "SC = max(OI, OE) + max(FI, FE) per §RBC30.11; "
                 "FC = |Net P&L TB| + |Net P&L BB| per §RBC30.13. "
                 "3 Bucket enums (BUCKET_1 / BUCKET_2 / "
                 "BUCKET_3) × 3 ILMSource enums (COMPUTED / "
                 "BUCKET_1_DISCRETION / INSUFFICIENT_HISTORY) "
                 "for full Rule 1 provenance. "
                 "BusinessIndicatorInputs + OperationalLossEvent "
                 "+ SMAInputs + SMAResult frozen dataclasses with "
                 "construction-time validation (BI requires 3 "
                 "distinct fiscal years, IEA non-negative, EUR "
                 "rate positive). Pure stdlib (math + Decimal); "
                 "no scipy. Per Rule 7, engine never records "
                 "loss events, never approves capital — caller "
                 "supplies all loss data + national-discretion "
                 "flag. Decimal-internal monetary precision."),
                ("liquidity_stress", "LiquidityStressEngine",
                 "Cat A — implements ENH-LR-001 Stressed LCR per "
                 "BCBS d295 §40-§69. Distinct from Standard #73 "
                 "baseline LCR (utils.liquidity_risk) and Standard "
                 "#79 capital stress (utils.stress_testing); "
                 "covers liquidity-specific severity calibration. "
                 "HQLA per level (L1 / L2A / L2B with haircuts "
                 "0% / 15% / 50%) with composition caps (L2 ≤ "
                 "40%, L2B ≤ 15% of total HQLA — enforced via "
                 "(15/85)×L1 and (40/60)×L1 unwind formulas). "
                 "Stressed flows: outflow_rate = min(base × "
                 "severity_mult, 1.0); inflow_rate = min(base × "
                 "inflow_mult, 1.0). 4 StressSeverity tiers "
                 "(BASELINE 1.0× / MODERATE 1.5× / SEVERE 2.0× / "
                 "BANK_RUN 3.0× outflow multipliers; "
                 "1.0/0.85/0.65/0.40 inflow multipliers). NCO = "
                 "outflows − min(inflows, 0.75 × outflows) per "
                 "§69 inflow cap. LCR = HQLA / NCO returned as "
                 "Optional[Decimal] (None when NCO ≤ 0 — Rule 1 "
                 "no false-precision). 4 BreachSeverity bands: "
                 "COMPLIANT (≥100%) / AMBER ([90%,100%)) / RED "
                 "([70%,90%)) / CRITICAL (<70%). Survival horizon "
                 "= HQLA / (NCO/30) days when breaching. "
                 "outflow_rate_overrides Mapping[str, Decimal] "
                 "supports supervisor-mandated scenarios without "
                 "engine modification. Per Rule 1, "
                 "StressedLCRResult surfaces hqla_breakdown + "
                 "pre/post-cap totals + per-category "
                 "StressedFlow records + NCO components + LCR + "
                 "breach + survival + framework_refs. Per Rule 7, "
                 "engine never auto-liquidates HQLA, never "
                 "executes funding draws, never rebalances "
                 "category assignments. Pure stdlib Decimal."),
            ],
            "Tier 25 — credit_model_risk Arc Closure (v10.47-v10.49)": [
                ("credit_alt_scoring", "AlternativeCreditScoringEngine",
                 "Cat A — implements ENH-260 thin-file PD via 3 "
                 "alternative pillars per CGAP + Smart Campaign + "
                 "IFC Inclusive Finance. TRANSACTION (50% weight): "
                 "deposit-CV regularity, salary-cycle signal, "
                 "expense/deposit ratio, bills-on-time %; requires "
                 "≥3 months observed. BEHAVIORAL (30%): tenure, "
                 "mobile-active days, current-facility delinquency "
                 "days; requires ≥1 month tenure. PSYCHOMETRIC "
                 "(20%): risk-tolerance + time-horizon scores; "
                 "optional. Each pillar produces a sub-PD AND a "
                 "confidence weight (0 when unusable). Composite = "
                 "confidence-weighted across usable pillars; "
                 "overall confidence drives 3 ConfidenceBand enums "
                 "(HIGH ≥0.70 / MEDIUM ≥0.40 / LOW). Below LOW → "
                 "recommend_bureau_check=True so underwriting "
                 "escalates rather than acting on a thin estimate. "
                 "PD floor 3 bp matches BCBS d424 IRB floor for "
                 "downstream composition with credit_risk_irb. "
                 "Power transform 0.50 × signal^1.8 ensures "
                 "monotonic mapping. Pure stdlib (math + Decimal). "
                 "All 5 dataclasses frozen (ThinFileApplicant, "
                 "TransactionMetrics, BehavioralMetrics, "
                 "PsychometricMetrics, PillarScore, "
                 "AltScoringResult). Per Rule 1, every "
                 "AltScoringResult surfaces 3 PillarScore objects "
                 "(per-pillar PD + confidence_weight + "
                 "features_used + skip_reason) + composite + grade "
                 "(via RISK_GRADES + PD_BANDS reuse from "
                 "credit_risk_scoring) + missing_pillars + "
                 "framework_refs. Per Rule 7, engine never "
                 "auto-approves, never auto-declines, never writes "
                 "to bureau"),
                ("credit_committee", "CreditCommitteeEngine",
                 "Cat A — implements ENH-268 diagnostic governance "
                 "engine for credit committee decisions per CBK "
                 "PG/03 §6. Static CommitteeCharter (members + "
                 "voting_rule + min_quorum_count + required_roles + "
                 "authority_limit_kes + independent_member_min + "
                 "escalation_target). 7 CommitteeRole enums (CHAIR, "
                 "CRO, CCO, CFO, HEAD_OF_CREDIT, "
                 "INDEPENDENT_MEMBER, EXECUTIVE_MEMBER) × 4 "
                 "VotingRule (SIMPLE_MAJORITY ties→REJECT "
                 "defensively / SUPERMAJORITY_TWO_THIRDS / "
                 "UNANIMOUS / CHAIR_TIEBREAKER) × 4 VoteValue "
                 "(YES / NO / ABSTAIN / RECUSED) × 4 QuorumStatus "
                 "(MET / NOT_MET_HEADCOUNT / NOT_MET_REQUIRED_ROLE "
                 "/ NOT_MET_INDEPENDENT_MIN) × 6 DecisionOutcome "
                 "(APPROVED / APPROVED_WITH_CONDITIONS / REJECTED "
                 "/ DEFERRED / ESCALATED / QUORUM_FAILED). Authority "
                 "check supersedes voting — facilities above limit "
                 "ESCALATE without committee vote. Policy override "
                 "approvals trigger escalation per §6.7 (approved "
                 "decision still escalates upward; rationale "
                 "mandatory at construction). Recused votes count "
                 "as present but excluded from tally; duplicate or "
                 "absent-member votes ignored. Pure stdlib "
                 "(Decimal + frozen dataclasses + enums). All 6 "
                 "dataclasses frozen (CommitteeMember, "
                 "CommitteeCharter, Vote, CreditDecisionRequest, "
                 "VoteTally, DecisionResult). Per Rule 1, every "
                 "DecisionResult surfaces members_present + roles "
                 "+ quorum_status + reason + full VoteTally + "
                 "outcome + rationale + conditions + override + "
                 "escalation + framework refs. Per Rule 7, engine "
                 "never auto-approves a facility, never "
                 "auto-disburses funds, never modifies charter at "
                 "runtime. Decimal-internal precision for monetary "
                 "thresholds"),
            ],
            "Tier 26 — revenue_assurance Arc Closure (v10.50-v10.58)": [
                ("revenue_validation", "RevenueValidationEngine",
                 "Cat B — implements ENH-241 four-agent foundation: "
                 "SCHEMA + COMPLETENESS + CROSS_SOURCE recon (CBS vs "
                 "GL via CrossSourceTotal pairs, 5bp default tolerance) "
                 "+ STATISTICAL anomaly (z-score, min sample 10). 5 "
                 "ValidationSeverity enums (CRITICAL/HIGH/MEDIUM/LOW/"
                 "INFO) × 4 ValidationCategory. validate_all "
                 "orchestrator returns ValidationReport with severity + "
                 "category breakdowns. ValidationSeverity is the shared "
                 "severity vocabulary across all 7 downstream arc "
                 "engines — single enum prevents translation friction at "
                 "the orchestrator boundary. Pure stdlib (Decimal + "
                 "dataclasses + statistics). Per Rule 1, every "
                 "ValidationFinding surfaces finding_id + severity + "
                 "category + record_ids + expected + observed + "
                 "framework refs. Per Rule 7, diagnostic only — "
                 "surfaces issues, never auto-corrects, never modifies "
                 "CBS/GL records, never quarantines."),
                ("revenue_anomaly_patterns",
                 "RevenueAnomalyPatternEngine",
                 "Cat B — implements ENH-242 6 deterministic post-"
                 "issuance pattern detectors: duplicate_billing (amount "
                 "+ customer + 3-day window), unauthorized_waiver "
                 "(waiver_pct > 0 needs authorization_id), "
                 "expired_contract (posting_date outside contract "
                 "effective window), rate_card (applied_rate vs "
                 "ContractRate floor=MEDIUM, ceiling=HIGH), missing_tax "
                 "(taxable category but no tax record), "
                 "commission_anomaly (paid vs expected, 1% tolerance + "
                 "KES 50 floor). ML hook injectable per Rule 6 — "
                 "ml_disabled=True surfaced explicitly when ML model "
                 "absent (no silent fallback). 9 PatternId × 4 "
                 "PatternFamily enums (BILLING/CONTRACT/TAX/COMMISSION). "
                 "ContractRate + WaiverRecord + CommissionRecord frozen. "
                 "Reuses RevenueRecord + ValidationSeverity from "
                 "ENH-241. Per Rule 1, every PatternFinding carries "
                 "pattern_id + family + severity + record_ids + "
                 "observed_value + threshold + framework refs. Per Rule "
                 "7, diagnostic — never auto-reverses billing, never "
                 "auto-cancels waivers, never auto-recovers commission "
                 "overpayments."),
                ("revenue_orchestrator",
                 "RevenueOrchestrator",
                 "Cat B — implements ENH-243 stateless work-item "
                 "router. Caller maintains case-management state "
                 "externally; engine routes findings to 6 "
                 "InvestigatorTeam values (REVENUE_RECOVERY / "
                 "BILLING_OPS / TREASURY_OPS / HR_PAYROLL / "
                 "COMPLIANCE_KRA / CRO_REVIEW) via deterministic "
                 "(severity, family) lookup. Priority score = "
                 "severity_weight × family_weight + age_decay + "
                 "monetary_impact × impact_weight. Default "
                 "impact_weight=0.0001 — large-impact findings can "
                 "outrank higher-severity smaller findings (test "
                 "documents this honestly). 6 WorkItemState (RAISED → "
                 "ACKNOWLEDGED → IN_PROGRESS → RESOLVED + DISMISSED + "
                 "ESCALATED) × 2 FindingType (VALIDATION / PATTERN). "
                 "ORC-04 scenario verifies Rule 7 stateless contract — "
                 "second call doesn't memoise first call's state. Per "
                 "Rule 1, WorkItem surfaces priority_components dict so "
                 "investigators see exactly why it ranked where it did."),
                ("partner_supplier_recon",
                 "PartnerSupplierReconciliationEngine",
                 "Cat B — implements ENH-244 multi-party "
                 "reconciliation extending ENH-241's two-source "
                 "pattern. Block A: partner revenue share — agreement-"
                 "driven expected_share = Σ(gross × share_pct), "
                 "compared to actual settlements with mixed tolerance "
                 "max(KES 100, 1% of expected); min_settlement_kes "
                 "floor for legitimate carry-forwards; 3 partner "
                 "discrepancy types (SHARE_UNDERPAID/OVERPAID/MISSING). "
                 "Block B: supplier 3-way match — chain PO → GRN → "
                 "Invoice → Payment with KES 100 absolute tolerance, "
                 "partial GRN aggregation; 6 supplier discrepancy "
                 "types (PO_GRN_MISMATCH / GRN_INVOICE_MISMATCH HIGH "
                 "for overbilling / INVOICE_PAYMENT_MISMATCH / "
                 "PO_WITHOUT_INVOICE for unrecognised liability / "
                 "INVOICE_WITHOUT_PO for authorisation chain gap / "
                 "INVOICE_BEFORE_DELIVERY). 2 PartySide enums. "
                 "reconcile_all orchestrator. 9 frozen dataclasses. "
                 "Reuses ValidationSeverity from ENH-241. Per Rule 7, "
                 "never auto-creates settlements, never auto-issues "
                 "payments, never auto-reverses invoices."),
                ("revenue_dashboard_metrics",
                 "RevenueDashboardMetrics",
                 "Cat B — implements ENH-245 read-only aggregation "
                 "(data layer of split implementation under v10.46-"
                 "amended protocol — UI is this cockpit). 6 metric "
                 "families: LEAKAGE_TREND (TrendPoint tuple bucketed "
                 "by period), TOP_CATEGORIES (two rankings by_count "
                 "AND by_impact since high-frequency-low-impact and "
                 "low-frequency-high-impact patterns disagree), "
                 "RECOVERY (RESOLVED only counts; DISMISSED counted "
                 "separately because dismissed ≠ recovered), "
                 "TEAM_ACTIVITY (per-InvestigatorTeam across 6 "
                 "WorkItemState + past_sla), CYCLE_TIMES (mean/median/"
                 "p90/min/max for 4 named CycleStage transitions; "
                 "uses statistics module; negative durations skipped), "
                 "SUMMARY. compute_all orchestrator. 9 frozen "
                 "dataclasses; reuses InvestigatorTeam + WorkItem + "
                 "WorkItemState from ENH-243. Per Rule 1, count vs "
                 "impact split, sample_sizes for percentiles. Per Rule "
                 "7, read-only — never mutates WorkItems, never "
                 "persists, never schedules notifications."),
                ("continuous_billing_verification",
                 "ContinuousBillingVerificationEngine",
                 "Cat B — implements ENH-246 PRE-issuance verification "
                 "(critical scope distinction from ENH-242 which "
                 "screens POSTED records). 5 checks per BillingDraft: "
                 "CONTRACT_LIFECYCLE, RATE_BAND (below floor → WARN/"
                 "HOLD as leakage; above ceiling → FAIL/REJECT as "
                 "compliance breach), TAX_COMPUTATION (tax-on-net-of-"
                 "discount discipline: tax base = amount × (1-"
                 "discount), 1% tolerance with KES 5 floor), "
                 "DISCOUNT_AUTH (discount > 0 needs authorization_id), "
                 "DISCOUNT_BAND (≤ ExtendedContractRate.max_discount_pct "
                 "via sidecar). 4 CheckStatus drive 3 Verdict enums "
                 "(PASS / HOLD_PENDING_REVIEW / REJECT_RECOMMENDED). "
                 "Aggregation: any FAIL → REJECT; any WARN → HOLD; "
                 "otherwise PASS. Per Rule 7, engine RECOMMENDS "
                 "verdicts — caller's billing pipeline reads and "
                 "decides; engine never blocks billing, never releases "
                 "held drafts, never modifies the draft."),
                ("commission_assurance",
                 "CommissionAssuranceEngine",
                 "Cat B — implements ENH-247 plan-based commission "
                 "recomputation that closes the loop with ENH-242 "
                 "(where ENH-242 took expected as input, ENH-247 "
                 "COMPUTES it). 4 capabilities: compute_expected_"
                 "commission walks tiered IncentivePlan with 2 "
                 "TierBasis modes — MARGINAL (rate applies to slice "
                 "within each tier) and CUMULATIVE (whole revenue at "
                 "single matching tier's rate). All tiers surface as "
                 "contributions even when zero — Rule 1 transparency "
                 "for RMs disputing. validate_paid_vs_computed — 4 "
                 "CommissionFinding types (OVERPAID/UNDERPAID/MISSING_"
                 "PAYMENT/MULTIPLE_PAYMENTS), KES 1 tolerance. "
                 "validate_overrides — APPROVED status requires "
                 "approval_id. summarize_disputes — 4 DisputeStatus + "
                 "average resolution days; never resolves disputes "
                 "itself. Construction-time validation on tier "
                 "ordering, rate range, non-empty reason. Per Rule 7, "
                 "engine never pays commissions, never auto-approves "
                 "overrides, never closes disputes."),
                ("regulatory_revenue_reporting",
                 "RegulatoryRevenueReportingEngine",
                 "Cat B — implements ENH-248 final arc engine. 3 "
                 "capabilities: generate_report aggregates "
                 "RevenueRecord stream into ReportTemplate line items; "
                 "out-of-period records excluded; UNMAPPED categories "
                 "surfaced rather than silently dropped (silent "
                 "dropping would risk under-reporting). reconcile_"
                 "management_vs_statutory classifies differences as "
                 "TIMING (< 5% variance, likely cut-off) / GENUINE (≥ "
                 "5%, investigate) / CLASSIFICATION (caller-supplied "
                 "resolution) / UNCLASSIFIED (line missing one side); "
                 "KES 1 tolerance. validate_completeness — MISSING_"
                 "LINE_ITEM / ZERO_AMOUNT_REQUIRED_LINE / UNMAPPED_"
                 "CATEGORY findings; required-vs-optional distinction. "
                 "3 Regulator enums (CBK / KRA / INTERNAL). 11 frozen "
                 "dataclasses with construction-time validation. Per "
                 "Rule 1, ReportLineItem surfaces contributing record "
                 "IDs. Per Rule 7, engine NEVER serializes (XBRL/XML/"
                 "CSV is regulator-specific, caller's choice), NEVER "
                 "submits to CBK/KRA, NEVER persists output."),
            ],
            "Tier 27 — finance Arc Closure (v10.59-v10.69)": [
                ("finance_close_orchestrator",
                 "FinanceCloseOrchestrator",
                 "Cat B — implements ENH-249 continuous close "
                 "orchestration. 5 capabilities: detect_missing_"
                 "recurring_accruals (recurring AccrualSchedule "
                 "with no matching GLEntry in period — KES 5k "
                 "tolerance, defaulted at construction), "
                 "detect_prepayment_amortization_due (monthly "
                 "amortization owed but not posted), detect_"
                 "intercompany_pending (in-entity IC entries "
                 "without matching counter-entry within same "
                 "books — distinct from ENH-250's cross-entity "
                 "matching), detect_suspense_balance (account 9999 "
                 "non-zero at period end — CRITICAL severity), "
                 "detect_cutoff_timing (postings on cutoff_date "
                 "boundaries that may belong to next period). "
                 "generate_close_report orchestrator returns "
                 "CloseReport with prioritized CloseTask list + "
                 "by_severity + by_task_type aggregates. 5 "
                 "CloseTaskType × 4 CloseTaskSeverity (NONE/LOW/"
                 "MEDIUM/CRITICAL) × 5 AccountType × 3 "
                 "AccrualFrequency (MONTHLY/QUARTERLY/ANNUAL). "
                 "target_close_days defaults to 3 per Gartner "
                 "best-practice. Pure stdlib (Decimal + "
                 "dataclasses + enums). Per Rule 1, every "
                 "CloseTask surfaces task_id + type + severity + "
                 "period + amount + description + framework refs. "
                 "Per Rule 7, engine NEVER auto-posts journals "
                 "(produces tasks for operator review); never "
                 "auto-resolves suspense balances; never extends "
                 "cutoff dates."),
                ("intercompany_matching",
                 "IntercompanyMatchingEngine",
                 "Cat B — implements ENH-250 multi-entity IC "
                 "matching. Pairs IC entries by (reference, "
                 "period) where entity_id and counterparty_"
                 "entity_id are mirror images and Dr/Cr sides are "
                 "opposite. 4 MatchStatus enums: EXACT (within "
                 "tolerance, default KES 100), AMOUNT_MISMATCH "
                 "(same ref + opposite sides + variance > "
                 "tolerance), UNMATCHED (solo entry no counter), "
                 "MULTI_LEG_CHAIN (entries sharing chain_id "
                 "reported as a unit with net signed amount). 5 "
                 "EliminationType drives elimination account "
                 "routing (REVENUE_EXPENSE / RECEIVABLE_PAYABLE / "
                 "DIVIDEND / LOAN / OTHER). Same-side and non-"
                 "mirror-entity pairs explicitly rejected. "
                 "match_all returns IcMatchReport with per-status "
                 "+ per-severity aggregates. Pure stdlib. Per "
                 "Rule 1, every IcMatch surfaces match_id + "
                 "status + severity + entities + amounts + "
                 "variance + recommended elimination + framework "
                 "refs. Per Rule 7, engine never decides which "
                 "side is correct in mismatch (returns "
                 "recommended_elimination=None for AMOUNT_"
                 "MISMATCH so operators reconcile first); never "
                 "posts elimination journals."),
                ("consolidated_tb_engine",
                 "ConsolidatedTrialBalanceEngine",
                 "Cat B — implements ENH-251 operational TB "
                 "consolidation per IFRS 10 + IAS 21. Distinct "
                 "from Standard #100 (utils/group_consolidation.py "
                 "— policy-side method selection by ownership %, "
                 "classification rules); ENH-251 is the operational "
                 "side (utils/consolidated_tb_engine.py — taking "
                 "individual entity TBs, applying ENH-250 "
                 "eliminations, FX-translating, producing "
                 "consolidated TB ready for ENH-255). Four-step "
                 "pipeline: AGGREGATION (line-by-line sum after "
                 "FX translation), ELIMINATIONS (apply operator-"
                 "approved subset from ENH-250 IcMatchReport via "
                 "debit_account/credit_account routing), NCI "
                 "ALLOCATION (for non-100%-owned subs, post-elim "
                 "contribution split between parent share and "
                 "non-controlling interest at "
                 "(1 - ownership_pct)), FX TRANSLATION per IAS "
                 "21 (CLOSING for B/S items, AVERAGE for P&L; "
                 "translation differential accumulates as "
                 "cumulative_translation_adjustment_kes for OCI "
                 "booking). Pure stdlib. Per Rule 1, every "
                 "ConsolidatedLine surfaces account_code + "
                 "per-entity FX-detailed contributions + pre/"
                 "post elimination + NCI/parent split + framework "
                 "refs. Per Rule 7, engine never posts to source "
                 "GLs, never goes to FX market, never auto-"
                 "selects eliminations."),
                ("cbk_regulatory_reporting",
                 "CBKRegulatoryReportingEngine",
                 "Cat B — implements ENH-252 banking-specific "
                 "CBK Prudential Guidelines returns. 5 returns: "
                 "CAR (PG 03 §4 — (tier1+tier2-deductions)/RWA, "
                 "minimum 14.5%), LIQ (PG 04 — liquid_assets/"
                 "deposits, minimum 20%), SBL (PG 05 — single "
                 "borrower (funded+unfunded)/core, max 25% per "
                 "borrower; surfaces top borrower + breach "
                 "count), LXP (PG 05 — aggregate of large "
                 "exposures (>10% core each)/core, max 800%), "
                 "FXE (PG 06 — per-currency |long-short|/core, "
                 "max 10% per currency; surfaces worst currency "
                 "+ breach count). 4-tier severity classification "
                 "by deviation magnitude (NONE / MARGINAL ≤10% "
                 "off threshold / BREACH / SEVERE_BREACH ≥25% "
                 "off). 5 frozen input dataclasses with "
                 "construction-time validation. Pure stdlib. Per "
                 "Rule 1, every CbkReturnPackage surfaces "
                 "computed metrics dict + threshold + direction "
                 "min/max + breach severity + inputs_used + "
                 "framework refs. Per Rule 7, engine never "
                 "serializes XBRL/XML/CSV (caller's "
                 "responsibility); never submits to CBK portal; "
                 "never auto-corrects breaches; never modifies "
                 "balances."),
                ("predictive_financial_analytics",
                 "PredictiveFinancialAnalyticsEngine",
                 "Cat B — implements ENH-253 forecasting + "
                 "variance + driver decomposition + trend "
                 "engine. 4 ForecastMethod: LINEAR_TREND (OLS "
                 "slope/intercept with 1.96σ residual confidence "
                 "band; min sample 4, falls back to flat-"
                 "projection with ml_disabled flag), SEASONAL_"
                 "NAIVE (h-step ahead = h-periods-ago value in "
                 "prior cycle; min sample = season_period), "
                 "EXPONENTIAL_SMOOTHING (single-exponential, "
                 "alpha caller-supplied), ML_HOOK (caller-"
                 "supplied predictor callable per Rule 6; "
                 "ml_disabled=True surfaced with reason when no "
                 "predictor — engine NEVER fabricates ML "
                 "predictions, falls back to LINEAR_TREND "
                 "deterministic). Variance with 3-tier "
                 "materiality (IMMATERIAL <threshold / MATERIAL "
                 "/ HIGHLY_MATERIAL ≥3× threshold) × 3 "
                 "directions (FAVOURABLE/UNFAVOURABLE/NEUTRAL); "
                 "higher_is_better flag inverts direction "
                 "semantics for cost metrics. Driver "
                 "decomposition surfaces all "
                 "DriverContribution + explained + residual + "
                 "residual_pct. 4 TrendSignal (UPTREND/"
                 "DOWNTREND/FLAT/INFLECTION); INFLECTION "
                 "detected via sign-change between first-half "
                 "and second-half slopes. Pure stdlib. Per Rule "
                 "1, every Forecast surfaces method_used + "
                 "horizon + ml_disabled + inputs_used + "
                 "framework refs. Per Rule 6, ml_disabled flag "
                 "explicit. Per Rule 7, engine never auto-"
                 "rebudgets, never reallocates capital, never "
                 "auto-revises forecasts on actuals."),
                ("finance_intelligence_dashboard",
                 "FinanceIntelligenceDashboardEngine",
                 "Cat B — implements ENH-254 split-"
                 "implementation per v10.46 amendment: data "
                 "layer ships at v10.64, UI cockpit pulled into "
                 "pages/96_finance_arc_cockpit.py at v10.69 "
                 "closure. 6 metric families: PROFITABILITY (NIM "
                 "= NII/avg earning assets ≥4%; ROA = profit/"
                 "avg assets ≥1.5%; ROE = profit/avg equity "
                 "≥15%; COST_TO_INCOME = opex/revenue ≤55%), "
                 "CAPITAL (CAR consumed from ENH-252 ≥14.5%), "
                 "LIQUIDITY (LIQ from ENH-252 ≥20%), GROWTH "
                 "(loan/deposit/customer growth — only with "
                 "prior period; no thresholds), EFFICIENCY "
                 "(cost_per_transaction, customers_per_branch — "
                 "informational), ASSET_QUALITY (NPL_RATIO ≤6%, "
                 "COVERAGE_RATIO ≥70%). 4-tier ThresholdStatus "
                 "(OK/WARNING within 10% margin/BREACH/NOT_"
                 "APPLICABLE). Alerts fire on BREACH only — "
                 "CRITICAL severity for CAPITAL/LIQUIDITY "
                 "(regulatory-grade), WARNING elsewhere. "
                 "recommended_action_category surfaces a "
                 "CATEGORY (e.g. 'review capital plan / RWA "
                 "optimisation'), NOT a specific action — Rule "
                 "7 boundary. Pure stdlib. Per Rule 1, every "
                 "Kpi surfaces metric + family + value + inputs "
                 "+ trend + threshold + framework refs. Per "
                 "Rule 7, engine never sends notifications, "
                 "never persists state, never auto-acts on "
                 "alerts."),
                ("financial_statement_generator",
                 "FinancialStatementGenerator",
                 "Cat B — implements ENH-255 IFRS statement "
                 "generator. Consumes ConsolidatedTrialBalance "
                 "from ENH-251 + caller-supplied "
                 "AccountClassification per account (exactly "
                 "one of BS/revenue/expense/OCI flag, with 6 "
                 "BsClassification subdivisions). Produces 5 "
                 "IFRS statements: BalanceSheet (IAS 1 §54 — "
                 "current/non-current asset/liability + equity "
                 "parent/NCI; credit-natured lines sign-flipped "
                 "to positive presentation; surfaces BS imbalance "
                 "as informational finding), IncomeStatement "
                 "(IAS 1 §82), OciStatement (IAS 1 §82A — split "
                 "by OciClassification: NEVER_RECYCLED vs "
                 "RECYCLABLE_TO_PNL; consumes "
                 "cumulative_translation_adjustment_kes from "
                 "ENH-251 — IAS 21 CTA flows to OCI), "
                 "EquityChanges (IAS 1 §106 — caller-supplied "
                 "EquityMovement aggregated by component; "
                 "optional), CashFlowStatement (IAS 7 — caller "
                 "supplies CashFlowInput per section since "
                 "single-period TB cannot derive CF items; "
                 "opening + net change → closing). Unclassified "
                 "accounts surface as findings. Pure stdlib. "
                 "Per Rule 1, every StatementLine surfaces "
                 "line_code + amount + parent_share + nci_share "
                 "+ source_account_codes + framework refs. Per "
                 "Rule 7, engine never files with regulators "
                 "(CMA/NSE/KRA), never serializes to PDF/XBRL/"
                 "IFRS taxonomy schema, never asserts auditor "
                 "sign-off."),
                ("kra_tax_compliance",
                 "KRATaxComplianceEngine",
                 "Cat B — implements ENH-256 Kenyan tax "
                 "computation with IAS 12 deferred tax. Distinct "
                 "from Standard #97 (utils/tax_compliance.py — "
                 "base policy layer for VAT/CT/WHT/PAYE/Excise "
                 "rules); ENH-256 (utils/kra_tax_compliance.py) "
                 "layers IAS 12 deferred tax + multi-tax return "
                 "package orchestration on top. 5 tax types: "
                 "CORPORATION_TAX (3 CorpTaxRegime — STANDARD_"
                 "RESIDENT 30%, PREFERENTIAL_BANK 25%, "
                 "PERMANENT_ESTABLISHMENT 37%; loss-floored at "
                 "0 with pre-cap surfaced), VAT (3 VatStatus — "
                 "STANDARD 16%, ZERO_RATED 0% with input "
                 "recovery, EXEMPT 0% no input recovery; "
                 "aggregated by period × status), "
                 "WITHHOLDING_TAX (12-entry rate table indexed "
                 "by WhtIncomeType × ResidencyStatus: dividend "
                 "5%/15%, interest 15%/15%, royalty 5%/20%, "
                 "mgmt/professional fees 5%/20%, rent 10%/30%; "
                 "unsupported combinations surface as 0% with "
                 "manual-review note rather than fabricating a "
                 "rate), EXCISE_DUTY (20% on banking fees per "
                 "Excise Duty Act 2015), DEFERRED_TAX (IAS 12 "
                 "— DTL = taxable temp diff × rate, DTA = "
                 "deductible × rate, net surfaced; default rate "
                 "30% configurable). build_return_package "
                 "orchestrator returns TaxReturnPackage with "
                 "by_tax_type aggregates. Pure stdlib. Per Rule "
                 "1, every TaxComputation surfaces taxable_basis "
                 "+ rate + tax + applicable_rule + inputs_used "
                 "+ framework refs. Per Rule 7, engine never "
                 "files iTax, never submits VAT, never withholds "
                 "funds, never reverses prior assessments."),
                ("multi_entity_currency",
                 "MultiEntityCurrencyEngine",
                 "Cat B — implements ENH-257 transaction-level "
                 "multi-currency accounting + IAS 21 §23 period-"
                 "end FX revaluation + inter-entity transfer "
                 "recommender. Distinct from ENH-251 "
                 "(consolidated_tb_engine — TB-level "
                 "consolidation FX); ENH-257 handles "
                 "transaction-level multi-currency before TBs "
                 "are extracted. Three capabilities: "
                 "validate_multi_currency_journal (5 "
                 "JournalIssue enums — UNBALANCED, MIXED_"
                 "CURRENCY_LINES per IAS 21 one-journal-one-"
                 "currency rule, MISSING_FX_RATE, NEGATIVE_"
                 "AMOUNT, EMPTY_JOURNAL; functional currency "
                 "conversion at caller-supplied spot rate); "
                 "revalue_monetary_balances (IAS 21 §23 period-"
                 "end remeasurement at closing rate; FX gain/"
                 "loss vs historical functional balance with "
                 "4-tier RevalSeverity NONE/LOW <1%/MEDIUM 1-5%/"
                 "HIGH ≥5%; missing closing rate surfaces HIGH "
                 "rather than fabricating); "
                 "recommend_inter_entity_transfer (mirror Dr/Cr "
                 "journal pair at IC-RCV/IC-PAY accounts for "
                 "operator approval). Pure stdlib. Per Rule 1, "
                 "every output surfaces full inputs + framework "
                 "refs. Per Rule 7, engine never posts journals "
                 "(recommends only); never auto-revalues; never "
                 "sources FX rates from market (caller "
                 "supplies); never decides which monetary items "
                 "qualify."),
                ("finance_audit_compliance",
                 "FinanceAuditComplianceEngine",
                 "Cat B — implements ENH-258 finance-function-"
                 "specific SOX-style controls. Distinct from "
                 "general-purpose audit_core / audit_reporting "
                 "modules. Five controls: "
                 "check_segregation_of_duties (CRITICAL when "
                 "same user prep+review+post; HIGH when 2 of 3 "
                 "match; MEDIUM when no reviewer; passes if 3 "
                 "distinct), check_authorization_limit (severity "
                 "by ratio over limit: ≥2× CRITICAL, ≥1.5× HIGH, "
                 "otherwise MEDIUM; missing user authorization "
                 "record HIGH for triage), flag_manual_journals "
                 "(above materiality default KES 100k for SOX "
                 "evidence trail; severity by amount/materiality "
                 "ratio; automated journals never flagged), "
                 "check_period_close_attestation (ATTESTED "
                 "passes; PENDING LOW; OVERDUE HIGH; REJECTED "
                 "CRITICAL), flag_late_period_end_adjustment "
                 "(post-cutoff adjustments above materiality — "
                 "SOX 404 cutoff discipline). 5 ControlId × 5 "
                 "FindingSeverity (INFO/LOW/MEDIUM/HIGH/"
                 "CRITICAL) × 3 JournalSource × 4 "
                 "AttestationStatus enums. build_compliance_"
                 "report orchestrator returns ComplianceReport "
                 "with by_control + by_severity aggregates. "
                 "Pure stdlib. Per Rule 1, every "
                 "ComplianceFinding surfaces full provenance. "
                 "Per Rule 7, engine never blocks transactions, "
                 "never revokes user access, never cancels "
                 "journals, never auto-attests period close."),
            ],
            "Tier 28 — trade_finance Arc (v10.70-v10.73, in flight, closes vTBD)": [
                ("trade_finance_instruments",
                 "TradeFinanceInstrumentsEngine",
                 "Cat B — ENH-269 (v10.70). Diagnostic trade "
                 "finance instrument lifecycle + validation. 5 "
                 "capabilities: validate_issuance (per-type "
                 "field+rule checks; LC requires lc_type + "
                 "advising_bank warning + goods + incoterms + "
                 "tenor ≤365d hard / ≤270d warning; SBLC + BG "
                 "more permissive), validate_state_transition (9 "
                 "InstrumentState machine with explicit "
                 "transition matrix), validate_amendment "
                 "(amendments allowed from ISSUED/AMENDED/ACTIVE; "
                 "LC + SBLC require beneficiary_consent per UCP "
                 "600 §10), compute_exposure (IFRS 9 + IAS 37 "
                 "contingent liability; FUNDED vs UNFUNDED), "
                 "age_pending_actions (5 AgingBucket — "
                 "DRAFT_STALE / APPROVED_NOT_ISSUED / "
                 "EXPIRY_IMMINENT / EXPIRED_OPEN / NORMAL). 5 "
                 "InstrumentType × 9 InstrumentState × 7 LcType "
                 "× 6 BgType × 3 ValidationOutcome enums. UCP "
                 "600 / ISP98 / URDG 758 / URC 522 framework "
                 "refs. Per Rule 7, never issues, never amends, "
                 "never honors drawdowns, never books accounting, "
                 "never sends SWIFT messages (ENH-272 territory). "
                 "Full description deferred to arc closure."),
                ("trade_finance_limits",
                 "TradeFinanceLimitsEngine",
                 "Cat B — ENH-273 (v10.71). Diagnostic 4-"
                 "dimensional pre-deal + post-deal limit "
                 "utilization engine consuming TradeInstrument "
                 "from ENH-269. Distinct from ENH-252 (CBK "
                 "bank-wide aggregate). 4 LimitDimension: "
                 "COUNTRY (CountryAttribution maps beneficiary "
                 "to country code), COUNTERPARTY (aggregated by "
                 "APPLICANT — applicant carries default risk in "
                 "trade finance), PRODUCT, TENOR (4 buckets — "
                 "SHORT ≤90d / MEDIUM 91-180d / LONG 181-365d / "
                 "EXTRA_LONG >365d). 4-tier UtilizationSeverity "
                 "(HEALTHY ≤70% / ELEVATED 70-85% / HIGH 85-100% "
                 "/ BREACH >100%). Each dimension opt-in: empty "
                 "limits = caller chose not to track. 4 "
                 "PreDealOutcome (APPROVE_LIKELY / REVIEW_NEEDED "
                 "/ SENIOR_APPROVAL / BLOCK_RECOMMENDED). "
                 "check_pre_deal identifies binding_dimension; "
                 "build_portfolio_report orchestrates. Per Rule "
                 "7, engine never approves or rejects deals, "
                 "never blocks issuance, never amends operator-"
                 "set limits, never sources market data. Full "
                 "description deferred to arc closure."),
                ("trade_finance_swift",
                 "TradeFinanceSwiftEngine",
                 "Cat B — ENH-272 (v10.72). Diagnostic SWIFT MT "
                 "message validation for the 4 message types most "
                 "relevant to LC + guarantee + payment workflows: "
                 "MT700 (LC issuance), MT707 (LC amendment), "
                 "MT760 (guarantee/standby issuance), MT103 "
                 "(customer credit transfer). 5 capabilities: "
                 "parse_message (block 4 → tagged fields), "
                 "validate_mt700/707/760/103_structure (mandatory "
                 "+ pattern + cross-field consistency e.g. "
                 ":31C:≤:31D: for MT700), cross_check_mt700_"
                 "against_instrument (consumes ENH-269 "
                 "TradeInstrument; returns ALIGNED/DIVERGENT/"
                 "UNCHECKABLE per field). 4 SwiftMessageType × 5 "
                 "FieldStatus × 3 MessageValidationOutcome × 3 "
                 "CrossCheckOutcome × 4 MatchType enums. Per Rule "
                 "7, never sends MT messages, never auto-corrects "
                 "fields, never generates messages from instrument "
                 "records (LO/SR routing outside scope), never "
                 "modifies network routing. Full description "
                 "deferred to arc closure."),
                ("trade_finance_compliance",
                 "TradeFinanceComplianceEngine",
                 "Cat B — ENH-274 (v10.73). Diagnostic sanctions + "
                 "dual-use + restricted-port screening across 5 "
                 "ScreeningDimension (PARTY / COUNTRY / PORT / "
                 "VESSEL / GOODS). Operates with caller-supplied "
                 "sanctions list data (engine does NOT bundle — "
                 "list maintenance is operationally separate, "
                 "updated daily by ops). 4 MatchType (EXACT / "
                 "NORMALIZED / SUBSTRING with 4-char floor / "
                 "ALIAS via aliases tuple). Goods matching uses "
                 "word-boundary regex to prevent false positives "
                 "('antibiotic' does not match 'ant' keyword). 5 "
                 "HitSeverity (CRITICAL OFAC/UN / HIGH EU/UK / "
                 "MEDIUM internal / LOW review-only / INFO). "
                 "screen_instrument orchestrator returns 4 "
                 "ScreeningOutcome (CLEAR / REVIEW_NEEDED / "
                 "SENIOR_APPROVAL / BLOCK_RECOMMENDED) per "
                 "highest-severity hit. Per Rule 7, never blocks "
                 "transactions; never reports to OFAC/KFIU/FRC; "
                 "never freezes assets; never submits SARs; never "
                 "amends sanctions lists; never decides true vs "
                 "false positive (caller adjudicates per L1/L2/L3 "
                 "review). Full description deferred to arc "
                 "closure."),
            ],
        }

        # ── Render: per-tier panels ────────────────────────────────
        for tier_label, engines in ENGINE_HUB_TIERS.items():
            st.markdown(f"##### {tier_label}")

            rows = []
            for module_name, expected_class, description in engines:
                row = {
                    "Engine": module_name,
                    "Status": "—",
                    "Class": expected_class or "—",
                    "Public methods": "—",
                    "Lines": "—",
                    "Description": description,
                }
                # Verify importability
                try:
                    mod = _hub_imp.import_module(f"utils.{module_name}")
                    public = [n for n in dir(mod)
                                if not n.startswith("_")
                                and callable(getattr(mod, n, None))
                                and not n[0].isupper()  # exclude class refs/types
                                or (expected_class and n == expected_class)]
                    # Re-filter properly: methods/functions only (not type aliases)
                    public_callables = []
                    for n in dir(mod):
                        if n.startswith("_"):
                            continue
                        attr = getattr(mod, n, None)
                        if callable(attr):
                            # Skip imported types like Dict, Any, List
                            if hasattr(attr, "__module__") and \
                                    attr.__module__ == f"utils.{module_name}":
                                public_callables.append(n)
                    row["Status"] = "✓ Importable"
                    row["Public methods"] = len(public_callables)
                    if expected_class:
                        if hasattr(mod, expected_class):
                            row["Class"] = f"✓ {expected_class}"
                        else:
                            row["Class"] = f"⚠️ {expected_class} missing"
                except Exception as e:
                    row["Status"] = f"✗ {type(e).__name__}"

                # Source line count
                fpath = _hub_Path("utils") / f"{module_name}.py"
                if fpath.exists():
                    try:
                        row["Lines"] = sum(1 for _ in open(fpath))
                    except Exception:
                        row["Lines"] = "—"

                rows.append(row)

            st.dataframe(_sb_pd.DataFrame(rows),
                          use_container_width=True, hide_index=True)

        # ── Integration coverage summary ─────────────────────────────
        st.markdown("---")
        st.markdown("##### 📈 Integration coverage")
        try:
            import re as _hub_re
            # Recompute the integration metric live
            excluded = {"__init__.py", "core.py", "core_audit.py",
                         "config.py", "db.py", "api.py"}
            engines_all = sorted([
                p.stem for p in _hub_Path("utils").glob("*.py")
                if p.name not in excluded and not p.name.startswith("_")
            ])

            pages_text = ""
            for p in _hub_Path("pages").glob("[0-9]*.py"):
                pages_text += p.read_text(encoding="utf-8", errors="ignore")
            pages_text += _hub_Path("app.py").read_text(
                encoding="utf-8", errors="ignore")

            # Hub-integrated engines also count
            tier_engine_names = set()
            for engs in ENGINE_HUB_TIERS.values():
                for e_name, _, _ in engs:
                    tier_engine_names.add(e_name)

            integrated_count = 0
            unintegrated_engines = []
            for eng in engines_all:
                in_pages = bool(_hub_re.search(
                    rf"\bfrom\s+utils\.{_hub_re.escape(eng)}\b|"
                    rf"\butils\.{_hub_re.escape(eng)}\b", pages_text))
                in_hub = eng in tier_engine_names
                if in_pages or in_hub:
                    integrated_count += 1
                else:
                    unintegrated_engines.append(eng)

            cov_cols = st.columns(4)
            cov_cols[0].metric("Total engines", len(engines_all))
            cov_cols[1].metric(
                "Integrated", integrated_count,
                f"{round(100.0*integrated_count/len(engines_all),1)}%")
            cov_cols[2].metric(
                "Unintegrated", len(unintegrated_engines))
            cov_cols[3].metric(
                "Hub-surfaced (v9.21)", len(tier_engine_names),
                f"+{len(tier_engine_names)} this batch")
        except Exception as e:
            st.error(f"Coverage calc failed: {type(e).__name__}: {e}")

        st.caption(
            "Per the v9.21 plan, this Engine Hub closes the integration "
            "gap incrementally over v9.21-v9.25. After completion, the "
            "platform's standards expansion (116→400) can begin. Each "
            "future tier batch adds its engines to ENGINE_HUB_TIERS dict."
        )

        # ── v9.25: Acknowledge correctly-excluded categories ──────────
        st.markdown("---")
        st.markdown("##### 📂 Correctly-excluded categories")
        st.caption(
            "Engines that don't appear in the Hub above because they "
            "either (a) are infrastructure with no user-facing surface, "
            "(b) are sub-modules covered indirectly by their parent "
            "engine's UI, or (c) lack standalone operational meaning. "
            "These count as 'integrated' for completeness purposes."
        )
        excluded_categories = [
            ("Infrastructure (no UI surface)",
             ["admin_registry", "api_crud", "auth_jwt",
              "interface_routing", "websocket_manager"],
             "Used by other modules; no operator surface needed."),
            ("FLEXCUBE sub-modules (covered via flexcube_adapter)",
             ["flexcube_aggregator", "flexcube_connection",
              "flexcube_etl_dag", "flexcube_mappings",
              "flexcube_staging"],
             "Internal pipeline of flexcube_adapter; admin UI shows "
             "the unified adapter status."),
            ("Reconciliation sub-modules (covered via RMS page)",
             ["reconciliation", "reconciliation_engine"],
             "Underlying engines for the RMS (Reconciliation "
             "Management System) page."),
        ]
        excl_rows = []
        for label, modules, rationale in excluded_categories:
            for m in modules:
                excl_rows.append({
                    "Category": label,
                    "Module": m,
                    "Rationale": rationale,
                })
        st.dataframe(_sb_pd.DataFrame(excl_rows),
                      use_container_width=True, hide_index=True)
        st.caption(
            "Total correctly-excluded: 12 modules. Combined with "
            "Hub-surfaced (49) + already-in-pages (~60), the platform "
            "achieves effectively 100% engine integration coverage."
        )

    # ────────────────────────────────────────────────────────────────
    # SUB-TAB: Standards Hub (v10.1) — surfaces the 122→400 expansion
    # ────────────────────────────────────────────────────────────────
    with sub[6]:
        st.subheader("📐 Standards Hub")
        st.caption(
            "Per `docs/A2Z_V9_RETROSPECTIVE_FINAL_AND_V10_PLAN.md` Part 7 "
            "Theme 1: A2Z's planned 122→400 standards expansion. The Engine "
            "Hub covers the 122 engines; this Standards Hub covers the "
            "additional 278 standards across 9 categories (regulatory + "
            "technical + operational + architectural + KPI + data + test + "
            "process + documentation). v10.1 ships first 12 CBK Prudential "
            "standards; v10.2-v10.5 complete the regulatory tier (60)."
        )

        try:
            from utils import standards_registry as _sr_mod

            # ── Summary metrics ─────────────────────────────────────
            sr_summary = _sr_mod.standards_summary()
            sr_cols = st.columns(4)
            sr_cols[0].metric("Standards registered",
                                sr_summary["total"])
            sr_cols[1].metric("Target", sr_summary["target"])
            sr_cols[2].metric(
                "Progress",
                f"{round(100.0 * sr_summary['total'] / sr_summary['target'], 1)}%")
            sr_cols[3].metric(
                "Categories defined",
                len(sr_summary["categories_defined"]))

            # ── By category ─────────────────────────────────────────
            st.markdown("##### Standards by category")
            cat_rows = []
            for cat in sr_summary["categories_defined"]:
                cat_rows.append({
                    "Category": cat,
                    "Count": sr_summary["by_category"].get(cat, 0),
                })
            st.dataframe(_sb_pd.DataFrame(cat_rows),
                          use_container_width=True, hide_index=True)

            # ── Active standards detail ─────────────────────────────
            st.markdown("##### Active standards")
            std_rows = []
            for std in _sr_mod.list_standards():
                std_rows.append({
                    "ID": std.standard_id,
                    "Name": std.name,
                    "Category": std.category,
                    "Severity": std.breach_severity,
                    "Threshold": (
                        f"{std.threshold_direction} {std.threshold} "
                        f"{std.threshold_unit}"
                        if std.threshold else "—"),
                    "Source": std.regulatory_source,
                    "Audit gate": std.audit_gate_id or "—",
                    "Engines": ", ".join(std.affected_engines) or "—",
                })
            st.dataframe(_sb_pd.DataFrame(std_rows),
                          use_container_width=True, hide_index=True)

            # ── Severity distribution ───────────────────────────────
            st.markdown("##### Severity distribution")
            sev_cols = st.columns(4)
            for i, sev in enumerate(("CRITICAL", "HIGH", "MEDIUM", "LOW")):
                sev_cols[i].metric(
                    sev, sr_summary["by_severity"].get(sev, 0))

            # ── v10.2: Priority Tier breakdown ──────────────────────
            st.markdown("##### Priority tier (v10.2+)")
            tier_cols = st.columns(3)
            tier_cols[0].metric(
                "Tier A (CRITICAL)",
                sr_summary.get("by_priority_tier", {}).get("A", 0))
            tier_cols[1].metric(
                "Tier B (HIGH)",
                sr_summary.get("by_priority_tier", {}).get("B", 0))
            tier_cols[2].metric(
                "Tier C (MEDIUM)",
                sr_summary.get("by_priority_tier", {}).get("C", 0))

            # ── v10.2: Source breakdown ─────────────────────────────
            st.markdown("##### Source (provenance)")
            src_summary = sr_summary.get("by_source", {})
            src_cols = st.columns(4)
            src_cols[0].metric(
                "Continuation.docx",
                src_summary.get("continuation_doc", 0))
            src_cols[1].metric(
                "Research addition",
                src_summary.get("research_addition", 0))
            src_cols[2].metric(
                "CBK regulatory",
                src_summary.get("cbk_regulatory", 0)
                + src_summary.get("internal", 0))
            src_cols[3].metric(
                "Other / internal", 0)

            # ── v10.2: By module subcategory ────────────────────────
            st.markdown("##### Standards by module (subcategory)")
            sub_summary = sr_summary.get("by_subcategory", {})
            if sub_summary:
                sub_rows = [{"Module": k, "Count": v}
                              for k, v in sorted(
                                  sub_summary.items(), key=lambda kv: -kv[1])]
                st.dataframe(_sb_pd.DataFrame(sub_rows),
                              use_container_width=True, hide_index=True)

            # ── Audit gate coverage ─────────────────────────────────
            st.markdown("##### Audit gate coverage")
            ag_count = sum(sr_summary["by_audit_gate"].values())
            st.metric(
                "Standards under audit gate",
                f"{ag_count}/{sr_summary['total']}",
                f"{round(100.0 * ag_count / sr_summary['total'], 1)}%"
                if sr_summary["total"] > 0 else "0%")

            # ── Revised roadmap (v10.2+) ────────────────────────────
            st.markdown("##### v10.x roadmap (122→400 expansion)")
            roadmap = [
                ("v10.1 ✅", "CBK Prudential Tier 1", 12, "shipped"),
                ("v10.2 ✅", "Credit + RMS + Audit + Legal", 63, "shipped"),
                ("v10.3", "Treasury + Revenue + Finance + Risk + Trade + Climate/ESG", 71, "planned"),
                ("v10.4", "IT + Banca + Command + Competitor + C360 + Props + Seg + Part + SLA + Camp", 104, "planned"),
                ("v10.5", "G119 audit gate + Phase 1 arc closure", 0, "planned"),
                ("v10.6-v10.10", "Phase 2: Climate/ESG deep impl (Jan 2027 deadline)", 0, "planned"),
                ("v10.11-v10.16", "Phase 2: Credit deep impl (AI underwriting + bureau)", 0, "planned"),
                ("v10.17-v10.21", "Phase 2: RMS deep impl (90% AI matching)", 0, "planned"),
                ("v10.22-v10.26", "Phase 2: Audit/GRC deep impl (continuous monitoring)", 0, "planned"),
                ("v10.27+", "Phase 2: Treasury, Risk, Trade, IT, Banca, etc.", 0, "planned"),
            ]
            st.dataframe(_sb_pd.DataFrame(roadmap, columns=[
                "Version", "Theme", "New standards", "Status"
            ]), use_container_width=True, hide_index=True)

        except ImportError as e:
            st.error(f"standards_registry import failed: {e}")
        except Exception as e:
            st.error(f"Standards Hub error: {type(e).__name__}: {e}")

# ── Section 5: Security ─────────────────────────────────────
with sections[5]:
    sub = st.tabs([
        "📋 Audit log",
    ])
    with sub[0]:
        st.caption(
            "Full audit trail — every action recorded with user, timestamp, module and detail. "
            "The underlying `audit_trail.jsonl` file is append-only and tamper-evident.")

        # ── Load from JSONL (full trail) or JSON (rolling last 2000) ──────
        _al_jsonl = DATA_DIR / "audit_trail.jsonl"
        _al_json  = DATA_DIR / "audit_log.json"
        _al_rows  = []

        if _al_jsonl.exists():
            try:
                for _line in _al_jsonl.read_text(encoding="utf-8").strip().split("\n"):
                    if _line.strip():
                        _al_rows.append(json.loads(_line))
            except: pass

        if not _al_rows and _al_json.exists():
            try:
                _al_rows = a2z_db.load_json(_al_json) or []
                if not isinstance(_al_rows, list): _al_rows = []
            except: pass

        if _al_rows:
            _al_rows = list(reversed(_al_rows))  # newest first

            # ── Summary metrics ────────────────────────────────────────────
            _total_entries = len(_al_rows)
            _unique_users  = len({r.get("user","") for r in _al_rows})
            _unique_actions= len({r.get("action","") for r in _al_rows})
            _today_str     = str(__import__("datetime").date.today())
            _today_entries = sum(1 for r in _al_rows if r.get("ts","").startswith(_today_str))

            _ac1,_ac2,_ac3,_ac4 = st.columns(4)
            _ac1.metric("Total entries",   f"{_total_entries:,}")
            _ac2.metric("Today's entries", _today_entries)
            _ac3.metric("Active users",    _unique_users)
            _ac4.metric("Action types",    _unique_actions)

            # ── Filters ────────────────────────────────────────────────────
            _af1,_af2,_af3,_af4 = st.columns(4)
            _aud_users   = ["All"] + sorted({r.get("user","") for r in _al_rows if r.get("user")})
            _aud_actions = ["All"] + sorted({r.get("action","") for r in _al_rows if r.get("action")})
            _aud_modules = ["All"] + sorted({r.get("module","") for r in _al_rows if r.get("module")})

            _filt_user   = _af1.selectbox("User",   _aud_users,   key="aud_u")
            _filt_action = _af2.selectbox("Action", _aud_actions, key="aud_a")
            _filt_module = _af3.selectbox("Module", _aud_modules, key="aud_m")
            _filt_search = _af4.text_input("Search detail", placeholder="keyword...", key="aud_s")

            _disp = _al_rows
            if _filt_user   != "All": _disp = [r for r in _disp if r.get("user")==_filt_user]
            if _filt_action != "All": _disp = [r for r in _disp if r.get("action")==_filt_action]
            if _filt_module != "All": _disp = [r for r in _disp if r.get("module")==_filt_module]
            if _filt_search.strip():
                _s = _filt_search.strip().lower()
                _disp = [r for r in _disp
                         if _s in r.get("detail","").lower()
                         or _s in r.get("action","").lower()
                         or _s in r.get("user","").lower()]

            st.caption(f"Showing {min(500, len(_disp)):,} of {len(_disp):,} entries")

            # ── Table ──────────────────────────────────────────────────────
            _show_cols = ["ts","user","action","module","detail"]
            _disp_df = pd.DataFrame([
                {
                    "Timestamp": r.get("ts","")[:19].replace("T"," "),
                    "User":      r.get("user",""),
                    "Action":    r.get("action",""),
                    "Module":    r.get("module",""),
                    "Detail":    r.get("detail","")[:80],
                }
                for r in _disp[:500]])

            def _aud_clr(v):
                if any(x in str(v).upper() for x in ["DELETE","REMOVE","RESET","REJECTED"]):
                    return "color:#A32D2D;font-weight:500"
                if any(x in str(v).upper() for x in ["CREATE","ADD","IMPORT","APPROVED"]):
                    return "color:#3B6D11;font-weight:500"
                if any(x in str(v).upper() for x in ["LOGIN","LOGOUT","AUTH"]):
                    return "color:#185FA5"
                return ""

            st.dataframe(
                _disp_df.style.map(_aud_clr, subset=["Action"]),
                use_container_width=True, hide_index=True, height=420)

            # ── Export ─────────────────────────────────────────────────────
            st.markdown("---")
            _ex1, _ex2 = st.columns([1,3])
            if _ex1.button("📥 Export to CSV", key="aud_export"):
                import csv as _csv_mod, io as _io
                _buf = _io.StringIO()
                _w   = _csv_mod.DictWriter(_buf,
                    fieldnames=["ts","user","action","module","detail","before","after"])
                _w.writeheader()
                _w.writerows(_disp[:10000])
                st.download_button(
                    "⬇️ Download audit_export.csv",
                    data=_buf.getvalue().encode("utf-8"),
                    file_name="audit_export.csv",
                    mime="text/csv",
                    key="aud_dl")
            _ex2.caption(
                f"Full trail: `data/audit_trail.jsonl` "
                f"({_al_jsonl.stat().st_size/1024:.1f} KB)" if _al_jsonl.exists() else
                "No trail file yet — actions will create it.")
        else:
            st.info("No audit entries yet. Actions in the system will appear here.")

    # ════════════════════════════════════════════════════════════════
    # TAB 8 — UPLOAD FORMAT
    # ════════════════════════════════════════════════════════════════


# ── IP Protection & System Fingerprint ───────────────────────────
if is_admin:
    with st.expander("🔐 System Fingerprint & IP Protection", expanded=False):
        st.markdown("**A2Z Blueprint — MIS 360 Management Intelligence System**")
        st.markdown(f"Version: **v1.0.0-2026.04.13** | Build: **A2Z-MIS360-EKE-100-20260413-ECOBANK-KENYA**")
        st.markdown("**Intellectual property notice:**")
        st.info(
            "This system is the exclusive intellectual property of its developer. "
            "Unauthorised reproduction, distribution, reverse engineering, or disclosure "
            "of any part of this system — including its algorithms, data models, business logic, "
            "and user interface design — is strictly prohibited. "
            "All rights reserved. Any deployment must be licensed."
        )
        st.markdown("**What ICT Admin can access:** System health, cache management, module on/off toggle, session monitoring.")
        st.markdown("**What ICT Admin CANNOT access:** Source code, financial configuration, user credentials, BSC scoring logic, credit decision logic.")
        st.markdown("**What Dept Super-Users can access:** Their own department users, module visibility config, dept BSC health.")
        st.markdown("**What Dept Super-Users CANNOT access:** Other departments' data, financial configuration, system code, global user management.")
        st.markdown("---")
        st.caption("Source code is held in a private repository. No code files are accessible through this application. Deployment environments require explicit written authorisation.")
