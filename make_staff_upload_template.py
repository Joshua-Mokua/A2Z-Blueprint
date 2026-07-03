#!/usr/bin/env python3
"""scripts/make_staff_upload_template.py — generate the once-off staff org-tree
upload template (.xlsx).

Produces data/STAFF_UPLOAD_TEMPLATE.xlsx with:
  - 'Staff' sheet: the rows you fill in (MD -> EXCO -> ... -> DSA), with
    dropdown validation on Role (from org_config hierarchy) and Branch (the 16),
    and a Reports-To-Code column (the specific manager's staff code) that wires
    the real person-to-person reporting tree.
  - 'Roles (reference)' sheet: every valid role name + its parent role(s) from
    the configured hierarchy — so you fill Role with EXACT valid values.
  - 'Branches (reference)' sheet: the 16 (+ Head Office).
  - 'Instructions' sheet: how to fill it, incl. the DSA dual-line rule.

The pipeline (upload_staff_register) validates uploads against THIS contract.
Run:  python scripts\\make_staff_upload_template.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from utils.core import get_org_config

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "STAFF_UPLOAD_TEMPLATE.xlsx"

cfg = get_org_config() or {}
hierarchy = cfg.get("hierarchy", {})
roles = list(hierarchy.keys())
branches = [b["name"] for b in cfg.get("branches", [])]

BRAND = "0082BB"; BRAND_DK = "005B82"; LIME = "BED600"; GREY = "EDEDED"
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=BRAND_DK)
inp_fill = PatternFill("solid", fgColor="FFFDE7")  # pale yellow = you fill
ref_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Sheet 1: Staff (the upload) ───────────────────────────────────────
ws = wb.active; ws.title = "Staff"
COLS = ["Staff Code", "Staff Name", "Role", "Branch", "Region (DSA only)",
        "Reports To Code", "Dotted Line Code 1", "Dotted Line Code 2",
        "Band", "Gender", "Email", "Date of Employment"]
ws.append(COLS)
for i, _ in enumerate(COLS, 1):
    c = ws.cell(row=1, column=i); c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 30
widths = [12, 24, 38, 20, 18, 16, 18, 18, 8, 8, 26, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

# Pre-stub the top of the tree so the operator sees the shape (MD + EXCO roots).
# These are EXAMPLES to overwrite with real people; Reports To Code blank = root.
stub = [
    ["300001", "<MD name>", "Managing Director", "Head Office", "", "", "", "", "E1", "", "", ""],
    ["", "<Director CCB>", "Director Consumer & Commercial Banking (CCB)", "Head Office", "", "300001", "", "", "E1", "", "", ""],
    ["", "<Director CIB>", "Director Corporate & Investment Banking (CIB)", "Head Office", "", "300001", "", "", "E1", "", "", ""],
    ["", "<CFO>", "Chief Finance Officer", "Head Office", "", "300001", "", "", "E1", "", "", ""],
]
for r in stub:
    ws.append(r)
# mark input area with pale fill for a few hundred rows
for row in range(2, 1200):
    for col in range(1, len(COLS) + 1):
        ws.cell(row=row, column=col).fill = inp_fill
        ws.cell(row=row, column=col).border = border
ws.freeze_panes = "A2"

# Data validation dropdowns
# Role dropdown — from hierarchy. (Excel list limit ~255 chars per formula, so
# we point at the reference sheet range instead of inline list.)
role_dv = DataValidation(type="list", formula1="='Roles (reference)'!$A$2:$A$200", allow_blank=False)
role_dv.error = "Pick a role from the Roles (reference) sheet."
role_dv.errorTitle = "Invalid role"
ws.add_data_validation(role_dv); role_dv.add(f"C2:C1199")

branch_dv = DataValidation(type="list", formula1="='Branches (reference)'!$A$2:$A$60", allow_blank=False)
branch_dv.error = "Pick a branch from the Branches (reference) sheet."
branch_dv.errorTitle = "Invalid branch"
ws.add_data_validation(branch_dv); branch_dv.add(f"D2:D1199")

gender_dv = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
ws.add_data_validation(gender_dv); gender_dv.add("J2:J1199")

# ── Sheet 2: Roles (reference) ────────────────────────────────────────
rs = wb.create_sheet("Roles (reference)")
rs.append(["Role (use EXACTLY)", "Reports to role(s)"])
for c in range(1, 3):
    cell = rs.cell(row=1, column=c); cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
for role in roles:
    parents = hierarchy.get(role, [])
    rs.append([role, " / ".join(parents) if parents else "(root)"])
rs.column_dimensions["A"].width = 42; rs.column_dimensions["B"].width = 50
for row in rs.iter_rows(min_row=2):
    for cell in row:
        cell.font = ref_font; cell.border = border

# ── Sheet 3: Branches (reference) ─────────────────────────────────────
bs = wb.create_sheet("Branches (reference)")
bs.append(["Branch (use EXACTLY)"])
bs.cell(row=1, column=1).font = hdr_font; bs.cell(row=1, column=1).fill = hdr_fill
for b in branches:
    bs.append([b])
bs.column_dimensions["A"].width = 28
for row in bs.iter_rows(min_row=2):
    for cell in row:
        cell.font = ref_font; cell.border = border

# ── Sheet 4: Instructions ─────────────────────────────────────────────
ins = wb.create_sheet("Instructions")
lines = [
    ("A2Z MIS 360 — Staff Org-Tree Upload Template", True),
    ("", False),
    ("Fill the 'Staff' sheet, one row per person, from the MD down to every DSA.", False),
    ("", False),
    ("COLUMNS:", True),
    ("• Staff Code — unique numeric code (also becomes the login username).", False),
    ("• Staff Name — full name.", False),
    ("• Role — MUST be one of the values on the 'Roles (reference)' sheet (dropdown).", False),
    ("• Branch — MUST be one of the 16 (+ Head Office) on 'Branches (reference)' (dropdown).", False),
    ("• Region (DSA only) — the DSA region (Nairobi CBD/Metro, etc.). Leave blank for non-DSA.", False),
    ("• Reports To Code — the Staff Code of this person's PRIMARY (solid-line) manager.", False),
    ("     The MD's Reports To Code is blank (root). Everyone else points up the tree.", False),
    ("• Dotted Line Code 1 / 2 — for DSAs ONLY: the Staff Codes of the Branch DSA Team", False),
    ("     Lead and the DSA Regional Head (functional/dotted oversight). Blank for others.", False),
    ("     (A DSA's PRIMARY Reports To Code = their Branch Manager's code.)", False),
    ("• Band, Gender, Email, Date of Employment — optional metadata.", False),
    ("", False),
    ("REPORTING-TREE RULES (validated on upload):", True),
    ("• Every Reports To Code must exist as a Staff Code elsewhere in the sheet.", False),
    ("• Exactly ONE row may have a blank Reports To Code (the MD / root).", False),
    ("• No cycles (A reports to B reports to A) — rejected.", False),
    ("• Role must be valid; Branch must be one of the 16; else the upload is REJECTED.", False),
    ("", False),
    ("This is a STRICT, wipe-and-replace load: on upload, the staff table is rebuilt", False),
    ("from this sheet (your designated test logins are preserved). Get it right here.", False),
]
for i, (txt, bold) in enumerate(lines, 1):
    cell = ins.cell(row=i, column=1, value=txt)
    cell.font = Font(name="Arial", bold=bold, size=13 if (bold and i == 1) else 11)
ins.column_dimensions["A"].width = 95

wb.save(OUT)
print(f"  wrote {OUT.relative_to(ROOT)}")
print(f"  roles in dropdown: {len(roles)} | branches in dropdown: {len(branches)}")
print("  Fill the 'Staff' sheet, then upload via the staff-upload endpoint.")
