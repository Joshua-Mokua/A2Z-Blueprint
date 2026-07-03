#!/usr/bin/env python3
"""scripts/upload_staff_register.py — STRICT staff org-tree upload pipeline.

Ingests a filled STAFF_UPLOAD template (.xlsx), validates the entire tree, and
(unless --dry-run) wipe-and-replaces the staff register + Postgres users table,
preserving designated test logins.

VALIDATION (whole upload REJECTED if any fails):
  - Role  ∈ configured hierarchy roles
  - Branch ∈ configured branches (the 16 + Head Office)
  - Exactly ONE root (blank Reports To Code) = the MD
  - Every non-blank Reports To Code resolves to a Staff Code in the sheet
  - No reporting cycles
  - Staff Code unique and non-blank
  - DSA dotted-line codes (if present) resolve to existing Staff Codes

DRY-RUN (--dry-run): validates + prints the resolved tree (who reports to MD
down) + the redistribution summary. WRITES NOTHING.

APPLY (default): writes data/staff_register.xlsx (scope/cascade source) AND
rebuilds Postgres users (auth/login), preserving --keep logins. Backs up both.

Usage:
  python scripts\\upload_staff_register.py <filled.xlsx> --dry-run
  python scripts\\upload_staff_register.py <filled.xlsx> --apply --keep william001,admin
"""
import sys, json, shutil
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

ROOT = Path(__file__).resolve().parent.parent
REGISTER = ROOT / "data" / "staff_register.xlsx"

REG_COLS = ["Staff Code", "Staff Name", "Role", "Unit", "Region", "Category",
            "Department", "Band", "Gender", "Reports To", "Date of Employment"]

def load_config_sets():
    from utils.core import get_org_config
    cfg = get_org_config() or {}
    roles = set(cfg.get("hierarchy", {}).keys())
    branches = set(b["name"] for b in cfg.get("branches", []))
    return roles, branches

def read_upload(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb["Staff"] if "Staff" in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}
    need = ["Staff Code", "Staff Name", "Role", "Branch", "Reports To Code"]
    for n in need:
        if n not in idx:
            raise SystemExit(f"REJECT: template missing required column '{n}'")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):  # skip blank rows
            continue
        sc = str(r[idx["Staff Code"]]).strip() if r[idx["Staff Code"]] is not None else ""
        name = str(r[idx["Staff Name"]]).strip() if r[idx["Staff Name"]] is not None else ""
        if name.startswith("<") or not sc:  # skip the example stubs
            if name.startswith("<"):
                continue
        rows.append({
            "code": sc,
            "name": name,
            "role": str(r[idx["Role"]]).strip() if r[idx["Role"]] is not None else "",
            "branch": str(r[idx["Branch"]]).strip() if r[idx["Branch"]] is not None else "",
            "region": str(r[idx.get("Region (DSA only)", -1)]).strip() if idx.get("Region (DSA only)") is not None and r[idx["Region (DSA only)"]] is not None else "",
            "reports_to": str(r[idx["Reports To Code"]]).strip() if r[idx["Reports To Code"]] is not None else "",
            "dotted1": str(r[idx.get("Dotted Line Code 1", -1)]).strip() if idx.get("Dotted Line Code 1") is not None and r[idx["Dotted Line Code 1"]] is not None else "",
            "dotted2": str(r[idx.get("Dotted Line Code 2", -1)]).strip() if idx.get("Dotted Line Code 2") is not None and r[idx["Dotted Line Code 2"]] is not None else "",
            "band": str(r[idx.get("Band", -1)]).strip() if idx.get("Band") is not None and r[idx["Band"]] is not None else "",
            "gender": str(r[idx.get("Gender", -1)]).strip() if idx.get("Gender") is not None and r[idx["Gender"]] is not None else "",
            "email": str(r[idx.get("Email", -1)]).strip() if idx.get("Email") is not None and r[idx["Email"]] is not None else "",
            "doe": r[idx.get("Date of Employment", -1)] if idx.get("Date of Employment") is not None else "",
        })
    wb.close()
    return rows

def validate(rows, roles, branches):
    errs = []
    codes = [r["code"] for r in rows]
    code_set = set(codes)
    # unique codes
    dups = [c for c, n in Counter(codes).items() if n > 1]
    if dups:
        errs.append(f"Duplicate Staff Codes: {dups[:10]}")
    # roles + branches
    for r in rows:
        if r["role"] not in roles:
            errs.append(f"{r['code']} ({r['name']}): invalid Role '{r['role']}'")
        if r["branch"] not in branches:
            errs.append(f"{r['code']} ({r['name']}): invalid Branch '{r['branch']}'")
        if r["reports_to"] and r["reports_to"] not in code_set:
            errs.append(f"{r['code']} ({r['name']}): Reports To Code '{r['reports_to']}' not found")
        for d in (r["dotted1"], r["dotted2"]):
            if d and d not in code_set:
                errs.append(f"{r['code']} ({r['name']}): dotted-line code '{d}' not found")
    # exactly one root
    roots = [r for r in rows if not r["reports_to"]]
    if len(roots) == 0:
        errs.append("No root: exactly one row (the MD) must have a blank Reports To Code")
    elif len(roots) > 1:
        errs.append(f"Multiple roots ({len(roots)}): only the MD may have blank Reports To Code -> {[r['code'] for r in roots][:10]}")
    # cycle detection
    parent = {r["code"]: r["reports_to"] for r in rows}
    for r in rows:
        seen, cur, steps = set(), r["code"], 0
        while cur and cur in parent and parent[cur]:
            cur = parent[cur]
            if cur in seen or steps > 10000:
                errs.append(f"Cycle detected involving {r['code']}")
                break
            seen.add(cur); steps += 1
    return errs, roots

def print_tree(rows, roots):
    children = defaultdict(list)
    by_code = {r["code"]: r for r in rows}
    for r in rows:
        children[r["reports_to"]].append(r)
    def walk(code, depth, limit_depth=3):
        for ch in sorted(children.get(code, []), key=lambda x: x["role"]):
            indent = "  " * depth
            extra = f" [{ch['branch']}]" if ch["branch"] != "Head Office" else ""
            print(f"  {indent}- {ch['name']} ({ch['role']}){extra}")
            if depth < limit_depth:
                walk(ch["code"], depth + 1, limit_depth)
            elif children.get(ch["code"]):
                print(f"  {indent}    ...{len(_descendants(children, ch['code']))} below")
    for root in roots:
        print(f"  {root['name']} ({root['role']})  <- ROOT")
        walk(root["code"], 1)

def _descendants(children, code):
    out, stack = [], list(children.get(code, []))
    while stack:
        n = stack.pop(); out.append(n); stack.extend(children.get(n["code"], []))
    return out

def main():
    args = sys.argv[1:]
    if not args or args[0].startswith("--"):
        raise SystemExit("usage: upload_staff_register.py <filled.xlsx> [--dry-run|--apply] [--keep a,b]")
    upload = Path(args[0])
    dry = "--dry-run" in args or "--apply" not in args
    keep = []
    if "--keep" in args:
        keep = args[args.index("--keep") + 1].split(",")

    roles, branches = load_config_sets()
    rows = read_upload(upload)
    print(f"  read {len(rows)} staff rows from {upload.name}")
    print(f"  config: {len(roles)} valid roles, {len(branches)} valid branches")

    errs, roots = validate(rows, roles, branches)
    if errs:
        print(f"\n  ✗ VALIDATION FAILED — {len(errs)} error(s). NOTHING written.\n")
        for e in errs[:40]:
            print(f"      - {e}")
        if len(errs) > 40:
            print(f"      ... and {len(errs)-40} more")
        sys.exit(1)

    print(f"\n  ✓ VALIDATION PASSED. Root: {roots[0]['name']} ({roots[0]['role']})")
    # redistribution summary
    by_branch = Counter(r["branch"] for r in rows)
    by_role = Counter(r["role"] for r in rows)
    direct_to_md = [r for r in rows if r["reports_to"] == roots[0]["code"]]
    print(f"\n  REDISTRIBUTION:")
    print(f"    reporting directly to MD: {len(direct_to_md)}")
    for r in sorted(direct_to_md, key=lambda x: x["role"]):
        print(f"      - {r['name']} ({r['role']})")
    print(f"    staff per branch:")
    for b, n in sorted(by_branch.items()):
        print(f"      {n:>4}  {b}")

    print(f"\n  RESOLVED TREE (top 3 levels):")
    print_tree(rows, roots)

    if dry:
        print(f"\n  --dry-run: validated + tree resolved. NOTHING written.")
        print(f"  To apply: rerun with --apply --keep <logins-to-preserve>")
        return

    # APPLY: write register + rebuild PG users (preserve keep-list)
    print(f"\n  APPLYING (preserve logins: {keep})...")
    _write_register(rows)
    _rebuild_pg_users(rows, keep)
    print("  ✓ applied. Restart API + invalidate roster cache.")

def _write_register(rows):
    from openpyxl import Workbook
    if REGISTER.exists():
        shutil.copy2(REGISTER, REGISTER.with_suffix(f".xlsx.pre_upload_{datetime.now():%Y%m%d-%H%M%S}"))
    wb = Workbook(); ws = wb.active; ws.title = "Staff"; ws.append(REG_COLS)
    by_code = {r["code"]: r for r in rows}
    for r in rows:
        mgr = by_code.get(r["reports_to"])
        reports_to_name = mgr["role"] if mgr else ""  # register stores role-name style
        ws.append([r["code"], r["name"], r["role"], r["branch"], r["region"], "",
                   "", r["band"], r["gender"], reports_to_name, r["doe"]])
    wb.save(REGISTER)
    print(f"    wrote {REGISTER.name} ({len(rows)} staff)")

def _rebuild_pg_users(rows, keep):
    from utils.db import db
    from utils.core_audit import _hash_password
    keep_set = set(keep)
    # backup current users
    cur = db.fetch_all("SELECT username FROM users") or []
    print(f"    PG users before: {len(cur)}")
    # delete all except keep-list
    if keep_set:
        ph = ",".join(["%s"] * len(keep_set))
        db.execute(f"DELETE FROM users WHERE username NOT IN ({ph})", tuple(keep_set))
    else:
        db.execute("DELETE FROM users", ())
    # insert the uploaded staff (username defaults to staff_code)
    for r in rows:
        uname = r["code"]
        if uname in keep_set:
            continue
        pw = _hash_password(f"EcoStaff{r['code'][-4:]}")
        db.execute(
            """INSERT INTO users (username, password_hash, full_name, role, unit, region,
                                  staff_code, band, gender, active, is_admin, must_change_password)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,true,false,true)
               ON CONFLICT (username) DO UPDATE SET
                 full_name=EXCLUDED.full_name, role=EXCLUDED.role, unit=EXCLUDED.unit,
                 region=EXCLUDED.region, staff_code=EXCLUDED.staff_code, active=true""",
            (uname, pw, r["name"], r["role"], r["branch"], r["region"],
             r["code"], r["band"], r["gender"]))
    after = db.fetch_all("SELECT username FROM users") or []
    print(f"    PG users after: {len(after)}")

if __name__ == "__main__":
    main()
