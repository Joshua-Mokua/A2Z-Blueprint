"""pages/8_export.py — Export module."""
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
from utils.core import *
try:
    from utils.core import get_fiscal_year as _gfy
except: _gfy = lambda: _gfy()


from pages._shared import load_shared_state
from pages._access import require_access, get_my_scope
require_access("export")


# Load session state
um, ud, uname, em, ri_pm, prod_m, pm, lm, hr_m, casc, vm, rlm = load_shared_state()

# Shared data
uploaded_file = st.session_state.get("uploaded_file")
staff_scores  = st.session_state.get("staff_scores", pd.DataFrame())
df_proc       = st.session_state.get("df_processed", pd.DataFrame())
filtered      = st.session_state.get("filtered_staff", pd.DataFrame())
all_months    = st.session_state.get("all_months", [])

st.subheader("Export reports")

# ── Header ────────────────────────────────────────────────────────────
st.markdown(
    "<div style='padding:14px 18px;background:var(--brand-primary,#006B3F);border-radius:10px;"
    "margin-bottom:16px'>"
    "<div style='color:var(--color-background-primary);font-size:14px;font-weight:600'>📥 Data exports</div>"
    "<div style='color:rgba(255,255,255,0.65);font-size:11px;margin-top:2px'>"
    "Download BSC data, pipeline, HR records, cascade status and more as Excel or CSV</div>"
    "</div>", unsafe_allow_html=True)

ex1, ex2 = st.columns([2,1])
report_type = ex1.selectbox("Report type", [
    "Staff Rankings",
    "Detailed KPIs",
    "Monthly Data",
    "Validation Summary",
    "Cascade Status",
    "Pipeline Deals",
    "HR — Leave Records",
    "HR — Exits",
    "Branch Log Summary",
    "Commission Summary",
])
export_fmt  = ex2.radio("Format", ["Excel (.xlsx)","CSV"], horizontal=True, key="exp_fmt")
stamp = datetime.now().strftime("%Y%m%d_%H%M")

if report_type == "Staff Rankings":
    exp = filtered[['Overall_Rank','Staff Name','Role','Unit','Final_BSC_Score',
                     'Avg_Achievement_Pct','Performance_Remark','Percentile']].copy()
    exp.columns = ['Rank','Name','Role','Unit','BSC Score','Avg Achievement %','Status','Percentile']
    fname = f"A2Z_Rankings_{stamp}.csv"

elif report_type == "Detailed KPIs":
    exp = df_proc[df_proc['Staff Name'].isin(filtered['Staff Name'])][
        ['Staff Name','Role','KPI','Pillar','Annual Target','YTD_Actual',
         'Percent_Achieved','Score','Weight','Weighted_Score']].copy()
    fname = f"A2Z_KPIs_{stamp}.csv"

elif report_type == "Monthly Data":
    rows = []
    filt_names = filtered['Staff Name'].tolist()
    for _, r in df_proc[df_proc['Staff Name'].isin(filt_names)].iterrows():
        for col in all_months:
            if col in r.index:
                dt = parse_month_column(col)
                rows.append({"Staff Name": r['Staff Name'], "Role": r['Role'],
                             "KPI": r['KPI'],
                             "Month": dt.strftime("%b %Y") if dt else col,
                             "Actual": r[col]})
    exp = pd.DataFrame(rows) if rows else pd.DataFrame()
    fname = f"A2Z_Monthly_{stamp}.csv"

else:  # Validation Summary
    vrows = []
    for name in filtered['Staff Name'].tolist():
        v  = vm.get(name, datetime.now().strftime("%b %Y"))
        sc = filtered[filtered['Staff Name']==name]['Final_BSC_Score'].values[0]
        vrows.append({
            "Staff": name, "BSC Score": fmt_score(sc),
            "Validation Status": v['status'] if v else "Pending",
            "Action Plan": v.get('action_plan','') if v else "",
            "Comments": v.get('comments','') if v else "",
            "Validated By": v['manager'] if v else "",
            "Date": v['validated_at'][:10] if v else "",
        })
    exp = pd.DataFrame(vrows)
    fname = f"A2Z_Validation_{stamp}.csv"

# ── Additional report types ──────────────────────────────────────────
if report_type == "Cascade Status":
    rows = []
    if casc:
        for uname_c, udata_c in um.users.items():
            sc_c   = str(udata_c.get("staff_code", uname_c))
            name_c = udata_c.get("full_name", uname_c)
            given  = casc.get_what_i_was_given(sc_c, _gfy(), name_c)
            locked = casc.targets_locked(sc_c, _gfy(), name_c)
            rows.append({
                "Username":      uname_c,
                "Name":          name_c,
                "Role":          udata_c.get("role",""),
                "KPIs received": len(given),
                "Locked":        "Yes" if locked else "No",
                "Status":        ("Locked" if locked else
                                  "Received" if given else "Pending"),
            })
    exp   = pd.DataFrame(rows)
    fname = f"A2Z_Cascade_Status_{stamp}"

elif report_type == "Pipeline Deals":
    deals = pm.get_deals() if pm else []
    exp   = pd.DataFrame(deals) if deals else pd.DataFrame()
    if not exp.empty:
        exp["deal_value"] = exp["deal_value"].apply(lambda x: float(x or 0))
        exp["probability"] = exp["probability"].apply(lambda x: float(x or 0))
        exp["weighted_value"] = exp["deal_value"] * exp["probability"]
    fname = f"A2Z_Pipeline_{stamp}"

elif report_type == "HR — Leave Records":
    recs  = lm.records if lm else []
    exp   = pd.DataFrame(recs) if recs else pd.DataFrame()
    fname = f"A2Z_Leave_{stamp}"

elif report_type == "HR — Exits":
    exits = hr_m.get_exits(months=120) if hr_m else []
    exp   = pd.DataFrame(exits) if exits else pd.DataFrame()
    fname = f"A2Z_Exits_{stamp}"

elif report_type == "Branch Log Summary":
    blm_inst = st.session_state.get("branch_log_manager")
    logs     = blm_inst.get_all() if blm_inst else []
    exp      = pd.DataFrame(logs) if logs else pd.DataFrame()
    fname    = f"A2Z_BranchLog_{stamp}"

elif report_type == "Commission Summary":
    comm_rows = []
    if not df_proc.empty:
        for staff_nm in df_proc["Staff Name"].unique():
            sr2 = filtered[filtered["Staff Name"]==staff_nm]
            if len(sr2):
                remark2 = sr2.iloc[0].get("Performance_Remark","Met")
                from utils.core import COMMISSION_TIERS, BONUS_MULTIPLIERS
                total_c = sum(
                    __import__("utils.core", fromlist=["compute_commission"])
                    .compute_commission(kpi, float(kdf["YTD_Actual"].values[0]),
                                        float(kdf["Annual Target"].values[0]), remark2)
                    for kpi in COMMISSION_TIERS
                    for kdf in [df_proc[(df_proc["Staff Name"]==staff_nm)&(df_proc["KPI"]==kpi)]]
                    if len(kdf) and float(kdf["Annual Target"].values[0]) > 0
                )
                comm_rows.append({"Staff":staff_nm,"Remark":remark2,"Commission KES":total_c})
    exp   = pd.DataFrame(comm_rows)
    fname = f"A2Z_Commission_{stamp}"

# ── Preview + download ───────────────────────────────────────────────
if not exp.empty:
    st.markdown(f"**Preview** — {len(exp)} row(s)")
    st.dataframe(exp.head(50), use_container_width=True, hide_index=True)

    if export_fmt == "Excel (.xlsx)":
        try:
            import io as _io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            _wb = Workbook()
            _ws = _wb.active
            _ws.title = report_type[:31]
            # Header
            _ws.append(list(exp.columns))
            for cell in _ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor="006B3F")
                cell.alignment = Alignment(horizontal="center")
            # Data
            for row in exp.itertuples(index=False):
                _ws.append(list(row))
            # Widths
            for ci in range(1, len(exp.columns)+1):
                _ws.column_dimensions[get_column_letter(ci)].width = 18
            _ws.freeze_panes = "A2"
            _buf = _io.BytesIO()
            _wb.save(_buf)
            _buf.seek(0)
            st.download_button(
                f"⬇️ Download {report_type} (.xlsx)",
                data=_buf.getvalue(),
                file_name=f"{fname}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as _e:
            st.error(f"Excel export error: {_e}. Falling back to CSV.")
            st.download_button(f"⬇️ Download {report_type} (.csv)",
                               exp.to_csv(index=False).encode(),
                               f"{fname}.csv", "text/csv")
    else:
        st.download_button(f"⬇️ Download {report_type} (.csv)",
                           exp.to_csv(index=False).encode(),
                           f"{fname}.csv", "text/csv")
else:
    st.info("No data to export for the selected report type.")

