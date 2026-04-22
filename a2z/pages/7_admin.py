"""pages/7_admin.py — Administration: users, permissions, reporting lines, audit."""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
from utils.core import *
try:
    from utils.core import get_fiscal_year as _gfy
except: _gfy = lambda: _gfy()

from pages._shared import load_shared_state
from pages._access import require_access, get_my_scope
require_access("admin")


um, ud, uname, em, ri_pm, prod_m, pm, lm, hr_m, casc, vm, rlm = load_shared_state()

role_low = str(ud.get("role","")).lower()
if not (ud.get("can_view_all") or "admin" in role_low):
    st.error("⛔ Access restricted to administrators.")
    st.stop()

staff_scores = st.session_state.get("staff_scores",  pd.DataFrame())
registry     = st.session_state.get("staff_registry", pd.DataFrame())
# Roles and units from org_config (source of truth), fallback to staff_scores
try:
    from utils.core import get_org_config as _goc2, get_all_branches as _gab2
    _oc2 = _goc2()
    avail_roles = sorted(_oc2.get("roles", []))
    avail_units = ["Head Office"] + sorted(
        b["name"] for b in _gab2() if b.get("type","") != "HO")
    if not avail_roles: raise ValueError("empty")
except:
    avail_roles = sorted(staff_scores["Role"].unique().tolist()) if len(staff_scores) and "Role" in staff_scores.columns else ["Staff"]
    avail_units = sorted(staff_scores["Unit"].unique().tolist()) if len(staff_scores) and "Unit" in staff_scores.columns else ["Head Office"]

st.markdown(
    "<div style=\'padding:16px 22px;background:var(--brand-primary,#006B3F);border-radius:12px;margin-bottom:20px;box-shadow:0 2px 12px rgba(0,0,0,0.15)\'><div style=\'display:flex;align-items:center;justify-content:space-between\'><div><div style=\'color:var(--color-background-primary);font-size:16px;font-weight:700;letter-spacing:-0.2px\'>System Administration</div><div style=\'color:rgba(255,255,255,0.65);font-size:11px;margin-top:3px;font-weight:400\'>Users · Permissions · Reporting lines · Audit · Upload format</div></div><div style=\'opacity:0.12;font-size:36px;line-height:1;color:white\'>◆</div></div></div>",
    unsafe_allow_html=True)

tabs = st.tabs([
    "🏦 Organisation",
    "👤 Users",
    "🔑 Permissions",
    "🗂️ Staff movements",
    "🌳 Org tree",
    "📋 Audit log",
    "📤 Upload format",
    "🏖️ Leave settings",
    "📚 KPI Library",
    "💼 Pipeline settings",
    "⚙️ System health",
])

# ════════════════════════════════════════════════════════════════
# TAB 0 — ORGANISATION SETUP
# ════════════════════════════════════════════════════════════════
with tabs[0]:
    from utils.core import get_org_config, save_org_config, DATA_DIR as _org_DATA_DIR
    _data_dir = _org_DATA_DIR

    st.subheader("🏦 Organisation Setup")
    st.caption(
        "Configure this bank's structure — name, branches, regions, reporting hierarchy, and roles. "
        "All other modules (BSC, Cascade, Pipeline, CBS) read from this configuration automatically.")

    _org = get_org_config()

    _org_view = st.radio("",
        ["🏦 Bank identity","🏢 Branches & regions","🌳 Hierarchy & roles","📥 Bulk staff import","🔧 Reset & migrate"],
        horizontal=True, key="org_view")
    st.markdown("---")

    # ══════════════════════════════════════════════════════════════
    if "Bank identity" in _org_view:
    # ══════════════════════════════════════════════════════════════
        st.markdown("**Configure the bank's identity and display settings.**")
        with st.form("org_identity_form"):
            _c1, _c2 = st.columns(2)
            _bank_name = _c1.text_input("Bank name *", value=_org.get("bank_name",""))
            _bank_code = _c2.text_input("Bank code (e.g. ECO)", value=_org.get("bank_code","ECO"),
                                         help="Short code used for account number prefixes")
            _country   = _c1.text_input("Country", value=_org.get("country","Kenya"))
            _currency  = _c2.text_input("Currency code (e.g. KES)", value=_org.get("currency","KES"))
            _curr_sym  = _c1.text_input("Currency symbol", value=_org.get("currency_symbol","KES"))

            st.markdown("**Pillar weights** — how the BSC score is computed bank-wide")
            _pw1,_pw2,_pw3,_pw4 = st.columns(4)
            _pillar_wts = _org.get("pillar_weights",{})
            def _pct(key, default_pct):
                v = _pillar_wts.get(key, default_pct/100)
                # Normalise: if stored as decimal (0.40), convert to int pct (40)
                return int(v * 100) if v <= 1.0 else int(v)
            _fin_wt  = _pw1.number_input("Financial %",  0, 100, _pct("Financial", 40))
            _cust_wt = _pw2.number_input("Customer %",   0, 100, _pct("Customer Focus", 25))
            _ops_wt  = _pw3.number_input("Operations %", 0, 100, _pct("Operational Excellence", 25))
            _ppl_wt  = _pw4.number_input("People %",     0, 100, _pct("People & Learning", 10))
            _wt_total = _fin_wt+_cust_wt+_ops_wt+_ppl_wt
            _wt_clr = "#10B981" if _wt_total==100 else "#EF4444"
            _wt_bg   = "#ECFDF5" if _wt_total==100 else "#FEF2F2"
            _wt_icon = "✅" if _wt_total==100 else "— must equal 100%"
            st.markdown(
                f"<div style='padding:6px 12px;background:{_wt_bg};"
                f"border-radius:6px;font-size:12px;font-weight:700;color:{_wt_clr}'>"
                f"Total: {_wt_total}% {_wt_icon}</div>",
                unsafe_allow_html=True)

            if st.form_submit_button("💾 Save bank identity", type="primary"):
                if _wt_total != 100:
                    st.error("Pillar weights must total 100%")
                else:
                    _org["bank_name"] = _bank_name.strip()
                    _org["bank_code"] = _bank_code.strip().upper()
                    _org["country"]   = _country.strip()
                    _org["currency"]  = _currency.strip().upper()
                    _org["currency_symbol"] = _curr_sym.strip()
                    _org["pillar_weights"] = {
                        "Financial": _fin_wt/100, "Customer Focus": _cust_wt/100,
                        "Operational Excellence": _ops_wt/100, "People & Learning": _ppl_wt/100}
                    save_org_config(_org)
                    audit_log("ORG_IDENTITY_SAVED", uname, _bank_name)
                    st.success(f"✅ Bank identity saved — {_bank_name}")
                    st.rerun()

    # ══════════════════════════════════════════════════════════════
    elif "Branches" in _org_view:
    # ══════════════════════════════════════════════════════════════
        st.markdown("**Manage branches and regions.** Changes apply immediately to BSC scoping, cascade, and CBS lookup.")

        _branches = list(_org.get("branches", []))
        _regions  = list(_org.get("regions", []))

        # Summary strip
        _bm1,_bm2,_bm3,_bm4 = st.columns(4)
        _bm1.metric("Branches", len(_branches))
        _bm2.metric("Regions",  len(set(b.get("region","") for b in _branches)))
        _tc = {}
        for _b in _branches: _tc[_b.get("tier",3)] = _tc.get(_b.get("tier",3),0)+1
        _bm3.metric("Flagship (T1)", _tc.get(1,0))
        _bm4.metric("Light (T4)",    _tc.get(4,0))

        # Live table
        if _branches:
            _bdf_show = pd.DataFrame([{
                "Code":b["code"],"Branch Name":b["name"],"Region":b.get("region",""),
                "County":b.get("county",""),"Type":b.get("type",""),"Tier":b.get("tier",3)
            } for b in _branches])
            st.dataframe(_bdf_show, use_container_width=True, hide_index=True, height=280)

        st.markdown("---")
        _br_action = st.radio("Action",
            ["➕ Add branch","✏️ Edit branch","✏️ Rename branch","🗑️ Delete branch",
             "🗺️ Regions"],
            horizontal=True, key="br_action")

        if "Add branch" in _br_action:
            with st.form("add_branch_form"):
                _ab1,_ab2 = st.columns(2)
                _new_bcode = _ab1.text_input("Branch code *", placeholder="e.g. BRN036")
                _new_bname = _ab2.text_input("Branch name *", placeholder="e.g. Kisumu Mega Branch")
                _new_breg  = _ab1.selectbox("Region *",
                    _regions + (["+ New region"] if _regions else ["Head Office","+ New region"]))
                _new_bnewreg = _ab2.text_input("New region name (if selected above)", placeholder="e.g. South West")
                _new_bcnty = _ab1.text_input("County", placeholder="e.g. Kisumu")
                _new_btype = _ab2.selectbox("Type", ["Flagship","Main","Standard","Light","HO"])
                _new_btier = _ab1.selectbox("Tier", [1,2,3,4], help="1=Flagship 2=Main 3=Standard 4=Light")
                if st.form_submit_button("➕ Add branch", type="primary"):
                    if _new_bcode.strip() and _new_bname.strip():
                        _reg_to_use = _new_bnewreg.strip() if _new_breg=="+ New region" and _new_bnewreg.strip() else _new_breg
                        if _new_bcode.strip().upper() in [b["code"] for b in _branches]:
                            st.error(f"Code {_new_bcode} already exists.")
                        else:
                            _branches.append({
                                "code":  _new_bcode.strip().upper(),
                                "name":  _new_bname.strip(),
                                "region":_reg_to_use,
                                "county":_new_bcnty.strip(),
                                "type":  _new_btype,
                                "tier":  int(_new_btier),
                            })
                            if _reg_to_use not in _regions:
                                _regions.append(_reg_to_use)
                            _org["branches"] = _branches
                            _org["regions"]  = _regions
                            save_org_config(_org)
                            audit_log("BRANCH_ADDED", uname, _new_bname)
                            st.success(f"✅ '{_new_bname}' added.")
                            st.rerun()
                    else:
                        st.error("Branch code and name are required.")

        elif "Edit branch" in _br_action:
            _sel_branch = st.selectbox("Select branch", [b["name"] for b in _branches], key="br_edit_sel")
            _sel_b = next((b for b in _branches if b["name"]==_sel_branch), None)
            if _sel_b:
                with st.form("edit_branch_form"):
                    _eb1,_eb2 = st.columns(2)
                    _eb_region = _eb1.selectbox("Region",
                        _regions if _regions else ["Head Office"],
                        index=_regions.index(_sel_b.get("region","")) if _sel_b.get("region","") in _regions else 0)
                    _eb_county = _eb2.text_input("County", value=_sel_b.get("county",""))
                    _eb_type   = _eb1.selectbox("Type",
                        ["Flagship","Main","Standard","Light","HO"],
                        index=["Flagship","Main","Standard","Light","HO"].index(_sel_b.get("type","Standard"))
                        if _sel_b.get("type","Standard") in ["Flagship","Main","Standard","Light","HO"] else 2)
                    _eb_tier = _eb2.selectbox("Tier",[1,2,3,4],
                        index=max(0,min(3,_sel_b.get("tier",3)-1)))
                    if st.form_submit_button("💾 Save", type="primary"):
                        for _b in _branches:
                            if _b["code"]==_sel_b["code"]:
                                _b["region"]=_eb_region; _b["county"]=_eb_county.strip()
                                _b["type"]=_eb_type; _b["tier"]=int(_eb_tier); break
                        _org["branches"]=_branches; save_org_config(_org)
                        audit_log("BRANCH_EDITED",uname,_sel_branch)
                        st.success("✅ Updated."); st.rerun()

        elif "Rename branch" in _br_action:
            with st.form("rename_branch_form"):
                _ren_branch = st.selectbox("Branch to rename", [b["name"] for b in _branches])
                _ren_bname  = st.text_input("New name *", placeholder="e.g. Kisumu Central Branch")
                if st.form_submit_button("✏️ Rename", type="primary"):
                    _ren_bname = _ren_bname.strip()
                    if _ren_bname:
                        for _b in _branches:
                            if _b["name"]==_ren_branch:
                                _b["name"]=_ren_bname; break
                        _org["branches"]=_branches; save_org_config(_org)
                        # Also update users who have this branch as their unit
                        _ren_count=0
                        for _u,_ud2 in um.users.items():
                            if _ud2.get("unit","")==_ren_branch:
                                _ud2["unit"]=_ren_bname; _ren_count+=1
                        if _ren_count: um.save()
                        audit_log("BRANCH_RENAMED",uname,f"{_ren_branch}→{_ren_bname}")
                        st.success(f"✅ Renamed. {_ren_count} user(s) updated."); st.rerun()

        elif "Delete branch" in _br_action:
            with st.form("del_branch_form"):
                _del_branch = st.selectbox("Branch to delete", [b["name"] for b in _branches])
                st.caption("⚠️ Staff assigned to this branch will retain the branch name until reassigned.")
                if st.form_submit_button("🗑️ Delete branch", type="secondary"):
                    _org["branches"]=[b for b in _branches if b["name"]!=_del_branch]
                    save_org_config(_org)
                    audit_log("BRANCH_DELETED",uname,_del_branch)
                    st.success(f"✅ {_del_branch} deleted."); st.rerun()

        elif "Regions" in _br_action:
            st.markdown("**Manage regions** — used for Regional Head scoping and branch grouping.")
            _reg_action = st.radio("",["Add region","Rename region","Delete region"],
                horizontal=True, key="reg_action")

            if "Add" in _reg_action:
                with st.form("add_region_form"):
                    _new_reg = st.text_input("New region name *", placeholder="e.g. North Eastern")
                    if st.form_submit_button("➕ Add region", type="primary"):
                        if _new_reg.strip() and _new_reg.strip() not in _regions:
                            _regions.append(_new_reg.strip())
                            _org["regions"]=_regions; save_org_config(_org)
                            st.success(f"✅ '{_new_reg}' added."); st.rerun()
                        elif _new_reg.strip() in _regions:
                            st.error("Region already exists.")

            elif "Rename" in _reg_action:
                with st.form("rename_region_form"):
                    _ren_reg_from = st.selectbox("Region to rename",
                        _regions if _regions else ["— none —"])
                    _ren_reg_to   = st.text_input("New name *")
                    if st.form_submit_button("✏️ Rename", type="primary"):
                        _ren_reg_to = _ren_reg_to.strip()
                        if _ren_reg_to:
                            _regions = [_ren_reg_to if r==_ren_reg_from else r for r in _regions]
                            for _b in _branches:
                                if _b.get("region")==_ren_reg_from: _b["region"]=_ren_reg_to
                            _org["regions"]=_regions; _org["branches"]=_branches
                            save_org_config(_org)
                            st.success(f"✅ '{_ren_reg_from}' → '{_ren_reg_to}'. Branches updated."); st.rerun()

            else:  # Delete
                with st.form("del_region_form"):
                    _del_reg = st.selectbox("Region to delete",
                        _regions if _regions else ["— none —"])
                    st.caption("Branches in this region will show 'Head Office' until reassigned.")
                    if st.form_submit_button("🗑️ Delete region", type="secondary"):
                        _regions=[r for r in _regions if r!=_del_reg]
                        _org["regions"]=_regions; save_org_config(_org)
                        st.success(f"✅ '{_del_reg}' removed."); st.rerun()

            # Live region list
            if _regions:
                st.markdown("**Current regions:**")
                _rdf = pd.DataFrame([{
                    "Region": r,
                    "Branches": sum(1 for b in _branches if b.get("region")==r)
                } for r in _regions])
                st.dataframe(_rdf, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════
    elif "Hierarchy" in _org_view:
    # ══════════════════════════════════════════════════════════════
        st.markdown("**Define who reports to whom.** This drives cascade allocation, BSC scoping, and backup staff selection.")
        st.caption("⚠️ Changing the hierarchy takes effect immediately on the next page load.")

        _hierarchy = _org.get("hierarchy", {})
        _roles     = _org.get("roles", [])

        # Display current hierarchy as a clean table
        import pandas as pd
        _hier_rows = []
        for role, parents in _hierarchy.items():
            _hier_rows.append({
                "Role": role,
                "Reports to": ", ".join(parents) if parents else "— (Top of hierarchy)"
            })
        if _hier_rows:
            st.dataframe(pd.DataFrame(_hier_rows), use_container_width=True,
                         hide_index=True, height=350)

        st.markdown("---")
        _hc1, _hc2 = st.columns(2)

        # Edit reporting line for a role
        with _hc1:
            st.markdown("**Edit a role's reporting line**")
            with st.form("edit_hierarchy_form"):
                _sel_hier_role = st.selectbox("Role to configure", _roles, key="hier_role_sel")
                _cur_parents = _hierarchy.get(_sel_hier_role, [])
                _new_parents = st.multiselect(
                    "Reports to (select one or more)",
                    [r for r in _roles if r != _sel_hier_role],
                    default=[p for p in _cur_parents if p in _roles])
                if st.form_submit_button("💾 Save reporting line", type="primary"):
                    _hierarchy[_sel_hier_role] = _new_parents
                    _org["hierarchy"] = _hierarchy
                    save_org_config(_org)
                    audit_log("HIERARCHY_CHANGED", uname,
                              f"{_sel_hier_role}→{','.join(_new_parents)}")
                    st.success(f"✅ {_sel_hier_role} now reports to: {', '.join(_new_parents) or 'nobody (top)'}")
                    st.rerun()

        with _hc2:
            _role_action = st.radio("Action",
                ["➕ Add role","✏️ Rename role","🗑️ Remove role"],
                horizontal=True, key="role_action")

            if "Add" in _role_action:
                with st.form("add_role_form"):
                    _new_role = st.text_input("New role name *",
                        placeholder="e.g. Digital Banking Manager / CEO / Chief")
                    _new_role_parent = st.selectbox(
                        "Reports to", ["— (Top level)"] + _roles)
                    _new_role_kpis = st.multiselect(
                        "Clone KPIs from similar role (optional)",
                        ["— None (configure later)"] + _roles,
                        max_selections=1, key="new_role_kpi_src")
                    if st.form_submit_button("➕ Add role", type="primary"):
                        if _new_role.strip() and _new_role.strip() not in _roles:
                            _nr = _new_role.strip()
                            _roles.append(_nr)
                            _hierarchy[_nr] = ([] if _new_role_parent=="— (Top level)"
                                               else [_new_role_parent])
                            # Clone KPIs from similar role
                            if _new_role_kpis and _new_role_kpis[0] != "— None (configure later)":
                                try:
                                    from utils.core import get_kpi_library, save_kpi_library
                                    _klib = get_kpi_library()
                                    _src_kpis = _klib.get("role_kpis",{}).get(_new_role_kpis[0],[])
                                    if _src_kpis:
                                        _klib["role_kpis"][_nr] = list(_src_kpis)
                                        save_kpi_library(_klib)
                                except: pass
                            _org["roles"] = _roles
                            _org["hierarchy"] = _hierarchy
                            save_org_config(_org)
                            audit_log("ROLE_ADDED", uname, _nr)
                            st.success(f"✅ Role '{_nr}' added.")
                            st.rerun()
                        elif _new_role.strip() in _roles:
                            st.error("That role already exists.")

            elif "Rename" in _role_action:
                st.caption("Renames the role everywhere: users, staff register, KPI library, hierarchy, cascade — all updated atomically.")
                with st.form("rename_role_form"):
                    _ren_from = st.selectbox("Role to rename", _roles, key="ren_from")
                    _ren_to   = st.text_input("New name *",
                        placeholder="e.g. rename 'Managing Director' → 'CEO'")
                    if st.form_submit_button("✏️ Rename role", type="primary"):
                        _ren_to = _ren_to.strip()
                        if not _ren_to:
                            st.error("New name cannot be empty.")
                        elif _ren_to in _roles and _ren_to != _ren_from:
                            st.error(f"'{_ren_to}' already exists.")
                        else:
                            # 1. Roles list
                            _roles = [_ren_to if r==_ren_from else r for r in _roles]
                            # 2. Hierarchy keys and values
                            _new_hier = {}
                            for r, parents in _hierarchy.items():
                                _new_key = _ren_to if r==_ren_from else r
                                _new_parents = [_ren_to if p==_ren_from else p for p in parents]
                                _new_hier[_new_key] = _new_parents
                            _hierarchy = _new_hier
                            _org["roles"] = _roles
                            _org["hierarchy"] = _hierarchy
                            save_org_config(_org)
                            # 3. KPI library role_kpis
                            try:
                                from utils.core import get_kpi_library, save_kpi_library
                                _klib = get_kpi_library()
                                _rk = _klib.get("role_kpis", {})
                                if _ren_from in _rk:
                                    _rk[_ren_to] = _rk.pop(_ren_from)
                                _klib["role_kpis"] = _rk
                                save_kpi_library(_klib)
                            except: pass
                            # 4. Users — rename role field
                            _ren_users = 0
                            for _u, _ud2 in um.users.items():
                                if _ud2.get("role","") == _ren_from:
                                    _ud2["role"] = _ren_to
                                    _ren_users += 1
                            if _ren_users: um.save()
                            # 5. Use rename_role_everywhere for org_config + kpi_library
                            try:
                                from utils.core import rename_role_everywhere as _rre
                                _rre(_ren_from, _ren_to)
                            except: pass
                            # 6. Update staff_register.xlsx
                            _sr_updated = 0
                            try:
                                import openpyxl as _ox
                                from pathlib import Path as _Pth
                                _sr_path = _Pth("a2z/data/staff_register.xlsx")
                                if _sr_path.exists():
                                    _sr_wb = _ox.load_workbook(str(_sr_path))
                                    _sr_ws = _sr_wb.active
                                    _sr_hdrs = [_sr_ws.cell(1,c).value for c in range(1,_sr_ws.max_column+1)]
                                    _sr_rc = _sr_hdrs.index("Role")+1 if "Role" in _sr_hdrs else None
                                    if _sr_rc:
                                        for _sr_row in _sr_ws.iter_rows(min_row=2):
                                            if _sr_row[_sr_rc-1].value == _ren_from:
                                                _sr_row[_sr_rc-1].value = _ren_to
                                                _sr_updated += 1
                                    _sr_wb.save(str(_sr_path))
                            except: pass
                            # 7. Update actuals xlsx
                            _act_updated = 0
                            try:
                                import glob as _gl
                                for _af in _gl.glob("a2z/data/actuals_*.xlsx"):
                                    _act_wb = _ox.load_workbook(_af)
                                    _act_ws = _act_wb.active
                                    _act_hdrs = [_act_ws.cell(2,c).value for c in range(1,_act_ws.max_column+1)]
                                    _act_rc = _act_hdrs.index("Role")+1 if "Role" in _act_hdrs else None
                                    if _act_rc:
                                        for _ar in _act_ws.iter_rows(min_row=3):
                                            if _ar[_act_rc-1].value == _ren_from:
                                                _ar[_act_rc-1].value = _ren_to
                                                _act_updated += 1
                                    _act_wb.save(_af)
                            except: pass
                            audit_log("ROLE_RENAMED", uname, f"{_ren_from}→{_ren_to}")
                            st.success(
                                f"✅ '{_ren_from}' renamed to '{_ren_to}' across all data — "
                                f"{_ren_users} user(s), {_sr_updated} staff register rows, "
                                f"{_act_updated} actuals rows, hierarchy, KPI library, "
                                f"role categories — all updated.")
                            st.rerun()

            else:  # Remove
                with st.form("remove_role_form"):
                    _del_role = st.selectbox("Role to remove", _roles, key="del_role_sel")
                    st.caption("⚠️ Staff with this role will retain it in their user record until reassigned.")
                    if st.form_submit_button("🗑️ Remove role", type="secondary"):
                        _roles = [r for r in _roles if r != _del_role]
                        _hierarchy.pop(_del_role, None)
                        for r in _hierarchy:
                            _hierarchy[r] = [p for p in _hierarchy[r] if p != _del_role]
                        _org["roles"] = _roles
                        _org["hierarchy"] = _hierarchy
                        save_org_config(_org)
                        audit_log("ROLE_REMOVED", uname, _del_role)
                        st.success(f"✅ '{_del_role}' removed."); st.rerun()

    # ══════════════════════════════════════════════════════════════
    elif "Bulk staff import" in _org_view:
    # ══════════════════════════════════════════════════════════════
        st.markdown("**Import staff roster from Excel or CSV.** Creates login accounts for all staff automatically.")
        st.info(
            "Expected columns: **Staff Name**, **Role**, **Unit** (branch name), **Staff Code** "
            "(optional — auto-assigned if missing), **Email** (optional), **Reports To** (optional).")

        _imp_file = st.file_uploader("Upload staff roster (.xlsx or .csv)",
                                      type=["xlsx","csv"], key="org_import_file")
        if _imp_file:
            try:
                import pandas as _pd, hashlib as _hl
                if _imp_file.name.endswith(".csv"):
                    _imp_df = _pd.read_csv(_imp_file)
                else:
                    _imp_df = _pd.read_excel(_imp_file)
                _imp_df.columns = [str(c).strip() for c in _imp_df.columns]

                st.success(f"✅ {len(_imp_df)} rows loaded")
                st.dataframe(_imp_df.head(10), use_container_width=True, hide_index=True)

                _req_cols = ["Staff Name","Role","Unit"]
                _missing  = [c for c in _req_cols if c not in _imp_df.columns]
                if _missing:
                    st.error(f"Missing required columns: {', '.join(_missing)}")
                else:
                    _next_code = max(
                        (int(str(u.get("staff_code","0")).strip() or 0)
                         for u in um.users.values()
                         if str(u.get("staff_code","0")).isdigit()),
                        default=300000) + 1

                    _preview = []
                    for _, row in _imp_df.iterrows():
                        name = str(row.get("Staff Name","")).strip()
                        if not name: continue
                        _sc = str(row.get("Staff Code","")).strip()
                        if not _sc or not _sc.isdigit():
                            _sc = str(_next_code); _next_code += 1
                        uname_gen = name.split()[0].lower() + _sc[-3:]
                        pwd_gen = f"{_org.get(chr(112)+ chr(97)+ chr(115)+ chr(115)+ chr(119)+ chr(111)+ chr(114)+ chr(100)+chr(95)+chr(112)+chr(114)+chr(101)+chr(102)+chr(105)+chr(120), chr(83)+chr(116)+chr(97)+chr(102)+chr(102))}{_sc[-4:]}"
                        _preview.append({
                            "Staff Name": name, "Role": row.get("Role",""),
                            "Unit": row.get("Unit",""), "Staff Code": _sc,
                            "Username": uname_gen, "Password": pwd_gen,
                        })
                    _prev_df = _pd.DataFrame(_preview)
                    st.markdown(f"**Preview — {len(_prev_df)} accounts will be created:**")
                    st.dataframe(_prev_df, use_container_width=True, hide_index=True, height=280)

                    if st.button(f"✅ Import {len(_prev_df)} staff", type="primary", key="org_import_run"):
                        _created = 0
                        for row in _preview:
                            _uname = row["Username"]
                            # Avoid duplicates
                            _base  = _uname
                            _sfx   = 1
                            while _uname in um.users:
                                _uname = f"{_base}{_sfx}"; _sfx += 1
                            pwd = row["Password"]
                            um.users[_uname] = {
                                "password": _hl.sha256(pwd.encode()).hexdigest(),
                                "full_name": row["Staff Name"],
                                "role": row["Role"],
                                "unit": row["Unit"],
                                "staff_code": row["Staff Code"],
                                "active": True,
                                "must_change_password": True,
                                "is_admin": False,
                                "can_view_all": False,
                            }
                            _created += 1
                        um.save()
                        audit_log("BULK_IMPORT", uname, f"{_created} staff imported")
                        st.success(f"✅ {_created} user accounts created. Default passwords: {_org.get('password_prefix','Staff')} + last 4 of staff code.")
                        st.rerun()
            except Exception as _ie:
                st.error(f"Import error: {_ie}")

    # ══════════════════════════════════════════════════════════════
    elif "Reset" in _org_view:
    # ══════════════════════════════════════════════════════════════
        st.markdown("**System reset and migration tools.**")
        st.warning("⚠️ These actions are irreversible. Use with caution.")

        with st.expander("🔄 Reset org config to defaults"):
            st.caption("Resets branches, regions, hierarchy and roles to the system defaults.")
            if st.button("Reset to system defaults", key="reset_org"):
                from utils.core import DEFAULT_ORG_CONFIG
                save_org_config(DEFAULT_ORG_CONFIG.copy())
                st.success("✅ Org config reset to defaults.")
                st.rerun()

        with st.expander("📤 Export all config as ZIP"):
            st.caption("Download all configuration files for backup or migration.")
            import io, zipfile, json as _json
            _zip_buf = io.BytesIO()
            with zipfile.ZipFile(_zip_buf, "w") as _zf:
                for _cfg_file in ["org_config.json","kpi_library.json","pipeline_settings.json",
                                  "leave_settings.json","bank_targets.json","target_cascade.json"]:
                    _cfg_path = _data_dir / _cfg_file
                    if _cfg_path.exists():
                        _zf.write(_cfg_path, _cfg_file)
            _zip_buf.seek(0)
            st.download_button(
                "⬇️ Download config backup",
                _zip_buf.read(),
                file_name=f"a2z_config_backup.zip",
                mime="application/zip")

        with st.expander("📥 Import config from backup ZIP"):
            st.caption("Restore configuration from a previously exported backup.")
            _imp_zip = st.file_uploader("Upload config ZIP", type=["zip"], key="cfg_import_zip")
            if _imp_zip:
                import io as _io, zipfile as _zf2
                with _zf2.ZipFile(_io.BytesIO(_imp_zip.read())) as _z:
                    _names = _z.namelist()
                    st.write("Files in backup:", _names)
                    if st.button("📥 Restore from backup", type="primary"):
                        for _n in _names:
                            (_data_dir / _n).write_bytes(_z.read(_n))
                        st.success("✅ Config restored. Restart the app to apply.")


# ════════════════════════════════════════════════════════════════
# TAB 1 — USERS
# ════════════════════════════════════════════════════════════════
with tabs[1]:
    st.subheader("User management")

    # ── Quick admin restore notice ────────────────────────────────
    if "admin" not in um.users:
        st.error(
            "⚠️ **Admin account missing.** Click below to restore it. "
            "Username: `admin` · Default password: `admin123`")
        if st.button("🔧 Restore admin account", type="primary", key="restore_admin"):
            um.users["admin"] = {
                "password":   __import__("hashlib").sha256(b"admin123").hexdigest(),
                "full_name":  "System Admin", "role": "Admin",
                "department": "All", "can_view_all": True,
                "managed_roles": [], "managed_units": [],
                "managed_staff_codes": [], "staff_code": "ADMIN001",
                "email": "admin@bank.com", "active": True, "_protected": True,
            }
            um.save()
            audit_log("ADMIN_RESTORED", uname, "admin")
            st.success("✅ Admin account restored. Password: admin123")
            st.rerun()
    else:
        # Show protected badge next to admin
        _admin_info = um.users.get("admin", {})
        st.markdown(
            f"<div style='padding:6px 12px;background:var(--brand-light,#E8F5EE);border:1px solid #BBF7D0;"
            f"border-radius:6px;font-size:11px;color:#166534;margin-bottom:12px'>"
            f"🔒 <b>admin</b> account is active and protected — cannot be deleted "
            f"· last active account: {_admin_info.get('full_name','System Admin')}"
            f"</div>", unsafe_allow_html=True)

    # ── Auto-sync staff codes from BSC data ────────────────────────
    df_bsc = st.session_state.get("df_processed", pd.DataFrame())
    if not df_bsc.empty and um.users:
        unsynced = []
        for uname_k, udata in um.users.items():
            fname = udata.get("full_name","")
            sc    = str(udata.get("staff_code","")).strip()
            if fname and (not sc or not sc.isdigit()):
                # Try to find in BSC data
                hits = df_bsc[df_bsc["Staff Name"] == fname]
                if hits.empty:
                    surname = fname.strip().split()[-1]
                    hits = df_bsc[df_bsc["Staff Name"].str.contains(surname, case=False, na=False)]
                if not hits.empty:
                    bsc_code = str(hits["Staff Code"].iloc[0]).strip()
                    if bsc_code and bsc_code != "nan":
                        unsynced.append((uname_k, fname, bsc_code))

        if unsynced:
            st.markdown(
                f"<div style='padding:10px 14px;background:#FEF3C7;border:1px solid #FDE68A;"
                f"border-radius:8px;margin-bottom:10px;font-size:12px;color:#92400E'>"
                f"⚠️ <b>{len(unsynced)} user account(s)</b> are missing their numeric staff code "
                f"(needed for cascade to work). Click below to auto-sync from BSC data."
                f"</div>", unsafe_allow_html=True)
            sc1, sc2 = st.columns([3,1])
            with sc1:
                for _, fname, bsc_code in unsynced:
                    st.markdown(f"  - **{fname}** → will set staff code to `{bsc_code}`")
            if sc2.button("🔄 Sync staff codes", type="primary", use_container_width=True):
                synced = 0
                for uname_k, fname, bsc_code in unsynced:
                    um.users[uname_k]["staff_code"] = bsc_code
                    synced += 1
                um.save()
                audit_log("STAFF_CODE_SYNC", "admin", f"{synced} users synced")
                st.toast(f"✅ {synced} staff codes synced from BSC data", icon="✅")
                st.rerun()

    # ── Permission safety check ──────────────────────────────────────
    try:
        from utils.core import fix_view_all_permissions, _ALL_VIEW_ROLES
        _danger = [(u, d.get("full_name",u), d.get("role",""))
                   for u,d in um.users.items()
                   if d.get("can_view_all")
                   and not d.get("is_admin")
                   and str(d.get("role","")).lower() not in _ALL_VIEW_ROLES]
        if _danger:
            st.markdown(
                f"<div style='padding:10px 14px;background:#FEF2F2;border:1px solid #FECACA;"
                f"border-radius:8px;margin-bottom:10px;font-size:12px;color:#991B1B'>"
                f"<b>⚠️ Privacy risk detected:</b> "
                f"{len(_danger)} account(s) have 'View ALL staff' enabled but are not MD/Admin: "
                f"{', '.join(n for _,n,_ in _danger)}. "
                f"Click below to fix immediately.</div>", unsafe_allow_html=True)
            if st.button("🔒 Fix — remove 'View all' from non-MD accounts",
                          type="primary", key="fix_view_all_btn"):
                n = fix_view_all_permissions(um)
                st.session_state["_filtered_for"] = None
                st.toast(f"✅ Fixed {n} accounts — view-all removed", icon="🔒")
                st.rerun()
    except Exception as _pe:
        pass

    mode = st.radio("Action", ["Create new user","Edit existing user"],
                    horizontal=True, key="admin_mode")

    if mode == "Edit existing user":
        existing = list(um.users.keys())
        if not existing:
            st.info("No users yet.")
        else:
            sel_user = st.selectbox("Select user", existing, key="edit_sel")
            eu = um.users.get(sel_user, {})
            with st.form("edit_user_form"):
                ec1, ec2 = st.columns(2)
                e_fname  = ec1.text_input("Full name",  value=eu.get("full_name",""))
                e_email  = ec2.text_input("Email",      value=eu.get("email",""))
                e_role   = ec1.selectbox("Role", avail_roles,
                    index=avail_roles.index(eu["role"]) if eu.get("role") in avail_roles else 0)
                e_unit   = ec2.selectbox("Unit", avail_units,
                    index=avail_units.index(eu["unit"]) if eu.get("unit") in avail_units else 0)
                e_sc     = ec1.text_input("Staff Code", value=str(eu.get("staff_code","")))
                e_active = ec2.checkbox("Active", value=eu.get("active", True))

                st.markdown("**Profile photo**")
                ph1, ph2 = st.columns([1,3])
                try:
                    from utils.core import photo_avatar_html, save_profile_photo, get_photo_b64
                    _sc_edit = eu.get("staff_code","") or sel_user
                    _av = photo_avatar_html(_sc_edit, eu.get("full_name",""), size=60)
                    ph1.markdown(_av, unsafe_allow_html=True)
                except: ph1.markdown("")

                new_photo = ph2.file_uploader(
                    "Upload photo (JPG/PNG, max 2MB)",
                    type=["jpg","jpeg","png","webp"],
                    key=f"photo_{sel_user}",
                    help="Square crop recommended · Will appear on scorecard and cascade pages")
                if new_photo:
                    _sc_save = eu.get("staff_code","") or sel_user
                    ext = new_photo.name.split(".")[-1].lower()
                    save_profile_photo(_sc_save, new_photo.read(), ext)
                    ph2.success("✅ Photo uploaded")

                st.markdown("**Permissions**")
                pc1,pc2,pc3 = st.columns(3)
                e_all    = pc1.checkbox("Can view all staff", value=eu.get("can_view_all",False))
                e_exec   = pc2.checkbox("Can manage Execute",  value=eu.get("can_execute",False))
                e_admin  = pc3.checkbox("Admin privileges",    value=eu.get("is_admin",False))

                if st.form_submit_button("Save changes", type="primary"):
                    um.users[sel_user].update({
                        "full_name":    e_fname,
                        "email":        e_email,
                        "role":         e_role,
                        "unit":         e_unit,
                        "staff_code":   e_sc,
                        "active":       e_active,
                        "can_view_all": e_all,
                        "can_execute":  e_exec,
                        "is_admin":     e_admin,
                    })
                    um.save()
                    audit_log("USER_EDITED", uname, sel_user)
                    st.success(f"User '{sel_user}' updated.")
                    st.rerun()
    else:
        # ── Staff code lookup (outside form so rerun works) ──────────
        df_for_lookup = st.session_state.get("df_processed", pd.DataFrame())
        registry_lu   = st.session_state.get("staff_registry", {})

        st.markdown(
            "<div style='padding:10px 14px;background:#F0FDF4;border:1px solid #BBF7D0;"
            "border-radius:8px;margin-bottom:12px;font-size:12px;color:#166534'>"
            "💡 <b>Quick create from BSC data</b> — enter the staff code below and the system "
            "will auto-fill name, role, unit and email from the uploaded file."
            "</div>", unsafe_allow_html=True)

        lookup_col, _ = st.columns([2,3])
        sc_input = lookup_col.text_input(
            "Staff Code (auto-fill)",
            key="new_user_sc_lookup",
            placeholder="e.g. 300130",
            help="Type the staff code and press Enter — details load automatically")

        # Resolve staff details from BSC data or staff registry
        auto_fname  = ""
        auto_role   = avail_roles[0] if avail_roles else ""
        auto_unit   = avail_units[0] if avail_units else ""
        auto_email  = ""
        sc_found    = False

        if sc_input.strip():
            sc_str = sc_input.strip()
            # Search df_processed first
            if not df_for_lookup.empty and "Staff Code" in df_for_lookup.columns:
                hits = df_for_lookup[df_for_lookup["Staff Code"].astype(str).str.strip()==sc_str]
                if not hits.empty:
                    row          = hits.iloc[0]
                    auto_fname   = str(row.get("Staff Name",""))
                    auto_role    = str(row.get("Role",""))
                    auto_unit    = str(row.get("Unit",""))
                    sc_found     = True
            # Fallback: search staff registry dict
            if not sc_found and isinstance(registry_lu, dict):
                for code, info in registry_lu.items():
                    if str(code).strip() == sc_str or str(info.get("Staff Code","")).strip()==sc_str:
                        auto_fname = info.get("Staff Name","")
                        auto_role  = info.get("Role","")
                        auto_unit  = info.get("Unit","")
                        auto_email = info.get("Email","")
                        sc_found   = True
                        break

            if sc_found:
                st.markdown(
                    f"<div style='padding:8px 12px;background:#EFF6FF;border:1px solid #BFDBFE;"
                    f"border-radius:8px;margin-bottom:8px;font-size:12px;color:#1E40AF'>"
                    f"✅ Found: <b>{auto_fname}</b> · {auto_role} · {auto_unit}"
                    f"</div>", unsafe_allow_html=True)
            elif sc_input.strip():
                st.markdown(
                    "<div style='padding:8px 12px;background:#FEF2F2;border:1px solid #FCA5A5;"
                    "border-radius:8px;margin-bottom:8px;font-size:12px;color:#991B1B'>"
                    "⚠️ Staff code not found in uploaded data. Fill details manually below."
                    "</div>", unsafe_allow_html=True)

        # Create user form — pre-populated from lookup
        # Staff code is the default username — use sc_input if found, else blank
        default_username = sc_input.strip() if sc_input.strip() else ""

        with st.form("create_user_form"):
            nc1, nc2 = st.columns(2)

            # Username defaults to staff code — admin can override
            n_user = nc1.text_input(
                "Username *",
                value=default_username,
                placeholder="Auto-filled from staff code",
                help="Defaults to the staff code. Staff use this to log in.")
            n_pass = nc2.text_input("Password *", type="password")
            n_fname = nc1.text_input("Full name *", value=auto_fname)
            n_email = nc2.text_input("Email", value=auto_email)

            if avail_roles:
                role_idx = avail_roles.index(auto_role) if auto_role in avail_roles else 0
                n_role = nc1.selectbox("Role", avail_roles, index=role_idx)
            else:
                n_role = nc1.text_input("Role", value=auto_role)

            if avail_units:
                unit_idx = avail_units.index(auto_unit) if auto_unit in avail_units else 0
                n_unit = nc2.selectbox("Unit", avail_units, index=unit_idx)
            else:
                n_unit = nc2.text_input("Unit", value=auto_unit)

            # Staff code shown and editable — already pre-filled from lookup
            n_sc = nc1.text_input(
                "Staff Code",
                value=sc_input,
                help="The numeric staff code from the BSC data. Also used as the default username.")

            if sc_found:
                nc2.markdown(
                    f"<div style='padding:8px 6px;background:#F0FDF4;"
                    f"border:1px solid #BBF7D0;border-radius:6px;"
                    f"font-size:11px;color:#166534;margin-top:24px'>"
                    f"✅ Username & staff code auto-filled from BSC data</div>",
                    unsafe_allow_html=True)

            st.markdown("**Permissions**")
            pc1,pc2,pc3,pc4 = st.columns(4)
            n_all   = pc1.checkbox("Can view all staff")
            n_exec  = pc2.checkbox("Can manage Execute")
            n_admin = pc3.checkbox("Admin privileges")
            n_val   = pc4.checkbox("Can validate")

            if st.form_submit_button("✅ Create user account", type="primary",
                                      use_container_width=True):
                # Use staff code as username if username field left as default
                final_user = n_user.strip() or n_sc.strip()
                final_sc   = n_sc.strip() or final_user
                if final_user and n_pass and n_fname:
                    if final_user in um.users:
                        st.error(f"Username '{final_user}' already exists.")
                    else:
                        um.add_user(final_user, n_pass, n_fname, n_email, n_role,
                                    n_unit, final_sc, n_all, n_exec, n_admin)
                        um.users[final_user]["can_validate"] = n_val
                        um.save()
                        audit_log("USER_CREATED", uname, f"{final_user}|sc:{final_sc}")
                        st.success(f"✅ User '{final_user}' created for {n_fname}. "
                                   f"Login: username `{final_user}`, staff code `{final_sc}`.")
                        st.rerun()
                else:
                    st.error("Username (or staff code), password and full name are required.")

    st.markdown("---")
    st.markdown("#### All users")
    if um.users:
        udf = pd.DataFrame([{
            "Username":   u,
            "Full name":  d.get("full_name",""),
            "Role":       d.get("role",""),
            "Unit":       d.get("unit",""),
            "Staff Code": d.get("staff_code",""),
            "Active":     "✅" if d.get("active",True) else "❌",
            "View all ⚠️": "⚠️ ALL" if d.get("can_view_all") else "Tree only",
            "Admin":      "✅" if d.get("is_admin") else "—",
            "Modules":    str(len(d.get("accessible_modules",[]))) + " set",
        } for u,d in um.users.items()])

        # Highlight rows where can_view_all is set on non-admin roles
        def _flag_view_all(v):
            return "background-color:#FEF2F2;color:#991B1B;font-weight:600" if "ALL" in str(v) else ""

        st.dataframe(
            udf.style.map(_flag_view_all, subset=["View all ⚠️"]),
            use_container_width=True, hide_index=True)

        # Warn if any non-admin/non-MD has can_view_all set
        _dangerous = [d.get("full_name",u) for u,d in um.users.items()
                      if d.get("can_view_all") and not d.get("is_admin")
                      and "managing" not in str(d.get("role","")).lower()
                      and "admin" not in str(d.get("role","")).lower()]
        if _dangerous:
            st.warning(
                f"⚠️ **Data privacy risk:** {', '.join(_dangerous)} "
                f"{'has' if len(_dangerous)==1 else 'have'} 'View all staff' enabled. "
                f"This grants access to all staff data. Remove unless MD/Admin.")

        # ── Bulk create users from Staff Register ────────────────
        with st.expander("📥 Bulk create users from Staff Register"):
            st.caption(
                "Create user accounts for all staff in the uploaded Staff Register. "
                "Username = staff_code.lower(), default password = first 6 chars of "
                "staff code + '!' (e.g. 300803! for staff 300803). "
                "Existing accounts are NOT overwritten.")

            _sr_df = st.session_state.get("staff_register_df")
            if _sr_df is None:
                st.warning("Upload the BSC Excel file in the sidebar first — "
                           "the Staff Register sheet will be detected automatically.")
            else:
                _sr_roles = _sr_df["Role"].unique().tolist() if "Role" in _sr_df.columns else []
                _bk_role_f = st.multiselect(
                    "Filter by role (leave blank = all roles)",
                    sorted(_sr_roles), key="bk_role_filter")

                _bk_df = _sr_df.copy()
                if _bk_role_f:
                    _bk_df = _bk_df[_bk_df["Role"].isin(_bk_role_f)]

                st.info(f"Will create accounts for {len(_bk_df)} staff members.")

                _preview_rows = []
                for _, _sr_row in _bk_df.head(5).iterrows():
                    _sc = str(_sr_row.get("Staff Code","")).strip()
                    _nm = str(_sr_row.get("Staff Name","")).strip()
                    _rl = str(_sr_row.get("Role","")).strip()
                    _un = str(_sc).lower()
                    _pw = _sc[:6] + "!" if len(_sc) >= 6 else _sc + "!"
                    _exists = "✅ exists" if _un in um.users else "🆕 new"
                    _preview_rows.append({"Username":_un,"Name":_nm,
                                          "Role":_rl,"Password":_pw,"Status":_exists})

                if _preview_rows:
                    st.markdown("**Preview (first 5):**")
                    st.dataframe(pd.DataFrame(_preview_rows),
                                 use_container_width=True, hide_index=True)

                if st.button("🚀 Create accounts", type="primary", key="bk_create"):
                    _created = 0; _skipped = 0
                    for _, _sr_row in _bk_df.iterrows():
                        _sc  = str(_sr_row.get("Staff Code","")).strip()
                        _nm  = str(_sr_row.get("Staff Name","")).strip()
                        _rl  = str(_sr_row.get("Role","")).strip()
                        _unt = str(_sr_row.get("Unit","")).strip()
                        _em  = str(_sr_row.get("Email","")).strip()
                        _rg  = str(_sr_row.get("Region","")).strip()
                        _un  = str(_sc).lower()
                        _pw  = _sc[:6] + "!" if len(_sc) >= 6 else _sc + "!"
                        if _un in um.users:
                            _skipped += 1
                            continue
                        um.users[_un] = {
                            "password":             um.hash_pw(_pw),
                            "full_name":            _nm,
                            "role":                 _rl,
                            "department":           _unt,
                            "unit":                 _unt,
                            "region":               _rg,
                            "email":                _em,
                            "staff_code":           _sc,
                            "can_view_all":         False,
                            "is_admin":             False,
                            "managed_roles":        [],
                            "managed_units":        [],
                            "managed_staff_codes":  [],
                            "active":               True,
                            "must_change_password": True,
                        }
                        _created += 1
                    um.save()
                    audit_log("BULK_USER_CREATE", uname, f"{_created} created, {_skipped} skipped")
                    st.success(f"✅ Created {_created} accounts, skipped {_skipped} existing. "
                               f"Default passwords: staff_code + '!'")
                    st.rerun()

        with st.expander("⚠️ Delete user — protected action"):
            # Admin cannot be deleted; any admin-level delete requires OTP verification
            _all_users = list(um.users.keys())
            _deletable = [u2 for u2 in _all_users if um.can_delete_user(u2)[0]]

            if not _deletable:
                st.info("🔒 No deletable accounts. The admin account is permanently protected.")
            else:
                del_sel = st.selectbox("User to delete", _deletable, key="del_sel")
                _can_del, _del_reason = um.can_delete_user(del_sel)

                if not _can_del:
                    st.error(f"🔒 {_del_reason}")
                else:
                    st.warning(
                        f"You are about to delete **{del_sel}** "
                        f"({um.users[del_sel].get('full_name','')}) permanently. "
                        f"This cannot be undone.")

                    # OTP flow — generate 6-digit code, show in UI
                    # (also send via email if SMTP configured)
                    _otp_key  = f"del_otp_{del_sel}"
                    _otp_sent = f"del_otp_sent_{del_sel}"

                    if st.button("🔑 Generate verification code", key="gen_otp_btn"):
                        import random
                        _otp = str(random.randint(100000, 999999))
                        st.session_state[_otp_key]  = _otp
                        st.session_state[_otp_sent] = True
                        # Try send by email
                        _admin_email = um.users.get(uname, {}).get("email", "")
                        _email_cfg   = load_email_config()
                        _sent_email  = False
                        if _admin_email and _email_cfg.get("smtp_host"):
                            try:
                                from email.mime.text import MIMEText
                                from email.mime.multipart import MIMEMultipart
                                import smtplib
                                _msg = MIMEMultipart("alternative")
                                _msg["Subject"] = "A2Z Blueprint — User deletion verification code"
                                _msg["From"]    = _email_cfg["sender_email"]
                                _msg["To"]      = _admin_email
                                _html = (
                                    f"<div style='font-family:Arial;max-width:400px;margin:auto'>"
                                    f"<div style='background:var(--brand-primary,#006B3F);padding:16px;border-radius:8px 8px 0 0'>"
                                    f"<div style='color:var(--color-background-primary);font-size:16px;font-weight:700'>A2Z Blueprint</div></div>"
                                    f"<div style='padding:20px;background:var(--color-background-primary);border:0.5px solid var(--color-border-tertiary)'>"
                                    f"<p style='color:var(--color-text-primary)'>Your verification code to delete user "
                                    f"<b>{del_sel}</b> is:</p>"
                                    f"<div style='font-size:36px;font-weight:700;letter-spacing:8px;"
                                    f"color:var(--brand-primary,#006B3F);text-align:center;padding:20px'>{_otp}</div>"
                                    f"<p style='color:var(--color-text-tertiary);font-size:12px'>This code expires when you close the page. "
                                    f"If you did not request this, contact your system administrator.</p>"
                                    f"</div></div>"
                                )
                                _msg.attach(MIMEText(_html, "html"))
                                with smtplib.SMTP(_email_cfg["smtp_host"],
                                                  int(_email_cfg.get("smtp_port",587))) as _s:
                                    _s.starttls()
                                    _s.login(_email_cfg["sender_email"], _email_cfg["sender_password"])
                                    _s.sendmail(_email_cfg["sender_email"], _admin_email, _msg.as_string())
                                _sent_email = True
                            except: pass

                        if _sent_email:
                            st.success(f"✅ Verification code sent to {_admin_email}")
                        else:
                            st.info(
                                "📋 Email not configured. Your verification code is: "
                                f"**{_otp}** — copy it and enter below.")

                    if st.session_state.get(_otp_sent):
                        _entered_otp = st.text_input(
                            "Enter 6-digit verification code",
                            max_chars=6, key="del_otp_input",
                            placeholder="______")
                        if st.button("🗑️ Confirm delete", type="secondary", key="del_confirm_btn"):
                            if _entered_otp == st.session_state.get(_otp_key, ""):
                                ok, msg = um.delete_user(del_sel, uname)
                                if ok:
                                    audit_log("USER_DELETED", uname, del_sel)
                                    st.session_state.pop(_otp_key, None)
                                    st.session_state.pop(_otp_sent, None)
                                    st.success(f"✅ {msg}")
                                    st.rerun()
                                else:
                                    st.error(f"❌ {msg}")
                            else:
                                st.error("❌ Incorrect verification code. Please try again.")

# ════════════════════════════════════════════════════════════════
# TAB 2 — PERMISSIONS
# ════════════════════════════════════════════════════════════════
with tabs[2]:
    st.subheader("Role-based access control")
    st.caption(
        "This matrix shows which modules each role can access by default. "
        "Use the override section below to grant or restrict access for individual users.")

    try:
        from utils.core import MODULE_ACCESS
        # Build matrix — rows=roles, cols=modules
        ALL_ROLES = [
            "Managing Director","Director Retail Banking","Director Commercial Banking",
            "Head Of Retail","Head Of Corporate","Head Of SME","Head Of Digital Innovation",
            "Head Of Strategy","Head Of Internal Audit","Head Of Marketing",
            "Chief Finance Officer","Chief Risk Officer","Chief Operations Officer",
            "Chief Compliance Officer","Chief Human Resources Officer","Chief Credit Officer",
            "Debt Recovery Unit Manager","Procurement Manager",
            "Regional Head","Branch Manager","Branch Operations Manager","Branch Credit Manager",
            "IT Manager","Operations Manager","HR Business Partner",
            "Relationship Manager Corporate","Relationship Manager SME",
            "Relationship Officer Personal Banking","Direct Sales Officer",
            "Teller","Customer Service Officer","Recovery Officer",
        ]
        MODULE_LABELS = {
            "perform":"Perform","people":"People","pipeline":"Pipeline",
            "execute":"Execute","products":"Products","integrate":"Integrate",
            "cascade":"Cascade","sla":"SLA","branch_log":"Branch Log",
            "optimize":"Optimize","commission":"Commission","campaigns":"Campaigns",
            "cims":"CIMS","sbu":"SBU","opex":"Opex","competitor":"Competitor",
            "export":"Export","admin":"Admin",
        }
        rows = []
        for role in ALL_ROLES:
            row = {"Role": role}
            for mod, lbl in MODULE_LABELS.items():
                cfg = MODULE_ACCESS.get(mod, {"min":"public","roles_all":[]})
                has = (role in cfg["roles_all"] or cfg["min"] == "self"
                       or cfg["min"] == "public")
                row[lbl] = "✅" if has else "—"
            rows.append(row)
        perm_df = pd.DataFrame(rows)
        # Style: green for ✅
        def _clr(v):
            return "background-color:#F0FDF4;color:#166534;font-weight:600" if v=="✅" else "color:#D1D5DB"
        st.dataframe(
            perm_df.style.map(_clr, subset=list(MODULE_LABELS.values())),
            use_container_width=True, hide_index=True,
            height=min(38*len(rows)+38, 520))
    except Exception as _e:
        st.error(f"Could not load module access: {_e}")

    st.markdown("---")
    st.markdown("#### Override permissions for a specific user")
    st.caption(
        "Changes here override the role defaults above for this individual user. "
        "Granting Admin gives full system access. "
        "Can view all gives access to all staff data across modules.")

    if um.users:
        pu_sel = st.selectbox("Select user to configure",
                               list(um.users.keys()), key="perm_sel")
        pu     = um.users.get(pu_sel, {})
        pu_role= pu.get("role","")

        st.markdown(
            f"<div style='padding:8px 12px;background:var(--color-background-secondary);border:0.5px solid var(--color-border-tertiary);"
            f"border-radius:8px;font-size:12px;margin-bottom:10px'>"
            f"<b>{pu.get('full_name',pu_sel)}</b> · Role: <b>{pu_role}</b> · "
            f"Unit: {pu.get('unit','—')} · Staff code: {pu.get('staff_code','—')}</div>",
            unsafe_allow_html=True)

        # When user selection changes, clear all cached widget states for permissions
        if st.session_state.get("_perm_user") != pu_sel:
            st.session_state["_perm_user"] = pu_sel
            for _kk in list(st.session_state.keys()):
                if _kk.startswith(("_p_", "_mod_")):
                    del st.session_state[_kk]

        _k = pu_sel

        # Use NO FORM — direct checkboxes with explicit session_state defaults
        # This is the only reliable way to show live values that can be toggled
        if f"_p_all_{_k}" not in st.session_state:
            st.session_state[f"_p_all_{_k}"]  = bool(pu.get("can_view_all", False))
            st.session_state[f"_p_adm_{_k}"]  = bool(pu.get("is_admin", False))
            st.session_state[f"_p_exec_{_k}"] = bool(pu.get("can_execute", False))
            st.session_state[f"_p_val_{_k}"]  = bool(pu.get("can_validate", False))
            st.session_state[f"_p_hr_{_k}"]   = bool(pu.get("can_hr", False))

        if True:  # scope block (replaces "with st.form")
            st.markdown("**System permissions**")
            fc1,fc2,fc3,fc4,fc5 = st.columns(5)
            p_all  = fc1.checkbox("View ALL staff (MD/Admin only)",
                                   key=f"_p_all_{_k}",
                                   help="⚠️ Grants visibility of ALL 380+ staff. "
                                        "Only tick for MD or System Admin. "
                                        "Directors see their reporting tree automatically — "
                                        "do NOT tick this for them.")
            p_adm  = fc2.checkbox("Admin privileges",
                                   key=f"_p_adm_{_k}",
                                   help="Full system access including Admin module.")
            p_exec = fc3.checkbox("Can manage Execute",
                                   key=f"_p_exec_{_k}")
            p_val  = fc4.checkbox("Can validate staff",
                                   key=f"_p_val_{_k}")
            p_hr   = fc5.checkbox("HR module access",
                                   key=f"_p_hr_{_k}",
                                   help="Access to People / HR module.")

            st.markdown("---")
            st.markdown(
                "**Module access** — tick modules this user can access. "
                "Expand each module to restrict to specific pages within it.")

            try:
                from utils.core import MODULE_ACCESS, MODULE_PAGES
                cur_accessible = set(pu.get("accessible_modules", []))
                cur_pages      = dict(pu.get("accessible_pages", {}))
                if not cur_accessible:
                    cur_accessible = {m for m, cfg in MODULE_ACCESS.items()
                                      if pu_role in cfg.get("roles_all",[])
                                      or cfg["min"] == "self"}
                mods_list = list(MODULE_ACCESS.keys())
                for _m in mods_list:
                    if f"_mod_{_k}_{_m}" not in st.session_state:
                        st.session_state[f"_mod_{_k}_{_m}"] = _m in cur_accessible

                new_accessible = set()
                new_pages      = {}

                # Render each module as a row with optional page sub-controls
                for m in mods_list:
                    mod_lbl  = m.replace("_"," ").title()
                    mod_pages = MODULE_PAGES.get(m, [])
                    has_pages = len(mod_pages) > 0

                    m_col1, m_col2 = st.columns([2, 3])
                    mod_on = m_col1.checkbox(
                        f"**{mod_lbl}**",
                        key=f"_mod_{_k}_{m}",
                        help=f"{len(mod_pages)} pages within this module" if has_pages else "")
                    if mod_on:
                        new_accessible.add(m)
                        if has_pages and mod_pages:
                            cur_mod_pages = set(cur_pages.get(m, mod_pages))
                            sel_pages = m_col2.multiselect(
                                f"Pages in {mod_lbl}",
                                options=mod_pages,
                                default=[p for p in cur_mod_pages if p in mod_pages],
                                key=f"_pages_{_k}_{m}",
                                label_visibility="collapsed",
                                placeholder=f"All {len(mod_pages)} pages (no restriction)")
                            if sel_pages and len(sel_pages) < len(mod_pages):
                                new_pages[m] = sel_pages
                    else:
                        m_col2.markdown(
                            "<span style='color:var(--color-text-tertiary);font-size:11px'>no access</span>",
                            unsafe_allow_html=True)
            except Exception as _me:
                st.error(f"Module access error: {_me}")
                new_accessible = set()
                new_pages      = {}

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("💾 Save permissions", type="primary",
                          use_container_width=True,
                          key=f"save_perm_{_k}"):
                um.users[pu_sel].update({
                    "can_view_all":       p_all,
                    "can_execute":        p_exec,
                    "is_admin":           p_adm,
                    "can_validate":       p_val,
                    "can_hr":             p_hr,
                    "accessible_modules": sorted(new_accessible),
                    "accessible_pages":   new_pages,
                })
                um.save()
                # Clear cached checkbox states so next render reads fresh from DB
                for _kk in list(st.session_state.keys()):
                    if _kk.startswith(("_p_", "_mod_")):
                        del st.session_state[_kk]
                st.session_state.pop("_perm_user", None)
                audit_log("PERM_CHANGED", uname,
                          f"{pu_sel}:admin={p_adm},view_all={p_all},"
                          f"modules={len(new_accessible)}")
                st.success(f"✅ Permissions saved for {pu.get('full_name',pu_sel)}")
                st.rerun()

# TAB 4 — STAFF MOVEMENTS
# ════════════════════════════════════════════════════════════════
with tabs[3]:
    st.subheader("🗂️ Staff movements")
    st.caption(
        "Remap reporting lines, transfer staff between branches, change roles, "
        "and manage acting appointments. All changes are audited and take effect immediately.")

    if rlm is None:
        st.error("ReportingLineManager not available. Restart the app.")
    else:
        _mv_view = st.radio("",
            ["🔗 Remap reporting line","🚀 Branch transfer","🎭 Role change","📋 Active overrides"],
            horizontal=True, key="mv_view")
        st.markdown("---")

        def _build_staff_opts(df):
            opts = {}
            if len(df) == 0: return opts
            for role in sorted(df["Role"].dropna().unique()):
                for _, row in df[df["Role"]==role].sort_values("Staff Name").iterrows():
                    label = f"[{role}]  {row['Staff Name']}  ·  {row.get('Unit','')}"
                    opts[label] = str(row["Staff Code"])
            return opts

        # ── REMAP REPORTING LINE ─────────────────────────────────────
        if "Remap reporting" in _mv_view:
            st.markdown("**Change who a staff member reports to** — acting appointment, individual exception.")
            if len(staff_scores) == 0:
                st.info("Load CBS data or upload BSC file to see staff.")
            else:
                rf1,rf2,rf3 = st.columns(3)
                _all_roles_mv = ["All roles"] + sorted(staff_scores["Role"].dropna().unique().tolist())
                _all_units_mv = ["All units"] + sorted(staff_scores["Unit"].dropna().unique().tolist())
                _filt_role = rf1.selectbox("Filter role", _all_roles_mv, key="mv_filt_role")
                _filt_unit = rf2.selectbox("Filter unit", _all_units_mv, key="mv_filt_unit")
                _filt_df   = staff_scores.copy()
                if _filt_role != "All roles": _filt_df = _filt_df[_filt_df["Role"]==_filt_role]
                if _filt_unit != "All units": _filt_df = _filt_df[_filt_df["Unit"]==_filt_unit]
                rf3.markdown(f"<div style='padding:10px 6px;font-size:12px'><b>{len(_filt_df)}</b> staff match</div>",unsafe_allow_html=True)

                # Manager pool from org_config hierarchy
                try:
                    from utils.core import get_org_config as _goc3
                    _parent_roles = set()
                    for _r3,_p3 in _goc3().get("hierarchy",{}).items():
                        _parent_roles.update(_p3)
                    _mgr_df = staff_scores[staff_scores["Role"].isin(_parent_roles)] if _parent_roles else staff_scores
                except:
                    _mgr_df = staff_scores

                _staff_opts = _build_staff_opts(_filt_df) or _build_staff_opts(staff_scores)
                _mgr_opts   = _build_staff_opts(_mgr_df)

                with st.form("remap_form"):
                    rl1,rl2 = st.columns(2)
                    _staff_lbl = rl1.selectbox(f"Staff member ({len(_staff_opts)} shown)",
                        list(_staff_opts.keys()), key="remap_staff")
                    _mgr_lbl   = rl2.selectbox(f"New line manager ({len(_mgr_opts)})",
                        list(_mgr_opts.keys()), key="remap_mgr")
                    _remap_reason = st.text_input("Reason *",
                        placeholder="e.g. Promotion, acting appointment, restructure")
                    if st.form_submit_button("✅ Apply remap", type="primary", use_container_width=True):
                        _sc_r  = _staff_opts.get(_staff_lbl,"")
                        _mc_r  = _mgr_opts.get(_mgr_lbl,"")
                        _sn    = _staff_lbl.split("]")[-1].strip().split("·")[0].strip()
                        _mn    = _mgr_lbl.split("]")[-1].strip().split("·")[0].strip()
                        if _sc_r == _mc_r:
                            st.error("Staff and manager cannot be the same person.")
                        elif not _sc_r or not _mc_r:
                            st.error("Could not resolve codes.")
                        else:
                            rlm.remap(_sc_r, _mc_r, uname, _remap_reason)
                            audit_log("REPORTING_LINE_REMAP", uname, f"{_sn} → {_mn}")
                            st.success(f"✅ {_sn} now reports to {_mn}")
                            st.rerun()

        # ── BRANCH TRANSFER ──────────────────────────────────────────
        elif "Branch transfer" in _mv_view:
            st.markdown("**Move a staff member to a different branch** — updates unit, region, and manager.")
            if len(staff_scores) == 0:
                st.info("Load CBS data or upload BSC file to see staff.")
            else:
                _all_staff_opts = _build_staff_opts(staff_scores)
                try:
                    from utils.core import get_all_branches as _gab4, get_org_config as _goc4
                    _branch_list = [b["name"] for b in _gab4() if b.get("type","")!="HO"]
                    _mgr_opts2   = _build_staff_opts(staff_scores)  # any staff can be manager
                except:
                    _branch_list = avail_units

                with st.form("transfer_form"):
                    tr1,tr2 = st.columns(2)
                    _tr_staff = tr1.selectbox("Staff member to transfer",
                        list(_all_staff_opts.keys()), key="tr_staff")
                    _tr_branch= tr2.selectbox("New branch", _branch_list, key="tr_branch")
                    _tr_mgr   = tr1.selectbox("New line manager (optional)",
                        ["— Keep current —"] + list(_build_staff_opts(staff_scores).keys()),
                        key="tr_mgr")
                    _tr_role  = tr2.selectbox("New role (if changing)",
                        ["— Keep current —"] + avail_roles, key="tr_role")
                    _tr_reason= st.text_input("Reason *",
                        placeholder="e.g. Branch-to-branch transfer, secondment")
                    _tr_eff   = st.date_input("Effective date", value=date.today())

                    if st.form_submit_button("🚀 Execute transfer", type="primary", use_container_width=True):
                        if _tr_reason.strip():
                            _tr_sc = _all_staff_opts.get(_tr_staff,"")
                            _tr_mn = ""
                            if _tr_mgr != "— Keep current —":
                                _tr_mc = _build_staff_opts(staff_scores).get(_tr_mgr,"")
                                rlm.remap(_tr_sc, _tr_mc, uname, f"Transfer: {_tr_reason}")
                                _tr_mn = _tr_mgr.split("]")[-1].strip().split("·")[0].strip()
                            # Update user account unit
                            for _un2, _ud3 in um.users.items():
                                if str(_ud3.get("staff_code","")) == _tr_sc:
                                    _ud3["unit"] = _tr_branch
                                    if _tr_role != "— Keep current —": _ud3["role"] = _tr_role
                                    um.save(); break
                            audit_log("BRANCH_TRANSFER", uname,
                                f"{_tr_staff.split(']')[-1].strip().split('·')[0].strip()} → {_tr_branch}")
                            st.success(f"✅ Transfer complete → {_tr_branch}"
                                       + (f" | Manager: {_tr_mn}" if _tr_mn else ""))
                            st.rerun()
                        else:
                            st.error("Reason is required.")

        # ── ROLE CHANGE ──────────────────────────────────────────────
        elif "Role change" in _mv_view:
            st.markdown("**Promote or reassign a staff member's role.** Updates user account and cascade position.")
            if len(staff_scores) == 0:
                st.info("Load CBS data or upload BSC file to see staff.")
            else:
                _rc_opts = {f"{r['Staff Name']} [{r['Role']}] · {r.get('Unit','')}": str(r['Staff Code'])
                            for _, r in staff_scores.iterrows()}
                with st.form("role_change_form"):
                    rc1,rc2 = st.columns(2)
                    _rc_staff   = rc1.selectbox("Staff member", list(_rc_opts.keys()), key="rc_staff_sel")
                    _rc_newrole = rc2.selectbox("New role", avail_roles, key="rc_new_role")
                    _rc_newunit = rc1.selectbox("New branch/unit", avail_units, key="rc_new_unit")
                    _rc_reason  = rc2.text_input("Reason *",
                        placeholder="e.g. Promotion, acting, secondment")
                    st.date_input("Effective date", value=date.today(), key="rc_eff_date")
                    if st.form_submit_button("✅ Apply role change", type="primary", use_container_width=True):
                        if _rc_reason.strip():
                            _rc_sc = _rc_opts.get(_rc_staff,"")
                            _changed = False
                            for _un3,_ud4 in um.users.items():
                                if str(_ud4.get("staff_code","")) == _rc_sc:
                                    _ud4["role"] = _rc_newrole
                                    _ud4["unit"] = _rc_newunit
                                    um.save(); _changed = True; break
                            if _changed:
                                audit_log("ROLE_CHANGE", uname,
                                    f"{_rc_staff.split('[')[0].strip()} → {_rc_newrole} @ {_rc_newunit}")
                                st.success(f"✅ Role changed to {_rc_newrole} at {_rc_newunit}.")
                                st.rerun()
                            else:
                                st.warning("Could not find matching user account. Check staff code.")
                        else:
                            st.error("Reason is required.")

        # ── ACTIVE OVERRIDES ─────────────────────────────────────────
        else:
            ov_summary = rlm.summary()
            _oa,_ob,_oc_ = st.columns(3)
            _oa.metric("Active overrides",   ov_summary.get("total_overrides",0))
            _ob.metric("Transfer overrides", ov_summary.get("total_transfers",0))
            _oc_.metric("Last change", str(ov_summary.get("last_updated","Never"))[:10])
            overrides = rlm.get_all_overrides()
            if not overrides:
                st.info("No overrides active.")
            else:
                _cn = {}
                if len(staff_scores) and "Staff Code" in staff_scores.columns:
                    for _,_r4 in staff_scores.iterrows():
                        _cn[str(_r4["Staff Code"])] = _r4["Staff Name"]
                _ov_rows = [{"Staff":_cn.get(o.get("staff_code",""),o.get("staff_code","")),
                             "New Manager":_cn.get(o.get("manager_code",""),o.get("manager_code","")),
                             "New Unit":o.get("unit","—"),"Type":o.get("type","remap"),
                             "Reason":o.get("reason",""),"By":o.get("updated_by",""),
                             "When":str(o.get("updated_at",""))[:16]} for o in overrides]
                st.dataframe(pd.DataFrame(_ov_rows), use_container_width=True, hide_index=True)
                _cc1,_cc2 = st.columns([3,1])
                _clr_opts = {f"{r['Staff']} ({r['Type']})": overrides[i].get("staff_code","")
                             for i,r in enumerate(_ov_rows)}
                _clr_sel = _cc1.selectbox("Remove override", list(_clr_opts.keys()), key="clr_ov3")
                if _cc2.button("Remove", type="secondary", key="clr_btn3"):
                    rlm.clear_override(_clr_opts[_clr_sel], uname)
                    audit_log("OVERRIDE_CLEARED", uname, _clr_sel)
                    st.success("Override removed."); st.rerun()


# TAB 6 — ORG TREE
# ════════════════════════════════════════════════════════════════
with tabs[4]:
    st.subheader("Organisation chart")
    st.caption("Live org tree combining uploaded data and all active overrides.")

    reg = st.session_state.get("staff_registry", pd.DataFrame())
    if rlm and len(reg):
        applied = rlm.apply_to_registry(reg)
        tree    = rlm.get_org_tree(applied)

        # Build code → name map
        code_name = {}
        code_role = {}
        code_unit = {}
        if "Staff Code" in applied.columns:
            for _, r in applied.iterrows():
                sc = str(r["Staff Code"])
                code_name[sc] = r.get("Staff Name", sc)
                code_role[sc] = r.get("Role","")
                code_unit[sc] = r.get("Unit","")

        # Build Plotly org chart
        def flatten_tree(node_code, depth=0, parent_idx=None, rows=None, edges=None):
            if rows is None: rows=[]; edges=[]
            idx = len(rows)
            name = code_name.get(node_code, node_code)
            role = code_role.get(node_code,"")
            unit = code_unit.get(node_code,"")
            rows.append({'id':node_code,'name':name,'role':role,'unit':unit,'depth':depth,'idx':idx})
            if parent_idx is not None:
                edges.append((parent_idx, idx))
            for child in tree.get(node_code,[]):
                flatten_tree(child, depth+1, idx, rows, edges)
            return rows, edges

        # Find roots (no parent in tree values)
        all_children = set(c for children in tree.values() for c in children)
        roots = [sc for sc in tree.keys() if sc not in all_children]
        if not roots and tree:
            roots = [list(tree.keys())[0]]

        # Org view filters
        ot1, ot2 = st.columns(2)
        view_unit = ot1.selectbox("Filter by unit", ["All"] + sorted(set(code_unit.values())), key="org_unit")
        max_depth = ot2.slider("Max depth shown", 1, 6, 3, key="org_depth")

        all_rows, all_edges = [], []
        for root in roots:
            flatten_tree(root, 0, None, all_rows, all_edges)

        if all_rows:
            # Filter by unit
            if view_unit != "All":
                keep_ids = {r['id'] for r in all_rows if r['unit']==view_unit}
                all_rows = [r for r in all_rows if r['id'] in keep_ids]

            # Filter by depth
            all_rows = [r for r in all_rows if r['depth'] <= max_depth]

            # Summary stats
            total_mapped = len(all_rows)
            n_overrides  = len(rlm.get_all_overrides())

            os1, os2, os3 = st.columns(3)
            os1.metric("Staff in tree", total_mapped)
            os2.metric("Active overrides", n_overrides)
            os3.metric("Depth levels", max_depth)

            # Render as collapsible tree rows
            DEPTH_COLORS = ['var(--brand-primary,#006B3F)','#F5A623','#185FA5','#7F8C8D','#9B59B6','#E24B4A']
            for row in all_rows[:100]:  # cap at 100 for performance
                indent  = '&nbsp;' * (row['depth'] * 5)
                clr     = DEPTH_COLORS[min(row['depth'], 5)]
                border  = 2 + row['depth']
                is_mgr  = row['id'] in tree
                icon    = '👔' if row['depth']==0 else ('📋' if is_mgr else '👤')
                has_ov  = row['id'] in rlm.overrides
                ov_badge= "<span style='background:#F5A623;color:var(--color-background-primary);padding:1px 5px;border-radius:8px;font-size:9px;margin-left:4px'>override</span>" if has_ov else ''

                st.markdown(
                    f"<div style='padding:5px 10px;background:var(--color-background-secondary);"
                    f"border-left:{border}px solid {clr};"
                    f"margin:1px 0;border-radius:0 3px 3px 0;font-size:11px'>"
                    f"{indent}{icon} <b>{row['name']}</b> "
                    f"<span style='color:#888;font-size:10px'>{row['role']}</span> "
                    f"<span style='color:#aaa;font-size:10px'>· {row['unit']}</span>"
                    f"{ov_badge}</div>",
                    unsafe_allow_html=True)

            if len(all_rows) > 100:
                st.caption(f"Showing 100 of {len(all_rows)} staff. Use the unit filter to narrow down.")
        else:
            st.info("No org tree data available. Upload BSC data with Staff Register sheet.")
    elif len(staff_scores):
        # Build from staff_scores directly
        st.info("Load the Staff Register (upload BSC Excel with Staff Register sheet) for the full org tree.")
        # Show a simple role hierarchy
        role_counts = staff_scores.groupby(['Role','Unit'])['Staff Name'].count().reset_index()
        role_counts.columns = ['Role','Unit','Count']
        fig = px.treemap(role_counts, path=['Unit','Role'], values='Count',
                          title='Staff distribution by unit and role')
        fig.update_layout(height=420)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Upload BSC data to view the organisation chart.")

# ════════════════════════════════════════════════════════════════
# TAB 7 — AUDIT LOG
# ════════════════════════════════════════════════════════════════
with tabs[5]:
    st.subheader("Audit trail")
    log_path = DATA_DIR / "audit.log"
    if log_path.exists():
        lines = log_path.read_text().strip().split("\n")[-200:]  # last 200 lines
        lines.reverse()  # newest first
        audit_rows = []
        for line in lines:
            parts = line.split("|")
            if len(parts) >= 3:
                audit_rows.append({
                    "Timestamp": parts[0].strip(),
                    "User":      parts[1].strip(),
                    "Action":    parts[2].strip(),
                    "Detail":    parts[3].strip() if len(parts)>3 else "",
                })
        if audit_rows:
            af1, af2 = st.columns(2)
            filter_user   = af1.selectbox("Filter by user", ["All"] +
                list(dict.fromkeys(r["User"] for r in audit_rows)), key="aud_u")
            filter_action = af2.selectbox("Filter by action", ["All"] +
                list(dict.fromkeys(r["Action"] for r in audit_rows)), key="aud_a")

            disp = audit_rows
            if filter_user   != "All": disp = [r for r in disp if r["User"]==filter_user]
            if filter_action != "All": disp = [r for r in disp if r["Action"]==filter_action]

            st.dataframe(pd.DataFrame(disp), use_container_width=True, hide_index=True)
        else:
            st.info("No audit entries yet.")
    else:
        st.info("Audit log will appear here once actions are performed.")

# ════════════════════════════════════════════════════════════════
# TAB 8 — UPLOAD FORMAT
# ════════════════════════════════════════════════════════════════
with tabs[6]:
    st.subheader("Upload format guide")
    st.caption("The system accepts two upload modes: Full BSC (targets + actuals) and Actuals-only (monthly update).")

    mode_tabs = st.tabs(["📊 Full BSC upload","📥 Actuals-only upload","📋 Column reference","✅ Validation rules"])

    with mode_tabs[0]:
        st.markdown("#### Full BSC upload — KPI Data sheet")
        st.markdown(
            "Use this when setting up the system or updating targets. "
            "The file must have two sheets: **KPI Data** and **Staff Register**.")

        st.markdown(
            "<div style='padding:12px 16px;background:#EFF6FF;border:1px solid #BFDBFE;"
            "border-radius:8px;margin-bottom:12px;font-size:12px;color:#1E40AF'>"
            "<b>📌 Targets come from the Cascade module — not from this file.</b><br>"
            "The upload file no longer includes an <code>Annual Target</code> column. "
            "Once the MD sets bank targets and managers cascade to every staff member, "
            "the system auto-populates each person's target directly into their BSC scorecard. "
            "The file now only contains KPI structure, weights, and monthly actuals."
            "</div>", unsafe_allow_html=True)

        full_cols = pd.DataFrame([
            {"Column":"Unit",          "Required":"✅","Type":"Text",   "Example":"Nairobi CBD Branch","Notes":"Branch name or HO department"},
            {"Column":"Category",      "Required":"✅","Type":"Text",   "Example":"Branch",            "Notes":"Branch or Head Office"},
            {"Column":"Staff Code",    "Required":"✅","Type":"Number", "Example":"300130",            "Notes":"Unique employee ID — numeric, no apostrophe"},
            {"Column":"Staff Name",    "Required":"✅","Type":"Text",   "Example":"Grace K. Kamau",    "Notes":"Full name as registered in HR system"},
            {"Column":"Role",          "Required":"✅","Type":"Text",   "Example":"Teller",            "Notes":"Must match role hierarchy exactly"},
            {"Column":"Role Function", "Required":"✅","Type":"Text",   "Example":"Support",           "Notes":"Business | Support | Executive"},
            {"Column":"Pillar",        "Required":"✅","Type":"Text",   "Example":"Financial",         "Notes":"Financial | Customer Focus | Operational Excellence"},
            {"Column":"KPI",           "Required":"✅","Type":"Text",   "Example":"Deposit Growth",    "Notes":"KPI name — must be consistent per role"},
            {"Column":"Annual Actual", "Required":"✅","Type":"Number", "Example":"45000000",          "Notes":"Cumulative YTD actual — updated by monthly upload"},
            {"Column":"Weight",        "Required":"✅","Type":"Decimal","Example":"0.08",              "Notes":"All weights per staff must sum to exactly 1.00"},
            {"Column":"Jan-26 Actual", "Required":"✅","Type":"Number", "Example":"15000000",          "Notes":"Format: Mon-YY Actual (e.g. Apr-26 Actual)"},
            {"Column":"Feb-26 Actual", "Required":"⬜","Type":"Number", "Example":"14500000",          "Notes":"Add new month columns each month — auto-detected"},
            {"Column":"Mar-26 Actual", "Required":"⬜","Type":"Number", "Example":"15500000",          "Notes":"Prior months can be left blank if not yet uploaded"},
            {"Column":"FY-25 Actual",  "Required":"⬜","Type":"Number", "Example":"380000000",         "Notes":"Full prior year actual — used for AI target suggestions"},
            {"Column":"Dec-25 Actual", "Required":"⬜","Type":"Number", "Example":"32000000",          "Notes":"December prior year monthly — used for trend analysis"},
        ])
        st.dataframe(full_cols, use_container_width=True, hide_index=True)

        st.markdown(
            "<div style='padding:12px 16px;background:#F0FDF4;border:1px solid #BBF7D0;"
            "border-radius:8px;margin-top:8px;font-size:12px;color:#166534'>"
            "<b>How targets flow into the BSC:</b><br>"
            "1. MD sets bank-level targets in <b>Target Cascade → Bank targets</b><br>"
            "2. MD allocates to directors, directors to managers, managers to staff<br>"
            "3. Staff accept and lock their targets — tracking begins immediately<br>"
            "4. The BSC scorecard reads the <b>cascaded target</b> as the Annual Target for scoring<br>"
            "5. Staff cannot see their target until their manager has cascaded and they have accepted"
            "</div>", unsafe_allow_html=True)

        st.markdown("#### Staff Register sheet")
        reg_cols = pd.DataFrame([
            {"Column":"Staff Code","Required":"✅","Example":"300130"},
            {"Column":"Staff Name","Required":"✅","Example":"Grace K. Kamau"},
            {"Column":"Email","Required":"⬜","Example":"g.kamau@ecobank.co.ke"},
            {"Column":"Phone","Required":"⬜","Example":"0712345678"},
            {"Column":"Role","Required":"✅","Example":"Teller"},
            {"Column":"Role Function","Required":"✅","Example":"Support"},
            {"Column":"Unit","Required":"✅","Example":"Nairobi CBD Branch"},
            {"Column":"Category","Required":"✅","Example":"Branch"},
            {"Column":"Staff Status","Required":"✅","Example":"Existing | New | Probation"},
            {"Column":"Hire Date","Required":"✅","Example":"01/01/2022"},
            {"Column":"Reports To Code","Required":"✅","Example":"300105","Notes":"Staff Code of direct line manager"},
            {"Column":"Region","Required":"⬜","Example":"Central","Notes":"Auto-derived from unit if blank"},
        ])
        st.dataframe(reg_cols, use_container_width=True, hide_index=True)

    with mode_tabs[1]:
        st.markdown("#### Actuals-only upload — monthly update format")
        st.markdown(
            "Once targets are set (via full upload or cascade), use this lighter format "
            "to upload monthly actuals only. The system merges actuals into the existing targets.")

        st.markdown(
            "<div style='padding:12px 16px;background:var(--brand-light,#E8F5EE);"
            "border-left:4px solid var(--brand-primary,#006B3F);border-radius:0 6px 6px 0;margin:8px 0'>"
            "<b>How it works:</b> Targets come from the cascade allocation or the last full upload. "
            "Each month you upload a file with only Staff Code, KPI, and the month's actual. "
            "The system matches on Staff Code + KPI and updates that month's column."
            "</div>", unsafe_allow_html=True)

        actuals_cols = pd.DataFrame([
            {"Column":"Staff Code","Required":"✅","Example":"300130","Notes":"Must match exactly"},
            {"Column":"KPI","Required":"✅","Example":"Deposit Growth","Notes":"Must match exactly"},
            {"Column":"Apr-26 Actual","Required":"✅","Example":"18500000","Notes":"The column header drives which month is updated"},
            {"Column":"Staff Name","Required":"⬜","Example":"Grace K. Kamau","Notes":"Optional — used for validation only"},
            {"Column":"Unit","Required":"⬜","Example":"Nairobi CBD Branch","Notes":"Optional — used for validation only"},
        ])
        st.dataframe(actuals_cols, use_container_width=True, hide_index=True)

        st.markdown("#### Sample actuals file (copy this format)")
        sample_actuals = pd.DataFrame({
            "Staff Code": [300130,300131,300132],
            "Staff Name":  ["Grace K. Kamau","John M. Otieno","Mary A. Wanjiku"],
            "KPI":         ["Deposit Growth","Deposit Growth","New Customer Acquisition"],
            "Apr-26 Actual":[16500000, 19200000, 38],
            "Unit":        ["Nairobi CBD Branch","Westlands Branch","Kisumu Branch"],
        })
        st.dataframe(sample_actuals, use_container_width=True, hide_index=True)

        st.markdown(
            "<div style='padding:10px 14px;background:#FFFBF0;"
            "border-left:3px solid #F5A623;font-size:12px;margin-top:8px'>"
            "⚠️ <b>For KPIs auto-calculated by the system</b> (Diligence Score, Initiative Score, "
            "SLA Adherence Score, Branch Optimization Score, CX Score, Diligence Score) — "
            "do NOT include these in the actuals upload. The system computes them automatically "
            "from Execute, Daily Log, and SLA modules. Including them will overwrite the system calculation."
            "</div>", unsafe_allow_html=True)

        st.markdown("#### Auto-calculated KPIs — source modules")
        auto_kpis = pd.DataFrame([
            {"KPI":"Diligence Score","Source module":"Execute → milestones + People → discipline/PIP","Frequency":"Real-time"},
            {"KPI":"Initiative Score","Source module":"Execute → gate progression","Frequency":"Real-time"},
            {"KPI":"SLA Adherence Score","Source module":"SLA Tracker (coming)","Frequency":"Daily"},
            {"KPI":"CX Score","Source module":"Daily Branch Log validation + SLA","Frequency":"Daily"},
            {"KPI":"Branch Optimization Score","Source module":"Branch Optimization Engine (coming)","Frequency":"Daily"},
            {"KPI":"Campaign Conversion Rate","Source module":"Campaigns module (coming)","Frequency":"Per campaign"},
            {"KPI":"Digital Acquiring","Source module":"Daily Branch Log","Frequency":"Daily"},
            {"KPI":"Digital Transaction Migration","Source module":"Daily Branch Log","Frequency":"Daily"},
            {"KPI":"Transactions","Source module":"Daily Branch Log","Frequency":"Daily"},
        ])
        st.dataframe(auto_kpis, use_container_width=True, hide_index=True)

    with mode_tabs[2]:
        st.markdown("#### Full column reference — all supported KPIs")
        st.caption("These are all KPIs currently mapped in the system with their BSC pillar and role mapping.")

        kpi_ref = pd.DataFrame([
            {"KPI":"Deposit Growth","Pillar":"Financial","Roles":"All business roles","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Loan Book Growth","Pillar":"Financial","Roles":"Business roles","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Loans Disbursement","Pillar":"Financial","Roles":"RM, DSO, BCM","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Fees and Commission","Pillar":"Financial","Roles":"All business","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"DFS Revenue","Pillar":"Financial","Roles":"Business + Digital","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Digital Acquiring","Pillar":"Financial/Customer","Roles":"Branch staff","Type":"Count","Unit":"Accounts"},
            {"KPI":"Transactions","Pillar":"Financial","Roles":"Tellers, CSO","Type":"Count","Unit":"Count"},
            {"KPI":"Trade Finance","Pillar":"Financial","Roles":"Corporate, SME","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Treasury","Pillar":"Financial","Roles":"Corporate, Treasury","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"PBT","Pillar":"Financial","Roles":"Managers/Directors","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"Bancassurance","Pillar":"Financial","Roles":"Branch, DSO","Type":"Amount (KES)","Unit":"KES 000"},
            {"KPI":"NPL Ratio","Pillar":"Financial","Roles":"Credit, Business","Type":"Ratio","Unit":"Decimal (0.05 = 5%)"},
            {"KPI":"PAR","Pillar":"Financial","Roles":"Credit, BCM","Type":"Ratio","Unit":"Decimal"},
            {"KPI":"CIR","Pillar":"Financial","Roles":"Directors, CFO","Type":"Ratio","Unit":"Decimal (0.65 = 65%)"},
            {"KPI":"ROE","Pillar":"Financial","Roles":"MD, Directors","Type":"Ratio","Unit":"Decimal"},
            {"KPI":"New Customer Acquisition","Pillar":"Customer Focus","Roles":"All business","Type":"Count","Unit":"Customers"},
            {"KPI":"Dormancy Reactivation","Pillar":"Customer Focus","Roles":"Branch, DSO","Type":"Count","Unit":"Accounts"},
            {"KPI":"CX Score","Pillar":"Customer Focus","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"SLA Adherence Score","Pillar":"Customer Focus","Roles":"All staff","Type":"Score","Unit":"0.00–1.00 (auto)"},
            {"KPI":"Credit TAT Score","Pillar":"Customer Focus","Roles":"Credit staff","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"Diligence Score","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0–100 (auto)"},
            {"KPI":"Initiative Score","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00 (auto)"},
            {"KPI":"Compliance","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"Audit Closure","Pillar":"Operational Excellence","Roles":"All staff","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"Timely Reconciliations","Pillar":"Operational Excellence","Roles":"Ops, Tellers","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"Branch Optimization Score","Pillar":"Operational Excellence","Roles":"BOM, BM","Type":"Score","Unit":"0.00–1.00 (auto)"},
            {"KPI":"Campaign Conversion Rate","Pillar":"Operational Excellence","Roles":"Marketing, DSO","Type":"Score","Unit":"0.00–1.00 (auto)"},
            {"KPI":"Recovery Rate","Pillar":"Financial","Roles":"DRU","Type":"Score","Unit":"0.00–1.00"},
            {"KPI":"Loan Recovery Amount","Pillar":"Financial","Roles":"DRU, Recovery Officer","Type":"Amount (KES)","Unit":"KES 000"},
        ])
        st.dataframe(kpi_ref, use_container_width=True, hide_index=True)

    with mode_tabs[3]:
        st.markdown("#### Validation rules — what the system checks on upload")
        rules = [
            ("✅","Weights sum to 1.00","Per staff member, all KPI weights must sum to exactly 1.00 (±0.005 tolerance)"),
            ("✅","No duplicate KPI rows","Each staff + KPI combination must appear only once"),
            ("✅","Valid pillar names","Pillar must be exactly: Financial, Customer Focus, Operational Excellence"),
            ("✅","Staff Code format","Staff codes must be numeric, no leading apostrophe, no spaces"),
            ("✅","Month column format","Month columns must be: Mon-YY Actual (e.g. Apr-26 Actual)"),
            ("ℹ️","Annual Target optional","Targets come from the cascade module — Annual Target column is not required"),
            ("✅","Weight > 0","Each KPI weight must be > 0"),
            ("✅","Role Function","Must be: Business, Support, or Executive"),
            ("⚠️","Auto-KPIs excluded","Diligence Score, SLA Adherence, etc. should not be in actuals upload"),
            ("⚠️","Reports To Code","Must match an existing Staff Code in the same file or overridden in Admin"),
            ("⚠️","Category","Must be: Branch or Head Office"),
        ]
        rules_df = pd.DataFrame(rules, columns=["Status","Rule","Detail"])
        st.dataframe(rules_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### Live validation check")
        st.caption("Upload your file here to check for errors before the main upload.")

        check_file = st.file_uploader("Upload file to validate", type=["xlsx","xls"], key="val_upload")
        if check_file:
            raw_chk = check_file.getvalue()
            try:
                chk_df = pd.read_excel(io.BytesIO(raw_chk), sheet_name='KPI Data', header=1)
                errors = []
                warnings = []

                # Weight check
                chk_df['Weight'] = pd.to_numeric(chk_df['Weight'], errors='coerce').fillna(0)
                chk_df['Weight'] = chk_df['Weight'].apply(lambda x: x/100 if x > 1 else x)
                wt_sums = chk_df.groupby('Staff Code')['Weight'].sum()
                bad_wt  = wt_sums[(wt_sums - 1.0).abs() > 0.005]
                if len(bad_wt):
                    errors.append(f"❌ {len(bad_wt)} staff have weights not summing to 1.00: {bad_wt.index.tolist()[:5]}")
                else:
                    st.success(f"✅ Weights: all {len(wt_sums)} staff sum to 1.00")

                # Pillar check
                valid_pillars = {'Financial','Customer Focus','Operational Excellence'}
                bad_pillars = chk_df[~chk_df['Pillar'].isin(valid_pillars)]['Pillar'].unique()
                if len(bad_pillars):
                    errors.append(f"❌ Invalid pillars: {bad_pillars.tolist()}")
                else:
                    st.success("✅ Pillars: all valid")

                # Duplicates
                dups = chk_df.duplicated(subset=['Staff Code','KPI'])
                if dups.sum():
                    errors.append(f"❌ {dups.sum()} duplicate Staff Code + KPI rows")
                else:
                    st.success(f"✅ No duplicates: {len(chk_df)} unique rows")

                # Annual Target is optional — targets come from cascade
                if 'Annual Target' in chk_df.columns:
                    zero_tgt = (pd.to_numeric(chk_df['Annual Target'], errors='coerce').fillna(0) == 0).sum()
                    if zero_tgt:
                        st.info(f"ℹ️ {zero_tgt} rows have zero/missing Annual Target — "
                                f"targets populate from the cascade module.")
                else:
                    st.success("✅ Annual Target column not present — targets will come from cascade (correct format)")

                st.metric("Total rows",    len(chk_df))
                st.metric("Unique staff",  chk_df['Staff Code'].nunique())
                st.metric("Unique KPIs",   chk_df['KPI'].nunique())
                st.metric("Errors",        len(errors),   delta=f"-{len(errors)}" if errors else "0", delta_color="inverse")
                st.metric("Warnings",      len(warnings), delta=f"-{len(warnings)}" if warnings else "0", delta_color="inverse")

                for e in errors:
                    st.error(e)
                for w in warnings:
                    st.warning(w)
                if not errors and not warnings:
                    st.success("✅ File passed all validation checks. Ready to upload via the sidebar.")
            except Exception as ex:
                st.error(f"Could not read file: {ex}")


# ════════════════════════════════════════════════════════════════
# TAB 9 — LEAVE SETTINGS
# ════════════════════════════════════════════════════════════════
with tabs[7]:
    st.subheader("Leave entitlement settings")
    st.caption(
        "Configure the number of leave days per leave type for your organisation. "
        "These settings apply to all staff and override the system defaults.")

    try:
        from utils.core import LEAVE_TYPES_DEFAULT, save_leave_settings, DATA_DIR
        import json as _json

        # Load current settings
        _lf = DATA_DIR / "leave_settings.json"
        cur_settings = {}
        if _lf.exists():
            try: cur_settings = _json.loads(_lf.read_text())
            except: pass

        st.markdown(
            "<div style='padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;"
            "border-radius:8px;font-size:12px;color:#1E40AF;margin-bottom:14px'>"
            "ℹ️ Leave days are set per your company policy. "
            "Changes take effect immediately for all new leave requests."
            "</div>", unsafe_allow_html=True)

        new_settings = {}
        ls_c1, ls_c2, ls_c3 = st.columns(3)
        cols_cycle = [ls_c1, ls_c2, ls_c3]

        for i, (lt, defaults) in enumerate(LEAVE_TYPES_DEFAULT.items()):
            col = cols_cycle[i % 3]
            saved_days = cur_settings.get(lt, {}).get("days_entitled", defaults["days_entitled"])
            saved_paid = cur_settings.get(lt, {}).get("paid", defaults["paid"])

            with col:
                st.markdown(
                    f"<div style='padding:10px 12px;background:{defaults['color']}12;"
                    f"border-left:3px solid {defaults['color']};border-radius:0 8px 8px 0;"
                    f"margin-bottom:4px'>"
                    f"<b style='font-size:12px'>{lt}</b></div>",
                    unsafe_allow_html=True)
                days = st.number_input(
                    f"Days — {lt}",
                    value=int(saved_days), min_value=0, max_value=365,
                    step=1, key=f"leave_days_{lt}",
                    label_visibility="collapsed",
                    help=f"Number of {lt} days per year (0 = discretionary/unlimited)")
                paid = st.checkbox(
                    "Paid leave", value=bool(saved_paid),
                    key=f"leave_paid_{lt}")
                new_settings[lt] = {
                    **defaults,
                    "days_entitled": days,
                    "max_days":      days,
                    "paid":          paid,
                }

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("💾 Save leave settings", type="primary",
                      key="save_leave_settings"):
            save_leave_settings(new_settings)
            # Reload into session state
            import importlib, utils.core as _core
            _core.LEAVE_TYPES = new_settings
            audit_log("LEAVE_SETTINGS_UPDATED", uname, f"{len(new_settings)} leave types")
            st.toast("✅ Leave settings saved", icon="🏖️")
            st.rerun()

    except Exception as _le:
        st.error(f"Leave settings error: {_le}")

# ════════════════════════════════════════════════════════════════
# TAB 10-KPI — KPI LIBRARY
# ════════════════════════════════════════════════════════════════
with tabs[8]:
    from utils.core import (get_kpi_library, save_kpi_library, DEFAULT_KPI_LIBRARY,
                             DEFAULT_ROLE_KPIS, CBS_SOURCE_LABELS, get_active_kpis)

    st.subheader("📚 KPI Library")
    st.caption(
        "Configure which KPIs this bank tracks, their pillars, weights, "
        "data sources, and which roles carry each KPI. "
        "This drives the BSC, cascade, and CBS actuals engine.")

    _lib = get_kpi_library()
    _pillars        = _lib.get("pillars", DEFAULT_KPI_LIBRARY)
    _role_kpis      = _lib.get("role_kpis", DEFAULT_ROLE_KPIS)
    _active_kpis    = set(_lib.get("active_kpis", []))
    _pillar_weights = _lib.get("pillar_weights", {
        "Financial":0.40,"Customer Focus":0.25,
        "Operational Excellence":0.25,"People & Learning":0.10})
    _pillar_colours = {
        "Financial":"#2563EB","Customer Focus":"#059669",
        "Operational Excellence":"#7C3AED","People & Learning":"#F59E0B"}
    _pillar_icons = {
        "Financial":"💰","Customer Focus":"👥",
        "Operational Excellence":"⚙️","People & Learning":"🎓"}

    _kl_view = st.radio("",
        ["🎯 Activate KPIs","⚖️ Pillar weights","👥 Role assignments",
         "➕ Add custom KPI","📋 Summary"],
        horizontal=True, key="kl_view")
    st.markdown("---")

    # ══ ACTIVATE KPIs ══════════════════════════════════════════════
    if "Activate KPIs" in _kl_view:
        st.markdown(
            "<div style='padding:10px 14px;background:#EFF6FF;border:1px solid #BFDBFE;"
            "border-radius:10px;font-size:12px;color:#1E40AF;margin-bottom:14px'>"
            "Tick the KPIs this bank tracks. Click <b>Save KPI selection</b> when done. "
            "Unticked KPIs are hidden from BSC, cascade and reporting."
            "</div>", unsafe_allow_html=True)

        with st.form("kpi_activate_form"):
            _selections = {}
            for _pillar, _kpis in _pillars.items():
                _pclr = _pillar_colours.get(_pillar,"#6B7280")
                _pico = _pillar_icons.get(_pillar,"📋")
                st.markdown(
                    f"<div style='font-weight:800;font-size:13px;color:{_pclr};"
                    f"margin:16px 0 8px;padding-bottom:4px;"
                    f"border-bottom:2px solid {_pclr}40'>{_pico} {_pillar}</div>",
                    unsafe_allow_html=True)
                _kpi_cols = st.columns(2)
                for _ki, _kpi in enumerate(_kpis):
                    _col = _kpi_cols[_ki % 2]
                    _default = _kpi["id"] in _active_kpis if _active_kpis else True
                    _auto  = _kpi.get("cbs_source","manual") != "manual"
                    _fixed = _kpi.get("fixed",False)
                    _val = _col.checkbox(
                        f"**{_kpi['name']}**", value=_default, key=f"kla_{_kpi['id']}")
                    _selections[_kpi["id"]] = _val
                    _src_badge = (
                        "<span style='background:#ECFDF5;color:#065F46;font-size:9px;"
                        "padding:1px 6px;border-radius:10px;font-weight:700'>🔄 CBS Auto</span>"
                        if _auto else
                        "<span style='background:#FEF3C7;color:#92400E;font-size:9px;"
                        "padding:1px 6px;border-radius:10px;font-weight:700'>✏️ Manual</span>")
                    _fix_badge = (" <span style='background:#EFF6FF;color:#1E40AF;font-size:9px;"
                        "padding:1px 6px;border-radius:10px;font-weight:700'>🔒 Fixed</span>"
                        if _fixed else "")
                    _col.markdown(
                        f"<div style='margin:-6px 0 10px 26px;font-size:10px;color:var(--color-text-tertiary)'>"
                        f"{_src_badge}{_fix_badge}<br>{_kpi.get('description','')[:80]}</div>",
                        unsafe_allow_html=True)

            if st.form_submit_button("💾 Save KPI selection", type="primary", use_container_width=True):
                _new_ids = [kid for kid,val in _selections.items() if val]
                _lib["active_kpis"] = _new_ids
                save_kpi_library(_lib)
                audit_log("KPI_LIBRARY_SAVED", uname, f"{len(_new_ids)} KPIs active")
                st.success(f"✅ {len(_new_ids)} KPIs saved.")
                st.rerun()

    # ══ PILLAR WEIGHTS ═════════════════════════════════════════════
    elif "Pillar weights" in _kl_view:
        st.markdown("**Set pillar weights.** Must total 100%.")
        with st.form("pillar_weights_form"):
            _new_pw = {}
            _pw_cols = st.columns(len(_pillar_weights))
            _pw_total = 0
            for _pi, (_pname, _pw) in enumerate(_pillar_weights.items()):
                _pclr = _pillar_colours.get(_pname,"#6B7280")
                _v = int(_pw*100) if _pw<=1.0 else int(_pw)
                _nv = _pw_cols[_pi].number_input(_pname, 0, 100, _v, key=f"pw_{_pname}")
                _pw_cols[_pi].markdown(
                    f"<div style='text-align:center;font-size:20px;font-weight:800;color:{_pclr}'>{_nv}%</div>",
                    unsafe_allow_html=True)
                _new_pw[_pname] = _nv/100
                _pw_total += _nv
            _pwclr = "#10B981" if _pw_total==100 else "#EF4444"
            st.markdown(
                f"<div style='padding:8px;border-radius:6px;font-weight:700;color:{_pwclr}'>"
                f"Total: {_pw_total}% {'✅' if _pw_total==100 else '— must equal 100%'}</div>",
                unsafe_allow_html=True)
            if st.form_submit_button("💾 Save weights", type="primary"):
                if _pw_total == 100:
                    _lib["pillar_weights"] = _new_pw
                    save_kpi_library(_lib)
                    st.success("✅ Pillar weights saved."); st.rerun()
                else:
                    st.error("Pillar weights must total 100%.")

    # ══ ROLE ASSIGNMENTS ═══════════════════════════════════════════
    elif "Role assignments" in _kl_view:
        st.markdown("**Assign KPIs to roles.** Configure one role, then clone to similar roles.")
        _ss_roles = sorted(DEFAULT_ROLE_KPIS.keys())
        if len(st.session_state.get("staff_scores",pd.DataFrame())):
            _sr = st.session_state["staff_scores"]["Role"].dropna().unique().tolist()
            if _sr: _ss_roles = sorted(set(_ss_roles + _sr))

        _sel_role = st.selectbox("Select role to configure", _ss_roles, key="kl_role_sel")
        if _sel_role:
            _cur_assigned = set(_role_kpis.get(_sel_role,
                DEFAULT_ROLE_KPIS.get(_sel_role,[])))
            _cur_names = [k["name"] for p,ks in _pillars.items() for k in ks if k["id"] in _cur_assigned]
            if _cur_names:
                st.markdown(
                    f"<div style='padding:8px 12px;background:var(--color-background-secondary);border:0.5px solid var(--color-border-tertiary);"
                    f"border-radius:8px;margin-bottom:8px;font-size:11px'>"
                    f"<b>Currently assigned ({len(_cur_names)}):</b> {', '.join(_cur_names)}</div>",
                    unsafe_allow_html=True)
            with st.form(f"kl_role_form"):
                _selections2 = {}
                for _pillar, _kpis in _pillars.items():
                    _active_in_p = [k for k in _kpis if not _active_kpis or k["id"] in _active_kpis]
                    if not _active_in_p: continue
                    _pclr = _pillar_colours.get(_pillar,"#6B7280")
                    st.markdown(
                        f"<div style='font-weight:700;font-size:11px;color:{_pclr};"
                        f"text-transform:uppercase;letter-spacing:.5px;margin:12px 0 4px'>{_pillar}</div>",
                        unsafe_allow_html=True)
                    _rc = st.columns(3)
                    for _ki, _kpi in enumerate(_active_in_p):
                        _v = _rc[_ki%3].checkbox(_kpi["name"],
                            value=_kpi["id"] in _cur_assigned, key=f"kl_ra2_{_sel_role}_{_kpi['id']}")
                        _selections2[_kpi["id"]] = _v
                _n_sel2 = sum(1 for v in _selections2.values() if v)
                st.markdown(f"**{_n_sel2} KPIs selected for {_sel_role}**")
                if st.form_submit_button(f"💾 Save for {_sel_role}", type="primary", use_container_width=True):
                    _new_ids2 = [kid for kid,val in _selections2.items() if val]
                    _role_kpis[_sel_role] = _new_ids2
                    _lib["role_kpis"] = _role_kpis
                    save_kpi_library(_lib)
                    audit_log("KPI_ROLE_SAVED", uname, f"{_sel_role}|{_n_sel2}")
                    st.success(f"✅ {_n_sel2} KPIs saved for {_sel_role}"); st.rerun()

            # Clone to other roles
            st.markdown("---")
            st.markdown(f"**Clone KPIs from {_sel_role} to other roles**")
            _clone_src = set(_role_kpis.get(_sel_role,[]))
            if _clone_src:
                _clone_names = [k["name"] for p,ks in _pillars.items() for k in ks if k["id"] in _clone_src]
                st.caption(f"Will clone: {', '.join(_clone_names[:6])}" + (f" +{len(_clone_names)-6} more" if len(_clone_names)>6 else ""))
                _other = [r for r in _ss_roles if r != _sel_role]
                _clone_tgts = st.multiselect("Apply to these roles", _other, key="kl_clone_tgts")
                _clone_mode = st.radio("Mode",["Replace existing","Merge with existing"], horizontal=True, key="kl_clone_mode")
                if _clone_tgts:
                    _prev_rows = [{"Role":r,"Before":len(_role_kpis.get(r,[])),"After":len(_clone_src if "Replace" in _clone_mode else _clone_src|set(_role_kpis.get(r,[])))} for r in _clone_tgts]
                    st.dataframe(pd.DataFrame(_prev_rows), use_container_width=True, hide_index=True)
                    if st.button(f"📋 Apply to {len(_clone_tgts)} role(s)", type="primary", key="kl_clone_apply"):
                        for _tr in _clone_tgts:
                            _exist = set(_role_kpis.get(_tr,[]))
                            _role_kpis[_tr] = list(_clone_src if "Replace" in _clone_mode else _clone_src|_exist)
                        _lib["role_kpis"] = _role_kpis
                        save_kpi_library(_lib)
                        audit_log("KPI_ROLE_CLONED",uname,f"{_sel_role}→{','.join(_clone_tgts)}")
                        st.success(f"✅ Applied to {len(_clone_tgts)} role(s)"); st.rerun()

    # ══ ADD CUSTOM KPI ═════════════════════════════════════════════
    elif "Add custom KPI" in _kl_view:
        st.markdown("**Add a KPI not in the standard library.**")
        with st.form("add_custom_kpi_form"):
            _ca,_cb = st.columns(2)
            _new_kpi_name   = _ca.text_input("KPI name *", placeholder="e.g. Foreign Currency Deposits")
            _new_kpi_pillar = _cb.selectbox("Pillar *", list(_pillars.keys()))
            _new_kpi_unit   = _ca.selectbox("Unit", ["KES","Count","%","Score","Days","USD","Other"])
            _new_kpi_dir    = _cb.selectbox("Direction", ["Higher is better","Lower is better"])
            _new_kpi_src    = _ca.selectbox("Data source",
                ["manual"] + list(CBS_SOURCE_LABELS.keys()),
                format_func=lambda x: CBS_SOURCE_LABELS.get(x,x))
            _new_kpi_wt     = _cb.number_input("Default weight",0.0,1.0,0.10,0.01,format="%.2f")
            _new_kpi_fixed  = _ca.checkbox("Fixed KPI (bank-wide, not cascaded)")
            _new_kpi_desc   = st.text_area("Description",height=60)
            if st.form_submit_button("➕ Add to library", type="primary"):
                if _new_kpi_name.strip():
                    _new_id = "CUSTOM_"+_new_kpi_name.strip().upper().replace(" ","_")[:20]
                    _new_entry = {"id":_new_id,"name":_new_kpi_name.strip(),
                        "unit":_new_kpi_unit,"direction":"higher" if "Higher" in _new_kpi_dir else "lower",
                        "cbs_source":_new_kpi_src,"fixed":_new_kpi_fixed,
                        "default_weight":_new_kpi_wt,"description":_new_kpi_desc.strip(),"custom":True}
                    if _new_kpi_pillar not in _lib["pillars"]: _lib["pillars"][_new_kpi_pillar]=[]
                    _lib["pillars"][_new_kpi_pillar].append(_new_entry)
                    if _lib.get("active_kpis"): _lib["active_kpis"].append(_new_id)
                    save_kpi_library(_lib)
                    audit_log("KPI_CUSTOM_ADDED",uname,f"{_new_id}|{_new_kpi_pillar}")
                    st.success(f"✅ '{_new_kpi_name}' added."); st.rerun()

    # ══ SUMMARY ════════════════════════════════════════════════════
    elif "Summary" in _kl_view:
        _all_active = get_active_kpis()
        _auto_k = [k for k in _all_active if k.get("cbs_source","manual")!="manual"]
        _sc = st.columns(4)
        _sc[0].metric("Active KPIs",    len(_all_active))
        _sc[1].metric("CBS Auto",       len(_auto_k))
        _sc[2].metric("Manual entry",   len(_all_active)-len(_auto_k))
        _sc[3].metric("Pillars",        len(_pillars))
        for _pillar, _kpis in _pillars.items():
            _active_in_p = [k for k in _kpis if not _active_kpis or k["id"] in _active_kpis]
            if not _active_in_p: continue
            _pclr = _pillar_colours.get(_pillar,"#6B7280")
            _pw = _pillar_weights.get(_pillar,0)
            _pwv = int(_pw*100) if _pw<=1.0 else int(_pw)
            st.markdown(
                f"<div style='padding:8px 14px;background:{_pclr}15;"
                f"border-left:4px solid {_pclr};border-radius:0 8px 8px 0;margin:8px 0'>"
                f"<span style='font-weight:800;color:{_pclr}'>{_pillar}</span>"
                f"<span style='float:right;font-weight:700;color:{_pclr}'>{_pwv}% of BSC</span>"
                f"</div>", unsafe_allow_html=True)
            _kdf = pd.DataFrame([{
                "KPI":k["name"],"Unit":k["unit"],
                "Direction":"↑ Higher" if k.get("direction")=="higher" else "↓ Lower",
                "Source":"🔄 CBS" if k.get("cbs_source","manual")!="manual" else "✏️ Manual",
                "Fixed":"🔒" if k.get("fixed") else "",
            } for k in _active_in_p])
            st.dataframe(_kdf,use_container_width=True,hide_index=True)


# TAB 10 — PIPELINE SETTINGS
# ════════════════════════════════════════════════════════════════
with tabs[9]:
    st.subheader("Pipeline settings")
    st.caption("Customise the product catalogue, stages, and deal configuration for your bank.")

    _ps = get_pipeline_settings()

    # ── Product catalogue editor ──────────────────────────────────────
    st.markdown("#### Product catalogue")
    st.caption("Rename, add, or remove products in each category. Changes take effect immediately in the pipeline module.")

    _custom_cats = _ps.get("product_catalogue", {cat: list(prods) for cat, prods in PRODUCT_CATALOGUE.items()})

    _edited_cats = {}
    for _cat, _prods in _custom_cats.items():
        with st.expander(f"📦 {_cat} ({len(_prods)} products)", expanded=False):
            _prod_text = st.text_area(
                f"Products in {_cat} (one per line)",
                value="\n".join(_prods),
                height=150,
                key=f"ps_cat_{_cat.replace(' ','_')}",
                help="One product per line. Edit, add or remove as needed.")
            _edited_cats[_cat] = [p.strip() for p in _prod_text.split("\n") if p.strip()]

    # Add new category
    with st.expander("➕ Add new product category"):
        _nc_name  = st.text_input("Category name", key="ps_newcat_name",
                                    placeholder="e.g. Islamic Banking")
        _nc_prods = st.text_area("Products (one per line)", key="ps_newcat_prods",
                                  height=100, placeholder="e.g. Murabaha / Ijara / Mudarabah (one per line)")
        if st.button("Add category", key="ps_addcat"):
            if _nc_name.strip():
                _edited_cats[_nc_name.strip()] = [p.strip() for p in _nc_prods.split("\n") if p.strip()]
                _ps["product_catalogue"] = _edited_cats
                save_pipeline_settings(_ps)
                st.success(f"Category '{_nc_name}' added.")
                st.rerun()

    if st.button("💾 Save product catalogue", type="primary", key="ps_save_cat"):
        _ps["product_catalogue"] = _edited_cats
        # Also update flat list
        _ps["product_types"] = [p for prods in _edited_cats.values() for p in prods]
        save_pipeline_settings(_ps)
        audit_log("PIPELINE_SETTINGS", uname, "Product catalogue updated")
        st.success("✅ Product catalogue saved. Pipeline module will use new products immediately.")

    st.markdown("---")

    # ── Stage configuration ───────────────────────────────────────────
    st.markdown("#### Stage names & descriptions")
    st.caption("Rename stages to match your internal sales methodology.")
    _custom_stages = _ps.get("stages", [
        {"stage": s["stage"], "description": s["description"]}
        for s in PIPELINE_STAGES])
    _edited_stages = []
    for _si in _custom_stages:
        _s1, _s2 = st.columns([1,2])
        _sname = _s1.text_input(f"Stage name", value=_si["stage"],
                                  key=f"ps_stg_{_si['stage']}")
        _sdesc = _s2.text_input(f"Description", value=_si.get("description",""),
                                  key=f"ps_stgd_{_si['stage']}")
        _edited_stages.append({"stage":_sname,"description":_sdesc})
    if st.button("💾 Save stage names", key="ps_save_stages"):
        _ps["stages"] = _edited_stages
        save_pipeline_settings(_ps)
        st.success("✅ Stage names saved.")

    st.markdown("---")

    # ── Delete/cancel policy ──────────────────────────────────────────
    st.markdown("#### Delete & cancel policy")
    _del_stage = st.selectbox(
        "Deals can be self-deleted up to and including this stage",
        STAGE_NAMES,
        index=STAGE_NAMES.index(_ps.get("delete_self_max","Lead")) if _ps.get("delete_self_max","Lead") in STAGE_NAMES else 0,
        key="ps_delstage",
        help="Beyond this stage, deletion requires manager approval.")
    _val_stage = st.selectbox(
        "Manager validation required from this stage onwards",
        STAGE_NAMES,
        index=STAGE_NAMES.index(_ps.get("validate_from","Contacted")) if _ps.get("validate_from","Contacted") in STAGE_NAMES else 1,
        key="ps_valstage",
        help="Unvalidated deals are excluded from management forecast reports.")
    if st.button("💾 Save policy", key="ps_save_policy"):
        _ps["delete_self_max"] = _del_stage
        _ps["validate_from"]   = _val_stage
        save_pipeline_settings(_ps)
        st.success(f"✅ Policy saved. Self-delete allowed up to {_del_stage}. Validation required from {_val_stage}.")

# ════════════════════════════════════════════════════════════════
# TAB 11 — SYSTEM HEALTH
# ════════════════════════════════════════════════════════════════
with tabs[10]:
    st.subheader("System health")
    st.caption("Data integrity checks, cascade status, storage usage and active users.")

    import os as _os
    from pathlib import Path as _Path

    _data_dir = DATA_DIR
    _sh1, _sh2, _sh3 = st.columns(3)

    # Data files status
    _files_info = []
    for _fn in ["users.json","target_cascade.json","bank_targets.json",
                "fixed_kpis.json","leaves.json","hr_records.json",
                "pipeline_deals.json","audit_log.json","sbu_action_plans.json"]:
        _fp = _data_dir / _fn
        _exists = _fp.exists()
        _size   = _fp.stat().st_size if _exists else 0
        _files_info.append({
            "File":   _fn,
            "Status": "✅ OK" if _exists else "⚠️ Missing",
            "Size":   f"{_size/1024:.1f} KB" if _size else "0 KB",
        })

    _sh1.markdown("**Data files**")
    _sh1.dataframe(pd.DataFrame(_files_info), hide_index=True, use_container_width=True)

    # User stats
    _active_users = sum(1 for u in um.users.values() if u.get("active"))
    _total_users  = len(um.users)
    _admin_users  = sum(1 for u in um.users.values() if u.get("is_admin") or u.get("role","").lower()=="admin")
    _must_pw      = sum(1 for u in um.users.values() if u.get("must_change_password"))

    _sh2.markdown("**User accounts**")
    _sh2.metric("Total users", _total_users)
    _sh2.metric("Active",      _active_users)
    _sh2.metric("Admins",      _admin_users)
    if _must_pw:
        _sh2.warning(f"⚠️ {_must_pw} user(s) must change password on next login")

    # Cascade health
    _sh3.markdown("**Cascade health**")
    if casc:
        try:
            _casc_data = getattr(casc, "cascade", {}) or {}
            _bt_data   = getattr(casc, "bank_targets", {}) or {}
            _allocs    = sum(1 for k in _casc_data if not k.startswith("deadline|") and not k.startswith("global_"))
            _deadlines = sum(1 for k in _casc_data if k.startswith("deadline|"))
            _bt_count  = len(_bt_data)
            try:    _fixed_k = len(casc.get_fixed_kpis(_gfy()))
            except: _fixed_k = 0
            _sh3.metric("Allocations saved", _allocs)
            _sh3.metric("Deadlines set",     _deadlines)
            _sh3.metric("Bank targets set",  _bt_count)
            _sh3.metric("Fixed KPIs (2026)", _fixed_k)
        except Exception as _ce:
            _sh3.warning(f"Cascade data unavailable: {_ce}")
    else:
        _sh3.info("Cascade manager not loaded.")

    # Audit log preview
    st.markdown("---")
    st.markdown("**Recent audit entries**")
    _al_file = _data_dir / "audit_log.json"
    if _al_file.exists():
        try:
            _al = json.loads(_al_file.read_text())
            _al_df = pd.DataFrame(_al[-20:] if len(_al)>20 else _al)
            if not _al_df.empty:
                st.dataframe(_al_df[::-1], hide_index=True,
                             use_container_width=True, height=250)
        except:
            st.info("Could not read audit log.")
    else:
        st.info("No audit log yet.")

    # Quick fix actions
    st.markdown("---")
    st.markdown("**Quick fix actions**")
    _qc1, _qc2 = st.columns(2)
    if _qc1.button("🔄 Reload cascade manager", key="sh_reload_casc"):
        try:
            st.session_state["cascade_manager"] = CascadeManager()
            st.success("✅ Cascade manager reloaded from disk.")
        except Exception as _e:
            st.error(f"Error: {_e}")
    if _qc2.button("🧹 Clear session cache", key="sh_clear_cache"):
        _keep = {"logged_in","username","user_data","user_manager"}
        _to_del = [k for k in st.session_state if k not in _keep]
        for k in _to_del: del st.session_state[k]
        st.success(f"✅ Cleared {len(_to_del)} session keys. Data managers will reload.")
        st.rerun()
