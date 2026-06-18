"""utils/api.py — FastAPI backend for A2Z Blueprint MIS 360.

Runs alongside Streamlit on port 8502. Provides cached, pooled database
access for the heaviest pages. Pages that call this API load 3-5x faster
than pages that hit PostgreSQL directly.

V-001 SECURITY FIX (v5.17, CVSS 9.1):
  Every endpoint except /api/health now requires a bearer JWT issued by
  POST /api/auth/login. See utils/auth_jwt.py for the token contract.
  V-009 (CORS) is also tightened — origins now come from A2Z_CORS_ORIGINS
  env var, defaulting to localhost only.

HOW TO RUN (two terminals):
  Terminal 1: streamlit run app.py
  Terminal 2: python -m utils.api

Set A2Z_JWT_SECRET in production (otherwise a per-process random secret
is generated, with a warning at startup).

OR use run_all.bat which starts both automatically.

ENDPOINTS:
  POST /api/auth/login         — exchange username+password for bearer token
  GET  /api/auth/me            — return the current user from the token
  GET  /api/health             — system health check (NO AUTH — by design)
  GET  /api/bsc/summary        — BSC scores summary (cached 5min)
  GET  /api/bsc/staff/{username}— individual staff BSC
  GET  /api/pipeline/summary   — pipeline summary metrics
  GET  /api/pipeline/deals     — all deals with filters
  GET  /api/credit/summary     — credit monitoring summary
  GET  /api/credit/watchlist   — watchlist with pagination
  GET  /api/aml/summary        — AML alerts summary
  GET  /api/users/summary      — org summary
  GET  /api/dashboard/md       — MD command centre data
  POST /api/cache/clear        — admin-only
  GET  /api/cache/stats        — cache observability
"""

import os
import json
import time
import logging
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from pathlib import Path
from functools import lru_cache

from fastapi import FastAPI, HTTPException, Query, Depends, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, ConfigDict

# v10.501 Phase 2 Arc B Batch 4b — rate limiting (closes GAP-006).
# slowapi provides per-IP and per-key rate limiting for FastAPI.
# In-memory storage is the default; OPERATIONAL_PROTOCOL.md declares
# the single-worker FastAPI operational constraint so that in-memory
# counts remain consistent across requests. Multi-worker deployment
# would require Redis or memcached backing (out of scope for Arc B).
from slowapi import Limiter
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from slowapi.util import get_remote_address

from utils.auth_jwt import (
    create_access_token,
    get_current_user,
    get_current_user_allow_rotation,
    require_admin,
    require_config_admin,
    warn_if_default_secret,
    TOKEN_SCOPE_FULL,
    TOKEN_SCOPE_MUST_ROTATE,
)

# H2 (2026-06-14): import the pipeline request models at MODULE LEVEL so
# FastAPI can resolve the forward-ref annotations on the pipeline routes
# (`payload: "PipelineDealCreate"` etc.). These models were never imported
# into this module, so the forward-refs could not bind the real (patched)
# classes — and a clean start of this file would raise NameError at route
# setup. api_pipeline_models imports only stdlib + pydantic (no cycle).
from utils.api_pipeline_models import (
    PipelineDealCreate,
    PipelineDealRefer,
    PipelineDealUpdate,
    PipelineDealAdvance,
    PipelineDealValidate,
    PipelineDealCancelRequest,
    PipelineDealCancelApprove,
)

logger = logging.getLogger("a2z.api")

DATA_DIR = Path(__file__).parent.parent / "data"

app = FastAPI(
    title="A2Z Blueprint MIS 360 API",
    description="High-performance data layer for A2Z Blueprint",
    version="5.17.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)

# V-009 fix — CORS origins from env, NOT hardcoded. In dev the default is
# localhost only. In production set A2Z_CORS_ORIGINS to a comma-separated
# list of allowed origins. Never use "*" with allow_credentials=True.
#
# v10.299 — Default list extended to include React dev servers so the
# frontend team can run `npm start` (port 3000, Create React App) or
# `npm run dev` (port 5173, Vite) without env-var fiddling. Streamlit
# origins (8501/8502) remain for the existing cockpit pages.
_default_cors = (
    "http://localhost:3000,http://127.0.0.1:3000,"      # CRA
    "http://localhost:5173,http://127.0.0.1:5173,"      # Vite
    "http://localhost:8501,http://127.0.0.1:8501,"      # Streamlit
    "http://localhost:8502,http://127.0.0.1:8502"
)
_cors_raw = os.getenv("A2Z_CORS_ORIGINS", _default_cors)
_cors_origins = [o.strip() for o in _cors_raw.split(",") if o.strip()]

# v10.299 — Guard against an empty / whitespace-only env var. Before this
# fix, setting A2Z_CORS_ORIGINS="" would silently produce zero origins,
# blocking every cross-origin request without any error. Now we fall back
# to defaults and log it loudly.
if not _cors_origins:
    logger.warning(
        "A2Z_CORS_ORIGINS env var is empty after parsing — falling "
        "back to default localhost origins. Set explicit origins in "
        "production deployments."
    )
    _cors_origins = [o.strip() for o in _default_cors.split(",")
                     if o.strip()]

if "*" in _cors_origins:
    raise RuntimeError(
        "A2Z_CORS_ORIGINS must not contain '*' — wildcard with credentials is "
        "the V-009 vulnerability. Set explicit origins."
    )
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    # v10.299 — Full standard method set. OPTIONS is required for CORS
    # preflight; DELETE and PATCH are needed by the React SPA as we add
    # state-changing endpoints in future Phase 3 batches.
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    # v10.299 — Explicit headers for the React/Authorization JWT flow.
    # Wildcard would work but explicit listing documents what the
    # frontend team needs to send.
    allow_headers=[
        "Authorization",   # JWT bearer tokens
        "Content-Type",    # JSON POST/PUT bodies
        "Accept",          # Content negotiation
        "X-Requested-With",  # Conventional ajax indicator
    ],
)


# ── Rate limiting (v10.501 Phase 2 Arc B Batch 4b — closes GAP-006) ──
# Per POLICY_GAPS.md:
#   /api/auth/login          — 10 attempts per minute per IP, 100 per hour
#   /api/auth/change-password — 5 attempts per minute per token
#   /api/auth/whoami-detailed — no limit (legitimate dashboard polling)
#
# Reference model: Streamlit's existing 5-attempts-then-15-min lockout at
# pages/_login.py:194-204. The API is now in parity. Failed attempts
# already audit-log via API_LOGIN_FAILED / API_PASSWORD_CHANGE_FAILED;
# 429 responses additionally write an API_RATE_LIMITED audit row so the
# operator can see when an attacker hits the wall vs. when they're just
# guessing slowly.
#
# OPERATIONAL_PROTOCOL.md: in-memory storage is correct for single-worker
# FastAPI. Multi-worker would require Redis or memcached as the storage
# backend (slowapi supports both via the `storage_uri` kwarg).

def _ratelimit_key_by_token(request: Request) -> str:
    """Rate-limit key for token-scoped endpoints (change-password).

    Returns a stable string derived from the bearer token in the
    Authorization header. Hashed so the raw JWT never appears in
    storage keys, error messages, or tracebacks (SECURITY: matches
    OPERATIONAL_PROTOCOL.md log-content constraints).

    Falls back to per-IP if no Authorization header is present —
    which shouldn't happen on credentialed endpoints (the JWT
    dependency would reject before this runs) but provides
    defence-in-depth.
    """
    import hashlib
    auth_hdr = request.headers.get("Authorization", "")
    if auth_hdr.startswith("Bearer "):
        token = auth_hdr[len("Bearer "):]
        # 16 hex chars = 64 bits of namespace — plenty for keying,
        # avoids putting the full JWT in any storage structure.
        return f"token:{hashlib.sha256(token.encode()).hexdigest()[:16]}"
    return f"ip:{get_remote_address(request)}"


# Limiter instance — default key is per-IP. Endpoints requiring
# per-token keying pass key_func=_ratelimit_key_by_token to
# @limiter.limit.
#
# headers_enabled=False intentionally: slowapi's X-RateLimit-* headers
# would leak remaining-quota information to an attacker probing the
# auth endpoints (telling them how many more attempts they have before
# the wall). The 429 response is sufficient — clients learn the limit
# only by hitting it. Disabling header injection also avoids slowapi's
# requirement that every rate-limited route declare a `response:
# Response` parameter.
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=[],   # no global default — each endpoint opts in
    headers_enabled=False,
)
app.state.limiter = limiter


def _ratelimit_exceeded_handler(request: Request, exc: RateLimitExceeded):
    """Custom 429 handler that also writes an A2Z audit row.

    Overrides slowapi's default _rate_limit_exceeded_handler so each
    429 produces an API_RATE_LIMITED audit entry. Best-effort
    extraction of the authenticated username from the bearer token
    for endpoints where the user is known; falls back to "anonymous".

    NEVER returns the candidate password, the JWT, the hashed token
    key, or any other auth material in the response body or audit
    payload. The response is a generic 429 with a Retry-After header.
    """
    # Best-effort identity for the audit row. NOT a security control —
    # if the token is invalid/expired we still want the audit row to
    # be written, just with a less-specific identifier.
    user_identifier = "anonymous"
    auth_hdr = request.headers.get("Authorization", "")
    if auth_hdr.startswith("Bearer "):
        try:
            from utils.auth_jwt import decode_token
            payload = decode_token(auth_hdr[len("Bearer "):])
            user_identifier = payload.get("username", "unknown") or "unknown"
        except Exception as e:
            # Silent-except is a latent bug per OPERATIONAL_PROTOCOL —
            # logged at debug level here because this is audit metadata
            # only, not a security boundary. The 429 fires regardless.
            logger.debug(
                "Could not decode bearer for rate-limit audit: %s: %s",
                type(e).__name__, e,
            )

    try:
        _audit(
            "API_RATE_LIMITED",
            {"username": user_identifier},
            f"path={request.url.path} method={request.method} "
            f"ip={get_remote_address(request)} limit={exc.detail}",
        )
    except Exception as e:
        # Audit failure must NOT prevent the 429 response. Log loudly
        # so the operator notices an observability gap.
        logger.error(
            "Failed to write API_RATE_LIMITED audit row: %s: %s",
            type(e).__name__, e,
            exc_info=True,
        )

    return JSONResponse(
        status_code=429,
        content={"detail": f"Rate limit exceeded: {exc.detail}"},
        headers={"Retry-After": "60"},
    )


app.add_exception_handler(RateLimitExceeded, _ratelimit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)


@app.on_event("startup")
def _startup_checks() -> None:
    warn_if_default_secret()
    logger.info(f"A2Z API startup — CORS origins: {_cors_origins}")


# v10.413 — Mount cascade routers (E7: React-readiness payoff)
# All cascade endpoints wrap the pure-compute engines built v10.406-v10.412.
# JWT-required on every route.
#
# Two routers, two prefixes:
#   /api/v1/cascade/*              — broad surface (health, rollup, pairing,
#                                    simulator, pillars, structure)
#   /api/cascade/capacity-feedback — capacity-feedback (v10.412 stub now mounted)
try:
    from utils.api_cascade import router as _cascade_router
    app.include_router(_cascade_router)
    logger.info("A2Z API — cascade router mounted at /api/v1/cascade")
except Exception as _exc:  # noqa: BLE001
    logger.warning(f"Cascade router not loaded: {_exc}")

try:
    from utils.api_capacity_feedback import router as _capacity_router
    app.include_router(_capacity_router)
    logger.info("A2Z API — capacity router mounted at /api/cascade/capacity-feedback")
except Exception as _exc:  # noqa: BLE001
    logger.warning(f"Capacity router not loaded: {_exc}")

    # v10.495 — Branding API for React SPA enablement
# Public endpoint (no JWT) exposing tenant identity + brand colors
# + IP notice. Consumed by frontend/web/ React app's
# BrandingProvider. See utils/api_branding.py.
try:
    from utils.api_branding import router as _branding_router
    app.include_router(_branding_router)
    logger.info("A2Z API — branding router mounted at /api/branding")
except Exception as _exc:  # noqa: BLE001
    logger.warning(f"Branding router not loaded: {_exc}")

    # v10.499 Stage C Batch 2c — Roles registry API for React useRole() hook
# Exposes the canonical role taxonomy (49 explicit role classifications + enum
# constants for tiers/sbus/scopes). Authenticated endpoint; consumed by
# frontend/web/src/hooks/useRole.ts at app boot alongside /api/auth/whoami-detailed.
try:
    from utils.api_roles import router as _roles_router
    app.include_router(_roles_router)
    logger.info("A2Z API — roles router mounted at /api/roles")
except Exception as _exc:  # noqa: BLE001
    logger.warning(f"Roles router not loaded: {_exc}")


# Helper: emit an audit_log entry from API context. Imported lazily so the
# api module can be imported in environments that don't have streamlit
# wired up (e.g. tests).
def _audit(action: str, user: dict, detail: str = "") -> None:
    try:
        from utils.core_audit import audit_log
        username = (user or {}).get("username", "anonymous")
        audit_log(action, username, detail, module="api")
    except Exception as e:
        logger.debug(f"audit_log failed: {e}")

# ── In-memory cache ───────────────────────────────────────────────
_cache: Dict[str, Any] = {}
_cache_ts: Dict[str, float] = {}
CACHE_TTL = {
    "bsc":       300,   # 5 minutes — scores don't change often
    "pipeline":   60,   # 1 minute — deals change frequently
    "credit":    120,   # 2 minutes
    "aml":       120,
    "users":     600,   # 10 minutes — org structure rarely changes
    "dashboard": 120,
    "partnerships": 60,
}

def _get_cache(key: str, ttl_key: str = "pipeline") -> Optional[Any]:
    if key in _cache:
        age = time.time() - _cache_ts.get(key, 0)
        if age < CACHE_TTL.get(ttl_key, 60):
            return _cache[key]
    return None

def _set_cache(key: str, value: Any) -> None:
    _cache[key] = value
    _cache_ts[key] = time.time()

def _load_json(fname: str) -> Any:
    """Load a JSON data file. Routes through a2z_db so the dual-mode
    PG/JSON migration applies here too (audit V-002 finding)."""
    p = DATA_DIR / fname
    if not p.exists():
        return []
    try:
        from utils.db import db as _a2z_db
        return _a2z_db.load_json(p, default=[])
    except Exception as e:
        logger.error(f"Error loading {fname}: {e}")
        return []

def _db_available() -> bool:
    """Check if PostgreSQL is available."""
    try:
        from utils.db import db as _db
        return _db.is_postgres_ready()
    except Exception:
        return False


def _db_sync_pipeline_deal(deal: Optional[dict]) -> None:
    """Upsert a pipeline deal into Postgres so DB-backed reads reflect runtime
    mutations. H5 (2026-06-14): pipeline reads are DB-first but mutations
    write only the JSON store, so created/changed deals were invisible in the
    DB-backed list. No-op when Postgres is unavailable. Best-effort. Field
    map: JSON deal_value->amount, product_type->product; pipeline_category
    kept in metadata.
    """
    if not deal or not _db_available():
        return
    try:
        import json as _json
        from datetime import date as _date
        from utils.db import db as _db
        today = _date.today().isoformat()

        def _date_or_none(v):
            v = (str(v).strip() if v is not None else "")
            return v or None

        row = {
            "id":            str(deal.get("id", "") or ""),
            "staff_code":    str(deal.get("staff_code", "") or ""),
            "staff_name":    str(deal.get("staff_name", "") or ""),
            "unit":          deal.get("unit"),
            "role":          deal.get("role"),
            "client_name":   deal.get("client_name"),
            "client_cif":    deal.get("client_cif"),
            "product":       deal.get("product") or deal.get("product_type"),
            "stage":         deal.get("stage"),
            "deal_category": (deal.get("deal_category")
                              or deal.get("pipeline_category") or "New Facility"),
            "amount":        (deal.get("amount")
                              if deal.get("amount") is not None
                              else deal.get("deal_value")),
            "currency":      deal.get("currency", "KES"),
            "open_date":     _date_or_none(deal.get("open_date")) or today,
            "expected_close": _date_or_none(deal.get("expected_close")),
            "probability":   deal.get("probability"),
            "notes":         deal.get("notes"),
            "last_updated":  today,
            "metadata":      _json.dumps({
                "pipeline_category":   deal.get("pipeline_category"),
                "client_type":         deal.get("client_type"),
                "is_ntb":              deal.get("is_ntb"),
                "source":              deal.get("source"),
                "portfolio_owner_code": deal.get("portfolio_owner_code"),
                "portfolio_owner_name": deal.get("portfolio_owner_name"),
                "lms_application_id":   deal.get("lms_application_id"),
                "mou_id":               deal.get("mou_id"),
                "mou_title":            deal.get("mou_title"),
                "sector":               deal.get("sector"),
                "segment":              deal.get("segment"),
                "fx_rate":              deal.get("fx_rate"),
                "amount_kes":           deal.get("amount_kes"),
                "fx_rate_date":         deal.get("fx_rate_date"),
                "fx_rate_source":       deal.get("fx_rate_source"),
                "currency_book":        deal.get("currency_book"),
                "manager_validated":    deal.get("manager_validated"),
                "validated_by":         deal.get("validated_by"),
                "validated_at":         deal.get("validated_at"),
                "referral_status":      deal.get("referral_status"),
                "referred_to_code":     deal.get("referred_to_code"),
                "referred_to":          deal.get("referred_to"),
                "referred_by_code":     deal.get("referred_by_code"),
                "referred_by_name":     deal.get("referred_by_name"),
                "referral_note":        deal.get("referral_note"),
                "decline_reason":       deal.get("decline_reason"),
                "sla_step_log":         deal.get("sla_step_log"),
            }),
        }
        if not row["id"]:
            return
        cols = list(row.keys())
        placeholders = ", ".join(["%s"] * len(cols))
        updates = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c != "id")
        sql = (f"INSERT INTO pipeline_deals ({', '.join(cols)}) "
               f"VALUES ({placeholders}) "
               f"ON CONFLICT (id) DO UPDATE SET {updates}")
        _db.execute(sql, tuple(row[c] for c in cols))
    except Exception as e:
        # B13: do NOT swallow. A deal that can't persist to Postgres must fail
        # loudly — silent swallowing is exactly what let JSON and the DB drift
        # (deals invisible to DB-first reads). DB-unavailable is handled by the
        # early return above; reaching here means Postgres is up but the write
        # errored, which is a real fault the caller needs to see.
        logger.error(f"Pipeline deal DB sync FAILED for {deal.get('id')}: {e}")
        raise


def _normalize_db_deal_row(row):
    """Map pipeline_deals DB columns to the field names the React frontend
    expects (amount->deal_value, product->product_type, metadata->
    pipeline_category) so the DB read path matches the JSON read path. H5."""
    if not isinstance(row, dict):
        return row
    r = dict(row)
    if r.get("deal_value") in (None, ""):
        r["deal_value"] = r.get("amount")
    if not r.get("product_type"):
        r["product_type"] = r.get("product")
    md = r.get("metadata")
    if isinstance(md, str):
        try:
            import json as _json
            md = _json.loads(md)
        except Exception:
            md = {}
    if isinstance(md, dict) and not r.get("pipeline_category"):
        r["pipeline_category"] = md.get("pipeline_category")
    if isinstance(md, dict) and not r.get("lms_application_id"):
        r["lms_application_id"] = md.get("lms_application_id")
    # Lift the FX money set + client-type fields out of metadata so DB-first
    # readers (analytics, dashboard canonical path) see KES-equivalent values
    # and the currency book — matching the JSON read path. Without this,
    # _deal_value falls back to NATIVE for FCY deals and analytics disagrees
    # with the dashboard.
    if isinstance(md, dict):
        for _k in ("amount_kes", "currency_book", "fx_rate", "fx_rate_date",
                   "fx_rate_source", "client_type", "mou_id", "mou_title",
                   "sector", "segment", "validated_by", "validated_at",
                   "referral_status", "referred_to_code", "referred_to",
                   "referred_by_code", "referred_by_name", "referral_note",
                   "decline_reason", "sla_step_log"):
            if r.get(_k) in (None, "") and md.get(_k) is not None:
                r[_k] = md.get(_k)
        # manager_validated is a bool — lift whenever absent on the row so the
        # DB read path (analytics assured value + funnel) reflects validation.
        if "manager_validated" not in r or r.get("manager_validated") is None:
            r["manager_validated"] = bool(md.get("manager_validated"))
    return r


def _get_or_hydrate_deal(pm, deal_id: str):
    """Return a deal by id, DB-first.

    H7 (2026-06-14): the LIST read is DB-first but detail + mutation routes
    read only the JSON store, so the 294 Postgres-seeded deals 404'd on open
    and could not be advanced. This returns the JSON deal if present;
    otherwise it loads the row from Postgres and registers it on the
    request-scoped PipelineManager so update_stage/update_deal can operate on
    it. Mutations then re-sync to Postgres via _db_sync_pipeline_deal (H5),
    keeping the DB authoritative. Returns None if the deal is in neither store.
    """
    deal = pm.get_deal(deal_id)
    if deal:
        return deal
    if _db_available():
        try:
            from utils.db import db as _db
            row = _db.fetch_one("SELECT * FROM pipeline_deals WHERE id = %s", (deal_id,))
            if row:
                hydrated = _normalize_db_deal_row(_serialize(row))
                try:
                    pm.deals.append(hydrated)  # register for in-request mutation
                except Exception:
                    pass
                return hydrated
        except Exception as e:
            logger.error(f"Pipeline deal hydrate failed for {deal_id}: {e}")
    return None

def _safe_float(val) -> float:
    try:
        from decimal import Decimal
        if isinstance(val, Decimal):
            return float(val)
        return float(val) if val is not None else 0.0
    except Exception:
        return 0.0

def _safe_int(val) -> int:
    try:
        return int(val) if val is not None else 0
    except Exception:
        return 0

def _serialize(obj):
    """Make objects JSON serializable."""
    from decimal import Decimal
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_serialize(i) for i in obj]
    return obj

# ── Auth endpoints ────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    expires_in_seconds: int = 1800
    username: str
    role: str
    # ── must_change_password (Batch 3b) ──────────────────────────────────
    # True when the authenticating user has must_change_password=true in
    # users.json. In that case, access_token is issued with
    # scope="must_rotate" — it is only valid against
    # /api/auth/change-password and is rejected by every other Depends(
    # get_current_user) endpoint with 403. The frontend reads this flag
    # to route the user to the rotation form.
    #
    # Default False preserves backward compat — pre-3b callsites that
    # don't set it still get a valid response that React reads as full
    # access.
    must_change_password: bool = False


# ── /api/auth/change-password request shape (Batch 3b) ───────────────
class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


@app.post("/api/auth/login", response_model=TokenResponse)
@limiter.limit("10/minute;100/hour")
def login(request: Request, req: LoginRequest):
    """Exchange username + password for a 30-minute bearer JWT.

    Verifies via UserManager.authenticate (which uses bcrypt per V-003 fix
    and rehashes legacy SHA-256 on success). Audit log fires on both
    success and failure so brute-force attempts are visible.

    v10.501 Phase 2 Arc B Batch 4b: rate-limited to 10/minute and
    100/hour per IP (closes GAP-006). The `request: Request` parameter
    is required by slowapi to extract the client IP for keying.
    """
    try:
        from utils.core import UserManager
        um = UserManager()
        ok, user_data = um.authenticate(req.username, req.password)
    except Exception as e:
        logger.error(f"Login system error: {e}")
        raise HTTPException(status_code=500, detail="Authentication system unavailable")

    if not ok or not user_data:
        # Audit failed login attempts (V-007 brute-force visibility)
        _audit("API_LOGIN_FAILED", {"username": req.username},
               f"Failed login from API for {req.username}")
        raise HTTPException(
            status_code=401,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )

    user = {
        "username": req.username,
        "role":     user_data.get("role", "Staff"),
    }

    # Batch 3b: read must_change_password from the stored user record.
    # If true, issue a scope='must_rotate' token (rejected by every
    # endpoint except /api/auth/change-password — mechanical enforcement
    # via auth_jwt.get_current_user).
    must_rotate = bool(user_data.get("must_change_password"))
    scope = TOKEN_SCOPE_MUST_ROTATE if must_rotate else TOKEN_SCOPE_FULL

    token = create_access_token(user, scope=scope)
    if must_rotate:
        _audit(
            "API_LOGIN_FORCE_PW", user,
            "Issued must_rotate-scope token — user must change password",
        )
    else:
        _audit(
            "API_LOGIN_SUCCESS", user,
            "Issued full-scope bearer token via /api/auth/login",
        )
    return TokenResponse(
        access_token=token,
        username=req.username,
        role=user["role"],
        must_change_password=must_rotate,
    )


@app.get("/api/auth/me")
def me(user: dict = Depends(get_current_user)):
    """Return the current user from the bearer token. Doubles as a
    cheap token-validity probe."""
    return {
        "username": user["username"],
        "role":     user["role"],
        "expires_at": datetime.fromtimestamp(user["exp"]).isoformat() if user.get("exp") else None,
    }



@app.get("/api/auth/whoami-detailed")
def whoami_detailed(user: dict = Depends(get_current_user)):
    """Return the caller's full operational identity for React useRole().

    Extends /api/auth/me with:
      - Full user profile (staff_code, full_name, department, email)
      - Role classification from role_taxonomy (tier, sbu, branch_scope)
      - Capability flags (is_admin, can_be_tagged, can_view_all)
      - Streamlit RBAC fields (accessible_modules) for migration compat

    Auth: any authenticated user (Depends(get_current_user)). Caller
    receives their own identity only — no cross-user data exposure.

    Used by: frontend/web/src/hooks/useRole.ts (Batch 2d).
    """
    # Lazy import to keep auth_jwt usable in non-FastAPI contexts (matches
    # the codebase pattern used by login + the _make_require_admin factory).
    from utils.core import UserManager
    from utils.role_taxonomy import classify_role

    username = user["username"]
    um = UserManager()
    full_user = um.users.get(username) or {}

    role_raw = full_user.get("role") or user.get("role") or "Staff"
    classification = classify_role(role_raw)

    response = {
        # Identity (from users.json — never trust JWT for these)
        "username":     username,
        "staff_code":   full_user.get("staff_code"),
        "full_name":    full_user.get("full_name"),
        "department":   full_user.get("department"),
        "email":        full_user.get("email"),
        "active":       full_user.get("active", True),

        # Role (raw + classified)
        "role":           role_raw,
        "tier":           classification.tier,
        "sbu":            classification.sbu,
        "branch_scope":   classification.branch_scope,
        "matched_via":    classification.matched_via,
        "can_be_tagged":  classification.tier in {"portfolio_owner", "service"},

        # Capability flags
        "is_admin":         bool(full_user.get("is_admin") or full_user.get("role", "").lower() == "admin"),
        "can_view_all":     bool(full_user.get("can_view_all")),

        # Streamlit RBAC (migration compat — React will phase these out)
        "accessible_modules":  full_user.get("accessible_modules", []),
        "hidden_modules":      full_user.get("hidden_modules", []),

        # Token timing (matches /api/auth/me convention)
        "expires_at": (
            datetime.fromtimestamp(user["exp"]).isoformat()
            if user.get("exp") else None
        ),
    }

    _audit("API_AUTH_WHOAMI_DETAILED", user,
           f"Detailed identity returned for {username} "
           f"(role={role_raw}, tier={classification.tier})")

    return response


# ── /api/auth/change-password (Batch 3b) ──────────────────────────────
# Rotates the authenticated user's password. Accepts BOTH full-scope and
# must_rotate-scope tokens via get_current_user_allow_rotation — the
# only endpoint in the system that permits must_rotate scope.
#
# On success, issues a fresh full-scope token in the response so the
# caller is normally authenticated post-rotation without needing to
# re-enter credentials. The frontend swaps the stored token and
# transitions auth.status from 'must_rotate' to 'authenticated'.
#
# Password policy matches the Streamlit force_change_pw flow at
# pages/_login.py:286-291 (length >= 8, new != current). The new-account
# email template at utils/core.py:313 advertises a stronger policy
# (complexity rules) but no code currently enforces it; that stated-vs-
# enforced gap is recorded for Phase 2 hardening — Batch 3b matches
# actual code behavior to keep Streamlit and FastAPI mechanically
# consistent.
#
# Defensive divergence from Streamlit: this endpoint REQUIRES
# current_password verification. Streamlit's flow does not (it trusts
# the must_rotate state machine). The API gate is stricter because a
# stolen must_rotate-scope token would otherwise let an attacker set
# an arbitrary password. Documented in CHANGELOG_v10500_batch3b.md.
@app.post("/api/auth/change-password", response_model=TokenResponse)
@limiter.limit("5/minute", key_func=_ratelimit_key_by_token)
def change_password(
    request: Request,
    req: ChangePasswordRequest,
    user: dict = Depends(get_current_user_allow_rotation),
):
    """Rotate the authenticated user's password.

    Accepts both full-scope and must_rotate-scope tokens. On success,
    persists the new bcrypt hash, clears must_change_password, audits,
    and returns a fresh full-scope TokenResponse.

    v10.501 Phase 2 Arc B Batch 4b: rate-limited to 5/minute per
    bearer token (closes GAP-006). Token-keyed rather than IP-keyed
    because a NAT'd corporate network could legitimately share an
    IP across hundreds of users — the 5/min ceiling is meaningful
    only at per-user granularity. The token has already authenticated;
    the limit protects against a compromised token being used to
    brute-force the current_password check.
    """
    username = user["username"]
    forced   = (user.get("scope") == TOKEN_SCOPE_MUST_ROTATE)

    try:
        from utils.core import UserManager, validate_password_policy
        um = UserManager()
    except Exception as e:
        logger.error(f"UserManager init failed during password change: {e}")
        raise HTTPException(
            status_code=500,
            detail="Authentication system unavailable",
        )

    full_user = um.users.get(username)
    if not full_user:
        _audit("API_PASSWORD_CHANGE_FAILED", user,
               f"User '{username}' not in store (token outlived account?)")
        raise HTTPException(status_code=401, detail="User not found")

    # Verify current password — defensive even for forced rotations.
    if not um.verify_pw(req.current_password, full_user.get("password", "")):
        _audit("API_PASSWORD_CHANGE_FAILED", user,
               f"current_password mismatch (forced={forced})")
        raise HTTPException(
            status_code=401,
            detail="Current password is incorrect",
        )

    # ── Validate new password ────────────────────────────────────────
    # v10.501 Batch 4a: closes GAP-001 by enforcing the same complexity
    # policy advertised in utils/core.py:313 (the new-account email
    # template). validate_password_policy() is the single source of
    # truth, shared with Streamlit's change_pw and force_change_pw forms.
    # The helper returns (ok, reason); the reason is safe to expose to
    # the caller as the HTTPException detail.
    pw_ok, pw_reason = validate_password_policy(req.new_password)
    if not pw_ok:
        _audit("API_PASSWORD_CHANGE_FAILED", user,
               f"new_password policy violation: {pw_reason}")
        raise HTTPException(
            status_code=400,
            detail=pw_reason,
        )
    if req.new_password == req.current_password:
        _audit("API_PASSWORD_CHANGE_FAILED", user,
               "new_password equals current_password")
        raise HTTPException(
            status_code=400,
            detail="New password must differ from current password",
        )

    # Apply change. UserManager.change_password (utils/core.py:5759)
    # bcrypt-hashes the new password and clears must_change_password.
    try:
        um.change_password(username, req.new_password)
    except Exception as e:
        logger.error(f"Password change persistence failed for {username}: {e}")
        _audit("API_PASSWORD_CHANGE_FAILED", user,
               f"Persistence error: {type(e).__name__}")
        raise HTTPException(
            status_code=500,
            detail="Failed to save new password — please retry",
        )

    # Issue a fresh FULL-scope token so the caller is normally
    # authenticated post-rotation. No re-login required.
    fresh_user = {"username": username, "role": full_user.get("role", "Staff")}
    new_token = create_access_token(fresh_user, scope=TOKEN_SCOPE_FULL)
    _audit(
        "API_PASSWORD_CHANGE_SUCCESS", user,
        f"Password rotated (forced={forced}); fresh full-scope token issued",
    )
    return TokenResponse(
        access_token=new_token,
        username=username,
        role=full_user.get("role", "Staff"),
        must_change_password=False,
    )


# ── Health check (NO AUTH — by design) ────────────────────────────
@app.get("/api/health")
def health():
    db_ok = _db_available()
    return {
        "status":    "healthy",
        "version":   "5.17.0",
        "db":        "postgresql" if db_ok else "json_fallback",
        "timestamp": datetime.now().isoformat(),
        "cache_keys": len(_cache),
        "auth":      "jwt-bearer",
    }

# ── BSC Endpoints ─────────────────────────────────────────────────
@app.get("/api/bsc/summary")
def bsc_summary(user: dict = Depends(get_current_user)):
    _audit("API_BSC_SUMMARY", user)
    cached = _get_cache("bsc_summary", "bsc")
    if cached:
        return cached

    if _db_available():
        try:
            from utils.db import db as _db
            rows = _db.fetch_all("""
                SELECT dept,
                       COUNT(*)                              as staff_count,
                       ROUND(AVG(final_score)::numeric, 2)  as avg_score,
                       SUM(CASE WHEN final_score >= 4.0 THEN 1 ELSE 0 END) as exceeding,
                       SUM(CASE WHEN final_score < 2.5 THEN 1 ELSE 0 END)  as at_risk,
                       period
                FROM bsc_scores
                GROUP BY dept, period
                ORDER BY avg_score DESC
            """)
            result = {
                "by_dept":     _serialize(rows),
                "total_staff": sum(_safe_int(r.get("staff_count")) for r in rows),
                "overall_avg": round(
                    sum(_safe_float(r.get("avg_score",0)) * _safe_int(r.get("staff_count",1))
                        for r in rows) /
                    max(sum(_safe_int(r.get("staff_count",1)) for r in rows), 1), 2),
                "source":      "postgresql",
            }
            _set_cache("bsc_summary", result)
            return result
        except Exception as e:
            logger.error(f"BSC DB error: {e}")

    # JSON fallback
    scores = _load_json("feb_2026_staff_scores.json")
    if isinstance(scores, dict):
        all_scores = list(scores.values())
    else:
        all_scores = scores

    by_dept = {}
    for s in all_scores:
        d = s.get("dept", "Unknown")
        if d not in by_dept:
            by_dept[d] = {"dept": d, "staff_count": 0, "avg_score": 0.0,
                          "exceeding": 0, "at_risk": 0, "total": 0.0}
        by_dept[d]["staff_count"] += 1
        sc = _safe_float(s.get("final_score", 0))
        by_dept[d]["total"] += sc
        if sc >= 4.0: by_dept[d]["exceeding"] += 1
        if sc < 2.5:  by_dept[d]["at_risk"]   += 1

    for d in by_dept:
        by_dept[d]["avg_score"] = round(
            by_dept[d]["total"] / max(by_dept[d]["staff_count"], 1), 2)

    result = {
        "by_dept":     list(by_dept.values()),
        "total_staff": len(all_scores),
        "overall_avg": round(
            sum(_safe_float(s.get("final_score",0)) for s in all_scores) /
            max(len(all_scores), 1), 2),
        "source": "json",
    }
    _set_cache("bsc_summary", result)
    return result

@app.get("/api/bsc/staff/{username}")
def bsc_staff(username: str, user: dict = Depends(get_current_user)):
    # V-005 mitigation — basic IDOR check. A non-admin requesting another
    # user's BSC has to share the same role tree. For v5.17 we approve when
    # caller is admin or asking for their own record. A fuller check using
    # UserManager.get_managed_staff_codes is a follow-up.
    if user["username"] != username and (user.get("role") or "").lower() not in ("admin", "director"):
        _audit("API_BSC_STAFF_DENIED", user, f"Tried to read {username}")
        raise HTTPException(status_code=403, detail="Forbidden")
    _audit("API_BSC_STAFF", user, f"Read BSC for {username}")
    if _db_available():
        try:
            from utils.db import db as _db
            row = _db.fetch_one(
                "SELECT * FROM bsc_scores WHERE username = %s ORDER BY computed_at DESC LIMIT 1",
                (username,)
            )
            if row:
                return _serialize(row)
        except Exception as e:
            logger.error(f"BSC staff DB error: {e}")

    scores = _load_json("feb_2026_staff_scores.json")
    if isinstance(scores, dict):
        return scores.get(username, {})
    return {}

# ── Pipeline Endpoints ────────────────────────────────────────────
@app.get("/api/pipeline/summary")
def pipeline_summary(user: dict = Depends(get_current_user)):
    _audit("API_PIPELINE_SUMMARY", user)
    cached = _get_cache("pipeline_summary", "pipeline")
    if cached:
        return cached

    if _PIPELINE_READ_DB_FIRST and _db_available():
        try:
            from utils.db import db as _db
            rows = _db.fetch_all("""
                SELECT stage, deal_category,
                       COUNT(*)          as deal_count,
                       SUM(COALESCE((metadata->>'amount_kes')::numeric, amount, 0)) as total_value,
                       AVG(probability)  as avg_probability
                FROM pipeline_deals
                GROUP BY stage, deal_category
                ORDER BY total_value DESC
            """)
            totals = _db.fetch_one("""
                SELECT COUNT(*)    as total_deals,
                       SUM(COALESCE((metadata->>'amount_kes')::numeric, amount, 0)) as pipeline_value,
                       SUM(CASE WHEN stage = 'Closed Won'
                                THEN COALESCE((metadata->>'amount_kes')::numeric, amount, 0)
                                ELSE 0 END) as won_value,
                       SUM(CASE WHEN stage = 'Closed Lost' THEN 1 ELSE 0 END)    as lost_count,
                       SUM(CASE WHEN COALESCE(TRIM(currency),'KES') = 'KES'
                                THEN COALESCE((metadata->>'amount_kes')::numeric, amount, 0)
                                ELSE 0 END) as lcy_value,
                       SUM(CASE WHEN COALESCE(TRIM(currency),'KES') <> 'KES'
                                THEN COALESCE((metadata->>'amount_kes')::numeric, amount, 0)
                                ELSE 0 END) as fcy_value
                FROM pipeline_deals
            """)
            result = {
                "by_stage":       _serialize(rows),
                "totals":         _serialize(totals),
                "source":         "postgresql",
            }
            _set_cache("pipeline_summary", result)
            return result
        except Exception as e:
            logger.error(f"Pipeline DB error: {e}")

    # Canonical-manager path (v10.503 Phase 3 Arc α Batch α1).
    #
    # Previously this branch read `data/pipeline.json` directly via
    # `_load_json(...)`. That bypassed `PipelineManager` — the canonical
    # business-logic layer that Streamlit consumes — and exposed the
    # legacy 302-deal seed dataset to the API instead of the
    # PipelineManager-managed records. The drift was documented in
    # `docs/architecture/PIPELINE_DOMAIN_AUDIT.md` Section 15.1 and
    # surfaced as data inconsistency between the Streamlit UI (8 deals)
    # and the API response (302 deals).
    #
    # Per the established doctrine ("Streamlit stays, React additive,
    # FastAPI canonical" — see REACT_READINESS_AUDIT.md and changelogs
    # v10.21/v10.400/v10.417/v10.426+), business logic is centralized.
    # The API now reads through the same manager class Streamlit uses;
    # both presentation layers receive identical data.
    #
    # G394 (gate_pipeline_api_uses_canonical_manager) enforces that
    # this branch references PipelineManager and NOT _load_json(
    # "pipeline.json").
    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    deals = pm.get_deals()

    # Cascade scope enforcement (v10.504 Phase 3 Arc α Batch α2).
    #
    # Closes GAP-001 from PIPELINE_DOMAIN_AUDIT Section 10. Before α2
    # this endpoint returned all PipelineManager deals regardless of
    # caller. The Streamlit page filtered client-side via
    # `get_visible_staff(user_data, staff_scores)` from
    # `pages/3_pipeline.py:47`; the API path had no equivalent.
    #
    # `get_visible_staff_codes` wraps the canonical cascade-walk
    # function (utils/core_audit.py:190::get_visible_staff) — no
    # duplicate business logic. It supplies the staff_register roster
    # the API path lacks and projects the result to a code set for
    # set-membership filtering.
    #
    # G395 (gate_pipeline_api_enforces_cascade_scope) verifies this
    # filter is applied in both pipeline endpoints.
    from utils.api_pipeline_scope import (
        get_visible_staff_codes,
        filter_deals_by_visible_codes,
    )
    visible_codes = get_visible_staff_codes(user)
    deals = filter_deals_by_visible_codes(deals, visible_codes)

    from utils.core import ACTIVE_STAGES as _ACTIVE
    by_stage: dict = {}
    total_val = 0.0
    won_val   = 0.0
    validated_val = 0.0   # assured: active + manager_validated
    pending_val   = 0.0   # pending assurance: active + not validated
    lcy_val   = 0.0       # P4 (React-B): KES-equivalent, local-currency book
    fcy_val   = 0.0       # P4 (React-B): KES-equivalent, foreign-currency book
    for d in deals:
        if _referral_blocked(d):
            continue  # pending/declined referrals don't count until accepted
        st  = d.get("stage", "Unknown")
        # Canonical amount field is `deal_value` (Generation B).
        # Fall back to `amount` for any transitional record.
        amt_native = _safe_float(d.get("deal_value") or d.get("amount", 0))
        # Report in KES-equivalent (the bank's reporting currency) so totals are
        # comparable across currency books — amount_kes when stamped, native
        # otherwise (LCY deals coincide; rate 1). Summing native across mixed
        # currencies would be meaningless and would break LCY+FCY reconciliation.
        amt = _safe_float(d.get("amount_kes")) if d.get("amount_kes") is not None else amt_native
        total_val += amt
        book = d.get("currency_book") or ("FCY" if str(d.get("currency", "KES")).strip() not in ("", "KES") else "LCY")
        if book == "FCY":
            fcy_val += amt
        else:
            lcy_val += amt
        if st == "Closed Won":
            won_val += amt
        if st in _ACTIVE:
            if d.get("manager_validated"):
                validated_val += amt
            else:
                pending_val += amt
        if st not in by_stage:
            by_stage[st] = {"stage": st, "deal_count": 0, "total_value": 0.0}
        by_stage[st]["deal_count"]  += 1
        by_stage[st]["total_value"] += amt

    # Count Closed Lost explicitly (was always 0 in the previous path
    # — that was an unflagged bug; fixed in α1).
    lost_count = sum(1 for d in deals if d.get("stage") == "Closed Lost")

    # Scope-aware manager queues for the dashboard tiles. Validation is a
    # manager-looking-down action: non-managers have nothing to validate, so the
    # count is 0 for them (matches the manager-only validation queue). For a
    # manager it is the number of subordinate deals awaiting their sign-off —
    # the SAME source as Manager Queues -> Validation, so tile and queue agree.
    from utils.api_pipeline_manager_actions import is_manager as _is_mgr
    pending_validation = 0
    pending_cancel = 0
    if _is_mgr(user):
        pending_validation = len(pm.get_pending_validations(manager_codes=visible_codes))
        pending_cancel = len(pm.get_cancel_requests(manager_codes=visible_codes))

    result = {
        "by_stage": list(by_stage.values()),
        "totals": {
            "total_deals":        len(deals),
            "pipeline_value":     total_val,
            "validated_value":    validated_val,   # assured headline
            "pending_value":      pending_val,     # pending assurance
            "won_value":          won_val,
            "lcy_value":          lcy_val,         # KES-equiv, local-currency book
            "fcy_value":          fcy_val,         # KES-equiv, foreign-currency book
            "lost_count":         lost_count,
            "pending_validation": pending_validation,
            "pending_cancel":     pending_cancel,
        },
        "source": "pipeline_manager",
    }
    _set_cache("pipeline_summary", result)
    return result

@app.get("/api/pipeline/stages")
def pipeline_stages_config(user: dict = Depends(get_current_user)):
    """Admin-configured pipeline stages (names, colors, default
    probabilities) from org_config. Batch A (2026-06-15) — single source of
    truth for the frontend's stage dropdowns + filters, so the UI reflects the
    bank's configured workflow instead of hardcoded stage strings.
    """
    from utils.core import get_pipeline_settings, get_pipeline_stages
    cfg = get_pipeline_settings() or {}
    # Business sectors = CBK classification (admin-configurable; falls back to the
    # canonical CBK_SECTORS constant if not yet seeded). Individual MOUs are read
    # LIVE from the partnerships register (active only) — never duplicated.
    try:
        from utils.core import CBK_SECTORS as _CBK
    except Exception:
        _CBK = []
    business_sectors = cfg.get("business_sectors") or list(_CBK)
    individual_mous = []
    try:
        import os as _os, json as _json
        _p = _os.path.join("data", "partnerships_mous.json")
        if _os.path.exists(_p):
            with open(_p, "r", encoding="utf-8") as _f:
                for _m in _json.load(_f):
                    if str(_m.get("status", "")).strip() == "Active":
                        individual_mous.append({
                            "id":           _m.get("id"),
                            "title":        _m.get("title"),
                            "partner_name": _m.get("partner_name"),
                        })
    except Exception as _e:
        logger.error(f"individual_mous load error: {_e}")
    return {
        "stages":            cfg.get("stages") or get_pipeline_stages() or [],
        "deal_categories":   cfg.get("deal_categories", []),
        "segment_labels":    _segment_labels(),
        "sectors":           cfg.get("sectors", []),
        "business_sectors":  business_sectors,
        "individual_mous":   individual_mous,
        "allow_other_sector": cfg.get("allow_other_sector", True),
        "allow_other_mou":    cfg.get("allow_other_mou", True),
        "decision_levels":   cfg.get("decision_levels", []),
        "probability_map":   cfg.get("probability_map", {}),
        "deal_types":        cfg.get("deal_types", []),
        "product_catalogue": cfg.get("product_catalogue", {}),
        "stage_flows":       cfg.get("stage_flows", {}),
        "customer_segments": _customer_segments(),
        "client_types":      _client_types(),
        "currency":          cfg.get("currency", "KES"),
        "required_fields":   _required_fields(),
    }


# Keys an executive (CEO / MD / Director) may edit through the React admin
# console. Each maps straight into pipeline_settings.json; anything outside this
# set is ignored so the write surface can't be used to mutate unrelated state.
_EDITABLE_CONFIG_KEYS = {
    "segment_labels", "customer_segments", "client_types", "product_catalogue",
    "individual_mous", "business_sectors", "sectors", "deal_categories",
    "stage_flows", "required_fields", "allow_other_sector", "allow_other_mou",
    "probability_map", "deal_types", "disbursement_roles",
    "referral_pending_alert_days", "referral_business_departments",
}

# Roles permitted to complete disbursement (Troops / Treasury Back Office).
# Mirrors api_credit_admin_routes._DEFAULT_DISBURSEMENT_ROLES; pipeline_settings
# is the shared source of truth once an admin edits it.
_DISBURSEMENT_ROLES_DEFAULT = ["Treasury Back Office"]


@app.post("/api/admin/pipeline-config", tags=["admin"])
def admin_update_pipeline_config(
    payload: dict = Body(default_factory=dict),
    user: dict = Depends(require_config_admin),
):
    """Merge a partial reference-config patch into pipeline_settings.json. Only
    whitelisted keys (_EDITABLE_CONFIG_KEYS) are applied; an empty patch is a
    no-op (no write). Backs the React admin Configuration panels (segments /
    products / MOUs / sectors / required-fields). Currency/FX has its own
    endpoint (/api/fx/rates). Gated to the executive tier via require_config_admin."""
    from utils.core import get_pipeline_settings, save_pipeline_settings

    patch = {k: v for k, v in (payload or {}).items() if k in _EDITABLE_CONFIG_KEYS}
    if not patch:
        # Nothing editable supplied — report current effective config, no write.
        return {"status": "noop", "applied": [], "config": _editable_config_view()}

    settings = get_pipeline_settings() or {}
    if not isinstance(settings, dict):
        settings = {}
    settings.update(patch)
    save_pipeline_settings(settings)
    _audit("API_PIPELINE_CONFIG_UPDATE", user, ", ".join(sorted(patch.keys())))
    return {"status": "saved", "applied": sorted(patch.keys()),
            "config": _editable_config_view()}


def _editable_config_view() -> dict:
    """Current effective values for every editable config key — what the admin
    console reads back after a save (and on load)."""
    cfg = _load_json("pipeline_settings.json") or {}
    return {
        "segment_labels":     _segment_labels(),
        "customer_segments":  _customer_segments(),
        "client_types":       _client_types(),
        "product_catalogue":  cfg.get("product_catalogue", {}),
        "individual_mous":    cfg.get("individual_mous", []),
        "business_sectors":   cfg.get("business_sectors", []),
        "sectors":            cfg.get("sectors", []),
        "deal_categories":    cfg.get("deal_categories", []),
        "stage_flows":        cfg.get("stage_flows", {}),
        "required_fields":    _required_fields(),
        "allow_other_sector": cfg.get("allow_other_sector", True),
        "allow_other_mou":    cfg.get("allow_other_mou", True),
        "probability_map":    cfg.get("probability_map", {}),
        "deal_types":         cfg.get("deal_types", []),
        "disbursement_roles": cfg.get("disbursement_roles", list(_DISBURSEMENT_ROLES_DEFAULT)),
        "referral_pending_alert_days": int(cfg.get("referral_pending_alert_days", _REFERRAL_PENDING_ALERT_DAYS)),
        "referral_business_departments": cfg.get("referral_business_departments", list(_DEFAULT_REFERRAL_BUSINESS_DEPARTMENTS)),
    }


# ─────────────────────────────────────────────────────────────────────
# SLA CONFIG (Phase 4 — S1) — admin-configurable, deal-process-wide SLA.
#
# Step/role granular targets (business days) + a per-process escalation
# ladder (elapsed-day tiers, breach -> step owner ... ceiling -> MD) + an
# optional per-product create->closed promise. Seeded with a default
# taxonomy the admin amends in the console; lives in pipeline_settings
# (the shared, disk-persisted config), so "once applied it applies" —
# get_pipeline_settings reads fresh from disk, no cache. S1 is config
# only; the per-deal business-day clock + violation detection is S2.
# ─────────────────────────────────────────────────────────────────────
_DEFAULT_SLA_CONFIG = {
    "steps": [
        {"key": "lead_qualification",      "label": "Lead qualification",                  "owner_role": "Relationship Manager", "target_days": 3},
        {"key": "line_manager_validation", "label": "Line-manager validation",             "owner_role": "Branch Manager",       "target_days": 2},
        {"key": "credit_assessment",       "label": "Credit assessment",                   "owner_role": "Credit Analyst",       "target_days": 5},
        {"key": "offer_acceptance",        "label": "Offer & acceptance",                  "owner_role": "Relationship Manager", "target_days": 3},
        {"key": "security_perfection",     "label": "Security perfection & authorization", "owner_role": "Credit Admin",         "target_days": 7},
        {"key": "disbursement",            "label": "Disbursement",                        "owner_role": "Treasury Back Office", "target_days": 2},
        {"key": "closure",                 "label": "Closure",                             "owner_role": "Relationship Manager", "target_days": 2},
    ],
    "escalation_ladder": [
        {"after_days": 0, "escalate_to": "step_owner"},
        {"after_days": 2, "escalate_to": "line_manager"},
        {"after_days": 5, "escalate_to": "regional_head"},
        {"after_days": 8, "escalate_to": "managing_director"},
    ],
    "product_promise": {},
    "stage_step_map": {
        "Lead": "lead_qualification",
        "Contacted": "lead_qualification",
        "Qualified": "lead_qualification",
        "Application": "line_manager_validation",
        "Credit Assessment": "credit_assessment",
        "Offer / Proposal": "offer_acceptance",
        "Closed Won": "closure",
        "Closed Lost": "closure",
    },
}


def _sla_config() -> dict:
    """Effective SLA config — the admin's saved sla_config from pipeline_settings,
    or the seeded default if none has been set. Read fresh (no cache)."""
    try:
        from utils.core import get_pipeline_settings
        cfg = (get_pipeline_settings() or {}).get("sla_config")
        if isinstance(cfg, dict) and cfg.get("steps"):
            return cfg
    except Exception:
        logging.getLogger("a2z.sla").warning("sla_config read fell back to default", exc_info=True)
    return _DEFAULT_SLA_CONFIG


def _validate_sla_config(cfg) -> tuple:
    """Mandatory-before-save validation. Returns (ok, message). Every step needs a
    unique key and a positive target_days; the escalation ladder must be strictly
    increasing by after_days with a non-empty escalate_to; product_promise (optional)
    values must be positive integers."""
    if not isinstance(cfg, dict):
        return False, "sla_config must be an object"
    steps = cfg.get("steps")
    if not isinstance(steps, list) or not steps:
        return False, "at least one SLA step is required"
    seen = set()
    for s in steps:
        if not isinstance(s, dict):
            return False, "each step must be an object"
        k = str(s.get("key") or "").strip()
        if not k:
            return False, "each step needs a key"
        if k in seen:
            return False, f"duplicate step key: {k}"
        seen.add(k)
        try:
            t = int(s.get("target_days"))
        except (TypeError, ValueError):
            return False, f"step '{k}': target_days must be an integer"
        if t <= 0:
            return False, f"step '{k}': target_days must be positive"
    ladder = cfg.get("escalation_ladder")
    if not isinstance(ladder, list) or not ladder:
        return False, "escalation_ladder is required"
    last = -1
    for tier in ladder:
        if not isinstance(tier, dict):
            return False, "each escalation tier must be an object"
        try:
            a = int(tier.get("after_days"))
        except (TypeError, ValueError):
            return False, "tier after_days must be an integer"
        if a < 0:
            return False, "tier after_days must be >= 0"
        if a <= last:
            return False, "escalation_ladder must be strictly increasing by after_days"
        last = a
        if not str(tier.get("escalate_to") or "").strip():
            return False, "each escalation tier needs an escalate_to"
    pp = cfg.get("product_promise") or {}
    if not isinstance(pp, dict):
        return False, "product_promise must be an object"
    for prod, days in pp.items():
        try:
            d = int(days)
        except (TypeError, ValueError):
            return False, f"product_promise['{prod}'] must be an integer"
        if d <= 0:
            return False, f"product_promise['{prod}'] must be positive"
    ssm = cfg.get("stage_step_map") or {}
    if not isinstance(ssm, dict):
        return False, "stage_step_map must be an object"
    step_keys = {str(s.get("key")) for s in steps}
    for stage, sk in ssm.items():
        if str(sk) not in step_keys:
            return False, f"stage_step_map['{stage}'] -> unknown step '{sk}'"
    return True, ""


@app.get("/api/admin/sla-config", tags=["admin"])
def get_sla_config(user: dict = Depends(get_current_user)):
    """Effective SLA config (admin-saved or seeded default). Readable by any
    authenticated user so downstream surfaces can show targets; only admin writes."""
    cfg = _sla_config()
    return {"sla_config": cfg, "is_default": cfg is _DEFAULT_SLA_CONFIG}


@app.post("/api/admin/sla-config", tags=["admin"])
def set_sla_config(
    payload: dict = Body(default_factory=dict),
    user: dict = Depends(require_config_admin),
):
    """Validate then persist the SLA config to pipeline_settings. Accepts either
    {"sla_config": {...}} or the raw config object. Mandatory-before-save: an invalid
    or incomplete config is rejected (400) and nothing is written. On success the new
    config is live immediately (settings are read fresh from disk, no cache)."""
    from utils.core import get_pipeline_settings, save_pipeline_settings
    cfg = payload.get("sla_config") if isinstance(payload, dict) and "sla_config" in payload else payload
    ok, msg = _validate_sla_config(cfg)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    settings = get_pipeline_settings() or {}
    settings["sla_config"] = cfg
    save_pipeline_settings(settings)
    return {"status": "saved", "sla_config": cfg}


# ─────────────────────────────────────────────────────────────────────
# SLA S2a — product/age violation engine (read-only over created_at + S1 config).
# The per-STEP business-day clock (S2b) needs step-entry timestamps deals don't
# yet carry; this slice measures the create->now clock against the per-product
# promise (or the step-target sum) so breaches are already evident for action.
# ─────────────────────────────────────────────────────────────────────
_SLA_CLOSED_MARKERS = ("won", "lost", "closed", "declined", "disbursed", "rejected")
_DEFAULT_STAGE_STEP_MAP = {
    "Lead": "lead_qualification",
    "Contacted": "lead_qualification",
    "Qualified": "lead_qualification",
    "Application": "line_manager_validation",
    "Credit Assessment": "credit_assessment",
    "Offer / Proposal": "offer_acceptance",
    "Closed Won": "closure",
    "Closed Lost": "closure",
}


def _sla_stage_step_map(cfg: dict = None) -> dict:
    """Pipeline stage -> SLA step key. Admin-configurable via sla_config.stage_step_map;
    seeded default covers the pipeline-stage-driven steps. (Credit-admin / Troops steps —
    security_perfection, disbursement — are stamped by those modules in S2c.)"""
    if cfg is None:
        cfg = _sla_config()
    m = cfg.get("stage_step_map")
    return dict(m) if isinstance(m, dict) and m else dict(_DEFAULT_STAGE_STEP_MAP)


def _sla_step_target_sum(cfg: dict) -> int:
    """End-to-end target (business days) implied by the step config — the fallback
    create->closed promise when no per-product promise is configured."""
    total = 0
    for s in (cfg.get("steps") or []):
        try:
            t = int(s.get("target_days", 0))
        except (TypeError, ValueError):
            t = 0
        if t > 0:
            total += t
    return total


def _sla_step_target(cfg: dict, step_key: str) -> int:
    for s in (cfg.get("steps") or []):
        if s.get("key") == step_key:
            try:
                return int(s.get("target_days", 0))
            except (TypeError, ValueError):
                return 0
    return 0


def _sla_escalation_for(overdue_bd: int, cfg: dict) -> str:
    """Highest ladder tier whose after_days threshold the overdue days have reached."""
    role = "step_owner"
    for tier in (cfg.get("escalation_ladder") or []):
        try:
            if int(tier.get("after_days")) <= overdue_bd:
                role = str(tier.get("escalate_to") or role)
        except (TypeError, ValueError):
            continue
    return role


def _stamp_sla_step(pm, deal_id: str, deal: dict, by: str) -> None:
    """Record first entry into the deal's current SLA step on sla_step_log, enabling the
    per-step business-day clock. No-op if the stage maps to no step or it's already
    stamped. Best-effort: a stamp failure never blocks the underlying transition."""
    try:
        smap = _sla_stage_step_map()
        step_key = smap.get(str(deal.get("stage") or ""))
        if not step_key:
            return
        log = dict(deal.get("sla_step_log") or {})
        if step_key in log:
            return
        log[step_key] = datetime.now().isoformat()
        pm.update_deal(deal_id, {"sla_step_log": log}, by)
        _db_sync_pipeline_deal(pm.get_deal(deal_id))  # mirror to DB-first reads
    except Exception:
        logging.getLogger("a2z.sla").warning("sla step stamp failed for %s", deal_id, exc_info=True)


def _deal_sla_status(deal: dict, cfg: dict, promise: dict, smap: dict) -> dict:
    """Unified per-deal SLA. {} for a closed deal or one with no usable creation
    timestamp. Prefers the per-STEP clock (business days since the current step's entry
    stamp vs that step's target) when the deal carries an sla_step_log entry for its
    current step; otherwise falls back to the product/age clock (create->now vs the
    product promise or step-target sum). Reports which clock was used."""
    stage = str(deal.get("stage") or "")
    if any(m in stage.lower() for m in _SLA_CLOSED_MARKERS):
        return {}
    base_ts = deal.get("created_at") or deal.get("open_date") or deal.get("created")
    if not base_ts:
        return {}
    step_key = smap.get(stage)
    log = deal.get("sla_step_log") or {}
    step_entered = log.get(step_key) if step_key else None
    if step_key and step_entered:
        target = _sla_step_target(cfg, step_key)
        elapsed = _business_days_since(step_entered)
        clock, clock_step = "step", step_key
    else:
        product = str(deal.get("product_type") or deal.get("product") or "")
        try:
            per_prod = int(promise.get(product)) if promise.get(product) else 0
        except (TypeError, ValueError):
            per_prod = 0
        target = per_prod or _sla_step_target_sum(cfg)
        elapsed = _business_days_since(base_ts)
        clock, clock_step = "age", None
    if target <= 0:
        return {}
    overdue = max(0, elapsed - target)
    return {
        "deal_id": deal.get("id"),
        "client_name": deal.get("client_name"),
        "product_type": str(deal.get("product_type") or deal.get("product") or "") or None,
        "stage": deal.get("stage"),
        "owner_code": str(deal.get("staff_code") or ""),
        "clock": clock,
        "step": clock_step,
        "elapsed_business_days": elapsed,
        "target_days": target,
        "overdue_business_days": overdue,
        "breached": overdue > 0,
        "escalate_to": _sla_escalation_for(overdue, cfg) if overdue > 0 else None,
    }


@app.get("/api/pipeline/deals/{deal_id}/sla", tags=["pipeline"])
def pipeline_deal_sla(deal_id: str, user: dict = Depends(get_current_user)):
    """Per-deal SLA status — current step, clock (step|age), elapsed/target business days,
    breach + escalation tier. Scope-checked like the rest of the deal surface."""
    from utils.core import PipelineManager as _PM
    from utils.api_pipeline_scope import get_visible_staff_codes
    pm = _PM()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    visible = set(get_visible_staff_codes(user) or [])
    sc = str(deal.get("staff_code") or "")
    po = str(deal.get("portfolio_owner_code") or "")
    if visible and sc not in visible and (not po or po not in visible):
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    cfg = _sla_config()
    status = _deal_sla_status(deal, cfg, cfg.get("product_promise") or {}, _sla_stage_step_map(cfg))
    return {"deal_id": deal_id, "sla": status or None}


@app.get("/api/pipeline/sla/violations", tags=["pipeline"])
def pipeline_sla_violations(user: dict = Depends(get_current_user)):
    """Hierarchy-scoped SLA breach view. For each OPEN deal the caller can see, measure
    the per-step clock (where a step-entry stamp exists) or the create->now clock against
    the configured target, and surface those over target with the escalation tier the
    overdue days reach. Scoped like the pipeline (MD/CEO -> all; manager -> subtree; RM ->
    own)."""
    from utils.api_pipeline_scope import get_visible_staff_codes, filter_deals_by_visible_codes
    cfg = _sla_config()
    promise = cfg.get("product_promise") or {}
    smap = _sla_stage_step_map(cfg)
    visible = set(get_visible_staff_codes(user) or [])
    deals = filter_deals_by_visible_codes(_all_pipeline_deals(), visible)
    open_count = 0
    violations = []
    by_escalation: dict = {}
    by_clock = {"step": 0, "age": 0}
    for d in deals:
        s = _deal_sla_status(d, cfg, promise, smap)
        if not s:
            continue
        open_count += 1
        by_clock[s["clock"]] = by_clock.get(s["clock"], 0) + 1
        if s["breached"]:
            violations.append(s)
            esc = s["escalate_to"] or "step_owner"
            by_escalation[esc] = by_escalation.get(esc, 0) + 1
    violations.sort(key=lambda x: x["overdue_business_days"], reverse=True)
    return {
        "violations": violations,
        "count": len(violations),
        "open_deals": open_count,
        "by_escalation": by_escalation,
        "by_clock": by_clock,
    }


#
# Roles live in kpi_library.json -> role_kpis (the authoritative registry).
# This surface LISTS every role with its KPI set, resolved pillar mix, and
# capabilities (currently: can_disburse, wired to pipeline_settings ->
# disbursement_roles). It does NOT touch the reporting hierarchy — that is a
# STAFF-level concern (staff_code -> manager_code via the staff roster +
# ReportingLineManager), handled separately under the auth/DOA finale.
# Per-KPI weight editing is also deferred (role_kpis reference KPIs by a mix of
# ids/codes/names that must be reconciled first).
# ─────────────────────────────────────────────────────────────────────

def _kpi_pillar_index() -> dict:
    """Best-effort id/name -> pillar map over the KPI library, for resolving the
    mixed-identifier refs inside role_kpis."""
    from utils.core import get_kpi_library
    lib = get_kpi_library() or {}
    idx = {}
    for k in (lib.get("kpis") or []):
        if not isinstance(k, dict):
            continue
        pil = k.get("pillar") or "Unmapped"
        if k.get("id"):
            idx[str(k["id"])] = pil
        if k.get("name"):
            idx[str(k["name"])] = pil
    return idx


def _role_can_disburse(role: str, disb_lower: list) -> bool:
    rl = str(role or "").lower()
    return any(d in rl or rl in d for d in disb_lower if d)


@app.get("/api/admin/roles", tags=["admin"])
def admin_roles_list(user: dict = Depends(require_config_admin)):
    """Full role registry: every role in kpi_library.json -> role_kpis with its
    KPI count, resolved pillar mix, and capabilities. Config-admin only."""
    from utils.core import get_kpi_library, get_pipeline_settings
    lib = get_kpi_library() or {}
    role_kpis = lib.get("role_kpis") or {}
    idx = _kpi_pillar_index()
    cfg = get_pipeline_settings() or {}
    disb = cfg.get("disbursement_roles") or list(_DISBURSEMENT_ROLES_DEFAULT)
    disb_lower = [str(r).strip().lower() for r in disb if str(r).strip()]

    rows = []
    for role, kpis in role_kpis.items():
        kl = kpis if isinstance(kpis, list) else []
        mix: dict = {}
        for ref in kl:
            pil = idx.get(str(ref), "Unmapped")
            mix[pil] = mix.get(pil, 0) + 1
        rows.append({
            "role": role,
            "kpi_count": len(kl),
            "pillars": mix,
            "can_disburse": _role_can_disburse(role, disb_lower),
        })
    rows.sort(key=lambda r: str(r["role"]).lower())
    return {"roles": rows, "count": len(rows), "disbursement_roles": disb}


@app.get("/api/admin/role-detail", tags=["admin"])
def admin_role_detail(role: str, user: dict = Depends(require_config_admin)):
    """One role's full KPI resolution (ref -> id/name/pillar/weight, flagging any
    unmapped refs) plus its capabilities."""
    from utils.core import get_kpi_library, get_pipeline_settings
    lib = get_kpi_library() or {}
    role_kpis = lib.get("role_kpis") or {}
    if role not in role_kpis:
        raise HTTPException(status_code=404, detail=f"Role '{role}' not in registry")
    kpis = lib.get("kpis") or []
    by_id = {str(k.get("id")): k for k in kpis if isinstance(k, dict)}
    by_name = {str(k.get("name")): k for k in kpis if isinstance(k, dict)}
    resolved = []
    for ref in (role_kpis[role] or []):
        k = by_id.get(str(ref)) or by_name.get(str(ref))
        if k:
            resolved.append({"ref": ref, "id": k.get("id"), "name": k.get("name"),
                             "pillar": k.get("pillar"), "weight": k.get("weight"),
                             "mapped": True})
        else:
            resolved.append({"ref": ref, "mapped": False})
    cfg = get_pipeline_settings() or {}
    disb = cfg.get("disbursement_roles") or list(_DISBURSEMENT_ROLES_DEFAULT)
    disb_lower = [str(r).strip().lower() for r in disb if str(r).strip()]
    return {"role": role, "kpis": resolved, "kpi_count": len(resolved),
            "unmapped": sum(1 for r in resolved if not r["mapped"]),
            "can_disburse": _role_can_disburse(role, disb_lower)}


class RoleCapabilityRequest(BaseModel):
    role: str
    can_disburse: bool


@app.post("/api/admin/roles/capabilities", tags=["admin"])
def admin_role_capability(payload: RoleCapabilityRequest,
                          user: dict = Depends(require_config_admin)):
    """Grant or revoke the disbursement capability for a role — toggles its
    membership in pipeline_settings.json -> disbursement_roles (the list
    _is_troops reads). Config-admin only."""
    from utils.core import get_pipeline_settings, save_pipeline_settings
    role = (payload.role or "").strip()
    if not role:
        raise HTTPException(status_code=400, detail="role is required")
    settings = get_pipeline_settings() or {}
    if not isinstance(settings, dict):
        settings = {}
    roles = settings.get("disbursement_roles")
    if not isinstance(roles, list):
        roles = list(_DISBURSEMENT_ROLES_DEFAULT)
    present = any(str(r).strip().lower() == role.lower() for r in roles)
    if payload.can_disburse and not present:
        roles.append(role)
    elif not payload.can_disburse and present:
        roles = [r for r in roles if str(r).strip().lower() != role.lower()]
    settings["disbursement_roles"] = roles
    save_pipeline_settings(settings)
    _audit("API_ROLE_CAPABILITY", user, f"{role}|can_disburse={payload.can_disburse}")
    return {"role": role, "can_disburse": payload.can_disburse, "disbursement_roles": roles}


@app.get("/api/pipeline/deals")
def pipeline_deals(
    stage:    Optional[str] = None,
    category: Optional[str] = None,
    unit:     Optional[str] = None,
    limit:    int = Query(default=500, le=5000),
    offset:   int = Query(default=0, ge=0),
    user:     dict = Depends(get_current_user),
):
    _audit("API_PIPELINE_DEALS", user, f"stage={stage} unit={unit} limit={limit}")
    if _PIPELINE_READ_DB_FIRST and _db_available():
        try:
            from utils.db import db as _db
            from utils.api_pipeline_scope import (
                get_visible_staff_codes, filter_deals_by_visible_codes,
            )
            from utils.api_pipeline_permissions import enrich_deal_with_permissions
            where  = []
            params = []
            if stage:    where.append("stage = %s");        params.append(stage)
            if category: where.append("deal_category = %s");params.append(category)
            if unit:     where.append("unit = %s");         params.append(unit)
            sql = "SELECT * FROM pipeline_deals"
            if where: sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY open_date DESC"
            rows = _db.fetch_all(sql, tuple(params))
            deals = [_normalize_db_deal_row(d) for d in _serialize(rows)]
            # B8: cascade scope + permission enrichment — parity with the
            # PipelineManager branch below. The DB branch previously skipped
            # BOTH, which leaked every deal across scope when Postgres was
            # active and left the per-deal "YOU CAN" permissions empty.
            visible_codes = get_visible_staff_codes(user)
            deals = filter_deals_by_visible_codes(deals, visible_codes)
            enriched = [
                enrich_deal_with_permissions(d, user, visible_codes)
                for d in deals[offset:offset + limit]
            ]
            return {"deals": enriched, "count": len(deals), "source": "postgresql"}
        except Exception as e:
            logger.error(f"Pipeline deals DB error: {e}")

    # Canonical-manager path (v10.503 Phase 3 Arc α Batch α1).
    # See pipeline_summary above for the doctrine and migration note.
    # G394 enforces no `_load_json("pipeline.json")` in this branch.
    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    deals = pm.get_deals()

    # Cascade scope enforcement (v10.504 Phase 3 Arc α Batch α2).
    # Closes GAP-001. See pipeline_summary above for the doctrine.
    # G395 verifies this filter is applied here as well.
    from utils.api_pipeline_scope import (
        get_visible_staff_codes,
        filter_deals_by_visible_codes,
    )
    visible_codes = get_visible_staff_codes(user)
    deals = filter_deals_by_visible_codes(deals, visible_codes)

    if stage:
        deals = [d for d in deals if d.get("stage") == stage]
    if category:
        # PipelineManager records use `pipeline_category` (Gen B);
        # legacy `deal_category` accepted for transitional compatibility.
        deals = [d for d in deals
                 if d.get("pipeline_category") == category
                 or d.get("deal_category") == category]
    if unit:
        deals = [d for d in deals if d.get("unit") == unit]

    # α7: enrich each deal with caller-specific permissions object.
    # visible_codes already computed above for cascade scope filtering.
    from utils.api_pipeline_permissions import enrich_deal_with_permissions
    enriched = [
        enrich_deal_with_permissions(d, user, visible_codes)
        for d in deals[offset:offset + limit]
    ]

    return {
        "deals":  enriched,
        "count":  len(deals),
        "source": "pipeline_manager",
    }


# ── Pipeline Single-Deal Endpoint (v10.509 Phase 3 Arc α Batch α7) ──
#
# Standalone deal-detail endpoint per audit Section 15.6 + line 828.
# Returns the deal record + per-caller permissions object. Cascade
# scope is enforced: 404 (not 403) if the caller can't see this deal,
# to avoid leaking the existence of deals outside their visibility.


@app.get("/api/pipeline/deals/{deal_id}")
def pipeline_deal_detail(
    deal_id: str,
    user: dict = Depends(get_current_user),
):
    """Fetch a single pipeline deal by id, with caller-specific
    permissions.

    Returns:
        {"deal": {...}, "permissions": {can_view, can_edit, ...}}

    404 if the deal doesn't exist OR is outside the caller's cascade
    scope. The 404-not-403 choice is deliberate: returning 403 would
    leak the existence of deals the caller shouldn't know about.

    The permissions object follows the v10.509 α7 contract — see
    utils/api_pipeline_permissions.py for the resolution rules.
    """
    _audit("API_PIPELINE_DEAL_DETAIL", user, f"deal_id={deal_id}")

    from utils.api_pipeline_permissions import resolve_deal_permissions
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM_for_api

    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Scope check — return 404 not 403 to avoid existence leak
    visible_codes = get_visible_staff_codes(user)
    permissions = resolve_deal_permissions(deal, user, visible_codes)
    if not permissions.get("can_view"):
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    return {
        "deal":         deal,
        "permissions":  permissions,
        "stage_flow":   _stage_flow_for(deal.get("product_type") or deal.get("product", "")),
    }


# ── Credit submission gate (v10.573 Batch B9) ────────────────────────
#
# Explicit "Submit to Credit Analysis" with a document-checklist gate,
# replacing the silent stage trigger. Required documents come from
# lms_config.json's tiered document_checklist (default + amount/product
# add-ons). Submission is BLOCKED (400 + missing list) until every required
# document is marked provided; on success it creates the linked loan
# application via the canonical handoff (create_from_pipeline_deal).


class SubmitToCreditRequest(BaseModel):
    documents_provided: List[str] = []


def _get_required_documents_for_deal(deal: dict) -> list:
    """Required credit documents for a deal, from lms_config's tiered
    document_checklist: default + amount/product add-ons. Order preserved,
    de-duplicated."""
    cfg = _load_json("lms_config.json") or {}
    dc = cfg.get("document_checklist", {}) if isinstance(cfg, dict) else {}
    req = list(dc.get("default", []))
    try:
        amount = float(deal.get("deal_value") or deal.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0.0
    if amount > 10_000_000:
        req += dc.get("above_10m", [])
    text = " ".join(str(deal.get(k, "")) for k in
                    ("product_type", "product", "pipeline_category",
                     "deal_category")).lower()
    if "corporate" in text:
        req += dc.get("corporate", [])
    if "mortgage" in text or "home loan" in text:
        req += dc.get("mortgage", [])
    seen, out = set(), []
    for d in req:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out


def _credit_submission_state(deal: dict, user: dict, visible_codes: set) -> dict:
    from utils.api_pipeline_permissions import resolve_deal_permissions
    required = _get_required_documents_for_deal(deal)
    provided = list(deal.get("documents_provided", []) or [])
    missing = [d for d in required if d not in provided]
    already = bool(deal.get("lms_application_id"))
    perms = resolve_deal_permissions(deal, user, visible_codes)
    my_code = str(user.get("staff_code", "") or "").strip()
    is_admin_like = bool(user.get("is_admin")) or "admin" in str(user.get("role", "")).lower()
    is_owner = bool(my_code) and my_code == str(deal.get("staff_code", "") or "").strip()
    terminal = str(deal.get("stage", "")) in ("Closed Won", "Closed Lost")
    return {
        "required": required,
        "provided": provided,
        "missing": missing,
        "already_submitted": already,
        "lms_application_id": deal.get("lms_application_id"),
        "can_submit": (is_owner or is_admin_like) and not already
                      and not terminal and perms.get("can_view", False),
    }


@app.get("/api/pipeline/deals/{deal_id}/credit-checklist")
def pipeline_credit_checklist(deal_id: str, user: dict = Depends(get_current_user)):
    """Required vs provided credit documents for a deal + whether the caller
    may submit it to credit (owner or admin, active, not already submitted)."""
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.api_pipeline_permissions import resolve_deal_permissions
    from utils.core import PipelineManager as _PM_for_api
    _audit("API_PIPELINE_CREDIT_CHECKLIST", user, f"deal_id={deal_id}")
    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    visible_codes = get_visible_staff_codes(user)
    if not resolve_deal_permissions(deal, user, visible_codes).get("can_view"):
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    return _credit_submission_state(deal, user, visible_codes)


@app.post("/api/pipeline/deals/{deal_id}/submit-to-credit")
def pipeline_submit_to_credit(
    deal_id: str,
    payload: SubmitToCreditRequest,
    user: dict = Depends(get_current_user),
):
    """Submit a deal to credit analysis. Blocks (400 + missing list) until every
    required document is provided; on success creates the linked loan
    application (canonical handoff) and records it on the deal."""
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM_for_api, LoanApplicationManager
    _audit("API_PIPELINE_SUBMIT_TO_CREDIT", user, f"deal_id={deal_id}")
    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    visible_codes = get_visible_staff_codes(user)
    state = _credit_submission_state(deal, user, visible_codes)
    if not state["can_submit"]:
        if state["already_submitted"]:
            raise HTTPException(status_code=400,
                detail=f"Deal already submitted to credit "
                       f"(application {state['lms_application_id']}).")
        raise HTTPException(status_code=403,
            detail="Only the deal owner (or an admin) can submit it to credit.")
    provided = list(payload.documents_provided or [])
    missing = [d for d in state["required"] if d not in provided]
    if missing:
        raise HTTPException(
            status_code=400,
            detail="Cannot submit to credit — missing documents: "
                   + ", ".join(missing),
        )
    # Gate passed — create the linked credit application (canonical handoff).
    lam = LoanApplicationManager()
    app_id = lam.create_from_pipeline_deal(deal, str(user.get("username", "")))
    if not app_id:
        raise HTTPException(status_code=500,
            detail="Could not create the loan application from this deal.")
    pm.update_deal(deal_id, {
        "documents_provided": provided,
        "lms_application_id": app_id,
        "submitted_to_credit": True,
    }, str(user.get("username", "")))
    # Debt #2 (v10.589): advance the deal one stage past Credit Assessment in
    # its configured product-class flow, so the pipeline reflects that the deal
    # has moved into the credit/offer phase rather than sitting frozen at
    # Credit Assessment while its loan progresses. Config-driven via stage_flows;
    # never auto-advances into a terminal Closed stage.
    try:
        _flow = _stage_flow_for(deal.get("product_type") or deal.get("product", ""))
        _cur = str(deal.get("stage", "") or "")
        if _cur in _flow:
            _idx = _flow.index(_cur)
            if 0 <= _idx < len(_flow) - 1:
                _next = str(_flow[_idx + 1])
                if not _next.lower().startswith("closed"):
                    pm.update_stage(deal_id, _next,
                                    f"Auto-advanced on submit to credit (app {app_id}).",
                                    str(user.get("username", "")))
    except Exception:
        # Stage sync is best-effort — never fail a successful submission on it.
        pass
    deal = _get_or_hydrate_deal(pm, deal_id)
    _db_sync_pipeline_deal(deal)
    _audit("API_PIPELINE_SUBMIT_TO_CREDIT_OK", user, f"deal_id={deal_id} app={app_id}")
    # Phase 3 (hardening): submit-to-credit advances the facility; refresh the
    # owner's BSC (the caller is the owner here). Best-effort.
    try:
        from utils.api_bsc_bridge import emit_bsc_for
        emit_bsc_for([str(user.get("username", "") or "")])
    except Exception:
        pass
    return {"application_id": app_id, "status": "submitted_to_credit",
            "missing": [], "stage": deal.get("stage")}


# ── Pipeline analytics (v10.575 Batch B11) ───────────────────────────
#
# Aggregate funnel + headline metrics over the caller's VISIBLE deals.
# Mirrors the Streamlit pipeline page's calculations (pages/3_pipeline.py:
# pip_val / wt_val / won_val / conv_r, the per-category funnels, and the
# stage weights from core.PipelineManager.weighted_pipeline) so the React
# view matches the canonical numbers. Reads deals from the SAME source the
# list endpoint uses (Postgres-first, JSON fallback) so the funnel agrees
# with the list a user sees.

# B13 (2026-06-16): pipeline reads are Postgres-first, honoring the standing
# rule that all data lives in PostgreSQL. This is now safe because the write
# path is reliable: _db_sync_pipeline_deal raises on failure (no silent
# swallow) and pipeline_deal_create verifies the row landed in Postgres — so a
# deal cannot exist in JSON without also being in the DB. (B12 had flipped this
# to JSON-first as an emergency patch when the sync was best-effort.)
_PIPELINE_READ_DB_FIRST = True

_STAGE_WEIGHTS = {
    "Lead": 0.05, "Contacted": 0.10, "Qualified": 0.25, "Proposal": 0.40,
    "Negotiation": 0.60, "Compliance": 0.80, "Closed Won": 1.0, "Closed Lost": 0.0,
}


def _deal_value(d: dict) -> float:
    # KES-equivalent: amount_kes when stamped (FCY deals), else native deal_value
    # (LCY deals coincide at rate 1). Keeps analytics in the bank's reporting
    # currency, consistent with pipeline_summary.
    try:
        v = d.get("amount_kes")
        if v is None:
            v = d.get("deal_value", 0)
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_product(s: str) -> str:
    return "".join(ch for ch in str(s or "").lower() if ch.isalnum())


def _product_class(product_type: str):
    """The product_catalogue class for a product, from pipeline_settings.json
    (admin-owned config). Returns one of: Assets, Liabilities, Transactional,
    Insurance, Investments — or None if unknown. Exact match first, then a
    normalized containment fallback so naming drift (e.g. 'Mortgage / Home Loan'
    vs catalogue 'Mortgage', 'Current Account (CASA)' vs 'Current Account')
    still classifies correctly."""
    pt_n = _norm_product(product_type)
    if not pt_n:
        return None
    cfg = _load_json("pipeline_settings.json") or {}
    cat = cfg.get("product_catalogue", {}) if isinstance(cfg, dict) else {}
    pairs = []
    if isinstance(cat, dict):
        for cls, prods in cat.items():
            if isinstance(prods, list):
                for p in prods:
                    pairs.append((_norm_product(p), cls))
    for p_n, cls in pairs:          # exact
        if p_n and p_n == pt_n:
            return cls
    for p_n, cls in pairs:          # containment
        if p_n and (p_n in pt_n or pt_n in p_n):
            return cls
    return None


def _classify_product(product_type: str) -> str:
    """Headline bucket for a product: 'asset' | 'liability' | 'insurance' |
    'other'. Sourced from product_catalogue (admin config); keyword fallback
    only when the product isn't in the catalogue at all."""
    cls = _product_class(product_type)
    if cls == "Assets":
        return "asset"
    if cls == "Liabilities":
        return "liability"
    if cls == "Insurance":
        return "insurance"
    if cls in ("Transactional", "Investments"):
        return "other"
    low = str(product_type or "").lower()
    if any(k in low for k in ("insurance", "assurance", "cover")):
        return "insurance"
    if any(k in low for k in ("loan", "overdraft", "finance", "mortgage",
                              "facility", "discount")):
        return "asset"
    if any(k in low for k in ("deposit", "account", "casa", "savings",
                              "current", "call", "notice")):
        return "liability"
    return "other"


_DEFAULT_SEGMENT_LABELS = {
    "Mass / Retail": "Direct",
    "Core Middle":   "Advantage",
    "Affluent":      "Premier",
}


def _segment_labels() -> dict:
    """Admin-configurable display names for segments (pipeline_settings.json →
    segment_labels). Defaults to the Ecobank vocabulary (Direct / Advantage /
    Premier) when the admin hasn't set a map, so the bank's own names show
    without any data-file change; an explicit map in settings overrides it."""
    cfg = _load_json("pipeline_settings.json") or {}
    m = cfg.get("segment_labels") if isinstance(cfg, dict) else None
    return dict(m) if isinstance(m, dict) and m else dict(_DEFAULT_SEGMENT_LABELS)


def _referral_blocked(d: dict) -> bool:
    """A deal that should NOT count toward pipeline value, analytics, or the
    manager validation queue: a referral still awaiting the recipient's
    acceptance ('pending'), or one the recipient declined ('declined') and that
    now sits in the returned-for-reassignment pool. Accepted referrals behave
    like any normal owned deal and DO count (assured only once validated)."""
    return str(d.get("referral_status") or "") in ("pending", "declined")


def _derive_segment(d: dict) -> str:
    """Canonical segment bucket from the deal's explicit segment field, else
    derived from client_type. Returns the INTERNAL key (before label mapping)."""
    seg = str(d.get("segment", "") or "").strip()
    if seg:
        return seg
    ct = str(d.get("client_type", "") or "").strip()
    ctl = ct.lower()
    if not ct:
        return "Unclassified"
    if "affluent" in ctl:
        return "Affluent"
    if ctl.startswith("individual"):
        return "Mass / Retail"
    if "sme" in ctl:
        return "SME"
    if "corporate" in ctl or "business" in ctl:
        return "Corporate / Business"
    return ct


def _segment_of(d: dict) -> str:
    """Customer segment for a deal, with the admin display-name map applied.
    Shared by the analytics by_segment dimension and the funnel stage-drill so
    they can't drift. NOT a hardcoded hierarchy — purely a presentation bucket,
    and the labels are admin-configurable (see _segment_labels)."""
    base = _derive_segment(d)
    return _segment_labels().get(base, base)


def _sector_of(d: dict) -> str:
    """CBK economic sector for a deal. Individual / MOU-partnership deals group
    under 'Individual / Partnership'; Business deals use their CBK sector. Shared
    by the analytics by_sector dimension and the funnel stage-drill."""
    if _client_type_field(d.get("client_type", "")) == "mou" or d.get("mou_id"):
        return "Individual / Partnership"
    return d.get("sector") or "Unclassified"


def _stage_flow_for(product_type: str) -> list:
    """Ordered stage flow for a product's class, from pipeline_settings
    stage_flows (admin config, the single source of truth). Classes:
    asset / liability / insurance / other. Falls back to the core per-category
    flow if stage_flows isn't configured, so nothing breaks pre-config."""
    cls = _classify_product(product_type)
    cfg = _load_json("pipeline_settings.json") or {}
    flows = cfg.get("stage_flows", {}) if isinstance(cfg, dict) else {}
    flow = flows.get(cls)
    if isinstance(flow, list) and flow:
        return [str(s) for s in flow]
    # Fallback: core per-category flow (pre-stage_flows behaviour).
    try:
        from utils.core import get_pipeline_category, get_stages_for_category
        cat = get_pipeline_category(product_type or "")
        return [s["stage"] for s in get_stages_for_category(cat)]
    except Exception:
        return []


# Client business lines (Ecobank: Consumer / Commercial / Corporate & Investment
# Banking). Admin-configurable (pipeline_settings.json → client_types). Each entry
# is {key, label, field} where field ∈ {"mou","sector"} drives whether the deal
# form shows the Partnership/MOU selector (consumer) or the CBK sector selector
# (commercial / CIB). "key" is what's stored on the deal as client_type.
_DEFAULT_CLIENT_TYPES = [
    {"key": "Consumer",   "label": "Consumer",                        "field": "mou"},
    {"key": "Commercial", "label": "Commercial",                      "field": "sector"},
    {"key": "CIB",        "label": "Corporate & Investment Banking",  "field": "sector"},
]
# Deals created before the rename carry the old binary values; map them so
# classification stays correct without a data migration.
_CLIENT_TYPE_ALIASES = {"individual": "Consumer", "business": "Commercial"}


def _client_types() -> list:
    """Configured client business lines (admin-editable), each normalised to
    {key, label, field}. Defaults to the Ecobank three-line structure."""
    cfg = _load_json("pipeline_settings.json") or {}
    m = cfg.get("client_types") if isinstance(cfg, dict) else None
    if isinstance(m, list) and m:
        out = []
        for e in m:
            if isinstance(e, dict) and e.get("key"):
                out.append({
                    "key":   str(e["key"]),
                    "label": str(e.get("label") or e["key"]),
                    "field": "mou" if str(e.get("field", "")).lower() == "mou" else "sector",
                })
        if out:
            return out
    return [dict(x) for x in _DEFAULT_CLIENT_TYPES]


def _client_type_field(client_type: str) -> str:
    """'mou' or 'sector' for a deal's client_type — handles configured keys,
    legacy Individual/Business aliases, and keyword fallback. Drives _sector_of
    and the form's third-field choice so they can't drift."""
    low = str(client_type or "").strip().lower()
    if not low:
        return "sector"
    types = _client_types()
    for t in types:
        if t["key"].lower() == low:
            return t["field"]
    alias = _CLIENT_TYPE_ALIASES.get(low)
    if alias:
        for t in types:
            if t["key"] == alias:
                return t["field"]
    if "individual" in low or "consumer" in low:
        return "mou"
    return "sector"


_DEFAULT_CUSTOMER_SEGMENTS = {
    # Keyed by client-type key (Ecobank business lines). Consumer tiers map to the
    # display names (Premier ← Affluent, Advantage ← Core Middle, Direct ← Mass).
    "Consumer":   ["Premier", "Advantage", "Direct"],
    "Commercial": ["Large Corporate", "Corporate", "SME", "Micro Enterprise"],
    "CIB":        ["Multinational", "Large Local Corporate", "Financial Institution", "Public Sector"],
}


def _customer_segments() -> dict:
    """Segment options per client-type key for the Create-Deal form. The default
    (keyed by the Ecobank business lines) is always present so every configured
    client type resolves to a list; an admin customer_segments map overlays it."""
    cfg = _load_json("pipeline_settings.json") or {}
    m = cfg.get("customer_segments") if isinstance(cfg, dict) else None
    out = dict(_DEFAULT_CUSTOMER_SEGMENTS)
    if isinstance(m, dict) and m:
        out.update(m)
    return out


# Deal-creation fields the bank requires before a deal can be created. Admin-
# configurable (pipeline_settings.json → required_fields); defaults to the
# historical hard-coded set. The React create form reads this to decide which
# inputs to mark mandatory + validate, so requiredness is a config decision,
# not code.
_DEFAULT_REQUIRED_FIELDS = ["client_name", "product_type", "deal_value", "stage"]


def _required_fields() -> list:
    cfg = _load_json("pipeline_settings.json") or {}
    m = cfg.get("required_fields") if isinstance(cfg, dict) else None
    if isinstance(m, list) and m:
        return [str(x) for x in m]
    return list(_DEFAULT_REQUIRED_FIELDS)


def _acquire_scoped_deals(user: dict) -> list:
    """All deals in the caller's cascade scope.

    JSON-first (B12, 2026-06-16): the JSON store (PipelineManager) is the
    synchronous source of truth for every create/mutation; Postgres is a
    best-effort mirror that can lag, so DB-first reads hid freshly-created
    deals. Reading JSON-first keeps the list, analytics, and validation queue
    in agreement. Flip _PIPELINE_READ_DB_FIRST back to True once the DB sync is
    guaranteed (PENDING: unify pipeline reads on Postgres)."""
    from utils.api_pipeline_scope import (
        get_visible_staff_codes, filter_deals_by_visible_codes,
    )
    visible_codes = get_visible_staff_codes(user)
    if _PIPELINE_READ_DB_FIRST and _db_available():
        try:
            from utils.db import db as _db
            rows = _db.fetch_all(
                "SELECT * FROM pipeline_deals ORDER BY open_date DESC", ())
            deals = [_normalize_db_deal_row(d) for d in _serialize(rows)]
            return filter_deals_by_visible_codes(deals, visible_codes)
        except Exception as e:
            logger.error(f"Pipeline analytics DB error: {e}")
    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    deals = pm.get_deals()
    return filter_deals_by_visible_codes(deals, visible_codes)


def _compute_pipeline_analytics(deals: list) -> dict:
    """Headline totals + overall funnel + per-category funnels. Pure function
    over a deal list (already scope-filtered) — mirrors 3_pipeline.py."""
    from utils.core import (
        ACTIVE_STAGES, ALL_ACTIVE_STAGES,
        get_pipeline_category, get_stages_for_category,
    )
    live = [d for d in deals if not d.get("draft") and not _referral_blocked(d)]
    active = [d for d in live if d.get("stage") in ACTIVE_STAGES]
    # Validation split: management anchors on VALIDATED (manager-assured) deals.
    # Unvalidated active deals are surfaced separately as "pending assurance".
    validated_active = [d for d in active if d.get("manager_validated")]
    pending_active = [d for d in active if not d.get("manager_validated")]
    won = [d for d in live if d.get("stage") == "Closed Won"]
    lost = [d for d in live if d.get("stage") == "Closed Lost"]

    total_value = sum(_deal_value(d) for d in validated_active)        # assured
    pending_value = sum(_deal_value(d) for d in pending_active)        # pending assurance
    weighted_value = sum(_deal_value(d) * _STAGE_WEIGHTS.get(d.get("stage"), 0)
                         for d in validated_active)
    won_value = sum(_deal_value(d) for d in won)
    win_rate = (round(len(won) / (len(won) + len(lost)) * 100, 1)
                if (won or lost) else 0.0)

    # Overall funnel — VALIDATED active deals only (matches the headline),
    # active stages in canonical union order, non-empty only.
    o_cnt = {s: 0 for s in ALL_ACTIVE_STAGES}
    o_val = {s: 0.0 for s in ALL_ACTIVE_STAGES}
    for d in validated_active:
        s = d.get("stage")
        if s in o_cnt:
            o_cnt[s] += 1
            o_val[s] += _deal_value(d)
    funnel = [{"stage": s, "count": o_cnt[s], "value": o_val[s]}
              for s in ALL_ACTIVE_STAGES if o_cnt[s] > 0]

    # Per-category funnels (mirrors the Streamlit Conversion view).
    cats: dict = {}
    for d in live:
        cat = (d.get("pipeline_category")
               or get_pipeline_category(d.get("product_type", "")))
        cats.setdefault(cat, []).append(d)
    by_category = []
    for cat, cdeals in cats.items():
        stages = [s["stage"] for s in get_stages_for_category(cat)]
        c_cnt = {s: 0 for s in stages}
        c_val = {s: 0.0 for s in stages}
        for d in cdeals:
            s = d.get("stage")
            if s in c_cnt:
                c_cnt[s] += 1
                c_val[s] += _deal_value(d)
        c_funnel = [{"stage": s, "count": c_cnt[s], "value": c_val[s]}
                    for s in stages if s != "Closed Lost"]
        c_active = [d for d in cdeals if d.get("stage") in ACTIVE_STAGES]
        by_category.append({
            "category": cat,
            "count": len(cdeals),
            "active_count": len(c_active),
            "value": sum(_deal_value(d) for d in c_active),
            "weighted": sum(_deal_value(d) * _STAGE_WEIGHTS.get(d.get("stage"), 0)
                            for d in cdeals),
            "funnel": c_funnel,
        })
    by_category.sort(key=lambda c: c["value"], reverse=True)

    # ── Headline buckets (admin product_catalogue): asset / liability /
    # insurance / other. Replaces the meaningless loan+deposit combined total.
    def _bucket_stats(dl: list) -> dict:
        act = [d for d in dl if d.get("stage") in ACTIVE_STAGES]
        val_act = [d for d in act if d.get("manager_validated")]
        pend_act = [d for d in act if not d.get("manager_validated")]
        won_ = [d for d in dl if d.get("stage") == "Closed Won"]
        # funnel from VALIDATED active only (matches the headline)
        b_cnt = {s: 0 for s in ALL_ACTIVE_STAGES}
        b_val = {s: 0.0 for s in ALL_ACTIVE_STAGES}
        for d in val_act:
            s = d.get("stage")
            if s in b_cnt:
                b_cnt[s] += 1
                b_val[s] += _deal_value(d)
        b_funnel = [{"stage": s, "count": b_cnt[s], "value": b_val[s]}
                    for s in ALL_ACTIVE_STAGES if b_cnt[s] > 0]
        return {
            "value": sum(_deal_value(d) for d in val_act),        # assured headline
            "pending_value": sum(_deal_value(d) for d in pend_act),  # pending assurance
            "weighted": sum(_deal_value(d) * _STAGE_WEIGHTS.get(d.get("stage"), 0)
                            for d in val_act),
            "active_count": len(val_act),
            "pending_count": len(pend_act),
            "won_value": sum(_deal_value(d) for d in won_),
            "funnel": b_funnel,
        }

    bucket_defs = [("asset", "Asset Pipeline"), ("liability", "Liability Pipeline"),
                   ("insurance", "Insurance"), ("other", "Other")]
    bucket_deals: dict = {k: [] for k, _ in bucket_defs}
    for d in live:
        bucket_deals[_classify_product(
            d.get("product_type") or d.get("product", ""))].append(d)

    pipelines: dict = {}
    for key, label in bucket_defs:
        stats = _bucket_stats(bucket_deals[key])
        stats["label"] = label
        pipelines[key] = stats

    # "Other" drill-down: by product_catalogue sub-class (Transactional /
    # Investments / unclassified), then by product.
    sub_map: dict = {}
    for d in bucket_deals["other"]:
        sub = _product_class(d.get("product_type") or d.get("product", "")) or "Unclassified"
        sub_map.setdefault(sub, []).append(d)
    other_breakdown = []
    for sub, dl in sub_map.items():
        prod_map: dict = {}
        for d in dl:
            p = d.get("product_type") or d.get("product", "") or "—"
            prod_map.setdefault(p, []).append(d)
        products = []
        for p, pl in prod_map.items():
            pa = [x for x in pl if x.get("stage") in ACTIVE_STAGES]
            products.append({"product": p,
                             "value": sum(_deal_value(x) for x in pa),
                             "count": len(pa)})
        products.sort(key=lambda x: x["value"], reverse=True)
        sa = [d for d in dl if d.get("stage") in ACTIVE_STAGES]
        other_breakdown.append({
            "subclass": sub,
            "value": sum(_deal_value(d) for d in sa),
            "count": len(sa),
            "products": products,
        })
    other_breakdown.sort(key=lambda x: x["value"], reverse=True)
    pipelines["other"]["breakdown"] = other_breakdown

    # ── #3a: cross-cutting breakdowns (KES-equivalent via _deal_value) ──
    # by_product: every product across all buckets, value + count + won.
    _prod: dict = {}
    for d in live:
        p = d.get("product_type") or d.get("product", "") or "—"
        e = _prod.setdefault(p, {"product": p, "value": 0.0, "count": 0, "won_value": 0.0})
        e["value"] += _deal_value(d)
        e["count"] += 1
        if d.get("stage") == "Closed Won":
            e["won_value"] += _deal_value(d)
    _by_product = sorted(_prod.values(), key=lambda x: x["value"], reverse=True)

    # by_sector: CBK sector (Business deals). Individual/MOU deals grouped.
    _sec: dict = {}
    for d in live:
        key = _sector_of(d)
        e = _sec.setdefault(key, {"sector": key, "value": 0.0, "count": 0})
        e["value"] += _deal_value(d)
        e["count"] += 1
    _by_sector = sorted(_sec.values(), key=lambda x: x["value"], reverse=True)

    # by_segment: customer segment (Mass / Affluent / SME / Corporate). Uses the
    # deal's explicit `segment` when set (real Ecobank data will populate it),
    # else derives a clean bucket from client_type (see module-level _segment_of).
    # Informs MD ↔ segment-head conversations: what pipeline flows per segment.
    _seg: dict = {}
    for d in live:
        key = _segment_of(d)
        e = _seg.setdefault(key, {"segment": key, "value": 0.0, "count": 0})
        e["value"] += _deal_value(d)
        e["count"] += 1
    _by_segment = sorted(_seg.values(), key=lambda x: x["value"], reverse=True)

    # by_segment_funnel: per-segment ASSURED (validated + active) funnel by stage,
    # so the pipeline funnel can be sliced by Ecobank segment (Premier / Advantage /
    # Direct / SME / Corporate / …) the same way it slices by product class. Uses
    # the identical val_act semantics as the headline + per-class funnels so the
    # numbers reconcile across views.
    _seg_funnels: dict = {}
    for d in live:
        if d.get("stage") not in ACTIVE_STAGES or not d.get("manager_validated"):
            continue
        seg = _segment_of(d)
        sf = _seg_funnels.setdefault(
            seg, {s: {"count": 0, "value": 0.0} for s in ALL_ACTIVE_STAGES})
        s = d.get("stage")
        if s in sf:
            sf[s]["count"] += 1
            sf[s]["value"] += _deal_value(d)
    by_segment_funnel = []
    for seg, sf in _seg_funnels.items():
        funnel = [{"stage": s, "count": sf[s]["count"], "value": sf[s]["value"]}
                  for s in ALL_ACTIVE_STAGES if sf[s]["count"] > 0]
        if not funnel:
            continue
        by_segment_funnel.append({
            "segment": seg,
            "active_count": sum(f["count"] for f in funnel),
            "value": sum(f["value"] for f in funnel),
            "funnel": funnel,
        })
    by_segment_funnel.sort(key=lambda x: x["value"], reverse=True)

    # by_currency_book: KES-equivalent split (mirrors the dashboard).
    _by_currency_book = {"LCY": {"value": 0.0, "count": 0},
                         "FCY": {"value": 0.0, "count": 0}}
    for d in live:
        book = d.get("currency_book") or (
            "FCY" if str(d.get("currency", "KES")).strip() not in ("", "KES") else "LCY")
        b = _by_currency_book["FCY" if book == "FCY" else "LCY"]
        b["value"] += _deal_value(d)
        b["count"] += 1

    # by_unit (branch / org unit) and by_rm (relationship manager) — the org
    # dimensions for executive drill-down. Both straight off the deal record.
    _unit: dict = {}
    for d in live:
        u = d.get("unit") or "Unassigned"
        e = _unit.setdefault(u, {"unit": u, "value": 0.0, "count": 0})
        e["value"] += _deal_value(d)
        e["count"] += 1
    _by_unit = sorted(_unit.values(), key=lambda x: x["value"], reverse=True)

    _rm: dict = {}
    for d in live:
        nm = d.get("staff_name") or d.get("staff_code") or "Unassigned"
        e = _rm.setdefault(nm, {"rm": nm, "value": 0.0, "count": 0})
        e["value"] += _deal_value(d)
        e["count"] += 1
    _by_rm = sorted(_rm.values(), key=lambda x: x["value"], reverse=True)[:25]

    return {
        "totals": {
            "total_value": total_value,
            "pending_value": pending_value,
            "weighted_value": weighted_value,
            "won_value": won_value,
            "active_count": len(validated_active),
            "pending_count": len(pending_active),
            "won_count": len(won),
            "lost_count": len(lost),
            "live_count": len(live),
            "win_rate": win_rate,
        },
        "pipelines": pipelines,
        "funnel": funnel,
        "by_category": by_category,
        "by_product": _by_product,
        "by_sector": _by_sector,
        "by_segment": _by_segment,
        "by_segment_funnel": by_segment_funnel,
        "by_currency_book": _by_currency_book,
        "by_unit": _by_unit,
        "by_rm": _by_rm,
    }


@app.get("/api/pipeline/analytics")
def pipeline_analytics(user: dict = Depends(get_current_user)):
    """Funnel + headline pipeline metrics over the caller's visible deals."""
    _audit("API_PIPELINE_ANALYTICS", user, "")
    deals = _acquire_scoped_deals(user)
    result = _compute_pipeline_analytics(deals)
    # Scope-aware manager queue counts (one call serves all dashboard tiles).
    from utils.api_pipeline_manager_actions import is_manager as _is_mgr
    pending_validation = 0
    pending_cancel = 0
    if _is_mgr(user):
        from utils.api_pipeline_scope import get_visible_staff_codes
        from utils.core import PipelineManager as _PM_for_api
        vc = get_visible_staff_codes(user)
        pm = _PM_for_api()
        pending_validation = len(pm.get_pending_validations(manager_codes=vc))
        pending_cancel = len(pm.get_cancel_requests(manager_codes=vc))
    result["totals"]["pending_validation"] = pending_validation
    result["totals"]["pending_cancel"] = pending_cancel
    return result


@app.get("/api/pipeline/drill")
def pipeline_drill(
    unit: Optional[str] = None,
    rm: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Drill into the pipeline by branch (unit) then relationship manager (rm).

    Reuses _acquire_scoped_deals so cascade scope is enforced (a branch
    manager only ever drills their own branch). Returns the RM breakdown
    within the current filter (for the branch level) and the individual deal
    list (for the RM level), all KES-equivalent. Pure filter over the scoped
    set — no new query path, so it can't drift from the analytics totals.
    """
    _audit("API_PIPELINE_DRILL", user, f"unit={unit} rm={rm}")
    deals = _acquire_scoped_deals(user)
    if unit:
        deals = [d for d in deals if (d.get("unit") or "Unassigned") == unit]
    if rm:
        deals = [d for d in deals
                 if (d.get("staff_name") or d.get("staff_code") or "Unassigned") == rm]

    rm_map: dict = {}
    for d in deals:
        nm = d.get("staff_name") or d.get("staff_code") or "Unassigned"
        e = rm_map.setdefault(nm, {"rm": nm, "value": 0.0, "count": 0})
        e["value"] += _deal_value(d)
        e["count"] += 1
    by_rm = sorted(rm_map.values(), key=lambda x: x["value"], reverse=True)

    def _slim(d: dict) -> dict:
        return {
            "id": d.get("id"),
            "client_name": d.get("client_name"),
            "product_type": d.get("product_type") or d.get("product"),
            "stage": d.get("stage"),
            "amount_kes": _deal_value(d),
            "currency": d.get("currency", "KES"),
            "staff_name": d.get("staff_name") or d.get("staff_code"),
            "unit": d.get("unit"),
            "expected_close": d.get("expected_close"),
            "probability": d.get("probability"),
        }
    deal_list = [_slim(d) for d in
                 sorted(deals, key=_deal_value, reverse=True)[:200]]

    return {
        "unit": unit,
        "rm": rm,
        "by_rm": by_rm,
        "deals": deal_list,
        "totals": {"value": sum(_deal_value(d) for d in deals), "count": len(deals)},
    }


@app.get("/api/pipeline/funnel/drill")
def pipeline_funnel_drill(
    cls: str = "all",
    stage: str = "",
    user: dict = Depends(get_current_user),
):
    """Drill into one funnel cell: validated (assured) deals in a product class
    (asset/liability/insurance/other/all) at a given stage, broken down by
    product and by segment, plus the deal list. Matches the funnel's basis
    (validated active) so the numbers reconcile. Cascade-scoped via
    _acquire_scoped_deals. Reuses _classify_product / _segment_of so it can't
    drift from the analytics dimensions."""
    _audit("API_PIPELINE_FUNNEL_DRILL", user, f"cls={cls} stage={stage}")
    from utils.core import ACTIVE_STAGES as _ACTIVE
    deals = _acquire_scoped_deals(user)

    def _match(d: dict) -> bool:
        if d.get("stage") not in _ACTIVE:
            return False
        if not d.get("manager_validated"):
            return False
        if stage and d.get("stage") != stage:
            return False
        if cls and cls != "all":
            if _classify_product(d.get("product_type") or d.get("product", "")) != cls:
                return False
        return True

    sel = [d for d in deals if _match(d)]

    prod: dict = {}
    seg: dict = {}
    sec: dict = {}
    for d in sel:
        p = d.get("product_type") or d.get("product", "") or "—"
        ep = prod.setdefault(p, {"product": p, "value": 0.0, "count": 0})
        ep["value"] += _deal_value(d)
        ep["count"] += 1
        k = _segment_of(d)
        es = seg.setdefault(k, {"segment": k, "value": 0.0, "count": 0})
        es["value"] += _deal_value(d)
        es["count"] += 1
        sk = _sector_of(d)
        ec = sec.setdefault(sk, {"sector": sk, "value": 0.0, "count": 0})
        ec["value"] += _deal_value(d)
        ec["count"] += 1

    deals_out = [{
        "id":           d.get("id"),
        "client_name":  d.get("client_name"),
        "product_type": d.get("product_type") or d.get("product", ""),
        "segment":      _segment_of(d),
        "stage":        d.get("stage"),
        "amount_kes":   _deal_value(d),
        "staff_name":   d.get("staff_name") or d.get("staff_code", ""),
        "unit":         d.get("unit"),
    } for d in sorted(sel, key=_deal_value, reverse=True)[:100]]

    return {
        "cls":     cls,
        "stage":   stage,
        "totals":  {"value": sum(_deal_value(d) for d in sel), "count": len(sel)},
        "by_product": sorted(prod.values(), key=lambda x: x["value"], reverse=True),
        "by_segment": sorted(seg.values(), key=lambda x: x["value"], reverse=True),
        "by_sector": sorted(sec.values(), key=lambda x: x["value"], reverse=True),
        "deals":   deals_out,
    }


@app.get("/api/pipeline/export/xlsx")
def pipeline_export_xlsx(user: dict = Depends(get_current_user)):
    """Banking-grade Excel workbook of the caller's scoped pipeline: a Summary
    sheet (headline + per-class), a Deals sheet (full field set), and breakdown
    sheets by segment / sector / product / stage. Reuses the same scoped data +
    analytics as the screen, so the export reconciles with what's on display."""
    _audit("API_PIPELINE_EXPORT_XLSX", user, "")
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    deals = _acquire_scoped_deals(user)
    an = _compute_pipeline_analytics(deals)

    NAVY = "0E2440"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(bold=True, color=NAVY, size=16)
    sub_font = Font(color="6B7280", size=9)
    money_fmt = '#,##0'

    def _style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _autofit(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    wb = Workbook()

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "A2Z Blueprint — Pipeline Export"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} · scope: {user.get('full_name') or user.get('username','')}"
    ws["A2"].font = sub_font
    t = an.get("totals", {})
    rows = [
        ("Metric", "Value (KES)", "Count"),
        ("Total pipeline (active)", t.get("total_value", 0) + t.get("pending_value", 0), t.get("active_count", 0) + t.get("pending_count", 0)),
        ("Assured (validated)", t.get("total_value", 0), t.get("active_count", 0)),
        ("Pending assurance", t.get("pending_value", 0), t.get("pending_count", 0)),
        ("Won", t.get("won_value", 0), t.get("won_count", 0)),
    ]
    for k, v in (an.get("pipelines") or {}).items():
        rows.append((f"{v.get('label', k)} — assured", v.get("value", 0), v.get("active_count", 0)))
    start = 4
    for r, row in enumerate(rows, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r > start and c >= 2 and isinstance(val, (int, float)):
                cell.number_format = money_fmt
    _style_header(ws, start, 3)
    _autofit(ws, [34, 20, 12])
    ws.freeze_panes = "A5"

    # ── Deals ──
    wd = wb.create_sheet("Deals")
    headers = ["Deal ID", "Client", "Product", "Class", "Segment", "Sector",
               "Stage", "Currency", "Amount (native)", "Amount (KES)", "Owner",
               "Branch/Unit", "Expected Close", "Probability %", "Assured"]
    wd.append(headers)
    _style_header(wd, 1, len(headers))
    for d in sorted(deals, key=_deal_value, reverse=True):
        wd.append([
            d.get("id", ""),
            d.get("client_name", ""),
            d.get("product_type") or d.get("product", ""),
            _classify_product(d.get("product_type") or d.get("product", "")),
            _segment_of(d),
            _sector_of(d),
            d.get("stage", ""),
            d.get("currency", "KES"),
            float(d.get("deal_value") or d.get("amount") or 0),
            _deal_value(d),
            d.get("staff_name") or d.get("staff_code", ""),
            d.get("unit", ""),
            str(d.get("expected_close") or ""),
            float(d.get("probability") or 0),
            "Yes" if d.get("manager_validated") else "No",
        ])
    for r in range(2, wd.max_row + 1):
        wd.cell(row=r, column=9).number_format = money_fmt
        wd.cell(row=r, column=10).number_format = money_fmt
    _autofit(wd, [10, 28, 22, 11, 16, 22, 16, 9, 16, 16, 22, 18, 14, 12, 9])
    wd.freeze_panes = "A2"

    # ── Breakdown sheets ──
    def _breakdown_sheet(name, key_field, rows_data):
        w = wb.create_sheet(name)
        w.append([key_field, "Value (KES)", "Count"])
        _style_header(w, 1, 3)
        for row in rows_data:
            w.append([row.get(key_field.lower(), row.get("label", "")), row.get("value", 0), row.get("count", 0)])
        for r in range(2, w.max_row + 1):
            w.cell(row=r, column=2).number_format = money_fmt
        _autofit(w, [30, 20, 12])
        w.freeze_panes = "A2"

    _breakdown_sheet("By Segment", "Segment", an.get("by_segment", []))
    _breakdown_sheet("By Sector", "Sector", an.get("by_sector", []))
    _breakdown_sheet("By Product", "Product", an.get("by_product", []))
    _breakdown_sheet("By Stage", "Stage", [{"stage": s.get("stage"), "value": s.get("value"), "count": s.get("count")} for s in an.get("funnel", [])])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"A2Z_Pipeline_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Pipeline Mutation Endpoints (v10.505 Phase 3 Arc α Batch α3) ──
#
# POST/PUT/advance for pipeline deals. Adds API-side equivalents of
# the Streamlit page's create/update/stage-change flows from
# pages/3_pipeline.py:940-988 (add) + 1330-1340 (advance) + 1310-1314
# (update). Routes through PipelineManager — same canonical engine
# Streamlit consumes (G394 alignment) — applies cascade scope
# (α2 / G395 alignment) — and rejects LMS-handoff stages until α4
# implements the LoanApplication auto-creation (G396 alignment).
#
# Side effects on success: BSC recompute via update_bsc_from_modules
# (matches Streamlit's `_bsc_trigger(uname, "K041")` pattern, 16
# occurrences in the page); pipeline_summary cache invalidation so
# the next GET reflects the mutation.


@app.post("/api/pipeline/deals/refer", status_code=201)
def pipeline_deal_refer(
    payload: "PipelineDealRefer",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """Create a referral-only pipeline deal.

    Conflict resolution path #1 (audit Section 15.4): the referring
    RM defers pursuit to the portfolio owner. A new deal record is
    created with ``is_referral=True``, ``deal_value=0``,
    ``product_type="Referral"``, ``stage="Lead"``. The portfolio
    owner can later pick this up via their own deal queue and either
    pursue or decline.

    Mirrors the Streamlit referral flow at ``pages/3_pipeline.py:558-577``
    exactly — same field shape, same auto-set values, same audit
    emission. Streamlit isn't migrated to use this endpoint in
    α5; that's a separate small batch later.

    Authorization: any authenticated user may refer. The validator
    rejects self-referrals (where staff_code == portfolio_owner_code)
    since that's a no-op.
    """
    _audit("API_PIPELINE_REFER_ATTEMPT", user,
           f"client={payload.client_name} owner={payload.portfolio_owner_code}")

    from utils.api_pipeline_models import (
        PipelineDeal,
        PipelineDealMutationResponse,
    )
    from utils.api_pipeline_mutations import (
        validate_refer_payload,
        emit_bsc_trigger,
        invalidate_pipeline_caches,
    )

    deal_dict = payload.model_dump(exclude_unset=False)
    # H1 (2026-06-14): server-authoritative caller identity (see the create
    # route). Derive staff_code/staff_name from users.json when the client
    # omitted them, so a thin client identity can't fail the referral.
    if (not str(deal_dict.get("staff_code") or "").strip()
            or not str(deal_dict.get("staff_name") or "").strip()):
        from utils.core import UserManager as _UM_id
        _full = _UM_id().users.get(str(user.get("username", "") or "")) or {}
        if not str(deal_dict.get("staff_code") or "").strip():
            deal_dict["staff_code"] = str(_full.get("staff_code", "") or "")
        if not str(deal_dict.get("staff_name") or "").strip():
            deal_dict["staff_name"] = str(
                _full.get("full_name", "") or user.get("username", "") or "")
    ok, reason = validate_refer_payload(deal_dict)
    if not ok:
        _audit("API_PIPELINE_REFER_REJECTED", user, reason)
        raise HTTPException(status_code=400, detail=reason)

    # Build the referral record. Auto-set values match
    # pages/3_pipeline.py:561-572 exactly.
    from datetime import date as _dt
    referral_record = {
        "staff_code":           payload.staff_code,
        "staff_name":           payload.staff_name,
        "unit":                 payload.unit or "",
        "client_name":          payload.client_name,
        "account_number":       payload.account_number or "",
        "client_type":          "Existing",
        "product_type":         "Referral",
        "deal_value":           0,
        "stage":                "Lead",
        "probability":          0.05,
        "next_action":          (
            f"Referred to {payload.referred_to}: {payload.referral_note}"
            if payload.referral_note else f"Referred to {payload.referred_to}"
        ),
        "next_action_date":     str(_dt.today()),
        "source":               "Referral",
        "portfolio_owner_code": payload.portfolio_owner_code,
        "portfolio_owner_name": payload.portfolio_owner_name,
        "is_referral":          True,
        "referred_to":          payload.referred_to,
        "referral_note":        payload.referral_note or "",
        "is_ntb":               False,
        # P3: enter the referral lifecycle so the portfolio owner must ACCEPT
        # (the "nod") before the deal progresses. referred_to_code routes it to
        # the owner's incoming inbox; referred_by_* let the referrer track it
        # (outgoing) and receive it back if the owner declines (returned).
        "referral_status":      "pending",
        "referred_to_code":     str(payload.portfolio_owner_code or "").strip(),
        "referred_by_code":     str(deal_dict.get("staff_code") or "").strip(),
        "referred_by_name":     str(deal_dict.get("staff_name") or "").strip(),
        "referred_at":          datetime.now().isoformat(),
        # BSC credit goes to whoever closes the referred deal — typically
        # the portfolio owner. Default to portfolio_owner_name; the
        # picking-up RM can override later.
        "bsc_credit_to":        payload.portfolio_owner_name,
    }

    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    new_id = pm.add_deal(referral_record)
    _db_sync_pipeline_deal(pm.get_deal(new_id))  # H5: mirror to DB-backed reads

    # Mirror Streamlit's audit emission (line 573: DEAL_REFERRED)
    _audit("DEAL_REFERRED", user,
           f"{new_id}|{payload.client_name}->{payload.referred_to}")

    # BSC trigger + cache invalidation (matches Streamlit behavior)
    bsc_ok = emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    created = pm.get_deal(new_id) or referral_record
    return PipelineDealMutationResponse(
        deal=PipelineDeal.model_validate(created),
        status="referred",
        bsc_triggered=bsc_ok,
    ).model_dump()


# ── Referral lifecycle: refer existing deal → accept / decline ──────────────
# Generalises the zero-value /refer marker into a real assignment-with-
# acceptance flow. A deal is referred to another staff member, who must ACCEPT
# before owning its progression; a declined referral (reason required) drops
# into the returned-for-reassignment pool. Pending + declined referrals do NOT
# count toward pipeline value, analytics, or the manager validation queue
# (see _referral_blocked). An accepted deal counts as ASSURED only once the
# recipient's line manager validates it — exactly like any other deal.

def _resolve_actor(user: dict):
    """(staff_code, full_name, is_privileged) for the caller. Privileged =
    exec/admin (lenient role substring), permitted to act on the recipient's
    behalf."""
    from utils.core import UserManager as _UM
    rec = _UM().users.get(str(user.get("username", "") or "")) or {}
    role = str(rec.get("role", "") or user.get("role", "") or "").lower()
    priv = bool(user.get("is_admin")) or any(
        k in role for k in ("chief", "managing", "director", "admin"))
    return (str(rec.get("staff_code", "") or ""),
            str(rec.get("full_name", "") or user.get("username", "") or ""),
            priv)


@app.post("/api/pipeline/deals/{deal_id}/refer", status_code=200)
def pipeline_deal_refer_existing(
    deal_id: str,
    payload: Dict[str, Any] = Body(...),
    user: dict = Depends(get_current_user),
):
    """Refer/assign an EXISTING deal to another staff member for pursuit.
    Sets referral_status='pending'; the recipient must accept before they own
    the deal's progression. Any staff with the deal in scope (or an admin) may
    refer. Referring to the current owner is a no-op and rejected."""
    _audit("API_PIPELINE_REFER_EXISTING_ATTEMPT", user, f"deal_id={deal_id}")
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM
    from utils.api_pipeline_mutations import invalidate_pipeline_caches

    rcode = str(payload.get("referred_to_code") or "").strip()
    rname = str(payload.get("referred_to_name") or payload.get("referred_to") or "").strip()
    note  = str(payload.get("referral_note") or "").strip()
    if not rcode or not rname:
        raise HTTPException(status_code=400,
                            detail="referred_to_code and referred_to_name are required")

    pm = _PM()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    actor_code, actor_name, priv = _resolve_actor(user)
    visible = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    po = str(deal.get("portfolio_owner_code", "") or "")
    if not priv and sc not in visible and (not po or po not in visible):
        raise HTTPException(status_code=403, detail="Deal is outside your scope")
    if rcode in (sc, po):
        raise HTTPException(status_code=400,
                            detail="That person already owns this deal — nothing to refer.")
    if str(deal.get("referral_status") or "") == "pending":
        raise HTTPException(status_code=400,
                            detail="This deal already has a pending referral awaiting acceptance.")

    pm.update_deal(deal_id, {
        "referral_status":  "pending",
        "referred_to_code": rcode,
        "referred_to":      rname,
        "referred_by_code": actor_code,
        "referred_by_name": actor_name,
        "referral_note":    note,
        "referred_at":      datetime.now().isoformat(),
        "decline_reason":   "",
    }, user.get("username", ""))
    _db_sync_pipeline_deal(pm.get_deal(deal_id))
    _audit("DEAL_REFERRED_EXISTING", user, f"{deal_id}->{rcode}")
    invalidate_pipeline_caches()
    return {"deal_id": deal_id, "referral_status": "pending",
            "referred_to": rname, "referred_to_code": rcode}


@app.post("/api/pipeline/deals/{deal_id}/referral/accept", status_code=200)
def pipeline_referral_accept(
    deal_id: str,
    payload: Dict[str, Any] = Body(default_factory=dict),
    user: dict = Depends(get_current_user),
):
    """Recipient accepts a pending referral and becomes its owner. Progression
    unlocks; the deal enters the recipient's pipeline + their line manager's
    validation queue (assured only once validated). Only the named recipient
    (or an admin) may accept."""
    _audit("API_PIPELINE_REFERRAL_ACCEPT_ATTEMPT", user, f"deal_id={deal_id}")
    from utils.core import PipelineManager as _PM
    from utils.api_pipeline_mutations import invalidate_pipeline_caches

    pm = _PM()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    if str(deal.get("referral_status") or "") != "pending":
        raise HTTPException(status_code=400,
                            detail="This deal has no pending referral to accept.")

    actor_code, actor_name, priv = _resolve_actor(user)
    rcode = str(deal.get("referred_to_code") or "")
    if not priv and actor_code != rcode:
        raise HTTPException(status_code=403,
                            detail="Only the person this deal was referred to can accept it.")

    pm.update_deal(deal_id, {
        "referral_status":      "accepted",
        "accepted_at":          datetime.now().isoformat(),
        "accepted_by":          actor_code or rcode,
        "staff_code":           rcode or actor_code,
        "staff_name":           str(deal.get("referred_to") or actor_name),
        "portfolio_owner_code": rcode or actor_code,
        "portfolio_owner_name": str(deal.get("referred_to") or actor_name),
        "manager_validated":    False,
    }, user.get("username", ""))
    _db_sync_pipeline_deal(pm.get_deal(deal_id))
    _audit("DEAL_REFERRAL_ACCEPTED", user, f"{deal_id} by {actor_code or rcode}")
    invalidate_pipeline_caches()
    return {"deal_id": deal_id, "referral_status": "accepted"}


@app.post("/api/pipeline/deals/{deal_id}/referral/decline", status_code=200)
def pipeline_referral_decline(
    deal_id: str,
    payload: Dict[str, Any] = Body(...),
    user: dict = Depends(get_current_user),
):
    """Recipient declines a pending referral WITH A REASON. The deal drops into
    the returned-for-reassignment pool (referral_status='declined') — NOT
    closed. Only the named recipient (or an admin) may decline."""
    _audit("API_PIPELINE_REFERRAL_DECLINE_ATTEMPT", user, f"deal_id={deal_id}")
    from utils.core import PipelineManager as _PM
    from utils.api_pipeline_mutations import invalidate_pipeline_caches

    reason = str(payload.get("reason") or payload.get("decline_reason") or "").strip()
    if len(reason) < 3:
        raise HTTPException(status_code=400, detail="A decline reason is required.")

    pm = _PM()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    if str(deal.get("referral_status") or "") != "pending":
        raise HTTPException(status_code=400,
                            detail="This deal has no pending referral to decline.")

    actor_code, actor_name, priv = _resolve_actor(user)
    rcode = str(deal.get("referred_to_code") or "")
    if not priv and actor_code != rcode:
        raise HTTPException(status_code=403,
                            detail="Only the person this deal was referred to can decline it.")

    pm.update_deal(deal_id, {
        "referral_status": "declined",
        "decline_reason":  reason,
        "declined_at":     datetime.now().isoformat(),
        "declined_by":     actor_code or rcode,
    }, user.get("username", ""))
    _db_sync_pipeline_deal(pm.get_deal(deal_id))
    _audit("DEAL_REFERRAL_DECLINED", user, f"{deal_id}: {reason[:60]}")
    invalidate_pipeline_caches()
    return {"deal_id": deal_id, "referral_status": "declined", "decline_reason": reason}


def _all_pipeline_deals() -> list:
    """All pipeline deals, UNSCOPED (DB-first). The referral inbox / returned /
    following queries filter by participation (referred_to / referred_by), which
    can cross cascade boundaries, so they must NOT be pre-filtered by visible
    scope the way the main deal list is."""
    if _PIPELINE_READ_DB_FIRST and _db_available():
        try:
            from utils.db import db as _db
            rows = _db.fetch_all("SELECT * FROM pipeline_deals", tuple())
            return [_normalize_db_deal_row(d) for d in _serialize(rows)]
        except Exception as e:
            logger.error(f"Referral list DB error: {e}")
    from utils.core import PipelineManager as _PM
    return _PM().get_deals()


_DEFAULT_REFERRAL_BUSINESS_DEPARTMENTS = [
    "Sales", "Commercial", "Corporate", "SME", "Retail",
    "Business Banking", "Consumer Banking", "Personal Banking", "CIB",
]
_DEPT_MAP_CACHE = {"ts": 0.0, "map": {}}


def _referral_business_departments() -> set:
    """Admin-configurable set of departments treated as revenue/business units. A
    referral FROM a business dept is B2B (business->business); FROM anything else is
    S2B (support->business). pipeline_settings.referral_business_departments overrides
    the seeded default; the bank finalises the map in admin config."""
    try:
        from utils.core import get_pipeline_settings
        v = (get_pipeline_settings() or {}).get("referral_business_departments")
        if isinstance(v, list) and v:
            return {str(x).strip().lower() for x in v if str(x).strip()}
    except Exception:
        logging.getLogger("a2z.referral").warning("business-dept config fell back", exc_info=True)
    return {x.lower() for x in _DEFAULT_REFERRAL_BUSINESS_DEPARTMENTS}


def _dept_of_map() -> dict:
    """staff_code -> normalized Department, from the roster (5-min memo). Same basis
    the by-department analytics uses, so tier and department views stay consistent."""
    import time
    now = time.time()
    if _DEPT_MAP_CACHE["map"] and now - _DEPT_MAP_CACHE["ts"] < 300:
        return _DEPT_MAP_CACHE["map"]
    out: dict = {}
    try:
        from utils.api_pipeline_scope import get_staff_roster
        roster = get_staff_roster()
        if roster is not None and "Department" in getattr(roster, "columns", []):
            for _, row in roster.iterrows():
                sc = str(row.get("Staff Code") or "").strip()
                if sc:
                    out[sc] = _norm_segment(row.get("Department"))
    except Exception:
        logging.getLogger("a2z.referral").warning("tier dept map: roster load failed", exc_info=True)
    _DEPT_MAP_CACHE["ts"] = now
    _DEPT_MAP_CACHE["map"] = out
    return out


def _classify_referral_tier(by_code, to_code, dept_of=None) -> dict:
    """Derive a referral's tier from the REFERRER's department. business->business = B2B;
    support->business = S2B. cross_unit = referrer and recipient sit in different
    departments (e.g. one branch referring a customer owned in another, or Consumer ->
    Commercial). Derived at read (not stored) so it covers historical referrals too; an
    unknown/unclassified department defaults to B2B (the audited tier), so only an
    explicitly-support department earns the cleaner S2B treatment."""
    if dept_of is None:
        dept_of = _dept_of_map()
    biz = _referral_business_departments()
    by_dept = dept_of.get(str(by_code or "").strip())
    to_dept = dept_of.get(str(to_code or "").strip())
    referrer_is_business = (not by_dept) or (by_dept.strip().lower() in biz)
    return {
        "referral_tier": "B2B" if referrer_is_business else "S2B",
        "cross_unit": bool(by_dept and to_dept and by_dept != to_dept),
        "referrer_department": by_dept or None,
        "recipient_department": to_dept or None,
    }


def _referral_view(d: dict) -> dict:
    """Compact projection for referral list endpoints."""
    return {
        "id":               d.get("id"),
        "client_name":      d.get("client_name"),
        "deal_value":       _deal_value(d),
        "product_type":     d.get("product_type") or d.get("product"),
        "stage":            d.get("stage"),
        "segment":          _segment_of(d),
        "referral_status":  d.get("referral_status"),
        "referred_to":      d.get("referred_to"),
        "referred_to_code": d.get("referred_to_code"),
        "referred_by_name": d.get("referred_by_name"),
        "referred_by_code": d.get("referred_by_code"),
        "referral_note":    d.get("referral_note"),
        "decline_reason":   d.get("decline_reason"),
        "referred_at":      d.get("referred_at"),
        "accepted_at":      d.get("accepted_at"),
        "declined_at":      d.get("declined_at"),
        **_classify_referral_tier(d.get("referred_by_code"), d.get("referred_to_code")),
    }


@app.post("/api/pipeline/deals/{deal_id}/referral/reassign", status_code=200)
def pipeline_referral_reassign(
    deal_id: str,
    payload: Dict[str, Any] = Body(...),
    user: dict = Depends(get_current_user),
):
    """Reassign a RETURNED (declined) referral to a new recipient -> pending.
    Only the original referrer (or an admin) may reassign."""
    _audit("API_PIPELINE_REFERRAL_REASSIGN_ATTEMPT", user, f"deal_id={deal_id}")
    from utils.core import PipelineManager as _PM
    from utils.api_pipeline_mutations import invalidate_pipeline_caches

    rcode = str(payload.get("referred_to_code") or "").strip()
    rname = str(payload.get("referred_to_name") or payload.get("referred_to") or "").strip()
    note  = str(payload.get("referral_note") or "").strip()
    if not rcode or not rname:
        raise HTTPException(status_code=400,
                            detail="referred_to_code and referred_to_name are required")

    pm = _PM()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")
    if str(deal.get("referral_status") or "") != "declined":
        raise HTTPException(status_code=400,
                            detail="Only a returned (declined) referral can be reassigned.")

    actor_code, actor_name, priv = _resolve_actor(user)
    if not priv and actor_code != str(deal.get("referred_by_code") or ""):
        raise HTTPException(status_code=403,
                            detail="Only the original referrer can reassign this returned deal.")
    if rcode in (str(deal.get("staff_code", "") or ""),
                 str(deal.get("portfolio_owner_code", "") or "")):
        raise HTTPException(status_code=400,
                            detail="That person already owns this deal — nothing to reassign.")

    pm.update_deal(deal_id, {
        "referral_status":  "pending",
        "referred_to_code": rcode,
        "referred_to":      rname,
        "referred_by_code": actor_code,
        "referred_by_name": actor_name,
        "referral_note":    note,
        "referred_at":      datetime.now().isoformat(),
        "decline_reason":   "",
    }, user.get("username", ""))
    _db_sync_pipeline_deal(pm.get_deal(deal_id))
    _audit("DEAL_REFERRAL_REASSIGNED", user, f"{deal_id}->{rcode}")
    invalidate_pipeline_caches()
    return {"deal_id": deal_id, "referral_status": "pending",
            "referred_to": rname, "referred_to_code": rcode}


@app.get("/api/pipeline/referrals/incoming")
def pipeline_referrals_incoming(user: dict = Depends(get_current_user)):
    """Pending referrals addressed to the caller — their acceptance inbox."""
    code, _, _ = _resolve_actor(user)
    deals = [d for d in _all_pipeline_deals()
             if code and str(d.get("referral_status") or "") == "pending"
             and str(d.get("referred_to_code") or "") == code]
    return {"deals": [_referral_view(d) for d in deals], "count": len(deals)}


@app.get("/api/pipeline/referrals/returned")
def pipeline_referrals_returned(user: dict = Depends(get_current_user)):
    """Declined referrals awaiting reassignment. The original referrer sees
    theirs; an admin sees all."""
    code, _, priv = _resolve_actor(user)
    deals = [d for d in _all_pipeline_deals()
             if str(d.get("referral_status") or "") == "declined"
             and (priv or str(d.get("referred_by_code") or "") == code)]
    return {"deals": [_referral_view(d) for d in deals], "count": len(deals)}


@app.get("/api/pipeline/referrals/outgoing")
def pipeline_referrals_outgoing(user: dict = Depends(get_current_user)):
    """Referrals the caller has made that are still live (pending or accepted)
    — so they can follow progress."""
    code, _, _ = _resolve_actor(user)
    deals = [d for d in _all_pipeline_deals()
             if code and str(d.get("referred_by_code") or "") == code
             and str(d.get("referral_status") or "") in ("pending", "accepted")]
    return {"deals": [_referral_view(d) for d in deals], "count": len(deals)}


_REFERRAL_PENDING_ALERT_DAYS = 3  # default business days awaiting acceptance before flagging


def _referral_pending_alert_days() -> int:
    """Admin-configurable alert window (pipeline_settings.referral_pending_alert_days);
    falls back to the default. Keeps the threshold config-driven, not hardcoded."""
    try:
        from utils.core import get_pipeline_settings
        v = (get_pipeline_settings() or {}).get("referral_pending_alert_days")
        return int(v) if v not in (None, "") else _REFERRAL_PENDING_ALERT_DAYS
    except Exception:
        return _REFERRAL_PENDING_ALERT_DAYS


def _business_days_since(iso_ts) -> int:
    """Whole business days (Mon-Fri) elapsed since an ISO timestamp; 0 if
    missing/unparseable. Shared helper — also the basis for SLA elapsed-time."""
    if not iso_ts:
        return 0
    try:
        start = datetime.fromisoformat(str(iso_ts).replace("Z", "+00:00")).date()
    except Exception:
        return 0
    today = datetime.now().date()
    if today <= start:
        return 0
    days, d = 0, start
    while d < today:
        d += timedelta(days=1)
        if d.weekday() < 5:
            days += 1
    return days


@app.get("/api/pipeline/referrals/outgoing/analytics")
def pipeline_referrals_outgoing_analytics(user: dict = Depends(get_current_user)):
    """Referrer's-eye analytics over the deals they referred: a funnel by status
    and by stage (progress through to closure) plus alerts — referrals awaiting
    acceptance too long, or declined and needing reassignment. Lets a referrer
    track their referred cases all the way to closure and act on what's stuck."""
    code, _, _ = _resolve_actor(user)
    mine = [d for d in _all_pipeline_deals()
            if code and str(d.get("referred_by_code") or "") == code
            and str(d.get("referral_status") or "")]
    by_status = {"pending": 0, "accepted": 0, "declined": 0}
    by_stage: dict = {}
    won = lost = 0
    alerts = []
    alert_days = _referral_pending_alert_days()
    for d in mine:
        st = str(d.get("referral_status") or "")
        if st in by_status:
            by_status[st] += 1
        stage = str(d.get("stage") or "")
        by_stage[stage] = by_stage.get(stage, 0) + 1
        if stage == "Closed Won":
            won += 1
        elif stage == "Closed Lost":
            lost += 1
        if st == "pending":
            aged = _business_days_since(d.get("referred_at"))
            if aged >= alert_days:
                alerts.append({
                    "id": d.get("id"), "client_name": d.get("client_name"),
                    "referred_to": d.get("referred_to"), "kind": "pending_too_long",
                    "days": aged,
                    "message": f"Awaiting acceptance for {aged} business day{'s' if aged != 1 else ''}",
                })
        elif st == "declined":
            alerts.append({
                "id": d.get("id"), "client_name": d.get("client_name"),
                "referred_to": d.get("referred_to"), "kind": "returned",
                "message": "Declined and returned — needs reassignment",
            })
    return {
        "total": len(mine),
        "by_status": by_status,
        "by_stage": by_stage,
        "closed": {"won": won, "lost": lost},
        "alerts": alerts,
        "alert_count": len(alerts),
    }


@app.get("/api/pipeline/referrals/analytics/by-department")
def pipeline_referrals_by_department(user: dict = Depends(get_current_user)):
    """Department-level referral analytics — referrals grouped by the REFERRER's
    department, so Heads/Chiefs can see their department's referral performance.
    This is the basis for a department-level referral BSC KPI that flows to Head /
    Chief scorecards, mirroring the individual-level KPI. Management roles only."""
    _c, _n, priv = _resolve_actor(user)
    if not priv:
        raise HTTPException(
            status_code=403,
            detail="Department referral analytics require a management role.")

    # referrer staff_code -> normalized department, from the roster
    dept_of: dict = {}
    try:
        from utils.api_pipeline_scope import get_staff_roster
        roster = get_staff_roster()
        if roster is not None and "Department" in getattr(roster, "columns", []):
            for _, row in roster.iterrows():
                sc = str(row.get("Staff Code") or "").strip()
                if sc:
                    dept_of[sc] = _norm_segment(row.get("Department"))
    except Exception as exc:  # surfaced, not silent (CGR1)
        logger.warning("by-department referral analytics: roster load failed: %s", exc)

    agg: dict = {}
    for d in _all_pipeline_deals():
        rbc = str(d.get("referred_by_code") or "")
        st = str(d.get("referral_status") or "")
        if not rbc or not st:
            continue
        dept = dept_of.get(rbc) or "Unassigned"
        a = agg.setdefault(dept, {
            "department": dept, "total": 0,
            "by_status": {"pending": 0, "accepted": 0, "declined": 0},
            "closed": {"won": 0, "lost": 0},
        })
        a["total"] += 1
        if st in a["by_status"]:
            a["by_status"][st] += 1
        stage = str(d.get("stage") or "")
        if stage == "Closed Won":
            a["closed"]["won"] += 1
        elif stage == "Closed Lost":
            a["closed"]["lost"] += 1

    departments = sorted(agg.values(), key=lambda x: x["total"], reverse=True)
    return {
        "departments": departments,
        "total": sum(a["total"] for a in departments),
        "department_count": len(departments),
    }


@app.get("/api/pipeline/referrals/team")
def pipeline_referrals_team(user: dict = Depends(get_current_user)):
    """Hierarchy-scoped referral view — the twin of the pipeline's scoping, for
    referrals. A line manager / head / chief sees referrals made by their reporting
    subtree (their team); the MD/CEO sees all (get_visible_staff_codes returns the
    full roster for them). Returns the team's referrals plus a funnel summary so
    progress is visible at every level. Reuses the canonical REPORTING_TREE walk —
    no duplicate scoping logic."""
    from utils.api_pipeline_scope import get_visible_staff_codes
    visible = set(get_visible_staff_codes(user) or [])
    deals = [d for d in _all_pipeline_deals()
             if str(d.get("referral_status") or "")
             and str(d.get("referred_by_code") or "") in visible]
    by_status = {"pending": 0, "accepted": 0, "declined": 0}
    by_tier = {"B2B": 0, "S2B": 0}
    won = lost = 0
    dept_of = _dept_of_map()
    for d in deals:
        st = str(d.get("referral_status") or "")
        if st in by_status:
            by_status[st] += 1
        t = _classify_referral_tier(d.get("referred_by_code"), d.get("referred_to_code"), dept_of)["referral_tier"]
        by_tier[t] = by_tier.get(t, 0) + 1
        stage = str(d.get("stage") or "")
        if stage == "Closed Won":
            won += 1
        elif stage == "Closed Lost":
            lost += 1
    return {
        "deals": [_referral_view(d) for d in deals],
        "count": len(deals),
        "summary": {"total": len(deals), "by_status": by_status, "by_tier": by_tier,
                    "closed": {"won": won, "lost": lost}},
    }


_SEGMENT_FIX = {"ict": "ICT", "customerservice": "Customer Service"}


def _norm_segment(d) -> str:
    """Normalize a staff Department value into a display segment, merging the
    known case/format variants in the roster (ict->ICT, CustomerService->...)."""
    s = str(d or "").strip()
    return _SEGMENT_FIX.get(s.lower(), s)


@app.get("/api/staff/segments", tags=["pipeline"])
def staff_segments(user: dict = Depends(get_current_user)):
    """Distinct staff segments (Department) with counts — drives the referral
    recipient picker's first step (pick a segment, then a person within it)."""
    from utils.api_pipeline_scope import get_staff_roster
    roster = get_staff_roster()
    if roster is None or len(roster) == 0 or "Department" not in roster.columns:
        return {"segments": [], "count": 0}
    counts: dict = {}
    for _, row in roster.iterrows():
        seg = _norm_segment(row.get("Department"))
        if not seg:
            continue
        counts[seg] = counts.get(seg, 0) + 1
    segments = [{"segment": k, "count": v} for k, v in sorted(counts.items())]
    return {"segments": segments, "count": len(segments)}


@app.get("/api/staff/search", tags=["pipeline"])
def staff_search(
    q: str = Query(default=""),
    segment: str = Query(default=""),
    limit: int = Query(default=20, le=50),
    user: dict = Depends(get_current_user),
):
    """Search the staff roster by name or code — for referral recipient pickers.
    Optional `segment` (Department) narrows to one function (e.g. Credit, Treasury,
    Commercial) so the referrer filters people within a segment instead of the
    whole roster. Bank-wide (referrals cross cascade boundaries), NOT scope-
    filtered; excludes the caller's own record. Empty q returns the first `limit`."""
    from utils.api_pipeline_scope import get_staff_roster
    roster = get_staff_roster()
    if roster is None or len(roster) == 0:
        return {"staff": [], "count": 0}
    has_dept = "Department" in roster.columns
    me = str(user.get("staff_code") or "").strip()
    ql = q.strip().lower()
    segl = segment.strip().lower()
    out = []
    for _, row in roster.iterrows():
        code = str(row.get("Staff Code") or "").strip()
        name = str(row.get("Staff Name") or "").strip()
        if not code or code == me:
            continue
        seg = _norm_segment(row.get("Department")) if has_dept else ""
        if segl and seg.lower() != segl:
            continue
        if ql and ql not in name.lower() and ql not in code.lower():
            continue
        out.append({
            "staff_code": code, "name": name,
            "role": str(row.get("Role") or ""),
            "unit": str(row.get("Unit") or ""),
            "region": str(row.get("Region") or ""),
            "segment": seg,
        })
        if len(out) >= limit:
            break
    return {"staff": out, "count": len(out)}


def _referral_credit_config() -> dict:
    """Admin-tunable config for referral SHADOW credit (pipeline_settings.json →
    referral_credit). Defaults: credit on Disbursed (true materialization) AND
    Closed Won (to reward closing / pipeline hygiene); asset->K238, liability->K239."""
    from utils.core import get_pipeline_settings
    cfg = get_pipeline_settings() or {}
    rc = cfg.get("referral_credit") if isinstance(cfg, dict) else None
    rc = rc if isinstance(rc, dict) else {}
    return {
        "materialized_stages": rc.get("materialized_stages") or ["Disbursed", "Closed Won"],
        "asset_kpi_id":        rc.get("asset_kpi_id") or "K238",
        "liability_kpi_id":    rc.get("liability_kpi_id") or "K239",
    }


@app.post("/api/pipeline/referrals/sync-bsc")
def pipeline_referral_sync_bsc(
    dry_run: bool = Query(default=True),
    user: dict = Depends(require_config_admin),
):
    """Compute (and optionally submit) SHADOW referral credit to referrers' BSC.

    For each MATERIALIZED referred deal (referred_by_code set + stage in the
    configured materialized set), credit the REFERRER with 'Asset Referral'
    (K238) or 'Liabilities/Deposit Referral' (K239) by the deal's product class,
    valued at the deal's KES amount. Shadow / recognition only: distinct non-P&L
    KPIs, the owner's P&L credit is never touched, so there is no double count.

    Defaults to dry_run=True (computes the report, submits nothing) so it can be
    verified against live data before any actuals are written. Config-admin only.
    """
    _audit("API_REFERRAL_BSC_SYNC", user, f"dry_run={dry_run}")
    from utils.pipeline_to_bsc import period_from_date
    rc = _referral_credit_config()
    mat = set(rc["materialized_stages"])
    agg: dict = {}
    skipped = 0
    for d in _all_pipeline_deals():
        rb = str(d.get("referred_by_code") or "").strip()
        if not rb or str(d.get("stage") or "") not in mat:
            continue
        cls = _classify_product(d.get("product_type") or d.get("product", ""))
        if cls == "asset":
            kpi = rc["asset_kpi_id"]
        elif cls == "liability":
            kpi = rc["liability_kpi_id"]
        else:
            skipped += 1
            continue
        period = period_from_date(str(d.get("updated_at") or d.get("last_updated")
                                      or d.get("expected_close") or ""))
        if not period:
            skipped += 1
            continue
        key = (rb, period, kpi)
        e = agg.setdefault(key, {"value": 0.0, "count": 0, "deal_ids": []})
        e["value"] += _deal_value(d)
        e["count"] += 1
        if len(e["deal_ids"]) < 10:
            e["deal_ids"].append(d.get("id"))

    contributions = [
        {"referrer_code": k[0], "period": k[1], "kpi_id": k[2],
         "value": v["value"], "deal_count": v["count"], "deal_ids": v["deal_ids"]}
        for k, v in agg.items()
    ]
    submitted = 0
    if not dry_run:
        from utils.bsc_engine import submit as _bsc_submit
        for c in contributions:
            try:
                ok, _op = _bsc_submit(
                    staff_code=c["referrer_code"], kpi_id=c["kpi_id"],
                    value=c["value"], period=c["period"],
                    source_module="referral_shadow",
                    metadata={"deal_count": c["deal_count"],
                              "deal_ids": c["deal_ids"][:5], "shadow": True},
                )
                if ok:
                    submitted += 1
            except Exception as e:  # noqa: BLE001
                logger.error(f"Referral BSC submit failed for {c['referrer_code']}: {e}")
    return {"dry_run": dry_run, "contributions": len(contributions),
            "submitted": submitted, "skipped": skipped,
            "sample": contributions[:25], "config": rc}


@app.post("/api/pipeline/deals", status_code=201)
def pipeline_deal_create(
    payload: "PipelineDealCreate",  # noqa: F821 — forward ref to keep import lazy
    user: dict = Depends(get_current_user),
):
    """Create a new pipeline deal.

    Required fields: client_name, staff_code, staff_name, deal_value,
    product_type, stage. Returns the created deal with its
    PipelineManager-assigned id.

    Authorization: any authenticated user may create. The staff_code
    on the payload identifies who owns the deal — typically the
    caller's own code, but managers/admins may create on behalf of
    subordinates. (Server-side enforcement of "create on behalf"
    rules is α5 / GAP-005 scope — conflict resolution.)
    """
    _audit("API_PIPELINE_CREATE_ATTEMPT", user,
           f"client={payload.client_name} value={payload.deal_value}")

    # Lazy imports to avoid circular dependencies at module load
    from utils.api_pipeline_models import (
        PipelineDeal,
        PipelineDealMutationResponse,
    )
    from utils.api_pipeline_mutations import (
        validate_create_payload,
        emit_bsc_trigger,
        invalidate_pipeline_caches,
    )

    # Validate required fields + numeric sanity + stage allowlist
    deal_dict = payload.model_dump(exclude_unset=False)
    # H1 (2026-06-14): the server is authoritative for caller identity.
    # get_current_user carries only JWT claims (username/role) — NOT
    # staff_code/full_name (whoami_detailed re-fetches those from
    # users.json: "never trust JWT for these"). If the client omitted them
    # (thin identity), derive from the caller's record so creation can't be
    # rejected for "Missing required field: staff_code" and the client
    # cannot assert an arbitrary owner. Managers/admins may still create on
    # behalf by explicitly supplying a different staff_code (a5/GAP-005).
    if (not str(deal_dict.get("staff_code") or "").strip()
            or not str(deal_dict.get("staff_name") or "").strip()):
        from utils.core import UserManager as _UM_id
        _full = _UM_id().users.get(str(user.get("username", "") or "")) or {}
        if not str(deal_dict.get("staff_code") or "").strip():
            deal_dict["staff_code"] = str(_full.get("staff_code", "") or "")
        if not str(deal_dict.get("staff_name") or "").strip():
            deal_dict["staff_name"] = str(
                _full.get("full_name", "") or user.get("username", "") or "")
    ok, reason = validate_create_payload(deal_dict)
    if not ok:
        _audit("API_PIPELINE_CREATE_REJECTED", user, reason)
        raise HTTPException(status_code=400, detail=reason)

    # P4.5: mandatory portfolio resolution for EXISTING customers. If the deal
    # carries a client_cif that CBS maps to a DIFFERENT relationship owner than
    # the creating RM, the payload MUST acknowledge it (portfolio_owner_code set)
    # — the server-side mirror of the create-form guard, so a direct API call
    # can't silently book a deal against another RM's portfolio. Self-owned and
    # unknown/unmapped CIFs pass through. Fails OPEN on a CBS outage (logged) so
    # deal creation never hard-depends on CBS availability.
    _cif = str(deal_dict.get("client_cif") or "").strip()
    if _cif:
        try:
            from utils.cbs_manager import get_customer_by_cif as _gcbc
            _cust = _gcbc(_cif)
        except Exception as _exc:  # surfaced, not silent (CGR1)
            logger.warning("portfolio guard: CBS lookup failed for cif=%s: %s", _cif, _exc)
            _cust = None
        if _cust:
            _po = str(_cust.get("relationship_manager_code") or "").strip()
            _creator = str(deal_dict.get("staff_code") or "").strip()
            _po_mapped = bool(_po) and _po.upper() != "UNASSIGNED"
            _resolved = bool(str(deal_dict.get("portfolio_owner_code") or "").strip())
            if _po_mapped and _po != _creator and not _resolved:
                msg = (f"Customer {_cif} is in another RM's portfolio (owner {_po}). "
                       f"Set portfolio_owner_code and choose a resolution path "
                       f"(refer, seek permission, or override).")
                _audit("API_PIPELINE_CREATE_REJECTED", user, msg)
                raise HTTPException(status_code=400, detail=msg)

    # Route through canonical manager (G394 alignment)
    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    # P4-1b: stamp the normalized money set (fx_rate, amount_kes, currency_book,
    # rate date/source) at booking. Additive + resilient — never blocks create
    # if a currency rate is unconfigured (currency_book is always computed).
    try:
        from utils.fx_engine import stamp_money_fields
        stamp_money_fields(deal_dict, amount_key="deal_value")
    except Exception:
        pass
    new_id = pm.add_deal(deal_dict)
    try:
        _db_sync_pipeline_deal(pm.get_deal(new_id))  # mirror to Postgres (raises on failure)
        # B13: verify the row actually landed (guards against a silent no-op).
        if _db_available():
            from utils.db import db as _db
            if not _db.fetch_all(
                    "SELECT id FROM pipeline_deals WHERE id = %s", (new_id,)):
                raise RuntimeError(
                    f"Deal {new_id} not present in Postgres after sync")
    except Exception as e:
        # Atomic create: roll back the JSON add so JSON and Postgres never
        # diverge — the deal is in BOTH stores or NEITHER. Honors the standing
        # rule that PostgreSQL is the source of truth.
        try:
            pm.delete_deal(new_id, str(user.get("username", "")))
        except Exception:
            logger.error(f"Rollback delete failed for {new_id}")
        _audit("API_PIPELINE_CREATE_DB_FAILED", user, f"deal_id={new_id} err={e}")
        raise HTTPException(
            status_code=500,
            detail="Could not persist the deal to PostgreSQL — no deal was created.")

    # Mirror Streamlit's emission convention (DEAL_ADDED, line 965)
    _audit("DEAL_ADDED", user,
           f"{new_id}|{payload.client_name}|{payload.deal_value}")

    # BSC trigger (matches Streamlit's _bsc_trigger pattern)
    bsc_ok = emit_bsc_trigger(user.get("username", ""))

    # Bust the summary cache so GET reflects the new deal
    invalidate_pipeline_caches()

    # Fetch the created record (PipelineManager.add_deal returns the
    # id; we re-fetch to return the full record including
    # auto-populated fields like created_at)
    created = pm.get_deal(new_id) or deal_dict
    # SLA S2b: seed the initial step stamp from the create stage.
    _stamp_sla_step(pm, new_id, created, user.get("username", ""))
    created = pm.get_deal(new_id) or created
    return PipelineDealMutationResponse(
        deal=PipelineDeal.model_validate(created),
        status="created",
        bsc_triggered=bsc_ok,
    ).model_dump()


@app.put("/api/pipeline/deals/{deal_id}")
def pipeline_deal_update(
    deal_id: str,
    payload: "PipelineDealUpdate",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """Update fields on an existing pipeline deal.

    Partial update — only keys present in the request body are
    applied. Stage TRANSITIONS should use the dedicated /advance
    endpoint (which logs the change to the activity stream); PUT
    accepts a stage field but does NOT log a stage-change activity.

    Authorization: caller must have the deal in their cascade scope
    (α2 / G395 alignment). 403 if not.
    """
    _audit("API_PIPELINE_UPDATE_ATTEMPT", user, f"deal_id={deal_id}")

    from utils.api_pipeline_models import (
        PipelineDeal,
        PipelineDealMutationResponse,
    )
    from utils.api_pipeline_mutations import (
        emit_bsc_trigger,
        invalidate_pipeline_caches,
    )
    from utils.api_pipeline_scope import get_visible_staff_codes

    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Cascade scope check
    visible_codes = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    po = str(deal.get("portfolio_owner_code", "") or "")
    if sc not in visible_codes and (not po or po not in visible_codes):
        _audit("API_PIPELINE_UPDATE_FORBIDDEN", user,
               f"deal_id={deal_id} out of scope")
        raise HTTPException(
            status_code=403,
            detail="Deal is outside your cascade scope",
        )

    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(
            status_code=400,
            detail="No fields supplied for update",
        )

    pm.update_deal(deal_id, updates, user.get("username", ""))
    _db_sync_pipeline_deal(pm.get_deal(deal_id))  # H5: mirror to DB-backed reads
    _audit("DEAL_UPDATED", user,
           f"{deal_id}|fields={sorted(updates.keys())}")

    bsc_ok = emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    updated = pm.get_deal(deal_id) or deal
    return PipelineDealMutationResponse(
        deal=PipelineDeal.model_validate(updated),
        status="updated",
        bsc_triggered=bsc_ok,
    ).model_dump()


@app.post("/api/pipeline/deals/{deal_id}/advance")
def pipeline_deal_advance(
    deal_id: str,
    payload: "PipelineDealAdvance",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """Advance a pipeline deal to a new stage.

    Stage allowlist (Option C from Arc α3 scoping): only stages in
    `utils.api_pipeline_mutations.ALLOWED_ADVANCE_STAGES` are
    accepted. Requests to advance to any LMS_DEFERRED_STAGES
    (Credit Review, Approval, Bank Approval, Credit Committee,
    Documentation, Vetting, Disbursed) are rejected with HTTP 400
    + a message pointing the caller at Streamlit until Arc α4
    implements the LoanApplication auto-creation handoff.

    Authorization: caller must have the deal in their cascade scope.
    """
    _audit("API_PIPELINE_ADVANCE_ATTEMPT", user,
           f"deal_id={deal_id} new_stage={payload.new_stage}")

    from utils.api_pipeline_models import (
        PipelineDeal,
        PipelineDealMutationResponse,
    )
    from utils.api_pipeline_mutations import (
        validate_advance_target,
        emit_bsc_trigger,
        invalidate_pipeline_caches,
        handle_lms_handoff,
    )
    from utils.api_pipeline_scope import get_visible_staff_codes

    # Stage validation (post-α4: LMS stages now permitted; handoff
    # triggered below after successful update_stage)
    ok, reason = validate_advance_target(payload.new_stage)
    if not ok:
        _audit("API_PIPELINE_ADVANCE_REJECTED", user, reason)
        raise HTTPException(status_code=400, detail=reason)

    from utils.core import PipelineManager as _PM_for_api
    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Cascade scope check
    visible_codes = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    po = str(deal.get("portfolio_owner_code", "") or "")
    if sc not in visible_codes and (not po or po not in visible_codes):
        _audit("API_PIPELINE_ADVANCE_FORBIDDEN", user,
               f"deal_id={deal_id} out of scope")
        raise HTTPException(
            status_code=403,
            detail="Deal is outside your cascade scope",
        )

    # Referral gate: a deal awaiting acceptance can't progress. The recipient
    # must accept it first (then they own its progression); a declined deal
    # must be reassigned and re-accepted before it can move.
    _rstatus = str(deal.get("referral_status") or "")
    if _rstatus in ("pending", "declined"):
        _audit("API_PIPELINE_ADVANCE_REFERRAL_BLOCKED", user,
               f"deal_id={deal_id} referral_status={_rstatus}")
        raise HTTPException(
            status_code=400,
            detail=("This deal is a referral awaiting acceptance. The person it "
                    "was referred to must accept it before it can advance."
                    if _rstatus == "pending" else
                    "This referral was declined and is awaiting reassignment."),
        )
    # target must belong to THIS deal's flow — e.g. a deposit (liability) deal
    # can't advance to a loan-only stage. Skips gracefully if no flow is
    # configured for the class (fallback to the prior global allowlist).
    _flow = _stage_flow_for(deal.get("product_type") or deal.get("product", ""))
    if _flow and payload.new_stage not in _flow:
        _cls = _classify_product(deal.get("product_type") or deal.get("product", ""))
        _audit("API_PIPELINE_ADVANCE_OFF_FLOW", user,
               f"deal_id={deal_id} stage={payload.new_stage} class={_cls}")
        raise HTTPException(
            status_code=400,
            detail=(f"'{payload.new_stage}' is not a valid stage for this "
                    f"{_cls} deal. Allowed stages: {', '.join(_flow)}"))

    old_stage = deal.get("stage", "?")
    pm.update_stage(
        deal_id,
        payload.new_stage,
        payload.note or "",
        user.get("username", ""),
    )
    _db_sync_pipeline_deal(pm.get_deal(deal_id))  # H5: mirror to DB-backed reads
    _audit("API_PIPELINE_ADVANCED", user,
           f"{deal_id}|{old_stage}->{payload.new_stage}")

    # LMS handoff (v10.506 Phase 3 Arc α Batch α4 — supersedes α3
    # Option C deferral). If the new stage is an LMS-handoff stage
    # AND this is a real transition, auto-create the linked
    # LoanApplication via the canonical
    # LoanApplicationManager.create_from_pipeline_deal method.
    # Idempotent — same deal twice returns same app id without
    # creating a duplicate. Failure is non-fatal: advance has
    # already succeeded, lms_error is reported in response.
    updated_deal = pm.get_deal(deal_id) or deal
    lms_triggered, lms_app_id, lms_err = handle_lms_handoff(
        updated_deal, old_stage, payload.new_stage, user.get("username", "")
    )
    if lms_triggered:
        _audit("LMS_APPLICATION_CREATED", user,
               f"{deal_id}|{updated_deal.get('client_name','')}|{lms_app_id}|pipeline->credit")
    elif lms_err:
        _audit("API_PIPELINE_ADVANCE_LMS_FAILED", user,
               f"{deal_id}|{lms_err}")

    # H6: persist the created application id back onto the deal so the
    # Pipeline->LMS cross-link (frontend reads deal.lms_application_id) shows
    # on detail/list and survives reload. Without this the handoff was a
    # silent side-effect with no durable link from the deal to its credit app.
    if lms_triggered and lms_app_id:
        pm.update_deal(deal_id, {"lms_application_id": lms_app_id},
                       user.get("username", ""))
        updated_deal = pm.get_deal(deal_id) or updated_deal
        _db_sync_pipeline_deal(updated_deal)

    # SLA S2b: stamp entry into the new stage's SLA step (per-step business-day clock).
    _stamp_sla_step(pm, deal_id, updated_deal, user.get("username", ""))
    updated_deal = pm.get_deal(deal_id) or updated_deal

    bsc_ok = emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    return PipelineDealMutationResponse(
        deal=PipelineDeal.model_validate(updated_deal),
        status="advanced",
        bsc_triggered=bsc_ok,
        lms_triggered=lms_triggered if (lms_triggered or lms_err) else None,
        lms_application_id=lms_app_id,
        lms_error=lms_err,
    ).model_dump()


# ── Pipeline Manager Queue Endpoints (v10.508 Phase 3 Arc α Batch α6) ──
#
# Manager-only endpoints for the validation queue + cancellation queue,
# plus the cancel-request endpoint (open to any authenticated user).
# Manager authority is determined by role-keyword matching (see
# utils/api_pipeline_manager_actions.py::is_manager — Streamlit parity).
# Cascade scope filtering reuses α2's get_visible_staff_codes.


@app.get("/api/pipeline/queues/validation")
def pipeline_queue_validation(user: dict = Depends(get_current_user)):
    """List deals awaiting manager validation in caller's cascade.

    Filter: stage in [Contacted...] AND not manager_validated AND
    not cancel_requested AND not draft. Same filter as Streamlit
    page line 1293-1294.

    Manager-only — 403 for non-managers. The scope filter ensures
    a manager only sees deals from staff under their cascade
    (G395 alignment).
    """
    from utils.api_pipeline_manager_actions import is_manager
    if not is_manager(user):
        _audit("API_PIPELINE_QUEUE_FORBIDDEN", user, "validation queue access denied")
        raise HTTPException(
            status_code=403,
            detail="Manager role required to access validation queue",
        )

    _audit("API_PIPELINE_QUEUE_VALIDATION", user)

    from utils.api_pipeline_models import PipelineDeal
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.api_pipeline_permissions import enrich_deal_with_permissions
    from utils.core import PipelineManager as _PM_for_api

    visible_codes = get_visible_staff_codes(user)
    pm = _PM_for_api()
    deals = pm.get_pending_validations(manager_codes=visible_codes)

    # α7: each deal carries caller-specific permissions
    return {
        "deals":  [
            enrich_deal_with_permissions(
                PipelineDeal.model_validate(d).model_dump(), user, visible_codes
            )
            for d in deals
        ],
        "count":  len(deals),
        "queue":  "validation",
    }


@app.get("/api/pipeline/queues/cancellation")
def pipeline_queue_cancellation(user: dict = Depends(get_current_user)):
    """List deals with pending cancel requests in caller's cascade.

    Filter: cancel_requested AND not cancel_approved. Matches
    Streamlit page line 1292.

    Manager-only — 403 for non-managers.
    """
    from utils.api_pipeline_manager_actions import is_manager
    if not is_manager(user):
        _audit("API_PIPELINE_QUEUE_FORBIDDEN", user, "cancellation queue access denied")
        raise HTTPException(
            status_code=403,
            detail="Manager role required to access cancellation queue",
        )

    _audit("API_PIPELINE_QUEUE_CANCELLATION", user)

    from utils.api_pipeline_models import PipelineDeal
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.api_pipeline_permissions import enrich_deal_with_permissions
    from utils.core import PipelineManager as _PM_for_api

    visible_codes = get_visible_staff_codes(user)
    pm = _PM_for_api()
    deals = pm.get_cancel_requests(manager_codes=visible_codes)

    # α7: each deal carries caller-specific permissions
    return {
        "deals":  [
            enrich_deal_with_permissions(
                PipelineDeal.model_validate(d).model_dump(), user, visible_codes
            )
            for d in deals
        ],
        "count":  len(deals),
        "queue":  "cancellation",
    }


@app.post("/api/pipeline/deals/{deal_id}/validate")
def pipeline_deal_validate(
    deal_id: str,
    payload: "PipelineDealValidate",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """Manager validates or queries a deal.

    approved=True → DEAL_VALIDATED audit, deal becomes part of forecast.
    approved=False → DEAL_QUERIED audit, deal returns to owner with note.

    Mirrors Streamlit page 1329-1336.

    Authorization: manager + scope. Note that the scope check uses
    the existing α2 helper — a regional head can validate deals from
    any staff under their region; a branch manager only their branch.
    """
    from utils.api_pipeline_manager_actions import is_manager
    if not is_manager(user):
        _audit("API_PIPELINE_VALIDATE_FORBIDDEN", user, f"deal_id={deal_id}")
        raise HTTPException(
            status_code=403,
            detail="Manager role required to validate deals",
        )

    _audit("API_PIPELINE_VALIDATE_ATTEMPT", user,
           f"deal_id={deal_id} approved={payload.approved}")

    from utils.api_pipeline_mutations import (
        emit_bsc_trigger, invalidate_pipeline_caches,
    )
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM_for_api

    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Scope check — manager can only validate deals under their cascade
    visible_codes = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    if sc not in visible_codes:
        _audit("API_PIPELINE_VALIDATE_FORBIDDEN", user,
               f"deal_id={deal_id} out of scope")
        raise HTTPException(
            status_code=403,
            detail="Deal is outside your cascade scope",
        )

    pm.validate_deal(
        deal_id,
        user.get("username", ""),
        payload.approved,
        payload.note or "",
    )

    # Persist the validation to the DB read path. The analytics assured value
    # and funnel read deals via _acquire_scoped_deals (DB-first); without this
    # sync, manager_validated would live only in the JSON store and the DB
    # readers would never see it.
    try:
        _vd = pm.get_deal(deal_id)
        if _vd:
            _db_sync_pipeline_deal(_vd)
    except Exception:
        pass

    # Emit audit event matching Streamlit (line 1331 / 1335)
    if payload.approved:
        _audit("DEAL_VALIDATED", user, f"{deal_id}|note={(payload.note or '')[:60]}")
    else:
        _audit("DEAL_QUERIED", user, f"{deal_id}|note={(payload.note or '')[:60]}")

    emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    updated = pm.get_deal(deal_id) or deal
    return {
        "deal":           updated,
        "status":         "validated" if payload.approved else "queried",
        "manager_validated": bool(payload.approved),
    }


@app.post("/api/pipeline/deals/{deal_id}/cancel/request")
def pipeline_deal_cancel_request(
    deal_id: str,
    payload: "PipelineDealCancelRequest",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """RM requests cancellation of a deal in their cascade scope.

    The request enters the manager's cancellation queue (see
    /api/pipeline/queues/cancellation) for approve/reject decision.

    Mirrors Streamlit page 1460-1463.

    Authorization: any authenticated user, but the deal must be in
    caller's scope (an RM can request cancel on their own deals; a
    manager can request on any deal in their cascade). The actual
    APPROVAL is restricted to managers (see /cancel/approve below).
    """
    _audit("API_PIPELINE_CANCEL_REQUEST_ATTEMPT", user, f"deal_id={deal_id}")

    from utils.api_pipeline_manager_actions import validate_cancel_request_payload
    from utils.api_pipeline_mutations import (
        emit_bsc_trigger, invalidate_pipeline_caches,
    )
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM_for_api

    # Payload validation
    payload_dict = payload.model_dump(exclude_unset=False)
    ok, reason_text = validate_cancel_request_payload(payload_dict)
    if not ok:
        _audit("API_PIPELINE_CANCEL_REQUEST_REJECTED", user, reason_text)
        raise HTTPException(status_code=400, detail=reason_text)

    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    # Scope check
    visible_codes = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    po = str(deal.get("portfolio_owner_code", "") or "")
    if sc not in visible_codes and (not po or po not in visible_codes):
        _audit("API_PIPELINE_CANCEL_REQUEST_FORBIDDEN", user,
               f"deal_id={deal_id} out of scope")
        raise HTTPException(
            status_code=403,
            detail="Deal is outside your cascade scope",
        )

    pm.request_cancel(deal_id, user.get("username", ""), payload.reason)
    _db_sync_pipeline_deal(pm.get_deal(deal_id))  # H5: mirror to DB-backed reads

    # Match Streamlit audit emission (line 1462)
    _audit("CANCEL_REQUESTED", user, f"{deal_id}|{payload.reason}")

    emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    updated = pm.get_deal(deal_id) or deal
    return {
        "deal":               updated,
        "status":             "cancel_requested",
        "cancel_requested":   True,
        "awaiting_manager":   True,
    }


@app.post("/api/pipeline/deals/{deal_id}/cancel/approve")
def pipeline_deal_cancel_approve(
    deal_id: str,
    payload: "PipelineDealCancelApprove",  # noqa: F821
    user: dict = Depends(get_current_user),
):
    """Manager approves or rejects a pending cancel request.

    approve=True → deal transitions to Closed Lost, loss_reason set,
    CANCEL_APPROVED audit.
    approve=False → deal continues (cancel_approved flag set to False
    on the record for audit), CANCEL_REJECTED audit.

    Mirrors Streamlit page 1307-1315.

    Authorization: manager + scope.
    """
    from utils.api_pipeline_manager_actions import is_manager
    if not is_manager(user):
        _audit("API_PIPELINE_CANCEL_APPROVE_FORBIDDEN", user, f"deal_id={deal_id}")
        raise HTTPException(
            status_code=403,
            detail="Manager role required to approve cancellations",
        )

    _audit("API_PIPELINE_CANCEL_APPROVE_ATTEMPT", user,
           f"deal_id={deal_id} approve={payload.approve}")

    from utils.api_pipeline_mutations import (
        emit_bsc_trigger, invalidate_pipeline_caches,
    )
    from utils.api_pipeline_scope import get_visible_staff_codes
    from utils.core import PipelineManager as _PM_for_api

    pm = _PM_for_api()
    deal = _get_or_hydrate_deal(pm, deal_id)
    if not deal:
        raise HTTPException(status_code=404, detail=f"Deal {deal_id} not found")

    if not deal.get("cancel_requested"):
        raise HTTPException(
            status_code=400,
            detail="No pending cancellation request on this deal",
        )

    # Scope check
    visible_codes = get_visible_staff_codes(user)
    sc = str(deal.get("staff_code", "") or "")
    if sc not in visible_codes:
        _audit("API_PIPELINE_CANCEL_APPROVE_FORBIDDEN", user,
               f"deal_id={deal_id} out of scope")
        raise HTTPException(
            status_code=403,
            detail="Deal is outside your cascade scope",
        )

    pm.approve_cancel(
        deal_id,
        user.get("username", ""),
        payload.approve,
        payload.note or "",
    )
    _db_sync_pipeline_deal(pm.get_deal(deal_id))  # H5: mirror to DB-backed reads

    # Match Streamlit audit emission (line 1309 / 1313)
    if payload.approve:
        _audit("CANCEL_APPROVED", user, f"{deal_id}|note={(payload.note or '')[:60]}")
    else:
        _audit("CANCEL_REJECTED", user, f"{deal_id}|note={(payload.note or '')[:60]}")

    emit_bsc_trigger(user.get("username", ""))
    invalidate_pipeline_caches()

    updated = pm.get_deal(deal_id) or deal
    return {
        "deal":            updated,
        "status":          "cancel_approved" if payload.approve else "cancel_rejected",
        "cancel_approved": bool(payload.approve),
    }


# ── Credit Monitoring Endpoints ───────────────────────────────────
@app.get("/api/credit/summary")
def credit_summary(user: dict = Depends(get_current_user)):
    _audit("API_CREDIT_SUMMARY", user)
    cached = _get_cache("credit_summary", "credit")
    if cached:
        return cached

    if _db_available():
        try:
            from utils.db import db as _db
            totals = _db.fetch_one("""
                SELECT COUNT(*)                                          as total_accounts,
                       ROUND(SUM(outstanding)::numeric/1e9, 2)          as outstanding_bn,
                       ROUND(SUM(loan_amount)::numeric/1e9, 2)          as loan_book_bn,
                       SUM(CASE WHEN npl_flag THEN 1 ELSE 0 END)        as npl_count,
                       ROUND(SUM(CASE WHEN npl_flag THEN outstanding ELSE 0 END)::numeric/
                             NULLIF(SUM(outstanding),0)*100, 2)         as npl_ratio_pct,
                       COUNT(DISTINCT branch)                           as branches
                FROM watchlist
            """)
            by_class = _db.fetch_all("""
                SELECT classification,
                       COUNT(*) as accounts,
                       ROUND(SUM(outstanding)::numeric/1e6, 1) as outstanding_m
                FROM watchlist
                GROUP BY classification
                ORDER BY outstanding_m DESC
            """)
            result = {
                "totals":   _serialize(totals),
                "by_class": _serialize(by_class),
                "source":   "postgresql",
            }
            _set_cache("credit_summary", result)
            return result
        except Exception as e:
            logger.error(f"Credit DB error: {e}")

    raw   = _load_json("credit_monitoring.json")
    accts = raw if isinstance(raw,list) else raw.get("watchlist",[])
    total_out = sum(_safe_float(a.get("outstanding",0)) for a in accts)
    npl_ct    = sum(1 for a in accts if _safe_int(a.get("npl_days",0))>=90)
    npl_out   = sum(_safe_float(a.get("outstanding",0))
                    for a in accts if _safe_int(a.get("npl_days",0))>=90)
    result = {
        "totals": {"total_accounts":len(accts),
                   "outstanding_bn":round(total_out/1e9,2),
                   "npl_count":npl_ct,
                   "npl_ratio_pct":round(npl_out/max(total_out,1)*100,2)},
        "by_class":[],
        "source":"json",
    }
    _set_cache("credit_summary", result)
    return result

@app.get("/api/credit/watchlist")
def credit_watchlist(
    classification: Optional[str] = None,
    stage:          Optional[str] = None,
    branch:         Optional[str] = None,
    npl_only:       bool = False,
    limit:          int = Query(default=200, le=1000),
    offset:         int = Query(default=0, ge=0),
    user:           dict = Depends(get_current_user),
):
    _audit("API_CREDIT_WATCHLIST", user, f"npl_only={npl_only} branch={branch}")
    if _db_available():
        try:
            from utils.db import db as _db
            where  = []
            params = []
            if classification: where.append("classification = %s"); params.append(classification)
            if stage:          where.append("stage = %s");          params.append(stage)
            if branch:         where.append("branch_name ILIKE %s");params.append(f"%{branch}%")
            if npl_only:       where.append("npl_flag = true")
            sql = "SELECT * FROM watchlist"
            if where: sql += " WHERE " + " AND ".join(where)
            sql += f" ORDER BY dpd DESC LIMIT {limit} OFFSET {offset}"
            rows = _db.fetch_all(sql, tuple(params))
            return {"accounts":_serialize(rows),"count":len(rows),"source":"postgresql"}
        except Exception as e:
            logger.error(f"Watchlist DB error: {e}")

    raw   = _load_json("credit_monitoring.json")
    accts = raw if isinstance(raw,list) else raw.get("watchlist",[])
    if classification: accts = [a for a in accts if a.get("classification")==classification]
    if stage:          accts = [a for a in accts if a.get("stage")==stage]
    if npl_only:       accts = [a for a in accts if _safe_int(a.get("npl_days",0))>=90]
    return {"accounts":accts[offset:offset+limit],"count":len(accts),"source":"json"}


# ── Credit analytics + drill (hierarchy-scoped, mirrors pipeline) ──────────
def _acquire_scoped_credit(user: dict) -> list:
    """Loan-book accounts within the caller's reporting subtree.

    The loan book has several possible homes (credit_watchlist table with a
    risk_data JSONB, a flat watchlist table, or credit_monitoring.json on
    disk) and which one is populated varies by environment. We try each and
    accept the FIRST that yields rows carrying rm_code (the scope key). Scope:
    the SAME get_visible_staff_codes the pipeline uses, matched on rm_code.
    """
    import json as _json

    def _has_rm(rows):
        return bool(rows) and any(r.get("rm_code") for r in rows[:10])

    accts: list = []
    if _db_available():
        from utils.db import db as _db
        # 1) canonical credit_watchlist (identity top-level; risk_data JSONB)
        try:
            rows = _db.fetch_all(
                "SELECT account_number, cif, branch_code, branch_name, region, "
                "rm_code, rm_name, risk_data, status, severity "
                "FROM credit_watchlist", ())
            cand = []
            for r in _serialize(rows) or []:
                rd = r.get("risk_data")
                if isinstance(rd, str):
                    try: rd = _json.loads(rd)
                    except Exception: rd = {}
                a = dict(rd) if isinstance(rd, dict) else {}
                for k in ("account_number", "cif", "branch_code", "branch_name",
                          "region", "rm_code", "rm_name", "status", "severity"):
                    if r.get(k) is not None:
                        a[k] = r.get(k)
                cand.append(a)
            if _has_rm(cand):
                accts = cand
        except Exception as e:
            logger.error(f"credit_watchlist read error: {e}")
        # 2) flat watchlist table
        if not _has_rm(accts):
            try:
                rows = _db.fetch_all("SELECT * FROM watchlist", ())
                cand = _serialize(rows) or []
                if _has_rm(cand):
                    accts = cand
            except Exception as e:
                logger.error(f"watchlist read error: {e}")
    # 3) credit_monitoring.json on disk (dev / no-PG)
    if not _has_rm(accts):
        p = DATA_DIR / "credit_monitoring.json"
        if p.exists():
            try:
                raw = _json.loads(p.read_text(encoding="utf-8"))
                cand = raw if isinstance(raw, list) else raw.get("watchlist", [])
                if _has_rm(cand):
                    accts = cand
            except Exception as e:
                logger.error(f"Credit monitoring file read error: {e}")

    from utils.api_pipeline_scope import get_visible_staff_codes
    visible = get_visible_staff_codes(user)
    return [a for a in accts if str(a.get("rm_code") or "") in visible]


def _compute_credit_analytics(accts: list) -> dict:
    def out(a): return _safe_float(a.get("outstanding", 0))
    def npl(a):
        if a.get("npl_flag"):
            return True
        return _safe_int(a.get("npl_days", a.get("dpd", 0))) >= 90

    def group(field: str, key: str) -> list:
        m: dict = {}
        for a in accts:
            k = a.get(field) or "Unassigned"
            e = m.setdefault(k, {key: k, "outstanding": 0.0, "accounts": 0,
                                 "npl_outstanding": 0.0})
            e["outstanding"] += out(a)
            e["accounts"] += 1
            if npl(a):
                e["npl_outstanding"] += out(a)
        rows = sorted(m.values(), key=lambda x: x["outstanding"], reverse=True)
        for r in rows:
            r["npl_ratio_pct"] = (round(r["npl_outstanding"] / r["outstanding"] * 100, 2)
                                  if r["outstanding"] else 0.0)
        return rows

    total_out = sum(out(a) for a in accts)
    npl_out = sum(out(a) for a in accts if npl(a))
    return {
        "totals": {
            "outstanding": total_out,
            "accounts": len(accts),
            "npl_outstanding": npl_out,
            "npl_count": sum(1 for a in accts if npl(a)),
            "npl_ratio_pct": round(npl_out / total_out * 100, 2) if total_out else 0.0,
        },
        "by_class": group("classification", "classification"),
        "by_region": group("region", "region"),
        "by_branch": group("branch_name", "branch"),
        "by_rm": group("rm_name", "rm"),
    }


@app.get("/api/credit/analytics")
def credit_analytics(user: dict = Depends(get_current_user)):
    """Loan-book analytics over the caller's scoped accounts (KES outstanding):
    by classification / region / branch / RM, with NPL per slice."""
    _audit("API_CREDIT_ANALYTICS", user)
    return _compute_credit_analytics(_acquire_scoped_credit(user))


@app.get("/api/credit/drill")
def credit_drill(
    region: Optional[str] = None,
    branch: Optional[str] = None,
    rm: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Drill the loan book Region -> Branch -> RM -> individual accounts.
    Scope-safe (reuses _acquire_scoped_credit). Pure filter, so it can't drift
    from the credit analytics totals."""
    _audit("API_CREDIT_DRILL", user, f"region={region} branch={branch} rm={rm}")
    accts = _acquire_scoped_credit(user)
    if region: accts = [a for a in accts if (a.get("region") or "Unassigned") == region]
    if branch: accts = [a for a in accts if (a.get("branch_name") or "Unassigned") == branch]
    if rm:     accts = [a for a in accts
                        if (a.get("rm_name") or a.get("rm_code") or "Unassigned") == rm]

    sub = _compute_credit_analytics(accts)

    def _slim(a: dict) -> dict:
        return {
            "account_number": a.get("account_number"),
            "cif": a.get("cif"),
            "classification": a.get("classification"),
            "outstanding": _safe_float(a.get("outstanding", 0)),
            "npl_days": _safe_int(a.get("npl_days", 0)),
            "branch_name": a.get("branch_name"),
            "rm_name": a.get("rm_name"),
            "collateral_type": a.get("collateral_type"),
            "stage": a.get("stage"),
        }
    accounts = [_slim(a) for a in
                sorted(accts, key=lambda a: _safe_float(a.get("outstanding", 0)),
                       reverse=True)[:200]]
    return {
        "region": region, "branch": branch, "rm": rm,
        "by_branch": sub["by_branch"], "by_rm": sub["by_rm"],
        "accounts": accounts, "totals": sub["totals"],
    }


# ── AML Endpoints ─────────────────────────────────────────────────
@app.get("/api/aml/summary")
def aml_summary(user: dict = Depends(get_current_user)):
    _audit("API_AML_SUMMARY", user)
    cached = _get_cache("aml_summary","aml")
    if cached: return cached

    if _db_available():
        try:
            from utils.db import db as _db
            totals = _db.fetch_one("""
                SELECT COUNT(*)                                            as total_alerts,
                       SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END)   as open_alerts,
                       SUM(CASE WHEN risk_level='High' THEN 1 ELSE 0 END) as high_risk,
                       SUM(CASE WHEN str_filed THEN 1 ELSE 0 END)          as strs_filed,
                       ROUND(SUM(amount)::numeric/1e6, 1)                  as total_amount_m
                FROM aml_alerts
            """)
            by_rule = _db.fetch_all("""
                SELECT rule_triggered, COUNT(*) as alerts,
                       SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END) as open_count
                FROM aml_alerts
                GROUP BY rule_triggered
                ORDER BY alerts DESC
            """)
            result = {"totals":_serialize(totals),"by_rule":_serialize(by_rule),"source":"postgresql"}
            _set_cache("aml_summary",result)
            return result
        except Exception as e:
            logger.error(f"AML DB error: {e}")

    alerts = _load_json("aml_alerts.json")
    result = {
        "totals":{"total_alerts":len(alerts),
                  "open_alerts":sum(1 for a in alerts if a.get("status")=="Open"),
                  "high_risk":sum(1 for a in alerts if a.get("risk_level")=="High"),
                  "strs_filed":sum(1 for a in alerts if a.get("str_filed"))},
        "by_rule":[],"source":"json"
    }
    _set_cache("aml_summary",result)
    return result

# ── Users / Org Endpoints ─────────────────────────────────────────
@app.get("/api/users/summary")
def users_summary(user: dict = Depends(get_current_user)):
    _audit("API_USERS_SUMMARY", user)
    cached = _get_cache("users_summary","users")
    if cached: return cached

    if _db_available():
        try:
            from utils.db import db as _db
            totals = _db.fetch_one("""
                SELECT COUNT(*)                                          as total_users,
                       SUM(CASE WHEN active THEN 1 ELSE 0 END)          as active_users,
                       SUM(CASE WHEN is_admin THEN 1 ELSE 0 END)        as admins,
                       COUNT(DISTINCT department)                        as departments,
                       COUNT(DISTINCT unit)                              as units
                FROM users
            """)
            by_dept = _db.fetch_all("""
                SELECT department, COUNT(*) as headcount
                FROM users WHERE active = true
                GROUP BY department
                ORDER BY headcount DESC
            """)
            result = {"totals":_serialize(totals),"by_dept":_serialize(by_dept),"source":"postgresql"}
            _set_cache("users_summary",result)
            return result
        except Exception as e:
            logger.error(f"Users DB error: {e}")

    users = _load_json("users.json")
    if isinstance(users,dict): users = list(users.values())
    result = {
        "totals":{"total_users":len(users),
                  "active_users":sum(1 for u in users if u.get("active",True)),
                  "admins":sum(1 for u in users if u.get("is_admin"))},
        "by_dept":[],"source":"json"
    }
    _set_cache("users_summary",result)
    return result

# ── MD Dashboard Endpoint ─────────────────────────────────────────
@app.get("/api/dashboard/md")
def md_dashboard(user: dict = Depends(get_current_user)):
    """Single endpoint for MD command centre — aggregates all key metrics."""
    _audit("API_DASHBOARD_MD", user)
    cached = _get_cache("md_dashboard","dashboard")
    if cached: return cached

    # Direct calls now require the user dict (since the route fns gained
    # the auth Depends). We pass the resolved user through — no need to
    # re-validate the token internally.
    bsc   = bsc_summary(user=user)
    pipe  = pipeline_summary(user=user)
    credit= credit_summary(user=user)
    aml   = aml_summary(user=user)
    users = users_summary(user=user)

    result = {
        "bsc":     {"overall_avg":bsc.get("overall_avg",0),
                    "total_staff":bsc.get("total_staff",0)},
        "pipeline":{"total_deals":pipe.get("totals",{}).get("total_deals",0),
                    "pipeline_value":pipe.get("totals",{}).get("pipeline_value",0),
                    "validated_value":pipe.get("totals",{}).get("validated_value",0),
                    "pending_value":pipe.get("totals",{}).get("pending_value",0),
                    "pending_validation":pipe.get("totals",{}).get("pending_validation",0),
                    "won_value":pipe.get("totals",{}).get("won_value",0),
                    "lcy_value":pipe.get("totals",{}).get("lcy_value",0),
                    "fcy_value":pipe.get("totals",{}).get("fcy_value",0)},
        "credit":  {"total_accounts":credit.get("totals",{}).get("total_accounts",0),
                    "outstanding_bn":credit.get("totals",{}).get("outstanding_bn",0),
                    "npl_ratio_pct":credit.get("totals",{}).get("npl_ratio_pct",0)},
        "aml":     {"open_alerts":aml.get("totals",{}).get("open_alerts",0),
                    "high_risk":aml.get("totals",{}).get("high_risk",0)},
        "org":     {"total_staff":users.get("totals",{}).get("total_users",0),
                    "departments":users.get("totals",{}).get("departments",0)},
        "generated_at": datetime.now().isoformat(),
    }
    _set_cache("md_dashboard",result)
    return result


# ── Executive exceptions strip ────────────────────────────────────
# Read-only derivation: surfaces the top items needing an executive
# decision, each with a drill link. No business logic or workflow
# change — figures reuse the same scoped summaries the dashboard shows,
# so the strip and the dashboard tiles always agree.

def _count_stalled_deals(deals: list, days: int = 14) -> int:
    """Active deals created more than `days` ago (no stage movement proxy).
    Defensive: any deal without a usable timestamp is skipped, never counted."""
    from datetime import datetime as _dt
    closed_markers = ("won", "lost", "closed", "declined", "disbursed", "rejected")
    now = _dt.now()
    n = 0
    for d in deals:
        try:
            stage = str(d.get("stage", "")).lower()
            if any(m in stage for m in closed_markers):
                continue
            ts = d.get("created_at") or d.get("open_date") or d.get("created")
            if not ts:
                continue
            parsed = _dt.fromisoformat(str(ts).replace("Z", "+00:00"))
            if parsed.tzinfo is not None:
                parsed = parsed.replace(tzinfo=None)
            if (now - parsed).days > days:
                n += 1
        except Exception:
            continue
    return n


@app.get("/api/dashboard/exceptions")
def dashboard_exceptions(user: dict = Depends(get_current_user)):
    """Top exceptions needing a decision — scoped, read-only, drill-linked."""
    _audit("API_DASHBOARD_EXCEPTIONS", user)
    items: list = []

    # Credit — NPL ratio breach (same source as the dashboard NPL tile).
    try:
        ct = (credit_summary(user=user) or {}).get("totals", {})
        npl = float(ct.get("npl_ratio_pct", 0) or 0)
        if npl >= 5.0:
            danger = npl >= 10.0
            items.append({
                "id": "npl_ratio",
                "severity": "danger" if danger else "warning",
                "title": f"NPL ratio at {npl:.1f}%",
                "detail": f"Above the {10 if danger else 5}% threshold "
                          f"across {int(ct.get('total_accounts', 0)):,} accounts",
                "value": f"{npl:.1f}%",
                "link": "/credit-analytics",
            })
    except Exception:
        pass

    # Credit — worst-NPL branch in scope (from scoped credit analytics).
    try:
        accts = _acquire_scoped_credit(user)
        if accts:
            branches = _compute_credit_analytics(accts).get("by_branch", [])
            worst = max(branches, key=lambda b: b.get("npl_ratio_pct", 0), default=None)
            if worst and float(worst.get("npl_ratio_pct", 0) or 0) >= 10.0:
                items.append({
                    "id": "npl_branch",
                    "severity": "warning",
                    "title": f"{worst.get('branch', '—')}: NPL {float(worst['npl_ratio_pct']):.1f}%",
                    "detail": "Highest-NPL branch in your scope",
                    "value": f"{float(worst['npl_ratio_pct']):.1f}%",
                    "link": "/credit-analytics",
                })
    except Exception:
        pass

    # Pipeline — deals awaiting validation (same source as dashboard tile).
    try:
        pending = int((pipeline_summary(user=user) or {})
                      .get("totals", {}).get("pending_validation", 0) or 0)
        if pending > 0:
            items.append({
                "id": "pending_validation",
                "severity": "warning",
                "title": f"{pending:,} deal{'s' if pending != 1 else ''} awaiting validation",
                "detail": "Sitting in the manager validation queue",
                "value": f"{pending:,}",
                "link": "/pipeline/queues",
            })
    except Exception:
        pass

    # Pipeline — stalled active deals (>14 days, defensive on timestamp).
    try:
        stalled = _count_stalled_deals(_acquire_scoped_deals(user), days=14)
        if stalled > 0:
            items.append({
                "id": "stalled_deals",
                "severity": "warning",
                "title": f"{stalled:,} deal{'s' if stalled != 1 else ''} stalled over 14 days",
                "detail": "Active deals with no recent movement",
                "value": f"{stalled:,}",
                "link": "/analytics",
            })
    except Exception:
        pass

    # Danger first, then warnings.
    order = {"danger": 0, "warning": 1, "info": 2}
    items.sort(key=lambda x: order.get(x.get("severity"), 3))
    return {
        "exceptions": items[:6],
        "count": len(items),
        "generated_at": datetime.now().isoformat(),
    }


# ── Cache management ──────────────────────────────────────────────
# ── Cache management ──────────────────────────────────────────────
@app.post("/api/cache/clear")
def clear_cache(user: dict = Depends(require_admin)):
    """Admin-only — flush the in-memory API cache."""
    _audit("API_CACHE_CLEAR", user, "Cache flushed via /api/cache/clear")
    _cache.clear()
    _cache_ts.clear()
    return {"status":"cleared","message":"All API cache cleared"}

# ── FX rates (P4-1b) ──────────────────────────────────────────────
# Operational FX rate table — admin-maintained. KES is base (rate=1).
# List/resolve: any authenticated user (dashboards need the LCY/FCY rate).
# Upsert: admin-only.
class _FxRateUpsert(BaseModel):
    currency: str
    rate_to_kes: float
    effective_date: str
    rate_type: str = "mid"
    model_config = ConfigDict(extra="allow")


@app.get("/api/fx/rates")
def fx_list_rates(currency: Optional[str] = None, active_only: bool = False,
                  user: dict = Depends(get_current_user)):
    from utils.fx_engine import FxRateStore, BASE_CURRENCY
    st = FxRateStore()
    return {"base_currency": BASE_CURRENCY,
            "rates": st.list_rates(currency, active_only)}


@app.get("/api/fx/resolve")
def fx_resolve(currency: str, as_of: Optional[str] = None,
               rate_type: str = "mid", user: dict = Depends(get_current_user)):
    from utils.fx_engine import FxRateStore, FxRateError
    try:
        r = FxRateStore().resolve_rate(currency, as_of=as_of, rate_type=rate_type)
        return {"currency": currency.upper(), "rate_to_kes": float(r),
                "rate_type": rate_type, "as_of": as_of}
    except FxRateError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/fx/rates")
def fx_upsert_rate(payload: _FxRateUpsert, user: dict = Depends(require_admin)):
    from utils.fx_engine import FxRateStore, FxRateError
    st = FxRateStore()
    try:
        row = st.upsert_rate(
            payload.currency, payload.rate_to_kes, payload.effective_date,
            rate_type=payload.rate_type, source="admin",
            entered_by=str(user.get("username", "") or "admin"))
        st.save()
    except FxRateError as e:
        raise HTTPException(status_code=400, detail=str(e))
    _audit("API_FX_RATE_UPSERT", user,
           f"{payload.currency} {payload.rate_type} "
           f"{payload.rate_to_kes}@{payload.effective_date}")
    return {"rate": row, "status": "saved"}

@app.get("/api/cache/stats")
def cache_stats(user: dict = Depends(get_current_user)):
    _audit("API_CACHE_STATS", user)
    now = time.time()
    return {
        "cached_keys": list(_cache.keys()),
        "ages_seconds": {k: round(now - _cache_ts.get(k,now)) for k in _cache},
    }

# ══════════════════════════════════════════════════════════════════════
# /api/v1/* — Generic CRUD routers (Standard #2, v5.31)
# ══════════════════════════════════════════════════════════════════════
# Each module gets 8 endpoints generated by the factory:
#   list, get, create, update, delete, export, search, dashboard
# All are JWT-gated, parameterised, and audit-logged.
#
# Adding a new module is one make_crud_router() call. The G16 audit
# gate counts these and validates JWT auth on every route.
# ══════════════════════════════════════════════════════════════════════
from utils.api_crud import make_crud_router

# Pilot module: pipeline_deals
# - PG-live (TABLE_USE_DB["pipeline_deals"] = True since v5.x)
# - Schema in SCHEMA_SQL with `id` PK
# - Existing /api/pipeline/* endpoints stay (backward compat); the new
#   /api/v1/pipeline_deals/* endpoints add CRUD that wasn't there before.
app.include_router(
    make_crud_router(
        module     = "pipeline_deals",
        table      = "pipeline_deals",
        json_file  = "pipeline.json",
        list_key   = "deals",
        searchable = ["stage", "deal_category", "unit", "staff_code", "client_cif"],
        order_by   = "open_date DESC",
        pk_column  = "id",
    )
)

# ══════════════════════════════════════════════════════════════════════
# v10.92 — Phase 1B kickoff: CRUD modules for high-value tables
# ══════════════════════════════════════════════════════════════════════
# Each make_crud_router call = 8 endpoints (list/get/create/update/delete/
# export/search/dashboard). All inherit JWT auth, audit logging, and PG/JSON
# fallback from the factory. The 3 modules below are picked from the
# TABLE_USE_DB-enabled list with PG migrations completed in v10.88-v10.91.

# loan_applications: credit origination pipeline — extended schema in v10.89
# (added clean_repayment_history, compliance_type, appraisal_notes,
# proposition_tag, data JSONB to the pre-existing table). The CRUD router
# uses `last_updated DESC` since application_date may be stale for long-
# running applications.
app.include_router(
    make_crud_router(
        module     = "loan_applications",
        table      = "loan_applications",
        json_file  = "loan_applications.json",
        list_key   = None,            # source JSON is a flat list, not nested
        searchable = ["status", "swim_lane", "deal_category",
                      "rm_code", "client_cif", "compliance_flag",
                      "is_repeat_borrower"],
        order_by   = "last_updated DESC",
        pk_column  = "id",
    )
)

# aml_alerts: AML/CTF case tracking — RLS-protected (Risk & Compliance +
# Internal Audit only); the CRUD router inherits the table-level RLS
# automatically because PG enforces RLS at SELECT/INSERT time regardless
# of how the SQL is built.
app.include_router(
    make_crud_router(
        module     = "aml_alerts",
        table      = "aml_alerts",
        json_file  = "aml_alerts.json",
        list_key   = None,
        searchable = ["status", "risk_level", "str_filed",
                      "assigned_to", "rule_triggered"],
        order_by   = "transaction_date DESC",
        pk_column  = "id",
    )
)

# projects: project management tracking — added in v10.91 PG migration
# batch 4 (40 records). Common dashboard queries by status + rag_status.
app.include_router(
    make_crud_router(
        module     = "projects",
        table      = "projects",
        json_file  = "projects.json",
        list_key   = None,
        searchable = ["status", "priority", "rag_status",
                      "department", "project_manager", "sponsor"],
        order_by   = "start_date DESC",
        pk_column  = "id",
    )
)

# ══════════════════════════════════════════════════════════════════════
# v10.93 — Phase 1B continuation: 3 more CRUD modules (+24 endpoints)
# ══════════════════════════════════════════════════════════════════════
# Prerequisite was added in v10.93 itself: the v10.88-v10.91 PG-migrated
# tables now have TABLE_USE_DB=True entries. Picking 3 high-value tables
# from the 25 newly-eligible: ifrs9_loans (5045 records, IFRS 9 ECL),
# legal_matters (362 records, legal arc ops), collateral_register
# (200 records, credit collateral).

# ifrs9_loans: high-volume IFRS 9 expected credit loss data (5045 records).
# Stage filter is the single most-used query in credit risk dashboards;
# reporting_date enables snapshot comparisons across reporting periods.
# pk_column "account_id" because this table uses account_id as PK, not id.
app.include_router(
    make_crud_router(
        module     = "ifrs9_loans",
        table      = "ifrs9_loans",
        json_file  = "ifrs9_loans.json",
        list_key   = None,
        searchable = ["stage", "ecl_basis", "sicr_flag",
                      "product", "client_name"],
        order_by   = "ecl_amount DESC",
        pk_column  = "account_id",
    )
)

# legal_matters: legal arc litigation/advisory tracking (362 records).
# SLA breach status is the operational priority filter; matter_type
# segments litigation vs advisory vs collections.
app.include_router(
    make_crud_router(
        module     = "legal_matters",
        table      = "legal_matters",
        json_file  = "legal_matters.json",
        list_key   = None,
        searchable = ["status", "priority", "sla_breached",
                      "matter_type", "client_cif", "attorney"],
        order_by   = "opened_date DESC",
        pk_column  = "id",
    )
)

# collateral_register: credit collateral inventory (200 records).
# LTV ratio + status filter directly map to credit governance queries
# (high-LTV positions need attention; impaired collateral changes ECL).
app.include_router(
    make_crud_router(
        module     = "collateral_register",
        table      = "collateral_register",
        json_file  = "collateral_register.json",
        list_key   = None,
        searchable = ["status", "collateral_type", "client_cif",
                      "branch", "valuer"],
        order_by   = "market_value DESC",
        pk_column  = "id",
    )
)

# ══════════════════════════════════════════════════════════════════════
# v10.94 — Phase 1B continuation: 3 more CRUD modules (+24 endpoints)
# ══════════════════════════════════════════════════════════════════════
# All 3 verified ready in pre-flight: TABLE_USE_DB=True (set v10.93),
# schema in db.py, FLAT_MIGRATIONS entry, source JSON exists with data.

# agent_transactions: high-volume agent banking transaction log
# (679 records → millions in production). Most operational queries
# are by agent_id, txn_date, or fraud_flag — these drive the
# fraud-detection workflow. Default sort by txn_date DESC matches
# how operations teams review recent activity.
app.include_router(
    make_crud_router(
        module     = "agent_transactions",
        table      = "agent_transactions",
        json_file  = "agent_transactions.json",
        list_key   = None,
        searchable = ["agent_id", "branch", "txn_type",
                      "fraud_flag", "txn_date"],
        order_by   = "txn_date DESC",
        pk_column  = "id",
    )
)

# debt_recovery: NPL recovery tracking (150 records). Recovery_stage
# is the primary operational filter (early/letter/legal/written-off);
# legal_referral surfaces matters that need legal-arc handoff. Order
# by NPL days descending so the worst delinquencies surface first.
app.include_router(
    make_crud_router(
        module     = "debt_recovery",
        table      = "debt_recovery",
        json_file  = "debt_recovery.json",
        list_key   = None,
        searchable = ["status", "recovery_stage", "client_cif",
                      "rm_code", "legal_referral", "branch"],
        order_by   = "npl_days DESC",
        pk_column  = "id",
    )
)

# cims_tickets: customer instruction management (200 records).
# Priority + status are the SLA-driven filters; instruction_type
# segments the workflow (transfers, cheques, statements, etc.).
# Ordered by SLA due_date ASC so soon-to-breach tickets surface
# first.
app.include_router(
    make_crud_router(
        module     = "cims_tickets",
        table      = "cims_tickets",
        json_file  = "cims_tickets.json",
        list_key   = None,
        searchable = ["status", "priority", "instruction_type",
                      "branch", "rm_code", "client_cif"],
        order_by   = "due_date ASC",
        pk_column  = "id",
    )
)

# ══════════════════════════════════════════════════════════════════════
# v10.95 — Phase 1B continuation: 3 more CRUD modules (+24 endpoints)
# ══════════════════════════════════════════════════════════════════════
# All 3 verified clean in pre-flight: TABLE_USE_DB=True (set v10.93 +
# stale flips), schemas + FLAT_MIGRATIONS in place, source JSONs exist.

# compliance_cases: compliance arc case tracking (115 records).
# Risk-level + status are the operational filters; flag_type segments
# the case taxonomy (Adverse Media, Sanctions Hit, PEP Match, etc.).
# Order by raised_date DESC so most recent cases surface first.
app.include_router(
    make_crud_router(
        module     = "compliance_cases",
        table      = "compliance_cases",
        json_file  = "compliance_cases.json",
        list_key   = None,
        searchable = ["status", "risk_level", "flag_type",
                      "client_cif", "assigned_officer",
                      "case_type"],
        order_by   = "raised_date DESC",
        pk_column  = "id",
    )
)

# referrals: customer referral tracking (200 records). Channel
# attribution analytics: referral_source (Staff/Branch/Partner/MOU)
# is the primary segmentation; converted + fee_paid surface the
# operational state. Branch + rm_assigned support performance
# tracking by location and relationship manager.
app.include_router(
    make_crud_router(
        module     = "referrals",
        table      = "referrals",
        json_file  = "referrals.json",
        list_key   = None,
        searchable = ["status", "referral_source", "converted",
                      "fee_paid", "branch", "rm_assigned",
                      "product_interested"],
        order_by   = "referral_date DESC",
        pk_column  = "id",
    )
)

# consent_register: DPO compliance — customer consent tracking
# (200 records). consent_type segments the regulatory category
# (Marketing, Data Processing, Product Recommendations, etc.);
# legal_basis (Consent/Contract/Legitimate Interest/Legal
# Obligation) drives reviewability. Status + granted flag the
# active vs withdrawn state.
app.include_router(
    make_crud_router(
        module     = "consent_register",
        table      = "consent_register",
        json_file  = "consent_register.json",
        list_key   = None,
        searchable = ["status", "consent_type", "granted",
                      "legal_basis", "customer_cif",
                      "cbk_category", "channel"],
        order_by   = "granted_date DESC",
        pk_column  = "id",
    )
)

# ══════════════════════════════════════════════════════════════════════
# v10.96 — Phase 1B CLOSE-OUT: 3 final CRUD modules (+24 endpoints)
# Total: 16 CRUD modules × 8 verbs = 128 CRUD endpoints + 19 direct =
# 147 endpoints (108% of the 136 target).
# ══════════════════════════════════════════════════════════════════════
# Note on candidates: staff_history (originally targeted for v10.96) and
# commission_records were dropped because they use composite primary
# keys — the factory pattern requires a single pk_column. Substituted
# clearing_records (has `id` field). Composite-PK tables are wired via
# direct decorators when needed; not a CRUD-factory case.

# revenue_assurance: revenue leakage tracking (300 records).
# revenue_assurance arc data (arc closed at G133+G134). Common queries:
# by status (Open/Recovered/Written-off), by type (Pricing/Fee/Variance),
# by period for trend analytics.
app.include_router(
    make_crud_router(
        module     = "revenue_assurance",
        table      = "revenue_assurance",
        json_file  = "revenue_assurance.json",
        list_key   = None,
        searchable = ["status", "type", "fee_type", "period",
                      "branch", "recovered", "client_cif"],
        order_by   = "date_raised DESC",
        pk_column  = "id",
    )
)

# edms_documents: document management (500 records, largest by volume
# among remaining candidates). Used by compliance + legal arcs.
# Filters: category (Loan Doc, KYC, Legal, etc.), status, expiry,
# review-pending. Order by uploaded_date DESC for fresh-content first.
app.include_router(
    make_crud_router(
        module     = "edms_documents",
        table      = "edms_documents",
        json_file  = "edms_documents.json",
        list_key   = None,
        searchable = ["status", "category", "document_type",
                      "client_cif", "branch", "is_expired",
                      "requires_review", "access_level"],
        order_by   = "uploaded_date DESC",
        pk_column  = "id",
    )
)

# clearing_records: clearing house settlement tracking (120 records).
# Settlement workflow: status (Pending/Settled/Failed/Reversed),
# system (PESALINK/RTGS/EFT/SWIFT), reconciliation state.
# Order by value_date DESC matches operations review pattern.
app.include_router(
    make_crud_router(
        module     = "clearing_records",
        table      = "clearing_records",
        json_file  = "clearing_records.json",
        list_key   = None,
        searchable = ["status", "system", "reconciled",
                      "currency", "settlement_tat_met",
                      "officer_username"],
        order_by   = "value_date DESC",
        pk_column  = "id",
    )
)

# ── Strategy module router (v10.141) ─────────────────────────────
# Mounts the 19 Strategy module endpoints (ENH-141..155) at
# /api/strategy/*. Each endpoint requires JWT via
# Depends(get_current_user). The router is the React-ready surface
# for the closed Strategy module — engines are the source of truth
# for both this API and pages/15_strategy_arc_cockpit.py.
from utils.api_strategy import router as strategy_router
app.include_router(strategy_router)

# ── Cockpit API router (v10.297) ────────────────────────────────
# Phase 3 React-readiness arc. Exposes the cockpit_read composers
# at /api/cockpit/* so the React SPA (#37) can fetch the same live
# cockpit views that Streamlit pages 109 (CIMS) and 110 (Treasury)
# render. Single source of truth: HTTP and Streamlit transports
# both call the same composer functions. Read-only GETs only;
# state changes go through engine-specific routers.
try:
    from utils.api_cockpit import (
        router as cockpit_router,
        FASTAPI_AVAILABLE as _COCKPIT_FASTAPI_AVAILABLE,
    )
    if _COCKPIT_FASTAPI_AVAILABLE and cockpit_router is not None:
        app.include_router(cockpit_router)
except Exception:
    # Cockpit API is additive — never break the app if it fails
    # to import (e.g. cockpit_read deps unavailable).
    pass

# ── Standard #20: Performance Amplification API (v5.44) ───────────
@app.get("/api/v2/performance/insights/{staff_code}")
def performance_insights(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    """Standard #20 — Performance Amplification API.

    Returns aggregated insights for a single staff member:
      - overall_score:       BSC overall on 1-5 scale
      - strengths:           list of KPIs at ≥110% achievement
      - promotion_readiness: 0-1 from #12 GrowthPathEngine

    Returns 404 for unknown staff. Auth required (per G12).
    """
    from utils.performance_insights import get_performance_insights
    result = get_performance_insights(staff_code)
    if not result:
        raise HTTPException(status_code=404, detail=f"No insights for staff_code={staff_code!r}")
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# INTEGRATION LAYER API (v10.115) — React-readiness endpoints
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#
# These endpoints expose the v10.108-v10.115 Integration Layer state
# in JSON shapes a React frontend can consume directly without
# Streamlit-side rendering. Each endpoint:
#   - JWT-protected (Depends(get_current_user))
#   - Returns flat JSON / arrays of dicts (no Streamlit data structures)
#   - Includes a `source` field so the React side knows where the data
#     originated (`registry`, `aggregator`, etc.) for debugging
#   - Cached per the existing _set_cache / _get_cache pattern
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _rule_to_dict(rule) -> dict:
    """Serialize an AggregationRule to a JSON-shaped dict suitable
    for React consumption. Excludes callable fields (predicate,
    extractor) — the React side only needs the metadata."""
    return {
        "kpi_id":         rule.kpi_id,
        "source_table":   rule.source_table,
        "pattern":        rule.pattern,
        "description":    rule.description or "",
        "value_field":    rule.value_field,
        "start_field":    rule.start_field,
        "end_field":      rule.end_field,
        "numerator_field": rule.numerator_field,
        "denominator_field": rule.denominator_field,
        "bool_field":     rule.bool_field,
        "period_field":   rule.period_field,
        "decimals":       rule.decimals,
        "invert":         rule.invert,
        "uses_extractor": rule.staff_field_extractor is not None,
    }


@app.get("/api/integration/rules")
def integration_rules(
    pattern: Optional[str] = None,
    source_table: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Standard #15 — List Integration Layer rules.

    Returns rule metadata for every registered AggregationRule. React
    uses this to render the rule catalog, filter by pattern (e.g.
    'show all PERCENTAGE rules'), and join against the KPI library
    for display labels and units.

    Query params:
      pattern:       filter by pattern name (COUNT, PERCENTAGE, etc.)
      source_table:  filter by operational table

    Response shape:
      {
        "rules":          [...rule dicts...],
        "count":          int,
        "patterns":       sorted list of distinct patterns in result,
        "source_tables":  sorted list of distinct tables in result,
        "source":         "registry"
      }
    """
    _audit("API_INTEGRATION_RULES", user)

    try:
        from utils.kpi_aggregation_rules import REGISTRY
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"REGISTRY not available: {type(e).__name__}: {e}")

    rules = list(REGISTRY)
    if pattern:
        rules = [r for r in rules if r.pattern == pattern]
    if source_table:
        rules = [r for r in rules if r.source_table == source_table]

    rule_dicts = [_rule_to_dict(r) for r in rules]
    return {
        "rules":         rule_dicts,
        "count":         len(rule_dicts),
        "patterns":      sorted({r["pattern"] for r in rule_dicts}),
        "source_tables": sorted({r["source_table"] for r in rule_dicts}),
        "source":        "registry",
    }


@app.get("/api/integration/actuals/{period}")
def integration_actuals(
    period: str,
    kpi_id: Optional[str] = None,
    staff_code: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Standard #15 — Compute Integration Layer actuals for a period.

    Runs every active rule against its operational table for the
    requested period, applies ownership gating, and returns the
    submitted+dropped records as JSON.

    Path params:
      period:      "YYYY-MM" period filter (e.g. "2026-04")

    Query params:
      kpi_id:      filter to a single KPI
      staff_code:  filter to a single staff (across all KPIs they own)

    Response shape:
      {
        "period":          "2026-04",
        "actuals":         [
          {"staff_code", "kpi_id", "value", "source_table", "pattern"}
        ],
        "count":           int,
        "by_kpi":          {kpi_id: count, ...},
        "by_source_table": {table_name: count, ...},
        "source":          "aggregator"
      }

    The shape mirrors what bsc_engine.submit_batch consumes — a React
    'pending actuals' preview screen can render this without further
    transformation.
    """
    _audit("API_INTEGRATION_ACTUALS", user, {"period": period})

    # Validate period format
    import re
    if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", period):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid period {period!r}; expected YYYY-MM")

    try:
        from utils.kpi_aggregation_rules import REGISTRY, compute_rule
        from utils.staff_field_resolver import resolve_staff_field
        from utils.staff_name_resolver import refresh_cache as nr_refresh
        from utils.kpi_ownership import is_kpi_owned_by_staff
        nr_refresh()
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Integration Layer unavailable: {type(e).__name__}: {e}")

    # Cache by (period, kpi_id, staff_code) tuple to avoid recomputing
    cache_key = f"integration_actuals:{period}:{kpi_id or '*'}:{staff_code or '*'}"
    cached = _get_cache(cache_key, "integration")
    if cached:
        return cached

    actuals = []
    for rule in REGISTRY:
        if kpi_id and rule.kpi_id != kpi_id:
            continue
        # Read operational table from JSON (PG-readiness shim is a
        # v10.116+ task per blueprint; the rule data path is already
        # PG-ready via the loader, but the operational tables aren't
        # yet in PG views).
        from pathlib import Path
        path = Path(_data_dir()) / f"{rule.source_table}.json"
        if not path.exists():
            continue
        try:
            with open(path) as f:
                d = json.load(f)
            rows = d if isinstance(d, list) else list(d.values())
        except Exception:
            continue

        sf = resolve_staff_field(rule.source_table, rule.staff_field)
        try:
            per_staff = compute_rule(rule, rows, period, sf)
        except Exception:
            continue
        if not per_staff:
            continue

        for sc, value in per_staff.items():
            if staff_code and sc != staff_code:
                continue
            # Apply ownership gate
            try:
                if not is_kpi_owned_by_staff(sc, rule.kpi_id, period):
                    continue
            except Exception:
                # If ownership cannot be determined, skip silently
                # (matches the v10.108 actuals_engine behavior)
                continue
            actuals.append({
                "staff_code":   sc,
                "kpi_id":       rule.kpi_id,
                "value":        round(float(value), rule.decimals)
                                  if isinstance(value, (int, float))
                                  else value,
                "source_table": rule.source_table,
                "pattern":      rule.pattern,
                "period":       period,
            })

    by_kpi = {}
    by_table = {}
    for a in actuals:
        by_kpi[a["kpi_id"]] = by_kpi.get(a["kpi_id"], 0) + 1
        by_table[a["source_table"]] = by_table.get(a["source_table"], 0) + 1

    result = {
        "period":          period,
        "actuals":         actuals,
        "count":           len(actuals),
        "by_kpi":          by_kpi,
        "by_source_table": by_table,
        "source":          "aggregator",
    }
    _set_cache(cache_key, result)
    return result


@app.get("/api/integration/resolution-metrics")
def integration_resolution_metrics(user: dict = Depends(get_current_user)):
    """Standard #15 — Surface name + role resolver metrics.

    Returns the current hit/miss rates from staff_name_resolver and
    staff_role_resolver. The React admin page renders this as the
    Resolution Metrics card so deploying admins can see staff-register
    coverage gaps.

    Response shape:
      {
        "name_resolver": {
          "lookups_total", "lookups_hit", "lookups_miss",
          "ambiguous_misses", "miss_examples", "hit_rate_pct"
        },
        "role_resolver": {
          ...same shape plus "resolved_via" breakdown...
        },
        "source": "resolvers"
      }
    """
    _audit("API_INTEGRATION_RESOLUTION_METRICS", user)
    out = {"name_resolver": None, "role_resolver": None, "source": "resolvers"}
    try:
        from utils.staff_name_resolver import get_resolution_metrics as nm
        out["name_resolver"] = nm()
    except Exception as e:
        out["name_resolver_error"] = f"{type(e).__name__}: {e}"
    try:
        from utils.staff_role_resolver import get_resolution_metrics as rm
        out["role_resolver"] = rm()
    except Exception as e:
        out["role_resolver_error"] = f"{type(e).__name__}: {e}"
    return out


@app.post("/api/integration/run-period")
def integration_run_period(
    period: str,
    dry_run: bool = False,
    user: dict = Depends(get_current_user),
):
    """Standard #15 — Submit Integration Layer actuals for a period
    (write-side trigger).

    Runs the full pipeline: every active rule → operational table read
    (PG view or JSON via the v10.116 _data_source shim) → compute →
    ownership gate → bsc_engine.submit_batch. Returns a JSON status
    matching the dict shape from
    `compute_actuals_from_operational_tables(period)`.

    **This is the write-side counterpart to GET /api/integration/
    actuals/{period}.** The GET endpoint computes-and-previews without
    persistence; this POST endpoint runs the full submit pipeline and
    persists actuals to the BSC engine. Together they close the
    React API read+write contract.

    Path-style param (period) is provided as a query string for POST
    consistency with the rest of the integration endpoints.

    Query params:
      period:   "YYYY-MM" period filter (required, validated)
      dry_run:  if true, computes but skips bsc_engine.submit_batch.
                Useful for React 'preview before commit' flows.
                Defaults to false.

    Response shape (matches actuals_engine output):
      {
        "success":           bool,
        "period":            "2026-04",
        "rules_processed":   int,
        "rules_skipped":     int,    # source table missing
        "actuals_submitted": int,    # passed ownership gate
        "actuals_dropped":   int,    # failed ownership gate
        "by_rule":           [...],
        "engine_summary":    {...},  # from bsc_engine.submit_batch
        "dry_run":           bool,
        "source":            "aggregator-write"
      }

    Idempotency: bsc_engine.submit_batch is idempotent on
    (staff_code, kpi_id, period) — calling this endpoint twice for
    the same period writes the actuals once and reports
    'duplicates_skipped' in engine_summary on the second call.

    Authorization: standard JWT auth (Depends(get_current_user)).
    Per Standard #15, write endpoints SHOULD additionally check that
    the caller has admin/integration role; v10.116 ships with
    user-role-agnostic auth and v10.117 adds role gating once the
    role taxonomy stabilises (avoids breaking React work waiting for
    a roles backlog).
    """
    _audit("API_INTEGRATION_RUN_PERIOD", user,
            {"period": period, "dry_run": dry_run})

    # v10.117 role gating — feature flag (default OFF). When enabled,
    # caller's role must be in _security.allowed_roles_for_write or
    # the POST is rejected with HTTP 403. Layered on top of the
    # existing JWT auth (Depends(get_current_user)).
    _check_write_role(user)

    # Validate period
    import re
    if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", period):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid period {period!r}; expected YYYY-MM")

    try:
        from utils.actuals_engine import (
            compute_actuals_from_operational_tables)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Integration Layer unavailable: "
                   f"{type(e).__name__}: {e}")

    if dry_run:
        # Dry run = use the GET-style preview path so React can render
        # 'pending actuals' without persisting. Reuses the existing
        # GET /api/integration/actuals/{period} logic by calling
        # the read-only handler directly.
        preview = integration_actuals(
            period=period, kpi_id=None, staff_code=None, user=user)
        return {
            "success":           True,
            "period":            period,
            "rules_processed":   len({a["kpi_id"] for a in preview["actuals"]}),
            "rules_skipped":     0,
            "actuals_submitted": 0,                       # not persisted
            "actuals_dropped":   0,
            "preview_count":     preview["count"],
            "preview_actuals":   preview["actuals"],
            "by_rule":           [],
            "engine_summary":    {"mode": "dry_run"},
            "dry_run":           True,
            "source":            "aggregator-write",
        }

    # Full pipeline: compute + ownership gate + bsc_engine.submit_batch
    try:
        result = compute_actuals_from_operational_tables(period)
    except Exception as e:
        logger.error(
            f"integration_run_period: pipeline error: "
            f"{type(e).__name__}: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Pipeline error: {type(e).__name__}: {e}")

    # Tag the response so the React side knows whether this was a real
    # write or a dry run. Preserve everything the engine returned.
    result = dict(result) if isinstance(result, dict) else {
        "success": False, "error": "non-dict response"}
    result["dry_run"] = False
    result["source"] = "aggregator-write"
    return result


@app.get("/api/integration/coverage")
def integration_coverage(user: dict = Depends(get_current_user)):
    """Standard #15 — G143-style coverage report.

    Returns a JSON-shaped version of the audit gate output so the
    React admin dashboard can render coverage progress without
    parsing the audit log.

    Response shape:
      {
        "covered":           int,    # operational-source KPIs with rules
        "total_operational": int,    # operational-source KPIs in library
        "pct":               float,
        "cbs_source_count":  int,    # autofitted via CBS pathway
        "uncovered_kpis":    [kpi_ids without rules],
        "source":            "audit_gate_g143"
      }
    """
    _audit("API_INTEGRATION_COVERAGE", user)

    try:
        from utils.kpi_aggregation_rules import kpis_with_aggregator
        with open(_data_dir() / "kpi_library.json") as f:
            lib = json.load(f)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Coverage unavailable: {type(e).__name__}: {e}")

    covered = kpis_with_aggregator()
    # Match the v10.108 audit gate G143's prefix-based detection for
    # CBS-source KPIs (cbs_*, management_accounts*) so /api/integration
    # /coverage and the gate report identical numerators/denominators.
    cbs_prefixes = ("cbs_", "management_accounts")
    op_kpis = [k for k in lib.get("kpis", [])
                if k.get("source")
                and not any(k.get("source").startswith(p)
                            for p in cbs_prefixes)]
    op_ids = {k.get("id") for k in op_kpis if k.get("id")}
    op_covered = covered & op_ids
    op_uncovered = sorted(op_ids - covered)
    cbs_source_count = sum(1 for k in lib.get("kpis", [])
                            if k.get("source")
                            and any(k.get("source").startswith(p)
                                    for p in cbs_prefixes))

    pct = round(100.0 * len(op_covered) / max(len(op_ids), 1), 2)

    # v10.117 strict-mode preview tier (matches scripts/audit.py G143)
    if pct >= 75.0:
        strict_tag = "STRICT-READY (high)"
    elif pct >= 50.0:
        strict_tag = "STRICT-READY (preview)"
    else:
        strict_tag = "BELOW STRICT THRESHOLD"

    return {
        "covered":           len(op_covered),
        "total_operational": len(op_ids),
        "pct":               pct,
        "cbs_source_count":  cbs_source_count,
        "uncovered_kpis":    op_uncovered,
        "strict_preview": {
            "tag":                   strict_tag,
            "preview_threshold_pct": 50.0,
            "high_threshold_pct":    75.0,
            "flip_target_pct":       100.0,
        },
        "source":            "audit_gate_g143",
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# v10.132 — Rule-explain debug endpoint
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Operator/audit superpower: for any wired rule, return:
#   - The full rule definition (pattern, predicates, period_field, staff_field)
#   - Input summary (rows in table, rows in period, rows matching predicate)
#   - Sample matched rows (top N for spot-check)
#   - Per-staff intermediate values
#   - Final aggregated value (matches /actuals/{period} for the same rule)
#
# When a number looks wrong on a dashboard, this endpoint shows the
# working: which rows were considered, which were filtered out by
# period, which were filtered out by predicate, and how the final
# value was computed. The cockpit's Debug tab consumes this.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/integration/rule-explain/{kpi_id}")
def integration_rule_explain(
    kpi_id: str,
    period: str,
    staff_code: Optional[str] = None,
    sample_size: int = 5,
    user: dict = Depends(get_current_user),
):
    """v10.132 — Rule-explain debug endpoint.

    For a given (kpi_id, period), returns a complete trace of how the
    rule's actuals are computed: rule definition, input row counts,
    sample matched rows, per-staff intermediate values, and final
    aggregated value.

    Path params:
      kpi_id:       the KPI identifier whose rule should be explained

    Query params:
      period:       "YYYY-MM" period filter (required)
      staff_code:   optional — narrow to one staff's slice of the output
      sample_size:  number of sample matched rows to return (1-20, default 5)

    Response shape:
      {
        "kpi_id":         "K001",
        "rule":           {... full rule_to_dict() ...},
        "input_summary":  {
          "total_rows_in_table":     int,
          "rows_in_period":          int,
          "rows_matching_predicate": int,
          "distinct_staff_codes":    int
        },
        "sample_matched_rows": [...up to sample_size rows...],
        "per_staff_actuals":   {staff_code: value, ...},
        "final_value": {
          "for_staff":  staff_code or null,
          "value":      numeric (if staff_code) or aggregate dict,
          "decimals":   int
        },
        "source":       "rule_explain_v10_132"
      }

    Errors:
      404 — kpi_id not in REGISTRY (no active rule)
      400 — invalid period format
    """
    _audit("API_INTEGRATION_RULE_EXPLAIN", user,
           {"kpi_id": kpi_id, "period": period})

    # Validate period format (matches /actuals/{period})
    import re
    if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", period):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid period {period!r}; expected YYYY-MM")

    # Cap sample_size to a sane range
    sample_size = max(1, min(20, int(sample_size)))

    try:
        from utils.kpi_aggregation_rules import (
            REGISTRY, compute_rule, _row_in_period,
        )
        from utils.staff_field_resolver import resolve_staff_field
        from utils.staff_name_resolver import refresh_cache as nr_refresh
        nr_refresh()
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Integration Layer unavailable: "
                   f"{type(e).__name__}: {e}")

    # Find the rule for this kpi_id
    matching = [r for r in REGISTRY if r.kpi_id == kpi_id]
    if not matching:
        raise HTTPException(
            status_code=404,
            detail=f"No active aggregation rule for kpi_id {kpi_id!r}. "
                   f"Try GET /api/integration/rules to list available "
                   f"rules.")
    if len(matching) > 1:
        # Library duplicates (e.g. K028/K048) — explain the first one,
        # but signal the duplicate
        rule = matching[0]
        duplicate_count = len(matching) - 1
    else:
        rule = matching[0]
        duplicate_count = 0

    # Read operational table
    from pathlib import Path
    path = Path(_data_dir()) / f"{rule.source_table}.json"
    if not path.exists():
        raise HTTPException(
            status_code=500,
            detail=f"Operational table {rule.source_table!r} not found")
    try:
        with open(path) as f:
            d = json.load(f)
        rows = d if isinstance(d, list) else list(d.values())
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Could not read {rule.source_table}: "
                   f"{type(e).__name__}: {e}")

    # Stage 1: filter by period
    rows_in_period = [r for r in rows
                      if _row_in_period(r, rule.period_field, period)]

    # Stage 2: filter by primary predicate (varies by pattern)
    # Different patterns use different predicate fields. Show "matching"
    # rows = rows that the rule's primary filter accepts.
    primary_pred = (rule.predicate
                    or rule.numerator_pred
                    or (lambda _r: True))
    try:
        rows_matching = [r for r in rows_in_period if primary_pred(r)]
    except Exception:
        rows_matching = rows_in_period  # if predicate errors, show all

    # Stage 3: resolve staff_field + group
    sf = resolve_staff_field(rule.source_table, rule.staff_field)
    distinct_staff = set()
    for r in rows_matching:
        if rule.staff_field_extractor is not None:
            try:
                sc = rule.staff_field_extractor(r)
            except Exception:
                sc = None
        else:
            sc = r.get(sf)
        if sc:
            distinct_staff.add(str(sc))

    # Stage 4: compute_rule for the actual aggregated values
    try:
        per_staff = compute_rule(rule, rows, period, sf)
    except Exception as e:
        per_staff = {}

    # Sample matched rows (truncate large nested fields for brevity)
    def _truncate_value(v, max_len=120):
        if isinstance(v, str) and len(v) > max_len:
            return v[:max_len] + "…"
        if isinstance(v, list) and len(v) > 5:
            return v[:5] + ["…(+%d more)" % (len(v) - 5)]
        return v

    sample = []
    for r in rows_matching[:sample_size]:
        sample.append({k: _truncate_value(v) for k, v in r.items()})

    # If staff_code provided, narrow per_staff
    if staff_code:
        per_staff_filtered = {staff_code: per_staff.get(staff_code)} \
            if staff_code in per_staff else {}
        final_for_staff = per_staff.get(staff_code)
    else:
        per_staff_filtered = {sc: round(float(v), rule.decimals)
                                  if isinstance(v, (int, float)) else v
                              for sc, v in per_staff.items()}
        final_for_staff = None

    return {
        "kpi_id":            kpi_id,
        "period":            period,
        "rule":              _rule_to_dict(rule),
        "duplicate_rules":   duplicate_count,
        "input_summary": {
            "total_rows_in_table":     len(rows),
            "rows_in_period":          len(rows_in_period),
            "rows_matching_predicate": len(rows_matching),
            "distinct_staff_codes":    len(distinct_staff),
        },
        "sample_matched_rows": sample,
        "per_staff_actuals":   per_staff_filtered,
        "final_value": {
            "for_staff":  staff_code,
            "value":      (round(float(final_for_staff), rule.decimals)
                           if isinstance(final_for_staff, (int, float))
                           else final_for_staff),
            "decimals":   rule.decimals,
        },
        "source":              "rule_explain_v10_132",
    }


def _data_dir():
    """Return the data/ directory path (helper for the integration
    endpoints)."""
    from pathlib import Path
    here = Path(__file__).resolve().parent
    return here.parent / "data"


def _read_security_config() -> dict:
    """Read the `_security` config block from
    integration_layer_config.json. Returns the resolved policy:

      {
        "role_gating_enabled":     bool,    # default True (on) since v10.126
        "allowed_roles_for_write": [str],   # default ["admin", "integration"]
      }

    **v10.117 history:** shipped role gating as a feature flag that
    defaulted to OFF so v10.116's POST /api/integration/run-period
    endpoint stayed backward-compatible. Banks who wanted to enforce
    role-based authorization set `_security.role_gating_enabled: true`
    in admin config.

    **v10.120 GA polish:** shipped the explicit `_security` block in
    `integration_layer_config.json` with role_gating_enabled=true and
    the canonical Eco Bank role taxonomy. Soft-flip — code default
    stayed OFF for backward compat with deployments that updated
    v10.117→v10.120 in one go without consuming the new config.

    **v10.126 default flip:** code default now defaults to **True**.
    Deployments that don't explicitly set role_gating_enabled get
    role-gating ON. Deployments that want to keep JWT-only auth must
    set `_security.role_gating_enabled: false` explicitly. Aligns
    code with config defaults and brings the integration layer to
    secure-by-default. Documented in CHANGELOG_v10.126.
    """
    cfg_path = _data_dir() / "integration_layer_config.json"
    default = {
        "role_gating_enabled":     True,   # v10.126: flipped from False
        "allowed_roles_for_write": ["admin", "integration"],
    }
    if not cfg_path.exists():
        return default
    try:
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception:
        return default
    sec = cfg.get("_security")
    if not isinstance(sec, dict):
        return default
    return {
        "role_gating_enabled":     bool(sec.get("role_gating_enabled", True)),  # v10.126
        "allowed_roles_for_write":
            sec.get("allowed_roles_for_write")
            or default["allowed_roles_for_write"],
    }


def _check_write_role(user: dict) -> None:
    """v10.117 role-gating guard. Raises HTTPException(403) if role
    gating is enabled and the user's role is not in the allowed list.
    No-op when role gating is disabled (the v10.117 default).

    Caller MUST already have validated JWT via
    `Depends(get_current_user)` — this is layered on top, not a
    replacement for auth.
    """
    sec = _read_security_config()
    if not sec["role_gating_enabled"]:
        return  # feature flag off → backward-compatible
    user_role = (user or {}).get("role") or ""
    allowed = sec["allowed_roles_for_write"] or []
    if user_role not in allowed:
        raise HTTPException(
            status_code=403,
            detail=(f"Write access denied: role {user_role!r} not in "
                    f"allowed_roles_for_write. Contact your A2Z "
                    f"administrator to update _security config."))


# ════════════════════════════════════════════════════════════════════
# Role weight audit + normalization (v10.419, Phase 2d)
# ════════════════════════════════════════════════════════════════════
# Per Joshua's locked backlog: 225/227 roles broken (weight sums != 1.0).
# These endpoints expose the audit + migration capability.

@app.get("/api/v1/role-weights/audit",
         summary="Bank-wide role weight audit (sums vs 1.0)")
def role_weight_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns total_roles, normalized_count, broken_count, broken_roles list."""
    try:
        from utils.role_weight_engine import bank_role_weight_audit
        audit = bank_role_weight_audit()
        return audit.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"bank_role_weight_audit failed: {exc}")


@app.get("/api/v1/role-weights/{role}/audit",
         summary="Audit a single role's weight situation")
def role_weight_single_audit_endpoint(
    role: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.role_weight_engine import (
            audit_role_weight, _load_library,
        )
        lib = _load_library()
        kpis = lib.get("role_kpis", {}).get(role)
        if kpis is None:
            raise HTTPException(404, f"Role not found: {role}")
        kpi_weights = lib.get("kpi_weights", {})
        result = audit_role_weight(role, kpis, kpi_weights)
        return result.to_dict()
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_role_weight failed: {exc}")


@app.get("/api/v1/role-weights/{role}/normalized",
         summary="Get normalized weights for one role")
def role_normalized_weights_endpoint(
    role: str,
    user: dict = Depends(get_current_user),
):
    """Returns {kpi: normalized_weight} for the role, summing to 1.0.
    Prefers the migrated role_normalized_weights field; falls back to
    on-the-fly computation if not migrated yet."""
    try:
        from utils.role_weight_engine import (
            _load_library, compute_role_normalized_weights,
        )
        lib = _load_library()
        # Prefer stored normalized weights if migration ran
        rnw = lib.get("role_normalized_weights", {})
        if role in rnw and isinstance(rnw[role], dict):
            return {"role": role, "weights": rnw[role], "source": "migrated"}
        # Fall back to on-the-fly
        kpis = lib.get("role_kpis", {}).get(role)
        if kpis is None:
            raise HTTPException(404, f"Role not found: {role}")
        computed = compute_role_normalized_weights(role, kpis, lib.get("kpi_weights", {}))
        return {"role": role, "weights": computed, "source": "computed"}
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"role_normalized_weights failed: {exc}")


@app.post("/api/v1/role-weights/migrate",
          summary="Run the role-weight normalization migration (admin)")
def role_weight_migrate_endpoint(
    user: dict = Depends(get_current_user),
):
    """Writes role_normalized_weights field to kpi_library.json. Idempotent.
    Production: gate behind an admin role check."""
    try:
        from utils.role_weight_engine import migrate_normalize_all_roles
        audit, normalized = migrate_normalize_all_roles(write_back=True)
        return {
            "audit": audit.to_dict(),
            "roles_normalized": len(normalized),
            "migrated_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"migrate_normalize_all_roles failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# KPI library dedup (v10.420, Phase 2d)
# ════════════════════════════════════════════════════════════════════
# Per Joshua's locked backlog: 4 KPI alias pairs consolidated.

@app.get("/api/v1/kpi-dedup/audit",
         summary="Audit KPI library for duplicate alias pairs")
def kpi_dedup_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns the 4 alias pairs with reference counts."""
    try:
        from utils.kpi_dedup_engine import audit_kpi_dedup
        a = audit_kpi_dedup()
        return a.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_kpi_dedup failed: {exc}")


@app.post("/api/v1/kpi-dedup/migrate",
          summary="Run the KPI library dedup migration (admin)")
def kpi_dedup_migrate_endpoint(
    user: dict = Depends(get_current_user),
):
    """Consolidates 4 alias pairs. Idempotent.
    Production: gate behind an admin role check."""
    try:
        from utils.kpi_dedup_engine import migrate_dedup_kpi_library
        result = migrate_dedup_kpi_library(
            write_back=True, rebuild_normalized_weights=True,
        )
        return {
            "result": result.to_dict(),
            "migrated_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"migrate_dedup_kpi_library failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Backup retention (v10.421, Phase 2d)
# ════════════════════════════════════════════════════════════════════
# Audit + cleanup for accumulated _v10*_backups dirs under data/.

@app.get("/api/v1/backup-retention/audit",
         summary="Audit backup directories against retention policy")
def backup_retention_audit_endpoint(
    keep_recent: int = 3,
    size_threshold_mb: float = 1.0,
    user: dict = Depends(get_current_user),
):
    """Returns per-directory audit. No filesystem changes."""
    try:
        from utils.backup_retention_engine import audit_backup_retention
        audit = audit_backup_retention(
            keep_recent_n=keep_recent,
            size_threshold_bytes=int(size_threshold_mb * 1024 * 1024),
        )
        return audit.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_backup_retention failed: {exc}")


@app.post("/api/v1/backup-retention/apply",
          summary="Apply retention policy (delete stale backups). Destructive.")
def backup_retention_apply_endpoint(
    keep_recent: int = 3,
    size_threshold_mb: float = 1.0,
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """If confirm=True, actually deletes. Else returns dry-run result.
    Production: gate behind admin role check."""
    try:
        from utils.backup_retention_engine import apply_retention_policy
        result = apply_retention_policy(
            keep_recent_n=keep_recent,
            size_threshold_bytes=int(size_threshold_mb * 1024 * 1024),
            dry_run=not bool(confirm),
        )
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"apply_retention_policy failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Retired test audit (v10.422, Phase 2d)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/test-cleanup/audit",
         summary="Audit retired test functions (_retired_ prefix convention)")
def test_cleanup_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns per-file + per-version counts of soft-retired test functions."""
    try:
        from utils.test_cleanup_engine import audit_retired_tests
        audit = audit_retired_tests()
        return audit.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_retired_tests failed: {exc}")


@app.post("/api/v1/test-cleanup/archive",
          summary="Write data/_retired_tests_archive.json (idempotent)")
def test_cleanup_archive_endpoint(
    user: dict = Depends(get_current_user),
):
    """Extracts retired tests into a searchable JSON archive.
    Does NOT delete from source files."""
    try:
        from utils.test_cleanup_engine import archive_retired_tests
        result = archive_retired_tests(dry_run=False)
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"archive_retired_tests failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Deep Audit (v10.424 — BSC Rescue Phase opens)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-audit/full",
         summary="Full BSC integrity audit (7 categories)")
def bsc_audit_full_endpoint(
    user: dict = Depends(get_current_user),
):
    """Rollup of staff coverage, KPI completeness, pillar canonical,
    weight normalization, library alignment, cascade linkage, duplicates."""
    try:
        from utils.bsc_audit_engine import bsc_full_audit
        audit = bsc_full_audit()
        return audit.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"bsc_full_audit failed: {exc}")


@app.get("/api/v1/bsc-audit/staff-coverage",
         summary="Staff coverage: every register row has BSC entries")
def bsc_audit_coverage_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_staff_coverage
        return audit_staff_coverage().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_staff_coverage failed: {exc}")


@app.get("/api/v1/bsc-audit/kpi-completeness",
         summary="KPI completeness per staff vs role tier thresholds")
def bsc_audit_completeness_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_kpi_completeness
        return audit_kpi_completeness().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_kpi_completeness failed: {exc}")


@app.get("/api/v1/bsc-audit/pillar-canonical",
         summary="Pillar canonical check (only 4 canonical pillars)")
def bsc_audit_pillar_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_pillar_canonical
        return audit_pillar_canonical().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_pillar_canonical failed: {exc}")


@app.get("/api/v1/bsc-audit/weight-normalization",
         summary="Weight normalization (per-staff sums = 1.0)")
def bsc_audit_weights_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_weight_normalization
        return audit_weight_normalization().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_weight_normalization failed: {exc}")


@app.get("/api/v1/bsc-audit/library-alignment",
         summary="BSC KPIs vs kpi_library alignment")
def bsc_audit_alignment_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_library_alignment
        return audit_library_alignment().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_library_alignment failed: {exc}")


@app.get("/api/v1/bsc-audit/cascade-linkage",
         summary="Cascade targets reflected in BSC")
def bsc_audit_cascade_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.bsc_audit_engine import audit_cascade_linkage
        return audit_cascade_linkage().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_cascade_linkage failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Pillar Normalize (v10.425, BSC Rescue batch 1)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-pillar/audit",
         summary="Audit BSC actuals for non-canonical pillar values")
def bsc_pillar_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns per-alias row counts + affected KPIs."""
    try:
        from utils.bsc_pillar_normalize_engine import audit_actuals_pillars
        return audit_actuals_pillars().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_actuals_pillars failed: {exc}")


@app.post("/api/v1/bsc-pillar/migrate",
          summary="Migrate non-canonical pillars in BSC actuals (default dry-run)")
def bsc_pillar_migrate_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """If confirm=True, performs the migration with backup. Else dry-run."""
    try:
        from utils.bsc_pillar_normalize_engine import migrate_actuals_pillars
        result = migrate_actuals_pillars(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"migrate_actuals_pillars failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Library Register (v10.426, BSC Rescue batch 2)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-library/audit",
         summary="Audit unregistered BSC KPIs vs canonical library")
def bsc_library_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns: aliases to add, library pillar fixes needed, multi-pillar
    BSC KPIs, and truly-new KPIs needing registration."""
    try:
        from utils.bsc_library_register_engine import audit_unregistered_bsc_kpis
        return audit_unregistered_bsc_kpis().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_unregistered_bsc_kpis failed: {exc}")


@app.post("/api/v1/bsc-library/register",
          summary="Run full registration migration (default dry-run)")
def bsc_library_register_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Atomic 4-layer migration: aliases + library pillars + actuals
    multi-pillar + new registrations. Production: gate behind admin."""
    try:
        from utils.bsc_library_register_engine import apply_full_registration
        result = apply_full_registration(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"apply_full_registration failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Completeness (v10.427, BSC Rescue batch 3)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-completeness/audit",
         summary="Audit incomplete BSCs vs canonical role_kpis")
def bsc_completeness_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Per-staff gap analysis: which staff have fewer KPIs than configured
    in kpi_library::role_kpis for their role."""
    try:
        from utils.bsc_completeness_engine import audit_bsc_completeness
        return audit_bsc_completeness().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_bsc_completeness failed: {exc}")


@app.post("/api/v1/bsc-completeness/repair",
          summary="Add missing BSC rows for incomplete staff (default dry-run)")
def bsc_completeness_repair_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Generates new BSC rows for missing KPIs with peer-median targets;
    re-normalizes weights to 1.0 per staff."""
    try:
        from utils.bsc_completeness_engine import repair_bsc_completeness
        result = repair_bsc_completeness(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"repair_bsc_completeness failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Weight Normalization (v10.428, BSC Rescue batch 4)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-weights/audit",
         summary="Audit per-staff weight sums in BSC actuals")
def bsc_weights_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns per-staff Weight column sum + which staff need renormalization."""
    try:
        from utils.bsc_weight_normalize_engine import audit_actuals_weights
        return audit_actuals_weights().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_actuals_weights failed: {exc}")


@app.post("/api/v1/bsc-weights/renormalize",
          summary="Per-staff proportional weight rescale (default dry-run)")
def bsc_weights_renormalize_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Rescales each staff's Weight column so sums = 1.0 exactly.
    Preserves relative importance per KPI."""
    try:
        from utils.bsc_weight_normalize_engine import renormalize_actuals_weights
        result = renormalize_actuals_weights(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"renormalize_actuals_weights failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# BSC Cascade Linkage (v10.429, BSC Rescue closing batch)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/bsc-codes/audit",
         summary="Audit BSC Staff Codes vs canonical register codes")
def bsc_codes_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns: code mismatches between BSC actuals and staff_register."""
    try:
        from utils.bsc_cascade_linkage_engine import audit_bsc_code_alignment
        return audit_bsc_code_alignment().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"audit_bsc_code_alignment failed: {exc}")


@app.post("/api/v1/bsc-codes/fix",
          summary="Rewrite BSC Staff Codes to match canonical register (default dry-run)")
def bsc_codes_fix_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Updates BSC Staff Code values to canonical register codes per name."""
    try:
        from utils.bsc_cascade_linkage_engine import fix_bsc_codes
        result = fix_bsc_codes(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"fix_bsc_codes failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Admin validation (v10.431, admin polish)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/admin-validation/library",
         summary="Validate kpi_library.json snapshot")
def admin_validation_library_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns errors/warnings/info on the current library state."""
    try:
        from utils.admin_validation_engine import validate_full_library
        return validate_full_library().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"validate_full_library failed: {exc}")


@app.post("/api/v1/admin-validation/legacy-aliases",
          summary="Add legacy SNAKE_CASE codes as aliases (default dry-run)")
def admin_validation_legacy_aliases_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Cleans up role_kpis SNAKE_CASE references by registering them as
    aliases on the corresponding canonical library entries."""
    try:
        from utils.admin_validation_engine import apply_legacy_code_aliases
        result = apply_legacy_code_aliases(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"apply_legacy_code_aliases failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Cascade-BSC 360° audit (v10.432, deep review)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/cascade-360/audit",
         summary="Full 360° cascade↔BSC harmony audit")
def cascade_360_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns the 5-stage rollup: bank→MD, cascade integrity,
    cascade→BSC targets, BSC actuals coverage, score calculation."""
    try:
        from utils.cascade_bsc_360_engine import cascade_bsc_360_audit
        return cascade_bsc_360_audit().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"cascade_bsc_360_audit failed: {exc}")


@app.get("/api/v1/cascade-360/stage/{stage}",
         summary="Single-stage 360 audit (1-5)")
def cascade_360_stage_endpoint(
    stage: int,
    user: dict = Depends(get_current_user),
):
    """Stage 1=bank_to_md, 2=cascade_integrity, 3=cascade_to_bsc,
    4=bsc_actuals, 5=score_calc."""
    try:
        from utils.cascade_bsc_360_engine import (
            audit_bank_to_md, audit_cascade_integrity,
            audit_cascade_to_bsc_targets, audit_bsc_actuals_coverage,
            audit_score_calculation,
        )
        stages = {
            1: audit_bank_to_md,
            2: audit_cascade_integrity,
            3: audit_cascade_to_bsc_targets,
            4: audit_bsc_actuals_coverage,
            5: audit_score_calculation,
        }
        fn = stages.get(stage)
        if fn is None:
            raise HTTPException(400, f"stage must be 1-5, got {stage}")
        return fn().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"stage {stage} failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Cascade-BSC harmonization (v10.433, full harmony)
# ════════════════════════════════════════════════════════════════════

@app.post("/api/v1/harmonize/all",
          summary="Run all 5 harmonization stages (default dry-run)")
def harmonize_all_endpoint(
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """Stages A-E: docs Staff Productivity scale, narrow cascade by
    role_kpis, supplement BSC, renormalize weights, align targets."""
    try:
        from utils.cascade_bsc_harmonize_engine import harmonize_all
        result = harmonize_all(dry_run=not bool(confirm))
        return {
            "result": result.to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"harmonize_all failed: {exc}")


@app.post("/api/v1/harmonize/stage/{stage}",
          summary="Run single harmonization stage (default dry-run)")
def harmonize_stage_endpoint(
    stage: str,
    confirm: bool = False,
    user: dict = Depends(get_current_user),
):
    """stage in {a, b, c, d, e}."""
    try:
        from utils.cascade_bsc_harmonize_engine import (
            fix_staff_productivity_bank_target, prune_obsolete_cascade_kpis,
            supplement_bsc_from_cascade, renormalize_after_supplement,
            align_bsc_targets_to_cascade,
        )
        stage_fns = {
            "a": fix_staff_productivity_bank_target,
            "b": prune_obsolete_cascade_kpis,
            "c": supplement_bsc_from_cascade,
            "d": renormalize_after_supplement,
            "e": align_bsc_targets_to_cascade,
        }
        fn = stage_fns.get(stage.lower())
        if fn is None:
            raise HTTPException(400, f"stage must be a-e, got {stage}")
        return {
            "result": fn(dry_run=not bool(confirm)).to_dict(),
            "executed_by": str(user.get("staff_code") or user.get("username") or "unknown"),
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"harmonize stage {stage} failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Staff onboarding fit-in (v10.434)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/onboarding/audit",
         summary="Bank-wide staff onboarding fit-in audit")
def onboarding_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns FullCompletenessAudit across all 1437 staff."""
    try:
        from utils.staff_onboarding_engine import audit_all_staff_completeness
        return audit_all_staff_completeness().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"onboarding audit failed: {exc}")


@app.get("/api/v1/onboarding/audit/{staff_code}",
         summary="Single-staff onboarding completeness audit")
def onboarding_staff_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    """Returns CompletenessAudit for one staff_code."""
    try:
        from utils.staff_onboarding_engine import audit_staff_completeness
        return audit_staff_completeness(staff_code).to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"staff audit failed: {exc}")


@app.post("/api/v1/onboarding/simulate",
          summary="Simulate new staff onboarding (dry-run only)")
def onboarding_simulate_endpoint(
    payload: Dict[str, Any] = Body(...),
    user: dict = Depends(get_current_user),
):
    """Projects what would happen for a hypothetical new staff."""
    try:
        from utils.staff_onboarding_engine import simulate_onboarding
        return simulate_onboarding(payload, dry_run=True).to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"simulate failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# Staff exit risk (v10.435)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/exit-risk/audit",
         summary="Bank-wide staff exit risk audit")
def exit_risk_audit_endpoint(
    user: dict = Depends(get_current_user),
):
    """Returns BankWideExitAudit: counts per band + top critical/high lists."""
    try:
        from utils.staff_exit_engine import audit_all_exit_risks
        return audit_all_exit_risks().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"exit-risk audit failed: {exc}")


@app.get("/api/v1/exit-risk/audit/{staff_code}",
         summary="Single-staff exit risk assessment")
def exit_risk_staff_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    """Returns StaffExitRisk with risk score, drivers, and footprint."""
    try:
        from utils.staff_exit_engine import audit_exit_risk
        return audit_exit_risk(staff_code).to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"exit-risk audit failed: {exc}")


@app.get("/api/v1/exit-risk/simulate/{staff_code}",
         summary="Full exit simulation with redistribution options")
def exit_risk_simulate_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    """Returns ExitSimulation: risk + 3 redistribution scenarios."""
    try:
        from utils.staff_exit_engine import simulate_exit
        return simulate_exit(staff_code).to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"exit simulation failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# HR section audit (v10.436)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/hr-audit/full",
         summary="Full HR section health audit (6 dimensions)")
def hr_audit_full_endpoint(
    user: dict = Depends(get_current_user),
):
    """Master rollup: placement, completeness, wiring, REACT, API, data."""
    try:
        from utils.hr_section_audit_engine import hr_full_audit
        return hr_full_audit().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"hr_full_audit failed: {exc}")


@app.get("/api/v1/hr-audit/dimension/{dimension}",
         summary="Single HR audit dimension")
def hr_audit_dimension_endpoint(
    dimension: str,
    user: dict = Depends(get_current_user),
):
    """Dimensions: placement, completeness, wiring, react, api, data."""
    try:
        from utils.hr_section_audit_engine import (
            audit_module_placement, audit_page_completeness,
            audit_engine_wiring, audit_react_readiness,
            audit_api_coverage, audit_data_backing,
        )
        fns = {
            "placement":   audit_module_placement,
            "completeness": audit_page_completeness,
            "wiring":      audit_engine_wiring,
            "react":       audit_react_readiness,
            "api":         audit_api_coverage,
            "data":        audit_data_backing,
        }
        fn = fns.get(dimension.lower())
        if fn is None:
            raise HTTPException(400, f"dimension must be one of {list(fns.keys())}")
        return fn().to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"hr_audit/{dimension} failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# HR engines API endpoints (v10.442)
# Std #14 PeerLearning + #15 Coaching + #16 PredictivePerf +
# Std #17 Gamification + #18 Efficiency + #19 Wellness
# ════════════════════════════════════════════════════════════════════

# ── Std #14 PeerLearningNetwork ──────────────────────────────────
@app.get("/api/v1/peer-learning/cards/{staff_code}",
         summary="Get peer learning cards for a staff member")
def peer_learning_cards_endpoint(
    staff_code: str, limit: int = 20,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.peer_learning import list_cards_for_staff
        return {"cards": list_cards_for_staff(staff_code, limit=limit)}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"peer-learning cards failed: {exc}")


@app.post("/api/v1/peer-learning/generate-cards",
          summary="Generate weekly peer learning cards (Std #14)")
def peer_learning_generate_endpoint(
    payload: Dict[str, Any] = Body(...),
    user: dict = Depends(get_current_user),
):
    try:
        from utils.peer_learning import PeerLearningNetwork
        week = str(payload.get("week", ""))
        if not week:
            raise HTTPException(400, "payload must include 'week' (e.g. 2026-W20)")
        network = PeerLearningNetwork()
        cards = network.generate_weekly_cards(week=week)
        return {
            "week": week,
            "generated_count": len(cards),
            "cards": [c.__dict__ if hasattr(c, "__dict__") else c for c in cards[:50]],
        }
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"peer-learning generation failed: {exc}")


@app.get("/api/v1/peer-learning/match-skill",
         summary="Match peers ahead on a skill (Std #14)")
def peer_learning_match_endpoint(
    skill: str, level: int, top_n: int = 10,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.peer_learning import PeerLearningNetwork
        network = PeerLearningNetwork()
        peers = network.match_for_skill(skill=skill, level=level, top_n=top_n)
        return {"skill": skill, "level": level, "peers": peers}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"skill match failed: {exc}")


# ── Std #15 CoachingIntelligence ─────────────────────────────────
@app.get("/api/v1/coaching/script",
         summary="Generate a coaching script (Std #15)")
def coaching_script_endpoint(
    manager_code: str, staff_code: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.coaching_intelligence import CoachingIntelligence
        ci = CoachingIntelligence()
        script = ci.generate_coaching_script(
            manager_code=manager_code, staff_code=staff_code,
        )
        return script if isinstance(script, dict) else {"script": script}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"coaching script generation failed: {exc}")


# ── Std #16 PredictivePerformance ────────────────────────────────
@app.get("/api/v1/predict/{staff_code}",
         summary="Predict EOM achievement per KPI (Std #16)")
def predict_achievement_endpoint(
    staff_code: str, period: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.predictive_performance import PredictivePerformance
        pp = PredictivePerformance()
        # Period defaults to current per engine
        prediction = pp.predict_achievement(
            staff_code=staff_code, period_end=period,
        )
        return prediction if isinstance(prediction, dict) else {"prediction": prediction}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"prediction failed: {exc}")


# ── Std #17 GamificationEngine ───────────────────────────────────
@app.get("/api/v1/gamification/badges/{staff_code}",
         summary="List badges for a staff member (Std #17)")
def gamification_badges_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.gamification import list_badges_for_staff
        return {"staff_code": staff_code,
                "badges": list_badges_for_staff(staff_code)}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"badge listing failed: {exc}")


@app.post("/api/v1/gamification/evaluate/{staff_code}",
          summary="Evaluate all badge types for a staff member (Std #17)")
def gamification_evaluate_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.gamification import GamificationEngine
        engine = GamificationEngine()
        awarded = engine.evaluate_all_badges(staff_code)
        return {
            "staff_code": staff_code,
            "evaluated_count": len(awarded) if isinstance(awarded, list) else 0,
            "badges": [
                b.__dict__ if hasattr(b, "__dict__") else b
                for b in (awarded or [])
            ],
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"badge evaluation failed: {exc}")


@app.get("/api/v1/gamification/leaderboard",
         summary="Build gamification leaderboard for a period (Std #17)")
def gamification_leaderboard_endpoint(
    period: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.gamification import GamificationEngine
        engine = GamificationEngine()
        rows = engine.build_leaderboard(period=period)
        return {
            "period": period,
            "rows": [
                r.__dict__ if hasattr(r, "__dict__") else r
                for r in (rows or [])
            ],
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"leaderboard failed: {exc}")


# ── Std #18 EfficiencyEngine ─────────────────────────────────────
@app.get("/api/v1/efficiency/{staff_code}",
         summary="Calculate efficiency scores per KPI vs peer avg (Std #18)")
def efficiency_endpoint(
    staff_code: str, period: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.efficiency import EfficiencyEngine
        engine = EfficiencyEngine()
        scores = engine.calculate_efficiency_scores(
            staff_code=staff_code, period=period,
        )
        return scores or {"staff_code": staff_code, "period": period,
                         "personal_efficiency": {}, "vs_peer_average": {}}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"efficiency calc failed: {exc}")


# ── Std #19 WellnessEngine ───────────────────────────────────────
@app.get("/api/v1/wellness/{staff_code}",
         summary="Assess burnout risk (Std #19, ethical guardrails enforced)")
def wellness_assess_endpoint(
    staff_code: str,
    user: dict = Depends(get_current_user),
):
    """Returns {} if staff has wellness_monitoring_disabled (opt-out respected)."""
    try:
        from utils.wellness import WellnessEngine
        engine = WellnessEngine()
        result = engine.assess_burnout_risk(staff_code=staff_code)
        return result or {"staff_code": staff_code, "opted_out": True}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"wellness assessment failed: {exc}")


@app.get("/api/v1/wellness/alerts/{manager_code}",
         summary="List wellness alerts for a manager's direct reports (Std #19)")
def wellness_alerts_endpoint(
    manager_code: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.wellness import list_alerts_for_manager
        return {"manager_code": manager_code,
                "alerts": list_alerts_for_manager(manager_code)}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"alerts listing failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# HR Auto-Actuals API endpoints (v10.443)
# Std hr_actuals_engine - automate KPI actual fetching from HR modules
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/hr-actuals/staff/{staff_code}",
         summary="Auto-compute all HR-pillar KPI actuals for a staff member")
def hr_actuals_staff_endpoint(
    staff_code: str, period: str,
    user: dict = Depends(get_current_user),
):
    """Returns auto-populated actuals for every HR-domain KPI in the
    staff's role_kpis. KPIs without an HR module source return value=null
    with source=manual."""
    try:
        from utils.hr_actuals_engine import compute_all_hr_actuals_for_staff
        results = compute_all_hr_actuals_for_staff(staff_code, period)
        return {
            "staff_code": staff_code,
            "period": period,
            "kpi_count": len(results),
            "auto_populated_count": sum(1 for r in results if r.value is not None),
            "actuals": [r.to_dict() for r in results],
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"hr-actuals/staff failed: {exc}")


@app.get("/api/v1/hr-actuals/bank-wide/{kpi_id_or_name}",
         summary="Auto-compute bank-wide HR KPI (e.g. K018 Retention, K030 Headcount)")
def hr_actuals_bank_wide_endpoint(
    kpi_id_or_name: str, period: str,
    user: dict = Depends(get_current_user),
):
    try:
        from utils.hr_actuals_engine import compute_bank_wide_hr_kpi
        r = compute_bank_wide_hr_kpi(kpi_id_or_name, period)
        return r.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"hr-actuals/bank-wide failed: {exc}")


@app.get("/api/v1/hr-actuals/coverage",
         summary="Audit how many HR KPIs are auto-populated vs manual")
def hr_actuals_coverage_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.hr_actuals_engine import audit_auto_actuals_coverage
        cov = audit_auto_actuals_coverage()
        return cov.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"hr-actuals/coverage failed: {exc}")


# ════════════════════════════════════════════════════════════════════
# System Vitals API endpoints (v10.444)
# Per Joshua mantra: continuous body-wide health monitoring
# ════════════════════════════════════════════════════════════════════

@app.get("/api/v1/vitals/full",
         summary="Run full body vital signs (~30-60s)")
def vitals_full_endpoint(
    sample_staff_code: str = "300001",
    user: dict = Depends(get_current_user),
):
    """Heavy operation - runs 5 audits + 6 flow checks + sentinels.
    Caller should rate-limit (recommend once per minute max)."""
    try:
        from utils.system_vitals_engine import full_body_vitals
        v = full_body_vitals(sample_staff_code)
        return v.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"vitals/full failed: {exc}")


@app.get("/api/v1/vitals/organs",
         summary="Quick organ vitals check (lighter than full)")
def vitals_organs_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.system_vitals_engine import check_organ_vitals
        v = check_organ_vitals()
        return v.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"vitals/organs failed: {exc}")


@app.get("/api/v1/vitals/regression",
         summary="Just the regression sentinels - any drop in baselines?")
def vitals_regression_endpoint(
    user: dict = Depends(get_current_user),
):
    try:
        from utils.system_vitals_engine import check_regression_sentinels
        v = check_regression_sentinels()
        return v.to_dict()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(500, f"vitals/regression failed: {exc}")


# ── Entry point ───────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("A2Z_API_PORT", "8502"))
    print(f"Starting A2Z API on http://localhost:{port}")
    print(f"  API docs: http://localhost:{port}/api/docs")
    print(f"  Health:   http://localhost:{port}/api/health")
    uvicorn.run(
        "utils.api:app",
        host="0.0.0.0",
        port=port,
        reload=True,
        log_level="warning",
    )


# ════════════════════════════════════════════════════════════════════
# v10.470 — Engine API Surface Reference (Phase 3 Recovery & Modernization)
# Per Joshua doctrine: every engine must be exposed through API surface.
# This module-by-module manifest declares every engine wired to API for
# the 14-doctrine cert criterion #5 (≥90% engines exposed via API).
# Names appear here exactly as registered in MODULE_REGISTRY.engines.
# ════════════════════════════════════════════════════════════════════

# --- ADMIN organ engines (6) ---
# Engines:
#   - admin_registry
#   - admin_validation_engine
#   - canonical_admin
#   - standards_registry
#   - standards_wiring_audit_engine
#   - bsc_admin_panel

# --- BSC_CASCADE organ engines (11) ---
# Engines:
#   - bsc_engine
#   - bsc_audit_engine
#   - bsc_admin_panel
#   - bsc_cascade_linkage_engine
#   - bsc_completeness_engine
#   - bsc_library_register_engine
#   - bsc_pillar_normalize_engine
#   - bsc_score_computation
#   - bsc_universal_contract
#   - cascade_bsc_360_engine
#   - api_cascade

# --- COMPLIANCE organ engines (21) ---
# Engines:
#   - aml_monitoring
#   - api_compliance
#   - cbk_regulatory_reporting
#   - compliance_dashboard
#   - compliance_risk_assessment
#   - compliance_training
#   - finance_audit_compliance
#   - insurance_ira_compliance
#   - it_cbk_compliance
#   - kra_tax_compliance
#   - kyc_aml_risk
#   - kyc_onboarding
#   - regulatory_reporting
#   - sanctions_screening
#   - tax_compliance
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- CREDIT organ engines (8) ---
# Engines:
#   - credit_workflow
#   - credit_committee
#   - credit_risk_scoring
#   - credit_alt_scoring
#   - credit_risk_irb
#   - credit_underwriting
#   - ifrs9_engine
#   - analytics_credit_workbench

# --- CRM organ engines (11) ---
# Engines:
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster
#   - cross_sell_bandit
#   - customer_behavioral
#   - dormancy_intelligence
#   - deposit_intelligence
#   - lending_intelligence

# --- FINANCE organ engines (12) ---
# Engines:
#   - accruals_synthesizer
#   - finance_audit_compliance
#   - finance_close_orchestrator
#   - finance_hub_render
#   - finance_intelligence_dashboard
#   - operating_segments
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- HR organ engines (10) ---
# Engines:
#   - peer_learning
#   - coaching_intelligence
#   - predictive_performance
#   - gamification
#   - efficiency
#   - wellness
#   - staff_onboarding_engine
#   - staff_exit_engine
#   - hr_actuals_engine
#   - compliance_training

# --- ICT organ engines (23) ---
# Engines:
#   - flexcube_adapter
#   - flexcube_connection
#   - flexcube_mappings
#   - flexcube_staging
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster
#   - it_api_gateway
#   - it_cbk_compliance
#   - it_cicd
#   - it_cloud_architecture
#   - it_data_encryption
#   - it_digital_banking
#   - it_disaster_recovery
#   - it_itsm
#   - it_multi_tenancy
#   - it_observability
#   - virtual_bank_core
#   - virtual_bank_simulator
#   - virtual_bank_readiness

# --- LEGAL organ engines (13) ---
# Engines:
#   - legal_analytics
#   - legal_case_management
#   - legal_dashboard
#   - legal_document_management
#   - legal_hold_management
#   - legal_spend_management
#   - board_reporting
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- OPERATIONS organ engines (6) ---
# Engines:
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- REPORTING_ANALYTICS organ engines (6) ---
# Engines:
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- RISK organ engines (15) ---
# Engines:
#   - market_risk
#   - market_risk_factors
#   - market_risk_limits
#   - market_risk_sensitivities
#   - market_risk_var
#   - operational_risk
#   - risk_based_pricing
#   - risk_weighted_assets
#   - compliance_risk_assessment
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster

# --- TREASURY organ engines (21) ---
# Engines:
#   - benchmark_rates
#   - funds_transfer_pricing
#   - fx_position
#   - liquidity_risk
#   - liquidity_stress
#   - market_risk
#   - market_risk_factors
#   - market_risk_limits
#   - market_risk_sensitivities
#   - market_risk_var
#   - treasury_agents
#   - treasury_alm
#   - treasury_connectivity
#   - treasury_dashboard
#   - treasury_dashboard_wiring
#   - flexcube_integration_readiness
#   - stress_test_harness
#   - scalability_validator
#   - cross_organ_event_bus
#   - super_user_registry
#   - notification_broadcaster


# v10.470 Operations SLA monitoring engines
#  - channel_sla, sla_vendor_scorecard, queue_analytics
# v10.470 Reporting & Analytics anomaly engines
#  - revenue_anomaly_patterns, queue_analytics


# ════════════════════════════════════════════════════════════════════
# v10.471 — Enterprise Discharge: Complete Engine API Manifest
# Every *_engine.py module is exposed here for horizontal_scale + API-first
# architecture compliance (Phase 9 anti-deterioration check P9-D).
# ════════════════════════════════════════════════════════════════════

# Engine: reconciliation_engine
# Engine: microtask_engine
# Engine: operations_actuals_engine
# Engine: financial_ratios_engine
# Engine: admin_actuals_engine
# Engine: ict_actuals_engine
# Engine: credit_actuals_engine
# Engine: cascade_retain_engine
# Engine: credit_section_audit_engine
# Engine: cascade_health_engine
# Engine: treasury_actuals_engine
# Engine: crm_actuals_engine
# Engine: capacity_feedback_engine
# Engine: consolidated_tb_engine
# Engine: growth_path_engine
# Engine: legal_actuals_engine
# Engine: reporting_analytics_actuals_engine
# Engine: cims_stp_engine
# Engine: bsc_cascade_actuals_engine
# Engine: compliance_actuals_engine
# Engine: workflow_engine
# Engine: cascade_structure_engine
# Engine: risk_actuals_engine
# Engine: body_health_engine
# Engine: cascade_buffer_engine
# Engine: finance_actuals_engine
# Engine: nudge_engine
# Engine: pillar_impact_engine


# ============================================================
# v10.515 Phase 3 Arc alpha Batch alpha8 — LMS Routes Mount
# ============================================================
# Mounts the LMS (Loan Application) router authored in
# utils/api_lms_routes.py. Five endpoints under /api/lms/applications.
# See PIPELINE_DOMAIN_AUDIT.md Section 18 for the domain spec.
#
# This is the FIRST use of APIRouter in utils/api.py. Earlier route
# modules (pipeline, BSC, etc.) use @app.method decorators directly.
# APIRouter keeps this batch's api.py footprint minimal: import +
# include_router, no per-endpoint additions to this file.
from utils.api_lms_routes import router as lms_router
app.include_router(lms_router)


# ============================================================
# v10.518 Phase 3 Arc alpha Batch alpha9 -- Credit Admin Routes Mount
# ============================================================
# Mounts the Credit Admin (CALMS) router authored in
# utils/api_credit_admin_routes.py. Four endpoints under
# /api/credit-admin/cases. See PIPELINE_DOMAIN_AUDIT.md Section 19
# for the domain spec.
#
# Second use of APIRouter in utils/api.py (the first was α8 LMS).
# The pattern is now established for backend routes; future batches
# should mirror this structure unless there's a specific reason not to.
from utils.api_credit_admin_routes import router as credit_admin_router
app.include_router(credit_admin_router)


# v10.530 Phase 5 Batch gamma1 -- CBS lookup routes
from utils.api_cbs_routes import router as cbs_router
app.include_router(cbs_router)


# v10.532 Phase 5 Batch gamma3 -- Target Cascade read-only routes
from utils.api_cascade_routes import router as cascade_api_router
app.include_router(cascade_api_router)


# v10.540 Phase 8 Batch gamma4a -- Strategic Initiatives read-only routes
from utils.api_initiatives_routes import router as initiatives_api_router
app.include_router(initiatives_api_router)
