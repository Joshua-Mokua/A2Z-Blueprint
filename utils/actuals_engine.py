"""
utils/actuals_engine.py — Live Actuals Engine
Computes KPI actuals from CBS data on demand or on schedule.
Called by app.py on startup and by the Admin refresh button.
No manual Excel upload needed once CBS data exists.
"""
import json, csv, openpyxl
import logging
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

logger = logging.getLogger(__name__)


def _root():
    """Project root — the folder containing a2z/."""
    return Path(__file__).parent.parent


def get_cbs_paths():
    """Return CBS data folder and actuals output folder."""
    root = _root()
    # CBS data can be in cbs_data/ (project root) or a2z/data/
    for candidate in [root / "cbs_data", root / "a2z" / "data"]:
        if candidate.exists():
            return candidate, root / "a2z" / "data"
    return root / "a2z" / "data", root / "a2z" / "data"


def get_period_label():
    """Current period label e.g. 'Mar-26'."""
    from utils.core import get_org_config
    try:
        cfg = get_org_config()
        fy  = str(cfg.get("fiscal_year", date.today().year))
        # Use last 2 digits
        yr2 = fy[-2:]
        return f"{date.today().strftime('%b')}-{yr2}"
    except:
        return date.today().strftime("%b-%y")



def inject_cascade_targets(xlsx_path: Path) -> int:
    """
    Read target_cascade.json and bank_targets.json and write Annual Target
    into the actuals xlsx for every matching Staff Code + KPI row.
    Returns count of rows updated.
    """
    data_dir = xlsx_path.parent
    tc_path  = data_dir / "target_cascade.json"
    bt_path  = data_dir / "bank_targets.json"

    if not tc_path.exists():
        return 0

    tc_data = json.loads(tc_path.read_text())
    bt_data = json.loads(bt_path.read_text()) if bt_path.exists() else {}

    # Build staff×KPI target map
    per_staff: dict = {}
    for key, entry in tc_data.items():
        if key.startswith("deadline") or key.startswith("global"):
            continue
        kpi = entry.get("kpi", "")
        fc  = entry.get("from_code", "")
        if not kpi:
            continue
        if fc and (fc, kpi) not in per_staff:
            per_staff[(fc, kpi)] = float(entry.get("total_target", 0) or 0)
        for alloc in entry.get("allocations", []):
            to_code = str(alloc.get("to_code", ""))
            amount  = alloc.get("amount", 0)
            if to_code and amount:
                per_staff[(to_code, kpi)] = float(amount)

    # Bank-level fallback
    bank_by_kpi = {k.split("|")[0]: float(v.get("target", 0) or 0)
                   for k, v in bt_data.items()}

    # Qualitative baselines (not in CBS or cascade)
    baselines = {
        "Compliance Score": 95.0,
        "Audit Score":      90.0,
        "CX Score":          4.0,
        "Staff Productivity": 3.0,
        "Diligence Score":   3.0,
        "Initiative Implementation Score": 100.0,
        "Active Initiatives Count":         5.0,
    }

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    if "Staff Code" not in headers or "KPI" not in headers:
        return 0

    sc_col  = headers.index("Staff Code")    + 1
    kpi_col = headers.index("KPI")           + 1
    tgt_col = headers.index("Annual Target") + 1

    updated = 0
    for row_idx in range(3, ws.max_row + 1):
        sc  = str(ws.cell(row_idx, sc_col).value or "").strip()
        kpi = str(ws.cell(row_idx, kpi_col).value or "").strip()
        if not sc or not kpi:
            continue
        # Look up: staff-specific allocation → bank-level → baseline
        # Always overwrite — cascade is source of truth; supports re-allocation after lock
        target = (per_staff.get((sc, kpi))
                  or bank_by_kpi.get(kpi)
                  or baselines.get(kpi))

        if target and float(target) > 0:
            current = float(ws.cell(row_idx, tgt_col).value or 0)
            if abs(float(target) - current) > 0.01:   # only write if changed
                ws.cell(row_idx, tgt_col).value = float(target)
                updated += 1

    # ── Create missing rows for cascaded KPIs not yet in xlsx ───────────
    existing = {}
    for row_idx in range(3, ws.max_row + 1):
        sc_v  = str(ws.cell(row_idx, sc_col).value  or "").strip()
        kpi_v = str(ws.cell(row_idx, kpi_col).value or "").strip()
        if sc_v and kpi_v:
            existing[(sc_v, kpi_v)] = row_idx

    kpi_meta: dict = {}
    kpi_lib_path = data_dir / "kpi_library.json"
    if kpi_lib_path.exists():
        try:
            kpi_lib = json.loads(kpi_lib_path.read_text())
            for pillar, kpis in kpi_lib.get("pillars", {}).items():
                for k in kpis:
                    kpi_meta[k["name"]] = {"pillar": pillar,
                                            "weight": k.get("default_weight", 0.05)}
        except Exception:
            pass

    staff_meta: dict = {}
    sr_path = data_dir / "staff_register.xlsx"
    if sr_path.exists():
        try:
            wb_sr = openpyxl.load_workbook(str(sr_path))
            ws_sr = wb_sr.active
            sr_hdr = [ws_sr.cell(1, c).value for c in range(1, ws_sr.max_column + 1)]
            sc_sr = sr_hdr.index("Staff Code") if "Staff Code" in sr_hdr else 0
            nc_sr = sr_hdr.index("Staff Name") if "Staff Name" in sr_hdr else 1
            rc_sr = sr_hdr.index("Role")       if "Role"       in sr_hdr else 2
            uc_sr = sr_hdr.index("Unit")       if "Unit"       in sr_hdr else 3
            cc_sr = sr_hdr.index("Category")   if "Category"   in sr_hdr else 4
            for row in ws_sr.iter_rows(min_row=2, values_only=True):
                if row[sc_sr]:
                    staff_meta[str(row[sc_sr])] = {
                        "name":     str(row[nc_sr] or ""),
                        "role":     str(row[rc_sr] or ""),
                        "unit":     str(row[uc_sr] or ""),
                        "category": str(row[cc_sr] or ""),
                    }
        except Exception:
            pass

    period_col = next((i + 1 for i, h in enumerate(headers)
                       if h and (str(h).endswith("-25") or str(h).endswith("-26"))), None)

    rows_added = 0
    for (sc_v, kpi_v), target in per_staff.items():
        if not sc_v or not kpi_v or float(target or 0) <= 0:
            continue
        if (sc_v, kpi_v) in existing:
            continue
        sm = staff_meta.get(sc_v)
        if not sm:
            continue
        km = kpi_meta.get(kpi_v, {"pillar": "Financial", "weight": 0.05})
        new_row = [""] * len(headers)
        def _set(col_name, val):
            if col_name in headers:
                new_row[headers.index(col_name)] = val
        _set("Staff Code",    sc_v)
        _set("Staff Name",    sm["name"])
        _set("Role",          sm["role"])
        _set("Unit",          sm["unit"])
        _set("Category",      sm["category"])
        _set("Staff Status",  "Active")
        _set("KPI",           kpi_v)
        _set("Pillar",        km["pillar"])
        _set("Weight",        km["weight"])
        _set("Annual Target", float(target))
        _set("YTD_Actual",    0.0)
        _set("Annual Actual", 0.0)
        if period_col:
            new_row[period_col - 1] = 0.0
        ws.append(new_row)
        existing[(sc_v, kpi_v)] = ws.max_row
        rows_added += 1
        updated += 1

    if updated:
        wb.save(str(xlsx_path))
    return updated

def compute_actuals_from_cbs(force: bool = False) -> dict:
    """
    Compute KPI actuals from CBS data and write actuals_YYYY_MM.xlsx.
    Returns status dict: {success, path, rows, duration_s, message}
    
    force=True: recompute even if actuals file is newer than CBS data.
    """
    t0 = datetime.now()
    cbs_dir, data_dir = get_cbs_paths()
    period = get_period_label()

    # ── Check if refresh needed ────────────────────────────────────────
    out_name   = f"actuals_{date.today().strftime('%Y_%b_%d')}.xlsx"
    out_path   = data_dir / out_name
    cbs_exists = any(cbs_dir.glob("*.csv")) or any(cbs_dir.glob("*.xlsx"))

    if not cbs_exists:
        return {"success": False, "message": "No CBS data found in cbs_data/",
                "path": None, "rows": 0, "duration_s": 0}

    if not force and out_path.exists():
        # Check if CBS is newer than actuals
        cbs_mtime    = max(p.stat().st_mtime for p in list(cbs_dir.glob("*.csv")) + list(cbs_dir.glob("*.xlsx")))
        actuals_mtime= out_path.stat().st_mtime
        if actuals_mtime >= cbs_mtime:
            return {"success": True, "message": "Actuals are current",
                    "path": out_path, "rows": 0, "duration_s": 0, "cached": True}

    # ── Load staff register ────────────────────────────────────────────
    sr_path = data_dir / "staff_register.xlsx"
    if not sr_path.exists():
        sr_path = cbs_dir / "staff_register.xlsx"
    if not sr_path.exists():
        return {"success": False, "message": "staff_register.xlsx not found",
                "path": None, "rows": 0, "duration_s": 0}

    staff_list = []
    try:
        wb = openpyxl.load_workbook(str(sr_path))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = dict(zip(headers, row))
            if r.get("Staff Code"):
                staff_list.append(r)
    except Exception as e:
        return {"success": False, "message": f"Error reading staff register: {e}",
                "path": None, "rows": 0, "duration_s": 0}

    # ── Load KPI library ───────────────────────────────────────────────
    try:
        from utils.core_kpi import get_kpi_library
        from utils.core import get_org_config
        lib = get_kpi_library()
        cfg = get_org_config()
        pillar_weights = cfg.get("pillar_weights", {
            "Financial": 0.68, "Customer Focus": 0.14,
            "Operational Excellence": 0.06, "People & Learning": 0.12})
    except Exception as e:
        return {"success": False, "message": f"KPI library error: {e}",
                "path": None, "rows": 0, "duration_s": 0}

    role_kpis    = lib.get("role_kpis", {})
    kpi_weights  = lib.get("kpi_weights", {})
    pillars_data = lib.get("pillars", {})

    # Build KPI maps
    id_to_kpi = {}
    for pillar, kpis in pillars_data.items():
        for k in kpis:
            id_to_kpi[k["id"]] = {**k, "pillar": pillar}

    # ── Load CBS actuals ───────────────────────────────────────────────
    # Prefer pre-computed actuals file from compute_actuals.py if available
    existing_actuals = sorted(
        list(cbs_dir.glob("actuals_*.xlsx")) + list(data_dir.glob("actuals_*.xlsx")),
        key=lambda p: p.stat().st_mtime, reverse=True)

    if existing_actuals and existing_actuals[0] != out_path:
        # Most recent existing actuals — copy and add initiative KPIs
        import shutil
        shutil.copy2(str(existing_actuals[0]), str(out_path))
        _inject_initiative_kpis(out_path, staff_list, lib)
        inject_cascade_targets(out_path)
        elapsed = (datetime.now() - t0).total_seconds()
        return {"success": True, "path": out_path,
                "rows": _count_rows(out_path),
                "duration_s": round(elapsed, 1),
                "message": f"Refreshed from {existing_actuals[0].name}"}

    # ── Build from CBS CSVs ────────────────────────────────────────────
    rows = _build_from_cbs(staff_list, role_kpis, id_to_kpi, kpi_weights,
                            pillar_weights, cbs_dir, period)

    # ── Add initiative KPIs ────────────────────────────────────────────
    rows = _add_initiative_kpis(rows, staff_list, lib, period)

    # ── Write to Excel ─────────────────────────────────────────────────
    _write_actuals(rows, out_path, period)

    # ── Inject cascade targets ────────────────────────────────────────
    tgt_updated = inject_cascade_targets(out_path)

    # ── Pilot: route through utils.bsc_engine (Standards #1 + #2) ─────
    # Every BSC contribution from the CBS-derived actuals is now stamped
    # through the central engine. Failures here are non-fatal — the
    # legacy xlsx output still ships even if the engine rejects records.
    bsc_summary = _submit_to_bsc_engine(rows, period)

    # v10.355: YoY sidecar is refreshed by callers (app.py / admin
    # refresh) AFTER this function returns. Inverted from the original
    # v10.355 placement to break the actuals_engine → live_actuals →
    # cbs_baseline → actuals_engine cycle that G128 flagged. The lower
    # data-extraction layer (this module) no longer depends on the
    # higher orchestration layer (live_actuals).

    elapsed = (datetime.now() - t0).total_seconds()
    return {"success": True, "path": out_path,
            "rows": len(rows),
            "targets_set": tgt_updated,
            "bsc_engine":  bsc_summary,
            "duration_s": round(elapsed, 1),
            "message": (f"Computed {len(rows):,} KPI rows, "
                        f"{tgt_updated:,} targets injected, "
                        f"BSC engine: {bsc_summary.get('ok',0)} ok / "
                        f"{bsc_summary.get('rejected',0)} rejected "
                        f"in {elapsed:.1f}s")}


def _period_to_engine_format(legacy_period: str) -> str:
    """Translate actuals_engine's 'Mar-26' label into the BSC engine's
    canonical 'YYYY-MM' contract format.

    'Mar-26' → '2026-03'
    Falls back to current YYYY-MM if parsing fails.
    """
    from datetime import datetime as _dt
    try:
        # Parse with %y (2-digit year, current century by default)
        parsed = _dt.strptime(legacy_period, "%b-%y")
        return parsed.strftime("%Y-%m")
    except Exception:
        return _dt.today().strftime("%Y-%m")


def _submit_to_bsc_engine(rows: list, legacy_period: str) -> dict:
    """Translate actuals rows into the universal BSC contract and submit
    them through utils.bsc_engine. Returns the engine's batch summary.

    This is the addendum Standards #1 + #2 pilot — the actuals engine is
    the first module wired to the central BSC integration engine.
    """
    try:
        from utils.bsc_engine import submit_batch as _bsc_submit_batch
    except Exception as e:
        return {"ok": 0, "rejected": 0, "errors": [{"index": -1, "error": f"bsc_engine import failed: {e}"}]}

    period = _period_to_engine_format(legacy_period)
    contract_records = []
    for r in rows:
        sc  = str(r.get("Staff Code", "")).strip()
        kid = str(r.get("kpi_id", "")).strip()
        val = r.get(legacy_period, r.get("Annual Actual", 0))
        if not sc or not kid:
            continue
        contract_records.append({
            "staff_code": sc,
            "kpi_id":     kid,
            "value":      val,
            "period":     period,
        })

    return _bsc_submit_batch(
        records       = contract_records,
        source_module = "actuals_engine",
        actor         = "actuals_engine_etl",
    )



def aggregate_cbs_by_rm(cbs_dir: Path) -> dict:
    """
    Single-pass read of accounts.csv.  Groups every field that matters
    by relationship_manager_code.  Returns dict: rm_code → KPI-ready dict.

    Loan segmentation (for Retail / MSME / Corporate split):
        Retail  : Personal Loan, Mortgage / Home Loan, Salary Advance,
                  Graduate Loan, Consumer Loan, Staff Loan
        MSME    : Business Loan, LPO Finance, Asset Finance, Invoice Discounting,
                  Business Overdraft
        Corporate: Syndicated Loan, Corporate Term Loan, Trade Finance Loan,
                  Project Finance
    """
    import csv as _csv

    RETAIL_LOAN_TYPES = {
        "Personal Loan", "Mortgage / Home Loan", "Salary Advance",
        "Graduate Loan", "Consumer Loan", "Staff Loan",
    }
    MSME_LOAN_TYPES = {
        "Business Loan", "LPO Finance", "Asset Finance",
        "Invoice Discounting", "Business Overdraft",
    }
    CORPORATE_LOAN_TYPES = {
        "Syndicated Loan", "Corporate Term Loan",
        "Trade Finance Loan", "Project Finance",
    }
    # Everything else (Overdraft Facility etc.) → MSME by default

    acct_csv = None
    for name in ["accounts.csv", "cbs_accounts.csv"]:
        p = cbs_dir / name
        if p.exists():
            acct_csv = p
            break
    if not acct_csv:
        return {}

    rm: dict = {}   # rm_code → aggregation buckets

    def _get(rm_dict, code):
        if code not in rm_dict:
            rm_dict[code] = {
                # Deposits
                "casa_balance":      0.0,   # CASA = Current + Savings
                "term_balance":      0.0,   # Term deposits
                "total_deposits":    0.0,   # CASA + Term
                "retail_deposits":   0.0,   # personal / salary accounts
                "commercial_deposits": 0.0, # business / corporate accounts
                # Loans
                "loan_outstanding":  0.0,   # all loans outstanding
                "loan_disbursed":    0.0,   # original loan amounts (proxy for disbursement)
                "retail_loans":      0.0,
                "msme_loans":        0.0,
                "corporate_loans":   0.0,
                # NPL
                "npl_outstanding":   0.0,
                "par_outstanding":   0.0,   # NPL days > 0 (any overdue)
                # Income
                "interest_income":   0.0,
                "fee_income":        0.0,
                # Accounts
                "account_count":     0,
                "loan_count":        0,
                "new_accounts_2025": 0,    # opened in 2025 fiscal year
                "new_accounts_2026": 0,    # opened in 2026 fiscal year
                "dormant_count":     0,
                "business_borrowers":0,    # distinct CIFs with business loans
                "business_cifs":     set(),
                # For CASA ratio
                "casa_total":        0.0,
                "deposits_total":    0.0,
                # Top 100 tracking
                "cif_deposits":      {},   # cif → total deposits
            }
        return rm_dict[code]

    with open(str(acct_csv), encoding="utf-8") as f:
        for row in _csv.DictReader(f):
            rm_code = (row.get("relationship_manager_code") or "").strip()
            if not rm_code:
                continue

            r     = _get(rm, rm_code)
            cat   = (row.get("category") or "").strip()
            atype = (row.get("account_type_name") or "").strip()
            bal   = float(row.get("current_balance") or 0)
            cif   = (row.get("cif") or "").strip()
            dopen = (row.get("date_opened") or "").strip()

            r["account_count"]   += 1
            r["interest_income"] += float(row.get("interest_income_ytd") or 0)
            r["fee_income"]      += float(row.get("fee_income_ytd") or 0)

            # Dormancy
            dorm = (row.get("dormancy_status") or "").strip()
            if dorm.lower() == "dormant":
                r["dormant_count"] += 1

            # New accounts
            if dopen.startswith("2025"):
                r["new_accounts_2025"] += 1
            elif dopen.startswith("2026"):
                r["new_accounts_2026"] += 1

            # ── DEPOSITS ──────────────────────────────────────────────
            if cat == "CASA":
                r["casa_balance"]   += bal
                r["total_deposits"] += bal
                r["deposits_total"] += bal
                r["casa_total"]     += bal
                # Segment: salary / personal → retail; business → commercial
                atype_l = atype.lower()
                if any(x in atype_l for x in
                       ["salary", "junior", "savings", "personal", "graduate",
                        "student", "women", "diaspora"]):
                    r["retail_deposits"] += bal
                else:
                    r["commercial_deposits"] += bal
                # Track per-CIF for Top 100
                if cif:
                    r["cif_deposits"][cif] = r["cif_deposits"].get(cif, 0.0) + bal

            elif cat == "Term Deposit":
                r["term_balance"]   += bal
                r["total_deposits"] += bal
                r["deposits_total"] += bal
                if cif:
                    r["cif_deposits"][cif] = r["cif_deposits"].get(cif, 0.0) + bal

            # ── LOANS ──────────────────────────────────────────────────
            elif cat == "Loan":
                loan_out  = float(row.get("loan_outstanding") or 0)
                loan_orig = float(row.get("loan_amount") or 0)
                npl_stat  = (row.get("npl_status") or "").strip()
                npl_days  = int(float(row.get("npl_days") or 0))

                r["loan_outstanding"] += loan_out
                r["loan_disbursed"]   += loan_orig
                r["loan_count"]       += 1

                # Segment
                if atype in RETAIL_LOAN_TYPES:
                    r["retail_loans"] += loan_out
                elif atype in MSME_LOAN_TYPES:
                    r["msme_loans"]   += loan_out
                    if cif:
                        r["business_cifs"].add(cif)
                elif atype in CORPORATE_LOAN_TYPES:
                    r["corporate_loans"] += loan_out
                else:
                    # default unclassified → MSME
                    r["msme_loans"]   += loan_out
                    if cif:
                        r["business_cifs"].add(cif)

                # NPL / PAR
                if npl_stat == "NPL":
                    r["npl_outstanding"] += loan_out
                if npl_days > 0:
                    r["par_outstanding"] += loan_out

    # ── Post-process derived fields ────────────────────────────────────
    for code, r in rm.items():
        loan_total = r["loan_outstanding"] or 1
        dep_total  = r["deposits_total"]   or 1

        r["npl_ratio"]     = round(r["npl_outstanding"] / loan_total * 100, 2)
        r["par_ratio"]     = round(r["par_outstanding"]  / loan_total * 100, 2)
        r["casa_ratio"]    = round(r["casa_total"]       / dep_total  * 100, 2)
        r["dormancy_pct"]  = round(r["dormant_count"]    / max(r["account_count"], 1) * 100, 2)
        r["business_borrowers"] = len(r["business_cifs"])
        r["nfi"]           = round(r["fee_income"] + r["interest_income"] * 0.15, 2)
        r["pbt"]           = round(
            r["interest_income"] + r["fee_income"]
            - r["loan_outstanding"] * 0.02,   # 2% provision proxy
            2)
        # Top 100 deposits: sum of top-100 CIF balances this RM manages
        sorted_cif_deps = sorted(r["cif_deposits"].values(), reverse=True)
        r["top100_deposits"] = sum(sorted_cif_deps[:100])
        r["new_accounts"]    = r["new_accounts_2025"] + r["new_accounts_2026"]
        # Remove sets (not JSON-serialisable if needed later)
        del r["business_cifs"]
        del r["cif_deposits"]

    return rm


def _map_cbs_to_kpi(kpi_id: str, kpi_name: str, rm_data: dict) -> float:
    """Map KPI ID to pre-aggregated rm_data field from aggregate_cbs_by_rm()."""
    ID_MAP = {
        "RETAIL_MSME_DEPOSIT": "retail_deposits",
        "DEP_GROWTH":          "total_deposits",
        "COMMERCIAL_DEPOSIT":  "commercial_deposits",
        "CASA_RATIO":          "casa_ratio",
        "TOP100_CUSTOMERS":    "top100_deposits",
        "LOAN_GROWTH":         "loan_outstanding",
        "LOAN_DISB":           "loan_disbursed",
        "DISB_RETAIL":         "retail_loans",
        "DISB_MSME":           "msme_loans",
        "DISB_CORPORATE":      "corporate_loans",
        "BUSINESS_BORROWERS":  "business_borrowers",
        "NPL_RATIO":           "npl_ratio",
        "PAR":                 "par_ratio",
        "COLLECTION_THROUGHPUT": "npl_ratio",
        "TOTAL_NFI":           "nfi",
        "FEES_COMM":           "fee_income",
        "NIM":                 "nfi",
        "PBT":                 "pbt",
        "ACCOUNT_DORMANCY":    "dormancy_pct",
        "CHANNEL_DORMANCY":    "dormancy_pct",
        "NEW_ACCOUNTS":        "new_accounts",
    }
    # Try exact KPI ID first
    field = ID_MAP.get(kpi_id.upper())
    if field:
        return float(rm_data.get(field, 0) or 0)
    # Name-based fallback
    name_l = kpi_name.lower()
    if "retail" in name_l and "deposit" in name_l:
        return float(rm_data.get("retail_deposits", 0))
    if "commercial" in name_l and "deposit" in name_l:
        return float(rm_data.get("commercial_deposits", 0))
    if "deposit" in name_l:
        return float(rm_data.get("total_deposits", 0))
    if "casa ratio" in name_l:
        return float(rm_data.get("casa_ratio", 0))
    if "top 100" in name_l or "top100" in name_l:
        return float(rm_data.get("top100_deposits", 0))
    if "disburs" in name_l and "retail" in name_l:
        return float(rm_data.get("retail_loans", 0))
    if "disburs" in name_l and "msme" in name_l:
        return float(rm_data.get("msme_loans", 0))
    if "disburs" in name_l and "corporate" in name_l:
        return float(rm_data.get("corporate_loans", 0))
    if "loan" in name_l or "disburs" in name_l:
        return float(rm_data.get("loan_outstanding", 0))
    if "npl" in name_l:
        return float(rm_data.get("npl_ratio", 0))
    if name_l.strip() == "par" or "portfolio at risk" in name_l:
        return float(rm_data.get("par_ratio", 0))
    if "collection" in name_l or "throughput" in name_l:
        return float(rm_data.get("npl_ratio", 0))
    if "nfi" in name_l or ("non" in name_l and "funded" in name_l):
        return float(rm_data.get("nfi", 0))
    if "fee" in name_l or "commission" in name_l:
        return float(rm_data.get("fee_income", 0))
    if "pbt" in name_l or "profit" in name_l:
        return float(rm_data.get("pbt", 0))
    if "dormancy" in name_l or "dormant" in name_l:
        return float(rm_data.get("dormancy_pct", 0))
    if "new account" in name_l:
        return float(rm_data.get("new_accounts", 0))
    if "business borrower" in name_l:
        return float(rm_data.get("business_borrowers", 0))
    return 0.0

def _build_from_cbs(staff_list, role_kpis, id_to_kpi, kpi_weights,
                     pillar_weights, cbs_dir, period):
    """
    Build KPI actuals rows per staff from CBS data.
    Uses aggregate_cbs_by_rm() for a clean single-pass read, then
    maps each staff member's CBS data to their assigned KPIs.
    Only staff who are relationship managers (have accounts in CBS)
    get CBS actuals; all other staff get zeros which the qualitative
    baselines (Compliance, Audit etc.) will fill separately.
    """
    # ── Single-pass CBS aggregation ────────────────────────────────────
    rm_data        = aggregate_cbs_by_rm(cbs_dir)
    branch_kpis    = aggregate_cbs_by_branch(cbs_dir)
    unit_to_branch = _build_unit_to_branch(cbs_dir)
    # Bank-wide averages for branches not in CBS simulation
    _branch_vals = list(branch_kpis.values())
    bank_avg_kpis: dict = {}
    if _branch_vals:
        from collections import defaultdict as _dd_avg
        _agg = _dd_avg(list)
        for _bv in _branch_vals:
            for _k, _v in _bv.items():
                if _v and float(_v) > 0:
                    _agg[_k].append(float(_v))
        bank_avg_kpis = {_k: sum(_vs)/len(_vs) for _k, _vs in _agg.items()}

    rows = []
    for staff in staff_list:
        sc   = str(staff.get("Staff Code", "") or "").strip()
        role = str(staff.get("Role", "")       or "").strip()
        if not sc or not role:
            continue

        kpi_ids = role_kpis.get(role, [])
        if not kpi_ids:
            continue

        # This staff member's CBS data (empty dict if not an RM)
        cbs = rm_data.get(sc, {})
        # Branch-level data for non-RM staff (Tellers, CSOs, BOS etc.)
        unit = str(staff.get("Unit", "") or "").strip()
        bc   = unit_to_branch.get(unit, "")
        if bc:
            branch_data = branch_kpis.get(bc, {})
        elif unit not in ("Head Office", "HO", ""):
            # Unmatched branch — use bank-wide average as proxy
            branch_data = bank_avg_kpis
        else:
            branch_data = {}

        for kpi_id in kpi_ids:
            kpi_def = id_to_kpi.get(kpi_id)
            if not kpi_def:
                continue
            kpi_name = kpi_def["name"]
            pillar   = kpi_def["pillar"]
            weight   = kpi_weights.get(kpi_id, kpi_def.get("default_weight", 0.05))

            # RM has personal portfolio; branch staff use branch aggregate
            if cbs:
                actual = _map_cbs_to_kpi(kpi_id, kpi_name, cbs)
            else:
                actual = float(branch_data.get(kpi_name, 0) or 0)

            # Disbursements Corporate Loans: only for corporate banking roles
            if kpi_name == "Disbursements Corporate Loans":
                _corp_role_keywords = ("corporate", "institutional", "government")
                _is_corp_role = any(x in role.lower() for x in _corp_role_keywords)
                if not _is_corp_role:
                    actual = 0.0

            rows.append({
                "Staff Code":    sc,
                "Staff Name":    str(staff.get("Staff Name", "")),
                "Role":          role,
                "Unit":          str(staff.get("Unit", "")),
                "Category":      str(staff.get("Category", "")),
                "Staff Status":  str(staff.get("Staff Status", "Active")),
                "KPI":           kpi_name,
                "kpi_id":        kpi_id,  # for bsc_engine contract submission
                "Pillar":        pillar,
                "Weight":        weight,
                "Annual Target": 0,      # filled by inject_cascade_targets()
                "YTD_Actual":    actual,
                period:          actual,
                "Annual Actual": actual,
            })
    return rows



def _add_initiative_kpis(rows, staff_list, lib, period):
    """Inject initiative KPIs for HO roles."""
    try:
        from utils.core import compute_initiative_kpis
        role_kpis = lib.get("role_kpis", {})
        kpi_weights = lib.get("kpi_weights", {})
        seen = set()

        for r in list(rows):
            sn   = r["Staff Name"]
            role = r["Role"]
            if (sn, "INIT_STATUS") in seen:
                continue
            if "INIT_STATUS" not in role_kpis.get(role, []):
                continue
            seen.add((sn, "INIT_STATUS"))
            ia = compute_initiative_kpis(sn)
            base = {k: v for k, v in r.items()
                    if k not in ("KPI", "Pillar", "Weight", "Annual Target",
                                 "YTD_Actual", period, "Annual Actual")}
            for kid, kname, pillar, tgt in [
                ("INIT_STATUS", "Initiative Implementation Score",
                 "Operational Excellence", 100),
                ("INIT_COUNT", "Active Initiatives Count",
                 "Operational Excellence", 5),
            ]:
                if kid in role_kpis.get(role, []):
                    v = ia.get(kname, 0)
                    rows.append({**base, "KPI": kname, "Pillar": pillar,
                                  "Weight": kpi_weights.get(kid, 0.03),
                                  "Annual Target": tgt,
                                  "YTD_Actual": v,
                                  period: v, "Annual Actual": v})
    except Exception:
        pass
    # ── Inject DRS recovery actuals ──────────────────────────────────
    try:
        # v10.352 — DATA_DIR was undefined here (latent NameError). Resolve
        # the data directory via the same helper other functions in this
        # module use.
        _cbs_dir, _data_dir = get_cbs_paths()
        ra_file = _data_dir.parent / "data" / "recovery_actuals.json"
        if not ra_file.exists():
            ra_file = _data_dir / "recovery_actuals.json"
        if ra_file.exists():
            ra_data = json.loads(ra_file.read_text())
            for row in rows:
                staff_name = row.get("Staff Name","")
                if staff_name in ra_data and row.get("KPI") == "Collection Throughput":
                    row["YTD_Actual"]    = ra_data[staff_name].get("COLLECTION_THROUGHPUT", 0)
                    row[period]          = row["YTD_Actual"]
                    row["Annual Actual"] = row["YTD_Actual"]
    except: pass

    # ── Bank-aggregate actuals for HO/exec roles ─────────────────────
    try:
        from utils.core import get_org_config as _goc_ae
        _org_cfg = _goc_ae()
        _bank_roles = _get_bank_aggregate_roles(_org_cfg)
        _bank_agg   = compute_bank_aggregates(cbs_dir)
        for row in rows:
            if row.get("Role","") not in _bank_roles:
                continue
            kpi_name = row.get("KPI","")
            new_val  = _bank_agg.get(kpi_name)
            if new_val is not None and float(new_val) > 0:
                row["YTD_Actual"]    = float(new_val)
                row[period]          = float(new_val)
                row["Annual Actual"] = float(new_val)
    except: pass

    # ── Fix Diligence Score scale (must be on 1-5 BSC scale, not %) ──
    for row in rows:
        if row.get("KPI") == "Diligence Score":
            act = row.get("YTD_Actual", 0)
            if act > 5.0:   # wrong scale detected
                row["YTD_Actual"]    = 2.8
                row[period]          = 2.8
                row["Annual Actual"] = 2.8

    # ── LMS actuals — disbursements, loan growth, business borrowers ────
    try:
        from utils.core import LoanApplicationManager as _LAM, ComplianceManager as _CM
        _lam = _LAM()
        _lms_kpis = _lam.bsc_actuals()   # {rm_code: {kpi_name: value}}
        for row in rows:
            sc_v     = str(row.get("Staff Code",""))
            kpi_name = row.get("KPI","")
            rm_kpis  = _lms_kpis.get(sc_v, {})
            lms_val  = rm_kpis.get(kpi_name)
            if lms_val and float(lms_val) > 0:
                row["YTD_Actual"]    = float(lms_val)
                row[period]          = float(lms_val)
                row["Annual Actual"] = float(lms_val)
        # Compliance score from ComplianceManager
        _comp_score = _CM().bsc_compliance_score()
        for row in rows:
            if row.get("KPI") == "Compliance Score":
                row["YTD_Actual"]    = _comp_score
                row[period]          = _comp_score
                row["Annual Actual"] = _comp_score
    except Exception:
        pass

    # ── Active Initiatives Count — from execute_initiatives.json ──────
    try:
        ini_path = Path(__file__).parent.parent / "data" / "execute_initiatives.json"
        if ini_path.exists():
            from collections import defaultdict as _dd_ini
            _ei = json.loads(ini_path.read_text())
            _staff_ini = _dd_ini(int)
            for _ini in _ei:
                _status = str(_ini.get("status","") or "").lower()
                _owner  = str(_ini.get("owner_code","") or _ini.get("owner","") or "")
                if _status in ("active","in progress","g1","g2","g3","g4","g5") and _owner:
                    _staff_ini[_owner] += 1
                for _m in (_ini.get("team_members") or []):
                    _msc = str(_m.get("staff_code","") or "")
                    if _msc: _staff_ini[_msc] += 1
            for row in rows:
                if row.get("KPI") == "Active Initiatives Count":
                    sc_v  = str(row.get("Staff Code",""))
                    count = _staff_ini.get(sc_v, 0)
                    row["YTD_Actual"]    = float(count)
                    row[period]          = float(count)
                    row["Annual Actual"] = float(count)
    except Exception:
        pass

    return rows


def _inject_initiative_kpis(xlsx_path, staff_list, lib):
    """Add initiative KPIs to an existing actuals xlsx."""
    try:
        import openpyxl
        from utils.core import compute_initiative_kpis
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        nc = headers.index("Staff Name") if "Staff Name" in headers else 1
        rc = headers.index("Role")       if "Role"       in headers else 2
        kc = headers.index("KPI")        if "KPI"        in headers else 6

        existing_kpis = set()
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:
                existing_kpis.add((str(row[nc] or ""), str(row[kc] or "")))

        role_kpis   = lib.get("role_kpis", {})
        kpi_weights = lib.get("kpi_weights", {})

        added = 0
        for row_idx in range(3, ws.max_row + 1):
            sn   = str(ws.cell(row_idx, nc + 1).value or "")
            role = str(ws.cell(row_idx, rc + 1).value or "")
            if not sn or "INIT_STATUS" not in role_kpis.get(role, []):
                continue
            if (sn, "Initiative Implementation Score") in existing_kpis:
                continue
            ia = compute_initiative_kpis(sn)
            # Get template row values
            tmpl = [ws.cell(row_idx, c + 1).value for c in range(len(headers))]
            for kid, kname, pillar, tgt in [
                ("INIT_STATUS", "Initiative Implementation Score", "Operational Excellence", 100),
                ("INIT_COUNT",  "Active Initiatives Count",        "Operational Excellence", 5),
            ]:
                if kid in role_kpis.get(role, []):
                    new_row = list(tmpl)
                    new_row[kc]     = kname
                    new_row[kc - 1] = pillar  # Pillar col before KPI
                    v = ia.get(kname, 0)
                    ws.append(new_row[:kc] + [kname, kpi_weights.get(kid, 0.03),
                                               tgt, v, v, v])
                    existing_kpis.add((sn, kname))
                    added += 1

        if added:
            wb.save(str(xlsx_path))
    except Exception:
        pass




def compute_bank_aggregates(cbs_dir: Path) -> dict:
    """
    Compute bank-wide totals from accounts.csv.
    Used for HO/exec roles who should see bank-level actuals,
    not just their personal RM portfolio.
    Returns dict: kpi_name → value
    """
    import csv as _csv2
    from collections import defaultdict as _dd

    acct_csv = None
    for name in ["accounts.csv", "cbs_accounts.csv"]:
        p = cbs_dir / name
        if p.exists():
            acct_csv = p
            break
    if not acct_csv:
        return {}

    bank = _dd(float);  bi = _dd(int)
    RETAIL_L = {"personal loan","mortgage / home loan","salary advance",
                "graduate loan","consumer loan","staff loan"}
    MSME_L   = {"business loan","lpo finance","asset finance",
                "invoice discounting","business overdraft"}

    with open(str(acct_csv), encoding="utf-8") as f:
        for row in _csv2.DictReader(f):
            cat   = row.get("category","")
            atype = (row.get("account_type_name") or "").lower().strip()
            bal   = float(row.get("current_balance") or 0)
            lo    = float(row.get("loan_outstanding") or 0)
            la    = float(row.get("loan_amount") or 0)
            npl   = row.get("npl_status","")
            nd    = int(float(row.get("npl_days") or 0))
            dopen = row.get("date_opened","")
            dorm  = row.get("dormancy_status","")
            fee   = float(row.get("fee_income_ytd") or 0)
            intc  = float(row.get("interest_income_ytd") or 0)

            bi["total"] += 1
            bank["fee"] += fee;  bank["int"] += intc
            if dopen[:4] in ("2025","2026"): bi["new_accts"] += 1
            if dorm.lower() == "dormant":    bi["dormant"]   += 1

            if cat == "CASA":
                bank["dep"] += bal;  bank["casa"] += bal
                if any(x in atype for x in ["salary","junior","savings","personal",
                                              "graduate","student","women","diaspora"]):
                    bank["retail_dep"] += bal
                else:
                    bank["comm_dep"] += bal
            elif cat == "Term Deposit":
                bank["dep"] += bal
            elif cat == "Loan":
                bank["loans"] += lo;  bank["disb"] += la
                if npl == "NPL":  bank["npl"] += lo
                if nd > 0:        bank["par"] += lo
                if atype in RETAIL_L:   bank["retail_l"] += lo
                elif atype in MSME_L:   bank["msme_l"] += lo;  bi["biz_borrow"] += 1
                else:                   bank["msme_l"] += lo

    lt = bank["loans"] or 1;  dt = bank["dep"] or 1;  at = bi["total"] or 1

    # ── v10.364: compute PBT via proper P&L breakdown ─────────────────
    # Replaces the v10.34x naive placeholder (bank[int] + bank[fee] - bank[loans]*0.02)
    # with a real Operating Income - OpEx - Impairment computation. All
    # factors configurable in data/opex_data.json + data/pbt_assumptions.json.
    # See utils/pbt_computation.py for the full P&L drill-down.
    try:
        from utils.pbt_computation import compute_pbt_from_cbs as _compute_pbt
        _pbt_components = _compute_pbt(cbs_dir)
        _pbt_value = float(_pbt_components.pbt)
        _nii_value = float(_pbt_components.nii)
        _cir_value = (
            float(_pbt_components.total_opex / _pbt_components.operating_income * 100)
            if _pbt_components.operating_income > 0 else 0.0
        )
    except Exception:
        # Fallback to legacy naive estimate if pbt_computation is unavailable
        # (won't happen normally — module is imported lazily for clean separation)
        _pbt_value = round(bank["int"] + bank["fee"] - bank["loans"] * 0.02, 2)
        _nii_value = bank["int"]
        _cir_value = 0.0

    return {
        "Retail & MSME Deposit Growth":  bank["retail_dep"],
        "Commercial Deposit Growth":      bank["comm_dep"],
        "Deposit Growth":                 bank["dep"],
        "Loan Book Growth":               bank["loans"],
        "Loans Disbursement":             bank["disb"],
        "Disbursements Retail Loans":     bank["retail_l"],
        "Disbursements MSME Loans":       bank["msme_l"],
        "Disbursements Corporate Loans":  bank.get("corp_l", 0),
        "NPL Ratio":                      round(bank["npl"] / lt * 100, 2),
        "PAR":                            round(bank["par"] / lt * 100, 2),
        "CASA Ratio":                     round(bank["casa"] / dt * 100, 2),
        "Total NFI":                      round(bank["fee"] + bank["int"] * 0.15, 2),
        "Fees and Commission":            bank["fee"],
        "PBT":                            _pbt_value,
        "NII":                            _nii_value,
        "CIR":                            round(_cir_value, 2),
        "Account Dormancy":               round(bi["dormant"] / at * 100, 2),
        "Channel Dormancy":               round(bi["dormant"] / at * 100, 2),
        "New Accounts":                   bi["new_accts"],
        "Number of Business Borrowers":   bi["biz_borrow"],
        "Top 100 Customers Deposit":      bank["dep"],
        "Staff Productivity":             3.0,
        "CX Score":                       3.2,
        "Compliance Score":               85.0,
        "Audit Score":                    82.0,
        "Diligence Score":                2.8,
    }


def _get_bank_aggregate_roles(org_config: dict) -> set:
    """CEO + all direct reports of CEO = bank-aggregate roles."""
    hier  = org_config.get("hierarchy", {})
    roots = [r for r, p in hier.items() if not p]
    if not roots:
        return set()
    ceo = roots[0]
    return {ceo} | {r for r, p in hier.items() if ceo in p}



def aggregate_cbs_by_branch(cbs_dir: Path) -> dict:
    """
    Aggregate CBS accounts.csv by branch_code.
    Returns dict: branch_code → KPI-name → value.
    Used to give branch staff (Tellers, CSOs, BOS etc.) their branch-level actuals.
    """
    import csv as _csv3
    from collections import defaultdict as _dd3

    acct_csv = None
    for name in ["accounts.csv", "cbs_accounts.csv"]:
        p = cbs_dir / name
        if p.exists():
            acct_csv = p
            break
    if not acct_csv:
        return {}

    RETAIL_L = {"personal loan", "mortgage / home loan", "salary advance",
                "graduate loan", "consumer loan", "staff loan"}
    MSME_L   = {"business loan", "lpo finance", "asset finance",
                "invoice discounting", "business overdraft"}

    branch: _dd3 = _dd3(lambda: _dd3(float))

    with open(str(acct_csv), encoding="utf-8") as f:
        for row in _csv3.DictReader(f):
            bc    = (row.get("branch_code") or "").strip()
            if not bc:
                continue
            cat   = row.get("category", "")
            atype = (row.get("account_type_name") or "").lower()
            bal   = float(row.get("current_balance") or 0)
            lo    = float(row.get("loan_outstanding") or 0)
            npl   = row.get("npl_status", "")
            nd    = int(float(row.get("npl_days") or 0))
            dopen = row.get("date_opened", "")
            dorm  = row.get("dormancy_status", "")
            fee   = float(row.get("fee_income_ytd") or 0)
            intc  = float(row.get("interest_income_ytd") or 0)
            cif   = row.get("cif", "")

            b = branch[bc]
            b["total_accts"] += 1
            b["fee"] += fee
            b["int"] += intc
            if dopen[:4] in ("2025", "2026"): b["new_accts"] += 1
            if dorm.lower() == "dormant":     b["dormant"]   += 1

            if cat == "CASA":
                b["dep"] += bal
                b["casa"] += bal
                if any(x in atype for x in ["salary","junior","savings","personal",
                                              "graduate","student","women","diaspora"]):
                    b["retail_dep"] += bal
                else:
                    b["comm_dep"] += bal
            elif cat == "Term Deposit":
                b["dep"] += bal
            elif cat == "Loan":
                b["loans"] += lo
                if npl == "NPL": b["npl"] += lo
                if nd > 0:       b["par"] += lo
                if atype in RETAIL_L:   b["retail_l"] += lo
                elif atype in MSME_L:   b["msme_l"]   += lo;  b["biz_count"] += 1
                else:                   b["msme_l"]   += lo

    # Build KPI-name keyed dict
    result = {}
    for bc, b in branch.items():
        lt = b["loans"] or 1;  dt = b["dep"] or 1;  at = b["total_accts"] or 1
        result[bc] = {
            "Retail & MSME Deposit Growth": b["retail_dep"],
            "Commercial Deposit Growth":     b["comm_dep"],
            "Deposit Growth":                b["dep"],
            "Loan Book Growth":              b["loans"],
            "Disbursements Retail Loans":    b["retail_l"],
            "Disbursements MSME Loans":      b["msme_l"],
            "Disbursements Corporate Loans": b.get("corp_l", 0),
            "Number of Business Borrowers":  b["biz_count"],
            "NPL Ratio":                     round(b["npl"] / lt * 100, 2),
            "PAR":                           round(b["par"] / lt * 100, 2),
            "CASA Ratio":                    round(b["casa"] / dt * 100, 2),
            "Total NFI":                     round(b["fee"] + b["int"] * 0.15, 2),
            "Fees and Commission":           b["fee"],
            "PBT":                           round(b["int"] + b["fee"] - b["loans"] * 0.02, 2),
            "Account Dormancy":              round(b["dormant"] / at * 100, 2),
            "Channel Dormancy":              round(b["dormant"] / at * 100, 2),
            "New Accounts":                  b["new_accts"],
            "Top 100 Customers Deposit":     b["dep"],
            "Collection Throughput":         round(b["npl"] / lt * 100, 2),
        }
    return result


def _build_unit_to_branch(cbs_dir: Path) -> dict:
    """Map staff register Unit name → branch_code via CBS branch_name."""
    import csv as _csv4
    acct_csv = cbs_dir / "accounts.csv"
    if not acct_csv.exists():
        return {}
    br_map = {}
    with open(str(acct_csv), encoding="utf-8") as f:
        for row in _csv4.DictReader(f):
            bn = (row.get("branch_name") or "").strip()
            bc = (row.get("branch_code") or "").strip()
            if bn and bc:
                br_map[bn] = bc
                short = bn.replace(" Branch", "").replace(" Main", "").strip()
                br_map[short] = bc
    return br_map

def _write_actuals(rows, out_path, period):
    """Write rows to Excel with formatted header."""
    if not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Data"

    headers = list(rows[0].keys())
    # Header row 1 — title
    ws.append(["A2Z Blueprint — KPI Actuals"] + [""] * (len(headers) - 1))
    # Header row 2 — column names
    ws.append(headers)
    # Data rows
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[2]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="006B3F")
        cell.alignment = Alignment(horizontal="center")

    wb.save(str(out_path))


def _count_rows(xlsx_path):
    """Count data rows in actuals file."""
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        return wb.active.max_row - 2
    except:
        return 0


# ─── v10.108 — Integration Layer: operational-table autofit tributary ──
#
# This is the second pathway alongside compute_actuals_from_cbs(). The
# CBS pathway feeds ~24 strategic-tier KPIs from CBS aggregations. This
# operational pathway feeds ~87 KPIs that are computed from operational
# tables (loan_applications, debt_recovery, pipeline, aml_screenings,
# etc.) using rules registered in utils/kpi_aggregation_rules.py.
#
# The ownership gate (utils/kpi_ownership.is_kpi_owned_by_staff) ensures
# we only submit actuals for staff who actually own the KPI for the
# given period — either through role_kpis or through cascade lock.

def _read_data_source_config(data_dir: Path) -> dict:
    """Read the `_data_source` config knob from
    integration_layer_config.json. Returns the resolved configuration:

      {
        "default":     "json"|"pg_view"|"auto",
        "per_table":   {"<table_name>": "json"|"pg_view"|"auto", ...}
      }

    The default for missing config is `"json"` (current behavior — keeps
    every existing deployment working unchanged). v10.116 closes the
    JSON-deprecation blueprint gap by making this knob honor PG views
    without a code change.

    Mode semantics:
      json     — read data/<table>.json (current default)
      pg_view  — SELECT * FROM <table>; raise on failure
      auto     — try pg_view first; fall back to json on any failure

    The per_table dict allows progressive migration — a bank can move
    one table at a time from JSON to a PG view, leaving others on JSON
    until ready.
    """
    cfg_path = data_dir / "integration_layer_config.json"
    if not cfg_path.exists():
        return {"default": "json", "per_table": {}}
    try:
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception:
        return {"default": "json", "per_table": {}}

    raw = cfg.get("_data_source", "json")
    # Allow the simple shorthand `"_data_source": "json"` (string)
    if isinstance(raw, str):
        return {"default": raw, "per_table": {}}
    # Or the structured form `{"default": ..., "per_table": {...}}`
    if isinstance(raw, dict):
        return {
            "default":   raw.get("default", "json"),
            "per_table": raw.get("per_table", {}) or {},
        }
    return {"default": "json", "per_table": {}}


def _read_operational_table(table: str, data_dir: Path) -> list[dict]:
    """Read all rows from an operational table.

    Per the A2Z storage convention, operational tables are JSON files
    in data/<table>.json (post-Phase-1A migration, PostgreSQL is dual-
    write but JSON is still authoritative for read paths until the cut-
    over). Returns [] for missing files — the autofit pipeline treats
    missing tables as "no actuals to submit", not as errors.

    **v10.116 PG-readiness shim**: respects the `_data_source` knob in
    integration_layer_config.json. Default behavior (json) is unchanged;
    setting `_data_source: "pg_view"` for a table reads from a PG view
    of the same name. `auto` mode tries PG first and silently falls
    back to JSON. This is the loader-side support for the JSON
    deprecation roadmap from the architectural blueprint.
    """
    cfg = _read_data_source_config(data_dir)
    mode = cfg["per_table"].get(table, cfg["default"])

    if mode in ("pg_view", "auto"):
        rows = _try_read_from_pg_view(table)
        if rows is not None:
            return rows
        if mode == "pg_view":
            # Strict mode: caller asked for PG explicitly and we couldn't
            # produce data. Don't silently downgrade to JSON — that
            # masks deployment misconfiguration. Return [] so the rule
            # is reported as 'no actuals' rather than served stale data.
            logger.warning(
                f"_read_operational_table({table}): pg_view mode but "
                f"PG read failed/unavailable — returning empty rows. "
                f"Switch _data_source to 'auto' to allow JSON fallback.")
            return []
        # auto mode falls through to JSON fallback

    # Default / fallback path — read from JSON file
    path = data_dir / f"{table}.json"
    if not path.exists():
        return []
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.warning(f"_read_operational_table({table}): {e}")
        return []

    # A2Z JSON tables are sometimes a list, sometimes a dict keyed by id.
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(data, dict):
        # Dict-keyed: yield the values that are dicts
        return [r for r in data.values() if isinstance(r, dict)]
    return []


def _try_read_from_pg_view(table: str) -> list[dict] | None:
    """Attempt to read an operational table from a PG view of the same
    name. Returns the list of row-dicts on success, None on any failure
    (no PG configured, view doesn't exist, query error). Used by the
    PG-readiness shim in `_read_operational_table`.

    **Safety**: the table identifier is validated against a
    conservative whitelist regex before being interpolated into SQL.
    SQL injection via this path is not possible because:
      1. Table names come from the rule registry (curated, not user
         input)
      2. The whitelist regex `^[a-z][a-z0-9_]{0,62}$` rejects anything
         that isn't a plain lowercase identifier
      3. We use psycopg2.sql.Identifier composition for the actual
         interpolation, not f-strings
    """
    import re
    if not re.match(r"^[a-z][a-z0-9_]{0,62}$", table):
        logger.warning(
            f"_try_read_from_pg_view({table}): identifier rejected by "
            f"whitelist; falling back")
        return None
    try:
        from utils.db import db as a2z_db, _USE_DB
        if not _USE_DB:
            return None
    except Exception:
        return None
    try:
        from psycopg2 import sql as _pg_sql
        sql = _pg_sql.SQL("SELECT * FROM {}").format(
            _pg_sql.Identifier(table))
        rows = a2z_db.fetch_all(sql.as_string(a2z_db._connection_template())
                                 if hasattr(a2z_db, '_connection_template')
                                 else f"SELECT * FROM {table}")
        return rows if isinstance(rows, list) else None
    except Exception as e:
        logger.debug(
            f"_try_read_from_pg_view({table}): {type(e).__name__}: {e}")
        return None


def compute_actuals_from_operational_tables(period: str) -> dict:
    """Second autofit tributary — operational tables → BSC actuals.

    For each registered rule in utils/kpi_aggregation_rules.REGISTRY:
        1. Read the operational table.
        2. Apply the rule (compute_rule groups by staff and aggregates).
        3. For each (staff, value) pair: check ownership via
           kpi_ownership.is_kpi_owned_by_staff.
        4. If owned: submit to bsc_engine via _submit_to_bsc_engine.
        5. If not owned: drop silently (this is the gate that prevents
           cascade-misallocated actuals from polluting the BSC).

    Returns a status dict:
        {
          "success":           bool,
          "period":            <period>,
          "rules_processed":   <int>,
          "rules_skipped":     <int>,        # tables not found
          "actuals_submitted": <int>,        # passed ownership gate
          "actuals_dropped":   <int>,        # failed ownership gate
          "by_rule":           [
              {"kpi_id": K011,
               "table": loan_applications,
               "staff_count": 17,
               "submitted": 12,
               "dropped": 5},
              ...
          ],
          "engine_summary":    <dict from bsc_engine.submit_batch>,
          "duration_s":        <float>,
        }
    """
    from datetime import datetime as _dt
    t0 = _dt.now()

    try:
        from utils import kpi_aggregation_rules as _rules
        from utils import kpi_ownership as _own
        from utils.staff_field_resolver import resolve_staff_field
        from utils.bsc_engine import submit_batch as _bsc_submit_batch
    except ImportError as e:
        return {
            "success": False,
            "period": period,
            "message": f"v10.108 Integration Layer modules not available: {e}",
            "rules_processed": 0,
            "actuals_submitted": 0,
            "duration_s": 0.0,
        }

    _, data_dir = get_cbs_paths()

    contract_records: list[dict] = []
    by_rule: list[dict] = []
    rules_processed = 0
    rules_skipped = 0
    actuals_dropped = 0

    for rule in _rules.REGISTRY:
        rows = _read_operational_table(rule.source_table, data_dir)
        if not rows:
            rules_skipped += 1
            by_rule.append({
                "kpi_id": rule.kpi_id,
                "table": rule.source_table,
                "staff_count": 0,
                "submitted": 0,
                "dropped": 0,
                "skip_reason": "table_missing_or_empty",
            })
            continue

        rules_processed += 1
        staff_field = resolve_staff_field(rule.source_table, rule.staff_field)
        per_staff = _rules.compute_rule(rule, rows, period, staff_field)

        rule_submitted = 0
        rule_dropped = 0
        for staff_code, value in per_staff.items():
            if _own.is_kpi_owned_by_staff(staff_code, rule.kpi_id, period):
                contract_records.append({
                    "staff_code": str(staff_code),
                    "kpi_id":     rule.kpi_id,
                    "value":      value,
                    "period":     str(period),
                })
                rule_submitted += 1
            else:
                rule_dropped += 1
                actuals_dropped += 1

        by_rule.append({
            "kpi_id": rule.kpi_id,
            "table": rule.source_table,
            "staff_count": len(per_staff),
            "submitted": rule_submitted,
            "dropped": rule_dropped,
        })

    # Submit the batch through the central BSC engine
    if contract_records:
        engine_summary = _bsc_submit_batch(
            records       = contract_records,
            source_module = "actuals_engine.operational",
            actor         = "operational_autofit",
        )
    else:
        engine_summary = {"ok": 0, "rejected": 0,
                          "errors": [], "skipped": "no_records"}

    duration = (_dt.now() - t0).total_seconds()
    return {
        "success": True,
        "period": period,
        "rules_processed": rules_processed,
        "rules_skipped": rules_skipped,
        "actuals_submitted": len(contract_records),
        "actuals_dropped": actuals_dropped,
        "by_rule": by_rule,
        "engine_summary": engine_summary,
        "duration_s": duration,
    }
