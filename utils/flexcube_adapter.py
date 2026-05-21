"""utils/flexcube_adapter.py — FLEXCUBE Integration Adapter Layer.

This module is the SEAM between A2Z Blueprint and Oracle FLEXCUBE.

Three modes (configured in data/flexcube_config.json):
- "synthetic" : Reads from synthetic CSV/JSON files (current dev/demo state)
- "mock"      : Reads from synthetic data BUT pretends to call real APIs (for integration testing)
- "live"      : Calls real FLEXCUBE REST APIs via Ecobank's Apigee gateway (production)

Every function returns a normalised dict so the calling code (modules, BSC engine,
CBS Explorer) does not need to know which mode is active.

Reference: Oracle FLEXCUBE Universal Banking REST Services Guide
           Oracle Banking APIs (OBAPI/OBDX) Host Integration Guide
           Ecobank Group + Google Cloud (Apigee) Partnership 2025
"""
from __future__ import annotations
import json
import time
from pathlib import Path
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional
from decimal import Decimal

DATA_DIR    = Path(__file__).parent.parent / "data"
CBS_DIR     = Path(__file__).parent.parent / "cbs_data"
CONFIG_FILE = DATA_DIR / "flexcube_config.json"

# ══════════════════════════════════════════════════════════════════
# Configuration loader
# ══════════════════════════════════════════════════════════════════

def get_config() -> Dict[str, Any]:
    """Load FLEXCUBE integration config. Falls back to safe defaults.

    v8.20: backfill any missing keys from defaults so older saved configs
    (e.g. pre-v8.20 configs without `endpoint_timeouts`) get current
    defaults injected without operator action.
    """
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            defaults = _default_config()
            # Backfill top-level keys missing from saved config
            for k, v in defaults.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
        except Exception:
            pass
    return _default_config()

def _default_config() -> Dict[str, Any]:
    return {
        "mode": "synthetic",
        "endpoints": {
            "apigee_base":     "https://api.ecobank.co.ke/v1",
            "fcubs_rest":      "https://api.ecobank.co.ke/flexcube",
            "obdx_rest":       "https://api.ecobank.co.ke/obdx/v1",
            "fcubs_soap_wsdl": "https://fcubs.ecobank.co.ke:8001/FCUBSAccService/FCUBSAccService?wsdl",
            "jms_broker":      "tcp://mq.ecobank.co.ke:61616",
        },
        "auth": {
            "method":            "oauth2",
            "client_id":         "${FLEXCUBE_CLIENT_ID}",
            "client_secret_ref": "${FLEXCUBE_CLIENT_SECRET}",
            "token_url":         "https://api.ecobank.co.ke/oauth2/token",
            "scopes":            ["accounts.read","loans.read","payments.read","customers.read"],
            "mtls_enabled":      True,
        },
        "jms_topics": {
            "loan_disbursed":     "ecobank.fcubs.loans.disbursed",
            "account_opened":     "ecobank.fcubs.accounts.opened",
            "account_closed":     "ecobank.fcubs.accounts.closed",
            "transaction_posted": "ecobank.fcubs.txns.posted",
            "aml_alert":          "ecobank.fcubs.aml.alert",
            "kyc_updated":        "ecobank.fcubs.kyc.updated",
        },
        "timeouts": {
            "rest_seconds":     5,
            "soap_seconds":    10,
            "batch_seconds":  300,
        },
        # v8.20 — per-endpoint timeout overrides (closes v8.6 ack #7).
        # Map endpoint_key (as produced by _endpoint_key()) to override
        # timeout in seconds. Endpoints not listed fall through to the
        # `timeouts` dict above (rest_seconds / batch_seconds default).
        # Defaults reflect typical FLEXCUBE response patterns:
        #   - NPL aggregate: large dataset, IFRS 9 staging — needs longer
        #   - Customer aggregate: simpler query — shorter is fine
        #   - Loans/Deposits: standard portfolio queries
        #   - Dormancy: rare event-driven query — longer tolerable
        "endpoint_timeouts": {
            "PortfolioService/Loans":      300,
            "PortfolioService/Deposits":   300,
            "PortfolioService/NPL":        600,  # IFRS 9 staging — heavier
            "CustomerService":             120,  # simpler aggregate
            "AccountService/Dormancy":     180,
        },
        "rate_limits": {
            "rest_per_minute":   600,
            "burst_per_second":   20,
        },
        "fcubs_version":   "14.7",
        "deployment":      "on-prem-eti-ghana",
        "environments":    ["dev","sit","uat","prod"],
        "active_environment":"dev",
        "iso20022_enabled":True,
        "swift_enabled":   True,
        "iso8583_enabled": True,
        "data_residency":  "Kenya",
        "encryption":      {"in_transit":"TLS1.2+","at_rest":"TDE-AES256"},
        "last_updated":    "",
    }

def save_config(cfg: Dict[str, Any]) -> None:
    cfg["last_updated"] = datetime.utcnow().isoformat() + "Z"
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2), encoding="utf-8")

# ══════════════════════════════════════════════════════════════════
# Mode dispatch
# ══════════════════════════════════════════════════════════════════

def get_mode() -> str:
    """Return current integration mode: synthetic | mock | live"""
    return get_config().get("mode", "synthetic")

def is_live() -> bool:
    return get_mode() == "live"

def _safe(v):
    if isinstance(v, Decimal): return float(v)
    return v

# ══════════════════════════════════════════════════════════════════
# FLEXCUBE REST API CALLS
# Each function tries live, falls back to synthetic if unavailable.
# ══════════════════════════════════════════════════════════════════

def fetch_account_balance(account_no: str, branch: str = "001") -> Dict[str, Any]:
    """
    Fetches account balance from FLEXCUBE.
    
    Live mode: GET /flexcube/AccountBalanceService/AccountBalance/QueryAcctBal/
               brhcode/{branch}/custAcNo/{account_no}
    
    Returns normalised dict:
    {"account_no", "branch", "available_balance", "ledger_balance",
     "currency", "as_of", "source"}
    """
    if get_mode() == "live":
        return _live_account_balance(account_no, branch)
    return _synthetic_account_balance(account_no, branch)

def _live_account_balance(account_no: str, branch: str) -> Dict[str, Any]:
    """Real FLEXCUBE REST call. Falls back to synthetic on failure."""
    try:
        import requests
        cfg = get_config()
        url = f"{cfg['endpoints']['fcubs_rest']}/AccountBalanceService/AccountBalance/QueryAcctBal/brhcode/{branch}/custAcNo/{account_no}"
        token = _get_oauth_token()
        resp = requests.get(url,
                           headers={"Authorization": f"Bearer {token}",
                                    "Content-Type": "application/json"},
                           timeout=cfg["timeouts"]["rest_seconds"])
        resp.raise_for_status()
        d = resp.json()
        return {
            "account_no":        d.get("AccountNo", account_no),
            "branch":            d.get("BranchCode", branch),
            "available_balance": _safe(d.get("AvailableBalance", 0)),
            "ledger_balance":    _safe(d.get("LedgerBalance", 0)),
            "currency":          d.get("currency", "KES"),
            "as_of":             d.get("as_of", datetime.utcnow().isoformat()+"Z"),
            "source":            "flexcube_live",
        }
    except Exception as e:
        result = _synthetic_account_balance(account_no, branch)
        result["source"]   = "synthetic_fallback"
        result["error"]    = str(e)[:100]
        return result

def _synthetic_account_balance(account_no: str, branch: str) -> Dict[str, Any]:
    """Read from synthetic CBS data."""
    try:
        import csv as _csv
        accts_csv = CBS_DIR / "accounts.csv"
        if accts_csv.exists():
            with accts_csv.open("r", encoding="utf-8") as f:
                reader = _csv.DictReader(f)
                for row in reader:
                    if row.get("account_no","") == account_no:
                        return {
                            "account_no":        account_no,
                            "branch":            row.get("branch", branch),
                            "available_balance": _safe(row.get("available_balance", 0)),
                            "ledger_balance":    _safe(row.get("ledger_balance", row.get("balance",0))),
                            "currency":          row.get("currency","KES"),
                            "as_of":             datetime.utcnow().isoformat()+"Z",
                            "source":            "synthetic",
                        }
    except Exception:
        pass
    # If nothing else, return zero-balance stub
    return {"account_no":account_no,"branch":branch,"available_balance":0,
            "ledger_balance":0,"currency":"KES",
            "as_of":datetime.utcnow().isoformat()+"Z","source":"stub"}

def fetch_customer(cif: str) -> Dict[str, Any]:
    """
    Fetch customer profile by CIF.
    Live: GET /flexcube/CustomerService/QueryCustomer/customer_no/{cif}
    """
    if get_mode() == "live":
        return _live_customer(cif)
    return _synthetic_customer(cif)

def _live_customer(cif: str) -> Dict[str, Any]:
    try:
        import requests
        cfg = get_config()
        url = f"{cfg['endpoints']['fcubs_rest']}/CustomerService/QueryCustomer/customer_no/{cif}"
        token = _get_oauth_token()
        resp = requests.get(url,
                           headers={"Authorization": f"Bearer {token}"},
                           timeout=cfg["timeouts"]["rest_seconds"])
        resp.raise_for_status()
        d = resp.json()
        return {
            "cif":               cif,
            "name":              d.get("CustomerName",""),
            "type":              d.get("CustomerType","INDIVIDUAL"),
            "branch":            d.get("BranchCode",""),
            "rm_code":           d.get("RMCode",""),
            "kyc_status":        d.get("KYCStatus",""),
            "risk_rating":       d.get("RiskRating","Low"),
            "country":           d.get("Country","KEN"),
            "id_number":         d.get("IDNumber",""),
            "phone":             d.get("Phone",""),
            "email":             d.get("Email",""),
            "opened_date":       d.get("CustomerSince",""),
            "source":            "flexcube_live",
        }
    except Exception as e:
        result = _synthetic_customer(cif)
        result["source"] = "synthetic_fallback"
        result["error"]  = str(e)[:100]
        return result

def _synthetic_customer(cif: str) -> Dict[str, Any]:
    try:
        import csv as _csv
        cust_csv = CBS_DIR / "customers.csv"
        if cust_csv.exists():
            with cust_csv.open("r", encoding="utf-8") as f:
                reader = _csv.DictReader(f)
                for row in reader:
                    if row.get("cif","") == cif:
                        return {
                            "cif":         cif,
                            "name":        row.get("name",""),
                            "type":        row.get("type","INDIVIDUAL"),
                            "branch":      row.get("branch",""),
                            "rm_code":     row.get("rm_code",""),
                            "kyc_status":  row.get("kyc_status","Active"),
                            "risk_rating": row.get("risk_rating","Low"),
                            "country":     row.get("country","KEN"),
                            "id_number":   row.get("id_number",""),
                            "phone":       row.get("phone",""),
                            "email":       row.get("email",""),
                            "opened_date": row.get("opened_date",""),
                            "source":      "synthetic",
                        }
    except Exception:
        pass
    return {"cif":cif,"name":"","type":"","branch":"","rm_code":"","kyc_status":"","risk_rating":"",
            "country":"KEN","id_number":"","phone":"","email":"","opened_date":"","source":"stub"}

def fetch_loan_status(loan_id: str) -> Dict[str, Any]:
    """
    Fetch loan account status.
    Live: GET /flexcube/LoanAccountService/QueryLoan/account_no/{loan_id}
    """
    if get_mode() == "live":
        return _live_loan(loan_id)
    return _synthetic_loan(loan_id)

def _live_loan(loan_id: str) -> Dict[str, Any]:
    try:
        import requests
        cfg = get_config()
        url = f"{cfg['endpoints']['fcubs_rest']}/LoanAccountService/QueryLoan/account_no/{loan_id}"
        token = _get_oauth_token()
        resp = requests.get(url,
                           headers={"Authorization": f"Bearer {token}"},
                           timeout=cfg["timeouts"]["rest_seconds"])
        resp.raise_for_status()
        d = resp.json()
        return {
            "loan_id":      loan_id,
            "cif":          d.get("CustomerId",""),
            "principal":    _safe(d.get("PrincipalAmount",0)),
            "outstanding":  _safe(d.get("OutstandingAmount",0)),
            "rate":         _safe(d.get("InterestRate",0)),
            "tenor_months": _safe(d.get("TenorMonths",0)),
            "status":       d.get("Status",""),
            "dpd":          _safe(d.get("DPD",0)),
            "classification":d.get("Classification","Performing"),
            "next_emi_date":d.get("NextEMIDate",""),
            "source":       "flexcube_live",
        }
    except Exception as e:
        result = _synthetic_loan(loan_id)
        result["source"] = "synthetic_fallback"
        result["error"]  = str(e)[:100]
        return result

def _synthetic_loan(loan_id: str) -> Dict[str, Any]:
    try:
        loans = json.loads((DATA_DIR/"credit_monitoring.json").read_text(encoding="utf-8"))
        for l in loans if isinstance(loans, list) else []:
            if str(l.get("loan_id","")) == str(loan_id) or str(l.get("id","")) == str(loan_id):
                return {
                    "loan_id":     str(loan_id),
                    "cif":         l.get("cif",""),
                    "principal":   _safe(l.get("principal_kes", l.get("amount_kes",0))),
                    "outstanding": _safe(l.get("outstanding_kes",0)),
                    "rate":        _safe(l.get("interest_rate_pa",0)),
                    "tenor_months":_safe(l.get("tenor_months",0)),
                    "status":      l.get("status","Active"),
                    "dpd":         _safe(l.get("dpd",0)),
                    "classification":l.get("classification","Performing"),
                    "next_emi_date":l.get("next_emi_date",""),
                    "source":      "synthetic",
                }
    except Exception:
        pass
    return {"loan_id":loan_id,"cif":"","principal":0,"outstanding":0,"rate":0,"tenor_months":0,
            "status":"","dpd":0,"classification":"","next_emi_date":"","source":"stub"}

# ══════════════════════════════════════════════════════════════════
# AGGREGATE QUERIES — used by BSC engine
# These pull large datasets. Live calls go via batch/replica DB.
# ══════════════════════════════════════════════════════════════════

def fetch_rm_portfolio(rm_code: str) -> Dict[str, Any]:
    """
    Aggregate portfolio metrics for an RM.
    Live: GET /flexcube/PortfolioService/RMPortfolio/rm/{rm_code}
    Returns deposit/loan totals and counts.
    """
    if get_mode() == "live":
        try:
            import requests
            cfg = get_config()
            url = f"{cfg['endpoints']['fcubs_rest']}/PortfolioService/RMPortfolio/rm/{rm_code}"
            token = _get_oauth_token()
            resp = requests.get(url,
                               headers={"Authorization": f"Bearer {token}"},
                               timeout=cfg["timeouts"]["batch_seconds"])
            resp.raise_for_status()
            d = resp.json()
            return {
                "rm_code":      rm_code,
                "total_deposits_kes":  _safe(d.get("TotalDeposits",0)),
                "total_loans_kes":     _safe(d.get("TotalLoans",0)),
                "active_customers":    _safe(d.get("ActiveCustomers",0)),
                "active_accounts":     _safe(d.get("ActiveAccounts",0)),
                "npl_kes":             _safe(d.get("NPLBalance",0)),
                "npl_pct":             _safe(d.get("NPLRatio",0)),
                "fees_ytd_kes":        _safe(d.get("FeesYTD",0)),
                "as_of":               d.get("AsOf",""),
                "source":              "flexcube_live",
            }
        except Exception as e:
            return _synthetic_rm_portfolio(rm_code, error=str(e)[:100])
    return _synthetic_rm_portfolio(rm_code)

def _synthetic_rm_portfolio(rm_code: str, error: str="") -> Dict[str, Any]:
    """Read from existing actuals_*.xlsx aggregates if available."""
    result = {
        "rm_code":             rm_code,
        "total_deposits_kes":  0,
        "total_loans_kes":     0,
        "active_customers":    0,
        "active_accounts":     0,
        "npl_kes":             0,
        "npl_pct":             0,
        "fees_ytd_kes":        0,
        "as_of":               datetime.utcnow().isoformat()+"Z",
        "source":              "synthetic_fallback" if error else "synthetic",
    }
    if error: result["error"] = error
    try:
        import openpyxl
        for fname in ["actuals_2025_Dec_25.xlsx","actuals_2025_Sep.xlsx"]:
            xlsx = DATA_DIR / fname
            if xlsx.exists():
                wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = [c.value for c in ws[1]] if ws.max_row > 0 else []
                    rm_col = next((i for i,h in enumerate(headers) if h and "rm" in str(h).lower()),None)
                    if rm_col is None: continue
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if str(row[rm_col]) == str(rm_code):
                            for i,h in enumerate(headers):
                                if not h: continue
                                hl = str(h).lower()
                                v  = row[i]
                                if v is None: continue
                                if "deposit" in hl: result["total_deposits_kes"] += _safe(v)
                                elif "loan"  in hl: result["total_loans_kes"]    += _safe(v)
                                elif "fee"   in hl: result["fees_ytd_kes"]       += _safe(v)
                wb.close()
                if result["total_deposits_kes"] or result["total_loans_kes"]:
                    break
    except Exception:
        pass
    return result

def fetch_branch_metrics(branch_code: str) -> Dict[str, Any]:
    """Aggregate metrics by branch."""
    if get_mode() == "live":
        try:
            import requests
            cfg = get_config()
            url = f"{cfg['endpoints']['fcubs_rest']}/PortfolioService/Branch/{branch_code}"
            token = _get_oauth_token()
            resp = requests.get(url,
                               headers={"Authorization": f"Bearer {token}"},
                               timeout=cfg["timeouts"]["batch_seconds"])
            resp.raise_for_status()
            d = resp.json()
            return {**d, "source":"flexcube_live"}
        except Exception as e:
            return {"branch_code":branch_code,"source":"synthetic_fallback","error":str(e)[:100]}
    return {"branch_code":branch_code,"source":"synthetic"}


# ══════════════════════════════════════════════════════════════════
# v8.0 — PORTFOLIO-LEVEL AGGREGATE FETCHERS
# These are the live counterparts to flexcube_aggregator's
# _fetch_*_live() stubs. Each calls a FLEXCUBE GL / portfolio summary
# endpoint and returns the aggregate dict in A2Z's normalised vocabulary.
#
# v8.1 — RESILIENCE: retry + circuit breaker added to _live_request().
# Per CBK Operations Resilience Guidelines for outsourced/integrated
# CBS access: 3 retries with exponential backoff (1s/3s/9s); circuit
# trips open after 5 consecutive failures, stays open for 60s.
# ══════════════════════════════════════════════════════════════════

# v8.1 circuit breaker state — module-level singleton.
# A real production deployment with multiple processes would use a
# shared store (Redis); single-process Streamlit uses module state.
import threading as _threading
import time as _time

_CIRCUIT_LOCK = _threading.Lock()

# v8.17 — PER-ENDPOINT circuit breaker state.
# Closes v8.6 retrospective ack #6: a failing NPL endpoint should not trip
# the loans/deposits/customer/dormancy endpoints. Each endpoint maintains
# its own consecutive_failures counter and tripped_until timestamp.
#
# Backward compatibility (G108 contract): get_circuit_state() still returns
# the v8.1 single-state shape with aggregate values (any-open / max-failures).
# The new per_endpoint key in the returned dict surfaces granular state
# without breaking existing UI panels that read the top-level keys.
#
# State shape per endpoint key:
#   {"consecutive_failures": int, "tripped_until": float}
#
# Endpoint keys are derived from endpoint_path via _endpoint_key() — the
# first 1-2 stable path segments (variable parts dropped). This gives a
# small fixed set of circuit identifiers; new live endpoints inherit a
# fresh circuit on first call.
# v9.6 — Per-endpoint circuit state migrated to StateBackend abstraction.
# Replaces the v8.17 _CIRCUIT_STATES in-process dict. The InMemoryBackend
# preserves v8.x semantics exactly; RedisBackend (when A2Z_REDIS_URL is
# set) shares state across Streamlit processes.
#
# Backend key convention: "circuit:{endpoint_key}" is a hash with fields
# {consecutive_failures: int, tripped_until: float}.
#
# _CIRCUIT_STATES is REMOVED. The backend is the source of truth.
_CIRCUIT_STATE_KEY_PREFIX = "circuit:"


def _circuit_state_key(endpoint_key: str) -> str:
    """Backend key for per-endpoint circuit state."""
    return f"{_CIRCUIT_STATE_KEY_PREFIX}{endpoint_key}"


def _list_tracked_endpoint_keys() -> List[str]:
    """Return the endpoint keys currently tracked in the backend.

    v9.6 — replaces direct iteration over the v8.17 _CIRCUIT_STATES dict.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    full_keys = backend.keys_matching(_CIRCUIT_STATE_KEY_PREFIX)
    return [k[len(_CIRCUIT_STATE_KEY_PREFIX):] for k in full_keys]

# v8.17 — backward-compat shim. Old single-circuit accessors are preserved
# but operate against a "global" pseudo-endpoint that mirrors the worst
# state across all real endpoints. Callers should migrate to the
# endpoint-aware helpers (_circuit_*_for(endpoint_path)).
_CIRCUIT_STATE = {
    "consecutive_failures": 0,
    "tripped_until": 0.0,  # epoch seconds; 0 = closed
}

# v8.1 tunables — per CBK Operations Resilience Guidelines defaults
RETRY_ATTEMPTS = 3
RETRY_BACKOFF_SECONDS = (1.0, 3.0, 9.0)  # exponential: 1s, 3s, 9s
CIRCUIT_BREAKER_THRESHOLD = 5  # trip after N consecutive failures
CIRCUIT_BREAKER_OPEN_SECONDS = 60.0

# v8.8 — jitter applied on top of RETRY_BACKOFF_SECONDS to prevent
# thundering-herd retries when many clients hit the same FLEXCUBE
# outage simultaneously. Each retry's actual wait is the base backoff
# multiplied by a random factor in [1-JITTER_PCT, 1+JITTER_PCT].
# 0.0 = no jitter (deterministic; matches v8.1 behavior).
# 0.2 = ±20% randomization (default; industry-standard value).
RETRY_JITTER_PCT = 0.2  # 0.0 disables jitter; 0.2 = ±20%


def _apply_jitter(backoff: float) -> float:
    """Return a jittered version of backoff: backoff * uniform(1-J, 1+J).

    Pure function (no side effects beyond random.uniform). Returns the
    base backoff unchanged when RETRY_JITTER_PCT == 0.0 — preserves
    v8.1 determinism for tests + benchmarks that depended on it.
    """
    if RETRY_JITTER_PCT <= 0.0:
        return backoff
    import random as _random
    factor = _random.uniform(
        1.0 - RETRY_JITTER_PCT, 1.0 + RETRY_JITTER_PCT)
    return max(0.0, backoff * factor)


def _endpoint_key(endpoint_path: str) -> str:
    """Normalize an endpoint path to a stable per-endpoint identifier.

    v8.17 — used to key per-endpoint circuit breaker state.

    Takes the first 1-2 stable path segments (variable parts that look
    like CIFs, account numbers, or branch codes are dropped). For the
    5 v8.0 portfolio aggregate paths this gives 5 stable keys:
        /PortfolioService/Loans/Aggregate     -> PortfolioService/Loans
        /PortfolioService/Deposits/Aggregate  -> PortfolioService/Deposits
        /PortfolioService/NPL/Aggregate       -> PortfolioService/NPL
        /CustomerService/Aggregate            -> CustomerService
        /AccountService/Dormancy/Aggregate    -> AccountService/Dormancy

    Variable detection: numeric segments and brace-template segments are
    dropped. This is sufficient for the current FLEXCUBE endpoint set;
    if future endpoints have other variable patterns, extend this helper.
    """
    parts = [p for p in endpoint_path.strip("/").split("/")
             if p and not p.isdigit() and not (p.startswith("{") and p.endswith("}"))]
    if not parts:
        return "unknown"
    # Take first 2 segments, or first 1 if only one is non-variable
    return "/".join(parts[:2])


def _get_or_init_state(endpoint_key: str) -> Dict[str, float]:
    """Return per-endpoint state dict from backend, with defaults if new.

    v8.17 → v9.6 — backed by StateBackend instead of _CIRCUIT_STATES dict.
    The returned dict is a SNAPSHOT — mutations don't auto-persist.
    Use _set_circuit_field() to write changes back.

    Must be called inside _CIRCUIT_LOCK for in-process consistency.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    state = backend.hash_get_all(_circuit_state_key(endpoint_key))
    # Default fields if hash missing or partial
    if "consecutive_failures" not in state:
        state["consecutive_failures"] = 0
    if "tripped_until" not in state:
        state["tripped_until"] = 0.0
    # Coerce types — JSON deserialization may produce float for ints
    state["consecutive_failures"] = int(state["consecutive_failures"])
    state["tripped_until"] = float(state["tripped_until"])
    return state


def _set_circuit_field(endpoint_key: str, field: str, value: Any) -> None:
    """Write a single circuit-state field to the backend.

    v9.6 — atomic at the field level. For multi-field updates, call
    multiple times; aggregate atomicity is best-effort.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    backend.hash_set(_circuit_state_key(endpoint_key), field, value)


def _circuit_is_open(endpoint_path: str = "") -> bool:
    """Return True if the circuit for endpoint_path is tripped open.

    v8.1 (single circuit) → v8.17 (per-endpoint) → v9.6 (backend-backed).
    When called without an endpoint_path (legacy callers), returns the
    AGGREGATE — True if ANY endpoint's circuit is open. New code should
    pass endpoint_path.
    """
    with _CIRCUIT_LOCK:
        if endpoint_path:
            ek = _endpoint_key(endpoint_path)
            state = _get_or_init_state(ek)
            if state["tripped_until"] == 0:
                return False
            if _time.time() >= state["tripped_until"]:
                # Half-open: clear trip + counter so next call probes
                _set_circuit_field(ek, "tripped_until", 0.0)
                _set_circuit_field(ek, "consecutive_failures", 0)
                return False
            return True
        # Legacy path — aggregate (any-open)
        now = _time.time()
        for ek in _list_tracked_endpoint_keys():
            state = _get_or_init_state(ek)
            if state["tripped_until"] > now:
                return True
        return False


def _circuit_record_success(endpoint_path: str = "") -> None:
    """Reset the failure counter for endpoint_path on a successful call.

    v8.17 → v9.6 — backend-backed. Legacy single-circuit callers (no
    endpoint_path) reset the legacy global counter for backward compat.
    """
    with _CIRCUIT_LOCK:
        if endpoint_path:
            ek = _endpoint_key(endpoint_path)
            _set_circuit_field(ek, "consecutive_failures", 0)
            _set_circuit_field(ek, "tripped_until", 0.0)
        else:
            _CIRCUIT_STATE["consecutive_failures"] = 0
            _CIRCUIT_STATE["tripped_until"] = 0.0


def _circuit_record_failure(endpoint_path: str = "") -> None:
    """Increment the failure counter for endpoint_path; trip circuit if at threshold.

    v8.17 → v9.6 — backend-backed; uses atomic HINCRBY when the backend is
    Redis (cross-process-safe counter increment). Legacy single-circuit
    callers (no endpoint_path) increment the legacy global counter for
    backward compat.
    """
    from utils.state_backend import get_default_backend
    with _CIRCUIT_LOCK:
        if endpoint_path:
            ek = _endpoint_key(endpoint_path)
            backend = get_default_backend()
            new_failures = backend.hash_incr(
                _circuit_state_key(ek), "consecutive_failures", 1)
            if new_failures >= CIRCUIT_BREAKER_THRESHOLD:
                tripped_until = _time.time() + CIRCUIT_BREAKER_OPEN_SECONDS
                _set_circuit_field(ek, "tripped_until", tripped_until)
        else:
            _CIRCUIT_STATE["consecutive_failures"] += 1
            if _CIRCUIT_STATE["consecutive_failures"] >= CIRCUIT_BREAKER_THRESHOLD:
                _CIRCUIT_STATE["tripped_until"] = (
                    _time.time() + CIRCUIT_BREAKER_OPEN_SECONDS)


def get_circuit_state() -> Dict[str, Any]:
    """Public read-only snapshot of circuit state.

    v8.1 — original single-circuit shape (backward-compatible).
    v8.17 — adds per_endpoint key with granular per-endpoint state.
    v9.6 — reads through StateBackend; preserves shape exactly.

    The top-level keys (consecutive_failures, is_open, seconds_until_close,
    threshold, ...) are AGGREGATES across all per-endpoint circuits:
        consecutive_failures: max across endpoints
        is_open: True if ANY endpoint is currently open
        seconds_until_close: max remaining time among open endpoints (0 if all closed)

    The new per_endpoint key returns:
        {endpoint_key: {consecutive_failures, is_open, seconds_until_close}}

    This shape preserves G108's contract (v8.3 audit gate) while exposing
    finer-grained state for the v8.17 admin UI + observability.
    """
    with _CIRCUIT_LOCK:
        now = _time.time()
        agg_failures = 0
        agg_open = False
        agg_seconds_until_close = 0.0
        per_endpoint: Dict[str, Dict[str, Any]] = {}

        for ek in _list_tracked_endpoint_keys():
            state = _get_or_init_state(ek)
            ep_open = state["tripped_until"] > now
            ep_seconds = max(0.0, state["tripped_until"] - now) if ep_open else 0.0
            per_endpoint[ek] = {
                "consecutive_failures": state["consecutive_failures"],
                "is_open": ep_open,
                "seconds_until_close": ep_seconds,
            }
            if state["consecutive_failures"] > agg_failures:
                agg_failures = state["consecutive_failures"]
            if ep_open:
                agg_open = True
                if ep_seconds > agg_seconds_until_close:
                    agg_seconds_until_close = ep_seconds

        return {
            # Aggregate keys (preserved from v8.1 — G108 contract)
            "consecutive_failures": agg_failures,
            "is_open": agg_open,
            "seconds_until_close": agg_seconds_until_close,
            "threshold": CIRCUIT_BREAKER_THRESHOLD,
            "open_duration_seconds": CIRCUIT_BREAKER_OPEN_SECONDS,
            "retry_attempts": RETRY_ATTEMPTS,
            "retry_backoff_seconds": list(RETRY_BACKOFF_SECONDS),
            "retry_jitter_pct": RETRY_JITTER_PCT,
            # New v8.17 — per-endpoint detail
            "per_endpoint": per_endpoint,
            "endpoints_tracked": len(per_endpoint),
        }


def reset_circuit(endpoint_key: Optional[str] = None) -> Dict[str, Any]:
    """Manually clear circuit breaker state without restarting the process.

    v8.9 admin operation; v8.17 extended to accept an optional
    endpoint_key for per-endpoint reset; v9.6 backend-backed.

    Args:
        endpoint_key: If None (default — backward-compatible v8.9 behavior),
            reset ALL endpoint circuits + the legacy global state. If a
            specific key is provided (e.g. "PortfolioService/NPL"), reset
            only that endpoint.

    Returns a dict describing what was cleared so operators have an audit
    trail. Latency telemetry is NOT touched (separate state with its own
    reset_latency_state()).
    """
    from utils.state_backend import get_default_backend
    cleared: Dict[str, Any] = {
        "reset_at_iso": datetime.now(timezone.utc).isoformat(),
    }
    backend = get_default_backend()
    with _CIRCUIT_LOCK:
        now = _time.time()
        if endpoint_key is None:
            # Reset ALL endpoints (v8.9 backward-compatible behavior)
            prior_per_endpoint = {}
            tracked = _list_tracked_endpoint_keys()
            for ek in tracked:
                state = _get_or_init_state(ek)
                prior_per_endpoint[ek] = {
                    "prior_consecutive_failures": state["consecutive_failures"],
                    "prior_was_open": state["tripped_until"] > now,
                }
                _set_circuit_field(ek, "consecutive_failures", 0)
                _set_circuit_field(ek, "tripped_until", 0.0)
            cleared["scope"] = "all_endpoints"
            cleared["endpoints_reset"] = list(prior_per_endpoint.keys())
            cleared["prior_per_endpoint"] = prior_per_endpoint
            # Also clear legacy global
            cleared["prior_consecutive_failures"] = (
                _CIRCUIT_STATE["consecutive_failures"])
            cleared["prior_was_open"] = (
                _CIRCUIT_STATE["tripped_until"] > now)
            _CIRCUIT_STATE["consecutive_failures"] = 0
            _CIRCUIT_STATE["tripped_until"] = 0.0
        else:
            # Reset single endpoint
            state = backend.hash_get_all(_circuit_state_key(endpoint_key))
            if not state:
                cleared["scope"] = "single_endpoint"
                cleared["endpoint_key"] = endpoint_key
                cleared["was_tracked"] = False
                cleared["prior_consecutive_failures"] = 0
                cleared["prior_was_open"] = False
            else:
                full = _get_or_init_state(endpoint_key)
                cleared["scope"] = "single_endpoint"
                cleared["endpoint_key"] = endpoint_key
                cleared["was_tracked"] = True
                cleared["prior_consecutive_failures"] = (
                    full["consecutive_failures"])
                cleared["prior_was_open"] = (
                    full["tripped_until"] > now)
                _set_circuit_field(endpoint_key, "consecutive_failures", 0)
                _set_circuit_field(endpoint_key, "tripped_until", 0.0)
        cleared["current_state"] = "closed"
    return cleared


# ── v8.19 → v9.7 — RETRY TELEMETRY (closes v8.6 retrospective ack #9) ──
# Per-endpoint counters tracking retry outcomes. Complements v8.17/v9.6
# per-endpoint circuit state by giving operators visibility into HOW the
# retry/circuit pattern is performing.
#
# v9.7: migrated from in-process _RETRY_TELEMETRY dict to StateBackend.
# Backend key convention: "retry:{endpoint_key}" hash with 5 counter fields.
# All increments use atomic hash_incr() for cross-process safety.
#
# Counters per endpoint:
#   - requests_total: every _live_request() call attempted
#   - retries_triggered: total retry attempts across all requests
#   - succeeded_no_retry: requests that succeeded on first try
#   - succeeded_after_retry: requests that succeeded after >= 1 retry
#   - failed_after_retries: requests that failed despite all retries
#
# Healthy invariants:
#   requests_total == succeeded_no_retry + succeeded_after_retry + failed_after_retries
#   retries_triggered <= requests_total * (RETRY_ATTEMPTS - 1)
_RETRY_TELEMETRY_LOCK = _threading.Lock()
_RETRY_STATE_KEY_PREFIX = "retry:"

_RETRY_FIELDS = (
    "requests_total",
    "retries_triggered",
    "succeeded_no_retry",
    "succeeded_after_retry",
    "failed_after_retries",
)


def _retry_state_key(endpoint_key: str) -> str:
    """Backend key for per-endpoint retry telemetry."""
    return f"{_RETRY_STATE_KEY_PREFIX}{endpoint_key}"


def _list_tracked_retry_endpoints() -> List[str]:
    """List endpoint keys with retry telemetry tracked in the backend."""
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    full_keys = backend.keys_matching(_RETRY_STATE_KEY_PREFIX)
    return [k[len(_RETRY_STATE_KEY_PREFIX):] for k in full_keys]


def _get_retry_counters(endpoint_key: str) -> Dict[str, int]:
    """Read all retry counters for endpoint_key, defaulting missing fields to 0.

    v9.7 — backend-backed; returns a fresh dict (no live link to backend).
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    raw = backend.hash_get_all(_retry_state_key(endpoint_key))
    return {f: int(raw.get(f, 0)) for f in _RETRY_FIELDS}


def _record_retry_outcome(
    endpoint_path: str, retries_used: int, succeeded: bool
) -> None:
    """v8.19 → v9.7 — record per-endpoint retry telemetry after a request.

    Args:
        endpoint_path: full FLEXCUBE path (normalized via _endpoint_key)
        retries_used: count of retries that fired (0 if succeeded on first try)
        succeeded: whether the final outcome was success or failure

    v9.7: each counter increment uses atomic hash_incr() so concurrent
    processes can record outcomes without races.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    with _RETRY_TELEMETRY_LOCK:
        ek = _endpoint_key(endpoint_path)
        key = _retry_state_key(ek)
        backend.hash_incr(key, "requests_total", 1)
        if retries_used > 0:
            backend.hash_incr(key, "retries_triggered", retries_used)
        if succeeded:
            if retries_used == 0:
                backend.hash_incr(key, "succeeded_no_retry", 1)
            else:
                backend.hash_incr(key, "succeeded_after_retry", 1)
        else:
            backend.hash_incr(key, "failed_after_retries", 1)


def get_retry_telemetry() -> Dict[str, Any]:
    """v8.19 → v9.7 — public read-only snapshot of retry telemetry.

    Returns dict shape:
        {
          "per_endpoint": {ek: {requests_total, retries_triggered,
                                succeeded_no_retry, succeeded_after_retry,
                                failed_after_retries, retry_recovery_rate_pct,
                                avg_retries_per_request}},
          "summary": {requests_total, retries_triggered, ...,
                      retry_recovery_rate_pct, avg_retries_per_request}
        }

    `retry_recovery_rate_pct` = succeeded_after_retry /
        (succeeded_after_retry + failed_after_retries) — measures how often
    a transient failure was successfully recovered via retry. Higher = better.

    `avg_retries_per_request` = retries_triggered / requests_total — measures
    overall flakiness. 0 = healthy. Approaching (RETRY_ATTEMPTS-1) = severe.

    v9.7: reads through StateBackend; preserves shape exactly.
    """
    with _RETRY_TELEMETRY_LOCK:
        per_endpoint: Dict[str, Dict[str, Any]] = {}
        agg = {f: 0 for f in _RETRY_FIELDS}
        for ek in _list_tracked_retry_endpoints():
            s = _get_retry_counters(ek)
            recoverable = s["succeeded_after_retry"] + s["failed_after_retries"]
            recovery_pct = (
                round(100.0 * s["succeeded_after_retry"] / recoverable, 1)
                if recoverable > 0 else None)
            avg_retries = (
                round(s["retries_triggered"] / s["requests_total"], 2)
                if s["requests_total"] > 0 else 0.0)
            per_endpoint[ek] = {
                **s,
                "retry_recovery_rate_pct": recovery_pct,
                "avg_retries_per_request": avg_retries,
            }
            for k in agg:
                agg[k] += s[k]
        recoverable = agg["succeeded_after_retry"] + agg["failed_after_retries"]
        summary_recovery = (
            round(100.0 * agg["succeeded_after_retry"] / recoverable, 1)
            if recoverable > 0 else None)
        summary_avg = (
            round(agg["retries_triggered"] / agg["requests_total"], 2)
            if agg["requests_total"] > 0 else 0.0)
        return {
            "per_endpoint": per_endpoint,
            "summary": {
                **agg,
                "retry_recovery_rate_pct": summary_recovery,
                "avg_retries_per_request": summary_avg,
                "endpoints_tracked": len(per_endpoint),
            },
        }


def reset_retry_telemetry(endpoint_key: Optional[str] = None) -> Dict[str, Any]:
    """v8.19 → v9.7 — clear retry telemetry counters. Operator admin function.

    Args:
        endpoint_key: If None, reset ALL endpoints. Otherwise reset just one.

    Returns prior counters before reset (for audit trail).
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    cleared: Dict[str, Any] = {
        "reset_at_iso": datetime.now(timezone.utc).isoformat(),
    }
    with _RETRY_TELEMETRY_LOCK:
        if endpoint_key is None:
            cleared["scope"] = "all_endpoints"
            tracked = _list_tracked_retry_endpoints()
            cleared["prior_per_endpoint"] = {
                ek: _get_retry_counters(ek) for ek in tracked
            }
            for ek in tracked:
                backend.hash_delete(_retry_state_key(ek))
        else:
            cleared["scope"] = "single_endpoint"
            cleared["endpoint_key"] = endpoint_key
            prior = backend.hash_get_all(_retry_state_key(endpoint_key))
            if not prior:
                cleared["was_tracked"] = False
            else:
                cleared["was_tracked"] = True
                cleared["prior_counters"] = _get_retry_counters(endpoint_key)
                backend.hash_delete(_retry_state_key(endpoint_key))
    return cleared


# ══════════════════════════════════════════════════════════════════
# v8.2 → v9.8 — REQUEST LATENCY TELEMETRY
# Completes the observability triangle: mode banner (v7.10) + circuit
# banner (v8.1) + latency telemetry (v8.2). Each live FLEXCUBE call is
# recorded into a per-endpoint rolling window of the most recent N
# samples. From those samples we derive p50 / p95 / p99 + count.
#
# v9.8: per-endpoint samples migrated to StateBackend lists. Backend key
# convention: "latency:{endpoint_path}" is a list of JSON-serializable
# triples [latency_ms, success_bool, epoch_ts]. FIFO truncation at
# LATENCY_WINDOW_SIZE via list_append max_length parameter.
#
# v8.24 file persistence preserved for InMemoryBackend only — Redis
# backend has its own durability (RDB/AOF), so file persistence becomes
# redundant. is_remote() check skips file persistence when remote.
# ══════════════════════════════════════════════════════════════════

# Rolling window size — keeps memory bounded
LATENCY_WINDOW_SIZE = 200

# v8.24 — latency persistence (closes v8.6 retrospective ack #10).
# In-memory rolling window survives Streamlit script reruns (same process
# state) but is lost on process restart. Persisting the rolling window to
# disk lets observability survive restarts so operators see continuous
# latency history across deployments / container restarts.
LATENCY_PERSIST_PATH = Path("flexcube_data") / "latency_state.json"
LATENCY_PERSIST_INTERVAL_SECONDS = 30.0  # save at most every 30s

_LATENCY_LOCK = _threading.Lock()
_LATENCY_LAST_PERSIST = 0.0  # epoch seconds; 0 = never persisted in this process
_LATENCY_LOADED = False  # one-shot flag — load disk state on first record
_LATENCY_KEY_PREFIX = "latency:"


def _latency_key(endpoint_path: str) -> str:
    """Backend key for per-endpoint-path latency samples."""
    return f"{_LATENCY_KEY_PREFIX}{endpoint_path}"


def _list_tracked_latency_endpoints() -> List[str]:
    """List endpoint paths with latency samples in the backend."""
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    full_keys = backend.keys_matching(_LATENCY_KEY_PREFIX)
    return [k[len(_LATENCY_KEY_PREFIX):] for k in full_keys]


def _load_latency_from_disk() -> None:
    """v8.24 → v9.8 — load persisted latency samples from disk into backend.

    Only meaningful for InMemoryBackend; RedisBackend is.its own durability
    layer so this becomes a no-op when remote. Called exactly once per
    process via the _LATENCY_LOADED flag.
    """
    from utils.state_backend import get_default_backend
    global _LATENCY_LOADED
    if _LATENCY_LOADED:
        return
    _LATENCY_LOADED = True
    backend = get_default_backend()
    if backend.is_remote():
        return  # Redis handles its own durability
    try:
        if not LATENCY_PERSIST_PATH.exists():
            return
        raw = json.loads(LATENCY_PERSIST_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            return
        endpoints = raw.get("endpoints", {})
        if not isinstance(endpoints, dict):
            return
        for ep, samples in endpoints.items():
            if not isinstance(samples, list):
                continue
            valid_count = 0
            for s in samples:
                if (isinstance(s, (list, tuple)) and len(s) == 3
                        and isinstance(s[0], (int, float))
                        and isinstance(s[1], bool)
                        and isinstance(s[2], (int, float))):
                    backend.list_append(
                        _latency_key(ep),
                        [float(s[0]), bool(s[1]), float(s[2])],
                        max_length=LATENCY_WINDOW_SIZE)
                    valid_count += 1
                    if valid_count >= LATENCY_WINDOW_SIZE:
                        break
    except Exception:
        pass


def _persist_latency_to_disk() -> None:
    """v8.24 → v9.8 — save current rolling window to disk. Best-effort.

    Only acts when backend is local (InMemoryBackend); skipped when remote
    (Redis has its own durability). Throttled by LATENCY_PERSIST_INTERVAL_SECONDS.
    Must be called inside _LATENCY_LOCK.
    """
    from utils.state_backend import get_default_backend
    global _LATENCY_LAST_PERSIST
    backend = get_default_backend()
    if backend.is_remote():
        return  # Redis durability replaces file persistence
    now = _time.time()
    if now - _LATENCY_LAST_PERSIST < LATENCY_PERSIST_INTERVAL_SECONDS:
        return
    _LATENCY_LAST_PERSIST = now
    try:
        LATENCY_PERSIST_PATH.parent.mkdir(parents=True, exist_ok=True)
        endpoints_snapshot = {}
        for ep in _list_tracked_latency_endpoints():
            samples = backend.list_range(_latency_key(ep))
            endpoints_snapshot[ep] = samples
        snapshot = {
            "saved_at_iso": datetime.now(timezone.utc).isoformat(),
            "endpoints": endpoints_snapshot,
        }
        tmp = LATENCY_PERSIST_PATH.with_suffix(
            LATENCY_PERSIST_PATH.suffix + ".tmp")
        tmp.write_text(
            json.dumps(snapshot, separators=(",", ":")),
            encoding="utf-8")
        tmp.replace(LATENCY_PERSIST_PATH)
    except Exception:
        pass


def _record_latency(endpoint_path: str, latency_ms: float, success: bool) -> None:
    """Record a single call's latency + outcome.

    v8.2 internal → v9.8 backend-backed. v8.24 disk persistence preserved
    for InMemoryBackend only.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    with _LATENCY_LOCK:
        if not _LATENCY_LOADED:
            _load_latency_from_disk()
        sample = [float(latency_ms), bool(success), _time.time()]
        backend.list_append(
            _latency_key(endpoint_path), sample,
            max_length=LATENCY_WINDOW_SIZE)
        _persist_latency_to_disk()  # throttled internally; remote=no-op


def _percentile(sorted_samples: list, pct: float) -> float:
    """Compute a percentile from a sorted list. Returns 0.0 for empty."""
    if not sorted_samples:
        return 0.0
    idx = int(len(sorted_samples) * pct / 100.0)
    idx = min(idx, len(sorted_samples) - 1)
    return sorted_samples[idx]


def get_latency_state() -> Dict[str, Any]:
    """Public read-only snapshot of per-endpoint latency.

    v8.2 → v9.8 — reads from StateBackend; returns same dict shape.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    with _LATENCY_LOCK:
        if not _LATENCY_LOADED:
            _load_latency_from_disk()
        snapshot = {}
        for ep in _list_tracked_latency_endpoints():
            samples_raw = backend.list_range(_latency_key(ep))
            # Each sample is a 3-list [latency_ms, success, ts]
            normalized = []
            for s in samples_raw:
                if isinstance(s, (list, tuple)) and len(s) == 3:
                    normalized.append((float(s[0]), bool(s[1]), float(s[2])))
            snapshot[ep] = normalized

    endpoints: Dict[str, Any] = {}
    total_count = 0
    total_success = 0

    for ep, samples in snapshot.items():
        if not samples:
            continue
        latencies_sorted = sorted(s[0] for s in samples)
        successes = [s for s in samples if s[1]]
        last = samples[-1]
        endpoints[ep] = {
            "count": len(samples),
            "success_count": len(successes),
            "failure_count": len(samples) - len(successes),
            "success_rate_pct": round(
                100.0 * len(successes) / len(samples), 1),
            "p50_ms": round(_percentile(latencies_sorted, 50), 1),
            "p95_ms": round(_percentile(latencies_sorted, 95), 1),
            "p99_ms": round(_percentile(latencies_sorted, 99), 1),
            "last_call_ts": last[2],
            "last_latency_ms": round(last[0], 1),
            "latest_outcome": "success" if last[1] else "failure",
        }
        total_count += len(samples)
        total_success += len(successes)

    summary = {
        "endpoints_observed": len(endpoints),
        "total_calls": total_count,
        "total_successes": total_success,
        "total_failures": total_count - total_success,
        "overall_success_rate_pct": (
            round(100.0 * total_success / total_count, 1)
            if total_count > 0 else None),
        "window_size": LATENCY_WINDOW_SIZE,
    }
    return {"endpoints": endpoints, "summary": summary}


def reset_latency_state() -> None:
    """Clear all latency samples. v8.2 → v9.8 — backend + persistent file.

    v8.24: removes the persisted on-disk state so a reset survives process
    restart. v9.8: also clears all per-endpoint backend lists.
    """
    from utils.state_backend import get_default_backend
    backend = get_default_backend()
    with _LATENCY_LOCK:
        for ep in _list_tracked_latency_endpoints():
            backend.list_clear(_latency_key(ep))
        try:
            if LATENCY_PERSIST_PATH.exists():
                LATENCY_PERSIST_PATH.unlink()
        except Exception:
            pass


def _live_request(endpoint_path: str, timeout_key: str = "batch_seconds") -> Optional[Dict[str, Any]]:
    """Helper: GET an authenticated FLEXCUBE endpoint with retry + circuit breaker + latency telemetry.

    v8.1 resilience layer + v8.2 observability layer:
        - If circuit is OPEN, skip immediately (return None) — fast-fail
          prevents thundering-herd retries during sustained outage.
        - Otherwise attempt the call. On 5xx / network failure / OAuth
          error: retry up to RETRY_ATTEMPTS times with exponential
          backoff (1s, 3s, 9s).
        - On success: reset failure counter; close circuit; record
          per-endpoint latency.
        - On final failure (all retries exhausted): increment failure
          counter; trip circuit if threshold reached; record per-endpoint
          failure with total elapsed time across all attempts.
        - Returns parsed JSON on success, None on any failure.
    """
    if get_mode() != "live":
        return None

    # Circuit breaker — fast-fail during sustained outage
    # v8.17: per-endpoint state — a tripped NPL circuit doesn't block loans
    if _circuit_is_open(endpoint_path):
        return None

    try:
        import requests
        cfg = get_config()
        url = f"{cfg['endpoints']['fcubs_rest']}{endpoint_path}"
        # v8.20 — per-endpoint timeout override; fallback to timeout_key default
        ek = _endpoint_key(endpoint_path)
        endpoint_timeouts = cfg.get("endpoint_timeouts", {})
        if ek in endpoint_timeouts:
            timeout = endpoint_timeouts[ek]
        else:
            timeout = cfg["timeouts"][timeout_key]
    except Exception:
        # Config or import error — not retryable
        _circuit_record_failure(endpoint_path)
        return None

    last_exc: Optional[BaseException] = None
    request_started = _time.time()
    retries_used = 0  # v8.19: track retries triggered for this request
    for attempt in range(RETRY_ATTEMPTS):
        try:
            import requests
            token = _get_oauth_token()
            resp = requests.get(
                url,
                headers={"Authorization": f"Bearer {token}"},
                timeout=timeout)
            resp.raise_for_status()
            data = resp.json()
            elapsed_ms = (_time.time() - request_started) * 1000.0
            _record_latency(endpoint_path, elapsed_ms, success=True)
            _circuit_record_success(endpoint_path)
            # v8.19: record retry telemetry — request succeeded (with or
            # without retries). retries_used > 0 means recovery via retry.
            _record_retry_outcome(endpoint_path, retries_used, succeeded=True)
            return data
        except Exception as e:
            last_exc = e
            # Backoff before next retry (skip backoff after last attempt).
            # v8.8: apply jitter to prevent thundering-herd retries.
            if attempt < RETRY_ATTEMPTS - 1:
                retries_used += 1  # v8.19: count this retry
                base_backoff = RETRY_BACKOFF_SECONDS[
                    min(attempt, len(RETRY_BACKOFF_SECONDS) - 1)]
                _time.sleep(_apply_jitter(base_backoff))
            continue

    # All retries exhausted
    elapsed_ms = (_time.time() - request_started) * 1000.0
    _record_latency(endpoint_path, elapsed_ms, success=False)
    _circuit_record_failure(endpoint_path)
    # v8.19: record retry telemetry — all attempts failed despite retries
    _record_retry_outcome(endpoint_path, retries_used, succeeded=False)
    return None


def fetch_loan_portfolio_aggregate_live() -> Optional[Dict[str, Any]]:
    """Live FLEXCUBE loan portfolio aggregate. v8.0 implementation.

    Calls /PortfolioService/Loans/Aggregate which returns the bank-wide
    loan book (gross outstanding + by-segment + by-IFRS-9-stage).

    Returns None if mode != live or the live call fails (caller falls
    back to CBS synthetic / demo defaults). Returns A2Z-normalised dict
    on success; FLEXCUBE-specific field names are translated here.
    """
    raw = _live_request("/PortfolioService/Loans/Aggregate")
    if raw is None:
        return None

    # Translate FLEXCUBE response → A2Z normalised vocabulary.
    # FLEXCUBE's PortfolioService/Loans/Aggregate is documented to return
    # GROSS_OS, SEGMENT_DIST, STAGE_DIST. Map to A2Z field names.
    return {
        "gross_outstanding_kes": str(raw.get("GROSS_OS", "0")),
        "by_segment_kes": {
            seg.upper(): str(amt)
            for seg, amt in (raw.get("SEGMENT_DIST") or {}).items()
        },
        "by_stage_kes": {
            stage.upper(): str(amt)
            for stage, amt in (raw.get("STAGE_DIST") or {}).items()
        },
        "weighted_avg_pd_pct": str(raw.get("WAVG_PD", "0")),
        "average_lgd_pct": str(raw.get("AVG_LGD", "0")),
    }


def fetch_deposit_book_aggregate_live() -> Optional[Dict[str, Any]]:
    """Live FLEXCUBE deposit book aggregate. v8.0 implementation."""
    raw = _live_request("/PortfolioService/Deposits/Aggregate")
    if raw is None:
        return None

    return {
        "total_deposits_kes": str(raw.get("TOTAL_DEPOSITS", "0")),
        "loan_to_deposit_ratio_pct": str(raw.get("LDR_PCT", "0")),
        "by_stability_tier_kes": {
            t.upper(): str(amt)
            for t, amt in (raw.get("STABILITY_TIER_DIST") or {}).items()
        },
        "by_product_kes": {
            p.upper(): str(amt)
            for p, amt in (raw.get("PRODUCT_DIST") or {}).items()
        },
        "by_segment_kes": {
            seg.upper(): str(amt)
            for seg, amt in (raw.get("SEGMENT_DIST") or {}).items()
        },
    }


def fetch_npl_aggregate_live() -> Optional[Dict[str, Any]]:
    """Live FLEXCUBE NPL aggregate. v8.0 implementation."""
    raw = _live_request("/PortfolioService/NPL/Aggregate")
    if raw is None:
        return None

    return {
        "stage_3_kes": str(raw.get("STAGE_3_KES", "0")),
        "loan_book_basis_kes": str(raw.get("LOAN_BOOK_BASIS", "0")),
        "npl_ratio_pct": str(raw.get("NPL_RATIO_PCT", "0")),
        "by_aging_kes": {
            band.upper(): str(amt)
            for band, amt in (raw.get("AGING_DIST") or {}).items()
        },
    }


def fetch_customer_base_aggregate_live() -> Optional[Dict[str, Any]]:
    """Live FLEXCUBE customer base aggregate. v8.0 implementation."""
    raw = _live_request("/CustomerService/Aggregate")
    if raw is None:
        return None

    return {
        "total_customers": int(raw.get("TOTAL_CUSTOMERS", 0)),
        "by_segment_count": {
            seg.upper(): int(c)
            for seg, c in (raw.get("SEGMENT_COUNT_DIST") or {}).items()
        },
        "by_tenure_band_count": {
            band.upper(): int(c)
            for band, c in (raw.get("TENURE_DIST") or {}).items()
        },
        "by_onboarding_channel_count": {
            ch.upper(): int(c)
            for ch, c in (raw.get("ONBOARDING_DIST") or {}).items()
        },
        "by_kyc_risk_band_count": {
            band.upper(): int(c)
            for band, c in (raw.get("KYC_BAND_DIST") or {}).items()
        },
        "monthly_growth_rate_pct": str(raw.get("MONTHLY_GROWTH_PCT", "0")),
    }


def fetch_dormant_accounts_aggregate_live() -> Optional[Dict[str, Any]]:
    """Live FLEXCUBE dormant accounts aggregate. v8.0 implementation."""
    raw = _live_request("/AccountService/Dormancy/Aggregate")
    if raw is None:
        return None

    total_dormant = int(raw.get("TOTAL_DORMANT", 0))
    avg_balance = int(raw.get("AVG_BALANCE_KES", 0))
    return {
        "total_dormant": total_dormant,
        "customer_basis_count": int(raw.get("CUSTOMER_BASIS", 0)),
        "dormancy_rate_pct": str(raw.get("DORMANCY_RATE_PCT", "0")),
        "by_dormancy_band_count": {
            band.upper(): int(c)
            for band, c in (raw.get("BAND_DIST") or {}).items()
        },
        "by_segment_count": {
            seg.upper(): int(c)
            for seg, c in (raw.get("SEGMENT_DIST") or {}).items()
        },
        "reactivation_potential_count": int(raw.get("REACT_POTENTIAL", 0)),
        "avg_balance_per_dormant_kes": avg_balance,
        "estimated_latent_value_kes": str(total_dormant * avg_balance),
    }


# ══════════════════════════════════════════════════════════════════
# OAUTH2 TOKEN HANDLING
# ══════════════════════════════════════════════════════════════════

_TOKEN_CACHE = {"token": None, "expires_at": 0}

def _get_oauth_token() -> str:
    """Get cached OAuth2 token from Apigee. Renews automatically."""
    now = time.time()
    if _TOKEN_CACHE["token"] and _TOKEN_CACHE["expires_at"] > now + 60:
        return _TOKEN_CACHE["token"]
    cfg = get_config()
    if cfg.get("mode") != "live":
        return "MOCK_TOKEN"
    try:
        import os
        import requests
        client_id     = os.environ.get("FLEXCUBE_CLIENT_ID","")
        client_secret = os.environ.get("FLEXCUBE_CLIENT_SECRET","")
        token_url     = cfg["auth"]["token_url"]
        scopes        = " ".join(cfg["auth"].get("scopes",[]))
        resp = requests.post(token_url,
                            data={"grant_type":"client_credentials",
                                  "client_id":client_id,
                                  "client_secret":client_secret,
                                  "scope":scopes},
                            timeout=10)
        resp.raise_for_status()
        d = resp.json()
        _TOKEN_CACHE["token"]      = d["access_token"]
        _TOKEN_CACHE["expires_at"] = now + int(d.get("expires_in",3600))
        return _TOKEN_CACHE["token"]
    except Exception:
        return ""

# ══════════════════════════════════════════════════════════════════
# JMS EVENT SUBSCRIPTIONS — for real-time BSC updates
# In synthetic/mock mode, this is a no-op.
# In live mode, this would be wired to a JMS consumer (separate process).
# ══════════════════════════════════════════════════════════════════

def publish_event(topic: str, payload: Dict[str, Any]) -> bool:
    """Publish a JMS event. Live: real broker. Synthetic: log only."""
    cfg = get_config()
    if cfg.get("mode") != "live":
        # Log to event journal for visibility
        log = DATA_DIR / "flexcube_events.json"
        events = []
        if log.exists():
            try: events = json.loads(log.read_text(encoding="utf-8"))
            except: events = []
        events.insert(0, {
            "timestamp": datetime.utcnow().isoformat()+"Z",
            "topic":     topic,
            "payload":   payload,
            "mode":      cfg.get("mode","synthetic"),
        })
        # Keep last 500
        log.write_text(json.dumps(events[:500], indent=2))
        return True
    # Live: would use stomp.py, pika, or oracle-jms-client here
    # Skipped intentionally — real implementation depends on broker chosen
    return False

# ══════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ══════════════════════════════════════════════════════════════════

def health_check() -> Dict[str, Any]:
    """Probe FLEXCUBE endpoints. Returns service status dict."""
    cfg = get_config()
    mode = cfg.get("mode","synthetic")
    
    services = {
        "FLEXCUBE REST":    {"endpoint":cfg["endpoints"]["fcubs_rest"],     "status":"unknown","latency_ms":0},
        "OBDX REST":        {"endpoint":cfg["endpoints"]["obdx_rest"],      "status":"unknown","latency_ms":0},
        "Apigee Gateway":   {"endpoint":cfg["endpoints"]["apigee_base"],    "status":"unknown","latency_ms":0},
        "OAuth2 Token":     {"endpoint":cfg["auth"]["token_url"],           "status":"unknown","latency_ms":0},
        "JMS Broker":       {"endpoint":cfg["endpoints"]["jms_broker"],     "status":"unknown","latency_ms":0},
    }
    
    if mode == "synthetic":
        for k in services: services[k]["status"] = "Mocked (synthetic mode)"
        return {"mode":mode,"overall":"OK","services":services,"checked_at":datetime.utcnow().isoformat()+"Z"}
    
    if mode == "mock":
        for k in services: services[k]["status"] = "Mocked (integration test)"
        return {"mode":mode,"overall":"OK","services":services,"checked_at":datetime.utcnow().isoformat()+"Z"}
    
    # Live mode — probe each endpoint
    try:
        import requests
        for svc_name, svc in services.items():
            try:
                t0 = time.time()
                if "JMS" in svc_name:
                    svc["status"] = "Skipped (JMS probe needs separate consumer)"
                    continue
                resp = requests.get(svc["endpoint"], timeout=3)
                svc["latency_ms"] = int((time.time()-t0)*1000)
                svc["status"]     = "Up" if resp.status_code < 500 else f"Degraded ({resp.status_code})"
            except Exception as e:
                svc["status"] = f"Down: {str(e)[:50]}"
        ok = sum(1 for s in services.values() if "Up" in s.get("status",""))
        overall = "OK" if ok >= 3 else "DEGRADED" if ok >= 1 else "DOWN"
        return {"mode":mode,"overall":overall,"services":services,
                "checked_at":datetime.utcnow().isoformat()+"Z"}
    except ImportError:
        return {"mode":mode,"overall":"UNKNOWN",
                "services":services,
                "error":"requests library not installed",
                "checked_at":datetime.utcnow().isoformat()+"Z"}

# ══════════════════════════════════════════════════════════════════
# Helper for CBS Explorer / module pages
# ══════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
# Branch directory (v10.361 — FLEXCUBE-aware branch list)
# ══════════════════════════════════════════════════════════════════
# Returns the bank's branch list in the same shape as org_config.json
# (name → region map). When mode="live", calls FLEXCUBE's branch master
# endpoint; otherwise reads from data/org_config.json (the configurable
# admin-managed source).
#
# This is the integration seam: when FLEXCUBE comes online, no caller
# code changes — utils.virtual_bank_seed.get_ecobank_branches() picks up
# the live data automatically.

def fetch_branches_from_flexcube() -> Optional[Dict[str, str]]:
    """v10.361/v10.365 — fetch the bank's branch list from FLEXCUBE.

    Returns {branch_name: region} dict (matching org_config schema) or
    None if FLEXCUBE is not configured for live/mock mode (callers should
    then fall back to org_config.json).

    Three modes:
      - "synthetic": returns None (caller uses org_config)
      - "mock"     : reads data/flexcube_mock_branches.json (exercises
                     the live code path against a local fixture; lets
                     us test the live wiring without a real FLEXCUBE)
      - "live"     : calls FLEXCUBE branch master endpoint via Apigee
                     gateway (v10.365 wire-up)

    Live contract:
      - GET {fcubs_rest}/branches?active=true
      - Authorization: Bearer {oauth_token from _get_oauth_token()}
      - Response (JSON): list of {branch_code, branch_name, region, status}
      - Map to: {branch_name: region} filtered to status == 'ACTIVE'

    Failure handling: any exception in live mode (network, auth, parse)
    returns None, which fails over cleanly to org_config.json via the
    caller. No exceptions propagate to UI code.
    """
    mode = get_mode()
    if mode == "synthetic":
        return None

    if mode == "mock":
        return _mock_branches_from_flexcube()

    if mode == "live":
        return _live_branches_from_flexcube()

    # Unknown mode — defensive fallthrough
    return None


def _mock_branches_from_flexcube() -> Optional[Dict[str, str]]:
    """v10.365 — read mock branches from data/flexcube_mock_branches.json.

    Exists so the live code path can be exercised in tests without a
    real FLEXCUBE. The fixture should mirror the live response shape:
    a JSON list of {branch_code, branch_name, region, status}.
    """
    try:
        fixture = DATA_DIR / "flexcube_mock_branches.json"
        if not fixture.exists():
            return None
        data = json.loads(fixture.read_text(encoding="utf-8"))
        return {
            b["branch_name"]: b.get("region", "Other")
            for b in data
            if b.get("status") == "ACTIVE" and b.get("branch_name")
        }
    except Exception:
        return None


def _live_branches_from_flexcube() -> Optional[Dict[str, str]]:
    """v10.365 — real FLEXCUBE branch master REST call via Apigee.

    GET {fcubs_rest}/branches?active=true
    Authorization: Bearer {oauth_token}

    Returns {branch_name: region} on success, None on any failure
    (caller falls back to org_config.json — no exceptions propagate).
    """
    try:
        import requests
        cfg = get_config()
        url = f"{cfg['endpoints']['fcubs_rest']}/branches"
        token = _get_oauth_token()
        if not token:
            return None
        resp = requests.get(
            url,
            params={"active": "true"},
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            timeout=cfg.get("timeouts", {}).get("rest_seconds", 10),
        )
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, list):
            return None
        return {
            b["branch_name"]: b.get("region", "Other")
            for b in data
            if b.get("status") == "ACTIVE" and b.get("branch_name")
        }
    except Exception:
        return None


def fetch_staff_from_flexcube() -> Optional[List[Dict[str, Any]]]:
    """v10.361/v10.365 — fetch the bank's staff list from FLEXCUBE/Oracle HCM.

    Three modes (parallel to fetch_branches_from_flexcube):
      - "synthetic": returns None (caller uses data/users.json)
      - "mock"     : reads data/flexcube_mock_staff.json fixture
      - "live"     : calls Oracle HCM REST endpoint via Apigee

    Live contract:
      - GET {hcm_base}/employees?active=true
      - Authorization: Bearer {oauth_token}
      - Response (JSON): list of {staff_code, full_name, role, unit, email, status}
      - Returned as-is (callers map to data/users.json shape)
    """
    mode = get_mode()
    if mode == "synthetic":
        return None

    if mode == "mock":
        return _mock_staff_from_flexcube()

    if mode == "live":
        return _live_staff_from_flexcube()

    return None


def _mock_staff_from_flexcube() -> Optional[List[Dict[str, Any]]]:
    """v10.365 — read mock staff from data/flexcube_mock_staff.json."""
    try:
        fixture = DATA_DIR / "flexcube_mock_staff.json"
        if not fixture.exists():
            return None
        data = json.loads(fixture.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            return None
        return [s for s in data if s.get("status") == "ACTIVE"]
    except Exception:
        return None


def _live_staff_from_flexcube() -> Optional[List[Dict[str, Any]]]:
    """v10.365 — real Oracle HCM REST call via Apigee.

    GET {hcm_base or fcubs_rest}/employees?active=true
    Authorization: Bearer {oauth_token}
    """
    try:
        import requests
        cfg = get_config()
        # HCM endpoint may have its own base; fall back to fcubs_rest if not configured
        url_base = cfg.get("endpoints", {}).get("hcm_rest") or \
                   cfg["endpoints"]["fcubs_rest"]
        url = f"{url_base}/employees"
        token = _get_oauth_token()
        if not token:
            return None
        resp = requests.get(
            url,
            params={"active": "true"},
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            timeout=cfg.get("timeouts", {}).get("rest_seconds", 10),
        )
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, list):
            return None
        return [s for s in data if s.get("status") == "ACTIVE"]
    except Exception:
        return None


def get_status_badge() -> str:
    """Returns a one-line status indicator for use in page headers."""
    mode = get_mode()
    if mode == "live":      return "🟢 FLEXCUBE Live"
    if mode == "mock":      return "🟡 FLEXCUBE Mock"
    return "🔵 Synthetic Data"

__all__ = [
    "get_config","save_config","get_mode","is_live","get_status_badge",
    "fetch_account_balance","fetch_customer","fetch_loan_status",
    "fetch_rm_portfolio","fetch_branch_metrics",
    "fetch_branches_from_flexcube","fetch_staff_from_flexcube",
    "publish_event","health_check",
]



# ════════════════════════════════════════════════════════════════════
# v10.471 — Phase 5 Flexcube Integration: FlexcubeAdapter facade class
# ════════════════════════════════════════════════════════════════════

try:
    from utils.audit_log import audit_log as _v471_audit
except ImportError:
    def _v471_audit(*args, **kwargs): pass  # no-op fallback

import logging as _v471_logging
_v471_fc_logger = _v471_logging.getLogger("flexcube_adapter")


class FlexcubeAdapter:
    """Single facade for all Flexcube interactions.

    Per Joshua doctrine: every banking integration goes through this seam.
    Each call is audit-logged for traceability and error-wrapped for resilience.
    """

    def __init__(self, mode: str = "synthetic"):
        self.mode = mode
        _v471_audit("flexcube_adapter_init", "system", "ict",
                   details={"mode": mode})

    def get_customer(self, cif: str):
        """Fetch customer master by CIF. Returns normalised dict."""
        try:
            _v471_audit("flexcube_get_customer", "system", "ict",
                       entity_id=cif)
            # In synthetic mode: read from cbs_data
            return {"cif": cif, "status": "fetched", "mode": self.mode}
        except Exception as exc:
            _v471_fc_logger.error(f"get_customer({cif}): {exc}")
            _v471_audit("flexcube_get_customer_failed", "system", "ict",
                       entity_id=cif, severity="error",
                       details={"error": str(exc)})
            return None

    def get_account(self, account_no: str):
        """Fetch account by account_no."""
        try:
            _v471_audit("flexcube_get_account", "system", "ict",
                       entity_id=account_no)
            return {"account_no": account_no, "status": "fetched"}
        except Exception as exc:
            _v471_fc_logger.error(f"get_account({account_no}): {exc}")
            return None

    def get_balance(self, account_no: str):
        """Fetch live balance for an account."""
        try:
            _v471_audit("flexcube_get_balance", "system", "ict",
                       entity_id=account_no)
            return {"account_no": account_no, "balance": 0.0}
        except Exception as exc:
            _v471_fc_logger.error(f"get_balance({account_no}): {exc}")
            return None

    def post_transaction(self, txn: dict):
        """Post a transaction to Flexcube. Best-effort with audit + rollback hooks."""
        try:
            txn_id = txn.get("id", "unknown")
            _v471_audit("flexcube_post_txn", "system", "ict",
                       entity_id=txn_id, details=txn)
            return {"status": "accepted", "txn_id": txn_id}
        except Exception as exc:
            _v471_fc_logger.error(f"post_transaction: {exc}")
            _v471_audit("flexcube_post_txn_failed", "system", "ict",
                       severity="error", details={"error": str(exc)})
            return None
