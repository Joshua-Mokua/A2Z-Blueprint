"""utils/live_actuals.py — v10.355 Live Actuals Engine + YoY Growth.

Builds on v10.354's CBS baseline foundation. The actuals pipeline already
exists in utils.actuals_engine (compute_actuals_from_cbs). This module adds
the missing piece: comparing current actuals against the frozen baseline
to produce YoY growth deltas that the BSC can surface.

Flow on app startup (already wired in app.py:_auto_load_cbs_data):
  CBS data refreshed?
    → actuals_engine.compute_actuals_from_cbs() rebuilds the actuals xlsx
    → live_actuals.refresh_yoy() reads the new actuals + the baseline
    → live_actuals.save_yoy_sidecar() writes data/actuals_yoy.json
  BSC page reads YoY values via get_yoy_for(staff_code, kpi_name)

The sidecar pattern keeps the xlsx canonical for downstream readers while
making YoY data efficiently queryable for the BSC. Sidecar is regenerated
on every actuals refresh; never edited by hand.

KPI → baseline metric mapping lives in data/kpi_baseline_mapping.json
(config, editable). Default mappings cover the deposit / loan / NPL /
customer-count KPIs that have direct CBS-aggregate equivalents.
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data"
YOY_SIDECAR_PATH = DATA_DIR / "actuals_yoy.json"
MAPPING_PATH = DATA_DIR / "kpi_baseline_mapping.json"
SCHEMA_VERSION = "1.0"


# Default mappings — KPI name patterns (case-insensitive substring) to
# baseline metric paths. Used when data/kpi_baseline_mapping.json doesn't
# exist or doesn't override a particular pattern.
#
# IMPORTANT — order matters: first match wins. List specific patterns
# BEFORE generic ones (e.g. "SME Loan Book" before "Loan Book").
DEFAULT_MAPPINGS: List[Dict[str, str]] = [
    # ── Segment-specific loans (specific before generic) ──
    {
        "kpi_pattern": "sme loan book",
        "baseline_path": "loans_aggregate.by_segment_kes.SME",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "corporate loan book",
        "baseline_path": "loans_aggregate.by_segment_kes.CORPORATE",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "retail loan book",
        "baseline_path": "loans_aggregate.by_segment_kes.RETAIL_INDIVIDUAL",
        "direction": "higher_is_better",
    },
    # ── Segment-specific deposits ──
    {
        "kpi_pattern": "retail & msme deposit",
        "baseline_path": "deposits_aggregate.by_segment_kes.RETAIL_INDIVIDUAL",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "retail deposit",
        "baseline_path": "deposits_aggregate.by_segment_kes.RETAIL_INDIVIDUAL",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "sme deposit",
        "baseline_path": "deposits_aggregate.by_segment_kes.SME",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "commercial deposit",
        "baseline_path": "deposits_aggregate.by_segment_kes.CORPORATE",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "corporate deposit",
        "baseline_path": "deposits_aggregate.by_segment_kes.CORPORATE",
        "direction": "higher_is_better",
    },
    # ── Generic deposit/loan KPIs ──
    {
        "kpi_pattern": "customer deposits",
        "baseline_path": "deposits_aggregate.total_deposits_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "total deposits",
        "baseline_path": "deposits_aggregate.total_deposits_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "deposit growth",
        "baseline_path": "deposits_aggregate.total_deposits_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "customer loans",
        "baseline_path": "loans_aggregate.gross_outstanding_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "loan book growth",
        "baseline_path": "loans_aggregate.gross_outstanding_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "loan book",
        "baseline_path": "loans_aggregate.gross_outstanding_kes",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "loan growth",
        "baseline_path": "loans_aggregate.gross_outstanding_kes",
        "direction": "higher_is_better",
    },
    # ── NPL ──
    {
        "kpi_pattern": "npl ratio",
        "baseline_path": "npl_aggregate.npl_ratio_pct",
        "direction": "lower_is_better",
    },
    {
        "kpi_pattern": "stage 3",
        "baseline_path": "npl_aggregate.stage_3_kes",
        "direction": "lower_is_better",
    },
    # ── Customer counts ──
    {
        "kpi_pattern": "active customers",
        "baseline_path": "customer_aggregate.total_customers",
        "direction": "higher_is_better",
    },
    {
        "kpi_pattern": "customer growth",
        "baseline_path": "customer_aggregate.total_customers",
        "direction": "higher_is_better",
    },
    # ── Loan-to-deposit ──
    {
        "kpi_pattern": "loan to deposit",
        "baseline_path": "deposits_aggregate.loan_to_deposit_ratio_pct",
        "direction": "neutral",
    },
]


def load_mapping() -> List[Dict[str, str]]:
    """Load KPI → baseline mappings. Falls back to DEFAULT_MAPPINGS if
    data/kpi_baseline_mapping.json doesn't exist or is malformed."""
    if MAPPING_PATH.exists():
        try:
            payload = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
            mappings = payload.get("mappings", DEFAULT_MAPPINGS)
            if isinstance(mappings, list) and mappings:
                return mappings
        except Exception:
            pass
    return DEFAULT_MAPPINGS


def _resolve_baseline_metric(baseline: Dict[str, Any], path: str) -> Optional[float]:
    """Resolve a dotted path into baseline['bank_aggregates']. Coerces
    numeric strings (some CBS JSON aggregates use strings for very-large
    integers to preserve precision)."""
    parts = path.split(".")
    node: Any = baseline.get("bank_aggregates", {})
    for p in parts:
        if not isinstance(node, dict):
            return None
        node = node.get(p)
    if isinstance(node, (int, float)):
        return float(node)
    if isinstance(node, str):
        try:
            return float(node)
        except ValueError:
            return None
    return None


def _find_mapping_for_kpi(
    kpi_name: str, mappings: List[Dict[str, str]]
) -> Optional[Dict[str, str]]:
    """Return the first mapping whose kpi_pattern matches the KPI name."""
    name_lc = (kpi_name or "").lower()
    for m in mappings:
        pat = (m.get("kpi_pattern") or "").lower()
        if pat and pat in name_lc:
            return m
    return None


def _safe_growth(current: float, baseline: float) -> Optional[float]:
    """YoY growth %. None when baseline is 0 (undefined)."""
    if baseline is None or baseline == 0:
        return None
    return (current - baseline) / baseline * 100.0


def compute_yoy_for_rows(
    rows: List[Dict[str, Any]],
    baseline: Dict[str, Any],
    mappings: Optional[List[Dict[str, str]]] = None,
) -> Dict[str, Any]:
    """For each actuals row, attach YoY growth if the KPI maps to a
    baseline metric.

    Args:
      rows: list of dicts with keys 'Staff Code', 'KPI', 'Annual Actual'
        (or similar). Tolerates legacy column names.
      baseline: a baseline dict loaded via cbs_baseline.load_baseline().
      mappings: optional override; defaults to load_mapping().

    Returns:
      Sidecar dict structured as:
        {
          "_doc": "...",
          "_schema_version": "1.0",
          "computed_at": ISO timestamp,
          "baseline_date": baseline's snapshot_date,
          "mapped_count": int,
          "entries": {
            "<staff_code>__<kpi_name>": {
              "staff_code": str,
              "kpi_name": str,
              "current_value": float,
              "baseline_value": float,
              "growth_pct": float | None,
              "direction": "higher_is_better" | "lower_is_better" | "neutral",
              "baseline_path": str,
            }
          }
        }
    """
    if mappings is None:
        mappings = load_mapping()

    baseline_date = baseline.get("snapshot_date", "unknown")
    entries: Dict[str, Dict[str, Any]] = {}
    mapped = 0

    for row in rows:
        sc = str(row.get("Staff Code") or row.get("staff_code") or "").strip()
        kpi = str(row.get("KPI") or row.get("kpi_name") or "").strip()
        if not sc or not kpi:
            continue
        m = _find_mapping_for_kpi(kpi, mappings)
        if m is None:
            continue
        baseline_val = _resolve_baseline_metric(baseline, m["baseline_path"])
        if baseline_val is None:
            continue
        cur_val = row.get("Annual Actual") or row.get("YTD_Actual") or 0
        try:
            cur_val_f = float(cur_val)
        except (TypeError, ValueError):
            continue
        growth = _safe_growth(cur_val_f, baseline_val)
        key = f"{sc}__{kpi}"
        entries[key] = {
            "staff_code": sc,
            "kpi_name": kpi,
            "current_value": cur_val_f,
            "baseline_value": baseline_val,
            "growth_pct": growth,
            "direction": m.get("direction", "neutral"),
            "baseline_path": m["baseline_path"],
        }
        mapped += 1

    return {
        "_doc": (
            "YoY growth sidecar for current actuals. Regenerated on every "
            "actuals refresh. Keyed by '<staff_code>__<kpi_name>'. "
            "Read via utils.live_actuals.get_yoy_for(staff_code, kpi_name)."
        ),
        "_schema_version": SCHEMA_VERSION,
        "computed_at": datetime.now(timezone.utc).isoformat(),
        "baseline_date": baseline_date,
        "mapped_count": mapped,
        "entries": entries,
    }


def save_yoy_sidecar(sidecar: Dict[str, Any], path: Optional[Path] = None) -> Path:
    """Atomic write of the YoY sidecar JSON."""
    if path is None:
        path = YOY_SIDECAR_PATH

    # Pattern Q — validate-before-save
    try:
        from utils.schema_validator import validate_before_save
        result = validate_before_save("actuals_yoy.json", sidecar)
        if not result.get("valid"):
            errs = result.get("errors", [])
            raise ValueError(
                f"Refusing to save invalid YoY sidecar: "
                f"{len(errs)} schema error(s). First: {errs[:3]}"
            )
    except ImportError:
        pass

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(sidecar, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)
    return path


def load_yoy_sidecar(path: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    """Load the YoY sidecar. Returns None if absent."""
    if path is None:
        path = YOY_SIDECAR_PATH
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def get_yoy_for(staff_code: str, kpi_name: str) -> Optional[Dict[str, Any]]:
    """Public lookup. Returns YoY entry for (staff_code, kpi_name) or
    None if no mapping/baseline is available."""
    sidecar = load_yoy_sidecar()
    if not sidecar:
        return None
    entries = sidecar.get("entries", {})
    if not isinstance(entries, dict):
        return None
    key = f"{staff_code}__{kpi_name}"
    return entries.get(key)


def _read_actuals_xlsx(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read an actuals xlsx into row dicts. Tolerates the project's
    convention where row 2 is the actual header row."""
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    headers: List[Any] = []
    for cells in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        if cells and "Staff Code" in cells:
            headers = list(cells)
            break
    if not headers:
        # Fallback: assume row 1
        headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    start_row = 3 if "Staff Code" in (
        next(ws.iter_rows(min_row=2, max_row=2, values_only=True)) or []
    ) else 2
    rows: List[Dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=start_row, values_only=True):
        if not cells or all(c is None for c in cells):
            continue
        rows.append(dict(zip(headers, cells)))
    return rows


def discover_newest_actuals() -> Optional[Path]:
    """Return the most recent data/actuals_*.xlsx file, or None."""
    if not DATA_DIR.exists():
        return None
    candidates = sorted(DATA_DIR.glob("actuals_*.xlsx"), reverse=True)
    return candidates[0] if candidates else None


def refresh_yoy(
    actuals_path: Optional[Path] = None,
    baseline_path: Optional[Path] = None,
) -> Dict[str, Any]:
    """Convenience: load newest actuals + baseline, compute YoY, save sidecar.

    Args:
      actuals_path: override actuals source; defaults to discover_newest_actuals().
      baseline_path: override baseline source; defaults to most recent baseline.

    Returns the saved sidecar dict, or a status dict on error.
    """
    if actuals_path is None:
        actuals_path = discover_newest_actuals()
    if actuals_path is None or not actuals_path.exists():
        return {
            "_doc": "no actuals file found",
            "_schema_version": SCHEMA_VERSION,
            "computed_at": datetime.now(timezone.utc).isoformat(),
            "baseline_date": "n/a",
            "mapped_count": 0,
            "entries": {},
        }

    from utils.cbs_baseline import load_baseline
    baseline = load_baseline(baseline_path) if baseline_path else load_baseline()
    if baseline is None:
        return {
            "_doc": "no baseline available — run scripts/snapshot_cbs_baseline.py",
            "_schema_version": SCHEMA_VERSION,
            "computed_at": datetime.now(timezone.utc).isoformat(),
            "baseline_date": "n/a",
            "mapped_count": 0,
            "entries": {},
        }

    rows = _read_actuals_xlsx(actuals_path)
    sidecar = compute_yoy_for_rows(rows, baseline)
    save_yoy_sidecar(sidecar)
    return sidecar


def format_yoy_label(entry: Dict[str, Any]) -> str:
    """Human-readable YoY label for BSC display.

    Example: '+12.4% vs baseline (110.0B → 123.6B)' for positive deposit growth
             '-2.3% vs baseline (11.1% → 10.8%)' for an improvement in NPL ratio
             'baseline 0 — growth undefined' when denominator is zero
    """
    if not entry:
        return ""
    growth = entry.get("growth_pct")
    cur = entry.get("current_value", 0)
    base = entry.get("baseline_value", 0)
    if growth is None:
        return "baseline 0 — growth undefined"
    direction = entry.get("direction", "neutral")
    sign = "+" if growth >= 0 else ""
    cur_disp = _format_kes(cur)
    base_disp = _format_kes(base)
    label = f"{sign}{growth:.1f}% vs baseline ({base_disp} → {cur_disp})"
    if direction == "higher_is_better" and growth >= 0:
        label = "📈 " + label
    elif direction == "lower_is_better" and growth < 0:
        label = "📈 " + label  # negative growth on lower-is-better is improvement
    elif direction in ("higher_is_better", "lower_is_better"):
        label = "📉 " + label
    return label


def _format_kes(value: float) -> str:
    """Format a numeric value in KES-friendly scale. 1B+ → 'X.XB', 1M+ → 'X.XM'."""
    if value is None:
        return "—"
    abs_v = abs(value)
    if abs_v >= 1e9:
        return f"{value/1e9:.1f}B"
    if abs_v >= 1e6:
        return f"{value/1e6:.1f}M"
    if abs_v >= 1e3:
        return f"{value/1e3:.1f}K"
    return f"{value:.1f}"
