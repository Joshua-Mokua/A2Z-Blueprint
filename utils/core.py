"""utils/core.py — Shared state: all managers, constants, helpers, data processing."""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
import datetime as _dt
import hashlib
import json
import logging
import re
import io
import smtplib
import secrets
import string
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart




import smtplib
import secrets
import string
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# P-AUTH-c: anchor DATA_DIR to an ABSOLUTE path derived from this file's
# location, NOT the process working directory. Previously `Path("data")`
# resolved against the CWD, so launching the app from a different folder
# (e.g. a nested duplicate tree) silently pointed UserManager at a DIFFERENT
# users.json — a root cause of vanishing test logins. This ties the data
# dir to wherever core.py physically lives (<repo>/data).
DATA_DIR = (Path(__file__).resolve().parent.parent / "data")
DATA_DIR.mkdir(exist_ok=True)

# Batch 3c — module logger for auth-path observability (envelope-verify
# success log, auto-upgrade failure instrumentation). Uses the standard
# a2z.<module> namespace; no handler attached here (FastAPI/Streamlit
# entry points configure handlers globally).
logger = logging.getLogger("a2z.core")
for f in ["users.json", "validations.json", "audit_log.json", "calendar_events.json",
          "staff_history.json", "email_config.json", "pending_tokens.json"]:
    p = DATA_DIR / f
    if not p.exists():
        p.write_text("{}")

# ─── PASSWORD UTILITIES ──────────────────────────────────────────────
def check_password_strength(pw):
    """Returns (ok: bool, issues: list[str])"""
    issues = []
    if len(pw) < 8:            issues.append("At least 8 characters")
    if not re.search(r'[A-Z]', pw): issues.append("At least one uppercase letter")
    if not re.search(r'[a-z]', pw): issues.append("At least one lowercase letter")
    if not re.search(r'\d', pw):    issues.append("At least one number")
    if not re.search(r'[^A-Za-z0-9]', pw): issues.append("At least one special character (!@#$%^&*)")
    return (len(issues) == 0, issues)

def generate_temp_password(length=12):
    alpha  = string.ascii_letters
    digits = string.digits
    spec   = "!@#$%^&*"
    chars  = alpha + digits + spec
    while True:
        pw = ''.join(secrets.choice(chars) for _ in range(length))
        ok, _ = check_password_strength(pw)
        if ok: return pw

def generate_token():
    return secrets.token_urlsafe(32)

# ─── EMAIL SENDER ────────────────────────────────────────────────────
def load_email_config():
    try:
        raw = (DATA_DIR/"email_config.json").read_text()
        return json.loads(raw) if raw.strip() else {}
    except: return {}

def save_email_config(cfg):
    (DATA_DIR/"email_config.json").write_text(json.dumps(cfg, indent=2))

def send_milestone_alert_email(to_email, recipient_name, ms_name, init_name,
                               due_date, days_info, esc_level, workstream, io_name):
    """Send milestone escalation/due-soon email."""
    cfg = load_email_config()
    if not cfg.get("smtp_host") or not cfg.get("sender_email"):
        return False, "Email not configured"
    esc_cfg = ESC_CONFIG.get(esc_level, ESC_CONFIG[1])
    subject = f"A2Z Execute — {esc_cfg['icon']} Milestone alert: {ms_name}"
    body = f"""
<html><body style="font-family:Arial,sans-serif;max-width:520px;margin:auto">
<div style="background:#1B4F72;padding:20px;border-radius:8px 8px 0 0">
  <h2 style="color:var(--color-background-primary);margin:0">A2Z Execute</h2>
  <p style="color:#AED6F1;margin:4px 0 0">Milestone Alert</p>
</div>
<div style="padding:20px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px">
  <div style="padding:10px 14px;background:{esc_cfg['bg']};border-left:4px solid {esc_cfg['color']};
              border-radius:0 6px 6px 0;margin-bottom:16px">
    <b style="color:{esc_cfg['color']}">{esc_cfg['icon']} {esc_cfg['label']}</b>
  </div>
  <p>Hi <b>{recipient_name}</b>,</p>
  <table style="width:100%;border-collapse:collapse;font-size:13px;margin:12px 0">
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Milestone</td>
        <td style="padding:6px 10px"><b>{ms_name}</b></td></tr>
    <tr><td style="padding:6px 10px;color:#555">Initiative</td>
        <td style="padding:6px 10px">{init_name}</td></tr>
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Workstream</td>
        <td style="padding:6px 10px">{workstream}</td></tr>
    <tr><td style="padding:6px 10px;color:#555">Due date</td>
        <td style="padding:6px 10px"><b>{due_date}</b> — {days_info}</td></tr>
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">IO</td>
        <td style="padding:6px 10px">{io_name}</td></tr>
  </table>
  <p style="font-size:13px;color:#555">Please update the milestone status on A2Z Execute immediately.</p>
  <p style="font-size:11px;color:#bbb;margin-top:24px">A2Z Perform · Automated alert</p>
</div></body></html>"""
    try:
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import smtplib
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = cfg["sender_email"]
        msg["To"]      = to_email
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port",587))) as s:
            s.starttls()
            s.login(cfg["sender_email"], cfg["sender_password"])
            s.sendmail(cfg["sender_email"], to_email, msg.as_string())
        return True, "Sent"
    except Exception as e:
        return False, str(e)

def send_structural_delay_email(to_emails, ms_name, init_name, workstream,
                                 delay_category, delay_reason, owner, io_name):
    """Immediate all-parties alert for structural/regulatory delay."""
    cfg = load_email_config()
    if not cfg.get("smtp_host"): return False, "Email not configured"
    subject = f"🚨 A2Z Execute — IMMEDIATE: Structural delay on {ms_name}"
    body = f"""
<html><body style="font-family:Arial,sans-serif;max-width:520px;margin:auto">
<div style="background:#A32D2D;padding:20px;border-radius:8px 8px 0 0">
  <h2 style="color:var(--color-background-primary);margin:0">A2Z Execute — Immediate escalation</h2>
  <p style="color:#F7C1C1;margin:4px 0 0">Structural / Regulatory Delay — All parties notified</p>
</div>
<div style="padding:20px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px">
  <div style="padding:10px 14px;background:#FCEBEB;border-left:4px solid #A32D2D;border-radius:0 6px 6px 0;margin-bottom:16px">
    <b style="color:#A32D2D">Category: {delay_category}</b><br>
    <span style="font-size:13px;color:#791F1F">{delay_reason}</span>
  </div>
  <table style="width:100%;border-collapse:collapse;font-size:13px;margin:12px 0">
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Milestone</td><td style="padding:6px 10px"><b>{ms_name}</b></td></tr>
    <tr><td style="padding:6px 10px;color:#555">Initiative</td><td style="padding:6px 10px">{init_name}</td></tr>
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Workstream</td><td style="padding:6px 10px">{workstream}</td></tr>
    <tr><td style="padding:6px 10px;color:#555">Milestone owner</td><td style="padding:6px 10px">{owner}</td></tr>
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Initiative owner</td><td style="padding:6px 10px">{io_name}</td></tr>
  </table>
  <p style="color:#A32D2D;font-weight:bold;font-size:13px">This delay requires immediate management intervention. Please review and take action on A2Z Execute.</p>
  <p style="font-size:11px;color:#bbb;margin-top:24px">A2Z Perform · Automated alert</p>
</div></body></html>"""
    try:
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import smtplib
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = cfg["sender_email"]
        msg["To"]      = ", ".join(to_emails)
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port",587))) as s:
            s.starttls()
            s.login(cfg["sender_email"], cfg["sender_password"])
            s.sendmail(cfg["sender_email"], to_emails, msg.as_string())
        return True, "Sent"
    except Exception as e:
        return False, str(e)


def send_start_alert_email(to_email, recipient_name, ms_name, init_name, start_date, workstream):
    """Email to milestone owner when start date arrives."""
    cfg = load_email_config()
    if not cfg.get("smtp_host"): return False, "Email not configured"
    subject = f"A2Z Execute — ⏰ Milestone starting today: {ms_name}"
    body = f"""
<html><body style="font-family:Arial,sans-serif;max-width:520px;margin:auto">
<div style="background:#185FA5;padding:20px;border-radius:8px 8px 0 0">
  <h2 style="color:var(--color-background-primary);margin:0">A2Z Execute</h2>
  <p style="color:#B5D4F4;margin:4px 0 0">Milestone starting today</p>
</div>
<div style="padding:20px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px">
  <p>Hi <b>{recipient_name}</b>,</p>
  <p>Your milestone is scheduled to <b>start today</b>.</p>
  <table style="width:100%;border-collapse:collapse;font-size:13px;margin:12px 0">
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Milestone</td><td style="padding:6px 10px"><b>{ms_name}</b></td></tr>
    <tr><td style="padding:6px 10px;color:#555">Initiative</td><td style="padding:6px 10px">{init_name}</td></tr>
    <tr style="background:#F8F9FA"><td style="padding:6px 10px;color:#555">Workstream</td><td style="padding:6px 10px">{workstream}</td></tr>
    <tr><td style="padding:6px 10px;color:#555">Start date</td><td style="padding:6px 10px"><b>{start_date}</b></td></tr>
  </table>
  <p style="font-size:13px;color:#555">Please mark the milestone as <b>In Progress</b> on A2Z Execute to confirm it has started.</p>
  <p style="font-size:11px;color:#bbb;margin-top:24px">A2Z Perform · Automated alert</p>
</div></body></html>"""
    try:
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import smtplib
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = cfg["sender_email"]
        msg["To"]      = to_email
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port",587))) as s:
            s.starttls()
            s.login(cfg["sender_email"], cfg["sender_password"])
            s.sendmail(cfg["sender_email"], to_email, msg.as_string())
        return True, "Sent"
    except Exception as e:
        return False, str(e)


def run_escalation_scan(em, user_registry):
    """
    Called once per session on app load.
    Scans all active milestones and sends alerts where needed.
    user_registry: dict of {username: {email, full_name, role}}
    Returns list of alert log entries.
    """
    alerts = []
    today  = date.today()

    def get_email(username):
        u = user_registry.get(username, {})
        return u.get('email',''), u.get('full_name', username)

    for init in em.get_initiatives(status='All'):
        if init.get('gate') not in ('G3','G4'): continue
        for ms in init.get('milestones', []):
            if ms['status'] == 'Complete': continue
            esc = ExecuteManager._escalation_level(ms)
            if esc == 0: continue

            owner_email, owner_name = get_email(ms.get('owner',''))
            io_email,    io_name    = get_email(init.get('io',''))
            workstream   = init.get('workstream','')

            try:
                due       = date.fromisoformat(ms.get('due_date',''))
                days_over = (today - due).days
                days_to   = (due - today).days
            except:
                days_over, days_to = 0, 0

            # ── Due in exactly 2 days — due-soon email to owner ────
            if days_to == 2:
                if owner_email:
                    send_milestone_alert_email(
                        owner_email, owner_name, ms['name'], init['name'],
                        ms.get('due_date',''), f"due in 2 days", 1, workstream, io_name)
                    alerts.append({'type':'due_soon','ms':ms['name'],'to':owner_email})

            # ── Start date = today — start alert ──────────────────
            if (ms.get('start_date') == str(today) and
                ms['status'] == 'Not Started' and not ms.get('has_started')):
                if owner_email:
                    send_start_alert_email(
                        owner_email, owner_name, ms['name'], init['name'],
                        ms['start_date'], workstream)
                    alerts.append({'type':'start_today','ms':ms['name'],'to':owner_email})

            # ── Structural/regulatory delay → all parties immediately ─
            if (ms.get('delay_category','') in STRUCTURAL_CATEGORIES and
                ms['status'] == 'Delayed'):
                all_emails = [e for e in [owner_email, io_email] if e]
                # Try to find lead/sponsor from workstream
                if all_emails:
                    send_structural_delay_email(
                        all_emails, ms['name'], init['name'], workstream,
                        ms['delay_category'], ms.get('delay_reason',''),
                        ms.get('owner',''), io_name)
                    alerts.append({'type':'structural','ms':ms['name'],'level':4})

            # ── Day 2 overdue → IO email ───────────────────────────
            elif days_over >= 2 and esc >= 2 and io_email:
                send_milestone_alert_email(
                    io_email, io_name, ms['name'], init['name'],
                    ms.get('due_date',''), f"{days_over}d overdue", esc, workstream, io_name)
                alerts.append({'type':'overdue_io','ms':ms['name'],'days':days_over})

    return alerts


def should_run_scan():
    """Run escalation scan once per day per session."""
    last = st.session_state.get('last_esc_scan')
    today_str = str(date.today())
    if last != today_str:
        st.session_state['last_esc_scan'] = today_str
        return True
    return False


def send_welcome_email(to_email, full_name, username, temp_password):
    cfg = load_email_config()
    if not cfg.get("smtp_host") or not cfg.get("sender_email"):
        return False, "Email not configured. Go to Admin → Email Settings."
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "A2Z Perform — Your account has been created"
        msg["From"]    = cfg["sender_email"]
        msg["To"]      = to_email
        body = f"""
<html><body style="font-family:Arial,sans-serif;max-width:500px;margin:auto">
<div style="background:#1B4F72;padding:20px;border-radius:8px 8px 0 0">
  <h2 style="color:var(--color-background-primary);margin:0">🏦 A2Z Perform</h2>
  <p style="color:#AED6F1;margin:4px 0 0">Performance Management System</p>
</div>
<div style="padding:24px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 8px 8px">
  <p>Hi <strong>{full_name}</strong>,</p>
  <p>Your account has been created on <strong>A2Z Perform</strong>.</p>
  <table style="background:#F8F9FA;padding:16px;border-radius:6px;width:100%;border-collapse:collapse">
    <tr><td style="padding:4px 8px;color:#555">Username</td><td style="padding:4px 8px"><strong>{username}</strong></td></tr>
    <tr><td style="padding:4px 8px;color:#555">Temp password</td><td style="padding:4px 8px"><strong style="font-size:18px;letter-spacing:2px">{temp_password}</strong></td></tr>
  </table>
  <p style="color:#E74C3C;font-weight:bold">⚠️ You will be prompted to set a new password on first login.</p>
  <p style="font-size:12px;color:#999">Your new password must be at least 8 characters and include uppercase, lowercase, a number, and a special character.</p>
  <p style="font-size:12px;color:#bbb;margin-top:32px">This is an automated message from A2Z Perform. Do not reply.</p>
</div>
</body></html>"""
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port", 587))) as s:
            s.starttls()
            s.login(cfg["sender_email"], cfg["sender_password"])
            s.sendmail(cfg["sender_email"], to_email, msg.as_string())
        return True, "Email sent"
    except Exception as e:
        return False, str(e)

# ─── HELPER FUNCTIONS ────────────────────────────────────────────────
def get_performance_remarks(score):
    if pd.isna(score):   return "No Data"
    elif score < 2.5:    return "Unmet"
    elif score < 3.0:    return "Partially Met"
    elif score < 3.1:    return "Met"
    elif score < 4.0:    return "Exceeded"
    else:                return "Exceeded By Far"

def highlight_performance(val):
    colors = {
        "Exceeded By Far": "background-color:#90EE90;color:#000",
        "Exceeded":        "background-color:#C1E1C1;color:#000",
        "Met":             "background-color:#FFE4B5;color:#000",
        "Partially Met":   "background-color:#FFD700;color:#000",
        "Unmet":           "background-color:#FFB6C1;color:#000",
        "No Data":         "background-color:#D3D3D3;color:#000",
    }
    return colors.get(val, "")

# Count-type KPIs — whole number display, no KES prefix
COUNT_KPIS_CORE = {
    "Transactions", "New Customer Acquisition", "Dormancy Reactivation",
    "Cards Issued", "Loans Referred", "Bancassurance",
    "New Accounts", "Number of Business Borrowers",
    "Active Initiatives Count", "Number of New Customers",
}

def is_count_kpi_core(k):
    """True if KPI is a count — display as integer not KES."""
    return (str(k) in COUNT_KPIS_CORE or
            any(x in str(k).lower() for x in
                ("accounts opened","cards issued","registrations",
                 "referred","reactivat","acquisition","transactions processed",
                 "number of","borrower","initiatives count","new accounts")))

def fmt_kpi_value(v, kpi="", short=False):
    """Format a KPI value — BSC score / % / count / KES (5-tier)."""
    try:
        if pd.isna(v) or v is None: return "—"
        f = float(v)
        if f == 0: return "—"
        kl = str(kpi).lower()

        # Tier 1: BSC score KPIs (1–5 scale) — plain decimal, never %
        _bsc_score_kpis = {
            "staff productivity", "diligence score", "cx score",
            "nps score", "employee satisfaction score",
            "ideation score", "initiative score",
        }
        if kl in _bsc_score_kpis:
            return f"{f:.2f}"

        # Tier 2: Always KES — never treated as % even if name has ratio/rate
        _kes_kpi_names = {
            "disbursements corporate loans", "disbursements retail loans",
            "disbursements msme loans", "loans disbursement",
            "loan book growth", "retail & msme deposit growth",
            "commercial deposit growth", "deposit growth",
            "total nfi", "fees and commission", "pbt",
            "top 100 customers deposit", "collection throughput",
            "trade finance", "treasury revenue", "bancassurance", "dfs revenue",
        }
        if kl in _kes_kpi_names:
            if short:
                if abs(f) >= 1e9: return f"KES {f/1e9:,.2f}B"
                if abs(f) >= 1e6: return f"KES {f/1e6:,.2f}M"
                return f"KES {f:,.0f}"
            if abs(f) >= 1e9: return f"KES {f/1e9:,.3f}B"
            if abs(f) >= 1e6: return f"KES {f/1e6:,.2f}M"
            return f"KES {f:,.0f}"

        # Tier 3: Percentage/ratio KPIs
        _pct_explicit = {
            "compliance score", "audit score", "audit closure",
            "timely reconciliations", "sla adherence score",
            "credit tat score", "campaign conversion rate",
            "digital transaction migration", "branch optimization score",
            "strategic initiative completion rate", "training completion rate",
            "staff retention rate", "recovery rate", "case resolution rate",
            "account dormancy", "channel dormancy",
        }
        if (kl in _pct_explicit or
                any(x in kl for x in ("ratio", " rate", "margin", "tat", "closure", "%",
                                       "cir", "npl", "par", "roe", "roa", "nim", "ldr", "car",
                                       "compliance", "reconciliation", "dormancy",
                                       "uptime", "coverage", "utilisation"))):
            disp = f * 100 if abs(f) <= 1.5 else f
            return f"{disp:,.2f}%"

        # Tier 4: Count KPIs
        if is_count_kpi_core(kpi):
            return f"{int(round(f)):,}"

        # Tier 5: Financial KES
        if short:
            if abs(f) >= 1e9: return f"KES {f/1e9:,.2f}B"
            if abs(f) >= 1e6: return f"KES {f/1e6:,.2f}M"
            if abs(f) >= 1e3: return f"KES {f/1e3:,.1f}K"
            return f"KES {f:,.0f}"
        if abs(f) >= 1e9: return f"KES {f/1e9:,.3f}B"
        if abs(f) >= 1e6: return f"KES {f/1e6:,.2f}M"
        return f"KES {f:,.0f}"
    except: return str(v)
def fmt_num(v, short=False):
    """Accounting format: comma separator, 2 decimal places."""
    try:
        if pd.isna(v): return "—"
        f = float(v)
        if short:
            if abs(f) >= 1_000_000_000: return f"{f/1e9:,.2f}B"
            if abs(f) >= 1_000_000:     return f"{f/1e6:,.2f}M"
            if abs(f) >= 1_000:         return f"{f/1e3:,.2f}K"
        return f"{f:,.2f}"
    except: return str(v)

def fmt_pct(v):
    try:
        if pd.isna(v): return "—"
        return f"{float(v):.1f}%"
    except: return str(v)

def fmt_score(v):
    try:
        if pd.isna(v): return "—"
        return f"{float(v):.2f}"
    except: return str(v)

def clean_code(v):
    """Normalize staff code — strip spaces and Excel apostrophe on both sides."""
    return str(v).strip().strip("'").strip()




# ══════════════════════════════════════════════════════════════════
# BSC MONTH-END LOCK
# Once locked, scores are frozen for audit. Only MD/Admin can unlock.
# ══════════════════════════════════════════════════════════════════
def is_bsc_locked(data_path="data") -> bool:
    """Return True if BSC is locked for the current period."""
    import json
    from pathlib import Path
    p = Path(data_path) / "bsc_lock.json"
    if not p.exists(): return False
    lock = json.loads(p.read_text())
    return lock.get("locked", False)

def lock_bsc(locked_by: str, period: str, data_path="data"):
    """Lock BSC scores for the period. Only callable by admin/MD."""
    import json
    from datetime import date
    from pathlib import Path
    p = Path(data_path) / "bsc_lock.json"
    lock_data = {
        "locked":     True,
        "locked_by":  locked_by,
        "period":     period,
        "locked_at":  str(date.today()),
        "unlock_log": []
    }
    p.write_text(json.dumps(lock_data, indent=2))

def unlock_bsc(unlocked_by: str, reason: str, data_path="data"):
    """Unlock BSC scores. Requires reason for audit."""
    import json
    from datetime import date
    from pathlib import Path
    p = Path(data_path) / "bsc_lock.json"
    lock = json.loads(p.read_text()) if p.exists() else {}
    lock["locked"] = False
    lock.setdefault("unlock_log", []).append({
        "unlocked_by": unlocked_by,
        "reason": reason,
        "date": str(date.today())
    })
    p.write_text(json.dumps(lock, indent=2))

# ═══════════════════════════════════════════════════════════════════════
# PRODUCT LIFECYCLE MODULE — Banking product registry
# Assets · Liabilities · NFI · Channels
# ═══════════════════════════════════════════════════════════════════════

PRODUCT_CATEGORIES = {
    "Assets": {
        "color": "#185FA5", "bg": "#E6F1FB", "icon": "📈",
        "description": "Lending products — income generating",
        "kpi_links": ["Loans Disbursement","Loan Book Growth","Trade Finance","NPL","PAR"],
        "lifecycle": ["Concept","Pilot","Active","Growth","Mature","Sunset"],
        "sub_categories": {
            "Retail Lending":    ["Personal Loan","Salary Advance","Mortgage / Home Loan",
                                   "Auto Finance","Education Loan","Emergency Loan"],
            "Business Lending":  ["Business Loan","Overdraft (OD)","Invoice Discounting",
                                   "Asset Finance","Contract Finance","Working Capital Loan"],
            "Corporate & SME":   ["Term Loan","Revolving Credit Facility","Trade Finance",
                                   "Supply Chain Finance","Syndicated Loan"],
            "Specialised":       ["Agricultural Loan","Green Finance","Government / County Loan"],
        },
    },
    "Liabilities": {
        "color": "var(--brand-hover,#0F6E56)", "bg": "#E1F5EE", "icon": "🏦",
        "description": "Deposit-taking products — balance sheet funding",
        "kpi_links": ["Deposit Growth"],
        "lifecycle": ["Concept","Pilot","Active","Growth","Mature","Sunset"],
        "sub_categories": {
            "Retail Deposits":   ["Current Account","Savings Account","Fixed Deposit",
                                   "Junior Account","Senior Account"],
            "Business Deposits": ["Business Current Account","Business Savings",
                                   "Business Fixed Deposit","Salary Processing Account"],
            "Corporate":         ["Call Account","Treasury Placement","Structured Deposit"],
            "Digital Deposits":  ["Mobile Wallet","Digital Savings (App)","Virtual Account"],
        },
    },
    "NFI": {
        "color": "#534AB7", "bg": "#EEEDFE", "icon": "💰",
        "description": "Non-funded income — fees, commissions, insurance",
        "kpi_links": ["Fees and Commission","DFS Revenue","Bancassurance","Treasury","Acquiring"],
        "lifecycle": ["Concept","Pilot","Active","Growth","Mature","Sunset"],
        "sub_categories": {
            "Bancassurance":  ["Life Cover","Credit Life","Motor Insurance",
                               "Home Insurance","Business Insurance"],
            "DFS":            ["Mobile Banking","Internet Banking","USSD Banking",
                               "Agency Banking","Merchant Acquiring"],
            "Transactional":  ["Forex / FX","Wire Transfers","Trade LC","Bank Guarantees"],
            "Treasury":       ["T-Bills","Bonds","FX Trading","Structured Products"],
            "Cards":          ["Debit Card","Credit Card","Prepaid Card"],
        },
    },
    "Channels": {
        "color": "#BA7517", "bg": "#FAEEDA", "icon": "📡",
        "description": "Delivery mechanisms — how products reach customers",
        "kpi_links": ["Digital Transaction Migration","Transactions","Customer Growth"],
        "lifecycle": ["Planning","Development","Launch","Active","Optimising","Decommissioned"],
        "sub_categories": {
            "Physical":    ["Branch Network","ATM Network","Agent Banking","Kiosks"],
            "Digital":     ["Mobile App","Internet Banking Portal","USSD (*XXX#)","WhatsApp Banking"],
            "Partnership": ["Bancassurance Partners","Telco Partnerships",
                            "Fintech Integrations","Government Portals"],
        },
    },
}

PRODUCT_LIFECYCLE_STAGES = {
    "Concept":         {"color":"#888780","bg":"#F1EFE8","desc":"Idea stage — being evaluated"},
    "Pilot":           {"color":"#185FA5","bg":"#E6F1FB","desc":"Limited rollout — testing"},
    "Active":          {"color":"var(--brand-mid,#1D9E75)","bg":"#E1F5EE","desc":"Live and available to customers"},
    "Growth":          {"color":"#3B6D11","bg":"#EAF3DE","desc":"Scaling — high acquisition focus"},
    "Mature":          {"color":"#BA7517","bg":"#FAEEDA","desc":"Stable — optimise & retain"},
    "Sunset":          {"color":"#A32D2D","bg":"#FCEBEB","desc":"Being phased out"},
    "Planning":        {"color":"#534AB7","bg":"#EEEDFE","desc":"Under development"},
    "Development":     {"color":"#185FA5","bg":"#E6F1FB","desc":"Being built"},
    "Launch":          {"color":"var(--brand-hover,#0F6E56)","bg":"#E1F5EE","desc":"Recently launched"},
    "Optimising":      {"color":"#3B6D11","bg":"#EAF3DE","desc":"Active improvement cycle"},
    "Decommissioned":  {"color":"#5F5E5A","bg":"#F1EFE8","desc":"Retired"},
}

PRODUCT_HEALTH_SIGNALS = ["On track","Needs review","At risk","Suspended"]
PRODUCT_OWNERS         = []  # populated from staff registry at runtime



# ─── FILE UPLOAD CACHE HELPER ─────────────────────────────────────────
def cache_upload(uploaded_file, session_key: str) -> bytes | None:
    """
    Safely extract bytes from a Streamlit UploadedFile and persist them
    in session state so they survive reruns.
    Returns the cached bytes, or None if nothing has been uploaded yet.
    """
    if uploaded_file is not None:
        try:
            raw = uploaded_file.getvalue()
        except Exception:
            try:
                raw = uploaded_file.read()
            except Exception:
                raw = None
        if raw:
            st.session_state[session_key] = raw
            return raw
    return st.session_state.get(session_key)

# ─── ECOBANK BRAND THEME ────────────────────────────────────────────
BRAND = {
    'primary':      'var(--brand-primary,#006B3F)',   # brand primary (configurable)
    'secondary':    '#F5A623',   # brand secondary (configurable)
    'primary_light':'var(--brand-light,#E8F5EE)',   # light green surface
    'secondary_light':'#FEF6E4', # light gold surface
    'dark':         '#004A2B',   # deep dark green
    'text_on_primary': '#FFFFFF',
    'text_on_secondary': '#3D2600',
    'app_name':     'A2Z Blueprint',
    'tagline':      'Perform · Execute · Integrate',
}

# ─── REGION MAPPING ───────────────────────────────────────────────────
# ─── BRANCH_REGION (v10.361 — fully configurable, no hardcoded fallback) ──
# Per Rule N1: tenant identity must be configured, never hardcoded.
# v10.360 made BRANCH_REGION dynamically sourced from data/org_config.json
# but retained a 21-entry hardcoded "_BRANCH_REGION_FALLBACK" for degraded
# environments. v10.361 deletes that fallback — the system is bank-agnostic
# and must not carry any tenant-specific branch list, even as a fallback.
#
# If org_config.json is missing or malformed, BRANCH_REGION resolves to an
# empty dict. Downstream consumers will see "0 branches", which surfaces
# the configuration error in the admin module rather than masking it with
# a stale fallback list. Configuration is enforced upstream — not patched
# downstream.
#
# Production integration path: pages/7_admin.py (render_branch_manager)
# provides full CRUD (add + edit + soft-delete via active=False). Future
# FLEXCUBE integration: utils.flexcube_adapter.fetch_branches_from_flexcube
# pulls the branch list from core banking when mode="live"; org_config.json
# is the authoritative source until then.
#
# G246 + G247 lock the configurability: no hardcoded branch dict literals
# of >0 entries permitted in utils/.

def _build_branch_region_from_org_config() -> dict:
    """v10.361 — build BRANCH_REGION from data/org_config.json.

    Returns an empty dict if the config is unavailable or malformed.
    This is deliberate: a missing config is a configuration error that
    should surface, not be masked by a hardcoded fallback.
    """
    try:
        import json as _json
        from pathlib import Path as _Path
        _path = _Path(__file__).parent.parent / "data" / "org_config.json"
        if not _path.exists():
            return {}
        _cfg = _json.loads(_path.read_text(encoding="utf-8"))
        _branches = _cfg.get("branches", [])
        if not _branches:
            return {}
        return {
            b["name"]: b.get("region", "Other")
            for b in _branches
            if b.get("active", True) and b.get("name")
        }
    except Exception:
        # Empty dict surfaces configuration errors upstream.
        # No hardcoded tenant data — Rule N1.
        return {}


BRANCH_REGION: dict = _build_branch_region_from_org_config()
def _build_regions_from_org_config() -> list:
    """v10.361 — REGIONS derived from active branches in org_config.

    Returns sorted unique regions. Empty list if config unavailable —
    surfaces configuration errors instead of masking with stale defaults.
    """
    try:
        regions = sorted(set(BRANCH_REGION.values()))
        # Prefer non-"Other" regions where possible (better admin UX)
        non_other = [r for r in regions if r and r != "Other"]
        return non_other if non_other else regions
    except Exception:
        return []


REGIONS: list = _build_regions_from_org_config()
REGIONAL_HEAD_ROLE = 'Regional Head'

def get_unit_region(unit: str) -> str:
    return BRANCH_REGION.get(unit, 'Head Office')

# ─── BANKING METRICS HELPERS ─────────────────────────────────────────
def calc_nim(nii, avg_earning_assets):
    """Net Interest Margin = NII / Avg Earning Assets × 100"""
    return round(nii / avg_earning_assets * 100, 2) if avg_earning_assets else 0.0

def calc_roa(pbt, avg_total_assets):
    """Return on Assets = PBT / Avg Total Assets × 100"""
    return round(pbt / avg_total_assets * 100, 2) if avg_total_assets else 0.0

def calc_roe(pbt, avg_equity):
    """Return on Equity = PBT / Avg Equity × 100"""
    return round(pbt / avg_equity * 100, 2) if avg_equity else 0.0

def calc_cir(opex, operating_income):
    """Cost-to-Income Ratio = Opex / Operating Income × 100"""
    return round(abs(opex) / operating_income * 100, 1) if operating_income else 0.0

def calc_jaws(rev_growth_pct, cost_growth_pct):
    """Jaws ratio = Revenue growth % - Cost growth %. Positive = expanding margins."""
    return round(rev_growth_pct - cost_growth_pct, 1)

def calc_rev_per_staff(operating_income, staff_count):
    """Revenue per head — productivity metric"""
    return round(operating_income / staff_count, 0) if staff_count else 0.0

def calc_npl_coverage(provisions, npl_balance):
    """Provision coverage ratio = Cumulative provisions / NPL balance × 100"""
    return round(provisions / npl_balance * 100, 1) if npl_balance else 0.0

def calc_pipeline_win_rate(won_deals, total_closed):
    """Win rate = Won deals / (Won + Lost) × 100"""
    return round(won_deals / total_closed * 100, 1) if total_closed else 0.0

def calc_avg_deal_size(total_pipeline_value, deal_count):
    """Average deal size"""
    return round(total_pipeline_value / deal_count, 0) if deal_count else 0.0

# ─── ROLE FUNCTION CLASSIFIER ───────────────────────────────────────
# Business roles = revenue-generating / client-facing
# Support roles  = enablement / infrastructure
BUSINESS_ROLES: frozenset = frozenset([
    'Relationship Manager SME','Relationship Manager Corporate',
    'Relationship Officer Business Banking','Relationship Officer Personal Banking',
    'Direct Sales Officer','Head Of SME','Head Of Corporate','Head Of Retail',
    'Director Commercial Banking','Director Retail',
    'Head Of Digital Innovation','Digital Innovation Manager','Digital Innovation Officer',
    'Head Of Products','Products Manager','Products Officer',
    'Marketing Manager','Marketing Officer','Head Of Marketing',
])
def _get_branch_management_roles() -> frozenset:
    """Load branch management roles from org_config — never hardcoded."""
    try:
        _cfg = get_org_config()
        _hier = _cfg.get('hierarchy', {})
        _br_staff = set(_cfg.get('role_categories', {}).get('branch_staff', []))
        # Branch management = branch staff that have direct reports
        from collections import defaultdict
        _ch = defaultdict(list)
        for r, parents in _hier.items():
            for p in parents:
                _ch[p].append(r)
        _mgmt = {r for r in _br_staff if _ch.get(r)}
        return frozenset(_mgmt) if _mgmt else frozenset(['Branch Manager','Branch Operations Manager'])
    except:
        return frozenset(['Branch Manager','Branch Operations Manager'])
# Lazy: populated on first use, cleared when org_config changes
_BRANCH_MANAGEMENT_CACHE: frozenset = frozenset()

def _ensure_branch_management():
    global _BRANCH_MANAGEMENT_CACHE
    if not _BRANCH_MANAGEMENT_CACHE:
        _BRANCH_MANAGEMENT_CACHE = _get_branch_management_roles()
    return _BRANCH_MANAGEMENT_CACHE

# Keep the name for backward compatibility
class _BranchMgmtProxy:
    def __contains__(self, item):
        return item in _ensure_branch_management()
    def __iter__(self):
        return iter(_ensure_branch_management())
    def __len__(self):
        return len(_ensure_branch_management())

BRANCH_MANAGEMENT = _BranchMgmtProxy()

def get_role_function(role, category):
    """
    Returns 'Business', 'Support', or 'Branch Management'.
    Branch staff are all treated as Business (sales-focused).
    """
    if str(category).strip() == 'Branch':
        if role in BRANCH_MANAGEMENT:  return 'Branch Management'
        return 'Business'
    if role in BUSINESS_ROLES:         return 'Business'
    return 'Support'


# ─── EXECUTE / WORKSTREAM HELPERS ────────────────────────────────────

def get_workstreams_from_hierarchy() -> dict:
    """
    Return workstreams. Saved execute_workstreams.json is the primary source —
    it was seeded from the hierarchy and may have been customised via Admin.
    Falls back to rebuilding from org_config if the file is missing.
    """
    ws_file = DATA_DIR / "execute_workstreams.json"
    # Saved file is authoritative
    if ws_file.exists():
        try:
            saved = json.loads(ws_file.read_text())
            if saved and isinstance(saved, dict):
                return saved
        except: pass

    # Fallback: derive from org_config hierarchy
    try:
        cfg   = get_org_config()
        hier  = cfg.get('hierarchy', {})
        roots = [r for r, p in hier.items() if not p]
        if not roots: return {}
        ceo   = roots[0]
        chiefs= [r for r, p in hier.items() if ceo in p]
        result = {}
        for i, chief_role in enumerate(chiefs):
            ws_id = f"WS{i+1:02d}"
            short = (chief_role
                     .replace('Chief','').replace('Officer','')
                     .replace('& Managing Director','')
                     .replace('  ',' ').strip())
            result[ws_id] = {
                'name': short, 'full_role': chief_role,
                'sponsor_username': '', 'sponsor_name': chief_role,
                'sub_workstreams': [], 'cross_functional_pool': [],
            }
        return result
    except:
        return {}

def get_workstream_staff(workstream_id: str, include_cross_functional: bool = True) -> list:
    """
    Return all staff codes reachable from the workstream's chief role.
    Includes cross_functional_pool members if include_cross_functional=True.
    Returns list of {'code', 'name', 'role', 'unit'} dicts.
    """
    try:
        wss   = get_workstreams_from_hierarchy()
        ws    = wss.get(workstream_id, {})
        chief_role = ws.get('full_role', '')
        cfg   = get_org_config()
        hier  = cfg.get('hierarchy', {})

        # Build children map
        from collections import defaultdict
        children = defaultdict(list)
        for role, parents in hier.items():
            for p in parents:
                children[p].append(role)

        # BFS from chief_role to collect all descendant roles
        reachable_roles = set()
        queue = [chief_role] if chief_role else []
        while queue:
            role = queue.pop(0)
            reachable_roles.add(role)
            for child in children.get(role, []):
                if child not in reachable_roles:
                    queue.append(child)

        # Load staff register and filter
        sr_file = DATA_DIR / "staff_register.xlsx"
        if not sr_file.exists():
            return []
        import openpyxl
        wb = openpyxl.load_workbook(str(sr_file))
        ws_sheet = wb.active
        headers = [ws_sheet.cell(1,c).value for c in range(1, ws_sheet.max_column+1)]
        rc = headers.index('Role')      if 'Role'       in headers else 2
        nc = headers.index('Staff Name')if 'Staff Name' in headers else 1
        sc = headers.index('Staff Code')if 'Staff Code' in headers else 0
        uc = headers.index('Unit')      if 'Unit'       in headers else 3

        staff = []
        for row in ws_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            role = str(row[rc] or '')
            if role in reachable_roles:
                staff.append({
                    'code': str(row[sc] or ''),
                    'name': str(row[nc] or ''),
                    'role': role,
                    'unit': str(row[uc] or ''),
                    'source': 'hierarchy',
                })

        # Add cross-functional pool
        if include_cross_functional:
            pool = ws.get('cross_functional_pool', [])
            existing_codes = {s['code'] for s in staff}
            for uname_cf in pool:
                if uname_cf not in existing_codes:
                    # Look up from users.json
                    try:
                        users_data = json.loads((DATA_DIR/"users.json").read_text())
                        ud = users_data.get(uname_cf, {})
                        if ud:
                            staff.append({
                                'code': str(ud.get('staff_code', uname_cf)),
                                'name': ud.get('full_name', uname_cf),
                                'role': ud.get('role', ''),
                                'unit': ud.get('unit', ''),
                                'source': 'cross_functional',
                            })
                    except: pass

        return sorted(staff, key=lambda x: x['name'])
    except Exception as _gws_exc:
        import traceback as _tb
        _tb.print_exc()   # visible in Streamlit logs
        return []

# ─── PIPELINE CONSTANTS (Banking CRM) ────────────────────────────────
# ── Category-specific pipeline stage paths ────────────────────────────
# Each pipeline category follows its own progression
PIPELINE_STAGES_ACCOUNT = [
    {"stage":"Lead",            "icon":"🎯", "description":"Identified prospect — account not yet opened"},
    {"stage":"Contacted",       "icon":"📞", "description":"Initial engagement — relationship started"},
    {"stage":"KYC / Documentation","icon":"📋","description":"Customer documents collected"},
    {"stage":"Account Opening", "icon":"🏦", "description":"Account opening form submitted to operations"},
    {"stage":"Compliance Review","icon":"🔏","description":"AML / KYC / compliance screening"},
    {"stage":"Closed Won",      "icon":"🏆", "description":"Account opened and activated"},
    {"stage":"Closed Lost",     "icon":"❌", "description":"Customer did not open account"},
]
PIPELINE_STAGES_LOAN = [
    {"stage":"Lead",            "icon":"🎯", "description":"Identified credit need — no formal application"},
    {"stage":"Contacted",       "icon":"📞", "description":"Discussion of credit need and product options"},
    {"stage":"Qualified",       "icon":"✅", "description":"Basic eligibility confirmed — credit appetite assessed"},
    {"stage":"Application",     "icon":"📝", "description":"Loan application form submitted"},
    {"stage":"Credit Assessment","icon":"🔍","description":"Credit committee / risk assessment underway"},
    {"stage":"Offer / Proposal","icon":"📄", "description":"Term sheet / offer letter issued to customer"},
    {"stage":"Negotiation",     "icon":"🤝", "description":"Terms being negotiated — legal docs in progress"},
    {"stage":"Compliance",      "icon":"🔏", "description":"KYC / AML / security perfection underway"},
    {"stage":"Closed Won",      "icon":"🏆", "description":"Facility signed and disbursed"},
    {"stage":"Closed Lost",     "icon":"❌", "description":"Customer declined or credit declined"},
]
PIPELINE_STAGES_DEPOSIT = [
    {"stage":"Lead",            "icon":"🎯", "description":"Identified deposit opportunity — no commitment"},
    {"stage":"Contacted",       "icon":"📞", "description":"Initial discussion — rates and terms presented"},
    {"stage":"Proposal",        "icon":"📄", "description":"Formal deposit proposal / term sheet presented"},
    {"stage":"Negotiation",     "icon":"🤝", "description":"Rate / tenor negotiation underway"},
    {"stage":"Documentation",   "icon":"📋", "description":"Deposit agreement and KYC documentation"},
    {"stage":"Closed Won",      "icon":"🏆", "description":"Deposit placed / funds received"},
    {"stage":"Closed Lost",     "icon":"❌", "description":"Customer placed funds elsewhere"},
]
# Generic fallback (other products — insurance, digital, treasury)
PIPELINE_STAGES_GENERIC = [
    {"stage":"Lead",            "icon":"🎯", "description":"Identified opportunity"},
    {"stage":"Contacted",       "icon":"📞", "description":"Initial engagement"},
    {"stage":"Qualified",       "icon":"✅", "description":"Need confirmed and qualified"},
    {"stage":"Proposal",        "icon":"📄", "description":"Offer / proposal presented"},
    {"stage":"Negotiation",     "icon":"🤝", "description":"Terms being agreed"},
    {"stage":"Closed Won",      "icon":"🏆", "description":"Deal concluded"},
    {"stage":"Closed Lost",     "icon":"❌", "description":"Opportunity lost"},
]

# Master list — all unique stages across all paths (for filtering / reporting)
PIPELINE_STAGES = PIPELINE_STAGES_LOAN   # default for backward compat
STAGE_NAMES    = [s["stage"] for s in PIPELINE_STAGES]
ACTIVE_STAGES  = [s["stage"] for s in PIPELINE_STAGES if s["stage"] not in ("Closed Won","Closed Lost")]

# All unique stage names across all categories
ALL_STAGE_NAMES = list(dict.fromkeys(
    s["stage"] for stages in [
        PIPELINE_STAGES_ACCOUNT, PIPELINE_STAGES_LOAN,
        PIPELINE_STAGES_DEPOSIT, PIPELINE_STAGES_GENERIC]
    for s in stages))
ALL_ACTIVE_STAGES = [s for s in ALL_STAGE_NAMES if s not in ("Closed Won","Closed Lost")]

def get_pipeline_category(product_type: str) -> str:
    """Map a product type to its pipeline category."""
    _ACCT_PRODS = {
        "Current Account (CASA)","Savings Account (CASA)","Fixed Deposit",
        "Call Deposit","Notice Deposit","Junior Account",
        "Business Current Account","Business Savings","Other Deposit",
    }
    _LOAN_PRODS = {
        "Business Loan","Personal Loan","Mortgage / Home Loan","Overdraft",
        "Trade Finance","Asset Finance","Invoice Discounting","LPO Finance",
        "Agricultural Loan","Staff Loan","Credit Card","Other Loan",
    }
    if product_type in _ACCT_PRODS:  return "Account"
    if product_type in _LOAN_PRODS:  return "Loan"
    if "deposit" in product_type.lower() or "casa" in product_type.lower(): return "Account"
    if "loan" in product_type.lower() or "credit" in product_type.lower() or        "overdraft" in product_type.lower() or "finance" in product_type.lower(): return "Loan"
    if "deposit" in product_type.lower(): return "Deposit"
    return "Other"

def get_stages_for_category(category: str) -> list:
    """Return stage list for a pipeline category."""
    return {
        "Account": PIPELINE_STAGES_ACCOUNT,
        "Loan":    PIPELINE_STAGES_LOAN,
        "Deposit": PIPELINE_STAGES_DEPOSIT,
    }.get(category, PIPELINE_STAGES_GENERIC)

# ── Contact person standard positions ─────────────────────────────────
CONTACT_POSITIONS = [
    "— Select position —",
    # C-Suite
    "Managing Director (MD) / CEO",
    "Chief Finance Officer (CFO)",
    "Chief Operations Officer (COO)",
    "Chief Commercial Officer (CCO)",
    "Chief Executive Officer (CEO)",
    # Board / Directors
    "Chairman / Board Member",
    "Director",
    # Senior Management
    "Finance Director / VP Finance",
    "Head of Finance",
    "Head of Treasury",
    "Head of Procurement",
    "General Manager",
    # Management
    "Finance Manager",
    "Treasury Manager",
    "Procurement Manager",
    "Operations Manager",
    "HR Manager",
    "Branch Manager",
    # Key Influencers
    "Finance Officer / Accountant",
    "Procurement Officer",
    "Business Development Manager",
    "Sales Manager",
    "Owner / Proprietor",
    "Partner",
    # Other
    "Other (specify below)",
]
DECISION_LEVELS = [
    "— Select —",
    "Ultimate decision maker — signs off",
    "Key influencer — recommends to board",
    "Evaluator — reviews options",
    "Gatekeeper — controls access",
    "End user — no signing authority",
]

# ─── REVENUE INTELLIGENCE — KPI CATEGORY MAPPING ────────────────────
RI_CATEGORIES = {
    'Deposits': {
        'kpis':      ['Deposit Growth'],
        'color':     'var(--brand-hover,#0F6E56)',   # teal-600
        'bg':        '#E1F5EE',   # teal-50
        'label':     'Liabilities',
        'unit':      'KES',
        'pipeline_product_types': ['Fixed Deposit','Savings Account','Current Account',
                                    'Salary Processing','SACCO Placement','Corporate Deposit','Other Deposit'],
        'stages':    ['Prospect','Engaged','Proposal','Negotiation','Approval','Funds In'],
        'closed_won':'Funds In',
        'closed_lost':'Lost',
        'stage_weights': {'Prospect':0.05,'Engaged':0.15,'Proposal':0.35,
                          'Negotiation':0.60,'Approval':0.85,'Funds In':1.0,'Lost':0.0},
    },
    'Loans': {
        'kpis':      ['Loans Disbursement','Loan Book Growth','Trade Finance'],
        'color':     '#185FA5',   # blue-600
        'bg':        '#E6F1FB',   # blue-50
        'label':     'Assets',
        'unit':      'KES',
        'pipeline_product_types': ['Business Loan','Personal Loan','Mortgage','Overdraft',
                                    'Asset Finance','Invoice Discounting','Trade Finance','Other Loan'],
        'stages':    ['Prospect','Application','Credit Appraisal','Approval','Documentation','Disbursed'],
        'closed_won':'Disbursed',
        'closed_lost':'Declined',
        'stage_weights': {'Prospect':0.05,'Application':0.20,'Credit Appraisal':0.45,
                          'Approval':0.75,'Documentation':0.90,'Disbursed':1.0,'Declined':0.0},
    },
    'NFI': {
        'kpis':      ['Fees and Commission','Bancassurance','DFS Revenue','Treasury'],
        'color':     '#534AB7',   # purple-600
        'bg':        '#EEEDFE',   # purple-50
        'label':     'Non-Funded Income',
        'unit':      'KES',
        'pipeline_product_types': ['Bancassurance Policy','DFS Onboarding','Treasury Deal',
                                    'Trade Finance Fee','Forex','Account Fee','Other NFI'],
        'stages':    ['Prospect','Pitched','Proposal','Committed','Converted'],
        'closed_won':'Converted',
        'closed_lost':'Lost',
        'stage_weights': {'Prospect':0.10,'Pitched':0.25,'Proposal':0.50,
                          'Committed':0.80,'Converted':1.0,'Lost':0.0},
    },
    'Customers': {
        'kpis':      ['Customer Growth'],
        'color':     '#BA7517',   # amber-600
        'bg':        '#FAEEDA',   # amber-50
        'label':     'New Customers',
        'unit':      'No.',
        'pipeline_product_types': ['Individual','SME','Corporate','Institutional','Other'],
        'stages':    ['Lead','Contacted','KYC Submitted','Account Opened'],
        'closed_won':'Account Opened',
        'closed_lost':'Lost',
        'stage_weights': {'Lead':0.10,'Contacted':0.30,'KYC Submitted':0.75,
                          'Account Opened':1.0,'Lost':0.0},
    },
}

RI_KPI_TO_CATEGORY = {}
for cat, cfg in RI_CATEGORIES.items():
    for kpi in cfg['kpis']:
        RI_KPI_TO_CATEGORY[kpi] = cat

# Months remaining in year (for forecast)
def months_remaining():
    return max(1, 12 - datetime.now().month)

def months_elapsed():
    return max(1, datetime.now().month)


# ═══════════════════════════════════════════════════════════════════════
# EXECUTE MODULE — STRATEGY EXECUTION TRACKING
# ═══════════════════════════════════════════════════════════════════════

EXECUTE_GATES = {
    'G0': {'label': 'Idea',           'color': '#888780', 'bg': '#F1EFE8',
           'desc': 'Initiative created — not yet submitted'},
    'G1': {'label': 'Validated',      'color': '#185FA5', 'bg': '#E6F1FB',
           'desc': 'Initiative approved as submitted'},
    'G2': {'label': 'Business case',  'color': 'var(--brand-hover,#0F6E56)', 'bg': '#E1F5EE',
           'desc': 'Business case approved, prioritised'},
    'G3': {'label': 'Implementation', 'color': '#BA7517', 'bg': '#FAEEDA',
           'desc': 'Milestone plan locked, executing'},
    'G4': {'label': 'Impact tracking','color': '#993C1D', 'bg': '#FAECE7',
           'desc': 'Money step reached, tracking impact'},
    'G5': {'label': 'Embedded',       'color': '#3B6D11', 'bg': '#EAF3DE',
           'desc': 'Complete — impact sustained'},
    'On Hold':  {'label': 'On hold',  'color': '#5F5E5A', 'bg': '#F1EFE8', 'desc': 'Paused'},
    'Dropped':  {'label': 'Dropped',  'color': '#A32D2D', 'bg': '#FCEBEB', 'desc': 'Discontinued'},
}

GATE_ORDER = ['G0','G1','G2','G3','G4','G5']

INITIATIVE_CATEGORIES = [
    'Impact Generation',   # Revenue / PBT focused
    'Cost Optimisation',   # Cost reduction
    'Enabler',             # Infrastructure / capability
    'Customer Experience', # NPS / retention
    'Compliance & Risk',   # Regulatory
    'Digital & Innovation',
    'People & Culture',
    'Other',
]

MILESTONE_TYPES = ['Implementation', 'Health Check', 'Money Step']

# ── ESCALATION DISPLAY HELPERS ──────────────────────────────────────
# Banking-grade escalation — tight timelines, immediate visibility
ESC_CONFIG = {
    0: {'label': 'On track',          'color': '#3B6D11', 'bg': '#EAF3DE', 'icon': '✅'},
    1: {'label': 'Due soon — alert',  'color': '#BA7517', 'bg': '#FAEEDA', 'icon': '🟡'},
    2: {'label': 'Overdue — IO',      'color': '#993C1D', 'bg': '#FAECE7', 'icon': '🔴'},
    3: {'label': 'Escalated — Lead',  'color': '#A32D2D', 'bg': '#FCEBEB', 'icon': '🚨'},
    4: {'label': 'CRITICAL — Sponsor','color': '#791F1F', 'bg': '#F7C1C1', 'icon': '💥'},
    5: {'label': 'NOT STARTED',       'color': '#5F5E5A', 'bg': '#F1EFE8', 'icon': '⏰'},
}
# Banking escalation tiers:
#   L0  On track        — start date future, due date > 2 days away
#   L1  Due soon        — due within 2 days (email kicks in here)
#   L2  Overdue (day 1) — 1–2 days overdue → IO notified immediately
#   L3  Escalated       — 3–7 days overdue OR blocker raised → Lead notified
#   L4  Critical        — >7 days overdue OR Structural delay → Sponsor immediately
#   L5  Not started     — start date passed, status still "Not Started"

# Delay categories — determines who gets notified immediately
DELAY_CATEGORIES = [
    'Structural',          # Org/resource issue — immediate sponsor alert regardless of days
    'Dependency blocked',  # Waiting on another team/milestone
    'Regulatory / compliance', # External constraint
    'Resource constraint', # Capacity / budget issue
    'Technical challenge', # System / data issue
    'Stakeholder alignment', # Sign-off pending
    'Data / information gap',
    'Other / in progress', # Normal delay — follow standard escalation
]

STRUCTURAL_CATEGORIES = frozenset({'Structural', 'Regulatory / compliance'})
# These trigger immediate Level 4 (Sponsor) escalation regardless of days overdue

def escalation_badge(level):
    cfg = ESC_CONFIG.get(level, ESC_CONFIG[0])
    return (f"<span style='background:{cfg['bg']};color:{cfg['color']};"
            f"padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500'>"
            f"{cfg['icon']} {cfg['label']}</span>")

def days_label(days):
    if days < -7:   return f"<span style='color:#791F1F;font-weight:500'>{-days}d overdue</span>"
    elif days < 0:  return f"<span style='color:#A32D2D;font-weight:500'>{-days}d overdue</span>"
    elif days == 0: return "<span style='color:#A32D2D;font-weight:500'>Due today</span>"
    elif days <= 2: return f"<span style='color:#BA7517;font-weight:500'>Due in {days}d ⚡</span>"
    else:           return f"<span style='color:var(--color-text-secondary)'>Due in {days}d</span>"

def start_label(days_to_start):
    """Label for milestone start date status."""
    if days_to_start < -1:  return f"<span style='color:#791F1F;font-weight:500'>Should have started {-days_to_start}d ago</span>"
    elif days_to_start < 0: return "<span style='color:#A32D2D;font-weight:500'>Should have started yesterday</span>"
    elif days_to_start == 0:return "<span style='color:#BA7517;font-weight:500'>Starts today</span>"
    elif days_to_start <= 3:return f"<span style='color:#BA7517;font-weight:500'>Starts in {days_to_start}d</span>"
    else:                    return f"<span style='color:var(--color-text-secondary)'>Starts in {days_to_start}d</span>"

# Approvers required per gate transition
GATE_APPROVERS = {
    'G0→G1': ['workstream_lead', 'sponsor'],
    'G1→G2': ['workstream_lead', 'sponsor', 'finance'],
    'G2→G3': ['workstream_lead', 'sponsor', 'finance'],   # after milestone owner confirmations
    'G3→G4': ['workstream_lead', 'sponsor', 'finance'],
    'G4→G5': ['workstream_lead', 'sponsor', 'finance'],
}

# Impact KPI options for business case
IMPACT_KPI_OPTIONS = [
    'PBT', 'Revenue', 'Cost Savings', 'Deposit Growth', 'Loan Book Growth',
    'Customer Growth', 'NPS / Cx Score', 'Fees and Commission', 'DFS Revenue',
    'NPL Reduction', 'Staff Productivity', 'Process Efficiency', 'Other',
]

# ── Pipeline product catalogue (grouped) ──────────────────────────────
PRODUCT_CATALOGUE = {
    "Loans & Credit": [
        "Business Loan","Personal Loan","Mortgage / Home Loan","Overdraft",
        "Trade Finance","Asset Finance","Invoice Discounting","LPO Finance",
        "Agricultural Loan","Staff Loan","Credit Card","Other Loan",
    ],
    "Deposits & CASA": [
        "Current Account (CASA)","Savings Account (CASA)",
        "Salary Account","Fixed Deposit","Call Deposit","Notice Deposit",
        "Junior Account","Business Current Account","Business Savings","Other Deposit",
    ],
    "Insurance & Bancassurance": [
        "Life Insurance","Credit Life","General Insurance",
        "Bancassurance — Medical","Bancassurance — Motor",
        "Pension / Investment","Other Insurance",
    ],
    "Digital & Transactional": [
        "Mobile Banking","Internet Banking","DFS Onboarding",
        "Agency Banking","Point of Sale (POS)","International Transfer",
        "Other Digital",
    ],
    "Treasury & Investments": [
        "Treasury Bill","Treasury Bond","Foreign Exchange","Money Market",
        "Other Treasury",
    ],
    "Other Facilities": [
        "Bank Guarantee","Letter of Credit","Bid Bond","Performance Bond",
        "Other",
    ],
}
# Flat list for dropdowns — with category labels as separators (prefixed ---)
PRODUCT_TYPES = []
for _cat, _prods in PRODUCT_CATALOGUE.items():
    PRODUCT_TYPES.extend(_prods)

# ── Customer segmentation ─────────────────────────────────────────────
CUSTOMER_SEGMENTS = {
    "Individual": ["Affluent","Core Middle","Mass / Retail"],
    "Business":   ["Large Corporate","Corporate","SME","Micro Enterprise"],
}

# ── CBK Economic sectors ──────────────────────────────────────────────
# ── Individual customer profile (non-CBK) ─────────────────────────────
INDIVIDUAL_SECTORS = {
    "Employment / Profession": [
        "Salaried — Civil Servant / Government",
        "Salaried — Private Sector",
        "Salaried — NGO / International Organisation",
        "Self-Employed Professional (Doctor, Lawyer, Engineer, Accountant)",
        "Teacher / Lecturer",
        "Military / Police / Uniformed Services",
        "Diplomat / Foreign National",
    ],
    "Business Owner": [
        "Sole Trader / Hawker",
        "Small Business Owner",
        "Medium Business Owner",
        "Farmer / Agri-preneur",
        "Landlord / Property Owner",
    ],
    "Social / Life Stage": [
        "Student",
        "Retired",
        "Housewife / Homemaker",
        "Unemployed / Job Seeker",
        "Diaspora / Returning Resident",
    ],
}
# Flat list for dropdown
INDIVIDUAL_SECTOR_LIST = [
    f"{cat} — {item}"
    for cat, items in INDIVIDUAL_SECTORS.items()
    for item in items
]

CBK_SECTORS = [
    "Agriculture, Forestry & Fishing",
    "Mining & Quarrying",
    "Manufacturing",
    "Electricity, Gas & Water Supply",
    "Building & Construction",
    "Trade (Wholesale & Retail)",
    "Tourism, Restaurant & Hotels",
    "Transport & Communication",
    "Real Estate & Business Services",
    "Financial Services",
    "Community, Social & Personal Services",
    "Government & Public Sector",
    "Non-Profit / NGO",
    "Other / Not Classified",
]

# ── Kenya commercial banks (CBK licensed) ─────────────────────────────
KENYA_BANKS = [
    "Absa Bank Kenya","Access Bank Kenya","African Banking Corporation (ABC)",
    "Bank of Africa Kenya","Bank of Baroda Kenya","Bank of India",
    "Citibank NA Kenya","Co-operative Bank of Kenya","Consolidated Bank",
    "Credit Bank","DTB — Diamond Trust Bank","Ecobank Kenya",
    "Equity Bank Kenya","Family Bank","First Community Bank",
    "Guaranty Trust Bank (GTB)","Gulf African Bank","HFC Bank",
    "Housing Finance (HFC)","I&M Bank Kenya","KCB Group",
    "Kingdom Bank","Mayfair CIB Bank","NCBA Bank Kenya",
    "National Bank of Kenya","Paramount Bank","Premier Bank Kenya",
    "Prime Bank Kenya","SBM Bank Kenya","Sidian Bank",
    "Spire Bank","Stanbic Bank Kenya","Standard Chartered Bank Kenya",
    "UBA Kenya Bank","Victoria Commercial Bank",
]

ACTIVITY_TYPES = [
    "Cold Call","Discovery Meeting","Follow-up Call","Product Presentation",
    "Site Visit","Credit Committee","Proposal Submitted","Contract Signing",
    "Account Opening","Referral","Email","WhatsApp / Messenger","Other",
]
LOSS_REASONS   = [
    "Pricing too high","Competitor offer (lower rate)",
    "Competitor offer (better terms)","Credit declined by bank",
    "Customer withdrew / no longer interested",
    "Documentation / KYC issues","Relationship breakdown",
    "Regulatory / compliance issue","Deal superseded by another product",
    "Other",
]

# Pipeline delete/cancel approval levels (by stage reached)
# Stages up to this index can be self-deleted; beyond requires manager approval
PIPELINE_DELETE_SELF_MAX = "Lead"     # can self-delete only at Lead stage
PIPELINE_VALIDATE_STAGE  = "Lead"  # validate at creation: a newly created deal
                                   # (Lead) must be confirmed real by the line
                                   # manager before it counts — first defense
                                   # against ghost deals inflating the pipeline

# ── KPI Library — persistent configuration ───────────────────────────
KPI_LIBRARY_FILE = DATA_DIR / "kpi_library.json"

# Default KPI library — full banking KPI catalogue
DEFAULT_KPI_LIBRARY = {
    "Financial": [
        {"id":"DEP_GROWTH",    "name":"Deposit Growth",             "unit":"KES",   "direction":"higher","cbs_source":"accounts.deposit_bal",      "fixed":False,"default_weight":0.15,"description":"Growth in customer deposit balances vs baseline"},
        {"id":"LOAN_GROWTH",   "name":"Loan Book Growth",           "unit":"KES",   "direction":"higher","cbs_source":"accounts.loan_outstanding",   "fixed":False,"default_weight":0.15,"description":"Growth in net loans and advances vs baseline"},
        {"id":"LOAN_DISB",     "name":"Loans Disbursement",         "unit":"KES",   "direction":"higher","cbs_source":"accounts.loan_new",           "fixed":False,"default_weight":0.10,"description":"Value of new loans disbursed in period"},
        {"id":"FEES_COMM",     "name":"Fees and Commission",        "unit":"KES",   "direction":"higher","cbs_source":"accounts.fee_income",         "fixed":False,"default_weight":0.10,"description":"Non-interest income from fees and commissions"},
        {"id":"DFS_REV",       "name":"DFS Revenue",                "unit":"KES",   "direction":"higher","cbs_source":"transactions.mobile_volume",  "fixed":False,"default_weight":0.05,"description":"Revenue from digital financial services"},
        {"id":"BANCASSURANCE", "name":"Bancassurance",              "unit":"KES",   "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.05,"description":"Premium income from insurance products sold"},
        {"id":"PBT",           "name":"PBT",                        "unit":"KES",   "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.05,"description":"Profit Before Tax contribution"},
        {"id":"NPL_RATIO",     "name":"NPL Ratio",                  "unit":"%",     "direction":"lower", "cbs_source":"accounts.npl_ratio",         "fixed":False,"default_weight":0.10,"description":"Non-performing loans as % of total loan book"},
        {"id":"PAR",           "name":"PAR",                        "unit":"%",     "direction":"lower", "cbs_source":"accounts.par_ratio",         "fixed":False,"default_weight":0.05,"description":"Portfolio at risk — loans overdue >30 days"},
        {"id":"CIR",           "name":"Cost-to-Income Ratio",       "unit":"%",     "direction":"lower", "cbs_source":"manual",                     "fixed":True, "default_weight":0.05,"description":"Operating costs / operating income (bank-wide)"},
        {"id":"NIM",           "name":"Net Interest Margin",        "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":True, "default_weight":0.05,"description":"Net interest income as % of earning assets (HO only)"},
        {"id":"ROE",           "name":"Return on Equity",           "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":True, "default_weight":0.05,"description":"Net profit as % of shareholders equity (HO only)"},
        {"id":"TRADE_FIN",     "name":"Trade Finance",              "unit":"KES",   "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.05,"description":"Revenue from trade finance products"},
        {"id":"TREASURY",      "name":"Treasury Revenue",           "unit":"KES",   "direction":"higher","cbs_source":"manual",                     "fixed":True, "default_weight":0.05,"description":"Net revenue from treasury operations"},
    ],
    "Customer Focus": [
        {"id":"NEW_CUST",      "name":"New Customer Acquisition",   "unit":"Count", "direction":"higher","cbs_source":"customers.new_onboarded",    "fixed":False,"default_weight":0.20,"description":"Net-new customers onboarded in period"},
        {"id":"ACTIVE_ACCTS",  "name":"Active Account Growth",      "unit":"Count", "direction":"higher","cbs_source":"accounts.active_count",      "fixed":False,"default_weight":0.20,"description":"Net growth in active accounts (opens minus dormancies)"},
        {"id":"DORMANCY_REACT","name":"Dormancy Reactivation",      "unit":"Count", "direction":"higher","cbs_source":"accounts.reactivated",       "fixed":False,"default_weight":0.15,"description":"Dormant accounts reactivated in period"},
        {"id":"DIGITAL_ACT",   "name":"Digital Active Customers",   "unit":"Count", "direction":"higher","cbs_source":"transactions.digital_cifs",  "fixed":False,"default_weight":0.15,"description":"Customers transacting digitally ≥1 time per month"},
        {"id":"NPS",           "name":"NPS Score",                  "unit":"Score", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.15,"description":"Net Promoter Score — customer advocacy measure"},
        {"id":"CX_SCORE",      "name":"CX Score",                   "unit":"Score", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"Customer experience composite score"},
        {"id":"COMPLAINT_RES", "name":"Complaint Resolution Rate",  "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"% complaints resolved within SLA"},
        {"id":"CUST_RETENTION","name":"Customer Retention Rate",    "unit":"%",     "direction":"higher","cbs_source":"customers.retention",        "fixed":False,"default_weight":0.10,"description":"% of customers retained in period"},
    ],
    "Operational Excellence": [
        {"id":"TRANSACTIONS",  "name":"Transactions Volume",        "unit":"Count", "direction":"higher","cbs_source":"transactions.count",         "fixed":False,"default_weight":0.15,"description":"Total customer transactions across all channels"},
        {"id":"COMPLIANCE",    "name":"Compliance Score",           "unit":"%",     "direction":"higher","cbs_source":"customers.kyc_verified_pct", "fixed":False,"default_weight":0.15,"description":"KYC verification rate across portfolio"},
        {"id":"AUDIT_SCORE",   "name":"Audit Score",                "unit":"Score", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.15,"description":"Internal audit score for branch/unit"},
        {"id":"SLA_SCORE",     "name":"SLA Adherence Score",        "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"% of service requests fulfilled within SLA"},
        {"id":"REG_TIMELINESS","name":"Regulatory Reporting Timeliness","unit":"%", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"% of regulatory submissions filed on time"},
        {"id":"DIGITAL_MIG",   "name":"Digital Transaction Migration","unit":"%",   "direction":"higher","cbs_source":"transactions.digital_pct",   "fixed":False,"default_weight":0.10,"description":"% transactions on digital vs branch channels"},
        {"id":"CREDIT_TAT",    "name":"Credit Approval TAT",        "unit":"Days",  "direction":"lower", "cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"Avg days from loan application to decision"},
        {"id":"RECONCILIATION","name":"Timely Reconciliations",     "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.10,"description":"% of reconciliations completed within deadline"},
        {"id":"AML_SAR",       "name":"AML SAR Filing Rate",        "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.05,"description":"% of SARs filed within regulatory deadline"},
    ],
    "People & Learning": [
        {"id":"DILIGENCE",     "name":"Diligence Score",            "unit":"Score", "direction":"higher","cbs_source":"system.diligence",           "fixed":False,"default_weight":0.30,"description":"Composite score: attendance, reporting, deadlines"},
        {"id":"TRAINING",      "name":"Training Completion Rate",   "unit":"%",     "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.25,"description":"% of mandatory training completed"},
        {"id":"STAFF_SAT",     "name":"Staff Satisfaction Index",   "unit":"Score", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.25,"description":"Employee engagement score from periodic survey"},
        {"id":"SUCCESSION",    "name":"Succession Readiness",       "unit":"Score", "direction":"higher","cbs_source":"manual",                     "fixed":False,"default_weight":0.20,"description":"% of critical roles with identified successors"},
    ],
}

# Default role → KPI assignments
DEFAULT_ROLE_KPIS = {
    "Managing Director":                ["DEP_GROWTH","LOAN_GROWTH","FEES_COMM","PBT","NPL_RATIO","CIR","NIM","ROE","NEW_CUST","DIGITAL_ACT","NPS","DILIGENCE"],
    "Director Retail Banking":          ["DEP_GROWTH","LOAN_GROWTH","FEES_COMM","NPL_RATIO","NEW_CUST","DIGITAL_ACT","TRANSACTIONS","COMPLIANCE","DILIGENCE"],
    "Director Commercial Banking":      ["LOAN_GROWTH","LOAN_DISB","FEES_COMM","TRADE_FIN","NPL_RATIO","PAR","NEW_CUST","DILIGENCE"],
    "Regional Head":                    ["DEP_GROWTH","LOAN_GROWTH","FEES_COMM","NPL_RATIO","NEW_CUST","ACTIVE_ACCTS","TRANSACTIONS","COMPLIANCE","DILIGENCE"],
    "Branch Manager":                   ["DEP_GROWTH","LOAN_GROWTH","FEES_COMM","DFS_REV","NPL_RATIO","NEW_CUST","ACTIVE_ACCTS","TRANSACTIONS","COMPLIANCE","DORMANCY_REACT","DILIGENCE"],
    "Branch Credit Manager":            ["LOAN_GROWTH","LOAN_DISB","FEES_COMM","NPL_RATIO","PAR","NEW_CUST","COMPLIANCE","DILIGENCE"],
    "Branch Operations Manager":        ["TRANSACTIONS","ACTIVE_ACCTS","DORMANCY_REACT","COMPLIANCE","DIGITAL_MIG","SLA_SCORE","DILIGENCE"],
    "Branch Operations Supervisor":     ["TRANSACTIONS","DORMANCY_REACT","COMPLIANCE","DILIGENCE"],
    "Relationship Officer Personal Banking": ["DEP_GROWTH","LOAN_GROWTH","FEES_COMM","NEW_CUST","ACTIVE_ACCTS","COMPLIANCE","DILIGENCE"],
    "Relationship Officer Business Banking": ["LOAN_GROWTH","LOAN_DISB","DEP_GROWTH","FEES_COMM","NPL_RATIO","NEW_CUST","COMPLIANCE","DILIGENCE"],
    "Direct Sales Officer":             ["NEW_CUST","DEP_GROWTH","LOAN_GROWTH","ACTIVE_ACCTS","DIGITAL_ACT","DILIGENCE"],
    "Customer Service Officer":         ["TRANSACTIONS","NEW_CUST","DORMANCY_REACT","COMPLIANCE","DILIGENCE"],
    "Teller":                           ["TRANSACTIONS","COMPLIANCE","DILIGENCE"],
    "Relationship Manager Corporate":   ["LOAN_GROWTH","LOAN_DISB","FEES_COMM","TRADE_FIN","DEP_GROWTH","NPL_RATIO","NEW_CUST","DILIGENCE"],
    "Relationship Manager SME":         ["LOAN_GROWTH","LOAN_DISB","DEP_GROWTH","FEES_COMM","NPL_RATIO","NEW_CUST","COMPLIANCE","DILIGENCE"],
}

CBS_SOURCE_LABELS = {
    "accounts.deposit_bal":      "CBS — Sum of deposit balances in portfolio",
    "accounts.loan_outstanding": "CBS — Sum of loan outstanding in portfolio",
    "accounts.loan_new":         "CBS — New loans disbursed in period",
    "accounts.fee_income":       "CBS — Fee income from portfolio accounts",
    "accounts.npl_ratio":        "CBS — NPL / total loans in portfolio",
    "accounts.par_ratio":        "CBS — Loans overdue >30d / total portfolio",
    "accounts.active_count":     "CBS — Count of active accounts in portfolio",
    "accounts.reactivated":      "CBS — Dormant accounts with recent activity",
    "customers.new_onboarded":   "CBS — New CIFs opened in period",
    "customers.kyc_verified_pct":"CBS — % of portfolio with verified KYC",
    "customers.retention":       "CBS — Customers retained vs prior period",
    "transactions.count":        "CBS — Transaction count at branch",
    "transactions.mobile_volume":"CBS — Mobile banking transaction volume",
    "transactions.digital_cifs": "CBS — Distinct CIFs transacting digitally",
    "transactions.digital_pct":  "CBS — % transactions on digital channels",
    "system.diligence":          "System — Auto-computed from attendance & deadlines",
    "manual":                    "Manual — Entered by HR / admin / manager",
}

_KPI_LIB_CACHE: dict = {}
_KPI_LIB_MTIME: float = 0.0

def get_kpi_library() -> dict:
    """Load the bank's active KPI library from disk. Returns default if not configured."""
    global _KPI_LIB_CACHE, _KPI_LIB_MTIME
    try:
        _f = KPI_LIBRARY_FILE
        if _f.exists():
            _mt = _f.stat().st_mtime
            if _mt == _KPI_LIB_MTIME and _KPI_LIB_CACHE:
                return _KPI_LIB_CACHE
    except: pass
    if not KPI_LIBRARY_FILE.exists():
        return {"pillars": DEFAULT_KPI_LIBRARY, "role_kpis": DEFAULT_ROLE_KPIS, "active_kpis": []}
    try:
        return json.loads(KPI_LIBRARY_FILE.read_text())
    except:
        return {"pillars": DEFAULT_KPI_LIBRARY, "role_kpis": DEFAULT_ROLE_KPIS, "active_kpis": []}

def save_kpi_library(library: dict):
    global _KPI_LIB_CACHE, _KPI_LIB_MTIME
    _KPI_LIB_CACHE = {}; _KPI_LIB_MTIME = 0.0
    """Persist the KPI library configuration to disk."""
    KPI_LIBRARY_FILE.write_text(json.dumps(library, indent=2))

def get_active_kpis() -> list:
    """Return only the KPIs the bank has activated."""
    lib = get_kpi_library()
    active_ids = set(lib.get("active_kpis", []))
    result = []
    for pillar, kpis in lib.get("pillars", DEFAULT_KPI_LIBRARY).items():
        for k in kpis:
            if not active_ids or k["id"] in active_ids:
                result.append({**k, "pillar": pillar})
    return result

def get_role_kpis(role: str) -> list:
    """Return KPI IDs assigned to a role."""
    lib = get_kpi_library()
    return lib.get("role_kpis", DEFAULT_ROLE_KPIS).get(role, [])

def get_pillar_weights() -> dict:
    """Return pillar-level weights {pillar_name: weight}."""
    lib = get_kpi_library()
    return lib.get("pillar_weights", {
        "Financial": 0.40,
        "Customer Focus": 0.25,
        "Operational Excellence": 0.25,
        "People & Learning": 0.10,
    })


# ── Organisation Configuration — admin-configurable ──────────────────
ORG_CONFIG_FILE = DATA_DIR / "org_config.json"

DEFAULT_ORG_CONFIG = {
    "bank_name":    "A2Z Blueprint",
    "bank_code":    "ECO",
    "country":      "Kenya",
    "currency":     "KES",
    "currency_symbol": "KES",
    "logo_url":     "",
    "branches": [
        {"code":"BRN001","name":"Head Office",        "region":"Head Office","county":"Nairobi",   "type":"HO",       "tier":1},
        {"code":"BRN002","name":"Upper Hill Branch",  "region":"Nairobi",   "county":"Nairobi",   "type":"Flagship", "tier":1},
        {"code":"BRN003","name":"Westlands Branch",   "region":"Nairobi",   "county":"Nairobi",   "type":"Flagship", "tier":1},
        {"code":"BRN004","name":"Sarit Centre Branch","region":"Nairobi",   "county":"Nairobi",   "type":"Main",     "tier":2},
        {"code":"BRN005","name":"Industrial Area Branch","region":"Nairobi","county":"Nairobi",   "type":"Main",     "tier":2},
        {"code":"BRN006","name":"Karen Branch",       "region":"Nairobi",   "county":"Nairobi",   "type":"Standard", "tier":3},
        {"code":"BRN007","name":"Eastleigh Branch",   "region":"Nairobi",   "county":"Nairobi",   "type":"Standard", "tier":3},
        {"code":"BRN008","name":"Gigiri Branch",      "region":"Nairobi",   "county":"Nairobi",   "type":"Standard", "tier":3},
        {"code":"BRN009","name":"Mombasa Road Branch","region":"Nairobi",   "county":"Nairobi",   "type":"Standard", "tier":3},
        {"code":"BRN010","name":"Thika Road Mall Branch","region":"Nairobi","county":"Nairobi",   "type":"Standard", "tier":3},
        {"code":"BRN011","name":"Mombasa Main Branch","region":"Coast",     "county":"Mombasa",   "type":"Flagship", "tier":1},
        {"code":"BRN012","name":"Nyali Branch",       "region":"Coast",     "county":"Mombasa",   "type":"Main",     "tier":2},
        {"code":"BRN013","name":"Diani Branch",       "region":"Coast",     "county":"Kwale",     "type":"Standard", "tier":3},
        {"code":"BRN014","name":"Malindi Branch",     "region":"Coast",     "county":"Kilifi",    "type":"Standard", "tier":3},
        {"code":"BRN015","name":"Kisumu Main Branch", "region":"Nyanza",    "county":"Kisumu",    "type":"Flagship", "tier":1},
        {"code":"BRN016","name":"Kisumu Mega Branch", "region":"Nyanza",    "county":"Kisumu",    "type":"Main",     "tier":2},
        {"code":"BRN017","name":"Migori Branch",      "region":"Nyanza",    "county":"Migori",    "type":"Standard", "tier":3},
        {"code":"BRN018","name":"Homabay Branch",     "region":"Nyanza",    "county":"Homabay",   "type":"Standard", "tier":3},
        {"code":"BRN019","name":"Nakuru Main Branch", "region":"Rift Valley","county":"Nakuru",   "type":"Flagship", "tier":1},
        {"code":"BRN020","name":"Nakuru West Branch", "region":"Rift Valley","county":"Nakuru",   "type":"Main",     "tier":2},
        {"code":"BRN021","name":"Eldoret Main Branch","region":"Rift Valley","county":"Uasin Gishu","type":"Flagship","tier":1},
        {"code":"BRN022","name":"Kitale Branch",      "region":"Rift Valley","county":"Trans Nzoia","type":"Main",   "tier":2},
        {"code":"BRN023","name":"Bungoma Branch",     "region":"Western",   "county":"Bungoma",   "type":"Main",     "tier":2},
        {"code":"BRN024","name":"Kakamega Branch",    "region":"Western",   "county":"Kakamega",  "type":"Standard", "tier":3},
        {"code":"BRN025","name":"Kisii Main Branch",  "region":"Nyanza",    "county":"Kisii",     "type":"Main",     "tier":2},
        {"code":"BRN026","name":"Nyeri Branch",       "region":"Central",   "county":"Nyeri",     "type":"Standard", "tier":3},
        {"code":"BRN027","name":"Thika Branch",       "region":"Central",   "county":"Kiambu",    "type":"Main",     "tier":2},
        {"code":"BRN028","name":"Kikuyu Branch",      "region":"Central",   "county":"Kiambu",    "type":"Standard", "tier":3},
        {"code":"BRN029","name":"Meru Branch",        "region":"Eastern",   "county":"Meru",      "type":"Standard", "tier":3},
        {"code":"BRN030","name":"Embu Branch",        "region":"Eastern",   "county":"Embu",      "type":"Standard", "tier":3},
        {"code":"BRN031","name":"Machakos Branch",    "region":"Eastern",   "county":"Machakos",  "type":"Standard", "tier":3},
        {"code":"BRN032","name":"Kitui Branch",       "region":"Eastern",   "county":"Kitui",     "type":"Light",    "tier":4},
        {"code":"BRN033","name":"Garissa Branch",     "region":"North Eastern","county":"Garissa", "type":"Light",   "tier":4},
        {"code":"BRN034","name":"Wajir Branch",       "region":"North Eastern","county":"Wajir",   "type":"Light",   "tier":4},
        {"code":"BRN035","name":"Lamu Branch",        "region":"Coast",     "county":"Lamu",      "type":"Light",    "tier":4},
    ],
    "regions": ["Nairobi","Coast","Nyanza","Rift Valley","Western","Central","Eastern","North Eastern","Head Office"],
    "hierarchy": {
        "Managing Director": [],
        "Director Retail Banking": ["Managing Director"],
        "Director Commercial Banking": ["Managing Director"],
        "Head Of Retail": ["Director Retail Banking"],
        "Head Of Corporate": ["Director Commercial Banking"],
        "Head Of SME": ["Director Commercial Banking"],
        "Regional Head": ["Head Of Retail","Director Retail Banking"],
        "Branch Manager": ["Regional Head"],
        "Branch Operations Manager": ["Branch Manager"],
        "Branch Credit Manager": ["Branch Manager"],
        "Branch Operations Supervisor": ["Branch Operations Manager"],
        "Customer Service Officer": ["Branch Operations Manager"],
        "Teller": ["Branch Operations Manager"],
        "Relationship Officer Personal Banking": ["Branch Credit Manager"],
        "Relationship Officer Business Banking": ["Branch Credit Manager"],
        "Direct Sales Officer": ["Branch Credit Manager"],
        "Relationship Manager Corporate": ["Head Of Corporate"],
        "Relationship Manager SME": ["Head Of SME"],
        "Chief Finance Officer": ["Managing Director"],
        "Chief Risk Officer": ["Managing Director"],
        "Chief Operations Officer": ["Managing Director"],
        "Chief Compliance Officer": ["Managing Director"],
        "Chief Human Resources Officer": ["Managing Director"],
    },
    "roles": [
        "Managing Director","Director Retail Banking","Director Commercial Banking",
        "Head Of Retail","Head Of Corporate","Head Of SME","Regional Head",
        "Branch Manager","Branch Operations Manager","Branch Credit Manager",
        "Branch Operations Supervisor","Customer Service Officer","Teller",
        "Relationship Officer Personal Banking","Relationship Officer Business Banking",
        "Direct Sales Officer","Relationship Manager Corporate","Relationship Manager SME",
        "Chief Finance Officer","Chief Risk Officer","Chief Operations Officer",
        "Chief Compliance Officer","Chief Human Resources Officer","Chief Credit Officer",
        "Head Of Digital Innovation","Head Of Strategy","Head Of Internal Audit",
        "Head Of Marketing","IT Manager","Operations Manager","Procurement Manager",
        "HR Business Partner","HR Officer","Compliance Officer","Risk Manager",
        "Financial Controller","Treasury Manager","Internal Auditor","Strategy Analyst",
        "Marketing Officer","IT Support Officer","Recovery Officer","Procurement Officer",
        "Credit Analyst","Credit Administrator","Debt Recovery Unit Manager",
    ],
}

# Module-level cache — avoids repeated disk reads on every Streamlit rerun
# Invalidated by save_org_config() and on explicit cache clear
_ORG_CONFIG_CACHE: dict = {}
_ORG_CONFIG_MTIME: float = 0.0

def get_org_config() -> dict:
    """Load org config with module-level cache. Reads disk only when file changes."""
    global _ORG_CONFIG_CACHE, _ORG_CONFIG_MTIME
    try:
        if not ORG_CONFIG_FILE.exists():
            return DEFAULT_ORG_CONFIG.copy()
        mtime = ORG_CONFIG_FILE.stat().st_mtime
        if mtime != _ORG_CONFIG_MTIME or not _ORG_CONFIG_CACHE:
            saved = json.loads(ORG_CONFIG_FILE.read_text())
            merged = DEFAULT_ORG_CONFIG.copy()
            merged.update(saved)
            _ORG_CONFIG_CACHE = merged
            _ORG_CONFIG_MTIME = mtime
        return _ORG_CONFIG_CACHE
    except:
        return DEFAULT_ORG_CONFIG.copy()

def save_org_config(cfg: dict):
    """Persist org config to disk and invalidate cache."""
    global _ORG_CONFIG_CACHE, _ORG_CONFIG_MTIME
    _ORG_CONFIG_CACHE = {}
    _ORG_CONFIG_MTIME = 0.0
    ORG_CONFIG_FILE.write_text(json.dumps(cfg, indent=2))

def get_branch_region_map() -> dict:
    """Dynamic branch→region map from org config. Replaces hardcoded BRANCH_REGION."""
    cfg = get_org_config()
    return {b["name"]: b["region"] for b in cfg.get("branches", [])}

def get_branch_list() -> list:
    """All branch names from org config."""
    cfg = get_org_config()
    return [b["name"] for b in cfg.get("branches", []) if b.get("type") != "HO"]

def get_all_branches() -> list:
    """All branch dicts from org config."""
    return get_org_config().get("branches", [])

def get_org_hierarchy() -> dict:
    """Role → list of parent roles. From org config."""
    return get_org_config().get("hierarchy", {})

def get_scoring_scale() -> list:
    """BSC scoring scale thresholds — from org_config, fully configurable."""
    try:
        cfg = get_org_config()
        return cfg.get("scoring_scale", {}).get("thresholds", [])
    except: return []

def bsc_score_from_pct(achievement_pct: float, reverse: bool = False) -> float:
    """
    Convert achievement % to BSC score using configurable scoring scale.
    reverse=True for KPIs where lower is better (NPL, PAR, Dormancy).
    """
    if achievement_pct is None: return 0.0
    pct = achievement_pct if not reverse else (200 - achievement_pct)
    thresholds = get_scoring_scale()
    if not thresholds:
        # Fallback built-in scale
        if pct > 130: return 5.0
        if pct > 120: return 4.5
        if pct > 110: return 4.0
        if pct > 100: return 3.5
        if pct >= 91: return 3.0
        if pct >= 61: return 2.5
        if pct >= 51: return 2.0
        if pct >= 31: return 1.5
        return 1.0
    for t in sorted(thresholds, key=lambda x: -x["min"]):
        if pct >= t["min"]: return float(t["score"])
    return 1.0

def get_performance_bands() -> list:
    """Performance band definitions — from org_config."""
    try:
        cfg = get_org_config()
        return cfg.get("performance_bands", [])
    except: return []

def score_to_band(score: float) -> dict:
    """Return the performance band dict for a given BSC score."""
    bands = get_performance_bands()
    if not bands:
        # Fallback
        if score >= 4.5: return {"label":"Exceeded By Far","color":"var(--brand-primary,#006B3F)","bg":"var(--brand-light,#E8F5EE)"}
        if score >= 3.5: return {"label":"Exceeded",       "color":"var(--brand-mid,#1D9E75)","bg":"#D1FAE5"}
        if score >= 3.0: return {"label":"Met",             "color":"#F5A623","bg":"#FFFBEB"}
        if score >= 2.0: return {"label":"Partially Met",   "color":"#E67E22","bg":"#FEF3C7"}
        return            {"label":"Unmet",            "color":"#E24B4A","bg":"#FEF2F2"}
    for b in sorted(bands, key=lambda x: -x["min_score"]):
        if score >= b["min_score"]: return b
    return bands[-1] if bands else {"label":"Unmet","color":"#E24B4A","bg":"#FEF2F2"}

def get_currency() -> str:
    """Currency symbol from org_config."""
    try:
        return get_org_config().get("currency_symbol", "KES")
    except: return "KES"

def get_fiscal_year() -> str:
    """Active fiscal year / period from org_config."""
    try:
        return str(get_org_config().get("active_period", str(get_org_config().get("active_period","2026"))))
    except: return str(get_org_config().get("active_period","2026"))

def get_pipeline_stages() -> list:
    """Pipeline stages from org_config — names, colors, default probabilities."""
    try:
        cfg = get_org_config()
        stages = cfg.get("pipeline_stages", [])
        if stages: return stages
    except: pass
    return [
        {"stage":"Prospecting","color":"#6366F1","prob_default":20},
        {"stage":"Needs Analysis","color":"#8B5CF6","prob_default":35},
        {"stage":"Proposal","color":"#3B82F6","prob_default":50},
        {"stage":"Credit Review","color":"#F59E0B","prob_default":55},
        {"stage":"Approval","color":"#10B981","prob_default":85},
        {"stage":"Disbursed","color":"var(--brand-mid,#1D9E75)","prob_default":100},
        {"stage":"Closed Lost","color":"#6B7280","prob_default":0},
    ]

def get_product_types() -> list:
    """Product types for pipeline from org_config."""
    try:
        return get_org_config().get("product_types", [])
    except: return []

def get_leave_types() -> dict:
    """Leave types from org_config — overrides LEAVE_TYPES constant."""
    try:
        lt = get_org_config().get("leave_types", {})
        if lt: return lt
    except: pass
    return LEAVE_TYPES  # fallback to core constant

def get_org_roles() -> list:
    """All role names at this bank."""
    return get_org_config().get("roles", [])

def get_root_roles() -> list:
    """Top-of-hierarchy roles (CEO/MD) — from org_config hierarchy."""
    try:
        hier = get_org_config().get('hierarchy', {})
        return [r for r, parents in hier.items() if not parents]
    except:
        return []

def is_top_management(role: str) -> bool:
    """True if this role is at or near the root of the hierarchy.
    Matches exact root roles AND common aliases (MD, CEO, Chief Executive)."""
    roots = get_root_roles()
    if role in roots:
        return True
    # Fuzzy: any root whose name contains this role string or vice versa
    role_l = role.lower()
    for r in roots:
        r_l = r.lower()
        if role_l in r_l or r_l in role_l:
            return True
    # Common aliases
    aliases = {'md','ceo','chief executive','managing director'}
    return any(a in role_l for a in aliases)

def is_branch_role(role: str) -> bool:
    """True if this role belongs to branch staff (Category=Branch)."""
    try:
        br = get_org_config().get('role_categories', {}).get('branch_staff', [])
        return role in br if br else False
    except:
        return False

def rename_role_everywhere(old_name: str, new_name: str) -> dict:
    """
    Atomically rename a role across: org_config (roles, hierarchy,
    role_categories), kpi_library (role_kpis), users.json.
    Returns summary of what was updated.
    """
    summary = {}
    try:
        cfg = get_org_config()
        cfg['roles'] = [new_name if r == old_name else r
                        for r in cfg.get('roles', [])]
        new_hier = {}
        for r, parents in cfg.get('hierarchy', {}).items():
            new_key    = new_name if r == old_name else r
            new_parents= [new_name if p == old_name else p for p in parents]
            new_hier[new_key] = new_parents
        cfg['hierarchy'] = new_hier
        new_cats = {}
        for cat, roles in cfg.get('role_categories', {}).items():
            new_cats[cat] = [new_name if r == old_name else r for r in roles]
        cfg['role_categories'] = new_cats
        save_org_config(cfg)
        summary['org_config'] = True
    except Exception as e:
        summary['org_config'] = str(e)
    try:
        lib = get_kpi_library()
        rk  = lib.get('role_kpis', {})
        if old_name in rk:
            rk[new_name] = rk.pop(old_name)
            lib['role_kpis'] = rk
            save_kpi_library(lib)
        summary['kpi_library'] = True
    except Exception as e:
        summary['kpi_library'] = str(e)
    return summary

def get_performance_bands() -> list:
    """Performance band definitions from org_config."""
    try:
        return get_org_config().get('performance_bands', [])
    except:
        return []

def get_scoring_scale() -> list:
    """BSC scoring thresholds from org_config."""
    try:
        return get_org_config().get('scoring_scale', {}).get('thresholds', [])
    except:
        return []


def get_pipeline_settings() -> dict:
    """Load pipeline settings (custom products, stages etc.) from disk."""
    _f = DATA_DIR / "pipeline_settings.json"
    if not _f.exists():
        return {}
    try:
        return json.loads(_f.read_text())
    except:
        return {}

def save_pipeline_settings(settings: dict):
    _f = DATA_DIR / "pipeline_settings.json"
    _f.write_text(json.dumps(settings, indent=2))

def get_all_pipeline_stage_names() -> set:
    """Every stage name across the configured pipeline — top-level stages plus
    each deal_category's flow — from pipeline_settings.json. Batch A2
    (2026-06-15): the advance gate accepts any configured stage so all 9
    deal-category flows (17+ stages) work without hardcoding.
    """
    names = set()
    try:
        cfg = get_pipeline_settings()
        for st in cfg.get("stages", []):
            n = str(st.get("stage", "")).strip() if isinstance(st, dict) else str(st).strip()
            if n:
                names.add(n)
        for cat in cfg.get("deal_categories", []):
            for st in cat.get("stages", []):
                n = str(st).strip()
                if n:
                    names.add(n)
        # B17: per-product-class flows (asset/liability/insurance/other) so the
        # advance gate accepts any stage the bank configures in stage_flows.
        for flow in cfg.get("stage_flows", {}).values():
            if isinstance(flow, list):
                for st in flow:
                    n = str(st).strip()
                    if n:
                        names.add(n)
    except Exception:
        pass
    return names


def get_custom_product_types() -> list:
    """Return product types, using custom list if admin has configured one."""
    settings = get_pipeline_settings()
    custom = settings.get("product_types", [])
    return custom if custom else PRODUCT_TYPES

# ─── KPI INSIGHT ENGINE ──────────────────────────────────────────────
def get_kpi_insights(kpis_df):
    """
    Analyse a staff member's KPI rows and return structured insights.
    Returns dict: strengths, weaknesses, critical, summary_text
    """
    if kpis_df.empty:
        return {"strengths":[], "weaknesses":[], "critical":[], "summary":"No KPI data available."}

    strengths  = []
    weaknesses = []
    critical   = []

    for _, r in kpis_df.iterrows():
        score = r.get('Score', np.nan)
        ach   = r.get('Percent_Achieved', np.nan)
        kpi   = str(r.get('KPI',''))
        pillar= str(r.get('Pillar','General'))
        w     = r.get('Weight', 0)
        tgt   = r.get('Annual Target', 0) or 0
        act   = r.get('YTD_Actual', np.nan)

        if pd.isna(score): continue

        gap = ""
        if pd.notna(tgt) and pd.notna(act) and tgt > 0:
            shortfall = tgt - act
            if shortfall > 0:
                gap = f" (gap: {fmt_num(shortfall, short=True)})"

        entry = {"kpi": kpi, "pillar": pillar, "score": score,
                 "achievement": ach, "weight": w, "gap": gap,
                 "weighted": round(score * w, 3)}

        if score >= 3.5:
            strengths.append(entry)
        elif score < 2.5:
            critical.append(entry)
        elif score < 3.0:
            weaknesses.append(entry)

    # Sort by weight × impact descending
    critical.sort(key=lambda x: x['weight'], reverse=True)
    weaknesses.sort(key=lambda x: x['weight'], reverse=True)
    strengths.sort(key=lambda x: x['weighted'], reverse=True)

    # Build narrative summary
    bsc = kpis_df['Weighted_Score'].sum() if 'Weighted_Score' in kpis_df.columns else 0
    remark = get_performance_remarks(bsc)

    if critical:
        top_crit = critical[0]
        crit_names = ", ".join(x['kpi'] for x in critical[:3])
        summary = (f"{remark}. Critical attention needed on: {crit_names}. "
                   f"These carry {sum(x['weight'] for x in critical)*100:.0f}% of total weight.")
    elif weaknesses:
        weak_names = ", ".join(x['kpi'] for x in weaknesses[:3])
        summary = (f"{remark}. Improvement needed in: {weak_names}. "
                   f"Addressing these could add ~{sum(x['weight']*(3-x['score']) for x in weaknesses):.2f} to BSC score.")
    elif strengths:
        str_names = ", ".join(x['kpi'] for x in strengths[:3])
        summary = f"{remark}. Performing strongly on: {str_names}. Maintain momentum."
    else:
        summary = f"{remark}. Performance is broadly on target."

    return {"strengths": strengths, "weaknesses": weaknesses,
            "critical": critical, "summary": summary, "bsc": round(bsc, 2)}


def render_insight_card(insights, staff_name=""):
    """Render a colour-coded insight card using st.markdown."""
    summary = insights.get("summary","")
    critical = insights.get("critical",[])
    weaknesses = insights.get("weaknesses",[])
    strengths  = insights.get("strengths",[])

    # Summary banner
    bsc = insights.get("bsc", 0)
    remark = get_performance_remarks(bsc)
    banner_colour = {
        "Exceeded By Far":"#2ECC71","Exceeded":"#58D68D","Met":"#F39C12",
        "Partially Met":"#E67E22","Unmet":"#E74C3C","No Data":"#BDC3C7"
    }.get(remark,"#BDC3C7")

    st.markdown(
        f"<div style='padding:12px 16px;background:{banner_colour}22;"
        f"border-left:5px solid {banner_colour};border-radius:6px;margin-bottom:12px'>"
        f"<strong>{staff_name + ' — ' if staff_name else ''}{remark} ({fmt_score(bsc)})</strong><br>"
        f"<span style='font-size:13px'>{summary}</span></div>",
        unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("##### 🔴 Needs attention")
        if critical:
            for x in critical[:4]:
                st.markdown(
                    f"<div style='padding:7px 10px;background:#FFB6C122;border-left:3px solid #E74C3C;"
                    f"border-radius:4px;margin:3px 0;font-size:13px'>"
                    f"<b>{x['kpi']}</b><br>"
                    f"Score {fmt_score(x['score'])} | {fmt_pct(x['achievement'])} achieved"
                    f"{x['gap']}<br>"
                    f"<span style='color:#888'>Weight {x['weight']*100:.0f}% | Pillar: {x['pillar']}</span>"
                    f"</div>", unsafe_allow_html=True)
        else:
            st.success("No critical KPIs")

    with c2:
        st.markdown("##### 🟡 Below target")
        if weaknesses:
            for x in weaknesses[:4]:
                st.markdown(
                    f"<div style='padding:7px 10px;background:#FFD70022;border-left:3px solid #F39C12;"
                    f"border-radius:4px;margin:3px 0;font-size:13px'>"
                    f"<b>{x['kpi']}</b><br>"
                    f"Score {fmt_score(x['score'])} | {fmt_pct(x['achievement'])} achieved"
                    f"{x['gap']}<br>"
                    f"<span style='color:#888'>Weight {x['weight']*100:.0f}% | Pillar: {x['pillar']}</span>"
                    f"</div>", unsafe_allow_html=True)
        else:
            st.success("No below-target KPIs")

    with c3:
        st.markdown("##### 🟢 Strengths")
        if strengths:
            for x in strengths[:4]:
                st.markdown(
                    f"<div style='padding:7px 10px;background:#90EE9022;border-left:3px solid #2ECC71;"
                    f"border-radius:4px;margin:3px 0;font-size:13px'>"
                    f"<b>{x['kpi']}</b><br>"
                    f"Score {fmt_score(x['score'])} | {fmt_pct(x['achievement'])} achieved<br>"
                    f"<span style='color:#888'>Weight {x['weight']*100:.0f}% | Pillar: {x['pillar']}</span>"
                    f"</div>", unsafe_allow_html=True)
        else:
            st.info("No standout strengths yet")

def parse_month_column(col_name):
    col_str = str(col_name).strip()
    col_str = col_str.replace(" Actual","").replace(" actual","")
    month_map = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,
                 'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
    for name, num in month_map.items():
        if name in col_str.upper():
            digits = ''.join(filter(str.isdigit, col_str))
            year = 2000 + int(digits) if digits and int(digits) < 100 else (
                   int(digits) if digits else datetime.now().year)
            return datetime(year, num, 1)
    return None

def detect_month_actual_columns(df):
    results = []
    month_abbr = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                  'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    skip = ['target','ytd','modulator','unnamed','annual']
    for col in df.columns:
        col_str = str(col).strip().lower()
        if any(x in col_str for x in skip): continue
        for abbr, num in month_abbr.items():
            if abbr in col_str:
                digits = re.findall(r'\d+', col_str)
                yr = 2000 + int(digits[0]) if digits and int(digits[0]) < 100 else (
                     int(digits[0]) if digits else datetime.now().year)
                results.append((datetime(yr, num, 1), col))
                break
    results.sort(key=lambda x: x[0])
    return [col for _, col in results]

# ─── DATA PROCESSING ─────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def process_kpi_data(df):
    # ── Strip any repeated header rows (row where KPI == 'KPI') ──────
    # Happens when Excel has a title row + header row → header=1 picks
    # up row 2 as headers, but row 2 is sometimes duplicated as data.
    if 'KPI' in df.columns:
        mask_hdr = df['KPI'].astype(str).str.strip().str.lower() == 'kpi'
        if mask_hdr.any():
            df = df[~mask_hdr].copy()
    # Also drop rows where Staff Name == 'Staff Name' (same issue)
    if 'Staff Name' in df.columns:
        mask_sn = df['Staff Name'].astype(str).str.strip().str.lower() == 'staff name'
        if mask_sn.any():
            df = df[~mask_sn].copy()

    # ── Normalise column names from either Excel format ───────────────
    # New format: 'Jan-26 Actual' → rename to 'Jan-26'
    rename_map = {}
    for col in df.columns:
        col_str = str(col)  # guard against integer column names
        for month in ['Jan','Feb','Mar','Apr','May','Jun',
                      'Jul','Aug','Sep','Oct','Nov','Dec']:
            if col_str.startswith(month) and col_str.endswith(' Actual'):
                rename_map[col] = col_str.replace(' Actual','')
    if rename_map:
        df = df.rename(columns=rename_map)

    # Drop any purely-numeric unnamed columns (openpyxl artefacts)
    df = df[[c for c in df.columns
             if not (isinstance(c, int) or str(c).startswith('Unnamed'))]]

    df = df.copy()
    # Ensure Staff Code column exists and is string
    if 'Staff Code' not in df.columns:
        # Try to find it with different capitalisation
        sc_candidates = [c for c in df.columns if str(c).lower().replace(' ','') == 'staffcode']
        if sc_candidates:
            df = df.rename(columns={sc_candidates[0]: 'Staff Code'})
        else:
            df['Staff Code'] = '0'
    df['Staff Code'] = df['Staff Code'].astype(str).str.strip()

    month_cols    = detect_month_actual_columns(df)
    now           = datetime.now()
    active_months = [c for c in month_cols
                     if (lambda d: d and (d.year < now.year or
                         (d.year == now.year and d.month <= now.month))
                        )(parse_month_column(c))]

    # YTD_Actual: sum of active month actuals if available, else use Annual Actual column
    if active_months:
        df['YTD_Actual'] = df[active_months].apply(
            pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
    elif 'Annual Actual' in df.columns:
        df['YTD_Actual'] = pd.to_numeric(df['Annual Actual'], errors='coerce').fillna(0)
    elif 'YTD_Actual' in df.columns:
        df['YTD_Actual'] = pd.to_numeric(df['YTD_Actual'], errors='coerce').fillna(0)
    else:
        df['YTD_Actual'] = 0.0

    # Always keep Annual Actual as a separate column for reference
    if 'Annual Actual' not in df.columns and 'YTD_Actual' in df.columns:
        df['Annual Actual'] = df['YTD_Actual']

    if 'Weight' in df.columns:
        df['Weight'] = (
            df['Weight'].astype(str).str.replace('%','',regex=False)
            .apply(lambda x: float(x) if x not in ['nan','','None'] else np.nan)
            .apply(lambda x: x/100 if pd.notna(x) and x > 1 else x)
            .fillna(0)
        )
    else:
        df['Weight'] = 1.0

    if 'Annual Target' not in df.columns and 'Target' in df.columns:
        df['Annual Target'] = df['Target']
    elif 'Annual Target' not in df.columns:
        # No target column — targets come from cascade module
        # Create placeholder zeros; cascade write-back fills these in
        df['Annual Target'] = 0.0

    reverse_kpis = ['PAR','NPL','PORTFOLIO AT RISK','DELINQUENCY','COST','EXPENSE']
    # Vectorized scoring — no iterrows, runs 50x faster on 2782 rows
    target_raw = pd.to_numeric(df['Annual Target'], errors='coerce').fillna(0)
    actual     = pd.to_numeric(df['YTD_Actual'],    errors='coerce').fillna(0)
    kpi_up     = df['KPI'].astype(str).str.upper()

    # When Annual Target is zero (e.g. pre-cascade or new format),
    # fall back to Annual Actual as a proxy target so scoring still works
    annual_actual_col = pd.to_numeric(
        df.get('Annual Actual', pd.Series(0, index=df.index)), errors='coerce').fillna(0)
    target = target_raw.where(target_raw != 0, annual_actual_col)
    df['Annual Target'] = target  # keep the resolved value in df

    is_rev  = kpi_up.apply(lambda k: any(t in k for t in reverse_kpis))
    valid   = target.notna() & (target != 0)

    ach = pd.Series(np.nan, index=df.index)
    # Forward KPIs
    fwd = valid & ~is_rev
    ach[fwd] = (actual[fwd] / target[fwd]).clip(0, 1.5)
    # Reverse KPIs
    rev = valid & is_rev
    has_actual = actual > 0
    ach[rev & has_actual] = (target[rev & has_actual] / actual[rev & has_actual]).clip(0, 1.5)
    ach[rev & ~has_actual] = 0.0

    # Score from achievement using vectorized np.select
    conditions = [
        ach.isna(),
        ach < 0.30, ach <= 0.50, ach <= 0.60, ach <= 0.90,
        ach <= 1.00, ach <= 1.10, ach <= 1.20, ach <= 1.30,
    ]
    choices = [np.nan, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5]
    score = pd.Series(np.select(conditions, choices, default=5.0), index=df.index)

    df['Score']            = score.round(2)
    df['Weighted_Score']   = (score * df['Weight'].fillna(0)).round(2)
    df['Percent_Achieved'] = (ach * 100).round(1)
    return df

@st.cache_data(show_spinner=False)

def compute_initiative_kpis(staff_name: str) -> dict:
    """
    Compute initiative KPI actuals for a staff member from execute_initiatives.json.
    Returns: {
      'Initiative Implementation Score': float (0-100 weighted avg gate progress),
      'Active Initiatives Count': int
    }
    Gate weights: G0=0, G1=20, G2=40, G3=60, G4=80, G5=100
    """
    GATE_SCORES = {'G0':0,'G1':20,'G2':40,'G3':60,'G4':80,'G5':100}
    try:
        init_file = DATA_DIR / "execute_initiatives.json"
        if not init_file.exists():
            return {'Initiative Implementation Score': 0, 'Active Initiatives Count': 0}
        initiatives = json.loads(init_file.read_text())
        if not isinstance(initiatives, list):
            initiatives = []

        owned = [i for i in initiatives
                 if (i.get('io','') == staff_name or
                     i.get('io_backup','') == staff_name or
                     any(ms.get('owner','') == staff_name
                         for ms in i.get('milestones',[])))
                 and i.get('gate','G0') not in ('Dropped',)]

        active = [i for i in owned if i.get('gate','G0') in ('G1','G2','G3','G4')]

        if not owned:
            return {'Initiative Implementation Score': 0, 'Active Initiatives Count': 0}

        # Weighted implementation score
        gate_scores = [GATE_SCORES.get(i.get('gate','G0'), 0) for i in owned]
        impl_score  = round(sum(gate_scores) / len(gate_scores), 1)

        return {
            'Initiative Implementation Score': impl_score,
            'Active Initiatives Count': len(active),
        }
    except:
        return {'Initiative Implementation Score': 0, 'Active Initiatives Count': 0}

def build_staff_scores(df):
    grp = [c for c in ['Staff Name','Role','Unit','Category','Staff Code','Staff Status'] if c in df.columns]
    staff = df.groupby(grp, as_index=False).agg(
        Final_BSC_Score     = ('Weighted_Score','sum'),
        Avg_KPI_Score       = ('Score','mean'),
        Avg_Achievement_Pct = ('Percent_Achieved','mean'),
        KPI_Count           = ('KPI','count'),
    )
    if 'Role' in staff.columns and 'Category' in staff.columns:
        staff['Role_Function'] = staff.apply(
            lambda r: get_role_function(r['Role'], r['Category']), axis=1)
    staff['Final_BSC_Score']     = staff['Final_BSC_Score'].round(2)
    staff['Avg_KPI_Score']       = staff['Avg_KPI_Score'].round(2)
    staff['Avg_Achievement_Pct'] = staff['Avg_Achievement_Pct'].round(1)
    staff['Performance_Remark']  = staff['Final_BSC_Score'].apply(get_performance_remarks)
    staff['Overall_Rank']        = staff['Final_BSC_Score'].rank(method='dense', ascending=False).astype(int)
    staff['Role_Rank']           = staff.groupby('Role')['Final_BSC_Score'].rank(method='dense', ascending=False).astype(int)
    staff['Role_Total']          = staff.groupby('Role')['Staff Name'].transform('count')
    staff['Percentile']          = (staff['Final_BSC_Score'].rank(pct=True)*100).round(1)
    if 'Unit' in staff.columns:
        staff['Region'] = staff['Unit'].map(BRANCH_REGION).fillna('Head Office')
    return staff.sort_values(by='Overall_Rank').reset_index(drop=True)


@st.cache_data(show_spinner=False)
def build_staff_registry(_raw_bytes: bytes):
    """Extract staff master data — accepts raw bytes. Handles multi-sheet and single-sheet formats."""
    try:
        raw = _raw_bytes
        xl  = pd.ExcelFile(io.BytesIO(raw))

        # ── Try new multi-sheet format (Staff Register sheet) ─────────
        if 'Staff Register' in xl.sheet_names:
            df = pd.read_excel(io.BytesIO(raw), sheet_name='Staff Register', header=1)
            # Normalise column names
            df.columns = [str(c).strip() for c in df.columns]
            # Strip any repeated header row (where Staff Code == 'Staff Code')
            if 'Staff Code' in df.columns:
                df = df[df['Staff Code'].astype(str).str.strip().str.lower() != 'staff code'].copy()
            rename = {'Reports To Code': 'Reports_To', 'Staff Status': 'Staff Status',
                      'Hire Date': 'Hire Date', 'Role Function': 'Role Function'}
            df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
            if 'Region' not in df.columns and 'Unit' in df.columns:
                df['Region'] = df['Unit'].map(BRANCH_REGION).fillna('Head Office')
            return df

        # ── Fallback: single-sheet format (original system build file) ─
        for sheet in xl.sheet_names:
            df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet)
            if 'Hire Date' in df.columns or 'Staff Status' in df.columns:
                df.columns = [str(c).strip() for c in df.columns]
                if 'Region' not in df.columns and 'Unit' in df.columns:
                    df['Region'] = df['Unit'].map(BRANCH_REGION).fillna('Head Office')
                return df

        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()



# ─── KENYA LEAVE TYPES (Employment Act 2007 & amendments) ────────────
# Default leave entitlements — admin can override these via Admin → Leave Settings
# Days are configurable per company policy, not tied to any specific legislation
LEAVE_TYPES_DEFAULT = {
    "Annual Leave":        {"days_entitled": 21, "max_days": 21, "paid": True,  "affects_performance": False, "compensation": "pro_rata",     "color": "var(--brand-primary,#006B3F)", "description": "Annual leave entitlement per year"},
    "Sick Leave":          {"days_entitled": 14, "max_days": 14, "paid": True,  "affects_performance": True,  "compensation": "exclude_month", "color": "#E24B4A", "description": "Sick leave — full pay period"},
    "Maternity Leave":     {"days_entitled": 91, "max_days": 91, "paid": True,  "affects_performance": True,  "compensation": "exclude_all",   "color": "#9B59B6", "description": "Fully paid maternity leave"},
    "Paternity Leave":     {"days_entitled": 14, "max_days": 14, "paid": True,  "affects_performance": False, "compensation": "pro_rata",      "color": "#185FA5", "description": "Fully paid paternity leave"},
    "Compassionate Leave": {"days_entitled": 5,  "max_days": 5,  "paid": True,  "affects_performance": False, "compensation": "pro_rata",      "color": "#7F8C8D", "description": "Bereavement/compassionate — immediate family"},
    "Study Leave":         {"days_entitled": 0,  "max_days": 0,  "paid": True,  "affects_performance": False, "compensation": "pro_rata",      "color": "#F5A623", "description": "Employer-discretionary; exam or professional courses"},
    "Leave Without Pay":   {"days_entitled": 0,  "max_days": 0,  "paid": False, "affects_performance": True,  "compensation": "exclude_month", "color": "#E67E22", "description": "Unpaid leave by mutual agreement"},
    "Garden Leave":        {"days_entitled": 0,  "max_days": 0,  "paid": True,  "affects_performance": True,  "compensation": "exclude_all",   "color": "#BDC3C7", "description": "Notice period served at home on full pay"},
    "Sabbatical Leave":    {"days_entitled": 0,  "max_days": 0,  "paid": False, "affects_performance": True,  "compensation": "exclude_all",   "color": "#1ABC9C", "description": "Extended leave for research or personal development"},
    "Public Holiday":      {"days_entitled": 0,  "max_days": 0,  "paid": True,  "affects_performance": False, "compensation": "none",          "color": "#95A5A6", "description": "Public/gazetted holidays"},
}

def _load_leave_settings() -> dict:
    """Load leave settings — admin-configured days override defaults."""
    _f = DATA_DIR / "leave_settings.json"
    if _f.exists():
        try:
            saved = json.loads(_f.read_text())
            merged = {}
            for lt, defaults in LEAVE_TYPES_DEFAULT.items():
                merged[lt] = {**defaults, **(saved.get(lt,{}))}
            return merged
        except: pass
    return LEAVE_TYPES_DEFAULT.copy()

LEAVE_TYPES = _load_leave_settings()

def save_leave_settings(settings: dict):
    """Persist admin-configured leave days to disk."""
    _f = DATA_DIR / "leave_settings.json"
    _f.write_text(json.dumps(settings, indent=2))
    # Reload global
    global LEAVE_TYPES
    LEAVE_TYPES = settings

# ─── EXIT REASONS ─────────────────────────────────────────────────────

COMPENSATION_LABELS: dict = {
    "none":          "✅ No impact — full score applies",
    "pro_rata":      "⚠️ Pro-rata — score weighted by days worked",
    "exclude_month": "⚠️ Month excluded — that period removed from BSC",
    "exclude_all":   "🔴 All affected months excluded from scoring",
}

EXIT_REASONS = [
    "Resignation — Better opportunity",
    "Resignation — Salary/compensation",
    "Resignation — Work environment",
    "Resignation — Career growth concerns",
    "Resignation — Personal reasons",
    "Resignation — Relocation",
    "Contract end — Not renewed",
    "Retirement — Mandatory",
    "Retirement — Voluntary (early)",
    "Dismissal — Gross misconduct",
    "Dismissal — Performance",
    "Dismissal — Redundancy",
    "Dismissal — Abscondment",
    "Mutual separation",
    "Medical — Incapacitation",
    "Death in service",
]

# ─── TRANSFER REASONS ─────────────────────────────────────────────────
TRANSFER_REASONS = [
    "Performance improvement — new environment",
    "Branch need — skills gap",
    "Staff request — personal reasons",
    "Staff request — proximity to home",
    "Disciplinary — conflict resolution",
    "Rotational development",
    "Promotion transfer",
    "Role restructuring",
]

# ─── DISCIPLINARY CATEGORIES ──────────────────────────────────────────
DISCIPLINARY_CATEGORIES = [
    "Gross misconduct",
    "Insubordination",
    "Absenteeism / lateness",
    "Fraud / financial irregularity",
    "Sexual harassment",
    "Conflict of interest",
    "Breach of confidentiality",
    "Performance negligence",
    "Policy violation",
    "Substance abuse at workplace",
    "Bullying / workplace violence",
    "Misrepresentation",
]

DISCIPLINARY_STAGES = [
    "Verbal warning",
    "Written warning (1st)",
    "Written warning (final)",
    "Show cause notice",
    "Suspension pending investigation",
    "Disciplinary hearing scheduled",
    "Disciplinary hearing held",
    "Decision — Dismissed",
    "Decision — Demoted",
    "Decision — Final warning",
    "Decision — Exonerated",
    "Appeal filed",
    "Appeal resolved",
]

# ─── PIP CONSTANTS ────────────────────────────────────────────────────
PIP_DURATIONS = [30, 60, 90]   # days
PIP_OUTCOMES  = [
    "In progress",
    "Successfully completed",
    "Extended",
    "Terminated — dismissal",
    "Terminated — resignation",
    "Converted to final warning",
]

PIP_REVIEW_FREQUENCIES = ["Weekly", "Fortnightly", "Monthly"]

# ─── DILIGENCE SCORE WEIGHTS ──────────────────────────────────────────
# Used to compute per-staff HR diligence score (0-100)
DILIGENCE_WEIGHTS = {
    "milestone_on_time_rate":   0.30,   # % milestones completed on time
    "action_acceptance_rate":   0.20,   # % action items accepted within 24h
    "leave_compliance":         0.15,   # no AWOL / unexplained absence
    "disciplinary_clean":       0.20,   # no active disciplinary = 1.0
    "pip_clear":                0.15,   # not on PIP = 1.0
}

class LeaveManager:
    def __init__(self):
        self.file = DATA_DIR / "leave_records.json"
        if not self.file.exists(): self.file.write_text("[]")
        try:
            raw = self.file.read_text()
            self.records = json.loads(raw) if raw.strip() else []
            if not isinstance(self.records, list): self.records = []
        except:
            self.records = []

    def save(self):
        self.file.write_text(json.dumps(self.records, indent=2))

    def add_leave(self, staff_code, staff_name, leave_type, start_date, end_date,
                  reason, approved_by, notify_suppress=True):
        days = (end_date - start_date).days + 1
        record = {
            "id":              len(self.records) + 1,
            "staff_code":      clean_code(staff_code),
            "staff_name":      staff_name,
            "leave_type":      leave_type,
            "start_date":      str(start_date),
            "end_date":        str(end_date),
            "days":            days,
            "reason":          reason,
            "approved_by":     approved_by,
            "notify_suppress": notify_suppress,
            "compensation":    LEAVE_TYPES.get(leave_type, LEAVE_TYPES["Annual Leave"]).get("compensation", "none"),
            "affects_perf":    LEAVE_TYPES.get(leave_type, LEAVE_TYPES["Annual Leave"]).get("affects_performance", False),
            "status":          "Active" if start_date <= datetime.now().date() <= end_date else (
                               "Upcoming" if start_date > datetime.now().date() else "Completed"),
            "recorded_at":     datetime.now().isoformat(),
        }
        self.records.append(record)
        self.save()
        return record

    # ── Leave application workflow ──────────────────────────────────
    def apply_leave(self, staff_code, staff_name, staff_role, staff_unit,
                    leave_type, start_date, end_date, reason, relief_staff=""):
        """Staff submits a leave application — status=Pending until manager approves."""
        import datetime as _dt
        _start = _dt.date.fromisoformat(str(start_date)[:10]) if isinstance(start_date,str) else start_date
        _end   = _dt.date.fromisoformat(str(end_date)[:10]) if isinstance(end_date,str) else end_date
        days   = (_end - _start).days + 1
        lt_cfg = LEAVE_TYPES.get(leave_type, LEAVE_TYPES.get("Annual Leave",{}))
        rec = {
            "id":             f"LV{len(self.records)+1:05d}",
            "staff_code":     clean_code(staff_code),
            "staff_name":     staff_name,
            "staff_role":     staff_role,
            "staff_unit":     staff_unit,
            "leave_type":     leave_type,
            "start_date":     str(_start),
            "end_date":       str(_end),
            "days":           days,
            "reason":         reason,
            "relief_staff":   relief_staff,
            "approved":       None,        # None=pending, True=approved, False=rejected
            "approved_by":    "",
            "approved_at":    "",
            "rejection_reason": "",
            "applied_at":     datetime.now().isoformat(),
            "compensation":   lt_cfg.get("compensation","none"),
            "affects_perf":   lt_cfg.get("affects_performance",False),
            "hr_notified":    False,
            "status":         "Pending",
        }
        self.records.append(rec)
        self.save()
        return rec["id"]

    def approve_leave(self, leave_id, approved_by, approve=True, reason=""):
        """Manager approves or rejects a leave application."""
        for r in self.records:
            if str(r.get("id","")) == str(leave_id):
                r["approved"]    = approve
                r["approved_by"] = approved_by
                r["approved_at"] = datetime.now().isoformat()
                r["status"]      = "Approved" if approve else "Rejected"
                r["rejection_reason"] = reason if not approve else ""
                r["hr_notified"] = True   # flag for HR records
                self.save()
                return True
        return False

    def get_pending_approvals(self, manager_unit=None, manager_role=None):
        """Get leave requests awaiting manager approval."""
        pending = [r for r in self.records if r.get("approved") is None]
        if manager_unit:
            pending = [r for r in pending
                       if r.get("staff_unit","") == manager_unit]
        return sorted(pending, key=lambda x: x.get("applied_at",""))

    def get_all_leaves(self):
        return self.records

    def get_active_leave(self, staff_code=None):
        today = datetime.now().date()
        active = [r for r in self.records
                  if r.get("approved") is not False  # not rejected
                  and r['start_date'] <= str(today) <= r['end_date']]
        if staff_code:
            active = [r for r in active if r['staff_code'] == clean_code(staff_code)]
        return active

    def get_staff_leave(self, staff_code):
        return [r for r in self.records if r['staff_code'] == clean_code(staff_code)]

    def is_on_leave(self, staff_code):
        return len(self.get_active_leave(staff_code)) > 0

    def get_leave_months(self, staff_code, year=None):
        """Return set of (year, month) tuples when staff was on leave — for score compensation."""
        year = year or datetime.now().year
        leave_months = set()
        for r in self.get_staff_leave(staff_code):
            if not r.get('affects_perf'): continue
            sd = datetime.strptime(r['start_date'], "%Y-%m-%d").date()
            ed = datetime.strptime(r['end_date'],   "%Y-%m-%d").date()
            current = sd
            while current <= ed:
                if current.year == year:
                    leave_months.add((current.year, current.month))
                current = (current.replace(day=1) + timedelta(days=32)).replace(day=1)
        return leave_months

    def compensated_score(self, staff_code, monthly_scores, year=None):
        """
        Apply leave compensation to monthly scores.
        Returns (adjusted_score, explanation).
        monthly_scores = {month_num: weighted_score}
        """
        year = year or datetime.now().year
        leave_months = self.get_leave_months(staff_code, year)
        if not leave_months:
            total = sum(monthly_scores.values())
            return total, "No leave compensation applied."

        # Find compensation type from most recent affecting leave
        comp_type = None
        for r in self.get_staff_leave(staff_code):
            if r.get('affects_perf'):
                comp_type = r['compensation']
                break

        active_months = {m: s for m, s in monthly_scores.items()
                         if (year, m) not in leave_months}

        if comp_type == "exclude_period":
            if not active_months:
                return 0.0, "All months on leave — no score."
            avg = sum(active_months.values()) / len(active_months)
            explanation = (f"Leave months excluded ({len(leave_months)} months). "
                           f"Score based on {len(active_months)} active months.")
            return round(avg * len(monthly_scores), 2), explanation

        elif comp_type == "pro_rata":
            if not active_months:
                return 0.0, "All months on leave — no score."
            pro_score = sum(active_months.values())
            explanation = (f"Pro-rata: scored on {len(active_months)}/{len(monthly_scores)} months "
                           f"({len(leave_months)} leave months excluded).")
            return round(pro_score, 2), explanation

        else:
            total = sum(monthly_scores.values())
            return total, "Leave noted but full score applied."

# ─── HR MANAGER (exits, transfers, disciplinary, PIP) ────────────────
class HRManager:
    """Central HR data manager — exits, transfers, disciplinary, PIP, diligence scores."""

    def __init__(self):
        self.exits_file  = DATA_DIR / "hr_exits.json"
        self.trans_file  = DATA_DIR / "hr_transfers.json"
        self.disc_file   = DATA_DIR / "hr_disciplinary.json"
        self.pip_file    = DATA_DIR / "hr_pip.json"
        self.exits       = self._load(self.exits_file)
        self.transfers   = self._load(self.trans_file)
        self.disciplinary= self._load(self.disc_file)
        self.pips        = self._load(self.pip_file)

    def _load(self, path):
        if not path.exists(): path.write_text("[]")
        try:
            raw = path.read_text()
            d = json.loads(raw) if raw.strip() else []
            return d if isinstance(d, list) else []
        except: return []

    def _save(self, path, data):
        path.write_text(json.dumps(data, indent=2, default=str))

    # ── EXITS ─────────────────────────────────────────────────────
    def record_exit(self, data: dict) -> dict:
        rec = {
            "id":           f"EX{len(self.exits)+1:04d}",
            "staff_code":   clean_code(data.get("staff_code","")),
            "staff_name":   data.get("staff_name",""),
            "unit":         data.get("unit",""),
            "region":       data.get("region",""),
            "role":         data.get("role",""),
            "exit_date":    str(data.get("exit_date", date.today())),
            "reason":       data.get("reason",""),
            "reason_detail":data.get("reason_detail",""),
            "final_bsc":    data.get("final_bsc", None),
            "tenure_years": data.get("tenure_years", 0),
            "exit_interview":data.get("exit_interview",""),
            "rehire_eligible":data.get("rehire_eligible", True),
            "recorded_by":  data.get("recorded_by",""),
            "recorded_at":  datetime.now().isoformat(),
        }
        self.exits.append(rec)
        self._save(self.exits_file, self.exits)
        return rec

    def get_exits(self, months_back=12):
        cutoff = (datetime.now() - timedelta(days=months_back*30)).date()
        return [e for e in self.exits
                if datetime.strptime(e['exit_date'][:10], "%Y-%m-%d").date() >= cutoff]

    def exit_analytics(self):
        exits = self.exits
        if not exits: return {}
        by_reason   = {}
        by_unit     = {}
        by_quarter  = {}
        avg_tenure  = []
        bsc_at_exit = []
        for e in exits:
            r = e.get("reason","Unknown"); by_reason[r]  = by_reason.get(r, 0)+1
            u = e.get("unit","Unknown");   by_unit[u]    = by_unit.get(u, 0)+1
            try:
                d = datetime.strptime(e['exit_date'][:10], "%Y-%m-%d")
                q = f"{d.year} Q{(d.month-1)//3+1}"
                by_quarter[q] = by_quarter.get(q, 0)+1
            except: pass
            if e.get("tenure_years"): avg_tenure.append(float(e["tenure_years"]))
            if e.get("final_bsc"):    bsc_at_exit.append(float(e["final_bsc"]))
        return {"by_reason": by_reason, "by_unit": by_unit, "by_quarter": by_quarter,
                "avg_tenure": sum(avg_tenure)/len(avg_tenure) if avg_tenure else 0,
                "avg_bsc_at_exit": sum(bsc_at_exit)/len(bsc_at_exit) if bsc_at_exit else None,
                "total": len(exits)}

    # ── TRANSFERS ─────────────────────────────────────────────────
    def record_transfer(self, data: dict) -> dict:
        rec = {
            "id":           f"TR{len(self.transfers)+1:04d}",
            "staff_code":   clean_code(data.get("staff_code","")),
            "staff_name":   data.get("staff_name",""),
            "from_unit":    data.get("from_unit",""),
            "to_unit":      data.get("to_unit",""),
            "from_region":  data.get("from_region",""),
            "to_region":    data.get("to_region",""),
            "transfer_date":str(data.get("transfer_date", date.today())),
            "reason":       data.get("reason",""),
            "bsc_before":   data.get("bsc_before", None),
            "bsc_after":    data.get("bsc_after", None),
            "initiated_by": data.get("initiated_by",""),
            "recorded_at":  datetime.now().isoformat(),
        }
        self.transfers.append(rec)
        self._save(self.trans_file, self.transfers)
        return rec

    def get_transfers(self, months_back=12):
        cutoff = (datetime.now() - timedelta(days=months_back*30)).date()
        return [t for t in self.transfers
                if datetime.strptime(t['transfer_date'][:10], "%Y-%m-%d").date() >= cutoff]

    # ── DISCIPLINARY ──────────────────────────────────────────────
    def open_case(self, data: dict) -> dict:
        rec = {
            "id":            f"DC{len(self.disciplinary)+1:04d}",
            "staff_code":    clean_code(data.get("staff_code","")),
            "staff_name":    data.get("staff_name",""),
            "unit":          data.get("unit",""),
            "category":      data.get("category",""),
            "incident_date": str(data.get("incident_date", date.today())),
            "description":   data.get("description",""),
            "stage":         DISCIPLINARY_STAGES[0],
            "stage_history": [{"stage": DISCIPLINARY_STAGES[0],
                                "date": str(date.today()), "by": data.get("opened_by","")}],
            "opened_by":     data.get("opened_by",""),
            "status":        "Open",
            "outcome":       None,
            "outcome_date":  None,
            "notes":         [],
            "recorded_at":   datetime.now().isoformat(),
        }
        self.disciplinary.append(rec)
        self._save(self.disc_file, self.disciplinary)
        return rec

    def advance_stage(self, case_id: str, new_stage: str, note: str, by: str):
        for case in self.disciplinary:
            if case["id"] == case_id:
                case["stage"] = new_stage
                case["stage_history"].append({"stage": new_stage,
                                               "date": str(date.today()), "by": by})
                if note: case["notes"].append({"note": note, "date": str(date.today()), "by": by})
                if "Decision" in new_stage:
                    case["status"]   = "Closed"
                    case["outcome"]  = new_stage
                    case["outcome_date"] = str(date.today())
                self._save(self.disc_file, self.disciplinary)
                return case
        return None

    def get_active_cases(self):
        return [c for c in self.disciplinary if c.get("status") == "Open"]

    def staff_has_active_case(self, staff_code: str) -> bool:
        sc = clean_code(staff_code)
        return any(c["staff_code"]==sc and c.get("status")=="Open"
                   for c in self.disciplinary)

    # ── PIP ───────────────────────────────────────────────────────
    def open_pip(self, data: dict) -> dict:
        start = data.get("start_date", date.today())
        dur   = int(data.get("duration_days", 90))
        end   = start + timedelta(days=dur) if isinstance(start, date) else                 datetime.strptime(start, "%Y-%m-%d").date() + timedelta(days=dur)
        rec = {
            "id":              f"PIP{len(self.pips)+1:04d}",
            "staff_code":      clean_code(data.get("staff_code","")),
            "staff_name":      data.get("staff_name",""),
            "unit":            data.get("unit",""),
            "role":            data.get("role",""),
            "manager":         data.get("manager",""),
            "hr_officer":      data.get("hr_officer",""),
            "start_date":      str(start),
            "end_date":        str(end),
            "duration_days":   dur,
            "reason":          data.get("reason",""),
            "performance_gaps":data.get("performance_gaps",[]),
            "objectives":      data.get("objectives",[]),
            "review_frequency":data.get("review_frequency","Fortnightly"),
            "support_offered": data.get("support_offered",""),
            "reviews":         [],
            "outcome":         "In progress",
            "outcome_date":    None,
            "current_bsc":     data.get("current_bsc", None),
            "target_bsc":      data.get("target_bsc", 3.0),
            "opened_by":       data.get("opened_by",""),
            "recorded_at":     datetime.now().isoformat(),
        }
        self.pips.append(rec)
        self._save(self.pip_file, self.pips)
        return rec

    def add_pip_review(self, pip_id: str, review: dict):
        for pip in self.pips:
            if pip["id"] == pip_id:
                pip["reviews"].append({
                    "date":     str(date.today()),
                    "score":    review.get("score"),
                    "progress": review.get("progress",""),
                    "concerns": review.get("concerns",""),
                    "support":  review.get("support",""),
                    "by":       review.get("by",""),
                })
                if review.get("outcome") and review["outcome"] != "In progress":
                    pip["outcome"]      = review["outcome"]
                    pip["outcome_date"] = str(date.today())
                self._save(self.pip_file, self.pips)
                return pip
        return None

    def get_active_pips(self):
        return [p for p in self.pips if p.get("outcome") == "In progress"]

    def staff_on_pip(self, staff_code: str) -> bool:
        sc = clean_code(staff_code)
        return any(p["staff_code"]==sc and p.get("outcome")=="In progress"
                   for p in self.pips)

    def pip_days_remaining(self, pip: dict) -> int:
        try:
            end = datetime.strptime(pip["end_date"][:10], "%Y-%m-%d").date()
            return max(0, (end - date.today()).days)
        except: return 0

    # ── DILIGENCE SCORE ───────────────────────────────────────────
    def compute_diligence(self, staff_code: str, em=None, action_plans: dict=None) -> float:
        """
        Compute an HR diligence score (0-100) based on:
        - Milestone on-time rate (from ExecuteManager)
        - Action plan acceptance rate
        - No active disciplinary case
        - Not on PIP
        - Leave compliance (no unexplained absence)
        Returns a float 0-100.
        """
        sc = clean_code(staff_code)
        score = 0.0

        # Milestone on-time rate
        if em:
            try:
                all_ms = [ms for i in em.get_initiatives(status='All')
                          for ms in i.get('milestones',[])
                          if clean_code(ms.get('owner','')) == sc]
                if all_ms:
                    on_time = sum(1 for m in all_ms if m.get('status')=='Complete'
                                  and em._escalation_level(m) == 0)
                    rate = on_time / len(all_ms)
                else:
                    rate = 1.0
            except: rate = 1.0
            score += rate * DILIGENCE_WEIGHTS["milestone_on_time_rate"] * 100

        # Action plan acceptance rate
        if action_plans:
            my_actions = [a for plans in action_plans.values()
                          for a in plans if a.get('owner') == sc or a.get('owner') == staff_code]
            if my_actions:
                accepted = sum(1 for a in my_actions if a.get('accepted'))
                rate = accepted / len(my_actions)
            else:
                rate = 1.0
            score += rate * DILIGENCE_WEIGHTS["action_acceptance_rate"] * 100
        else:
            score += DILIGENCE_WEIGHTS["action_acceptance_rate"] * 100

        # Disciplinary
        disc_score = 0.0 if self.staff_has_active_case(sc) else 1.0
        score += disc_score * DILIGENCE_WEIGHTS["disciplinary_clean"] * 100

        # PIP
        pip_score = 0.0 if self.staff_on_pip(sc) else 1.0
        score += pip_score * DILIGENCE_WEIGHTS["pip_clear"] * 100

        # Leave compliance — assume clean if no unexplained absence recorded
        score += DILIGENCE_WEIGHTS["leave_compliance"] * 100

        return min(100.0, round(score, 1))



# ─── TARGET CASCADING MANAGER ─────────────────────────────────────────
class CascadeManager:
    """
    Manages top-down target allocation:
    MD → Director → Head/Manager → Staff
    Each level allocates portions of their own target downwards.
    """
    def __init__(self):
        self.file        = DATA_DIR / "target_cascade.json"
        self.bank_file   = DATA_DIR / "bank_targets.json"
        self.fixed_file  = DATA_DIR / "fixed_kpis.json"
        self.cascade     = self._load()
        self.bank_targets= self._load_bank()
        self.fixed_kpis  = self._load_fixed()

    def __getattr__(self, name):
        """Return safe defaults for any missing attribute — prevents stale-instance errors."""
        _safe_defaults = {
            "targets_locked": False,
            "cascade": {},
            "bank_targets": {},
            "fixed_kpis": {},
            "cascade_deadlines": {},
            "global_timeline": {},
            "review_requests": [],
        }
        if name in _safe_defaults:
            return _safe_defaults[name]
        # For unknown methods, return a no-op callable
        return lambda *a, **k: None

    def _load(self):
        if not self.file.exists(): self.file.write_text("{}")
        try:
            raw = self.file.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _load_bank(self):
        if not self.bank_file.exists(): self.bank_file.write_text("{}")
        try:
            raw = self.bank_file.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _load_fixed(self):
        if not self.fixed_file.exists(): self.fixed_file.write_text("{}")
        try:
            raw = self.fixed_file.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _save(self):
        self.file.write_text(json.dumps(self.cascade, indent=2, default=str))

    # ── Deadline tracking ─────────────────────────────────────────────
    def set_cascade_deadline(self, from_code: str, period: str,
                              confirm_by: str, cascade_by: str, set_by: str):
        """
        MD/manager sets deadlines for a reportee:
        confirm_by  : ISO date string — deadline to acknowledge/confirm targets
        cascade_by  : ISO date string — deadline to cascade down to their reports
        Compliance tracked as Diligence score component.
        """
        dl_key = f"deadline|{from_code}|{period}"
        self.cascade[dl_key] = {
            "type":        "deadline",
            "from_code":   from_code,
            "period":      period,
            "confirm_by":  confirm_by,
            "cascade_by":  cascade_by,
            "set_by":      set_by,
            "set_at":      datetime.now().isoformat(),
            "confirmed_at": None,
            "cascaded_at":  None,
            "confirmed":    False,
            "cascaded":     False,
        }
        self._save()

    def get_cascade_deadline(self, staff_code: str, period: str,
                               staff_name: str = ""):
        sc = clean_code(staff_code)
        # Direct lookup first
        entry = self.cascade.get(f"deadline|{sc}|{period}")
        if entry:
            return entry
        # Fallback: scan all deadline entries for this period and name
        if staff_name:
            sn = str(staff_name).strip().lower()
            for key, e in self.cascade.items():
                if key.startswith("_"): continue
                if not key.startswith("deadline|"): continue
                if e.get("period","") != period: continue
                from_c = clean_code(e.get("from_code",""))
                # Check if this deadline was SET FOR this person via
                # their to_name in any allocation
                for ak, ae in self.cascade.items():
                    if ak.startswith("_") or ak.startswith("deadline|") or ak.startswith("global_"): continue
                    for alloc in ae.get("allocations",[]):
                        to_name = str(alloc.get("to_name","")).strip().lower()
                        to_code = clean_code(alloc.get("to_code",""))
                        if (sn in to_name or to_name in sn) and to_code:
                            direct = self.cascade.get(f"deadline|{to_code}|{period}")
                            if direct:
                                return direct
        # Return safe defaults — never crash pages that expect these keys
        from datetime import timedelta as _td
        _today = datetime.now()
        return {
            "staff_code":    staff_code,
            "period":        period,
            "confirm_by":    (_today + _td(days=30)).strftime("%Y-%m-%d"),
            "cascade_by":    (_today + _td(days=14)).strftime("%Y-%m-%d"),
            "confirmed":     False,
            "cascaded":      False,
            "targets_locked":False,
            "locked_at":     "",
        }

    def mark_confirmed(self, staff_code: str, period: str):
        dl = self.cascade.get(f"deadline|{staff_code}|{period}")
        if dl:
            dl["confirmed"]    = True
            dl["confirmed_at"] = datetime.now().isoformat()
            self._save()

    def mark_cascaded(self, staff_code: str, period: str):
        dl = self.cascade.get(f"deadline|{staff_code}|{period}")
        if dl:
            dl["cascaded"]    = True
            dl["cascaded_at"] = datetime.now().isoformat()
            self._save()

    def deadline_compliance_score(self, staff_code: str, period: str) -> float:
        """0–100 score for meeting cascade deadlines. Feeds Diligence Score."""
        dl = self.get_cascade_deadline(staff_code, period)
        if not dl:
            return 100.0  # No deadline set = not penalised
        score = 0.0
        today = datetime.now().date()

        # Confirmation compliance (50 points)
        if dl.get("confirmed"):
            conf_dt = datetime.fromisoformat(dl["confirmed_at"]).date()
            due_dt  = datetime.fromisoformat(dl["confirm_by"]).date()
            score  += 50.0 if conf_dt <= due_dt else 25.0  # partial for late
        elif datetime.fromisoformat(dl["confirm_by"]).date() < today:
            score  += 0.0  # overdue and not done

        # Cascade compliance (50 points)
        if dl.get("cascaded"):
            casc_dt = datetime.fromisoformat(dl["cascaded_at"]).date()
            due_dt  = datetime.fromisoformat(dl["cascade_by"]).date()
            score  += 50.0 if casc_dt <= due_dt else 25.0
        elif datetime.fromisoformat(dl["cascade_by"]).date() < today:
            score  += 0.0

        return score

    def all_deadlines_summary(self, period: str) -> list:
        """Summary of all deadlines for a period — for MD oversight."""
        today = datetime.now().date()
        result = []
        for key, entry in self.cascade.items():
            if key.startswith("_"): continue
            if not key.startswith("deadline|"): continue
            if entry.get("period") != period: continue
            sc = entry.get("from_code", entry.get("staff_code", ""))
            _cby = entry.get("confirm_by") or entry.get("locked_at", "")
            conf_due = datetime.fromisoformat(_cby[:10]).date() if _cby else (datetime.now() + timedelta(days=30)).date()
            _dby = entry.get("cascade_by") or entry.get("locked_at", "")
            casc_due = datetime.fromisoformat(_dby[:10]).date() if _dby else (datetime.now() + timedelta(days=14)).date()
            result.append({
                "staff_code":   sc,
                "confirm_by":   str(conf_due),
                "cascade_by":   str(casc_due),
                "confirmed":    entry.get("confirmed", False),
                "cascaded":     entry.get("cascaded", False),
                "conf_overdue": not entry.get("confirmed") and conf_due < today,
                "casc_overdue": not entry.get("cascaded") and casc_due < today,
                "score":        self.deadline_compliance_score(sc, period),
            })
        return sorted(result, key=lambda x: x["conf_overdue"], reverse=True)

    # ── Global cascade timeline ──────────────────────────────────────
    def set_global_timeline(self, period: str, cascade_end_date: str,
                             levels: list, set_by: str):
        """
        MD sets the master cascade timeline.
        levels = [{"role":"Director...","confirm_by":"2026-01-10","cascade_by":"2026-01-20"}, ...]
        cascade_end_date = when cascade must reach the last level (staff)
        """
        key = f"global_timeline|{period}"
        self.cascade[key] = {
            "type":             "global_timeline",
            "period":           period,
            "cascade_end_date": cascade_end_date,
            "levels":           levels,
            "set_by":           set_by,
            "set_at":           datetime.now().isoformat(),
        }
        self._save()

    def get_global_timeline(self, period: str):
        return self.cascade.get(f"global_timeline|{period}")

    def validate_deadline_against_global(self, period: str,
                                          role: str, confirm_by: str,
                                          cascade_by: str) -> tuple:
        """
        Check that proposed deadlines fit within the global timeline.
        Returns (is_valid: bool, message: str)
        """
        tl = self.get_global_timeline(period)
        if not tl:
            return True, ""  # no global timeline set — allow anything
        end_date   = _dt.date.fromisoformat(tl["cascade_end_date"])
        conf_dt    = _dt.date.fromisoformat(confirm_by)
        casc_dt    = _dt.date.fromisoformat(cascade_by)
        today      = _dt.date.today()
        # Find expected window for this role
        for lvl in tl.get("levels", []):
            if role.lower() in lvl.get("role","").lower():
                lvl_conf = _dt.date.fromisoformat(lvl["confirm_by"])
                lvl_casc = _dt.date.fromisoformat(lvl["cascade_by"])
                if conf_dt > lvl_conf:
                    return False, (f"Confirm-by ({confirm_by}) is after the "
                                   f"expected window for {role} ({lvl['confirm_by']}). "
                                   f"Please set an earlier date.")
                if casc_dt > lvl_casc:
                    return False, (f"Cascade-by ({cascade_by}) is after the "
                                   f"expected window for {role} ({lvl['cascade_by']}). "
                                   f"Please set an earlier date.")
        if casc_dt > end_date:
            return False, (f"Cascade-by ({cascade_by}) is after the master "
                           f"cascade end date ({tl['cascade_end_date']}). "
                           f"The whole bank cascade must complete by {tl['cascade_end_date']}.")
        return True, ""

    def time_remaining_analysis(self, period: str) -> dict:
        """
        Analyse time remaining for cascade to reach last level.
        Returns dict with days_remaining, levels_pending, is_on_track.
        """
        tl = self.get_global_timeline(period)
        if not tl:
            return {"days_remaining": None, "levels_pending": 0, "is_on_track": True}
        today    = _dt.date.today()
        end_date = _dt.date.fromisoformat(tl["cascade_end_date"])
        days_rem = (end_date - today).days
        # Count how many levels still have pending cascades
        levels_done    = 0
        levels_pending = 0
        for lvl in tl.get("levels", []):
            casc_date = _dt.date.fromisoformat(lvl["cascade_by"])
            # Count staff who have cascaded from this level
            done_count = sum(1 for k, e in self.cascade.items()
                             if not k.startswith("_")
                             and not k.startswith("deadline|")
                             and not k.startswith("global_")
                             and e.get("period") == period)
            if casc_date < today:
                levels_pending += 1
            else:
                levels_done += 1
        return {
            "days_remaining":  days_rem,
            "levels_pending":  levels_pending,
            "end_date":        str(end_date),
            "is_on_track":     days_rem > 0 and levels_pending == 0,
            "is_overdue":      days_rem < 0,
        }

    # ── Review requests ───────────────────────────────────────────────
    def request_review(self, staff_code: str, staff_name: str,
                        period: str, kpi: str, reason: str, requested_target: float):
        """Staff member requests a review of their cascaded target."""
        rr_file = DATA_DIR / "cascade_review_requests.json"
        if not rr_file.exists(): rr_file.write_text("[]")
        try:
            requests = json.loads(rr_file.read_text())
        except:
            requests = []
        requests.append({
            "id":               f"RR{len(requests)+1:04d}",
            "staff_code":       staff_code,
            "staff_name":       staff_name,
            "period":           period,
            "kpi":              kpi,
            "reason":           reason,
            "requested_target": requested_target,
            "status":           "Pending",  # Pending / Approved / Rejected
            "response":         "",
            "raised_at":        datetime.now().isoformat(),
            "resolved_at":      None,
            "resolved_by":      None,
        })
        rr_file.write_text(json.dumps(requests, indent=2, default=str))

    def get_review_requests(self, period: str = None, staff_code: str = None) -> list:
        rr_file = DATA_DIR / "cascade_review_requests.json"
        if not rr_file.exists(): return []
        try:
            requests = json.loads(rr_file.read_text())
        except:
            return []
        if period:     requests = [r for r in requests if r.get("period")==period]
        if staff_code: requests = [r for r in requests if r.get("staff_code")==staff_code]
        return requests

    def resolve_review(self, rr_id: str, status: str, response: str, by: str,
                       counter_target: float = None, escalate_to: str = "",
                       escalate_to_name: str = ""):
        """Resolve a review request — supports E4 negotiation escalation.

        v10.409 — Negotiation workflow:
          status values supported:
            - "Approved" — manager approves the request as-is
            - "Rejected" — manager rejects; review closes
            - "Counter-Proposed" — manager offers counter_target; staff can accept/escalate
            - "Escalated" — request routed to escalate_to (skip-level manager)

        counter_target: If status="Counter-Proposed", the manager's
                        proposed alternative target.
        escalate_to:    If status="Escalated", staff_code of the skip-
                        level manager to route to.
        escalate_to_name: Display name of the escalation recipient.
        """
        rr_file = DATA_DIR / "cascade_review_requests.json"
        if not rr_file.exists(): return
        try:
            requests = json.loads(rr_file.read_text())
        except:
            return
        for r in requests:
            if r["id"] == rr_id:
                r["status"]       = status
                r["response"]     = response
                r["resolved_by"]  = by
                r["resolved_at"]  = datetime.now().isoformat()
                # v10.409 — record escalation chain
                if status == "Counter-Proposed" and counter_target is not None:
                    r["counter_target"] = float(counter_target)
                if status == "Escalated":
                    r["escalated_to"] = escalate_to
                    r["escalated_to_name"] = escalate_to_name
                    r["escalated_at"] = datetime.now().isoformat()
                    # An escalated review is still actionable — reopen status
                    # for the new resolver
                    r["status"] = "Pending"   # pending from escalate_to
                    r["original_status_was"] = "Escalated"
                # Append to history
                hist = r.get("history", [])
                hist.append({
                    "at": datetime.now().isoformat(),
                    "by": by,
                    "status": status,
                    "response": response,
                    "counter_target": counter_target,
                    "escalate_to": escalate_to or None,
                })
                r["history"] = hist
        rr_file.write_text(json.dumps(requests, indent=2, default=str))

    def auto_escalate_overdue_reviews(self, sla_days: int = 7) -> int:
        """v10.409 — Auto-escalate Pending reviews older than `sla_days`.

        Returns count of reviews escalated. Stamps `auto_escalated_at`
        and changes status to indicate SLA breach.
        """
        rr_file = DATA_DIR / "cascade_review_requests.json"
        if not rr_file.exists(): return 0
        try:
            requests = json.loads(rr_file.read_text())
        except:
            return 0
        cutoff = datetime.now() - timedelta(days=sla_days)
        escalated = 0
        for r in requests:
            if r.get("status") != "Pending": continue
            if r.get("auto_escalated_at"): continue   # only once
            try:
                raised = datetime.fromisoformat(r["raised_at"])
            except (ValueError, KeyError):
                continue
            if raised < cutoff:
                r["auto_escalated_at"] = datetime.now().isoformat()
                r["auto_escalation_reason"] = (
                    f"SLA breach: {sla_days} days since raised without resolution"
                )
                r["sla_breached"] = True
                escalated += 1
        if escalated > 0:
            rr_file.write_text(json.dumps(requests, indent=2, default=str))
        return escalated

    # ── Target locking (on acceptance) ────────────────────────────────
    def lock_targets(self, staff_code: str, period: str):
        """
        Lock targets for this staff member and period.
        Writes targets_locked flag to the deadline entry in target_cascade.json,
        updates locked_targets.json for quick lookup, and triggers
        inject_cascade_targets so the actuals xlsx reflects the latest targets.
        """
        sc = clean_code(staff_code)
        # 1. Mark locked in cascade deadline entry
        dl = self.cascade.get(f"deadline|{sc}|{period}")
        if dl:
            dl["targets_locked"] = True
            dl["locked_at"]      = datetime.now().isoformat()
        else:
            # Create deadline entry if missing
            self.cascade[f"deadline|{sc}|{period}"] = {
                "staff_code":     sc,
                "period":         period,
                "targets_locked": True,
                "locked_at":      datetime.now().isoformat(),
                "confirmed":      True,
                "confirmed_at":   datetime.now().isoformat(),
            }
        self._save()

        # 2. Update locked_targets.json for fast lookup
        lt_file = DATA_DIR / "locked_targets.json"
        try:
            lt_data = json.loads(lt_file.read_text()) if lt_file.exists() else {}
            if not isinstance(lt_data, dict):
                lt_data = {}
            lt_data[f"{sc}|{period}"] = True
            lt_file.write_text(json.dumps(lt_data, indent=2))
        except Exception:
            pass

        # 3. Inject latest cascade targets into actuals xlsx
        try:
            act_files = sorted(
                [f for f in DATA_DIR.glob("actuals_*.xlsx") if "backup" not in f.name],
                reverse=True)
            if act_files:
                from utils.actuals_engine import inject_cascade_targets
                inject_cascade_targets(act_files[0])
        except Exception:
            pass

    def _resolve_staff_code(self, staff_code: str, staff_name: str = "") -> str:
        """Try to resolve a username to a numeric staff code by scanning deadlines."""
        sc = clean_code(staff_code)
        # If it looks numeric already, return as-is
        if sc.isdigit():
            return sc
        # Scan deadline entries for a name match
        sn = str(staff_name).strip().lower()
        for key, entry in self.cascade.items():
            if key.startswith("_"):
                continue
            if not key.startswith("deadline|"):
                continue
            for alloc_key, alloc_entry in self.cascade.items():
                if alloc_key.startswith("_") or alloc_key.startswith("deadline|") or alloc_key.startswith("global_"):
                    continue
                for alloc in alloc_entry.get("allocations",[]):
                    to_name = str(alloc.get("to_name","")).strip().lower()
                    if sn and to_name and (sn in to_name or to_name in sn):
                        return clean_code(alloc.get("to_code", sc))
        return sc

    def targets_locked(self, staff_code: str, period: str,
                        staff_name: str = "") -> bool:
        sc = clean_code(staff_code)
        # Try direct key
        dl = self.cascade.get(f"deadline|{sc}|{period}")
        if dl:
            return bool(dl.get("targets_locked", False))
        # Fallback: search all deadline entries for name match
        if staff_name:
            sn = str(staff_name).strip().lower()
            for key, entry in self.cascade.items():
                if key.startswith("_"): continue
                if not key.startswith("deadline|"): continue
                to_name = str(entry.get("to_name","")).strip().lower()
                en_name = str(entry.get("staff_name","")).strip().lower()
                if sn and (sn in to_name or to_name in sn or
                            sn in en_name or en_name in sn):
                    return bool(entry.get("targets_locked", False))
        # Final fallback — check locked_targets.json
        # (pre-populated entries + written by lock_targets on every new lock)
        try:
            lt_file = DATA_DIR / "locked_targets.json"
            if lt_file.exists():
                lt_data = json.loads(lt_file.read_text())
                if isinstance(lt_data, dict) and lt_data.get(f"{sc}|{period}"):
                    return True
        except Exception:
            pass
        return False

    def write_targets_to_df(self, df: "pd.DataFrame", period: str) -> "pd.DataFrame":
        """
        Write accepted+locked cascade targets into the BSC dataframe.
        Replaces Annual Target with the cascaded amount for staff who have
        accepted and locked their targets.

        Called after process_kpi_data so the scorecard reflects agreed targets.
        Returns the modified dataframe.
        """
        import pandas as _pd
        df = df.copy()
        if "Annual Target" not in df.columns:
            df["Annual Target"] = 0.0

        # Iterate all cascade allocations
        for key, entry in self.cascade.items():
            if key.startswith("_") or key.startswith("deadline|") or key.startswith("global_"):
                continue
            kpi   = entry.get("kpi","")
            per   = entry.get("period","")
            if per != period:
                continue
            for alloc in entry.get("allocations",[]):
                to_code = str(alloc.get("to_code",""))
                amount  = float(alloc.get("amount",0) or 0)
                if not to_code or amount==0:
                    continue
                # Check targets are locked for this person
                if not self.targets_locked(to_code, period):
                    continue
                # Write into df where Staff Code matches and KPI matches
                mask = ((df["Staff Code"].astype(str).str.strip()==to_code) &
                        (df["KPI"]==kpi))
                if mask.any():
                    df.loc[mask,"Annual Target"] = amount
        return df

    def set_bank_target(self, kpi: str, period: str, target: float, buffer_pct: float = 0):
        """MD sets the overall bank-level target for a KPI."""
        key = f"{kpi}|{period}"
        new_entry = {
            "kpi": kpi, "period": period,
            "target": target, "buffer_pct": buffer_pct,
            "stretch_target": round(target * (1 + buffer_pct/100), 2),
            "updated_at": datetime.now().isoformat(),
            "updated_by": "MD",
        }
        # v10.343 — schema-lock check before mutating in-memory state.
        # Refuses scalar or otherwise malformed entries (the v10.337 mistake).
        try:
            from utils.schema_validator import validate_value, load_schema
            schema = load_schema("bank_targets.json")
            if schema is not None:
                # Schema applies to the OVERALL dict — we validate a single
                # entry by checking its value against additionalProperties
                entry_schema = schema.get("additionalProperties", {})
                if entry_schema:
                    result = validate_value(new_entry, entry_schema)
                    if not result.get("valid", True):
                        # Don't crash the page — log + return False, let
                        # the caller handle. Caller is the cascade UI which
                        # already shows a toast on save.
                        return False
        except Exception:
            pass  # validator unavailable — fall through to existing behaviour

        self.bank_targets[key] = new_entry
        # noqa: a2z-bootstrap-fallback — CascadeManager bootstraps state
        self.bank_file.write_text(json.dumps(self.bank_targets, indent=2, default=str))
        return True

    def get_bank_target(self, kpi: str, period: str):
        return self.bank_targets.get(f"{kpi}|{period}")

    def set_fixed_kpis(self, period: str, kpis: list, values: dict = None):
        """MD marks KPIs as fixed with their locked values.
        kpis: list of KPI names
        values: {kpi: value} dict of locked targets
        Storage: {period: {"kpis": [...], "values": {kpi: val}}}
        """
        existing = self.fixed_kpis.get(period, {})
        if isinstance(existing, list):
            existing = {"kpis": existing, "values": {}}
        entry = {"kpis": kpis, "values": values or existing.get("values", {})}
        self.fixed_kpis[period] = entry
        self.fixed_file.write_text(json.dumps(self.fixed_kpis, indent=2, default=str))

    def get_fixed_kpis(self, period: str) -> list:
        entry = self.fixed_kpis.get(period, [])
        if isinstance(entry, list): return entry          # legacy format
        return entry.get("kpis", [])

    def get_fixed_value(self, kpi: str, period: str) -> float:
        """Return the locked value for a fixed KPI, or 0 if not set."""
        entry = self.fixed_kpis.get(period, {})
        if isinstance(entry, list): entry = {}
        val = entry.get("values", {}).get(kpi, 0)
        if val: return float(val)
        # Fallback: bank_targets
        bt = self.get_bank_target(kpi, period)
        return float(bt["target"]) if bt and bt.get("target") else 0.0

    def is_fixed(self, kpi: str, period: str) -> bool:
        return kpi in self.get_fixed_kpis(period)

    def set_allocation(self, from_code: str, kpi: str, period: str,
                       allocations: list, total_target: float,
                       updated_by: str = None):
        """
        Set target allocations from one person to their direct reports.
        allocations = [{"to_code": "xxx", "to_name": "...", "amount": 1000000}, ...]

        v10.404: stamps _v10404_manual=True + updated_by so the regenerator
        preserves this entry on admin regen.
        """
        key = f"{from_code}|{kpi}|{period}"
        self.cascade[key] = {
            "from_code":    from_code,
            "kpi":          kpi,
            "period":       period,
            "total_target": total_target,
            "allocations":  allocations,
            "allocated_sum":sum(a["amount"] for a in allocations),
            "updated_at":   datetime.now().isoformat(),
            "updated_by":   updated_by or from_code,
            "_v10404_manual": True,
        }
        self._save()
        return self.cascade[key]

    def get_allocation(self, from_code: str, kpi: str, period: str):
        key = f"{from_code}|{kpi}|{period}"
        return self.cascade.get(key)

    def get_my_allocations(self, staff_code: str, period: str):
        """What has this person cascaded down to their reports?"""
        sc = clean_code(staff_code)
        return {k: v for k, v in self.cascade.items()
                if not k.startswith("_") and isinstance(v, dict)
                and v.get("from_code") == sc and v.get("period") == period}

    def get_what_i_was_given(self, staff_code: str, period: str,
                               staff_name: str = "") -> list:
        """What targets have been cascaded TO this person?
        Matches by staff_code first; falls back to staff_name if code is
        a username string (not a numeric code) or if no matches found.
        """
        sc     = clean_code(staff_code)
        sn     = str(staff_name).strip().lower()
        result = []
        for key, entry in self.cascade.items():
            # v10.409 — skip ALL non-allocation keys (meta + deadline + global)
            if key.startswith("_") or key.startswith("deadline|") or key.startswith("global_"):
                continue
            if not isinstance(entry, dict):
                continue
            if period and entry.get("period","") != period:
                continue
            for alloc in entry.get("allocations", []):
                to_code = clean_code(alloc.get("to_code",""))
                to_name = str(alloc.get("to_name","")).strip().lower()
                # Match by code OR by name (handles case where user account
                # was created with username instead of staff code)
                code_match = (to_code == sc)
                name_match = (sn and to_name and (sn in to_name or to_name in sn))
                if code_match or name_match:
                    result.append({
                        "kpi":        entry.get("kpi",""),
                        "period":     entry.get("period",""),
                        "from_code":  entry.get("from_code",""),
                        "amount":     alloc.get("amount",0),
                        "total_pool": entry.get("total_target",0),
                        "my_share":   (alloc.get("amount",0)/entry["total_target"]*100
                                      if entry.get("total_target") else 0),
                    })
        return result

    def cascade_coverage(self, from_code: str, kpi: str, period: str):
        """
        % of target that has been allocated down. 
        Returns (allocated_sum, total_target, coverage_pct, unallocated).
        """
        entry = self.get_allocation(from_code, kpi, period)
        if not entry:
            return 0, 0, 0, 0
        total = entry["total_target"]
        alloc = entry["allocated_sum"]
        cov   = round(alloc/total*100, 1) if total else 0
        return alloc, total, cov, max(0, total-alloc)


# ─── INTELLIGENT TARGET SUGGESTION ENGINE ─────────────────────────────
def suggest_target(kpi: str, staff_name: str, df_proc: "pd.DataFrame",
                   period: str = str(get_org_config().get("active_period","2026")), growth_trajectory: float = 0.0) -> dict:
    """
    Analyse historical BSC data and suggest a scientifically grounded target range.

    Returns a dict with:
      - prior_year_target    : what was set last year
      - prior_year_actual    : what they actually achieved
      - prior_year_pct       : achievement %
      - avg_achievement_2yr  : rolling 2-year average achievement (if data exists)
      - is_new_hire          : True if < 6 months data available
      - suggested_min        : conservative floor (85th-pct of historical)
      - suggested_target     : recommended target
      - suggested_stretch    : stretch / ambitious ceiling
      - rationale            : plain-English explanation
      - confidence           : 'high' / 'medium' / 'low'
    """
    result = {
        "prior_year_target":   0.0,
        "prior_year_actual":   0.0,
        "prior_year_pct":      0.0,
        "avg_achievement_2yr": 0.0,
        "is_new_hire":         False,
        "suggested_min":       0.0,
        "suggested_target":    0.0,
        "suggested_stretch":   0.0,
        "rationale":           "",
        "confidence":          "low",
    }

    if df_proc is None or (hasattr(df_proc,'empty') and df_proc.empty):
        result["rationale"] = "No historical data available."
        result["is_new_hire"] = True
        return result

    # Filter to this staff member and KPI
    staff_kpi = df_proc[(df_proc["Staff Name"] == staff_name) &
                         (df_proc["KPI"] == kpi)].copy()

    if staff_kpi.empty:
        result["rationale"] = f"{staff_name} has no historical record for {kpi}. Treat as new hire."
        result["is_new_hire"] = True
        return result

    # Detect month columns (any column containing 'Jan','Feb','Mar' etc.)
    month_cols = [c for c in df_proc.columns
                  if any(m in str(c) for m in
                         ["Jan","Feb","Mar","Apr","May","Jun",
                          "Jul","Aug","Sep","Oct","Nov","Dec"])]
    has_months = len(month_cols) > 0

    # Count months with actual data (non-zero, non-null)
    months_with_data = 0
    if has_months:
        for mc in month_cols:
            if mc in staff_kpi.columns:
                v = pd.to_numeric(staff_kpi[mc].values[0], errors='coerce')
                if not pd.isna(v) and v > 0:
                    months_with_data += 1

    # New hire detection: fewer than 6 months of data
    is_new_hire = (months_with_data < 6 and has_months) or months_with_data == 0
    result["is_new_hire"] = is_new_hire

    # Extract annual target and actuals
    row = staff_kpi.iloc[0]
    cur_target = float(pd.to_numeric(row.get("Annual Target", 0), errors="coerce") or 0)
    cur_actual  = float(pd.to_numeric(row.get("YTD_Actual",
                         row.get("Annual Actual", 0)), errors="coerce") or 0)
    cur_pct = round(cur_actual / cur_target * 100, 1) if cur_target else 0

    result["prior_year_target"] = cur_target
    result["prior_year_actual"] = cur_actual
    result["prior_year_pct"]    = cur_pct

    if is_new_hire:
        # For new hires: use role peer average as baseline
        if "Role" in df_proc.columns:
            role = row.get("Role","")
            peers = df_proc[(df_proc["Role"]==role) & (df_proc["KPI"]==kpi)]
            if len(peers) > 1:
                peer_tgts = pd.to_numeric(peers["Annual Target"], errors="coerce").dropna()
                peer_acts = pd.to_numeric(peers.get("YTD_Actual",
                                          peers.get("Annual Actual", peer_tgts)),
                                          errors="coerce").dropna()
                avg_peer_tgt = float(peer_tgts.mean()) if len(peer_tgts) else cur_target
                avg_peer_ach = float(peer_acts.mean()) if len(peer_acts) else avg_peer_tgt * 0.85

                # New hire gets 70% of peer average target (ramp-up period)
                ramp = 0.70
                suggested = round(avg_peer_tgt * ramp * (1 + growth_trajectory / 100), 2)
                result.update({
                    "suggested_min":     round(suggested * 0.85, 2),
                    "suggested_target":  suggested,
                    "suggested_stretch": round(suggested * 1.20, 2),
                    "avg_achievement_2yr": round(avg_peer_ach / avg_peer_tgt * 100, 1)
                                          if avg_peer_tgt else 0,
                    "rationale": (
                        f"New hire — using {ramp*100:.0f}% of peer average target "
                        f"(KES {avg_peer_tgt:,.0f}) as ramp-up baseline. "
                        f"Peer average achievement: {result['avg_achievement_2yr']:.1f}%. "
                        f"Growth trajectory: {growth_trajectory:+.1f}%."
                        if growth_trajectory else
                        f"New hire — using {ramp*100:.0f}% of peer average target as ramp-up baseline."
                    ),
                    "confidence": "medium",
                })
                return result

        # No peers found — return current target at 80%
        result.update({
            "suggested_min":     round(cur_target * 0.65, 2),
            "suggested_target":  round(cur_target * 0.80, 2),
            "suggested_stretch": round(cur_target * 1.00, 2),
            "rationale": "New hire with no peer benchmark. Suggest 80% of current target as ramp-up.",
            "confidence": "low",
        })
        return result

    # EXISTING STAFF — science-based target setting
    # Achievement tier logic:
    #   >120% repeatedly → raise target aggressively
    #   90-120% → raise moderately by trajectory + small push
    #   70-90%  → hold or small increase (don't over-push)
    #   <70%    → diagnose before raising; suggest hold or slight increase

    traj_factor = 1 + (growth_trajectory / 100)
    confidence  = "high" if cur_pct > 0 else "low"

    if cur_pct >= 120:
        # Consistently over-delivering — raise meaningfully
        push = 1.15 * traj_factor
        rationale = (
            f"Achieved {cur_pct:.1f}% of target — significantly exceeding. "
            f"Recommend +15% above prior target (+ {growth_trajectory:+.1f}% trajectory). "
            f"Raising the ceiling will maintain motivation and grow the business."
        )
    elif cur_pct >= 100:
        push = 1.08 * traj_factor
        rationale = (
            f"Achieved {cur_pct:.1f}% — on or slightly above target. "
            f"Recommend +8% stretch (+ {growth_trajectory:+.1f}% trajectory). "
            f"Consistent achiever; push gently upward."
        )
    elif cur_pct >= 90:
        push = 1.04 * traj_factor
        rationale = (
            f"Achieved {cur_pct:.1f}% — near target. "
            f"Recommend +4% nudge (+ {growth_trajectory:+.1f}% trajectory). "
            f"Close to target; maintain pressure with modest increase."
        )
    elif cur_pct >= 70:
        push = 1.00 * traj_factor
        rationale = (
            f"Achieved {cur_pct:.1f}% — below target. "
            f"Recommend holding target flat (+ trajectory only: {growth_trajectory:+.1f}%). "
            f"Focus on closing the current gap before raising the ceiling."
        )
    else:
        push = 0.95 * traj_factor
        rationale = (
            f"Achieved only {cur_pct:.1f}% — well below target. "
            f"Recommend reducing target slightly (-5%) to set a realistic but achievable goal. "
            f"Investigate root cause before escalating targets."
        )
        confidence = "medium"

    suggested_tgt = round(cur_target * push, 2)
    result.update({
        "suggested_min":     round(cur_target * max(push - 0.05, 0.85), 2),
        "suggested_target":  suggested_tgt,
        "suggested_stretch": round(cur_target * (push + 0.10), 2),
        "avg_achievement_2yr": cur_pct,
        "rationale": rationale,
        "confidence": confidence,
    })
    return result


def get_bank_growth_trajectory(kpi: str, bank_targets_dict: dict) -> float:
    """
    Estimate the bank's YoY growth trajectory for a KPI from saved bank targets.
    Returns a % e.g. 15.0 means 15% growth expected.
    """
    # Find targets for this KPI across multiple periods
    kpi_targets = {k.split("|")[1]: v["target"]
                   for k, v in bank_targets_dict.items()
                   if k.startswith(f"{kpi}|") and v.get("target", 0) > 0}
    if len(kpi_targets) < 2:
        return 10.0  # Default 10% growth trajectory when no history

    periods = sorted(kpi_targets.keys())
    t_prev  = kpi_targets[periods[-2]]
    t_curr  = kpi_targets[periods[-1]]
    if t_prev <= 0: return 10.0
    return round((t_curr - t_prev) / t_prev * 100, 1)


# ─── VALIDATION MANAGER (performance sign-off) ───────────────────────
class ValidationManager:
    """Manager sign-off on staff performance per period."""
    def __init__(self):
        self.file    = DATA_DIR / "validations.json"
        self.records = self._load()

    def _load(self):
        if not self.file.exists(): self.file.write_text("{}")
        try:
            raw = self.file.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _save(self):
        self.file.write_text(json.dumps(self.records, indent=2, default=str))

    def validate(self, manager: str, staff_name: str, period: str,
                 status: str, action_plan: str = '', comments: str = ''):
        key = f"{staff_name}|{period}"
        self.records[key] = {
            "staff_name":   staff_name,
            "period":       period,
            "manager":      manager,
            "status":       status,
            "action_plan":  action_plan,
            "comments":     comments,
            "validated_at": datetime.now().isoformat(),
        }
        self._save()
        return self.records[key]

    def get(self, staff_name: str, period: str):
        return self.records.get(f"{staff_name}|{period}")

    def get_period_summary(self, period: str):
        return {k:v for k,v in self.records.items() if v.get('period')==period}


# ─── REPORTING LINE MANAGER ───────────────────────────────────────────
class ReportingLineManager:
    """
    Stores admin overrides to the reporting hierarchy.
    Overrides take precedence over the uploaded spreadsheet's 'Reports To Code'.
    Supports: remap staff → new manager, transfer staff between units,
    bulk unit reassignment, and full org-tree rebuild.
    """
    def __init__(self):
        self.file      = DATA_DIR / "reporting_lines.json"
        self.overrides = self._load()   # {staff_code: {manager_code, unit, region, updated_by, updated_at}}
        self.unit_map  = DATA_DIR / "unit_map.json"
        self.units     = self._load_units()  # {staff_code: {unit, region}}

    def _load(self):
        if not self.file.exists(): self.file.write_text("{}")
        try:
            raw = self.file.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _load_units(self):
        if not self.unit_map.exists(): self.unit_map.write_text("{}")
        try:
            raw = self.unit_map.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _save(self):
        self.file.write_text(json.dumps(self.overrides, indent=2, default=str))

    def _save_units(self):
        self.unit_map.write_text(json.dumps(self.units, indent=2, default=str))

    def remap(self, staff_code: str, new_manager_code: str,
              updated_by: str, reason: str = ''):
        """Remap a staff member to a new line manager."""
        sc = clean_code(staff_code)
        self.overrides[sc] = {
            'manager_code': clean_code(new_manager_code),
            'reason':       reason,
            'updated_by':   updated_by,
            'updated_at':   datetime.now().isoformat(),
        }
        self._save()

    def transfer(self, staff_code: str, new_unit: str, new_region: str,
                 new_manager_code: str, updated_by: str, reason: str = ''):
        """Transfer staff to a new unit + new manager."""
        sc = clean_code(staff_code)
        self.overrides[sc] = {
            'manager_code': clean_code(new_manager_code),
            'reason':       reason,
            'updated_by':   updated_by,
            'updated_at':   datetime.now().isoformat(),
        }
        self.units[sc] = {
            'unit':       new_unit,
            'region':     new_region,
            'updated_by': updated_by,
            'updated_at': datetime.now().isoformat(),
        }
        self._save()
        self._save_units()

    def bulk_remap_unit(self, old_manager_code: str, new_manager_code: str,
                        staff_codes: list, updated_by: str, reason: str = ''):
        """Remap multiple staff from one manager to another."""
        for sc in staff_codes:
            self.remap(sc, new_manager_code, updated_by, reason)

    def clear_override(self, staff_code: str, updated_by: str):
        """Remove override — staff reverts to uploaded data."""
        sc = clean_code(staff_code)
        if sc in self.overrides:
            self.overrides.pop(sc)
            self._save()
        if sc in self.units:
            self.units.pop(sc)
            self._save_units()

    def get_manager(self, staff_code: str) -> str:
        """Return overridden manager code, or None if not overridden."""
        return self.overrides.get(clean_code(staff_code), {}).get('manager_code')

    def get_unit(self, staff_code: str):
        """Return overridden unit info, or None."""
        return self.units.get(clean_code(staff_code))

    def apply_to_registry(self, registry_df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply all overrides to a staff registry DataFrame.
        Returns enriched DataFrame with overrides applied.
        """
        df = registry_df.copy()
        sc_col = 'Staff Code' if 'Staff Code' in df.columns else None
        if sc_col is None:
            return df
        df['Staff Code'] = df['Staff Code'].astype(str).str.strip()

        for sc, ov in self.overrides.items():
            mask = df['Staff Code'] == sc
            if mask.any() and 'Reports_To' in df.columns:
                df.loc[mask, 'Reports_To'] = ov['manager_code']

        for sc, unit_ov in self.units.items():
            mask = df['Staff Code'] == sc
            if mask.any():
                df.loc[mask, 'Unit']   = unit_ov['unit']
                df.loc[mask, 'Region'] = unit_ov['region']
        return df

    def get_all_overrides(self) -> list:
        """Return list of all current overrides for display."""
        result = []
        for sc, ov in self.overrides.items():
            rec = {'staff_code': sc, **ov}
            if sc in self.units:
                rec.update(self.units[sc])
            result.append(rec)
        return result

    def get_direct_reports(self, manager_code: str, registry_df: pd.DataFrame) -> list:
        """
        Return staff codes that report to this manager,
        combining uploaded data + overrides.
        """
        mc = clean_code(manager_code)
        applied = self.apply_to_registry(registry_df)
        rt_col = 'Reports_To' if 'Reports_To' in applied.columns else None
        if rt_col is None:
            return []
        mask = applied[rt_col].astype(str).str.strip() == mc
        return applied[mask]['Staff Code'].tolist()

    def get_org_tree(self, registry_df: pd.DataFrame) -> dict:
        """
        Build full org tree as nested dict {manager_code: [direct_report_codes]}.
        """
        applied = self.apply_to_registry(registry_df)
        tree = {}
        rt_col = 'Reports_To' if 'Reports_To' in applied.columns else None
        if rt_col is None:
            return {}
        for _, row in applied.iterrows():
            mgr = str(row.get(rt_col,'')).strip()
            sc  = str(row.get('Staff Code','')).strip()
            if mgr and mgr not in ('nan','None',''):
                tree.setdefault(mgr, []).append(sc)
        return tree

    def summary(self) -> dict:
        return {
            'total_overrides':  len(self.overrides),
            'total_transfers':  len(self.units),
            'last_updated':     max((v['updated_at'] for v in self.overrides.values()),
                                     default='Never'),
        }

# ─── PIPELINE MANAGER ────────────────────────────────────────────────
class PipelineManager:
    def __init__(self):
        self.deals_file      = DATA_DIR / "pipeline_deals.json"
        self.activities_file = DATA_DIR / "pipeline_activities.json"
        for f in [self.deals_file, self.activities_file]:
            if not f.exists(): f.write_text("[]")
        self.deals      = self._load(self.deals_file)
        self.activities = self._load(self.activities_file)

    def _load(self, f):
        try:
            raw = f.read_text()
            d = json.loads(raw) if raw.strip() else []
            return d if isinstance(d, list) else []
        except: return []

    def _save_deals(self):
        self.deals_file.write_text(json.dumps(self.deals, indent=2))

    def _save_activities(self):
        self.activities_file.write_text(json.dumps(self.activities, indent=2))

    def add_deal(self, d):
        d['id']         = f"D{len(self.deals)+1:04d}"
        d['created_at'] = datetime.now().isoformat()
        d['updated_at'] = datetime.now().isoformat()
        # Batch A: stamp open_date so new deals have a real date for the
        # open_date-DESC list ordering (was unset -> created deals sorted oddly).
        d.setdefault('open_date', datetime.now().strftime('%Y-%m-%d'))
        d['staff_code'] = clean_code(d.get('staff_code',''))
        self.deals.append(d)
        self._save_deals()
        return d['id']

    def update_stage(self, deal_id, new_stage, note, updated_by):
        for d in self.deals:
            if d['id'] == deal_id:
                old_stage = d['stage']
                d['stage']      = new_stage
                d['updated_at'] = datetime.now().isoformat()
                d['updated_by'] = updated_by
                if new_stage == 'Closed Won':
                    d['closed_date'] = str(datetime.now().date())
                self.add_activity({
                    'deal_id': deal_id, 'staff_code': d['staff_code'],
                    'staff_name': d['staff_name'], 'activity_type': 'Stage Change',
                    'note': f"Stage: {old_stage} → {new_stage}. {note}",
                    'outcome': new_stage,
                })
                break
        self._save_deals()

    def add_activity(self, a):
        a['id']          = f"A{len(self.activities)+1:04d}"
        a['recorded_at'] = datetime.now().isoformat()
        a['staff_code']  = clean_code(a.get('staff_code',''))
        self.activities.append(a)
        self._save_activities()
        return a['id']

    def get_deals(self, staff_code=None, stage=None, active_only=False):
        result = self.deals
        if staff_code: result = [d for d in result if d['staff_code'] == clean_code(staff_code)]
        if stage:      result = [d for d in result if d['stage'] == stage]
        if active_only:result = [d for d in result if d['stage'] in ACTIVE_STAGES]
        return result

    def get_activities(self, staff_code=None, deal_id=None, limit=50):
        result = self.activities
        if staff_code: result = [a for a in result if a['staff_code'] == clean_code(staff_code)]
        if deal_id:    result = [a for a in result if a.get('deal_id') == deal_id]
        return list(reversed(result))[:limit]

    def update_deal(self, deal_id: str, updates: dict, updated_by: str):
        """Edit deal fields. Logs change as an activity."""
        for d in self.deals:
            if d['id'] == deal_id:
                changed = []
                for k, v in updates.items():
                    if d.get(k) != v:
                        changed.append(f"{k}: {d.get(k)} → {v}")
                    d[k] = v
                d['updated_at'] = datetime.now().isoformat()
                d['updated_by'] = updated_by
                if changed:
                    self.add_activity({
                        'deal_id': deal_id,
                        'staff_code': d['staff_code'],
                        'staff_name': d['staff_name'],
                        'activity_type': 'Deal Updated',
                        'note': 'Fields updated: ' + ', '.join(changed),
                        'outcome': 'Updated',
                    })
                break
        self._save_deals()

    def get_deal(self, deal_id: str):
        """Get a single deal by ID."""
        return next((d for d in self.deals if d['id'] == deal_id), None)

    def delete_deal(self, deal_id: str, deleted_by: str):
        """Soft delete — mark as Closed Lost with reason 'Deleted'."""
        for d in self.deals:
            if d['id'] == deal_id:
                d['stage'] = 'Closed Lost'
                d['updated_at'] = datetime.now().isoformat()
                d['updated_by'] = deleted_by
                d['loss_reason'] = 'Deleted / Cancelled'
                d['notes'] = (d.get('notes','') + ' [Deleted by ' + deleted_by + ']').strip()
                break
        self._save_deals()

    def request_cancel(self, deal_id: str, requested_by: str, reason: str):
        """Request cancellation — for deals beyond Lead stage needs manager approval."""
        for d in self.deals:
            if d['id'] == deal_id:
                d['cancel_requested'] = True
                d['cancel_requested_by'] = requested_by
                d['cancel_requested_at'] = datetime.now().isoformat()
                d['cancel_reason'] = reason
                d['updated_at'] = datetime.now().isoformat()
                break
        self._save_deals()

    def approve_cancel(self, deal_id: str, approved_by: str, approve: bool, note: str = ""):
        """Manager approves or rejects a cancellation request."""
        for d in self.deals:
            if d['id'] == deal_id:
                if approve:
                    d['stage'] = 'Closed Lost'
                    d['loss_reason'] = 'Cancelled — ' + d.get('cancel_reason','')
                    d['updated_by'] = approved_by
                d['cancel_approved'] = approve
                d['cancel_approved_by'] = approved_by
                d['cancel_approved_at'] = datetime.now().isoformat()
                d['cancel_note'] = note
                d['updated_at'] = datetime.now().isoformat()
                break
        self._save_deals()

    def validate_deal(self, deal_id: str, validated_by: str, approved: bool, note: str = ""):
        """Manager validates a deal at Contacted stage before it counts in forecast."""
        for d in self.deals:
            if d['id'] == deal_id:
                d['manager_validated'] = approved
                d['validated_by'] = validated_by
                d['validated_at'] = datetime.now().isoformat()
                d['validation_note'] = note
                d['updated_at'] = datetime.now().isoformat()
                break
        self._save_deals()

    def get_pending_validations(self, manager_codes: set = None):
        """Deals at Contacted+ stage that need manager validation."""
        idx = STAGE_NAMES.index(PIPELINE_VALIDATE_STAGE) if PIPELINE_VALIDATE_STAGE in STAGE_NAMES else 1
        # STAGE_NAMES[idx:] runs to the end of the list, which includes the
        # terminal stages (Closed Won / Closed Lost). A closed deal never needs
        # validation, so intersect with ACTIVE_STAGES to drop terminal deals.
        result = [d for d in self.deals
                  if d['stage'] in STAGE_NAMES[idx:]
                  and d['stage'] in ACTIVE_STAGES
                  and not d.get('manager_validated')
                  and not d.get('cancel_requested')
                  and d.get('referral_status') not in ('pending', 'declined')]
        if manager_codes:
            result = [d for d in result if d.get('staff_code','') in manager_codes]
        return result

    def get_cancel_requests(self, manager_codes: set = None):
        """Deals with pending cancellation requests."""
        result = [d for d in self.deals if d.get('cancel_requested') and not d.get('cancel_approved')]
        if manager_codes:
            result = [d for d in result if d.get('staff_code','') in manager_codes]
        return result

    def get_actions_due(self, staff_code=None, days_window=0):
        """Deals with next_action_date on or before today + days_window."""
        cutoff = str((datetime.now() + timedelta(days=days_window)).date())
        today  = str(datetime.now().date())
        result = [d for d in self.deals
                  if d['stage'] in ACTIVE_STAGES
                  and d.get('next_action_date','') <= cutoff]
        if staff_code:
            result = [d for d in result if d['staff_code'] == clean_code(staff_code)]
        return result

    def pipeline_value(self, deals):
        return sum(float(d.get('deal_value', 0)) for d in deals if d['stage'] in ACTIVE_STAGES)

    def weighted_pipeline(self, deals):
        weights = {
            'Lead':0.05,'Contacted':0.10,'Qualified':0.25,'Proposal':0.40,
            'Negotiation':0.60,'Compliance':0.80,'Closed Won':1.0,'Closed Lost':0.0
        }
        return sum(float(d.get('deal_value',0)) * weights.get(d['stage'],0) for d in deals)


# ─── EXECUTE MANAGER ─────────────────────────────────────────────────
class ExecuteManager:
    def __init__(self):
        self.init_file    = DATA_DIR / "execute_initiatives.json"
        self.ideas_file   = DATA_DIR / "execute_ideas.json"
        self.ws_file      = DATA_DIR / "execute_workstreams.json"
        self.impact_file  = DATA_DIR / "execute_impact.json"
        for f in [self.init_file, self.ideas_file, self.ws_file, self.impact_file]:
            if not f.exists(): f.write_text("[]" if f != self.ws_file else "{}")
        self.initiatives  = self._load_list(self.init_file)
        self.ideas        = self._load_list(self.ideas_file)
        self.workstreams  = self._load_dict(self.ws_file)
        self.impact_data  = self._load_list(self.impact_file)

    def _load_list(self, f):
        try:
            raw = f.read_text()
            d = json.loads(raw) if raw.strip() else []
            return d if isinstance(d, list) else []
        except: return []

    def _load_dict(self, f):
        try:
            raw = f.read_text()
            d = json.loads(raw) if raw.strip() else {}
            return d if isinstance(d, dict) else {}
        except: return {}

    def _save_initiatives(self):
        self.init_file.write_text(json.dumps(self.initiatives, indent=2))

    def _save_ideas(self):
        self.ideas_file.write_text(json.dumps(self.ideas, indent=2))

    def _save_workstreams(self):
        self.ws_file.write_text(json.dumps(self.workstreams, indent=2))

    def _save_impact(self):
        self.impact_file.write_text(json.dumps(self.impact_data, indent=2))

    # ── WORKSTREAM MANAGEMENT ─────────────────────────────────────
    def upsert_workstream(self, ws_id, name, sponsor, sub_workstreams=None):
        self.workstreams[ws_id] = {
            'name': name, 'sponsor': sponsor,
            'sub_workstreams': sub_workstreams or [],
            'updated_at': datetime.now().isoformat()
        }
        self._save_workstreams()

    # ── INITIATIVE CRUD ───────────────────────────────────────────
    def create_initiative(self, data):
        init_id = f"INI{len(self.initiatives)+1:04d}"
        init = {
            'id':               init_id,
            'name':             data['name'],
            'objective':        data['objective'],
            'category':         data['category'],
            'workstream':       data['workstream'],
            'sub_workstream':   data.get('sub_workstream',''),
            'io':               data['io'],           # Initiative Owner
            'io_backup':        data.get('io_backup',''),
            'gate':             'G0',
            'gate_history':     [{'gate':'G0','date': str(datetime.now().date()),
                                  'by': data['created_by'], 'note':'Created'}],
            'approvals':        {},   # gate_key → {approver: {status, date, note}}
            'business_case':    {},
            'milestones':       [],
            'milestone_confirmations': {},  # milestone_id → {owner: confirmed bool}
            'impact_kpis':      [],
            'monthly_impacts':  [],
            'created_by':       data['created_by'],
            'created_at':       datetime.now().isoformat(),
            'updated_at':       datetime.now().isoformat(),
            'status':           'Active',
            'priority_score':   None,
            'estimated_impact': data.get('estimated_impact', 0),
            'tags':             data.get('tags', []),
        }
        self.initiatives.append(init)
        self._save_initiatives()
        return init_id

    def get_initiative(self, init_id):
        return next((i for i in self.initiatives if i['id'] == init_id), None)

    def update_initiative(self, init_id, updates):
        for i in self.initiatives:
            if i['id'] == init_id:
                i.update(updates)
                i['updated_at'] = datetime.now().isoformat()
                break
        self._save_initiatives()

    # ── GATE TRANSITIONS ──────────────────────────────────────────
    def submit_for_gate(self, init_id, target_gate, submitted_by, note=''):
        init = self.get_initiative(init_id)
        if not init: return False, "Initiative not found"
        current = init['gate']
        transition_key = f"{current}→{target_gate}"
        # Mark as pending approval
        if 'pending_gate' not in init:
            init['pending_gate'] = {}
        init['pending_gate'] = {
            'target': target_gate, 'submitted_by': submitted_by,
            'submitted_at': datetime.now().isoformat(), 'note': note,
            'approvals': {r: {'status':'Pending','date':None,'note':''}
                          for r in GATE_APPROVERS.get(transition_key, [])}
        }
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()
        return True, f"Submitted to {target_gate} — awaiting approvals"

    def approve_gate(self, init_id, approver_role, approver_name, approved, note=''):
        init = self.get_initiative(init_id)
        if not init or 'pending_gate' not in init: return False, "No pending gate"
        pg = init['pending_gate']
        if approver_role not in pg['approvals']: return False, "You are not an approver for this gate"
        pg['approvals'][approver_role] = {
            'status': 'Approved' if approved else 'Rejected',
            'by': approver_name, 'date': str(datetime.now().date()), 'note': note
        }
        # Check if all approved
        statuses = [v['status'] for v in pg['approvals'].values()]
        if all(s == 'Approved' for s in statuses):
            # Advance gate
            new_gate = pg['target']
            init['gate'] = new_gate
            init['gate_history'].append({
                'gate': new_gate, 'date': str(datetime.now().date()),
                'by': approver_name, 'note': f"All approvers confirmed. {note}"
            })
            init['approvals'][f"{init['gate_history'][-2]['gate']}→{new_gate}"] = pg['approvals']
            init.pop('pending_gate', None)
        elif 'Rejected' in statuses:
            init['gate_history'].append({
                'gate': init['gate'], 'date': str(datetime.now().date()),
                'by': approver_name, 'note': f"REJECTED at {pg['target']}. {note}"
            })
            init.pop('pending_gate', None)
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()
        return True, "Approval recorded"

    # ── BUSINESS CASE ─────────────────────────────────────────────
    def save_business_case(self, init_id, bc):
        init = self.get_initiative(init_id)
        if not init: return
        init['business_case'] = bc
        init['priority_score']   = bc.get('priority_score')
        init['estimated_impact'] = bc.get('estimated_impact', 0)
        init['impact_kpis']      = bc.get('impact_kpis', [])
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()

    # ── MILESTONES ────────────────────────────────────────────────
    def add_milestone(self, init_id, milestone):
        init = self.get_initiative(init_id)
        if not init: return None
        ms_id = f"MS{len(init['milestones'])+1:03d}"
        ms = {
            'id':              ms_id,
            'name':            milestone['name'],
            'type':            milestone['type'],
            'owner':           milestone['owner'],
            'owner_workstream': milestone.get('owner_workstream', ''),  # WS of owner if cross-functional
            'depends_on_workstream': milestone.get('depends_on_workstream', ''),  # cross-WS dependency
            'depends_on_description': milestone.get('depends_on_description', ''),
            'co_owners':       milestone.get('co_owners', []),
            'due_date':        milestone['due_date'],
            'start_date':      milestone.get('start_date', str(datetime.now().date())),
            'description':     milestone.get('description', ''),
            'dependencies':    milestone.get('dependencies', []),  # list of ms_ids
            'status':          'Not Started',
            'completion_date': None,
            'confirmed':       False,
            'confirmed_at':    None,
            'notes':           [],
            'delay_reason':    '',
            'blockers':        [],   # list of {blocker, raised_by, raised_at, resolved}
            'escalation_level': 0,   # 0=none 1=IO 2=lead 3=sponsor
            'escalation_history': [],
            'last_update_by':  '',
            'last_update_at':  '',
        }
        init['milestones'].append(ms)
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()
        return ms_id

    def confirm_milestone(self, init_id, ms_id, owner_name):
        init = self.get_initiative(init_id)
        if not init: return
        for ms in init['milestones']:
            if ms['id'] == ms_id:
                ms['confirmed']   = True
                ms['confirmed_at'] = datetime.now().isoformat()
                break
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()

    def update_milestone_status(self, init_id, ms_id, status, note='',
                                 delay_reason='', delay_category='',
                                 updated_by='', started=None):
        """
        started: True/False/None — whether the milestone has physically started.
        delay_category: one of DELAY_CATEGORIES — drives escalation tier.
        """
        init = self.get_initiative(init_id)
        if not init: return
        for ms in init['milestones']:
            if ms['id'] == ms_id:
                prev_status           = ms.get('status')
                ms['status']          = status
                ms['last_update_by']  = updated_by
                ms['last_update_at']  = datetime.now().isoformat()

                # Started flag
                if started is True and not ms.get('actual_start_date'):
                    ms['actual_start_date'] = str(datetime.now().date())
                    ms['has_started'] = True
                elif started is False:
                    ms['has_started'] = False

                if status == 'Complete':
                    ms['completion_date'] = str(datetime.now().date())
                    ms['delay_reason']    = ''
                    ms['delay_category']  = ''
                if status == 'Delayed':
                    if delay_reason:   ms['delay_reason']   = delay_reason
                    if delay_category: ms['delay_category'] = delay_category

                if note:
                    ms['notes'].append({
                        'date': str(datetime.now().date()),
                        'note': note, 'by': updated_by,
                        'status_was': prev_status, 'status_now': status,
                    })

                # Recompute escalation
                ms['escalation_level'] = ExecuteManager._escalation_level(ms)

                # Log to escalation history if level changed
                new_esc = ms['escalation_level']
                if new_esc > 0:
                    prev_esc = ms.get('escalation_history',[-1])
                    prev_level = prev_esc[-1]['level'] if prev_esc else 0
                    if new_esc != prev_level:
                        if 'escalation_history' not in ms: ms['escalation_history'] = []
                        ms['escalation_history'].append({
                            'level': new_esc,
                            'label': ESC_CONFIG.get(new_esc,{}).get('label',''),
                            'reason': f"Status: {status}" + (f" | {delay_category}" if delay_category else ""),
                            'date': str(datetime.now().date()),
                            'by': updated_by,
                        })
                break
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()

    def raise_blocker(self, init_id, ms_id, blocker_text, raised_by):
        init = self.get_initiative(init_id)
        if not init: return
        for ms in init['milestones']:
            if ms['id'] == ms_id:
                if 'blockers' not in ms: ms['blockers'] = []
                ms['blockers'].append({
                    'blocker':    blocker_text,
                    'raised_by':  raised_by,
                    'raised_at':  datetime.now().isoformat(),
                    'resolved':   False,
                    'resolved_at': None,
                    'resolution': '',
                })
                ms['escalation_level'] = max(ms.get('escalation_level',0), 2)
                if 'escalation_history' not in ms: ms['escalation_history'] = []
                ms['escalation_history'].append({
                    'level': 2, 'reason': f'Blocker raised: {blocker_text}',
                    'date': str(datetime.now().date()), 'by': raised_by,
                })
                break
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()

    def resolve_blocker(self, init_id, ms_id, blocker_idx, resolution, resolved_by):
        init = self.get_initiative(init_id)
        if not init: return
        for ms in init['milestones']:
            if ms['id'] == ms_id and blocker_idx < len(ms.get('blockers',[])):
                ms['blockers'][blocker_idx].update({
                    'resolved': True,
                    'resolved_at': datetime.now().isoformat(),
                    'resolution': resolution,
                })
                break
        init['updated_at'] = datetime.now().isoformat()
        self._save_initiatives()

    @staticmethod
    def _escalation_level(ms):
        """
        Banking escalation — tight timelines:
          0 On track        : all good
          1 Due soon        : due ≤ 2 days (email notification)
          2 Overdue / IO    : 1-2 days overdue
          3 Lead escalated  : 3-7 days overdue OR any open blocker
          4 Sponsor critical: >7 days overdue OR structural/regulatory delay category
          5 Not started     : start_date passed, status still Not Started
        """
        if ms.get('status') == 'Complete': return 0

        today = date.today()

        # Not started alert — start date has passed
        if ms.get('status') == 'Not Started' and ms.get('start_date'):
            try:
                start = date.fromisoformat(ms['start_date'])
                if start < today: return 5
            except: pass

        # Structural / regulatory delay → immediate Sponsor escalation
        delay_cat = ms.get('delay_category', '')
        if delay_cat in STRUCTURAL_CATEGORIES and ms.get('status') == 'Delayed':
            return 4

        # Open blockers → Lead escalation (at minimum)
        has_blocker = ms.get('blockers') and any(not b['resolved'] for b in ms['blockers'])

        try:
            due = date.fromisoformat(ms.get('due_date',''))
            days_overdue = (today - due).days
            days_to_due  = (due - today).days
        except:
            return 3 if has_blocker else 0

        if days_overdue > 7:  return 4  # Critical — Sponsor
        if days_overdue >= 3: return 3  # Lead escalation
        if days_overdue >= 1: return 2  # IO notified immediately
        if has_blocker:       return 3  # Blocker regardless of due date
        if days_to_due <= 2:  return 1  # Due soon — email alert
        return 0

    @staticmethod
    def _needs_start_alert(ms):
        """True if milestone should have started but hasn't."""
        if ms.get('status') != 'Not Started': return False
        if not ms.get('start_date'): return False
        try:
            start = date.fromisoformat(ms['start_date'])
            return start <= date.today()
        except: return False

    def all_milestones_confirmed(self, init_id):
        init = self.get_initiative(init_id)
        if not init or not init['milestones']: return False
        return all(ms['confirmed'] for ms in init['milestones'])

    def money_step_complete(self, init_id):
        init = self.get_initiative(init_id)
        if not init: return False
        money_steps = [ms for ms in init['milestones'] if ms['type'] == 'Money Step']
        return money_steps and all(ms['status'] == 'Complete' for ms in money_steps)

    # ── IMPACT TRACKING ───────────────────────────────────────────
    def record_impact(self, init_id, month, kpi_name, target, actual, note=''):
        entry = {
            'initiative_id': init_id,
            'month':   month,
            'kpi':     kpi_name,
            'target':  target,
            'actual':  actual,
            'note':    note,
            'recorded_at': datetime.now().isoformat(),
        }
        self.impact_data.append(entry)
        self._save_impact()

    def get_impact(self, init_id):
        return [e for e in self.impact_data if e['initiative_id'] == init_id]

    # ── IDEAS ─────────────────────────────────────────────────────
    def submit_idea(self, data):
        idea_id = f"IDEA{len(self.ideas)+1:04d}"
        idea = {
            'id':          idea_id,
            'title':       data['title'],
            'description': data['description'],
            'submitted_by': data['submitted_by'],
            'workstream':  data.get('workstream',''),
            'status':      'Submitted',   # Submitted / Under Review / Adopted / Declined
            'votes':       [],
            'comments':    [],
            'created_at':  datetime.now().isoformat(),
            'adopted_as':  None,  # initiative ID if adopted
        }
        self.ideas.append(idea)
        self._save_ideas()
        return idea_id

    def vote_idea(self, idea_id, voter):
        for idea in self.ideas:
            if idea['id'] == idea_id:
                if voter not in idea['votes']:
                    idea['votes'].append(voter)
                    self._save_ideas()
                break

    # ── QUERIES ───────────────────────────────────────────────────
    def get_initiatives(self, workstream=None, gate=None, io=None, status='Active'):
        result = [i for i in self.initiatives if i.get('status','Active') == status or status == 'All']
        if workstream: result = [i for i in result if i.get('workstream') == workstream]
        if gate:       result = [i for i in result if i.get('gate') == gate]
        if io:         result = [i for i in result if i.get('io') == io or i.get('io_backup') == io]
        return result

    def gate_counts(self):
        counts = {g: 0 for g in GATE_ORDER}
        for i in self.initiatives:
            g = i.get('gate','G0')
            if g in counts: counts[g] += 1
        return counts

    def get_my_actions(self, username, role):
        """Items requiring action from this user."""
        actions = []
        role_lower = role.lower()
        for init in self.initiatives:
            pg = init.get('pending_gate')
            if pg:
                role_map = {
                    'workstream_lead': ['workstream lead','head of','branch manager','department head'],
                    'sponsor':         ['director','sponsor'],
                    'finance':         ['finance'],
                }
                for approver_role, role_keywords in role_map.items():
                    if (approver_role in pg.get('approvals',{}) and
                        pg['approvals'][approver_role]['status'] == 'Pending' and
                        any(kw in role_lower for kw in role_keywords)):
                        actions.append({
                            'type':         'gate_approval',
                            'init_id':       init['id'],
                            'name':         init['name'],
                            'gate':         pg['target'],
                            'approver_role': approver_role,
                            'workstream':   init.get('workstream',''),
                        })
            for ms in init.get('milestones',[]):
                is_owner = (ms.get('owner','') == username or username in ms.get('co_owners',[]))
                if not is_owner: continue
                # Confirmation needed
                if not ms['confirmed']:
                    actions.append({
                        'type': 'milestone_confirm', 'init_id': init['id'],
                        'ms_id': ms['id'], 'name': init['name'],
                        'ms_name': ms['name'], 'due_date': ms.get('due_date',''),
                        'workstream': init.get('workstream',''),
                    })
                # Due soon alerts
                elif ms['status'] not in ('Complete',):
                    esc = ExecuteManager._escalation_level(ms)
                    if esc > 0:
                        try:
                            due = date.fromisoformat(ms.get('due_date',''))
                            overdue = (date.today() - due).days
                        except: overdue = 0
                        actions.append({
                            'type':       'milestone_overdue' if overdue > 0 else 'milestone_due_soon',
                            'init_id':     init['id'], 'ms_id': ms['id'],
                            'name':       init['name'], 'ms_name': ms['name'],
                            'due_date':   ms.get('due_date',''),
                            'overdue_days': overdue,
                            'esc_level':  esc,
                            'workstream': init.get('workstream',''),
                            'has_blocker': any(not b['resolved'] for b in ms.get('blockers',[])),
                        })
        return actions

    def get_all_milestones_for_owner(self, username, full_name: str = ""):
        """
        Return every milestone where this person is owner or co-owner.
        Matches on BOTH login username AND full_name because milestone owner
        is stored as full_name (from the dropdown) but uname is the login key.
        Also accepts cross-workstream assignments.
        """
        result = []
        # Build set of all identifiers for this person
        identifiers = {username}
        if full_name:
            identifiers.add(full_name)
        # Also resolve via users.json if possible
        try:
            users_data = json.loads((DATA_DIR / "users.json").read_text())
            ud = users_data.get(username, {})
            if ud.get("full_name"):
                identifiers.add(ud["full_name"])
            # Reverse: if username looks like a full name, find the login key too
            for u, d in users_data.items():
                if d.get("full_name") == username:
                    identifiers.add(u)
        except: pass

        for init in self.initiatives:
            for ms in init.get('milestones', []):
                owner_match = (ms.get('owner','') in identifiers or
                               any(co in identifiers for co in ms.get('co_owners', [])))
                if not owner_match:
                    continue
                    esc = ExecuteManager._escalation_level(ms)
                    try:
                        due = date.fromisoformat(ms.get('due_date',''))
                        days_diff = (due - date.today()).days
                    except:
                        days_diff = 999
                    try:
                        days_to_start = (date.fromisoformat(ms.get('start_date','9999-12-31')) - date.today()).days
                    except: days_to_start = 999
                    needs_start = ExecuteManager._needs_start_alert(ms)
                    entry = dict(ms)   # explicit copy — avoids unhashable type warning
                    entry.update({
                        'initiative_id':    init['id'],
                        'initiative_name':  init['name'],
                        'workstream':       init.get('workstream',''),
                        'sub_workstream':   init.get('sub_workstream',''),
                        'gate':             init.get('gate',''),
                        'io':               init.get('io',''),
                        'escalation_level': esc,
                        'days_to_due':      days_diff,
                        'days_to_start':    days_to_start,
                        'needs_start_alert': needs_start,
                        'is_primary_owner': ms.get('owner','') in identifiers,
                    })
                    result.append(entry)
        result.sort(key=lambda x: (0 if int(x.get('escalation_level',0)) > 0 else 1, int(x.get('days_to_due',999))))
        return result


    def get_cross_ws_delays_for_workstream(self, workstream_name: str) -> list:
        """
        Return milestones in OTHER workstreams whose delay is blocking 
        initiatives that depend on this workstream's output.
        Used to show workstream WS-B: 'WS-A is waiting on you'.
        """
        blocking = []
        for init in self.initiatives:
            for ms in init.get('milestones', []):
                dep_ws = ms.get('depends_on_workstream', '')
                if not dep_ws: continue
                if workstream_name.lower() not in dep_ws.lower(): continue
                if ms.get('status') in ('Complete',): continue
                try:
                    due = date.fromisoformat(ms.get('due_date',''))
                    days_diff = (due - date.today()).days
                except: days_diff = 999
                blocking.append({
                    **ms,
                    'initiative_id':   init['id'],
                    'initiative_name': init['name'],
                    'blocking_workstream': init.get('workstream',''),
                    'needs_from_workstream': dep_ws,
                    'days_to_due': days_diff,
                    'io': init.get('io',''),
                })
        blocking.sort(key=lambda x: x.get('days_to_due', 999))
        return blocking

    def get_escalation_dashboard(self, scope_initiatives=None):
        """All milestones at risk — grouped by escalation level (banking timelines)."""
        inits = scope_initiatives or self.initiatives
        buckets = {4: [], 3: [], 2: [], 1: [], 5: [], 'cross_ws': []}  # level → list
        for init in inits:
            for ms in init.get('milestones',[]):
                if ms.get('status') == 'Complete': continue
                esc = ExecuteManager._escalation_level(ms)
                if esc > 0:
                    try:
                        due = date.fromisoformat(ms.get('due_date',''))
                        overdue = (date.today() - due).days
                    except: overdue = 0
                    try:
                        days_to_start = (date.fromisoformat(ms.get('start_date','9999-12-31')) - date.today()).days
                    except: days_to_start = 999
                    item = dict(ms)   # explicit copy — avoids unhashable type warning
                    item.update({
                        'initiative_id':   init['id'],
                        'initiative_name': init['name'],
                        'workstream':      init.get('workstream',''),
                        'io':              init.get('io',''),
                        'gate':            init.get('gate',''),
                        'overdue_days':    max(0, overdue),
                        'days_to_due':     (date.fromisoformat(ms.get('due_date','')) - date.today()).days
                                           if ms.get('due_date') else 0,
                        'days_to_start':   days_to_start,
                        'needs_start_alert': ExecuteManager._needs_start_alert(ms),
                    })
                    if esc in buckets:
                        buckets[esc].append(item)
                    # Cross-workstream dependency flag (separate, not mutually exclusive)
                    if item.get('depends_on_workstream'):
                        item['cross_ws_delayed'] = (ms.get('status') in ('Delayed','Not Started') and
                                                     ms.get('delay_category','') in ('Dependency on another department','Cross-functional dependency','External dependency'))
                        buckets['cross_ws'].append(item)
                    else:
                        buckets[max(buckets.keys())].append(item)
        for level in buckets:
            buckets[level].sort(key=lambda x: -x.get('overdue_days', 0))
        return buckets


# ─── PRODUCT MANAGER ─────────────────────────────────────────────────
class ProductManager:
    """
    Banking product registry — tracks products across their full lifecycle.
    Linked to KPIs (Perform), pipeline deals (RI), and initiatives (Execute).
    """
    def __init__(self):
        self.file = DATA_DIR / "product_registry.json"
        if not self.file.exists(): self.file.write_text("[]")
        try:
            raw = self.file.read_text()
            self.products = json.loads(raw) if raw.strip() else []
            if not isinstance(self.products, list): self.products = []
        except:
            self.products = []

    def save(self):
        self.file.write_text(json.dumps(self.products, indent=2))

    def add_product(self, data: dict) -> str:
        prod_id = f"PRD{len(self.products)+1:04d}"
        product = {
            "id":               prod_id,
            "name":             data["name"],
            "category":         data["category"],         # Assets/Liabilities/NFI/Channels
            "sub_category":     data.get("sub_category",""),
            "product_type":     data.get("product_type",""),
            "lifecycle_stage":  data.get("lifecycle_stage","Active"),
            "health":           data.get("health","On track"),
            "owner":            data.get("owner",""),     # username
            "sponsor":          data.get("sponsor",""),   # Director
            "description":      data.get("description",""),
            "launch_date":      data.get("launch_date",""),
            "review_date":      data.get("review_date",""),
            "linked_kpis":      data.get("linked_kpis",[]),
            "linked_initiatives": data.get("linked_initiatives",[]),
            "target_segment":   data.get("target_segment",""),  # Retail/SME/Corporate
            "channels":         data.get("channels",[]),
            "annual_target":    data.get("annual_target",0),
            "ytd_actual":       data.get("ytd_actual",0),
            "customer_count":   data.get("customer_count",0),
            "notes":            [],
            "stage_history":    [{"stage": data.get("lifecycle_stage","Active"),
                                  "date": str(datetime.now().date()),
                                  "by": data.get("created_by","")}],
            "created_by":       data.get("created_by",""),
            "created_at":       datetime.now().isoformat(),
            "updated_at":       datetime.now().isoformat(),
            "active":           True,
        }
        self.products.append(product)
        self.save()
        return prod_id

    def update_product(self, prod_id: str, updates: dict, updated_by: str = ""):
        for p in self.products:
            if p["id"] == prod_id:
                old_stage = p.get("lifecycle_stage")
                p.update(updates)
                p["updated_at"] = datetime.now().isoformat()
                # Log stage changes
                if "lifecycle_stage" in updates and updates["lifecycle_stage"] != old_stage:
                    if "stage_history" not in p: p["stage_history"] = []
                    p["stage_history"].append({
                        "stage": updates["lifecycle_stage"],
                        "date": str(datetime.now().date()),
                        "by": updated_by,
                    })
                break
        self.save()

    def get_products(self, category: str = None, stage: str = None,
                     health: str = None, active_only: bool = True):
        result = [p for p in self.products if not active_only or p.get("active", True)]
        if category: result = [p for p in result if p.get("category") == category]
        if stage:    result = [p for p in result if p.get("lifecycle_stage") == stage]
        if health:   result = [p for p in result if p.get("health") == health]
        return result

    def get_product(self, prod_id: str):
        return next((p for p in self.products if p["id"] == prod_id), None)

    def lifecycle_summary(self):
        """Count products per stage for funnel view."""
        counts: dict = {}
        for p in self.products:
            if not p.get("active", True): continue
            s = p.get("lifecycle_stage","Active")
            counts[s] = counts.get(s, 0) + 1
        return counts

    def category_summary(self):
        """Count and aggregate per category."""
        summary: dict = {}
        for cat in PRODUCT_CATEGORIES:
            prods = self.get_products(category=cat)
            summary[cat] = {
                "count":    len(prods),
                "active":   sum(1 for p in prods if p.get("lifecycle_stage") in ("Active","Growth","Mature","Optimising")),
                "at_risk":  sum(1 for p in prods if p.get("health") == "At risk"),
                "in_pilot": sum(1 for p in prods if p.get("lifecycle_stage") == "Pilot"),
                "sunset":   sum(1 for p in prods if p.get("lifecycle_stage") in ("Sunset","Decommissioned")),
            }
        return summary

    def link_to_initiative(self, prod_id: str, init_id: str):
        for p in self.products:
            if p["id"] == prod_id:
                if init_id not in p.get("linked_initiatives", []):
                    if "linked_initiatives" not in p: p["linked_initiatives"] = []
                    p["linked_initiatives"].append(init_id)
                break
        self.save()

    def add_note(self, prod_id: str, note: str, author: str):
        for p in self.products:
            if p["id"] == prod_id:
                if "notes" not in p: p["notes"] = []
                p["notes"].append({"note": note, "by": author,
                                    "date": str(datetime.now().date())})
                break
        self.save()


# ─── REVENUE INTELLIGENCE PIPELINE MANAGER ──────────────────────────
class RIPipelineManager:
    """
    Manages the 4-category revenue pipeline:
    Deposits · Loans · NFI · New Customers
    Each deal is tagged to a category and links back to a KPI.
    """
    def __init__(self):
        self.file = DATA_DIR / "ri_pipeline.json"
        if not self.file.exists(): self.file.write_text("[]")
        try:
            raw = self.file.read_text()
            self.deals = json.loads(raw) if raw.strip() else []
            if not isinstance(self.deals, list): self.deals = []
        except:
            self.deals = []

    def save(self):
        self.file.write_text(json.dumps(self.deals, indent=2))

    def add_deal(self, d):
        d['id']         = f"RI{len(self.deals)+1:05d}"
        d['created_at'] = datetime.now().isoformat()
        d['updated_at'] = datetime.now().isoformat()
        d['staff_code'] = clean_code(d.get('staff_code',''))
        d['history']    = [{'stage': d['stage'], 'date': str(datetime.now().date()),
                            'note': d.get('notes','')}]
        self.deals.append(d)
        self.save()
        return d['id']

    def update_stage(self, deal_id, new_stage, note, updated_by):
        for d in self.deals:
            if d['id'] == deal_id:
                d['stage']      = new_stage
                d['updated_at'] = datetime.now().isoformat()
                d['updated_by'] = updated_by
                if not isinstance(d.get('history'), list): d['history'] = []
                d['history'].append({'stage': new_stage, 'date': str(datetime.now().date()), 'note': note})
                if new_stage in [RI_CATEGORIES[d['category']]['closed_won']]:
                    d['closed_date'] = str(datetime.now().date())
                break
        self.save()

    def get_deals(self, staff_code=None, category=None, team_names=None, active_only=False):
        result = self.deals
        if staff_code:   result = [d for d in result if d.get('staff_code') == clean_code(staff_code)]
        if team_names:   result = [d for d in result if d.get('staff_name') in team_names]
        if category:     result = [d for d in result if d.get('category') == category]
        if active_only:
            result = [d for d in result if d.get('stage') not in
                      [RI_CATEGORIES.get(d.get('category',{}) or 'Loans',{}).get('closed_lost','Lost'),
                       RI_CATEGORIES.get(d.get('category',{}) or 'Loans',{}).get('closed_won','')]]
        return result

    def weighted_value(self, deals):
        total = 0
        for d in deals:
            cat  = d.get('category','Loans')
            cfg  = RI_CATEGORIES.get(cat, {})
            wts  = cfg.get('stage_weights', {})
            val  = float(d.get('deal_value', 0))
            wt   = wts.get(d.get('stage',''), 0)
            total += val * wt
        return total

    def pipeline_value(self, deals):
        """Raw total of active deals (not weighted)."""
        closed_lost_stages = set()
        for cat, cfg in RI_CATEGORIES.items():
            closed_lost_stages.add(cfg['closed_lost'])
        return sum(float(d.get('deal_value', 0)) for d in deals
                   if d.get('stage') not in closed_lost_stages)

    def won_value(self, deals):
        total = 0
        for d in deals:
            cat = d.get('category','Loans')
            cw  = RI_CATEGORIES.get(cat,{}).get('closed_won','')
            if d.get('stage') == cw:
                total += float(d.get('deal_value', 0))
        return total

    def category_summary(self, deals, kpi_actuals, kpi_targets):
        """
        For each RI category return:
        ytd_actual, annual_target, pipeline_raw, pipeline_weighted, won,
        gap_to_target, forecast_eoy, coverage_pct, conversion_needed
        """
        summary = {}
        for cat, cfg in RI_CATEGORIES.items():
            cat_deals   = [d for d in deals if d.get('category') == cat]
            won_stage   = cfg['closed_won']
            lost_stage  = cfg['closed_lost']
            active_d    = [d for d in cat_deals if d.get('stage') not in (won_stage, lost_stage)]
            won_d       = [d for d in cat_deals if d.get('stage') == won_stage]

            ytd_act  = kpi_actuals.get(cat, 0)
            ann_tgt  = kpi_targets.get(cat, 0)
            pip_raw  = self.pipeline_value(active_d)
            pip_wt   = self.weighted_value(active_d)
            won_val  = self.won_value(won_d)

            # Forecast: ytd_actual / months_elapsed * 12  +  weighted pipeline
            run_rate = (ytd_act / months_elapsed()) * 12 if ytd_act > 0 else 0
            forecast = run_rate + pip_wt
            gap      = max(0, ann_tgt - ytd_act)
            coverage = (pip_wt / gap * 100) if gap > 0 else 100

            # Conversion needed: if pipeline_raw > 0, what % must convert to close gap
            conv_needed = (gap / pip_raw * 100) if pip_raw > 0 else None

            # Current conversion rate (won / (won + lost) deals)
            lost_d = [d for d in cat_deals if d.get('stage') == lost_stage]
            total_closed = len(won_d) + len(lost_d)
            curr_conv = (len(won_d) / total_closed * 100) if total_closed > 0 else None

            summary[cat] = {
                'ytd_actual':     ytd_act,
                'annual_target':  ann_tgt,
                'pipeline_raw':   pip_raw,
                'pipeline_wtd':   pip_wt,
                'won_ytd':        won_val,
                'gap_to_target':  gap,
                'forecast_eoy':   forecast,
                'coverage_pct':   coverage,
                'conv_needed':    conv_needed,
                'curr_conv':      curr_conv,
                'active_deals':   len(active_d),
                'won_deals':      len(won_d),
                'unit':           cfg['unit'],
            }
        return summary

# ─── USER MANAGER ────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
# CENTRAL ACCESS CONTROL SYSTEM
# ══════════════════════════════════════════════════════════════════════

# Module-level access definitions
# Each module maps to the minimum permission needed and optional role list
# "public"   → any logged-in user
# "self"     → can see own data only
# "team"     → can see self + direct reports
# "unit"     → can see own unit
# "all"      → can_view_all or admin
# "admin"    → is_admin only
# Sub-page definitions per module — admin can restrict to specific pages
MODULE_PAGES = {
    "perform":    ["My Scorecard","Rankings","Individual View","Validation","Analytics","Leave"],
    "people":     ["Leave","Exits & Separations","Transfers","Disciplinary","PIP"],
    "pipeline":   ["My Pipeline","Team Pipeline","Summary"],
    "execute":    ["My Tasks","Workstreams","Finance Approvals","Gantt"],
    "cascade":    ["Bank Targets & Timeline","Fixed KPIs","Set Team Targets",
                   "My Targets","Cascade Tree","Coverage & Deadlines","Review Requests"],
    "sla":        ["SLA Dashboard","By Branch","By Staff","Violations"],
    "branch_log": ["Daily Log Entry","Log History","Supervisor Review"],
    "commission": ["My Commission","Team Commission","Payouts"],
    "cims":       ["Raise Instruction","My Instructions","Team Queue","Admin"],
    "admin":      ["Users","Permissions","Reporting Lines","Transfers","Org Tree",
                   "Audit Log","Upload Format","Leave Settings"],
    "people":     ["Leave Management","Exits","Transfers","Disciplinary","PIP"],
}



# ══════════════════════════════════════════════════════════════════
# MAKER-CHECKER: dual approval for high-value operations
# CBK requirement for transactions above defined thresholds
# ══════════════════════════════════════════════════════════════════
MAKER_CHECKER_LIMITS = {
    "fd_ratification":    10_000_000,   # FD > KES 10M needs dual approval
    "loan_approval":      50_000_000,   # Loan > KES 50M needs dual approval
    "waiver_approval":    500_000,      # Waiver > KES 500K needs dual approval
    "legal_settlement":   5_000_000,    # Settlement > KES 5M needs dual approval
    "recon_write_off":    1_000_000,    # Write-off > KES 1M needs dual approval
}



# ══════════════════════════════════════════════════════════════════
# DEPARTMENT DEFINITIONS
# Maps canonical dept names to navigation groups and modules
# ══════════════════════════════════════════════════════════════════
DEPARTMENTS = [
    "Retail Banking",
    "Commercial & Corporate",
    "Credit",
    "Treasury",
    "Finance",
    "Risk & Compliance",
    "Legal",
    "Operations",
    "People & HR",
    "IT & Digital",
    "Bancassurance",
    "Marketing",
    "Internal Audit",
    "Support Services",
    "Executive",
]

# Modules that are ALWAYS visible regardless of department
UNIVERSAL_MODULES = [
    "perform",        # BSC (own score)
    "smart_alerts",
    "approvals",
    "statement_analyzer",
    "customer360",
]

# Department → primary modules (what the sidebar shows first)
DEPT_PRIMARY_MODULES = {
    "Retail Banking": [
        "nps","crosssell",
        "perform","cascade","pipeline","loan_applications","cims",
        "branch_log","commission","sla","campaigns","propositions",
        "optimize","products","customer360",
    ],
    "Commercial & Corporate": [
        "crosssell",
        "perform","cascade","pipeline","loan_applications","cims",
        "commission","sla","campaigns","propositions","customer360",
    ],
    "Credit": [
        "ews","collateral",
        "perform","loan_applications","credit_analysis","credit_admin",
        "credit_monitoring","debt_recovery","ifrs9","statement_analyzer",
    ],
    "Treasury": [
        "perform","treasury","ifrs9","stress_testing","ra",
    ],
    "Finance": [
        "budget",
        "perform","sbu","opex","revenue_assurance","rms","ra","ifrs9",
    ],
    "Risk & Compliance": [
        "perform","compliance","credit_monitoring","stress_testing",
        "debt_recovery","legal","edms","ifrs9",
    ],
    "Legal": [
        "perform","legal","edms","compliance",
    ],
    "Operations": [
        "perform","rms","cims","edms","approvals","export",
    ],
    "People & HR": [
        "lms","pip",
        "perform","people","cascade",
    ],
    "IT & Digital": [
        "incidents",
        "perform","cbs","export","admin",
    ],
    "Bancassurance": [
        "perform","campaigns","pipeline","propositions",
    ],
    "Marketing": [
        "perform","campaigns","competitor","propositions",
    ],
    "Internal Audit": [
        "perform","ra","compliance","credit_monitoring",
    ],
    "Support Services": [
        "perform","edms",
    ],
    "Executive": [
        "perform","integrate","ra","sbu","opex","competitor",
        "stress_testing","people","cascade","customer360",
        "pipeline","loan_applications","treasury","compliance",
        "credit_monitoring","ifrs9","revenue_assurance",
    ],
}

# Dept super-user permissions
DEPT_SUPER_USER_MODULES = [
    "admin",          # limited dept admin sub-panel only
    "people",         # view their dept staff
    "perform",        # see dept BSC summary
]

# ICT admin permissions
ICT_ADMIN_MODULES = [
    "admin",          # system health sub-panel only
    "export",
    "cbs",
]


MODULE_ACCESS = {
    "perform":     {"min": "self",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Regional Head",
                                                   "Branch Manager","Head Of Retail","Head Of Corporate",
                                                   "Head Of SME","Chief Finance Officer","Chief Risk Officer",
                                                   "Chief Operations Officer","Chief Compliance Officer",
                                                   "Chief Human Resources Officer","Head Of Digital Innovation",
                                                   "Head Of Strategy","Head Of Internal Audit","Head Of Marketing"]},
    "people":      {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Regional Head","Branch Manager",
                                                   "Chief Human Resources Officer","HR Business Partner"]},
    "pipeline":    {"min": "self", "roles_all": []},
    "execute":     {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Head Of Corporate","Head Of SME",
                                                   "Head Of Digital Innovation","Head Of Strategy",
                                                   "Chief Operations Officer","Regional Head","Branch Manager"]},
    "products":    {"min": "all",   "roles_all": ["Admin","Managing Director","Director Commercial Banking",
                                                   "Head Of Corporate","Head Of SME","Chief Finance Officer",
                                                   "Director Retail Banking"]},
    "integrate":   {"min": "all",   "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Chief Finance Officer"]},
    "admin":       {"min": "admin", "roles_all": ["Admin"]},
    "export":      {"min": "all",   "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                   "Director Retail Banking","Director Commercial Banking"]},
    "sbu":         {"min": "all",   "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                   "Director Retail Banking","Director Commercial Banking",
                                                   "Regional Head","Branch Manager"]},
    "opex":        {"min": "all",   "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                   "Director Retail Banking","Director Commercial Banking"]},
    "competitor":  {"min": "all",   "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Head Of Strategy",
                                                   "Chief Finance Officer"]},
    "cascade":     {"min": "self",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Head Of Retail","Head Of Corporate",
                                                   "Head Of SME","Head Of Digital Innovation","Chief Finance Officer",
                                                   "Chief Risk Officer","Chief Operations Officer",
                                                   "Chief Compliance Officer","Chief Human Resources Officer",
                                                   "Head Of Strategy","Head Of Internal Audit","Head Of Marketing",
                                                   "Regional Head","Branch Manager","Branch Operations Manager",
                                                   "Branch Credit Manager","Direct Sales Officer",
                                                   "Relationship Officer Personal Banking","Teller",
                                                   "Customer Service Officer","Relationship Manager Corporate",
                                                   "Relationship Manager SME"]},
    "sla":         {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Regional Head","Branch Manager",
                                                   "Branch Operations Manager","Chief Operations Officer",
                                                   "Chief Compliance Officer"]},
    "branch_log":  {"min": "unit",  "roles_all": ["Admin","Managing Director","Regional Head",
                                                   "Branch Manager","Branch Operations Manager","Teller",
                                                   "Customer Service Officer","Direct Sales Officer",
                                                   "Relationship Officer Personal Banking"]},
    "optimize":    {"min": "all",   "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Regional Head","Branch Manager","Chief Operations Officer"]},
    "commission":  {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Regional Head","Branch Manager",
                                                   "Branch Credit Manager","Direct Sales Officer",
                                                   "Relationship Manager Corporate","Relationship Manager SME",
                                                   "Relationship Officer Personal Banking"]},
    "campaigns":   {"min": "all",   "roles_all": ["Admin","Managing Director","Head Of Marketing",
                                                   "Marketing Officer","Director Retail Banking",
                                                   "Director Commercial Banking","Regional Head","Branch Manager"]},
    "credit_monitoring":  {"min": "self", "roles_all": []},
    "debt_recovery":      {"min": "self", "roles_all": []},
    "loan_applications":  {"min": "self", "roles_all": []},
    "legal":              {"min": "self", "roles_all": []},
    "credit_analysis":    {"min": "self", "roles_all": []},
    "credit_admin":       {"min": "self", "roles_all": []},
    "compliance":         {"min": "self", "roles_all": []},
    "treasury":           {"min": "self", "roles_all": []},
    "propositions":       {"min": "self", "roles_all": []},
    "ra":                  {"min": "self", "roles_all": []},
    "revenue_assurance": {"min": "self", "roles_all": []},
    "ifrs9":               {"min": "self", "roles_all": []},
    "statement_analyzer":  {"min": "self", "roles_all": []},
    "customer360":    {"min": "self", "roles_all": []},
    "stress_testing":  {"min": "self", "roles_all": []},
    "smart_alerts":    {"min": "self", "roles_all": []},
    "approvals":       {"min": "self", "roles_all": []},
    "digital_analytics":{"min": "self", "roles_all": []},
    "rms":                 {"min": "self", "roles_all": []},
    "edms":                {"min": "self", "roles_all": []},
    "cims":        {"min": "self",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                   "Director Commercial Banking","Chief Operations Officer",
                                                   "Branch Manager","Branch Operations Manager",
                                                   "Branch Credit Manager","Regional Head",
                                                   "Customer Service Officer","Teller","Direct Sales Officer",
                                                   "Relationship Officer Personal Banking",
                                                   "Relationship Manager Corporate","Relationship Manager SME"]},

    # ── Phase 1 CRITICAL Regulatory ─────────────────────────────────
    "cbk_returns":           {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                              "Chief Risk Officer","Chief Compliance Officer",
                                                              "Head Of Internal Audit","Head Of Strategy",
                                                              "Director Retail Banking","Director Commercial Banking"]},
    "data_protection":       {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Risk Officer",
                                                              "Chief Compliance Officer","Chief Information Officer",
                                                              "Head Of Internal Audit"]},
    "sanctions_screening":   {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Compliance Officer",
                                                              "Chief Risk Officer","Head Of Internal Audit"]},
    "regulatory_capital":    {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                              "Chief Risk Officer","Head Of Strategy",
                                                              "Director Retail Banking","Director Commercial Banking"]},
    # ── Phase 2 HIGH Business ───────────────────────────────────────
    "customer_onboarding":   {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                              "Director Commercial Banking","Regional Head",
                                                              "Branch Manager","Head Of Retail","Head Of SME",
                                                              "Head Of Corporate","Head Of Digital Innovation",
                                                              "Chief Operations Officer"]},
    "card_management":       {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                              "Head Of Retail","Chief Operations Officer",
                                                              "Head Of Digital Innovation","Branch Manager",
                                                              "Regional Head"]},
    "merchant_acquiring":    {"min": "team",  "roles_all": ["Admin","Managing Director","Director Commercial Banking",
                                                              "Director Retail Banking","Head Of SME","Head Of Corporate",
                                                              "Head Of Digital Innovation","Chief Finance Officer",
                                                              "Regional Head","Branch Manager"]},
    "alm_liquidity":         {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                              "Chief Risk Officer","Head Of Strategy"]},
    "operational_risk":      {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Risk Officer",
                                                              "Chief Compliance Officer","Chief Operations Officer",
                                                              "Head Of Internal Audit","Director Retail Banking",
                                                              "Director Commercial Banking","Regional Head","Branch Manager"]},
    # ── Phase 3 STRATEGIC ───────────────────────────────────────────
    "strategic_initiatives": {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                              "Director Commercial Banking","Chief Finance Officer",
                                                              "Chief Risk Officer","Chief Operations Officer",
                                                              "Chief Information Officer","Chief Human Resources Officer",
                                                              "Head Of Strategy","Head Of Internal Audit",
                                                              "Head Of Marketing","Head Of Retail","Head Of SME",
                                                              "Head Of Corporate","Head Of Digital Innovation"]},
    "board_papers":          {"min": "team",  "roles_all": ["Admin","Managing Director","Director Retail Banking",
                                                              "Director Commercial Banking","Chief Finance Officer",
                                                              "Chief Risk Officer","Chief Operations Officer",
                                                              "Chief Information Officer","Chief Human Resources Officer",
                                                              "Chief Compliance Officer","Head Of Strategy",
                                                              "Head Of Internal Audit"]},
    "esg_climate":           {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Risk Officer",
                                                              "Chief Compliance Officer","Head Of Strategy",
                                                              "Director Commercial Banking","Director Retail Banking"]},
    # ── FLEXCUBE Integration ────────────────────────────────────────
    "flexcube_integration":  {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Information Officer",
                                                              "Chief Operations Officer","Chief Risk Officer",
                                                              "Head Of Digital Innovation","Head Of Internal Audit",
                                                              "Director Retail Banking","Director Commercial Banking"]},

    # ── Tier-1 Benchmarking ─────────────────────────────────────────
    "benchmarking":          {"min": "team",  "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                              "Chief Risk Officer","Head Of Strategy",
                                                              "Director Retail Banking","Director Commercial Banking",
                                                              "Chief Operations Officer","Head Of Internal Audit"]},

    # ── MD/CEO Executive Cockpit (v10.214) ─────────────────────────
    "md_cockpit":            {"min": "all",   "roles_all": ["Admin","Managing Director","Chief Finance Officer",
                                                              "Chief Risk Officer","Chief Operations Officer",
                                                              "Chief Compliance Officer","Chief Human Resources Officer",
                                                              "Director Retail Banking","Director Commercial Banking",
                                                              "Head Of Strategy","Head Of Internal Audit"]},
}


# ─── LOAN MANAGEMENT SYSTEM MANAGERS ─────────────────────────────────

class LoanApplicationManager:
    """Persist and manage loan_applications.json."""
    def __init__(self):
        self.file = DATA_DIR / "loan_applications.json"
        self.apps = self._load()

    def _load(self) -> list:
        try:
            return json.loads(self.file.read_text()) if self.file.exists() else []
        except Exception:
            return []

    def save(self):
        self.file.write_text(json.dumps(self.apps, indent=2, default=str))

    def get(self, app_id: str) -> dict:
        return next((a for a in self.apps if a["id"] == app_id), {})

    def update(self, app_id: str, fields: dict):
        for i, a in enumerate(self.apps):
            if a["id"] == app_id:
                self.apps[i].update(fields)
                self.apps[i]["last_updated"] = datetime.now().date().isoformat()
                self.save()
                return True
        return False

    def submit_to_credit(self, app_id: str, analyst_code: str = "",
                          analyst_name: str = "") -> bool:
        app = self.get(app_id)
        if not app:
            return False
        updates = {"status": "assigned", "last_updated": datetime.now().date().isoformat()}
        if analyst_code:
            updates["analyst"] = {"code": analyst_code, "name": analyst_name}
        return self.update(app_id, updates)

    def record_decision(self, app_id: str, verdict: str, authority: str,
                        reason: str = "", conditions: list = None,
                        comments: str = "") -> bool:
        new_status = {
            "approved": "approved", "decline": "declined",
            "declined": "declined", "return":  "returned",
            "returned": "returned",
        }.get(verdict.lower(), verdict)
        return self.update(app_id, {
            "status":   new_status,
            "decision": {
                "verdict":    verdict.lower(),
                "date":       datetime.now().date().isoformat(),
                "authority":  authority,
                "reason":     reason,
                "conditions": conditions or [],
                "comments":   comments,
            },
        })

    # ── Credit workflow state machine (v10.584) ──────────────────────
    # Hardcoded transitions live in utils/api_lms_mutations.py; these are
    # the data mutations + the shared event log. Every workflow action
    # appends a history event {event, by, at, note, ...} so the analyst,
    # the deal owner, and credit admin share one timeline (who's handling
    # it, when, time taken).
    def _log_event(self, app_id: str, event: str, by: str,
                   note: str = "", extra: dict = None) -> None:
        app = self.get(app_id)
        if not app:
            return
        hist = list(app.get("history") or [])
        entry = {"event": event, "by": by,
                 "at": datetime.now().isoformat(), "note": note}
        if extra:
            entry.update(extra)
        hist.append(entry)
        self.update(app_id, {"history": hist})

    def request_info(self, app_id: str, by: str, reasons: list = None,
                     documents: list = None, note: str = "") -> bool:
        """Analyst parks the case asking the deal owner for more docs
        (pre-decision). Routes visibility to the owner."""
        if not self.get(app_id):
            return False
        self.update(app_id, {
            "status": "info_requested",
            "info_request": {
                "by": by, "at": datetime.now().isoformat(),
                "reasons": reasons or [], "documents": documents or [],
                "note": note, "resolved": False,
            },
        })
        self._log_event(app_id, "info_requested", by, note,
                        {"reasons": reasons or [], "documents": documents or []})
        return True

    def provide_info(self, app_id: str, by: str, note: str = "",
                     documents: list = None) -> bool:
        """Deal owner supplies requested info; case returns to assigned."""
        app = self.get(app_id)
        if not app:
            return False
        ir = dict(app.get("info_request") or {})
        ir.update({"resolved": True, "resolved_by": by,
                   "resolved_at": datetime.now().isoformat()})
        if documents:
            ir["provided_documents"] = documents
        self.update(app_id, {"status": "assigned", "info_request": ir})
        self._log_event(app_id, "info_provided", by, note,
                        {"documents": documents or []})
        return True

    def issue_offer(self, app_id: str, by: str, note: str = "") -> bool:
        """Route an approved app back to the deal owner to issue the
        letter of offer (offer_issued)."""
        if not self.get(app_id):
            return False
        self.update(app_id, {
            "status": "offer_issued",
            "offer": {"issued_by": by,
                      "issued_at": datetime.now().isoformat(),
                      "note": note, "signed": False, "validated": None},
        })
        self._log_event(app_id, "offer_issued", by, note)
        return True

    def sign_offer(self, app_id: str, by: str, attachment: dict = None,
                   note: str = "") -> bool:
        """Deal owner marks the offer signed + attaches the signed copy
        (reference or file ref per config)."""
        app = self.get(app_id)
        if not app:
            return False
        offer = dict(app.get("offer") or {})
        offer.update({"signed": True, "signed_by": by,
                      "signed_at": datetime.now().isoformat()})
        if attachment:
            offer["signed_attachment"] = attachment
        self.update(app_id, {"status": "offer_signed", "offer": offer})
        self._log_event(app_id, "offer_signed", by, note,
                        {"attachment": attachment or {}})
        return True

    def validate_offer(self, app_id: str, by: str, approve: bool = True,
                       note: str = "") -> bool:
        """Line manager validates the signed offer (checks & balances).
        Reject sends it back to offer_signed for re-handling."""
        app = self.get(app_id)
        if not app:
            return False
        offer = dict(app.get("offer") or {})
        offer.update({"validated": bool(approve), "validated_by": by,
                      "validated_at": datetime.now().isoformat()})
        new_status = "offer_validated" if approve else "offer_signed"
        self.update(app_id, {"status": new_status, "offer": offer})
        self._log_event(app_id,
                        "offer_validated" if approve else "offer_validation_rejected",
                        by, note)
        return True

    def confirm_to_credit_admin(self, app_id: str, by: str,
                                note: str = "") -> bool:
        """Credit analyst confirms to credit admin to proceed
        (analyst_confirmed). The route then creates the CALMS case."""
        if not self.get(app_id):
            return False
        self.update(app_id, {
            "status": "analyst_confirmed",
            "analyst_confirmation": {"by": by,
                                     "at": datetime.now().isoformat(),
                                     "note": note},
        })
        self._log_event(app_id, "analyst_confirmed", by, note)
        return True

    def refer_to_committee(self, app_id: str, by: str, note: str = "") -> bool:
        """Route an application to the credit committee (committee_voting mode)."""
        if not self.get(app_id):
            return False
        self.update(app_id, {
            "status": "referred_to_committee",
            "committee": {"votes": [], "referred_by": by,
                          "referred_at": datetime.now().isoformat(),
                          "note": note, "resolved": False},
        })
        self._log_event(app_id, "referred_to_committee", by, note)
        return True

    def record_committee_vote(self, app_id: str, member_id: str, vote: str,
                              rationale: str = "", by: str = "") -> bool:
        """Record/replace one member's vote on the application's committee."""
        app = self.get(app_id)
        if not app:
            return False
        committee = dict(app.get("committee") or {"votes": []})
        votes = [v for v in (committee.get("votes") or [])
                 if v.get("member_id") != member_id]
        votes.append({"member_id": member_id, "vote": vote,
                      "rationale": rationale, "by": by,
                      "at": datetime.now().isoformat()})
        committee["votes"] = votes
        self.update(app_id, {"committee": committee})
        self._log_event(app_id, "committee_vote", by or member_id,
                        rationale, {"member_id": member_id, "vote": vote})
        return True

    def resolve_committee(self, app_id: str, result: dict, by: str,
                          authority: str = "Credit Committee",
                          note: str = "") -> bool:
        """Store the committee engine result + set the resulting status.
        approved -> 'approved' (the route then issues the offer);
        rejected -> 'declined'; otherwise stays referred_to_committee."""
        app = self.get(app_id)
        if not app:
            return False
        committee = dict(app.get("committee") or {})
        committee.update({"result": result, "resolved": bool(
            result.get("approved") or result.get("rejected")),
            "resolved_by": by, "resolved_at": datetime.now().isoformat()})
        upd: dict = {"committee": committee}
        if result.get("approved"):
            upd["status"] = "approved"
            upd["decision"] = {
                "verdict": "approved",
                "date": datetime.now().date().isoformat(),
                "authority": authority,
                "reason": result.get("rationale", ""),
                "conditions": result.get("conditions", []),
                "comments": note, "via": "committee",
            }
        elif result.get("rejected"):
            upd["status"] = "declined"
            upd["decision"] = {
                "verdict": "declined",
                "date": datetime.now().date().isoformat(),
                "authority": authority,
                "reason": result.get("rationale", ""),
                "conditions": [], "comments": note, "via": "committee",
            }
        self.update(app_id, upd)
        self._log_event(app_id,
                        f"committee_{str(result.get('outcome', 'resolved')).lower()}",
                        by, note)
        return True

    def bsc_actuals(self) -> dict:
        """Compute BSC actuals from LMS data for credit KPIs."""
        from collections import defaultdict as _dd
        rm_kpis: dict = _dd(lambda: _dd(float))
        for app in self.apps:
            if app.get("status") not in ("approved","credit_admin","disbursed"):
                continue
            rm  = str(app.get("rm_code",""))
            amt = float(app.get("amount",0) or 0)
            prod= str(app.get("product","") or "").lower()
            if not rm or not amt:
                continue
            # Route to correct disbursement KPI
            if any(x in prod for x in ("personal","salary","mortgage","asset","staff","advance")):
                rm_kpis[rm]["Disbursements Retail Loans"] += amt
            elif any(x in prod for x in ("corporate","trade finance","import","export","bond","syndic")):
                rm_kpis[rm]["Disbursements Corporate Loans"] += amt
            else:
                rm_kpis[rm]["Disbursements MSME Loans"] += amt
                rm_kpis[rm]["Number of Business Borrowers"] += 1
            rm_kpis[rm]["Loan Book Growth"] += amt
            rm_kpis[rm]["New Accounts"]      += 1
        return {rm: dict(kpis) for rm, kpis in rm_kpis.items()}

    def create_from_pipeline_deal(self, deal: dict, username: str = ""):
        """Create a LoanApplication record from a pipeline deal.

        Canonical handoff function — v10.506 Phase 3 Arc α Batch α4
        adds this. It replaces the inline `lam.apps.append(...)` block
        in `pages/3_pipeline.py:1239-1287` AS A CANONICAL TARGET; the
        Streamlit page is not migrated to use it in this batch (a
        separate small batch can do that without behavioral change).

        Behaviour
        ---------
        - **Idempotent.** If an application already exists with a
          matching `pipeline_deal_id`, returns the existing app's id
          and does NOT create a duplicate. The deal can be re-advanced
          to an LMS stage repeatedly without side effects.
        - **ID generation uses `max(existing_ids) + 1`** rather than
          `len(apps) + 1`. The latter formula has a latent collision
          bug — same-turn inspection at α4 found existing data state
          where `len + 1` would collide with an existing id (724
          apps but highest id = LMS00725, one gap somewhere).
        - **Field mapping prefers Generation B canonical names.**
          `product_type` (Gen B) is preferred over `product` (Gen A);
          this fixes a latent bug in the Streamlit handoff where
          `product` was always empty for Gen B deals, breaking the
          KPI routing in `bsc_actuals()`.
        - **Swim lane bands** match the Streamlit handoff exactly:
          `Express` if amount ≤ 5M, `Complex` if ≥ 100M, `Standard`
          otherwise.

        Parameters
        ----------
        deal : dict
            A pipeline deal record (typically from
            `PipelineManager.get_deal(id)`).
        username : str, optional
            For the audit breadcrumb. The endpoint caller's username.

        Returns
        -------
        Optional[str]
            The created application's id (e.g. `"LMS00726"`), OR the
            existing app's id if already linked, OR `None` if the
            deal lacks the minimum fields required to create an
            application.

        Raises
        ------
        IOError, OSError
            If saving to disk fails. The caller (endpoint) decides
            how to handle this — α4's pipeline_deal_advance endpoint
            catches and returns lms_error in the response.
        """
        if not deal:
            return None
        deal_id = str(deal.get("id", "") or "")
        if not deal_id:
            return None

        # Idempotency: check for an existing application with this
        # pipeline_deal_id. If found, return its id without creating
        # a duplicate. This makes the handoff safely re-runnable.
        for a in self.apps:
            if str(a.get("pipeline_deal_id", "") or "") == deal_id:
                return str(a.get("id", "") or "")

        # ID generation — use max + 1 not len + 1, so gaps in the
        # sequence don't cause collisions. Defensive against the
        # latent bug present in the Streamlit handoff.
        existing_nums = []
        for a in self.apps:
            aid = str(a.get("id", "") or "")
            if aid.startswith("LMS") and len(aid) > 3:
                try:
                    existing_nums.append(int(aid[3:]))
                except ValueError:
                    continue
        next_num = (max(existing_nums) + 1) if existing_nums else 1
        new_id = f"LMS{next_num:05d}"

        # Field mapping. product_type (Gen B canonical) preferred over
        # product (Gen A legacy). amount from canonical deal_value.
        amount = 0.0
        try:
            amount = float(deal.get("deal_value", 0) or 0)
        except (TypeError, ValueError):
            amount = 0.0
        if not amount:
            # Some legacy records have amount instead of deal_value
            try:
                amount = float(deal.get("amount", 0) or 0)
            except (TypeError, ValueError):
                amount = 0.0

        product = (deal.get("product_type") or deal.get("product") or "").strip()

        if amount <= 5_000_000:
            swim_lane = "Express"
        elif amount >= 100_000_000:
            swim_lane = "Complex"
        else:
            swim_lane = "Standard"

        today = datetime.now().date().isoformat()

        app = {
            "id":              new_id,
            "pipeline_deal_id": deal_id,
            "client_name":     deal.get("client_name", ""),
            "client_cif":      str(deal.get("client_cif", "") or ""),
            "product":         product,
            "amount":          amount,
            "currency":        str(deal.get("currency", "") or "KES"),
            # P4-1b: inherit the normalized money set stamped on the deal so the
            # FCY/LCY dimension flows pipeline -> LMS -> credit admin.
            "fx_rate":         deal.get("fx_rate"),
            "amount_kes":      deal.get("amount_kes"),
            "currency_book":   deal.get("currency_book"),
            "fx_rate_date":    deal.get("fx_rate_date"),
            "fx_rate_source":  deal.get("fx_rate_source"),
            # P4-2: facility security classification (default unsecured; credit
            # reclassifies). Drives perfection routing; enforced at gate (P4-6).
            "facility_security_type": str(deal.get("facility_security_type", "") or "unsecured"),
            "security_subtype":       deal.get("security_subtype"),
            "swim_lane":       swim_lane,
            "status":          "submitted",
            "application_date": today,
            "rm_code":         str(deal.get("staff_code", "") or ""),
            "rm_name":         deal.get("staff_name", ""),
            "rm_unit":         deal.get("unit", ""),
            "analyst":         None,
            "is_repeat_borrower":     False,
            "clean_repayment_history": False,
            "docs_required":   [],
            "docs_submitted":  [],
            "completeness_score": 0,
            "compliance_flag":    False,
            "compliance_type":    None,
            "decision":           None,
            "tat_days":           0,
            "sla_target_days":    10,
            "last_updated":       today,
            # Provenance breadcrumb — useful for forensics
            "created_by":         username or "",
            "created_via":        "api_pipeline_advance",
        }
        self.apps.append(app)
        self.save()
        return new_id


class CreditAdminManager:
    """Persist credit_admin.json — pre-disbursement conditions."""
    def __init__(self):
        self.file = DATA_DIR / "credit_admin.json"
        self.cases = self._load()

    def _load(self) -> list:
        try:
            return json.loads(self.file.read_text()) if self.file.exists() else []
        except Exception:
            return []

    def save(self):
        self.file.write_text(json.dumps(self.cases, indent=2, default=str))

    def get(self, case_id: str) -> dict:
        return next((c for c in self.cases if c["id"] == case_id), {})

    def fulfill_condition(self, case_id: str, condition_type: str,
                           officer_name: str = "") -> bool:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            for cond in case.get("conditions", []):
                if cond["type"] == condition_type and not cond["fulfilled"]:
                    cond["fulfilled"] = True
                    cond["date_met"]  = datetime.now().date().isoformat()
                    cond["officer"]   = officer_name
                    break
            all_met = all(c["fulfilled"] for c in case.get("conditions", []))
            case["all_conditions_met"]      = all_met
            # Two-layer (config): when on, conditions being met does NOT make
            # the case ready — a CA manager must authorize first. When off,
            # preserve the legacy behaviour (all-met -> ready).
            if not self._two_layer_enabled():
                case["ready_for_disbursement"]  = all_met
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    # ── P4-2: CP/CS + facility classification ────────────────────────
    def set_facility_classification(self, case_id: str,
                                    facility_security_type: str,
                                    security_subtype: str = "") -> bool:
        """Set the case's facility security type (unsecured/secured) and
        optional subtype. Additive; does not itself gate anything (the gate is
        enforced in P4-6)."""
        fst = str(facility_security_type or "").strip().lower()
        if fst not in ("unsecured", "secured"):
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            case["facility_security_type"] = fst
            if security_subtype:
                case["security_subtype"] = security_subtype
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def classify_condition(self, case_id: str, condition_type: str,
                           classification: str = None, mandatory: bool = None,
                           due_date: str = None) -> bool:
        """Reclassify a condition as Precedent or Subsequent, set whether it is
        mandatory, and (for Subsequent) an optional due date. Only the provided
        fields are changed."""
        if classification is not None and classification not in (
                "precedent", "subsequent"):
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            hit = False
            for cond in case.get("conditions", []):
                if cond.get("type") == condition_type:
                    if classification is not None:
                        cond["classification"] = classification
                    if mandatory is not None:
                        cond["mandatory"] = bool(mandatory)
                    if due_date is not None:
                        cond["due_date"] = due_date or None
                    hit = True
                    break
            if not hit:
                return False
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    @staticmethod
    def outstanding_mandatory_cp(case: dict) -> list:
        """Pure helper — the future disbursement-gate input (P4-6). Returns the
        mandatory Conditions Precedent that are not yet fulfilled. A condition
        defaults to mandatory precedent when classification/mandatory are
        absent (legacy/safe). Conditions Subsequent never appear here."""
        out = []
        for c in (case.get("conditions", []) or []):
            classification = c.get("classification", "precedent")
            mandatory = c.get("mandatory", True)
            if classification == "precedent" and mandatory and not c.get("fulfilled"):
                out.append(c.get("type"))
        return out

    def _recompute_coverage(self, case: dict) -> None:
        """Recompute coverage_ratio + security_classification on the case from
        its linked collateral, using the admin Credit Policy Matrix. Best-effort
        — never raises into a mutation path."""
        try:
            from utils.collateral_coverage import assess_facility, CreditPolicyMatrix
            linked = case.get("linked_collateral", []) or []
            # Facility exposure in KES (prefer the normalized amount_kes).
            fac_kes = case.get("amount_kes")
            if fac_kes is None:
                fac_kes = case.get("amount", 0)
            assessment = assess_facility(
                fac_kes, linked,
                subtype_override=case.get("security_subtype"),
                matrix=CreditPolicyMatrix())
            case["coverage_ratio"] = assessment["coverage_ratio"]
            case["required_ratio"] = assessment["required_ratio"]
            case["security_total_kes"] = assessment["security_total_kes"]
            case["security_classification"] = assessment["security_classification"]
        except Exception:
            pass

    def link_collateral(self, case_id: str, collateral_id: str,
                        collateral_type: str, forced_sale_value,
                        currency: str = "KES", market_value=None,
                        allocated_value_kes=None, valuation_date: str = None) -> bool:
        """Link a collateral item to the facility and recompute coverage +
        security classification. Stores a snapshot of the security value on the
        link (coverage is a point-in-time assessment + audit anchor)."""
        for case in self.cases:
            if case["id"] != case_id:
                continue
            links = case.setdefault("linked_collateral", [])
            # replace existing link for the same collateral_id
            links[:] = [l for l in links if l.get("collateral_id") != collateral_id]
            links.append({
                "collateral_id":     collateral_id,
                "collateral_type":   collateral_type,
                "forced_sale_value": forced_sale_value,
                "market_value":      market_value,
                "currency":          currency or "KES",
                "allocated_value_kes": allocated_value_kes,
                "valuation_date":    valuation_date,
                "linked_at":         datetime.now().date().isoformat(),
            })
            self._recompute_coverage(case)
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def unlink_collateral(self, case_id: str, collateral_id: str) -> bool:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            links = case.get("linked_collateral", []) or []
            before = len(links)
            links[:] = [l for l in links if l.get("collateral_id") != collateral_id]
            if len(links) == before:
                return False
            self._recompute_coverage(case)
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    # ── P4-4: Legal Review workflow ──────────────────────────────────
    @staticmethod
    def _ensure_legal_review(case: dict) -> dict:
        """Lazily initialize the legal_review object so existing cases work."""
        lr = case.get("legal_review")
        if not isinstance(lr, dict):
            lr = {
                "status":                "not_started",
                "assigned_officer_code": None,
                "assigned_officer_name": None,
                "outcome":               None,
                "comments":              [],
                "started_at":            None,
                "completed_at":          None,
            }
            case["legal_review"] = lr
        return lr

    def assign_legal_officer(self, case_id: str, officer_code: str,
                             officer_name: str = "") -> bool:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            lr = self._ensure_legal_review(case)
            lr["assigned_officer_code"] = officer_code
            lr["assigned_officer_name"] = officer_name
            if lr["status"] in ("not_started",):
                lr["status"] = "in_review"
                lr["started_at"] = datetime.now().date().isoformat()
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def add_legal_comment(self, case_id: str, author_code: str, text: str,
                          raises_query: bool = False) -> bool:
        if not str(text or "").strip():
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            lr = self._ensure_legal_review(case)
            lr["comments"].append({
                "author_code": author_code,
                "text":        text,
                "at":          datetime.now().isoformat(timespec="seconds"),
            })
            if raises_query:
                lr["status"] = "queries_raised"
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def set_legal_outcome(self, case_id: str, outcome: str,
                          by: str = "") -> bool:
        if outcome not in ("approved", "approved_with_conditions", "rejected"):
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            lr = self._ensure_legal_review(case)
            lr["outcome"] = outcome
            lr["status"] = "rejected" if outcome == "rejected" else "cleared"
            lr["completed_at"] = datetime.now().date().isoformat()
            lr["completed_by"] = by
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    @staticmethod
    def legal_blocks_disbursement(case: dict) -> bool:
        """Pure gate-input (P4-6): secured facilities require a cleared legal
        review (outcome approved or approved_with_conditions). Unsecured never
        blocked on legal."""
        if str(case.get("facility_security_type", "unsecured")) != "secured":
            return False
        lr = case.get("legal_review") or {}
        return lr.get("outcome") not in ("approved", "approved_with_conditions")

    # ── P4-5: Security Perfection + Insurance ────────────────────────
    def add_security_perfection(self, case_id: str, security_type: str,
                                registration_reference: str = "",
                                registration_status: str = "pending",
                                registration_date: str = None,
                                perfection_status: str = "unperfected",
                                officer_code: str = "", notes: str = "") -> str:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            perfs = case.setdefault("security_perfections", [])
            pid = f"{case_id}-P{len(perfs) + 1}"
            now = datetime.now().isoformat(timespec="seconds")
            perfs.append({
                "id":                     pid,
                "security_type":          security_type,
                "registration_status":    registration_status,
                "registration_reference": registration_reference,
                "registration_date":      registration_date,
                "perfection_status":      perfection_status,
                "perfecting_officer_code": officer_code,
                "notes":                  notes,
                "created_at":             now,
                "updated_at":             now,
            })
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return pid
        return ""

    def update_security_perfection(self, case_id: str, perfection_id: str,
                                   **fields) -> bool:
        allowed = {"registration_status", "registration_reference",
                   "registration_date", "perfection_status",
                   "perfecting_officer_code", "notes"}
        rs_ok = ("pending", "lodged", "registered", "failed")
        ps_ok = ("unperfected", "in_progress", "perfected", "lapsed")
        if "registration_status" in fields and fields["registration_status"] not in rs_ok:
            return False
        if "perfection_status" in fields and fields["perfection_status"] not in ps_ok:
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            for p in case.get("security_perfections", []):
                if p.get("id") == perfection_id:
                    for k, v in fields.items():
                        if k in allowed and v is not None:
                            p[k] = v
                    p["updated_at"] = datetime.now().isoformat(timespec="seconds")
                    case["last_updated"] = datetime.now().date().isoformat()
                    self.save()
                    return True
            return False
        return False

    def add_insurance_policy(self, case_id: str, insurer: str,
                             policy_number: str, sum_insured=None,
                             currency: str = "KES", effective_date: str = None,
                             expiry_date: str = None,
                             bank_interest_noted: bool = False,
                             collateral_id: str = "", status: str = "active",
                             renewal_alert_days: int = 30) -> str:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            pols = case.setdefault("insurance_policies", [])
            iid = f"{case_id}-I{len(pols) + 1}"
            now = datetime.now().isoformat(timespec="seconds")
            pols.append({
                "id":                  iid,
                "collateral_id":       collateral_id,
                "insurer":             insurer,
                "policy_number":       policy_number,
                "sum_insured":         sum_insured,
                "currency":            currency or "KES",
                "effective_date":      effective_date,
                "expiry_date":         expiry_date,
                "bank_interest_noted": bool(bank_interest_noted),
                "status":              status,
                "renewal_alert_days":  renewal_alert_days,
                "created_at":          now,
                "updated_at":          now,
            })
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return iid
        return ""

    def update_insurance_policy(self, case_id: str, policy_id: str,
                                **fields) -> bool:
        allowed = {"insurer", "policy_number", "sum_insured", "currency",
                   "effective_date", "expiry_date", "bank_interest_noted",
                   "status", "renewal_alert_days", "collateral_id"}
        st_ok = ("active", "expired", "cancelled", "pending")
        if "status" in fields and fields["status"] not in st_ok:
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            for pol in case.get("insurance_policies", []):
                if pol.get("id") == policy_id:
                    for k, v in fields.items():
                        if k in allowed and v is not None:
                            pol[k] = v
                    pol["updated_at"] = datetime.now().isoformat(timespec="seconds")
                    case["last_updated"] = datetime.now().date().isoformat()
                    self.save()
                    return True
            return False
        return False

    @staticmethod
    def perfection_blocks_disbursement(case: dict) -> bool:
        """Secured facilities require every security instrument perfected. A
        secured facility with NO perfection records also blocks (nothing
        perfected yet). Unsecured never blocked."""
        if str(case.get("facility_security_type", "unsecured")) != "secured":
            return False
        perfs = case.get("security_perfections", []) or []
        if not perfs:
            return True
        return any(p.get("perfection_status") != "perfected" for p in perfs)

    @staticmethod
    def has_valid_insurance(case: dict, as_of: str = None) -> bool:
        """True if the case has at least one active, unexpired policy with the
        bank's interest noted."""
        import datetime as _dt
        today = as_of or _dt.date.today().isoformat()
        for pol in (case.get("insurance_policies", []) or []):
            if (pol.get("status") == "active"
                    and bool(pol.get("bank_interest_noted"))
                    and str(pol.get("expiry_date") or "9999-12-31") >= today):
                return True
        return False

    @staticmethod
    def insurance_blocks_disbursement(case: dict, as_of: str = None,
                                      required: bool = True) -> bool:
        """Blocks when insurance is required (caller/P4-6 decides per policy)
        and there is no valid policy. Unsecured or not-required never blocks."""
        if str(case.get("facility_security_type", "unsecured")) != "secured":
            return False
        if not required:
            return False
        return not CreditAdminManager.has_valid_insurance(case, as_of)

    # ── P4-6: disbursement HARD-GATE + tiered override ───────────────
    @staticmethod
    def _insurance_required_for(case: dict, matrix) -> bool:
        """Insurance mandatory if any linked collateral type — or the facility
        subtype — is in the matrix's insurance_required_subtypes."""
        req = set(getattr(matrix, "insurance_required_subtypes", []) or [])
        if not req:
            return False
        subtype = case.get("security_subtype")
        if subtype and subtype in req:
            return True
        for link in (case.get("linked_collateral", []) or []):
            if link.get("collateral_type") in req:
                return True
        return False

    @staticmethod
    def _stale_valuations(case: dict, matrix) -> list:
        """Collateral links whose valuation_date is older than the configured
        max age. Lenient: links WITHOUT a valuation_date are not flagged."""
        import datetime as _dt
        max_age = int(getattr(matrix, "valuation_max_age_days", 365) or 365)
        cutoff = (_dt.date.today() - _dt.timedelta(days=max_age)).isoformat()
        stale = []
        for link in (case.get("linked_collateral", []) or []):
            vd = link.get("valuation_date")
            if vd and str(vd) < cutoff:
                stale.append(link.get("collateral_id"))
        return stale

    @staticmethod
    def evaluate_disbursement_gate(case: dict, matrix=None, as_of: str = None) -> dict:
        """The disbursement readiness gate. Returns {passed, failures[], secured}.
        Unsecured facilities check only mandatory Conditions Precedent (the
        affordability path is unchanged). Secured facilities additionally require
        cleared legal review, perfected security, valid insurance (where
        required), coverage at/above the policy threshold, and fresh valuations.
        """
        failures = []
        secured = str(case.get("facility_security_type", "unsecured")) == "secured"

        cp = CreditAdminManager.outstanding_mandatory_cp(case)
        if cp:
            failures.append({"check": "conditions_precedent",
                             "reason": f"{len(cp)} mandatory condition(s) precedent outstanding",
                             "needed": cp})
        if secured:
            from utils.collateral_coverage import CreditPolicyMatrix
            m = matrix or CreditPolicyMatrix()
            if CreditAdminManager.legal_blocks_disbursement(case):
                failures.append({"check": "legal_review",
                                 "reason": "legal review not cleared",
                                 "needed": "outcome approved or approved_with_conditions"})
            if CreditAdminManager.perfection_blocks_disbursement(case):
                failures.append({"check": "security_perfection",
                                 "reason": "security not fully perfected",
                                 "needed": "every instrument perfection_status=perfected"})
            ins_req = CreditAdminManager._insurance_required_for(case, m)
            if CreditAdminManager.insurance_blocks_disbursement(case, as_of, ins_req):
                failures.append({"check": "insurance",
                                 "reason": "no valid insurance (active, bank interest noted, unexpired)",
                                 "needed": "valid insurance policy"})
            cls = str(case.get("security_classification", "unsecured"))
            if cls in ("unsecured", "partially_secured"):
                failures.append({"check": "coverage",
                                 "reason": f"coverage classification is {cls}",
                                 "needed": "fully_secured or over_secured",
                                 "coverage_ratio": case.get("coverage_ratio"),
                                 "required_ratio": case.get("required_ratio")})
            stale = CreditAdminManager._stale_valuations(case, m)
            if stale:
                failures.append({"check": "valuation",
                                 "reason": "collateral valuation stale",
                                 "needed": f"revalue within {m.valuation_max_age_days} days",
                                 "items": stale})

        # An authorized override that covers the current failures clears the gate.
        ov = case.get("perfection_override") or {}
        overridden = False
        if failures and ov.get("status") == "authorized":
            bypassed = {f.get("check") for f in (ov.get("failures_bypassed") or [])}
            current = {f["check"] for f in failures}
            if current <= bypassed:
                overridden = True

        return {"passed": (len(failures) == 0) or overridden,
                "failures": failures, "secured": secured,
                "overridden": overridden}

    # Override authority tiers
    @staticmethod
    def _override_role(user: dict):
        role = str(user.get("role", "") or "").lower()
        if "head of credit" in role or "head_of_credit" in role:
            return "head_of_credit"
        if "chief risk" in role or role == "cro" or "risk officer" in role:
            return "cro"
        if "managing director" in role or "chief executive" in role:
            return "md"
        return None

    def is_high_value(self, case: dict, matrix=None) -> bool:
        from utils.collateral_coverage import CreditPolicyMatrix
        m = matrix or CreditPolicyMatrix()
        amt = case.get("amount_kes")
        if amt is None:
            amt = case.get("amount", 0)
        try:
            return float(amt or 0) >= float(m.high_value_threshold_kes)
        except Exception:
            return False

    def request_perfection_override(self, case_id: str, by: str,
                                    justification: str, failures: list) -> bool:
        if not str(justification or "").strip():
            return False
        for case in self.cases:
            if case["id"] != case_id:
                continue
            case["perfection_override"] = {
                "status":            "pending",
                "requested_by":      by,
                "requested_at":      datetime.now().isoformat(timespec="seconds"),
                "justification":     justification,
                "failures_bypassed": failures or [],
                "approvals":         [],
            }
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def add_override_approval(self, case_id: str, role: str, approver: str,
                              matrix=None) -> dict:
        """Add an approval. Returns {ok, status, required_roles, have_roles}.
        Authorizes when the tier's required role set is satisfied:
          standard   -> head_of_credit OR cro
          high_value -> head_of_credit AND cro AND md
        """
        for case in self.cases:
            if case["id"] != case_id:
                continue
            ov = case.get("perfection_override")
            if not ov or ov.get("status") not in ("pending", "authorized"):
                return {"ok": False, "reason": "no pending override request"}
            approvals = ov.setdefault("approvals", [])
            if not any(a.get("role") == role for a in approvals):
                approvals.append({"role": role, "approver": approver,
                                  "at": datetime.now().isoformat(timespec="seconds")})
            have = {a["role"] for a in approvals}
            hv = self.is_high_value(case, matrix)
            # Pilot affordance: an 'admin' approval is a documented superuser
            # override and satisfies any tier (flagged in design doc + audit).
            if "admin" in have:
                satisfied = True
                required = {"admin"}
            elif hv:
                # High-value: ALL THREE of Head of Credit, CRO, MD.
                required = {"head_of_credit", "cro", "md"}
                satisfied = required <= have
            else:
                # Standard: ANY ONE of Head of Credit, CRO, or MD. The MD
                # outranks HoC/CRO and is always an acceptable approver.
                required = {"head_of_credit", "cro", "md"}
                satisfied = bool(have & required)
            if satisfied:
                ov["status"] = "authorized"
                ov["authorized_at"] = datetime.now().isoformat(timespec="seconds")
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return {"ok": True, "status": ov["status"], "high_value": hv,
                    "required_roles": sorted(required), "have_roles": sorted(have)}
        return {"ok": False, "reason": "case not found"}

    def _two_layer_enabled(self) -> bool:
        """Whether the Credit-Admin two-layer authorization policy is on
        (admin config). Defaults to on. Best-effort; no hard dependency."""
        try:
            from utils.api_lms_mutations import get_credit_workflow_config
            return bool(get_credit_workflow_config().get(
                "credit_admin_two_layer_authorization", True))
        except Exception:
            return True

    def request_authorization(self, case_id: str, by: str,
                              note: str = "") -> bool:
        """Layer 1 — a credit-admin officer confirms the case is ready and
        requests manager authorization. Requires all conditions met."""
        for case in self.cases:
            if case["id"] != case_id:
                continue
            if not case.get("all_conditions_met"):
                return False
            if case.get("disbursed"):
                return False
            case["authorization_requested"]    = True
            case["authorization_requested_by"] = by
            case["authorization_requested_at"] = datetime.now().isoformat()
            case["authorization_note"]         = note
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def authorize(self, case_id: str, by: str, note: str = "") -> bool:
        """Layer 2 — a credit-admin MANAGER authorizes disbursement. Requires
        a pending authorization request. Sets ready_for_disbursement."""
        for case in self.cases:
            if case["id"] != case_id:
                continue
            if not case.get("authorization_requested"):
                return False
            if case.get("disbursed"):
                return False
            case["authorized"]              = True
            case["authorized_by"]           = by
            case["authorized_at"]           = datetime.now().isoformat()
            case["authorization_decision_note"] = note
            case["ready_for_disbursement"]  = True
            case["last_updated"] = datetime.now().date().isoformat()
            self.save()
            return True
        return False

    def clear_for_disbursement(self, case_id: str) -> bool:
        for case in self.cases:
            if case["id"] == case_id and case.get("ready_for_disbursement"):
                # RELEASE to Treasury Back Office (Troops) for actual
                # disbursement. The disbursed=True flip now happens in the
                # Troops workflow (book -> value-date -> disburse), not here —
                # matching this module's stated intent that the fund transfer is
                # a separate finance-system step. (Previously this set
                # disbursed=True + disbursement_date directly.)
                case["cleared_for_disbursement"] = True
                case["cleared_at"]         = datetime.now().isoformat()
                case["last_updated"]       = datetime.now().date().isoformat()
                self.save()
                return True
        return False

    def create_case_from_application(self, app: dict, conditions=None,
                                      authority: str = "") -> str:
        """Create a credit-admin case from an approved LMS application.

        P2 (2026-06-12): the live LMS-approval -> credit-admin handoff.
        Idempotent — the case id is deterministically 'CA'+application_id
        (matching generate_lms_data.py), so a re-approval or retry no-ops
        instead of duplicating or resetting an in-flight case. Returns the
        case id ('' if the app has no id).
        """
        app_id = str(app.get("id", "") or "")
        if not app_id:
            return ""
        case_id = f"CA{app_id}"
        if self.get(case_id):
            # Already handed off — never clobber fulfilled conditions or a
            # disbursed flag on retry/re-approval.
            return case_id

        today = datetime.now().date().isoformat()
        cond_src = conditions if conditions is not None else (
            app.get("decision", {}).get("conditions", []) or []
        )
        case_conditions = [{
            "type":      str(c),
            "required":  True,
            "fulfilled": False,
            # P4-2: CP/CS first-class. Default every condition to a mandatory
            # Condition Precedent (safest: blocks disbursement until credit
            # reclassifies). Credit can move a condition to "subsequent"
            # (tracked post-disbursement, non-blocking) or mark non-mandatory.
            "classification": "precedent",   # "precedent" | "subsequent"
            "mandatory":      True,
            "due_date":       None,          # used by subsequent conditions
            "date_set":  today,
            "date_met":  None,
            "officer":   None,
            "notes":     "",
        } for c in cond_src]
        all_met = all(c["fulfilled"] for c in case_conditions)  # True iff none

        self.cases.append({
            "id":                     case_id,
            "application_id":         app_id,
            "client_name":            app.get("client_name", ""),
            "product":                app.get("product", ""),
            "amount":                 app.get("amount", 0),
            "currency":               str(app.get("currency", "") or "KES"),
            "amount_kes":             app.get("amount_kes"),
            "currency_book":          app.get("currency_book"),
            "fx_rate":                app.get("fx_rate"),
            "rm_code":                app.get("rm_code", ""),
            "rm_name":                app.get("rm_name", ""),
            "approval_date":          today,
            "conditions":             case_conditions,
            "all_conditions_met":     all_met,
            # P4-2: facility security classification (drives perfection routing,
            # enforced at the disburse gate in P4-6). Derived default from the
            # application; "unsecured" until credit classifies.
            "facility_security_type": str(app.get("facility_security_type", "") or "unsecured"),
            "security_subtype":       app.get("security_subtype"),
            "ready_for_disbursement": False,  # set on authorize (two-layer) or all-met
            "authorization_requested": False,
            "authorization_requested_by": None,
            "authorization_requested_at": None,
            "authorized":             False,
            "authorized_by":          None,
            "authorized_at":          None,
            "disbursed":              False,
            "disbursement_date":      None,
            "last_updated":           today,
            "authority":              authority,
        })
        self.save()
        return case_id


class ComplianceManager:
    """Persist compliance_cases.json."""
    def __init__(self):
        self.file = DATA_DIR / "compliance_cases.json"
        self.cases = self._load()

    def _load(self) -> list:
        try:
            return json.loads(self.file.read_text()) if self.file.exists() else []
        except Exception:
            return []

    def save(self):
        self.file.write_text(json.dumps(self.cases, indent=2, default=str))

    def update_status(self, case_id: str, new_status: str,
                       officer: str = "", notes: str = "",
                       escalate_to: str = "") -> bool:
        for case in self.cases:
            if case["id"] != case_id:
                continue
            case["status"]       = new_status
            case["last_updated"] = datetime.now().date().isoformat()
            if notes:   case["review_notes"]  = notes
            if officer: case["assigned_officer"] = officer
            if new_status == "cleared":
                case["cleared_date"] = datetime.now().date().isoformat()
            if new_status == "escalated" and escalate_to:
                case["escalated_to"] = escalate_to
            self.save()
            return True
        return False

    def bsc_compliance_score(self) -> float:
        """Return a compliance score (0-100) for BSC based on case clearance rate."""
        if not self.cases:
            return 85.0
        cleared  = sum(1 for c in self.cases if c["status"] == "cleared")
        rejected = sum(1 for c in self.cases if c["status"] == "rejected")
        resolved = cleared + rejected
        total    = len(self.cases)
        if total == 0:
            return 85.0
        # Score = (resolved/total)*100, capped to reasonable range
        raw = (resolved / total) * 100
        return round(max(60.0, min(99.0, raw)), 1)


class TreasuryManager:
    """Persist treasury_fd.json — FD ratification queue."""
    def __init__(self):
        self.file = DATA_DIR / "treasury_fd.json"
        self.requests = self._load()

    def _load(self) -> list:
        try:
            return json.loads(self.file.read_text()) if self.file.exists() else []
        except Exception:
            return []

    def save(self):
        self.file.write_text(json.dumps(self.requests, indent=2, default=str))

    def get(self, req_id: str) -> dict:
        return next((r for r in self.requests if r["id"] == req_id), {})

    def ratify(self, req_id: str, ratified_rate: float, officer: str,
                counter: bool = False) -> bool:
        for r in self.requests:
            if r["id"] != req_id:
                continue
            r["ratified_rate"]   = ratified_rate
            r["treasury_officer"] = officer
            r["ratified_date"]   = datetime.now().date().isoformat()
            r["status"] = "counter_offered" if counter else "approved"
            if counter:
                r["counter_rate"] = ratified_rate
            self.save()
            return True
        return False

    def book(self, req_id: str) -> bool:
        for r in self.requests:
            if r["id"] == req_id and r["status"] in ("approved","counter_offered"):
                r["status"]      = "booked"
                r["booked_date"] = datetime.now().date().isoformat()
                self.save()
                return True
        return False


# Reporting tree — role → all roles that report into it downward
REPORTING_TREE = {
    "Managing Director": {"tree_roles": None, "units": None},
    "Director Retail Banking": {
        "tree_roles": [
            "Director Retail Banking","Head Of Retail","Regional Head",
            "Branch Manager","Branch Operations Manager","Branch Credit Manager",
            "Teller","Customer Service Officer","Direct Sales Officer",
            "Relationship Officer Personal Banking",
        ],
        "units": None,   # handled specially — all Branch units
    },
    "Director Commercial Banking": {
        "tree_roles": [
            "Director Commercial Banking","Head Of SME","Head Of Corporate",
            "Relationship Manager SME","Relationship Manager Corporate",
            "Credit Analyst","Credit Administrator",
        ],
        "units": ["Commercial Banking","Corporate Banking","SME Banking","Credit"],
    },
    "Head Of Retail": {
        "tree_roles": [
            "Head Of Retail","Regional Head","Branch Manager",
            "Branch Operations Manager","Branch Credit Manager",
            "Teller","Customer Service Officer","Direct Sales Officer",
            "Relationship Officer Personal Banking",
        ],
        "units": None,
    },
    "Head Of Corporate":          {"tree_roles":["Head Of Corporate","Relationship Manager Corporate"],"units":["Corporate Banking"]},
    "Head Of SME":                {"tree_roles":["Head Of SME","Relationship Manager SME"],"units":["SME Banking"]},
    "Head Of Digital Innovation": {"tree_roles":["Head Of Digital Innovation","IT Manager","IT Support Officer"],"units":["Digital & Channels","ICT"]},
    "Head Of Strategy":           {"tree_roles":["Head Of Strategy","Strategy Analyst"],"units":["Strategy"]},
    "Head Of Internal Audit":     {"tree_roles":["Head Of Internal Audit","Internal Auditor"],"units":["Internal Audit"]},
    "Head Of Marketing":          {"tree_roles":["Head Of Marketing","Marketing Officer"],"units":["Marketing"]},
    "Chief Finance Officer":      {"tree_roles":["Chief Finance Officer","Financial Controller","Treasury Manager"],"units":["Finance","Treasury"]},
    "Chief Risk Officer":         {"tree_roles":["Chief Risk Officer","Risk Manager"],"units":["Risk"]},
    "Chief Operations Officer":   {"tree_roles":["Chief Operations Officer","Operations Manager","Branch Operations Manager"],"units":["Operations"]},
    "Chief Compliance Officer":   {"tree_roles":["Chief Compliance Officer","Compliance Officer","Legal Counsel"],"units":["Compliance & Legal"]},
    "Chief Human Resources Officer":{"tree_roles":["Chief Human Resources Officer","HR Business Partner","HR Officer"],"units":["Human Resources"]},
    "Chief Credit Officer":       {"tree_roles":["Chief Credit Officer","Credit Analyst","Credit Administrator"],"units":["Credit"]},
    "Debt Recovery Unit Manager": {"tree_roles":["Debt Recovery Unit Manager","Recovery Officer"],"units":None},
    "Procurement Manager":        {"tree_roles":["Procurement Manager","Procurement Officer"],"units":["Procurement"]},
    "IT Manager":                 {"tree_roles":["IT Manager","IT Support Officer"],"units":["ICT"]},
    "Operations Manager":         {"tree_roles":["Operations Manager","Branch Operations Manager"],"units":None},
    "HR Business Partner":        {"tree_roles":["HR Business Partner","HR Officer"],"units":["Human Resources"]},
    "Regional Head": {
        "tree_roles": [
            "Regional Head","Branch Manager","Branch Operations Manager","Branch Credit Manager",
            "Branch Operations Supervisor","Teller","Customer Service Officer",
            "Direct Sales Officer","Relationship Officer Personal Banking",
            "Relationship Officer Business Banking",
        ],
        "units": None,   # scoped by region (matched via Region column)
    },
    "Branch Manager": {
        "tree_roles": [
            "Branch Manager","Branch Operations Manager","Branch Credit Manager",
            "Branch Operations Supervisor","Teller","Customer Service Officer",
            "Direct Sales Officer","Relationship Officer Personal Banking",
            "Relationship Officer Business Banking",
        ],
        "units": None,   # scoped by unit (branch) in _UNIT_SCOPED_ROLES
    },
    "Branch Operations Manager": {
        "tree_roles": [
            "Branch Operations Manager","Branch Operations Supervisor",
            "Teller","Customer Service Officer",
        ],
        "units": None,
    },
    "Branch Credit Manager": {
        "tree_roles": [
            "Branch Credit Manager","Relationship Officer Personal Banking",
            "Relationship Officer Business Banking","Direct Sales Officer",
        ],
        "units": None,
    },
    "Branch Operations Supervisor": {
        "tree_roles": ["Branch Operations Supervisor","Teller","Customer Service Officer"],
        "units": None,
    },
}

_UNIT_SCOPED_ROLES = {
    "branch manager","branch operations manager","branch credit manager",
    "branch operations supervisor","it manager","operations manager","hr business partner",
    "relationship officer personal banking","relationship officer business banking",
    "direct sales officer","customer service officer","teller",
}
# Regional Heads scope by Region column, not Unit
_REGION_SCOPED_ROLES = {"regional head"}
_ALL_VIEW_ROLES = {
    "managing director",
    "admin",
    # H4 (2026-06-14): the canonical top-exec role is "Chief Executive &
    # Managing Director" (role_taxonomy), NOT "Managing Director". Without
    # these entries the CEO fell through to self-only visibility and could
    # not see all deals. Variants included for resilience to title spelling.
    "chief executive & managing director",
    "chief executive officer",
    "chief executive",
}


# ─── PROFILE PHOTO HELPERS ────────────────────────────────────────────
import base64 as _b64

PHOTO_DIR = DATA_DIR / "profile_photos"

def save_profile_photo(staff_code: str, image_bytes: bytes, ext: str = "jpg") -> str:
    """Save a staff profile photo. Returns the relative path."""
    PHOTO_DIR.mkdir(exist_ok=True)
    safe_code = str(staff_code).strip().replace("/","_")
    path = PHOTO_DIR / f"{safe_code}.{ext}"
    path.write_bytes(image_bytes)
    return str(path)

def get_photo_b64(staff_code: str, staff_name: str = "") -> str:
    """Return base64-encoded photo for a staff member, or empty string if none."""
    PHOTO_DIR.mkdir(exist_ok=True)
    safe_code = str(staff_code).strip().replace("/","_")
    for ext in ("jpg","jpeg","png","webp"):
        p = PHOTO_DIR / f"{safe_code}.{ext}"
        if p.exists():
            data = _b64.b64encode(p.read_bytes()).decode()
            return f"data:image/{ext};base64,{data}"
    # Fallback: try by name slug
    if staff_name:
        slug = staff_name.lower().replace(" ","_").replace(".","")[:20]
        for ext in ("jpg","jpeg","png"):
            p = PHOTO_DIR / f"{slug}.{ext}"
            if p.exists():
                data = _b64.b64encode(p.read_bytes()).decode()
                return f"data:image/{ext};base64,{data}"
    return ""

def photo_avatar_html(staff_code: str, staff_name: str,
                       size: int = 44, initials_fallback: bool = True) -> str:
    """Return <img> or initials-circle HTML for a staff member's photo."""
    uri = get_photo_b64(staff_code, staff_name)
    if uri:
        return (f"<img src='{uri}' width='{size}' height='{size}' "
                f"style='border-radius:50%;object-fit:cover;flex-shrink:0' />")
    if initials_fallback:
        parts = str(staff_name).strip().split()
        ini   = ((parts[0][0]+parts[-1][0]).upper()
                 if len(parts)>=2 else (staff_name[:2].upper() if staff_name else "?"))
        return (f"<div style='width:{size}px;height:{size}px;border-radius:50%;"
                f"background:linear-gradient(135deg,var(--brand-primary,#006B3F),var(--brand-mid,#1D9E75);"
                f"display:flex;align-items:center;justify-content:center;"
                f"color:var(--color-background-primary);font-size:{size//3}px;font-weight:800;flex-shrink:0'>{ini}</div>")
    return ""


# ── Password policy validator (v10.501 Batch 4a) ─────────────────────
# Closes GAP-001: doctrine and email templates advertised a strong
# password policy (uppercase + lowercase + digit + special character,
# minimum 8) but enforcement at every call site was length-only. This
# helper is the SINGLE SOURCE OF TRUTH for the policy. Called from:
#
#   - pages/_login.py  (voluntary change_pw form + force_change_pw form)
#   - utils/api.py     (/api/auth/change-password endpoint)
#
# CGR1 doctrine: this validator matches what utils/core.py:313 advertises
# in the new-account email template. If the policy ever changes, the
# email template and this function must change together.
#
# SECURITY: never logs, returns, or includes the password in the reason
# string. Only the rule that was violated is named.

# Character classes used by the policy. Defined as module-level constants
# so tests can introspect what the policy enforces without parsing strings.
_PWD_MIN_LENGTH      = 8
_PWD_SPECIAL_CHARS   = "!@#$%^&*()_+-=[]{}|;:'\",.<>/?`~\\"

def validate_password_policy(pw: str) -> tuple:
    """Validate a candidate password against the A2Z policy.

    Returns:
        (ok: bool, reason: str)
        - (True, "")                         on accept
        - (False, "<human-readable reason>") on reject

    Policy (mirrors utils/core.py:313 new-account email template):
        - At least 8 characters
        - Contains at least one uppercase letter (A-Z)
        - Contains at least one lowercase letter (a-z)
        - Contains at least one digit (0-9)
        - Contains at least one special character from _PWD_SPECIAL_CHARS

    The reason string is suitable for direct display to end users
    (Streamlit st.error, FastAPI HTTPException detail). It never
    includes the password itself.
    """
    if not isinstance(pw, str):
        return False, "Password must be a string."
    if len(pw) < _PWD_MIN_LENGTH:
        return False, f"Password must be at least {_PWD_MIN_LENGTH} characters."
    if not any(c.isupper() for c in pw):
        return False, "Password must contain at least one uppercase letter."
    if not any(c.islower() for c in pw):
        return False, "Password must contain at least one lowercase letter."
    if not any(c.isdigit() for c in pw):
        return False, "Password must contain at least one digit."
    if not any(c in _PWD_SPECIAL_CHARS for c in pw):
        return False, ("Password must contain at least one special "
                       "character (e.g. !@#$%^&*).")
    return True, ""


class UserManager:
    # P-AUTH-c: canonical test logins that must ALWAYS be available during
    # the frontend phase. ensure_test_logins() recreates any that are missing
    # on construction — belt-and-suspenders alongside the absolute DATA_DIR
    # and the _load hardening. The full 49-role set is (re)seeded via
    # scripts/seed_test_logins.py; this guarantees only the CEO login can
    # never silently vanish.
    _CANONICAL_TEST_LOGINS = [
        # (username, password, full_name, role, staff_code)
        ("william001", "EcoStaff0001", "William Mwangi",
         "Chief Executive & Managing Director", "0001"),
    ]

    def __init__(self):
        self.users_file = DATA_DIR / "users.json"
        self.users = self._load()
        self.ensure_test_logins()
        self.ensure_branch_test_logins()

    def ensure_test_logins(self) -> int:
        """Recreate any missing canonical test logins. Returns count restored.

        Sources the full per-role set from utils.test_logins (the same source
        scripts/seed_test_logins.py uses, so they cannot drift). If that import
        fails for any reason, falls back to guaranteeing at least the CEO login
        so a taxonomy problem can never lock everyone out.

        Cheap when healthy: membership checks only, no writes.
        """
        try:
            from utils.test_logins import canonical_test_logins
            canon = canonical_test_logins()
        except Exception:
            canon = self._CANONICAL_TEST_LOGINS  # CEO-only fallback

        restored = 0
        for username, password, full_name, role, staff_code in canon:
            if username in self.users:
                continue
            self.add_user(
                username, password, full_name,
                email=f"{username}@bank.com", role=role,
                staff_code=staff_code, can_view_all=True, can_execute=True,
            )
            self.users[username]["_protected"] = True
            self.users[username]["must_change_password"] = False
            restored += 1
        if restored:
            self.save_users()
            try:
                logger.warning(
                    "UserManager self-heal: restored %d canonical test "
                    "login(s)", restored,
                )
            except Exception:
                pass
        return restored

    def ensure_branch_test_logins(self) -> int:
        """Recreate any missing register-staff branch test logins. Returns count.

        Parity with ensure_test_logins: the canonical role logins self-heal, but
        the register branch chain (300xxx, used for scope testing) did not — so
        any users.json reset wiped them while the canonical set returned. This
        restores them from utils.test_logins.branch_test_logins() with the
        correct can_view_all (False except the register root), unit and region.

        Cheap when healthy: membership checks only, no writes.
        """
        try:
            from utils.test_logins import branch_test_logins
            chain = branch_test_logins()
        except Exception:
            return 0
        restored = 0
        for uname, pw, full, role, code, unit, region, cva in chain:
            if uname in self.users:
                continue
            self.add_user(
                uname, pw, full, email=f"{uname}@bank.com", role=role,
                unit=unit, staff_code=code, can_view_all=cva, can_execute=True,
            )
            self.users[uname]["region"] = region
            self.users[uname]["_protected"] = True
            self.users[uname]["must_change_password"] = False
            restored += 1
        if restored:
            self.save_users()
            try:
                logger.warning(
                    "UserManager self-heal: restored %d branch test login(s)",
                    restored,
                )
            except Exception:
                pass
        return restored

    def _load(self):
        # Hardened (P-AUTH-b): a transient read error or a corrupt file must
        # NEVER cause a silent fall-back to defaults, because _defaults()
        # immediately re-saves and would overwrite a recoverable real file
        # (this is exactly what wiped the seeded test logins). We distinguish:
        #   - absent / empty file          -> first run, defaults are correct
        #   - exists but unreadable/corrupt -> back up + fail loud (no overwrite)
        file_exists = self.users_file.exists()
        raw = ""
        if file_exists:
            try:
                raw = self.users_file.read_text(encoding="utf-8")
            except Exception as e:
                self._backup_unreadable("read-error")
                raise RuntimeError(
                    f"users.json exists but could not be read ({e!r}). "
                    "Refusing to overwrite it with defaults; a backup was "
                    "written. Resolve the error or restore a backup."
                ) from e

        if not raw.strip():
            # Genuinely absent or empty -> first-run defaults are correct.
            return self._defaults()

        try:
            users = json.loads(raw)
        except Exception as e:
            self._backup_unreadable("parse-error")
            raise RuntimeError(
                f"users.json is present but not valid JSON ({e!r}). A backup "
                "was written; refusing to overwrite with defaults."
            ) from e

        if not isinstance(users, dict):
            self._backup_unreadable("shape-error")
            raise RuntimeError("users.json parsed to a non-object shape.")
        if not users:
            return self._defaults()

        # Always ensure admin account exists and cannot be permanently removed
        if 'admin' not in users:
            users['admin'] = {
                "password":   self.hash_pw("admin123"),
                "full_name":  "System Admin",
                "role":       "Admin",
                "department": "All",
                "can_view_all": True,
                "managed_roles": [], "managed_units": [],
                "managed_staff_codes": [],
                "staff_code": "ADMIN001",
                "email":      "admin@bank.com",
                "active":     True,
                "_protected": True,
            }
        else:
            users['admin'].update({
                'can_view_all': True,
                'role': 'Admin',
                'active': True,
                '_protected': True,
            })
        self._save(users)
        return users

    def _backup_unreadable(self, tag: str) -> None:
        """Copy an unreadable/corrupt users.json aside before failing loud.

        Best-effort: a backup failure must never mask the original error.
        """
        try:
            if self.users_file.exists():
                import datetime as _dt
                import shutil as _shutil
                ts = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
                bak = self.users_file.with_name(f"users.json.corrupt-{tag}-{ts}")
                _shutil.copy2(self.users_file, bak)
                print(f"[UserManager] users.json unreadable ({tag}); "
                      f"backup written -> {bak}")
        except Exception:
            pass

    def _defaults(self):
        u = {
            "admin": {
                "password": self.hash_pw("admin123"),
                "full_name": "System Admin", "role": "Admin", "department": "All",
                "can_view_all": True, "managed_roles": [], "managed_units": [],
                "managed_staff_codes": [], "staff_code": "ADMIN001",
                "email": "admin@bank.com", "active": True,
            },
            "manager1": {
                "password": self.hash_pw("manager123"),
                "full_name": "John Manager", "role": "Manager", "department": "Retail Banking",
                "can_view_all": False, "managed_roles": [], "managed_units": ["Retail Banking"],
                "managed_staff_codes": [], "staff_code": "MGR001",
                "email": "manager@bank.com", "active": True,
            },
            "staff1": {
                "password": self.hash_pw("staff123"),
                "full_name": "Jane Staff", "role": "Staff", "department": "Retail Banking",
                "can_view_all": False, "managed_roles": [], "managed_units": [],
                "managed_staff_codes": [], "staff_code": "STF001",
                "email": "staff1@bank.com", "active": True,
            },
        }
        self._save(u)
        return u

    def _save(self, u=None):
        self.users_file.write_text(json.dumps(u or self.users, indent=2))

    def save(self):
        """Alias for save_users — keeps all call sites consistent."""
        self.save_users()

    def save_users(self):
        self._save(self.users)

    def can_delete_user(self, username: str) -> tuple:
        """Returns (can_delete: bool, reason: str)."""
        if username == "admin":
            return False, "The admin account is protected and cannot be deleted."
        u = self.users.get(username, {})
        if u.get("_protected"):
            return False, f"'{username}' is a protected account."
        # Count active admins — must always have at least one
        admins = [u2 for u2 in self.users.values()
                  if u2.get("is_admin") or u2.get("role","").lower() == "admin"]
        if len(admins) <= 1 and (u.get("is_admin") or u.get("role","").lower()=="admin"):
            return False, "Cannot delete the last admin account."
        return True, ""

    def delete_user(self, username: str, verified_by: str) -> tuple:
        """Delete user after protection check. Returns (success, message)."""
        can, reason = self.can_delete_user(username)
        if not can:
            return False, reason
        self.users.pop(username, None)
        self.save_users()
        return True, f"User '{username}' deleted by {verified_by}."

    def hash_pw(self, pw: str) -> str:
        """bcrypt hash with SHA-256 fallback. Always use verify_pw to check.

        Delegates to module-level _hash_password() so bootstrap and runtime
        share a single implementation.

        Batch 3b: _hash_password lives in utils.core_audit (extracted there
        when this method was the last user). Deferred import here avoids the
        circular dependency that would result from a top-of-file import
        (core_audit imports back from core).
        """
        from utils.core_audit import _hash_password
        return _hash_password(pw)

    def verify_pw(self, pw: str, stored: str, username: str = "") -> bool:
        """Verify password — handles bcrypt, envelope-bcrypt, and legacy SHA-256.

        Three verification paths tried in order:

          1. Direct bcrypt: stored = bcrypt(password)
             — produced by hash_pw / change_password / add_user, or by
               a successful auto-upgrade in authenticate().
          2. Envelope bcrypt: stored = bcrypt(sha256(password).hex)
             — produced by the Batch 3c migration script (scripts/verify_bcrypt.py
               --upgrade) when wrapping legacy SHA-256 hashes without
               requiring plaintext password recovery.
          3. Legacy SHA-256 direct: stored = sha256(password).hex
             — pre-bcrypt tablestakes; still present for dormant accounts
               not yet logged-in or migrated.

        Returns True on first match. Each bcrypt check costs ~25ms; worst
        case (envelope-stored, wrong password) is ~50ms. Operationally
        irrelevant at this scale.

        Batch 3c additions:
          - $2y$ prefix support (some bcrypt libraries emit this; Python's
            bcrypt emits $2b$, but external systems may produce $2y$).
          - Envelope verification path.
          - INFO-level log on envelope-success (observability for ratchet
            planning — measures how many users still authenticate via the
            transitional envelope path so Phase 2 can plan deprecation).

        The optional `username` kwarg enables the envelope-success log to
        identify the user. Callers WITHOUT username available may omit it;
        existing call sites that don't pass it remain backward compatible.

        SECURITY: this method never logs plaintext password, hash, sha256
        derivation, or bcrypt string. Username + log message only.

        CGR1 doctrine (Batch 3c): envelope is a TRANSITIONAL stabilization
        layer, NOT canonical end-state. Phase 2 hardening may add forced
        normalization, Argon2 migration, etc.
        """
        if not pw or not stored:
            return False
        try:
            import bcrypt as _bc
            if stored.startswith('$2b$') or stored.startswith('$2a$') \
                    or stored.startswith('$2y$'):
                # Path 1: direct bcrypt
                if _bc.checkpw(pw.encode('utf-8'), stored.encode('utf-8')):
                    return True
                # Path 2: envelope — bcrypt of sha256(password) hex string
                sha_hex = hashlib.sha256(pw.encode('utf-8')).hexdigest()
                if _bc.checkpw(sha_hex.encode('utf-8'), stored.encode('utf-8')):
                    if username:
                        logger.info(
                            "Envelope-backed credential authenticated for "
                            "user '%s'", username
                        )
                    else:
                        logger.info("Envelope-backed credential authenticated")
                    return True
                return False
        except ImportError:
            # bcrypt missing — fall through to legacy SHA-256 path.
            pass
        # Path 3: legacy SHA-256 direct comparison
        return hashlib.sha256(pw.encode()).hexdigest() == stored

    def authenticate(self, username, password):
        u = self.users.get(username)
        if not u or not u.get('active'):
            return False, None
        if not self.verify_pw(password, u.get('password', ''), username=username):
            return False, None
        # Batch 3c: Auto-upgrade legacy SHA-256 (raw) to direct bcrypt on
        # successful login. Envelope-bcrypted hashes are already bcrypt
        # and skip this path.
        #
        # Pre-Batch-3b this swallow silently hid a NameError that masked
        # the absence of any auto-upgrade for ~2 years. Batch 3b's hotfix
        # (commit 2aab56b) made hash_pw actually work. This Batch 3c
        # change instruments the swallow so any future hygiene failure
        # surfaces in logs — auth availability remains primary (do NOT
        # re-raise), but migration failure becomes observable.
        #
        # SECURITY: log message contains username + exception class +
        # traceback ONLY. NEVER plaintext password, sha256 hex, bcrypt
        # string, or token material.
        stored = u.get('password', '')
        needs_upgrade = bool(stored) and not (
            stored.startswith('$2b$') or stored.startswith('$2a$')
            or stored.startswith('$2y$')
        )
        if needs_upgrade:
            try:
                u['password'] = self.hash_pw(password)
                self.save_users()
            except Exception as e:
                logger.error(
                    "Auto-upgrade SHA-256 -> bcrypt FAILED for user '%s': "
                    "%s: %s — auth allowed, migration deferred",
                    username, type(e).__name__, e,
                    exc_info=True,
                )
        return True, u

    def change_password(self, username, new_password):
        self.users[username]['password'] = self.hash_pw(new_password)
        self.users[username]['must_change_password'] = False
        self.save_users()

    def get_managed_staff_codes(self, username):
        u = self.users.get(username, {})
        role = str(u.get('role','')).lower()
        if role in ('admin','director') or u.get('can_view_all'):
            return "all"
        codes = u.get('managed_staff_codes', [])
        if codes: return [str(c).strip() for c in codes if c]
        sc = u.get('staff_code','')
        return [sc] if sc else []

    def add_user(self, username, password, full_name, email="",
                 role="Staff", unit="", staff_code="",
                 can_view_all=False, can_execute=False, is_admin=False):
        """Create a new user account with auto-derived module permissions."""
        # Auto-derive module access from role
        accessible = [m for m, cfg in MODULE_ACCESS.items()
                      if role in cfg.get("roles_all",[])
                      or cfg["min"] == "self"
                      or (is_admin)
                      or (can_view_all and cfg["min"] in ("all","team","unit","self"))]
        self.users[username] = {
            "password":    self.hash_pw(password),
            "full_name":   full_name,
            "email":       email,
            "role":        role,
            "unit":        unit,
            "department":  unit,
            "staff_code":  str(staff_code),
            "can_view_all":can_view_all,
            "can_execute": can_execute,
            "is_admin":    is_admin,
            "active":      True,
            "managed_roles": [],
            "managed_units": [unit] if unit else [],
            "managed_staff_codes": [str(staff_code)] if staff_code else [],
            "must_change_password": False,
            "accessible_modules":  accessible,
        }
        self.save_users()
        return self.users[username]

    def filter_data(self, user_data, staff_df):
        role = str(user_data.get('role','')).strip().lower()
        if user_data.get('can_view_all') or role in ('admin','director'):
            return staff_df, f"Full access — {len(staff_df)} staff"

        codes = [str(c).strip() for c in user_data.get('managed_staff_codes',[]) if c]
        roles = user_data.get('managed_roles', [])
        units = user_data.get('managed_units', [])
        sc    = user_data.get('staff_code','')

        mask = pd.Series([False]*len(staff_df), index=staff_df.index)
        if codes and 'Staff Code' in staff_df.columns:
            mask |= staff_df['Staff Code'].astype(str).isin(codes)
        if roles and 'Role' in staff_df.columns:
            mask |= staff_df['Role'].isin(roles)
        if units and 'Unit' in staff_df.columns:
            mask |= staff_df['Unit'].isin(units)
        if units and 'Category' in staff_df.columns:
            mask |= staff_df['Category'].isin(units)
        if sc and 'Staff Code' in staff_df.columns:
            mask |= staff_df['Staff Code'].astype(str) == str(sc)

        filtered = staff_df[mask]
        if len(filtered): return filtered, f"Filtered — {len(filtered)} staff"
        return pd.DataFrame(), "No matching staff — check permissions in Admin panel"

# ─── ADMIN PANEL ─────────────────────────────────────────────────────

# ── Performance: cached summary loaders ──────────────────────────
import streamlit as _st_perf

@_st_perf.cache_data(ttl=300, show_spinner=False)
def get_bsc_summary_cached(period: str = "Feb 2026"):
    """Cached BSC summary — refreshes every 5 minutes."""
    try:
        from utils.db import db as _db
        if _db.table_uses_db("bsc_scores"):
            rows = _db.fetch_all(
                "SELECT dept, COUNT(*) as n, ROUND(AVG(final_score)::numeric,2) as avg "
                "FROM bsc_scores WHERE period=%s GROUP BY dept", (period,))
            total = sum(r.get("n",0) for r in rows)
            avg   = sum(float(r.get("avg",0))*r.get("n",1) for r in rows)/max(total,1)
            return {"by_dept":rows,"total":total,"avg":round(avg,2)}
    except Exception:
        pass
    return {"by_dept":[],"total":0,"avg":0.0}

@_st_perf.cache_data(ttl=60, show_spinner=False)
def get_pipeline_summary_cached():
    """Cached pipeline summary — refreshes every minute."""
    try:
        from utils.db import db as _db
        if _db.table_uses_db("pipeline_deals"):
            return _db.fetch_all(
                "SELECT stage, COUNT(*) as deals, SUM(amount) as value "
                "FROM pipeline_deals GROUP BY stage ORDER BY value DESC")
    except Exception:
        pass
    return []

# ══════════════════════════════════════════════════════════════════
# BSC AUTO-SCORE ENGINE
# Reads from all operational modules and auto-computes KPI actuals
# for head office and operational staff
# ══════════════════════════════════════════════════════════════════

def compute_operational_kpi_actuals(username: str, period: str = "2026") -> dict:
    """
    Compute KPI actuals for a staff member from all operational modules.
    Returns dict of {kpi_id: {"actual": float, "source": str, "detail": str}}
    
    Called by the BSC page and by the nightly score refresh.
    """
    import json
    from pathlib import Path as _Path
    from datetime import date as _date, timedelta as _td
    from decimal import Decimal as _D

    DATA = _Path(__file__).parent.parent / "data"
    today = _date.today()
    actuals = {}

    def _safe_float(v):
        try:
            if isinstance(v, _D): return float(v)
            return float(v) if v is not None else 0.0
        except: return 0.0

    def _load(fname):
        p = DATA / fname
        return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []

    # ── K036: Projects On-Time Delivery (%) ──────────────────────
    # ── K037: Milestones Completed (count) ───────────────────────
    # ── K038: Project Budget Adherence (%) ───────────────────────
    try:
        projects = _load("projects.json")
        if not isinstance(projects, list):
            projects = projects.get("projects", [])
        
        # Filter to projects where this user is PM or milestone/action owner
        my_projects = [p for p in projects
                       if p.get("owner_username") == username
                       or p.get("project_manager","").lower() in username.lower()]
        
        # Milestones owned by this user
        my_milestones = []
        for proj in projects:
            for ms in proj.get("milestones", []):
                if ms.get("owner_username") == username:
                    my_milestones.append(ms)
        
        # K037: Milestones completed
        ms_done = sum(1 for ms in my_milestones if ms.get("status") == "Complete")
        actuals["K037"] = {
            "actual": float(ms_done),
            "source": "projects",
            "detail": f"{ms_done} milestones completed out of {len(my_milestones)} assigned"
        }
        
        # K036: On-time delivery
        if my_projects:
            completed = [p for p in my_projects if p.get("status") in ("Completed","Closed")]
            on_time   = [p for p in completed
                         if p.get("actual_end_date","") and p.get("planned_end_date","")
                         and p.get("actual_end_date","") <= p.get("planned_end_date","")]
            pct = round(len(on_time)/max(len(completed),1)*100, 1)
            actuals["K036"] = {
                "actual": pct,
                "source": "projects",
                "detail": f"{len(on_time)}/{len(completed)} projects delivered on time"
            }
        
        # K038: Budget adherence
        if my_projects:
            within_budget = [p for p in my_projects
                            if _safe_float(p.get("spent_m",0)) <= _safe_float(p.get("budget_m",1))]
            pct = round(len(within_budget)/max(len(my_projects),1)*100, 1)
            actuals["K038"] = {
                "actual": pct,
                "source": "projects",
                "detail": f"{len(within_budget)}/{len(my_projects)} projects within budget"
            }
        
        # Action items
        my_actions = []
        for proj in projects:
            for act in proj.get("action_items", []):
                if act.get("owner_username") == username:
                    my_actions.append(act)
        if my_actions:
            closed = sum(1 for a in my_actions if a.get("status") == "Closed")
            actuals["_action_items"] = {
                "actual": round(closed/max(len(my_actions),1)*100, 1),
                "source": "projects",
                "detail": f"{closed}/{len(my_actions)} action items closed"
            }
    except Exception as _e:
        pass

    # ── K039: Tickets Resolved Within SLA (%) ────────────────────
    # ── K040: Open Ticket Age (avg days) ─────────────────────────
    try:
        tickets_raw = _load("cims_tickets.json")
        tickets = tickets_raw if isinstance(tickets_raw, list) else tickets_raw.get("tickets", [])
        my_tickets = [t for t in tickets if t.get("assigned_to","") == username
                      or t.get("owner","") == username]
        
        if my_tickets:
            resolved = [t for t in my_tickets if t.get("status","") in ("Resolved","Closed")]
            sla_met   = [t for t in resolved if t.get("sla_status","") in ("Met","On Time","")]
            sla_pct   = round(len(sla_met)/max(len(resolved),1)*100, 1)
            actuals["K039"] = {
                "actual": sla_pct,
                "source": "sla_tickets",
                "detail": f"{len(sla_met)}/{len(resolved)} tickets resolved within SLA"
            }
            
            open_t = [t for t in my_tickets if t.get("status","") == "Open"]
            if open_t:
                ages = []
                for t in open_t:
                    try:
                        created = _date.fromisoformat(str(t.get("created_at",""))[:10])
                        ages.append((today - created).days)
                    except: ages.append(0)
                avg_age = round(sum(ages)/max(len(ages),1), 1)
                actuals["K040"] = {
                    "actual": avg_age,
                    "source": "sla_tickets",
                    "detail": f"Average {avg_age} days for {len(open_t)} open tickets"
                }
    except Exception: pass

    # ── K041: Pipeline Deals Progressed ──────────────────────────
    # ── K042: Deal Win Rate (%) ───────────────────────────────────
    try:
        from utils.db import db as _db
        if _db.table_uses_db("pipeline_deals"):
            deals = _db.fetch_all(
                "SELECT * FROM pipeline_deals WHERE staff_code = %s",
                (str(username),)
            )
        else:
            raw = _load("pipeline.json")
            deals = raw if isinstance(raw, list) else raw.get("deals", [])
            deals = [d for d in deals if d.get("staff_code","") == username
                     or d.get("rm_code","") == username]
        
        if deals:
            won   = [d for d in deals if d.get("stage","") == "Closed Won"]
            lost  = [d for d in deals if d.get("stage","") == "Closed Lost"]
            closed = won + lost
            win_rate = round(len(won)/max(len(closed),1)*100, 1)
            actuals["K041"] = {
                "actual": float(len(deals)),
                "source": "pipeline",
                "detail": f"{len(deals)} deals in pipeline, {len(won)} won"
            }
            actuals["K042"] = {
                "actual": win_rate,
                "source": "pipeline",
                "detail": f"{len(won)}/{len(closed)} deals closed won"
            }
    except Exception: pass

    # ── K043: MOU Activations ─────────────────────────────────────
    # ── K044: Referral Conversion Rate (%) ───────────────────────
    try:
        mous = _load("partnerships_mous.json")
        my_mous = [m for m in mous if m.get("relationship_manager","") == username]
        total_activations = sum(m.get("activations_ytd",0) for m in my_mous)
        actuals["K043"] = {
            "actual": float(total_activations),
            "source": "partnerships",
            "detail": f"{total_activations} MOU activations from {len(my_mous)} MOUs"
        }
        
        refs = _load("referrals.json")
        my_refs = [r for r in refs if r.get("rm_assigned","") == username
                   or r.get("referrer_code","") == username]
        if my_refs:
            converted = sum(1 for r in my_refs if r.get("converted"))
            conv_rate = round(converted/max(len(my_refs),1)*100, 1)
            actuals["K044"] = {
                "actual": conv_rate,
                "source": "referrals",
                "detail": f"{converted}/{len(my_refs)} referrals converted"
            }
    except Exception: pass

    # ── K045: Loan TAT Compliance (%) ────────────────────────────
    # ── K046: Credit Analysis Completeness (%) ───────────────────
    try:
        from utils.db import db as _db
        if _db.table_uses_db("loan_applications"):
            loans = _db.fetch_all(
                "SELECT * FROM loan_applications WHERE rm_code = %s OR analyst = %s",
                (username, username)
            )
        else:
            raw = _load("loan_applications.json")
            loans = [l for l in (raw if isinstance(raw,list) else [])
                     if l.get("rm_code","") == username or l.get("analyst","") == username]
        
        if loans:
            with_tat = [l for l in loans if l.get("tat_days") and l.get("sla_target_days")]
            tat_met  = [l for l in with_tat
                        if _safe_float(l.get("tat_days",999)) <= _safe_float(l.get("sla_target_days",1))]
            tat_pct  = round(len(tat_met)/max(len(with_tat),1)*100, 1)
            actuals["K045"] = {
                "actual": tat_pct,
                "source": "loan_applications",
                "detail": f"{len(tat_met)}/{len(with_tat)} loans within TAT"
            }
            completeness = [_safe_float(l.get("completeness_score",0)) for l in loans if l.get("completeness_score")]
            if completeness:
                avg_comp = round(sum(completeness)/len(completeness), 1)
                actuals["K046"] = {
                    "actual": avg_comp,
                    "source": "loan_applications",
                    "detail": f"Average {avg_comp}% completeness on {len(completeness)} applications"
                }
    except Exception: pass

    # ── K047: EWS Cases Resolved (%) ─────────────────────────────
    try:
        from utils.db import db as _db
        if _db.table_uses_db("ews_cases"):
            ews = _db.fetch_all("SELECT * FROM ews_cases WHERE rm_code = %s", (username,))
        else:
            ews = [e for e in _load("ews_cases.json") if e.get("rm_code","") == username]
        
        if ews:
            resolved = [e for e in ews if e.get("stage","") in ("Resolved","Upgraded","Closed")]
            res_pct  = round(len(resolved)/max(len(ews),1)*100, 1)
            actuals["K047"] = {
                "actual": res_pct,
                "source": "ews_cases",
                "detail": f"{len(resolved)}/{len(ews)} EWS cases resolved"
            }
    except Exception: pass

    # ── K049: AML Cases Closed (%) ───────────────────────────────
    # ── K050: STRs Filed ─────────────────────────────────────────
    try:
        from utils.db import db as _db
        if _db.table_uses_db("aml_alerts"):
            aml = _db.fetch_all("SELECT * FROM aml_alerts WHERE assigned_to ILIKE %s",
                               (f"%{username}%",))
        else:
            aml = [a for a in _load("aml_alerts.json")
                   if username in str(a.get("assigned_to",""))]
        
        if aml:
            closed = [a for a in aml if a.get("status","") in ("Cleared","Closed","Referred to FRC")]
            closed_pct = round(len(closed)/max(len(aml),1)*100, 1)
            strs = sum(1 for a in aml if a.get("str_filed"))
            actuals["K049"] = {
                "actual": closed_pct,
                "source": "aml_alerts",
                "detail": f"{len(closed)}/{len(aml)} AML alerts closed"
            }
            actuals["K050"] = {
                "actual": float(strs),
                "source": "aml_alerts",
                "detail": f"{strs} STRs filed"
            }
    except Exception: pass

    # ── K051: PRs Processed Within TAT (%) ───────────────────────
    try:
        prs = _load("purchase_requests.json")
        my_prs = [r for r in prs if r.get("requested_by","") == username
                  or r.get("approved_by","") == username]
        if my_prs:
            approved = [r for r in my_prs if r.get("status","") in ("Approved","Ordered","Paid")]
            actuals["K051"] = {
                "actual": round(len(approved)/max(len(my_prs),1)*100, 1),
                "source": "purchase_requests",
                "detail": f"{len(approved)}/{len(my_prs)} PRs approved"
            }
    except Exception: pass

    # ── K053: Branch Log Submission Rate (%) ─────────────────────
    try:
        branch_logs = _load("branch_logs.json")
        if isinstance(branch_logs, dict):
            branch_logs = list(branch_logs.values())
        my_logs = [l for l in branch_logs if l.get("submitted_by","") == username]
        if my_logs:
            submitted = [l for l in my_logs if l.get("status","") in ("Submitted","Validated")]
            pct = round(len(submitted)/max(len(my_logs),1)*100, 1)
            actuals["K053"] = {
                "actual": pct,
                "source": "branch_log",
                "detail": f"{len(submitted)}/{len(my_logs)} logs submitted on time"
            }
    except Exception: pass

    # ── K055-K057: Clearing & Settlement ─────────────────────────
    try:
        clearing = _load("clearing_records.json")
        my_clearing = [c for c in clearing if c.get("officer_username") == username
                       or c.get("created_by") == username]
        if my_clearing:
            total = len(my_clearing)
            exceptions = sum(1 for c in my_clearing if c.get("status") in ("Failed","Rejected","Exception"))
            reconciled = sum(1 for c in my_clearing if c.get("reconciled"))
            settled_tat = sum(1 for c in my_clearing if c.get("settlement_tat_met"))
            actuals["K055"] = {"actual": round(exceptions/max(total,1)*100,1),
                               "source":"clearing","detail":f"{exceptions}/{total} exceptions"}
            actuals["K056"] = {"actual": round(reconciled/max(total,1)*100,1),
                               "source":"clearing","detail":f"{reconciled}/{total} reconciled"}
            actuals["K057"] = {"actual": round(settled_tat/max(total,1)*100,1),
                               "source":"clearing","detail":f"{settled_tat}/{total} within TAT"}
    except Exception: pass

    # ── K058-K059: Consent Management ────────────────────────────
    try:
        consent = _load("consent_register.json")
        my_consent = [c for c in consent if c.get("recorded_by") == username]
        total_cust = len(set(c.get("customer_id","") for c in consent))
        valid = sum(1 for c in consent if c.get("status") == "Active")
        renewals = sum(1 for c in my_consent if c.get("action") == "Renewed")
        actuals["K058"] = {"actual": round(valid/max(total_cust,1)*100,1),
                           "source":"consent","detail":f"{valid}/{total_cust} customers with valid consent"}
        actuals["K059"] = {"actual": float(renewals),"source":"consent",
                           "detail":f"{renewals} consent renewals processed"}
    except Exception: pass

    # ── K060-K062: Retailer Finance ───────────────────────────────
    try:
        rf = _load("retailer_finance.json")
        my_rf = [r for r in rf if r.get("rm_username") == username
                 or r.get("created_by") == username]
        if my_rf:
            total_disb = sum(_safe_float(r.get("disbursed_m",0)) for r in my_rf)
            npl_count  = sum(1 for r in my_rf if _safe_float(r.get("dpd",0))>=90)
            buyers     = sum(1 for r in my_rf if r.get("status")=="Active" and r.get("new_buyer"))
            actuals["K060"] = {"actual":round(total_disb,1),"source":"retailer_finance",
                               "detail":f"KES {total_disb:.1f}M disbursed"}
            actuals["K061"] = {"actual":round(npl_count/max(len(my_rf),1)*100,1),
                               "source":"retailer_finance","detail":f"{npl_count}/{len(my_rf)} NPL"}
            actuals["K062"] = {"actual":float(buyers),"source":"retailer_finance",
                               "detail":f"{buyers} new buyers onboarded"}
    except Exception: pass

    # ── K063-K065: Bid Bond & Guarantees ─────────────────────────
    try:
        bonds = _load("bid_bonds.json")
        my_bonds = [b for b in bonds if b.get("officer_username") == username
                    or b.get("created_by") == username]
        if my_bonds:
            issued = len(my_bonds)
            commission = sum(_safe_float(b.get("commission_kes",0)) for b in my_bonds)/1e6
            managed = sum(1 for b in my_bonds
                         if b.get("status") in ("Expired","Returned","Renewed") or
                            (_date.today().isoformat() < b.get("expiry_date","")))
            actuals["K063"] = {"actual":float(issued),"source":"bid_bond",
                               "detail":f"{issued} bonds issued"}
            actuals["K064"] = {"actual":round(commission,2),"source":"bid_bond",
                               "detail":f"KES {commission:.1f}M commission"}
            actuals["K065"] = {"actual":round(managed/max(issued,1)*100,1),
                               "source":"bid_bond","detail":f"{managed}/{issued} managed proactively"}
    except Exception: pass

    # ── K066-K068: System Observability ──────────────────────────
    try:
        obs = _load("observability_metrics.json")
        my_obs = [o for o in obs if o.get("owner_username") == username
                  or o.get("assigned_to") == username]
        if obs:  # Bank-wide metrics — use all for IT staff
            uptime = sum(_safe_float(o.get("uptime_pct",0)) for o in obs)/max(len(obs),1)
            incidents = [o for o in obs if o.get("type") == "Incident"]
            sla_met = sum(1 for i in incidents if i.get("sla_met"))
            mttr_vals = [_safe_float(i.get("resolution_mins",0)) for i in incidents if i.get("resolution_mins")]
            mttr = sum(mttr_vals)/max(len(mttr_vals),1) if mttr_vals else 0
            actuals["K066"] = {"actual":round(uptime,2),"source":"observability",
                               "detail":f"Average uptime across {len(obs)} systems"}
            actuals["K067"] = {"actual":round(sla_met/max(len(incidents),1)*100,1),
                               "source":"observability","detail":f"{sla_met}/{len(incidents)} incidents within SLA"}
            actuals["K068"] = {"actual":round(mttr,0),"source":"observability",
                               "detail":f"Average {mttr:.0f} mins to resolve"}
    except Exception: pass

    # ── K069-K071: Channels Management ───────────────────────────
    try:
        channels = _load("channels_data.json")
        if channels:
            mobile = next((c for c in channels if "mobile" in c.get("channel","").lower()), {})
            digital_txn = [c for c in channels if c.get("type") == "Transaction"]
            successful  = sum(1 for c in digital_txn if c.get("success"))
            adoption = [c for c in channels if c.get("metric") == "adoption"]
            actuals["K069"] = {"actual":_safe_float(mobile.get("uptime_pct",99.0)),
                               "source":"channels","detail":"Mobile banking uptime"}
            actuals["K070"] = {"actual":round(successful/max(len(digital_txn),1)*100,1),
                               "source":"channels","detail":f"{successful}/{len(digital_txn)} txns successful"}
            adopt_rate = _safe_float(next((c.get("value",40) for c in adoption),40))
            actuals["K071"] = {"actual":adopt_rate,"source":"channels",
                               "detail":f"{adopt_rate:.0f}% customers on digital channels"}
    except Exception: pass

    # ── K109/K110/K111: FLEXCUBE Integration Health (auto-computed) ──
    try:
        from utils import flexcube_adapter as _fcx
        _h = _fcx.health_check()
        _services = _h.get("services",{})
        _n = len(_services)
        _up = sum(1 for _s in _services.values() if "Up" in str(_s.get("status","")) or "Mocked" in str(_s.get("status","")))
        _uptime = round(_up/max(_n,1)*100, 1)
        actuals["K109"] = {"actual": _uptime, "source": "flexcube",
                           "detail": f"{_up}/{_n} services up"}

        # K110 — error count from event journal
        _evt_log = DATA / "flexcube_events.json"
        _err_24h = 0
        if _evt_log.exists():
            try:
                _events = json.loads(_evt_log.read_text(encoding="utf-8"))
                from datetime import datetime as _dt
                _cutoff = _dt.utcnow() - _td(hours=24)
                for _e in _events:
                    try:
                        _t = _dt.fromisoformat(_e["timestamp"].replace("Z",""))
                        if _t > _cutoff and "error" in (_e.get("payload",{}) or {}):
                            _err_24h += 1
                    except: pass
            except: pass
        actuals["K110"] = {"actual": _err_24h, "source": "flexcube",
                           "detail": "errors in last 24h"}

        # K111 — minutes since last event
        _lag_min = 0
        if _evt_log.exists():
            try:
                _events = json.loads(_evt_log.read_text(encoding="utf-8"))
                if _events:
                    from datetime import datetime as _dt
                    _last = _dt.fromisoformat(_events[0]["timestamp"].replace("Z",""))
                    _lag_min = round((_dt.utcnow() - _last).total_seconds()/60, 1)
            except: pass
        actuals["K111"] = {"actual": _lag_min, "source": "flexcube",
                           "detail": "minutes since last event"}
    except Exception:
        pass

    return actuals


def _legacy_period_to_engine(legacy: str) -> str:
    """Translate legacy period strings into the BSC engine's canonical
    'YYYY-MM' format. Accepts 'Feb 2026', 'Feb-26', '2026-02', '2026-Q1',
    falling back to current YYYY-MM if parsing fails.

    Used by update_bsc_from_modules to bridge to bsc_engine.submit_batch.
    """
    from datetime import datetime as _dt
    if not isinstance(legacy, str) or not legacy.strip():
        return _dt.today().strftime("%Y-%m")
    s = legacy.strip()
    # Already canonical?
    if len(s) == 7 and s[4] == "-":
        return s  # "2026-02" or "2026-Q1"
    # "Feb 2026"
    for fmt in ("%b %Y", "%b-%y", "%b-%Y", "%B %Y"):
        try:
            return _dt.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return _dt.today().strftime("%Y-%m")


def update_bsc_from_modules(username: str, period: str = "Feb 2026") -> dict:
    """
    Update a staff member's BSC KPI scores from operational modules.
    Returns updated score dict. Call this when a staff member completes work.
    
    Used by:
    - Project module (milestone completed)
    - SLA/CIMS module (ticket closed)
    - Pipeline module (deal won/lost)
    - Loan Applications (application processed)
    - EWS (case resolved)
    - AML (alert closed)
    - Branch Log (log submitted)
    """
    import json
    from pathlib import Path as _Path
    
    DATA = _Path(__file__).parent.parent / "data"
    scores_file = DATA / "feb_2026_staff_scores.json"
    
    if not scores_file.exists():
        return {}
    
    scores = json.loads(scores_file.read_text(encoding="utf-8"))
    if username not in scores:
        return {}
    
    user_score = scores[username]
    kpi_lib_data = json.loads((DATA / "kpi_library.json").read_text(encoding="utf-8"))
    role_kpis_map = kpi_lib_data.get("role_kpis", {})
    all_kpis = {k["id"]: k for k in kpi_lib_data.get("kpis", [])}
    
    # Get this user's KPIs
    role = user_score.get("role", "")
    user_role_kpis = role_kpis_map.get(role, [])
    
    if not user_role_kpis:
        return user_score
    
    # Compute actuals from modules
    actuals = compute_operational_kpi_actuals(username, period[:4])
    
    if not actuals:
        return user_score
    
    # ── BSC engine pilot #2: stamp every actual through the contract ──
    # Each entry in `actuals` is already a (kpi_id, value) pair tagged with
    # a `source` field. Translate to the universal contract and submit
    # through utils/bsc_engine. Failures are non-blocking — the legacy
    # kpi_scores update below still runs even if the engine rejects records.
    try:
        from utils.bsc_engine import submit_batch as _bsc_submit_batch
        _staff_code = str(user_score.get("staff_code", "") or "").strip()
        _engine_period = _legacy_period_to_engine(period)
        if _staff_code and _engine_period:
            _bsc_records = []
            for _kid, _payload in actuals.items():
                _val = _payload.get("actual")
                if _val is None:
                    continue
                _bsc_records.append({
                    "staff_code": _staff_code,
                    "kpi_id":     _kid,
                    "value":      _val,
                    "period":     _engine_period,
                    "metadata":   {
                        "original_source": str(_payload.get("source", "")),
                        "detail":          str(_payload.get("detail", ""))[:200],
                    },
                })
            if _bsc_records:
                _bsc_submit_batch(
                    records       = _bsc_records,
                    source_module = "operational_modules",
                    actor         = username,
                )
    except Exception as _e:
        # Engine failures must never block the legacy update path.
        pass
    
    # Update kpi_scores with new actuals
    kpi_scores = user_score.get("kpi_scores", {})
    updated_kpis = []
    
    for kpi_id in user_role_kpis:
        if kpi_id not in actuals:
            continue
        
        kpi_def = all_kpis.get(kpi_id, {})
        actual  = actuals[kpi_id].get("actual", 0)
        
        # Get target from BSC targets (if set) or use default
        existing = kpi_scores.get(kpi_id, {})
        target   = existing.get("target", _get_default_target(kpi_id, kpi_def))
        
        if not target:
            continue
        
        # Compute achievement %
        direction = kpi_def.get("direction", "higher_better")
        if direction == "lower_better" and target > 0:
            ach_pct = round(target / max(actual, 0.001) * 100, 1)
        elif target > 0:
            ach_pct = round(actual / target * 100, 1)
        else:
            ach_pct = 0.0
        
        # Score 1-5
        score = (5.0 if ach_pct >= 120 else
                 4.5 if ach_pct >= 110 else
                 4.0 if ach_pct >= 100 else
                 3.5 if ach_pct >= 90  else
                 3.0 if ach_pct >= 80  else
                 2.5 if ach_pct >= 70  else
                 2.0 if ach_pct >= 60  else
                 1.5 if ach_pct >= 50  else 1.0)
        
        kpi_scores[kpi_id] = {
            **existing,
            "actual":       actual,
            "target":       target,
            "achievement_pct": ach_pct,
            "score":        score,
            "source":       actuals[kpi_id].get("source", "module"),
            "detail":       actuals[kpi_id].get("detail", ""),
            "auto_updated": True,
        }
        updated_kpis.append(kpi_id)
    
    if updated_kpis:
        user_score["kpi_scores"] = kpi_scores
        # Recompute final score
        weights = [all_kpis.get(k,{}).get("weight",0.05) for k in user_role_kpis if k in kpi_scores]
        scores_list = [kpi_scores[k].get("score",3.0) for k in user_role_kpis if k in kpi_scores]
        if weights and scores_list:
            total_w = sum(weights)
            if total_w > 0:
                weighted = sum(s*w for s,w in zip(scores_list, weights))
                user_score["final_score"] = round(min(weighted/total_w, 5.0), 2)
        
        scores[username] = user_score
        scores_file.write_text(json.dumps(scores, indent=2))
    
    return user_score


def _get_default_target(kpi_id: str, kpi_def: dict) -> float:
    """Default targets for operational KPIs when no cascade target is set."""
    defaults = {
        "K036": 80.0,   # 80% on-time delivery
        "K037": 5.0,    # 5 milestones per period
        "K038": 90.0,   # 90% within budget
        "K039": 85.0,   # 85% tickets within SLA
        "K040": 7.0,    # 7 days avg open ticket age
        "K041": 10.0,   # 10 deals progressed
        "K042": 40.0,   # 40% win rate
        "K043": 5.0,    # 5 MOU activations
        "K044": 50.0,   # 50% referral conversion
        "K045": 80.0,   # 80% loan TAT compliance
        "K046": 85.0,   # 85% completeness
        "K047": 75.0,   # 75% EWS resolution
        "K049": 80.0,   # 80% AML cases closed
        "K050": 2.0,    # 2 STRs filed
        "K051": 85.0,   # 85% PRs within TAT
        "K053": 90.0,   # 90% log submission rate
    }
    return defaults.get(kpi_id, kpi_def.get("default_target", 80.0))
