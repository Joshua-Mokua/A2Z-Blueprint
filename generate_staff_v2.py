"""
generate_staff_v2.py — Full bank staff generator based on real Kenyan bank structure.
Uses real roles and departments but 100% synthetic names.
Produces:
  cbs_data/staff_register_v2.xlsx
  cbs_data/users_v2.json  → copy to a2z/data/users.json
  cbs_data/staff_kpi_map_v2.json

Structure: 1,755 staff | 94 branches | 10 regions | 9 C-suite | ~60 HO roles
"""
import json, random, hashlib, csv
from pathlib import Path
from datetime import date, timedelta

random.seed(2026)
CBS  = Path(__file__).parent / "cbs_data"
DATA = Path(__file__).parent / "a2z" / "data"
CBS.mkdir(exist_ok=True); DATA.mkdir(exist_ok=True)

# ── Name pools (Kenyan, diverse) ────────────────────────────────────
FIRST_M = [
    "James","John","Peter","David","Paul","Joseph","Michael","Patrick","Francis","Daniel",
    "Samuel","Philip","George","Charles","Robert","William","Thomas","Richard","Edward","Kevin",
    "Dennis","Eric","Stephen","Brian","Victor","Moses","Aaron","Simon","Emmanuel","Kenneth",
    "Leonard","Alex","Mark","Anthony","Calvin","Edwin","Festus","Gilbert","Isaac","Joel",
    "Keith","Lawrence","Martin","Nathan","Oscar","Phillip","Raymond","Solomon","Timothy","Usman",
    "Vincent","Walter","Zachary","Baraka","Caleb","Derek","Elijah","Frank","Geoffrey","Hassan",
    "Ishmael","Julius","Kelvin","Levi","Maurice","Nicholas","Oliver","Pius","Rodgers","Stanley",
    "Bernard","Clinton","Dominic","Evans","Felix","Gregory","Herbert","Ibrahim","Jasper","Kurt",
    "Lewis","Mathew","Norman","Obed","Quinn","Reuben","Sebastian","Tobias","Ulrich","Vaughan",
]
FIRST_F = [
    "Mary","Grace","Faith","Ann","Rose","Jane","Susan","Patricia","Elizabeth","Catherine",
    "Agnes","Beatrice","Charity","Diana","Esther","Florence","Gloria","Hannah","Irene","Joyce",
    "Karen","Lydia","Margaret","Nancy","Olive","Priscilla","Rebecca","Sarah","Teresa","Alice",
    "Bernadette","Caroline","Daisy","Ellen","Fatuma","Gladys","Harriet","Ida","Jacqueline",
    "Kezia","Lilian","Miriam","Naomi","Olivia","Pauline","Rachael","Stella","Tabitha","Veronica",
    "Winnie","Yvonne","Zipporah","Aisha","Brenda","Cynthia","Dorcas","Eunice","Felicia",
    "Georgina","Helena","Isabella","Janet","Kemunto","Lucy","Mildred","Njeri","Peninah",
    "Ruth","Sharon","Tina","Vivian","Wendy","Xenia","Yasmin","Zainab","Amina","Blessing",
    "Christine","Doreen","Edith","Fridah","Hilda","Immaculate","Judith","Ketty","Lorna",
]
SURNAMES = [
    "Kamau","Odhiambo","Wanjiku","Mwangi","Otieno","Kariuki","Mutua","Njoroge","Akinyi","Ochieng",
    "Kiprotich","Chebet","Langat","Ruto","Koech","Bett","Sang","Kemboi","Chepkemoi","Mutai",
    "Waweru","Ngugi","Githu","Gicheru","Kabui","Njenga","Ndegwa","Mungai","Murage","Kinyanjui",
    "Ouma","Ogolla","Okello","Anyango","Adhiambo","Awino","Auma","Odongo","Omondi","Owino",
    "Hassan","Omar","Abdullahi","Mohamed","Ali","Sheikh","Issa","Abdi","Farah","Maalim",
    "Mwenda","Kithinji","Kirimi","Murithi","Njue","Mugambi","Karithi","Gakunju","Micheni","Ntiba",
    "Wafula","Simiyu","Barasa","Masinde","Wekesa","Khaemba","Nasambu","Nekesa","Nafula","Khisa",
    "Mwamba","Mwanzia","Munyao","Ndeti","Musyoka","Mutiso","Muema","Nzuki","Matheka","Mbatha",
    "Shimba","Mwanake","Katana","Kazungu","Ngumbao","Sidi","Mwaburi","Karisa","Baya","Charo",
    "Onyango","Okeyo","Achola","Awuor","Aluoch","Omolo","Odero","Oguda","Obiero","Asande",
    "Ndung'u","Kimani","Gitonga","Nderitu","Gachau","Muthoni","Njagi","Kuria","Weru","Maina",
    "Kibet","Rotich","Chirchir","Ngetich","Komen","Biwott","Yego","Lagat","Kiptoo","Sigei",
    "Wanjala","Namukhula","Andanje","Wekesa","Mulama","Makokha","Wanyama","Wetungu","Amolo",
]

used_names = set()

def gen_name(gender=None):
    if gender is None:
        gender = random.choice(['M','F'])
    for _ in range(200):
        f = random.choice(FIRST_M if gender=='M' else FIRST_F)
        s = random.choice(SURNAMES)
        full = f"{f} {s}"
        if full not in used_names:
            used_names.add(full)
            return full, gender
    # fallback with number suffix
    f = random.choice(FIRST_M if gender=='M' else FIRST_F)
    s = random.choice(SURNAMES)
    full = f"{f} {s}{random.randint(2,9)}"
    used_names.add(full)
    return full, gender

def rand_doe(start_yr=2005, end_yr=2024):
    s = date(start_yr,1,1); e = date(end_yr,12,31)
    return s + timedelta(days=random.randint(0,(e-s).days))

# ── BRANCH LIST with regions ─────────────────────────────────────────
BRANCHES = [
    # (branch_name, region, county, tier)
    # Tier: 1=large(flagship), 2=medium, 3=standard, 4=light
    # Nairobi CBD Region
    ("Kenyatta Avenue","Nairobi CBD","Nairobi",1),
    ("FB Towers Retail","Nairobi CBD","Nairobi",1),
    ("FB Towers Corporate","Nairobi CBD","Nairobi",1),
    ("River Road","Nairobi CBD","Nairobi",2),
    ("River Road West","Nairobi CBD","Nairobi",2),
    ("City Hall","Nairobi CBD","Nairobi",2),
    ("Cargen House","Nairobi CBD","Nairobi",2),
    ("Ngara","Nairobi CBD","Nairobi",2),
    ("Industrial Area","Nairobi CBD","Nairobi",2),
    ("Laptrust","Nairobi CBD","Nairobi",2),
    ("Moi Avenue","Nairobi CBD","Nairobi",2),
    # Nairobi Metro Region
    ("Kilimani","Nairobi Metro","Nairobi",2),
    ("Westlands","Nairobi Metro","Nairobi",2),
    ("Kasarani","Nairobi Metro","Nairobi",2),
    ("Donholm","Nairobi Metro","Nairobi",2),
    ("Eastleigh","Nairobi Metro","Nairobi",2),
    ("Kitengela","Nairobi Metro","Kajiado",2),
    ("Utawala","Nairobi Metro","Nairobi",2),
    ("Kayole","Nairobi Metro","Nairobi",3),
    ("Mlolongo","Nairobi Metro","Machakos",3),
    ("Gateway Mall","Nairobi Metro","Nairobi",2),
    ("Kahawa West","Nairobi Metro","Nairobi",3),
    ("Kangemi","Nairobi Metro","Nairobi",3),
    ("Kariobangi","Nairobi Metro","Nairobi",3),
    ("JKIA","Nairobi Metro","Nairobi",3),
    # Nairobi North Region
    ("Wangige","Nairobi North","Kiambu",3),
    ("Kikuyu","Nairobi North","Kiambu",3),
    ("Banana","Nairobi North","Kiambu",3),
    ("Dagoretti","Nairobi North","Nairobi",3),
    ("Ruaka","Nairobi North","Kiambu",3),
    ("KTDA Retail","Nairobi North","Kiambu",2),
    ("Gikomba","Nairobi North","Nairobi",2),
    ("Gikomba Area 42","Nairobi North","Nairobi",3),
    ("Rongai","Nairobi North","Kajiado",3),
    ("Naivasha","Nairobi North","Nakuru",3),
    ("Kajiado","Nairobi North","Kajiado",3),
    ("Narok","Nairobi North","Narok",3),
    ("Makongeni","Nairobi North","Nairobi",3),
    # Mt Kenya West Region
    ("Thika","Mt Kenya West","Kiambu",1),
    ("Ruiru","Mt Kenya West","Kiambu",2),
    ("Githurai","Mt Kenya West","Kiambu",2),
    ("Kiambu","Mt Kenya West","Kiambu",2),
    ("Limuru","Mt Kenya West","Kiambu",2),
    ("Githunguri","Mt Kenya West","Kiambu",3),
    ("Kangema","Mt Kenya West","Murang'a",3),
    ("Muranga","Mt Kenya West","Murang'a",3),
    ("Kangari","Mt Kenya West","Murang'a",3),
    # Mt Kenya East Region
    ("Nyeri","Mt Kenya East","Nyeri",2),
    ("Othaya","Mt Kenya East","Nyeri",3),
    ("Nyahururu","Mt Kenya East","Laikipia",3),
    ("Ol Kalau","Mt Kenya East","Nyandarua",3),
    ("Kiriaini","Mt Kenya East","Murang'a",3),
    ("Meru","Mt Kenya East","Meru",2),
    ("Nkubu","Mt Kenya East","Meru",3),
    ("Chuka","Mt Kenya East","Tharaka-Nithi",3),
    ("Karatina","Mt Kenya East","Nyeri",3),
    ("Nanyuki","Mt Kenya East","Laikipia",3),
    ("Kerugoya","Mt Kenya East","Kirinyaga",3),
    ("Kutus","Mt Kenya East","Kirinyaga",3),
    ("Maua","Mt Kenya East","Meru",3),
    # Eastern Region
    ("Embu","Eastern","Embu",2),
    ("Machakos","Eastern","Machakos",2),
    ("Kitui","Eastern","Kitui",3),
    ("Wote","Eastern","Makueni",3),
    ("Mwea","Eastern","Kirinyaga",3),
    ("Isiolo","Eastern","Isiolo",3),
    # Coast Region
    ("Mombasa Kenyatta Avenue","Coast","Mombasa",1),
    ("Mombasa Nkrumah","Coast","Mombasa",2),
    ("Mtwapa","Coast","Kilifi",2),
    ("Ukunda","Coast","Kwale",2),
    ("Digo","Coast","Mombasa",2),
    ("Kongowea","Coast","Mombasa",3),
    ("Malindi","Coast","Kilifi",2),
    ("Nyali","Coast","Mombasa",2),
    # North Rift Region
    ("Eldoret","North Rift","Uasin Gishu",1),
    ("Eldoret West","North Rift","Uasin Gishu",2),
    ("Kapsabet","North Rift","Nandi",2),
    ("Kitale","North Rift","Trans Nzoia",2),
    ("Nakuru Finance","North Rift","Nakuru",1),
    ("Nakuru Market","North Rift","Nakuru",2),
    # South Rift Region
    ("Kericho","South Rift","Kericho",2),
    ("Bomet","South Rift","Bomet",3),
    ("Litein","South Rift","Kericho",3),
    ("Molo","South Rift","Nakuru",3),
    ("Kisii","South Rift","Kisii",2),
    ("Nyamira","South Rift","Nyamira",3),
    # West Kenya Region
    ("Kisumu Express","West Kenya","Kisumu",1),
    ("Kakamega","West Kenya","Kakamega",2),
    ("Bungoma","West Kenya","Bungoma",2),
    ("Busia","West Kenya","Busia",3),
    ("Mumias","West Kenya","Kakamega",3),
    ("Migori","West Kenya","Migori",3),
    ("Kagwe","West Kenya","Kiambu",3),
    ("Gatundu","West Kenya","Kiambu",3),
]

print(f"Branches: {len(BRANCHES)}")

# ── BRANCH ROLE TEMPLATES by tier ───────────────────────────────────
# Based on real data averages
# (role, reports_to, kpi_category, band, T1, T2, T3, T4)
BRANCH_ROLES = [
    ("Senior Branch Manager",           "Area Manager",                    "Management","M5",1,0,0,0),
    ("Branch Manager",                  "Area Manager",                    "Management","M4",0,1,1,1),
    ("Branch Operations Manager",       "Branch Manager",                  "Operations","M3",1,1,1,1),
    ("Branch Operations Supervisor",    "Branch Operations Manager",       "Operations","M2",2,1,1,1),
    ("Teller",                          "Branch Operations Supervisor",    "Operations","M1",4,3,2,2),
    ("Customer Service Officer",        "Branch Operations Supervisor",    "Operations","M1",2,2,1,1),
    ("Senior Digital Channels Officer", "Branch Operations Manager",       "Digital",   "M2",1,1,1,0),
    ("Branch Senior Relationship Officer","Branch Manager",                "Sales",     "M3",1,1,0,0),
    ("Branch Relationship Manager",     "Branch Manager",                  "Sales",     "M2",1,1,1,0),
    ("Relationship Officer-Business Banker","Branch Relationship Manager", "Sales",     "M2",2,1,1,1),
    ("Relationship Officer-Personal Banker","Branch Relationship Manager", "Sales",     "M2",1,1,1,0),
    ("Direct Sales Representative - Assets & Liabilities","Branch Manager","Sales",    "M1",1,1,0,0),
    ("Relationship Officer Bancassurance","Branch Manager",                "Bancassurance","M1",0,0,1,0),
]

# ── HO STRUCTURE ────────────────────────────────────────────────────
HO_ROLES = [
    # C-Suite
    ("Chief Executive & Managing Director",  None,                              "Executive","E1",1),
    ("Chief Retail Banking Officer",         "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Commercial Officer",             "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Financial Officer",              "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Credit Officer",                 "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Risk Officer",                   "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Information Officer",            "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Operating Officer",              "Chief Executive & Managing Director","Executive","E2",1),
    ("Chief Human Resource Officer",         "Chief Executive & Managing Director","Executive","E2",1),
    ("Company Secretary and Chief Legal Officer","Chief Executive & Managing Director","Executive","E2",1),
    # Retail Banking
    ("Head of Branches",                     "Chief Retail Banking Officer",    "Retail","M5",1),
    ("Area Manager",                         "Head of Branches",                "Retail","M5",10),
    ("Senior Manager Direct Sales Force",    "Chief Retail Banking Officer",    "Retail","M5",1),
    ("Head Of Women Banking",                "Chief Retail Banking Officer",    "Retail","M4",1),
    ("Senior Manager Diaspora Banking",      "Chief Retail Banking Officer",    "Retail","M4",1),
    ("Relationship Manager - Diaspora",      "Senior Manager Diaspora Banking", "Retail","M3",2),
    # Commercial Banking
    ("Head Of Corporates & Trade Finance",   "Chief Commercial Officer",        "Commercial","M5",1),
    ("Head of MSME",                         "Chief Commercial Officer",        "Commercial","M5",1),
    ("Head of Government & Institutional Banking","Chief Commercial Officer",   "Commercial","M5",1),
    ("Senior Relationship Manager - Corporate Banking","Head Of Corporates & Trade Finance","Commercial","M4",4),
    ("Relationship Manager - Corporate Banking","Head Of Corporates & Trade Finance","Commercial","M3",5),
    ("Assistant Relationship Manager-Corporate","Relationship Manager - Corporate Banking","Commercial","M2",2),
    ("Senior Relationship Manager- SME",     "Head of MSME",                    "Commercial","M4",2),
    ("Relationship Manager - SME",           "Head of MSME",                    "Commercial","M3",5),
    ("Relationship Manager - Agribusiness",  "Head of MSME",                    "Commercial","M3",2),
    ("Relationship Manager- Trade Finance",  "Head Of Corporates & Trade Finance","Commercial","M3",3),
    ("Senior Relationship Manager-Trade Finance Specialist","Head Of Corporates & Trade Finance","Commercial","M4",2),
    ("Relationship Manager- Public Sector",  "Head of Government & Institutional Banking","Commercial","M3",4),
    ("Relationship Manager - Institutional Banking","Head of Government & Institutional Banking","Commercial","M3",2),
    # DFS
    ("Head of Digital Financial Services",   "Chief Commercial Officer",        "Digital","M5",1),
    ("Manager Agency Banking",               "Head of Digital Financial Services","Digital","M4",1),
    ("Manager Mobile Banking",               "Head of Digital Financial Services","Digital","M4",1),
    ("Manager Card Operations",              "Head of Digital Financial Services","Digital","M4",1),
    ("Operations Supervisor-DFS",            "Head of Digital Financial Services","Digital","M3",1),
    ("Senior Digital Channels Officer",      "Head of Digital Financial Services","Digital","M2",5),
    ("Acquiring Officer",                    "Manager Agency Banking",           "Digital","M1",4),
    # Credit
    ("Senior Manager -Credit Analysis",     "Chief Credit Officer",             "Credit","M5",1),
    ("Corporate Analysis Manager",           "Senior Manager -Credit Analysis",  "Credit","M4",1),
    ("Consumer and Staff Loan Analysis Manager","Senior Manager -Credit Analysis","Credit","M4",1),
    ("Credit Analyst",                       "Corporate Analysis Manager",       "Credit","M3",7),
    ("Assistant Manager -Credit Administration","Chief Credit Officer",          "Credit","M4",1),
    ("Credit Admin Officer",                 "Assistant Manager -Credit Administration","Credit","M2",8),
    ("Manager-Credit Monitoring",            "Chief Credit Officer",             "Credit","M4",1),
    ("Supervisor Credit Reporting",          "Manager-Credit Monitoring",        "Credit","M3",1),
    ("Senior Manager-Collections & Recoveries","Chief Credit Officer",           "Credit","M5",1),
    ("Collections and Recoveries Officer",   "Senior Manager-Collections & Recoveries","Credit","M2",4),
    ("Write-Off Officer",                    "Senior Manager-Collections & Recoveries","Credit","M2",3),
    # Finance
    ("Financial Controller-Senior Manager",  "Chief Financial Officer",          "Finance","M5",1),
    ("Finance Manager & Money Laundering Reporting Officer","Chief Financial Officer","Finance","M4",1),
    ("Finance Officer",                      "Financial Controller-Senior Manager","Finance","M2",5),
    ("Tax Manager",                          "Chief Financial Officer",          "Finance","M4",1),
    ("Business Analytics Manager",           "Chief Financial Officer",          "Finance","M4",1),
    ("Business Analytics Officer",           "Business Analytics Manager",       "Finance","M2",4),
    # Risk & Compliance
    ("Risk Manager",                         "Chief Risk Officer",               "Risk","M4",1),
    ("Operational Risk Manager",             "Chief Risk Officer",               "Risk","M4",1),
    ("Senior Manager- Compliance",           "Chief Risk Officer",               "Risk","M5",1),
    ("Regulatory Compliance Officer",        "Senior Manager- Compliance",       "Risk","M3",1),
    ("Reconciliation Supervisor",            "Chief Operating Officer",          "Operations","M3",1),
    ("Reconciliation Officer",               "Reconciliation Supervisor",        "Operations","M2",3),
    # ICT
    ("Head Of ICT",                          "Chief Information Officer",        "ICT","M5",1),
    ("Database Manager",                     "Head Of ICT",                      "ICT","M4",1),
    ("Network Manager",                      "Head Of ICT",                      "ICT","M4",1),
    ("Manager Core Banking Support",         "Head Of ICT",                      "ICT","M4",1),
    ("Core Banking Support Officer",         "Manager Core Banking Support",     "ICT","M2",4),
    ("ICT Support Officer",                  "Head Of ICT",                      "ICT","M1",6),
    ("PHP Software Developer",               "Head Of ICT",                      "ICT","M2",5),
    ("System Administrator",                 "Head Of ICT",                      "ICT","M3",1),
    ("Cyber Security SOC Analyst",           "Head Of ICT",                      "ICT","M3",4),
    ("Business Analyst Officer",             "Head Of ICT",                      "ICT","M2",3),
    # Operations
    ("Head of Operations",                   "Chief Operating Officer",          "Operations","M5",1),
    ("Central Processing Manager",           "Head of Operations",               "Operations","M4",1),
    ("Operations Officer",                   "Central Processing Manager",       "Operations","M2",3),
    ("Manager - Clearing",                   "Head of Operations",               "Operations","M4",1),
    ("Clearing Officer",                     "Manager - Clearing",               "Operations","M2",4),
    ("Cash Centre Manager",                  "Head of Operations",               "Operations","M4",1),
    ("Cash Centre Supervisor",               "Cash Centre Manager",              "Operations","M3",3),
    # Human Resources
    ("Human Resource Business Partner- Operations","Chief Human Resource Officer","HR","M4",1),
    ("Human Resource Business Partner-Payroll","Chief Human Resource Officer",   "HR","M4",1),
    ("Human Resource Business Partner- Adminstration","Chief Human Resource Officer","HR","M4",1),
    ("Senior Human Resource Business Partner -OSH","Chief Human Resource Officer","HR","M4",1),
    ("Senior Human Resource Business Partner- Performance & HRIS","Chief Human Resource Officer","HR","M5",1),
    ("Senior Human Resource Business Partner-Training","Chief Human Resource Officer","HR","M4",1),
    ("Human Resource Officer Admin",         "Human Resource Business Partner- Operations","HR","M2",2),
    # Treasury
    ("Head of Treasury",                     "Chief Financial Officer",          "Treasury","M5",1),
    ("Senior Manager Treasury",              "Head of Treasury",                 "Treasury","M4",1),
    ("Manager Forex Trader",                 "Head of Treasury",                 "Treasury","M4",1),
    ("Corporate Sales Dealer",               "Head of Treasury",                 "Treasury","M3",2),
    ("Treasury Dealer",                      "Head of Treasury",                 "Treasury","M3",1),
    ("Treasury Front Office Officer",        "Head of Treasury",                 "Treasury","M2",1),
    # Trade Finance
    ("Trade Finance Back Office Manager",    "Head Of Corporates & Trade Finance","Commercial","M4",1),
    ("Senior Trade Finance Officer",         "Trade Finance Back Office Manager", "Commercial","M3",2),
    ("Trade Finance Officer",                "Trade Finance Back Office Manager", "Commercial","M2",2),
    ("Trade Finance Operations Officer",     "Trade Finance Back Office Manager", "Commercial","M2",2),
    # Bancassurance
    ("General Manager - Bancassurance",      "Chief Commercial Officer",         "Bancassurance","M5",1),
    ("Manager Underwriting",                 "General Manager - Bancassurance",  "Bancassurance","M4",1),
    ("Bancassurance Officer",                "Manager Underwriting",             "Bancassurance","M2",5),
    # Customer Service / Contact Centre
    ("Head, Customer Experience",            "Chief Operating Officer",          "CustomerService","M5",1),
    ("Contact Centre Officer",               "Head, Customer Experience",        "CustomerService","M2",19),
    ("Senior Relationship Officer - Diaspora Banking","Chief Retail Banking Officer","Retail","M3",2),
    # Audit & Legal
    ("Senior Manager Internal Audit",        "Chief Executive & Managing Director","Audit","M5",1),
    ("Internal Auditor",                     "Senior Manager Internal Audit",    "Audit","M3",2),
    ("Manager- Legal",                       "Company Secretary and Chief Legal Officer","Legal","M4",1),
    ("Legal Officer",                        "Manager- Legal",                   "Legal","M2",4),
    # Procurement & Admin
    ("Head of Procurement",                  "Chief Operating Officer",          "Operations","M5",1),
    ("Procurement Officer",                  "Head of Procurement",              "Operations","M2",3),
    ("Facilities and Property Manager",      "Head of Procurement",              "Operations","M4",1),
    ("Facilities Officer",                   "Facilities and Property Manager",  "Operations","M2",2),
    # Marketing
    ("Head Of Marketing and Corporate Communication","Chief Executive & Managing Director","Marketing","M5",1),
    ("Marketing Officer",                    "Head Of Marketing and Corporate Communication","Marketing","M2",2),
    ("Marketing Assistant Manager",          "Head Of Marketing and Corporate Communication","Marketing","M4",1),
]

print(f"HO roles defined: {len(HO_ROLES)}")

# ── KPI ASSIGNMENTS — deposit growth for ALL branch/sales roles ─────
KPI_ASSIGNMENTS = {
    # BRANCH ROLES — every role that touches customers has deposit target
    "Senior Branch Manager": {
        "Deposit Growth":0.20,"Loan Book Growth":0.15,"Fees and Commission":0.10,
        "DFS Revenue":0.05,"NPL Ratio":0.10,"New Customer Acquisition":0.10,
        "Active Account Growth":0.05,"Transactions":0.05,"Compliance Score":0.05,
        "Dormancy Reactivation":0.05,"Diligence Score":0.10},
    "Branch Manager": {
        "Deposit Growth":0.20,"Loan Book Growth":0.15,"Fees and Commission":0.10,
        "DFS Revenue":0.05,"NPL Ratio":0.10,"New Customer Acquisition":0.10,
        "Active Account Growth":0.05,"Transactions":0.05,"Compliance Score":0.05,
        "Dormancy Reactivation":0.05,"Diligence Score":0.10},
    "Branch Operations Manager": {
        "Deposit Growth":0.15,"Transactions":0.20,"Active Account Growth":0.15,
        "Dormancy Reactivation":0.15,"Compliance Score":0.15,
        "SLA Adherence Score":0.10,"Diligence Score":0.10},
    "Branch Operations Supervisor": {
        "Deposit Growth":0.10,"Transactions":0.30,"Active Account Growth":0.15,
        "Dormancy Reactivation":0.20,"Compliance Score":0.15,"Diligence Score":0.10},
    "Teller": {
        "Deposit Growth":0.15,"Transactions":0.45,"Compliance Score":0.20,
        "New Customer Acquisition":0.10,"Diligence Score":0.10},
    "Customer Service Officer": {
        "Deposit Growth":0.15,"New Customer Acquisition":0.20,"Transactions":0.25,
        "Dormancy Reactivation":0.20,"Compliance Score":0.10,"Diligence Score":0.10},
    "Senior Digital Channels Officer": {
        "Deposit Growth":0.15,"Digital Active Customers":0.25,"DFS Revenue":0.25,
        "New Customer Acquisition":0.15,"Transactions":0.10,"Diligence Score":0.10},
    "Branch Senior Relationship Officer": {
        "Deposit Growth":0.25,"Loan Book Growth":0.20,"Fees and Commission":0.15,
        "New Customer Acquisition":0.15,"NPL Ratio":0.10,"Compliance Score":0.05,
        "Diligence Score":0.10},
    "Branch Relationship Manager": {
        "Deposit Growth":0.20,"Loan Book Growth":0.20,"Fees and Commission":0.15,
        "New Customer Acquisition":0.15,"NPL Ratio":0.10,"Active Account Growth":0.10,
        "Diligence Score":0.10},
    "Relationship Officer-Business Banker": {
        "Deposit Growth":0.20,"Loan Book Growth":0.25,"Fees and Commission":0.15,
        "NPL Ratio":0.15,"New Customer Acquisition":0.10,"Compliance Score":0.05,
        "Diligence Score":0.10},
    "Relationship Officer-Personal Banker": {
        "Deposit Growth":0.30,"Loan Book Growth":0.20,"Fees and Commission":0.15,
        "New Customer Acquisition":0.15,"Active Account Growth":0.05,
        "Compliance Score":0.05,"Diligence Score":0.10},
    "Direct Sales Representative - Assets & Liabilities": {
        "Deposit Growth":0.25,"Loan Book Growth":0.20,"New Customer Acquisition":0.30,
        "Active Account Growth":0.10,"Digital Active Customers":0.05,"Diligence Score":0.10},
    "Relationship Officer Bancassurance": {
        "Bancassurance":0.40,"Deposit Growth":0.15,"New Customer Acquisition":0.20,
        "Compliance Score":0.15,"Diligence Score":0.10},
    # COMMERCIAL / HO SALES
    "Relationship Manager - Corporate Banking": {
        "Deposit Growth":0.15,"Loan Book Growth":0.25,"Fees and Commission":0.20,
        "Trade Finance":0.15,"NPL Ratio":0.15,"New Customer Acquisition":0.05,
        "Diligence Score":0.05},
    "Senior Relationship Manager - Corporate Banking": {
        "Deposit Growth":0.15,"Loan Book Growth":0.20,"Fees and Commission":0.20,
        "Trade Finance":0.15,"NPL Ratio":0.15,"New Customer Acquisition":0.05,
        "Diligence Score":0.10},
    "Relationship Manager - SME": {
        "Deposit Growth":0.20,"Loan Book Growth":0.20,"Fees and Commission":0.15,
        "NPL Ratio":0.15,"New Customer Acquisition":0.15,"Compliance Score":0.05,
        "Diligence Score":0.10},
    "Relationship Manager - Agribusiness": {
        "Deposit Growth":0.20,"Loan Book Growth":0.25,"Fees and Commission":0.10,
        "NPL Ratio":0.20,"New Customer Acquisition":0.10,"Compliance Score":0.05,
        "Diligence Score":0.10},
    "Relationship Manager- Trade Finance": {
        "Deposit Growth":0.10,"Trade Finance":0.30,"Fees and Commission":0.25,
        "Loan Book Growth":0.15,"NPL Ratio":0.10,"Diligence Score":0.10},
    "Relationship Manager- Public Sector": {
        "Deposit Growth":0.30,"Fees and Commission":0.20,"Loan Book Growth":0.15,
        "New Customer Acquisition":0.15,"NPL Ratio":0.10,"Diligence Score":0.10},
    "Area Manager": {
        "Deposit Growth":0.20,"Loan Book Growth":0.20,"Fees and Commission":0.10,
        "NPL Ratio":0.10,"New Customer Acquisition":0.10,"Active Account Growth":0.10,
        "Transactions":0.05,"Compliance Score":0.05,"DFS Revenue":0.05,
        "Diligence Score":0.05},
    # HO SUPPORT ROLES
    "Credit Analyst":           {"NPL Ratio":0.30,"Compliance Score":0.30,"Diligence Score":0.40},
    "Credit Admin Officer":     {"NPL Ratio":0.20,"Compliance Score":0.40,"Diligence Score":0.40},
    "Finance Officer":          {"Compliance Score":0.50,"Diligence Score":0.50},
    "ICT Support Officer":      {"SLA Adherence Score":0.50,"Diligence Score":0.50},
    "Contact Centre Officer":   {"Deposit Growth":0.10,"New Customer Acquisition":0.20,"Compliance Score":0.35,"Diligence Score":0.35},
    "Internal Auditor":         {"Audit Score":0.60,"Diligence Score":0.40},
    "Acquiring Officer":        {"Deposit Growth":0.15,"Digital Active Customers":0.30,"DFS Revenue":0.30,"New Customer Acquisition":0.15,"Diligence Score":0.10},
    "Senior Digital Channels Officer": {
        "Deposit Growth":0.15,"Digital Active Customers":0.25,"DFS Revenue":0.25,
        "New Customer Acquisition":0.15,"Transactions":0.10,"Diligence Score":0.10},
    "Collections and Recoveries Officer": {"NPL Ratio":0.50,"Compliance Score":0.30,"Diligence Score":0.20},
}
DEFAULT_KPI = {"Compliance Score":0.50,"Diligence Score":0.50}

# ── STAFF GENERATION ────────────────────────────────────────────────
print("\n" + "="*60)
print("Generating staff...")
print("="*60)

staff_rows   = []   # for Excel
users_dict   = {}   # for users.json
kpi_map_rows = []   # for kpi_map.json
staff_code   = 300001

# ── EXEC & C-SUITE first (fixed staff codes 300001-300020) ──────────
exec_names = {
    "Chief Executive & Managing Director": ("300001","M"),
    "Chief Retail Banking Officer":        ("300002","M"),
    "Chief Commercial Officer":            ("300003","M"),
    "Chief Financial Officer":             ("300004","F"),
    "Chief Credit Officer":                ("300005","M"),
    "Chief Risk Officer":                  ("300006","F"),
    "Chief Information Officer":           ("300007","M"),
    "Chief Operating Officer":             ("300008","F"),
    "Chief Human Resource Officer":        ("300009","F"),
    "Company Secretary and Chief Legal Officer": ("300010","M"),
}

EXEC_PASSWORDS = {  # login: role-derived
    "CEO001": ("ceo001","EcoStaff0001"),
    "CEO002": ("crbo002","EcoStaff0002"),
    "CEO003": ("cco003","EcoStaff0003"),
}

for role_def in HO_ROLES:
    role_title, reports_to, dept, band, count = role_def
    if role_title in exec_names:
        sc_str, gender = exec_names[role_title]
        name, gender = gen_name(gender)
        sc_num = int(sc_str)  # now purely numeric
        sc_num_str = sc_str  # now a numeric string
        uname = name.split()[0].lower() + sc_str[-3:]
        base_u = uname; sfx_u = 1
        while uname in users_dict:
            uname = f"{base_u}{sfx_u}"; sfx_u += 1
        pwd   = f"EcoStaff{sc_str[-4:]}"
        doe   = rand_doe(2005, 2015)

        staff_rows.append({
            "Staff Code": sc_str, "Staff Name": name, "Role": role_title,
            "Unit": "Head Office", "Region": "Head Office", "Category": "Head Office",
            "Department": dept, "Band": band, "Gender": gender,
            "Reports To": reports_to or "", "Date of Employment": doe,
        })
        is_ceo = role_title == "Chief Executive & Managing Director"
        is_md  = is_ceo
        users_dict[uname] = {
            "password": hashlib.sha256(pwd.encode()).hexdigest(),
            "full_name": name, "role": role_title, "unit": "Head Office",
            "staff_code": sc_str, "active": True, "is_admin": is_md,
            "can_view_all": is_md, "must_change_password": False,
            "department": dept, "band": band, "gender": gender,
        }
        kpis = KPI_ASSIGNMENTS.get(role_title, DEFAULT_KPI)
        kpi_map_rows.append({"staff_code": sc_str, "name": name,
                              "role": role_title, "unit": "Head Office",
                              "kpis": json.dumps(kpis)})

print(f"  ✅ C-Suite: {len(exec_names)} executives")

# ── HO STAFF ─────────────────────────────────────────────────────────
for role_def in HO_ROLES:
    role_title, reports_to, dept, band, count = role_def
    if role_title in exec_names:
        continue  # already done above
    for _ in range(count):
        name, gender = gen_name()
        sc  = str(staff_code); staff_code += 1
        uname = name.split()[0].lower() + sc[-3:]
        # ensure unique username
        base = uname; sfx = 1
        while uname in users_dict:
            uname = f"{base}{sfx}"; sfx += 1
        pwd = f"EcoStaff{sc[-4:]}"
        doe = rand_doe(2005, 2024)

        staff_rows.append({
            "Staff Code": sc, "Staff Name": name, "Role": role_title,
            "Unit": "Head Office", "Region": "Head Office", "Category": "Head Office",
            "Department": dept, "Band": band, "Gender": gender,
            "Reports To": reports_to or "", "Date of Employment": doe,
        })
        users_dict[uname] = {
            "password": hashlib.sha256(pwd.encode()).hexdigest(),
            "full_name": name, "role": role_title, "unit": "Head Office",
            "staff_code": sc, "active": True, "is_admin": False,
            "can_view_all": False, "must_change_password": True,
            "department": dept, "band": band, "gender": gender,
        }
        kpis = KPI_ASSIGNMENTS.get(role_title, DEFAULT_KPI)
        kpi_map_rows.append({"staff_code": sc, "name": name,
                              "role": role_title, "unit": "Head Office", "kpis": json.dumps(kpis)})

ho_count = len([r for r in staff_rows if r["Unit"]=="Head Office"])
print(f"  ✅ Head Office: {ho_count} staff")

# ── BRANCH STAFF ─────────────────────────────────────────────────────
TIER_IDX = {1:4, 2:5, 3:6, 4:7}  # column index for tier count in BRANCH_ROLES

for branch_name, region, county, tier in BRANCHES:
    tier_col = TIER_IDX[tier]
    branch_mgr_code = None

    for role_def in BRANCH_ROLES:
        role_title, rep_to, kpi_cat, band = role_def[:4]
        n = role_def[tier_col]
        if n == 0:
            continue

        for i in range(n):
            name, gender = gen_name()
            sc  = str(staff_code); staff_code += 1
            uname = name.split()[0].lower() + sc[-3:]
            base = uname; sfx = 1
            while uname in users_dict:
                uname = f"{base}{sfx}"; sfx += 1
            pwd = f"EcoStaff{sc[-4:]}"
            doe = rand_doe(2008, 2024)

            # Resolve actual reports_to
            if rep_to == "Area Manager":
                actual_rep = "Area Manager"
            elif rep_to == "Branch Manager":
                actual_rep = "Branch Manager"
            else:
                actual_rep = rep_to

            staff_rows.append({
                "Staff Code": sc, "Staff Name": name, "Role": role_title,
                "Unit": branch_name, "Region": region, "Category": "Branch",
                "Department": kpi_cat, "Band": band, "Gender": gender,
                "Reports To": actual_rep, "Date of Employment": doe,
            })
            users_dict[uname] = {
                "password": hashlib.sha256(pwd.encode()).hexdigest(),
                "full_name": name, "role": role_title, "unit": branch_name,
                "staff_code": sc, "active": True, "is_admin": False,
                "can_view_all": False, "must_change_password": True,
                "department": kpi_cat, "band": band, "gender": gender,
                "region": region,
            }
            kpis = KPI_ASSIGNMENTS.get(role_title, DEFAULT_KPI)
            kpi_map_rows.append({"staff_code": sc, "name": name,
                                  "role": role_title, "unit": branch_name, "kpis": json.dumps(kpis)})

branch_count = len([r for r in staff_rows if r["Category"]=="Branch"])
print(f"  ✅ Branch staff: {branch_count} across {len(BRANCHES)} branches")
print(f"  ✅ Total staff:  {len(staff_rows)}")
print(f"  ✅ User accounts:{len(users_dict)}")

# ── VERIFY KPI weights sum to 1.0 ───────────────────────────────────
print("\nVerifying KPI weights...")
errors = 0
for role, kpis in KPI_ASSIGNMENTS.items():
    total = sum(kpis.values())
    if abs(total - 1.0) > 0.005:
        print(f"  ⚠️  {role}: {total:.3f}")
        errors += 1
if not errors:
    print("  ✅ All role KPI weights sum to 1.0")

# ── SAVE staff_register Excel ────────────────────────────────────────
print("\nSaving files...")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Staff Register"

HEADERS = ["Staff Code","Staff Name","Role","Unit","Region","Category",
           "Department","Band","Gender","Reports To","Date of Employment"]
# Header row
hdr_fill = PatternFill("solid", fgColor="006B3F")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
for ci, h in enumerate(HEADERS, 1):
    c = ws.cell(1, ci, h)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal="center")

# Data rows — alternate shading
fill_a = PatternFill("solid", fgColor="F0FDF4")
fill_b = PatternFill("solid", fgColor="FFFFFF")
for ri, row in enumerate(staff_rows, 2):
    fill = fill_a if ri % 2 == 0 else fill_b
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(ri, ci, row.get(h,""))
        c.fill = fill
        c.alignment = Alignment(horizontal="left")

# Column widths
col_widths = [14,28,42,28,16,12,16,8,8,32,20]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A2"

out_xlsx = CBS / "staff_register_v2.xlsx"
wb.save(out_xlsx)
print(f"  ✅ {out_xlsx} ({len(staff_rows)} rows)")

# Copy to a2z/data
import shutil
shutil.copy(out_xlsx, DATA / "staff_register.xlsx")
print(f"  ✅ Copied to a2z/data/staff_register.xlsx")

# ── SAVE users.json ──────────────────────────────────────────────────
# Add protected admin account
users_dict["admin"] = {
    "password": hashlib.sha256(b"admin123").hexdigest(),
    "full_name": "System Admin", "role": "Admin", "unit": "Head Office",
    "staff_code": "ADMIN001", "active": True, "is_admin": True,
    "can_view_all": True, "must_change_password": False, "_protected": True,
}

out_users = CBS / "users_v2.json"
out_users.write_text(json.dumps(users_dict, indent=2, default=str))
shutil.copy(out_users, DATA / "users.json")
print(f"  ✅ users.json ({len(users_dict)} accounts)")

# ── SAVE kpi_map ─────────────────────────────────────────────────────
out_kpi = CBS / "staff_kpi_map_v2.json"
out_kpi.write_text(json.dumps(kpi_map_rows, indent=2))
print(f"  ✅ staff_kpi_map_v2.json ({len(kpi_map_rows)} rows)")

# ── SAVE logins CSV ──────────────────────────────────────────────────
out_logins = CBS / "staff_logins_v2.csv"
with open(out_logins, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["username","password","staff_code",
                                            "name","role","unit","region"])
    writer.writeheader()
    for uname2, ud2 in users_dict.items():
        if uname2 == "admin": continue
        sc2 = str(ud2.get("staff_code",""))
        writer.writerow({
            "username":  uname2,
            "password":  f"EcoStaff{sc2[-4:]}",
            "staff_code":sc2,
            "name":      ud2.get("full_name",""),
            "role":      ud2.get("role",""),
            "unit":      ud2.get("unit",""),
            "region":    ud2.get("region","Head Office"),
        })
# logins already at out_logins
print(f"  ✅ staff_logins_v2.csv")

# ── UPDATE org_config.json ───────────────────────────────────────────
org_config = {
    "bank_name": "Your Bank",
    "bank_code": "YB",
    "country": "Kenya",
    "currency": "KES",
    "currency_symbol": "KES",
    "regions": ["Nairobi CBD","Nairobi Metro","Nairobi North",
                "Mt Kenya West","Mt Kenya East","Eastern",
                "Coast","North Rift","South Rift","West Kenya","Head Office"],
    "branches": [
        {"code": f"BR{str(i+1).zfill(3)}", "name": b[0], "region": b[1],
         "county": b[2], "type": {1:"Flagship",2:"Main",3:"Standard",4:"Light"}[b[3]],
         "tier": b[3]}
        for i, b in enumerate(BRANCHES)
    ],
    "roles": sorted(set(r["Role"] for r in staff_rows)),
    "hierarchy": {
        "Chief Executive & Managing Director": [],
        "Chief Retail Banking Officer":  ["Chief Executive & Managing Director"],
        "Chief Commercial Officer":      ["Chief Executive & Managing Director"],
        "Chief Financial Officer":       ["Chief Executive & Managing Director"],
        "Chief Credit Officer":          ["Chief Executive & Managing Director"],
        "Chief Risk Officer":            ["Chief Executive & Managing Director"],
        "Chief Information Officer":     ["Chief Executive & Managing Director"],
        "Chief Operating Officer":       ["Chief Executive & Managing Director"],
        "Chief Human Resource Officer":  ["Chief Executive & Managing Director"],
        "Company Secretary and Chief Legal Officer": ["Chief Executive & Managing Director"],
        "Head of Branches":              ["Chief Retail Banking Officer"],
        "Area Manager":                  ["Head of Branches"],
        "Senior Branch Manager":         ["Area Manager"],
        "Branch Manager":                ["Area Manager"],
        "Branch Operations Manager":     ["Branch Manager","Senior Branch Manager"],
        "Branch Operations Supervisor":  ["Branch Operations Manager"],
        "Teller":                        ["Branch Operations Supervisor"],
        "Customer Service Officer":      ["Branch Operations Supervisor"],
        "Senior Digital Channels Officer":["Branch Operations Manager"],
        "Branch Senior Relationship Officer":["Branch Manager","Senior Branch Manager"],
        "Branch Relationship Manager":   ["Branch Manager","Senior Branch Manager"],
        "Relationship Officer-Business Banker": ["Branch Relationship Manager","Branch Senior Relationship Officer"],
        "Relationship Officer-Personal Banker": ["Branch Relationship Manager","Branch Senior Relationship Officer"],
        "Direct Sales Representative - Assets & Liabilities": ["Branch Manager","Senior Branch Manager"],
        "Relationship Officer Bancassurance": ["Branch Manager"],
        "Head Of Corporates & Trade Finance": ["Chief Commercial Officer"],
        "Head of MSME":                  ["Chief Commercial Officer"],
        "Head of Government & Institutional Banking": ["Chief Commercial Officer"],
        "Relationship Manager - Corporate Banking": ["Head Of Corporates & Trade Finance"],
        "Senior Relationship Manager - Corporate Banking": ["Head Of Corporates & Trade Finance"],
        "Relationship Manager - SME":    ["Head of MSME"],
        "Relationship Manager - Agribusiness": ["Head of MSME"],
        "Relationship Manager- Trade Finance": ["Head Of Corporates & Trade Finance"],
        "Relationship Manager- Public Sector": ["Head of Government & Institutional Banking"],
        "Head of Digital Financial Services": ["Chief Commercial Officer"],
        "Senior Manager Direct Sales Force": ["Chief Retail Banking Officer"],
        "Head of Treasury":              ["Chief Financial Officer"],
        "Head Of ICT":                   ["Chief Information Officer"],
        "Head of Operations":            ["Chief Operating Officer"],
        "Head, Customer Experience":     ["Chief Operating Officer"],
        "Senior Manager Internal Audit": ["Chief Executive & Managing Director"],
        "Senior Manager- Compliance":    ["Chief Risk Officer"],
        "Senior Manager -Credit Analysis": ["Chief Credit Officer"],
        "Senior Manager-Collections & Recoveries": ["Chief Credit Officer"],
        "Head of Procurement":           ["Chief Operating Officer"],
        "Head Of Marketing and Corporate Communication": ["Chief Executive & Managing Director"],
        "General Manager - Bancassurance": ["Chief Commercial Officer"],
    },
    "pillar_weights": {
        "Financial": 0.40,
        "Customer Focus": 0.25,
        "Operational Excellence": 0.25,
        "People & Learning": 0.10,
    },
}

(DATA / "org_config.json").write_text(json.dumps(org_config, indent=2))
print(f"  ✅ org_config.json ({len(org_config['branches'])} branches, "
      f"{len(org_config['roles'])} roles)")

# ── SUMMARY ─────────────────────────────────────────────────────────
from collections import Counter
roles_counter = Counter(r["Role"] for r in staff_rows)
regions_counter = Counter(r["Region"] for r in staff_rows)
gender_counter  = Counter(r["Gender"] for r in staff_rows)

print(f"\n{'='*60}")
print("GENERATION SUMMARY")
print(f"{'='*60}")
print(f"  Total staff:        {len(staff_rows):,}")
print(f"  Total accounts:     {len(users_dict):,}")
print(f"  Branches:           {len(BRANCHES)}")
print(f"  Regions:            {len(set(b[1] for b in BRANCHES))}")
print(f"  Gender: M={gender_counter.get('M',0)} F={gender_counter.get('F',0)}")
print(f"\n  Top roles:")
for role, cnt in roles_counter.most_common(10):
    print(f"    {cnt:4d}  {role}")
print(f"\n  By region:")
for region, cnt in sorted(regions_counter.items(), key=lambda x:-x[1]):
    print(f"    {cnt:4d}  {region}")
print(f"\n  Sample logins:")
sample_roles = ["Chief Executive & Managing Director","Area Manager","Branch Manager","Teller"]
for r in staff_rows[:200]:
    if r["Role"] in sample_roles:
        sc = str(r["Staff Code"])
        uname_s = r["Staff Name"].split()[0].lower() + sc[-3:]
        print(f"    {uname_s} / EcoStaff{sc[-4:]}  →  {r['Role']} @ {r['Unit']}")
        sample_roles.remove(r["Role"])
    if not sample_roles:
        break

print(f"\n✅ Done. Next: run compute_baseline.py then compute_actuals.py")
