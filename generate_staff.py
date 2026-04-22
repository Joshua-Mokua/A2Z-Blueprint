"""
generate_staff.py — Generate full staff register from CBS branch structure.
Creates:
  cbs_data/staff_register.xlsx  — staff master (same format app reads)
  cbs_data/users.json           — login credentials for all staff
  cbs_data/staff_kpi_map.json  — which KPIs each role carries

Run AFTER generate_cbs.py
"""
import json, random, csv
from pathlib import Path
from datetime import date, timedelta

random.seed(99)
CBS  = Path(__file__).parent / "cbs_data"
DATA = Path(__file__).parent / "a2z" / "data"
DATA.mkdir(exist_ok=True)

# ── Load branches ─────────────────────────────────────────────────────
branches = json.loads((CBS/"branches.json").read_text())
branch_map = {b["branch_code"]: b for b in branches}

# ── Kenyan name pools ─────────────────────────────────────────────────
FIRST_M = ["James","John","Peter","David","Paul","Joseph","Michael","Patrick","Francis",
           "Daniel","Samuel","Philip","George","Charles","Robert","William","Thomas",
           "Kevin","Dennis","Eric","Stephen","Brian","Victor","Moses","Aaron","Simon",
           "Emmanuel","Kenneth","Leonard","Alex","Mark","Anthony","Calvin","Edwin",
           "Festus","Gilbert","Isaac","Joel","Keith","Lawrence","Martin","Nathan"]
FIRST_F = ["Mary","Grace","Faith","Ann","Rose","Jane","Susan","Patricia","Elizabeth",
           "Catherine","Agnes","Beatrice","Charity","Diana","Esther","Florence","Gloria",
           "Hannah","Irene","Joyce","Karen","Lydia","Margaret","Nancy","Olive","Priscilla",
           "Rebecca","Sarah","Teresa","Alice","Caroline","Daisy","Ellen","Fatuma","Gladys",
           "Harriet","Ida","Jacqueline","Kezia","Lilian","Miriam","Naomi","Olivia","Pauline"]
SURNAMES = ["Kamau","Odhiambo","Wanjiku","Mwangi","Otieno","Kariuki","Mutua","Njoroge",
            "Akinyi","Ochieng","Kiprotich","Chebet","Langat","Ruto","Koech","Bett","Sang",
            "Waweru","Ngugi","Githu","Gicheru","Kabui","Njenga","Ndegwa","Mungai","Murage",
            "Ouma","Ogolla","Okello","Anyango","Adhiambo","Awino","Hassan","Omar","Mohamed",
            "Mwenda","Kithinji","Kirimi","Murithi","Njue","Wafula","Simiyu","Barasa",
            "Masinde","Wekesa","Mwamba","Mwanzia","Munyao","Ndeti","Musyoka","Shimba"]

def random_name():
    g = random.choice(["M","F"])
    f = random.choice(FIRST_M if g=="M" else FIRST_F)
    s = random.choice(SURNAMES)
    return f"{f} {s}", g

def random_date(start_year=2010, end_year=2024):
    start = date(start_year, 1, 1)
    end   = date(end_year, 12, 31)
    return start + timedelta(days=random.randint(0,(end-start).days))

# ── Role definitions — exact match to cascade HIERARCHY ──────────────
# Format: (role, reports_to, kpi_category, salary_band, count_per_branch_tier)
# Tier: 1=flagship, 2=main, 3=standard, 4=light
ROLE_DEFS = [
    # role                                reports_to                      category      band  T1  T2  T3  T4
    ("Branch Manager",                    "Regional Head",                "Management", "M4", 1,  1,  1,  1),
    ("Branch Operations Manager",         "Branch Manager",               "Operations", "M3", 1,  1,  1,  1),
    ("Branch Credit Manager",             "Branch Manager",               "Credit",     "M3", 1,  1,  1,  0),
    ("Branch Operations Supervisor",      "Branch Operations Manager",    "Operations", "M2", 1,  1,  1,  0),
    ("Relationship Officer Personal Banking","Branch Credit Manager",     "Sales",      "M2", 3,  2,  1,  1),
    ("Relationship Officer Business Banking","Branch Credit Manager",     "Sales",      "M2", 2,  2,  1,  0),
    ("Direct Sales Officer",              "Branch Credit Manager",        "Sales",      "M1", 3,  2,  2,  1),
    ("Customer Service Officer",          "Branch Operations Manager",    "Operations", "M1", 2,  2,  1,  1),
    ("Teller",                            "Branch Operations Manager",    "Operations", "M1", 4,  3,  2,  1),
]

# HO roles
HO_ROLES = [
    ("Managing Director",             None,                          "Executive",  "E1", 1),
    ("Director Retail Banking",       "Managing Director",           "Executive",  "E2", 1),
    ("Director Commercial Banking",   "Managing Director",           "Executive",  "E2", 1),
    ("Head Of Retail",                "Director Retail Banking",     "Management", "M5", 1),
    ("Head Of Corporate",             "Director Commercial Banking", "Management", "M5", 1),
    ("Head Of SME",                   "Director Commercial Banking", "Management", "M5", 1),
    ("Regional Head",                 "Head Of Retail",              "Management", "M5", 3),
    ("Chief Finance Officer",         "Managing Director",           "Finance",    "E2", 1),
    ("Chief Risk Officer",            "Managing Director",           "Risk",       "E2", 1),
    ("Chief Operations Officer",      "Managing Director",           "Operations", "E2", 1),
    ("Chief Compliance Officer",      "Managing Director",           "Compliance", "E2", 1),
    ("Chief Human Resources Officer", "Managing Director",           "HR",         "E2", 1),
    ("Head Of Digital Innovation",    "Managing Director",           "Digital",    "M5", 1),
    ("Head Of Strategy",              "Managing Director",           "Strategy",   "M5", 1),
    ("Head Of Internal Audit",        "Managing Director",           "Audit",      "M5", 1),
    ("Head Of Marketing",             "Managing Director",           "Marketing",  "M5", 1),
    ("Relationship Manager Corporate","Head Of Corporate",           "Sales",      "M4", 5),
    ("Relationship Manager SME",      "Head Of SME",                 "Sales",      "M4", 5),
    ("Chief Credit Officer",          "Managing Director",           "Credit",     "E2", 1),
    ("Debt Recovery Unit Manager",    "Managing Director",           "Credit",     "M4", 1),
    ("Financial Controller",          "Chief Finance Officer",       "Finance",    "M4", 1),
    ("Treasury Manager",              "Chief Finance Officer",       "Finance",    "M4", 1),
    ("Risk Manager",                  "Chief Risk Officer",          "Risk",       "M4", 1),
    ("Compliance Officer",            "Chief Compliance Officer",    "Compliance", "M3", 2),
    ("HR Business Partner",           "Chief Human Resources Officer","HR",        "M3", 2),
    ("IT Manager",                    "Head Of Digital Innovation",  "Digital",    "M4", 1),
    ("Operations Manager",            "Chief Operations Officer",    "Operations", "M4", 1),
    ("Procurement Manager",           "Chief Operations Officer",    "Operations", "M4", 1),
    ("Internal Auditor",              "Head Of Internal Audit",      "Audit",      "M3", 2),
    ("Strategy Analyst",              "Head Of Strategy",            "Strategy",   "M2", 1),
    ("Marketing Officer",             "Head Of Marketing",           "Marketing",  "M2", 2),
    ("IT Support Officer",            "IT Manager",                  "Digital",    "M1", 2),
    ("Recovery Officer",              "Debt Recovery Unit Manager",  "Credit",     "M2", 2),
    ("Procurement Officer",           "Procurement Manager",         "Operations", "M2", 1),
    ("Credit Analyst",                "Chief Credit Officer",        "Credit",     "M3", 2),
    ("Credit Administrator",          "Chief Credit Officer",        "Credit",     "M2", 2),
    ("HR Officer",                    "Chief Human Resources Officer","HR",        "M2", 1),
]

# KPI map per role
KPI_MAP = {
    # Sales roles
    "Relationship Officer Personal Banking":  ["Loan Book Growth","Deposit Growth","New Customer Acquisition","Fees and Commission","Compliance Score"],
    "Relationship Officer Business Banking":  ["Loan Book Growth","Deposit Growth","New Customer Acquisition","Fees and Commission","NPL Ratio","Compliance Score"],
    "Direct Sales Officer":                   ["New Customer Acquisition","Deposit Growth","Loan Book Growth","Transactions","Compliance Score"],
    "Branch Credit Manager":                  ["Loan Book Growth","Loans Disbursement","NPL Ratio","Fees and Commission","New Customer Acquisition","Compliance Score"],
    "Branch Operations Manager":              ["Transactions","Dormancy Reactivation","Compliance Score","Digital Active Customers"],
    "Branch Operations Supervisor":           ["Transactions","Dormancy Reactivation","Compliance Score"],
    "Customer Service Officer":               ["Transactions","New Customer Acquisition","Dormancy Reactivation","Compliance Score"],
    "Teller":                                 ["Transactions","Compliance Score"],
    "Branch Manager":                         ["Loan Book Growth","Deposit Growth","Fees and Commission","New Customer Acquisition","NPL Ratio","Transactions","Compliance Score","Digital Active Customers"],
    "Regional Head":                          ["Loan Book Growth","Deposit Growth","Fees and Commission","New Customer Acquisition","NPL Ratio","Transactions","Compliance Score"],
    "Relationship Manager Corporate":         ["Loan Book Growth","Loans Disbursement","Deposit Growth","Fees and Commission","NPL Ratio","Compliance Score"],
    "Relationship Manager SME":               ["Loan Book Growth","Deposit Growth","Fees and Commission","New Customer Acquisition","NPL Ratio","Compliance Score"],
}
# Defaults for other roles
DEFAULT_KPIS = ["Compliance Score","Staff Satisfaction","Diligence Score"]

# ── Build staff list ──────────────────────────────────────────────────
print("Building staff register from CBS branches...")
staff_list = []
code_counter = [300001]
used_names   = set()

def next_code():
    c = code_counter[0]; code_counter[0] += 1; return str(c)

def unique_name():
    for _ in range(100):
        n,g = random_name()
        if n not in used_names:
            used_names.add(n); return n,g
    n,g = random_name(); return n,g  # fallback

# HO staff first
ho_branch = next((b for b in branches if b["branch_code"]=="BRN001"), branches[0])
region_heads_added = 0

for role, reports_to, category, band, count in HO_ROLES:
    for i in range(count):
        name,gender = unique_name()
        sc = next_code()
        hire = random_date(2008, 2022)
        region = "North" if region_heads_added==0 else ("Central" if region_heads_added==1 else "South")
        if role == "Regional Head": region_heads_added += 1
        staff_list.append({
            "Staff Code":    sc,
            "Staff Name":    name,
            "Gender":        gender,
            "Role":          role,
            "Reports To":    reports_to or "",
            "Category":      category,
            "Band":          band,
            "Unit":          "Head Office",
            "Branch Code":   "BRN001",
            "Branch Name":   "Head Office",
            "Region":        region if role=="Regional Head" else "Head Office",
            "County":        "Nairobi",
            "Staff Status":  "Active",
            "Hire Date":     str(hire),
            "Email":         name.split()[0].lower()+"."+name.split()[-1].lower()+"@ecobank.com",
        })

print(f"  HO staff: {len(staff_list)}")

# Branch staff
for b in branches:
    if b["branch_code"] == "BRN001": continue
    tier = b["tier"]
    for role, reports_to, category, band, t1,t2,t3,t4 in ROLE_DEFS:
        count = {1:t1, 2:t2, 3:t3, 4:t4}.get(tier, 0)
        for _ in range(count):
            name, gender = unique_name()
            sc = next_code()
            hire = random_date(2010, 2024)
            staff_list.append({
                "Staff Code":  sc,
                "Staff Name":  name,
                "Gender":      gender,
                "Role":        role,
                "Reports To":  reports_to or "",
                "Category":    category,
                "Band":        band,
                "Unit":        b["branch_name"],
                "Branch Code": b["branch_code"],
                "Branch Name": b["branch_name"],
                "Region":      b["region"],
                "County":      b["county"],
                "Staff Status":"Active",
                "Hire Date":   str(hire),
                "Email":       name.split()[0].lower()+"."+name.split()[-1].lower()+"@ecobank.com",
            })

print(f"  Total staff: {len(staff_list)}")

# ── Save staff_register.xlsx ──────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Register"

    headers = ["Staff Code","Staff Name","Gender","Role","Reports To","Category",
               "Band","Unit","Branch Code","Branch Name","Region","County",
               "Staff Status","Hire Date","Email"]
    # Header row
    for ci,h in enumerate(headers,1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for ri, s in enumerate(staff_list, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=s.get(h,""))

    # Column widths
    widths = [12,25,8,35,30,15,8,30,12,30,15,15,12,12,35]
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out_path = CBS/"staff_register.xlsx"
    wb.save(out_path)
    print(f"  ✅ staff_register.xlsx — {len(staff_list)} staff")
except Exception as e:
    print(f"  ⚠️ Excel save failed: {e} — saving as CSV instead")
    import csv as _csv
    out_path = CBS/"staff_register.csv"
    with open(out_path,"w",newline="",encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=list(staff_list[0].keys()))
        w.writeheader(); w.writerows(staff_list)
    print(f"  ✅ staff_register.csv — {len(staff_list)} staff")

# ── Save users.json ───────────────────────────────────────────────────
import hashlib, json as _json
existing_users = {}
users_path = DATA/"users.json"
if users_path.exists():
    existing_users = _json.loads(users_path.read_text())

users = dict(existing_users)  # keep existing admin accounts
for s in staff_list:
    uname = s["Staff Name"].split()[0].lower() + s["Staff Code"][-3:]
    # Default password: EcoStaff + last 4 digits of staff code
    pwd = "EcoStaff" + s["Staff Code"][-4:]
    role_l = s["Role"].lower()
    is_admin = "managing" in role_l or "director" in role_l or "chief" in role_l
    is_mgr   = is_admin or any(k in role_l for k in ("manager","supervisor","head","regional"))
    users[uname] = {
        "password":      hashlib.sha256(pwd.encode()).hexdigest(),
        "full_name":     s["Staff Name"],
        "role":          s["Role"],
        "unit":          s["Unit"],
        "staff_code":    s["Staff Code"],
        "branch_code":   s["Branch Code"],
        "is_admin":      is_admin,
        "can_view_all":  is_admin,
        "region":        s["Region"],
        "active":        True,
        "must_change_password": True,  # force password change on first login
    }

users_path.write_text(_json.dumps(users, indent=2))
print(f"  ✅ users.json — {len(users)} accounts (password: EcoStaff + last 4 of staff code)")

# ── Save KPI map ──────────────────────────────────────────────────────
(CBS/"staff_kpi_map.json").write_text(_json.dumps(KPI_MAP, indent=2))
print(f"  ✅ staff_kpi_map.json")

# ── Summary ───────────────────────────────────────────────────────────
by_role = {}
for s in staff_list:
    by_role[s["Role"]] = by_role.get(s["Role"],0)+1
print("\n  Staff by role (top 10):")
for role,cnt in sorted(by_role.items(),key=lambda x:-x[1])[:10]:
    print(f"    {role:<40} {cnt:>4}")
print(f"\n  Total: {len(staff_list)} staff across {len(set(s['Branch Code'] for s in staff_list))} branches")
