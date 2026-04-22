"""
compute_baseline.py — Snapshot CBS portfolio positions as at 31 Dec 2025.
This creates the baseline against which 2026 actuals growth is measured.

Run ONCE after generate_cbs.py and generate_staff.py.
Output: cbs_data/baseline_2025_Dec.json
        a2z/data/baseline_2025_Dec.json  (copy for app)
"""
import csv, json
from pathlib import Path
from collections import defaultdict
from datetime import date

CBS  = Path(__file__).parent / "cbs_data"
DATA = Path(__file__).parent / "a2z" / "data"
DATA.mkdir(exist_ok=True)

BASELINE_DATE = date(2025, 12, 31)
BASELINE_PERIOD = "2025_Dec"

print("="*60)
print("CBS Baseline Computation — 31 Dec 2025")
print("="*60)

# ── Load staff ────────────────────────────────────────────────────
print("\nLoading staff register...")
staff_list = []
try:
    import openpyxl
    wb = openpyxl.load_workbook(CBS/"staff_register.xlsx")
    ws = wb.active
    headers = [ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    for row in ws.iter_rows(min_row=2,values_only=True):
        r = dict(zip(headers,row))
        if r.get("Staff Code"): staff_list.append(r)
except Exception as e:
    print(f"  Excel failed: {e} — trying CSV")
    with open(CBS/"staff_register.csv",encoding="utf-8") as f:
        staff_list = list(csv.DictReader(f))

print(f"  {len(staff_list)} staff loaded")
staff_by_code  = {str(s["Staff Code"]): s for s in staff_list}

# ── Accumulators ──────────────────────────────────────────────────
rm_baseline    = defaultdict(lambda: {
    "deposit_bal":0.0, "loan_bal":0.0, "fee_income":0.0,
    "npl_bal":0.0, "total_loan":0.0,
    "active_accounts":0, "dormant_accounts":0, "total_accounts":0,
    "cif_count":0, "cifs":set(),
    "loan_count":0,
})
branch_baseline = defaultdict(lambda: {
    "deposit_bal":0.0, "loan_bal":0.0, "npl_bal":0.0,
    "active_accounts":0, "dormant_accounts":0, "new_customers_ytd":0,
    "kyc_verified":0, "kyc_total":0,
})

# ── Stream accounts ───────────────────────────────────────────────
print("\nStreaming accounts for baseline snapshot...")
with open(CBS/"accounts.csv",encoding="utf-8") as f:
    for i, row in enumerate(csv.DictReader(f)):
        rm  = row.get("relationship_manager_code","")
        bc  = row.get("branch_code","")
        cat = row.get("category","")
        bal = float(row.get("current_balance") or 0)
        loan_out = float(row.get("loan_outstanding") or 0)
        fee  = float(row.get("fee_income_ytd") or 0)
        npl  = row.get("npl_status","")
        status = row.get("account_status","")
        cif  = row.get("cif","")
        dormancy = row.get("dormancy_status","")

        # RM portfolio baseline
        r = rm_baseline[rm]
        r["cifs"].add(cif)
        r["total_accounts"] += 1
        if status == "Active": r["active_accounts"] += 1
        else: r["dormant_accounts"] += 1

        if cat in ("CASA","Term Deposit"):
            r["deposit_bal"] += bal
        if cat == "Loan":
            r["loan_bal"]   += loan_out
            r["total_loan"] += loan_out
            r["fee_income"] += fee
            r["loan_count"] += 1
            if npl == "NPL": r["npl_bal"] += loan_out
        else:
            r["fee_income"] += fee * 0.3

        # Branch baseline
        b = branch_baseline[bc]
        if cat in ("CASA","Term Deposit"): b["deposit_bal"] += bal
        if cat == "Loan":
            b["loan_bal"] += loan_out
            if npl == "NPL": b["npl_bal"] += loan_out
        if dormancy == "Active": b["active_accounts"] += 1
        else: b["dormant_accounts"] += 1

        if (i+1) % 200000 == 0:
            print(f"  {i+1:,} accounts...")

print("  ✅ Accounts processed")

# ── Stream customers for new customer count ───────────────────────
print("\nStreaming customers...")
with open(CBS/"customers.csv",encoding="utf-8") as f:
    for row in csv.DictReader(f):
        bc = row.get("branch_code","")
        kyc = row.get("kyc_status","")
        onboarded = row.get("date_onboarded","")
        b = branch_baseline[bc]
        b["kyc_total"] += 1
        if kyc == "Verified": b["kyc_verified"] += 1
        if onboarded and onboarded[:4] == "2025":
            b["new_customers_ytd"] += 1

print("  ✅ Customers processed")

# ── Build baseline dict ───────────────────────────────────────────
print("\nBuilding baseline...")

# Convert sets to counts
rm_baseline_clean = {}
for rm, data in rm_baseline.items():
    d = dict(data)
    d["cif_count"] = len(d.pop("cifs",set()))
    d["npl_ratio"] = round(d["npl_bal"]/max(d["total_loan"],1)*100,2)
    d["active_account_rate"] = round(d["active_accounts"]/max(d["total_accounts"],1)*100,1)
    # Round all floats
    for k,v in d.items():
        if isinstance(v,float): d[k] = round(v,2)
    rm_baseline_clean[rm] = d

branch_baseline_clean = {}
for bc, data in branch_baseline.items():
    d = dict(data)
    d["npl_ratio"] = round(d["npl_bal"]/max(d.get("loan_bal",1),1)*100,2)
    d["compliance_score"] = round(d["kyc_verified"]/max(d["kyc_total"],1)*100,1)
    for k,v in d.items():
        if isinstance(v,float): d[k] = round(v,2)
    branch_baseline_clean[bc] = d

baseline = {
    "period": BASELINE_PERIOD,
    "date":   str(BASELINE_DATE),
    "rm":     rm_baseline_clean,
    "branch": branch_baseline_clean,
}

# ── Save ──────────────────────────────────────────────────────────
out_name = f"baseline_{BASELINE_PERIOD}.json"
for dest in [CBS/out_name, DATA/out_name]:
    dest.write_text(json.dumps(baseline, indent=2))
    print(f"  ✅ {dest}")

# ── Summary ───────────────────────────────────────────────────────
total_dep  = sum(v["deposit_bal"] for v in rm_baseline_clean.values())
total_loan = sum(v["loan_bal"]    for v in rm_baseline_clean.values())
total_npl  = sum(v["npl_bal"]     for v in rm_baseline_clean.values())
print(f"\n{'='*60}")
print("BASELINE SUMMARY — 31 Dec 2025")
print(f"{'='*60}")
print(f"  RMs with portfolio:    {len(rm_baseline_clean):,}")
print(f"  Branches:              {len(branch_baseline_clean):,}")
print(f"  Total deposit book:    KES {total_dep/1e9:.1f}B")
print(f"  Total loan book:       KES {total_loan/1e9:.1f}B")
print(f"  NPL:                   KES {total_npl/1e9:.1f}B ({total_npl/max(total_loan,1)*100:.1f}%)")
print(f"{'='*60}")
print(f"\nBaseline saved. compute_actuals.py will now compute GROWTH vs this baseline.")
